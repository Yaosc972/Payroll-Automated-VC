#!/usr/bin/env python3
"""Run a real-material FBU performance E2E replay and offline-base comparison.

The script intentionally drives the FastAPI endpoints with real files instead of
calling parser internals directly. Output is isolated under output/ so it does
not mutate the normal local activity store.
"""
from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
from collections import defaultdict
from contextlib import ExitStack
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import msoffcrypto
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import bonus_platform.app as app_module  # noqa: E402
from bonus_platform.engine.fbu_performance.runs import FBURosterStore, FBURunManager  # noqa: E402


DATA_ROOT = Path("/Users/zt27532/Documents/FBU美洲大区激励方案")
DEFAULT_FILES = {
    "roster": DATA_ROOT / "4月绩效奖金/OEHR数据表/4月花名册5.20.xlsx",
    "attendance": DATA_ROOT / "4月绩效奖金/OEHR数据表/考勤日报表-20260520.xlsx",
    "salary": DATA_ROOT / "4月绩效奖金/OEHR数据表/1779277434142薪酬档案-（含离职）.xlsx",
    "supplemental_leave": DATA_ROOT / "临时数据/4月线下病假&年假材料.xlsx",
    "performance_book": DATA_ROOT / "4月绩效奖金/4月仓库管理+非仓/202604绩效-新泽西区仓库管理&非仓人员.xlsx",
    "offline_summary": DATA_ROOT / "【汇总】202604新泽西仓库管理绩效基数.xlsx",
}

DEFAULT_96_WORK_HOUR_RULES = [
    ("zt12979", "赵婉妍"),
    ("zt12988", "陈海冰"),
    ("zt14260", "陈炜"),
    ("zt17850", "韩勇"),
]


@dataclass
class WorkbookPassword:
    key: str
    password: str


PASSWORDS = [
    WorkbookPassword("performance_book", "FBU2026"),
    WorkbookPassword("offline_summary", "202604"),
]


def normalize_employee_id(value: Any) -> str:
    employee_id = str(value or "").strip()
    if employee_id.endswith("-1"):
        return employee_id[:-2]
    return employee_id


def as_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("$", "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def ensure_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(str(path))


def copy_or_decrypt(src: Path, dst: Path, password: str | None = None) -> Path:
    """Copy a workbook to dst, decrypting when a password is provided."""
    ensure_file(src)
    dst.parent.mkdir(parents=True, exist_ok=True)
    if not password:
        shutil.copyfile(src, dst)
        return dst

    with src.open("rb") as infile:
        office_file = msoffcrypto.OfficeFile(infile)
        office_file.load_key(password=password)
        with dst.open("wb") as outfile:
            office_file.decrypt(outfile)
    return dst


def extract_sheet(src: Path, sheet_name: str, dst: Path, output_sheet_name: str | None = None) -> Path:
    wb = load_workbook(src, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{src.name} 缺少工作表: {sheet_name}")

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = output_sheet_name or sheet_name
    in_ws = wb[sheet_name]
    for row in in_ws.iter_rows(values_only=True):
        out_ws.append(list(row))

    dst.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(dst)
    return dst


def build_work_hour_rule_payload() -> dict[str, list[dict[str, Any]]]:
    """Build the maintained rule-list payload used by the real E2E flow."""
    return {
        "work_hour_employees": [
            {"employee_id": employee_id, "name": name, "active": True}
            for employee_id, name in DEFAULT_96_WORK_HOUR_RULES
        ],
        "fixed_base_employees": [],
    }


def create_work_hour_rule_marker(path: Path, calc_month: str) -> Path:
    """Create a small workbook documenting the 96-hour rule rows used by E2E."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "规则名单"
    sheet.append(["工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注"])
    for employee_id, name in DEFAULT_96_WORK_HOUR_RULES:
        sheet.append([employee_id, name, "96工时制", None, calc_month, "启用", "E2E默认规则名单"])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def post_file(
    client: TestClient,
    url: str,
    path: Path,
    data: dict[str, Any] | None = None,
    extra_files: dict[str, Path] | None = None,
) -> dict[str, Any]:
    with ExitStack() as stack:
        fh = stack.enter_context(path.open("rb"))
        files = {
            "file": (
                path.name,
                fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        }
        for field_name, extra_path in (extra_files or {}).items():
            extra_fh = stack.enter_context(extra_path.open("rb"))
            files[field_name] = (
                extra_path.name,
                extra_fh,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        response = client.post(
            url,
            data=data or {},
            files=files,
        )
    if response.status_code != 200:
        raise RuntimeError(f"{url} failed: {response.status_code} {response.text[:1000]}")
    return response.json()


def read_offline_base(summary_path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(summary_path, data_only=True, read_only=True)
    if "基数汇总" not in wb.sheetnames:
        raise ValueError("线下汇总表缺少工作表: 基数汇总")

    ws = wb["基数汇总"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for idx, row in enumerate(rows[:20]):
        values = [str(value or "").strip() for value in row]
        if "工号" in values and any("合计所得" in value for value in values):
            header_idx = idx
            break
    if header_idx is None:
        raise ValueError("基数汇总未找到 工号/合计所得 表头")

    header = [str(value or "").strip() for value in rows[header_idx]]
    id_col = header.index("工号")
    name_col = header.index("中文名字") if "中文名字" in header else None
    amount_col = next(i for i, value in enumerate(header) if "合计所得" in value)

    offline: dict[str, dict[str, Any]] = {}
    for row in rows[header_idx + 1 :]:
        employee_id = normalize_employee_id(row[id_col] if id_col < len(row) else "")
        if not employee_id.startswith("zt"):
            continue
        amount = round(as_float(row[amount_col] if amount_col < len(row) else 0), 2)
        if amount == 0:
            continue
        bucket = offline.setdefault(
            employee_id,
            {
                "employee_id": employee_id,
                "name": row[name_col] if name_col is not None and name_col < len(row) else "",
                "offline_performance_base": 0.0,
                "offline_rows": 0,
            },
        )
        bucket["offline_performance_base"] = round(bucket["offline_performance_base"] + amount, 2)
        bucket["offline_rows"] += 1
    return offline


def read_final_performance_base(performance_path: Path) -> dict[str, dict[str, Any]]:
    """Read final performance bases from the actual bonus calculation output workbook."""
    wb = load_workbook(performance_path, data_only=True, read_only=True)
    final_base: dict[str, dict[str, Any]] = {}

    for sheet_name in wb.sheetnames:
        if (
            sheet_name in {"总金额", "调薪", "仓经调薪差额"}
            or "绩效报表" in sheet_name
            or "花名册" in sheet_name
        ):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = None
        employee_id_col = None
        name_col = None
        base_col = None
        for idx, row in enumerate(rows[:10]):
            headers = [str(value or "").strip() for value in row]
            employee_candidates = {"工号", "员工工号", "employee_id"}
            employee_id_col = next((i for i, value in enumerate(headers) if value in employee_candidates), None)
            base_col = next(
                (
                    i for i, value in enumerate(headers)
                    if "绩效基数" in value or "绩效奖金基数" in value
                ),
                None,
            )
            name_col = next((i for i, value in enumerate(headers) if value in {"姓名", "中文名字", "员工姓名"}), None)
            if employee_id_col is not None and base_col is not None:
                header_idx = idx
                break
        if header_idx is None or employee_id_col is None or base_col is None:
            continue

        for row in rows[header_idx + 1:]:
            employee_id = normalize_employee_id(row[employee_id_col] if employee_id_col < len(row) else "")
            if not employee_id.startswith("zt"):
                continue
            amount = round(as_float(row[base_col] if base_col < len(row) else 0), 2)
            if amount == 0:
                continue
            bucket = final_base.setdefault(
                employee_id,
                {
                    "employee_id": employee_id,
                    "name": row[name_col] if name_col is not None and name_col < len(row) else "",
                    "offline_performance_base": 0.0,
                    "offline_rows": 0,
                },
            )
            bucket["offline_performance_base"] = round(bucket["offline_performance_base"] + amount, 2)
            bucket["offline_rows"] += 1

    return final_base


def aggregate_platform_results(results: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    by_employee: dict[str, dict[str, Any]] = {}
    for row in results:
        source_id = normalize_employee_id(row.get("source_employee_id") or row.get("employee_id"))
        if not source_id:
            continue
        bucket = by_employee.setdefault(
            source_id,
            {
                "employee_id": source_id,
                "name": row.get("name", ""),
                "platform_performance_base": 0.0,
                "platform_bonus": 0.0,
                "row_count": 0,
                "calculation_paths": set(),
                "raw_employee_ids": set(),
                "sick_hours": 0.0,
                "sick_settlement_hours": 0.0,
                "annual_hours": 0.0,
                "base_override_types": set(),
                "hourly_rates": set(),
                "single_hourly_rate": 0.0,
                "has_segments": False,
            },
        )
        bucket["platform_performance_base"] += as_float(row.get("performance_base"))
        bucket["platform_bonus"] += as_float(row.get("performance_bonus"))
        bucket["row_count"] += 1
        if row.get("calculation_path"):
            bucket["calculation_paths"].add(row["calculation_path"])
        if row.get("employee_id"):
            bucket["raw_employee_ids"].add(row["employee_id"])
        bucket["sick_hours"] += as_float(row.get("sick_hours"))
        bucket["sick_settlement_hours"] += as_float(row.get("sick_settlement_hours"))
        bucket["annual_hours"] += as_float(row.get("annual_hours"))
        if row.get("base_override_type"):
            bucket["base_override_types"].add(row["base_override_type"])
        hourly_rate = as_float(row.get("hourly_rate"))
        if hourly_rate:
            bucket["hourly_rates"].add(round(hourly_rate, 2))
        if row.get("calculation_segments"):
            bucket["has_segments"] = True

    for bucket in by_employee.values():
        bucket["platform_performance_base"] = round(bucket["platform_performance_base"], 2)
        bucket["platform_bonus"] = round(bucket["platform_bonus"], 2)
        bucket["sick_hours"] = round(bucket["sick_hours"], 2)
        bucket["sick_settlement_hours"] = round(bucket["sick_settlement_hours"], 2)
        bucket["annual_hours"] = round(bucket["annual_hours"], 2)
        bucket["calculation_paths"] = " / ".join(sorted(bucket["calculation_paths"]))
        bucket["raw_employee_ids"] = " / ".join(sorted(bucket["raw_employee_ids"]))
        bucket["base_override_types"] = " / ".join(sorted(bucket["base_override_types"]))
        rates = sorted(bucket["hourly_rates"])
        bucket["hourly_rates"] = " / ".join(f"{rate:g}" for rate in rates)
        bucket["single_hourly_rate"] = rates[0] if len(rates) == 1 else 0.0
    return by_employee


def filter_platform_to_final_target(
    platform: dict[str, dict[str, Any]],
    final_target: dict[str, dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], list[str]]:
    target_ids = set(final_target)
    filtered = {employee_id: row for employee_id, row in platform.items() if employee_id in target_ids}
    platform_only = sorted(set(platform) - target_ids)
    return filtered, platform_only


def aggregate_supplemental_leave(preview: dict[str, Any] | None) -> dict[str, dict[str, Any]]:
    """Summarize supplemental leave rows by original employee id for mismatch diagnosis.

    The actual engine applies supplemental leave by exact employee id, including
    the day-shift "-1" suffix. The comparison workbook is employee-level, so the
    diagnostic output normalizes ids only for reporting.
    """
    by_employee: dict[str, dict[str, Any]] = {}
    if not preview:
        return by_employee

    for row in preview.get("rows", []):
        employee_id = normalize_employee_id(row.get("employee_id"))
        if not employee_id:
            continue

        bucket = by_employee.setdefault(
            employee_id,
            {
                "supp_total_rows": 0,
                "supp_confirmed_count": 0,
                "supp_pending_count": 0,
                "supp_excluded_count": 0,
                "supp_unmatched_count": 0,
                "supp_include_hours": 0.0,
                "supp_sick_hours": 0.0,
                "supp_annual_hours": 0.0,
                "supp_sick_settlement_hours": 0.0,
                "supp_row_summary": [],
            },
        )

        leave_type = str(row.get("leave_type") or row.get("raw_leave_type") or "").strip()
        status = str(row.get("confirmation_status") or "").strip()
        original_hours = as_float(row.get("hours"))
        included_hours = as_float(row.get("included_hours", original_hours))

        bucket["supp_total_rows"] += 1
        if status == "confirmed":
            bucket["supp_confirmed_count"] += 1
        elif status == "pending":
            bucket["supp_pending_count"] += 1
        elif status == "excluded":
            bucket["supp_excluded_count"] += 1
        if row.get("attendance_matched") is False:
            bucket["supp_unmatched_count"] += 1

        if row.get("include_in_base"):
            bucket["supp_include_hours"] += included_hours
            if "病假清算" in leave_type or "离职病假" in leave_type:
                bucket["supp_sick_settlement_hours"] += included_hours
            elif "年假" in leave_type:
                bucket["supp_annual_hours"] += included_hours
            elif "病假" in leave_type:
                bucket["supp_sick_hours"] += included_hours

        if len(bucket["supp_row_summary"]) < 10:
            exact_id = str(row.get("employee_id") or "").strip()
            allocation = str(row.get("allocation_period") or "").strip()
            reason = str(row.get("system_reason") or row.get("confirmation_note") or "").strip()
            summary = (
                f"{exact_id} {leave_type} {status} "
                f"计入{included_hours:g}/原始{original_hours:g} {allocation}"
            ).strip()
            if reason:
                summary = f"{summary} {reason[:36]}"
            bucket["supp_row_summary"].append(summary)

    for bucket in by_employee.values():
        for key in (
            "supp_include_hours",
            "supp_sick_hours",
            "supp_annual_hours",
            "supp_sick_settlement_hours",
        ):
            bucket[key] = round(bucket[key], 2)
        bucket["supp_row_summary"] = " | ".join(bucket["supp_row_summary"])

    return by_employee


def classify_diff(row: dict[str, Any]) -> str:
    if row["offline_performance_base"] == 0:
        return "线下无基数"
    if row["platform_performance_base"] == 0:
        return "平台无结果"
    if abs(row["diff"]) <= 0.05:
        return "一致"
    paths = row.get("calculation_paths", "")
    if "区长固定基数路径" in paths:
        return "区长固定基数路径"
    if "96工时制固定基数覆盖路径" in paths:
        return "96工时制固定基数覆盖路径"
    if "96工时制自动基数路径" in paths:
        return "96工时制自动基数路径"
    if "线下固定基数覆盖路径" in paths:
        return "线下固定基数覆盖路径"
    if row.get("has_segments"):
        return "调薪/转正拆分路径"
    if row.get("supp_unmatched_count", 0) > 0:
        return "补充假勤未匹配考勤路径"
    if row.get("supp_pending_count", 0) > 0:
        return "补充假勤待确认路径"
    if row.get("sick_settlement_hours", 0) > 0:
        return "病假清算路径"
    if row.get("sick_hours", 0) or row.get("annual_hours", 0):
        return "病假/年假路径"
    return "普通金额差异"


def compare_results(
    platform: dict[str, dict[str, Any]],
    offline: dict[str, dict[str, Any]],
    supplemental: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    supplemental = supplemental or {}
    for employee_id in sorted(set(platform) | set(offline) | set(supplemental)):
        p = platform.get(employee_id, {})
        o = offline.get(employee_id, {})
        s = supplemental.get(employee_id, {})
        platform_base = round(as_float(p.get("platform_performance_base")), 2)
        offline_base = round(as_float(o.get("offline_performance_base")), 2)
        diff = round(platform_base - offline_base, 2)
        row = {
            "employee_id": employee_id,
            "name": p.get("name") or o.get("name") or "",
            "platform_performance_base": platform_base,
            "offline_performance_base": offline_base,
            "diff": diff,
            "abs_diff": abs(diff),
            "platform_bonus": p.get("platform_bonus", 0),
            "row_count": p.get("row_count", 0),
            "raw_employee_ids": p.get("raw_employee_ids", ""),
            "calculation_paths": p.get("calculation_paths", ""),
            "base_override_types": p.get("base_override_types", ""),
            "hourly_rates": p.get("hourly_rates", ""),
            "single_hourly_rate": p.get("single_hourly_rate", 0),
            "has_segments": bool(p.get("has_segments")),
            "sick_hours": p.get("sick_hours", 0),
            "sick_settlement_hours": p.get("sick_settlement_hours", 0),
            "annual_hours": p.get("annual_hours", 0),
            "supp_total_rows": s.get("supp_total_rows", 0),
            "supp_confirmed_count": s.get("supp_confirmed_count", 0),
            "supp_pending_count": s.get("supp_pending_count", 0),
            "supp_excluded_count": s.get("supp_excluded_count", 0),
            "supp_unmatched_count": s.get("supp_unmatched_count", 0),
            "supp_include_hours": s.get("supp_include_hours", 0),
            "supp_sick_hours": s.get("supp_sick_hours", 0),
            "supp_annual_hours": s.get("supp_annual_hours", 0),
            "supp_sick_settlement_hours": s.get("supp_sick_settlement_hours", 0),
            "supp_row_summary": s.get("supp_row_summary", ""),
        }
        single_hourly_rate = as_float(row.get("single_hourly_rate"))
        has_supplemental_gap = (
            row.get("supp_pending_count", 0) > 0
            or row.get("supp_unmatched_count", 0) > 0
        )
        row["estimated_missing_hours"] = (
            round(abs(diff) / single_hourly_rate, 2)
            if diff < 0 and single_hourly_rate and has_supplemental_gap
            else ""
        )
        row["reason"] = classify_diff(row)
        rows.append(row)

    matched = [row for row in rows if abs(row["diff"]) <= 0.05 and row["offline_performance_base"]]
    mismatched = [row for row in rows if abs(row["diff"]) > 0.05]
    total_platform = round(sum(row["platform_performance_base"] for row in rows), 2)
    total_offline = round(sum(row["offline_performance_base"] for row in rows), 2)
    by_reason: dict[str, int] = defaultdict(int)
    by_reason_abs_diff: dict[str, float] = defaultdict(float)
    for row in mismatched:
        by_reason[row["reason"]] += 1
        by_reason_abs_diff[row["reason"]] += row["abs_diff"]

    summary = {
        "platform_employee_count": len(platform),
        "offline_employee_count": len(offline),
        "matched_count": len(matched),
        "mismatched_count": len(mismatched),
        "total_platform_performance_base": total_platform,
        "total_offline_performance_base": total_offline,
        "total_diff": round(total_platform - total_offline, 2),
        "by_reason": dict(sorted(by_reason.items())),
        "by_reason_abs_diff": {
            key: round(value, 2) for key, value in sorted(by_reason_abs_diff.items())
        },
    }
    return rows, summary


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def platform_result_export_fields() -> list[str]:
    return [
        "employee_id",
        "source_employee_id",
        "raw_employee_ids",
        "name",
        "job_type",
        "calculation_path",
        "base_override_type",
        "base_hours",
        "ot15_hours",
        "ot20_hours",
        "sick_hours",
        "sick_settlement_hours",
        "annual_hours",
        "holiday_hours",
        "performance_base",
        "performance_ratio",
        "performance_score",
        "performance_level",
        "performance_coefficient",
        "performance_bonus",
    ]


def run_e2e(args: argparse.Namespace) -> dict[str, Any]:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = PROJECT_ROOT / "output" / f"fbu_real_e2e_run_{timestamp}"
    runs_dir = output_dir / "runs"
    exports_dir = output_dir / "exports"
    source_dir = output_dir / "sources"
    output_dir.mkdir(parents=True, exist_ok=True)

    files = DEFAULT_FILES.copy()
    for key in files:
        override = getattr(args, key, None)
        if override:
            files[key] = Path(override)
        ensure_file(files[key])

    prepared = {
        "roster": copy_or_decrypt(files["roster"], source_dir / files["roster"].name),
        "attendance": copy_or_decrypt(files["attendance"], source_dir / files["attendance"].name),
        "salary": copy_or_decrypt(files["salary"], source_dir / files["salary"].name),
        "supplemental_leave": copy_or_decrypt(
            files["supplemental_leave"],
            source_dir / files["supplemental_leave"].name,
        ),
        "performance_book": copy_or_decrypt(
            files["performance_book"],
            source_dir / "decrypted_performance_book.xlsx",
            "FBU2026",
        ),
        "offline_summary": copy_or_decrypt(
            files["offline_summary"],
            source_dir / "decrypted_offline_summary.xlsx",
            "202604",
        ),
    }
    previous_attendance = getattr(args, "previous_attendance", None)
    if previous_attendance:
        previous_attendance_path = copy_or_decrypt(
            Path(previous_attendance),
            source_dir / Path(previous_attendance).name,
        )
        prepared["previous_attendance"] = previous_attendance_path

    performance_source = extract_sheet(
        prepared["performance_book"],
        "4月绩效报表",
        source_dir / "performance_source_4月绩效报表.xlsx",
        "4月绩效报表",
    )
    adjustments_source = extract_sheet(
        prepared["offline_summary"],
        "调薪拆分",
        source_dir / "adjustments_from_offline_summary.xlsx",
        "调薪拆分",
    )
    app_module.FBU_PERFORMANCE_RUNS_DIR = runs_dir
    app_module.EXPORT_DIR = exports_dir
    app_module.fbu_run_manager = FBURunManager(str(runs_dir))
    app_module.fbu_roster_store = FBURosterStore(str(runs_dir))
    client = TestClient(app_module.app)

    roster_payload = post_file(client, "/api/fbu-performance/roster", prepared["roster"])
    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": args.calc_month})
    if run_response.status_code != 200:
        raise RuntimeError(f"create run failed: {run_response.status_code} {run_response.text}")
    run_id = run_response.json()["run_id"]

    attendance_payload = post_file(
        client,
        "/api/fbu-performance/import-attendance",
        prepared["attendance"],
        {"calc_month": args.calc_month, "run_id": run_id},
        (
            {"previous_attendance": prepared["previous_attendance"]}
            if prepared.get("previous_attendance")
            else None
        ),
    )
    salary_payload = post_file(
        client,
        "/api/fbu-performance/import-salary",
        prepared["salary"],
        {"run_id": run_id},
    )
    performance_payload = post_file(
        client,
        "/api/fbu-performance/import-performance",
        performance_source,
        {"run_id": run_id},
    )
    adjustments_payload = post_file(
        client,
        "/api/fbu-performance/import-adjustments",
        adjustments_source,
        {"run_id": run_id},
    )
    supplemental_payload = post_file(
        client,
        "/api/fbu-performance/import-supplemental-leave",
        prepared["supplemental_leave"],
        {"run_id": run_id},
    )
    work_hour_rule_payload = build_work_hour_rule_payload()
    base_overrides_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/rule-lists/confirm",
        json=work_hour_rule_payload,
    )
    if base_overrides_response.status_code != 200:
        raise RuntimeError(
            "confirm rule lists failed: "
            f"{base_overrides_response.status_code} {base_overrides_response.text}"
        )
    base_overrides_payload = base_overrides_response.json()

    calculate_response = client.post(f"/api/fbu-performance/calculate/{run_id}")
    if calculate_response.status_code != 200:
        raise RuntimeError(f"calculate failed: {calculate_response.status_code} {calculate_response.text}")

    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    results_payload = client.get(f"/api/fbu-performance/runs/{run_id}/results")
    if results_payload.status_code != 200:
        raise RuntimeError(f"results failed: {results_payload.status_code} {results_payload.text}")
    results = results_payload.json()["results"]

    platform = aggregate_platform_results(results)
    offline = read_final_performance_base(prepared["performance_book"])
    comparison_platform, platform_only_employee_ids = filter_platform_to_final_target(platform, offline)
    supplemental = aggregate_supplemental_leave(supplemental_payload.get("preview", {}))
    comparison_supplemental = {
        employee_id: row for employee_id, row in supplemental.items()
        if employee_id in offline
    }
    comparison_rows, comparison_summary = compare_results(comparison_platform, offline, comparison_supplemental)
    comparison_summary["platform_all_employee_count"] = len(platform)
    comparison_summary["platform_only_employee_count"] = len(platform_only_employee_ids)
    top_diffs = sorted(comparison_rows, key=lambda row: row["abs_diff"], reverse=True)

    result_fields = platform_result_export_fields()
    comparison_fields = [
        "employee_id",
        "name",
        "platform_performance_base",
        "offline_performance_base",
        "diff",
        "abs_diff",
        "reason",
        "row_count",
        "raw_employee_ids",
        "calculation_paths",
        "base_override_types",
        "hourly_rates",
        "estimated_missing_hours",
        "has_segments",
        "sick_hours",
        "sick_settlement_hours",
        "annual_hours",
        "supp_total_rows",
        "supp_confirmed_count",
        "supp_pending_count",
        "supp_excluded_count",
        "supp_unmatched_count",
        "supp_include_hours",
        "supp_sick_hours",
        "supp_annual_hours",
        "supp_sick_settlement_hours",
        "supp_row_summary",
        "platform_bonus",
    ]

    write_csv(output_dir / "platform_results.csv", results, result_fields)
    write_csv(output_dir / "offline_base.csv", list(offline.values()), ["employee_id", "name", "offline_performance_base"])
    write_csv(output_dir / "comparison_all.csv", comparison_rows, comparison_fields)
    write_csv(output_dir / "comparison_top_diffs.csv", top_diffs[:100], comparison_fields)

    artifact = {
        "run_id": run_id,
        "calc_month": args.calc_month,
        "output_dir": str(output_dir),
        "inputs": {key: str(value) for key, value in files.items()},
        "prepared_sources": {key: str(value) for key, value in prepared.items()},
        "work_hour_rule_payload": work_hour_rule_payload,
        "api_summaries": {
            "roster": roster_payload.get("roster", {}),
            "attendance": attendance_payload.get("preview", {}).get("summary", {}),
            "salary": salary_payload.get("preview", {}).get("summary", {}),
            "performance": performance_payload.get("preview", {}).get("summary", {}),
            "adjustments": adjustments_payload.get("preview", {}).get("summary", {}),
            "supplemental_leave": supplemental_payload.get("preview", {}).get("summary", {}),
            "base_overrides": base_overrides_payload.get("preview", {}).get("summary", {}),
            "calculate": calculate_response.json(),
            "run": {
                "status": run_detail.get("status"),
                "current_step": run_detail.get("current_step"),
                "total_employees": run_detail.get("total_employees"),
                "total_bonus": run_detail.get("total_bonus"),
            },
        },
        "comparison": comparison_summary,
        "platform_only_employee_ids": platform_only_employee_ids[:100],
        "top_diffs": top_diffs[:20],
    }
    with (output_dir / "comparison_summary.json").open("w", encoding="utf-8") as fh:
        json.dump(artifact, fh, ensure_ascii=False, indent=2)

    return artifact


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--calc-month", default="2026-04")
    parser.add_argument("--previous-attendance", help="Optional previous-month attendance report for cross-period context.")
    for key in DEFAULT_FILES:
        parser.add_argument(f"--{key.replace('_', '-')}", dest=key)
    return parser.parse_args()


def main() -> None:
    artifact = run_e2e(parse_args())
    comparison = artifact["comparison"]
    print(json.dumps({
        "run_id": artifact["run_id"],
        "output_dir": artifact["output_dir"],
        "platform_employee_count": comparison["platform_employee_count"],
        "offline_employee_count": comparison["offline_employee_count"],
        "matched_count": comparison["matched_count"],
        "mismatched_count": comparison["mismatched_count"],
        "total_platform_performance_base": comparison["total_platform_performance_base"],
        "total_offline_performance_base": comparison["total_offline_performance_base"],
        "total_diff": comparison["total_diff"],
        "by_reason": comparison["by_reason"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
