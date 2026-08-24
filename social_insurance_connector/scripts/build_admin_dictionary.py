from __future__ import annotations

import json
from pathlib import Path
import sys

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: build_admin_dictionary.py TEMPLATE.xlsx OUTPUT.json")
    source = Path(sys.argv[1])
    destination = Path(sys.argv[2])
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook["行政区划字典"]
    sheet.reset_dimensions()
    values = [
        str(value).strip()
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None and str(value).strip()
    ]
    destination.write_text(
        json.dumps(values, ensure_ascii=False, separators=(",", ":")) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
