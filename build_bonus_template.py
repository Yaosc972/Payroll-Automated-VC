from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = Path(os.environ.get("BONUS_SOURCE_XLSX", ROOT / "outputs" / "source.xlsx"))
OUTPUT_DIR = ROOT / "outputs"
RULE_XLSX = OUTPUT_DIR / "招聘奖金核算_规则库.xlsx"
IMPORT_TEMPLATE_XLSX = OUTPUT_DIR / "招聘奖金核算_月度导入模板.xlsx"

MAX_ROWS = 2000


THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
RULE_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
IMPORT_FILL = PatternFill("solid", fgColor="F4B183")
CALC_FILL = PatternFill("solid", fgColor="9DC3E6")
MANUAL_FILL = PatternFill("solid", fgColor="FFD966")
DATE_FORMAT = "yyyy/m/d"


def normalize(v):
    return "" if v is None else v


def style_range(ws, min_row, max_row, min_col, max_col, header_rows=1, fill=HEADER_FILL):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row < min_row + header_rows:
                cell.fill = fill
                cell.font = Font(color="FFFFFF", bold=True)


def add_table(ws, name, ref):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def color_header_cells(ws, cols, fill, font_color="000000"):
    for col in cols:
        cell = ws.cell(1, col)
        cell.fill = fill
        cell.font = Font(color=font_color, bold=True)


def apply_date_format(ws, cols, start_row=2, end_row=None):
    max_row = end_row or ws.max_row
    for col in cols:
        for row in range(start_row, max_row + 1):
            ws.cell(row, col).number_format = DATE_FORMAT


def yyyymm_formula(date_ref):
    return f'IF({date_ref}="","",TEXT({date_ref},"yyyymm")*1)'


def edate_yyyymm_formula(date_ref, months):
    return f'IF({date_ref}="","",TEXT(EDATE({date_ref},{months}),"yyyymm")*1)'


def period_override_formula(row: int, override_col: str, fallback: str) -> str:
    override_ref = f"导入_月度数据!{override_col}{row}"
    return f'IF({override_ref}<>"",{override_ref},{fallback})'


def amount_override_formula(row: int, override_col: str, fallback: str) -> str:
    override_ref = f"导入_月度数据!{override_col}{row}"
    return f'IF({override_ref}<>"",{override_ref},{fallback})'


def adjusted_recruit_base_formula(row: int) -> str:
    return (
        f'IF(U{row}=0,0,MAX(U{row}*T{row}*30%,'
        f'MIN(U{row}*T{row}*130%,'
        f'U{row}*T{row}*(1-ROUNDDOWN(S{row}/10,0)*10%))))'
    )


def recruit_first_ratio_formula(row: int) -> str:
    return (
        f'IF(OR(AND(P{row}="国内",OR(G{row}="A类",G{row}="B类")),'
        f'AND(P{row}="海外",OR(G{row}="A类",G{row}="B类"))),'
        f'IF(P{row}="国内",20%,30%),'
        f'IF(OR(G{row}="C类",G{row}="C1类"),50%,0))'
    )


def recruit_probation_ratio_formula(row: int) -> str:
    return f'IF(P{row}="国内",IF(OR(G{row}="A类",G{row}="B类"),80%,IF(G{row}="C类",50%,0)),0)'


def recruit_three_ratio_formula(row: int) -> str:
    return f'IF(P{row}="海外",IF(G{row}="A类",30%,IF(G{row}="B类",70%,IF(G{row}="C1类",50%,0))),0)'


def recruit_six_ratio_formula(row: int) -> str:
    return f'IF(AND(P{row}="海外",G{row}="A类"),40%,0)'


def read_source_rules():
    wb = load_workbook(SOURCE_XLSX, data_only=True, read_only=False)
    ws = wb["参数设置"]

    grades = [ws.cell(4, c).value for c in range(6, 27)]
    cycle_rules = []
    for r in range(5, 11):
        region = ws.cell(r, 3).value
        category = ws.cell(r, 4).value
        if not region or not category:
            continue
        for idx, grade in enumerate(grades, start=6):
            value = ws.cell(r, idx).value
            if grade and value not in (None, "-", ""):
                cycle_rules.append([region, category, grade, value, "Excel参数设置!C5:Z10"])

    recruit_rules = []
    for r in range(19, 25):
        region = ws.cell(r, 3).value
        category = ws.cell(r, 4).value
        currency = ws.cell(r, 27).value
        if not region or not category:
            continue
        for idx, grade in enumerate(grades, start=6):
            value = ws.cell(r, idx).value
            if grade and value not in (None, "-", ""):
                recruit_rules.append([region, category, grade, value, currency, "Excel参数设置!C19:AA24"])

    referral_rules = []
    for r in range(37, 46):
        region_type = ws.cell(r, 3).value
        category = ws.cell(r, 4).value
        currency = ws.cell(r, 27).value
        if not region_type or not category:
            continue
        for idx, grade in enumerate(grades, start=6):
            value = ws.cell(r, idx).value
            if grade and value not in (None, "-", ""):
                referral_rules.append([region_type, category, grade, value, currency, "集团制度/Excel参数设置"])

    channel_rules = []
    for r in range(5, 14):
        channel = ws.cell(r, 28).value
        ratio = ws.cell(r, 29).value
        if channel not in (None, ""):
            channel_rules.append([channel, ratio, "Excel参数设置!AB5:AC13"])

    return grades, cycle_rules, recruit_rules, referral_rules, channel_rules


def add_instruction_sheet(wb):
    ws = wb.active
    ws.title = "使用说明"
    rows = [
        ["招聘奖金核算月度导入模板", ""],
        ["使用流程", "1. 在“导入_月度数据”粘贴本月 HR 数据。"],
        ["", "2. 保存后上传到招聘奖金核算平台。"],
        ["", "3. 在平台下载初算结果；如有待确认节点，再按平台流程回传确认表。"],
        ["规则口径", "本模板只承载月度导入数据；平台按内部当前规则库计算招聘奖金和内推奖金。"],
        ["颜色说明", "橙色表头=需从系统导入/人工填写；蓝色表头=模板辅助计算字段。"],
        ["重要口径", "月度导入模板不包含规则表和结果表，旧规则历史结转不进入当前主计算流程。"],
        ["需人工提供", "员工姓名/工号、入职/转正/离职日期、职位职级、招聘渠道、招聘负责人、协助招聘人、推荐人、汇报关系、特殊地区标记。"],
        ["特殊规则优先级", "FBU德国、FBU捷克优先于集团统一内推规则；未命中特殊地区时按集团规则计算。"],
        ["容量说明", "导入区和计算区默认覆盖 2000 行数据；如月度数据超过 2000 行，需先扩展 MAX_ROWS 后重新生成。"],
    ]
    for r, row in enumerate(rows, start=1):
        ws.append(row)
        ws.row_dimensions[r].height = 28
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["B1"].fill = HEADER_FILL
    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=2):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    set_widths(ws, {"A": 20, "B": 120})


def add_import_sheet(wb):
    ws = wb.create_sheet("导入_月度数据")
    headers = [
        "核算月份", "姓名", "工号", "证件号", "入职日期", "人员状态", "工作地", "最后工作日", "职级",
        "二级组织", "三级组织", "四级组织", "五级组织", "职位", "职族", "职类", "直线经理",
        "直线经理工号", "标签分类", "职族分类", "ABC类别", "招聘负责人工号", "招聘负责人姓名",
        "协助招聘人工号", "协助招聘人姓名", "招聘渠道", "海外offer流程编号", "海外招聘需求编号",
        "招聘启动日期", "候选人入职时间", "转正日期", "周期剔除天数", "奖金地区类型",
        "推荐人姓名", "推荐人工号", "推荐人职级", "推荐人人员状态", "推荐人职位",
        "直接上级工号", "间接上级1工号", "间接上级2工号", "间接上级3工号", "间接上级4工号",
        "特殊地区规则", "人工备注",
        "招聘负责人人员状态", "招聘负责人最后工作日", "协助招聘人人员状态", "协助招聘人最后工作日",
        "推荐人最后工作日",
    ]
    ws.append(headers)
    for row in range(2, MAX_ROWS + 2):
        ws.cell(row, 1).value = "=IF(B{0}=\"\",\"\",$A$2)".format(row)
        ws.cell(row, 19).value = '=IF(G{0}="","",IF(G{0}="中国大陆","国内","海外"))'.format(row)
        ws.cell(row, 33).value = (
            '=IF(G{0}="","",IF(ISNUMBER(SEARCH("中国",G{0})),"国内发展中国家",'
            'IF(SUMPRODUCT(--ISNUMBER(SEARCH({{"德国","捷克","美国","英国","法国","加拿大","日本","韩国","荷兰","意大利","西班牙","葡萄牙","奥地利","瑞士","比利时","卢森堡","丹麦","瑞典","挪威","芬兰","爱尔兰","澳大利亚","新西兰","新加坡","以色列","波兰"}},G{0})))>0,'
            '"海外发达国家","海外发展中国家")))'
        ).format(row)
    add_table(ws, "tbl_import", f"A1:{get_column_letter(len(headers))}{MAX_ROWS + 1}")
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1)
    color_header_cells(ws, range(1, len(headers) + 1), IMPORT_FILL)
    color_header_cells(ws, [19, 33], CALC_FILL)
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for col in ["B", "J", "K", "L", "M", "N", "O", "P", "Q", "W", "Y", "AI", "AS", "AT", "AV"]:
        ws.column_dimensions[col].width = 22
    apply_date_format(ws, [5, 8, 29, 30, 31, 47, 49, 50], 2, MAX_ROWS + 1)

    validations = {
        "F": '"正式,试用,离职,通知期"',
        "AT": '"正式,试用,离职,通知期,在职,正式在职,试用在职"',
        "AV": '"正式,试用,离职,通知期,在职,正式在职,试用在职"',
        "S": '"国内,海外,DBU海外"',
        "T": '"职能类,操作类,技术类,营销类"',
        "U": '"A类,B类,C类,C1类"',
        "Z": '"招聘网站,内推,校招,二次入职,猎头,返聘,体系外转体系内,外包转正式,其他,人员调动"',
        "AG": '"国内发展中国家,海外发达国家,海外发展中国家,FBU德国,FBU捷克"',
        "AR": '"FBU德国,FBU捷克"',
    }
    for col, formula1 in validations.items():
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{MAX_ROWS + 1}")


def add_rule_sheets(wb, cycle_rules, recruit_rules, referral_rules, channel_rules):
    ws = wb.create_sheet("规则_招聘周期")
    ws.append(["标签分类", "ABC类别", "职级", "标准招聘周期天数", "规则来源", "匹配键"])
    for row in cycle_rules:
        ws.append(row + [f"{row[0]}|{row[1]}|{row[2]}"])
    style_range(ws, 1, ws.max_row, 1, 6, fill=RULE_FILL)
    add_table(ws, "tbl_cycle_rules", f"A1:F{ws.max_row}")
    set_widths(ws, {"A": 14, "B": 12, "C": 12, "D": 18, "E": 26, "F": 24})

    ws = wb.create_sheet("规则_招聘奖金")
    ws.append(["标签分类", "ABC类别", "职级", "招聘奖金标准", "币种", "规则来源", "匹配键"])
    for row in recruit_rules:
        ws.append(row + [f"{row[0]}|{row[1]}|{row[2]}"])
    style_range(ws, 1, ws.max_row, 1, 7, fill=RULE_FILL)
    add_table(ws, "tbl_recruit_rules", f"A1:G{ws.max_row}")
    set_widths(ws, {"A": 14, "B": 12, "C": 12, "D": 16, "E": 14, "F": 26, "G": 24})

    ws = wb.create_sheet("规则_渠道系数")
    ws.append(["招聘渠道", "奖金系数", "规则来源"])
    for row in channel_rules:
        ws.append(row)
    style_range(ws, 1, ws.max_row, 1, 3, fill=RULE_FILL)
    add_table(ws, "tbl_channel_rules", f"A1:C{ws.max_row}")
    set_widths(ws, {"A": 22, "B": 12, "C": 26})

    ws = wb.create_sheet("规则_内推奖金")
    ws.append(["制度范围", "地区类型", "ABC类别", "职级", "内推奖金标准", "币种", "入职1月比例", "入职3月比例", "入职6月比例", "转正比例", "规则来源", "匹配键", "备注"])
    for region_type, category, grade, amount, currency, source in referral_rules:
        p1, p3, p6, probation = referral_split(region_type, category, grade)
        ws.append(["集团统一", region_type, category, grade, amount, currency, p1, p3, p6, probation, source, f"集团统一|{region_type}|{category}|{grade}", ""])
    for row in germany_referral_rules():
        ws.append(row)
    for row in czech_referral_rules():
        ws.append(row)
    style_range(ws, 1, ws.max_row, 1, 13, fill=RULE_FILL)
    add_table(ws, "tbl_referral_rules", f"A1:M{ws.max_row}")
    set_widths(ws, {"A": 14, "B": 18, "C": 12, "D": 12, "E": 16, "F": 12, "G": 14, "H": 14, "I": 14, "J": 12, "K": 24, "L": 34, "M": 34})

    ws = wb.create_sheet("规则_资格条件")
    rows = [
        ["规则类型", "判断项", "自动判断字段", "处理方式", "规则来源"],
        ["招聘奖金", "缺招聘启动日期/入职日期/职级/ABC类别/招聘渠道", "导入_月度数据", "进入异常清单，不直接填0", "Excel公式反推"],
        ["招聘奖金", "招聘渠道影响奖金系数", "招聘渠道", "按规则_渠道系数计算", "Excel参数设置"],
        ["内推奖金", "推荐人与被推荐人核算时须同时在职", "人员状态、推荐人人员状态", "任一非在职则异常/不计发", "集团制度/德国制度/捷克制度"],
        ["内推奖金", "推荐人是HRBP/招聘岗或高职级", "推荐人职位、推荐人职级", "进入异常清单，需薪酬确认", "集团制度"],
        ["内推奖金", "被推荐人成为推荐人直接或间接下属", "直接/间接上级工号", "不计发并进入异常", "集团制度/德国制度/捷克制度"],
        ["内推奖金", "近6个月或德国近3个月已有简历记录", "系统暂缺", "需人工确认字段或后续接入招聘系统", "集团制度/德国制度"],
        ["特殊地区", "FBU德国/FBU捷克", "特殊地区规则", "优先覆盖集团统一规则", "用户确认"],
    ]
    for row in rows:
        ws.append(row)
    style_range(ws, 1, ws.max_row, 1, 5, fill=WARN_FILL)
    add_table(ws, "tbl_eligibility_rules", f"A1:E{ws.max_row}")
    set_widths(ws, {"A": 14, "B": 36, "C": 24, "D": 34, "E": 24})

    ws = wb.create_sheet("规则说明")
    rows = [
        ["规则项", "已实现方式", "来源", "备注"],
        ["招聘周期", "按标签分类+ABC类别+职级查标准周期", "202604 Excel 参数设置", ""],
        ["招聘奖金标准", "按标签分类+ABC类别+职级查标准奖金和币种", "202604 Excel 参数设置", ""],
        ["招聘达成调整", "基础奖金×渠道系数后，每超期10天扣10%，提前则加成，最低30%、最高130%", "202604 Excel 公式反推", "原表使用ROUNDDOWN(周期差异/10)"],
        ["招聘发放比例", "有协助招聘人时，招聘负责人按调整后标准70%计算，协助招聘人按30%计算；国内/海外 A/B 首月20%或30%，C/C1首月50%；国内转正发放剩余部分", "202604 Excel 公式反推", ""],
        ["内推集团规则", "按地区类型+ABC类别+职级匹配标准及发放节点", "集团内推制度V2/Excel参数设置", ""],
        ["德国FBU", "特殊地区规则=FBU德国时优先匹配德国规则", "FBU德国区内部推荐激励制度202510-V2", "部分C类P3+发放节点按制度版式需复核，已在备注中标记"],
        ["捷克FBU", "特殊地区规则=FBU捷克时优先匹配捷克规则", "FBU捷克区内推奖金制度", "P1/P1-3按6个月发放，M2-1及以上提示按集团制度人工确认"],
        ["AI辅助", "AI可读取规则说明和异常清单，辅助解释差异和维护规则", "实施方案", "金额仍由公式计算"],
    ]
    for row in rows:
        ws.append(row)
    style_range(ws, 1, ws.max_row, 1, 4, fill=HEADER_FILL)
    add_table(ws, "tbl_rule_notes", f"A1:D{ws.max_row}")
    set_widths(ws, {"A": 22, "B": 54, "C": 36, "D": 54})


def referral_split(region_type, category, grade):
    domestic = region_type == "国内发展中国家"
    high = grade.startswith(("P3", "P4", "M3", "M4", "M5"))
    mid = grade.startswith(("P2", "M2"))
    low = grade.startswith("P1")
    if domestic:
        if category in ("A类", "B类"):
            return 0.30, 0, 0, 0.70
        if category == "C类":
            return (0.30, 0, 0, 0.70) if high else (0.50, 0, 0, 0.50)
    else:
        if category == "A类":
            return 0.30, 0.30, 0.40, 0
        if category == "B类":
            return (0.30, 0.30, 0.40, 0) if high else (0.30, 0.70, 0, 0)
        if category == "C1类":
            return (0.30, 0.70, 0, 0) if high else (0.50, 0.50, 0, 0)
    return 0, 0, 0, 0


def germany_referral_rules():
    rows = []
    specs = [
        ("B类", ["P3-1", "P3-2", "P3-3", "M3-1", "M3-2", "M3-3", "P4-1", "P4-2", "P4-3", "M4-1", "M4-2", "M4-3"], 1050, 0.30, 0.30, 0.40, ""),
        ("B类", ["P2-1", "P2-2", "P2-3", "M2-1", "M2-2", "M2-3"], 400, 0.70, 0, 0, ""),
        ("C类", ["P3-1", "P3-2", "P3-3", "M3-1", "M3-2", "M3-3"], 850, 0.50, 0.50, 0, "制度版式未完全明确，建议薪酬复核"),
        ("C类", ["P2-1", "P2-2", "P2-3", "M2-1", "M2-2", "M2-3"], 300, 0.50, 0.50, 0, ""),
        ("C类", ["P1-3"], 200, 0.50, 0.50, 0, ""),
        ("不占编", ["P1-1", "P1-2"], 100, 0.70, 0.30, 0, "操作工"),
        ("不占编", ["P1-1"], 80, 1.00, 0, 0, "学生工，需人工区分"),
    ]
    for category, grades, amount, p1, p3, p6, note in specs:
        for grade in grades:
            rows.append(["FBU德国", "FBU德国", category, grade, amount, "欧元", p1, p3, p6, 0, "德国制度", f"FBU德国|FBU德国|{category}|{grade}", note])
    return rows


def czech_referral_rules():
    rows = []
    specs = [
        ("不占编", ["P1-1", "P1-2"], 5000, "捷克克朗", "仓库操作员/叉车司机，入职满6个月发放"),
        ("C类", ["P1-2", "P1-3"], 8000, "捷克克朗", "副组长/仓库文员/职能，入职满6个月发放"),
    ]
    for category, grades, amount, currency, note in specs:
        for grade in grades:
            rows.append(["FBU捷克", "FBU捷克", category, grade, amount, currency, 0, 0, 1.00, 0, "捷克制度", f"FBU捷克|FBU捷克|{category}|{grade}", note])
    rows.append(["FBU捷克", "FBU捷克", "M2及以上", "M2-1", "", "", 0, 0, 1.00, 0, "捷克制度", "FBU捷克|FBU捷克|M2及以上|M2-1", "按集团制度金额，满6个月发放；需人工确认对应ABC类别"])
    return rows


def add_detail_sheet(wb):
    ws = wb.create_sheet("输出_招聘奖金明细")
    headers = [
        "唯一验证", "姓名", "工号", "人员状态", "工作地", "职级", "ABC类别", "招聘渠道",
        "招聘负责人工号", "招聘负责人姓名", "协助招聘人工号", "协助招聘人姓名",
        "招聘启动日期", "候选人入职时间", "转正日期", "标签分类", "标准入职周期", "实际入职周期",
        "入职周期差异", "渠道系数", "招聘奖金标准", "实际发放标准", "招聘人入职1月奖金",
        "招聘人入职1月周期", "招聘人转正奖金", "招聘人转正周期", "协助人入职1月奖金",
        "协助人入职1月周期", "协助人转正奖金", "协助人转正周期", "币种", "推荐人姓名",
        "推荐人工号", "推荐人状态", "推荐人是否上级", "内推规则范围", "内推奖金标准",
        "内推入职1月奖金", "内推入职1月周期", "内推入职3月奖金", "内推入职3月周期",
        "内推入职6月奖金", "内推入职6月周期", "内推转正奖金", "内推转正周期", "异常提示"
    ]
    ws.append(headers)
    for r in range(2, MAX_ROWS + 2):
        src = r
        ws.cell(r, 1).value = f'=IF(导入_月度数据!C{src}="","",IF(COUNTIF(导入_月度数据!C:C,导入_月度数据!C{src})>1,"重复","唯一"))'
        direct_refs = {
            2: "B", 3: "C", 4: "F", 5: "G", 6: "I", 7: "U", 8: "Z", 9: "V", 10: "W",
            11: "X", 12: "Y", 13: "AC", 14: "AD", 15: "AE", 16: "S", 32: "AH", 33: "AI", 34: "AK",
        }
        for col, src_col in direct_refs.items():
            ws.cell(r, col).value = f'=IF(导入_月度数据!{src_col}{src}="","",导入_月度数据!{src_col}{src})'
        ws.cell(r, 17).value = f'=IFERROR(INDEX(规则_招聘周期!$D:$D,MATCH(P{r}&"|"&G{r}&"|"&F{r},规则_招聘周期!$F:$F,0)),"")'
        ws.cell(r, 18).value = f'=IF(OR(M{r}="",N{r}=""),"",N{r}-M{r}-IFERROR(导入_月度数据!AF{src},0))'
        ws.cell(r, 19).value = f'=IF(OR(Q{r}="",R{r}=""),"",R{r}-Q{r})'
        ws.cell(r, 20).value = f'=IFERROR(INDEX(规则_渠道系数!$B:$B,MATCH(H{r},规则_渠道系数!$A:$A,0)),1)'
        ws.cell(r, 21).value = f'=IFERROR(INDEX(规则_招聘奖金!$D:$D,MATCH(P{r}&"|"&G{r}&"|"&F{r},规则_招聘奖金!$G:$G,0)),0)'
        base = adjusted_recruit_base_formula(r)
        first_ratio = recruit_first_ratio_formula(r)
        probation_ratio = recruit_probation_ratio_formula(r)
        ws.cell(r, 22).value = f'={base}*IF(K{r}="",1,70%)'
        ws.cell(r, 23).value = f'=IF(V{r}=0,0,V{r}*{first_ratio})'
        ws.cell(r, 24).value = f'={edate_yyyymm_formula(f"N{r}", 1)}'
        ws.cell(r, 25).value = f'=IF(V{r}=0,0,V{r}*{probation_ratio})'
        ws.cell(r, 26).value = f'=IF(O{r}="","未转正，待发放",{yyyymm_formula(f"O{r}")})'
        ws.cell(r, 27).value = f'=IF(K{r}="",0,{base}*30%*{first_ratio})'
        ws.cell(r, 28).value = f'=IF(K{r}="","",X{r})'
        ws.cell(r, 29).value = f'=IF(K{r}="",0,{base}*30%*{probation_ratio})'
        ws.cell(r, 30).value = f'=IF(K{r}="","",Z{r})'
        ws.cell(r, 31).value = f'=IFERROR(INDEX(规则_招聘奖金!$E:$E,MATCH(P{r}&"|"&G{r}&"|"&F{r},规则_招聘奖金!$G:$G,0)),"")'
        ws.cell(r, 35).value = f'=IF(OR(AG{r}="-",AG{r}=""),"否",IF(OR(AG{r}=导入_月度数据!AM{src},AG{r}=导入_月度数据!AN{src},AG{r}=导入_月度数据!AO{src},AG{r}=导入_月度数据!AP{src},AG{r}=导入_月度数据!AQ{src}),"是","否"))'
        ws.cell(r, 36).value = f'=IF(导入_月度数据!AR{src}<>"",导入_月度数据!AR{src},"集团统一")'
        ws.cell(r, 37).value = f'=IF(OR(AG{r}="",AG{r}="-",AI{r}="是"),0,IFERROR(INDEX(规则_内推奖金!$E:$E,MATCH(AJ{r}&"|"&IF(AJ{r}="集团统一",导入_月度数据!AG{src},AJ{r})&"|"&G{r}&"|"&F{r},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 38).value = f'=IF(AK{r}=0,0,AK{r}*IFERROR(INDEX(规则_内推奖金!$G:$G,MATCH(AJ{r}&"|"&IF(AJ{r}="集团统一",导入_月度数据!AG{src},AJ{r})&"|"&G{r}&"|"&F{r},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 39).value = f'=IF(AL{r}=0,"",{edate_yyyymm_formula(f"N{r}", 1)})'
        ws.cell(r, 40).value = f'=IF(AK{r}=0,0,AK{r}*IFERROR(INDEX(规则_内推奖金!$H:$H,MATCH(AJ{r}&"|"&IF(AJ{r}="集团统一",导入_月度数据!AG{src},AJ{r})&"|"&G{r}&"|"&F{r},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 41).value = f'=IF(AN{r}=0,"",{edate_yyyymm_formula(f"N{r}", 3)})'
        ws.cell(r, 42).value = f'=IF(AK{r}=0,0,AK{r}*IFERROR(INDEX(规则_内推奖金!$I:$I,MATCH(AJ{r}&"|"&IF(AJ{r}="集团统一",导入_月度数据!AG{src},AJ{r})&"|"&G{r}&"|"&F{r},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 43).value = f'=IF(AP{r}=0,"",{edate_yyyymm_formula(f"N{r}", 6)})'
        ws.cell(r, 44).value = f'=IF(AK{r}=0,0,AK{r}*IFERROR(INDEX(规则_内推奖金!$J:$J,MATCH(AJ{r}&"|"&IF(AJ{r}="集团统一",导入_月度数据!AG{src},AJ{r})&"|"&G{r}&"|"&F{r},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 45).value = f'=IF(AR{r}=0,"",IF(O{r}="","未转正，待发放",{yyyymm_formula(f"O{r}")}))'
        ws.cell(r, 46).value = (
            f'=IF(C{r}="","",TEXTJOIN("；",TRUE,'
            f'IF(A{r}="重复","工号重复",""),'
            f'IF(OR(F{r}="",G{r}="",H{r}="",M{r}="",N{r}=""),"招聘奖金缺关键字段",""),'
            f'IF(Q{r}="","未匹配招聘周期",""),'
            f'IF(U{r}=0,"未匹配招聘奖金标准",""),'
            f'IF(AND(AG{r}<>"",AG{r}<>"-",AI{r}<>"是",AK{r}=0),"未匹配内推规则或不满足资格",""),'
            f'IF(AI{r}="是","推荐人为直接/间接上级，不计内推",""),'
        f'IF(FALSE,"推荐人非正式在职，需确认","")))'
        )
    add_table(ws, "tbl_detail", f"A1:AT{MAX_ROWS + 1}")
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1)
    imported_cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 32, 33, 34]
    color_header_cells(ws, imported_cols, IMPORT_FILL)
    color_header_cells(ws, [c for c in range(1, len(headers) + 1) if c not in imported_cols], CALC_FILL)
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    for col in ["B", "J", "L", "AF", "AT"]:
        ws.column_dimensions[col].width = 24


def add_summary_sheets(wb):
    ws = wb.create_sheet("输出_招聘奖金汇总")
    headers = ["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "转正奖金", "合计发放"]
    ws.append(headers)
    summary_rows = MAX_ROWS * 2
    for r in range(2, summary_rows + 2):
        if r <= MAX_ROWS + 1:
            src = r
            id_col = "V"
            name_col = "W"
            role = "招聘负责人"
        else:
            src = r - MAX_ROWS
            id_col = "X"
            name_col = "Y"
            role = "协助招聘人"
        ws.cell(r, 1).value = f'=IF(OR(导入_月度数据!{id_col}{src}="",导入_月度数据!{id_col}{src}="-",COUNTIF(导入_月度数据!${id_col}$2:{id_col}{src},导入_月度数据!{id_col}{src})>1),"",导入_月度数据!{id_col}{src})'
        ws.cell(r, 2).value = f'=IF(A{r}="","",导入_月度数据!{name_col}{src})'
        ws.cell(r, 3).value = '="招聘负责人"'
        ws.cell(r, 3).value = f'="{role}"'
        ws.cell(r, 4).value = f'=IF(A{r}="","",INDEX(输出_招聘奖金明细!AE:AE,MATCH(A{r},输出_招聘奖金明细!I:I,0)))'
        ws.cell(r, 5).value = '=导入_月度数据!$A$2'
        ws.cell(r, 6).value = f'=IF(A{r}="","",SUMIFS(输出_招聘奖金明细!W:W,输出_招聘奖金明细!I:I,A{r},输出_招聘奖金明细!X:X,E{r})+SUMIFS(输出_招聘奖金明细!AA:AA,输出_招聘奖金明细!K:K,A{r},输出_招聘奖金明细!AB:AB,E{r}))'
        ws.cell(r, 7).value = f'=IF(A{r}="","",SUMIFS(输出_招聘奖金明细!Y:Y,输出_招聘奖金明细!I:I,A{r},输出_招聘奖金明细!Z:Z,E{r})+SUMIFS(输出_招聘奖金明细!AC:AC,输出_招聘奖金明细!K:K,A{r},输出_招聘奖金明细!AD:AD,E{r}))'
        ws.cell(r, 8).value = f'=SUM(F{r}:G{r})'
    style_range(ws, 1, summary_rows + 1, 1, len(headers), header_rows=1)
    color_header_cells(ws, range(1, len(headers) + 1), CALC_FILL)
    add_table(ws, "tbl_recruit_summary", f"A1:H{summary_rows + 1}")
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 16, "B": 18, "C": 14, "D": 12, "E": 12, "F": 16, "G": 16, "H": 16})

    ws = wb.create_sheet("输出_内推奖金汇总")
    headers = ["内推人员工号", "内推人员姓名", "人员状态", "币种", "核算月份", "本月发放"]
    ws.append(headers)
    for r in range(2, MAX_ROWS + 2):
        src = r
        ws.cell(r, 1).value = f'=IF(OR(导入_月度数据!AI{src}="",导入_月度数据!AI{src}="-",COUNTIF(导入_月度数据!$AI$2:AI{src},导入_月度数据!AI{src})>1),"",导入_月度数据!AI{src})'
        ws.cell(r, 2).value = f'=IF(A{r}="","",导入_月度数据!AH{src})'
        ws.cell(r, 3).value = f'=IF(A{r}="","",导入_月度数据!AK{src})'
        ws.cell(r, 4).value = f'=IF(A{r}="","",INDEX(规则_内推奖金!F:F,MATCH(INDEX(输出_招聘奖金明细!AJ:AJ,MATCH(A{r},输出_招聘奖金明细!AG:AG,0))&"|"&IF(INDEX(输出_招聘奖金明细!AJ:AJ,MATCH(A{r},输出_招聘奖金明细!AG:AG,0))="集团统一",INDEX(导入_月度数据!AG:AG,MATCH(A{r},导入_月度数据!AI:AI,0)),INDEX(输出_招聘奖金明细!AJ:AJ,MATCH(A{r},输出_招聘奖金明细!AG:AG,0)))&"|"&INDEX(输出_招聘奖金明细!G:G,MATCH(A{r},输出_招聘奖金明细!AG:AG,0))&"|"&INDEX(输出_招聘奖金明细!F:F,MATCH(A{r},输出_招聘奖金明细!AG:AG,0)),规则_内推奖金!L:L,0)))'
        ws.cell(r, 5).value = '=导入_月度数据!$A$2'
        ws.cell(r, 6).value = f'=IF(A{r}="","",SUMIFS(输出_招聘奖金明细!AL:AL,输出_招聘奖金明细!AG:AG,A{r},输出_招聘奖金明细!AM:AM,E{r})+SUMIFS(输出_招聘奖金明细!AN:AN,输出_招聘奖金明细!AG:AG,A{r},输出_招聘奖金明细!AO:AO,E{r})+SUMIFS(输出_招聘奖金明细!AP:AP,输出_招聘奖金明细!AG:AG,A{r},输出_招聘奖金明细!AQ:AQ,E{r})+SUMIFS(输出_招聘奖金明细!AR:AR,输出_招聘奖金明细!AG:AG,A{r},输出_招聘奖金明细!AS:AS,E{r}))'
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1)
    color_header_cells(ws, range(1, len(headers) + 1), CALC_FILL)
    add_table(ws, "tbl_referral_summary", f"A1:F{MAX_ROWS + 1}")
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 18, "B": 18, "C": 14, "D": 12, "E": 12, "F": 16})

    ws = wb.create_sheet("输出_异常清单")
    headers = ["行号", "姓名", "工号", "异常提示", "人工处理意见"]
    ws.append(headers)
    for r in range(2, MAX_ROWS + 2):
        src = r
        ws.cell(r, 1).value = f'=IF(输出_招聘奖金明细!AT{src}="","",ROW(导入_月度数据!A{src}))'
        ws.cell(r, 2).value = f'=IF(A{r}="","",导入_月度数据!B{src})'
        ws.cell(r, 3).value = f'=IF(A{r}="","",导入_月度数据!C{src})'
        ws.cell(r, 4).value = f'=IF(A{r}="","",输出_招聘奖金明细!AT{src})'
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1, fill=WARN_FILL)
    color_header_cells(ws, [1, 2, 3, 4], CALC_FILL)
    color_header_cells(ws, [5], MANUAL_FILL)
    add_table(ws, "tbl_exceptions", f"A1:E{MAX_ROWS + 1}")
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 10, "B": 18, "C": 16, "D": 80, "E": 28})


def add_detail_sheet_v2(wb):
    ws = wb.create_sheet("输出_招聘奖金明细")
    headers = [
        "唯一验证", "姓名", "工号", "人员状态", "工作地", "职级", "ABC类别", "招聘渠道",
        "招聘负责人工号", "招聘负责人姓名", "协助招聘人工号", "协助招聘人姓名",
        "招聘启动日期", "候选人入职时间", "转正日期", "标签分类", "标准入职周期", "实际入职周期",
        "入职周期差异", "渠道系数", "招聘奖金标准", "招聘人实际发放标准",
        "招聘人入职1月奖金", "招聘人入职1月周期", "招聘人入职3月奖金", "招聘人入职3月周期",
        "招聘人入职6月奖金", "招聘人入职6月周期", "招聘人转正奖金", "招聘人转正周期",
        "协助人实际发放标准", "协助人入职1月奖金", "协助人入职1月周期", "协助人入职3月奖金",
        "协助人入职3月周期", "协助人入职6月奖金", "协助人入职6月周期", "协助人转正奖金",
        "协助人转正周期", "币种", "推荐人姓名", "推荐人工号", "推荐人状态", "推荐人是否上级",
        "内推规则范围", "内推奖金标准", "内推入职1月奖金", "内推入职1月周期", "内推入职3月奖金",
        "内推入职3月周期", "内推入职6月奖金", "内推入职6月周期", "内推转正奖金", "内推转正周期",
        "异常提示"
    ]
    ws.append(headers)
    for r in range(2, MAX_ROWS + 2):
        src = r
        ws.cell(r, 1).value = f'=IF(导入_月度数据!C{src}="","",IF(COUNTIF(导入_月度数据!C:C,导入_月度数据!C{src})>1,"重复","唯一"))'
        for col, src_col in {
            2: "B", 3: "C", 4: "F", 5: "G", 6: "I", 7: "U", 8: "Z", 9: "V", 10: "W",
            11: "X", 12: "Y", 13: "AC", 14: "AD", 15: "AE", 16: "S", 41: "AH", 42: "AI", 43: "AK",
        }.items():
            ws.cell(r, col).value = f'=IF(导入_月度数据!{src_col}{src}="","",导入_月度数据!{src_col}{src})'

        base = adjusted_recruit_base_formula(r)
        first_ratio = recruit_first_ratio_formula(r)
        three_ratio = recruit_three_ratio_formula(r)
        six_ratio = recruit_six_ratio_formula(r)
        probation_ratio = recruit_probation_ratio_formula(r)
        probation_period_fallback = f'IF(O{r}="","未转正，待发放",{yyyymm_formula(f"O{r}")})'

        ws.cell(r, 17).value = f'=IFERROR(INDEX(规则_招聘周期!$D:$D,MATCH(P{r}&"|"&G{r}&"|"&F{r},规则_招聘周期!$F:$F,0)),"")'
        ws.cell(r, 18).value = f'=IF(OR(M{r}="",N{r}=""),"",N{r}-M{r}-IFERROR(导入_月度数据!AF{src},0))'
        ws.cell(r, 19).value = f'=IF(OR(Q{r}="",R{r}=""),"",R{r}-Q{r})'
        ws.cell(r, 20).value = f'=IFERROR(INDEX(规则_渠道系数!$B:$B,MATCH(H{r},规则_渠道系数!$A:$A,0)),1)'
        ws.cell(r, 21).value = f'=IFERROR(INDEX(规则_招聘奖金!$D:$D,MATCH(P{r}&"|"&G{r}&"|"&F{r},规则_招聘奖金!$G:$G,0)),0)'
        ws.cell(r, 22).value = f'={base}*IF(K{r}="",1,70%)'
        ws.cell(r, 23).value = f'={amount_override_formula(src, "BB", f"ROUND(IF(V{r}=0,0,V{r}*{first_ratio}),2)")}'
        ws.cell(r, 24).value = f'={period_override_formula(src, "AT", edate_yyyymm_formula(f"N{r}", 1))}'
        ws.cell(r, 25).value = f'={amount_override_formula(src, "BC", f"ROUND(IF(V{r}=0,0,V{r}*{three_ratio}),2)")}'
        ws.cell(r, 26).value = f'=IF(Y{r}=0,"",{period_override_formula(src, "AU", edate_yyyymm_formula(f"N{r}", 3))})'
        ws.cell(r, 27).value = f'={amount_override_formula(src, "BD", f"ROUND(IF(V{r}=0,0,V{r}*{six_ratio}),2)")}'
        ws.cell(r, 28).value = f'=IF(AA{r}=0,"",{period_override_formula(src, "AV", edate_yyyymm_formula(f"N{r}", 6))})'
        ws.cell(r, 29).value = f'={amount_override_formula(src, "BE", f"ROUND(IF(V{r}=0,0,V{r}*{probation_ratio}),2)")}'
        ws.cell(r, 30).value = f'=IF(AC{r}=0,"",{period_override_formula(src, "AW", probation_period_fallback)})'
        ws.cell(r, 31).value = f'=IF(K{r}="",0,{base}*30%)'
        ws.cell(r, 32).value = f'={amount_override_formula(src, "BF", f"ROUND(IF(AE{r}=0,0,AE{r}*{first_ratio}),2)")}'
        ws.cell(r, 33).value = f'=IF(AF{r}=0,"",{period_override_formula(src, "AX", f"X{r}")})'
        ws.cell(r, 34).value = f'={amount_override_formula(src, "BG", f"ROUND(IF(AE{r}=0,0,AE{r}*{three_ratio}),2)")}'
        ws.cell(r, 35).value = f'=IF(AH{r}=0,"",{period_override_formula(src, "AY", f"Z{r}")})'
        ws.cell(r, 36).value = f'={amount_override_formula(src, "BH", f"ROUND(IF(AE{r}=0,0,AE{r}*{six_ratio}),2)")}'
        ws.cell(r, 37).value = f'=IF(AJ{r}=0,"",{period_override_formula(src, "AZ", f"AB{r}")})'
        ws.cell(r, 38).value = f'={amount_override_formula(src, "BI", f"ROUND(IF(AE{r}=0,0,AE{r}*{probation_ratio}),2)")}'
        ws.cell(r, 39).value = f'=IF(AL{r}=0,"",{period_override_formula(src, "BA", f"AD{r}")})'
        ws.cell(r, 40).value = f'=IFERROR(INDEX(规则_招聘奖金!$E:$E,MATCH(P{r}&"|"&G{r}&"|"&F{r},规则_招聘奖金!$G:$G,0)),"")'
        ws.cell(r, 44).value = f'=IF(OR(AP{r}="-",AP{r}=""),"否",IF(OR(AP{r}=导入_月度数据!AM{src},AP{r}=导入_月度数据!AN{src},AP{r}=导入_月度数据!AO{src},AP{r}=导入_月度数据!AP{src},AP{r}=导入_月度数据!AQ{src}),"是","否"))'
        ws.cell(r, 45).value = f'=IF(导入_月度数据!AR{src}<>"",导入_月度数据!AR{src},"集团统一")'
        ref_category = f'IF(OR(AS{r}="FBU德国",AS{r}="FBU捷克"),IF(G{r}="C1类","C类",G{r}),G{r})'
        key = f'AS{r}&"|"&IF(AS{r}="集团统一",导入_月度数据!AG{src},AS{r})&"|"&{ref_category}&"|"&F{r}'
        ws.cell(r, 46).value = f'=IF(OR(AP{r}="",AP{r}="-",AR{r}="是"),0,IFERROR(INDEX(规则_内推奖金!$E:$E,MATCH({key},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 47).value = f'=IF(AT{r}=0,0,AT{r}*IFERROR(INDEX(规则_内推奖金!$G:$G,MATCH({key},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 48).value = f'=IF(AU{r}=0,"",{edate_yyyymm_formula(f"N{r}", 1)})'
        ws.cell(r, 49).value = f'=IF(AT{r}=0,0,AT{r}*IFERROR(INDEX(规则_内推奖金!$H:$H,MATCH({key},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 50).value = f'=IF(AW{r}=0,"",{edate_yyyymm_formula(f"N{r}", 3)})'
        ws.cell(r, 51).value = f'=IF(AT{r}=0,0,AT{r}*IFERROR(INDEX(规则_内推奖金!$I:$I,MATCH({key},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 52).value = f'=IF(AY{r}=0,"",{edate_yyyymm_formula(f"N{r}", 6)})'
        ws.cell(r, 53).value = f'=IF(AT{r}=0,0,AT{r}*IFERROR(INDEX(规则_内推奖金!$J:$J,MATCH({key},规则_内推奖金!$L:$L,0)),0))'
        ws.cell(r, 54).value = f'=IF(BA{r}=0,"",IF(O{r}="","未转正，待发放",{yyyymm_formula(f"O{r}")}))'
        ws.cell(r, 55).value = (
            f'=IF(C{r}="","",TEXTJOIN("；",TRUE,'
            f'IF(A{r}="重复","工号重复",""),'
            f'IF(OR(F{r}="",G{r}="",H{r}="",M{r}="",N{r}=""),"招聘奖金缺关键字段",""),'
            f'IF(Q{r}="","未匹配招聘周期",""),'
            f'IF(U{r}=0,"未匹配招聘奖金标准",""),'
            f'IF(AND(AP{r}<>"",AP{r}<>"-",AR{r}<>"是",AT{r}=0),"未匹配内推规则或不满足资格",""),'
            f'IF(AR{r}="是","推荐人为直接/间接上级，不计内推","")))'
        )

    add_table(ws, "tbl_detail", f"A1:BC{MAX_ROWS + 1}")
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1)
    imported_cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 41, 42, 43]
    color_header_cells(ws, imported_cols, IMPORT_FILL)
    color_header_cells(ws, [c for c in range(1, len(headers) + 1) if c not in imported_cols], CALC_FILL)
    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    for col in ["B", "J", "L", "AO", "BC"]:
        ws.column_dimensions[col].width = 24
    apply_date_format(ws, [13, 14, 15], 2, MAX_ROWS + 1)


def add_summary_sheets_v2(wb):
    ws = wb.create_sheet("输出_招聘奖金汇总")
    headers = ["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"]
    ws.append(headers)
    summary_rows = MAX_ROWS * 2
    for r in range(2, summary_rows + 2):
        if r <= MAX_ROWS + 1:
            src = r
            id_col = "V"
            name_col = "W"
            role = "招聘负责人"
            detail_id_col = "I"
            amt_cols = [("W", "X"), ("Y", "Z"), ("AA", "AB"), ("AC", "AD")]
        else:
            src = r - MAX_ROWS
            id_col = "X"
            name_col = "Y"
            role = "协助招聘人"
            detail_id_col = "K"
            amt_cols = [("AF", "AG"), ("AH", "AI"), ("AJ", "AK"), ("AL", "AM")]
        ws.cell(r, 1).value = f'=IF(OR(导入_月度数据!{id_col}{src}="",导入_月度数据!{id_col}{src}="-",COUNTIF(导入_月度数据!${id_col}$2:{id_col}{src},导入_月度数据!{id_col}{src})>1),"",导入_月度数据!{id_col}{src})'
        ws.cell(r, 2).value = f'=IF(A{r}="","",导入_月度数据!{name_col}{src})'
        ws.cell(r, 3).value = f'="{role}"'
        ws.cell(r, 4).value = f'=IF(A{r}="","",INDEX(输出_招聘奖金明细!AN:AN,MATCH(A{r},输出_招聘奖金明细!{detail_id_col}:{detail_id_col},0)))'
        ws.cell(r, 5).value = '=导入_月度数据!$A$2'
        for idx, (amount_col, period_col) in enumerate(amt_cols, start=6):
            ws.cell(r, idx).value = f'=IF(A{r}="","",SUMIFS(输出_招聘奖金明细!{amount_col}:{amount_col},输出_招聘奖金明细!{detail_id_col}:{detail_id_col},A{r},输出_招聘奖金明细!{period_col}:{period_col},E{r}))'
        ws.cell(r, 10).value = f'=SUM(F{r}:I{r})'
    style_range(ws, 1, summary_rows + 1, 1, len(headers), header_rows=1)
    color_header_cells(ws, range(1, len(headers) + 1), CALC_FILL)
    add_table(ws, "tbl_recruit_summary", f"A1:J{summary_rows + 1}")
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 16, "B": 18, "C": 14, "D": 12, "E": 12, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16})

    ws = wb.create_sheet("输出_内推奖金汇总")
    headers = ["内推人员工号", "内推人员姓名", "人员状态", "币种", "核算月份", "本月发放"]
    ws.append(headers)
    for r in range(2, MAX_ROWS + 2):
        src = r
        ws.cell(r, 1).value = f'=IF(OR(导入_月度数据!AI{src}="",导入_月度数据!AI{src}="-",COUNTIF(导入_月度数据!$AI$2:AI{src},导入_月度数据!AI{src})>1),"",导入_月度数据!AI{src})'
        ws.cell(r, 2).value = f'=IF(A{r}="","",导入_月度数据!AH{src})'
        ws.cell(r, 3).value = f'=IF(A{r}="","",导入_月度数据!AK{src})'
        ws.cell(r, 4).value = f'=IF(A{r}="","",INDEX(规则_内推奖金!F:F,MATCH(INDEX(输出_招聘奖金明细!AS:AS,MATCH(A{r},输出_招聘奖金明细!AP:AP,0))&"|"&IF(INDEX(输出_招聘奖金明细!AS:AS,MATCH(A{r},输出_招聘奖金明细!AP:AP,0))="集团统一",INDEX(导入_月度数据!AG:AG,MATCH(A{r},导入_月度数据!AI:AI,0)),INDEX(输出_招聘奖金明细!AS:AS,MATCH(A{r},输出_招聘奖金明细!AP:AP,0)))&"|"&INDEX(输出_招聘奖金明细!G:G,MATCH(A{r},输出_招聘奖金明细!AP:AP,0))&"|"&INDEX(输出_招聘奖金明细!F:F,MATCH(A{r},输出_招聘奖金明细!AP:AP,0)),规则_内推奖金!L:L,0)))'
        ws.cell(r, 5).value = '=导入_月度数据!$A$2'
        ws.cell(r, 6).value = f'=IF(A{r}="","",SUMIFS(输出_招聘奖金明细!AU:AU,输出_招聘奖金明细!AP:AP,A{r},输出_招聘奖金明细!AV:AV,E{r})+SUMIFS(输出_招聘奖金明细!AW:AW,输出_招聘奖金明细!AP:AP,A{r},输出_招聘奖金明细!AX:AX,E{r})+SUMIFS(输出_招聘奖金明细!AY:AY,输出_招聘奖金明细!AP:AP,A{r},输出_招聘奖金明细!AZ:AZ,E{r})+SUMIFS(输出_招聘奖金明细!BA:BA,输出_招聘奖金明细!AP:AP,A{r},输出_招聘奖金明细!BB:BB,E{r}))'
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1)
    color_header_cells(ws, range(1, len(headers) + 1), CALC_FILL)
    add_table(ws, "tbl_referral_summary", f"A1:F{MAX_ROWS + 1}")
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 18, "B": 18, "C": 14, "D": 12, "E": 12, "F": 16})

    ws = wb.create_sheet("输出_异常清单")
    headers = ["行号", "姓名", "工号", "异常提示", "人工处理意见"]
    ws.append(headers)
    for r in range(2, MAX_ROWS + 2):
        src = r
        ws.cell(r, 1).value = f'=IF(输出_招聘奖金明细!BC{src}="","",ROW(导入_月度数据!A{src}))'
        ws.cell(r, 2).value = f'=IF(A{r}="","",导入_月度数据!B{src})'
        ws.cell(r, 3).value = f'=IF(A{r}="","",导入_月度数据!C{src})'
        ws.cell(r, 4).value = f'=IF(A{r}="","",输出_招聘奖金明细!BC{src})'
    style_range(ws, 1, MAX_ROWS + 1, 1, len(headers), header_rows=1, fill=WARN_FILL)
    color_header_cells(ws, [1, 2, 3, 4], CALC_FILL)
    color_header_cells(ws, [5], MANUAL_FILL)
    add_table(ws, "tbl_exceptions", f"A1:E{MAX_ROWS + 1}")
    ws.freeze_panes = "A2"
    set_widths(ws, {"A": 10, "B": 18, "C": 16, "D": 80, "E": 28})


def add_regression_sheet(wb):
    ws = wb.create_sheet("回归测试_202604摘要")
    src = load_workbook(SOURCE_XLSX, data_only=True, read_only=False)
    detail = src["附件一-招聘奖金核算明细"]
    recruit_sum = src["附件二-招聘奖金汇总表"]
    referral_sum = src["附件三-内推奖金汇总"]
    rows = [
        ["检查项", "线下表结果", "说明"],
        ["线下明细行数", detail.max_row - 3, "附件一从第4行起为数据"],
        ["招聘汇总人数", recruit_sum.max_row - 2, "附件二从第3行起为数据"],
        ["内推汇总人数", referral_sum.max_row - 2, "附件三从第3行起为数据"],
        ["招聘奖金汇总合计", f"=SUM('输出_招聘奖金汇总'!H:H)", "导入同月数据后自动计算"],
        ["内推奖金汇总合计", f"=SUM('输出_内推奖金汇总'!F:F)", "导入同月数据后自动计算"],
        ["下一步校验", "导入202604原始数据后，对比附件二/附件三人工结果", "差异进入异常清单或规则调整"],
    ]
    for row in rows:
        ws.append(row)
    style_range(ws, 1, ws.max_row, 1, 3, fill=HEADER_FILL)
    add_table(ws, "tbl_regression", f"A1:C{ws.max_row}")
    set_widths(ws, {"A": 24, "B": 28, "C": 60})


def finalize(wb, output_path):
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = ws.freeze_panes or "A2"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def build_import_template():
    wb = Workbook()
    add_instruction_sheet(wb)
    add_import_sheet(wb)
    return finalize(wb, IMPORT_TEMPLATE_XLSX)


def build_rule_workbook(cycle_rules, recruit_rules, referral_rules, channel_rules):
    wb = Workbook()
    wb.remove(wb.active)
    add_rule_sheets(wb, cycle_rules, recruit_rules, referral_rules, channel_rules)
    del wb["规则_资格条件"]
    del wb["规则说明"]
    return finalize(wb, RULE_XLSX)


def main():
    _, cycle_rules, recruit_rules, referral_rules, channel_rules = read_source_rules()
    print(build_import_template())
    print(build_rule_workbook(cycle_rules, recruit_rules, referral_rules, channel_rules))


if __name__ == "__main__":
    main()
