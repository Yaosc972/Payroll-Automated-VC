from datetime import date
from io import BytesIO
import time

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from bonus_platform.app import app
from bonus_platform.engine.domestic_labor import night_shift_config as config_module


pytestmark = pytest.mark.usefixtures("bypass_domestic_labor_access_gate")


def _config():
    return {
        "shift_break_overrides": [
            {
                "shift_code": "DN06",
                "shift_category": "东南班次",
                "shift_name": "处理中心20:00晚班-当月调整",
                "shift_time": "20:00-29:00;",
                "regular_hours": 9,
                "break_periods": ["00:00-00:30"],
                "note": "当月临时调整",
            }
        ],
        "jinjiang_exclusions": [
            {
                "employee_id": "OWHN001",
                "employee_name": "张三",
                "reason": "轻松岗位",
                "start_date": "2026-08-01",
                "end_date": "",
                "note": "",
            }
        ],
        "jinjiang_list_confirmed": True,
    }


def test_platform_baseline_contains_unique_existing_shift_table():
    shifts = config_module.load_baseline_shift_breaks()
    by_code = {row["shift_code"]: row for row in shifts}

    assert len(shifts) == 122
    assert len(by_code) == 122
    assert by_code["DN06"]["break_periods"] == ["24:00-25:00"]
    assert by_code["HD024"]["break_periods"] == ["06:00-06:30"]
    assert by_code["DN06"]["break_segments"] == [
        {"period": "24:00-25:00", "category": "晚上休息"}
    ]
    assert by_code["HD024"]["break_segments"] == [
        {"period": "06:00-06:30", "category": "早上休息"}
    ]
    assert {category: sum(row["shift_category"] == category for row in shifts) for category in {
        "寮步班次", "华东班次", "东南班次",
    }} == {"寮步班次": 55, "华东班次": 39, "东南班次": 28}


def test_save_load_and_copy_only_monthly_jinjiang_list(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path)

    saved = config_module.save_night_shift_config("202608", _config(), updated_by="test")
    copied = config_module.copy_night_shift_config("202608", "202609", updated_by="test")
    changed = _config()
    changed["shift_break_overrides"][0]["shift_name"] = "东南晚班-二次调整"
    second = config_module.save_night_shift_config("202608", changed, updated_by="test")

    assert saved["revision"] == 1
    assert saved["counts"]["baseline_shift_count"] == 122
    assert saved["counts"]["shift_break_override_count"] == 1
    assert next(row for row in saved["effective_shift_breaks"] if row["shift_code"] == "DN06")["break_periods"] == ["00:00-00:30"]
    assert next(row for row in saved["effective_shift_breaks"] if row["shift_code"] == "DN06")["break_segments"] == [
        {"period": "00:00-00:30", "category": "晚上休息"}
    ]
    assert copied["month"] == "202609"
    assert copied["copied_from"] == "202608"
    assert copied["revision"] == 1
    assert copied["shift_break_overrides"] == []
    assert copied["jinjiang_exclusions"] == saved["jinjiang_exclusions"]
    assert copied["jinjiang_list_confirmed"] is True
    assert second["revision"] == 2
    assert [item["revision"] for item in config_module.list_night_shift_config_revisions("202608")] == [2, 1]
    assert config_module.load_night_shift_config_revision("202608", 1)["shift_break_overrides"][0]["shift_name"] == "处理中心20:00晚班-当月调整"
    assert config_module.load_night_shift_config("202608")["shift_break_overrides"][0]["shift_code"] == "DN06"
    assert "regional_positions" not in saved
    assert "continuous_shift_overrides" not in saved
    with pytest.raises(FileExistsError, match="已存在"):
        config_module.copy_night_shift_config("202608", "202609", updated_by="test")


def test_duplicate_business_keys_are_rejected(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path)
    payload = _config()
    payload["shift_break_overrides"].append(dict(payload["shift_break_overrides"][0]))

    with pytest.raises(ValueError, match="重复班次"):
        config_module.save_night_shift_config("202608", payload)


def test_config_template_keeps_examples_in_single_input_sheet(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path)

    workbook_bytes = config_module.generate_night_shift_config_workbook()
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    imported = config_module.parse_night_shift_config_workbook(workbook_bytes)

    assert workbook.sheetnames == ["晋江不享有名单"]
    assert workbook["晋江不享有名单"]["A2"].value == "示例0001"
    assert workbook["晋江不享有名单"]["A3"].value == "示例0002"
    assert imported["jinjiang_exclusions"] == []
    validations = list(workbook["晋江不享有名单"].data_validations.dataValidation)
    assert any("轻松岗位" in (validation.formula1 or "") for validation in validations)


def test_config_workbook_with_current_rows_does_not_add_examples(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path)

    workbook_bytes = config_module.generate_night_shift_config_workbook(_config())
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
    imported = config_module.parse_night_shift_config_workbook(workbook_bytes)

    assert workbook.sheetnames == ["晋江不享有名单"]
    assert workbook["晋江不享有名单"]["A2"].value == "OWHN001"
    assert imported == {
        "shift_break_overrides": [],
        "jinjiang_exclusions": _config()["jinjiang_exclusions"],
        "jinjiang_list_confirmed": True,
    }


def test_config_workbook_accepts_legacy_sheet_name():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "晋江特殊名单"
    sheet.append(["工号", "姓名", "排除原因", "生效日期", "失效日期", "备注"])
    sheet.append(["OWHN002", "李四", "轻松岗位", "2026-08-01", "", ""])
    output = BytesIO()
    workbook.save(output)

    imported = config_module.parse_night_shift_config_workbook(output.getvalue())

    assert imported["jinjiang_exclusions"][0]["employee_id"] == "OWHN002"


def test_config_workbook_rejects_reason_that_system_already_handles():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "晋江不享有名单"
    sheet.append(["工号", "姓名", "排除原因", "生效日期", "失效日期", "备注"])
    sheet.append(["OWHN003", "王五", "计件岗", "2026-08-01", "", ""])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="排除原因只能选择"):
        config_module.parse_night_shift_config_workbook(output.getvalue())


def test_config_workbook_rejects_reversed_dates():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "晋江不享有名单"
    sheet.append(["工号", "姓名", "排除原因", "生效日期", "失效日期", "备注"])
    sheet.append(["OWHN004", "赵六", "轻松岗位", "2026-08-02", "2026-08-01", ""])
    output = BytesIO()
    workbook.save(output)

    with pytest.raises(ValueError, match="失效日期不能早于生效日期"):
        config_module.parse_night_shift_config_workbook(output.getvalue())


def test_active_jinjiang_exclusion_obeys_effective_date():
    entries = _config()["jinjiang_exclusions"]

    assert config_module.find_active_jinjiang_exclusion(entries, "OWHN001", date(2026, 8, 15))["reason"] == "轻松岗位"
    assert config_module.find_active_jinjiang_exclusion(entries, "OWHN001", date(2026, 7, 31)) is None


def test_night_shift_config_api_import_copy_and_download(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path)
    client = TestClient(app)

    missing = client.get("/api/domestic-labor/night-shift/config/202608")
    blank_template = client.get("/api/domestic-labor/night-shift/config-template/download")
    saved = client.put("/api/domestic-labor/night-shift/config/202608", json=_config())
    history = client.get("/api/domestic-labor/night-shift/config/202608/history")
    downloaded = client.get("/api/domestic-labor/night-shift/config/202608/download")
    copied = client.post(
        "/api/domestic-labor/night-shift/config/202609/copy",
        json={"source_month": "202608"},
    )
    imported = client.post(
        "/api/domestic-labor/night-shift/config/202610/import",
        files={
            "file": (
                "night-shift-config.xlsx",
                config_module.generate_night_shift_config_workbook(_config()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert missing.status_code == 200
    assert missing.json()["exists"] is False
    assert missing.json()["counts"]["baseline_shift_count"] == 122
    assert missing.json()["counts"]["effective_shift_count"] == 122
    assert missing.json()["jinjiang_list_confirmed"] is False
    assert blank_template.status_code == 200
    assert load_workbook(BytesIO(blank_template.content)).sheetnames == ["晋江不享有名单"]
    assert saved.status_code == 200
    assert saved.json()["revision"] == 1
    assert saved.json()["counts"]["shift_break_override_count"] == 1
    assert saved.json()["counts"]["effective_shift_count"] == 122
    assert history.status_code == 200
    assert [item["revision"] for item in history.json()["revisions"]] == [1]
    assert downloaded.status_code == 200
    assert load_workbook(BytesIO(downloaded.content)).sheetnames == ["晋江不享有名单"]
    assert copied.status_code == 200
    assert copied.json()["copied_from"] == "202608"
    assert copied.json()["shift_break_overrides"] == []
    assert imported.status_code == 200
    assert imported.json()["month"] == "202610"
    assert imported.json()["jinjiang_list_confirmed"] is True


def test_night_shift_run_uses_immutable_config_snapshot(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path / "configs")
    initial = _config()
    initial["shift_break_overrides"] = [{
        "shift_code": "DG01", "shift_category": "测试班次", "shift_name": "东莞夜班",
        "shift_time": "22:00-32:00;", "regular_hours": 10, "break_periods": [], "note": "",
    }]
    initial["jinjiang_exclusions"] = []
    config_module.save_night_shift_config("202608", initial)

    workbook = Workbook()
    monthly = workbook.active
    monthly.title = "月考勤"
    monthly.append(["工号", "姓名", "考勤月份", "工作地区", "二级部门名称", "岗位名称"])
    monthly.append(["DG001", "张三", "202608", "东莞", "中国操作部", "操作员"])
    daily = workbook.create_sheet("日考勤")
    daily.append(["日期", "工号", "姓名", "工作地区", "岗位名称", "班次编号", "上班一", "下班一"])
    daily.append(["2026-08-01", "DG001", "张三", "东莞", "操作员", "DG01", "22:00", "08:00"])
    buffer = BytesIO()
    workbook.save(buffer)

    client = TestClient(app)
    created = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("night-shift.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "yeban_butie", "attendance_month": "202608"},
    )
    assert created.status_code == 200
    run_id = created.json()["run_id"]

    changed = {**initial, "shift_break_overrides": [{
        "shift_code": "DG01", "shift_category": "测试班次", "shift_name": "东莞夜班",
        "shift_time": "22:00-32:00;", "regular_hours": 10,
        "break_periods": ["00:00-08:00"], "note": "",
    }]}
    config_module.save_night_shift_config("202608", changed)

    for _ in range(30):
        metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
        if metadata["status"] in {"已完成", "失败"}:
            break
        time.sleep(0.1)

    assert metadata["status"] == "已完成"
    assert metadata["nightShiftConfigSnapshot"]["revision"] == 1
    assert config_module.load_night_shift_config("202608")["revision"] == 2
    assert metadata["results"][0]["yeban_butie"] == 25
    assert metadata["summary"]["total_yeban_butie"] == 25
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_night_shift_run_can_use_platform_baseline_without_monthly_config(monkeypatch, tmp_path):
    monkeypatch.setattr(config_module, "NIGHT_SHIFT_CONFIG_DIR", tmp_path / "configs")
    workbook = Workbook()
    monthly = workbook.active
    monthly.title = "月考勤"
    monthly.append(["工号", "姓名", "考勤月份", "工作地区", "二级部门名称", "岗位名称"])
    monthly.append(["DG001", "张三", "202608", "东莞", "中国操作部", "操作员"])
    daily = workbook.create_sheet("日考勤")
    daily.append(["日期", "工号", "姓名", "工作地区", "岗位名称", "班次编号", "上班一", "下班一"])
    daily.append(["2026-08-01", "DG001", "张三", "东莞", "操作员", "DN06", "22:00", "08:00"])
    buffer = BytesIO()
    workbook.save(buffer)

    client = TestClient(app)
    created = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("night-shift.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "yeban_butie", "attendance_month": "202608"},
    )

    assert created.status_code == 200
    snapshot = created.json()["night_shift_config_snapshot"]
    assert snapshot["exists"] is False
    assert snapshot["revision"] == 0
    assert snapshot["counts"]["effective_shift_count"] == 122
    assert len(snapshot["shift_breaks"]) == 122
    for _ in range(30):
        metadata = client.get(f"/api/domestic-labor/runs/{created.json()['run_id']}").json()
        if metadata["status"] in {"已完成", "失败"}:
            break
        time.sleep(0.1)
    assert metadata["status"] == "已完成"
    assert metadata["summary"]["total_yeban_butie"] == 25
    client.delete(f"/api/domestic-labor/runs/{created.json()['run_id']}")
