from bonus_platform.engine.labor.presentation import (
    build_labor_presentation,
    validate_labor_presentation,
)
from bonus_platform.engine.labor.report import build_labor_report
from openpyxl import load_workbook


def _row(name: str, amount_delta: float, *, status: str = "金额差异") -> dict:
    return {
        "employeeName": name,
        "matchStatus": status,
        "pdfHoursTotal": 8.0 if name not in {"site", "Gonesse"} else 0.0,
        "excelHoursTotal": 8.0 if name not in {"site", "Gonesse"} else 0.0,
        "hoursDelta": 0.0,
        "pdfAmountTotal": 100.0 + amount_delta if name not in {"site", "Gonesse"} else 0.0,
        "excelAmountTotal": 100.0 if name not in {"site", "Gonesse"} else 0.0,
        "amountDelta": amount_delta,
        "riskFlags": [],
    }


def test_presentation_contract_counts_real_employees_and_review_items_once():
    amount_deltas = [146.13, 123.53, 113.19, 29.20, 24.30, 19.44]
    rows = [
        _row(f"Employee {index}", amount_delta)
        for index, amount_delta in enumerate(amount_deltas, start=1)
    ]
    rows.extend(
        [
            _row("site", 0.0, status="Excel有PDF无"),
            _row("Gonesse", 0.0, status="通过"),
        ]
    )
    comparison = {
        "summary": {
            "pdfEmployeeCount": 6,
            "excelEmployeeCount": 8,
            "amountDiffCount": 6,
            "exceptionCount": 6,
        },
        "rows": rows,
        "candidateMatches": [
            {
                "pdfEmployeeName": "Employee 3 long name",
                "excelEmployeeName": "Employee 3",
                "amountDelta": 113.19,
            }
        ],
    }

    presentation = build_labor_presentation(comparison, excel_record_count=6)

    assert [row["employeeName"] for row in presentation["employeeRows"]] == [
        f"Employee {index}" for index in range(1, 7)
    ]
    assert presentation["summary"] == {
        "employeeCount": 6,
        "differenceEmployeeCount": 6,
        "passedEmployeeCount": 0,
        "amountDiffEmployeeCount": 6,
        "hoursDiffEmployeeCount": 0,
        "notInInvoiceEmployeeCount": 0,
        "candidateMatchCount": 1,
        "reviewItemCount": 7,
        "amountImpact": 455.79,
        "hoursImpact": 0.0,
        "excelRecordCount": 6,
        "sourceComparisonRowCount": 8,
        "excludedNonEmployeeRowCount": 2,
    }
    assert validate_labor_presentation(presentation) == []


def test_presentation_contract_validation_rejects_count_drift():
    presentation = build_labor_presentation(
        {
            "summary": {},
            "rows": [_row("Employee 1", 10.0)],
            "candidateMatches": [],
        },
        excel_record_count=1,
    )
    presentation["summary"]["employeeCount"] = 9
    presentation["summary"]["notInInvoiceEmployeeCount"] = 2

    errors = validate_labor_presentation(presentation)

    assert "employeeCount must equal len(employeeRows)" in errors
    assert "notInInvoiceEmployeeCount must equal missing-invoice employee rows" in errors


def test_report_consumes_presentation_rows_and_summary_without_recounting_candidates(tmp_path):
    comparison = {
        "summary": {
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 2,
            "amountDiffCount": 1,
            "exceptionCount": 1,
        },
        "rows": [
            _row("Employee 1", 10.0),
            _row("site", 0.0, status="通过"),
        ],
        "candidateMatches": [
            {
                "pdfEmployeeName": "Employee One",
                "excelEmployeeName": "Employee 1",
                "amountDelta": 10.0,
            }
        ],
    }
    presentation = build_labor_presentation(comparison, excel_record_count=1)
    output = tmp_path / "canonical-report.xlsx"

    build_labor_report(
        output,
        comparison,
        [],
        [],
        {"name": "姓名", "hours": "工时", "amount": "金额"},
        presentation=presentation,
    )

    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        assert workbook["全员对账明细"].max_row == 2
        summary = {
            row[0]: row[1]
            for row in workbook["核对摘要"].iter_rows(values_only=True)
            if row[0]
        }
        assert summary["displayEmployeeCount"] == 1
        assert summary["differenceEmployeeCount"] == 1
        assert summary["candidateMatchCount"] == 1
        assert summary["reviewItemCount"] == 2
        assert summary["amountImpact"] == 10.0
    finally:
        workbook.close()
