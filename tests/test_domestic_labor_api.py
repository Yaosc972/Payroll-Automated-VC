"""国内劳务工薪酬核算 API 测试"""
from datetime import date
from io import BytesIO
import re
import threading

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from bonus_platform.app import app, _domestic_labor_export_filename
from bonus_platform.engine.domestic_labor.engines.canbu import CanBuEngine
from bonus_platform.engine.domestic_labor.engines.gonglingjiang import GongLingJiangEngine
from bonus_platform.engine.domestic_labor.engines.quanqinjiang import QuanQinJiangEngine
from bonus_platform.engine.domestic_labor.engines.waisu_butie import WaiSuBuTieEngine
from bonus_platform.engine.domestic_labor.exporter import ExcelExporter
from bonus_platform.engine.domestic_labor import parser as domestic_parser
from bonus_platform.engine.domestic_labor.parser import ExcelParser


pytestmark = pytest.mark.usefixtures("bypass_domestic_labor_access_gate")


def _create_test_excel(sheet_names: dict[str, list[list]]) -> bytes:
    """创建测试 Excel 文件"""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name, rows in sheet_names.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_excel_parser_supports_legacy_xls(monkeypatch, tmp_path):
    """老式 .xls 文件走 xlrd 解析分支，避免被 openpyxl 拒绝"""
    class FakeCell:
        def __init__(self, value):
            self.value = value
            self.ctype = domestic_parser.xlrd.XL_CELL_NUMBER if isinstance(value, (int, float)) else domestic_parser.xlrd.XL_CELL_TEXT

    class FakeSheet:
        name = "Sheet"
        nrows = 2
        ncols = 4
        values = [
            ["工号", "姓名", "正班出勤天数", "旷工天数"],
            ["OWHN001", "张三", 20.0, 1.0],
        ]

        def cell_value(self, row, col):
            return self.values[row][col]

        def cell(self, row, col):
            return FakeCell(self.values[row][col])

    class FakeBook:
        datemode = 0

        def sheet_names(self):
            return ["Sheet"]

        def sheet_by_name(self, name):
            assert name == "Sheet"
            return FakeSheet()

    xls_path = tmp_path / "attendance.xls"
    xls_path.write_bytes(b"fake xls")
    monkeypatch.setattr(domestic_parser.xlrd, "open_workbook", lambda _: FakeBook())

    parser = ExcelParser(str(xls_path)).load()
    parsed = parser.parse_sheet("Sheet")

    assert parser.get_sheet_names() == ["Sheet"]
    assert parsed.headers == ["工号", "姓名", "正班出勤天数", "旷工天数"]
    assert parsed.rows == [{"工号": "OWHN001", "姓名": "张三", "正班出勤天数": 20, "旷工天数": 1}]


def test_excel_parser_skips_external_workbook_links(monkeypatch, tmp_path):
    """核算只读取当前工作簿数据，不应解析可能非常大的历史外链缓存。"""
    path = tmp_path / "attendance.xlsx"
    path.write_bytes(b"placeholder")
    load_calls = []

    def recording_load_workbook(*args, **kwargs):
        load_calls.append((args, kwargs))
        return object()

    monkeypatch.setattr(domestic_parser.openpyxl, "load_workbook", recording_load_workbook)

    ExcelParser(str(path)).load()

    assert load_calls[0][1]["keep_links"] is False


def test_excel_parser_detects_legacy_xls_with_xlsx_extension(monkeypatch, tmp_path):
    """东宝导出的错后缀文件应按真实文件头走 xlrd。"""
    class FakeBook:
        def sheet_names(self):
            return ["日考勤", "月考勤"]

    mislabeled_path = tmp_path / "东宝系统-7月东南考勤数据.xlsx"
    mislabeled_path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"legacy workbook")
    opened_paths = []
    monkeypatch.setattr(
        domestic_parser.xlrd,
        "open_workbook",
        lambda path: opened_paths.append(path) or FakeBook(),
    )

    parser = ExcelParser(str(mislabeled_path)).load()

    assert parser._engine == "xlrd"
    assert parser.get_sheet_names() == ["日考勤", "月考勤"]
    assert opened_paths == [str(mislabeled_path)]


def test_payroll_loader_supports_dormitory_sheet_and_date_aliases(tmp_path):
    """宿舍名单及入宿/离宿字段应统一为外宿引擎使用的标准字段。"""
    path = tmp_path / "dormitory-aliases.xlsx"
    path.write_bytes(_create_test_excel({
        "月考勤": [
            ["工号", "姓名", "考勤月份"],
            ["OWHN001", "张三", "202605"],
        ],
        "宿舍名单": [
            ["工号", "入宿时间", "离宿时间"],
            [" OWHN001 ", date(2026, 5, 1), date(2026, 5, 31)],
        ],
    }))

    with domestic_parser.PayrollDataLoader(str(path)) as loader:
        grouped = loader.group_housing_by_employee()

    assert grouped == {
        "OWHN001": [{
            "工号": "OWHN001",
            "入宿时间": date(2026, 5, 1),
            "离宿时间": date(2026, 5, 31),
            "入住时间": date(2026, 5, 1),
            "退宿时间": date(2026, 5, 31),
        }]
    }


def test_multi_file_loader_identifies_business_fields_after_columns_are_reordered(tmp_path):
    """输入列顺序变化时应始终按表头字段取值，不依赖固定列号。"""
    path = tmp_path / "reordered-columns.xlsx"
    path.write_bytes(_create_test_excel({
        "月考勤": [
            ["岗位名称", "考勤月份", "姓名", "工号", "工作地区", "实际在职工作日天数", "排班天数", "一级部门名称"],
            ["操作员", "202606", "张三", "OWHN001", "嘉善", 26, 26, "华东操作"],
        ],
        "日考勤": [
            ["刷卡加班", "工号", "日期", "正班时数", "岗位名称", "工作地区", "姓名", "工作状态"],
            [0, "OWHN001", date(2026, 6, 1), 8, "操作员", "嘉善", "张三", "工作日"],
        ],
        "住宿名单": [
            ["退宿时间", "姓名", "工号", "入住时间"],
            [date(2026, 6, 11), "张三", "OWHN001", date(2026, 5, 1)],
        ],
    }))

    with domestic_parser.MultiFilePayrollDataLoader([str(path)]) as loader:
        summary = loader.validate_inputs(["waisu_butie"], "202606")
        monthly = loader.monthly.rows[0]
        daily = loader.group_daily_by_employee()["OWHN001"][0]
        housing = loader.group_housing_by_employee()["OWHN001"][0]

    assert summary["present_types"] == ["daily", "housing", "monthly"]
    assert (monthly["工号"], monthly["工作地区"], monthly["岗位名称"]) == ("OWHN001", "嘉善", "操作员")
    assert (daily["工号"], daily["正班时数"], daily["刷卡加班"]) == ("OWHN001", 8, 0)
    assert (housing["工号"], housing["入住时间"], housing["退宿时间"]) == (
        "OWHN001", date(2026, 5, 1), date(2026, 6, 11),
    )


def test_multi_file_loader_opens_independent_workbooks_in_parallel(monkeypatch, tmp_path):
    paths = [tmp_path / "月考勤.xlsx", tmp_path / "日考勤.xlsx"]
    for path in paths:
        path.write_bytes(b"placeholder")

    load_barrier = threading.Barrier(len(paths), timeout=1)
    load_threads = set()
    load_threads_lock = threading.Lock()

    def synchronized_load(parser, password=None):
        load_barrier.wait()
        with load_threads_lock:
            load_threads.add(threading.get_ident())
        return parser

    monkeypatch.setattr(domestic_parser.ExcelParser, "load", synchronized_load)

    loader = domestic_parser.MultiFilePayrollDataLoader([str(path) for path in paths])
    loader.load()

    assert len(load_threads) == len(paths)


def _quanqinjiang_data() -> bytes:
    """全勤奖测试数据"""
    return _create_test_excel({
        "全勤奖": [
            ["工号", "姓名", "考勤月份", "入职日期", "最后工作日", "旷工天数", "正班迟到次数", "早退次数", "签卡次数", "工伤假天数", "事假时数", "病假时数", "入离职缺勤时数", "迟到早退30分钟内扣款"],
            ["OWHN001", "张三", "202606", "2023-01-15", None, 0, 0, 0, 0, 0, 0, 0, 0, 0],
            ["OWHN002", "李四", "202606", "2022-06-01", None, 1, 0, 0, 0, 0, 0, 0, 0, 0],
        ],
    })


def _multi_engine_data() -> bytes:
    """多引擎测试数据"""
    return _create_test_excel({
        "全勤奖": [
            ["工号", "姓名", "考勤月份", "入职日期", "最后工作日", "旷工天数", "正班迟到次数", "早退次数", "签卡次数", "工伤假天数", "事假时数", "病假时数", "入离职缺勤时数", "迟到早退30分钟内扣款"],
            ["OWHN001", "张三", "202606", "2023-01-15", None, 0, 0, 0, 0, 0, 0, 0, 0, 0],
        ],
        "餐补": [
            ["工号", "姓名", "餐补标准", "出勤天数", "正班时数合计", "预计算餐补"],
            ["OWHN001", "张三", "19元/天，封顶500元/月", 22, 176, None],
        ],
    })


def _canbu_data_with_blank_rows() -> bytes:
    """餐补数据包含空白行，空白行不应进入核算结果。"""
    return _create_test_excel({
        "月考勤": [
            ["工号", "姓名", "考勤月份", "工作地区", "一级部门名称", "二级部门名称", "岗位名称", "排班天数", "实际在职工作日天数", "事假时数", "病假时数", "旷工天数"],
            ["OWHN001", "张三", "202605", "东莞", "莞深操作", "中国操作部", "操作员", 22, 22, 0, 0, 0],
            ["工号", "姓名", "考勤月份", "工作地区", "一级部门名称", "二级部门名称", "岗位名称", "排班天数", "实际在职工作日天数", "事假时数", "病假时数", "旷工天数"],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            ["None", "None", None, None, None, None, None, None, None, None, None, None],
        ],
        "日考勤": [
            ["日期", "工号", "姓名", "工作地区", "工作状态", "正班时数", "刷卡加班"],
            ["2026-05-01", "OWHN001", "张三", "东莞", "工作日", 8, 0],
            ["2026-05-02", None, None, None, None, None, None],
            ["2026-05-03", "None", "None", None, None, None, None],
        ],
    })


def _gonglingjiang_data() -> bytes:
    """工龄奖 API 工作流测试数据"""
    return _create_test_excel({
        "月考勤": [
            ["工号", "姓名", "考勤月份", "工作地区", "二级部门名称", "岗位名称", "入职日期", "排班天数", "实际在职工作日天数", "正班出勤天数", "事假时数", "病假时数", "旷工天数", "排休请假天数"],
            ["OWHN001", "张三", "202606", "东莞", "中国操作部", "操作员", "2023-01-01", 26, 26, 26, 0, 0, 0, 0],
        ],
    })


def _gonglingjiang_collection_data(work_area: str = "东莞") -> bytes:
    """第四纵队工龄奖名单校验数据"""
    return _create_test_excel({
        "月考勤": [
            ["工号", "姓名", "考勤月份", "工作地区", "二级部门名称", "岗位名称", "入职日期", "排班天数", "实际在职工作日天数", "正班出勤天数", "事假时数", "病假时数", "旷工天数", "排休请假天数"],
            ["OWHN001", "张三", "202606", work_area, "第四纵队", "操作员", "2023-01-01", 26, 26, 26, 0, 0, 0, 0],
        ],
    })


def _waisu_butie_data() -> bytes:
    """外宿补贴平台端到端测试数据。"""
    return _create_test_excel({
        "月考勤": [
            ["工号", "姓名", "考勤月份", "工作地区", "一级部门名称", "岗位名称", "入职日期", "最后工作日", "外宿补贴标准", "休年假小时", "病假时数"],
            ["OWHN001", "张三", "202606", "嘉善", "华东操作", "操作员", date(2023, 1, 1), "", 999, 0, 0],
        ],
        "日考勤": [
            ["工号", "姓名", "工作地区", "岗位名称", "日期", "上班一", "下班一"],
            ["OWHN001", "张三", "嘉善", "操作员", date(2026, 6, 12), "09:00", "18:00"],
        ],
        "住宿名单": [
            ["工号", "姓名", "入住时间", "退宿时间"],
            ["OWHN001", "张三", date(2026, 5, 1), date(2026, 6, 11)],
        ],
    })


def _split_waisu_butie_files() -> tuple[bytes, bytes, bytes]:
    monthly = _create_test_excel({
        "Sheet1": [
            ["工号", "姓名", "考勤月份", "工作地区", "一级部门名称", "岗位名称", "入职日期", "最后工作日", "休年假小时", "病假时数"],
            ["OWHN001", "张三", "202606", "嘉善", "华东操作", "操作员", date(2023, 1, 1), "", 0, 0],
        ],
    })
    daily = _create_test_excel({
        "Sheet1": [
            ["工号", "姓名", "工作地区", "岗位名称", "日期", "上班一", "下班一"],
            ["OWHN001", "张三", "嘉善", "操作员", date(2026, 6, 12), "09:00", "18:00"],
        ],
    })
    housing = _create_test_excel({
        "Sheet1": [
            ["工号", "姓名", "入住时间", "退宿时间"],
            ["OWHN001", "张三", date(2026, 5, 1), date(2026, 6, 11)],
        ],
    })
    return monthly, daily, housing


def _split_canbu_files() -> tuple[bytes, bytes]:
    monthly = _create_test_excel({
        "月度汇总": [
            ["工号", "姓名", "考勤月份", "工作地区", "一级部门名称", "二级部门名称", "岗位名称", "排班天数", "实际在职工作日天数"],
            ["OWHN001", "张三", "202606", "东莞", "莞深操作", "中国操作部", "操作员", 26, 26],
        ],
    })
    daily = _create_test_excel({
        "考勤明细": [
            ["日期", "工号", "姓名", "工作地区", "岗位名称", "工作状态", "正班时数", "刷卡加班"],
            [date(2026, 6, 1), "OWHN001", "张三", "东莞", "操作员", "工作日", 8, 0],
        ],
    })
    return monthly, daily


# ── 测试用例 ──


def test_rule_package_only_publishes_verified_subjects():
    client = TestClient(app)

    response = client.get("/api/domestic-labor/rule-package")

    assert response.status_code == 200
    package = response.json()
    assert package["package_id"] == "DL-PAYROLL"
    assert package["version"] == "1.1.9"
    assert package["status"] == "已发布"
    assert {category["id"] for category in package["categories"]} == {"allowance", "bonus"}
    assert {subject["id"] for subject in package["subjects"]} == {"quanqinjiang", "canbu", "waisu_butie", "gonglingjiang"}
    assert all(subject["status"] == "已验证" for subject in package["subjects"])
    assert all(subject["version"].startswith("DL-") for subject in package["subjects"])
    assert all(subject["verification"] for subject in package["subjects"])
    assert all(subject["regions"] for subject in package["subjects"])
    assert all(subject["change_log"] for subject in package["subjects"])
    assert package["version_history"][0]["version"] == "1.1.9"
    assert package["version_history"][0]["subject_ids"] == ["quanqinjiang", "canbu", "waisu_butie", "gonglingjiang"]


def test_rule_package_publishes_verified_attendance_bonus_and_seniority():
    client = TestClient(app)

    package = client.get("/api/domestic-labor/rule-package").json()
    payload = str(package)

    quanqin = next(subject for subject in package["subjects"] if subject["id"] == "quanqinjiang")
    assert quanqin["status"] == "已验证"
    assert quanqin["version"] == "DL-QUANQIN.v1.0.0"
    assert "2,807" in str(quanqin)
    assert "OWHN9535" in str(quanqin)
    assert "OWHN9353" in str(quanqin)
    assert "OWHX0190" in str(quanqin)
    gongling = next(subject for subject in package["subjects"] if subject["id"] == "gonglingjiang")
    assert gongling["status"] == "已验证"
    assert gongling["version"] == "DL-GONGLING.v1.0.7"
    assert "B操作部" in str(gongling)
    assert "包含“安检员”" in str(gongling)
    assert "第四纵队" in str(gongling)
    assert "头程运营部" in str(gongling)
    assert "不限制工作地区" in str(gongling)
    assert "FBU不设56小时门槛" in str(gongling)
    assert "FBU折算不包含排休请假" in str(gongling)
    assert "FBU不使用正班出勤为0的通用归零特例" in str(gongling)
    assert "工龄奖标准/排班天数" in str(gongling)
    assert "东南 / 闽赣兼容区域" in str(gongling)
    assert "华东 / 华西不发放部门" in str(gongling)
    assert gongling["pending_confirmations"]
    assert "OWHN2187" in gongling["pending_confirmations"][0]


def test_rule_package_supports_immutable_version_lookup():
    client = TestClient(app)

    published = client.get("/api/domestic-labor/rule-package", params={"version": "1.0.0"})
    missing = client.get("/api/domestic-labor/rule-package", params={"version": "9.9.9"})

    assert published.status_code == 200
    assert published.json()["display_version"] == "DL-PAYROLL.v1.0.0"
    assert {subject["id"] for subject in published.json()["subjects"]} == {"canbu", "waisu_butie"}
    assert missing.status_code == 404
    assert missing.json()["detail"] == "规则包版本不存在: 9.9.9"


def test_rule_package_preserves_dongguan_restriction_in_previous_version():
    package = TestClient(app).get("/api/domestic-labor/rule-package", params={"version": "1.1.4"}).json()
    gongling = next(subject for subject in package["subjects"] if subject["id"] == "gonglingjiang")

    assert gongling["version"] == "DL-GONGLING.v1.0.3"
    assert "仅识别到东莞第四纵队" in str(gongling)
    assert "工作地区为东莞且二级部门为头程运营部" in str(gongling)


def test_rule_package_preserves_broad_wes_route_in_previous_version():
    package = TestClient(app).get("/api/domestic-labor/rule-package", params={"version": "1.1.5"}).json()
    gongling = next(subject for subject in package["subjects"] if subject["id"] == "gonglingjiang")

    assert gongling["version"] == "DL-GONGLING.v1.0.4"
    assert "华西 / 华东 / 东南兼容区域" in str(gongling)


def test_rule_package_preserves_operation_only_intermediate_version():
    package = TestClient(app).get("/api/domestic-labor/rule-package", params={"version": "1.1.2"}).json()
    gongling = next(subject for subject in package["subjects"] if subject["id"] == "gonglingjiang")

    assert "第四纵队" not in str(gongling)
    assert "头程运营部" not in str(gongling)


def test_rule_package_preserves_pre_fix_version_and_publishes_cross_month_fix():
    client = TestClient(app)

    current = client.get("/api/domestic-labor/rule-package").json()
    previous = client.get("/api/domestic-labor/rule-package", params={"version": "1.1.0"}).json()
    current_waisu = next(subject for subject in current["subjects"] if subject["id"] == "waisu_butie")
    previous_waisu = next(subject for subject in previous["subjects"] if subject["id"] == "waisu_butie")

    assert current_waisu["version"] == "DL-WAISU.v1.0.1"
    assert "最后工作日在核算月月末或之后" in "".join(current_waisu["common_rules"])
    assert previous_waisu["version"] == "DL-WAISU.v1.0.0"
    assert "最后工作日在核算月月末或之后" not in "".join(previous_waisu["common_rules"])


@pytest.mark.parametrize(
    ("engines", "subject_name"),
    [
        (["canbu"], "餐补"),
        (["waisu_butie"], "外宿补贴"),
        (["quanqinjiang"], "全勤奖"),
        (["gonglingjiang"], "工龄奖"),
        (["canbu", "waisu_butie"], "多科目"),
    ],
)
def test_domestic_labor_export_filename_uses_chinese_subject_and_timestamp(engines, subject_name):
    file_name = _domestic_labor_export_filename({
        "engines": engines,
        "attendanceMonth": "2026-05",
    })

    assert re.fullmatch(
        rf"{subject_name}核算结果_202605_\d{{8}}\.xlsx",
        file_name,
    )


def test_list_templates():
    """测试模板列表接口"""
    client = TestClient(app)
    response = client.get("/api/domestic-labor/templates")

    assert response.status_code == 200
    data = response.json()
    assert "templates" in data
    assert len(data["templates"]) == 4

    # 验证每个模板结构
    engines = {t["engine"] for t in data["templates"]}
    assert engines == {"quanqinjiang", "canbu", "waisu_butie", "gonglingjiang"}

    for template in data["templates"]:
        assert "name" in template
        assert "description" in template
        assert "columns" in template


def test_download_template():
    """测试模板下载接口"""
    client = TestClient(app)
    response = client.get("/api/domestic-labor/templates/quanqinjiang/download")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert len(response.content) > 0


def test_download_waisu_template_is_directly_uploadable_workbook():
    """外宿补贴模板应包含资格、日考勤和住宿记录三类输入，且不把说明写成数据行。"""
    client = TestClient(app)
    response = client.get("/api/domestic-labor/templates/waisu_butie/download")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["月考勤", "日考勤", "住宿名单"]

    monthly_headers = [cell.value for cell in workbook["月考勤"][1]]
    assert {"工号", "姓名", "考勤月份", "工作地区", "一级部门名称", "岗位名称"}.issubset(monthly_headers)
    assert {"入职日期", "最后工作日", "休年假小时", "病假时数"}.issubset(monthly_headers)
    assert workbook["月考勤"].max_row == 1

    daily_headers = [cell.value for cell in workbook["日考勤"][1]]
    assert {"日期", "工号", "姓名", "工作地区", "岗位名称", "上班一", "下班一"}.issubset(daily_headers)
    assert workbook["日考勤"].max_row == 1

    housing_headers = [cell.value for cell in workbook["住宿名单"][1]]
    assert housing_headers == ["工号", "姓名", "入住时间", "退宿时间"]
    assert workbook["住宿名单"].max_row == 1


def test_download_template_invalid_engine():
    """测试无效引擎模板下载"""
    client = TestClient(app)
    response = client.get("/api/domestic-labor/templates/invalid_engine/download")

    assert response.status_code == 404


def test_create_run_with_valid_data():
    """测试创建计算任务 - 有效数据"""
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "quanqinjiang", "attendance_month": "202606"},
    )

    assert response.status_code == 200
    data = response.json()
    assert "run_id" in data
    assert data["status"] == "已完成"
    assert "计算完成" in data["message"]

    results_response = client.get(f"/api/domestic-labor/runs/{data['run_id']}/results")
    assert results_response.status_code == 200
    assert results_response.json()["status"] == "已完成"
    assert results_response.json()["results"]

    # Cleanup
    client.delete(f"/api/domestic-labor/runs/{data['run_id']}")


def test_create_run_persists_uploaded_file_before_returning(monkeypatch):
    import bonus_platform.app as app_module

    persisted = []
    monkeypatch.setattr(
        app_module,
        "persist_payroll_file",
        lambda run_id, path: persisted.append((run_id, str(path))),
    )
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "quanqinjiang", "attendance_month": "202606"},
    )

    assert response.status_code == 200
    assert response.json()["status"] == "已完成"
    assert len(persisted) == 1
    assert persisted[0][0] == response.json()["run_id"]
    assert persisted[0][1].endswith(".xlsx")
    client.delete(f"/api/domestic-labor/runs/{response.json()['run_id']}")


def test_domestic_direct_upload_plan_returns_signed_url(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(
        app_module,
        "create_domestic_labor_signed_upload",
        lambda run_id, filename: {
            "signedUrl": "https://example.supabase.co/storage/v1/object/upload/sign/token",
            "objectPath": f"domestic-labor-runs/production/{run_id}/{filename}",
            "relativePath": filename,
        },
    )
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={
            "fileName": "考勤.xlsx",
            "fileSize": len(_quanqinjiang_data()),
            "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["runId"].startswith("payroll_")
    assert payload["upload"]["signedUrl"].startswith("https://example.supabase.co/")
    assert payload["upload"]["filename"].endswith(".xlsx")
    assert "service-role" not in str(payload)
    client.delete(f"/api/domestic-labor/runs/{payload['runId']}")


def test_domestic_direct_upload_plan_returns_signed_urls_for_multiple_files(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(
        app_module,
        "create_domestic_labor_signed_upload",
        lambda run_id, filename: {
            "signedUrl": f"https://example.supabase.co/storage/v1/object/upload/sign/{filename}",
            "objectPath": f"domestic-labor-runs/production/{run_id}/{filename}",
            "relativePath": filename,
        },
    )
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={
            "files": [
                {"fileName": "月考勤.xlsx", "fileSize": 1024, "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
                {"fileName": "日考勤.xlsx", "fileSize": 2048, "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert len(payload["uploads"]) == 2
    assert [upload["originalFilename"] for upload in payload["uploads"]] == ["月考勤.xlsx", "日考勤.xlsx"]
    assert len({upload["filename"] for upload in payload["uploads"]}) == 2
    assert payload["upload"] == payload["uploads"][0]
    metadata = client.get(f"/api/domestic-labor/runs/{payload['runId']}").json()
    assert metadata["fileNames"] == ["月考勤.xlsx", "日考勤.xlsx"]
    assert metadata["expectedFileSize"] == 3072
    client.delete(f"/api/domestic-labor/runs/{payload['runId']}")


def test_domestic_direct_upload_plan_rejects_invalid_file(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={"fileName": "考勤.csv", "fileSize": 10, "contentType": "text/csv"},
    )

    assert response.status_code == 400
    assert "Excel" in response.json()["detail"]


def test_domestic_direct_upload_plan_rejects_oversized_file(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(app_module, "_domestic_labor_direct_upload_max_bytes", lambda: 100)
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={"fileName": "考勤.xlsx", "fileSize": 101},
    )

    assert response.status_code == 413
    assert "100" in response.json()["detail"]


def test_domestic_direct_upload_complete_materializes_and_calculates(monkeypatch):
    import bonus_platform.app as app_module

    file_bytes = _quanqinjiang_data()
    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(
        app_module,
        "create_domestic_labor_signed_upload",
        lambda run_id, filename: {
            "signedUrl": "https://example.supabase.co/storage/v1/object/upload/sign/token",
            "objectPath": f"domestic-labor-runs/production/{run_id}/{filename}",
            "relativePath": filename,
        },
    )
    client = TestClient(app)
    plan = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={"fileName": "考勤.xlsx", "fileSize": len(file_bytes)},
    ).json()
    run_id = plan["runId"]
    filename = plan["upload"]["filename"]

    def fake_materialize(target_run_id, target_filename):
        assert target_run_id == run_id
        assert target_filename == filename
        target = app_module.get_payroll_run_dir(run_id) / filename
        target.write_bytes(file_bytes)
        return target

    calculated = []

    def fake_calculate(
        target_run_id,
        file_path,
        month,
        engines,
        password,
        hrbp,
        validate_inputs=False,
        initial_metadata=None,
    ):
        calculated.append((target_run_id, file_path, month, engines, password, hrbp))
        return app_module.update_payroll_metadata(target_run_id, {
            "status": "已完成",
            "results": [],
            "inputSummary": {"file_count": 1},
        })

    monkeypatch.setattr(app_module, "materialize_payroll_file", fake_materialize)
    monkeypatch.setattr(app_module, "_run_payroll_calculation", fake_calculate)

    response = client.post(
        f"/api/domestic-labor/runs/{run_id}/direct-upload-complete",
        json={
            "engines": "quanqinjiang",
            "attendanceMonth": "202606",
            "password": "",
            "hrbpList": [{"工号": "OWHN001"}],
        },
    )

    assert response.status_code == 200
    assert response.json()["status"] == "已完成"
    assert calculated[0][2:] == ("202606", ["quanqinjiang"], None, ["OWHN001"])
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_domestic_direct_upload_complete_materializes_multiple_files(monkeypatch):
    import bonus_platform.app as app_module

    monthly, daily = _split_canbu_files()
    file_payloads = {"月考勤.xlsx": monthly, "日考勤.xlsx": daily}
    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(
        app_module,
        "create_domestic_labor_signed_upload",
        lambda run_id, filename: {
            "signedUrl": f"https://example.supabase.co/storage/v1/object/upload/sign/{filename}",
            "objectPath": f"domestic-labor-runs/production/{run_id}/{filename}",
            "relativePath": filename,
        },
    )
    client = TestClient(app)
    plan = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={
            "files": [
                {"fileName": name, "fileSize": len(content), "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
                for name, content in file_payloads.items()
            ]
        },
    ).json()
    run_id = plan["runId"]
    content_by_saved_name = {
        upload["filename"]: file_payloads[upload["originalFilename"]]
        for upload in plan["uploads"]
    }

    def fake_materialize(target_run_id, target_filename):
        target = app_module.get_payroll_run_dir(target_run_id) / target_filename
        target.write_bytes(content_by_saved_name[target_filename])
        return target

    calculated = []

    def fake_calculate(
        target_run_id,
        file_paths,
        month,
        engines,
        password,
        hrbp,
        validate_inputs=False,
        initial_metadata=None,
    ):
        calculated.append((target_run_id, file_paths, month, engines, password, hrbp, initial_metadata))
        return {
            "status": "已完成",
            "results": [],
            "inputSummary": {"file_count": len(file_paths)},
        }

    monkeypatch.setattr(app_module, "materialize_payroll_file", fake_materialize)
    monkeypatch.setattr(app_module, "_run_payroll_calculation", fake_calculate)

    response = client.post(
        f"/api/domestic-labor/runs/{run_id}/direct-upload-complete",
        json={"engines": ["canbu"], "attendanceMonth": "202606"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["status"] == "已完成"
    assert payload["input_summary"]["file_count"] == 2
    assert len(calculated[0][1]) == 2
    assert calculated[0][6] == {
        "engines": ["canbu"],
        "attendanceMonth": "202606",
        "filePath": calculated[0][1][0],
        "filePaths": calculated[0][1],
        "fileSize": len(monthly) + len(daily),
    }
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_domestic_direct_upload_parses_each_workbook_only_once(monkeypatch):
    import bonus_platform.app as app_module

    monthly, daily = _split_canbu_files()
    file_payloads = {"月考勤.xlsx": monthly, "日考勤.xlsx": daily}
    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(
        app_module,
        "create_domestic_labor_signed_upload",
        lambda run_id, filename: {
            "signedUrl": f"https://example.supabase.co/storage/v1/object/upload/sign/{filename}",
            "objectPath": f"domestic-labor-runs/production/{run_id}/{filename}",
            "relativePath": filename,
        },
    )
    client = TestClient(app)
    plan = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={
            "files": [
                {"fileName": name, "fileSize": len(content)}
                for name, content in file_payloads.items()
            ]
        },
    ).json()
    run_id = plan["runId"]
    content_by_saved_name = {
        upload["filename"]: file_payloads[upload["originalFilename"]]
        for upload in plan["uploads"]
    }

    def fake_materialize(target_run_id, target_filename):
        target = app_module.get_payroll_run_dir(target_run_id) / target_filename
        target.write_bytes(content_by_saved_name[target_filename])
        return target

    load_calls = []
    original_load = domestic_parser.ExcelParser.load

    def counted_load(parser, *args, **kwargs):
        load_calls.append(parser.file_path.name)
        return original_load(parser, *args, **kwargs)

    monkeypatch.setattr(app_module, "materialize_payroll_file", fake_materialize)
    monkeypatch.setattr(domestic_parser.ExcelParser, "load", counted_load)

    response = client.post(
        f"/api/domestic-labor/runs/{run_id}/direct-upload-complete",
        json={"engines": ["canbu"], "attendanceMonth": "202606"},
    )

    assert response.status_code == 200
    assert response.json()["status"] == "已完成"
    assert len(load_calls) == 2
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_domestic_direct_upload_complete_rejects_size_mismatch(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "domestic_labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(
        app_module,
        "create_domestic_labor_signed_upload",
        lambda run_id, filename: {
            "signedUrl": "https://example.supabase.co/storage/v1/object/upload/sign/token",
            "objectPath": f"domestic-labor-runs/production/{run_id}/{filename}",
            "relativePath": filename,
        },
    )
    client = TestClient(app)
    plan = client.post(
        "/api/domestic-labor/runs/direct-upload-plan",
        json={"fileName": "考勤.xlsx", "fileSize": 100},
    ).json()
    run_id = plan["runId"]

    def fake_materialize(target_run_id, filename):
        target = app_module.get_payroll_run_dir(target_run_id) / filename
        target.write_bytes(b"short")
        return target

    monkeypatch.setattr(app_module, "materialize_payroll_file", fake_materialize)
    monkeypatch.setattr(
        app_module,
        "_run_payroll_calculation",
        lambda *args, **kwargs: pytest.fail("size mismatch must not start calculation"),
    )

    response = client.post(
        f"/api/domestic-labor/runs/{run_id}/direct-upload-complete",
        json={"engines": "quanqinjiang", "attendanceMonth": "202606"},
    )

    assert response.status_code == 400
    assert "大小不一致" in response.json()["detail"]
    client.delete(f"/api/domestic-labor/runs/{run_id}")



def test_create_run_supports_split_monthly_daily_and_housing_workbooks():
    """拆分为三个Excel时应自动识别、合并并完成外宿补贴核算。"""
    client = TestClient(app)
    monthly, daily, housing = _split_waisu_butie_files()

    response = client.post(
        "/api/domestic-labor/runs",
        files=[
            ("files", ("月考勤.xlsx", monthly, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("日考勤.xlsx", daily, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("住宿名单.xlsx", housing, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
        data={"engines": "waisu_butie", "attendance_month": "202606"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["input_summary"] == {
        "file_count": 3,
        "monthly_rows": 1,
        "daily_rows": 1,
        "housing_rows": 1,
        "present_types": ["daily", "housing", "monthly"],
        "sources": payload["input_summary"]["sources"],
    }
    assert [source["sheets"][0]["type"] for source in payload["input_summary"]["sources"]] == [
        "monthly", "daily", "housing",
    ]

    import time
    for _ in range(20):
        time.sleep(0.25)
        metadata = client.get(f"/api/domestic-labor/runs/{payload['run_id']}").json()
        if metadata["status"] in ["已完成", "失败"]:
            break

    assert metadata["status"] == "已完成"
    assert metadata["summary"]["total_waisu_butie"] == 100
    assert metadata["fileNames"] == ["月考勤.xlsx", "日考勤.xlsx", "住宿名单.xlsx"]
    client.delete(f"/api/domestic-labor/runs/{payload['run_id']}")


def test_create_run_supports_split_monthly_and_daily_canbu_workbooks():
    """餐补拆分月考勤和日考勤后应自动合并核算。"""
    client = TestClient(app)
    monthly, daily = _split_canbu_files()

    response = client.post(
        "/api/domestic-labor/runs",
        files=[
            ("files", ("月考勤.xlsx", monthly, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("日考勤.xlsx", daily, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
        data={"engines": "canbu", "attendance_month": "202606"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["input_summary"]["file_count"] == 2
    assert payload["input_summary"]["monthly_rows"] == 1
    assert payload["input_summary"]["daily_rows"] == 1

    import time
    for _ in range(20):
        time.sleep(0.25)
        metadata = client.get(f"/api/domestic-labor/runs/{payload['run_id']}").json()
        if metadata["status"] in ["已完成", "失败"]:
            break

    assert metadata["status"] == "已完成"
    assert metadata["summary"]["total_canbu"] == 19
    client.delete(f"/api/domestic-labor/runs/{payload['run_id']}")


def test_waisu_split_upload_rejects_missing_housing_workbook():
    """外宿补贴缺少住宿名单时应在计算前明确阻断。"""
    client = TestClient(app)
    monthly, daily, _ = _split_waisu_butie_files()

    response = client.post(
        "/api/domestic-labor/runs",
        files=[
            ("files", ("月考勤.xlsx", monthly, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("日考勤.xlsx", daily, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
        data={"engines": "waisu_butie", "attendance_month": "202606"},
    )

    assert response.status_code == 400
    assert "缺少住宿名单" in response.json()["detail"]




def test_create_run_without_engines():
    """测试创建计算任务 - 无引擎"""
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "", "attendance_month": "202606"},
    )

    assert response.status_code == 400
    assert "请至少选择一个计算引擎" in response.json()["detail"]


def test_create_run_with_invalid_engine():
    """测试创建计算任务 - 无效引擎"""
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "invalid_engine", "attendance_month": "202606"},
    )

    assert response.status_code == 400
    assert "未知引擎" in response.json()["detail"]


def test_create_run_with_invalid_file():
    """测试创建计算任务 - 无效文件类型"""
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.txt", b"not an excel", "text/plain")},
        data={"engines": "quanqinjiang", "attendance_month": "202606"},
    )

    assert response.status_code == 400
    assert "请上传 Excel 文件" in response.json()["detail"]


def test_get_run_status():
    """测试获取任务状态"""
    client = TestClient(app)

    # Create run
    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "quanqinjiang", "attendance_month": "202606"},
    )
    run_id = create_response.json()["run_id"]

    # Get status
    response = client.get(f"/api/domestic-labor/runs/{run_id}")
    assert response.status_code == 200
    data = response.json()
    assert "status" in data
    assert data["status"] in ["已上传", "计算中", "已完成", "失败"]

    # Cleanup
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_get_run_compact_status_excludes_large_results():
    """轮询状态不应重复传输完整员工结果。"""
    client = TestClient(app)
    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "quanqinjiang", "attendance_month": "202606"},
    )
    run_id = create_response.json()["run_id"]

    import time
    for _ in range(20):
        time.sleep(0.1)
        response = client.get(f"/api/domestic-labor/runs/{run_id}?response_mode=status")
        if response.json()["status"] in ["已完成", "失败"]:
            break

    assert response.status_code == 200
    assert response.json()["id"] == run_id
    assert "results" not in response.json()
    assert "filePath" not in response.json()

    results_response = client.get(f"/api/domestic-labor/runs/{run_id}/results")
    assert results_response.status_code == 200
    assert "results" in results_response.json()
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_get_run_not_found():
    """测试获取不存在的任务"""
    client = TestClient(app)
    response = client.get("/api/domestic-labor/runs/nonexistent_id")

    assert response.status_code == 404


def test_list_runs():
    """测试任务列表"""
    client = TestClient(app)
    response = client.get("/api/domestic-labor/runs")

    assert response.status_code == 200
    data = response.json()
    assert "runs" in data
    assert isinstance(data["runs"], list)
    assert all("results" not in run for run in data["runs"])
    assert all("filePath" not in run for run in data["runs"])


def test_delete_run():
    """测试删除任务"""
    client = TestClient(app)

    # Create run
    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _quanqinjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "quanqinjiang", "attendance_month": "202606"},
    )
    run_id = create_response.json()["run_id"]

    # Delete
    response = client.delete(f"/api/domestic-labor/runs/{run_id}")
    assert response.status_code == 200
    assert "已删除" in response.json()["message"]

    # Verify deleted
    get_response = client.get(f"/api/domestic-labor/runs/{run_id}")
    assert get_response.status_code == 404


def test_full_workflow():
    """完整工作流测试：创建 → 等待 → 结果 → 导出 → 下载"""
    client = TestClient(app)

    # 1. 创建任务
    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("test.xlsx", _multi_engine_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "quanqinjiang,canbu", "attendance_month": "202606"},
    )
    assert create_response.status_code == 200
    run_id = create_response.json()["run_id"]

    # 2. 等待计算完成 (轮询)
    import time
    for _ in range(20):
        time.sleep(0.5)
        status_response = client.get(f"/api/domestic-labor/runs/{run_id}")
        status = status_response.json()["status"]
        if status in ["已完成", "失败"]:
            break

    assert status == "已完成"

    # 3. 获取结果
    results_response = client.get(f"/api/domestic-labor/runs/{run_id}/results")
    assert results_response.status_code == 200
    results = results_response.json()
    assert results["status"] == "已完成"
    assert len(results["results"]) == 1
    assert results["results"][0]["employee_id"] == "OWHN001"

    # 4. 导出 Excel
    export_response = client.get(f"/api/domestic-labor/runs/{run_id}/export")
    assert export_response.status_code == 200
    file_name = export_response.json()["file_name"]
    assert re.fullmatch(r"多科目核算结果_202606_\d{8}\.xlsx", file_name)

    # 5. 下载文件
    download_response = client.get(f"/api/domestic-labor/runs/{run_id}/download/{file_name}")
    assert download_response.status_code == 200
    assert len(download_response.content) > 0

    # Cleanup
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_canbu_api_ignores_blank_employee_rows():
    """空白行或工号为 None 的行不应进入餐补核算结果"""
    client = TestClient(app)

    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("canbu_blank.xlsx", _canbu_data_with_blank_rows(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "canbu", "attendance_month": "202605"},
    )
    assert create_response.status_code == 200
    run_id = create_response.json()["run_id"]

    import time
    for _ in range(20):
        time.sleep(0.5)
        status_response = client.get(f"/api/domestic-labor/runs/{run_id}")
        metadata = status_response.json()
        if metadata["status"] in ["已完成", "失败"]:
            break

    assert metadata["status"] == "已完成"
    assert [row["employee_id"] for row in metadata["results"]] == ["OWHN001"]
    assert metadata["summary"]["total_employees"] == 1

    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_canbu_export_outputs_business_reconciliation_sheet(tmp_path):
    """餐补导出只保留人员信息、适用规则、逐日餐补和月合计。"""
    output_path = tmp_path / "canbu_export.xlsx"
    exporter = ExcelExporter(str(output_path))
    results = [
        {
            "employee_id": "OWHN001",
            "employee_name": "张三",
            "department": "中国操作部",
            "canbu": 500,
            "total": 500,
            "warnings": "触发封顶: 累计551.00元 > 500.0元",
            "exceptions": [],
            "subject_details": {
                "canbu": {
                    "amount": 500,
                    "details": {
                        "地区规则": "东莞",
                        "日餐补明细": [19, 19, 9.5],
                        "月累计": 551,
                        "封顶金额": 500,
                        "是否触发封顶": True,
                    },
                    "audit_explanation": {
                        "rule_name": "东莞餐补逐日折算与封顶",
                        "formula": "min(Σ单日餐补, 500)",
                        "inputs": {
                            "工作地区": "东莞",
                            "岗位名称": "操作员",
                            "日考勤记录数": 31,
                        },
                        "intermediate_values": {
                            "日标准": 19,
                            "日餐补合计": 551,
                            "最终金额": 500,
                        },
                        "steps": [
                            "按日考勤正班时数和刷卡加班取较大值折算",
                            "最终餐补=min(551, 500)=500",
                        ],
                    },
                }
            },
        },
        {
            "employee_id": "None",
            "employee_name": "None",
            "department": "None",
            "canbu": 0,
            "total": 0,
            "subject_details": {},
        },
    ]

    exporter.export(results, "202605", {"total_employees": 2})

    wb = load_workbook(output_path)
    ws = wb["计算详情"]
    headers = [cell.value for cell in ws[1]]
    assert headers == [
        "工号", "姓名", "工作地区", "部门", "岗位", "餐补口径",
        "01日餐补", "02日餐补", "03日餐补", "餐补合计",
    ]
    assert wb.sheetnames == ["计算详情"]
    assert ws.max_row == 2
    row = [cell.value for cell in ws[2]]
    assert row[0] == "OWHN001"
    assert row[2] == "东莞"
    assert row[4] == "操作员"
    assert row[5] == "东莞"
    assert row[6:9] == [19, 19, 9.5]
    assert row[9] == 500
    wb.close()


def test_waisu_butie_export_outputs_business_reconciliation_sheet(tmp_path):
    """外宿补贴单科目导出包含住宿、缺勤和补贴天数审计字段。"""
    output_path = tmp_path / "waisu_export.xlsx"
    exporter = ExcelExporter(str(output_path))
    results = [{
        "employee_id": "OWHN001",
        "employee_name": "张三",
        "department": "华东操作",
        "position": "操作员",
        "waisu_butie": 100,
        "total": 100,
        "warnings": "",
        "exceptions": [],
        "subject_details": {
            "waisu_butie": {
                "amount": 100,
                "details": {
                    "在职天数": 30,
                    "住宿扣除天数": 10,
                    "外宿补贴天数": 20,
                    "缺勤时数": 0,
                    "补贴标准": 150,
                },
                "audit_explanation": {
                    "rule_name": "嘉善外宿补贴住宿与缺勤折算",
                    "formula": "补贴标准/月天数 × 有效补贴天数",
                    "inputs": {"工作地区": "嘉善", "岗位名称": "操作员"},
                    "intermediate_values": {"在职天数": 30, "住宿扣除天数": 10, "外宿补贴天数": 20, "缺勤时数": 0, "补贴标准": 150},
                    "steps": ["住宿名单扣除10天", "最终外宿补贴为100"],
                },
            }
        },
    }, {
        "employee_id": "OWHN002",
        "employee_name": "李四",
        "department": "华东操作",
        "position": "保洁",
        "waisu_butie": 0,
        "total": 0,
        "warnings": "",
        "exceptions": [],
        "subject_details": {
            "waisu_butie": {
                "amount": 0,
                "details": {"reason": "嘉善外宿补贴资格不满足"},
                "audit_explanation": {
                    "rule_name": "嘉善外宿补贴资格判断",
                    "inputs": {"工作地区": "嘉善", "岗位名称": "保洁"},
                },
            }
        },
    }]

    exporter.export(results, "202606", {"total_employees": 1})

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["计算详情"]
    ws = wb["计算详情"]
    assert [cell.value for cell in ws[1]] == [
        "工号", "姓名", "工作地区", "部门", "岗位", "外宿补贴口径",
        "在职天数", "住宿扣除天数", "外宿补贴天数", "缺勤时数",
        "补贴标准", "应发外宿补贴", "异常/提示",
    ]
    assert [cell.value for cell in ws[2]][:12] == [
        "OWHN001", "张三", "嘉善", "华东操作", "操作员", "嘉善外宿补贴",
        30, 10, 20, 0, 150, 100,
    ]
    assert ws.cell(3, 6).value == "嘉善外宿补贴"
    assert ws.cell(3, 13).value == "嘉善外宿补贴资格不满足"
    wb.close()


def test_waisu_butie_full_api_workflow_exposes_audit_and_export():
    """外宿补贴完整流程：上传、计算、审计、导出和下载。"""
    client = TestClient(app)
    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("waisu.xlsx", _waisu_butie_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "waisu_butie", "attendance_month": "202606"},
    )
    assert create_response.status_code == 200
    run_id = create_response.json()["run_id"]

    import time
    for _ in range(20):
        time.sleep(0.25)
        metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
        if metadata["status"] in ["已完成", "失败"]:
            break

    assert metadata["status"] == "已完成"
    assert metadata["summary"]["total_waisu_butie"] == 100
    row = metadata["results"][0]
    assert row["waisu_butie"] == 100
    assert row["subject_details"]["waisu_butie"]["details"]["外宿补贴天数"] == 20
    assert row["subject_details"]["waisu_butie"]["details"]["补贴标准"] == 150
    assert row["subject_details"]["waisu_butie"]["audit_explanation"]["rule_name"] == "嘉善外宿补贴住宿与缺勤折算"

    export_response = client.get(f"/api/domestic-labor/runs/{run_id}/export")
    assert export_response.status_code == 200
    file_name = export_response.json()["file_name"]
    assert re.fullmatch(r"外宿补贴核算结果_202606_\d{8}\.xlsx", file_name)
    download_response = client.get(f"/api/domestic-labor/runs/{run_id}/download/{file_name}")
    assert download_response.status_code == 200
    wb = load_workbook(BytesIO(download_response.content))
    assert wb.sheetnames == ["计算详情"]
    assert wb["计算详情"]["L2"].value == 100
    wb.close()

    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_gonglingjiang_api_exposes_subject_details_and_audit_explanation():
    """API 结果暴露引擎详情，供前端解释抽屉使用"""
    client = TestClient(app)

    create_response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("gongling.xlsx", _gonglingjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "engines": "gonglingjiang",
            "attendance_month": "202606",
            "hrbp_list": '[{"employee_id":"OWHN001","employee_name":"张三"}]',
        },
    )
    assert create_response.status_code == 200
    run_id = create_response.json()["run_id"]

    import time
    for _ in range(20):
        time.sleep(0.5)
        status_response = client.get(f"/api/domestic-labor/runs/{run_id}")
        metadata = status_response.json()
        if metadata["status"] in ["已完成", "失败"]:
            break

    assert metadata["status"] == "已完成"
    assert metadata["collectionSeniorityRoster"] == [{"employee_id": "OWHN001", "employee_name": "张三"}]
    assert metadata["collectionSeniorityRosterCount"] == 1
    assert metadata["inputSummary"]["requires_collection_seniority_roster"] is False
    row = metadata["results"][0]
    subject_detail = row["subject_details"]["gonglingjiang"]
    explanation = subject_detail["audit_explanation"]

    assert row["gonglingjiang"] == 450
    assert subject_detail["amount"] == 450
    assert subject_detail["exceptions"] == []
    assert explanation["subject"] == "gonglingjiang"
    assert explanation["intermediate_values"]["工龄(年)"] == 3
    assert row["exceptions"] == []

    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_gonglingjiang_api_rejects_invalid_hrbp_list():
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("gongling.xlsx", _gonglingjiang_data(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "engines": "gonglingjiang",
            "attendance_month": "202606",
            "hrbp_list": '{"OWHN001": true}',
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "揽收线工龄奖名单必须为人员列表"


def test_gonglingjiang_api_requires_named_roster_for_fourth_column_in_any_region():
    client = TestClient(app)

    missing_name = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("collection.xlsx", _gonglingjiang_collection_data("深圳"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "engines": "gonglingjiang",
            "attendance_month": "202606",
            "hrbp_list": '[{"employee_id":"OWHN001","employee_name":""}]',
        },
    )

    assert missing_name.status_code == 400
    assert missing_name.json()["detail"] == "已识别到第四纵队，请维护包含工号和姓名的揽收线工龄奖名单"

    accepted = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("collection.xlsx", _gonglingjiang_collection_data("深圳"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "engines": "gonglingjiang",
            "attendance_month": "202606",
            "hrbp_list": '[{"employee_id":"OWHN001","employee_name":"张三"}]',
        },
    )
    assert accepted.status_code == 200
    run_id = accepted.json()["run_id"]

    import time
    for _ in range(20):
        time.sleep(0.25)
        metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
        if metadata["status"] in ["已完成", "失败"]:
            break

    assert metadata["status"] == "已完成"
    assert metadata["inputSummary"]["requires_collection_seniority_roster"] is True
    assert metadata["inputSummary"]["collection_seniority_employee_count"] == 1
    assert metadata["results"][0]["gonglingjiang"] == 450
    client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_quanqinjiang_returns_audit_explanation():
    """全勤奖返回计算公式和审计解释"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "考勤月份": "202606",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "旷工天数": 0,
        "正班迟到次数": 0,
        "早退次数": 0,
        "签卡次数": 0,
        "工伤假天数": 0,
        "事假时数": 0,
        "病假时数": 0,
        "入离职缺勤时数": 0,
        "迟到早退30分钟内扣款": 0,
    }

    result = QuanQinJiangEngine().calculate(employee)
    explanation = result.details["audit_explanation"]

    assert result.amount == 100
    assert explanation["subject"] == "quanqinjiang"
    assert explanation["formula"] == "满足全勤条件 = 100"
    assert explanation["intermediate_values"]["旷工天数"] == 0
    assert explanation["steps"]


def test_canbu_dongguan_uses_platform_rule_without_meal_standard():
    """东莞餐补使用平台规则，不依赖月报餐补标准"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "",
        "一级部门名称": "莞深操作",
        "岗位名称": "操作员",
        "餐补标准": "",
    }
    daily_attendance = [
        {"工号": "OWHN001", "工作地区": "东莞", "工作状态": "工作日", "正班时数": 8, "刷卡加班": 0},
        {"工号": "OWHN001", "工作地区": "东莞", "工作状态": "工作日", "正班时数": 4, "刷卡加班": 0},
        {"工号": "OWHN001", "工作地区": "东莞", "工作状态": "星期天休息", "正班时数": 0, "刷卡加班": 3},
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)
    explanation = result.details["audit_explanation"]

    assert result.amount == 35.62
    assert explanation["subject"] == "canbu"
    assert explanation["formula"] == "min(Σ单日餐补, 500)"
    assert explanation["inputs"]["工作地区"] == "东莞"
    assert explanation["intermediate_values"]["日餐补合计"] == 35.62
    assert explanation["steps"]


def test_canbu_dongguan_uses_max_regular_hours_and_overtime():
    """东莞按正班时数和刷卡加班较大值计算餐补"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "一级部门名称": "莞深操作",
        "岗位名称": "操作员",
    }
    daily_attendance = [
        {"工号": "OWHN001", "工作地区": "东莞", "工作状态": "工作日", "正班时数": 3, "刷卡加班": 4},
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)

    assert result.amount == 9.5
    assert result.details["日餐补明细"] == [9.5]


def test_canbu_dongguan_piecework_tally_is_not_excluded():
    """东莞不再按理货操作组计件单独排除餐补"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "一级部门名称": "莞深操作",
        "岗位名称": "操作员",
    }
    daily_attendance = [
        {
            "工号": "OWHN001",
            "工作地区": "东莞",
            "工作状态": "工作日",
            "正班时数": 8,
            "刷卡加班": 0,
            "四级部门名称": "理货操作组",
            "计时": "计件",
        },
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)

    assert result.amount == 19
    assert result.details["日餐补明细"] == [19]


def test_canbu_dongguan_abnormal_absenteeism_formats_are_excluded():
    """东莞旷工异常兼容数字异常标记和复合异常原因"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "一级部门名称": "莞深操作",
        "岗位名称": "操作员",
    }
    daily_attendance = [
        {
            "工号": "OWHN001",
            "工作地区": "东莞",
            "工作状态": "工作日",
            "是否异常": "是",
            "异常原因": "旷工",
            "正班时数": 8,
            "刷卡加班": 0,
        },
        {
            "工号": "OWHN001",
            "工作地区": "东莞",
            "工作状态": "工作日",
            "是否异常": 1,
            "异常原因": "早退;旷工",
            "正班时数": 8,
            "刷卡加班": 0,
        },
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)

    assert result.amount == 0
    assert result.details["日餐补明细"] == [0, 0]


def test_canbu_dongguan_operation_clerk_is_eligible():
    """东莞操作文员享有餐补，普通文员不享有"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "一级部门名称": "莞深操作",
        "岗位名称": "操作文员",
    }
    daily_attendance = [
        {"工号": "OWHN001", "工作地区": "东莞", "工作状态": "工作日", "正班时数": 8, "刷卡加班": 0},
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)
    explanation = result.details["audit_explanation"]

    assert result.amount == 19
    assert explanation["rule_name"] == "东莞餐补逐日折算与封顶"


def test_canbu_dongguan_explicit_ineligible_position_is_not_eligible():
    """东莞不享有岗位按线下口径排除"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "一级部门名称": "莞深操作",
        "岗位名称": "保洁",
    }
    daily_attendance = [
        {"工号": "OWHN001", "工作地区": "东莞", "工作状态": "工作日", "正班时数": 8, "刷卡加班": 0},
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)
    explanation = result.details["audit_explanation"]

    assert result.amount == 0
    assert result.details["reason"] == "东莞餐补资格不满足"
    assert explanation["intermediate_values"]["岗位是否明确不享有"] is True


def test_canbu_jinjiang_is_not_eligible():
    """晋江区域不享有餐补"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "晋江",
        "一级部门名称": "东南区",
        "岗位名称": "操作员",
    }

    result = CanBuEngine().calculate(employee, daily_attendance=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 0
    assert result.details["reason"] == "晋江区域不享有餐补"
    assert explanation["formula"] == "晋江区域 = 0"


def test_canbu_jiashan_uses_monthly_attendance_formula():
    """嘉善按月报字段和线下公式计算餐补"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "嘉善",
        "岗位名称": "操作员",
        "排班天数": 20,
        "实际在职工作日天数": 20,
        "事假时数": 8,
        "病假时数": 16,
        "旷工天数": 1,
    }

    result = CanBuEngine().calculate(employee, daily_attendance=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 258
    assert explanation["formula"] == "300/排班天数×(实际在职工作日天数-事假天数-旷工天数-病假天数×0.4)"
    assert explanation["intermediate_values"]["事假天数"] == 1
    assert explanation["intermediate_values"]["病假天数"] == 2
    assert explanation["intermediate_values"]["有效餐补天数"] == 17.2


def test_canbu_audit_department_fields_are_deduplicated():
    """计算解释中的部门层级不应按每日考勤重复拼接。"""
    employee = {
        "工号": "OWHD2511",
        "姓名": "周月明",
        "工作地区": "嘉善",
        "一级部门名称": "华东区",
        "二级部门名称": "华东枢纽",
        "三级部门名称": "后勤保障组",
        "岗位名称": "设备维修专员",
        "排班天数": 20,
        "实际在职工作日天数": 20,
    }
    daily_attendance = [
        {
            "工号": "OWHD2511",
            "工作地区": "嘉善",
            "一级部门名称": "华东区",
            "二级部门名称": "华东枢纽",
            "三级部门名称": "后勤保障组",
        }
        for _ in range(3)
    ]

    result = CanBuEngine().calculate(employee, daily_attendance)

    assert result.details["audit_explanation"]["inputs"]["部门字段"] == "华东区 华东枢纽 后勤保障组"


def test_canbu_yiwu_uses_jiashan_monthly_attendance_formula():
    """义乌与嘉善使用同一套餐补月报公式"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "义乌",
        "岗位名称": "操作员",
        "排班天数": 22,
        "实际在职工作日天数": 22,
        "事假时数": 8,
        "病假时数": 0,
        "旷工天数": 0,
    }

    result = CanBuEngine().calculate(employee, daily_attendance=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 286.36
    assert result.details["地区规则"] == "义乌"
    assert explanation["rule_name"] == "义乌餐补月报折算"
    assert explanation["intermediate_values"]["有效餐补天数"] == 21


def test_canbu_jiashan_yiwu_include_cleaner_and_maintenance_alias():
    """嘉善/义乌保洁享有餐补，设备维护与设备维修岗位名称均兼容。"""
    base_employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "排班天数": 24,
        "实际在职工作日天数": 24,
        "事假时数": 0,
        "病假时数": 0,
        "旷工天数": 0,
    }

    cleaner = {**base_employee, "工作地区": "嘉善", "岗位名称": "保洁"}
    maintenance = {**base_employee, "工作地区": "义乌", "岗位名称": "设备维护员"}
    repair_specialist = {**base_employee, "工作地区": "嘉善", "岗位名称": "设备维修专员"}

    cleaner_result = CanBuEngine().calculate(cleaner, daily_attendance=[])
    maintenance_result = CanBuEngine().calculate(maintenance, daily_attendance=[])
    repair_specialist_result = CanBuEngine().calculate(repair_specialist, daily_attendance=[])

    assert cleaner_result.amount == 300
    assert cleaner_result.details["地区规则"] == "嘉善"
    assert maintenance_result.amount == 300
    assert maintenance_result.details["地区规则"] == "义乌"
    assert repair_specialist_result.amount == 300
    assert repair_specialist_result.details["地区规则"] == "嘉善"


def test_canbu_jiashan_unknown_position_is_not_eligible():
    """嘉善未在享有名单的岗位兜底不享有"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "嘉善",
        "岗位名称": "财务专员",
        "排班天数": 20,
        "实际在职工作日天数": 20,
    }

    result = CanBuEngine().calculate(employee, daily_attendance=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 0
    assert result.details["reason"] == "嘉善餐补资格不满足"
    assert explanation["intermediate_values"]["岗位是否在享有名单"] is False


def test_waisu_butie_returns_audit_explanation():
    """外宿补贴返回按天折算公式"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "考勤月份": "202606",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "事假时数": 0,
        "排休请假时数": 0,
        "病假时数": 0,
        "旷工时数": 0,
        "入离职缺勤时数": 0,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 150
    assert explanation["subject"] == "waisu_butie"
    assert explanation["formula"] == "补贴标准/月天数 × 有效补贴天数"
    assert explanation["intermediate_values"]["外宿补贴天数"] == 30
    assert explanation["steps"]


def test_waisu_butie_dongguan_ineligible_position_gets_zero():
    """东莞外宿补贴按岗位不享有名单排除"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "保洁",
        "考勤月份": "202606",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 0
    assert result.details["reason"] == "东莞外宿补贴资格不满足"
    assert result.details["audit_explanation"]["intermediate_values"]["岗位是否明确不享有"] is True


def test_waisu_butie_dongguan_absence_excludes_annual_leave_hours():
    """东莞缺勤满56小时不包含休年假小时"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "考勤月份": "202606",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "排班天数": 26,
        "实际在职工作日天数": 26,
        "事假时数": 16,
        "排休请假时数": 16,
        "病假时数": 8,
        "旷工时数": 0,
        "休年假小时": 16,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 150
    assert result.details["缺勤时数"] == 40
    assert explanation["rule_name"] == "东莞外宿补贴住宿与缺勤折算"
    assert explanation["intermediate_values"]["休年假小时"] == 16


def test_waisu_butie_dongguan_next_month_exit_still_prorates_current_month_absence():
    """次月离职不影响本月全月在职判断，东莞缺勤仍需折算。"""
    employee = {
        "工号": "OWHN12750",
        "姓名": "黄海明",
        "工作地区": "东莞",
        "岗位名称": "揽收充电司机",
        "考勤月份": "202606",
        "入职日期": date(2025, 8, 22),
        "最后工作日": date(2026, 7, 7),
        "事假时数": 8,
        "排休请假时数": 0,
        "病假时数": 80,
        "旷工时数": 0,
        "入离职缺勤时数": 0,
    }
    daily_attendance = [{"工号": "OWHN12750", "上班一": "22:00", "下班一": "06:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 95
    assert result.details["在职天数"] == 30
    assert result.details["全月在职"] is True
    assert result.details["缺勤时数"] == 88


def test_waisu_butie_dongguan_slash_standard_does_not_block_rule_calculation():
    """外宿补贴标准为/时不再作为资格依据"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "考勤月份": "202606",
        "外宿补贴标准": "/",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 150


def test_waisu_butie_ignores_uploaded_standard_value():
    """上传的外宿补贴标准不参与资格或金额计算。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "考勤月份": "202606",
        "外宿补贴标准": "999",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 150
    assert result.details["补贴标准"] == 150


def test_waisu_butie_jiashan_mid_month_checkout_keeps_checkout_day_external():
    """嘉善退宿当天开始享有外宿补贴。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "嘉善",
        "岗位名称": "操作员",
        "考勤月份": "202606",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]
    housing_records = [{"工号": "OWHN001", "入住时间": date(2026, 5, 1), "退宿时间": date(2026, 6, 11)}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records)

    assert result.amount == 100
    assert result.details["外宿补贴天数"] == 20
    assert result.details["audit_explanation"]["rule_name"] == "嘉善外宿补贴住宿与缺勤折算"


def test_waisu_butie_yiwu_mid_month_checkout_keeps_checkout_day_external():
    """义乌沿用原规则，退宿当天开始享有外宿补贴。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "义乌",
        "岗位名称": "仓库文员",
        "考勤月份": "202606",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]
    housing_records = [{"工号": "OWHN001", "入住时间": date(2026, 5, 1), "退宿时间": date(2026, 6, 11)}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records)

    assert result.amount == 100
    assert result.details["外宿补贴天数"] == 20


def test_waisu_butie_jiashan_absence_includes_annual_leave_and_weights_sick_leave():
    """嘉善请假总时数包含年假小时，病假小时按60%计入。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "嘉善",
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "排班天数": 22,
        "实际在职工作日天数": 22,
        "休年假小时": 8,
        "病假时数": 80,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.details["缺勤时数"] == 56
    assert result.amount == 116.13


@pytest.mark.parametrize("work_area", ["嘉善", "义乌"])
def test_waisu_butie_jiashan_yiwu_absence_includes_paid_leave_types(work_area):
    """嘉善/义乌请假总时数包含年假、婚假、陪产假和工伤假。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": work_area,
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "排班天数": 22,
        "实际在职工作日天数": 22,
        "休年假小时": 8,
        "婚假天数": 2,
        "陪产假天数": 2,
        "工伤假天数": 2,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.details["缺勤时数"] == 56
    assert result.amount == 116.13


@pytest.mark.parametrize("position", ["数据专员", "保洁", "操作文员", "设备维养专员"])
def test_waisu_butie_jiashan_confirmed_positions_are_eligible(position):
    """薪酬确认的嘉善岗位名称均享有外宿补贴。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "嘉善",
        "岗位名称": position,
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 150


def test_waisu_butie_jinjiang_confirmed_safety_officer_is_eligible():
    """薪酬确认晋江安全员享有外宿补贴。"""
    employee = {
        "工号": "OWDN0255",
        "姓名": "李磊",
        "工作地区": "晋江",
        "岗位名称": "安全员",
        "考勤月份": "202606",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "排班天数": 22,
        "实际在职工作日天数": 22,
    }

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance=[], housing_records=[])

    assert result.amount == 150


@pytest.mark.parametrize(
    ("work_area", "position"),
    [("东莞", "操作员"), ("嘉善", "操作员"), ("义乌", "操作员"), ("晋江", "操作员")],
)
@pytest.mark.parametrize("attendance_days", [0, 1])
def test_waisu_butie_all_regions_at_most_one_attendance_day_with_absence_gets_zero(
    work_area, position, attendance_days
):
    """各地区正班出勤不超过1天且存在旷工时不发外宿补贴。"""
    employee = {
        "工号": "OWHD8098",
        "姓名": "梁林",
        "工作地区": work_area,
        "岗位名称": position,
        "考勤月份": "202606",
        "入职日期": date(2026, 6, 27),
        "最后工作日": None,
        "正班出勤天数": attendance_days,
        "旷工天数": 1,
    }
    housing_records = [{
        "工号": "OWHD8098",
        "入住时间": date(2026, 6, 26),
        "退宿时间": date(2026, 6, 30),
    }]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance=[], housing_records=housing_records)

    assert result.amount == 0
    assert result.details["reason"] == "正班出勤不超过1天且旷工至少1天"


@pytest.mark.parametrize(("employee_id", "attendance_days"), [("OWHD8092", 0), ("OWHD8021", 0.38)])
def test_waisu_butie_half_day_absence_does_not_trigger_low_attendance_zero_rule(
    employee_id, attendance_days
):
    """低出勤但仅旷工0.5天时仍按正常外宿补贴规则计算。"""
    employee = {
        "工号": employee_id,
        "姓名": "测试员工",
        "工作地区": "嘉善",
        "岗位名称": "操作员",
        "考勤月份": "202606",
        "入职日期": date(2026, 6, 30),
        "最后工作日": None,
        "正班出勤天数": attendance_days,
        "旷工天数": 0.5,
    }

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance=[], housing_records=[])

    assert result.amount == 5


def test_waisu_butie_jiashan_mid_month_exit_still_deducts_absence_over_56_hours():
    """嘉善当月离职人员也按请假总时数满56小时扣减。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "嘉善",
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": date(2026, 5, 18),
        "旷工天数": 9,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.details["缺勤时数"] == 72
    assert result.amount == 43.55


def test_waisu_butie_jinjiang_deducts_entry_exit_and_leave_days():
    """晋江外宿补贴按入离职自然日和请假旷工天数扣减"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "晋江",
        "岗位名称": "操作员",
        "考勤月份": "202606",
        "外宿补贴标准": "150",
        "入职日期": date(2026, 6, 6),
        "最后工作日": None,
        "事假时数": 24,
        "病假时数": 16,
        "旷工天数": 2,
        "排休请假天数": 1,
        "休年假小时": 16,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 75
    assert explanation["rule_name"] == "晋江外宿补贴月考勤扣减"
    assert explanation["intermediate_values"]["入离职缺勤自然日天数"] == 5
    assert explanation["intermediate_values"]["请假旷工天数"] == 10


def test_waisu_butie_jinjiang_same_month_hire_and_exit_uses_exit_days_only():
    """晋江同月入职又离职时线下只扣最后工作日后的自然日"""
    employee = {
        "工号": "OWDN0243",
        "姓名": "吴绍阳",
        "工作地区": "晋江",
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2026, 5, 5),
        "最后工作日": date(2026, 5, 15),
        "事假时数": 0,
        "病假时数": 0,
        "旷工天数": 0,
        "排休请假天数": 0,
        "休年假小时": 0,
    }
    daily_attendance = [{"工号": "OWDN0243", "上班一": "09:00", "下班一": "18:00"}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])
    explanation = result.details["audit_explanation"]

    assert result.amount == 72.58
    assert explanation["intermediate_values"]["入离职缺勤自然日天数"] == 16
    assert explanation["intermediate_values"]["入离职扣减"] == 77.42


def test_waisu_butie_mid_month_exit_without_punches_still_prorates():
    """月中离职无打卡时仍按线下入离职天数折算"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "监察员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2024, 7, 24),
        "最后工作日": date(2026, 5, 4),
        "排班天数": 22,
        "实际在职工作日天数": 3,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "", "下班一": ""} for _ in range(4)]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 19.35
    assert result.details["在职天数"] == 4


def test_waisu_butie_first_day_exit_without_punches_gets_zero():
    """当月首日离职且无打卡时线下不发外宿补贴"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2024, 10, 22),
        "最后工作日": date(2026, 5, 1),
        "排班天数": 22,
        "实际在职工作日天数": 0,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "", "下班一": ""}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records=[])

    assert result.amount == 0
    assert result.details["reason"] == "首日离职且无打卡"


def test_waisu_butie_checkout_on_last_workday_keeps_checkout_day_external():
    """退宿日等于最后工作日，当天仍享有外宿补贴。"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "/",
        "入职日期": date(2026, 5, 14),
        "最后工作日": date(2026, 5, 18),
        "排班天数": 22,
        "实际在职工作日天数": 3,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]
    housing_records = [{"工号": "OWHN001", "入住时间": date(2026, 5, 12), "退宿时间": date(2026, 5, 18)}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records)

    assert result.amount == 4.84
    assert result.details["外宿补贴天数"] == 1


def test_waisu_butie_active_housing_and_absence_over_56_gets_zero():
    """入住未退宿且缺勤满56小时按线下结果不发外宿补贴"""
    employee = {
        "工号": "OWHN001",
        "姓名": "张三",
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "考勤月份": "202605",
        "外宿补贴标准": "150",
        "入职日期": date(2023, 1, 1),
        "最后工作日": None,
        "排班天数": 22,
        "实际在职工作日天数": 22,
        "病假时数": 64,
    }
    daily_attendance = [{"工号": "OWHN001", "上班一": "09:00", "下班一": "18:00"}]
    housing_records = [{"工号": "OWHN001", "入住时间": date(2026, 5, 13), "退宿时间": None}]

    result = WaiSuBuTieEngine().calculate(employee, daily_attendance, housing_records)

    assert result.amount == 0
    assert result.details["reason"] == "在宿且缺勤满56小时"


def test_gonglingjiang_fourth_column_collection_uses_hrbp_list():
    employee = {**_gongling_employee(), "二级部门名称": "第四纵队", "岗位名称": "内勤专员"}

    result = GongLingJiangEngine().calculate(employee, hrbp_list=["OWHN001"])

    assert result.amount == 450
    assert result.details["部门类别"] == "揽收"
    assert result.details["标准"] == 150
    assert result.details["上限"] == 600


def test_gonglingjiang_fourth_column_collection_requires_hrbp_list():
    employee = {**_gongling_employee(), "二级部门名称": "第四纵队", "岗位名称": "内勤专员"}

    result = GongLingJiangEngine().calculate(employee, hrbp_list=[])

    assert result.amount == 0
    assert result.details["exceptions"][0]["code"] == "MISSING_HRBP_LIST"


@pytest.mark.parametrize(
    ("employee_id", "employee_name", "work_area"),
    [
        ("OWHN11388", "夏雷", "深圳"),
        ("OWHN6172", "邓军洋", "深圳"),
        ("OWHN2248", "赖志强", "惠州"),
        ("OWHN0474", "曾威", "深圳"),
    ],
)
def test_gonglingjiang_fourth_column_roster_applies_in_any_region(employee_id, employee_name, work_area):
    employee = {
        **_gongling_employee(),
        "工号": employee_id,
        "姓名": employee_name,
        "工作地区": work_area,
        "二级部门名称": "第四纵队",
        "岗位名称": "内勤专员",
    }

    result = GongLingJiangEngine().calculate(employee, hrbp_list=[employee_id])

    assert result.amount == 450
    assert result.details["部门类别"] == "揽收"
    assert result.details["标准"] == 150
    assert result.details["上限"] == 600


def test_gonglingjiang_fourth_column_group_leader_remains_ineligible_outside_dongguan():
    employee = {
        **_gongling_employee(),
        "工作地区": "深圳",
        "二级部门名称": "第四纵队",
        "岗位名称": "操作组长",
    }

    result = GongLingJiangEngine().calculate(employee, hrbp_list=["OWHN001"])

    assert result.amount == 0


def test_gonglingjiang_headhaul_fbu_uses_fbu_rate_and_cap():
    employee = {**_gongling_employee(), "二级部门名称": "头程运营部"}

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 300
    assert result.details["部门类别"] == "FBU"
    assert result.details["标准"] == 100
    assert result.details["上限"] == 500


@pytest.mark.parametrize(
    ("employee_id", "employee_name", "work_area"),
    [
        ("SFXNLS014", "黄亚博", "宁波"),
        ("zt07308", "袁冲伟", "广州"),
    ],
)
def test_gonglingjiang_headhaul_fbu_applies_in_any_region(employee_id, employee_name, work_area):
    employee = {
        **_gongling_employee(),
        "工号": employee_id,
        "姓名": employee_name,
        "工作地区": work_area,
        "二级部门名称": "头程运营部",
    }

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 300
    assert result.details["部门类别"] == "FBU"
    assert result.details["标准"] == 100
    assert result.details["上限"] == 500


def test_gonglingjiang_headhaul_fbu_prorates_absence_without_56_hour_threshold():
    employee = {
        **_gongling_employee(),
        "二级部门名称": "头程运营部",
        "入职日期": date(2019, 8, 15),
        "考勤月份": "202602",
        "排班天数": 21,
        "实际在职工作日天数": 21,
        "事假时数": 37,
    }

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 389.88


def test_gonglingjiang_headhaul_fbu_does_not_count_rest_leave_as_absence():
    employee = {
        **_gongling_employee(),
        "二级部门名称": "头程运营部",
        "排休请假时数": 80,
    }

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 300


def test_gonglingjiang_headhaul_fbu_combines_reported_entry_exit_absence_once():
    employee = {
        **_gongling_employee(),
        "二级部门名称": "头程运营部",
        "入职日期": date(2024, 10, 1),
        "考勤月份": "202602",
        "排班天数": 20,
        "实际在职工作日天数": 20,
        "入离职缺勤时数": 72,
    }

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 55


def test_gonglingjiang_headhaul_fbu_uses_absenteeism_days_times_eight():
    employee = {
        **_gongling_employee(),
        "二级部门名称": "头程运营部",
        "旷工天数": 1,
        "旷工时数": 80,
    }

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 288.46


def test_gonglingjiang_headhaul_fbu_does_not_use_full_month_leave_override():
    employee = {
        **_gongling_employee(),
        "二级部门名称": "头程运营部",
        "排班天数": 20,
        "实际在职工作日天数": 20,
        "正班出勤天数": 0,
        "事假时数": 8,
    }

    result = GongLingJiangEngine().calculate(employee)

    assert result.amount == 285


@pytest.mark.parametrize(
    ("department", "position"),
    [
        ("华东枢纽", "操作员"),
        ("华东揽收组", "揽收操作员"),
        ("华东B2B枢纽", "操作员"),
        ("华西区操作部", "操作员"),
    ],
)
def test_gonglingjiang_east_west_departments_return_zero(department, position):
    employee = {
        **_gongling_employee(),
        "工作地区": "",
        "二级部门名称": department,
        "岗位名称": position,
    }

    result = GongLingJiangEngine().calculate(employee, region="wes")

    assert result.amount == 0
    assert result.details["reason"] == "华东/华西指定二级部门无工龄奖"
    assert result.details["audit_explanation"]["rule_name"] == "工龄奖二级部门判断"


@pytest.mark.parametrize(
    ("department", "position"),
    [
        ("东南枢纽", "操作员"),
        ("闽赣揽收组", "内勤专员"),
    ],
)
def test_gonglingjiang_keeps_southeast_compatibility_route(department, position):
    employee = {
        **_gongling_employee(),
        "工作地区": "",
        "二级部门名称": department,
        "岗位名称": position,
    }

    result = GongLingJiangEngine().calculate(employee, region="wes")

    assert result.amount == 150
    assert result.details["标准"] == 50
    assert result.details["上限"] == 150


def test_gonglingjiang_uses_raw_absence_fields_instead_of_leave_hours():
    """工龄奖缺勤折算不读取月报聚合的请假时数，按规则卡原始字段计算"""
    employee = {
        **_gongling_employee(),
        "请假时数": 80,
        "事假时数": 0,
        "病假时数": 0,
        "旷工天数": 0,
        "排休请假天数": 0,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=["OWHN001"], region="gsdg")

    assert result.amount == 450
    explanation = result.details["audit_explanation"]
    assert explanation["intermediate_values"]["事病旷排休时数"] == 0
    assert "未达到56小时门槛" in " ".join(explanation["steps"])


def test_gonglingjiang_absence_components_trigger_proration():
    """事假+病假+旷工天数×8+排休请假天数×8 达到56小时后按天折算"""
    employee = {
        **_gongling_employee(),
        "事假时数": 8,
        "病假时数": 8,
        "旷工天数": 1,
        "排休请假天数": 4,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=["OWHN001"], region="gsdg")

    assert result.amount == 328.85
    explanation = result.details["audit_explanation"]
    assert explanation["intermediate_values"]["事病旷排休时数"] == 56
    assert explanation["intermediate_values"]["旷工折算时数"] == 8
    assert explanation["intermediate_values"]["排休请假折算时数"] == 32
    assert "达到56小时门槛" in " ".join(explanation["steps"])


def test_gonglingjiang_prefers_offline_hour_fields_for_dongguan_proration():
    """东莞月报的旷工、排休小时字段按线下工资表口径直接参与56小时判断"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "排班天数": 22,
        "实际在职工作日天数": 22,
        "事假时数": 0,
        "病假时数": 0,
        "旷工天数": 0,
        "旷工时数": 0,
        "排休请假天数": 0,
        "排休请假时数": 80,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 245.45
    explanation = result.details["audit_explanation"]
    assert explanation["intermediate_values"]["事病旷排休时数"] == 80
    assert explanation["intermediate_values"]["排休字段口径"] == "排休请假时数"


def test_gonglingjiang_uses_absenteeism_hours_when_days_field_is_missing():
    """东莞仅提供旷工时数时不得漏算"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "排班天数": 22,
        "实际在职工作日天数": 22,
        "事假时数": 31,
        "病假时数": 0,
        "旷工天数": 0,
        "旷工时数": 68,
        "排休请假天数": 0,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 196.88
    explanation = result.details["audit_explanation"]
    assert explanation["intermediate_values"]["事病旷排休时数"] == 99
    assert explanation["intermediate_values"]["旷工字段口径"] == "旷工时数"


def test_domestic_loader_normalizes_rest_leave_day_alias():
    """晋江月报的排休请假列按天数兼容"""
    loader = domestic_parser.PayrollDataLoader.__new__(domestic_parser.PayrollDataLoader)

    row = loader._normalize_row({"工号": "OWDN001", "排休请假": "2"})

    assert row["排休请假天数"] == 2
    assert "排休请假时数" not in row


def test_excel_parser_preserves_first_nonempty_duplicate_header_value(tmp_path):
    """重复月报表头不得用后面的空值覆盖前面的真实值"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "月考勤"
    sheet.append(["工号", "正班出勤天数", "正班出勤天数"])
    sheet.append(["OWHN001", 16.25, None])
    path = tmp_path / "duplicate-header.xlsx"
    workbook.save(path)

    with ExcelParser(str(path)) as parser:
        row = parser.parse_sheet("月考勤").rows[0]

    assert row["正班出勤天数"] == 16.25


def test_gonglingjiang_keeps_negative_amount_from_offline_formula():
    """线下工资表未设置最低0元兜底，折算结果可为负数"""
    employee = {
        **_gongling_employee(),
        "实际在职工作日天数": -10,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=["OWHN001"], region="gsdg")

    assert result.amount == -173.08
    assert result.details["audit_explanation"]["intermediate_values"]["入离职折算后金额"] < 0


def test_gonglingjiang_full_month_personal_leave_is_zero():
    """正班出勤为0且存在事假时按线下工资表人工归零口径处理"""
    employee = {
        **_gongling_employee(),
        "正班出勤天数": 0,
        "事假时数": 128,
        "排班天数": 20,
        "实际在职工作日天数": 20,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=["OWHN001"], region="gsdg")

    assert result.amount == 0
    assert result.details["audit_explanation"]["rule_name"] == "工龄奖标准与缺勤折算"
    assert result.details["audit_explanation"]["intermediate_values"]["事病旷排休时数"] == 128
    assert result.details["audit_explanation"]["intermediate_values"]["全月事假未出勤归零"] is True
    assert result.details["exceptions"][0]["code"] == "ZERO_REGULAR_ATTENDANCE_DAYS"
    assert result.details["exceptions"][0]["level"] == "info"
    assert "正班出勤天数为0" in result.warnings[0]


def test_gonglingjiang_dongguan_operation_uses_work_area_position_rules():
    """东莞操作按工作地区和新岗位范围发放"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "操作员",
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 450
    explanation = result.details["audit_explanation"]
    assert explanation["inputs"]["工作地区"] == "东莞"
    assert explanation["intermediate_values"]["标准"] == 150
    assert explanation["intermediate_values"]["上限"] == 600


def test_gonglingjiang_b_operation_department_uses_china_operation_rules():
    """B操作部与中国操作部使用同一套工龄奖规则"""
    employee = {
        **_operation_employee(),
        "二级部门名称": "B操作部",
        "工作地区": "东莞",
        "岗位名称": "操作员",
    }

    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 450
    assert result.details["audit_explanation"]["intermediate_values"]["标准"] == 150


@pytest.mark.parametrize("position", ["内部初级安检员", "民航中级安检员", "内部高级安检员"])
def test_gonglingjiang_security_inspector_position_uses_contains_match(position):
    """岗位名称包含“安检员”字样即按安检员资格判断"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": position,
    }

    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 450
    assert result.details["audit_explanation"]["intermediate_values"]["标准"] == 150


@pytest.mark.parametrize("position", ["理货员", "揽收充电司机"])
def test_gonglingjiang_dongguan_operation_includes_offline_eligible_positions(position):
    """东莞线下规则表明确列出的岗位均享有工龄奖，资格不依赖职级"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": position,
        "职级": "",
    }

    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 450
    assert result.details["audit_explanation"]["intermediate_values"]["标准"] == 150


def test_gonglingjiang_dongguan_clerk_stays_pending_confirmation():
    """规则图与6月线下结果冲突，文员确认前不开放"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "文员",
    }

    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 0
    assert result.details["reason"] == "东莞操作岗位不享有工龄奖"


def test_gonglingjiang_dongguan_operation_excludes_non_allowed_position():
    """东莞操作内勤专员不再按操作岗位发放工龄奖"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "内勤专员",
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 0
    assert result.details["reason"] == "东莞操作岗位不享有工龄奖"
    assert result.details["audit_explanation"]["rule_name"] == "工龄奖资格判断"


def test_gonglingjiang_jiashan_east_hub_returns_zero_by_department():
    """华东枢纽优先按二级部门返回0，不再进入嘉善地区判断"""
    employee = {
        **_operation_employee(),
        "工作地区": "嘉善",
        "二级部门名称": "华东枢纽",
        "岗位名称": "操作员",
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 0
    assert result.details["reason"] == "华东/华西指定二级部门无工龄奖"
    assert result.details["audit_explanation"]["rule_name"] == "工龄奖二级部门判断"


def test_gonglingjiang_yiwu_b2b_returns_zero_by_department():
    """华东B2B枢纽优先按二级部门返回0，不再进入义乌地区判断"""
    employee = {
        **_operation_employee(),
        "工作地区": "义乌",
        "二级部门名称": "华东B2B枢纽",
        "岗位名称": "操作员",
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="wes")

    assert result.amount == 0
    assert result.details["reason"] == "华东/华西指定二级部门无工龄奖"
    assert result.details["audit_explanation"]["rule_name"] == "工龄奖二级部门判断"


def test_gonglingjiang_jinjiang_operation_uses_50_rate_and_150_cap():
    """晋江操作员按50元/年且150封顶"""
    employee = {
        **_operation_employee(),
        "工作地区": "晋江",
        "一级部门名称": "东南区",
        "二级部门名称": "东南枢纽",
        "岗位名称": "操作员",
        "入职日期": date(2020, 1, 1),
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 150
    explanation = result.details["audit_explanation"]
    assert explanation["intermediate_values"]["工龄(年)"] == 6
    assert explanation["intermediate_values"]["标准"] == 50
    assert explanation["intermediate_values"]["上限"] == 150


def test_gonglingjiang_jinjiang_gate_guard_is_eligible():
    """晋江门禁员按跨月实际工资表享有工龄奖"""
    employee = {
        **_operation_employee(),
        "工作地区": "晋江",
        "一级部门名称": "东南区",
        "二级部门名称": "东南枢纽",
        "岗位名称": "门禁员",
        "入职日期": date(2025, 1, 2),
        "考勤月份": "202602",
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="wes")

    assert result.amount == 50


def test_gonglingjiang_uses_excel_half_up_rounding():
    """线下Excel ROUND的0.005按远离0方向四舍五入"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "入职日期": date(2024, 12, 4),
        "考勤月份": "202602",
        "排班天数": 20,
        "实际在职工作日天数": 20,
        "事假时数": 2,
        "病假时数": 0,
        "旷工天数": 1,
        "排休请假天数": 7,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 88.13


def test_gonglingjiang_operation_does_not_deduct_work_injury_days():
    """线下工龄奖公式未配置工伤假额外扣减"""
    employee = {
        **_operation_employee(),
        "工作地区": "东莞",
        "岗位名称": "操作员",
        "排班天数": 20,
        "实际在职工作日天数": 20,
        "工伤假天数": 2,
    }
    result = GongLingJiangEngine().calculate(employee, hrbp_list=[], region="gsdg")

    assert result.amount == 450
    explanation = result.details["audit_explanation"]
    assert "工伤假天数" not in explanation["intermediate_values"]
    assert "工伤" not in " ".join(explanation["steps"])


def _gongling_employee() -> dict:
    return {
        "工号": "OWHN001",
        "姓名": "张三",
        "一级部门名称": "莞深操作",
        "二级部门名称": "中国操作部",
        "岗位名称": "操作员",
        "工作地区": "东莞",
        "入职日期": date(2023, 1, 1),
        "考勤月份": "202606",
        "排班天数": 26,
        "实际在职工作日天数": 26,
        "正班出勤天数": 26,
        "请假时数": 0,
        "事假时数": 0,
        "病假时数": 0,
        "旷工天数": 0,
        "排休请假天数": 0,
    }


def _operation_employee() -> dict:
    return {
        "工号": "OWHN888",
        "姓名": "李四",
        "一级部门名称": "莞深操作",
        "二级部门名称": "中国操作部",
        "岗位名称": "操作员",
        "工作地区": "东莞",
        "入职日期": date(2023, 1, 1),
        "考勤月份": "202606",
        "排班天数": 26,
        "实际在职工作日天数": 26,
        "正班出勤天数": 26,
        "事假时数": 0,
        "病假时数": 0,
        "旷工天数": 0,
        "排休请假天数": 0,
        "工伤假天数": 0,
    }
