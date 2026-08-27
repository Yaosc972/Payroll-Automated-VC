from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime, timezone
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import json
import logging
from pathlib import Path
import subprocess
import sys
import threading
import time

from openpyxl import Workbook, load_workbook
import httpx
import pytest
from fastapi.testclient import TestClient

from bonus_platform.app import app
from bonus_platform.engine.social_insurance.adapter import sync_beisen_candidates
from bonus_platform.engine.social_insurance.baseline import (
    capture_monthly_baseline,
    ensure_monthly_baseline_confirmation_date,
    list_monthly_baseline_subjects,
    load_monthly_baseline,
    merge_monthly_baseline,
)
from bonus_platform.engine.social_insurance import field_metadata
from bonus_platform.engine.social_insurance.coverage import build_coverage_tasks
from bonus_platform.engine.social_insurance.report import build_audit_export
from bonus_platform.engine.social_insurance.rule_catalog import RULE_VERSION, public_rule_catalog
from bonus_platform.engine.social_insurance.sync_snapshot import (
    capture_reporting_snapshot,
    load_reporting_snapshot,
)
from bonus_platform.engine.social_insurance.runs import (
    RunValidationError,
    add_supplement_employee,
    confirm_run,
    create_run,
    default_reporting_window,
    list_runs,
    load_run,
    load_run_index,
    persist_run_index,
    update_employee,
)


BASE_REPORT = {
    "证件号码": "TEST-ID-001",
    "姓名": "测试员工",
    "户籍": "广东省外户籍",
    "入深户时间": "",
    "民族": "汉族",
    "手机号码": "13000000000",
    "通讯地址": "深圳市测试地址",
    "电脑号": "",
    "岗位类别": "工人岗位",
    "个人身份": "工人",
    "用工形式": "合同工",
    "学历": "大学专科",
    "职称": "无",
    "国家职业资格或职业技能等级": "无",
    "医疗缴费档次": "职工二档",
    "部门": "",
    "户籍地类别": "农业",
    "户口所在地行政区划代码": "450801.市辖区",
    "灵活就业人员就业形式": "雇佣就业",
    "就业前个人身份": "其他",
}


def _record(*, identity: str, name: str, status: str = "ready", issues: list[dict] | None = None) -> dict:
    report = deepcopy(BASE_REPORT)
    report["证件号码"] = identity
    report["姓名"] = name
    return {
        "status": status,
        "reason": "规则校验通过" if status == "ready" else "需要业务确认",
        "issues": issues or [],
        "report": report,
        "entryDate": "2026-07-20",
        "source": {"subject": "深圳市前海云途物流有限公司", "place": "深圳", "employType": "内部员工"},
    }


def test_default_reporting_window_uses_previous_16th_to_current_15th():
    assert default_reporting_window(date(2026, 8, 16)) == ("2026-07-16", "2026-08-15")
    assert default_reporting_window(date(2026, 8, 15)) == ("2026-06-16", "2026-07-15")


def test_default_reporting_window_uses_shanghai_date_at_the_monthly_boundary():
    assert default_reporting_window(
        now=datetime(2026, 9, 15, 16, 30, tzinfo=timezone.utc)
    ) == ("2026-08-16", "2026-09-15")


def test_default_confirmation_date_is_the_day_after_the_period_end():
    from bonus_platform.engine.social_insurance import runs as social_runs

    assert social_runs.default_confirmation_date("2026-08-15") == "2026-08-16"
    assert social_runs.default_confirmation_date("2026-02-28") == "2026-03-01"


def test_scheduled_reporting_context_uses_the_day_after_the_period_end(
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import prefetch

    monkeypatch.setattr(prefetch, "default_reporting_window", lambda: ("2026-07-16", "2026-08-15"))

    assert prefetch._current_reporting_context() == {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-16",
        "subject": "*",
    }


def test_legacy_numeric_social_place_falls_back_to_work_place_for_template_routing():
    tasks = build_coverage_tasks(
        coverage_source={"socialPlace": "246", "socialMedicalStatus": "社保待审核，医保待审核"},
        source={"place": "四川省/成都市"},
        employee_status="ready",
        decision="include",
    )

    assert tasks["social"]["route"] == "chengdu-social"
    assert tasks["medical"]["route"] == "chengdu-medical"


def test_rule_history_starts_with_current_rule_version():
    catalog = public_rule_catalog()

    assert catalog["version"] == RULE_VERSION
    assert catalog["history"][0]["version"] == RULE_VERSION


def test_list_runs_prefers_newest_created_batch_even_when_older_batch_was_updated_later(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    runs_root = tmp_path / "runs"
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs_root))
    older_path = runs_root / "sir_older_000001" / "run.json"
    newer_path = runs_root / "sir_newer_000002" / "run.json"
    older_path.parent.mkdir(parents=True)
    newer_path.parent.mkdir(parents=True)
    older_path.write_text(json.dumps({
        "id": "sir_older_000001",
        "createdAt": "2026-08-21T07:19:11Z",
        "updatedAt": "2026-08-21T09:13:12Z",
        "employees": [],
    }), encoding="utf-8")
    newer_path.write_text(json.dumps({
        "id": "sir_newer_000002",
        "createdAt": "2026-08-21T09:12:34Z",
        "updatedAt": "2026-08-21T09:12:34Z",
        "employees": [],
    }), encoding="utf-8")
    older_path.touch()

    assert [run["id"] for run in list_runs(limit=2)] == ["sir_newer_000002", "sir_older_000001"]


def test_runs_endpoint_filters_batches_by_period_confirmation_date_and_subject(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    for subject in ("测试主体甲", "测试主体乙"):
        record = _record(identity=f"TEST-{subject}", name=f"{subject}员工")
        record["source"]["subject"] = subject
        create_run(
            records=[record],
            period_start="2026-07-16",
            period_end="2026-08-15",
            confirmation_date="2026-08-21",
            subject=subject,
            source="fixture",
        )

    response = TestClient(app).get(
        "/api/social-insurance/runs",
        params={
            "limit": 10,
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-21",
            "subject": "测试主体乙",
        },
    )

    assert response.status_code == 200
    assert [run["subject"] for run in response.json()["runs"]] == ["测试主体乙"]


def test_current_run_endpoint_uses_the_context_index_and_returns_preflight(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import router as social_router

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    record = _record(identity="TEST-CURRENT-RUN-001", name="索引批次员工")
    record["source"]["subject"] = "索引测试主体"
    created = create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject="索引测试主体",
        source="fixture",
    )

    def fail_if_scanned(*_args, **_kwargs):
        pytest.fail("精确主体查询不应扫描全部批次")

    monkeypatch.setattr(social_router, "list_runs", fail_if_scanned)
    response = TestClient(app).get(
        "/api/social-insurance/runs/current",
        params={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-25",
            "subject": "索引测试主体",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["lookupSource"] == "run-index"
    assert payload["run"]["id"] == created["id"]
    assert payload["preflight"]["runId"] == created["id"]
    assert payload["run"]["employees"][0]["report"]["姓名"] == "索引批次员工"


def test_saving_an_older_batch_does_not_replace_the_current_context_index(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import runs as social_runs

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    timestamps = iter(("2026-08-25T04:00:00.000001Z", "2026-08-25T04:00:00.000002Z"))
    monkeypatch.setattr(social_runs, "current_timestamp", lambda: next(timestamps))
    record = _record(identity="TEST-CURRENT-ORDER-001", name="索引顺序员工")
    record["source"]["subject"] = "索引顺序主体"
    context = {
        "period_start": "2026-07-16",
        "period_end": "2026-08-15",
        "confirmation_date": "2026-08-25",
        "subject": "索引顺序主体",
    }
    older = create_run(records=[record], source="fixture", **context)
    newer = create_run(records=[record], source="fixture", **context)

    assert older["createdAt"] < newer["createdAt"]
    assert persist_run_index(older) is True

    indexed = load_run_index(**context)
    assert indexed is not None
    assert indexed["runId"] == newer["id"]


def test_current_run_endpoint_backfills_an_index_for_an_existing_batch(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import router as social_router

    runs_root = tmp_path / "runs"
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs_root))
    run_id = "sir_existing_without_index"
    run_dir = runs_root / run_id
    run_dir.mkdir(parents=True)
    run_dir.joinpath("run.json").write_text(json.dumps({
        "id": run_id,
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-25",
        "subject": "历史索引主体",
        "status": "draft",
        "employees": [],
        "summary": {"total": 0, "ready": 0, "needsReview": 0, "included": 0, "excluded": 0},
        "processingPlan": [],
        "createdAt": "2026-08-24T01:00:00Z",
        "updatedAt": "2026-08-24T01:00:00Z",
    }, ensure_ascii=False), encoding="utf-8")
    params = {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-25",
        "subject": "历史索引主体",
    }

    first = TestClient(app).get("/api/social-insurance/runs/current", params=params)
    assert first.status_code == 200
    assert first.json()["lookupSource"] == "run-list-fallback"

    monkeypatch.setattr(social_router, "list_runs", lambda *_args, **_kwargs: pytest.fail("索引回填后不应再次扫描"))
    second = TestClient(app).get("/api/social-insurance/runs/current", params=params)
    assert second.status_code == 200
    assert second.json()["lookupSource"] == "run-index"


def test_sync_all_fetches_beisen_once_and_creates_one_batch_per_subject(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    from bonus_platform.engine.social_insurance import router as social_router

    records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙"), start=1):
        record = _record(identity=f"TEST-ID-ALL-{index}", name=f"测试员工{index}")
        record["source"]["subject"] = subject
        records.append(record)
    calls: list[str] = []

    def one_live_sync(**kwargs):
        calls.append(kwargs["subject"])
        return records, {"provider": "beisen-open-platform", "candidateCount": 2, "warnings": []}

    monkeypatch.setattr(social_router, "sync_beisen_candidates", one_live_sync)
    monkeypatch.setattr(social_router, "cached_beisen_contract_subjects", lambda **_kwargs: [
        {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
        {"value": "测试主体乙", "label": "测试主体乙", "candidateCount": 1},
        {"value": "测试主体丙", "label": "测试主体丙", "candidateCount": 0},
    ])
    response = TestClient(app).post(
        "/api/social-insurance/runs/sync-all",
        json={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-21",
            "subject": "测试主体乙",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert calls == ["*"]
    assert payload["batchCount"] == 3
    assert {run["subject"] for run in payload["runs"]} == {"测试主体甲", "测试主体乙", "测试主体丙"}
    assert payload["selectedRun"]["subject"] == "测试主体乙"
    assert payload["selectedRun"]["summary"]["total"] == 1
    assert payload["runs"][-1]["subject"] == "测试主体乙"
    empty_run = next(run for run in payload["runs"] if run["subject"] == "测试主体丙")
    assert empty_run["summary"]["total"] == 0
    assert list_runs(1)[0]["subject"] == "测试主体乙"


def test_sync_all_keeps_a_subject_that_only_exists_in_the_monthly_baseline(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    from bonus_platform.engine.social_insurance import router as social_router

    historical = _record(identity="TEST-ID-BASELINE-SUBJECT", name="基线主体员工")
    historical["source"]["subject"] = "仅基线主体"
    capture_monthly_baseline(
        records=[historical],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="仅基线主体",
        source="beisen-monthly-snapshot",
    )
    invalid = _record(identity="", name="无唯一身份员工")
    invalid["source"]["subject"] = "无有效身份基线主体"
    capture_monthly_baseline(
        records=[invalid],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="无有效身份基线主体",
        source="beisen-monthly-snapshot",
    )
    current = _record(identity="TEST-ID-CURRENT-SUBJECT", name="当前主体员工")
    current["source"]["subject"] = "当前主体"

    monkeypatch.setattr(
        social_router,
        "sync_beisen_candidates",
        lambda **_kwargs: ([current], {"provider": "beisen-open-platform", "warnings": []}),
    )
    monkeypatch.setattr(
        social_router,
        "cached_beisen_contract_subjects",
        lambda **_kwargs: [{"value": "当前主体", "label": "当前主体", "candidateCount": 1}],
    )

    assert list_monthly_baseline_subjects(
        period_start="2026-07-16",
        period_end="2026-08-15",
    ) == ["仅基线主体"]

    response = TestClient(app).post(
        "/api/social-insurance/runs/sync-all",
        json={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-21",
            "subject": "当前主体",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["batchCount"] == 2
    baseline_run = next(run for run in payload["runs"] if run["subject"] == "仅基线主体")
    assert baseline_run["summary"]["total"] == 1
    assert baseline_run["summary"]["ready"] == 0
    assert baseline_run["summary"]["needsReview"] == 1
    assert baseline_run["summary"]["excluded"] == 0
    loaded = load_run(baseline_run["id"])
    assert loaded["employees"][0]["report"]["姓名"] == "基线主体员工"
    assert loaded["employees"][0]["status"] == "needs_review"
    assert loaded["sourceSummary"]["monthlyBaseline"]["baselineOnlyCount"] == 1
    assert any("未出现在北森当前任职结果" in warning for warning in loaded["sourceSummary"]["warnings"])


def test_sync_all_uses_fresh_period_snapshot_without_waiting_for_beisen(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    from bonus_platform.engine.social_insurance import router as social_router

    records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙"), start=1):
        record = _record(identity=f"TEST-ID-CACHED-ALL-{index}", name=f"快照员工{index}")
        record["source"]["subject"] = subject
        records.append(record)
    capture_reporting_snapshot(
        records=records,
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-21",
        subject="*",
    )

    def unexpected_live_sync(**_kwargs):
        raise AssertionError("新鲜的全主体周期快照不应再次请求北森")

    monkeypatch.setattr(social_router, "sync_beisen_candidates", unexpected_live_sync)
    monkeypatch.setattr(social_router, "cached_beisen_contract_subjects", lambda **_kwargs: [
        {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
        {"value": "测试主体乙", "label": "测试主体乙", "candidateCount": 1},
        {"value": "测试主体丙", "label": "测试主体丙", "candidateCount": 0},
    ])

    response = TestClient(app).post(
        "/api/social-insurance/runs/sync-all",
        json={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-21",
            "subject": "测试主体乙",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["batchCount"] == 3
    assert payload["selectedRun"]["subject"] == "测试主体乙"
    assert payload["selectedRun"]["sourceSummary"]["dataMode"] == "background-all-subject-snapshot"
    assert payload["selectedRun"]["sourceSummary"]["snapshotAgeSeconds"] >= 0


def test_sync_all_reuses_a_rebasable_period_snapshot_for_a_new_confirmation_date(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    from bonus_platform.engine.social_insurance import router as social_router

    record = _record(identity="TEST-ID-REBASED-ALL", name="跨确认日快照员工", status="excluded")
    record["source"]["subject"] = "测试主体甲"
    record["confirmationRuleContext"] = {
        "version": 1,
        "baseStatus": "ready",
        "baseIssues": [],
        "currentEntryDate": "2026-07-20",
        "dimissionRecords": [{
            "lastWorkDate": "2026-08-20",
            "voluntaryStopFlag": "自愿停保",
            "processCreatedTime": "2026-08-20T10:00:00+08:00",
        }],
    }
    capture_reporting_snapshot(
        records=[record],
        source_summary={"provider": "beisen-open-platform", "confirmationDate": "2026-08-25"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject="*",
    )

    monkeypatch.setattr(
        social_router,
        "sync_beisen_candidates",
        lambda **_kwargs: pytest.fail("可重算的周期快照不应因确认日变化再次请求北森"),
    )
    monkeypatch.setattr(social_router, "cached_beisen_contract_subjects", lambda **_kwargs: [
        {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
    ])

    response = TestClient(app).post(
        "/api/social-insurance/runs/sync-all",
        json={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-16",
            "subject": "测试主体甲",
        },
    )

    assert response.status_code == 200
    selected = response.json()["selectedRun"]
    assert selected["confirmationDate"] == "2026-08-16"
    assert selected["employees"][0]["status"] == "ready"
    assert selected["summary"]["excluded"] == 0
    assert selected["sourceSummary"]["dataMode"] == "background-all-subject-snapshot"
    assert selected["sourceSummary"]["snapshotRebasedFromConfirmationDate"] == "2026-08-25"
    assert "confirmationRuleContext" not in selected["employees"][0]

    later_response = TestClient(app).post(
        "/api/social-insurance/runs/sync-all",
        json={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-21",
            "subject": "测试主体甲",
        },
    )

    assert later_response.status_code == 200
    later_selected = later_response.json()["selectedRun"]
    assert later_selected["employees"][0]["status"] == "excluded"
    assert later_selected["summary"]["excluded"] == 1


def test_sync_all_refreshes_an_expired_period_snapshot_once_before_creating_batches(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_SNAPSHOT_FRESH_SECONDS", "900")
    from bonus_platform.engine.social_insurance import router as social_router

    expired = _record(identity="TEST-ID-EXPIRED-ALL", name="过期快照员工")
    expired["source"]["subject"] = "测试主体甲"
    capture_reporting_snapshot(
        records=[expired],
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-21",
        subject="*",
        captured_at="2026-08-20T00:00:00Z",
    )
    live = _record(identity="TEST-ID-REFRESHED-ALL", name="实时刷新员工")
    live["source"]["subject"] = "测试主体甲"
    calls: list[str] = []

    def one_live_sync(**kwargs):
        calls.append(kwargs["subject"])
        return [live], {"provider": "beisen-open-platform", "warnings": []}

    monkeypatch.setattr(social_router, "sync_beisen_candidates", one_live_sync)
    monkeypatch.setattr(social_router, "cached_beisen_contract_subjects", lambda **_kwargs: [
        {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
    ])

    response = TestClient(app).post(
        "/api/social-insurance/runs/sync-all",
        json={
            "periodStart": "2026-07-16",
            "periodEnd": "2026-08-15",
            "confirmationDate": "2026-08-21",
            "subject": "测试主体甲",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert calls == ["*"]
    assert payload["selectedRun"]["employees"][0]["report"]["姓名"] == "实时刷新员工"
    assert payload["selectedRun"]["sourceSummary"]["dataMode"] == "live-beisen-all-subjects"
    refreshed_snapshot = load_reporting_snapshot(
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-21",
        subject="*",
    )
    assert refreshed_snapshot is not None
    assert refreshed_snapshot["stale"] is False
    assert refreshed_snapshot["records"][0]["report"]["姓名"] == "实时刷新员工"


def test_run_preserves_business_audit_fields_from_source(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    record = _record(identity="TEST-ID-AUDIT-FIELDS", name="审核字段员工")
    record["source"].update({
        "socialContributionBase": "4775",
        "housingContributionBase": "9200",
        "housingFundAccount": "HF-001",
        "householdAddress": "广东省深圳市南山区测试街道",
        "jobNumber": "TEST-001",
        "gender": "女",
        "lastWorkDate": "",
        "housingContributionRate": "5%",
        "socialContributionPlace": "深圳市",
        "birthplace": "广东省深圳市",
        "domicileType": "城市户口",
        "education": "本科/学士",
        "currentAddress": "广东省深圳市福田区测试地址",
        "nation": "汉族",
        "employeeStatus": "试用",
        "email": "test@example.com",
        "employmentPlace": "集团深圳",
        "changeDescription": "测试变动",
    })
    run = create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="深圳市前海云途物流有限公司",
        source="fixture",
    )

    report = run["employees"][0]["report"]
    assert report["社保缴交基数"] == "4775"
    assert report["公积金缴交基数"] == "9200"
    assert report["公积金号"] == "HF-001"
    assert report["户口具体地址"] == "广东省深圳市南山区测试街道"
    source = run["employees"][0]["source"]
    expected_source = {
        "subject": "深圳市前海云途物流有限公司",
        "jobNumber": "TEST-001",
        "place": "深圳",
        "employType": "内部员工",
        "gender": "女",
        "mobile": "",
        "lastWorkDate": "",
        "housingContributionRate": "5%",
        "socialContributionPlace": "深圳市",
        "birthplace": "广东省深圳市",
        "domicileType": "城市户口",
        "education": "本科/学士",
        "currentAddress": "广东省深圳市福田区测试地址",
        "nation": "汉族",
        "employeeStatus": "试用",
        "email": "test@example.com",
        "employmentPlace": "集团深圳",
        "changeDescription": "测试变动",
        "socialPlace": "",
        "socialMedicalStatus": "",
        "housingStatus": "",
    }
    assert {key: source[key] for key in expected_source} == expected_source
    assert source["virtualEmployee"] == ""
    assert source["householdAddress"] == "广东省深圳市南山区测试街道"


def test_audit_export_contains_every_employee_and_all_reporting_fields_in_one_sheet(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    included = _record(identity="TEST-ID-AUDIT-001", name="纳入员工")
    included["source"].update({
        "jobNumber": "ZT-AUDIT-001",
        "gender": "女",
        "mobile": "13800138000",
        "lastWorkDate": "",
        "employeeStatus": "试用",
        "socialContributionPlace": "深圳市",
        "birthplace": "广东省深圳市",
        "domicileType": "城市户口",
        "email": "audit@example.com",
        "employmentPlace": "集团深圳",
        "housingContributionRate": "5%",
    })
    excluded = _record(identity="TEST-ID-AUDIT-002", name="排除员工", status="excluded")
    run = create_run(
        records=[included, excluded],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-19",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )

    response = TestClient(app).get(
        f"/api/social-insurance/runs/{run['id']}/audit-export"
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["全部增员人员"]
    sheet = workbook["全部增员人员"]
    headers = [cell.value for cell in sheet[4]]
    assert sheet.max_row == 6
    assert sheet["A1"].value == "社保增员审核清单（业务核对版）"
    assert sheet["A3"].value == "处理"
    assert sheet["C3"].value == "人员与入职"
    assert sheet.freeze_panes == "E5"
    assert "处理结果" in headers
    assert "校验问题" in headers
    assert "证件号码" in headers
    assert "工号" in headers
    assert "户籍所在地" in headers
    assert "社保缴纳地" in headers
    assert "公积金个人比例" in headers
    assert "医疗缴费档次" in headers
    assert "户口所在地行政区划代码" in headers
    assert tuple(headers[:12]) == (
        "序号", "处理结果", "姓名", "工号", "证件号码", "性别", "手机号码",
        "入职日期", "离职日期", "在职状态", "雇佣关系", "合同主体",
    )
    rows = list(sheet.iter_rows(min_row=5, values_only=True))
    assert {row[headers.index("姓名")] for row in rows} == {"纳入员工", "排除员工"}
    assert {row[headers.index("处理结果")] for row in rows} == {"纳入", "排除"}
    assert {row[headers.index("证件号码")] for row in rows} == {
        "TEST-ID-AUDIT-001",
        "TEST-ID-AUDIT-002",
    }
    included_row = next(row for row in rows if row[headers.index("姓名")] == "纳入员工")
    assert included_row[headers.index("工号")] == "ZT-AUDIT-001"
    assert included_row[headers.index("户籍所在地")] == "广东省深圳市"
    assert included_row[headers.index("社保缴纳地")] == "深圳市"
    assert sheet.cell(5, headers.index("证件号码") + 1).quotePrefix is True
    assert sheet.auto_filter.ref.endswith("6")


def test_live_sync_requires_current_departure_snapshot(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    engine_dir = tmp_path / "engine"
    engine_dir.mkdir()
    (engine_dir / "mvp.mjs").write_text("// test entrypoint\n", encoding="utf-8")
    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_DIMISSION_FILE", raising=False)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_NODE", sys.executable)

    with pytest.raises(RunValidationError, match="一键同步已锁定"):
        sync_beisen_candidates(
            period_start="2026-07-16",
            period_end="2026-08-15",
            confirmation_date="2026-08-17",
            subject="深圳市前海云途物流有限公司",
            output_dir=tmp_path / "output",
        )


def test_live_sync_uses_configured_monthly_source_as_historical_baseline(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    engine_dir = tmp_path / "engine"
    engine_dir.mkdir()
    (engine_dir / "mvp.mjs").write_text("// test entrypoint\n", encoding="utf-8")
    dimission = tmp_path / "全部离职记录_20260817.xlsx"
    dimission.touch()
    baseline_source = tmp_path / "0616-0715北森原始名单.xlsx"
    baseline_source.touch()
    captured: dict[str, list[str]] = {}

    def fake_run(command, **_kwargs):
        captured["command"] = command
        output_dir = Path(command[command.index("--output") + 1])
        (output_dir / "深圳社保增员_候选与人工确认.xlsx").touch()
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=json.dumps({"candidate_count": 1, "diagnostics": {"combined_match_count": 1}}),
            stderr="",
        )

    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_NODE", sys.executable)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_DIMISSION_FILE", str(dimission))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINE_SOURCE_FILE", str(baseline_source))
    monkeypatch.setattr("bonus_platform.engine.social_insurance.adapter.subprocess.run", fake_run)
    monkeypatch.setattr(
        "bonus_platform.engine.social_insurance.adapter._read_review_workbook",
        lambda _path: [_record(identity="TEST-ID-SEED-001", name="历史基线员工")],
    )

    _records, summary = sync_beisen_candidates(
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        output_dir=tmp_path / "output",
    )

    assert captured["command"][captured["command"].index("--source") + 1] == "hybrid"
    assert captured["command"][captured["command"].index("--source-file") + 1] == str(baseline_source)
    assert captured["command"][captured["command"].index("--cutoff") + 1] == "2026-07-17T23:59:59+08:00"
    assert summary["historicalBaselineSeedUsed"] is True
    assert summary["historicalBaselineSourceName"] == baseline_source.name


def test_live_sync_enriches_candidates_with_offline_coverage_status_without_copying_raw_rows(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    engine_dir = tmp_path / "engine"
    engine_dir.mkdir()
    (engine_dir / "mvp.mjs").write_text("// test entrypoint\n", encoding="utf-8")
    dimission = tmp_path / "全部离职记录_20260820.xlsx"
    dimission.touch()
    baseline_source = tmp_path / "北森员工社保报表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["parent_Name", "parent_IDNumber", "", "", ""])
    sheet.append(["姓名", "身份证号码", "社保、医保", "公积金", "社保缴纳地"])
    sheet.append(["规则样例", "TEST-ID-OFFLINE-001", "社保待审核，医保补缴7月", "未封存", "成都"])
    workbook.save(baseline_source)

    def fake_run(command, **_kwargs):
        output_dir = Path(command[command.index("--output") + 1])
        (output_dir / "候选记录.json").write_text(json.dumps([
            _record(identity="TEST-ID-OFFLINE-001", name="规则样例")
        ], ensure_ascii=False), encoding="utf-8")
        return subprocess.CompletedProcess(command, 0, stdout='{"candidate_count":1}\n', stderr="")

    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_NODE", sys.executable)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_DIMISSION_FILE", str(dimission))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINE_SOURCE_FILE", str(baseline_source))
    monkeypatch.setattr("bonus_platform.engine.social_insurance.adapter.subprocess.run", fake_run)

    records, _summary = sync_beisen_candidates(
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="测试合同主体",
        output_dir=tmp_path / "output",
    )

    assert records[0]["coverageSource"] == {
        "socialPlace": "成都",
        "socialMedicalStatus": "社保待审核，医保补缴7月",
        "housingStatus": "未封存",
    }


def test_monthly_baseline_prevents_historical_employee_from_silently_disappearing(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    baseline_record = _record(identity="TEST-ID-HISTORY-001", name="历史名单员工")

    metadata = capture_monthly_baseline(
        records=[baseline_record],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen-monthly-snapshot",
    )
    loaded = load_monthly_baseline(
        period_start="2026-06-16",
        period_end="2026-07-15",
        subject="深圳市前海云途物流有限公司",
    )
    merged, summary = merge_monthly_baseline([], loaded["records"])

    assert metadata["recordCount"] == 1
    assert summary == {
        "baselineCount": 1,
        "currentCount": 0,
        "baselineOnlyCount": 1,
        "baselineDecisionReuseCount": 0,
        "mergedCount": 1,
    }
    assert merged[0]["status"] == "needs_review"
    assert merged[0]["confirmed"] is False
    assert merged[0]["issues"] == [
        {
            "field": "",
            "severity": "blocking",
            "message": "月度名单基线中有此人，但当前北森任职接口未返回；请结合最新离职信息确认是否仍纳入。",
        }
    ]
    baseline_files = list((tmp_path / "baselines").glob("*.json"))
    assert len(baseline_files) == 1
    assert baseline_files[0].stat().st_mode & 0o777 == 0o600


def test_current_beisen_record_overrides_same_person_in_monthly_baseline():
    baseline_record = _record(identity="TEST-ID-HISTORY-002", name="旧姓名")
    current_record = _record(identity="TEST-ID-HISTORY-002", name="当前姓名")
    current_record["report"]["手机号码"] = "13100000000"

    merged, summary = merge_monthly_baseline([current_record], [baseline_record])

    assert summary["baselineOnlyCount"] == 0
    assert len(merged) == 1
    assert merged[0]["report"]["姓名"] == "当前姓名"
    assert merged[0]["report"]["手机号码"] == "13100000000"


def test_legacy_monthly_baseline_can_add_confirmation_date_without_changing_records(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    original = _record(identity="TEST-ID-LEGACY-001", name="历史基线员工")
    capture_monthly_baseline(
        records=[original],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen-api-plus-historical-source",
    )
    baseline_path = next((tmp_path / "baselines").glob("*.json"))
    legacy_payload = json.loads(baseline_path.read_text(encoding="utf-8"))
    legacy_payload.pop("confirmationDate")
    baseline_path.write_text(json.dumps(legacy_payload, ensure_ascii=False), encoding="utf-8")

    upgraded = ensure_monthly_baseline_confirmation_date(
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
    )

    assert upgraded["confirmationDate"] == "2026-07-17"
    assert upgraded["records"] == [original]


def test_blocking_exception_requires_human_confirmation(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path))
    run = create_run(
        records=[
            _record(identity="TEST-ID-001", name="正常员工"),
            _record(
                identity="TEST-ID-002",
                name="待确认员工",
                status="needs_review",
                issues=[{"field": "医疗缴费档次", "severity": "blocking", "message": "档次待确认"}],
            ),
        ],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )

    with pytest.raises(RunValidationError, match="仍有1人需要人工确认"):
        confirm_run(run["id"])

    pending = next(item for item in run["employees"] if item["status"] == "needs_review")
    update_employee(
        run["id"],
        pending["id"],
        {"confirmed": True, "report": {"医疗缴费档次": "职工二档"}, "reviewNote": "业务已确认"},
    )
    confirmed = confirm_run(run["id"])

    assert confirmed["status"] == "confirmed"
    assert confirmed["summary"]["included"] == 2
    assert confirmed["summary"]["needsReview"] == 0
    assert confirmed["summary"]["ready"] == 2
    assert load_run(run["id"])["employees"][1]["reviewNote"] == "业务已确认"


def test_reincluding_an_excluded_employee_restores_ready_status(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    record = _record(identity="TEST-ID-DECISION-001", name="重新纳入员工")
    record["confirmed"] = False
    run = create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-26",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    employee_id = run["employees"][0]["id"]

    excluded = update_employee(run["id"], employee_id, {"decision": "exclude"})
    included = update_employee(run["id"], employee_id, {"decision": "include"})

    assert excluded["employees"][0]["status"] == "excluded"
    assert included["employees"][0]["decision"] == "include"
    assert included["employees"][0]["status"] == "ready"
    assert included["summary"]["included"] == 1
    assert included["summary"]["excluded"] == 0


def test_loading_a_legacy_include_excluded_mismatch_repairs_the_visible_status(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    run = create_run(
        records=[_record(identity="TEST-ID-DECISION-002", name="历史错位员工")],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-26",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    run_path = tmp_path / "runs" / run["id"] / "run.json"
    stored = json.loads(run_path.read_text(encoding="utf-8"))
    stored["employees"][0]["decision"] = "include"
    stored["employees"][0]["status"] = "excluded"
    run_path.write_text(json.dumps(stored, ensure_ascii=False), encoding="utf-8")

    repaired = load_run(run["id"])

    assert repaired["employees"][0]["decision"] == "include"
    assert repaired["employees"][0]["status"] == "ready"
    assert repaired["summary"]["excluded"] == 0


def test_address_history_difference_is_information_only(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path))
    run = create_run(
        records=[
            _record(
                identity="TEST-ID-003",
                name="地址变更员工",
                issues=[{"field": "通讯地址", "severity": "info", "message": "历史信息变更"}],
            )
        ],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )

    confirmed = confirm_run(run["id"])
    assert confirmed["status"] == "confirmed"
    assert confirmed["summary"]["infoOnly"] == 1


def test_fixture_sync_api_supports_page_review_flow(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setattr(
        field_metadata,
        "load_administrative_divisions",
        lambda: ["450801.市辖区", "450802.港北区"],
    )
    fixture_path = tmp_path / "beisen-fixture.json"
    fixture_path.write_text(
        json.dumps(
            {
                "records": [
                    _record(identity="TEST-ID-010", name="接口员工"),
                    _record(
                        identity="TEST-ID-011",
                        name="接口待确认员工",
                        status="needs_review",
                        issues=[{"field": "行政区划", "severity": "blocking", "message": "行政区划待确认"}],
                    ),
                ],
                "sourceSummary": {"provider": "fixture", "rawApiResponseSaved": True},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))

    client = TestClient(app)
    response = client.post(
        "/api/social-insurance/runs/sync",
        json={
            "periodStart": "2026-07-16",
                "periodEnd": "2026-08-15",
                "confirmationDate": "2026-08-17",
                "subject": "深圳市前海云途物流有限公司",
        },
    )
    assert response.status_code == 200
    run = response.json()
    assert run["summary"]["total"] == 2
    assert run["ruleVersion"] == "2026.08.24-06"
    assert run["sourceSummary"]["rawApiResponseSaved"] is False
    assert run["sourceSummary"]["monthlyBaseline"]["created"] is True
    assert run["confirmationDate"] == "2026-08-17"

    pending = next(item for item in run["employees"] if item["status"] == "needs_review")
    edit = client.patch(
        f"/api/social-insurance/runs/{run['id']}/employees/{pending['id']}",
        json={"confirmed": True, "report": {"户口所在地行政区划代码": "450801.市辖区"}},
    )
    assert edit.status_code == 200

    confirm = client.post(f"/api/social-insurance/runs/{run['id']}/confirm")
    assert confirm.status_code == 200
    assert confirm.json()["status"] == "confirmed"


def test_repeated_user_sync_uses_background_snapshot_without_waiting_for_beisen(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    fixture_path = tmp_path / "beisen-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [_record(identity="TEST-ID-SNAPSHOT-001", name="快照员工")]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    from bonus_platform.engine.social_insurance import router as social_router

    live_call_count = 0
    original_sync = social_router.sync_beisen_candidates

    def counted_sync(**kwargs):
        nonlocal live_call_count
        live_call_count += 1
        return original_sync(**kwargs)

    monkeypatch.setattr(social_router, "sync_beisen_candidates", counted_sync)
    payload = {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-17",
        "subject": "深圳市前海云途物流有限公司",
    }

    with TestClient(app) as client:
        first = client.post("/api/social-insurance/runs/sync", json=payload)
        second = client.post("/api/social-insurance/runs/sync", json=payload)

    assert first.status_code == 200
    assert second.status_code == 200
    assert live_call_count == 1
    assert second.json()["sourceSummary"]["dataMode"] == "background-snapshot"
    assert second.json()["sourceSummary"]["snapshotAgeSeconds"] >= 0


def test_app_startup_seeds_reporting_snapshot_before_the_user_clicks_sync(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_ENABLED", "1")
    legacy_record = _record(
        identity="TEST-ID-STARTUP-001",
        name="定时快照员工",
        status="needs_review",
        issues=[{
            "field": "医疗缴费档次",
            "severity": "blocking",
            "message": "医疗缴费档次：北森档次备注无法识别，需人工确认",
        }],
    )
    legacy_record["report"]["医疗缴费档次"] = ""
    existing = create_run(
        records=[legacy_record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
        source_summary={"provider": "beisen-open-platform"},
    )
    from bonus_platform.engine.social_insurance import router as social_router

    def unexpected_live_sync(**_kwargs):
        raise AssertionError("用户点击不应再同步等待北森")

    monkeypatch.setattr(social_router, "sync_beisen_candidates", unexpected_live_sync)
    with TestClient(app) as client:
        response = client.post(
            "/api/social-insurance/runs/sync",
            json={
                "periodStart": existing["periodStart"],
                "periodEnd": existing["periodEnd"],
                "confirmationDate": existing["confirmationDate"],
                "subject": existing["subject"],
            },
        )

    assert response.status_code == 200
    assert response.json()["sourceSummary"]["dataMode"] == "background-snapshot"
    assert response.json()["summary"]["total"] == 1
    assert response.json()["summary"]["ready"] == 1
    assert response.json()["employees"][0]["report"]["医疗缴费档次"] == "职工二档"


def test_second_sync_keeps_person_from_first_monthly_baseline(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    fixture_path = tmp_path / "beisen-fixture.json"
    fixture_path.write_text(
        json.dumps(
            {"records": [
                _record(identity="TEST-ID-KEEP-001", name="仍在接口员工"),
                _record(identity="TEST-ID-KEEP-002", name="后来接口缺失员工"),
            ]},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    client = TestClient(app)
    payload = {
        "periodStart": "2026-06-16",
        "periodEnd": "2026-07-15",
        "confirmationDate": "2026-07-17",
        "subject": "深圳市前海云途物流有限公司",
    }
    assert client.post("/api/social-insurance/runs/sync", json=payload).status_code == 200

    fixture_path.write_text(
        json.dumps(
            {"records": [_record(identity="TEST-ID-KEEP-001", name="仍在接口员工")]},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    second = client.post("/api/social-insurance/runs/sync", json={**payload, "forceRefresh": True})

    assert second.status_code == 200
    run = second.json()
    assert run["summary"]["total"] == 2
    assert run["summary"]["ready"] == 2
    assert run["summary"]["needsReview"] == 0
    assert run["sourceSummary"]["monthlyBaseline"]["baselineOnlyCount"] == 1
    assert run["sourceSummary"]["monthlyBaseline"]["baselineDecisionReuseCount"] == 1
    missing = next(item for item in run["employees"] if item["report"]["姓名"] == "后来接口缺失员工")
    assert missing["status"] == "ready"

    payload["confirmationDate"] = "2026-07-18"
    third = client.post("/api/social-insurance/runs/sync", json=payload)
    assert third.status_code == 200
    third_run = third.json()
    assert third_run["summary"]["needsReview"] == 1
    assert third_run["sourceSummary"]["monthlyBaseline"]["baselineDecisionReuseCount"] == 0


def test_confirmation_date_cannot_be_before_period_end(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    with pytest.raises(RunValidationError, match="名单确认日不能早于增员周期结束日"):
        create_run(
            records=[_record(identity="TEST-ID-CUTOFF-001", name="确认日测试员工")],
            period_start="2026-06-16",
            period_end="2026-07-15",
            confirmation_date="2026-07-14",
            subject="深圳市前海云途物流有限公司",
            source="beisen",
        )


def test_supplement_employee_requires_supported_reason_and_manual_confirmation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    run = create_run(
        records=[_record(identity="TEST-ID-CURRENT-001", name="正常周期员工")],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    supplement = _record(identity="TEST-ID-SUPPLEMENT-001", name="上期漏报员工")
    supplement["entryDate"] = "2026-06-03"

    updated = add_supplement_employee(
        run["id"],
        supplement,
        reason_type="prior_period_omission",
        note="业务核实为上期遗漏",
    )

    employee = next(item for item in updated["employees"] if item["report"]["姓名"] == "上期漏报员工")
    assert updated["summary"]["supplemental"] == 1
    assert employee["status"] == "needs_review"
    assert employee["confirmed"] is False
    assert employee["supplemental"] == {
        "type": "prior_period_omission",
        "label": "上期漏报",
        "note": "业务核实为上期遗漏",
        "addedAt": employee["supplemental"]["addedAt"],
    }
    with pytest.raises(RunValidationError, match="仍有1人需要人工确认"):
        confirm_run(run["id"])
    with pytest.raises(RunValidationError, match="已在当前批次"):
        add_supplement_employee(
            run["id"],
            supplement,
            reason_type="delayed_enrollment",
            note="重复补充测试",
        )


def test_supplement_employee_rejects_normal_period_person_and_unknown_reason(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    run = create_run(
        records=[_record(identity="TEST-ID-CURRENT-010", name="当前批次员工")],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    in_period = _record(identity="TEST-ID-SUPPLEMENT-010", name="周期内员工")
    in_period["entryDate"] = "2026-06-20"

    with pytest.raises(RunValidationError, match="本期正常增员周期内"):
        add_supplement_employee(
            run["id"],
            in_period,
            reason_type="prior_period_omission",
            note="不应人工补充",
        )
    in_period["entryDate"] = "2026-06-03"
    with pytest.raises(RunValidationError, match="补充增员原因"):
        add_supplement_employee(
            run["id"],
            in_period,
            reason_type="back_payment",
            note="补缴暂不支持",
        )


def test_supplement_search_and_add_api_uses_beisen_record_without_exposing_full_identity(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    current = _record(identity="TEST-ID-SEARCH-001", name="当前名单员工")
    supplement = _record(identity="TEST-ID-SEARCH-002", name="待补充员工")
    supplement["entryDate"] = "2026-06-03"
    fixture_path = tmp_path / "supplement-search-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [current, supplement]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    from bonus_platform.engine.social_insurance import supplements

    query_call_count = 0
    original_query_records = supplements._query_records

    def counted_query_records(active_run):
        nonlocal query_call_count
        query_call_count += 1
        return original_query_records(active_run)

    monkeypatch.setattr(supplements, "_query_records", counted_query_records)
    run = create_run(
        records=[current],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    client = TestClient(app)

    search = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "待补充"},
    )

    assert search.status_code == 200
    candidates = search.json()["candidates"]
    assert len(candidates) == 1
    assert candidates[0]["name"] == "待补充员工"
    assert "TEST-ID-SEARCH-002" not in search.text
    second_search = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "完全不存在"},
    )
    assert second_search.status_code == 200
    assert second_search.json()["candidates"] == []
    cache_status = client.get(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/status"
    )
    assert cache_status.status_code == 200
    assert cache_status.json()["state"] == "ready"
    assert cache_status.json()["recordCount"] == 2
    assert cache_status.json()["rawApiResponseSaved"] is False
    assert "records" not in cache_status.json()
    add = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplements",
        json={
            "candidateId": candidates[0]["id"],
            "reasonType": "delayed_enrollment",
            "note": "员工资料延迟完善后补充",
        },
    )

    assert add.status_code == 200
    payload = add.json()
    assert payload["summary"]["total"] == 2
    assert payload["summary"]["supplemental"] == 1
    added = next(item for item in payload["employees"] if item["report"]["姓名"] == "待补充员工")
    assert added["supplemental"]["label"] == "延迟增员"
    assert added["status"] == "needs_review"
    repeated_search = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "待补充"},
    )
    assert repeated_search.status_code == 200
    assert repeated_search.json()["candidates"] == []
    assert query_call_count == 1


def test_supplement_search_reuses_recent_beisen_run_before_building_a_year_pool(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    historical = _record(identity="TEST-ID-HISTORY-002", name="历史同步可补充员工")
    historical["entryDate"] = "2026-06-03"
    current = _record(identity="TEST-ID-HISTORY-001", name="当前名单员工")
    create_run(
        records=[historical],
        period_start="2026-05-16",
        period_end="2026-06-15",
        confirmation_date="2026-06-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    active_run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    fixture_path = tmp_path / "current-only-fixture.json"
    fixture_path.write_text(json.dumps({"records": [current]}, ensure_ascii=False), encoding="utf-8")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))

    response = TestClient(app).post(
        f"/api/social-insurance/runs/{active_run['id']}/supplement-candidates/search",
        json={"query": "历史同步"},
    )

    assert response.status_code == 200
    assert [item["name"] for item in response.json()["candidates"]] == ["历史同步可补充员工"]
    assert response.json()["candidates"][0]["lookupSource"] == "recent-beisen-run"
    assert "TEST-ID-HISTORY-002" not in response.text


def test_supplement_pool_is_reused_for_repeated_sync_of_the_same_period(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    current = _record(identity="TEST-ID-POOL-001", name="当前名单员工")
    current["source"]["subject"] = "测试主体甲"
    supplement = _record(identity="TEST-ID-POOL-002", name="缓存候选员工甲")
    supplement["entryDate"] = "2026-06-03"
    supplement["source"]["subject"] = "测试主体甲"
    supplement_b = _record(identity="TEST-ID-POOL-003", name="缓存候选员工乙")
    supplement_b["entryDate"] = "2026-06-04"
    supplement_b["source"]["subject"] = "测试主体乙"
    fixture_path = tmp_path / "pool-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [current, supplement, supplement_b]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    first = create_run(
        records=[current], period_start="2026-07-16", period_end="2026-08-15",
        confirmation_date="2026-08-18", subject="测试主体甲", source="beisen",
    )
    client = TestClient(app)
    assert client.post(
        f"/api/social-insurance/runs/{first['id']}/supplement-candidates/search",
        json={"query": "员工甲"},
    ).status_code == 200
    repeated = create_run(
        records=[current], period_start="2026-07-16", period_end="2026-08-15",
        confirmation_date="2026-08-18", subject="测试主体乙", source="beisen",
    )

    status = client.get(
        f"/api/social-insurance/runs/{repeated['id']}/supplement-candidates/status"
    ).json()

    assert status["state"] == "ready"
    assert status["recordCount"] == 3
    second_search = client.post(
        f"/api/social-insurance/runs/{repeated['id']}/supplement-candidates/search",
        json={"query": "员工乙"},
    )
    assert [item["name"] for item in second_search.json()["candidates"]] == ["缓存候选员工乙"]


def test_supplement_search_filters_subject_before_deduplicating_identity(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-SUBJECT-POOL-001", name="当前名单员工")
    current["source"]["subject"] = "测试主体甲"
    wrong_subject = _record(identity="TEST-ID-SUBJECT-POOL-002", name="跨主体候选员工")
    wrong_subject["entryDate"] = "2026-06-03"
    wrong_subject["source"]["subject"] = "测试主体乙"
    matching_subject = deepcopy(wrong_subject)
    matching_subject["source"]["subject"] = "测试主体甲"
    run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="测试主体甲",
        source="beisen",
    )
    capture_reporting_snapshot(
        records=[wrong_subject, matching_subject],
        source_summary={"dataMode": "supplement-candidate-pool-v1"},
        period_start="2025-07-14",
        period_end="2026-07-15",
        confirmation_date="2026-08-18",
        subject="*",
    )

    response = TestClient(app).post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "跨主体候选"},
    )

    assert response.status_code == 200
    assert [item["subject"] for item in response.json()["candidates"]] == ["测试主体甲"]


def test_supplement_search_uses_lightweight_context_without_loading_the_full_run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    runs_root = tmp_path / "runs"
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs_root))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-LIGHTWEIGHT-001", name="当前名单员工")
    candidate = _record(identity="TEST-ID-LIGHTWEIGHT-002", name="轻量查找候选")
    candidate["entryDate"] = "2026-06-03"
    run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    capture_reporting_snapshot(
        records=[current, candidate],
        source_summary={"dataMode": "supplement-candidate-pool-v1"},
        period_start="2025-07-14",
        period_end="2026-07-15",
        confirmation_date="2026-08-18",
        subject="*",
    )
    context_text = (runs_root / ".supplement-context" / f"{run['id']}.json").read_text(
        encoding="utf-8"
    )
    assert "existingCandidateIds" in context_text
    assert "TEST-ID-LIGHTWEIGHT-001" not in context_text
    (runs_root / run["id"] / "run.json").unlink()

    response = TestClient(app).post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "轻量查找"},
    )

    assert response.status_code == 200
    assert [item["name"] for item in response.json()["candidates"]] == ["轻量查找候选"]


def test_supplement_status_uses_lightweight_context_without_loading_the_full_run(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    runs_root = tmp_path / "runs"
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs_root))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-LIGHTWEIGHT-STATUS-001", name="当前名单员工")
    candidate = _record(identity="TEST-ID-LIGHTWEIGHT-STATUS-002", name="轻量状态候选")
    candidate["entryDate"] = "2026-06-03"
    run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    capture_reporting_snapshot(
        records=[candidate],
        source_summary={"dataMode": "supplement-candidate-pool-v1"},
        period_start="2025-07-14",
        period_end="2026-07-15",
        confirmation_date="2026-08-18",
        subject="*",
    )
    (runs_root / run["id"] / "run.json").unlink()

    response = TestClient(app).get(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/status"
    )

    assert response.status_code == 200
    assert response.json()["state"] == "ready"
    assert response.json()["recordCount"] == 1


def test_precomputed_supplement_index_serves_status_and_search_without_runtime_pool_reads(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    runs_root = tmp_path / "runs"
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs_root))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-PRECOMPUTED-001", name="当前名单员工")
    candidate = _record(identity="TEST-ID-PRECOMPUTED-002", name="预计算候选员工")
    candidate["entryDate"] = "2026-06-03"
    run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    from bonus_platform.engine.social_insurance import supplements

    supplements.publish_supplement_search_indexes(
        [run],
        records=[current, candidate],
        pool_status={
            "state": "ready",
            "cachedAt": "2026-08-25T10:00:00Z",
            "expiresAt": "2026-08-25T14:00:00Z",
            "recordCount": 2,
        },
    )
    supplements.clear_supplement_search_index_cache()
    (runs_root / run["id"] / "run.json").unlink()
    (runs_root / ".supplement-context" / f"{run['id']}.json").unlink()

    client = TestClient(app)
    status = client.get(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/status"
    )
    search = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "预计算候选"},
    )

    assert status.status_code == 200
    assert status.json()["state"] == "ready"
    assert status.json()["storage"] == "precomputed-search-index"
    assert status.json()["recordCount"] == 2
    assert search.status_code == 200
    assert [item["name"] for item in search.json()["candidates"]] == ["预计算候选员工"]
    assert "TEST-ID-PRECOMPUTED-002" not in search.text


def test_precomputed_supplement_index_removes_candidate_after_successful_add(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-INDEX-ADD-001", name="当前名单员工")
    candidate = _record(identity="TEST-ID-INDEX-ADD-002", name="待加入索引候选")
    candidate["entryDate"] = "2026-06-03"
    fixture_path = tmp_path / "precomputed-index-add.json"
    fixture_path.write_text(
        json.dumps({"records": [current, candidate]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    run = create_run(
        records=[current],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    from bonus_platform.engine.social_insurance import supplements

    records, status = supplements.prepare_beisen_supplement_pool(run)
    supplements.publish_supplement_search_indexes([run], records=records, pool_status=status)
    client = TestClient(app)
    search = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "待加入索引"},
    )
    candidate_id = search.json()["candidates"][0]["id"]

    added = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplements",
        json={
            "candidateId": candidate_id,
            "reasonType": "delayed_enrollment",
            "note": "员工资料延迟完善后补充",
        },
    )
    repeated = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "待加入索引"},
    )

    assert added.status_code == 200
    assert repeated.status_code == 200
    assert repeated.json()["candidates"] == []
    index = supplements.load_supplement_search_index(run["id"])
    assert index is not None
    assert index["candidateCount"] == 0


def test_identity_edit_invalidates_precomputed_supplement_index(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    current = _record(identity="TEST-ID-INDEX-PATCH-001", name="待修改证件员工")
    candidate = _record(identity="TEST-ID-INDEX-PATCH-002", name="历史候选员工")
    run = create_run(
        records=[current],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    from bonus_platform.engine.social_insurance import supplements

    supplements.publish_supplement_search_indexes(
        [run],
        records=[candidate],
        pool_status={"cachedAt": "2026-08-25T10:00:00Z", "recordCount": 1},
    )
    employee_id = run["employees"][0]["id"]

    response = TestClient(app).patch(
        f"/api/social-insurance/runs/{run['id']}/employees/{employee_id}",
        json={"report": {"证件号码": "TEST-ID-INDEX-PATCH-NEW"}},
    )

    assert response.status_code == 200
    assert supplements.load_supplement_search_index(run["id"]) is None


def test_decision_patch_can_bundle_preflight_without_a_second_run_read(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    run = create_run(
        records=[_record(identity="TEST-ID-DECISION-BUNDLE-001", name="决策校验员工")],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-26",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    employee_id = run["employees"][0]["id"]

    with caplog.at_level(
        logging.INFO,
        logger="bonus_platform.social_insurance.performance",
    ):
        response = TestClient(app).patch(
            f"/api/social-insurance/runs/{run['id']}/employees/{employee_id}?includePreflight=true",
            json={"decision": "exclude"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert set(payload) == {"run", "preflight"}
    assert payload["run"]["employees"][0]["decision"] == "exclude"
    assert payload["preflight"]["runId"] == run["id"]
    assert payload["preflight"]["summary"]["employeeCount"] == 0
    request_id = response.headers["x-sigma-request-id"]
    assert request_id
    server_timings = {
        name: float(duration)
        for item in response.headers["server-timing"].split(",")
        for name, duration in [item.strip().split(";dur=", 1)]
    }
    assert set(server_timings) == {
        "snapshot-load",
        "state-mutation",
        "snapshot-save",
        "preflight",
        "total",
    }
    assert all(duration >= 0 for duration in server_timings.values())

    performance_records = [
        record
        for record in caplog.records
        if record.name == "bonus_platform.social_insurance.performance"
    ]
    assert len(performance_records) == 1
    performance = json.loads(performance_records[0].message)
    assert performance["event"] == "social_insurance_decision_performance"
    assert performance["request_id"] == request_id
    assert performance["snapshot_bytes"] > 0
    assert 0 < performance["persisted_bytes"] <= performance["snapshot_bytes"]
    assert performance["include_preflight"] is True
    for field in (
        "snapshot_load_ms",
        "state_mutation_ms",
        "snapshot_save_ms",
        "preflight_ms",
        "total_ms",
    ):
        assert isinstance(performance[field], float)
        assert performance[field] >= 0
    assert "TEST-ID-DECISION-BUNDLE-001" not in performance_records[0].message
    assert "决策校验员工" not in performance_records[0].message


def test_supplement_search_falls_back_to_full_run_for_legacy_batches(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    runs_root = tmp_path / "runs"
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs_root))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-LEGACY-CONTEXT-001", name="当前名单员工")
    candidate = _record(identity="TEST-ID-LEGACY-CONTEXT-002", name="旧批次候选")
    candidate["entryDate"] = "2026-06-03"
    run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    capture_reporting_snapshot(
        records=[candidate],
        source_summary={"dataMode": "supplement-candidate-pool-v1"},
        period_start="2025-07-14",
        period_end="2026-07-15",
        confirmation_date="2026-08-18",
        subject="*",
    )
    (runs_root / ".supplement-context" / f"{run['id']}.json").unlink()

    response = TestClient(app).post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "旧批次"},
    )

    assert response.status_code == 200
    assert [item["name"] for item in response.json()["candidates"]] == ["旧批次候选"]


def test_supplement_search_reuses_published_pool_across_serverless_instances(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    current = _record(identity="TEST-ID-SHARED-POOL-001", name="当前名单员工")
    supplement = _record(identity="TEST-ID-SHARED-POOL-002", name="共享候选员工")
    supplement["entryDate"] = "2026-06-03"
    fixture_path = tmp_path / "shared-pool-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [current, supplement]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    run = create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    client = TestClient(app)

    first = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "共享候选"},
    )

    assert first.status_code == 200
    assert [item["name"] for item in first.json()["candidates"]] == ["共享候选员工"]

    from bonus_platform.engine.social_insurance import supplements

    with supplements._POOL_CONDITION:
        supplements._POOL_CACHE.clear()
        supplements._POOL_STATUS.clear()
        supplements._POOL_REFRESHING.clear()

    def unexpected_source_query(*_args, **_kwargs):
        raise AssertionError("共享候选快照存在时不得扫描历史批次或重新查询北森")

    monkeypatch.setattr(supplements, "list_runs", unexpected_source_query)
    monkeypatch.setattr(supplements, "_query_records", unexpected_source_query)

    second = client.post(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
        json={"query": "共享候选"},
    )

    assert second.status_code == 200
    assert [item["name"] for item in second.json()["candidates"]] == ["共享候选员工"]
    status = client.get(
        f"/api/social-insurance/runs/{run['id']}/supplement-candidates/status"
    ).json()
    assert status["state"] == "ready"
    assert status["storage"] == "shared-snapshot"


def test_app_startup_prefetches_latest_supplement_pool_before_user_search(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_ENABLED", "1")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_STARTUP_DELAY_SECONDS", "0")
    current = _record(identity="TEST-ID-PREFETCH-001", name="当前名单员工")
    supplement = _record(identity="TEST-ID-PREFETCH-002", name="后台预热员工")
    supplement["entryDate"] = "2026-06-03"
    fixture_path = tmp_path / "prefetch-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [current, supplement]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))
    from bonus_platform.engine.social_insurance import supplements

    query_call_count = 0
    original_query_records = supplements._query_records

    def counted_query_records(active_run):
        nonlocal query_call_count
        query_call_count += 1
        return original_query_records(active_run)

    monkeypatch.setattr(supplements, "_query_records", counted_query_records)
    run = create_run(
        records=[current],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )

    with TestClient(app) as client:
        deadline = time.monotonic() + 3
        status_payload = {}
        while time.monotonic() < deadline:
            status_response = client.get(
                f"/api/social-insurance/runs/{run['id']}/supplement-candidates/status"
            )
            status_payload = status_response.json()
            if status_payload.get("state") == "ready":
                break
            time.sleep(0.05)

        assert status_payload["state"] == "ready"
        assert status_payload["scheduler"]["enabled"] is True
        assert status_payload["scheduler"]["running"] is True
        search = client.post(
            f"/api/social-insurance/runs/{run['id']}/supplement-candidates/search",
            json={"query": "后台预热"},
        )

    assert search.status_code == 200
    assert [item["name"] for item in search.json()["candidates"]] == ["后台预热员工"]
    assert query_call_count == 1


def test_sync_defers_supplement_pool_until_the_user_searches(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_ENABLED", "1")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_STARTUP_DELAY_SECONDS", "7200")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_INTERACTIVE_DELAY_SECONDS", "0")
    current = _record(identity="TEST-ID-SYNC-PREFETCH-001", name="本批员工")
    supplement = _record(identity="TEST-ID-SYNC-PREFETCH-002", name="可补充员工")
    supplement["entryDate"] = "2026-06-03"
    fixture_path = tmp_path / "sync-prefetch-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [current]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))

    with TestClient(app) as client:
        sync_response = client.post(
            "/api/social-insurance/runs/sync",
            json={
                "periodStart": "2026-07-16",
                "periodEnd": "2026-08-15",
                "confirmationDate": "2026-08-19",
                "subject": "深圳市前海云途物流有限公司",
            },
        )
        assert sync_response.status_code == 200
        run_id = sync_response.json()["id"]

        status_payload = client.get(
            f"/api/social-insurance/runs/{run_id}/supplement-candidates/status"
        ).json()
        assert status_payload["state"] == "empty"
        fixture_path.write_text(
            json.dumps({"records": [current, supplement]}, ensure_ascii=False),
            encoding="utf-8",
        )
        search = client.post(
            f"/api/social-insurance/runs/{run_id}/supplement-candidates/search",
            json={"query": "可补充"},
        )

    assert search.status_code == 200
    assert [item["name"] for item in search.json()["candidates"]] == ["可补充员工"]


def test_default_prefetch_waits_for_the_regular_interval_instead_of_racing_page_load(
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import prefetch

    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_STARTUP_DELAY_SECONDS", raising=False)

    assert prefetch._startup_delay_seconds() == prefetch._interval_minutes() * 60


def test_post_sync_prefetch_runs_immediately_by_default(monkeypatch: pytest.MonkeyPatch):
    from bonus_platform.engine.social_insurance import prefetch

    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_PREFETCH_INTERACTIVE_DELAY_SECONDS", raising=False)

    assert prefetch._interactive_delay_seconds() == 0


def test_reporting_snapshot_refresh_deduplicates_the_same_context(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import prefetch

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    entered = threading.Event()
    release = threading.Event()
    call_count = 0

    def slow_sync(**_kwargs):
        nonlocal call_count
        call_count += 1
        entered.set()
        assert release.wait(timeout=2)
        return [_record(identity="TEST-ID-DEDUPE-001", name="去重员工")], {"provider": "fixture"}

    monkeypatch.setattr(prefetch, "sync_beisen_candidates", slow_sync)
    context = {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-18",
        "subject": "深圳市前海云途物流有限公司",
    }
    with ThreadPoolExecutor(max_workers=2) as executor:
        first = executor.submit(prefetch._refresh_reporting_context, context)
        assert entered.wait(timeout=1)
        second = executor.submit(prefetch._refresh_reporting_context, context)
        second_result = second.result(timeout=1)
        release.set()
        first_result = first.result(timeout=2)

    assert call_count == 1
    assert first_result["state"] == "ready"
    assert second_result["state"] == "warming"


def test_reporting_refresh_returns_safe_stage_diagnostics_for_connector_timeout(
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
):
    from bonus_platform.engine.social_insurance import prefetch

    sensitive_marker = "SECRET-EMPLOYEE-IDENTITY-440301199001010011"

    def fail_sync(**_kwargs):
        try:
            raise httpx.ReadTimeout(sensitive_marker)
        except httpx.ReadTimeout as exc:
            raise RunValidationError(sensitive_marker) from exc

    monkeypatch.setattr(prefetch, "sync_beisen_candidates", fail_sync)
    monkeypatch.setattr(prefetch, "_current_reporting_context", lambda: {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-26",
        "subject": "*",
    })

    with caplog.at_level(logging.WARNING, logger="bonus_platform.social_insurance.prefetch"):
        result = prefetch.refresh_latest_reporting_snapshot()

    assert result["state"] == "error"
    assert result["failedStage"] == "current_sync"
    assert result["errorCategory"] == "connector_timeout"
    assert set(result["stageTimingsMs"]) == {"current_sync"}
    assert isinstance(result["stageTimingsMs"]["current_sync"], int)
    assert result["stageTimingsMs"]["current_sync"] >= 0
    assert isinstance(result["elapsedMs"], int)
    assert result["elapsedMs"] >= result["stageTimingsMs"]["current_sync"]
    serialized = json.dumps(result, ensure_ascii=False)
    assert sensitive_marker not in serialized
    assert sensitive_marker not in caplog.text
    assert "stage=current_sync" in caplog.text
    assert "category=connector_timeout" in caplog.text


@pytest.mark.parametrize(
    ("target_name", "expected_stage", "expected_stages"),
    [
        (
            "capture_reporting_snapshot",
            "snapshot_persist",
            {"current_sync", "snapshot_persist"},
        ),
        (
            "prepare_beisen_supplement_pool",
            "supplement_pool",
            {"current_sync", "snapshot_persist", "supplement_pool"},
        ),
        (
            "cached_beisen_contract_subjects",
            "subjects",
            {"current_sync", "snapshot_persist", "supplement_pool", "subjects"},
        ),
    ],
)
def test_reporting_refresh_identifies_safe_prepublication_failure_stage(
    target_name: str,
    expected_stage: str,
    expected_stages: set[str],
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
):
    from bonus_platform.engine.social_insurance import prefetch

    sensitive_marker = "SECRET-REPORTING-STAGE-440301199001010022"
    record = _record(identity="TEST-ID-STAGE-001", name="阶段诊断员工")
    record["source"]["subject"] = "测试主体甲"

    monkeypatch.setattr(
        prefetch,
        "sync_beisen_candidates",
        lambda **_kwargs: ([record], {"provider": "beisen-open-platform"}),
    )
    monkeypatch.setattr(
        prefetch,
        "capture_reporting_snapshot",
        lambda **_kwargs: {"capturedAt": "2026-08-26T00:00:00Z", "recordCount": 1},
    )
    monkeypatch.setattr(
        prefetch,
        "prepare_beisen_supplement_pool",
        lambda *_args, **_kwargs: ([], {"state": "ready", "recordCount": 0}),
    )
    monkeypatch.setattr(
        prefetch,
        "cached_beisen_contract_subjects",
        lambda **_kwargs: [{"value": "测试主体甲", "label": "测试主体甲"}],
    )

    def fail_stage(*_args, **_kwargs):
        raise RunValidationError(sensitive_marker)

    monkeypatch.setattr(prefetch, target_name, fail_stage)
    monkeypatch.setattr(prefetch, "_current_reporting_context", lambda: {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-26",
        "subject": "*",
    })

    with caplog.at_level(logging.WARNING, logger="bonus_platform.social_insurance.prefetch"):
        result = prefetch.refresh_latest_reporting_snapshot()

    assert result["state"] == "error"
    assert result["failedStage"] == expected_stage
    assert result["errorCategory"] == "validation"
    assert set(result["stageTimingsMs"]) == expected_stages
    assert sensitive_marker not in json.dumps(result, ensure_ascii=False)
    assert sensitive_marker not in caplog.text


@pytest.mark.parametrize(
    ("expected_stage", "expected_category"),
    [
        ("batch_materialize", "validation"),
        ("index_publish", "storage_timeout"),
        ("release_publish", "storage_permission"),
    ],
)
def test_reporting_refresh_keeps_previous_release_and_safely_identifies_publication_failure(
    expected_stage: str,
    expected_category: str,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    caplog: pytest.LogCaptureFixture,
):
    from bonus_platform.engine.labor.blob_storage import LaborBlobError
    from bonus_platform.engine.social_insurance import prefetch, publication
    from bonus_platform.engine.social_insurance.persistent_storage import (
        SocialInsuranceStorageError,
    )

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    previous = _record(identity="TEST-ID-DIAGNOSTIC-PREVIOUS", name="上一成功版本员工")
    previous["source"]["subject"] = "测试主体甲"
    previous_release = publication.materialize_all_subject_runs(
        records=[previous],
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject_options=[{"value": "测试主体甲", "label": "测试主体甲"}],
    )

    current = _record(identity="TEST-ID-DIAGNOSTIC-CURRENT", name="本次诊断员工")
    current["source"]["subject"] = "测试主体甲"
    monkeypatch.setattr(
        prefetch,
        "sync_beisen_candidates",
        lambda **_kwargs: ([current], {"provider": "beisen-open-platform"}),
    )
    monkeypatch.setattr(
        prefetch,
        "prepare_beisen_supplement_pool",
        lambda *_args, **_kwargs: (
            [current],
            {"state": "ready", "recordCount": 1, "cachedAt": "2026-08-26T00:00:00Z"},
        ),
    )
    monkeypatch.setattr(
        prefetch,
        "cached_beisen_contract_subjects",
        lambda **_kwargs: [{"value": "测试主体甲", "label": "测试主体甲"}],
    )
    monkeypatch.setattr(prefetch, "_current_reporting_context", lambda: {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-26",
        "subject": "*",
    })
    sensitive_marker = "SECRET-PUBLICATION-STAGE-440301199001010033"

    def fail_publication(*_args, **_kwargs):
        if expected_stage == "batch_materialize":
            raise RunValidationError(sensitive_marker)
        code = (
            "LABOR_BLOB_TIMEOUT"
            if expected_stage == "index_publish"
            else "LABOR_BLOB_PERMISSION_DENIED"
        )
        try:
            raise LaborBlobError(
                code,
                sensitive_marker,
                retryable=expected_stage == "index_publish",
            )
        except LaborBlobError as exc:
            raise SocialInsuranceStorageError(sensitive_marker) from exc

    target_name = {
        "batch_materialize": "materialize_subject_run",
        "index_publish": "publish_supplement_search_indexes",
        "release_publish": "_persist_release",
    }[expected_stage]
    monkeypatch.setattr(publication, target_name, fail_publication)

    with caplog.at_level(logging.WARNING, logger="bonus_platform.social_insurance.prefetch"):
        result = prefetch.refresh_latest_reporting_snapshot()

    assert result["state"] == "error"
    assert result["failedStage"] == expected_stage
    assert result["errorCategory"] == expected_category
    assert list(result["stageTimingsMs"])[-1] == expected_stage
    assert result["elapsedMs"] >= sum(result["stageTimingsMs"].values())
    assert sensitive_marker not in json.dumps(result, ensure_ascii=False)
    assert sensitive_marker not in caplog.text
    latest = publication.load_latest_reporting_release()
    assert latest is not None
    assert latest["id"] == previous_release["releaseId"]


def test_reporting_scheduler_refreshes_one_shared_all_subject_period_snapshot(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import prefetch

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    create_run(
        records=[_record(identity="TEST-ID-SCHEDULE-CONTEXT", name="已有批次员工")],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-21",
        subject="测试主体甲",
        source="beisen",
        source_summary={"provider": "beisen-open-platform"},
    )
    records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙"), start=1):
        record = _record(identity=f"TEST-ID-SCHEDULE-ALL-{index}", name=f"定时员工{index}")
        record["source"]["subject"] = subject
        records.append(record)
    supplement_records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙", "测试主体丙"), start=1):
        record = _record(identity=f"TEST-ID-SCHEDULE-SUPPLEMENT-{index}", name=f"历史候选{index}")
        record["source"]["subject"] = subject
        supplement_records.append(record)
    calls: list[str] = []
    supplement_prefetch_calls: list[tuple[dict, bool]] = []

    def scheduled_sync(**kwargs):
        calls.append(kwargs["subject"])
        return records, {"provider": "beisen-open-platform", "warnings": []}

    monkeypatch.setattr(prefetch, "sync_beisen_candidates", scheduled_sync)
    monkeypatch.setattr(
        prefetch,
        "prepare_beisen_supplement_pool",
        lambda run, force=False: (
            supplement_prefetch_calls.append((deepcopy(run), force))
            or (
                supplement_records,
                {
                    "state": "ready",
                    "cachedAt": "2026-08-24T00:00:00Z",
                    "recordCount": len(supplement_records),
                },
            )
        ),
    )
    monkeypatch.setattr(prefetch, "cached_beisen_contract_subjects", lambda **_kwargs: [
        {"value": "测试主体甲", "label": "测试主体甲", "code": "A", "candidateCount": 1},
        {"value": "测试主体乙", "label": "测试主体乙", "code": "B", "candidateCount": 1},
        {"value": "测试主体丙", "label": "测试主体丙", "code": "C", "candidateCount": 0},
    ])
    monkeypatch.setattr(
        prefetch,
        "list_beisen_contract_subjects",
        lambda **_kwargs: pytest.fail("定时发布不应重复实时查询合同主体"),
    )
    monkeypatch.setattr(prefetch, "_current_reporting_context", lambda: {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-24",
        "subject": "*",
    }, raising=False)

    result = prefetch.refresh_latest_reporting_snapshot()

    assert result["state"] == "ready"
    assert calls == ["*"]
    assert len(supplement_prefetch_calls) == 1
    supplement_context, supplement_force = supplement_prefetch_calls[0]
    assert supplement_force is True
    assert supplement_context["periodStart"] == "2026-07-16"
    assert supplement_context["periodEnd"] == "2026-08-15"
    assert supplement_context["confirmationDate"] == "2026-08-24"
    assert supplement_context["subject"] == "*"
    assert result["supplementCandidateCount"] == 3
    assert result["supplementSearchIndexCount"] == 3
    assert set(result["stageTimingsMs"]) == {
        "current_sync",
        "snapshot_persist",
        "supplement_pool",
        "subjects",
        "batch_materialize",
        "index_publish",
        "release_publish",
    }
    assert all(
        isinstance(duration, int) and duration >= 0
        for duration in result["stageTimingsMs"].values()
    )
    assert result["elapsedMs"] >= sum(result["stageTimingsMs"].values())
    assert "failedStage" not in result
    assert "errorCategory" not in result
    snapshot = load_reporting_snapshot(
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-24",
        subject="*",
    )
    assert snapshot is not None
    assert snapshot["stale"] is False
    assert {record["source"]["subject"] for record in snapshot["records"]} == {"测试主体甲", "测试主体乙"}

    from bonus_platform.engine.social_insurance.publication import load_latest_reporting_release

    release = load_latest_reporting_release()
    assert release is not None
    assert result["releaseId"] == release["id"]
    assert result["batchCount"] == 3
    assert release["periodStart"] == "2026-07-16"
    assert release["periodEnd"] == "2026-08-15"
    assert release["confirmationDate"] == "2026-08-24"
    assert [subject["value"] for subject in release["subjects"]] == ["测试主体甲", "测试主体乙", "测试主体丙"]
    assert [subject["candidateCount"] for subject in release["subjects"]] == [1, 1, 0]
    assert all(subject.get("runId") for subject in release["subjects"])
    assert all(load_run(subject["runId"])["subject"] == subject["value"] for subject in release["subjects"])
    from bonus_platform.engine.social_insurance import supplements

    subject_b = next(subject for subject in release["subjects"] if subject["value"] == "测试主体乙")
    supplements.clear_supplement_search_index_cache()
    index = supplements.load_supplement_search_index(subject_b["runId"])
    assert index is not None
    assert [candidate["name"] for candidate in index["candidates"]] == ["历史候选2"]
    assert "定时员工" not in json.dumps(release, ensure_ascii=False)
    assert "TEST-ID-SCHEDULE-ALL" not in json.dumps(release, ensure_ascii=False)


def test_reporting_subject_options_merge_current_records_and_cache_without_a_live_query():
    from bonus_platform.engine.social_insurance import prefetch

    current_a = _record(identity="TEST-ID-SUBJECT-OPTION-A", name="当前员工甲")
    current_a["source"].update({"subject": "测试主体甲", "subjectCode": "A"})
    current_b = _record(identity="TEST-ID-SUBJECT-OPTION-B", name="当前员工乙")
    current_b["source"].update({"subject": "测试主体乙", "subjectCode": "B"})

    options = prefetch._publication_subject_options(
        records=[current_a, current_b],
        cached_subjects=[
            {"value": "旧主体乙标识", "label": "主体乙", "code": "B", "candidateCount": 99},
            {"value": "测试主体丙", "label": "主体丙", "code": "C", "candidateCount": 0},
        ],
    )

    assert options == [
        {"value": "测试主体甲", "label": "测试主体甲", "code": "A", "candidateCount": 1},
        {"value": "测试主体乙", "label": "主体乙", "code": "B", "candidateCount": 1},
        {"value": "测试主体丙", "label": "主体丙", "code": "C", "candidateCount": 0},
    ]


def test_failed_all_subject_publication_keeps_the_previous_successful_release(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import publication

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    first = _record(identity="TEST-ID-RELEASE-001", name="上一成功版本员工")
    first["source"]["subject"] = "测试主体甲"
    published = publication.materialize_all_subject_runs(
        records=[first],
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject_options=[{"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1}],
    )
    previous_release_id = published["releaseId"]

    original_materialize = publication.materialize_subject_run
    calls = 0

    def fail_on_second_subject(**kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RunValidationError("模拟第二个主体生成失败")
        return original_materialize(**kwargs)

    monkeypatch.setattr(publication, "materialize_subject_run", fail_on_second_subject)
    second_records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙"), start=1):
        record = _record(identity=f"TEST-ID-RELEASE-FAIL-{index}", name=f"失败版本员工{index}")
        record["source"]["subject"] = subject
        second_records.append(record)

    with pytest.raises(RunValidationError, match="模拟第二个主体生成失败"):
        publication.materialize_all_subject_runs(
            records=second_records,
            source_summary={"provider": "beisen-open-platform"},
            period_start="2026-07-16",
            period_end="2026-08-15",
            confirmation_date="2026-08-25",
            subject_options=[
                {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
                {"value": "测试主体乙", "label": "测试主体乙", "candidateCount": 1},
            ],
        )

    latest = publication.load_latest_reporting_release()
    assert latest is not None
    assert latest["id"] == previous_release_id


def test_bootstrap_reads_one_published_release_without_scanning_or_live_beisen(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import publication
    from bonus_platform.engine.social_insurance import router as social_router

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙"), start=1):
        record = _record(identity=f"TEST-ID-BOOTSTRAP-{index}", name=f"首屏员工{index}")
        record["source"]["subject"] = subject
        records.append(record)
    published = publication.materialize_all_subject_runs(
        records=records,
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        preferred_subject="测试主体乙",
        subject_options=[
            {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
            {"value": "测试主体乙", "label": "测试主体乙", "candidateCount": 1},
        ],
    )

    monkeypatch.setattr(social_router, "list_runs", lambda *_args, **_kwargs: pytest.fail("首屏不应扫描批次"))
    monkeypatch.setattr(
        social_router,
        "list_beisen_contract_subjects",
        lambda **_kwargs: pytest.fail("首屏不应实时查询北森合同主体"),
    )
    response = TestClient(app).get("/api/social-insurance/bootstrap")

    assert response.status_code == 200
    payload = response.json()
    assert payload["state"] == "ready"
    assert payload["release"]["id"] == published["releaseId"]
    assert payload["config"] == {
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "confirmationDate": "2026-08-16",
    }
    assert [subject["value"] for subject in payload["subjects"]] == ["测试主体甲", "测试主体乙"]
    assert payload["selectedSubject"] == "测试主体乙"
    assert payload["run"] is None
    assert payload["preflight"] is None


def test_bootstrap_returns_the_published_batch_when_it_uses_the_period_default_confirmation(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import publication
    from bonus_platform.engine.social_insurance import router as social_router

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    record = _record(identity="TEST-ID-BOOTSTRAP-DEFAULT", name="默认确认日员工")
    record["source"]["subject"] = "测试主体甲"
    publication.materialize_all_subject_runs(
        records=[record],
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-16",
        preferred_subject="测试主体甲",
        subject_options=[{"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1}],
    )

    monkeypatch.setattr(social_router, "list_runs", lambda *_args, **_kwargs: pytest.fail("首屏不应扫描批次"))
    monkeypatch.setattr(
        social_router,
        "list_beisen_contract_subjects",
        lambda **_kwargs: pytest.fail("首屏不应实时查询北森合同主体"),
    )

    payload = TestClient(app).get("/api/social-insurance/bootstrap").json()

    assert payload["config"]["confirmationDate"] == "2026-08-16"
    assert payload["run"]["subject"] == "测试主体甲"
    assert payload["preflight"]["runId"] == payload["run"]["id"]


def test_subject_switch_stays_on_the_same_release_when_a_partial_new_run_exists(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import publication

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SNAPSHOTS_DIR", str(tmp_path / "snapshots"))
    released_records = []
    for index, subject in enumerate(("测试主体甲", "测试主体乙"), start=1):
        record = _record(identity=f"TEST-ID-ATOMIC-{index}", name=f"发布版本员工{index}")
        record["source"]["subject"] = subject
        released_records.append(record)
    published = publication.materialize_all_subject_runs(
        records=released_records,
        source_summary={"provider": "beisen-open-platform"},
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject_options=[
            {"value": "测试主体甲", "label": "测试主体甲", "candidateCount": 1},
            {"value": "测试主体乙", "label": "测试主体乙", "candidateCount": 1},
        ],
    )
    released_subject = next(
        subject for subject in published["release"]["subjects"] if subject["value"] == "测试主体乙"
    )
    partial = _record(identity="TEST-ID-PARTIAL-NEW", name="尚未发布的新员工")
    partial["source"]["subject"] = "测试主体乙"
    newer = create_run(
        records=[partial],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject="测试主体乙",
        source="beisen",
    )
    assert load_run_index(
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-25",
        subject="测试主体乙",
    )["runId"] == newer["id"]

    response = TestClient(app).get(
        f"/api/social-insurance/releases/{published['releaseId']}/runs/current",
        params={"subject": "测试主体乙"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["lookupSource"] == "reporting-release"
    assert payload["run"]["id"] == released_subject["runId"]
    assert payload["run"]["employees"][0]["report"]["姓名"] == "发布版本员工2"
    assert "尚未发布的新员工" not in response.text


def test_contract_subject_options_are_derived_without_employee_details(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    first = _record(identity="TEST-ID-SUBJECT-001", name="主体枚举员工甲")
    second = _record(identity="TEST-ID-SUBJECT-002", name="主体枚举员工乙")
    second["source"]["subject"] = "测试合同主体二"
    fixture_path = tmp_path / "beisen-subject-fixture.json"
    fixture_path.write_text(
        json.dumps({"records": [first, second]}, ensure_ascii=False),
        encoding="utf-8",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", str(fixture_path))

    response = TestClient(app).get(
        "/api/social-insurance/subjects",
        params={"periodStart": "2026-07-16", "periodEnd": "2026-08-15"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["subjects"] == [
        {"value": "深圳市前海云途物流有限公司", "label": "深圳市前海云途物流有限公司", "code": "", "candidateCount": 1},
        {"value": "测试合同主体二", "label": "测试合同主体二", "code": "", "candidateCount": 1},
    ]
    serialized = response.text
    assert "主体枚举员工" not in serialized
    assert "TEST-ID-SUBJECT" not in serialized


def test_contract_subjects_do_not_scan_recent_runs_or_call_live_beisen_without_a_release(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import router as social_router

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    current = _record(identity="TEST-ID-SAVED-SUBJECT-001", name="已同步员工")
    create_run(
        records=[current],
        period_start="2026-06-16",
        period_end="2026-07-15",
        confirmation_date="2026-07-17",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(tmp_path / "missing-engine"))
    monkeypatch.setattr(
        social_router,
        "list_runs",
        lambda *_args, **_kwargs: pytest.fail("没有发布版本时也不应扫描历史批次"),
    )
    monkeypatch.setattr(
        social_router,
        "list_beisen_contract_subjects",
        lambda **_kwargs: pytest.fail("普通页面读取不应实时查询北森"),
    )

    response = TestClient(app).get(
        "/api/social-insurance/subjects",
        params={"periodStart": "2026-07-16", "periodEnd": "2026-08-15"},
    )

    assert response.status_code == 200
    assert response.json() == {
        "subjects": [],
        "periodStart": "2026-07-16",
        "periodEnd": "2026-08-15",
        "source": "no-published-release",
    }


def test_contract_subjects_prefer_complete_cached_options_to_partial_recent_runs(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import adapter

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    current = _record(identity="TEST-ID-CACHED-SUBJECT-001", name="已同步员工")
    create_run(
        records=[current],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-18",
        subject="深圳市前海云途物流有限公司",
        source="beisen",
    )
    engine_dir = tmp_path / "engine"
    (engine_dir / "lib").mkdir(parents=True)
    (engine_dir / "lib" / "beisen-client.mjs").write_text("", encoding="utf-8")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setattr(adapter, "_node_binary", lambda: "node")
    monkeypatch.setattr(
        adapter.subprocess,
        "run",
        lambda *args, **kwargs: subprocess.CompletedProcess(
            args=args,
            returncode=0,
            stdout=json.dumps({
                "subjects": [
                    {"value": "深圳市前海云途物流有限公司", "label": "深圳市前海云途物流有限公司", "candidateCount": 34},
                    {"value": "深圳市测试合同主体二", "label": "深圳市测试合同主体二", "candidateCount": 8},
                ],
            }, ensure_ascii=False) + "\n",
            stderr="",
        ),
    )
    adapter.clear_contract_subject_cache()
    assert len(adapter.list_beisen_contract_subjects(
        period_start="2026-07-16",
        period_end="2026-08-15",
        force_refresh=True,
    )) == 2

    response = TestClient(app).get(
        "/api/social-insurance/subjects",
        params={"periodStart": "2026-07-16", "periodEnd": "2026-08-15"},
    )

    assert response.status_code == 200
    assert response.json()["source"] == "beisen-contract-cache"
    assert len(response.json()["subjects"]) == 2


def test_contract_subject_options_include_a_subject_only_present_in_monthly_baseline(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import router as social_router

    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    historical = _record(identity="TEST-ID-BASELINE-OPTION", name="基线选项员工")
    historical["source"]["subject"] = "仅基线主体"
    capture_monthly_baseline(
        records=[historical],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="仅基线主体",
        source="beisen-monthly-snapshot",
    )
    monkeypatch.setattr(
        social_router,
        "cached_beisen_contract_subjects",
        lambda **_kwargs: [{"value": "当前主体", "label": "当前主体", "candidateCount": 4}],
    )

    response = TestClient(app).get(
        "/api/social-insurance/subjects",
        params={"periodStart": "2026-07-16", "periodEnd": "2026-08-15"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["subjects"] == [
        {"value": "当前主体", "label": "当前主体", "candidateCount": 4},
        {"value": "仅基线主体", "label": "仅基线主体", "code": "", "candidateCount": 1},
    ]
    assert payload["baselineSubjectCount"] == 1


def test_manual_contract_subject_refresh_failure_keeps_complete_cached_options(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import adapter
    from bonus_platform.engine.social_insurance import router as social_router

    engine_dir = tmp_path / "engine"
    (engine_dir / "lib").mkdir(parents=True)
    (engine_dir / "lib" / "beisen-client.mjs").write_text("", encoding="utf-8")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_BASELINES_DIR", str(tmp_path / "baselines"))
    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    monkeypatch.setattr(
        adapter.subprocess,
        "run",
        lambda *args, **kwargs: subprocess.CompletedProcess(
            args=args,
            returncode=0,
            stdout='{"subjects":[{"value":"缓存主体一","label":"缓存主体一","candidateCount":12}]}' + "\n",
            stderr="",
        ),
    )
    adapter.clear_contract_subject_cache()
    adapter.list_beisen_contract_subjects(
        period_start="2026-07-16",
        period_end="2026-08-15",
        force_refresh=True,
    )
    monkeypatch.setattr(
        social_router,
        "list_beisen_contract_subjects",
        lambda **kwargs: (_ for _ in ()).throw(RunValidationError("北森合同主体加载失败")),
    )

    response = TestClient(app).get(
        "/api/social-insurance/subjects",
        params={"periodStart": "2026-07-16", "periodEnd": "2026-08-15", "refresh": "true"},
    )

    assert response.status_code == 200
    assert response.json()["source"] == "beisen-contract-cache"
    assert response.json()["subjects"] == [
        {"value": "缓存主体一", "label": "缓存主体一", "code": "", "candidateCount": 12}
    ]
    assert response.json()["refreshWarning"] == "北森实时刷新暂时失败，已保留最近缓存主体"


def test_live_contract_subject_queries_are_serialized_and_cached(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import adapter

    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    engine_dir = tmp_path / "engine"
    (engine_dir / "lib").mkdir(parents=True)
    (engine_dir / "lib" / "beisen-client.mjs").write_text("", encoding="utf-8")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setattr(adapter, "_node_binary", lambda: "node")

    active = 0
    max_active = 0
    call_count = 0
    guard = threading.Lock()

    def fake_run(*args, **kwargs):
        nonlocal active, max_active, call_count
        with guard:
            active += 1
            max_active = max(max_active, active)
            call_count += 1
        time.sleep(0.05)
        with guard:
            active -= 1
        return subprocess.CompletedProcess(args=args, returncode=0, stdout='{"subjects": []}\n', stderr="")

    monkeypatch.setattr(adapter.subprocess, "run", fake_run)
    adapter.clear_contract_subject_cache()
    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(
            lambda _: adapter.list_beisen_contract_subjects(
                period_start="2026-07-16",
                period_end="2026-08-15",
            ),
            range(2),
        ))

    assert results == [[], []]
    assert max_active == 1
    assert call_count == 1

    with adapter._SUBJECT_CACHE_LOCK:
        adapter._SUBJECT_CACHE.clear()
    assert adapter.list_beisen_contract_subjects(
        period_start="2026-07-16",
        period_end="2026-08-15",
    ) == []
    assert call_count == 1


def test_concurrent_forced_contract_subject_refreshes_share_one_live_query(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    from bonus_platform.engine.social_insurance import adapter

    monkeypatch.delenv("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE", raising=False)
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    engine_dir = tmp_path / "engine"
    (engine_dir / "lib").mkdir(parents=True)
    (engine_dir / "lib" / "beisen-client.mjs").write_text("", encoding="utf-8")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR", str(engine_dir))
    monkeypatch.setattr(adapter, "_node_binary", lambda: "node")
    call_count = 0
    guard = threading.Lock()

    def fake_run(*args, **kwargs):
        nonlocal call_count
        with guard:
            call_count += 1
        time.sleep(0.05)
        return subprocess.CompletedProcess(
            args=args,
            returncode=0,
            stdout='{"subjects":[{"value":"测试主体","label":"测试主体","candidateCount":1}]}\n',
            stderr="",
        )

    monkeypatch.setattr(adapter.subprocess, "run", fake_run)
    adapter.clear_contract_subject_cache()
    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(
            lambda _: adapter.list_beisen_contract_subjects(
                period_start="2026-07-16",
                period_end="2026-08-15",
                force_refresh=True,
            ),
            range(2),
        ))

    assert results == [[{"value": "测试主体", "label": "测试主体", "code": "", "candidateCount": 1}]] * 2
    assert call_count == 1



def test_business_rules_api_is_versioned_and_uses_business_language():
    response = TestClient(app).get("/api/social-insurance/rules")

    assert response.status_code == 200
    payload = response.json()
    assert payload["version"]
    assert payload["updatedAt"] == "2026-08-24"
    assert payload["scope"] == "全国社保增员"
    assert any(section["id"] == "candidate-list" for section in payload["sections"])
    serialized = response.text
    assert "人员怎么进入本期名单" in serialized
    assert "月度名单基线" in serialized
    assert "虚拟员工不进入候选" in serialized
    assert "上期漏报" in serialized
    assert "延迟增员" in serialized
    assert "补缴" not in serialized
    assert "API" not in serialized
    assert "JSON" not in serialized
    assert "全国候选不按工作地点提前过滤" in serialized
    assert "少量无模板城市进入线下办理清单" in serialized


def test_social_insurance_module_exposes_validated_flow():
    static_dir = Path(__file__).resolve().parents[1] / "bonus_platform" / "static"
    html = (static_dir / "social-insurance.html").read_text(encoding="utf-8")
    script = (static_dir / "social-insurance.js").read_text(encoding="utf-8")

    for label in ("生成本批名单", "人工确认", "导入政务模板", "生成报盘", "人工最终确认"):
        assert label in html
    assert "自动化待接入" in html
    assert '<select id="subject"' in html
    assert "/api/social-insurance/subjects" in script
    assert "payload.refreshWarning" in script
    assert "/api/social-insurance/runs/sync" in script
    assert "/api/social-insurance/runs/" in script


def test_run_infers_independent_social_and_medical_tasks_from_offline_status_text(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    record = _record(identity="TEST-ID-COVERAGE-001", name="成都规则样例")
    record["coverageSource"] = {
        "socialPlace": "成都",
        "socialMedicalStatus": "社保待审核，医保补缴6、7月份",
        "housingStatus": "未封存",
    }

    run = create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="测试合同主体",
        source="offline-replay",
    )

    employee = run["employees"][0]
    assert employee["coverageTasks"]["social"] == {
        "coverage": "social",
        "label": "社保",
        "status": "needs_review",
        "statusLabel": "待确认",
        "actionMonth": "",
        "supplementMonths": [],
        "route": "chengdu-social",
        "routeLabel": "成都社保模板",
        "handling": "template",
        "reason": "线下源表标记社保待审核",
    }
    assert employee["coverageTasks"]["medical"]["status"] == "supplement"
    assert employee["coverageTasks"]["medical"]["supplementMonths"] == ["6", "7"]
    assert employee["coverageTasks"]["medical"]["route"] == "chengdu-medical"
    assert employee["coverageTasks"]["housing"]["status"] == "deferred"
    assert run["summary"]["coverageNeedsReview"] == 2


def test_run_routes_major_template_families_and_keeps_unknown_city_as_manual_handling(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    shenzhen = _record(identity="TEST-ID-COVERAGE-002", name="深圳规则样例")
    shenzhen["coverageSource"] = {
        "socialPlace": "深圳",
        "socialMedicalStatus": "社保ok，医保ok",
    }
    xiamen = _record(identity="TEST-ID-COVERAGE-003", name="小量城市样例")
    xiamen["coverageSource"] = {
        "socialPlace": "厦门",
        "socialMedicalStatus": "社保待审核，医保待审核",
    }

    run = create_run(
        records=[shenzhen, xiamen],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="测试合同主体",
        source="offline-replay",
    )

    first, second = run["employees"]
    assert first["coverageTasks"]["social"]["route"] == "shenzhen-social-medical"
    assert first["coverageTasks"]["medical"]["route"] == "shenzhen-social-medical"
    assert first["coverageTasks"]["social"]["status"] == "completed"
    assert first["coverageTasks"]["medical"]["status"] == "completed"
    assert second["coverageTasks"]["social"]["handling"] == "manual"
    assert second["coverageTasks"]["medical"]["handling"] == "manual"
    assert second["coverageTasks"]["social"]["routeLabel"] == "线下办理（暂不生成模板）"
    assert run["summary"]["manualHandling"] == 2
    assert run["status"] == "draft"


def test_run_carries_explicit_future_enrollment_month_without_blocking_other_people(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    record = _record(identity="TEST-ID-COVERAGE-004", name="跨月规则样例")
    record["coverageSource"] = {
        "socialPlace": "成都",
        "socialMedicalStatus": "社保9月购买，医保ok",
    }

    run = create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="测试合同主体",
        source="offline-replay",
    )

    tasks = run["employees"][0]["coverageTasks"]
    assert tasks["social"]["status"] == "scheduled"
    assert tasks["social"]["actionMonth"] == "9"
    assert tasks["medical"]["status"] == "completed"
    assert run["summary"]["coverageScheduled"] == 1


def test_audit_export_contains_coverage_status_and_route_in_the_single_sheet(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(tmp_path / "runs"))
    record = _record(identity="TEST-ID-COVERAGE-005", name="审核导出样例")
    record["coverageSource"] = {
        "socialPlace": "杭州",
        "socialMedicalStatus": "社保待审核，医保ok",
    }
    run = create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="测试合同主体",
        source="offline-replay",
    )

    workbook = load_workbook(build_audit_export(run["id"]), read_only=True, data_only=True)
    assert workbook.sheetnames == ["全部增员人员"]
    sheet = workbook["全部增员人员"]
    headers = [cell.value for cell in sheet[4]]
    row = [cell.value for cell in sheet[5]]
    assert headers[headers.index("社保状态")] == "社保状态"
    assert row[headers.index("社保状态")] == "待确认"
    assert row[headers.index("医保状态")] == "已完成"
    assert row[headers.index("办理去向")] == "浙江社保医保合并模板"
    workbook.close()
