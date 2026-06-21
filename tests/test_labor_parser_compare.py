import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bonus_platform.engine.labor.parser_compare import (
    _pymupdf4llm_text,
    build_expected_results_template_for_materials,
    build_parser_comparison,
    build_parser_comparison_for_materials,
    main as parser_compare_main,
    summarize_business_signals,
    write_parser_comparison_report,
)


def test_summarize_business_signals_detects_amount_name_and_hours():
    text = """
    Invoice Total $1,234.56
    Mucu, Pablo 40.00 hours $812.80
    Lopez Elizabeth 8 hrs $160.00
    """

    signals = summarize_business_signals(text)

    assert signals["hasAmountSignal"] is True
    assert signals["hasEmployeeNameSignal"] is True
    assert signals["hasHoursSignal"] is True
    assert signals["moneyCandidateCount"] >= 3
    assert signals["employeeNameSignalCount"] >= 2


def test_build_parser_comparison_lists_target_parsers_without_touching_formal_flow(tmp_path):
    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\nInvoice Total $100.00\nMucu, Pablo 40 hours $800.00\n")

    report = build_parser_comparison([pdf])

    parser_names = {row["parser"] for row in report["parserResults"]}
    assert {"LiteParse", "PyMuPDF4LLM", "Docling", "Marker", "pdfplumber", "Camelot"}.issubset(parser_names)
    assert report["summary"]["fileCount"] == 1
    assert report["summary"]["parserCount"] >= 6
    assert report["recommendation"]["formalFlowChanged"] is False


def test_build_parser_comparison_for_materials_supports_deterministic_sample(tmp_path):
    materials = tmp_path / "materials"
    materials.mkdir()
    for name in ["c.pdf", "a.pdf", "b.pdf"]:
        (materials / name).write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")

    report = build_parser_comparison_for_materials(materials, sample_size=2)

    assert report["summary"]["fileCount"] == 2
    assert report["summary"]["totalFileCount"] == 3
    assert report["summary"]["sampled"] is True
    assert report["summary"]["sampleSize"] == 2
    assert report["summary"]["selectedFiles"] == ["a.pdf", "b.pdf"]


def test_build_expected_results_template_for_materials_uses_same_sample_order(tmp_path):
    materials = tmp_path / "materials"
    materials.mkdir()
    for name in ["c.pdf", "a.pdf", "b.pdf"]:
        (materials / name).write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")

    template = build_expected_results_template_for_materials(materials, sample_size=2)

    assert template["source"] == "labor_parser_expected_results_template"
    assert template["summary"] == {"totalFileCount": 3, "templateFileCount": 2, "sampled": True}
    assert list(template["files"]) == ["a.pdf", "b.pdf"]
    assert template["files"]["a.pdf"]["invoice_total"] is None
    assert "填写人工确认后的 PDF 发票总额" in template["instructions"][0]


def test_build_parser_comparison_marks_expected_invoice_total_when_candidate_amount_matches(tmp_path):
    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\nInvoice Total $100.00\nMucu, Pablo 40 hours $800.00\n")

    report = build_parser_comparison(
        [pdf],
        expected_results={"files": {"invoice.pdf": {"invoice_total": 100.0}}},
        parser_defs=[{"name": "TestParser", "module": "json", "runner": lambda _: "Invoice Total $100.00"}],
    )

    assert report["accuracySummary"]["expectedFileCount"] == 1
    assert report["accuracySummary"]["invoiceTotalMatchedFileCount"] == 1
    assert report["parserResults"][0]["accuracy"]["invoiceTotalMatched"] is True


def test_empty_expected_invoice_total_is_not_counted_as_confirmed_answer(tmp_path):
    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")

    report = build_parser_comparison(
        [pdf],
        expected_results={"files": {"invoice.pdf": {"invoice_total": None}}},
        parser_defs=[{"name": "TestParser", "module": "json", "runner": lambda _: "Invoice Total $100.00"}],
    )

    assert report["accuracySummary"]["expectedFileCount"] == 0
    assert report["parserResults"][0]["accuracy"]["hasExpectedResult"] is False


def test_business_report_explains_invoice_total_accuracy_when_expected_results_exist(tmp_path):
    report = {
        "summary": {
            "fileCount": 1,
            "parserCount": 1,
            "availableParserCount": 1,
            "unavailableParserCount": 0,
            "resultCount": 1,
            "usableResultCount": 1,
            "totalFileCount": 1,
            "sampled": False,
            "sampleSize": 1,
            "selectedFiles": ["invoice.pdf"],
        },
        "accuracySummary": {
            "expectedFileCount": 1,
            "invoiceTotalMatchedFileCount": 1,
            "invoiceTotalMissedFileCount": 0,
        },
        "recommendation": {"formalFlowChanged": False, "nextStep": "先用真实材料对比解析效果。"},
        "parserResults": [
            {
                "parser": "TestParser",
                "file": "invoice.pdf",
                "available": True,
                "status": "可继续评估总金额",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
                "accuracy": {"hasExpectedResult": True, "invoiceTotalMatched": True},
            }
        ],
    }
    output_dir = tmp_path / "parser_compare"

    write_parser_comparison_report(report, output_dir)

    markdown = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    assert "## 已确认答案对照" in markdown
    assert "已确认总额命中：1 个" in markdown


def test_build_parser_comparison_explains_when_no_candidate_parser_is_installed(tmp_path, monkeypatch):
    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\nInvoice Total $100.00\nMucu, Pablo 40 hours $800.00\n")

    monkeypatch.setattr("bonus_platform.engine.labor.parser_compare._module_available", lambda _: False)

    report = build_parser_comparison([pdf])

    assert report["summary"]["availableParserCount"] == 0
    assert report["summary"]["unavailableParserCount"] >= 6
    assert "当前本机未安装候选解析器" in report["recommendation"]["nextStep"]
    assert "正式核对流程未改变" in report["recommendation"]["nextStep"]


def test_business_report_collapses_file_rows_when_no_parser_is_available(tmp_path, monkeypatch):
    pdf_a = tmp_path / "invoice-a.pdf"
    pdf_b = tmp_path / "invoice-b.pdf"
    pdf_a.write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    pdf_b.write_bytes(b"%PDF-1.4\nInvoice Total $200.00\n")
    monkeypatch.setattr("bonus_platform.engine.labor.parser_compare._module_available", lambda _: False)

    report = build_parser_comparison([pdf_a, pdf_b])
    output_dir = tmp_path / "parser_compare"
    write_parser_comparison_report(report, output_dir)

    markdown = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    assert "当前没有可运行的候选解析器，暂不展开逐文件明细。" in markdown
    assert "| LiteParse | invoice-a.pdf |" not in markdown


def test_business_report_summarizes_available_parsers_instead_of_listing_every_missing_file(tmp_path):
    report = {
        "summary": {
            "fileCount": 2,
            "parserCount": 2,
            "availableParserCount": 1,
            "unavailableParserCount": 1,
            "resultCount": 4,
            "usableResultCount": 1,
        },
        "recommendation": {"formalFlowChanged": False, "nextStep": "先用真实材料对比解析效果。"},
        "parserResults": [
            {
                "parser": "LiteParse",
                "file": "invoice-a.pdf",
                "available": False,
                "status": "本机未安装",
                "signals": {"hasAmountSignal": False, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
            },
            {
                "parser": "LiteParse",
                "file": "invoice-b.pdf",
                "available": False,
                "status": "本机未安装",
                "signals": {"hasAmountSignal": False, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
            },
            {
                "parser": "pdfplumber",
                "file": "invoice-a.pdf",
                "available": True,
                "status": "可继续评估员工明细",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": True, "hasHoursSignal": True},
            },
            {
                "parser": "pdfplumber",
                "file": "invoice-b.pdf",
                "available": True,
                "status": "可继续评估总金额",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
            },
        ],
    }
    output_dir = tmp_path / "parser_compare"

    write_parser_comparison_report(report, output_dir)

    markdown = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    assert "## 解析器表现" in markdown
    assert "| pdfplumber | 2 | 2 | - | 1 | 1 |" in markdown
    assert "建议优先复测 pdfplumber" in markdown
    assert "未安装的解析器不展开逐文件明细" in markdown
    assert "| LiteParse | invoice-a.pdf |" not in markdown


def test_business_report_recommends_parallel_retest_when_top_parsers_are_tied(tmp_path):
    report = {
        "summary": {
            "fileCount": 1,
            "parserCount": 2,
            "availableParserCount": 2,
            "unavailableParserCount": 0,
            "resultCount": 2,
            "usableResultCount": 2,
        },
        "recommendation": {"formalFlowChanged": False, "nextStep": "先用真实材料对比解析效果。"},
        "parserResults": [
            {
                "parser": "PyMuPDF4LLM",
                "file": "invoice-a.pdf",
                "available": True,
                "status": "可继续评估员工明细",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": True, "hasHoursSignal": True},
            },
            {
                "parser": "pdfplumber",
                "file": "invoice-a.pdf",
                "available": True,
                "status": "可继续评估员工明细",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": True, "hasHoursSignal": True},
            },
        ],
    }
    output_dir = tmp_path / "parser_compare"

    write_parser_comparison_report(report, output_dir)

    markdown = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    assert "建议同时复测 PyMuPDF4LLM、pdfplumber" in markdown


def test_business_report_shows_confirmed_invoice_total_hits_by_parser(tmp_path):
    report = {
        "summary": {
            "fileCount": 2,
            "parserCount": 2,
            "availableParserCount": 2,
            "unavailableParserCount": 0,
            "resultCount": 4,
            "usableResultCount": 3,
        },
        "accuracySummary": {
            "expectedFileCount": 2,
            "invoiceTotalMatchedFileCount": 2,
            "invoiceTotalMissedFileCount": 0,
        },
        "recommendation": {"formalFlowChanged": False, "nextStep": "先用真实材料对比解析效果。"},
        "parserResults": [
            {
                "parser": "pdfplumber",
                "file": "invoice-a.pdf",
                "available": True,
                "status": "可继续评估总金额",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
                "accuracy": {"hasExpectedResult": True, "invoiceTotalMatched": True},
            },
            {
                "parser": "pdfplumber",
                "file": "invoice-b.pdf",
                "available": True,
                "status": "可继续评估总金额",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
                "accuracy": {"hasExpectedResult": True, "invoiceTotalMatched": False},
            },
            {
                "parser": "PyMuPDF4LLM",
                "file": "invoice-a.pdf",
                "available": True,
                "status": "可继续评估总金额",
                "signals": {"hasAmountSignal": True, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
                "accuracy": {"hasExpectedResult": True, "invoiceTotalMatched": True},
            },
            {
                "parser": "PyMuPDF4LLM",
                "file": "invoice-b.pdf",
                "available": True,
                "status": "暂未发现有效线索",
                "signals": {"hasAmountSignal": False, "hasEmployeeNameSignal": False, "hasHoursSignal": False},
                "accuracy": {"hasExpectedResult": True, "invoiceTotalMatched": False},
            },
        ],
    }
    output_dir = tmp_path / "parser_compare"

    write_parser_comparison_report(report, output_dir)

    markdown = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    assert "| 解析器 | 可运行文件 | 读出金额 | 人工确认总额命中 | 读出员工姓名 | 读出工时 |" in markdown
    assert "| pdfplumber | 2 | 2 | 1 / 2 | 0 | 0 |" in markdown
    assert "| PyMuPDF4LLM | 2 | 1 | 1 / 2 | 0 | 0 |" in markdown


def test_pymupdf4llm_text_handles_short_pdf_without_requesting_missing_pages(tmp_path):
    import fitz

    pdf = tmp_path / "one-page.pdf"
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), "Invoice Total $100.00\nMucu, Pablo 40 hours $800.00")
    doc.save(pdf)
    doc.close()

    text = _pymupdf4llm_text(pdf)

    assert "Invoice Total" in text
    assert "Mucu" in text


def test_parser_compare_cli_writes_business_report(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    (materials / "invoice.pdf").write_bytes(b"%PDF-1.4\nInvoice Total $100.00\nMucu, Pablo 40 hours $800.00\n")
    output_dir = tmp_path / "parser_compare"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--output-dir",
            str(output_dir),
        ]
    )

    assert exit_code == 0
    payload = json.loads((output_dir / "parser_comparison_summary.json").read_text(encoding="utf-8"))
    summary = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    assert payload["summary"]["fileCount"] == 1
    assert "# 免费解析器对比业务摘要" in summary
    assert "正式核对流程未改变" in summary
    assert "LiteParse" in summary
    stdout = capsys.readouterr().out
    assert "已检查 PDF 文件" in stdout
    assert "parserResults" not in stdout


def test_parser_compare_cli_supports_sample_size(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    for name in ["c.pdf", "a.pdf", "b.pdf"]:
        (materials / name).write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    output_dir = tmp_path / "parser_compare"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--output-dir",
            str(output_dir),
            "--sample-size",
            "2",
        ]
    )

    assert exit_code == 0
    payload = json.loads((output_dir / "parser_comparison_summary.json").read_text(encoding="utf-8"))
    summary = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    stdout = capsys.readouterr().out
    assert payload["summary"]["sampled"] is True
    assert payload["summary"]["selectedFiles"] == ["a.pdf", "b.pdf"]
    assert "抽样检查 PDF 文件：2 / 3 个" in stdout
    assert "本报告是抽样结果" in summary


def test_parser_compare_cli_accepts_expected_results_file(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    (materials / "invoice.pdf").write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    expected = tmp_path / "expected.json"
    expected.write_text(
        json.dumps({"files": {"invoice.pdf": {"invoice_total": 100.0}}}, ensure_ascii=False),
        encoding="utf-8",
    )
    output_dir = tmp_path / "parser_compare"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--output-dir",
            str(output_dir),
            "--expected-results",
            str(expected),
        ]
    )

    assert exit_code == 0
    payload = json.loads((output_dir / "parser_comparison_summary.json").read_text(encoding="utf-8"))
    summary = (output_dir / "PARSER_COMPARISON_SUMMARY.md").read_text(encoding="utf-8")
    stdout = capsys.readouterr().out
    assert payload["accuracySummary"]["expectedFileCount"] == 1
    assert "已确认答案对照" in summary
    assert "已载入人工确认答案：1 个文件" in stdout


def test_parser_compare_cli_writes_expected_results_template(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    for name in ["c.pdf", "a.pdf", "b.pdf"]:
        (materials / name).write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    template_path = tmp_path / "expected_template.json"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--sample-size",
            "2",
            "--write-expected-template",
            str(template_path),
        ]
    )

    assert exit_code == 0
    template = json.loads(template_path.read_text(encoding="utf-8"))
    stdout = capsys.readouterr().out
    assert list(template["files"]) == ["a.pdf", "b.pdf"]
    assert "人工确认答案模板已生成" in stdout


def test_parser_compare_cli_writes_business_excel_expected_results_template(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    for name in ["c.pdf", "a.pdf", "b.pdf"]:
        (materials / name).write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    template_path = tmp_path / "expected_template.json"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--sample-size",
            "2",
            "--write-expected-template",
            str(template_path),
        ]
    )

    assert exit_code == 0
    excel_path = tmp_path / "expected_template.xlsx"
    assert excel_path.exists()
    workbook = load_workbook(excel_path, read_only=True)
    rows = list(workbook["人工确认发票总额"].iter_rows(values_only=True))
    assert rows[0] == ("PDF文件", "人工确认发票总额", "复核人", "复核日期", "证据备注")
    assert rows[1][0] == "a.pdf"
    assert rows[1][1] is None
    assert rows[2][0] == "b.pdf"
    stdout = capsys.readouterr().out
    assert "Excel 模板" in stdout


def test_parser_compare_cli_writes_json_sidecar_when_expected_template_path_is_xlsx(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    for name in ["c.pdf", "a.pdf", "b.pdf"]:
        (materials / name).write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    excel_template_path = tmp_path / "expected_template.xlsx"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--sample-size",
            "2",
            "--write-expected-template",
            str(excel_template_path),
        ]
    )

    assert exit_code == 0
    json_template_path = tmp_path / "expected_template.json"
    assert excel_template_path.exists()
    assert json_template_path.exists()
    workbook = load_workbook(excel_template_path, read_only=True)
    rows = list(workbook["人工确认发票总额"].iter_rows(values_only=True))
    assert rows[0] == ("PDF文件", "人工确认发票总额", "复核人", "复核日期", "证据备注")
    template = json.loads(json_template_path.read_text(encoding="utf-8"))
    assert list(template["files"]) == ["a.pdf", "b.pdf"]
    stdout = capsys.readouterr().out
    assert f"人工确认答案模板已生成：{json_template_path}" in stdout
    assert f"Excel 模板已生成：{excel_template_path}" in stdout


def test_parser_compare_cli_accepts_excel_expected_results_file(tmp_path, capsys):
    materials = tmp_path / "materials"
    materials.mkdir()
    (materials / "invoice-a.pdf").write_bytes(b"%PDF-1.4\nInvoice Total $100.00\n")
    (materials / "invoice-b.pdf").write_bytes(b"%PDF-1.4\nInvoice Total $200.00\n")
    expected = tmp_path / "expected.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工确认发票总额"
    sheet.append(["PDF文件", "人工确认发票总额", "复核人", "复核日期", "证据备注"])
    sheet.append(["invoice-a.pdf", 100.0, "ops", "2026-06-21", "首页总额"])
    sheet.append(["invoice-b.pdf", 200.0, "ops", "2026-06-21", "首页总额"])
    workbook.save(expected)
    output_dir = tmp_path / "parser_compare"

    exit_code = parser_compare_main(
        [
            "--materials-root",
            str(materials),
            "--output-dir",
            str(output_dir),
            "--expected-results",
            str(expected),
        ]
    )

    assert exit_code == 0
    payload = json.loads((output_dir / "parser_comparison_summary.json").read_text(encoding="utf-8"))
    stdout = capsys.readouterr().out
    assert payload["accuracySummary"]["expectedFileCount"] == 2
    assert "已载入人工确认答案：2 个文件" in stdout
