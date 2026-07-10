from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from bonus_platform.engine.recruitment_import_validation import (
    CompareResult,
    STANDARD_IMPORT_HEADERS,
    build_auto_import_rows,
    compare_import_rows,
    write_validation_report,
    write_standard_import_workbook,
)


def _save(workbook: Workbook, path: Path) -> Path:
    workbook.save(path)
    return path


def _history_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "附件一-招聘奖金核算明细"
    sheet.append([""] * 60)
    sheet.append([
        "验证唯一", "姓名", "工号", "身份证", "入职日期", "人员状态", "工作地", "最后工作日", "职级",
        "二级组织", "三级组织", "四级组织", "五级组织", "职位", "职族", "职类", "直线经理",
        "直线经理工号", "标签分类\n（国内/海外）", "职族分类\n（职能类/操作类/技术类）", "ABC类别",
        "招聘负责人工号", "招聘负责人姓名", "协助招聘人工号", "协助招聘人姓名",
        "招聘渠道\n（招聘网站、内推、校招、二次入职、猎头）", "海外offer流程编号", "海外招聘需求编号",
        "招聘启动日期", "候选人入职时间", "转正日期",
    ] + [""] * 29)
    sheet.append([""] * 32 + ["周期天数剔除"] + [""] * 24 + ["推荐人姓名", "推荐人工号", "推荐人是否直接/间接上级"])
    sheet.append([
        "唯一", "历史员工", "zt100", "ID100", datetime(2025, 4, 2), "正式", "中国大陆", "", "P1-3",
        "二级", "三级", "", "", "专员", "", "", "经理", "zt900", "国内", "", "C类",
        "海外-zt901", "招聘一", "", "", "招聘网站", "", "", datetime(2025, 3, 20), datetime(2025, 4, 2),
        datetime(2025, 7, 2),
    ] + ["", 7] + [""] * 24 + ["推荐一", "zt902", "否"])
    return _save(workbook, path)


def _ehr_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EHR"
    sheet.append([
        "JobNumber", "parent_Name", "EmployeeStatus", "parent_IDNumber", "EntryDate",
        "RegularizationDate", "LastWorkDate", "LookupPrefix_OIdDepartment_SecondLevelOrganization",
        "LookupPrefix_OIdDepartment_ThirdLevelOrganization", "LookupPrefix_OIdDepartment_FourthLevelOrganization",
        "LookupPrefix_OIdDepartment_FifthLevelOrganization", "OIdJobPosition_Name", "OIdJobLevel",
        "POIdEmpAdmin-ExportName", "POIdEmpAdmin-StaffCode",
    ])
    sheet.append(["zt100", "历史员工", "正式", "ID100", datetime(2025, 4, 2), datetime(2025, 7, 2), "", "二级", "三级", "", "", "专员", "P1-3", "经理", "zt900"])
    sheet.append(["zt200", "新增员工", "试用", "ID200", datetime(2025, 5, 5), "", "", "新二级", "新三级", "", "", "助理", "P1-1", "新经理", "zt901"])
    sheet.append(["zt901", "招聘一", "正式", "ID901", datetime(2024, 1, 1), "", "", "招聘部", "", "", "", "招聘主管", "P2-1", "招聘总监", "zt920"])
    sheet.append(["zt920", "招聘总监", "正式", "ID920", datetime(2023, 1, 1), "", "", "招聘部", "", "", "", "招聘经理", "P3-1", "", ""])
    sheet.append(["zt912", "推荐二", "正式", "ID912", datetime(2024, 1, 1), "", "", "业务部", "", "", "", "业务专员", "P1-3", "业务主管", "zt920"])
    sheet.append(["abc1", "非ZT员工", "试用", "ID201", datetime(2025, 5, 8), "", "", "二级", "", "", "", "助理", "P1-1", "", ""])
    return _save(workbook, path)


def _ehr_resign_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EHR离职"
    sheet.append(["parent_Name", "JobNumber", "LastWorkDate", "ApprovalStatus"])
    sheet.append(["历史员工", "zt100", datetime(2025, 5, 20), "审批中"])
    sheet.append(["推荐二", "zt912", datetime(2025, 6, 15), "审批中"])
    return _save(workbook, path)


def _special_approval_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "流程管理搜索结果导出"
    sheet.append([
        "申请单编号", "申请人工号", "申请人姓名", "申请说明", "事项分类", "文档状态",
        "招聘需求单号", "招聘周期特殊剔除时间（天）",
    ])
    sheet.append([
        "ZTJT-LC-1", "zt200", "新增员工", "申请剔除招聘周期12天",
        "招聘：海外招聘周期时长剔除", "结束", "REQ-1", 12,
    ])
    sheet.append([
        "ZTJT-LC-2", "zt201", "未结束员工", "申请剔除招聘周期20天",
        "招聘：海外招聘周期时长剔除", "审批中", "REQ-2", 20,
    ])
    return _save(workbook, path)


def _domestic_offer_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "国内offer"
    sheet.append([
        "LookupPrefix_LURecord_JobNumber", "Name", "IDNumber",
        "LookupPrefix_RecruitmentRequirement_extzhiweifenlei_109025_933701443",
        "HrDutyUserId-ExportName", "HrDutyUserId-StaffCode",
        "extxiezhuzhaopinfuzerengonghao_109025_801104894", "extxiezhuzhaopinfuzeren_109025_1163526586",
        "LookupPrefix_RecruitmentRequirement_CreateDate", "extQ20210635_109025_2041445622",
        "extQ202106334_109025_847145517-ExportName", "extQ202106334_109025_847145517-StaffCode",
        "ApprovalStatus",
    ])
    sheet.append([
        "zt200", "新增员工", "ID200", "C类", "招聘二", "zt910", "zt911", "协助二",
        datetime(2025, 4, 15), "Boss直聘", "推荐二", "zt912", "审批通过",
    ])
    return _save(workbook, path)


def _empty_source_workbook(path: Path, title: str, headers: list[str]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    return _save(workbook, path)


def test_build_auto_import_rows_inherits_history_adds_new_hires_and_applies_resignations(tmp_path):
    rows, diagnostics = build_auto_import_rows(
        target_month=202505,
        previous_workbook=_history_workbook(tmp_path / "history.xlsx"),
        ehr_roster=_ehr_workbook(tmp_path / "ehr.xlsx"),
        oehr_roster=_empty_source_workbook(tmp_path / "oehr.xlsx", "OEHR", ["姓名", "工号", "offer流程编号", "入职日期"]),
        domestic_offer=_domestic_offer_workbook(tmp_path / "domestic_offer.xlsx"),
        overseas_offer=_empty_source_workbook(tmp_path / "overseas_offer.xlsx", "海外offer", ["offer编码"]),
        ehr_resignations=_ehr_resign_workbook(tmp_path / "ehr_resign.xlsx"),
        oehr_resignations=_empty_source_workbook(tmp_path / "oehr_resign.xlsx", "OEHR离职", ["工号", "最后工作日", "审批状态"]),
        special_approvals=_special_approval_workbook(tmp_path / "special_approval.xlsx"),
    )

    by_employee = {row["工号"]: row for row in rows}
    assert set(by_employee) == {"zt100", "zt200"}
    assert by_employee["zt100"]["人员状态"] == "离职"
    assert by_employee["zt100"]["最后工作日"] == datetime(2025, 5, 20)
    assert by_employee["zt100"]["招聘负责人人员状态"] == "正式在职"
    assert by_employee["zt100"]["周期剔除天数"] == 7
    assert by_employee["zt200"]["姓名"] == "新增员工"
    assert by_employee["zt200"]["招聘渠道"] == "招聘网站"
    assert by_employee["zt200"]["ABC类别"] == "C类"
    assert by_employee["zt200"]["招聘负责人工号"] == "zt910"
    assert by_employee["zt200"]["推荐人工号"] == "zt912"
    assert by_employee["zt200"]["推荐人人员状态"] == "正式在职"
    assert by_employee["zt200"]["推荐人最后工作日"] == datetime(2025, 6, 15)
    assert by_employee["zt200"]["周期剔除天数"] == 12
    assert by_employee["zt200"]["直接上级工号"] == "zt901"
    assert by_employee["zt200"]["间接上级1工号"] == "zt920"
    assert diagnostics["counts"]["historyRows"] == 1
    assert diagnostics["counts"]["newDomesticRows"] == 1
    assert diagnostics["counts"]["cycleExclusionRows"] == 1


def test_compare_import_rows_reports_field_differences_and_missing_rows(tmp_path):
    generated = [
        {"工号": "zt100", "姓名": "历史员工", "核算月份": 202505, "人员状态": "离职"},
        {"工号": "zt200", "姓名": "新增员工", "核算月份": 202505, "人员状态": "试用"},
    ]
    target = [
        {"工号": "zt100", "姓名": "历史员工", "核算月份": 202505, "人员状态": "正式"},
        {"工号": "zt300", "姓名": "人工多出", "核算月份": 202505, "人员状态": "试用"},
    ]

    result = compare_import_rows(generated, target, headers=["核算月份", "姓名", "工号", "人员状态"])

    assert result.summary["generatedRows"] == 2
    assert result.summary["targetRows"] == 2
    assert result.summary["fieldDiffCount"] == 1
    assert result.missing_in_target[0]["工号"] == "zt200"
    assert result.missing_in_generated[0]["工号"] == "zt300"
    assert result.field_differences[0]["字段"] == "人员状态"
    assert result.field_rates["人员状态"]["matchCount"] == 0


def test_compare_import_rows_normalizes_template_display_values():
    generated = [{
        "工号": "zt100",
        "姓名": "海外员工",
        "核算月份": 202505,
        "标签分类": "海外",
        "工作地": "非中国/英国",
        "ABC类别": "C",
        "招聘负责人工号": "zt0018397",
        "推荐人姓名": "",
        "推荐人工号": "",
        "协助招聘人工号": "",
        "协助招聘人人员状态": "",
        "周期剔除天数": "",
    }]
    target = [{
        "工号": "zt100",
        "姓名": "海外员工",
        "核算月份": 202505,
        "标签分类": "海外",
        "工作地": "非中国/英国",
        "奖金地区类型": "海外发达国家",
        "ABC类别": "C1类",
        "招聘负责人工号": "海外-zt0018397",
        "推荐人姓名": "-",
        "推荐人工号": "-",
        "协助招聘人工号": "",
        "协助招聘人人员状态": "正式在职",
        "周期剔除天数": 0,
    }]

    result = compare_import_rows(
        generated,
        target,
        headers=[
            "核算月份", "姓名", "工号", "标签分类", "工作地", "奖金地区类型", "ABC类别",
            "招聘负责人工号", "推荐人姓名", "推荐人工号", "协助招聘人工号", "协助招聘人人员状态",
            "周期剔除天数",
        ],
    )

    assert result.summary["fieldDiffCount"] == 0


def test_compare_import_rows_derives_fbu_special_region_rule_from_department():
    generated = [
        {
            "工号": "zt001",
            "姓名": "德国内推员工",
            "核算月份": 202605,
            "二级组织": "FBU仓储事业部",
            "三级组织": "欧洲区",
            "四级组织": "德国区",
            "招聘渠道": "内推",
            "推荐人工号": "zt100",
            "直接上级工号": "zt900",
            "特殊地区规则": "",
        },
        {
            "工号": "zt002",
            "姓名": "捷克HRBP员工",
            "核算月份": 202605,
            "二级组织": "HRAS人力综合条线",
            "三级组织": "FBU HRBP Dept.",
            "四级组织": "捷克区HRBP部",
            "招聘渠道": "内推",
            "推荐人工号": "zt200",
            "直接上级工号": "zt901",
            "特殊地区规则": "",
        },
        {
            "工号": "zt005",
            "姓名": "德国HRBP员工",
            "核算月份": 202605,
            "二级组织": "HRAS人力综合条线",
            "三级组织": "FBU HRBP Dept.",
            "四级组织": "德国区HRBP部",
            "招聘渠道": "内推",
            "推荐人工号": "zt500",
            "直接上级工号": "zt902",
            "特殊地区规则": "",
        },
        {
            "工号": "zt003",
            "姓名": "无推荐员工",
            "核算月份": 202605,
            "二级组织": "FBU仓储事业部",
            "三级组织": "欧洲区",
            "四级组织": "德国区",
            "招聘渠道": "招聘网站",
            "推荐人工号": "",
            "特殊地区规则": "",
        },
        {
            "工号": "zt004",
            "姓名": "推荐人为上级员工",
            "核算月份": 202605,
            "二级组织": "FBU仓储事业部",
            "三级组织": "欧洲区",
            "四级组织": "德国区",
            "招聘渠道": "内推",
            "推荐人工号": "zt400",
            "直接上级工号": "zt400",
            "特殊地区规则": "",
        },
    ]
    target = [
        {**generated[0], "特殊地区规则": "FBU德国"},
        {**generated[1], "特殊地区规则": "FBU捷克"},
        {**generated[2], "特殊地区规则": "FBU德国"},
        {**generated[3], "推荐人工号": "-", "特殊地区规则": ""},
        {**generated[4], "特殊地区规则": ""},
    ]

    result = compare_import_rows(generated, target, headers=["核算月份", "姓名", "工号", "特殊地区规则"])

    assert result.summary["fieldDiffCount"] == 0
    assert result.field_rates["特殊地区规则"]["matchCount"] == 5


def test_compare_import_rows_separates_manual_blank_id_duplicate_rows():
    generated = [{
        "工号": "zt0022125",
        "姓名": "Vasile Cojocar",
        "核算月份": 202605,
        "入职日期": datetime(2026, 5, 4),
        "职位": "卡车司机",
    }]
    target = [{
        "工号": "",
        "姓名": "Vasile Cojocar",
        "核算月份": 202605,
        "入职日期": datetime(2026, 5, 4),
        "职位": "卡车司机",
    }]

    result = compare_import_rows(generated, target, headers=["核算月份", "姓名", "工号", "入职日期", "职位"])

    assert result.summary["missingInGenerated"] == 0
    assert result.summary["manualBlankIdMatchedByAlias"] == 1


def test_compare_import_rows_does_not_parse_large_identifier_numbers_as_dates():
    rows = [{"工号": "zt100", "姓名": "测试", "证件号": 440825199207081172}]

    result = compare_import_rows(rows, rows, headers=["工号", "姓名", "证件号"])

    assert result.summary["fieldDiffCount"] == 0
    assert result.field_rates["证件号"]["matchCount"] == 1


def test_write_standard_import_workbook_uses_expected_header_order(tmp_path):
    output = tmp_path / "auto.xlsx"
    write_standard_import_workbook([{"核算月份": 202505, "姓名": "测试", "工号": "zt001"}], output)

    workbook = Workbook()
    workbook = __import__("openpyxl").load_workbook(output, read_only=True)
    sheet = workbook["导入_月度数据"]
    headers = [sheet.cell(1, column).value for column in range(1, len(STANDARD_IMPORT_HEADERS) + 1)]
    assert headers == STANDARD_IMPORT_HEADERS
    assert sheet.cell(2, 1).value == 202505


def test_validation_report_lists_pending_confirmation_items(tmp_path):
    compare = CompareResult(
        summary={"missingInGenerated": 1, "missingInTarget": 1, "fieldDiffCount": 1},
        field_rates={
            "周期剔除天数": {
                "field": "周期剔除天数",
                "total": 2,
                "matchCount": 1,
                "diffCount": 1,
                "matchRate": 0.5,
            }
        },
        field_differences=[],
        missing_in_generated=[{
            "姓名": "Kejian YIN 尹克俭",
            "工号": "",
            "招聘渠道": "人员调动",
            "入职日期": datetime(2026, 5, 1),
        }],
        missing_in_target=[{
            "姓名": "何小强",
            "工号": "zt27772",
            "招聘渠道": "二次入职",
            "入职日期": datetime(2026, 5, 21),
        }],
        duplicate_keys=[],
        alias_resolved_missing=[],
        auto_enrichments=[],
    )
    output = tmp_path / "report.xlsx"

    write_validation_report(compare, {"counts": {"cycleExclusionRows": 1}, "issues": [], "changes": [], "newRows": []}, output)

    workbook = load_workbook(output, read_only=True)
    sheet = workbook["待确认事项"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("类别", "姓名", "工号", "字段", "当前结论", "建议动作")
    assert rows[1] == (
        "待接入数据源",
        "Kejian YIN 尹克俭",
        None,
        "招聘渠道",
        "人员调动/异动审批源未接入，暂不自动生成",
        "等待人员调动或异动审批导出后再验证",
    )
    assert rows[2] == (
        "待业务确认",
        "何小强",
        "zt27772",
        "招聘渠道",
        "自动源存在二次入职记录，人工模板未包含",
        "确认是否应进入人工导入模板",
    )
    assert rows[3] == (
        "待接入数据源",
        None,
        None,
        "周期剔除天数",
        "已接入本次OA特殊事项审批明细，PPT已列明节假日及海外特殊剔除条件，该字段仍有 1 条差异",
        "继续补充国内节假日历、完整OA审批明细，或复核人工模板差异",
    )
