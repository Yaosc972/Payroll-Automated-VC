import json
from pathlib import Path

from openpyxl import load_workbook


SCENARIOS_PATH = Path("docs/labor_generalization_scenarios.json")
REQUIRED_FAMILIES = {
    "standard_table",
    "cross_language",
    "tax_scope",
    "multipage_attachment",
    "degraded_image",
    "evidence_failure",
}


def test_generalization_scenarios_cover_six_families_with_twelve_unique_cases():
    from tools.labor_generalization_fixtures import load_scenarios

    scenarios = load_scenarios(SCENARIOS_PATH)

    assert len(scenarios) == 12
    assert len({row["id"] for row in scenarios}) == 12
    assert {row["family"] for row in scenarios} == REQUIRED_FAMILIES
    assert all(sum(1 for row in scenarios if row["family"] == family) == 2 for family in REQUIRED_FAMILIES)


def test_generalization_truth_is_independent_and_amounts_are_arithmetically_closed(tmp_path):
    from tools.labor_generalization_fixtures import generate_fixture_set

    summary = generate_fixture_set(SCENARIOS_PATH, tmp_path)
    truth = json.loads(Path(summary["truthPath"]).read_text(encoding="utf-8"))

    assert summary["scenarioCount"] == 12
    for case in truth["cases"]:
        employee_total = round(sum(float(row["amount"]) for row in case["employees"]), 2)
        assert employee_total == case["employeeSubtotal"]
        assert round(case["netAmount"] + case["taxAmount"] + case["feeAmount"], 2) == case["grossAmount"]
        assert case["truthSource"] == "scenario_definition"
        workbook = load_workbook(tmp_path / case["workbookFile"], read_only=True)
        headers = [cell.value for cell in next(workbook["Employee Billing"].iter_rows(min_row=1, max_row=1))]
        assert case["amountColumn"] in headers
        assert case["amountColumn"] == ("Amount (Net)" if case["amountScope"] == "net" else "Amount (Gross)")
        assert case["warehouseId"]
        assert "Physical Warehouse" in headers


def test_generated_fixtures_contain_no_known_real_material_tokens(tmp_path):
    from tools.labor_generalization_fixtures import generate_fixture_set

    summary = generate_fixture_set(SCENARIOS_PATH, tmp_path)
    forbidden = {
        "onesource",
        "prompt",
        "armz",
        "us elogistics",
        "alvarez",
        "benavides",
        "gulf coast bank",
    }
    text = "\n".join(
        path.read_bytes().decode("latin-1", errors="ignore")
        for path in tmp_path.rglob("*")
        if path.is_file()
    ).lower()

    assert forbidden.isdisjoint(token for token in forbidden if token in text)
    assert len(summary["files"]) >= 24


def test_fixture_generation_is_deterministic_and_includes_required_variants(tmp_path):
    from tools.labor_generalization_fixtures import generate_fixture_set

    first = generate_fixture_set(SCENARIOS_PATH, tmp_path / "first")
    second = generate_fixture_set(SCENARIOS_PATH, tmp_path / "second")

    assert first["truthSha256"] == second["truthSha256"]
    truth = json.loads(Path(first["truthPath"]).read_text(encoding="utf-8"))
    assert {case["language"] for case in truth["cases"]} >= {"en", "es", "de"}
    assert any(case["taxAmount"] > 0 for case in truth["cases"])
    assert any(len(case["pages"]) > 1 for case in truth["cases"])
    assert {case["recognizable"] for case in truth["cases"]} == {True, False}
