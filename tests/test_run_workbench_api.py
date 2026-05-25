from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from bonus_platform.app import app
from bonus_platform.engine.workbook_io import PENDING_CONFIRMATION_HEADERS


def _monthly_workbook_bytes(employee_no: str = "zt-run-001") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入_月度数据"
    sheet.append(["核算月份", "工号", "姓名", "工作地", "标签分类", "职级", "ABC类别", "招聘渠道"])
    sheet.append([202510, employee_no, "批次测试", "中国大陆", "国内", "P1-3", "C类", "招聘网站"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _confirmation_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "待确认_发放判断"
    sheet.append(PENDING_CONFIRMATION_HEADERS)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _offline_review_workbook_bytes() -> bytes:
    workbook = Workbook()
    recruitment = workbook.active
    recruitment.title = "附件二-招聘奖金汇总表"
    recruitment.append(["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    recruitment.append(["zt-offline", "线下招聘", "招聘负责人", "CNY", 202510, 10, 0, 0, 0, 10])
    referral = workbook.create_sheet("附件三-内推奖金汇总")
    referral.append(["推荐人工号", "推荐人姓名", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    referral.append(["zt-referrer", "线下内推", "CNY", 202510, 20, 0, 0, 0, 20])
    detail = workbook.create_sheet("附件一-招聘奖金核算明细")
    detail.append(["工号", "姓名", "核算月份", "招聘人入职1月奖金", "内推入职1月奖金"])
    detail.append(["zt-run-001", "批次测试", 202510, 10, 20])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_run_calculation_creates_queryable_batch_with_rule_info():
    client = TestClient(app)

    response = client.post(
        "/api/runs/calculate",
        files={"file": ("monthly.xlsx", _monthly_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["id"]
    assert data["month"] == 202510
    assert data["status"] in {"已初算", "待确认"}
    assert data["downloadUrl"].endswith(".xlsx")
    assert data["files"]["initialResult"]["downloadUrl"] == data["downloadUrl"]
    assert data["ruleInfo"]["workbook"].endswith("招聘奖金核算_规则库.xlsx")

    detail_response = client.get(f"/api/runs/{data['id']}")
    assert detail_response.status_code == 200
    assert detail_response.json()["id"] == data["id"]

    list_response = client.get("/api/runs")
    assert list_response.status_code == 200
    assert any(run["id"] == data["id"] for run in list_response.json()["runs"])


def test_same_month_run_calculations_do_not_overwrite_each_other():
    client = TestClient(app)

    first = client.post(
        "/api/runs/calculate",
        files={"file": ("monthly.xlsx", _monthly_workbook_bytes("zt-run-a"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()
    second = client.post(
        "/api/runs/calculate",
        files={"file": ("monthly.xlsx", _monthly_workbook_bytes("zt-run-b"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()

    assert first["id"] != second["id"]
    assert Path(first["files"]["initialResult"]["path"]).exists()
    assert Path(second["files"]["initialResult"]["path"]).exists()


def test_run_finalize_uses_saved_initial_result_and_updates_batch():
    client = TestClient(app)
    run = client.post(
        "/api/runs/calculate",
        files={"file": ("monthly.xlsx", _monthly_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()

    response = client.post(
        f"/api/runs/{run['id']}/finalize",
        files={"confirmation_file": ("confirmation.xlsx", _confirmation_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "已最终确认"
    assert data["finalDownloadUrl"].endswith(".xlsx")
    assert Path(data["files"]["finalResult"]["path"]).exists()


def test_run_compare_generates_difference_report_for_batch():
    client = TestClient(app)
    run = client.post(
        "/api/runs/calculate",
        files={"file": ("monthly.xlsx", _monthly_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    ).json()

    response = client.post(
        f"/api/runs/{run['id']}/compare",
        files={"offline_file": ("offline.xlsx", _offline_review_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "已生成差异报告"
    assert data["diffDownloadUrl"].endswith(".xlsx")
    assert data["diffMetrics"]["recruitmentSummaryDiffCount"] >= 1
    workbook = load_workbook(Path(data["files"]["diffReport"]["path"]), read_only=True)
    assert workbook.sheetnames == ["摘要", "招聘汇总差异", "内推汇总差异", "招聘明细差异", "内推明细差异"]
