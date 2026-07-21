from datetime import date

import openpyxl
from openpyxl import Workbook

from bonus_platform.engine.fbu_performance.engines.base import CalculationSegment, EmployeeData
from bonus_platform.engine.fbu_performance.engines.bonus import BonusCalculator
from bonus_platform.engine.fbu_performance.engines.coefficient import CoefficientCalculator
from bonus_platform.engine.fbu_performance.engines.salary import SalaryProcessor
from bonus_platform.engine.fbu_performance.parser import FBUPerformanceParser
from bonus_platform.engine.fbu_performance.runs import FBURosterStore, FBURunManager, build_final_result_rows


def test_warehouse_coefficient_boundaries():
    assert CoefficientCalculator.calc_warehouse_coefficient(None) == 0
    assert CoefficientCalculator.calc_warehouse_coefficient(60) == 0
    assert CoefficientCalculator.calc_warehouse_coefficient(76) == 0.8
    assert CoefficientCalculator.calc_warehouse_coefficient(95) == 1
    assert CoefficientCalculator.calc_warehouse_coefficient(110) == 1.3
    assert CoefficientCalculator.calc_warehouse_coefficient(108.75) == 1.28
    assert CoefficientCalculator.calc_warehouse_coefficient(125) == 1.6
    assert CoefficientCalculator.calc_warehouse_coefficient(130) == 1.6


def test_functional_coefficient_normalizes_level_text():
    assert CoefficientCalculator.calc_functional_coefficient(" 符合预期+ ") == 1.2
    assert CoefficientCalculator.calc_functional_coefficient("未知等级") == 0


def test_bonus_formula_uses_performance_base_ratio_and_calculated_coefficient():
    emp = EmployeeData(
        employee_id="E001",
        name="Ana",
        hourly_rate=20,
        performance_ratio=0.1,
        performance_score=110,
        base_hours=160,
        ot15_hours=10,
        ot20_hours=5,
        sick_hours=8,
        annual_hours=8,
        holiday_hours=4,
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_base == 20 * (160 + 10 * 1.5 + 5 * 2 + 8 + 8 + 4)
    assert emp.performance_coefficient == 1.3
    assert emp.performance_bonus == emp.performance_base * 0.1 * 1.3


def test_bonus_formula_includes_sick_settlement_pay():
    emp = EmployeeData(
        employee_id="E001",
        name="Ana",
        hourly_rate=20,
        performance_ratio=0.1,
        performance_score=95,
        base_hours=160,
        sick_settlement_hours=6,
    )

    BonusCalculator.calculate(emp)

    assert emp.sick_settlement_pay == 120
    assert emp.performance_base == 3320
    assert emp.performance_bonus == 332


def test_96_hour_rule_calculates_base_from_hours_without_ot_premium():
    emp = EmployeeData(
        employee_id="zt12988",
        name="陈海冰",
        hourly_rate=20,
        performance_ratio=0.1,
        performance_score=95,
        work_hour_rule="96工时制",
        work_hour_rule_cap=96,
        base_hours=90,
        ot15_hours=10,
        annual_hours=8,
        holiday_hours=4,
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_base == 2000
    assert emp.performance_bonus == 200
    assert emp.calculation_path == "96工时制自动基数路径"


def test_96_hour_rule_uses_salary_file_rate_instead_of_night_premium_rate():
    emp = EmployeeData(
        employee_id="zt12988",
        name="陈海冰",
        hourly_rate=21,
        is_night_shift=True,
        performance_ratio=0.1,
        performance_score=95,
        work_hour_rule="96工时制",
        work_hour_rule_special_total_hours=100,
    )

    BonusCalculator.calculate(emp)

    assert emp.work_hour_rule_rounded_hourly_rate == 20
    assert emp.performance_base == 2000
    assert emp.performance_bonus == 200


def test_96_hour_rule_counts_zhao_wanyan_holiday_inside_cap():
    emp = EmployeeData(
        employee_id="zt12979",
        name="赵婉妍",
        hourly_rate=20,
        performance_ratio=0.1,
        performance_score=95,
        work_hour_rule="96工时制",
        work_hour_rule_cap=96,
        base_hours=80,
        ot15_hours=10,
        holiday_hours=16,
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_base == 1920
    assert emp.performance_bonus == 192


def test_96_hour_rule_uses_period_caps_when_period_rows_are_available():
    emp = EmployeeData(
        employee_id="zt12988",
        name="陈海冰",
        hourly_rate=20,
        performance_ratio=0.1,
        performance_score=95,
        work_hour_rule="96工时制",
        work_hour_rule_cap=96,
        work_hour_rule_periods=[
            {
                "period": "2026-04 / 4.12-4.25",
                "cap_hours": 96,
                "base_hours": 90,
                "ot15_hours": 10,
                "annual_hours": 8,
                "holiday_hours": 4,
            },
            {
                "period": "2026-04 / 4.26-4.30",
                "cap_hours": 40,
                "base_hours": 35,
                "annual_hours": 8,
                "holiday_hours": 2,
            },
        ],
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_base == 2840
    assert emp.performance_bonus == 284
    assert emp.work_hour_rule_periods[0]["included_hours"] == 100
    assert emp.work_hour_rule_periods[1]["included_hours"] == 42


def test_96_hour_rule_prefers_special_summary_hours_with_rounded_hourly_rate():
    emp = EmployeeData(
        employee_id="zt12988",
        name="陈海冰",
        hourly_rate=28.8481570512821,
        performance_ratio=0.4285714286,
        performance_score=101.58,
        work_hour_rule="96工时制",
        work_hour_rule_special_total_hours=201.19,
    )

    BonusCalculator.calculate(emp)

    assert emp.base_salary == 5804.3315
    assert emp.performance_base == 5804.3315
    assert emp.work_hour_rule_rounded_hourly_rate == 28.85


def test_uploaded_coefficient_is_preserved_for_audit_but_formula_is_authoritative():
    emp = EmployeeData(
        employee_id="E002",
        name="Ben",
        hourly_rate=10,
        performance_ratio=0.2,
        performance_score=95,
        uploaded_coefficient=1.6,
        base_hours=100,
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_coefficient == 1.0
    assert emp.uploaded_coefficient == 1.6
    assert any("上传绩效系数与系统计算系数不一致" in msg for msg in emp.exceptions)


def test_uploaded_coefficient_is_used_when_no_performance_result_exists():
    emp = EmployeeData(
        employee_id="E003",
        name="Cara",
        hourly_rate=20,
        performance_ratio=0.1,
        uploaded_coefficient=1.0,
        base_hours=100,
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_coefficient == 1.0
    assert emp.performance_bonus == 200


def test_salary_processor_accepts_percent_strings():
    rows = [
        ["Ana", "E001", None, None, None, None, None, None, None, "10%", None, "20.5"],
        ["Ben", "E002", None, None, None, None, None, None, None, 15, None, 18],
        ["Cara", "E003", None, None, None, None, None, None, None, None, None, 0],
    ]

    data = SalaryProcessor().load(rows)

    assert data["E001"]["hourly_rate"] == 20.5
    assert data["E001"]["ratio"] == 0.1
    assert data["E002"]["hourly_rate"] == 18.0
    assert data["E002"]["ratio"] == 0.15
    assert data["E003"]["hourly_rate"] == 0.0
    assert data["E003"]["ratio"] == 0.0


def test_district_manager_uses_fixed_base_and_uploaded_coefficient():
    emp = EmployeeData(
        employee_id="zt15638",
        name="万其鑫",
        hourly_rate=40.384615,
        performance_ratio=0,
        performance_score=112.72,
        performance_level="超出预期",
        uploaded_coefficient=1.35,
        job_type="district_manager",
        fixed_performance_base=3000,
        base_hours=171.5,
        sick_hours=4.5,
    )

    BonusCalculator.calculate(emp)

    assert emp.performance_base == 3000
    assert emp.performance_coefficient == 1.35
    assert round(emp.performance_bonus, 2) == 4050


def test_save_results_marks_district_manager_fixed_base_path(tmp_path):
    manager = FBURunManager(str(tmp_path))
    run = manager.create_run(calc_month="2026-04")
    emp = EmployeeData(
        employee_id="zt15638",
        name="万其鑫",
        hourly_rate=40.384615,
        performance_ratio=0,
        uploaded_coefficient=1.35,
        job_type="district_manager",
        fixed_performance_base=3000,
        base_hours=171.5,
        sick_hours=4.5,
    )

    BonusCalculator.calculate(emp)
    manager.save_results(run.run_id, [emp])

    saved = manager.get_run(run.run_id).results[0]
    assert saved["fixed_performance_base"] == 3000
    assert saved["calculation_path"] == "区长固定基数路径"
    base_detail = saved["base_calculation_details"][0]
    assert base_detail["path"] == "区长固定基数路径"
    assert base_detail["components"] == [{
        "label": "区长固定绩效基数",
        "amount": 3000.0,
        "multiplier": 1.0,
    }]


def test_save_results_preserves_standard_base_calculation_components(tmp_path):
    manager = FBURunManager(str(tmp_path))
    run = manager.create_run(calc_month="2026-04")
    emp = EmployeeData(
        employee_id="zt001",
        name="员工甲",
        hourly_rate=20,
        performance_ratio=0.05,
        performance_score=95,
        base_hours=80,
        ot15_hours=2,
        sick_hours=1,
        annual_hours=8,
        holiday_hours=4,
    )

    BonusCalculator.calculate(emp)
    manager.save_results(run.run_id, [emp])

    detail = manager.get_run(run.run_id).results[0]["base_calculation_details"][0]
    components = {component["label"]: component for component in detail["components"]}
    assert detail["path"] == "标准绩效基数路径"
    assert detail["performance_base"] == 1920.0
    assert components["基础工时"]["hours"] == 80.0
    assert components["基础工时"]["hourly_rate"] == 20.0
    assert components["OT 1.5"]["multiplier"] == 1.5
    assert components["OT 1.5"]["amount"] == 60.0
    assert components["年假"]["amount"] == 160.0
    assert components["节日补贴"]["amount"] == 80.0


def test_save_results_preserves_96_hour_base_formula(tmp_path):
    manager = FBURunManager(str(tmp_path))
    run = manager.create_run(calc_month="2026-04")
    emp = EmployeeData(
        employee_id="zt12988",
        name="员工乙",
        hourly_rate=29.8481570512821,
        performance_ratio=0.1,
        performance_score=95,
        work_hour_rule="96工时制",
        work_hour_rule_special_total_hours=201.19,
        is_night_shift=True,
    )

    BonusCalculator.calculate(emp)
    manager.save_results(run.run_id, [emp])

    detail = manager.get_run(run.run_id).results[0]["base_calculation_details"][0]
    component = detail["components"][0]
    assert detail["path"] == "96工时制自动基数路径"
    assert detail["display_label"] == "夜班"
    assert component["label"] == "96工时制计入工时"
    assert component["hours"] == 201.19
    assert component["hourly_rate"] == 28.85
    assert component["amount"] == 5804.33


def test_save_results_preserves_adjustment_split_base_segments(tmp_path):
    manager = FBURunManager(str(tmp_path))
    run = manager.create_run(calc_month="2026-04")
    emp = EmployeeData(
        employee_id="zt0020155",
        name="员工丙",
        performance_ratio=0.05,
        performance_score=95,
        calculation_segments=[
            CalculationSegment("4月1日-25日", "调薪前", 3000, 0, 1),
            CalculationSegment("4月26日-30日", "调薪后", 700, 0.05, 1),
        ],
    )

    BonusCalculator.calculate(emp)
    manager.save_results(run.run_id, [emp])

    detail = manager.get_run(run.run_id).results[0]["base_calculation_details"][0]
    assert detail["path"] == "调薪/转正拆分路径"
    assert detail["performance_base"] == 700.0
    assert [component["label"] for component in detail["components"]] == ["调薪前", "调薪后"]
    assert [component["amount"] for component in detail["components"]] == [3000.0, 700.0]


def test_final_rows_reconstruct_base_formula_for_historical_runs():
    rows = [{
        "employee_id": "zt001",
        "source_employee_id": "zt001",
        "name": "历史员工",
        "calculation_path": "标准绩效基数路径",
        "hourly_rate": 20,
        "base_hours": 80,
        "ot15_hours": 2,
        "ot20_hours": 0,
        "sick_hours": 1,
        "sick_settlement_hours": 0,
        "annual_hours": 8,
        "holiday_hours": 4,
        "performance_base": 1920,
        "performance_bonus": 115.2,
    }]

    final_row = build_final_result_rows(rows)[0]

    detail = final_row["base_calculation_details"][0]
    assert detail["path"] == "标准绩效基数路径"
    assert [component["label"] for component in detail["components"]] == [
        "基础工时",
        "OT 1.5",
        "病假",
        "年假",
        "节日补贴",
    ]
    assert detail["note"] == "根据历史批次保存的工时和时薪还原标准基数计算。"


def test_save_results_total_bonus_matches_rounded_employee_rows(tmp_path):
    manager = FBURunManager(str(tmp_path))
    run = manager.create_run(calc_month="2026-04")
    employees = [
        EmployeeData(employee_id="zt001", name="员工甲", performance_bonus=1.004),
        EmployeeData(employee_id="zt002", name="员工乙", performance_bonus=1.004),
    ]

    manager.save_results(run.run_id, employees)

    saved = manager.get_run(run.run_id)
    assert saved.total_bonus == 2.00


def test_save_results_total_bonus_rounds_after_source_employee_rollup(tmp_path):
    manager = FBURunManager(str(tmp_path))
    run = manager.create_run(calc_month="2026-04")
    employees = [
        EmployeeData(employee_id="zt001-1", source_employee_id="zt001", name="员工甲", performance_bonus=1.004),
        EmployeeData(employee_id="zt001", source_employee_id="zt001", name="员工甲", performance_bonus=1.004),
    ]

    manager.save_results(run.run_id, employees)

    saved = manager.get_run(run.run_id)
    assert saved.total_bonus == 2.01


def test_adjustment_split_preview_reads_zhang_haibing_style_segments(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "调薪拆分"
    sheet.append(["header"] * 32)
    for period, amount, reason in [
        ("4.1-4.11", 1404.9, "调薪前"),
        ("4.12-4.25", 1667.88, "调薪前"),
        ("4.26-4.30", 732.42, "调薪后"),
    ]:
        row = [""] * 32
        row[3] = "zt0021990"
        row[4] = "张海冰"
        row[9] = period
        row[28] = amount
        row[31] = reason
        sheet.append(row)
    path = tmp_path / "adjustments.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_adjustments_preview(str(path))

    assert preview["summary"]["total_employees"] == 1
    assert preview["summary"]["total_segments"] == 3
    assert preview["summary"]["active_performance_base"] == 732.42
    employee = preview["employees"][0]
    assert employee["employee_id"] == "zt0021990"
    assert employee["name"] == "张海冰"
    assert employee["segments"][2] == {
        "period": "4.26-4.30",
        "reason": "调薪后",
        "performance_base": 732.42,
    }


def test_adjustment_split_preview_accepts_platform_template_headers(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "调薪拆分"
    sheet.append(["填报说明："])
    sheet.append(["以下为脱敏示例，不参与导入"])
    sheet.append(["zt0000001", "花名一", "4.1-4.15", 1200, "调薪前", "示例"])
    sheet.append([])
    sheet.append([])
    sheet.append(["工号", "姓名", "分段期间", "分段绩效基数", "核算标识", "备注"])
    sheet.append(["zt0021990", "张海冰", "4.26-4.30", 732.42, "调薪后", "转正"])
    path = tmp_path / "adjustments_template.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_adjustments_preview(str(path))

    assert preview["summary"]["total_employees"] == 1
    assert preview["summary"]["active_performance_base"] == 732.42
    assert preview["employees"][0]["segments"] == [
        {"period": "4.26-4.30", "reason": "调薪后", "performance_base": 732.42}
    ]


def test_adjustment_preview_accepts_oehr_adjustment_export_as_events(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1779431330576调薪管理导出"
    sheet.append([
        "审批主题",
        "姓名",
        "工号",
        "审批状态",
        "二级部门",
        "三级部门",
        "四级部门",
        "五级部门",
        "六级部门",
        "七级部门",
        "八级部门",
        "调薪类型",
        "职级",
        "调薪原因",
        "调薪生效日期",
        "调薪后薪酬制度",
        "调薪后成本归属",
        "调薪后币种",
        "基本工资标准",
        "绩效奖金计算方式",
        "月度绩效奖金基数",
        "月度绩效奖金比例(%)",
        "是否考勤豁免人员",
        "时薪标准",
        "年终奖基数",
        "度假金",
        "圣诞金",
        "401k（仅LBU）",
        "月交通补贴（非FBU）",
        "目标年收",
        "月薪标准",
        "月电话补贴（非FBU）",
        "其他补贴（非FBU）",
        "其他补贴说明（非FBU）",
        "发薪周期（仅LBU）",
        "月标准工时（仅FBU）",
        "备注",
    ])
    sheet.append([
        "张海冰的调薪申请",
        "张海冰",
        "zt0021990",
        "已完成",
        "FBU仓储事业部",
        "美洲区",
        "新泽西区",
        "新泽西21号仓（SN）",
        "理货组",
        "",
        "",
        "非窗口期调薪",
        "",
        "转正调薪",
        "2026/04/26",
        "时薪制",
        "理货组",
        "美元(USD)",
        18,
        "固定比例核算",
        0,
        "5",
        "否",
        18,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        0,
        "",
        0,
        0,
        "P1-2转正，增加绩效占比",
    ])
    path = tmp_path / "oehr_adjustments.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_adjustments_preview(str(path))

    assert preview["employees"] == []
    assert preview["summary"]["total_employees"] == 0
    assert preview["summary"]["total_segments"] == 0
    assert preview["summary"]["active_performance_base"] == 0
    assert preview["summary"]["total_events"] == 1
    assert preview["summary"]["auto_split_ready"] == 1
    assert preview["summary"]["manual_split_required"] == 0
    assert preview["events"][0]["employee_id"] == "zt0021990"
    assert preview["events"][0]["effective_date"] == "2026-04-26"
    assert "自动拆分" in preview["warnings"][0]


def test_oehr_adjustment_event_auto_calculates_post_effective_base():
    parser = FBUPerformanceParser()
    adjustment_preview = {
        "employees": [],
        "events": [
            {
                "employee_id": "zt0021990",
                "name": "张海冰",
                "approval_status": "已完成",
                "adjustment_type": "非窗口期调薪",
                "adjustment_reason": "转正调薪",
                "effective_date": "2026-04-26",
                "hourly_rate": 18,
                "performance_ratio": 0.05,
            }
        ],
    }

    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt0021990",
                "name": "张海冰",
                "department": "新泽西21号仓（SN）",
                "area": "新泽西区",
                "personnel_status": "正式",
                "position": "Tallyman 理货员",
                "job_type": "warehouse",
                "has_night_shift": False,
                "day_shift": {"计薪出勤": 183.95, "OT1.5": 18.3, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "attendance_daily_rows": [
                    {"date": "2026-04-01", "shift_type": "白班", "base_hours": 63.95, "ot15_hours": 9.4},
                    {"date": "2026-04-12", "shift_type": "白班", "base_hours": 80, "ot15_hours": 8.44},
                    {"date": "2026-04-26", "shift_type": "白班", "base_hours": 40, "ot15_hours": 0.46},
                ],
            }
        ],
        salary_data=[
            {
                "employee_id": "zt0021990",
                "hourly_rate": 18,
                "ratio": 0,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[],
        adjustment_data=adjustment_preview,
        calc_month="2026-04",
    )

    emp = engine.get_employee("zt0021990")
    assert round(emp.performance_base, 2) == 732.42
    assert emp.performance_ratio == 0.05
    assert emp.performance_coefficient == 1.0
    assert round(emp.performance_bonus, 2) == 36.62
    assert [segment.reason for segment in emp.calculation_segments] == ["调薪前", "调薪后"]
    assert [round(segment.performance_base, 2) for segment in emp.calculation_segments] == [3072.78, 732.42]


def test_oehr_adjustment_event_on_first_day_applies_ratio_without_split():
    parser = FBUPerformanceParser()
    adjustment_preview = {
        "employees": [],
        "events": [
            {
                "employee_id": "zt0021862",
                "name": "孙嘉鸣",
                "approval_status": "已完成",
                "adjustment_type": "非窗口期调薪",
                "adjustment_reason": "转正调薪",
                "effective_date": "2026-04-01",
                "hourly_rate": 18,
                "performance_ratio": 0.05,
            }
        ],
    }

    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt0021862",
                "name": "孙嘉鸣",
                "department": "新泽西21号仓（SN）",
                "area": "新泽西区",
                "personnel_status": "正式",
                "position": "Sorting Clerk 分播员",
                "job_type": "warehouse",
                "has_night_shift": False,
                "day_shift": {"计薪出勤": 100, "OT1.5": 10, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "attendance_daily_rows": [
                    {"date": "2026-04-01", "shift_type": "白班", "base_hours": 100, "ot15_hours": 10},
                ],
            }
        ],
        salary_data=[
            {
                "employee_id": "zt0021862",
                "hourly_rate": 18,
                "ratio": 0,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[
            {
                "employee_id": "zt0021862",
                "score": 88.7,
                "level": "符合预期-",
                "coefficient": 0.93,
            }
        ],
        adjustment_data=adjustment_preview,
        calc_month="2026-04",
    )

    emp = engine.get_employee("zt0021862")
    assert emp.calculation_segments == []
    assert emp.performance_ratio == 0.05
    assert emp.performance_coefficient == 0.93
    assert round(emp.performance_base, 2) == 2070.0
    assert round(emp.performance_bonus, 2) == 96.26


def test_adjustment_split_uses_post_adjustment_base_ratio_and_fixed_coefficient():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt0021990",
                "name": "张海冰",
                "department": "新泽西21号仓（SN）",
                "area": "新泽西区",
                "job_type": "warehouse",
                "has_night_shift": False,
                "day_shift": {"计薪出勤": 183.95, "OT1.5": 18.3, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            }
        ],
        salary_data=[
            {
                "employee_id": "zt0021990",
                "hourly_rate": 18,
                "ratio": 0.05,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[],
        adjustment_data=[
            {
                "employee_id": "zt0021990",
                "segments": [
                    {"period": "4.1-4.11", "reason": "调薪前", "performance_base": 1404.9},
                    {"period": "4.12-4.25", "reason": "调薪前", "performance_base": 1667.88},
                    {"period": "4.26-4.30", "reason": "调薪后", "performance_base": 732.42},
                ],
            }
        ],
    )

    emp = engine.get_employee("zt0021990")

    assert round(emp.performance_base, 2) == 732.42
    assert emp.performance_ratio == 0.05
    assert emp.performance_coefficient == 1.0
    assert round(emp.performance_bonus, 2) == 36.62
    assert [round(segment.performance_bonus, 2) for segment in emp.calculation_segments] == [0, 0, 36.62]


def test_adjustment_split_keeps_supplemented_score_level_while_using_fixed_coefficient():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt0021990",
                "name": "张海冰",
                "department": "新泽西21号仓（SN）",
                "area": "新泽西区",
                "personnel_status": "正式",
                "confirmation_date": "2026-04-26",
                "position": "Tallyman 理货员",
                "job_type": "warehouse",
                "has_night_shift": False,
                "day_shift": {"计薪出勤": 183.95, "OT1.5": 18.3, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            }
        ],
        salary_data=[
            {
                "employee_id": "zt0021990",
                "hourly_rate": 18,
                "ratio": 0.05,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[
            {
                "employee_id": "zt0021990",
                "score": 95,
                "level": "符合预期+",
                "coefficient": 1,
            }
        ],
        adjustment_data=[
            {
                "employee_id": "zt0021990",
                "segments": [
                    {"period": "4.1-4.11", "reason": "调薪前", "performance_base": 1404.9},
                    {"period": "4.12-4.25", "reason": "调薪前", "performance_base": 1667.88},
                    {"period": "4.26-4.30", "reason": "调薪后", "performance_base": 732.42},
                ],
            }
        ],
        calc_month="2026-04",
    )

    emp = engine.get_employee("zt0021990")

    assert emp.performance_score == 95
    assert emp.performance_level == "符合预期+"
    assert "15号后转正" in emp.coefficient_override_reason
    assert emp.performance_coefficient == 1.0
    assert round(emp.performance_bonus, 2) == 36.62


def test_salary_preview_reports_total_valid_and_zero_hourly_counts(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["姓名", "工号", "", "", "", "", "", "", "", "月度绩效奖金比例(%)", "", "时薪标准"])
    sheet.append(["Ana", "E001", "", "", "", "", "", "", "", "10%", "", 20])
    sheet.append(["Ben", "E002", "", "", "", "", "", "", "", "", "", 0])
    sheet.append(["Cara", "E003", "", "", "", "", "", "", "", 20, "", 30])
    path = tmp_path / "salary.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_salary_preview(str(path))

    assert preview["summary"]["total_employees"] == 3
    assert preview["summary"]["valid_hourly_count"] == 2
    assert preview["summary"]["zero_hourly_count"] == 1
    assert preview["summary"]["avg_hourly_rate"] == 25


def test_salary_preview_uses_headers_when_new_columns_shift_hourly_rate(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "姓名",
        "工号",
        "人员状态",
        "划分区域",
        "成本归属",
        "薪酬制度",
        "基本工资标准",
        "绩效奖金计算方式",
        "月度绩效奖金基数",
        "月度绩效奖金比例(%)",
        "叉车补贴",
        "是否考勤豁免人员",
        "时薪标准",
        "年终奖基数",
        "度假金",
        "圣诞金",
        "401k（仅LBU）",
        "月交通补贴（非FBU）",
        "目标年收",
        "月薪标准",
        "月电话补贴（非FBU）",
        "其他补贴（非FBU）",
        "其他补贴说明（非FBU）",
        "发薪周期（仅LBU）",
        "月标准工时（仅FBU）",
        "备注",
        "币种",
        "发薪机构",
        "二级部门",
        "三级部门",
        "四级部门",
        "五级部门",
        "六级部门",
        "七级部门",
        "八级部门",
        "岗位",
    ])
    sheet.append([
        "Ana",
        "E001",
        "正式",
        "新泽西区",
        "入库组",
        "时薪制",
        23,
        "固定比例核算",
        "",
        0.13,
        "",
        "是",
        23,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "美元(USD)",
        "US ELOGISTICS SERVICE CORP",
        "FBU仓储事业部",
        "美洲区",
        "新泽西区",
        "新泽西6号仓",
        "入库组",
        "",
        "",
        "仓库操作员",
    ])
    path = tmp_path / "salary-new-columns.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_salary_preview(str(path))

    employee = preview["employees"][0]
    assert employee["employee_id"] == "E001"
    assert employee["hourly_rate"] == 23
    assert employee["ratio"] == 0.13
    assert employee["department"] == "FBU仓储事业部-美洲区-新泽西区-新泽西6号仓-入库组"
    assert employee["position"] == "仓库操作员"


def test_attendance_preview_uses_headers_when_new_columns_shift_paid_hours(tmp_path):
    from openpyxl import Workbook

    headers = [""] * 74
    for index, header in {
        0: "考勤日期",
        1: "姓名",
        2: "工号",
        15: "班次名称",
        25: "工作时长",
        29: "节假日时长",
        30: "年假时长",
        31: "病假时长",
        53: "OT1.5",
        56: "应出勤时长",
        59: "出勤时长",
        60: "计薪出勤时长",
        61: "缺勤时长",
        64: "计薪加班时长",
        65: "病假余额结算（离职结算）",
    }.items():
        headers[index] = header

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(headers)
    row = [""] * 74
    row[0] = "2026/05/01"
    row[1] = "Ana"
    row[2] = "E001"
    row[15] = "新泽西-9:15"
    row[25] = 8.5
    row[29] = 1
    row[30] = 2
    row[31] = 3
    row[53] = 4
    row[56] = 8
    row[59] = 8
    row[60] = 8.25
    row[61] = 0
    row[65] = 5
    sheet.append(row)
    path = tmp_path / "attendance-new-columns.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_attendance_preview(str(path), 5)

    employee = preview["employees"][0]
    assert employee["employee_id"] == "E001"
    assert employee["total_base_hours"] == 8.25
    assert employee["total_ot15"] == 4
    assert employee["sick_hours"] == 3
    assert employee["sick_settlement_hours"] == 5
    assert employee["annual_hours"] == 2
    assert employee["holiday_hours"] == 1
    assert preview["summary"]["total_base_hours"] == 8.25


def test_attendance_preview_infers_holiday_pay_hours_from_holiday_shift_name(tmp_path):
    from openpyxl import Workbook

    headers = [""] * 74
    for index, header in {
        0: "考勤日期",
        1: "姓名",
        2: "工号",
        15: "班次名称",
        29: "节假日时长",
        51: "标准工作时间",
        60: "计薪出勤时长",
    }.items():
        headers[index] = header

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(headers)

    holiday_row = [""] * 74
    holiday_row[0] = "2026/05/29"
    holiday_row[1] = "蒋庆莲"
    holiday_row[2] = "zt0021600"
    holiday_row[15] = "美东假期"
    holiday_row[29] = 0
    holiday_row[51] = 40
    holiday_row[60] = 0
    sheet.append(holiday_row)

    regular_off_row = [""] * 74
    regular_off_row[0] = "2026/05/30"
    regular_off_row[1] = "蒋庆莲"
    regular_off_row[2] = "zt0021600"
    regular_off_row[15] = "公休日-OFFICE"
    regular_off_row[51] = 40
    sheet.append(regular_off_row)

    path = tmp_path / "attendance-holiday.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_attendance_preview(str(path), target_month=5)

    employee = preview["employees"][0]
    assert employee["holiday_hours"] == 8
    daily_rows = employee["attendance_daily_rows"]
    assert daily_rows[0]["holiday_hours"] == 8
    assert daily_rows[1]["holiday_hours"] == 0


def test_attendance_preview_infers_night_shift_from_shift_name_when_start_column_missing(tmp_path):
    from openpyxl import Workbook

    headers = [""] * 74
    for index, header in {
        0: "考勤日期",
        1: "姓名",
        2: "工号",
        15: "班次名称",
        53: "OT1.5",
        60: "计薪出勤时长",
    }.items():
        headers[index] = header

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(headers)
    row = [""] * 74
    row[0] = "2026/05/12"
    row[1] = "郑增良"
    row[2] = "zt18731"
    row[15] = "新泽西-16:00"
    row[53] = 6.17
    row[60] = 40
    sheet.append(row)
    path = tmp_path / "attendance-night-name.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_attendance_preview(str(path), target_month=5)

    employee = preview["employees"][0]
    assert employee["employee_id"] == "zt18731"
    assert employee["shift_type"] == "夜班"
    assert employee["has_night_shift"] is True
    assert employee["night_shift"]["计薪出勤"] == 40


def test_night_shift_time_range_wraps_across_midnight():
    from bonus_platform.engine.fbu_performance.engines.attendance import AttendanceProcessor

    assert AttendanceProcessor.is_night_shift(None, "新泽西-14:00") is True
    assert AttendanceProcessor.is_night_shift(None, "新泽西-00:15am") is True
    assert AttendanceProcessor.is_night_shift(None, "新泽西-04:59") is True
    assert AttendanceProcessor.is_night_shift(None, "新泽西-05:00") is False
    assert AttendanceProcessor.is_night_shift(None, "新泽西-09:15") is False


def test_build_employees_splits_night_shift_and_flags_missing_matches():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 8, "OT1.5": 2, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": True,
        },
        "E002": {
            "白班": {"计薪出勤": 100, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 95, "level": None, "coefficient": None}}

    employees = parser.build_employees(attendance_data, salary_data, performance_data)

    by_id = {emp.employee_id: emp for emp in employees}
    assert by_id["E001-1"].source_employee_id == "E001"
    assert by_id["E001-1"].hourly_rate == 20
    assert by_id["E001"].source_employee_id == "E001"
    assert by_id["E001"].hourly_rate == 21
    assert "未匹配薪资档案" in by_id["E002"].exceptions
    assert "未匹配绩效报表" in by_id["E002"].exceptions


def test_build_employees_keeps_original_id_for_all_night_shift():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 160, "OT1.5": 8, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": True,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 95, "level": None, "coefficient": None}}

    employees = parser.build_employees(attendance_data, salary_data, performance_data)

    assert [emp.employee_id for emp in employees] == ["E001"]
    assert employees[0].source_employee_id == "E001"
    assert employees[0].hourly_rate == 21
    assert employees[0].is_night_shift is True


def test_management_trainee_without_performance_defaults_coefficient_to_one():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    employee_info = {
        "E001": {
            "name": "管培员工",
            "department": "FBU仓储事业部-美洲区-新泽西区-新泽西仓",
            "area": "新泽西区",
            "position": "Management Trainee 管培生",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(attendance_data, salary_data, {}, employee_info)
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert emp.uploaded_coefficient == 1.0
    assert emp.performance_coefficient == 1.0
    assert emp.performance_bonus == 320
    assert "未匹配绩效报表" not in emp.exceptions
    assert any("管培生" in msg for msg in emp.exceptions)


def test_resigned_management_trainee_without_performance_is_deferred():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 23.81, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 18, "ratio": 0.09}}
    employee_info = {
        "E001": {
            "name": "离职管培",
            "department": "FBU仓储事业部-美洲区-新泽西区-管培组",
            "area": "新泽西区",
            "personnel_status": "离职",
            "resignation_date": date(2026, 4, 3),
            "position": "Management Trainee 管培生",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(attendance_data, salary_data, {}, employee_info, calc_month="2026-04")
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert emp.is_deferred is True
    assert emp.performance_coefficient == 0
    assert emp.performance_bonus == 0
    assert "未匹配绩效报表" in emp.exceptions
    assert not any("管培生" in msg for msg in emp.exceptions)


def test_management_trainee_with_performance_still_uses_fixed_coefficient_one():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 125, "level": "超出预期", "coefficient": 1.6}}
    employee_info = {
        "E001": {
            "name": "管培员工",
            "department": "FBU仓储事业部-美洲区-新泽西区-新泽西仓",
            "area": "新泽西区",
            "position": "Management Trainee 管培生",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(attendance_data, salary_data, performance_data, employee_info)
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert emp.performance_score == 125
    assert emp.uploaded_coefficient == 1.0
    assert emp.performance_coefficient == 1.0
    assert emp.performance_bonus == 320
    assert not any("上传绩效系数与系统计算系数不一致" in msg for msg in emp.exceptions)
    assert any("管培生" in msg for msg in emp.exceptions)


def test_probation_status_uses_fixed_coefficient_one_even_with_performance():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 125, "level": "超出预期", "coefficient": 1.6}}
    employee_info = {
        "E001": {
            "name": "试用员工",
            "department": "FBU仓储事业部-美洲区-新泽西区-新泽西仓",
            "area": "新泽西区",
            "personnel_status": "试用",
            "position": "仓库专员",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(attendance_data, salary_data, performance_data, employee_info)
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert emp.personnel_status == "试用"
    assert emp.performance_score == 125
    assert emp.performance_coefficient == 1.0
    assert emp.performance_bonus == 320
    assert any("人员状态为试用" in msg for msg in emp.exceptions)


def test_resigned_status_uses_oehr_result_when_performance_exists():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 110, "level": "超出预期", "coefficient": 1.3}}
    employee_info = {
        "E001": {
            "name": "离职员工",
            "department": "FBU仓储事业部-美洲区-新泽西区-新泽西仓",
            "area": "新泽西区",
            "personnel_status": "离职",
            "position": "仓库专员",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(attendance_data, salary_data, performance_data, employee_info)
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert emp.personnel_status == "离职"
    assert emp.performance_coefficient == 1.3
    assert emp.performance_bonus == 416
    assert not any("人员状态为离职" in msg for msg in emp.exceptions)


def test_regular_employee_without_performance_is_deferred():
    emp = EmployeeData(
        employee_id="E001",
        name="正式员工",
        personnel_status="正式",
        hourly_rate=20,
        performance_ratio=0.1,
        base_hours=160,
    )

    BonusCalculator.calculate(emp)

    assert emp.is_deferred is True
    assert emp.performance_bonus == 0
    assert emp.deferred_reason == "OEHR绩效结果尚未出，延期发放"


def test_confirmation_before_or_on_15th_uses_oehr_result():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 125, "level": "超出预期", "coefficient": 1.6}}
    employee_info = {
        "E001": {
            "name": "月中转正员工",
            "department": "FBU仓储事业部-美洲区-新泽西区-新泽西仓",
            "area": "新泽西区",
            "personnel_status": "正式",
            "confirmation_date": date(2026, 4, 15),
            "position": "仓库专员",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(
        attendance_data,
        salary_data,
        performance_data,
        employee_info,
        calc_month="2026-04",
    )
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert emp.coefficient_override_reason == ""
    assert emp.performance_coefficient == 1.6
    assert emp.performance_bonus == 512


def test_confirmation_after_15th_defaults_coefficient_to_one():
    parser = FBUPerformanceParser()
    attendance_data = {
        "E001": {
            "白班": {"计薪出勤": 160, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "夜班": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            "has_night_shift": False,
        },
    }
    salary_data = {"E001": {"hourly_rate": 20, "ratio": 0.1}}
    performance_data = {"E001": {"score": 125, "level": "超出预期", "coefficient": 1.6}}
    employee_info = {
        "E001": {
            "name": "月中转正员工",
            "department": "FBU仓储事业部-美洲区-新泽西区-新泽西仓",
            "area": "新泽西区",
            "personnel_status": "正式",
            "confirmation_date": date(2026, 4, 16),
            "position": "仓库专员",
            "job_type": "warehouse",
        }
    }

    employees = parser.build_employees(
        attendance_data,
        salary_data,
        performance_data,
        employee_info,
        calc_month="2026-04",
    )
    emp = employees[0]
    BonusCalculator.calculate(emp)

    assert "15号后转正" in emp.coefficient_override_reason
    assert emp.performance_coefficient == 1.0
    assert emp.performance_bonus == 320


def test_confirmation_after_15th_without_adjustment_keeps_full_month_base():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt0021990",
                "name": "张海冰",
                "department": "新泽西21号仓（SN）",
                "area": "新泽西区",
                "personnel_status": "正式",
                "confirmation_date": "2026-04-26",
                "position": "Tallyman 理货员",
                "job_type": "warehouse",
                "has_night_shift": False,
                "day_shift": {"计薪出勤": 183.95, "OT1.5": 18.3, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "attendance_daily_rows": [
                    {"date": "2026-04-01", "shift_type": "白班", "base_hours": 63.95, "ot15_hours": 9.4},
                    {"date": "2026-04-12", "shift_type": "白班", "base_hours": 80, "ot15_hours": 8.44},
                    {"date": "2026-04-26", "shift_type": "白班", "base_hours": 40, "ot15_hours": 0.46},
                ],
            }
        ],
        salary_data=[
            {
                "employee_id": "zt0021990",
                "hourly_rate": 18,
                "ratio": 0.05,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[],
        calc_month="2026-04",
    )

    emp = engine.get_employee("zt0021990")
    assert round(emp.base_hours, 2) == 183.95
    assert round(emp.ot15_hours, 2) == 18.3
    assert round(emp.performance_base, 2) == 3805.2
    assert emp.performance_coefficient == 1.0
    assert round(emp.performance_bonus, 2) == 190.26


def test_adjustment_split_with_mixed_shift_is_not_duplicated():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "E001",
                "name": "Ana",
                "department": "新泽西仓",
                "area": "新泽西区",
                "job_type": "warehouse",
                "has_night_shift": True,
                "day_shift": {"计薪出勤": 120, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 40, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            }
        ],
        salary_data=[
            {
                "employee_id": "E001",
                "hourly_rate": 20,
                "ratio": 0.1,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[],
        adjustment_data=[
            {
                "employee_id": "E001",
                "segments": [
                    {"period": "4.1-4.15", "reason": "调薪前", "performance_base": 1000},
                    {"period": "4.16-4.30", "reason": "调薪后", "performance_base": 2000},
                ],
            }
        ],
    )

    employees = engine.get_all_employees()
    assert [emp.employee_id for emp in employees] == ["E001"]
    assert employees[0].performance_base == 2000
    assert round(employees[0].performance_bonus, 2) == 200


def test_performance_preview_average_uses_scored_employee_count(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append([""] * 19)
    row_with_score = [""] * 19
    row_with_score[3] = "E001"
    row_with_score[16] = 80
    row_with_score[17] = "符合预期"
    row_with_score[18] = 1
    sheet.append(row_with_score)
    row_without_score = [""] * 19
    row_without_score[3] = "E002"
    row_without_score[17] = "符合预期+"
    row_without_score[18] = 1.2
    sheet.append(row_without_score)
    path = tmp_path / "performance.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_performance_preview(str(path))

    assert preview["summary"]["total_employees"] == 2
    assert preview["summary"]["scored_employees"] == 1
    assert preview["summary"]["avg_score"] == 80


def test_performance_preview_accepts_resigned_employee_supplement_table(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "离职线下绩效考核表"
    sheet.append(["工号", "姓名", "绩效得分", "绩效等级", "绩效系数", "备注"])
    sheet.append(["zt0019943", "洪梓腾", 88.75, "符合预期-", 0.93, "5.29离职，线下绩效考核表"])
    path = tmp_path / "performance_supplement.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_performance_preview(str(path))

    assert preview["summary"]["source_type"] == "performance_supplement"
    assert preview["summary"]["total_employees"] == 1
    employee = preview["employees"][0]
    assert employee["employee_id"] == "zt0019943"
    assert employee["name"] == "洪梓腾"
    assert employee["score"] == 88.75
    assert employee["level"] == "符合预期-"
    assert employee["coefficient"] == 0.93
    assert employee["performance_source"] == "绩效补录"
    assert employee["note"] == "5.29离职，线下绩效考核表"


def test_performance_preview_header_parse_keeps_oehr_report_source_type(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "4月绩效报表"
    sheet.append(["姓名", "部门", "岗位", "工号"] + [""] * 12 + ["总分", "总等级", "绩效系数"])
    sheet.append(["Ana", "", "", "zt001", *[""] * 12, 95, "符合预期", 1])
    path = tmp_path / "performance_report.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_performance_preview(str(path))

    assert preview["summary"]["source_type"] == "performance_report"
    assert preview["employees"][0]["employee_id"] == "zt001"
    assert preview["employees"][0]["performance_source"] == "OEHR绩效报表"


def test_performance_supplement_row_can_drive_resigned_employee_bonus():
    parser = FBUPerformanceParser()

    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt0019943",
                "name": "洪梓腾",
                "department": "新泽西21号仓（SN）",
                "area": "新泽西区",
                "personnel_status": "离职",
                "position": "Warehouse Staff",
                "job_type": "warehouse",
                "has_night_shift": False,
                "day_shift": {"计薪出勤": 167.48, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "年假": 0, "节假日": 0},
            }
        ],
        salary_data=[
            {
                "employee_id": "zt0019943",
                "hourly_rate": 23,
                "ratio": 0.06,
                "calculation_method": "固定比例核算",
                "fixed_performance_base": 0,
            }
        ],
        performance_data=[
            {
                "employee_id": "zt0019943",
                "score": 88.75,
                "level": "符合预期-",
                "coefficient": 0.93,
                "performance_source": "绩效补录",
            }
        ],
        calc_month="2026-04",
    )

    emp = engine.get_employee("zt0019943")
    assert round(emp.performance_base, 2) == 3852.04
    assert emp.performance_coefficient == 0.93
    assert round(emp.performance_bonus, 2) == 214.94
    assert "未匹配绩效报表" not in emp.exceptions


def test_attendance_preview_uses_attendance_name_when_roster_is_missing(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["header"] * 118)
    row = [""] * 118
    row[0] = "2026-04-01"
    row[1] = "Ana Attendance"
    row[2] = "E001"
    row[21] = "08:00"
    row[117] = 8
    sheet.append(row)
    path = tmp_path / "attendance.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_attendance_preview(str(path), target_month=4)

    employee = preview["employees"][0]
    assert employee["name"] == "Ana Attendance"
    assert employee["roster_matched"] is False
    assert preview["summary"]["roster_missing"] == 1


def test_attendance_preview_daily_rows_keep_each_attendance_date(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["header"] * 149)
    for attendance_date, base_hours in [("2026/04/30", 8.25), ("2026/04/29", 7.75)]:
        row = [""] * 149
        row[0] = attendance_date
        row[1] = "Ana Attendance"
        row[2] = "E001"
        row[21] = "08:00"
        row[117] = base_hours
        sheet.append(row)
    path = tmp_path / "attendance.xlsx"
    workbook.save(path)

    preview = FBUPerformanceParser().parse_attendance_preview(str(path), target_month=4)

    daily_rows = preview["employees"][0]["attendance_daily_rows"]
    assert [row["date"] for row in daily_rows] == ["2026-04-30", "2026-04-29"]
    assert [row["base_hours"] for row in daily_rows] == [8.25, 7.75]


def test_attendance_preview_loads_large_workbook_in_read_only_mode(monkeypatch, tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["考勤日期", "姓名", "工号"])
    sheet.append(["2026-04-01", "测试员工", "zt1"])
    path = tmp_path / "attendance.xlsx"
    workbook.save(path)

    observed = {}
    original_load_workbook = openpyxl.load_workbook

    def capture_load_workbook(*args, **kwargs):
        observed.update(kwargs)
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", capture_load_workbook)

    FBUPerformanceParser().parse_attendance_preview(str(path), 4)

    assert observed["read_only"] is True


def test_roster_loader_finds_shifted_lingse_column_by_header(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    headers = [""] * 123
    headers[0] = "姓名"
    headers[3] = "工号"
    headers[19] = "二级部门"
    headers[20] = "三级部门"
    headers[89] = "划分区域"
    headers[122] = "领色"
    sheet.append(headers)

    for employee_id, name, department_parts, lingse in [
        ("E001", "Ana Roster", ("FBU", "Americas"), "蓝领"),
        ("E002", "Ben Roster", ("FBU", "Americas"), "白领"),
        ("E003", "Cara Roster", ("HRAS人力综合条线", "FBU HRBP Dept.", "新泽西区HRBP部"), "蓝领"),
        ("zt15638", "万其鑫", ("FBU仓储事业部", "美洲区", "新泽西区"), "白领"),
    ]:
        row = [""] * 123
        row[0] = name
        row[3] = employee_id
        row[7] = "试用" if employee_id == "E001" else "正式"
        row[19] = department_parts[0]
        row[20] = department_parts[1]
        if len(department_parts) > 2:
            row[21] = department_parts[2]
        row[30] = "Management Trainee 管培生" if employee_id == "E001" else ""
        row[89] = "US-West"
        row[122] = lingse
        sheet.append(row)

    path = tmp_path / "roster.xlsx"
    workbook.save(path)

    roster = FBUPerformanceParser().load_roster(str(path))

    assert roster["E001"]["name"] == "Ana Roster"
    assert roster["E001"]["department"] == "FBU-Americas"
    assert roster["E001"]["area"] == "US-West"
    assert roster["E001"]["personnel_status"] == "试用"
    assert roster["E001"]["position"] == "Management Trainee 管培生"
    assert roster["E001"]["job_type"] == "warehouse"
    assert roster["E002"]["job_type"] == "warehouse"
    assert roster["E003"]["job_type"] == "functional"
    assert roster["zt15638"]["job_type"] == "district_manager"


def test_base_roster_store_saves_metadata_and_copies_snapshot(tmp_path):
    store = FBURosterStore(str(tmp_path))
    content = b"fake roster bytes"

    metadata = store.save_active_roster(content, "roster.xlsx", total_employees=12)
    copied = store.copy_active_to_run("run123")

    assert metadata["filename"] == "roster.xlsx"
    assert metadata["total_employees"] == 12
    assert store.get_metadata()["has_roster"] is True
    assert copied.read_bytes() == content
    assert copied.name == "roster.xlsx"
