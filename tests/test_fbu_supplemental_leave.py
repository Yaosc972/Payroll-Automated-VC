from pathlib import Path
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

from bonus_platform import app as app_module
from bonus_platform.app import app
from bonus_platform.engine.fbu_performance.engines.base import EmployeeData
from bonus_platform.engine.fbu_performance.parser import FBUPerformanceParser


pytestmark = pytest.mark.usefixtures("bypass_fbu_access_gate")


def _leave_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假3.29-4.11"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/4/2 9:15", "2026/4/2 18:00", 8, "LC-001", "2026/4/17", "", "病假"])
    sheet.append(["zt002", "员工乙", "离职", None, None, 15, "LC-002", "2026/4/17", "", "病假"])
    sheet.append(["zt003", "员工丙", "离职", None, None, 12, "LC-003", "2026/4/17", "", "年假"])
    workbook.save(path)


def _leave_workbook_bytes() -> bytes:
    buffer = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假3.29-4.11"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/4/2 9:15", "2026/4/2 18:00", 8, "LC-001", "2026/4/17", "", "病假"])
    sheet.append(["zt002", "员工乙", "离职", None, None, 15, "LC-002", "2026/4/17", "", "病假"])
    sheet.append(["zt003", "员工丙", "离职", None, None, 12, "LC-003", "2026/4/17", "", "年假"])
    workbook.save(buffer)
    return buffer.getvalue()


def _cross_month_leave_workbook_bytes() -> bytes:
    buffer = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.26-5.09"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/4/27 9:15", "2026/5/9 18:00", 80, "LC-001", "2026/5/15", "", "年假"])
    workbook.save(buffer)
    return buffer.getvalue()


def _split_flow_leave_workbook_bytes() -> bytes:
    buffer = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.12-4.25"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt15556", "郑晓君", "正式", "2026/4/20 9:15", "2026/4/20 18:00", 0.15, "ZTJT-LC-20260417131933", "2026/5/1", "", "病假"])
    sheet.append(["zt15556", "郑晓君", "正式", "2026/4/20 9:15", "2026/4/20 18:00", 7.85, "ZTJT-LC-20260417131933", "2026/5/1", "", "病假"])
    workbook.save(buffer)
    return buffer.getvalue()


def _parser_with_roster() -> FBUPerformanceParser:
    parser = FBUPerformanceParser()
    parser.employee_roster = {
        "zt001": {"name": "员工甲", "termination_date": "", "employee_status": "正式"},
        "zt002": {"name": "员工乙", "termination_date": "2026-04-30", "employee_status": "离职"},
        "zt003": {"name": "员工丙", "termination_date": "2026-04-30", "employee_status": "离职"},
    }
    return parser


def _supplemental_preview_for_suggestions() -> dict:
    return {
        "rows": [
            {
                "row_id": "leave:2",
                "employee_id": "zt001",
                "name": "员工甲",
                "leave_type": "年假",
                "hours": 80,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "allocation_month": "2026-04",
                "attendance_matched": True,
            },
            {
                "row_id": "leave:3",
                "employee_id": "zt002",
                "name": "员工乙",
                "leave_type": "离职年假结算",
                "hours": 8,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "allocation_month": "2026-04",
                "attendance_matched": True,
            },
            {
                "row_id": "leave:4",
                "employee_id": "zt003",
                "name": "员工丙",
                "leave_type": "病假",
                "hours": 8,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "allocation_month": "2026-04",
                "attendance_matched": True,
            },
            {
                "row_id": "leave:5",
                "employee_id": "zt004-1",
                "name": "员工丁",
                "leave_type": "病假",
                "hours": 16,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "allocation_month": "2026-04",
                "attendance_matched": True,
            },
        ],
        "summary": {},
    }


def test_supplemental_leave_suggestion_uses_offline_gap_when_exact_row_is_safe():
    preview = _supplemental_preview_for_suggestions()
    employees = [
        EmployeeData("zt001", "员工甲", hourly_rate=17, performance_base=1000),
    ]

    updated = FBUPerformanceParser().apply_supplemental_leave_suggestions(
        preview,
        employees,
        offline_bases=[{"employee_id": "zt001", "offline_performance_base": 1408}],
        calc_month="2026-04",
    )

    row = updated["rows"][0]
    assert row["suggested_included_hours"] == 24
    assert row["suggestion_status"] == "suggested"
    assert "线下基数差异反推" in row["suggestion_reason"]
    assert row["confirmation_status"] == "pending"
    assert row["include_in_base"] is False


def test_supplemental_leave_suggestion_respects_rule_boundaries():
    preview = _supplemental_preview_for_suggestions()
    employees = [
        EmployeeData("zt002", "员工乙", hourly_rate=20, performance_base=1000),
        EmployeeData("zt003", "员工丙", hourly_rate=20, performance_base=1000, base_override_amount=1100, base_override_type="96工时制"),
    ]

    updated = FBUPerformanceParser().apply_supplemental_leave_suggestions(
        preview,
        employees,
        offline_bases=[
            {"employee_id": "zt002", "offline_performance_base": 1160},
            {"employee_id": "zt003", "offline_performance_base": 1160},
        ],
        calc_month="2026-04",
    )

    by_id = {row["employee_id"]: row for row in updated["rows"]}
    assert by_id["zt002"]["suggestion_status"] == "blocked"
    assert by_id["zt002"]["suggested_included_hours"] == ""
    assert "离职年假结算" in by_id["zt002"]["suggestion_reason"]
    assert by_id["zt003"]["suggestion_status"] == "blocked"
    assert "固定基数" in by_id["zt003"]["suggestion_reason"]


def test_supplemental_leave_suggestion_requires_exact_shift_row():
    preview = _supplemental_preview_for_suggestions()
    employees = [
        EmployeeData("zt004", "员工丁", hourly_rate=18, performance_base=900, is_night_shift=True),
    ]

    updated = FBUPerformanceParser().apply_supplemental_leave_suggestions(
        preview,
        employees,
        offline_bases=[{"employee_id": "zt004", "offline_performance_base": 1080}],
        calc_month="2026-04",
    )

    by_id = {row["employee_id"]: row for row in updated["rows"]}
    assert by_id["zt004-1"]["suggestion_status"] == "blocked"
    assert by_id["zt004-1"]["suggested_included_hours"] == ""
    assert "精确工号" in by_id["zt004-1"]["suggestion_reason"]


def test_base_summary_offline_bases_use_first_employee_base_block(tmp_path):
    rows = [
        (None, None, None, None, None, None, "绩效取数", None, None),
        ("工号", "中文名字", "求和项:合计所得", None, "工号", "姓名", "合计基数", "差异备注"),
        ("zt001", "员工甲", 100, None, "zt999", "错位员工", 999, "考勤日报块"),
    ]

    offline_bases = FBUPerformanceParser()._parse_offline_base_summary(
        "基数汇总",
        rows,
        1,
        rows[1],
    )

    assert offline_bases == [
        {
            "employee_id": "zt001",
            "name": "员工甲",
            "offline_performance_base": 100.0,
            "offline_rows": 1,
        }
    ]


def test_parse_supplemental_leave_preview_sets_defaults_for_regular_and_termination_rows(tmp_path):
    workbook_path = tmp_path / "leave.xlsx"
    _leave_workbook(workbook_path)

    preview = _parser_with_roster().parse_supplemental_leave_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["total_rows"] == 3
    assert preview["summary"]["pending_count"] == 1
    assert preview["summary"]["confirmed_count"] == 1
    assert preview["summary"]["excluded_count"] == 1

    regular, sick_settlement, annual_settlement = preview["rows"]
    assert regular["employee_id"] == "zt001"
    assert regular["leave_type"] == "病假"
    assert regular["include_in_base"] is True
    assert regular["confirmation_status"] == "confirmed"
    assert regular["allocation_month"] == "2026-04"
    assert regular["allocation_period"] == "4.1-4.11"
    assert regular["included_hours"] == 8
    assert regular["termination_date"] == ""

    assert sick_settlement["employee_id"] == "zt002"
    assert sick_settlement["leave_type"] == "离职病假结算"
    assert sick_settlement["include_in_base"] is False
    assert sick_settlement["confirmation_status"] == "pending"
    assert sick_settlement["allocation_month"] == "2026-04"
    assert sick_settlement["allocation_period"] == "4.1-4.11"
    assert sick_settlement["included_hours"] == 0
    assert sick_settlement["termination_date"] == "2026-04-30"
    assert "缺少申请时间和周期归属列" in sick_settlement["system_reason"]

    assert annual_settlement["employee_id"] == "zt003"
    assert annual_settlement["leave_type"] == "离职年假结算"
    assert annual_settlement["include_in_base"] is False
    assert annual_settlement["confirmation_status"] == "excluded"
    assert annual_settlement["included_hours"] == 0
    assert annual_settlement["termination_date"] == "2026-04-30"


def test_apply_supplemental_leave_batch_updates_selected_rows_only(tmp_path):
    workbook_path = tmp_path / "leave.xlsx"
    _leave_workbook(workbook_path)
    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][1]["row_id"]

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={
            "confirmation_status": "excluded",
            "include_in_base": False,
            "allocation_month": "2026-03",
            "allocation_period": "3.29-3.31",
            "confirmation_note": "已在3月计入",
        },
    )

    regular, termination, annual_settlement = updated["rows"]
    assert regular["confirmation_status"] == "confirmed"
    assert regular["include_in_base"] is True
    assert annual_settlement["confirmation_status"] == "excluded"
    assert annual_settlement["include_in_base"] is False

    assert termination["confirmation_status"] == "excluded"
    assert termination["include_in_base"] is False
    assert termination["allocation_month"] == "2026-03"
    assert termination["allocation_period"] == "3.29-3.31"
    assert termination["confirmation_note"] == "已在3月计入"
    assert updated["summary"]["excluded_count"] == 2
    assert updated["summary"]["include_count"] == 1
    assert updated["summary"]["include_hours"] == 8


def test_apply_supplemental_leave_batch_updates_included_hours(tmp_path):
    workbook_path = tmp_path / "leave.xlsx"
    _leave_workbook(workbook_path)
    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][1]["row_id"]

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={"included_hours": 4},
    )

    termination = updated["rows"][1]
    assert termination["hours"] == 15
    assert termination["included_hours"] == 4
    assert termination["include_in_base"] is True
    assert termination["confirmation_status"] == "confirmed"
    assert updated["summary"]["include_hours"] == 12


def test_apply_supplemental_leave_batch_status_excluded_defaults_to_not_counting(tmp_path):
    workbook_path = tmp_path / "leave.xlsx"
    _leave_workbook(workbook_path)
    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][0]["row_id"]

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={"confirmation_status": "excluded"},
    )

    regular = updated["rows"][0]
    assert regular["confirmation_status"] == "excluded"
    assert regular["include_in_base"] is False
    assert regular["included_hours"] == 0
    assert updated["summary"]["include_count"] == 0
    assert updated["summary"]["include_hours"] == 0


def test_apply_supplemental_leave_batch_status_confirmed_defaults_to_counting_without_guessing_hours(tmp_path):
    workbook_path = tmp_path / "leave_without_split.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假3.29-4.11"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/3/30 9:15", "2026/4/1 18:00", 24, "LC-001", "2026/4/3", "", "年假"])
    workbook.save(workbook_path)

    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][0]["row_id"]

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={"confirmation_status": "confirmed"},
    )

    row = updated["rows"][0]
    assert row["confirmation_status"] == "confirmed"
    assert row["include_in_base"] is True
    assert row["included_hours"] == 0
    assert updated["summary"]["include_hours"] == 0


def test_apply_supplemental_leave_batch_can_confirm_settlement_hours(tmp_path):
    workbook_path = tmp_path / "leave.xlsx"
    _leave_workbook(workbook_path)
    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][2]["row_id"]

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={
            "confirmation_status": "confirmed",
            "include_in_base": True,
            "included_hours": 4,
        },
    )

    annual_settlement = updated["rows"][2]
    assert annual_settlement["leave_type"] == "离职年假结算"
    assert annual_settlement["included_hours"] == 4
    assert annual_settlement["include_in_base"] is True
    assert updated["summary"]["include_hours"] == 12


def test_apply_supplemental_leave_batch_confirms_pending_cross_month_hours(tmp_path):
    workbook_path = tmp_path / "leave_without_split.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假3.29-4.11"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/3/30 9:15", "2026/4/1 18:00", 24, "LC-001", "2026/4/3", "", "年假"])
    workbook.save(workbook_path)

    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][0]["row_id"]

    assert preview["rows"][0]["confirmation_status"] == "pending"
    assert preview["rows"][0]["include_in_base"] is False

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={
            "included_hours": 8,
            "confirmation_status": "confirmed",
            "include_in_base": True,
        },
    )

    row = updated["rows"][0]
    assert row["included_hours"] == 8
    assert row["confirmation_status"] == "confirmed"
    assert row["include_in_base"] is True
    assert updated["summary"]["pending_count"] == 0
    assert updated["summary"]["confirmed_count"] == 1
    assert updated["summary"]["include_hours"] == 8


def test_apply_supplemental_leave_batch_excludes_pending_cross_month_zero_hours(tmp_path):
    workbook_path = tmp_path / "leave_without_split.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假3.29-4.11"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/3/30 9:15", "2026/4/1 18:00", 24, "LC-001", "2026/4/3", "", "年假"])
    workbook.save(workbook_path)

    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    target_id = preview["rows"][0]["row_id"]

    updated = parser.apply_supplemental_leave_batch(
        preview,
        row_ids=[target_id],
        updates={
            "included_hours": 0,
            "confirmation_status": "excluded",
            "include_in_base": False,
        },
    )

    row = updated["rows"][0]
    assert row["included_hours"] == 0
    assert row["confirmation_status"] == "excluded"
    assert row["include_in_base"] is False
    assert updated["summary"]["pending_count"] == 0
    assert updated["summary"]["excluded_count"] == 1
    assert updated["summary"]["include_hours"] == 0


def test_parse_supplemental_leave_uses_current_month_split_column(tmp_path):
    workbook_path = tmp_path / "split_leave.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.26-5.09"
    sheet.append([
        "工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号",
        "发放时间", "不含大区", "假期类型", "验算", 1, "备注", "主体", "", "",
        "4.26-4.30剔除离职结算年假", "5.1-5.9",
    ])
    sheet.append([
        "zt001", "员工甲", "年假-休假", "2026/4/27 9:15", "2026/5/9 18:00",
        80, "LC-003", "2026/5/15", "", "年假", "", True, "", "US EL", "", "",
        24, 56,
    ])
    sheet.append([
        "zt003", "员工丙", "正式", "2026/5/5 9:15", "2026/5/5 18:00",
        8, "LC-004", "2026/5/15", "", "病假", "", True, "", "US EL", "", "",
        None, 8,
    ])
    workbook.save(workbook_path)

    preview = _parser_with_roster().parse_supplemental_leave_preview(str(workbook_path), "2026-04")

    current_row, future_row = preview["rows"]
    assert current_row["included_hours"] == 24
    assert current_row["include_in_base"] is True
    assert current_row["allocation_period"] == "4.26-4.30"

    assert future_row["hours"] == 8
    assert future_row["included_hours"] == 0
    assert future_row["include_in_base"] is False
    assert future_row["confirmation_status"] == "excluded"
    assert preview["summary"]["include_hours"] == 24


def test_parse_supplemental_leave_counts_termination_sick_only_with_current_split_column(tmp_path):
    workbook_path = tmp_path / "termination_split_leave.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.26-5.09"
    sheet.append([
        "工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号",
        "发放时间", "不含大区", "假期类型", "验算", 1, "备注", "主体", "", "",
        "4.26-4.30剔除离职结算年假", "5.1-5.9",
    ])
    sheet.append([
        "zt002", "员工乙", "离职", None, None, 15, "LC-002", "2026/5/15", "",
        "病假", "", True, "", "US EL", "", "", 6, 9,
    ])
    workbook.save(workbook_path)

    preview = _parser_with_roster().parse_supplemental_leave_preview(str(workbook_path), "2026-04")

    row = preview["rows"][0]
    assert row["leave_type"] == "离职病假结算"
    assert row["included_hours"] == 6
    assert row["include_in_base"] is True
    assert row["confirmation_status"] == "confirmed"
    assert row["allocation_period"] == "4.26-4.30"
    assert preview["summary"]["include_hours"] == 6


def test_parse_supplemental_leave_merges_duplicate_flow_rows(tmp_path):
    workbook_path = tmp_path / "split_by_flow_leave.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.12-4.25"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt15556", "郑晓君", "正式", "2026/4/20 9:15", "2026/4/20 18:00", 0.15, "ZTJT-LC-20260417131933", "2026/5/1", "", "病假"])
    sheet.append(["zt15556", "郑晓君", "正式", "2026/4/20 9:15", "2026/4/20 18:00", 7.85, "ZTJT-LC-20260417131933", "2026/5/1", "", "病假"])
    workbook.save(workbook_path)

    preview = _parser_with_roster().parse_supplemental_leave_preview(str(workbook_path), "2026-04")

    assert len(preview["rows"]) == 1
    row = preview["rows"][0]
    assert row["employee_id"] == "zt15556"
    assert row["hours"] == 8
    assert row["included_hours"] == 8
    assert row["source_rows"] == [2, 3]
    assert preview["summary"]["total_rows"] == 1
    assert preview["summary"]["include_hours"] == 8


def test_validate_supplemental_leave_keeps_unmatched_rows_included_by_default(tmp_path):
    workbook_path = tmp_path / "leave.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.26-5.09"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/4/27 9:15", "2026/4/27 18:00", 8, "LC-001", "2026/5/15", "", "病假"])
    sheet.append(["zt999", "员工外", "正式", "2026/4/27 9:15", "2026/4/27 18:00", 8, "LC-999", "2026/5/15", "", "病假"])
    workbook.save(workbook_path)

    parser = _parser_with_roster()
    preview = parser.parse_supplemental_leave_preview(str(workbook_path), "2026-04")
    validated = parser.validate_supplemental_leave_against_attendance(
        preview,
        {
            "employees": [
                {"employee_id": "zt001-1"},
            ],
        },
    )

    matched, unmatched = validated["rows"]
    assert matched["attendance_matched"] is True
    assert matched["include_in_base"] is True
    assert matched["confirmation_status"] == "confirmed"

    assert unmatched["attendance_matched"] is False
    assert unmatched["include_in_base"] is True
    assert unmatched["included_hours"] == 8
    assert unmatched["confirmation_status"] == "confirmed"
    assert "未匹配本月考勤" in unmatched["system_reason"]
    assert "默认计入" in unmatched["system_reason"]
    assert validated["summary"]["attendance_unmatched_count"] == 1
    assert validated["summary"]["include_hours"] == 16


def test_parse_supplemental_leave_without_split_columns_holds_cross_month_rows(tmp_path):
    workbook_path = tmp_path / "leave_without_split.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sickpay&年假4.26-5.09"
    sheet.append(["工号", "姓名", "人员状态", "申请时间", "申请时间", "总时长", "流程号", "发放时间", "不含大区", "假期类型"])
    sheet.append(["zt001", "员工甲", "正式", "2026/4/27 9:15", "2026/5/9 18:00", 80, "LC-001", "2026/5/15", "", "年假"])
    sheet.append(["zt002", "员工乙", "正式", "2026/5/5 9:15", "2026/5/5 18:00", 8, "LC-002", "2026/5/15", "", "病假"])
    workbook.save(workbook_path)

    preview = _parser_with_roster().parse_supplemental_leave_preview(str(workbook_path), "2026-04")

    cross_month, future_month = preview["rows"]
    assert cross_month["confirmation_status"] == "pending"
    assert cross_month["include_in_base"] is False
    assert cross_month["included_hours"] == 0
    assert "缺少周期拆分列" in cross_month["system_reason"]

    assert future_month["confirmation_status"] == "excluded"
    assert future_month["include_in_base"] is False
    assert future_month["included_hours"] == 0
    assert "不属于当前核算月份" in future_month["system_reason"]


def test_supplemental_leave_import_and_batch_api_updates_run():
    client = TestClient(app)
    created = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()

    imported = client.post(
        "/api/fbu-performance/import-supplemental-leave",
        data={"run_id": created["run_id"]},
        files={"file": ("leave.xlsx", _leave_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert imported.status_code == 200
    preview = imported.json()["preview"]
    assert preview["summary"]["excluded_count"] == 1
    pending_row_id = next(row["row_id"] for row in preview["rows"] if row["leave_type"] == "离职病假结算")

    updated = client.post(
        f"/api/fbu-performance/runs/{created['run_id']}/supplemental-leave/batch",
        json={
            "row_ids": [pending_row_id],
            "confirmation_status": "excluded",
            "include_in_base": False,
            "allocation_month": "2026-03",
            "allocation_period": "3.29-3.31",
            "confirmation_note": "已在3月计入",
            "included_hours": 4,
        },
    )

    assert updated.status_code == 200
    assert updated.json()["preview"]["summary"]["excluded_count"] == 2
    assert updated.json()["preview"]["rows"][1]["included_hours"] == 4
    detail = client.get(f"/api/fbu-performance/runs/{created['run_id']}").json()
    assert detail["supplemental_leave_data"]["summary"]["excluded_count"] == 2


def test_supplemental_leave_api_infers_status_when_saving_included_hours_only():
    client = TestClient(app)
    created = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()

    imported = client.post(
        "/api/fbu-performance/import-supplemental-leave",
        data={"run_id": created["run_id"]},
        files={"file": ("leave.xlsx", _cross_month_leave_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert imported.status_code == 200
    row = imported.json()["preview"]["rows"][0]
    assert row["confirmation_status"] == "pending"
    assert row["include_in_base"] is False

    updated = client.post(
        f"/api/fbu-performance/runs/{created['run_id']}/supplemental-leave/batch",
        json={
            "row_ids": [row["row_id"]],
            "included_hours": 8,
        },
    )

    assert updated.status_code == 200
    updated_row = updated.json()["preview"]["rows"][0]
    assert updated_row["included_hours"] == 8
    assert updated_row["confirmation_status"] == "confirmed"
    assert updated_row["include_in_base"] is True
    assert updated.json()["preview"]["summary"]["pending_count"] == 0
    assert updated.json()["preview"]["summary"]["include_hours"] == 8


def test_supplemental_leave_api_supports_compact_row_response():
    client = TestClient(app)
    created = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()

    imported = client.post(
        "/api/fbu-performance/import-supplemental-leave",
        data={"run_id": created["run_id"]},
        files={"file": ("leave.xlsx", _cross_month_leave_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    row = imported.json()["preview"]["rows"][0]

    updated = client.post(
        f"/api/fbu-performance/runs/{created['run_id']}/supplemental-leave/batch",
        json={
            "row_ids": [row["row_id"]],
            "included_hours": 8,
            "response_mode": "row",
        },
    )

    assert updated.status_code == 200
    payload = updated.json()
    assert "preview" not in payload
    assert payload["row"]["row_id"] == row["row_id"]
    assert payload["row"]["included_hours"] == 8
    assert payload["summary"]["pending_count"] == 0
    assert payload["summary"]["include_hours"] == 8


def test_supplemental_leave_api_merges_duplicate_flow_rows_on_import():
    client = TestClient(app)
    created = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()

    imported = client.post(
        "/api/fbu-performance/import-supplemental-leave",
        data={"run_id": created["run_id"]},
        files={"file": ("leave.xlsx", _split_flow_leave_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert imported.status_code == 200
    preview = imported.json()["preview"]
    assert preview["summary"]["total_rows"] == 1
    assert preview["summary"]["include_hours"] == 8
    row = preview["rows"][0]
    assert row["hours"] == 8
    assert row["included_hours"] == 8
    assert row["source_rows"] == [2, 3]


def test_supplemental_leave_api_status_excluded_defaults_to_not_counting():
    client = TestClient(app)
    created = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()

    imported = client.post(
        "/api/fbu-performance/import-supplemental-leave",
        data={"run_id": created["run_id"]},
        files={"file": ("leave.xlsx", _leave_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert imported.status_code == 200
    regular_row = imported.json()["preview"]["rows"][0]
    assert regular_row["include_in_base"] is True
    assert regular_row["included_hours"] == 8

    updated = client.post(
        f"/api/fbu-performance/runs/{created['run_id']}/supplemental-leave/batch",
        json={
            "row_ids": [regular_row["row_id"]],
            "confirmation_status": "excluded",
        },
    )

    assert updated.status_code == 200
    updated_row = updated.json()["preview"]["rows"][0]
    assert updated_row["confirmation_status"] == "excluded"
    assert updated_row["include_in_base"] is False
    assert updated_row["included_hours"] == 0
    assert updated.json()["preview"]["summary"]["include_hours"] == 0


def test_supplemental_leave_api_applies_all_pending_suggestions_only():
    client = TestClient(app)
    created = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()
    run = app_module.fbu_run_manager.get_run(created["run_id"])
    run.supplemental_leave_data = {
        "rows": [
            {
                "row_id": "leave:2",
                "employee_id": "zt001",
                "leave_type": "年假",
                "hours": 16,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "suggestion_status": "suggested",
                "suggested_included_hours": 8,
            },
            {
                "row_id": "leave:3",
                "employee_id": "zt002",
                "leave_type": "离职年假结算",
                "hours": 16,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "suggestion_status": "",
                "suggested_included_hours": "",
            },
            {
                "row_id": "leave:4",
                "employee_id": "zt003",
                "leave_type": "病假",
                "hours": 24,
                "included_hours": 0,
                "confirmation_status": "confirmed",
                "include_in_base": False,
                "suggestion_status": "suggested",
                "suggested_included_hours": 12,
            },
            {
                "row_id": "leave:5",
                "employee_id": "zt004",
                "leave_type": "离职病假结算",
                "hours": 10,
                "included_hours": 0,
                "confirmation_status": "pending",
                "include_in_base": False,
                "suggestion_status": "suggested",
                "suggested_included_hours": 10,
            },
        ],
        "summary": {},
    }
    app_module.fbu_run_manager.update_run(created["run_id"], supplemental_leave_data=run.supplemental_leave_data)

    updated = client.post(
        f"/api/fbu-performance/runs/{created['run_id']}/supplemental-leave/batch",
        json={"apply_suggestions": True},
    )

    assert updated.status_code == 200
    body = updated.json()
    assert body["applied_count"] == 2
    rows = {row["row_id"]: row for row in body["preview"]["rows"]}
    assert rows["leave:2"]["included_hours"] == 8
    assert rows["leave:2"]["confirmation_status"] == "confirmed"
    assert rows["leave:2"]["include_in_base"] is True
    assert rows["leave:5"]["included_hours"] == 10
    assert rows["leave:5"]["confirmation_status"] == "confirmed"
    assert rows["leave:3"]["confirmation_status"] == "pending"
    assert rows["leave:3"]["included_hours"] == 0
    assert rows["leave:4"]["confirmation_status"] == "confirmed"
    assert rows["leave:4"]["included_hours"] == 0
    assert body["preview"]["summary"]["include_hours"] == 18


def test_step_calculation_uses_confirmed_supplemental_leave_hours():
    parser = FBUPerformanceParser()
    attendance_rows = [
        {
            "employee_id": "zt001",
            "name": "员工甲",
            "department": "仓库",
            "area": "新泽西区",
            "personnel_status": "正式",
            "job_type": "warehouse",
            "has_night_shift": False,
            "day_shift": {
                "计薪出勤": 160,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 99,
                "年假": 99,
                "节假日": 0,
            },
            "night_shift": {
                "计薪出勤": 0,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 0,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
        }
    ]
    salary_rows = [
        {
            "employee_id": "zt001",
            "hourly_rate": 20,
            "ratio": 0.1,
            "fixed_performance_base": 0,
        }
    ]
    performance_rows = [
        {
            "employee_id": "zt001",
            "score": 95,
            "level": "A",
            "coefficient": 1,
        }
    ]
    supplemental_leave_data = {
        "rows": [
            {
                "employee_id": "zt001",
                "leave_type": "病假",
                "included_hours": 8,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
            {
                "employee_id": "zt001",
                "leave_type": "年假",
                "included_hours": 4,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
            {
                "employee_id": "zt001",
                "leave_type": "离职病假结算",
                "included_hours": 3,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
            {
                "employee_id": "zt001",
                "leave_type": "离职年假结算",
                "included_hours": 5,
                "include_in_base": False,
                "confirmation_status": "excluded",
                "allocation_month": "2026-04",
            },
        ]
    }

    engine = parser.parse_all_from_step_data(
        attendance_rows,
        salary_rows,
        performance_rows,
        supplemental_leave_data=supplemental_leave_data,
        calc_month="2026-04",
    )

    employee = engine.get_all_employees()[0]
    assert employee.sick_hours == 8
    assert employee.annual_hours == 4
    assert employee.sick_settlement_hours == 3
    assert employee.performance_base == 20 * (160 + 8 + 4 + 3)


def test_supplemental_leave_keeps_day_and_night_shift_rows_separate():
    parser = FBUPerformanceParser()
    attendance_rows = [
        {
            "employee_id": "zt001-1",
            "name": "员工甲",
            "department": "仓库",
            "area": "新泽西区",
            "personnel_status": "正式",
            "job_type": "warehouse",
            "has_night_shift": False,
            "day_shift": {
                "计薪出勤": 80,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 99,
                "节假日": 0,
            },
            "night_shift": {
                "计薪出勤": 0,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 0,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
        },
        {
            "employee_id": "zt001",
            "name": "员工甲",
            "department": "仓库",
            "area": "新泽西区",
            "personnel_status": "正式",
            "job_type": "warehouse",
            "has_night_shift": True,
            "day_shift": {
                "计薪出勤": 0,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 0,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
            "night_shift": {
                "计薪出勤": 90,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 99,
                "节假日": 0,
            },
        },
    ]
    salary_rows = [
        {
            "employee_id": "zt001",
            "hourly_rate": 20,
            "ratio": 0.1,
            "fixed_performance_base": 0,
        }
    ]
    performance_rows = [
        {
            "employee_id": "zt001",
            "score": 95,
            "level": "A",
            "coefficient": 1,
        }
    ]
    supplemental_leave_data = {
        "rows": [
            {
                "employee_id": "zt001-1",
                "leave_type": "病假",
                "included_hours": 8,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
            {
                "employee_id": "zt001",
                "leave_type": "年假",
                "included_hours": 6,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
        ]
    }

    engine = parser.parse_all_from_step_data(
        attendance_rows,
        salary_rows,
        performance_rows,
        supplemental_leave_data=supplemental_leave_data,
        calc_month="2026-04",
    )

    by_id = {employee.employee_id: employee for employee in engine.get_all_employees()}
    assert by_id["zt001-1"].sick_hours == 8
    assert by_id["zt001-1"].annual_hours == 0
    assert by_id["zt001"].sick_hours == 0
    assert by_id["zt001"].annual_hours == 6
    assert by_id["zt001-1"].performance_base == 20 * (80 + 8)
    assert by_id["zt001"].performance_base == 21 * (90 + 6)


def test_supplemental_leave_with_original_id_uses_dated_shift_row_when_split():
    parser = FBUPerformanceParser()
    attendance_rows = [
        {
            "employee_id": "zt0020984",
            "name": "黄东俊",
            "department": "新泽西21号仓（SN）",
            "area": "新泽西区",
            "personnel_status": "正式",
            "job_type": "warehouse",
            "has_night_shift": True,
            "day_shift": {
                "计薪出勤": 143.76,
                "OT1.5": 25.75,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
            "night_shift": {
                "计薪出勤": 32.14,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
            "attendance_daily_rows": [
                {"date": "2026-04-10", "shift_type": "白班", "base_hours": 143.76, "ot15_hours": 25.75},
                {"date": "2026-04-26", "shift_type": "夜班", "base_hours": 32.14, "ot15_hours": 0},
            ],
        }
    ]
    salary_rows = [
        {
            "employee_id": "zt0020984",
            "hourly_rate": 18,
            "ratio": 0.05,
            "fixed_performance_base": 0,
        }
    ]
    performance_rows = [
        {
            "employee_id": "zt0020984",
            "score": 95,
            "level": "符合预期",
            "coefficient": 1,
        }
    ]
    supplemental_leave_data = {
        "rows": [
            {
                "employee_id": "zt0020984",
                "leave_type": "病假",
                "start_at": "2026-04-10T09:15",
                "end_at": "2026-04-10T18:00",
                "included_hours": 13,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
        ]
    }

    engine = parser.parse_all_from_step_data(
        attendance_rows,
        salary_rows,
        performance_rows,
        supplemental_leave_data=supplemental_leave_data,
        calc_month="2026-04",
    )

    by_id = {employee.employee_id: employee for employee in engine.get_all_employees()}
    assert by_id["zt0020984-1"].sick_hours == 13
    assert by_id["zt0020984"].sick_hours == 0
    assert round(by_id["zt0020984-1"].performance_base, 2) == 3516.93
    assert round(by_id["zt0020984"].performance_base, 2) == 610.66
    assert round(sum(emp.performance_base for emp in by_id.values()), 2) == 4127.59


def test_supplemental_leave_after_midnight_matches_cross_day_night_shift_row():
    parser = FBUPerformanceParser()
    attendance_rows = [
        {
            "employee_id": "zt18731",
            "name": "郑增良",
            "department": "新泽西21号仓（SN）",
            "area": "新泽西区",
            "personnel_status": "正式",
            "job_type": "warehouse",
            "has_night_shift": True,
            "day_shift": {
                "计薪出勤": 80,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
            "night_shift": {
                "计薪出勤": 40,
                "OT1.5": 6.17,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
            "attendance_daily_rows": [
                {
                    "date": "2026-05-12",
                    "shift_type": "夜班",
                    "base_hours": 40,
                    "ot15_hours": 6.17,
                    "sick_hours": 0,
                    "shift_start_time": "16:00",
                    "scheduled_hours": 8,
                    "work_hours": 8.6,
                },
                {
                    "date": "2026-05-13",
                    "shift_type": "白班",
                    "base_hours": 80,
                    "ot15_hours": 0,
                    "sick_hours": 0,
                    "shift_start_time": "09:15",
                    "scheduled_hours": 8,
                    "work_hours": 8,
                },
            ],
        }
    ]
    salary_rows = [
        {
            "employee_id": "zt18731",
            "hourly_rate": 27,
            "ratio": 0.25,
            "fixed_performance_base": 0,
        }
    ]
    performance_rows = [
        {
            "employee_id": "zt18731",
            "score": 101.43,
            "level": "符合预期+",
            "coefficient": 1.13,
        }
    ]
    supplemental_leave_data = {
        "rows": [
            {
                "employee_id": "zt18731",
                "leave_type": "病假",
                "start_at": "2026-05-13T00:05",
                "end_at": "2026-05-13T00:30",
                "included_hours": 8,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-05",
            },
        ]
    }

    engine = parser.parse_all_from_step_data(
        attendance_rows,
        salary_rows,
        performance_rows,
        supplemental_leave_data=supplemental_leave_data,
        calc_month="2026-05",
    )

    by_id = {employee.employee_id: employee for employee in engine.get_all_employees()}
    assert by_id["zt18731"].sick_hours == 8
    assert by_id["zt18731-1"].sick_hours == 0
    assert by_id["zt18731"].hourly_rate == 28
    assert round(by_id["zt18731"].performance_base, 2) == 1603.14


def test_supplemental_leave_original_id_targets_single_day_shift_row():
    parser = FBUPerformanceParser()
    attendance_rows = [
        {
            "employee_id": "zt002",
            "name": "员工乙",
            "department": "仓库",
            "area": "新泽西区",
            "personnel_status": "正式",
            "job_type": "warehouse",
            "has_night_shift": False,
            "day_shift": {
                "计薪出勤": 80,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 99,
                "病假清算": 0,
                "年假": 99,
                "节假日": 0,
            },
            "night_shift": {
                "计薪出勤": 0,
                "OT1.5": 0,
                "OT2.0": 0,
                "病假": 0,
                "病假清算": 0,
                "年假": 0,
                "节假日": 0,
            },
        },
    ]
    salary_rows = [
        {
            "employee_id": "zt002",
            "hourly_rate": 20,
            "ratio": 0.1,
            "fixed_performance_base": 0,
        }
    ]
    performance_rows = [
        {
            "employee_id": "zt002",
            "score": 95,
            "level": "A",
            "coefficient": 1,
        }
    ]
    supplemental_leave_data = {
        "rows": [
            {
                "employee_id": "zt002",
                "leave_type": "病假",
                "included_hours": 8,
                "include_in_base": True,
                "confirmation_status": "confirmed",
                "allocation_month": "2026-04",
            },
        ]
    }

    engine = parser.parse_all_from_step_data(
        attendance_rows,
        salary_rows,
        performance_rows,
        supplemental_leave_data=supplemental_leave_data,
        calc_month="2026-04",
    )

    employees = engine.get_all_employees()
    assert [employee.employee_id for employee in employees] == ["zt002"]
    assert employees[0].sick_hours == 8
    assert employees[0].annual_hours == 0
    assert employees[0].performance_base == 20 * (80 + 8)
