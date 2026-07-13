from openpyxl import Workbook

import tools.fbu_real_e2e as fbu_real_e2e


def test_read_final_performance_base_uses_warehouse_management_output_sheet(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "1.仓库管理人员"
    sheet.append(["新泽西区绩效考核与奖金核算——仓库管理人员"])
    sheet.append(["员工信息"])
    sheet.append(["绩效周期", "三级部门", "四级部门", "五级部门", "姓名", "员工工号", "职位", "4月绩效基数"])
    sheet.append(["4月", "美洲区", "新泽西区", "", "赵婉妍", "zt12979", "副总经理", 6499.7013])
    sheet.append(["4月", "美洲区", "新泽西区", "新泽西1号仓", "陈海冰", "zt12988", "仓库经理", 5804.3315])
    path = tmp_path / "final-performance.xlsx"
    workbook.save(path)

    assert hasattr(fbu_real_e2e, "read_final_performance_base")
    final_base = fbu_real_e2e.read_final_performance_base(path)

    assert final_base["zt12979"]["offline_performance_base"] == 6499.7
    assert final_base["zt12988"]["offline_performance_base"] == 5804.33
    assert final_base["zt12988"]["name"] == "陈海冰"


def test_read_final_performance_base_ignores_monthly_roster_sheet(tmp_path):
    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "3.区长"
    result_sheet.append(["工号", "姓名", "绩效奖金基数"])
    result_sheet.append(["zt15638", "万其鑫", 3000])
    roster_sheet = workbook.create_sheet("5月花名册6.22")
    roster_sheet.append(["工号", "姓名", "月度绩效奖金基数"])
    roster_sheet.append(["zt15638", "万其鑫", 3000])
    path = tmp_path / "final-performance.xlsx"
    workbook.save(path)

    final_base = fbu_real_e2e.read_final_performance_base(path)

    assert final_base["zt15638"]["offline_performance_base"] == 3000


def test_filter_platform_to_final_target_separates_platform_only_rows():
    platform = {
        "zt001": {"employee_id": "zt001", "platform_performance_base": 100},
        "zt002": {"employee_id": "zt002", "platform_performance_base": 200},
    }
    final_target = {
        "zt001": {"employee_id": "zt001", "offline_performance_base": 100},
    }

    filtered, platform_only = fbu_real_e2e.filter_platform_to_final_target(platform, final_target)

    assert filtered == {"zt001": platform["zt001"]}
    assert platform_only == ["zt002"]


def test_create_work_hour_rule_marker_only_marks_96_rule(tmp_path):
    path = tmp_path / "work-hour-rules.xlsx"

    fbu_real_e2e.create_work_hour_rule_marker(path, "2026-04")

    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0] == ("工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注")
    assert rows[1][0] == "zt12979"
    assert rows[1][2] == "96工时制"
    assert rows[1][3] is None
    assert {row[0] for row in rows[1:]} == {"zt12979", "zt12988", "zt14260", "zt17850"}


def test_platform_result_export_fields_match_final_result_view():
    fields = fbu_real_e2e.platform_result_export_fields()

    assert "hourly_rate" not in fields
    assert "performance_score" in fields
    assert "performance_level" in fields
