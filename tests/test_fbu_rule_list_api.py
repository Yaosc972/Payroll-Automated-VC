import json
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
import pytest

import bonus_platform.app as app_module
from bonus_platform.engine.fbu_performance.runs import FBURosterStore, FBURuleListStore, FBURunManager


pytestmark = pytest.mark.usefixtures("bypass_fbu_access_gate")


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _roster_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roster"
    sheet.append(["姓名", "工号", "二级部门", "三级部门", "四级部门", "划分区域"])
    sheet.append(["陈海冰（花名册）", "zt12988", "FBU", "仓储事业部", "新泽西仓", "新泽西区"])
    sheet.append(["万其鑫（花名册）", "zt15638", "FBU", "区域管理", "美东", "美东区"])
    return _workbook_bytes(workbook)


def _client_with_tmp_store(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_rule_list_store", FBURuleListStore(str(tmp_path)))
    return TestClient(app_module.app)


def test_rule_lists_return_empty_lists_for_new_region(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)

    response = client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload == {
        "region_code": "us_nj",
        "work_hour_employees": [],
        "fixed_base_employees": [],
    }


def test_legacy_global_rule_list_is_flagged_without_copying_it_into_a_region(monkeypatch, tmp_path):
    settings_dir = tmp_path / "_settings"
    settings_dir.mkdir(parents=True)
    (settings_dir / "rule_lists.json").write_text(
        json.dumps({
            "work_hour_employees": [
                {"employee_id": "LEGACY001", "name": "旧名单员工", "active": True},
            ],
            "fixed_base_employees": [],
        }),
        encoding="utf-8",
    )
    client = _client_with_tmp_store(monkeypatch, tmp_path)

    payload = client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
    ).json()

    assert payload["work_hour_employees"] == []
    assert payload["fixed_base_employees"] == []
    assert payload["legacy_unmigrated"] is True


def test_run_manager_quarantines_corrupt_runs_json(tmp_path):
    runs_file = tmp_path / "runs.json"
    runs_file.write_text("{not valid json", encoding="utf-8")

    manager = FBURunManager(str(tmp_path))

    assert manager.list_runs() == []
    assert not runs_file.exists()
    assert len(list(tmp_path.glob("runs.corrupt-*.json"))) == 1

    run = manager.create_run("2026-04")
    assert run.calc_month == "2026-04"
    assert runs_file.exists()


def test_rule_lists_can_be_saved_without_uploading_workbook(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)

    response = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
        json={
            "work_hour_employees": [
                {"employee_id": "zt12988", "name": "陈海冰", "active": True},
            ],
            "fixed_base_employees": [
                {"employee_id": "zt15638", "name": "万其鑫", "fixed_performance_base": 3000, "active": True},
            ],
        },
    )

    assert response.status_code == 200
    saved = json.loads(
        (tmp_path / "_settings" / "rule-lists" / "us_nj.json").read_text(encoding="utf-8")
    )
    assert saved["region_code"] == "us_nj"
    assert saved["work_hour_employees"][0]["employee_id"] == "zt12988"
    assert saved["fixed_base_employees"][0]["fixed_performance_base"] == 3000


def test_rule_lists_are_isolated_by_region_and_empty_region_stays_empty(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)

    new_jersey = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
        json={
            "work_hour_employees": [
                {"employee_id": "NJ001", "name": "新泽西员工", "active": True},
            ],
            "fixed_base_employees": [],
        },
    )
    california = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_ca"},
        json={
            "work_hour_employees": [],
            "fixed_base_employees": [
                {
                    "employee_id": "CA001",
                    "name": "加州员工",
                    "fixed_performance_base": 2800,
                    "active": True,
                },
            ],
        },
    )
    empty_east = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_atlanta"},
        json={"work_hour_employees": [], "fixed_base_employees": []},
    )

    assert new_jersey.status_code == 200
    assert california.status_code == 200
    assert empty_east.status_code == 200
    assert client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
    ).json()["work_hour_employees"][0]["employee_id"] == "NJ001"
    assert client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_ca"},
    ).json()["fixed_base_employees"][0]["employee_id"] == "CA001"
    assert client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_atlanta"},
    ).json() == {
        "region_code": "us_atlanta",
        "work_hour_employees": [],
        "fixed_base_employees": [],
    }


def test_rule_lists_keep_explicitly_saved_empty_lists(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)

    save_response = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
        json={"work_hour_employees": [], "fixed_base_employees": []},
    )

    assert save_response.status_code == 200
    saved_payload = save_response.json()
    assert saved_payload["work_hour_employees"] == []
    assert saved_payload["fixed_base_employees"] == []

    get_response = client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
    )
    assert get_response.status_code == 200
    payload = get_response.json()
    assert payload["work_hour_employees"] == []
    assert payload["fixed_base_employees"] == []


def test_confirm_rule_lists_writes_base_override_data_to_run(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]

    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/rule-lists/confirm",
        json={
            "work_hour_employees": [
                {"employee_id": "zt12988", "name": "陈海冰", "active": True},
            ],
            "fixed_base_employees": [
                {"employee_id": "zt15638", "name": "万其鑫", "fixed_performance_base": 3000, "active": True},
            ],
        },
    )

    assert response.status_code == 200
    detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    rows = detail["base_override_data"]["employees"]
    by_id = {row["employee_id"]: row for row in rows}
    assert by_id["zt12988"]["rule_type"] == "96工时制"
    assert by_id["zt12988"]["fixed_performance_base"] is None
    assert by_id["zt12988"]["source_employee_id"] == "zt12988"
    assert "area" in by_id["zt12988"]
    assert "department" in by_id["zt12988"]
    assert by_id["zt15638"]["rule_type"] == "线下固定基数覆盖"
    assert by_id["zt15638"]["fixed_performance_base"] == 3000
    assert by_id["zt15638"]["source_employee_id"] == "zt15638"
    assert "area" in by_id["zt15638"]
    assert "department" in by_id["zt15638"]
    assert detail["base_override_file"] == "页面维护"


def test_confirm_rule_lists_enriches_name_area_and_department_from_roster(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)
    upload = client.post(
        "/api/fbu-performance/roster",
        data={"calc_month": "2026-04"},
        files={
            "file": (
                "roster.xlsx",
                _roster_workbook_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 200

    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/rule-lists/confirm",
        json={
            "work_hour_employees": [
                {"employee_id": "zt12988", "name": "页面姓名", "active": True},
            ],
            "fixed_base_employees": [
                {"employee_id": "zt15638", "name": "页面姓名", "fixed_performance_base": 3000, "active": True},
            ],
        },
    )

    assert response.status_code == 200
    rows = response.json()["preview"]["employees"]
    by_id = {row["employee_id"]: row for row in rows}
    assert by_id["zt12988"]["name"] == "陈海冰（花名册）"
    assert by_id["zt12988"]["area"] == "新泽西区"
    assert by_id["zt12988"]["department"] == "FBU-仓储事业部-新泽西仓"
    assert "zt15638" not in by_id


def test_confirm_rule_lists_keeps_only_employees_in_activity_region(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roster"
    sheet.append(["姓名", "工号", "二级部门", "三级部门", "四级部门", "划分区域"])
    sheet.append(["新泽西员工", "zt12988", "FBU", "仓储事业部", "新泽西仓", "新泽西区"])
    sheet.append(["加州员工", "zt20001", "FBU", "仓储事业部", "加州仓", "加州区"])
    sheet.append(["区域待确认员工", "zt20002", "FBU", "仓储事业部", "待确认仓", ""])
    upload = client.post(
        "/api/fbu-performance/roster",
        data={"calc_month": "2026-07"},
        files={
            "file": (
                "roster.xlsx",
                _workbook_bytes(workbook),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 200
    run_id = client.post(
        "/api/fbu-performance/runs",
        json={"calc_month": "2026-07", "region_code": "us_ca"},
    ).json()["run_id"]

    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/rule-lists/confirm",
        json={
            "work_hour_employees": [
                {"employee_id": "zt12988", "name": "新泽西员工", "active": True},
                {"employee_id": "zt20001", "name": "加州员工", "active": True},
                {"employee_id": "zt20002", "name": "区域待确认员工", "active": True},
            ],
            "fixed_base_employees": [],
        },
    )

    assert response.status_code == 200
    preview = response.json()["preview"]
    rows = preview["employees"]
    assert [row["employee_id"] for row in rows] == ["zt20001"]
    assert rows[0]["area"] == "加州区"
    assert preview["summary"]["excluded_count"] == 2


def test_confirmed_activity_keeps_snapshot_when_region_list_changes(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roster"
    sheet.append(["姓名", "工号", "二级部门", "三级部门", "四级部门", "划分区域"])
    sheet.append(["新泽西员工", "NJ001", "FBU", "仓储事业部", "新泽西仓", "新泽西区"])
    upload = client.post(
        "/api/fbu-performance/roster",
        data={"calc_month": "2026-07"},
        files={
            "file": (
                "roster.xlsx",
                _workbook_bytes(workbook),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 200
    run_id = client.post(
        "/api/fbu-performance/runs",
        json={"calc_month": "2026-07", "region_code": "us_nj"},
    ).json()["run_id"]

    confirmed = client.post(
        f"/api/fbu-performance/runs/{run_id}/rule-lists/confirm",
        json={
            "work_hour_employees": [],
            "fixed_base_employees": [
                {
                    "employee_id": "NJ001",
                    "name": "新泽西员工",
                    "fixed_performance_base": 2500,
                    "active": True,
                },
            ],
        },
    )
    assert confirmed.status_code == 200

    changed_template = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
        json={
            "work_hour_employees": [],
            "fixed_base_employees": [
                {
                    "employee_id": "NJ001",
                    "name": "新泽西员工",
                    "fixed_performance_base": 3200,
                    "active": True,
                },
            ],
        },
    )
    assert changed_template.status_code == 200

    detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    assert detail["base_override_data"]["employees"][0]["fixed_performance_base"] == 2500
    assert client.get(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
    ).json()["fixed_base_employees"][0]["fixed_performance_base"] == 3200


def test_confirm_rule_lists_excludes_fixed_base_employee_missing_from_active_roster(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roster"
    sheet.append(["姓名", "工号", "二级部门", "三级部门", "四级部门", "划分区域"])
    sheet.append(["新泽西员工", "zt12988", "FBU", "仓储事业部", "新泽西仓", "新泽西区"])
    upload = client.post(
        "/api/fbu-performance/roster",
        data={"calc_month": "2026-07"},
        files={
            "file": (
                "roster.xlsx",
                _workbook_bytes(workbook),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert upload.status_code == 200
    run_id = client.post(
        "/api/fbu-performance/runs",
        json={"calc_month": "2026-07", "region_code": "us_nj"},
    ).json()["run_id"]

    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/rule-lists/confirm",
        json={
            "work_hour_employees": [
                {"employee_id": "zt12988", "name": "新泽西员工", "active": True},
            ],
            "fixed_base_employees": [
                {
                    "employee_id": "zt15638",
                    "name": "万其鑫",
                    "fixed_performance_base": 3000,
                    "active": True,
                },
            ],
        },
    )

    assert response.status_code == 200
    preview = response.json()["preview"]
    rows = preview["employees"]
    by_id = {row["employee_id"]: row for row in rows}
    assert by_id["zt12988"]["area"] == "新泽西区"
    assert "zt15638" not in by_id
    assert preview["summary"]["excluded_count"] == 1


def test_save_rule_lists_rejects_non_numeric_fixed_base(monkeypatch, tmp_path):
    client = _client_with_tmp_store(monkeypatch, tmp_path)

    response = client.post(
        "/api/fbu-performance/rule-lists",
        params={"region_code": "us_nj"},
        json={
            "work_hour_employees": [],
            "fixed_base_employees": [
                {"employee_id": "zt15638", "name": "万其鑫", "fixed_performance_base": "abc", "active": True},
            ],
        },
    )

    assert response.status_code == 400
    assert "固定绩效基数必须是数字" in response.json()["detail"]
