from io import BytesIO
from datetime import datetime

from fastapi.testclient import TestClient
from openpyxl import Workbook

import bonus_platform.app as app_module
from bonus_platform.engine.fbu_performance.engines.base import (
    NINETY_SIX_HOUR_FIXED_BASE_PATH,
)
from bonus_platform.engine.fbu_performance.engines.bonus import BonusCalculator
from bonus_platform.engine.fbu_performance.parser import FBUPerformanceParser
from bonus_platform.engine.fbu_performance.runs import FBURunManager


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _base_override_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "固定基数覆盖"
    sheet.append(["工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注"])
    sheet.append(["zt001", "员工甲", "96工时制", 1920, "2026-04", "启用", "96工时制固定基数"])
    sheet.append(["zt002", "员工乙", "线下固定基数覆盖", 1200, "2026-03", "启用", "非本月不计入"])
    return _workbook_bytes(workbook)


def test_parse_base_overrides_preview_keeps_only_current_enabled_rows(tmp_path):
    workbook_path = tmp_path / "base-overrides.xlsx"
    workbook_path.write_bytes(_base_override_workbook_bytes())

    preview = FBUPerformanceParser().parse_base_overrides_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["total_rows"] == 2
    assert preview["summary"]["active_count"] == 1
    assert preview["summary"]["excluded_count"] == 1
    assert preview["summary"]["active_fixed_base"] == 1920
    assert preview["employees"][0]["employee_id"] == "zt001"
    assert preview["employees"][0]["fixed_performance_base"] == 1920
    assert preview["employees"][0]["calculation_path"] == NINETY_SIX_HOUR_FIXED_BASE_PATH
    assert preview["employees"][1]["include_in_calculation"] is False


def test_parse_base_overrides_allows_blank_fixed_base_for_96_hour_rule(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "固定基数覆盖"
    sheet.append(["工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注"])
    sheet.append(["zt12988", "陈海冰", "96工时制", None, "2026-04", "启用", "自动按96工时制规则计算"])
    workbook_path = tmp_path / "base-overrides.xlsx"
    workbook_path.write_bytes(_workbook_bytes(workbook))

    preview = FBUPerformanceParser().parse_base_overrides_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["total_rows"] == 1
    assert preview["summary"]["active_count"] == 1
    assert preview["summary"]["active_fixed_base"] == 0
    row = preview["employees"][0]
    assert row["employee_id"] == "zt12988"
    assert row["rule_type"] == "96工时制"
    assert row["fixed_performance_base"] == 0
    assert row["calculation_path"] == "96工时制自动基数路径"


def test_parse_base_overrides_normalizes_excel_date_month(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "固定基数覆盖"
    sheet.append(["工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注"])
    sheet.append(["zt12979", "赵婉妍", "96工时制", None, datetime(2026, 4, 1), "启用", "Excel日期格式"])
    workbook_path = tmp_path / "base-overrides.xlsx"
    workbook_path.write_bytes(_workbook_bytes(workbook))

    preview = FBUPerformanceParser().parse_base_overrides_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["active_count"] == 1
    row = preview["employees"][0]
    assert row["allocation_month"] == "2026-04"
    assert row["include_in_calculation"] is True


def test_parse_base_overrides_normalizes_compact_month_text(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "固定基数覆盖"
    sheet.append(["工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注"])
    sheet.append(["zt12988", "陈海冰", "96工时制", None, "202604", "启用", "紧凑月份格式"])
    workbook_path = tmp_path / "base-overrides.xlsx"
    workbook_path.write_bytes(_workbook_bytes(workbook))

    preview = FBUPerformanceParser().parse_base_overrides_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["active_count"] == 1
    row = preview["employees"][0]
    assert row["allocation_month"] == "2026-04"
    assert row["include_in_calculation"] is True


def test_parse_base_overrides_from_offline_summary_diff_notes(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "基数汇总"
    sheet.append([None] * 18)
    sheet.append([
        "工号",
        "中文名字",
        "求和项:合计所得",
        None,
        "工号",
        "中文名字",
        "求和项:求和项:合计所得",
        "考勤日报",
        "验算",
        "差异备注",
    ])
    sheet.append(["zt001", "员工甲", 1888, None, "zt001", "员工甲", 1888, 2100, -212, "96工时制"])
    sheet.append(["zt002", "员工乙", 1200, None, "zt002", "员工乙", 1200, 1180, 20, "夜班"])
    workbook_path = tmp_path / "offline-summary.xlsx"
    workbook_path.write_bytes(_workbook_bytes(workbook))

    preview = FBUPerformanceParser().parse_base_overrides_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["total_rows"] == 1
    assert preview["summary"]["active_count"] == 1
    assert preview["employees"][0]["employee_id"] == "zt001"
    assert preview["employees"][0]["rule_type"] == "96工时制"
    assert preview["employees"][0]["fixed_performance_base"] == 1888
    assert preview["employees"][0]["calculation_path"] == NINETY_SIX_HOUR_FIXED_BASE_PATH
    assert "差异备注: 96工时制" in preview["employees"][0]["note"]


def test_parse_base_overrides_extracts_96_special_hours_from_special_summary(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "特殊工时汇总"
    sheet.append([
        "姓名",
        "员工工号",
        "职位",
        "2026/04/01",
        "4.1-4.11",
        "4.12-4.25",
        "4.26-4.30",
        "节假日",
        "合计时长",
        "",
        "时薪",
        "时薪*时长",
    ])
    sheet.append(["陈海冰", "zt12988", "仓库经理", 8, 0, 0, 0, 0, 201.19, "", 28.8481570512821, 5803.96])
    workbook_path = tmp_path / "offline-summary.xlsx"
    workbook_path.write_bytes(_workbook_bytes(workbook))

    preview = FBUPerformanceParser().parse_base_overrides_preview(str(workbook_path), "2026-04")

    assert preview["summary"]["active_count"] == 1
    row = preview["employees"][0]
    assert row["employee_id"] == "zt12988"
    assert row["rule_type"] == "96工时制"
    assert row["fixed_performance_base"] == 0
    assert row["special_total_hours"] == 201.19
    assert row["special_hourly_rate"] == 28.8481570512821
    assert row["calculation_path"] == "96工时制自动基数路径"


def test_base_override_replaces_performance_base_but_keeps_normal_bonus_formula(tmp_path):
    workbook_path = tmp_path / "base-overrides.xlsx"
    workbook_path.write_bytes(_base_override_workbook_bytes())
    parser = FBUPerformanceParser()
    preview = parser.parse_base_overrides_preview(str(workbook_path), "2026-04")

    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt001",
                "name": "员工甲",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 80, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
            }
        ],
        salary_data=[
            {"employee_id": "zt001", "name": "员工甲", "hourly_rate": 10, "ratio": 0.1},
        ],
        performance_data=[
            {"employee_id": "zt001", "name": "员工甲", "score": None, "level": None, "coefficient": 1},
        ],
        calc_month="2026-04",
        base_override_data=preview,
    )

    employee = engine.get_employee("zt001")
    assert employee.base_salary == 800
    assert employee.performance_base == 1920
    assert employee.performance_bonus == 192
    assert employee.base_override_reason == "96工时制固定基数"
    assert employee.calculation_path == NINETY_SIX_HOUR_FIXED_BASE_PATH
    assert "固定基数覆盖" in employee.exceptions[-1]


def test_blank_96_hour_rule_marker_applies_auto_base_without_fixed_amount():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt12988",
                "name": "陈海冰",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 90, "OT1.5": 10, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 8, "节假日": 4},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
            }
        ],
        salary_data=[
            {"employee_id": "zt12988", "name": "陈海冰", "hourly_rate": 20, "ratio": 0.1},
        ],
        performance_data=[
            {"employee_id": "zt12988", "name": "陈海冰", "score": 95, "level": None, "coefficient": 1},
        ],
        calc_month="2026-04",
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt12988",
                    "rule_type": "96工时制",
                    "fixed_performance_base": 0,
                    "allocation_month": "2026-04",
                    "include_in_calculation": True,
                    "note": "自动按96工时制规则计算",
                }
            ]
        },
    )

    employee = engine.get_employee("zt12988")
    assert employee.work_hour_rule == "96工时制"
    assert employee.base_override_amount in (None, 0)
    assert employee.performance_base == 2240
    assert employee.calculation_path == "96工时制自动基数路径"


def test_96_hour_rule_combines_day_and_night_daily_rows_before_calculation():
    parser = FBUPerformanceParser()
    employees = parser.build_employees(
        attendance_data={
            "zt12988": {
                "白班": {"计薪出勤": 8, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "夜班": {"计薪出勤": 8, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": True,
                "白班_daily_rows": [{"date": "2026-05-01", "shift_type": "白班", "base_hours": 8, "work_hours": 8}],
                "夜班_daily_rows": [{"date": "2026-05-02", "shift_type": "夜班", "base_hours": 8, "work_hours": 8}],
                "daily_rows": [
                    {"date": "2026-05-01", "shift_type": "白班", "base_hours": 8, "work_hours": 8},
                    {"date": "2026-05-02", "shift_type": "夜班", "base_hours": 8, "work_hours": 8},
                ],
            }
        },
        salary_data={"zt12988": {"hourly_rate": 20, "ratio": 0.1}},
        performance_data={"zt12988": {"score": 95, "level": "符合预期", "coefficient": 1}},
        employee_info={"zt12988": {"name": "陈海冰", "job_type": "warehouse"}},
        base_override_data={
            "employees": [{
                "employee_id": "zt12988",
                "rule_type": "96工时制",
                "include_in_calculation": True,
                "allocation_month": "2026-05",
            }]
        },
        calc_month="2026-05",
    )

    rule_employee = next(employee for employee in employees if employee.work_hour_rule == "96工时制")
    BonusCalculator.calculate(rule_employee)

    assert rule_employee.work_hour_rule_special_total_hours == 16
    assert rule_employee.work_hour_rule_rounded_hourly_rate == 20
    assert rule_employee.performance_base == 320


def test_blank_96_hour_rule_marker_builds_special_periods_from_daily_attendance_rows():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt12988",
                "name": "陈海冰",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 125, "OT1.5": 10, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 16, "节假日": 6},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
                "attendance_daily_rows": [
                    {
                        "date": "2026-04-12",
                        "shift_type": "白班",
                        "base_hours": 90,
                        "ot15_hours": 10,
                        "annual_hours": 8,
                        "holiday_hours": 4,
                    },
                    {
                        "date": "2026-04-26",
                        "shift_type": "白班",
                        "base_hours": 35,
                        "annual_hours": 8,
                        "holiday_hours": 2,
                    },
                ],
            }
        ],
        salary_data=[
            {"employee_id": "zt12988", "name": "陈海冰", "hourly_rate": 20, "ratio": 0.1},
        ],
        performance_data=[
            {"employee_id": "zt12988", "name": "陈海冰", "score": 95, "level": None, "coefficient": 1},
        ],
        calc_month="2026-04",
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt12988",
                    "rule_type": "96工时制",
                    "fixed_performance_base": 0,
                    "allocation_month": "2026-04",
                    "include_in_calculation": True,
                    "note": "自动按96工时制规则计算",
                }
            ]
        },
    )

    employee = engine.get_employee("zt12988")
    assert employee.work_hour_rule_special_total_hours == 145
    assert employee.performance_base == 2900
    assert [period["period"] for period in employee.work_hour_rule_periods] == [
        "2026-04 / 4.12-4.25",
        "2026-04 / 4.26-4.30",
    ]
    assert [period["included_hours"] for period in employee.work_hour_rule_periods] == [100, 45]


def test_blank_96_hour_rule_marker_calculates_special_hours_from_raw_attendance_rows():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt12988",
                "name": "陈海冰",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 169.19, "OT1.5": 20.1, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 16, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
                "attendance_daily_rows": [
                    {"date": "2026-03-29", "shift_type": "白班", "base_hours": 24},
                    {"date": "2026-03-30", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-03-31", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-01", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-02", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-03", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-04", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-06", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-07", "shift_type": "白班", "annual_hours": 8},
                    {"date": "2026-04-08", "shift_type": "白班", "annual_hours": 8},
                    {"date": "2026-04-09", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-10", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-11", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-24", "shift_type": "白班", "base_hours": 80, "ot15_hours": 20.1},
                    {"date": "2026-04-27", "shift_type": "白班", "base_hours": 33.19},
                ],
            }
        ],
        salary_data=[
            {"employee_id": "zt12988", "name": "陈海冰", "hourly_rate": 28.8481570512821, "ratio": 0.1},
        ],
        performance_data=[
            {"employee_id": "zt12988", "name": "陈海冰", "score": 95, "level": None, "coefficient": 1},
        ],
        calc_month="2026-04",
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt12988",
                    "rule_type": "96工时制",
                    "fixed_performance_base": 0,
                    "allocation_month": "2026-04",
                    "include_in_calculation": True,
                    "note": "自动按96工时制规则计算",
                }
            ]
        },
    )

    employee = engine.get_employee("zt12988")
    assert employee.work_hour_rule_special_total_hours == 201.19
    assert employee.work_hour_rule_rounded_hourly_rate == 28.85
    assert employee.performance_base == 5804.3315
    assert employee.work_hour_rule_periods[0]["regular_hours"] == 56
    assert employee.work_hour_rule_periods[0]["leave_hours"] == 16


def test_96_hour_rule_cross_month_first_segment_uses_full_biweekly_cap():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt12979",
                "name": "赵婉妍",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 97.75, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 8},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
                "attendance_daily_rows": [
                    {"date": "2026-04-26", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-27", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-28", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-29", "shift_type": "白班", "base_hours": 8},
                    {"date": "2026-04-30", "shift_type": "白班", "base_hours": 0.67},
                    {"date": "2026-05-01", "shift_type": "白班", "base_hours": 65.08, "holiday_hours": 8},
                    {"date": "2026-05-10", "shift_type": "白班", "base_hours": 96},
                    {"date": "2026-05-24", "shift_type": "白班", "base_hours": 40.57},
                ],
            }
        ],
        salary_data=[
            {"employee_id": "zt12979", "name": "赵婉妍", "hourly_rate": 32.3894230769231, "ratio": 0.13},
        ],
        performance_data=[
            {"employee_id": "zt12979", "name": "赵婉妍", "score": 95, "level": None, "coefficient": 1},
        ],
        calc_month="2026-05",
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt12979",
                    "rule_type": "96工时制",
                    "fixed_performance_base": 0,
                    "allocation_month": "2026-05",
                    "include_in_calculation": True,
                    "note": "自动按96工时制规则计算",
                }
            ]
        },
    )

    employee = engine.get_employee("zt12979")
    assert employee.work_hour_rule_periods[0]["mode"] == "跨月首段REG-双周96封顶"
    assert employee.work_hour_rule_periods[0]["regular_hours"] == 63.33
    assert employee.work_hour_rule_special_total_hours == 207.9
    assert employee.performance_base == 6733.881


def test_96_hour_rule_full_period_caps_regular_after_leave_hours():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt17850",
                "name": "韩勇",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 94.67, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 8, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
                "attendance_daily_rows": [
                    {"date": "2026-05-10", "shift_type": "白班", "base_hours": 94.67},
                    {"date": "2026-05-11", "shift_type": "白班", "annual_hours": 8, "scheduled_hours": 8},
                ],
            }
        ],
        salary_data=[
            {"employee_id": "zt17850", "name": "韩勇", "hourly_rate": 27.1654647435897, "ratio": 0.4285714286},
        ],
        performance_data=[
            {"employee_id": "zt17850", "name": "韩勇", "score": 113.72, "level": None, "coefficient": 1.37},
        ],
        calc_month="2026-05",
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt17850",
                    "rule_type": "96工时制",
                    "fixed_performance_base": 0,
                    "allocation_month": "2026-05",
                    "include_in_calculation": True,
                    "note": "自动按96工时制规则计算",
                }
            ]
        },
    )

    employee = engine.get_employee("zt17850")
    assert employee.work_hour_rule_periods[0]["regular_hours"] == 88
    assert employee.work_hour_rule_periods[0]["leave_hours"] == 8
    assert employee.work_hour_rule_periods[0]["included_hours"] == 96


def test_96_hour_rule_does_not_double_count_full_day_leave_with_raw_work_hours():
    parser = FBUPerformanceParser()
    engine = parser.parse_all_from_step_data(
        attendance_data=[
            {
                "employee_id": "zt12979",
                "name": "赵婉妍",
                "department": "FBU仓储事业部",
                "area": "新泽西区",
                "job_type": "warehouse",
                "day_shift": {"计薪出勤": 8, "OT1.5": 0, "OT2.0": 0, "病假": 8, "病假清算": 0, "年假": 0, "节假日": 0},
                "night_shift": {"计薪出勤": 0, "OT1.5": 0, "OT2.0": 0, "病假": 0, "病假清算": 0, "年假": 0, "节假日": 0},
                "has_night_shift": False,
                "attendance_daily_rows": [
                    {
                        "date": "2026-04-12",
                        "shift_type": "白班",
                        "base_hours": 8,
                        "sick_hours": 8,
                        "scheduled_hours": 8,
                        "work_hours": 8,
                    },
                ],
            }
        ],
        salary_data=[
            {"employee_id": "zt12979", "name": "赵婉妍", "hourly_rate": 20, "ratio": 0.1},
        ],
        performance_data=[
            {"employee_id": "zt12979", "name": "赵婉妍", "score": 95, "level": None, "coefficient": 1},
        ],
        calc_month="2026-04",
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt12979",
                    "rule_type": "96工时制",
                    "fixed_performance_base": 0,
                    "allocation_month": "2026-04",
                    "include_in_calculation": True,
                    "note": "自动按96工时制规则计算",
                }
            ]
        },
    )

    employee = engine.get_employee("zt12979")
    assert employee.work_hour_rule_special_total_hours == 8
    assert employee.performance_base == 160
    assert employee.work_hour_rule_periods[0]["regular_hours"] == 0
    assert employee.work_hour_rule_periods[0]["leave_hours"] == 8


def test_fbu_base_override_upload_is_saved_as_optional_run_data(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    client = TestClient(app_module.app)
    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    response = client.post(
        "/api/fbu-performance/import-base-overrides",
        data={"run_id": run_id},
        files={"file": ("base-overrides.xlsx", _base_override_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["preview"]["summary"]["active_count"] == 1
    assert payload["result_file"]["type"] == "base_overrides"

    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    assert run_detail["base_override_file"] == "base-overrides.xlsx"
    assert run_detail["base_override_data"]["employees"][0]["fixed_performance_base"] == 1920
