"""高温补贴规则、数据接入和导出回归。"""
from datetime import date
from io import BytesIO
from pathlib import Path
import time

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from fastapi.testclient import TestClient

from bonus_platform.app import app
from bonus_platform.engine.domestic_labor.engines.gaowen_butie import GaoWenBuTieEngine
from bonus_platform.engine.domestic_labor.exporter import ExcelExporter
from bonus_platform.engine.domestic_labor.parser import MultiFilePayrollDataLoader


pytestmark = pytest.mark.usefixtures("bypass_domestic_labor_access_gate")


def _employee(
    employee_id="OWHN001",
    name="张三",
    work_area="东莞",
    position="操作员",
    **extra,
):
    return {
        "工号": employee_id,
        "姓名": name,
        "工作地区": work_area,
        "岗位名称": position,
        "一级部门名称": "寮步区",
        "二级部门名称": "中国操作部",
        "三级部门名称": "华南1号枢纽2",
        **extra,
    }


def _day(day, *, schedule="19:00-28:00;", regular=8, overtime=0, actual=8, **extra):
    return {
        "出勤日期": day,
        "班次时间段": schedule,
        "正班时数": regular,
        "刷卡加班": overtime,
        "实际上班时数": actual,
        **extra,
    }


def _temperature(day, shift, temperature, site="华南1号枢纽-寮步仓"):
    return {
        "班次日期": day,
        "测温班次": shift,
        "测温网点": site,
        "测温温度": temperature,
    }


def test_same_site_date_and_shift_must_reach_33_degrees_before_payment():
    measurements = [
        _temperature(date(2026, 7, 1), "白班", 34.2),
        _temperature(date(2026, 7, 1), "夜班", 32.9),
        _temperature(date(2026, 7, 2), "夜班", 33.0),
    ]
    attendance = [
        _day(date(2026, 7, 1), regular=8),
        _day(date(2026, 7, 2), regular=4),
    ]

    result = GaoWenBuTieEngine(measurements).calculate(_employee(), attendance)

    assert result.amount == 6.90
    assert [row["reason_code"] for row in result.details["daily_results"]] == [
        "temperature_below_33",
        "calculated",
    ]
    assert result.details["daily_results"][1]["temperature"] == 33.0
    assert result.details["daily_results"][1]["attendance_hours"] == 4


def test_hours_use_larger_of_regular_and_swipe_overtime_with_daily_and_monthly_caps():
    days = [date(2026, 7, day) for day in range(1, 32)]
    measurements = [_temperature(day, "夜班", 35) for day in days]
    attendance = [
        _day(day, regular=4, overtime=10, actual=10)
        for day in days
    ]

    result = GaoWenBuTieEngine(measurements).calculate(_employee(), attendance)

    assert result.amount == 300
    assert result.details["高温出勤天数"] == 31
    assert result.details["月度封顶前金额"] == 427.8
    assert all(row["amount"] == 13.8 for row in result.details["daily_results"])


def test_explicit_zero_actual_attendance_does_not_pay_from_half_hour_swipe_overtime():
    measurements = [_temperature(date(2026, 7, 11), "夜班", 35)]
    attendance = [
        _day(date(2026, 7, 11), regular=0, overtime=0.5, actual=0),
    ]

    result = GaoWenBuTieEngine(measurements).calculate(_employee(), attendance)

    assert result.amount == 0
    assert result.details["daily_results"][0]["reason_code"] == "actual_attendance_zero"


def test_source_regular_hours_are_not_discarded_by_zero_actual_hours_cache():
    """源规则取正班/刷卡较大值；实际上班时数的零缓存不能抹掉8小时正班。"""
    day = date(2026, 7, 12)
    result = GaoWenBuTieEngine([_temperature(day, "夜班", 35)]).calculate(
        _employee(),
        [_day(day, regular=8, overtime=0, actual=0)],
    )

    assert result.amount == 13.8
    assert result.details["daily_results"][0]["attendance_hours"] == 8


def test_schedule_start_overrides_misleading_early_morning_continuous_shift_name():
    """“早上连班”仍是19点开班的夜班，不能因名称含“早”错配白班温度。"""
    day = date(2026, 7, 13)
    result = GaoWenBuTieEngine([
        _temperature(day, "白班", 31),
        _temperature(day, "夜班", 35),
    ]).calculate(
        _employee(),
        [_day(
            day,
            schedule="19:00-28:00;",
            班次类别名称="寮步班次",
            班次名称="寮步19:00-04:00(早上连班)",
        )],
    )

    assert result.amount == 13.8
    assert result.details["daily_results"][0]["shift"] == "夜班"


def test_dongguan_site_resolution_distinguishes_b2b_qingxi_and_liaobu_entities():
    day = date(2026, 7, 3)
    measurements = [
        _temperature(day, "夜班", 34, "华南1号枢纽-寮步仓"),
        _temperature(day, "夜班", 31, "华南B2B枢纽-清溪仓"),
    ]
    engine = GaoWenBuTieEngine(measurements)
    attendance = [_day(day)]

    liaobu = engine.calculate(
        _employee(一级部门名称="华南B2B枢纽", 二级部门名称="华南B2B枢纽"),
        attendance,
    )
    qingxi = engine.calculate(
        _employee(一级部门名称="华南B2B枢纽组", 二级部门名称="华南B2B枢纽组"),
        attendance,
    )

    assert liaobu.amount == 13.8
    assert liaobu.details["测温网点"] == "华南1号枢纽-寮步仓"
    assert qingxi.amount == 0
    assert qingxi.details["测温网点"] == "华南B2B枢纽-清溪仓"


def test_dongguan_safety_subgroup_uses_its_physical_warehouse_temperature():
    day = date(2026, 7, 4)
    measurements = [
        _temperature(day, "白班", 34, "中国仓组-东莞茶山仓"),
        _temperature(day, "白班", 31, "华南1号枢纽-寮步仓"),
    ]
    employee = _employee(
        一级部门名称="操作风控部",
        二级部门名称="操作支持组",
        五级部门名称="中国仓安全组",
    )

    result = GaoWenBuTieEngine(measurements).calculate(
        employee,
        [_day(day, schedule="09:00-18:00;")],
    )

    assert result.amount == 13.8
    assert result.details["测温网点"] == "中国仓组-东莞茶山仓"


def test_dongguan_shared_support_and_risk_groups_use_liaobu_measurement_site():
    day = date(2026, 7, 4)
    engine = GaoWenBuTieEngine([_temperature(day, "白班", 34)])

    for group in ("操作支持组", "操作风控组", "后勤保障组"):
        employee = _employee(
            一级部门名称="中国操作部",
            二级部门名称="中国操作部",
            三级部门名称=group,
        )
        result = engine.calculate(employee, [_day(day, schedule="09:00-18:00;")])
        assert result.details["测温网点"] == "华南1号枢纽-寮步仓"


def test_named_exclusions_and_jinjiang_position_scope_do_not_depend_on_grade_or_collar():
    day = date(2026, 7, 5)
    dongguan_measurement = [_temperature(day, "夜班", 35)]
    excluded = GaoWenBuTieEngine(dongguan_measurement).calculate(
        _employee(name="王雯雯", 职级="P9-9", 领色="蓝领"),
        [_day(day)],
    )

    jinjiang_measurement = [
        _temperature(day, "夜班", 35, "东南枢纽-晋江仓"),
    ]
    engine = GaoWenBuTieEngine(jinjiang_measurement)
    eligible = engine.calculate(
        _employee(work_area="晋江", position="门禁员", 一级部门名称="东南枢纽"),
        [_day(day, regular=6, actual=6)],
    )
    ineligible = engine.calculate(
        _employee(work_area="晋江", position="物流专员", 一级部门名称="东南枢纽"),
        [_day(day, regular=8, actual=8)],
    )

    assert excluded.amount == 0
    assert excluded.details["资格判断"] == "固定排除名单"
    assert eligible.amount == 9
    assert eligible.details["小时单价"] == 1.5
    assert ineligible.amount == 0
    assert ineligible.details["资格判断"] == "岗位不在晋江适用范围"


def test_may_attendance_is_outside_the_june_to_october_high_temperature_season():
    day = date(2026, 5, 31)
    result = GaoWenBuTieEngine([_temperature(day, "夜班", 36)]).calculate(
        _employee(),
        [_day(day)],
    )

    assert result.amount == 0
    assert result.details["daily_results"][0]["reason_code"] == "outside_high_temperature_season"


def test_missing_measurement_file_does_not_block_but_marks_result_for_confirmation():
    result = GaoWenBuTieEngine([]).calculate(
        _employee(),
        [_day(date(2026, 7, 1))],
    )

    assert result.amount == 0
    assert any("未提供高温测温登记" in warning for warning in result.warnings)
    assert result.details["exceptions"][0]["code"] == "HIGH_TEMPERATURE_MEASUREMENTS_MISSING"
    assert result.details["daily_results"][0]["reason_code"] == "no_matching_temperature"


def _workbook(sheet_name, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_multifile_loader_recognizes_temperature_registration_as_a_separate_source(tmp_path):
    monthly = _workbook(
        "202607月考勤",
        ["考勤月份", "工号", "姓名", "工作地区", "岗位名称", "排班天数"],
        [["202607", "OWHN001", "张三", "东莞", "操作员", 23]],
    )
    daily = _workbook(
        "202607日考勤",
        ["出勤日期", "工号", "姓名", "班次时间段", "正班时数", "刷卡加班", "实际上班时数"],
        [[date(2026, 7, 1), "OWHN001", "张三", "19:00-28:00;", 8, 0, 8]],
    )
    temperature = _workbook(
        "中国操作部测温登记",
        ["温度测温点", "班次日期", "测温班次", "测温网点", "测温温度"],
        [["寮步", date(2026, 7, 1), "夜班", "华南1号枢纽-寮步仓", 34.5]],
    )
    paths = []
    for name, payload in (("月考勤.xlsx", monthly), ("日考勤.xlsx", daily), ("测温.xlsx", temperature)):
        path = tmp_path / name
        path.write_bytes(payload)
        paths.append(str(path))

    with MultiFilePayrollDataLoader(paths) as loader:
        summary = loader.validate_inputs(["gaowen_butie"], "202607")

        assert loader.temperature.row_count == 1
        assert loader.temperature.rows[0]["测温温度"] == 34.5
        assert summary["temperature_rows"] == 1
        assert "temperature" in summary["present_types"]


def test_loader_ignores_embedded_ai_reconciliation_sheet_when_real_monthly_sheet_exists(tmp_path):
    """生产核对工作簿会同时带结果页，结果页不能再次并入月考勤。"""
    workbook = Workbook()
    comparison = workbook.active
    comparison.title = "AI&线下核算比对版"
    comparison.append(["考勤月份", "工号", "姓名", "入职日期", "高温补贴"])
    comparison.append(["202607", "OWHN001", "张三", date(2025, 1, 1), 13.8])
    monthly = workbook.create_sheet("202607月考勤")
    monthly.append(["考勤月份", "工号", "姓名", "工作地区", "岗位名称", "排班天数"])
    monthly.append(["202607", "OWHN001", "张三", "东莞", "操作员", 23])
    daily = workbook.create_sheet("202607日考勤")
    daily.append(["出勤日期", "工号", "姓名", "班次时间段", "正班时数", "刷卡加班"])
    daily.append([date(2026, 7, 1), "OWHN001", "张三", "19:00-28:00;", 8, 0])
    path = tmp_path / "生产核对工作簿.xlsx"
    workbook.save(path)

    with MultiFilePayrollDataLoader([str(path)]) as loader:
        summary = loader.validate_inputs(["gaowen_butie"], "202607")

    assert summary["monthly_rows"] == 1
    assert summary["sources"][0]["sheets"] == [
        {"sheet": "202607月考勤", "type": "monthly", "row_count": 1},
        {"sheet": "202607日考勤", "type": "daily", "row_count": 1},
    ]


def test_parser_trims_formatted_empty_tail_from_temperature_sheet(tmp_path):
    """WPS表格可能把格式扩展到百万行，解析时不应生成百万条空记录。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "高温补贴"
    sheet.append(["班次日期", "测温班次", "测温网点", "测温温度"])
    sheet.append([date(2026, 7, 1), "白班", "华东枢纽-嘉善仓", 34])
    sheet.cell(row=2005, column=4).fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    path = tmp_path / "带空白格式尾部的测温表.xlsx"
    workbook.save(path)

    with MultiFilePayrollDataLoader([str(path)]) as loader:
        temperature = loader.temperature

    assert temperature.row_count == 1
    assert temperature.rows[0]["测温温度"] == 34


def test_high_temperature_export_has_business_readable_summary_and_daily_sheets(tmp_path):
    result = GaoWenBuTieEngine([
        _temperature(date(2026, 7, 1), "夜班", 34),
    ]).calculate(_employee(), [_day(date(2026, 7, 1))])
    row = {
        "employee_id": "OWHN001",
        "employee_name": "张三",
        "department": "中国操作部",
        "gaowen_butie": result.amount,
        "total": result.amount,
        "warnings": "",
        "exceptions": [],
        "subject_details": {
            "gaowen_butie": {
                "amount": result.amount,
                "details": result.details,
                "exceptions": result.details["exceptions"],
                "audit_explanation": result.details["audit_explanation"],
            }
        },
    }
    path = tmp_path / "高温补贴.xlsx"

    ExcelExporter(str(path)).export([row], "202607")

    workbook = load_workbook(path, data_only=True)
    assert workbook.sheetnames[:2] == ["核算汇总", "每日明细"]
    assert [cell.value for cell in workbook["核算汇总"][1]][:8] == [
        "工号", "姓名", "工作地区", "部门", "岗位", "测温网点", "高温出勤天数", "应发高温补贴",
    ]
    daily_headers = [cell.value for cell in workbook["每日明细"][1]]
    assert "最高测温（℃）" in daily_headers
    assert "实际高温出勤时长（小时）" in daily_headers
    assert "当日高温补贴" in daily_headers
    workbook.close()


def test_api_runs_high_temperature_allowance_from_monthly_daily_and_temperature_files():
    monthly = _workbook(
        "202607月考勤",
        [
            "考勤月份", "工号", "姓名", "工作地区", "岗位名称", "排班天数",
            "一级部门名称", "二级部门名称", "三级部门名称",
        ],
        [["202607", "OWHN001", "张三", "东莞", "操作员", 23, "寮步区", "中国操作部", "华南1号枢纽2"]],
    )
    daily = _workbook(
        "202607日考勤",
        ["出勤日期", "工号", "姓名", "班次时间段", "正班时数", "刷卡加班", "实际上班时数"],
        [[date(2026, 7, 1), "OWHN001", "张三", "19:00-28:00;", 8, 0, 8]],
    )
    temperature = _workbook(
        "中国操作部测温登记",
        ["温度测温点", "班次日期", "测温班次", "测温网点", "测温温度"],
        [["寮步", date(2026, 7, 1), "夜班", "华南1号枢纽-寮步仓", 34.5]],
    )
    client = TestClient(app)

    response = client.post(
        "/api/domestic-labor/runs",
        files=[
            ("files", ("月考勤.xlsx", monthly, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("日考勤.xlsx", daily, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("测温登记.xlsx", temperature, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
        data={"engines": "gaowen_butie", "attendance_month": "202607"},
    )

    assert response.status_code == 200
    assert response.json()["input_summary"]["temperature_rows"] == 1
    run_id = response.json()["run_id"]
    for _ in range(50):
        metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
        if metadata["status"] in {"已完成", "失败"}:
            break
        time.sleep(0.05)
    assert metadata["status"] == "已完成"
    assert metadata["results"][0]["gaowen_butie"] == 13.8
    assert metadata["summary"]["total_gaowen_butie"] == 13.8


def test_rule_package_and_subject_card_publish_high_temperature_as_validating():
    package = TestClient(app).get("/api/domestic-labor/rule-package").json()
    high_temperature = next(subject for subject in package["subjects"] if subject["id"] == "gaowen_butie")
    payload = str(high_temperature)

    assert package["version"] == "1.4.0"
    assert high_temperature["status"] == "验证中"
    assert "同测温网点、同出勤日期、同白/夜班" in payload
    assert "MAX(正班时数,刷卡加班)" in payload
    assert "浙江室内" in payload
    assert "无测温区域" in payload
    assert "99.64%" in payload

    html = Path("bonus_platform/static/domestic-labor.html").read_text(encoding="utf-8")
    js = Path("bonus_platform/static/domestic-labor.js").read_text(encoding="utf-8")
    assert 'class="dl-subject-card primary validating" data-subject-entry="gaowen_butie"' in html
    assert "高温补贴核算" in html
    assert "验证中" in html
    assert "gaowen_butie" in js
    assert "renderGaowenResults" in js
    assert "renderHighTemperatureExplanation" in js
    assert "同仓同日同班次最高温" in js
