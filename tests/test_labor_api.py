from io import BytesIO
import asyncio
import json
import logging
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import bonus_platform.app as app_module
import bonus_platform.engine.labor.runs as labor_runs
import bonus_platform.engine.labor.structure as labor_structure
from bonus_platform.app import app
from bonus_platform.engine.labor.models import LaborLineItem


def test_runtime_suppresses_full_http_client_request_urls():
    assert logging.getLogger("httpx").level >= logging.WARNING
    assert logging.getLogger("httpcore").level >= logging.WARNING


def _excel_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种"])
    sheet.append(["WUS042586", "Rosa Alvarez Minchaca", 31.19, 701.90, "USD"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_bytes_with_warehouse() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS000001", "Alice Worker", 8, 100, "USD", "1号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_bytes_with_two_warehouses() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS000001", "Alice Worker", 8, 100, "USD", "1号仓"])
    sheet.append(["WUS000002", "Bob Worker", 10, 200, "USD", "2号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _excel_bytes_for_structure_fallback() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    for index, name in enumerate(["A One", "B Two", "C Three", "D Four", "E Five"], start=1):
        sheet.append([f"DE{index:06d}", name, 8, 20, "EUR", "15号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _prepare_labor_orchestration_run(
    monkeypatch,
    *,
    pdf_names: list[str],
    workbook_bytes: bytes,
) -> tuple[TestClient, dict]:
    monkeypatch.setattr(app_module, "_labor_cost_summaries", lambda *args, **kwargs: [])
    monkeypatch.setattr(app_module, "audit_ai_page_cache_candidates", lambda *args, **kwargs: {})
    monkeypatch.setattr(
        app_module,
        "build_ai_cache_reconciliation_preview",
        lambda *args, **kwargs: {"fileQuality": []},
    )
    monkeypatch.setattr(
        app_module,
        "build_reocr_candidate_plan",
        lambda *args, **kwargs: {"summary": {}, "tasks": [], "reviewableCandidates": []},
    )
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {
                "fileCount": len(paths),
                "textReadableFileCount": len(paths),
                "imageOnlyFileCount": 0,
                "textReadablePageCount": len(paths),
                "emptyTextPageCount": 0,
                "imageOnlyPdfFiles": [],
            },
            "files": [],
        },
    )
    monkeypatch.setattr(
        app_module,
        "calculate_extraction_quality",
        lambda *args, **kwargs: {
            "level": "ok",
            "message": "",
            "issues": [],
            "retryAttempted": False,
            "retryApplied": False,
        },
    )

    def fake_report(path, *args, **kwargs):
        Path(path).write_bytes(b"report")

    def fake_business_report(path, *args, **kwargs):
        Path(path).write_text("<html></html>", encoding="utf-8")

    monkeypatch.setattr(app_module, "build_labor_report", fake_report)
    monkeypatch.setattr(app_module, "build_labor_business_html_report", fake_business_report)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Task 4 Supplier",
            "period_start": "2026-06-01",
            "period_end": "2026-06-07",
            "currency": "USD",
        },
    ).json()
    files = [
        ("pdf_files", (name, b"%PDF-1.4\n", "application/pdf"))
        for name in pdf_names
    ]
    files.append(
        (
            "workbook_files",
            (
                "bill.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )
    )
    response = client.post(f"/api/labor/runs/{run['id']}/files", files=files)
    assert response.status_code == 200
    response = client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
                "warehouse": "物理仓",
            },
        },
    )
    assert response.status_code == 200
    return client, run


def _unresolved_quick_totals(paths) -> list[dict]:
    return [
        {
            "source_file": Path(path).name,
            "total_amount": 0.0,
            "warehouse_id": "",
            "authoritative": False,
            "evidence_status": "needs_review",
            "page_evidence": [{"page": 1, "role": "invoice_primary"}],
        }
        for path in paths
    ]


def test_labor_legacy_upload_can_append_files_in_separate_rounds(monkeypatch, tmp_path):
    monkeypatch.setenv("SIGMA_LABOR_AUTH_REQUIRED", "0")
    monkeypatch.setenv("SIGMA_LABOR_P1_REQUIRED", "0")
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", tmp_path / "runs")
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", tmp_path / "runs")
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Separate upload rounds",
            "period_start": "2026-07-20",
            "period_end": "2026-07-26",
            "currency": "USD",
        },
    ).json()

    first = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("first.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    second = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[("pdf_files", ("second.pdf", b"%PDF-1.4\n", "application/pdf"))],
    )

    assert first.status_code == 200
    assert second.status_code == 200
    assert [item["originalFilename"] for item in second.json()["files"]["pdfInvoices"]] == ["first.pdf", "second.pdf"]
    assert [item["originalFilename"] for item in second.json()["files"]["workbooks"]] == ["bill.xlsx"]


def _reocr_csv_bytes() -> bytes:
    return (
        "Employee,Hours,Amount,Page,Confidence,Evidence\n"
        "Alice Worker,8,100,p1,96%,Alice Worker 8 $100\n"
    ).encode("utf-8")


def test_labor_delete_run_removes_files_and_keeps_redacted_audit(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    audit_path = tmp_path / "audit" / "events.jsonl"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(audit_path))
    monkeypatch.setattr(
        app_module,
        "delete_labor_run_from_persistent",
        lambda _run_id, _owner_user_id: None,
    )
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Sensitive Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()
    run_dir = runs_dir / run["id"]
    (run_dir / "invoice.pdf").write_text("Sensitive Person $123.45", encoding="utf-8")
    app_module.update_labor_metadata(run["id"], {"status": "已生成差异报告", "ownerUserId": "user-1"})

    response = client.delete(f"/api/labor/runs/{run['id']}")

    assert response.status_code == 200
    assert not run_dir.exists()
    raw_audit = audit_path.read_text(encoding="utf-8")
    assert '"action":"run_deleted"' in raw_audit
    assert "Sensitive Person" not in raw_audit
    assert "123.45" not in raw_audit


def test_labor_audit_endpoint_returns_recent_redacted_events_for_run(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.audit import append_labor_audit_event

    audit_path = tmp_path / "audit" / "labor.jsonl"
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(audit_path))
    append_labor_audit_event(
        audit_path,
        action="run_created",
        run_id="labor_1",
        owner_user_id="user-1",
        outcome="success",
        details={"pdfFileCount": 2, "employeeName": "Sensitive Person", "amount": 123.45},
    )
    append_labor_audit_event(
        audit_path,
        action="files_uploaded",
        run_id="labor_1",
        owner_user_id="user-1",
        outcome="success",
        details={"pdfPageCount": 8},
    )
    append_labor_audit_event(
        audit_path,
        action="run_created",
        run_id="labor_2",
        owner_user_id="user-2",
        outcome="success",
    )

    response = TestClient(app).get("/api/labor/audit", params={"run_id": "labor_1", "limit": 10})

    assert response.status_code == 200
    body = response.json()
    assert [event["action"] for event in body["events"]] == ["files_uploaded", "run_created"]
    assert body["events"][0]["details"] == {"pdfPageCount": 8}
    assert "Sensitive Person" not in response.text
    assert "123.45" not in response.text


def test_labor_delete_run_blocks_active_task(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()
    app_module.update_labor_metadata(
        run["id"],
        {"status": "抽取中", "asyncTask": {"status": "running"}},
    )

    response = client.delete(f"/api/labor/runs/{run['id']}")

    assert response.status_code == 409
    assert response.json()["detail"]["errorCode"] == "LABOR_ACTIVE_RUN_DELETE_BLOCKED"
    assert (runs_dir / run["id"]).exists()


def test_labor_upload_rejects_pdf_count_limit_before_saving(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setenv("LABOR_MAX_PDF_FILES", "1")
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()

    response = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("a.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("b.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )

    assert response.status_code == 413
    assert response.json()["detail"]["errorCode"] == "LABOR_PDF_FILE_COUNT_LIMIT_EXCEEDED"
    assert list((runs_dir / run["id"]).glob("*.pdf")) == []


def test_labor_upload_accepts_four_workbooks_with_default_limit(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.delenv("LABOR_MAX_WORKBOOK_FILES", raising=False)
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()
    files = [("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf"))]
    files.extend(
        (
            "workbook_files",
            (
                f"bill-{index}.xlsx",
                _excel_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        )
        for index in range(1, 5)
    )

    response = client.post(f"/api/labor/runs/{run['id']}/files", files=files)

    assert response.status_code == 200
    assert len(response.json()["files"]["workbooks"]) == 4


def test_labor_upload_rolls_back_partial_files_when_byte_limit_exceeded(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setenv("LABOR_MAX_PDF_BYTES", "12")
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()

    response = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("small.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("large.pdf", b"%PDF-1.4\nTOO-LARGE", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )

    assert response.status_code == 413
    assert response.json()["detail"]["errorCode"] == "LABOR_PDF_SIZE_LIMIT_EXCEEDED"
    assert list((runs_dir / run["id"]).glob("*.pdf")) == []
    metadata = client.get(f"/api/labor/runs/{run['id']}").json()
    assert metadata["files"] == {}


def test_labor_upload_rejects_total_pdf_page_limit(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setenv("LABOR_MAX_PDF_PAGES", "3")
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    monkeypatch.setattr(app_module, "_labor_pdf_page_count", lambda _path: 4)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()

    response = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )

    assert response.status_code == 413
    assert response.json()["detail"]["errorCode"] == "LABOR_PDF_PAGE_LIMIT_EXCEEDED"
    assert list((runs_dir / run["id"]).glob("*.pdf")) == []


def test_labor_extract_endpoint_enforces_owner_concurrency_limit(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.hardening import LaborTaskLimiter
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "_LABOR_TASK_LIMITER", LaborTaskLimiter())
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    monkeypatch.setenv("LABOR_MAX_ACTIVE_TASKS_PER_OWNER", "1")
    monkeypatch.setenv("LABOR_MAX_ACTIVE_TASKS_GLOBAL", "2")
    monkeypatch.setattr(app_module, "_uses_request_scoped_labor_runtime", lambda: False)

    class FakeLoop:
        def run_in_executor(self, *_args, **_kwargs):
            return None

    monkeypatch.setattr(app_module.asyncio, "get_event_loop", lambda: FakeLoop())
    client = TestClient(app)

    def ready_run(supplier: str) -> dict:
        run = client.post(
            "/api/labor/runs",
            json={"supplier_name": supplier, "period_start": "2026-07-01", "period_end": "2026-07-07"},
        ).json()
        response = client.post(
            f"/api/labor/runs/{run['id']}/files",
            files=[
                ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
                ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ],
        )
        assert response.status_code == 200
        response = client.post(
            f"/api/labor/runs/{run['id']}/mapping",
            json={
                "sheet_name": "员工账单",
                "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"},
            },
        )
        assert response.status_code == 200
        return run

    first = ready_run("Supplier A")
    second = ready_run("Supplier B")

    first_response = client.post(f"/api/labor/runs/{first['id']}/extract-and-compare")
    second_response = client.post(f"/api/labor/runs/{second['id']}/extract-and-compare")

    assert first_response.status_code == 200
    assert second_response.status_code == 429
    assert second_response.json()["detail"]["errorCode"] == "LABOR_OWNER_CONCURRENCY_LIMIT_EXCEEDED"
    assert client.get(f"/api/labor/runs/{second['id']}").json()["status"] == "已确认字段"


def test_labor_extract_request_scoped_runtime_never_reserves_local_runner_slot(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.hardening import LaborTaskLimiter
    import bonus_platform.engine.labor.runs as labor_runs

    runs_dir = tmp_path / "labor_runs"
    limiter = LaborTaskLimiter()
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "LABOR_RUNS_DIR", runs_dir)
    monkeypatch.setattr(app_module, "_LABOR_TASK_LIMITER", limiter)
    monkeypatch.setenv("LABOR_AUDIT_PATH", str(tmp_path / "audit.jsonl"))
    monkeypatch.setenv("SIGMA_OVERSEAS_LABOR_ACCESS", "uat_full")
    monkeypatch.setattr(app_module, "_uses_request_scoped_labor_runtime", lambda: True)
    monkeypatch.setattr(app_module, "_run_labor_extract_compare", lambda _run_id: None)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Supplier", "period_start": "2026-07-01", "period_end": "2026-07-07"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"},
        },
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 409
    assert response.json()["detail"]["errorCode"] == "LABOR_UAT_EXTRACT_DISABLED"
    assert limiter.snapshot()["activeGlobalTasks"] == 0


def test_labor_safe_early_ocr_candidate_skips_legacy_detail_extraction(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    app_module.update_labor_metadata(run["id"], {"requireEmployeeDetail": True})
    monkeypatch.setitem(app_module.AI_CONFIG, "ocr_command", "python worker.py")
    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": Path(paths[0]).name,
                "total_amount": 100.0,
                "warehouse_id": "1",
                "authoritative": True,
                "evidence_status": "authoritative",
                "page_evidence": [{"page": 1, "role": "invoice_total", "total_amount": 100.0}],
            }
        ],
    )
    monkeypatch.setattr(app_module, "extract_structured_invoice_rows", lambda *args, **kwargs: [])

    def safe_candidate(_run_id, pdf_paths, *_args, **_kwargs):
        return {
            "decision": "auto_accept",
            "safeToUse": True,
            "runtimeStatus": "completed",
            "runtimeFiles": [{"sourceFile": pdf_paths[0].name, "cacheHit": True}],
            "rows": [
                {
                    "source_type": "pdf_invoice_candidate",
                    "source_file": pdf_paths[0].name,
                    "source_page_or_row": "p1",
                    "employee_name_raw": "Alice Worker",
                    "hours": 8,
                    "amount": 100,
                    "currency": "USD",
                    "confidence": 0.99,
                    "warehouse_id": "1",
                }
            ],
        }

    monkeypatch.setattr(app_module, "_run_labor_auto_ocr_candidate", safe_candidate)
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("legacy extraction should be skipped")),
    )

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert saved["autoOcrCandidate"]["safeToUse"] is True
    assert saved["autoOcrCandidate"]["runtimeFiles"][0]["cacheHit"] is True
    assert saved["pdfExtractedRows"][0]["employee_name_raw"] == "Alice Worker"


def test_labor_ocr_expected_totals_uses_unique_excel_warehouse_for_unresolved_invoice():
    excel_rows = [
        LaborLineItem(
            source_type="excel_bill",
            source_file="bill.xlsx",
            source_page_or_row="row 2",
            employee_id="1",
            employee_name_raw="Alice Worker",
            hours=8,
            amount=100,
            currency="USD",
            confidence=1,
            warehouse_id="20",
        ),
        LaborLineItem(
            source_type="excel_bill",
            source_file="bill.xlsx",
            source_page_or_row="row 3",
            employee_id="2",
            employee_name_raw="Bob Worker",
            hours=8,
            amount=120,
            currency="USD",
            confidence=1,
            warehouse_id="20",
        ),
    ]
    totals = [
        {
            "source_file": "elog20-4_20260520204256.pdf",
            "total_amount": 0,
            "warehouse_id": "20",
            "authoritative": False,
        }
    ]

    expected = app_module._labor_ocr_expected_totals(
        totals,
        {"elog20-4_20260520204256.pdf"},
        excel_rows=excel_rows,
    )

    assert expected == {"elog20-4_20260520204256.pdf": 220.0}


def test_labor_review_early_ocr_candidate_falls_back_to_legacy_once(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    app_module.update_labor_metadata(run["id"], {"requireEmployeeDetail": True})
    monkeypatch.setitem(app_module.AI_CONFIG, "ocr_command", "python worker.py")
    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": Path(paths[0]).name,
                "total_amount": 100.0,
                "warehouse_id": "1",
                "authoritative": True,
                "evidence_status": "authoritative",
                "page_evidence": [{"page": 1, "role": "invoice_total", "total_amount": 100.0}],
            }
        ],
    )
    monkeypatch.setattr(app_module, "extract_structured_invoice_rows", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        app_module,
        "_run_labor_auto_ocr_candidate",
        lambda *args, **kwargs: {
            "decision": "needs_review",
            "safeToUse": False,
            "blockers": ["strict_name_review_required"],
            "runtimeStatus": "completed",
            "runtimeFiles": [{"cacheHit": True}],
            "rows": [],
        },
    )
    calls = {"legacy": 0}

    def legacy_rows(pdf_paths, *args, **kwargs):
        calls["legacy"] += 1
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(pdf_paths[0]).name,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=100,
                currency="USD",
                confidence=0.99,
                warehouse_id="1",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", legacy_rows)

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert calls["legacy"] == 1
    assert saved["autoOcrCandidate"]["safeToUse"] is False
    assert saved["pdfExtractedRows"][0]["employee_name_raw"] == "Alice Worker"


def test_labor_review_ocr_candidate_with_explicit_pdf_total_skips_empty_legacy_fallback(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    app_module.update_labor_metadata(run["id"], {"requireEmployeeDetail": True})
    monkeypatch.setitem(app_module.AI_CONFIG, "ocr_command", "python worker.py")
    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": Path(paths[0]).name,
                "total_amount": 0.0,
                "warehouse_id": "1",
                "authoritative": False,
                "evidence_status": "needs_review",
            }
        ],
    )
    monkeypatch.setattr(app_module, "extract_structured_invoice_rows", lambda *args, **kwargs: [])

    def review_candidate(_run_id, pdf_paths, *_args, **_kwargs):
        source_file = pdf_paths[0].name
        return {
            "decision": "needs_review",
            "safeToUse": False,
            "blockers": ["strict_name_review_required"],
            "runtimeStatus": "completed",
            "runtimeFiles": [
                {
                    "sourceFile": source_file,
                    "pageCount": 1,
                    "successfulPageCount": 1,
                    "failedPageCount": 0,
                    "explicitTotalAmount": 100.0,
                    "explicitTotalEvidence": {
                        "page": 1,
                        "evidenceText": "TOTAL: $ 100.00",
                    },
                }
            ],
            "pdfTotalEvidence": {
                source_file: {
                    "amount": 100.0,
                    "page": 1,
                    "evidenceText": "TOTAL: $ 100.00",
                }
            },
            "rows": [
                {
                    "source_type": "pdf_invoice_candidate",
                    "source_file": source_file,
                    "source_page_or_row": "p1",
                    "employee_name_raw": "Alic Worker",
                    "hours": 8,
                    "amount": 100,
                    "currency": "USD",
                    "confidence": 0.9,
                    "warehouse_id": "1",
                }
            ],
        }

    monkeypatch.setattr(app_module, "_run_labor_auto_ocr_candidate", review_candidate)
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("page-complete OCR review rows should be retained")
        ),
    )

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert saved["autoOcrCandidate"]["reviewRowsApplied"] is True
    assert saved["pdfExtractedRows"][0]["employee_name_raw"] == "Alic Worker"
    assert saved["batchGuard"]["status"] == "ocr_candidate_review"
    assert saved["batchGuard"]["allowReleasableReport"] is False
    assert saved["warehouseComparison"]["summary"]["pdfAmountTotal"] == 100.0


def test_labor_review_candidate_retries_only_the_review_page(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    app_module.update_labor_metadata(run["id"], {"requireEmployeeDetail": True})
    monkeypatch.setitem(app_module.AI_CONFIG, "ocr_command", "python worker.py")
    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": Path(paths[0]).name,
                "total_amount": 100.0,
                "warehouse_id": "1",
                "authoritative": True,
                "evidence_status": "authoritative",
                "page_evidence": [{"page": 4, "role": "invoice_total", "total_amount": 100.0}],
            }
        ],
    )
    monkeypatch.setattr(app_module, "extract_structured_invoice_rows", lambda *args, **kwargs: [])
    def review_candidate(_run_id, pdf_paths, *_args, **_kwargs):
        source_file = pdf_paths[0].name
        return {
            "decision": "needs_review",
            "safeToUse": False,
            "blockers": ["strict_name_review_required"],
            "runtimeStatus": "completed",
            "runtimeFiles": [{"cacheHit": True}],
            "fileClosure": [{"sourceFile": source_file, "expectedAmount": 100, "closed": True}],
            "nameGate": {
                "matches": [
                    {"candidateName": "Alic Worker", "excelName": "Alice Worker", "status": "review"}
                ]
            },
            "rows": [
                {
                    "source_type": "pdf_invoice_candidate",
                    "source_file": source_file,
                    "source_page_or_row": "p4",
                    "employee_name_raw": "Alic Worker",
                    "hours": 8,
                    "amount": 100,
                    "currency": "USD",
                    "confidence": 0.9,
                    "warehouse_id": "1",
                }
            ],
        }

    monkeypatch.setattr(app_module, "_run_labor_auto_ocr_candidate", review_candidate)
    captured = {}

    def targeted_rows(pdf_paths, *args, **kwargs):
        captured["paths"] = [Path(path).name for path in pdf_paths]
        captured["pages"] = kwargs.get("allowed_pages_by_source")
        source_file = Path(pdf_paths[0]).name
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=source_file,
                source_page_or_row="p4",
                employee_id="",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=100,
                currency="USD",
                confidence=0.99,
                warehouse_id="1",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", targeted_rows)

    saved = app_module._perform_labor_extract_compare(run["id"])

    source_file = captured["paths"][0]
    assert captured["pages"] == {source_file: [4]}
    assert saved["pdfExtractedRows"][0]["employee_name_raw"] == "Alice Worker"
    assert saved["autoOcrCandidate"]["targetedRetry"]["applied"] is True


def _reocr_batch_csv_bytes() -> bytes:
    return (
        "SourceFile,WarehouseId,Employee,Hours,Amount,Page,Confidence,Currency,Evidence\n"
        "elog1-1_20260520204104.pdf,1,Alice Worker,8,100,p1,96%,USD,Alice Worker 8 $100\n"
        "elog2-2_20260520204104.pdf,2,Wrong Worker,10,200,p1,96%,USD,Wrong Worker 10 $200\n"
    ).encode("utf-8")


def test_labor_telemetry_records_sanitized_events_and_exports_jsonl(monkeypatch, tmp_path):
    telemetry_dir = tmp_path / "labor_telemetry"
    telemetry_file = telemetry_dir / "events.jsonl"
    monkeypatch.setattr(app_module, "LABOR_TELEMETRY_DIR", telemetry_dir)
    monkeypatch.setattr(app_module, "LABOR_TELEMETRY_FILE", telemetry_file)
    client = TestClient(app)

    response = client.post(
        "/api/labor/telemetry",
        json={
            "event": "labor.extract.completed",
            "runId": "labor_sample",
            "supplier": "Fairway Staffing Service",
            "step": "extract_compare",
            "status": "completed",
            "durationMs": 1234.4,
            "errorMessage": "first line\nsecond line",
            "summary": {
                "pdfAmountTotal": 100,
                "excelAmountTotal": 99.9,
                "exceptionCount": 1,
                "employeeRows": [{"employee": "Alice Worker", "amount": 100}],
            },
            "context": {
                "pdfCount": 2,
                "workbookCount": 1,
                "sourceFilename": "invoice.pdf",
            },
        },
    )

    assert response.status_code == 200
    event = json.loads(telemetry_file.read_text(encoding="utf-8").strip())
    assert event["schemaVersion"] == 1
    assert event["event"] == "labor.extract.completed"
    assert event["durationMs"] == 1234
    assert event["errorMessage"] == "first line second line"
    assert event["summary"] == {
        "pdfAmountTotal": 100.0,
        "excelAmountTotal": 99.9,
        "exceptionCount": 1.0,
    }
    assert event["context"] == {"pdfCount": 2.0, "workbookCount": 1.0}

    export = client.get("/api/labor/telemetry/export")
    assert export.status_code == 200
    assert export.text.strip() == telemetry_file.read_text(encoding="utf-8").strip()


def test_labor_access_endpoint_marks_uat_trial(monkeypatch):
    monkeypatch.delenv("SIGMA_OVERSEAS_LABOR_ACCESS", raising=False)
    client = TestClient(app)

    response = client.get("/api/labor/access")

    assert response.status_code == 200
    body = response.json()
    assert body["stage"] == "UAT试用版"
    assert body["access"] == "uat_trial"
    assert body["canUse"] is True
    assert "Payroll Admin" in body["allowedRoles"]
    assert body["uploadLimits"]["maxWorkbookFiles"] == 10


def test_labor_access_endpoint_exposes_release_contract(monkeypatch):
    monkeypatch.setenv("SIGMA_LABOR_BUILD_ID", "build-20260715-p0")
    monkeypatch.setenv("SIGMA_LABOR_SOURCE_REF", "codex/overseas-labor-p0")
    client = TestClient(app)

    response = client.get("/api/labor/access")

    assert response.status_code == 200
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["x-sigma-labor-build"] == "build-20260715-p0"
    body = response.json()
    assert body["version"] == "0.5-uat"
    assert body["apiContractVersion"] == 2
    assert body["buildId"] == "build-20260715-p0"
    assert "sourceRef" not in body
    assert "runtimeStartedAt" not in body
    assert body["build"]["schemaVersion"] == 1
    assert body["build"]["status"] == "current"
    assert "startupFingerprint" not in json.dumps(body)
    assert "currentFingerprint" not in json.dumps(body)
    assert body["runtimeGate"]["canStartFormalTask"] is True
    assert body["runtimeGate"]["runtimeSourceCurrent"] is True
    assert body["reconciliationScope"] == "employee_detail_required"
    assert body["manualReviewRequired"] is True
    assert body["directPaymentAllowed"] is False


def test_labor_access_allows_formal_uat_queue_only_through_personal_worker(monkeypatch):
    monkeypatch.setenv("VERCEL", "1")
    monkeypatch.setenv("SIGMA_LABOR_EXECUTION_MODE", "personal-worker")
    client = TestClient(app)

    response = client.get("/api/labor/access")

    assert response.status_code == 200
    gate = response.json()["formalTaskGate"]
    assert gate["canQueue"] is True
    assert gate["executionMode"] == "personal_worker"
    assert gate["reasonCode"] == ""


def test_labor_access_blocks_request_scoped_formal_task_without_personal_worker(monkeypatch):
    monkeypatch.setenv("VERCEL", "1")
    monkeypatch.delenv("SIGMA_LABOR_EXECUTION_MODE", raising=False)
    client = TestClient(app)

    response = client.get("/api/labor/access")

    assert response.status_code == 200
    gate = response.json()["formalTaskGate"]
    assert gate["canQueue"] is False
    assert gate["executionMode"] == "blocked"
    assert gate["reasonCode"] == "LABOR_PERSONAL_WORKER_REQUIRED"


def test_labor_runtime_gate_blocks_new_formal_batch_when_service_source_changed(monkeypatch):
    stale_build = {
        "status": "restart_required",
        "buildId": "startup-build",
        "sourceRef": "local-worktree",
        "processStartedAt": "2026-07-15T08:00:00Z",
    }
    monkeypatch.setattr(app_module, "_labor_build_snapshot", lambda: stale_build)
    client = TestClient(app)

    access = client.get("/api/labor/access")
    response = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Synthetic Supplier",
            "period_start": "2026-07-01",
            "period_end": "2026-07-07",
            "currency": "USD",
        },
    )

    assert access.status_code == 200
    assert access.json()["canUse"] is True
    assert access.json()["runtimeGate"]["canStartFormalTask"] is False
    assert response.status_code == 409
    assert response.json()["detail"]["errorCode"] == "LABOR_SERVICE_RESTART_REQUIRED"


def test_labor_formal_mutation_rejects_missing_or_stale_ui_contract_when_enforced(monkeypatch):
    monkeypatch.setenv("SIGMA_TEST_NO_LABOR_CONTRACT_HEADERS", "1")
    monkeypatch.setenv("SIGMA_LABOR_REQUIRE_CLIENT_CONTRACT", "true")
    monkeypatch.setenv("SIGMA_LABOR_BUILD_ID", "build-contract-p0")
    client = TestClient(app)
    payload = {
        "supplier_name": "Synthetic Supplier",
        "period_start": "2026-07-01",
        "period_end": "2026-07-07",
        "currency": "USD",
    }

    missing = client.post("/api/labor/runs", json=payload)
    stale = client.post(
        "/api/labor/runs",
        headers={
            "x-sigma-labor-api-contract": "2",
            "x-sigma-labor-ui-version": "0.5-uat",
            "x-sigma-labor-ui-build": "old-build",
        },
        json=payload,
    )
    current = client.post(
        "/api/labor/runs",
        headers={
            "x-sigma-labor-api-contract": "2",
            "x-sigma-labor-ui-version": "0.5-uat",
            "x-sigma-labor-ui-build": "build-contract-p0",
        },
        json=payload,
    )

    assert missing.status_code == 409
    assert missing.json()["detail"]["errorCode"] == "LABOR_CLIENT_UPGRADE_REQUIRED"
    assert stale.status_code == 409
    assert stale.json()["detail"]["errorCode"] == "LABOR_CLIENT_UPGRADE_REQUIRED"
    assert current.status_code == 200


def test_labor_cleanup_requires_operations_token(monkeypatch):
    monkeypatch.setenv("SIGMA_LABOR_OPERATIONS_TOKEN", "admin-secret")
    monkeypatch.setattr(app_module, "_cleanup_expired_labor_data", lambda: {"deleted": 0})
    client = TestClient(app)

    assert client.post("/api/labor/maintenance/cleanup").status_code == 401
    response = client.post(
        "/api/labor/maintenance/cleanup",
        headers={"x-admin-token": "admin-secret"},
    )

    assert response.status_code == 200
    assert response.json() == {"deleted": 0}


def test_labor_access_gate_can_disable_uat_module(monkeypatch):
    monkeypatch.setenv("SIGMA_OVERSEAS_LABOR_ACCESS", "disabled")
    client = TestClient(app)

    blocked = client.get("/api/labor/runs")
    access = client.get("/api/labor/access")

    assert access.status_code == 200
    assert access.json()["canUse"] is False
    assert blocked.status_code == 403
    assert blocked.json()["access"]["access"] == "disabled"


def test_labor_extract_is_blocked_in_vercel_uat_light_mode(monkeypatch):
    monkeypatch.setenv("VERCEL", "1")
    monkeypatch.setenv("SIGMA_LABOR_AUTH_REQUIRED", "0")
    monkeypatch.setenv("SIGMA_OVERSEAS_LABOR_ACCESS", "uat")
    monkeypatch.setenv("SIGMA_WORKBENCH_HOME", "/tmp/sigma-workbench")
    monkeypatch.setenv("SIGMA_LABOR_STORAGE_BACKEND", "blob")
    monkeypatch.setenv("BLOB_READ_WRITE_TOKEN", "vercel_blob_rw_qqD75P7a2QuwEh0S_abcd1234")
    monkeypatch.setattr(
        app_module,
        "_labor_metadata_or_404",
        lambda run_id: (_ for _ in ()).throw(
            app_module.HTTPException(status_code=404, detail="metadata read should not happen")
        ),
    )
    client = TestClient(app)

    response = client.post("/api/labor/runs/labor_synthetic/extract-and-compare")

    assert response.status_code == 409
    assert "Vercel UAT" in response.json()["detail"]["message"]
    assert "测试材料验证" in response.json()["detail"]["message"]


def test_labor_extract_vercel_uat_light_mode_returns_structured_next_action(monkeypatch):
    monkeypatch.setenv("VERCEL", "1")
    monkeypatch.setenv("SIGMA_LABOR_AUTH_REQUIRED", "0")
    monkeypatch.setenv("SIGMA_OVERSEAS_LABOR_ACCESS", "uat")
    monkeypatch.setenv("SIGMA_LABOR_STORAGE_BACKEND", "blob")
    monkeypatch.setenv("BLOB_READ_WRITE_TOKEN", "vercel_blob_rw_qqD75P7a2QuwEh0S_abcd1234")
    client = TestClient(app)

    response = client.post("/api/labor/runs/labor_synthetic/extract-and-compare")

    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_UAT_EXTRACT_DISABLED"
    assert detail["retryable"] is False
    assert "Vercel UAT" in detail["message"]
    assert "测试材料验证" in detail["nextAction"]


def test_labor_extract_never_runs_synchronously_on_vercel_full_uat(monkeypatch):
    monkeypatch.setenv("VERCEL", "1")
    monkeypatch.setenv("SIGMA_LABOR_AUTH_REQUIRED", "0")
    monkeypatch.setenv("SIGMA_OVERSEAS_LABOR_ACCESS", "uat_full")
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "OneSource UAT",
            "period_start": "2026-06-17",
            "period_end": "2026-06-17",
            "currency": "USD",
        },
    ).json()
    app_module.update_labor_metadata(
        run["id"],
        {
            "files": {
                "pdfInvoices": [{"path": "/tmp/invoice.pdf"}],
                "workbooks": [{"path": "/tmp/bill.xlsx"}],
            },
            "workbookSheet": "员工账单",
            "excelMapping": {"name": "姓名", "hours": "工时", "amount": "金额"},
        },
    )

    called = []

    def fake_extract(run_id):
        called.append(run_id)

    monkeypatch.setattr(app_module, "_run_labor_extract_compare", fake_extract)

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 409
    assert response.json()["detail"]["errorCode"] == "LABOR_UAT_EXTRACT_DISABLED"
    assert called == []


def test_labor_extract_never_runs_synchronously_when_only_vercel_env_is_present(monkeypatch):
    monkeypatch.delenv("VERCEL", raising=False)
    monkeypatch.setenv("VERCEL_ENV", "production")
    monkeypatch.setenv("SIGMA_LABOR_AUTH_REQUIRED", "0")
    monkeypatch.setenv("SIGMA_OVERSEAS_LABOR_ACCESS", "uat_full")
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "OneSource UAT",
            "period_start": "2026-06-17",
            "period_end": "2026-06-17",
            "currency": "USD",
        },
    ).json()
    app_module.update_labor_metadata(
        run["id"],
        {
            "files": {
                "pdfInvoices": [{"path": "/tmp/invoice.pdf"}],
                "workbooks": [{"path": "/tmp/bill.xlsx"}],
            },
            "workbookSheet": "员工账单",
            "excelMapping": {"name": "姓名", "hours": "工时", "amount": "金额"},
        },
    )

    called = []

    def fake_extract(run_id):
        called.append(run_id)

    monkeypatch.setattr(app_module, "_run_labor_extract_compare", fake_extract)

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 409
    assert response.json()["detail"]["errorCode"] == "LABOR_UAT_EXTRACT_DISABLED"
    assert called == []


def test_labor_upload_missing_run_returns_structured_next_action():
    client = TestClient(app)

    response = client.post(
        "/api/labor/runs/labor_missing/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )

    assert response.status_code == 404
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_RUN_NOT_FOUND"
    assert detail["retryable"] is False
    assert detail["requiresReupload"] is True
    assert "批次记录未找到" in detail["message"]
    assert "新建核对批次" in detail["nextAction"]


def test_labor_download_recovers_nested_blob_report_by_metadata(monkeypatch):
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Blob Supplier", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    run_dir = app_module.get_labor_run_dir(run["id"])
    report_path = run_dir / "reports" / "business.html"
    app_module.update_labor_metadata(
        run["id"],
        {
            "files": {
                "businessReport": {
                    "filename": "business.html",
                    "path": str(report_path),
                    "downloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
                }
            },
            "businessReportDownloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
        },
    )

    monkeypatch.setattr(app_module, "labor_blob_storage_enabled", lambda: True)

    def fake_sync_from_blob(run_id: str, target_dir: Path) -> bool:
        restored_report = target_dir / "reports" / "business.html"
        restored_report.parent.mkdir(parents=True, exist_ok=True)
        restored_report.write_text("<html>business report</html>", encoding="utf-8")
        return True

    monkeypatch.setattr(app_module, "sync_labor_run_from_blob", fake_sync_from_blob)

    response = client.get(f"/api/labor/runs/{run['id']}/download/business.html")

    assert response.status_code == 200
    assert response.text == "<html>business report</html>"


def test_labor_download_returns_structured_restore_failure_when_blob_sync_fails(monkeypatch):
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Blob Supplier", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    run_dir = app_module.get_labor_run_dir(run["id"])
    report_path = run_dir / "reports" / "business.html"
    app_module.update_labor_metadata(
        run["id"],
        {
            "files": {
                "businessReport": {
                    "filename": "business.html",
                    "path": str(report_path),
                    "downloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
                }
            },
            "businessReportDownloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
        },
    )

    monkeypatch.setattr(app_module, "labor_blob_storage_enabled", lambda: True)
    monkeypatch.setattr(app_module, "sync_labor_run_from_blob", lambda run_id, target_dir: False)

    response = client.get(f"/api/labor/runs/{run['id']}/download/business.html")

    assert response.status_code == 503
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_REPORT_RESTORE_FAILED"
    assert detail["retryable"] is True
    assert detail["requiresReupload"] is False
    assert "报告文件暂时无法恢复" in detail["message"]
    assert "稍后重试" in detail["nextAction"]
    assert "reports/business.html" not in str(detail)


def test_labor_download_masks_blob_restore_exception_details(monkeypatch):
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Blob Supplier", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    run_dir = app_module.get_labor_run_dir(run["id"])
    report_path = run_dir / "reports" / "business.html"
    app_module.update_labor_metadata(
        run["id"],
        {
            "files": {
                "businessReport": {
                    "filename": "business.html",
                    "path": str(report_path),
                    "downloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
                }
            },
            "businessReportDownloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
        },
    )

    monkeypatch.setattr(app_module, "labor_blob_storage_enabled", lambda: True)

    def raise_sensitive_restore_error(run_id: str, target_dir: Path) -> bool:
        raise RuntimeError(f"token=secret-token path={target_dir / 'reports' / 'business.html'}")

    monkeypatch.setattr(app_module, "sync_labor_run_from_blob", raise_sensitive_restore_error)

    response = client.get(f"/api/labor/runs/{run['id']}/download/business.html")

    assert response.status_code == 503
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_REPORT_RESTORE_FAILED"
    assert detail["retryable"] is True
    assert "报告文件暂时无法恢复" in detail["message"]
    response_body = str(detail)
    assert "secret-token" not in response_body
    assert "reports/business.html" not in response_body
    assert str(run_dir) not in response_body


def test_labor_download_returns_structured_missing_file_error():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Local Supplier", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    run_dir = app_module.get_labor_run_dir(run["id"])
    missing_report_path = run_dir / "reports" / "business.html"
    app_module.update_labor_metadata(
        run["id"],
        {
            "files": {
                "businessReport": {
                    "filename": "business.html",
                    "path": str(missing_report_path),
                    "downloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
                }
            },
            "businessReportDownloadUrl": f"/api/labor/runs/{run['id']}/download/business.html",
        },
    )

    response = client.get(f"/api/labor/runs/{run['id']}/download/business.html")

    assert response.status_code == 404
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_REPORT_FILE_MISSING"
    assert detail["retryable"] is False
    assert detail["requiresReupload"] is False
    assert detail["requiresHumanReview"] is True
    assert "报告文件不存在或已被清理" in detail["message"]
    assert "重新生成报告" in detail["nextAction"]
    assert "reports/business.html" not in str(detail)


def test_labor_runs_list_uses_bounded_recent_metadata(monkeypatch):
    observed: dict[str, object] = {}

    def fake_list_labor_metadata(*, limit=None) -> list[dict]:
        observed["limit"] = limit
        return [
            {
                "id": "labor_recent",
                "status": "已生成差异报告",
                "supplierName": "Synthetic Supplier",
                "updatedAt": "2026-06-20T10:00:00",
                "comparisonSummary": {"exceptionCount": 1},
                "comparisonRows": [{"employeeName": "Synthetic Worker"}],
                "pdfExtractedRows": [{"employeeNameRaw": "Synthetic Worker"}],
                "excelRows": [{"employeeNameRaw": "Synthetic Worker"}],
            }
        ]

    monkeypatch.setattr(app_module, "list_labor_metadata", fake_list_labor_metadata)
    client = TestClient(app)

    response = client.get("/api/labor/runs")

    assert response.status_code == 200
    assert response.json()["runs"] == [
        {
            "id": "labor_recent",
            "status": "已生成差异报告",
            "supplierName": "Synthetic Supplier",
            "periodStart": "",
            "periodEnd": "",
            "currency": "",
            "createdAt": "",
            "updatedAt": "2026-06-20T10:00:00",
            "stage": "",
            "diffDownloadUrl": "",
            "comparisonSummary": {"exceptionCount": 1},
            "readinessGate": {},
        }
    ]
    assert observed["limit"] == 50


def test_labor_run_api_accepts_explicit_employee_detail_audit_mode():
    client = TestClient(app)

    response = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Unknown Synthetic Supplier",
            "period_start": "2026-06-01",
            "period_end": "2026-06-07",
            "currency": "USD",
            "require_employee_detail": True,
        },
    )

    assert response.status_code == 200
    assert response.json()["requireEmployeeDetail"] is True


def test_labor_run_api_requires_employee_detail_for_formal_batches_even_when_client_disables_it():
    client = TestClient(app)

    response = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Unknown Synthetic Supplier",
            "period_start": "2026-06-01",
            "period_end": "2026-06-07",
            "currency": "USD",
            "require_employee_detail": False,
        },
    )

    assert response.status_code == 200
    assert response.json()["requireEmployeeDetail"] is True
    assert response.json()["reconciliationScope"] == "employee_detail_required"


def test_labor_employee_detail_scope_defaults_to_formal_and_requires_explicit_diagnostic_marker():
    assert app_module._labor_requires_employee_detail({}) is True
    assert app_module._labor_requires_employee_detail({"requireEmployeeDetail": False}) is True
    assert app_module._labor_requires_employee_detail({"reconciliationScope": "total_only_diagnostic"}) is True
    assert app_module._labor_requires_employee_detail(
        {"reconciliationScope": "total_only_diagnostic", "diagnosticOnly": True}
    ) is False


def test_labor_uat_result_always_waits_for_business_review_even_when_machine_passes():
    passed = app_module._labor_uat_review_state({"canRelease": True, "conclusionLevel": "pass"})
    blocked = app_module._labor_uat_review_state({"canRelease": False, "conclusionLevel": "critical"})

    assert passed == {
        "machineCheckStatus": "passed",
        "businessReviewStatus": "pending",
        "manualReviewRequired": True,
        "directPaymentAllowed": False,
        "requiresHumanReview": True,
    }
    assert blocked["machineCheckStatus"] == "blocked"
    assert blocked["businessReviewStatus"] == "pending"
    assert blocked["requiresHumanReview"] is True


def test_labor_run_api_creates_batch_uploads_files_and_suggests_mapping():
    client = TestClient(app)

    create = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Fairway Staffing Service", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD", "notes": "sample"},
    )

    assert create.status_code == 200
    run = create.json()
    assert run["id"].startswith("labor_")
    assert run["status"] == "已创建"

    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n% sample", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )

    assert upload.status_code == 200
    uploaded = upload.json()
    assert uploaded["status"] == "已上传文件"
    assert uploaded["files"]["workbook"]["filename"].endswith(".xlsx")
    assert uploaded["files"]["pdfInvoices"][0]["filename"].endswith(".pdf")

    sheets = client.get(f"/api/labor/runs/{run['id']}/workbook-sheets")
    assert sheets.status_code == 200
    assert sheets.json()["sheets"] == ["员工账单"]

    suggestion = client.post(f"/api/labor/runs/{run['id']}/field-suggestions", json={"sheet_name": "员工账单"})
    assert suggestion.status_code == 200
    assert suggestion.json()["suggestedMapping"]["name"] == "姓名"
    assert suggestion.json()["suggestedMapping"]["hours"] == "时长总计(H)"

    mapping = client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {"name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"},
            "manualNameMapping": {"Gamboa, Arilene": "Arlene Gamboa"},
        },
    )
    assert mapping.status_code == 200
    assert mapping.json()["excelMapping"]["name"] == "姓名"
    assert mapping.json()["manualNameMapping"]["Gamboa, Arilene"] == "Arlene Gamboa"


def test_labor_extract_before_upload_tells_user_to_upload_files_first():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Fairway Staffing Service",
            "period_start": "2026-05-11",
            "period_end": "2026-05-17",
            "currency": "USD",
        },
    ).json()

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_FILES_REQUIRED"
    assert detail["requiresReupload"] is True
    assert "请先上传本期 PDF 发票、Excel 账单" in detail["message"]
    assert "上传文件" in detail["nextAction"]


def test_labor_extract_after_upload_without_mapping_tells_user_to_confirm_mapping():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={
            "supplier_name": "Fairway Staffing Service",
            "period_start": "2026-05-11",
            "period_end": "2026-05-17",
            "currency": "USD",
        },
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n% sample", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    assert upload.status_code == 200

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 400
    detail = response.json()["detail"]
    assert detail["errorCode"] == "LABOR_MAPPING_REQUIRED"
    assert detail.get("requiresReupload") is False
    assert "请先确认 Excel 工作表和字段映射" in detail["message"]
    assert "字段映射" in detail["nextAction"]


def test_labor_material_index_api_lists_replay_ready_batches(tmp_path):
    batch = tmp_path / "oss 2"
    batch.mkdir()
    (batch / "US Elogis Service #7 Invoice W.E 05.24.26.pdf").write_bytes(b"%PDF-1.4\n")
    (batch / "员工账单明细 - 2026-06-04T094719.972.xlsx").write_bytes(b"fake workbook")
    (tmp_path / "handover").mkdir()
    (tmp_path / "handover" / "README_RECONCILIATION_METHOD.md").write_text("method", encoding="utf-8")

    client = TestClient(app)
    response = client.get("/api/labor/material-index", params={"root": str(tmp_path)})

    assert response.status_code == 200
    body = response.json()
    assert body["summary"]["candidateBatchCount"] == 1
    assert body["summary"]["suppliers"] == ["oss"]
    assert body["batches"] == body["candidateBatches"]
    assert body["candidateBatches"][0]["warehouseIds"] == ["7"]
    assert body["candidateBatches"][0]["invoicePdfCount"] == 1
    assert body["candidateBatches"][0]["workbookCount"] == 1
    assert body["candidateBatches"][0]["pdfFiles"] == body["candidateBatches"][0]["invoiceFiles"]


def test_labor_material_replay_plan_api_returns_mapping_candidates(tmp_path):
    batch = tmp_path / "workforce已报账"
    batch.mkdir()
    (batch / "Invoice-5058871.pdf").write_bytes(b"%PDF-1.4\n")
    (batch / "员工账单明细 - 2026-06-01T112149.990.xlsx").write_bytes(_excel_bytes())

    client = TestClient(app)
    response = client.get(
        "/api/labor/material-replay-plan",
        params={"root": str(tmp_path), "batchKey": "workforce已报账"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["summary"]["planCount"] == 1
    plan = body["plans"][0]
    assert plan["supplier"] == "workforce"
    assert plan["uploadPlan"]["pdfFiles"] == ["workforce已报账/Invoice-5058871.pdf"]
    assert plan["mappingCandidates"][0]["sheetName"] == "员工账单"
    assert plan["mappingCandidates"][0]["suggestedMapping"]["name"] == "姓名"
    assert plan["mappingCandidates"][0]["suggestedMapping"]["amount"] == "费用总计(含税)"
    assert plan["replayReady"] is True


def test_labor_material_dry_run_api_does_not_create_labor_run(monkeypatch, tmp_path):
    import bonus_platform.app as app_module

    before_count = len(app_module.list_labor_metadata())
    monkeypatch.setattr(
        app_module,
        "build_material_dry_run",
        lambda root, batch_key, **kwargs: {
            "decision": "dry_run_only",
            "mode": "deterministic_first_no_write",
            "batchKey": batch_key,
            "supplier": "oss",
            "summary": {"pdfRowCount": 1, "excelRowCount": 1},
            "writesRun": False,
            "aiInvoked": False,
        },
    )

    client = TestClient(app)
    response = client.post(
        "/api/labor/material-dry-run",
        json={"root": str(tmp_path), "batchKey": "oss_2"},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["decision"] == "dry_run_only"
    assert body["writesRun"] is False
    assert body["aiInvoked"] is False
    assert len(app_module.list_labor_metadata()) == before_count


def test_labor_material_run_api_copies_reference_files_and_prefills_mapping(tmp_path):
    batch = tmp_path / "workforce已报账"
    batch.mkdir()
    pdf_path = batch / "Invoice-5058871.pdf"
    workbook_path = batch / "员工账单明细 - 2026-06-01T112149.990.xlsx"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    workbook_path.write_bytes(_excel_bytes())
    cache_dir = batch / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "Invoice-5058871_p1_mimo-v2.5_v4.json").write_text(
        json.dumps([{"employee_name_raw": "Alice Worker", "amount": 100, "source_page": 1}]),
        encoding="utf-8",
    )

    client = TestClient(app)
    response = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "workforce已报账",
            "periodStart": "2026-05-11",
            "periodEnd": "2026-05-17",
            "currency": "USD",
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "已确认字段"
    assert body["supplierName"] == "workforce"
    assert body["periodStart"] == "2026-05-11"
    assert body["periodEnd"] == "2026-05-17"
    assert body["workbookSheet"] == "员工账单"
    assert body["excelMapping"]["name"] == "姓名"
    assert body["excelMapping"]["amount"] == "费用总计(含税)"
    assert body["materialReplayNextStep"]["action"] == "extract_compare"
    assert body["materialReplayNextStep"]["enabled"] is True
    assert "可直接执行抽取核对" in body["materialReplayNextStep"]["description"]
    assert body["materialReplaySource"]["batchKey"] == "workforce已报账"
    assert body["materialReplaySource"]["uploadPlan"]["pdfFiles"] == ["workforce已报账/Invoice-5058871.pdf"]
    assert body["files"]["pdfInvoices"][0]["filename"].endswith(".pdf")
    assert body["files"]["workbooks"][0]["filename"].endswith(".xlsx")
    copied_pdf = Path(body["files"]["pdfInvoices"][0]["path"])
    copied_workbook = Path(body["files"]["workbooks"][0]["path"])
    assert copied_pdf.exists()
    assert copied_workbook.exists()
    assert copied_pdf != pdf_path
    assert copied_workbook != workbook_path
    assert pdf_path.exists()
    assert workbook_path.exists()
    copied_cache = copied_pdf.parent / ".ai_extract_cache" / f"{copied_pdf.stem}_p1_mimo-v2.5_v4.json"
    assert copied_cache.exists()
    assert "Alice Worker" in copied_cache.read_text(encoding="utf-8")


def test_labor_material_run_preserves_mapping_for_each_workbook(tmp_path):
    batch = tmp_path / "Sovitrat groupe"
    batch.mkdir()
    (batch / "invoice.pdf").write_bytes(b"%PDF-1.4\n")

    with_total = Workbook()
    sheet = with_total.active
    sheet.title = "Sheet2"
    sheet.append(["员工名称", "总计", "时薪", "本周薪资", "本周餐补", "周日补贴", "总额"])
    sheet.append(["Alice One", 8, 12.5, 100, 10, 0, 110])
    with_total.save(batch / "巴黎1号仓.xlsx")

    components = Workbook()
    sheet = components.active
    sheet.title = "Sheet2"
    sheet.append(["员工名称", "总计", "时薪", "本周薪资", "本周餐补", "周日补贴"])
    sheet.append(["Alice Two", 8, 12.5, 100, 10, 5])
    components.save(batch / "巴黎2号仓.xlsx")

    client = TestClient(app)
    response = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "Sovitrat_groupe",
            "periodStart": "2026-05-25",
            "periodEnd": "2026-05-31",
            "currency": "EUR",
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert len(body["workbookMappings"]) == 2
    rows = app_module._labor_excel_rows_from_metadata(body)
    assert {row.employee_name_raw: row.amount for row in rows} == {
        "Alice One": 110.0,
        "Alice Two": 115.0,
    }


def test_labor_material_run_extracts_and_replays_name_mapping_candidate(monkeypatch, tmp_path):
    import bonus_platform.app as app_module

    batch = tmp_path / "29仓"
    batch.mkdir()
    (batch / "In291943.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS040020", "Deisi Pozo", 37.84, 847.84, "USD", "29号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    (batch / "员工账单明细 - 2026-05-28T141945.414.xlsx").write_bytes(buffer.getvalue())

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 900.0, "warehouse_id": "29"}],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file=paths[0].name, source_page_or_row="p1", employee_id="", employee_name_raw="Rozo Panche, Deisy V", hours=37.84, amount=847.84, currency="USD", confidence=0.98, evidence_text="Rozo Panche, Deisy V 37.84 $847.84")
        ],
    )

    client = TestClient(app)
    run = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "29仓",
            "supplierName": "29仓",
            "periodStart": "2026-05-18",
            "periodEnd": "2026-05-24",
            "currency": "USD",
        },
    ).json()

    compared = app_module._perform_labor_extract_compare(run["id"])

    candidates = compared["nameMappingGovernance"]["candidates"]
    assert compared["comparisonSummary"]["candidateMatchCount"] == 1
    assert len(candidates) == 1
    summary = compared["nameMappingGovernance"]["summary"]
    assert summary["candidateCount"] == 1
    assert summary["readyToReplayCount"] == 1
    assert summary["projectedFixedExceptionCount"] == 2
    candidate = candidates[0]
    assert candidate["cacheEmployeeName"] == "Rozo Panche, Deisy V"
    assert candidate["excelEmployeeName"] == "Deisi Pozo"
    assert candidate["confidence"] == "high"
    assert candidate["decision"] == "candidate_only"
    assert candidate["projectedFixedExceptionCount"] == 2
    assert candidate["matchReason"] == "姓名相似且金额/工时一致"
    assert "是否确认 PDF 名称 Rozo Panche, Deisy V 对应 Excel 员工 Deisi Pozo" in candidate["businessQuestion"]
    assert candidate["impactSummary"] == "金额和工时均一致"
    assert "必须预览影响" in candidate["cannotAutoResolveReason"]

    replay = client.post(
        f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/auto-replay",
        json={"limit": 10},
    )

    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["decision"] == "ready_for_user_confirmation"
    assert replay_body["summary"]["fixedCount"] == 1
    assert replay_body["summary"]["regressionCount"] == 0

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "material run replay passed", "recalculate": True},
    )

    assert confirmed.status_code == 200
    confirmed_body = confirmed.json()
    assert confirmed_body["manualNameMapping"] == {"Rozo Panche, Deisy V": "Deisi Pozo"}
    assert confirmed_body["recalculatedRun"]["status"] == "部分核对完成"
    assert confirmed_body["recalculatedRun"]["comparisonSummary"]["canRelease"] is False
    assert confirmed_body["recalculatedRun"]["comparisonSummary"]["exceptionCount"] == 0
    assert confirmed_body["recalculatedRun"]["diffDownloadUrl"]
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["comparisonSummary"]["exceptionCount"] == 0
    assert refreshed["diffDownloadUrl"] == refreshed["files"]["diffReport"]["downloadUrl"]


def test_labor_material_run_name_mapping_candidate_with_amount_gap_requires_review(monkeypatch, tmp_path):
    import bonus_platform.app as app_module

    batch = tmp_path / "29仓"
    batch.mkdir()
    (batch / "In291943.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS033570", "Freddy Moran (MOR47K)", 40.48, 830.72, "USD", "29号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    (batch / "员工账单明细 - 2026-05-28T141945.414.xlsx").write_bytes(buffer.getvalue())

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 1042.43, "warehouse_id": "29"}],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file=paths[0].name, source_page_or_row="p1", employee_id="", employee_name_raw="Moran Treminio, Freddy", hours=40.48, amount=1042.43, currency="USD", confidence=0.98, evidence_text="Moran Treminio, Freddy 40.48 $1042.43")
        ],
    )

    client = TestClient(app)
    run = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "29仓",
            "supplierName": "29仓",
            "periodStart": "2026-05-18",
            "periodEnd": "2026-05-24",
            "currency": "USD",
        },
    ).json()

    compared = app_module._perform_labor_extract_compare(run["id"])

    candidates = compared["nameMappingGovernance"]["candidates"]
    assert compared["comparisonSummary"]["candidateMatchCount"] == 1
    assert compared["comparisonSummary"]["exceptionCount"] == 1
    assert len(candidates) == 1
    summary = compared["nameMappingGovernance"]["summary"]
    assert summary["candidateCount"] == 1
    assert summary["readyToReplayCount"] == 0
    assert summary["projectedFixedExceptionCount"] == 0
    assert summary["amountStillDifferentCount"] == 1
    candidate = candidates[0]
    assert candidate["cacheEmployeeName"] == "Moran Treminio, Freddy"
    assert candidate["excelEmployeeName"] == "Freddy Moran (MOR47K)"
    assert candidate["confidence"] == "medium"
    assert candidate["projectedFixedExceptionCount"] == 0
    assert candidate["matchReason"] == "姓名相似，但金额或工时仍需复核"
    assert "需先复核差异口径" in candidate["businessQuestion"]
    assert "PDF 高于 Excel" in candidate["impactSummary"]
    assert "不能直接确认匹配" in candidate["cannotAutoResolveReason"]

    replay = client.post(
        f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/auto-replay",
        json={"limit": 10},
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["decision"] != "ready_for_user_confirmation"
    assert replay_body["summary"]["fixedCount"] == 0

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "reviewed but amount still differs", "recalculate": True},
    )
    assert confirmed.status_code == 400


def test_labor_material_run_keeps_employee_detail_when_totals_pass(monkeypatch, tmp_path):
    import bonus_platform.app as app_module

    batch = tmp_path / "oss 2"
    batch.mkdir()
    (batch / "US Elogis Service #1 Invoice W.E 05.24.26.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS045753", "Manuel Lozano", 16.09, 361.42, "USD", "1号仓"])
    sheet.append(["WUS045746", "Massiel Castillo", 3.50, 78.40, "USD", "1号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    (batch / "员工账单明细 - 2026-06-04T094719.972.xlsx").write_bytes(buffer.getvalue())

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 439.82, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=paths[0].name,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Lozano, Manuel",
                hours=19.59,
                amount=439.82,
                currency="USD",
                confidence=0.95,
                evidence_text="Lozano, Manuel 19.50 0.09 439.82",
            )
        ],
    )

    client = TestClient(app)
    run = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "oss_2",
            "supplierName": "oss",
            "periodStart": "2026-05-18",
            "periodEnd": "2026-05-24",
            "currency": "USD",
        },
    ).json()

    compared = app_module._perform_labor_extract_compare(run["id"])

    assert compared["warehouseComparison"]["summary"]["totalPassed"] is True
    assert compared["comparisonSummary"]["candidateMatchCount"] == 1
    assert compared["comparisonSummary"]["exceptionCount"] == 2
    assert compared["candidateMatches"][0]["issueType"] == "combined_pdf_row"
    combined = compared["combinedRowGovernance"]
    assert combined["summary"]["candidateCount"] == 1
    assert combined["summary"]["amountImpactTotal"] == 78.4
    assert combined["candidates"][0]["status"] == "pending_invoice_review"
    assert any("总金额已通过" in issue for issue in compared["extractionQuality"]["issues"])
    assert compared["diffDownloadUrl"]
    report_response = client.get(compared["diffDownloadUrl"])
    assert report_response.status_code == 200
    assert report_response.content


def test_labor_material_run_surfaces_amount_rate_review_queue(monkeypatch, tmp_path):
    import bonus_platform.app as app_module

    batch = tmp_path / "Grande"
    batch.mkdir()
    (batch / "Grande_Invoice_WH1.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS000501", "Alice Worker", 40.0, 1000.0, "USD", "1号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    (batch / "员工账单明细 - 2026-06-01T112149.990.xlsx").write_bytes(buffer.getvalue())

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 1040.0, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=paths[0].name,
                source_page_or_row="p1",
                employee_id="WUS000501",
                employee_name_raw="Alice Worker",
                hours=40.0,
                amount=1040.0,
                currency="USD",
                confidence=0.98,
                evidence_text="Alice Worker 40.00 $1040.00",
            )
        ],
    )

    client = TestClient(app)
    run = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "Grande",
            "supplierName": "Grande",
            "periodStart": "2026-05-18",
            "periodEnd": "2026-05-24",
            "currency": "USD",
        },
    ).json()

    compared = app_module._perform_labor_extract_compare(run["id"])

    assert compared["comparisonSummary"]["exceptionCount"] == 1
    assert compared["comparisonRows"][0]["matchStatus"] == "金额差异"
    assert compared["reviewQueues"]["primary"] == "amount_rate_review"
    amount_queue = compared["reviewQueues"]["amountRateReview"]
    assert amount_queue["count"] == 1
    assert amount_queue["reviewMode"] == "amount_basis"
    assert amount_queue["businessQuestion"] == "工时已经对齐，金额差来自费率、加班、服务费还是税费口径？"
    assert amount_queue["rows"][0]["employeeName"] == "Alice Worker"
    assert amount_queue["rows"][0]["reviewFocus"] == "先核金额口径"
    assert amount_queue["rows"][0]["amountDirectionLabel"] == "PDF 高于 Excel"
    assert "确认前不能由系统自动改金额或清账" in amount_queue["cannotAutoResolveReason"]
    assert compared["diffDownloadUrl"]
    report_response = client.get(compared["diffDownloadUrl"])
    assert report_response.status_code == 200
    assert report_response.content
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["reviewQueues"]["primary"] == "amount_rate_review"
    assert refreshed["readinessGate"]["status"] == "blocked"
    assert refreshed["readinessGate"]["summary"]["exceptionCount"] == 1
    assert any(issue["code"] == "comparison_exceptions" for issue in refreshed["readinessGate"]["issues"])


def test_labor_material_run_blocks_when_employee_detail_recognition_is_incomplete(monkeypatch, tmp_path):
    import bonus_platform.app as app_module

    batch = tmp_path / "oss"
    batch.mkdir()
    (batch / "elog1-1_20260520204104.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    sheet.append(["WUS045751", "Massiel Castillo", 30.92, 100.00, "USD", "1号仓"])
    buffer = BytesIO()
    workbook.save(buffer)
    (batch / "员工账单明细 - 2026-05-27T110404.877.xlsx").write_bytes(buffer.getvalue())

    ai_cache_audit = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "summary": {"fileCount": 1, "candidateFileCount": 1, "candidateAmountTotal": 100.0},
        "files": [],
    }
    ai_cache_preview = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "summary": {
            "candidateRowCount": 1,
            "excelRowCount": 1,
            "passedCount": 0,
            "exceptionCount": 1,
            "cacheAmountTotal": 100,
            "excelAmountTotal": 100,
            "amountDeltaTotal": 0,
            "matchRate": 0,
            "reviewableFileCount": 0,
            "needsReocrFileCount": 1,
        },
        "fileQuality": [],
        "rows": [],
        "exceptionRows": [],
        "candidateMatches": [],
    }
    reocr_plan = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "summary": {"taskCount": 1, "reviewableCandidateCount": 0, "totalExpectedExcelAmount": 100.0, "totalCurrentCacheAmount": 100.0},
        "tasks": [
            {
                "sourceFile": "elog1-1_20260520204104.pdf",
                "warehouseId": "1",
                "expectedExcelAmount": 100.0,
                "amountDelta": 0.0,
                "diagnostics": {
                    "suspectedNamePairs": [
                        {
                            "cacheEmployeeName": "Espinosa Manuel",
                            "excelEmployeeName": "Massiel Castillo",
                            "amountGap": 0.0,
                            "hoursGap": -0.02,
                            "sourceRefs": "elog1 p1; workbook row 2",
                            "cacheAmount": 100,
                            "excelAmount": 100,
                            "cacheHours": 30.90,
                            "excelHours": 30.92,
                        }
                    ]
                },
            }
        ],
        "reviewableCandidates": [],
    }

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 0.0, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(app_module, "audit_ai_page_cache_candidates", lambda *args, **kwargs: ai_cache_audit)
    monkeypatch.setattr(app_module, "build_ai_cache_reconciliation_preview", lambda *args, **kwargs: ai_cache_preview)
    monkeypatch.setattr(app_module, "build_reocr_candidate_plan", lambda *args, **kwargs: reocr_plan)
    monkeypatch.setattr(app_module, "extract_invoice_items", lambda *args, **kwargs: (_ for _ in ()).throw(ValueError("AI 抽取失败：network unavailable")))

    client = TestClient(app)
    run = client.post(
        "/api/labor/material-runs",
        json={
            "root": str(tmp_path),
            "batchKey": "oss",
            "supplierName": "oss",
            "periodStart": "2026-05-11",
            "periodEnd": "2026-05-17",
            "currency": "USD",
        },
    ).json()

    updated = app_module._perform_labor_extract_compare(run["id"])

    assert updated["status"] == "PDF识别未完成"
    assert updated["files"]["diffReport"]["filename"].endswith(".xlsx")
    assert updated["comparisonSummary"]["exceptionCount"] == 1
    assert updated["extractionQuality"]["level"] == "critical"
    assert any("AI 抽取失败" in issue for issue in updated["extractionQuality"]["issues"])
    assert updated["aiCacheAudit"] == ai_cache_audit
    assert updated["aiCacheReconciliationPreview"] == ai_cache_preview
    assert updated["reocrPlan"] == reocr_plan
    candidate = updated["nameMappingGovernance"]["candidates"][0]
    assert candidate["cacheEmployeeName"] == "Espinosa Manuel"
    assert candidate["excelEmployeeName"] == "Massiel Castillo"
    assert candidate["decision"] == "candidate_only"
    assert candidate["requiresConfirmation"] is True
    assert updated["diffDownloadUrl"] == updated["files"]["diffReport"]["downloadUrl"]


def test_labor_compare_keeps_otws_warehouse_result_when_employee_amount_rows_are_unavailable(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["NJ13 Invoice Report WE 051726 JF.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    monkeypatch.setattr(
        app_module,
        "_labor_cost_summaries",
        lambda paths: [
            {
                "sourceFile": paths[0].name,
                "warehouseId": "1",
                "supplier": "Strategic Staffing Solutions Corp.",
                "employeeCount": 1,
                "summary": {
                    "reportedTotal": 105.0,
                    "componentTotal": 105.0,
                    "componentDelta": 0.0,
                    "evidence": "Warehouse-information!2",
                },
                "details": {
                    "employeeExpenses": {"amount": 100.0, "hours": 4.0},
                    "employeeBenefits": {"amount": 5.0, "hours": 0.0},
                    "detailTotal": 105.0,
                    "summaryDelta": 0.0,
                },
            }
        ],
    )
    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": paths[0].name,
                "total_amount": 110.0,
                "warehouse_id": "1",
                "authoritative": True,
                "evidence_status": "authoritative",
                "page_evidence": [
                    {"page": 1, "role": "invoice_total", "evidence_text": "Total $110.00"}
                ],
            }
        ],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            ValueError("AI 抽取失败：AI 图片抽取返回 0 条员工明细")
        ),
    )

    updated = app_module._perform_labor_extract_compare(run["id"])

    assert updated["status"] == "部分核对完成"
    assert updated["pdfExtractedRows"] == []
    assert updated["warehouseComparison"]["summary"]["pdfAmountTotal"] == 110.0
    assert updated["warehouseComparison"]["summary"]["excelAmountTotal"] == 105.0
    assert updated["warehouseComparison"]["rows"][0]["excelAmountTotal"] == 105.0
    assert updated["warehouseComparison"]["rows"][0]["amountDelta"] == 5.0
    assert updated["extractionQuality"]["level"] == "warning"
    assert any("班组汇总" in issue for issue in updated["extractionQuality"]["issues"])
    assert updated["requiresHumanReview"] is True
    assert updated["files"]["diffReport"]["filename"].endswith(".xlsx")


def test_labor_compare_keeps_runtime_profile_config_immutable_when_pdf_extraction_returns_no_rows(monkeypatch):
    import bonus_platform.app as app_module

    def _unexpected_profile_mutation(*args, **kwargs):
        raise AssertionError("正式核对不得在运行时改写全局供应商 Profile 配置")

    monkeypatch.setattr(app_module, "record_profile_failure", _unexpected_profile_mutation, raising=False)
    monkeypatch.setattr(app_module, "reset_profile_failure", _unexpected_profile_mutation, raising=False)

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": paths[0].name,
                "total_amount": 100.0,
                "warehouse_id": "",
                "authoritative": True,
                "evidence_status": "authoritative",
            }
        ],
    )
    monkeypatch.setattr(app_module, "extract_invoice_items", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {
                "fileCount": len(paths),
                "textReadableFileCount": len(paths),
                "imageOnlyFileCount": 0,
                "textReadablePageCount": len(paths),
                "emptyTextPageCount": 0,
                "imageOnlyPdfFiles": [],
            },
            "files": [
                {
                    "sourceFile": path.name,
                    "pageCount": 1,
                    "readablePageCount": 1,
                    "emptyTextPageCount": 0,
                    "hasTextLayer": True,
                    "needsOcr": False,
                    "diagnostic": "text_readable_pdf",
                }
                for path in paths
            ],
        },
    )
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    assert upload.status_code == 200
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )
    app_module.update_labor_metadata(run["id"], {"errorMessage": "上一次抽取失败", "errorCode": "OLD_ERROR", "nextAction": "请重试"})

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    assert response.json()["status"] == "抽取中"
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert body["status"] == "抽取失败"
    assert "PDF 未抽取出员工明细" in body["errorMessage"]


def test_labor_compare_blocks_uploaded_image_only_pdfs_without_employee_detail(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": path.name,
                "total_amount": 0.0,
                "warehouse_id": str(index),
                "authoritative": False,
                "evidence_status": "needs_review",
                "page_evidence": [{"page": 1, "role": "invoice_primary"}],
            }
            for index, path in enumerate(paths, start=1)
        ],
    )
    monkeypatch.setattr(app_module, "extract_invoice_items", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {
                "fileCount": len(paths),
                "textReadableFileCount": 0,
                "imageOnlyFileCount": len(paths),
                "textReadablePageCount": 0,
                "emptyTextPageCount": len(paths),
                "imageOnlyPdfFiles": [path.name for path in paths],
            },
            "files": [
                {
                    "sourceFile": path.name,
                    "pageCount": 1,
                    "readablePageCount": 0,
                    "emptyTextPageCount": 1,
                    "hasTextLayer": False,
                    "needsOcr": True,
                    "diagnostic": "image_only_pdf",
                }
                for path in paths
            ],
        },
    )
    monkeypatch.setattr(
        app_module,
        "build_reocr_candidate_plan",
        lambda *args, **kwargs: {
            "summary": {"taskCount": 1, "reviewableCandidateCount": 0},
            "tasks": [{"sourceFile": "scan.pdf", "warehouseId": "1", "amountDelta": -100, "expectedExcelAmount": 100}],
            "reviewableCandidates": [],
        },
    )
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    assert upload.status_code == 200
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    assert response.json()["status"] == "抽取中"
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert body["status"] == "PDF识别未完成"
    assert body["stage"] == "生成报告"
    assert body["errorMessage"] == ""
    assert body["errorCode"] == ""
    assert body["nextAction"] == ""
    assert body["reviewQueues"]["primary"] == "reocr"
    assert body["files"]["diffReport"]["label"] == "PDF识别诊断报告"
    assert any("未连接图片识别服务" in issue for issue in body["extractionQuality"]["issues"])


def test_labor_compare_attempts_auto_image_extraction_before_reocr_fallback(monkeypatch):
    import bonus_platform.app as app_module

    extracted_paths: list[str] = []

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": path.name,
                "total_amount": 0.0,
                "warehouse_id": str(index),
                "authoritative": False,
                "evidence_status": "needs_review",
                "page_evidence": [{"page": 1, "role": "invoice_primary"}],
            }
            for index, path in enumerate(paths, start=1)
        ],
    )
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {
                "fileCount": len(paths),
                "textReadableFileCount": 0,
                "imageOnlyFileCount": len(paths),
                "textReadablePageCount": 0,
                "emptyTextPageCount": len(paths),
                "imageOnlyPdfFiles": [path.name for path in paths],
            },
            "files": [
                {
                    "sourceFile": path.name,
                    "pageCount": 1,
                    "readablePageCount": 0,
                    "emptyTextPageCount": 1,
                    "hasTextLayer": False,
                    "needsOcr": True,
                    "diagnostic": "image_only_pdf",
                }
                for path in paths
            ],
        },
    )
    monkeypatch.setattr(
        app_module,
        "build_reocr_candidate_plan",
        lambda *args, **kwargs: {
            "summary": {"taskCount": 2, "reviewableCandidateCount": 0},
            "tasks": [
                {"sourceFile": "scan-a.pdf", "warehouseId": "1", "amountDelta": -701.90, "expectedExcelAmount": 701.90},
                {"sourceFile": "scan-b.pdf", "warehouseId": "2", "amountDelta": 0, "expectedExcelAmount": 0},
            ],
            "reviewableCandidates": [],
        },
    )

    def fake_extract(pdf_paths, *args, **kwargs):
        extracted_paths.extend(Path(path).name for path in pdf_paths)
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(pdf_paths[0]).name,
                source_page_or_row="p1",
                employee_id="WUS042586",
                employee_name_raw="Rosa Alvarez Minchaca",
                hours=31.19,
                amount=701.90,
                currency="USD",
                confidence=0.96,
                evidence_text="Rosa Alvarez Minchaca 31.19 $701.90",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan-a.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("scan-b.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    assert upload.status_code == 200
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    assert response.json()["status"] == "抽取中"
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert any(name.startswith("scan-a_") for name in extracted_paths)
    assert any(name.startswith("scan-b_") for name in extracted_paths)
    assert body["status"] == "PDF识别未完成"
    assert body["batchGuard"]["status"] == "pdf_recognition_incomplete"
    assert body["reviewQueues"]["primary"] == "employee_exceptions"
    assert body["comparisonSummary"]["exceptionCount"] == 0
    assert body["comparisonSummary"]["pdfEmployeeCount"] == 1
    assert body["files"]["diffReport"]["label"] == "PDF识别诊断报告"


def test_labor_retry_keeps_original_rows_when_retry_is_less_complete(monkeypatch, tmp_path):
    retry_pdf = tmp_path / "scan.pdf"
    retry_pdf.write_bytes(b"%PDF-1.4\n")

    def line(name: str, amount: float) -> LaborLineItem:
        return LaborLineItem(
            source_type="pdf_invoice",
            source_file=retry_pdf.name,
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw=name,
            hours=8,
            amount=amount,
            currency="USD",
            confidence=0.95,
            evidence_text=f"{name} 8 ${amount}",
        )

    original_rows = [line(f"Worker {index}", 100 + index) for index in range(10)]
    retry_rows = [line(f"Worker {index}", 100 + index) for index in range(6)]
    original_quality = {"level": "warning", "issues": ["总金额差异较大"], "retryAttempted": False, "retryApplied": False}
    original_comparison = {"summary": {"exceptionCount": 10, "amountDeltaTotal": 1000}, "rows": []}
    retry_comparison = {"summary": {"exceptionCount": 1, "amountDeltaTotal": 10}, "rows": []}

    monkeypatch.setattr(app_module, "extract_invoice_items", lambda *args, **kwargs: retry_rows)
    monkeypatch.setattr(app_module, "compare_labor_items", lambda *args, **kwargs: retry_comparison)
    monkeypatch.setattr(app_module, "calculate_extraction_quality", lambda *args, **kwargs: {"level": "ok", "issues": []})

    rows, comparison, quality = app_module._retry_if_better(
        [retry_pdf],
        original_rows,
        [],
        original_quality,
        original_comparison,
    )

    assert rows == original_rows
    assert comparison == original_comparison
    assert quality["retryAttempted"] is True
    assert quality["retryApplied"] is False


def test_labor_compare_skips_quality_retry_for_uploaded_image_only_pdfs(monkeypatch):
    import bonus_platform.app as app_module

    calls = {"extract": 0}

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {"source_file": Path(path).name, "total_amount": 700.00, "warehouse_id": ""}
            for path in paths
        ],
    )
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {
                "fileCount": len(paths),
                "textReadableFileCount": 0,
                "imageOnlyFileCount": len(paths),
                "textReadablePageCount": 0,
                "emptyTextPageCount": len(paths),
                "imageOnlyPdfFiles": [path.name for path in paths],
            },
            "files": [
                {
                    "sourceFile": path.name,
                    "pageCount": 1,
                    "readablePageCount": 0,
                    "emptyTextPageCount": 1,
                    "hasTextLayer": False,
                    "needsOcr": True,
                    "diagnostic": "image_only_pdf",
                }
                for path in paths
            ],
        },
    )
    monkeypatch.setattr(
        app_module,
        "build_reocr_candidate_plan",
        lambda *args, **kwargs: {"summary": {"taskCount": 1}, "tasks": [], "reviewableCandidates": []},
    )

    def fake_extract(pdf_paths, *args, **kwargs):
        calls["extract"] += 1
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(pdf_paths[0]).name,
                source_page_or_row="p1",
                employee_id="WUS042586",
                employee_name_raw="Rosa Alvarez Minchaca",
                hours=31.19,
                amount=701.90,
                currency="USD",
                confidence=0.96,
                evidence_text="Rosa Alvarez Minchaca 31.19 $701.90",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    monkeypatch.setattr(
        app_module,
        "calculate_extraction_quality",
        lambda *args, **kwargs: {
            "level": "warning",
            "message": "需要业务确认。",
            "issues": ["图片识别结果需要确认。"],
            "metrics": {},
            "lowConfidenceRows": [],
        },
    )
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    assert upload.status_code == 200
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert calls["extract"] == 1
    assert body["extractionQuality"]["retryAttempted"] is False
    assert body["extractionQuality"]["retryApplied"] is False


def test_labor_recover_stuck_run_marks_retryable_system_interruption(monkeypatch):
    import bonus_platform.app as app_module

    captured: dict[str, dict] = {}
    monkeypatch.setattr(app_module, "list_labor_metadata", lambda: [{"id": "labor_stuck", "status": "抽取中"}])
    monkeypatch.setattr(
        app_module,
        "update_labor_metadata",
        lambda run_id, updates: captured.setdefault(run_id, updates),
    )

    app_module._recover_stuck_labor_runs()

    updates = captured["labor_stuck"]
    assert updates["status"] == "抽取失败"
    assert updates["stage"] == "系统中断"
    assert updates["failureType"] == "system_interrupted"
    assert updates["errorCode"] == "LABOR_EXTRACT_INTERRUPTED"
    assert updates["retryable"] is True
    assert updates["requiresReupload"] is False
    assert updates["businessReviewStatus"] == "pending"
    assert updates["manualReviewRequired"] is True
    assert updates["directPaymentAllowed"] is False
    assert updates["requiresHumanReview"] is True
    assert "重新点击" in updates["nextAction"]
    assert "服务器已重启" in updates["errorMessage"]


def test_labor_recover_stuck_run_does_not_interrupt_personal_worker(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setenv("SIGMA_LABOR_EXECUTION_MODE", "personal-worker")
    monkeypatch.setattr(app_module, "list_labor_metadata", lambda: [{"id": "labor_active", "status": "抽取中"}])
    updates: list[tuple[str, dict]] = []
    monkeypatch.setattr(
        app_module,
        "update_labor_metadata",
        lambda run_id, payload: updates.append((run_id, payload)),
    )

    app_module._recover_stuck_labor_runs()

    assert updates == []


def test_labor_stale_extracting_check_accepts_timezone_aware_postgres_timestamp():
    from datetime import datetime, timezone

    metadata = {
        "id": "labor_timezone_aware",
        "status": "抽取中",
        "updatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }

    assert app_module._check_stale_extracting(metadata) is metadata


def test_labor_compare_formal_scope_extracts_every_payable_pdf_when_warehouse_cannot_map(monkeypatch):
    import bonus_platform.app as app_module

    captured_paths = []

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [{"source_file": paths[0].name, "total_amount": 50, "warehouse_id": ""}],
    )
    monkeypatch.setattr(app_module, "_warehouse_id_from_text_path", lambda *args, **kwargs: False)

    def fake_extract(pdf_paths, *args, **kwargs):
        captured_paths.extend([Path(p).name for p in pdf_paths])
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(pdf_paths[0]).name,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=100,
                currency="USD",
                confidence=0.95,
                evidence_text="Total $100",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Invoice", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("Invoice-5058871.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("Invoice-5058872.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert body["status"] == "部分核对完成"
    assert body["batchGuard"]["status"] == "partial_review"
    assert any(name.startswith("Invoice-5058871_") for name in captured_paths)
    assert any(name.startswith("Invoice-5058872_") for name in captured_paths)


def test_labor_rule_extraction_reads_wage_code_invoice_rows():
    from bonus_platform.engine.labor.extract import _extract_wage_code_invoice_rows

    rows = _extract_wage_code_invoice_rows(
        {
            "source_file": "Invoice-5058871.pdf",
            "page": 1,
            "text": """
Reference
Employee
Wage Code
Type
Hours
Rate
Amount
Aguilar, Hortensia
Reg
REG
40.00
22.58
$903.20
Aguilar, Hortensia
Reg
OT
1.40
33.86
$47.40
Customer ID
Invoice Number
5058871
""",
        },
        supplier="Invoice",
        period_start="2026-05-17",
        period_end="2026-05-22",
        currency="USD",
    )

    assert len(rows) == 2
    assert rows[0].employee_name_raw == "Aguilar, Hortensia"
    assert rows[0].hours == 40
    assert rows[0].amount == 903.20
    assert rows[1].hours == 1.4
    assert rows[1].amount == 47.40


def test_labor_rule_extraction_combines_wrapped_wage_code_names():
    from bonus_platform.engine.labor.extract import _extract_wage_code_invoice_rows

    rows = _extract_wage_code_invoice_rows(
        {
            "source_file": "Invoice-5058877.pdf",
            "page": 1,
            "text": """
ROSEL DUARTE, YAIR
GUILLERMO
Reg
REG
40.00
22.58
$903.20
ROSEL DUARTE, YAIR
GUILLERMO
Reg
OT
0.97
33.86
$32.84
""",
        },
        supplier="Invoice",
        period_start="2026-05-17",
        period_end="2026-05-22",
        currency="USD",
    )

    assert len(rows) == 2
    assert {row.employee_name_raw for row in rows} == {"ROSEL DUARTE, YAIR GUILLERMO"}
    assert round(sum(row.amount for row in rows), 2) == 936.04


def test_labor_vertical_invoice_rows_include_ca_penalty_payable_lines():
    from bonus_platform.engine.labor.extract import _extract_vertical_invoice_rows

    rows = _extract_vertical_invoice_rows(
        {
            "source_file": "US_ELogistics_Service_Corp__35362.pdf",
            "page": 1,
            "text": """
6/14/2026
Duenas, Oscar
1.00
CAPenalty
REG
$20.00
26.00
$26.00
6/14/2026
Duenas, Oscar
40.00
Reg
REG
$20.00
26.00
$1,040.00
6/14/2026
Duenas, Oscar
6.46
OT
OT
$30.00
39.00
$251.94
6/14/2026
Duenas, Oscar
1.72
Reg
DT
$40.00
52.00
$89.44
6/14/2026
Duenas, Oscar
0.50
Meal Premium
REG
$20.00
26.00
$13.00
""",
        },
        supplier="OSI",
        period_start="2026-06-08",
        period_end="2026-06-14",
        currency="USD",
    )

    assert len(rows) == 5
    assert [row.amount for row in rows] == [26.00, 1040.00, 251.94, 89.44, 13.00]
    assert round(sum(row.amount for row in rows), 2) == 1420.38
    assert round(sum(row.hours for row in rows), 2) == 49.68


def test_labor_compare_response_includes_candidate_matches(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": paths[0].name,
                "total_amount": 50.0,
                "warehouse_id": "",
                "authoritative": True,
                "evidence_status": "authoritative",
            }
        ],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file="scan.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alvarez Mitrache, Ross", hours=30.5, amount=698.99, currency="USD", confidence=0.95, evidence_text="Total $698.99")
        ],
    )
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert "candidateMatches" in body
    assert isinstance(body["candidateMatches"], list)


def test_labor_compare_persists_diagnostics_and_ai_cache_audit_in_report_flow(monkeypatch):
    import bonus_platform.app as app_module

    captured_report_kwargs = {}
    cost_summaries = [
        {
            "sourceFile": "otws.xlsx",
            "warehouseId": "1",
            "summary": {"reportedTotal": 100.0, "componentTotal": 100.0, "componentDelta": 0.0},
            "details": {"detailTotal": 100.0, "summaryDelta": 0.0},
        }
    ]
    ai_cache_audit = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "summary": {"fileCount": 1, "candidateFileCount": 1, "candidateAmountTotal": 100.0},
        "files": [{"sourceFile": "invoice.pdf", "rowCount": 1, "candidateAmountTotal": 100.0}],
    }
    ai_cache_preview = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "summary": {"needsReocrFileCount": 1, "reviewableFileCount": 0},
        "fileQuality": [
            {
                "sourceFile": "invoice.pdf",
                "warehouseId": "1",
                "decision": "needs_reocr",
                "cacheAmountTotal": 80.0,
                "excelAmountTotal": 100.0,
                "amountDelta": -20.0,
                "excelRowCount": 1,
                "recommendation": "历史识别金额与账单不一致，建议重新识别后预览影响。",
            }
        ],
    }
    reocr_plan = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "summary": {"taskCount": 1, "reviewableCandidateCount": 0, "totalExpectedExcelAmount": 100.0, "totalCurrentCacheAmount": 80.0},
        "tasks": [{"sourceFile": "invoice.pdf", "warehouseId": "1", "expectedExcelAmount": 100.0, "amountDelta": -20.0}],
        "reviewableCandidates": [],
    }

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": Path(paths[0]).name,
                "total_amount": 100,
                "warehouse_id": "1",
            }
        ],
    )
    monkeypatch.setattr(app_module, "_labor_cost_summaries", lambda *args, **kwargs: cost_summaries)
    monkeypatch.setattr(app_module, "audit_ai_page_cache_candidates", lambda *args, **kwargs: ai_cache_audit)
    monkeypatch.setattr(app_module, "build_ai_cache_reconciliation_preview", lambda *args, **kwargs: ai_cache_preview)
    monkeypatch.setattr(app_module, "build_reocr_candidate_plan", lambda *args, **kwargs: reocr_plan)
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda paths, *args, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(paths[0]).name,
                source_page_or_row="p1",
                employee_id="WUS000001",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=100,
                currency="USD",
                confidence=0.99,
                warehouse_id="1",
            )
        ],
    )

    def fake_build_labor_report(*args, **kwargs):
        report_path = Path(args[0])
        report_path.write_bytes(b"report")
        captured_report_kwargs.update(kwargs)

    monkeypatch.setattr(app_module, "build_labor_report", fake_build_labor_report)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "SSS", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    updated = app_module._perform_labor_extract_compare(run["id"])

    assert updated["status"] == "已生成差异报告"
    assert updated["presentation"]["schemaVersion"] == 1
    assert updated["presentation"]["summary"]["employeeCount"] == 1
    assert updated["presentation"]["summary"]["differenceEmployeeCount"] == 0
    assert updated["presentation"]["summary"]["reviewItemCount"] == 0
    assert updated["presentation"]["summary"]["excelRecordCount"] == 1
    assert captured_report_kwargs["presentation"] == updated["presentation"]
    assert updated["batchGuard"]["status"] == "ok"
    assert updated["machineCheckStatus"] == "passed"
    assert updated["businessReviewStatus"] == "pending"
    assert updated["requiresHumanReview"] is True
    assert updated["directPaymentAllowed"] is False
    assert updated["resultInputFingerprint"] == app_module._labor_result_input_fingerprint(updated)
    business_report = updated["files"]["businessReport"]
    assert business_report["filename"].endswith(".html")
    assert business_report["label"] == "业务核对报告"
    assert updated["businessReportDownloadUrl"] == business_report["downloadUrl"]
    business_report_response = client.get(business_report["downloadUrl"])
    assert business_report_response.status_code == 200
    business_report_html = business_report_response.text
    assert "核对结论" in business_report_html
    assert "供应商：SSS" in business_report_html
    assert "核算周期：2026-05-11 ~ 2026-05-17" in business_report_html
    assert "发票编号或文件范围：invoice.pdf" in business_report_html
    for internal_term in ["AI 候选", "规则治理", "profile", "re-OCR", "回放", "Blob", "线程"]:
        assert internal_term not in business_report_html
    assert updated["costSummaries"] == cost_summaries
    assert updated["aiCacheAudit"] == ai_cache_audit
    assert updated["aiCacheReconciliationPreview"] == ai_cache_preview
    assert updated["reocrPlan"] == reocr_plan
    assert updated["reconciliationDiagnostics"]["signals"]["amountBasis"][0]["warehouseId"] == "1"
    assert captured_report_kwargs["reconciliation_diagnostics"] == updated["reconciliationDiagnostics"]
    assert captured_report_kwargs["ai_cache_audit"] == ai_cache_audit


def test_labor_compare_creates_profile_candidate_without_saving_profile(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "quick_extract_totals", lambda paths, *args, **kwargs: _unresolved_quick_totals(paths))
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="WUS042586", employee_name_raw="Rosa Alvarez Minchaca", hours=31.19, amount=701.90, currency="USD", confidence=0.95, evidence_text="Rosa Alvarez Minchaca 31.19 $701.90")
        ],
    )
    monkeypatch.setattr(
        app_module,
        "generate_profile_from_extraction",
        lambda **kwargs: {
            "key": "onesource",
            "aliases": ["onesource"],
            "prompt_notes": ["Use candidate profile only after confirmation."],
            "image_page_policy": "first_page_only",
            "version": 1,
        },
    )

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    updated = app_module._perform_labor_extract_compare(run["id"])

    candidates = updated["profileGovernance"]["candidates"]
    assert len(candidates) == 1
    assert candidates[0]["decision"] == "candidate_only"
    assert candidates[0]["requiresConfirmation"] is True
    assert candidates[0]["profileData"]["key"] == "onesource"
    assert candidates[0]["evidence"][0]["sourcePageOrRow"] == "p1"


def test_labor_profile_candidate_confirm_rejects_auto_generated_draft():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Draft Workforce", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    candidate = app_module._build_profile_candidate(
        run["id"],
        "Draft Workforce",
        {
            "key": "draft-workforce",
            "aliases": ["draft workforce"],
            "version": 1,
            "status": "draft",
            "created_from": "auto_generation",
        },
        [],
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "profileGovernance": {
                "candidates": [candidate],
                "replaySummaries": {
                    candidate["candidateId"]: {
                        "decision": "ready_for_user_confirmation",
                        "summary": {"compatibleCount": 1, "regressionCount": 0},
                    }
                },
                "activeProfiles": [],
                "rolledBackProfiles": [],
            }
        },
    )

    response = client.post(
        f"/api/labor/runs/{run['id']}/profile-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "metadata replay only"},
    )

    assert response.status_code == 400
    assert "draft" in response.json()["detail"].lower()


def test_labor_profile_candidate_api_confirms_and_rolls_back():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Workforce", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    candidate = app_module._build_profile_candidate(
        run["id"],
        "Workforce",
        {
            "key": "workforce",
            "aliases": ["workforce"],
            "prompt_notes": ["Extract wage code rows."],
            "image_page_policy": "first_page_only",
            "version": 1,
            "status": "approved",
            "approvedBy": "payroll-admin@example.com",
            "approvedAt": "2026-07-15T09:30:00+08:00",
            "created_from": "manual_review",
        },
        [
            LaborLineItem(source_type="pdf_invoice", source_file="Invoice-5058871.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=0.95, evidence_text="Alice Worker 8 $100")
        ],
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "profileGovernance": {
                "candidates": [candidate],
                "replaySummaries": {},
                "activeProfiles": [],
                "rolledBackProfiles": [],
            }
        },
    )

    blocked = client.post(
        f"/api/labor/runs/{run['id']}/profile-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "Profile evidence reviewed"},
    )

    assert blocked.status_code == 400
    assert "缺少历史回放摘要" in blocked.json()["detail"]

    monkeypatch_runs = [
        {
            "id": run["id"],
            "supplierName": "Workforce",
            "periodStart": "2026-05-18",
            "periodEnd": "2026-05-24",
            "reconciliationDiagnostics": {"level": "ok", "issues": []},
            "extractionQuality": {"level": "ok", "issues": []},
            "comparisonSummary": {
                "exceptionCount": 0,
                "conclusionLevel": "pass",
                "canRelease": True,
                "machineCheckStatus": "passed",
            },
            "machineCheckStatus": "passed",
            "batchGuard": {"status": "ok", "allowReleasableReport": True},
            "reconciliationDiagnostics": {"level": "ok", "issues": []},
            "extractionQuality": {"level": "ok", "issues": []},
        }
    ]
    original_list = app_module.list_labor_metadata
    app_module.list_labor_metadata = lambda: monkeypatch_runs
    try:
        replay = client.post(
            f"/api/labor/runs/{run['id']}/profile-candidates/{candidate['candidateId']}/auto-replay",
            json={"limit": 10},
        )
    finally:
        app_module.list_labor_metadata = original_list

    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["decision"] == "ready_for_user_confirmation"
    assert replay_body["summary"]["compatibleCount"] == 1
    assert replay_body["preflight"]["blockingAfterApply"] is False
    assert replay_body["preflight"]["delta"]["compatibleCount"] == 1
    assert replay_body["preflight"]["affectedScopeCount"] == 1
    assert replay_body["preflight"]["affectedSuppliers"] == ["Workforce"]
    assert "prompt_notes" in replay_body["preflight"]["changedFields"]

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/profile-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "Profile evidence reviewed"},
    )

    assert confirmed.status_code == 200
    assert confirmed.json()["decision"] == "active"
    assert confirmed.json()["requiresConfirmation"] is False
    assert confirmed.json()["confirmedBy"] == "ops-user"
    assert confirmed.json()["preflight"] == replay_body["preflight"]

    rolled_back = client.post(
        f"/api/labor/runs/{run['id']}/profile-candidates/{candidate['candidateId']}/rollback",
        json={"rolledBackBy": "ops-user", "reason": "Profile regression", "targetVersion": 0},
    )

    assert rolled_back.status_code == 200
    assert rolled_back.json()["decision"] == "rolled_back"
    assert rolled_back.json()["rollbackToVersion"] == 0
    governance = client.get(f"/api/labor/runs/{run['id']}").json()["profileGovernance"]
    assert governance["activeProfiles"] == []
    assert governance["rolledBackProfiles"][0]["candidateId"] == candidate["candidateId"]


def test_labor_name_mapping_candidate_confirm_and_rollback_updates_manual_mapping():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    candidate = {
        "candidateId": "name_map_sample",
        "decision": "candidate_only",
        "status": "pending_user_confirmation",
        "requiresConfirmation": True,
        "sourceFile": "elog27-1_20260520204231.pdf",
        "warehouseId": "27",
        "cacheEmployeeName": "Coria, Virgilio",
        "excelEmployeeName": "Brayan Gomez Vargas",
        "proposedMapping": {"Coria, Virgilio": "Brayan Gomez Vargas"},
        "recommendation": "金额/工时接近，优先人工确认是否为同一员工姓名映射。",
        "auditTrail": [{"action": "created", "actor": "system", "reason": "reocr_suspected_name_pair"}],
    }
    app_module.update_labor_metadata(
        run["id"],
        {
            "manualNameMapping": {},
            "nameMappingGovernance": {
                "candidates": [candidate],
                "replaySummaries": {
                    candidate["candidateId"]: {
                        "decision": "ready_for_user_confirmation",
                        "summary": {"fixedCount": 1, "regressionCount": 0},
                    }
                },
                "activeMappings": [],
                "rolledBackMappings": [],
            },
        },
    )

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "OSS27 amount and hours align"},
    )

    assert confirmed.status_code == 200
    confirmed_body = confirmed.json()
    assert confirmed_body["decision"] == "active"
    assert confirmed_body["manualNameMapping"] == {"Coria, Virgilio": "Brayan Gomez Vargas"}
    after_confirm = client.get(f"/api/labor/runs/{run['id']}").json()
    assert after_confirm["manualNameMapping"]["Coria, Virgilio"] == "Brayan Gomez Vargas"
    assert after_confirm["nameMappingGovernance"]["activeMappings"][0]["confirmedBy"] == "ops-user"

    rolled_back = client.post(
        f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/rollback",
        json={"rolledBackBy": "ops-user", "reason": "Later batch showed different employee"},
    )

    assert rolled_back.status_code == 200
    rolled_back_body = rolled_back.json()
    assert rolled_back_body["decision"] == "rolled_back"
    assert rolled_back_body["manualNameMapping"] == {}
    after_rollback = client.get(f"/api/labor/runs/{run['id']}").json()
    assert after_rollback["manualNameMapping"] == {}
    assert after_rollback["nameMappingGovernance"]["activeMappings"] == []
    assert after_rollback["nameMappingGovernance"]["candidates"][0]["status"] == "rolled_back"
    assert after_rollback["nameMappingGovernance"]["rolledBackMappings"][0]["candidateId"] == candidate["candidateId"]


def test_labor_allocation_governance_builds_candidates_from_warehouse_offsets():
    import bonus_platform.app as app_module

    governance = app_module._build_allocation_governance(
        "labor_allocation_sample",
        {
            "allocationIssues": [
                {
                    "employeeKey": "id:WUS043938",
                    "employeeName": "JIMENEZ, ENEAS",
                    "netAmountDelta": -0.01,
                    "warehouseCount": 2,
                    "warehouses": [
                        {"warehouseId": "25", "pdfAmount": 118.04, "excelAmount": 116.85, "amountDelta": 1.19},
                        {"warehouseId": "28", "pdfAmount": 928.67, "excelAmount": 929.87, "amountDelta": -1.2},
                    ],
                    "recommendation": "员工总额可抵消，但仓库归属金额不一致，需按仓库复核发票与账单归属。",
                }
            ]
        },
    )

    candidate = governance["candidates"][0]
    assert candidate["decision"] == "candidate_only"
    assert candidate["status"] == "pending_user_confirmation"
    assert candidate["requiresConfirmation"] is True
    assert candidate["issueType"] == "cross_warehouse_employee_allocation"
    assert candidate["employeeName"] == "JIMENEZ, ENEAS"
    assert candidate["warehouseCount"] == 2
    assert candidate["warehouses"][0]["warehouseId"] == "25"
    assert "不自动修改" in candidate["confirmationGate"]


def test_labor_excel_aggregation_preserves_cross_warehouse_employee_rows():
    import bonus_platform.app as app_module
    from bonus_platform.engine.labor.models import LaborLineItem

    rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="WUS041037", employee_name_raw="PEREZ, JOSE", hours=4.0, amount=100.67, currency="USD", confidence=1, evidence_text="", warehouse_id="25"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="WUS041037", employee_name_raw="PEREZ, JOSE", hours=40.0, amount=935.59, currency="USD", confidence=1, evidence_text="", warehouse_id="28"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!4", employee_id="WUS041037", employee_name_raw="PEREZ, JOSE", hours=1.0, amount=10.0, currency="USD", confidence=1, evidence_text="", warehouse_id="25"),
    ]

    aggregated = app_module._aggregate_excel_rows(rows)

    by_warehouse = {row.warehouse_id: row for row in aggregated}
    assert len(aggregated) == 2
    assert by_warehouse["25"].amount == 110.67
    assert by_warehouse["25"].hours == 5.0
    assert by_warehouse["28"].amount == 935.59
    assert by_warehouse["28"].hours == 40.0


def test_labor_allocation_candidate_confirm_and_rollback_updates_readiness(tmp_path):
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "fairway", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    candidate = {
        "candidateId": "allocation_sample",
        "decision": "candidate_only",
        "status": "pending_user_confirmation",
        "requiresConfirmation": True,
        "issueType": "cross_warehouse_employee_allocation",
        "employeeKey": "id:WUS041037",
        "employeeName": "PEREZ, JOSE",
        "netAmountDelta": 0.0,
        "warehouseCount": 2,
        "warehouses": [
            {"warehouseId": "25", "pdfAmount": 101.26, "excelAmount": 100.67, "amountDelta": 0.59},
            {"warehouseId": "28", "pdfAmount": 935.0, "excelAmount": 935.59, "amountDelta": -0.59},
        ],
        "recommendation": "员工总额可抵消，但仓库归属金额不一致，需按仓库复核发票与账单归属。",
        "auditTrail": [{"action": "created", "actor": "system", "reason": "cross_warehouse_employee_allocation_detected"}],
    }
    report_path = tmp_path / "report.xlsx"
    report_path.write_bytes(b"allocation-report")
    report = app_module.attach_labor_file(run["id"], report_path, "差异报告")
    baseline = {
            "status": "已生成差异报告",
            "comparisonSummary": {
                "exceptionCount": 0,
                "conclusionLevel": "pass",
                "canRelease": True,
                "machineCheckStatus": "passed",
                "pdfEmployeeCount": 1,
                "excelEmployeeCount": 1,
            },
            "comparisonRows": [{"employeeName": "Synthetic Worker", "matchStatus": "通过"}],
            "machineCheckStatus": "passed",
            "batchGuard": {"status": "ok", "allowReleasableReport": True},
            "reconciliationDiagnostics": {"level": "ok", "issues": []},
            "extractionQuality": {"level": "ok", "issues": []},
            "files": {"diffReport": report},
            "diffDownloadUrl": report["downloadUrl"],
            "allocationGovernance": {
                "candidates": [candidate],
                "activeAllocations": [],
                "rolledBackAllocations": [],
            },
        }
    baseline["resultInputFingerprint"] = app_module._labor_result_input_fingerprint(
        {**run, **baseline}
    )
    app_module.update_labor_metadata(run["id"], baseline)

    before = client.get(f"/api/labor/runs/{run['id']}").json()
    assert before["readinessGate"]["summary"]["pendingGovernanceCount"] == 1
    assert before["readinessGate"]["status"] == "needs_review"

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/allocation-candidates/{candidate['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "warehouse allocation reviewed", "decisionNote": "accepted as allocation split"},
    )

    assert confirmed.status_code == 200
    confirmed_body = confirmed.json()
    assert confirmed_body["decision"] == "confirmed"
    assert confirmed_body["requiresConfirmation"] is False
    assert confirmed_body["readinessGate"]["summary"]["pendingGovernanceCount"] == 0
    assert confirmed_body["readinessGate"]["status"] == "blocked"
    assert any(issue["code"] == "stale_result_inputs" for issue in confirmed_body["readinessGate"]["issues"])
    after_confirm = client.get(f"/api/labor/runs/{run['id']}").json()
    assert after_confirm["allocationGovernance"]["activeAllocations"][0]["confirmedBy"] == "ops-user"
    assert after_confirm["allocationGovernance"]["candidates"][0]["status"] == "confirmed"
    assert after_confirm["comparisonSummary"]["exceptionCount"] == 0

    rolled_back = client.post(
        f"/api/labor/runs/{run['id']}/allocation-candidates/{candidate['candidateId']}/rollback",
        json={"rolledBackBy": "ops-user", "reason": "review note was wrong"},
    )

    assert rolled_back.status_code == 200
    rolled_back_body = rolled_back.json()
    assert rolled_back_body["decision"] == "rolled_back"
    assert rolled_back_body["readinessGate"]["summary"]["pendingGovernanceCount"] == 0
    after_rollback = client.get(f"/api/labor/runs/{run['id']}").json()
    assert after_rollback["allocationGovernance"]["activeAllocations"] == []
    assert after_rollback["allocationGovernance"]["rolledBackAllocations"][0]["candidateId"] == candidate["candidateId"]
    assert after_rollback["allocationGovernance"]["candidates"][0]["status"] == "rolled_back"


def test_labor_name_mapping_candidate_requires_and_records_impact_replay(monkeypatch):
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    candidate = {
        "candidateId": "name_map_replay_sample",
        "decision": "candidate_only",
        "status": "pending_user_confirmation",
        "requiresConfirmation": True,
        "sourceFile": "elog27-1_20260520204231.pdf",
        "warehouseId": "27",
        "cacheEmployeeName": "Coria, Virgilio",
        "excelEmployeeName": "Brayan Gomez Vargas",
        "proposedMapping": {"Coria, Virgilio": "Brayan Gomez Vargas"},
        "recommendation": "金额/工时接近，优先人工确认是否为同一员工姓名映射。",
    }
    pdf_row = LaborLineItem(
        source_type="pdf_invoice",
        source_file="elog27-1_20260520204231.pdf",
        source_page_or_row="p1",
        employee_id="",
        employee_name_raw="Coria, Virgilio",
        hours=14.17,
        amount=353.69,
        currency="USD",
        confidence=0.95,
        evidence_text="Coria, Virgilio 14.17 $353.69",
        warehouse_id="27",
    )
    excel_row = LaborLineItem(
        source_type="offline_workbook",
        source_file="账单.xlsx",
        source_page_or_row="员工账单!2",
        employee_id="",
        employee_name_raw="Brayan Gomez Vargas",
        hours=14.17,
        amount=353.69,
        currency="USD",
        confidence=1.0,
        warehouse_id="27",
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "manualNameMapping": {},
            "pdfExtractedRows": [pdf_row.to_dict()],
            "excelRows": [excel_row.to_dict()],
            "nameMappingGovernance": {
                "candidates": [candidate],
                "replaySummaries": {},
                "activeMappings": [],
                "rolledBackMappings": [],
            },
        },
    )
    historical_run = {
        "id": "historical_oss_27",
        "supplierName": "OSS",
        "periodStart": "2026-05-11",
        "periodEnd": "2026-05-17",
        "manualNameMapping": {},
        "pdfExtractedRows": [pdf_row.to_dict()],
        "excelRows": [excel_row.to_dict()],
    }
    insufficient_run = {
        "id": "historical_oss_missing_detail",
        "supplierName": "OSS",
        "periodStart": "2026-05-04",
        "periodEnd": "2026-05-10",
        "manualNameMapping": {},
    }
    monkeypatch_runs = [client.get(f"/api/labor/runs/{run['id']}").json(), historical_run, insufficient_run]
    original_list = app_module.list_labor_metadata
    app_module.list_labor_metadata = lambda: monkeypatch_runs

    try:
        blocked = client.post(
            f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/confirm",
            json={"confirmedBy": "ops-user", "reason": "reviewed"},
        )
        assert blocked.status_code == 400
        assert "影响回放摘要" in blocked.json()["detail"]

        replay = client.post(
            f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/auto-replay",
            json={"limit": 10},
        )
        assert replay.status_code == 200
        replay_body = replay.json()
        assert replay_body["decision"] == "ready_for_user_confirmation"
        assert replay_body["mode"] == "current_and_historical_name_mapping_replay"
        assert replay_body["summary"]["fixedCount"] == 1
        assert replay_body["summary"]["regressionCount"] == 0
        assert replay_body["summary"]["historicalCheckedCount"] == 1
        assert replay_body["summary"]["historicalInsufficientCount"] == 1

        confirmed = client.post(
            f"/api/labor/runs/{run['id']}/name-mapping-candidates/{candidate['candidateId']}/confirm",
            json={"confirmedBy": "ops-user", "reason": "impact replay fixed the unmatched pair"},
        )
    finally:
        app_module.list_labor_metadata = original_list
    assert confirmed.status_code == 200
    assert confirmed.json()["manualNameMapping"] == {"Coria, Virgilio": "Brayan Gomez Vargas"}
    assert confirmed.json()["replaySummary"]["fixedCount"] == 1


def test_labor_name_mapping_candidates_are_built_from_reocr_suspected_pairs():
    import bonus_platform.app as app_module

    candidates = app_module._build_name_mapping_candidates_from_reocr_plan(
        "labor_sample",
        {
            "tasks": [
                {
                    "sourceFile": "elog27-1_20260520204231.pdf",
                    "warehouseId": "27",
                    "diagnostics": {
                        "suspectedNamePairs": [
                            {
                                "cacheEmployeeName": "Coria, Virgilio",
                                "excelEmployeeName": "Brayan Gomez Vargas",
                                "amountGap": -0.01,
                                "hoursGap": 0.3,
                                "sourceRefs": "elog27 p1; workbook row 34",
                                "cacheAmount": 353.68,
                                "excelAmount": 353.69,
                                "cacheHours": 14.47,
                                "excelHours": 14.17,
                            }
                        ]
                    },
                }
            ]
        },
    )

    assert len(candidates) == 1
    candidate = candidates[0]
    assert candidate["decision"] == "candidate_only"
    assert candidate["requiresConfirmation"] is True
    assert candidate["sourceFile"] == "elog27-1_20260520204231.pdf"
    assert candidate["warehouseId"] == "27"
    assert candidate["proposedMapping"] == {"Coria, Virgilio": "Brayan Gomez Vargas"}
    assert candidate["evidence"]["sourceRefs"] == "elog27 p1; workbook row 34"


def test_labor_name_mapping_candidates_are_built_from_candidate_matches():
    import bonus_platform.app as app_module

    candidates = app_module._build_name_mapping_candidates_from_candidate_matches(
        "labor_29_sample",
        [
            {
                "pdfEmployeeName": "Rozo Panche, Deisy V",
                "excelEmployeeName": "Deisi Pozo",
                "pdfAmountTotal": 847.84,
                "excelAmountTotal": 847.84,
                "pdfHoursTotal": 37.84,
                "excelHoursTotal": 37.84,
                "amountDelta": 0,
                "hoursDelta": 0,
                "nameSimilarity": 0.4,
                "sourceRefs": "In291943.pdf p1; 账单.xlsx 员工账单明细!3",
                "recommendation": "人工复核",
            },
            {
                "pdfEmployeeName": "Moran Treminio, Freddy",
                "excelEmployeeName": "Freddy Moran (MOR47K)",
                "pdfAmountTotal": 1042.43,
                "excelAmountTotal": 830.72,
                "pdfHoursTotal": 40.48,
                "excelHoursTotal": 40.48,
                "amountDelta": 211.71,
                "hoursDelta": 0,
                "nameSimilarity": 0.52,
                "sourceRefs": "In291943.pdf p1; 账单.xlsx 员工账单明细!4",
                "recommendation": "人工复核",
            }
        ],
    )

    assert len(candidates) == 2
    candidate = candidates[0]
    assert candidate["decision"] == "candidate_only"
    assert candidate["status"] == "pending_user_confirmation"
    assert candidate["requiresConfirmation"] is True
    assert candidate["sourceFile"] == "In291943.pdf"
    assert candidate["cacheEmployeeName"] == "Rozo Panche, Deisy V"
    assert candidate["excelEmployeeName"] == "Deisi Pozo"
    assert candidate["proposedMapping"] == {"Rozo Panche, Deisy V": "Deisi Pozo"}
    assert candidate["confidence"] == "high"
    assert candidate["amountGap"] == 0
    assert candidate["hoursGap"] == 0
    assert candidate["projectedFixedExceptionCount"] == 2
    assert candidate["matchReason"] == "姓名相似且金额/工时一致"
    assert "确认后预计减少 2 项异常" in candidate["businessQuestion"]
    assert candidate["impactSummary"] == "金额和工时均一致"
    assert "必须预览影响" in candidate["cannotAutoResolveReason"]
    assert candidate["evidence"]["nameSimilarity"] == 0.4
    assert candidate["auditTrail"][0]["reason"] == "candidate_match_name_pair"
    medium_candidate = candidates[1]
    assert medium_candidate["confidence"] == "medium"
    assert medium_candidate["projectedFixedExceptionCount"] == 0
    assert medium_candidate["matchReason"] == "姓名相似，但金额或工时仍需复核"
    assert "需先复核差异口径" in medium_candidate["businessQuestion"]
    assert "PDF 高于 Excel" in medium_candidate["impactSummary"]
    assert "不能直接确认匹配" in medium_candidate["cannotAutoResolveReason"]


def test_labor_name_mapping_merge_preserves_confirmed_and_rolled_back_candidates():
    import bonus_platform.app as app_module

    generated = app_module._build_name_mapping_candidates_from_reocr_plan(
        "labor_sample",
        {
            "tasks": [
                {
                    "sourceFile": "elog27-1_20260520204231.pdf",
                    "warehouseId": "27",
                    "diagnostics": {
                        "suspectedNamePairs": [
                            {
                                "cacheEmployeeName": "Coria, Virgilio",
                                "excelEmployeeName": "Brayan Gomez Vargas",
                                "amountGap": -0.01,
                                "hoursGap": 0.3,
                            }
                        ]
                    },
                },
                {
                    "sourceFile": "elog1-1_20260520204104.pdf",
                    "warehouseId": "1",
                    "diagnostics": {
                        "suspectedNamePairs": [
                            {
                                "cacheEmployeeName": "Espinosa Manuel",
                                "excelEmployeeName": "Massiel Castillo",
                                "amountGap": 0,
                                "hoursGap": 0,
                            }
                        ]
                    },
                },
            ]
        },
    )
    confirmed = {**generated[0], "status": "confirmed", "decision": "confirmed"}
    rolled_back = {**generated[1], "status": "rolled_back", "decision": "rolled_back"}

    merged = app_module._merge_name_mapping_candidates(
        {
            "candidates": [confirmed, rolled_back],
            "activeMappings": [{**confirmed, "decision": "active", "status": "active"}],
            "rolledBackMappings": [rolled_back],
        },
        generated,
    )

    by_id = {candidate["candidateId"]: candidate for candidate in merged}
    assert by_id[confirmed["candidateId"]]["status"] == "confirmed"
    assert by_id[rolled_back["candidateId"]]["status"] == "rolled_back"
    assert len(merged) == 2


def test_labor_compare_uses_active_profile_from_current_run(monkeypatch):
    import bonus_platform.app as app_module

    captured_profiles = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda paths, *args, **kwargs: _unresolved_quick_totals(paths))

    def fake_extract(*args, **kwargs):
        captured_profiles.append(kwargs.get("supplier_profile_override"))
        return [
            LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="WUS042586", employee_name_raw="Rosa Alvarez Minchaca", hours=31.19, amount=701.90, currency="USD", confidence=0.95, evidence_text="Rosa Alvarez Minchaca 31.19 $701.90")
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Workforce", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    app_module.update_labor_metadata(
        run["id"],
        {
            "profileGovernance": {
                "candidates": [],
                "replaySummaries": {},
                "activeProfiles": [
                    {
                        "candidateId": "profile_workforce_active",
                        "profileKey": "workforce",
                        "supplier": "Workforce",
                        "decision": "active",
                        "status": "active",
                        "requiresConfirmation": False,
                        "profileData": {
                            "key": "workforce",
                            "aliases": ["workforce"],
                            "prompt_notes": ["Active run profile guidance."],
                            "image_page_policy": "first_page_only",
                            "version": 2,
                            "status": "approved",
                            "approvedBy": "payroll-admin@example.com",
                            "approvedAt": "2026-07-15T09:30:00+08:00",
                            "created_from": "manual_review",
                        },
                    }
                ],
                "rolledBackProfiles": [],
            }
        },
    )
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    updated = app_module._perform_labor_extract_compare(run["id"])

    assert updated["status"] == "PDF识别未完成"
    assert updated["batchGuard"]["status"] == "pdf_recognition_incomplete"
    assert captured_profiles
    assert captured_profiles[0].key == "workforce"
    assert captured_profiles[0].version == 2
    assert captured_profiles[0].prompt_notes == ["Active run profile guidance."]


def test_labor_low_confidence_rows_create_correction_candidates(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "quick_extract_totals", lambda paths, *args, **kwargs: _unresolved_quick_totals(paths))
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="WUS042586", employee_name_raw="Rosa Alvarez Minchaca", hours=31.19, amount=701.90, currency="USD", confidence=0.6, evidence_text="blurred row Rosa 31.19 701.90")
        ],
    )

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    updated = app_module._perform_labor_extract_compare(run["id"])
    candidates = updated["correctionGovernance"]["candidates"]

    assert len(candidates) == 1
    assert candidates[0]["decision"] == "candidate_only"
    assert candidates[0]["requiresConfirmation"] is True
    assert candidates[0]["confidence"] == 0.6
    assert candidates[0]["evidence"]["evidenceText"] == "blurred row Rosa 31.19 701.90"

    blocked = client.post(
        f"/api/labor/runs/{run['id']}/correction-candidates/{candidates[0]['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "Evidence reviewed"},
    )
    assert blocked.status_code == 400
    assert "影响回放摘要" in blocked.json()["detail"]

    replay = client.post(
        f"/api/labor/runs/{run['id']}/correction-candidates/{candidates[0]['candidateId']}/auto-replay",
        json={},
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["decision"] == "ready_for_user_confirmation"
    assert replay_body["summary"]["affectedEmployees"] == ["Rosa Alvarez Minchaca"]
    assert replay_body["summary"]["regressionCount"] == 0
    assert replay_body["impact"][0]["riskReduced"] is True

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/correction-candidates/{candidates[0]['candidateId']}/confirm",
        json={"confirmedBy": "ops-user", "reason": "Evidence reviewed"},
    )
    assert confirmed.status_code == 200
    assert confirmed.json()["decision"] == "active"
    assert confirmed.json()["replaySummary"]["regressionCount"] == 0
    before_preview = client.get(f"/api/labor/runs/{run['id']}").json()["comparisonSummary"]
    before_files = client.get(f"/api/labor/runs/{run['id']}").json()["files"]

    preview = client.post(
        f"/api/labor/runs/{run['id']}/corrections/projected-preview",
        json={"candidateIds": [candidates[0]["candidateId"]], "generateReport": True},
    )
    assert preview.status_code == 200
    preview_body = preview.json()
    assert preview_body["decision"] == "preview_only"
    assert preview_body["summaryDelta"]["lowConfidenceCount"] == -1
    assert preview_body["summaryDelta"]["exceptionCount"] == -1
    assert preview_body["affectedRows"][0]["matchStatus"] == "通过"
    assert preview_body["preflight"]["willOverwriteOfficialResult"] is False
    assert preview_body["preflight"]["willRegenerateDiffReport"] is False
    assert preview_body["preflight"]["delta"]["lowConfidenceCount"] == -1
    assert preview_body["preflight"]["delta"]["exceptionCount"] == -1
    assert preview_body["preflight"]["affectedEmployeeCount"] == 1
    assert preview_body["preflight"]["blockingAfterApply"] is False
    assert preview_body["reportFile"]["label"] == "修正预览报告"
    assert Path(preview_body["reportFile"]["path"]).exists()
    downloaded_preview = client.get(preview_body["reportFile"]["downloadUrl"])
    assert downloaded_preview.status_code == 200
    after_preview = client.get(f"/api/labor/runs/{run['id']}").json()
    assert after_preview["comparisonSummary"] == before_preview
    assert after_preview["files"]["diffReport"] == before_files["diffReport"]
    assert after_preview["files"]["correctionPreviewReport"]["filename"].endswith(".xlsx")
    assert after_preview["correctionGovernance"]["activeCorrections"][0]["preflight"] == preview_body["preflight"]

    rolled_back = client.post(
        f"/api/labor/runs/{run['id']}/correction-candidates/{candidates[0]['candidateId']}/rollback",
        json={"rolledBackBy": "ops-user", "reason": "Correction no longer needed"},
    )
    assert rolled_back.status_code == 200
    assert rolled_back.json()["decision"] == "rolled_back"


def test_labor_governance_report_collects_rule_profile_correction_and_ai_evidence():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Workforce", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    run_dir = app_module.get_labor_run_dir(run["id"])
    diff_path = run_dir / "existing_diff.xlsx"
    Workbook().save(diff_path)
    correction = {
        "candidateId": "correction_sample",
        "decision": "active",
        "status": "active",
        "requiresConfirmation": False,
        "confidence": 0.6,
        "proposed": {
            "employeeName": "Rosa Alvarez Minchaca",
            "hours": 31.19,
            "amount": 701.9,
            "sourceFile": "invoice.pdf",
            "sourcePageOrRow": "p1",
        },
        "auditTrail": [{"action": "created", "actor": "system", "reason": "low confidence"}],
        "replaySummary": {"fixedCount": 1, "regressionCount": 0},
    }
    app_module.update_labor_metadata(
        run["id"],
        {
            "status": "已生成差异报告",
            "ruleGovernance": {
                "candidates": [
                    {
                        "ruleId": "workforce-low-confidence-v1",
                        "title": "低置信度复核规则",
                        "status": "pending_user_confirmation",
                        "decision": "candidate_only",
                        "version": 1,
                        "auditTrail": [{"action": "created", "actor": "ai", "reason": "diagnostic signal"}],
                    }
                ],
                "replaySummaries": {
                    "workforce-low-confidence-v1": {
                        "decision": "ready_for_user_confirmation",
                        "summary": {"replayedCount": 1, "fixedCount": 1, "regressionCount": 0},
                    }
                },
                "activeRules": [],
                "rolledBackRules": [],
            },
            "profileGovernance": {
                "candidates": [
                    {
                        "candidateId": "profile_workforce_v1",
                        "supplier": "Workforce",
                        "profileKey": "workforce",
                        "status": "pending_user_confirmation",
                        "decision": "candidate_only",
                        "profileData": {"key": "workforce", "version": 1},
                        "auditTrail": [{"action": "created", "actor": "system", "reason": "profile suggestion"}],
                    }
                ],
                "replaySummaries": {
                    "profile_workforce_v1": {
                        "decision": "ready_for_user_confirmation",
                        "summary": {"compatibleCount": 1, "regressionCount": 0},
                    }
                },
                "activeProfiles": [],
                "rolledBackProfiles": [],
            },
            "correctionGovernance": {
                "candidates": [correction],
                "replaySummaries": {
                    "correction_sample": {
                        "decision": "ready_for_user_confirmation",
                        "summary": {"affectedEmployees": ["Rosa Alvarez Minchaca"], "fixedCount": 1, "regressionCount": 0},
                    }
                },
                "activeCorrections": [correction],
                "rolledBackCorrections": [],
            },
            "nameMappingGovernance": {
                "candidates": [
                    {
                        "candidateId": "name_map_sample",
                        "decision": "candidate_only",
                        "status": "pending_user_confirmation",
                        "cacheEmployeeName": "Cache Rosa",
                        "excelEmployeeName": "Rosa Alvarez Minchaca",
                        "sourceFile": "invoice.pdf",
                        "warehouseId": "1",
                        "amountGap": 0,
                        "hoursGap": 0,
                        "recommendation": "Review name mapping",
                        "evidence": {"sourceRefs": "invoice.pdf p1; sheet row 2"},
                        "auditTrail": [{"action": "created", "actor": "system", "reason": "reocr_suspected_name_pair"}],
                    }
                ],
                "replaySummaries": {},
                "activeMappings": [],
                "rolledBackMappings": [],
            },
            "reocrReplayGovernance": {
                "replays": [
                    {
                        "decision": "ready_for_user_confirmation",
                        "mode": "new_ocr_candidate_replay",
                        "sourceFile": "invoice.pdf",
                        "warehouseId": "1",
                        "summary": {
                            "candidateRowCount": 1,
                            "candidateAmountTotal": 701.9,
                            "expectedExcelAmount": 701.9,
                            "exceptionCount": 0,
                        },
                        "diagnostics": {
                            "recommendedAction": "review_name_mapping_then_reocr_if_amounts_remain_unexplained",
                            "rootCauseHints": ["possible_name_mapping", "possible_missing_cache_rows"],
                            "suspectedNamePairs": [
                                {
                                    "cacheEmployeeName": "Cache Rosa",
                                    "excelEmployeeName": "Rosa Alvarez Minchaca",
                                    "cacheAmount": 701.9,
                                    "excelAmount": 701.9,
                                }
                            ],
                        },
                        "blockers": [],
                    }
                ],
                "activeCandidates": [
                    {
                        "candidateId": "reocr_sample",
                        "decision": "active",
                        "status": "active",
                        "sourceFile": "invoice.pdf",
                        "warehouseId": "1",
                        "confirmedBy": "ops-user",
                        "confirmationReason": "OCR replay passed",
                        "replay": {
                            "summary": {
                                "candidateRowCount": 1,
                                "candidateAmountTotal": 701.9,
                                "expectedExcelAmount": 701.9,
                                "exceptionCount": 0,
                            }
                        },
                        "auditTrail": [{"action": "confirmed", "actor": "ops-user", "reason": "OCR replay passed"}],
                    }
                ],
                "rolledBackCandidates": [
                    {
                        "candidateId": "reocr_rolled_back",
                        "decision": "rolled_back",
                        "status": "rolled_back",
                        "sourceFile": "invoice_old.pdf",
                        "warehouseId": "1",
                        "confirmedBy": "ops-user",
                        "rolledBackBy": "ops-user",
                        "rollbackReason": "Candidate was superseded by a cleaner OCR run.",
                        "replay": {
                            "summary": {
                                "candidateRowCount": 1,
                                "candidateAmountTotal": 680.0,
                                "expectedExcelAmount": 701.9,
                                "exceptionCount": 1,
                            }
                        },
                        "auditTrail": [
                            {"action": "confirmed", "actor": "ops-user", "reason": "initial OCR replay passed"},
                            {"action": "rolled_back", "actor": "ops-user", "reason": "Candidate was superseded by a cleaner OCR run."},
                        ],
                    }
                ],
            },
            "files": {
                "diffReport": app_module.attach_labor_file(run["id"], diff_path, "差异报告"),
                "reocrCandidateFiles": [
                    {
                        "filename": "partial_batch.csv",
                        "summary": {
                            "plannedTaskCount": 2,
                            "coveredTaskCount": 1,
                            "missingTaskCount": 1,
                            "extraScopeCount": 0,
                            "parsedRowCount": 1,
                        },
                        "coverage": {
                            "coverageComplete": False,
                            "plannedTaskCount": 2,
                            "coveredTaskCount": 1,
                            "missingTaskCount": 1,
                            "extraScopeCount": 0,
                            "uploadedScopes": [{"sourceFile": "invoice.pdf", "warehouseId": "1", "rowCount": 1}],
                            "missingTasks": [{"sourceFile": "invoice_2.pdf", "warehouseId": "2"}],
                            "extraScopes": [],
                        },
                    }
                ]
            },
            "aiCacheAudit": {
                "decision": "candidate_only",
                "requiresConfirmation": True,
                "message": "AI cache evidence only.",
                "summary": {"candidateFileCount": 1},
                "files": [
                    {
                        "sourceFile": "invoice.pdf",
                        "warehouseId": "1",
                        "rowCount": 1,
                        "candidateAmountTotal": 701.9,
                        "averageConfidence": 0.6,
                        "decision": "candidate_only",
                        "evidence": [{"employeeName": "Rosa Alvarez Minchaca", "amount": 701.9, "sourcePageOrRow": "p1"}],
                    }
                ],
            },
        },
    )

    before_files = client.get(f"/api/labor/runs/{run['id']}").json()["files"]
    response = client.post(f"/api/labor/runs/{run['id']}/governance-report")

    assert response.status_code == 200
    report = response.json()
    assert report["label"] == "治理审计报告"
    assert Path(report["path"]).exists()
    downloaded = client.get(report["downloadUrl"])
    assert downloaded.status_code == 200
    workbook = load_workbook(report["path"], read_only=True)
    assert workbook.sheetnames == ["治理总览", "规则治理", "姓名映射治理", "Profile治理", "修正治理", "图片识别治理", "图片识别上传覆盖", "AI候选治理", "审计记录"]
    overview_rows = list(workbook["治理总览"].iter_rows(values_only=True))
    assert any(row[:2] == ("姓名映射候选", 1) for row in overview_rows)
    name_mapping_rows = list(workbook["姓名映射治理"].iter_rows(values_only=True))
    assert any(row[0] == "候选" and row[1] == "name_map_sample" and row[2] == "Cache Rosa" and row[3] == "Rosa Alvarez Minchaca" for row in name_mapping_rows)
    reocr_rows = list(workbook["图片识别治理"].iter_rows(values_only=True))
    assert any(row[0] == "回放" and row[2] == "invoice.pdf" and row[5] == "ready_for_user_confirmation" for row in reocr_rows)
    assert any(
        row[0] == "回放"
        and row[10] == "review_name_mapping_then_reocr_if_amounts_remain_unexplained"
        and "possible_name_mapping" in str(row[11])
        and "Cache Rosa" in str(row[12])
        for row in reocr_rows
    )
    assert any(row[0] == "已确认" and row[1] == "reocr_sample" and row[13] == "ops-user" for row in reocr_rows)
    assert any(row[0] == "已回滚" and row[1] == "reocr_rolled_back" and row[14] == "Candidate was superseded by a cleaner OCR run." for row in reocr_rows)
    upload_rows = list(workbook["图片识别上传覆盖"].iter_rows(values_only=True))
    assert any(row[0] == "汇总" and row[1] == "partial_batch.csv" and row[2:7] == (2, 1, 1, 0, "否") for row in upload_rows)
    assert any(row[0] == "已上传范围" and row[7] == "invoice.pdf" and row[8] == "1" and row[9] == 1 for row in upload_rows)
    assert any(row[0] == "缺失计划任务" and row[7] == "invoice_2.pdf" and row[8] == "2" for row in upload_rows)
    audit_rows = list(workbook["审计记录"].iter_rows(values_only=True))
    assert any(row[0] == "姓名映射" and row[1] == "name_map_sample" and row[2] == "created" for row in audit_rows)
    after_files = client.get(f"/api/labor/runs/{run['id']}").json()["files"]
    assert after_files["diffReport"] == before_files["diffReport"]
    assert after_files["governanceAuditReport"]["filename"].endswith(".xlsx")


def test_labor_rule_candidate_api_blocks_confirmation_when_replay_regresses():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSI", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()

    created = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates",
        json={
            "ruleId": "osi-name-fuzzy-v1",
            "title": "OSI name fuzzy matching",
            "description": "Candidate rule from employee attribution diagnostics.",
            "source": "real replay: osi",
            "evidence": [{"sourceFile": "US ELogistics Service Corp. 34794.pdf", "sourcePageOrRow": "p1"}],
            "conditions": {"supplier": "OSI"},
        },
    )

    assert created.status_code == 200
    assert created.json()["candidates"][0]["requiresConfirmation"] is True

    replay = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/osi-name-fuzzy-v1/replay-summary",
        json={
            "replayResults": [
                {"runId": "osi_34794", "supplier": "OSI", "beforeStatus": "warning", "afterStatus": "ok", "beforeIssueCount": 1, "afterIssueCount": 0},
                {"runId": "fairway_135612", "supplier": "Fairway", "beforeStatus": "ok", "afterStatus": "warning", "beforeIssueCount": 0, "afterIssueCount": 1},
            ]
        },
    )

    assert replay.status_code == 200
    assert replay.json()["decision"] == "blocked_by_replay_regression"
    assert replay.json()["preflight"]["blockingAfterApply"] is True
    assert replay.json()["preflight"]["delta"]["regressionCount"] == 1
    assert replay.json()["preflight"]["regressionRuns"][0]["runId"] == "fairway_135612"

    confirm = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/osi-name-fuzzy-v1/confirm",
        json={"confirmedBy": "ops-user", "reason": "should not pass"},
    )

    assert confirm.status_code == 400
    assert "未通过历史影响预览" in confirm.json()["detail"]


def test_labor_rule_candidate_api_confirms_and_rolls_back_after_clean_replay():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()

    created = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates",
        json={
            "ruleId": "oss-hash-warehouse-v1",
            "title": "OSS # warehouse id extraction",
            "description": "Parse warehouse id from US Elogis Service #N invoice names.",
            "source": "real replay: oss 2",
            "proposedBy": "ai",
            "conditions": {"supplier": "OSS", "filenamePattern": "#N"},
        },
    )
    assert created.status_code == 200

    replay = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/oss-hash-warehouse-v1/replay-summary",
        json={
            "replayResults": [
                {"runId": "oss2_warehouse_7", "supplier": "OSS", "beforeStatus": "warning", "afterStatus": "ok", "beforeIssueCount": 1, "afterIssueCount": 0},
                {"runId": "oss2_warehouse_27", "supplier": "OSS", "beforeStatus": "ok", "afterStatus": "ok", "beforeIssueCount": 0, "afterIssueCount": 0},
            ]
        },
    )
    assert replay.status_code == 200
    replay_body = replay.json()
    assert replay_body["decision"] == "ready_for_user_confirmation"
    assert replay_body["preflight"]["blockingAfterApply"] is False
    assert replay_body["preflight"]["delta"]["fixedCount"] == 1
    assert replay_body["preflight"]["affectedScopeCount"] == 2
    assert replay_body["preflight"]["affectedSuppliers"] == ["OSS"]

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/oss-hash-warehouse-v1/confirm",
        json={"confirmedBy": "ops-user", "reason": "OSS2 replay clean"},
    )
    assert confirmed.status_code == 200
    assert confirmed.json()["decision"] == "active"
    assert confirmed.json()["requiresConfirmation"] is False
    assert confirmed.json()["preflight"] == replay_body["preflight"]

    rolled_back = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/oss-hash-warehouse-v1/rollback",
        json={"rolledBackBy": "ops-user", "reason": "future batch regression", "targetVersion": 0},
    )
    assert rolled_back.status_code == 200
    assert rolled_back.json()["decision"] == "rolled_back"
    assert rolled_back.json()["rollbackToVersion"] == 0

    governance = client.get(f"/api/labor/runs/{run['id']}/governance").json()
    assert governance["activeRules"] == []
    assert governance["rolledBackRules"][0]["ruleId"] == "oss-hash-warehouse-v1"


def test_labor_reocr_candidate_replay_api_returns_ready_for_confirmation():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay",
        json={
            "task": {
                "sourceFile": "elog1-1_20260520204104.pdf",
                "warehouseId": "1",
                "expectedExcelAmount": 100,
                "amountDelta": 50,
                "confirmationGate": "new OCR candidate must replay clean before replacing cache",
            },
            "candidateRows": [
                {
                    "employeeName": "Alice Worker",
                    "sourcePageOrRow": "p1",
                    "hours": 8,
                    "amount": 100,
                    "currency": "USD",
                    "confidence": 0.96,
                    "evidenceText": "Alice Worker 8.00 $100.00",
                }
            ],
            "amountTolerance": 0.1,
            "hoursTolerance": 0.1,
            "confidenceThreshold": 0.85,
        },
    )

    assert replay.status_code == 200
    body = replay.json()
    assert body["runId"] == run["id"]
    assert body["decision"] == "ready_for_user_confirmation"
    assert body["requiresConfirmation"] is True
    assert body["blockers"] == []
    assert body["summary"]["candidateAmountTotal"] == 100
    assert body["summary"]["expectedExcelAmount"] == 100
    assert body["summary"]["amountPassed"] is True
    assert body["summary"]["exceptionCount"] == 0
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["reocrReplayGovernance"]["replays"][0]["decision"] == "ready_for_user_confirmation"
    before_summary = refreshed.get("comparisonSummary") or {}

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/confirm",
        json={
            "sourceFile": "elog1-1_20260520204104.pdf",
            "warehouseId": "1",
            "confirmedBy": "ops-user",
            "reason": "new OCR candidate replay passed",
            "generateReport": True,
        },
    )

    assert confirmed.status_code == 200
    confirmed_body = confirmed.json()
    assert confirmed_body["decision"] == "active"
    assert confirmed_body["reportFile"]["label"] == "图片识别结果预览报告"
    assert Path(confirmed_body["reportFile"]["path"]).exists()
    downloaded = client.get(confirmed_body["reportFile"]["downloadUrl"])
    assert downloaded.status_code == 200
    after_confirm = client.get(f"/api/labor/runs/{run['id']}").json()
    assert (after_confirm.get("comparisonSummary") or {}) == before_summary
    assert after_confirm["files"]["reocrPreviewReport"]["filename"].endswith(".xlsx")
    assert after_confirm["reocrReplayGovernance"]["activeCandidates"][0]["decision"] == "active"

    candidate_id = after_confirm["reocrReplayGovernance"]["activeCandidates"][0]["candidateId"]
    applied = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/{candidate_id}/apply",
        json={"appliedBy": "ops-user", "reason": "preview report reviewed"},
    )

    assert applied.status_code == 200
    applied_body = applied.json()
    assert applied_body["decision"] == "applied"
    assert applied_body["comparisonSummary"]["reocrCandidateApplied"] is True
    assert applied_body["reportFile"]["label"] == "差异报告"
    assert applied_body["preflight"]["willOverwriteOfficialResult"] is True
    assert applied_body["preflight"]["willRegenerateDiffReport"] is True
    assert applied_body["preflight"]["projected"]["pdfAmountTotal"] == 100
    assert applied_body["preflight"]["projected"]["exceptionCount"] == 0
    assert applied_body["preflight"]["affectedScopeCount"] == 1
    assert applied_body["preflight"]["affectedEmployeeCount"] == 1
    assert applied_body["preflight"]["blockingAfterApply"] is True
    assert any("完整核对" in warning for warning in applied_body["preflight"]["postApplyWarnings"])
    assert applied_body["comparisonSummary"]["canRelease"] is False
    assert applied_body["comparisonSummary"]["machineCheckStatus"] == "needs_review"
    assert applied_body["readinessGate"]["status"] == "blocked"
    assert client.get(applied_body["reportFile"]["downloadUrl"]).status_code == 200
    after_apply = client.get(f"/api/labor/runs/{run['id']}").json()
    assert after_apply["status"] == "部分核对完成"
    assert after_apply["comparisonSummary"]["pdfAmountTotal"] == 100
    assert after_apply["comparisonSummary"]["exceptionCount"] == 0
    assert after_apply["comparisonSummary"]["conclusionLevel"] == "warning"
    assert after_apply["comparisonSummary"]["canRelease"] is False
    assert after_apply["machineCheckStatus"] == "needs_review"
    assert after_apply["businessReviewStatus"] == "pending"
    assert after_apply["directPaymentAllowed"] is False
    assert after_apply["batchGuard"]["status"] == "reocr_revalidation_required"
    assert after_apply["batchGuard"]["allowReleasableReport"] is False
    assert after_apply["comparisonRows"][0]["matchStatus"] == "通过"
    assert after_apply["presentation"]["summary"]["employeeCount"] == 1
    assert after_apply["presentation"]["summary"]["reviewItemCount"] == 0
    assert after_apply["reocrReplayGovernance"]["activeCandidates"][0]["status"] == "applied"
    assert after_apply["files"]["diffReport"]["filename"].endswith(".xlsx")
    assert after_apply["diffDownloadUrl"] == after_apply["files"]["diffReport"]["downloadUrl"]
    assert after_apply["reocrAdoption"]["preflight"] == applied_body["preflight"]

    rolled_back = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/{candidate_id}/rollback",
        json={"rolledBackBy": "ops-user", "reason": "candidate superseded by corrected OCR file"},
    )

    assert rolled_back.status_code == 200
    rolled_back_body = rolled_back.json()
    assert rolled_back_body["decision"] == "rolled_back"
    assert rolled_back_body["status"] == "rolled_back"
    assert rolled_back_body["rollbackReason"] == "candidate superseded by corrected OCR file"
    assert rolled_back_body["auditTrail"][-1]["action"] == "rolled_back"
    after_rollback = client.get(f"/api/labor/runs/{run['id']}").json()
    assert (after_rollback.get("comparisonSummary") or {}) == before_summary
    assert after_rollback.get("presentation") == {}
    assert "diffReport" not in after_rollback.get("files", {})
    assert after_rollback.get("diffDownloadUrl", "") == ""
    assert after_rollback["reocrReplayGovernance"]["activeCandidates"] == []
    assert after_rollback["reocrReplayGovernance"]["rolledBackCandidates"][0]["candidateId"] == candidate_id


def test_labor_material_replay_blocks_incomplete_employee_detail_and_preserves_pdf_provenance(monkeypatch):
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "Prompt Staffing", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("DEPT#1.pdf", b"%PDF-1.4\n% image only", "application/pdf")),
            ("pdf_files", ("DEPT#2.pdf", b"%PDF-1.4\n% image only", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {"name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种", "warehouse": "物理仓"},
        },
    )
    app_module.update_labor_metadata(run["id"], {"materialReplaySource": {"batchKey": "prompt"}})

    quick_calls: list[list[str]] = []

    def fake_quick_extract(paths, *args, **kwargs):
        quick_calls.append([Path(path).name for path in paths])
        return [
            {
                "source_file": Path(path).name,
                "warehouse_id": str(index),
                "total_amount": 0.0,
                "authoritative": False,
                "evidence_status": "needs_review",
                "page_evidence": [{"page": 1, "role": "invoice_primary"}],
            }
            for index, path in enumerate(paths, start=1)
        ]

    monkeypatch.setattr(app_module, "quick_extract_totals", fake_quick_extract)
    monkeypatch.setattr(app_module, "extract_invoice_items", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {
                "fileCount": len(paths),
                "textReadableFileCount": 0,
                "imageOnlyFileCount": len(paths),
                "textReadablePageCount": 0,
                "emptyTextPageCount": len(paths),
                "imageOnlyPdfFiles": [path.name for path in paths],
            },
            "files": [
                {
                    "sourceFile": path.name,
                    "pageCount": 1,
                    "readablePageCount": 0,
                    "emptyTextPageCount": 1,
                    "hasTextLayer": False,
                    "needsOcr": True,
                    "diagnostic": "image_only_pdf",
                }
                for path in paths
            ],
        },
    )
    monkeypatch.setattr(
        app_module,
        "build_reocr_candidate_plan",
        lambda *args, **kwargs: {
            "summary": {"taskCount": 1, "reviewableCandidateCount": 0},
            "tasks": [{"sourceFile": "DEPT#1.pdf", "warehouseId": "1", "amountDelta": -100, "expectedExcelAmount": 100}],
            "reviewableCandidates": [],
        },
    )

    updated = app_module._perform_labor_extract_compare(run["id"])

    assert len(quick_calls) == 1
    assert len(quick_calls[0]) == 2
    assert updated["status"] == "PDF识别未完成"
    assert updated["stage"] == "生成报告"
    assert updated["reviewQueues"]["primary"] == "reocr"
    assert updated["reviewQueues"]["reocr"]["taskCount"] == 1
    assert updated["extractionQuality"]["level"] == "critical"
    assert updated["pdfExtractedRows"] == []
    assert updated["files"]["diffReport"]["label"] == "PDF识别诊断报告"
    assert len(updated["invoiceEvidenceAudit"]) == 2
    assert {row["reconciliationStatus"] for row in updated["warehouseComparison"]["rows"]} == {"needs_review"}
    assert all(row["pdfEvidenceFile"] for row in updated["warehouseComparison"]["rows"])
    assert not any(row["reconciliationStatus"] == "missing_pdf_invoice" for row in updated["warehouseComparison"]["rows"])


def test_labor_reocr_batch_preview_and_apply_updates_official_report():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    replay_1 = {
        "decision": "ready_for_user_confirmation",
        "sourceFile": "elog1.pdf",
        "warehouseId": "1",
        "summary": {"candidateAmountTotal": 100, "expectedExcelAmount": 100, "exceptionCount": 0},
        "comparison": {
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 1,
            "pdfHoursTotal": 8,
            "excelHoursTotal": 8,
            "pdfAmountTotal": 100,
            "excelAmountTotal": 100,
            "exceptionCount": 0,
            "averageConfidence": 0.95,
        },
        "comparisonRows": [
            {
                "employeeKey": "name:ALICE WORKER",
                "employeeName": "Alice Worker",
                "pdfHoursTotal": 8,
                "excelHoursTotal": 8,
                "hoursDelta": 0,
                "pdfAmountTotal": 100,
                "excelAmountTotal": 100,
                "amountDelta": 0,
                "matchStatus": "通过",
                "riskFlags": [],
                "sourceRefs": "elog1.pdf p1; bill.xlsx row 2",
            }
        ],
        "candidateMatches": [],
    }
    replay_2 = {
        "decision": "ready_for_user_confirmation",
        "sourceFile": "elog2.pdf",
        "warehouseId": "2",
        "summary": {"candidateAmountTotal": 200, "expectedExcelAmount": 200, "exceptionCount": 0},
        "comparison": {
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 1,
            "pdfHoursTotal": 10,
            "excelHoursTotal": 10,
            "pdfAmountTotal": 200,
            "excelAmountTotal": 200,
            "exceptionCount": 0,
            "averageConfidence": 0.96,
        },
        "comparisonRows": [
            {
                "employeeKey": "name:BOB WORKER",
                "employeeName": "Bob Worker",
                "pdfHoursTotal": 10,
                "excelHoursTotal": 10,
                "hoursDelta": 0,
                "pdfAmountTotal": 200,
                "excelAmountTotal": 200,
                "amountDelta": 0,
                "matchStatus": "通过",
                "riskFlags": [],
                "sourceRefs": "elog2.pdf p1; bill.xlsx row 3",
            }
        ],
        "candidateMatches": [],
    }
    app_module.update_labor_metadata(
        run["id"],
        {
            "comparisonSummary": {"exceptionCount": 4, "pdfAmountTotal": 0, "excelAmountTotal": 300, "amountDeltaTotal": -300},
            "comparisonRows": [],
            "candidateMatches": [],
            "excelMapping": {"name": "姓名", "hours": "工时", "amount": "金额"},
            "reocrPlan": {
                "tasks": [
                    {"sourceFile": "elog1.pdf", "warehouseId": "1"},
                    {"sourceFile": "elog2.pdf", "warehouseId": "2"},
                    {"sourceFile": "elog3.pdf", "warehouseId": "3"},
                ],
                "reviewableCandidates": [],
            },
            "reocrReplayGovernance": {
                "replays": [replay_1, replay_2],
                "activeCandidates": [
                    {"candidateId": "reocr_1", "decision": "active", "status": "active", "sourceFile": "elog1.pdf", "warehouseId": "1", "replay": replay_1},
                    {"candidateId": "reocr_2", "decision": "active", "status": "active", "sourceFile": "elog2.pdf", "warehouseId": "2", "replay": replay_2},
                ],
                "rolledBackCandidates": [],
            },
        }
    )

    preview = client.post(f"/api/labor/runs/{run['id']}/reocr-candidates/batch-preview", json={})
    assert preview.status_code == 200
    preview_body = preview.json()
    assert preview_body["decision"] == "ready_for_batch_apply"
    assert preview_body["summary"]["candidateCount"] == 2
    assert preview_body["summary"]["plannedTaskCount"] == 3
    assert preview_body["summary"]["missingAppliedTaskCount"] == 1
    assert preview_body["coverage"]["missingAppliedTasks"] == [{"sourceFile": "elog3.pdf", "warehouseId": "3"}]
    assert preview_body["comparisonSummary"]["pdfAmountTotal"] == 300
    assert preview_body["preflight"]["willOverwriteOfficialResult"] is True
    assert preview_body["preflight"]["willRegenerateDiffReport"] is True
    assert preview_body["preflight"]["current"]["exceptionCount"] == 4
    assert preview_body["preflight"]["projected"]["exceptionCount"] == 0
    assert preview_body["preflight"]["delta"]["exceptionCount"] == -4
    assert preview_body["preflight"]["delta"]["amountDeltaTotal"] == 300
    assert preview_body["preflight"]["affectedScopeCount"] == 2
    assert preview_body["preflight"]["affectedEmployeeCount"] == 2
    assert preview_body["preflight"]["coverageCompleteAfterApply"] is False
    assert preview_body["preflight"]["blockingAfterApply"] is True
    assert preview_body["preflight"]["postApplyWarnings"] == [
        "仍有 1 个图片识别复核任务未采纳，交付状态将保持阻断。",
        "采纳后必须重新执行整批完整核对，重新生成员工明细覆盖、金额闭合和诊断门禁。",
    ]

    applied = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/batch-apply",
        json={"appliedBy": "ops-user", "reason": "all reocr previews reviewed"},
    )
    assert applied.status_code == 200
    body = applied.json()
    assert body["decision"] == "batch_applied"
    assert body["summary"]["candidateCount"] == 2
    assert body["summary"]["missingAppliedTaskCount"] == 1
    assert body["coverage"]["coverageComplete"] is False
    assert body["preflight"] == preview_body["preflight"]
    assert body["reportFile"]["label"] == "差异报告"
    assert client.get(body["reportFile"]["downloadUrl"]).status_code == 200
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["status"] == "部分核对完成"
    assert refreshed["comparisonSummary"]["pdfAmountTotal"] == 300
    assert refreshed["comparisonSummary"]["exceptionCount"] == 0
    assert refreshed["comparisonSummary"]["conclusionLevel"] == "warning"
    assert refreshed["comparisonSummary"]["canRelease"] is False
    assert refreshed["machineCheckStatus"] == "needs_review"
    assert refreshed["batchGuard"]["status"] == "reocr_revalidation_required"
    assert len(refreshed["comparisonRows"]) == 2
    assert refreshed["presentation"]["summary"]["employeeCount"] == 2
    assert refreshed["presentation"]["summary"]["reviewItemCount"] == 0
    assert {item["status"] for item in refreshed["reocrReplayGovernance"]["activeCandidates"]} == {"applied"}
    assert refreshed["diffDownloadUrl"] == refreshed["files"]["diffReport"]["downloadUrl"]
    assert refreshed["reocrAdoption"]["preflight"] == preview_body["preflight"]


def test_labor_reocr_batch_confirm_promotes_ready_replays_without_applying_result():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    replay_1 = {
        "decision": "ready_for_user_confirmation",
        "sourceFile": "elog1.pdf",
        "warehouseId": "1",
        "replayedAt": "2026-06-15T00:00:00",
        "summary": {"candidateAmountTotal": 100, "expectedExcelAmount": 100, "exceptionCount": 0},
        "comparison": {"pdfAmountTotal": 100, "excelAmountTotal": 100, "exceptionCount": 0},
        "comparisonRows": [{"employeeName": "Alice Worker", "matchStatus": "通过", "pdfAmountTotal": 100, "excelAmountTotal": 100, "amountDelta": 0}],
        "candidateMatches": [],
    }
    replay_2 = {
        "decision": "ready_for_user_confirmation",
        "sourceFile": "elog2.pdf",
        "warehouseId": "2",
        "replayedAt": "2026-06-15T00:00:01",
        "summary": {"candidateAmountTotal": 200, "expectedExcelAmount": 200, "exceptionCount": 0},
        "comparison": {"pdfAmountTotal": 200, "excelAmountTotal": 200, "exceptionCount": 0},
        "comparisonRows": [{"employeeName": "Bob Worker", "matchStatus": "通过", "pdfAmountTotal": 200, "excelAmountTotal": 200, "amountDelta": 0}],
        "candidateMatches": [],
    }
    replay_blocked = {
        "decision": "blocked_by_replay",
        "sourceFile": "elog3.pdf",
        "warehouseId": "3",
        "summary": {"candidateAmountTotal": 50, "expectedExcelAmount": 75, "exceptionCount": 1},
    }
    app_module.update_labor_metadata(
        run["id"],
        {
            "comparisonSummary": {"exceptionCount": 9, "pdfAmountTotal": 0, "excelAmountTotal": 300},
            "reocrReplayGovernance": {"replays": [replay_1, replay_2, replay_blocked], "activeCandidates": [], "rolledBackCandidates": []},
        },
    )

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/confirm-batch",
        json={"confirmedBy": "ops-user", "reason": "batch replay reviewed", "generateReport": True},
    )

    assert confirmed.status_code == 200
    body = confirmed.json()
    assert body["decision"] == "batch_confirmed"
    assert body["summary"]["confirmedCount"] == 2
    assert body["summary"]["skippedCount"] == 1
    assert body["reportFile"]["label"] == "图片识别批量结果预览报告"
    assert client.get(body["reportFile"]["downloadUrl"]).status_code == 200
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert len(refreshed["reocrReplayGovernance"]["activeCandidates"]) == 2
    assert {candidate["status"] for candidate in refreshed["reocrReplayGovernance"]["activeCandidates"]} == {"active"}
    assert refreshed["comparisonSummary"]["exceptionCount"] == 9
    assert refreshed["files"]["reocrPreviewReport"]["filename"].endswith(".xlsx")


def test_labor_readiness_gate_blocks_until_reocr_plan_is_fully_applied():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    report_url = f"/api/labor/runs/{run['id']}/download/report.xlsx"
    (app_module.get_labor_run_dir(run["id"]) / "report.xlsx").write_bytes(b"report")
    baseline = {
            "status": "已生成差异报告",
            "comparisonSummary": {
                "conclusionLevel": "pass",
                "conclusionMessage": "核对通过",
                "canRelease": True,
                "machineCheckStatus": "passed",
                "exceptionCount": 0,
                "pdfAmountTotal": 300,
                "excelAmountTotal": 300,
                "amountDeltaTotal": 0,
            },
            "machineCheckStatus": "passed",
            "batchGuard": {"status": "ok", "allowReleasableReport": True},
            "reconciliationDiagnostics": {"level": "ok", "issues": []},
            "extractionQuality": {"level": "ok", "issues": []},
            "files": {"diffReport": {"filename": "report.xlsx", "downloadUrl": report_url}},
            "diffDownloadUrl": report_url,
            "reocrPlan": {
                "tasks": [
                    {"sourceFile": "elog1.pdf", "warehouseId": "1"},
                    {"sourceFile": "elog2.pdf", "warehouseId": "2"},
                ]
            },
            "reocrReplayGovernance": {
                "activeCandidates": [
                    {"candidateId": "reocr_1", "sourceFile": "elog1.pdf", "warehouseId": "1", "status": "applied", "decision": "applied"},
                    {"candidateId": "reocr_2", "sourceFile": "elog2.pdf", "warehouseId": "2", "status": "active", "decision": "active"},
                ],
                "replays": [],
                "rolledBackCandidates": [],
            },
        }
    baseline["resultInputFingerprint"] = app_module._labor_result_input_fingerprint(
        {**run, **baseline}
    )
    app_module.update_labor_metadata(run["id"], baseline)

    blocked = client.get(f"/api/labor/runs/{run['id']}").json()["readinessGate"]
    assert blocked["status"] == "blocked"
    assert blocked["summary"]["reocrPlannedTaskCount"] == 2
    assert blocked["summary"]["reocrAppliedTaskCount"] == 1
    assert any(issue["code"] == "reocr_coverage_incomplete" for issue in blocked["issues"])
    assert any(issue["code"] == "pending_governance_candidates" for issue in blocked["issues"])

    app_module.update_labor_metadata(
        run["id"],
        {
            "reocrReplayGovernance": {
                "activeCandidates": [
                    {"candidateId": "reocr_1", "sourceFile": "elog1.pdf", "warehouseId": "1", "status": "applied", "decision": "applied"},
                    {"candidateId": "reocr_2", "sourceFile": "elog2.pdf", "warehouseId": "2", "status": "applied", "decision": "applied"},
                ],
                "replays": [],
                "rolledBackCandidates": [],
            },
        },
    )

    stale = client.get(f"/api/labor/runs/{run['id']}").json()["readinessGate"]
    assert stale["status"] == "blocked"
    assert stale["ready"] is False
    assert stale["businessReviewRequired"] is True
    assert stale["directPaymentAllowed"] is False
    assert stale["summary"]["pendingGovernanceCount"] == 0
    assert any(issue["code"] == "stale_result_inputs" for issue in stale["issues"])


def test_labor_readiness_gate_blocks_confirmed_reocr_when_no_plan_requires_apply_or_rollback():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    report_url = f"/api/labor/runs/{run['id']}/download/report.xlsx"
    app_module.update_labor_metadata(
        run["id"],
        {
            "status": "已生成差异报告",
            "comparisonSummary": {
                "conclusionLevel": "pass",
                "conclusionMessage": "核对通过",
                "exceptionCount": 0,
                "pdfAmountTotal": 100,
                "excelAmountTotal": 100,
                "amountDeltaTotal": 0,
            },
            "files": {"diffReport": {"filename": "report.xlsx", "downloadUrl": report_url}},
            "diffDownloadUrl": report_url,
            "reocrReplayGovernance": {
                "activeCandidates": [
                    {"candidateId": "reocr_1", "sourceFile": "elog1.pdf", "warehouseId": "1", "status": "active", "decision": "active"}
                ],
                "replays": [],
                "rolledBackCandidates": [],
            },
        },
    )

    gate = client.get(f"/api/labor/runs/{run['id']}").json()["readinessGate"]

    assert gate["status"] == "blocked"
    assert gate["summary"]["confirmedReocrNotAppliedCount"] == 1
    assert any(issue["code"] == "confirmed_reocr_not_applied" for issue in gate["issues"])


def test_labor_readiness_gate_blocks_when_report_file_is_missing():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    report_url = f"/api/labor/runs/{run['id']}/download/missing-report.xlsx"
    missing_report_path = app_module.get_labor_run_dir(run["id"]) / "missing-report.xlsx"
    app_module.update_labor_metadata(
        run["id"],
        {
            "status": "已生成差异报告",
            "comparisonSummary": {
                "conclusionLevel": "pass",
                "conclusionMessage": "核对通过",
                "exceptionCount": 0,
                "pdfAmountTotal": 100,
                "excelAmountTotal": 100,
                "amountDeltaTotal": 0,
            },
            "files": {
                "diffReport": {
                    "filename": "missing-report.xlsx",
                    "path": str(missing_report_path),
                    "downloadUrl": report_url,
                }
            },
            "diffDownloadUrl": report_url,
        },
    )

    gate = client.get(f"/api/labor/runs/{run['id']}").json()["readinessGate"]

    assert gate["status"] == "blocked"
    assert gate["ready"] is False
    assert any(issue["code"] == "report_file_missing" for issue in gate["issues"])


def test_labor_readiness_gate_blocks_missing_report_even_when_url_mismatches():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    missing_report_path = app_module.get_labor_run_dir(run["id"]) / "missing-report.xlsx"
    app_module.update_labor_metadata(
        run["id"],
        {
            "status": "已生成差异报告",
            "comparisonSummary": {
                "conclusionLevel": "pass",
                "conclusionMessage": "核对通过",
                "exceptionCount": 0,
                "pdfAmountTotal": 100,
                "excelAmountTotal": 100,
                "amountDeltaTotal": 0,
            },
            "files": {
                "diffReport": {
                    "filename": "missing-report.xlsx",
                    "path": str(missing_report_path),
                    "downloadUrl": f"/api/labor/runs/{run['id']}/download/other-report.xlsx",
                }
            },
            "diffDownloadUrl": f"/api/labor/runs/{run['id']}/download/missing-report.xlsx",
        },
    )

    gate = client.get(f"/api/labor/runs/{run['id']}").json()["readinessGate"]

    assert gate["status"] == "blocked"
    assert any(issue["code"] == "report_url_mismatch" for issue in gate["issues"])
    assert any(issue["code"] == "report_file_missing" for issue in gate["issues"])


def test_labor_reocr_candidate_replay_api_blocks_employee_level_exceptions():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay",
        json={
            "task": {
                "sourceFile": "elog1-1_20260520204104.pdf",
                "warehouseId": "1",
                "expectedExcelAmount": 100,
                "amountDelta": 50,
            },
            "candidateRows": [
                {
                    "employeeName": "Wrong Worker",
                    "sourcePageOrRow": "p1",
                    "hours": 8,
                    "amount": 100,
                    "currency": "USD",
                    "confidence": 0.96,
                    "evidenceText": "Wrong Worker 8.00 $100.00",
                }
            ],
            "amountTolerance": 0.1,
            "hoursTolerance": 0.1,
            "confidenceThreshold": 0.85,
        },
    )

    assert replay.status_code == 200
    body = replay.json()
    assert body["decision"] == "blocked_by_replay"
    assert body["summary"]["amountPassed"] is True
    assert body["summary"]["exceptionCount"] > 0
    assert body["blockers"] == ["employee_level_exceptions"]
    assert body["exceptionRows"]
    assert body["nameGate"]["summary"]["confirmed"] == 0
    assert body["nameGate"]["matches"][0]["status"] in {"review", "unmatched"}
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["reocrReplayGovernance"]["replays"][0]["decision"] == "blocked_by_replay"
    assert refreshed["reocrReplayGovernance"]["replays"][0]["nameGate"]["summary"]["confirmed"] == 0

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/confirm",
        json={"sourceFile": "elog1-1_20260520204104.pdf", "warehouseId": "1", "confirmedBy": "ops-user"},
    )
    assert confirmed.status_code == 400
    assert "未通过影响预览" in confirmed.json()["detail"]


def test_labor_reocr_candidate_replay_file_api_parses_upload_and_records_candidate_file():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay-file",
        data={
            "task": json.dumps(
                {
                    "sourceFile": "elog1-1_20260520204104.pdf",
                    "warehouseId": "1",
                    "expectedExcelAmount": 100,
                    "amountDelta": 50,
                }
            ),
            "amount_tolerance": "0.1",
            "hours_tolerance": "0.1",
            "confidence_threshold": "0.85",
        },
        files={"candidate_file": ("reocr.csv", _reocr_csv_bytes(), "text/csv")},
    )

    assert replay.status_code == 200
    body = replay.json()
    assert body["decision"] == "ready_for_user_confirmation"
    assert body["parsedCandidateRowCount"] == 1
    assert body["parsedCandidateRows"][0]["employeeName"] == "Alice Worker"
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["reocrReplayGovernance"]["replays"][0]["candidateFile"] == "reocr.csv"
    assert refreshed["files"]["reocrCandidateFiles"][0]["filename"].endswith(".csv")


def test_labor_reocr_candidate_replay_cache_uses_local_ai_cache_without_applying_result():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )
    pdf_path = Path(upload["files"]["pdfInvoices"][0]["path"])
    cache_dir = pdf_path.parent / ".ai_extract_cache"
    cache_dir.mkdir(exist_ok=True)
    cache_path = cache_dir / f"{pdf_path.stem}_p1_mimo-v2.5_v4.json"
    cache_path.write_text(
        json.dumps(
            [
                {
                    "source_file": pdf_path.name,
                    "source_page_or_row": "1",
                    "employee_name_raw": "Alice Worker",
                    "hours": 8,
                    "amount": 100,
                    "currency": "USD",
                    "confidence": 0.95,
                    "evidence_text": "Alice Worker | Reg Time 8.00 | TOTAL $100.00",
                    "warehouse_id": "1",
                }
            ]
        ),
        encoding="utf-8",
    )

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay-cache",
        json={
            "task": {
                "sourceFile": pdf_path.name,
                "warehouseId": "1",
                "expectedExcelAmount": 100,
                "amountDelta": 0,
            }
        },
    )

    assert replay.status_code == 200
    body = replay.json()
    assert body["mode"] == "ai_cache_candidate_replay"
    assert body["candidateSource"] == "local_ai_page_cache"
    assert body["decision"] == "ready_for_user_confirmation"
    assert body["summary"]["candidateRowCount"] == 1
    assert body["cacheFiles"] == [cache_path.name]
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["reocrReplayGovernance"]["replays"][0]["mode"] == "ai_cache_candidate_replay"
    assert refreshed["comparisonSummary"] == {}


def test_labor_reocr_candidate_batch_replay_cache_records_each_candidate_without_confirming():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    upload = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("elog2-2_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_two_warehouses(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )
    pdf_paths = [Path(record["path"]) for record in upload["files"]["pdfInvoices"]]
    for pdf_path, name, hours, amount, warehouse in [
        (pdf_paths[0], "Alice Worker", 8, 100, "1"),
        (pdf_paths[1], "Wrong Worker", 10, 200, "2"),
    ]:
        cache_dir = pdf_path.parent / ".ai_extract_cache"
        cache_dir.mkdir(exist_ok=True)
        (cache_dir / f"{pdf_path.stem}_p1_mimo-v2.5_v4.json").write_text(
            json.dumps(
                [
                    {
                        "source_file": pdf_path.name,
                        "source_page_or_row": "1",
                        "employee_name_raw": name,
                        "hours": hours,
                        "amount": amount,
                        "currency": "USD",
                        "confidence": 0.95,
                        "evidence_text": f"{name} | TOTAL ${amount:.2f}",
                        "warehouse_id": warehouse,
                    }
                ]
            ),
            encoding="utf-8",
        )

    app_module.update_labor_metadata(
        run["id"],
        {
            "reocrPlan": {
                "reviewableCandidates": [
                    {"sourceFile": pdf_paths[0].name, "warehouseId": "1", "currentCacheAmount": 100, "expectedExcelAmount": 100, "amountDelta": 0},
                    {"sourceFile": pdf_paths[1].name, "warehouseId": "2", "currentCacheAmount": 200, "expectedExcelAmount": 200, "amountDelta": 0},
                ]
            }
        },
    )

    replay = client.post(f"/api/labor/runs/{run['id']}/reocr-candidates/replay-cache-batch", json={})

    assert replay.status_code == 200
    body = replay.json()
    assert body["decision"] == "batch_cache_replay_completed"
    assert body["summary"]["candidateCount"] == 2
    assert body["summary"]["replayedCount"] == 2
    assert body["summary"]["readyCount"] == 1
    assert body["summary"]["blockedCount"] == 1
    assert body["summary"]["errorCount"] == 0
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert [item["mode"] for item in refreshed["reocrReplayGovernance"]["replays"]] == ["ai_cache_candidate_replay", "ai_cache_candidate_replay"]
    assert refreshed["reocrReplayGovernance"].get("activeCandidates", []) == []
    assert refreshed["comparisonSummary"] == {}


def test_labor_reocr_candidate_template_api_generates_downloadable_csv_from_excel_rows():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )

    response = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/template",
        json={"task": {"sourceFile": "elog1-1_20260520204104.pdf", "warehouseId": "1"}},
    )

    assert response.status_code == 200
    template = response.json()
    assert template["label"] == "图片识别结果模板"
    assert template["filename"].endswith(".csv")
    downloaded = client.get(template["downloadUrl"])
    assert downloaded.status_code == 200
    text = downloaded.content.decode("utf-8-sig")
    assert "SourceFile,WarehouseId,EmployeeId,Employee,Hours,Amount,Page,Confidence,Currency,Evidence,ExcelRef,ExpectedHours,ExpectedAmount" in text
    assert "elog1-1_20260520204104.pdf,1,WUS000001,Alice Worker,8.00,100.00,,0.95,USD" in text
    assert "员工账单!2,8.00,100.00" in text
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["files"]["reocrCandidateTemplates"][0]["filename"] == template["filename"]


def test_labor_reocr_candidate_batch_template_api_exports_all_task_rows():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("elog2-2_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_two_warehouses(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "reocrPlan": {
                "tasks": [
                    {"sourceFile": "elog1-1_20260520204104.pdf", "warehouseId": "1"},
                    {"sourceFile": "elog2-2_20260520204104.pdf", "warehouseId": "2"},
                ]
            }
        },
    )

    response = client.post(f"/api/labor/runs/{run['id']}/reocr-candidates/template-batch", json={})

    assert response.status_code == 200
    template = response.json()
    assert template["label"] == "图片识别批量结果模板"
    assert template["summary"] == {"taskCount": 2, "rowCount": 2, "missingTaskCount": 0}
    downloaded = client.get(template["downloadUrl"])
    assert downloaded.status_code == 200
    text = downloaded.content.decode("utf-8-sig")
    assert "SourceFile,WarehouseId,EmployeeId,Employee,Hours,Amount,Page,Confidence,Currency,Evidence,ExcelRef,ExpectedHours,ExpectedAmount" in text
    assert "elog1-1_20260520204104.pdf,1,WUS000001,Alice Worker,8.00,100.00,,0.95,USD" in text
    assert "elog2-2_20260520204104.pdf,2,WUS000002,Bob Worker,10.00,200.00,,0.95,USD" in text
    assert "员工账单!2,8.00,100.00" in text
    assert "员工账单!3,10.00,200.00" in text
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["files"]["reocrCandidateTemplates"][0]["filename"] == template["filename"]


def test_labor_reocr_candidate_batch_upload_replays_each_source_scope_without_confirming():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("elog2-2_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_two_warehouses(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "reocrPlan": {
                "tasks": [
                    {"sourceFile": "elog1-1_20260520204104.pdf", "warehouseId": "1", "expectedExcelAmount": 100, "amountDelta": 0},
                    {"sourceFile": "elog2-2_20260520204104.pdf", "warehouseId": "2", "expectedExcelAmount": 200, "amountDelta": 0},
                ]
            }
        },
    )

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay-file-batch",
        files={"candidate_file": ("filled_batch.csv", _reocr_batch_csv_bytes(), "text/csv")},
    )

    assert replay.status_code == 200
    body = replay.json()
    assert body["decision"] == "batch_file_replay_completed"
    assert body["summary"]["groupCount"] == 2
    assert body["summary"]["replayedCount"] == 2
    assert body["summary"]["readyCount"] == 1
    assert body["summary"]["blockedCount"] == 1
    assert body["summary"]["errorCount"] == 0
    assert body["summary"]["parsedRowCount"] == 2
    assert body["summary"]["plannedTaskCount"] == 2
    assert body["summary"]["coveredTaskCount"] == 2
    assert body["summary"]["missingTaskCount"] == 0
    assert body["summary"]["extraScopeCount"] == 0
    assert body["coverage"]["coverageComplete"] is True
    assert body["coverage"]["missingTasks"] == []
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert len(refreshed["reocrReplayGovernance"]["replays"]) == 2
    assert {item["decision"] for item in refreshed["reocrReplayGovernance"]["replays"]} == {"ready_for_user_confirmation", "blocked_by_replay"}
    assert refreshed["reocrReplayGovernance"].get("activeCandidates", []) == []
    assert refreshed["files"]["reocrCandidateFiles"][0]["label"] == "图片识别批量结果文件"
    assert refreshed["comparisonSummary"] == {}


def test_labor_reocr_candidate_batch_upload_reports_missing_planned_tasks():
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("pdf_files", ("elog2-2_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_two_warehouses(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "reocrPlan": {
                "tasks": [
                    {"sourceFile": "elog1-1_20260520204104.pdf", "warehouseId": "1", "expectedExcelAmount": 100, "amountDelta": 0},
                    {"sourceFile": "elog2-2_20260520204104.pdf", "warehouseId": "2", "expectedExcelAmount": 200, "amountDelta": 0},
                ]
            }
        },
    )
    partial_csv = (
        "SourceFile,WarehouseId,EmployeeId,Employee,Hours,Amount,Page,Confidence,Currency,Evidence\n"
        "elog1-1_20260520204104.pdf,1,WUS000001,Alice Worker,8,100,p1,96%,USD,Alice Worker 8 $100\n"
    ).encode("utf-8")

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay-file-batch",
        files={"candidate_file": ("partial_batch.csv", partial_csv, "text/csv")},
    )

    assert replay.status_code == 200
    body = replay.json()
    assert body["summary"]["plannedTaskCount"] == 2
    assert body["summary"]["coveredTaskCount"] == 1
    assert body["summary"]["missingTaskCount"] == 1
    assert body["coverage"]["coverageComplete"] is False
    assert body["coverage"]["missingTasks"] == [{"sourceFile": "elog2-2_20260520204104.pdf", "warehouseId": "2"}]
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["files"]["reocrCandidateFiles"][0]["coverage"]["missingTasks"] == body["coverage"]["missingTasks"]


def test_labor_reocr_candidate_template_upload_confirm_end_to_end_flow():
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("elog1-1_20260520204104.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes_with_warehouse(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={
            "sheet_name": "员工账单",
            "mapping": {
                "employeeId": "工号",
                "name": "姓名",
                "hours": "时长总计(H)",
                "amount": "费用总计(含税)",
                "currency": "币种",
            },
        },
    )
    task = {"sourceFile": "elog1-1_20260520204104.pdf", "warehouseId": "1", "expectedExcelAmount": 100, "amountDelta": 50}

    template = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/template",
        json={"task": task},
    ).json()
    template_bytes = client.get(template["downloadUrl"]).content

    replay = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/replay-file",
        data={
            "task": json.dumps(task),
            "amount_tolerance": "0.1",
            "hours_tolerance": "0.1",
            "confidence_threshold": "0.85",
        },
        files={"candidate_file": ("filled_reocr_template.csv", template_bytes, "text/csv")},
    )

    assert replay.status_code == 200
    assert replay.json()["decision"] == "ready_for_user_confirmation"

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/reocr-candidates/confirm",
        json={
            "sourceFile": task["sourceFile"],
            "warehouseId": task["warehouseId"],
            "confirmedBy": "ops-user",
            "reason": "template upload replay passed",
            "generateReport": True,
        },
    )

    assert confirmed.status_code == 200
    body = confirmed.json()
    assert body["decision"] == "active"
    assert body["reportFile"]["label"] == "图片识别结果预览报告"
    assert client.get(body["reportFile"]["downloadUrl"]).status_code == 200
    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["files"]["reocrCandidateTemplates"][0]["filename"].endswith(".csv")
    assert refreshed["files"]["reocrCandidateFiles"][0]["filename"].endswith(".csv")
    assert refreshed["files"]["reocrPreviewReport"]["filename"].endswith(".xlsx")
    assert refreshed["reocrReplayGovernance"]["activeCandidates"][0]["confirmationReason"] == "template upload replay passed"


def test_labor_rule_candidate_auto_replay_uses_historical_run_metadata(monkeypatch):
    import bonus_platform.app as app_module

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "OSS", "period_start": "2026-05-18", "period_end": "2026-05-24", "currency": "USD"},
    ).json()
    monkeypatch.setattr(
        app_module,
        "list_labor_metadata",
        lambda: [
            {
                "id": run["id"],
                "supplierName": "OSS",
                "periodStart": "2026-05-18",
                "periodEnd": "2026-05-24",
                "reconciliationDiagnostics": {
                    "level": "warning",
                    "issues": [{"code": "missing_warehouse_id", "level": "warning"}],
                },
                "comparisonSummary": {"exceptionCount": 0},
            },
            {
                "id": "fairway_clean",
                "supplierName": "Fairway",
                "periodStart": "2026-05-11",
                "periodEnd": "2026-05-17",
                "reconciliationDiagnostics": {"level": "ok", "issues": []},
                "comparisonSummary": {"exceptionCount": 0},
            },
        ],
    )

    client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates",
        json={
            "ruleId": "oss-hash-warehouse-v1",
            "title": "OSS # warehouse id extraction",
            "description": "Parse #N warehouse id from invoice filenames.",
            "source": "real replay: oss 2",
            "conditions": {"supplier": "OSS", "fixIssueCodes": ["missing_warehouse_id"]},
        },
    )

    replay = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/oss-hash-warehouse-v1/auto-replay",
        json={"limit": 10},
    )

    assert replay.status_code == 200
    assert replay.json()["mode"] == "metadata_signal_replay"
    assert replay.json()["decision"] == "ready_for_user_confirmation"
    assert replay.json()["summary"]["fixedCount"] == 1
    assert replay.json()["summary"]["regressionCount"] == 0

    confirmed = client.post(
        f"/api/labor/runs/{run['id']}/rule-candidates/oss-hash-warehouse-v1/confirm",
        json={"confirmedBy": "ops-user", "reason": "auto replay clean"},
    )
    assert confirmed.status_code == 200
    assert confirmed.json()["decision"] == "active"


def test_labor_compare_records_extraction_quality_warning_for_misaligned_totals(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(
        app_module,
        "quick_extract_totals",
        lambda paths, *args, **kwargs: [
            {
                "source_file": paths[0].name,
                "total_amount": 50.0,
                "warehouse_id": "",
                "authoritative": True,
                "evidence_status": "authoritative",
            }
        ],
    )
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda *args, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file=Path(args[0][0]).name, source_page_or_row="p1", employee_id="", employee_name_raw="Alvarez Mitrache, Rosa", hours=10, amount=100, currency="USD", confidence=0.95, evidence_text="Total $100")
        ],
    )
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert body["status"] == "部分核对完成"
    assert body["batchGuard"]["status"] == "partial_review"
    assert body["extractionQuality"]["level"] == "critical"
    assert body["warehouseComparison"]["rows"][0]["reconciliationStatus"] == "needs_review"
    assert any("总金额差异" in issue for issue in body["extractionQuality"]["issues"])
    assert "必须人工复核" in body["extractionQuality"]["message"]


def test_labor_compare_uses_excel_candidates_on_initial_extract(monkeypatch):
    import bonus_platform.app as app_module

    monkeypatch.setattr(app_module, "quick_extract_totals", lambda paths, *args, **kwargs: _unresolved_quick_totals(paths))

    calls = []

    def fake_extract(*args, **kwargs):
        calls.append(kwargs)
        return [
            LaborLineItem(source_type="pdf_invoice", source_file="scan.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Rosa Alvarez Minchaca", hours=31.19, amount=701.90, currency="USD", confidence=0.9, evidence_text="initial")
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = client.get(f"/api/labor/runs/{run['id']}").json()
    assert len(calls) == 1
    assert calls[0]["expected_rows"][0]["employee_name"] == "Rosa Alvarez Minchaca"
    assert body["extractionQuality"]["level"] == "critical"
    assert body["warehouseComparison"]["rows"][0]["reconciliationStatus"] == "needs_review"
    assert body["extractionQuality"].get("retryApplied") is not True
    assert body["comparisonSummary"]["exceptionCount"] == 0


def test_labor_extraction_quality_passes_when_counts_and_totals_align():
    from bonus_platform.engine.labor.quality import calculate_extraction_quality

    quality = calculate_extraction_quality(
        [],  # No PDF rows needed for this test
        {
            "pdfEmployeeCount": 161,
            "excelEmployeeCount": 161,
            "pdfHoursTotal": 5912.62,
            "excelHoursTotal": 5912.62,
            "pdfAmountTotal": 150078.21,
            "excelAmountTotal": 150119.51,
            "unmatchedPdfCount": 0,
            "unmatchedExcelCount": 0,
        }
    )

    assert quality["level"] == "ok"
    assert quality["message"] == "抽取质量检查通过。"
    assert quality["issues"] == []


def test_labor_compare_endpoint_returns_running_status_before_polling(monkeypatch):
    import bonus_platform.app as app_module

    queued = {}

    # 后台任务现在通过 run_in_executor 运行，monkeypatch 替换为同步调用以便测试
    monkeypatch.setattr(app_module, "_run_labor_extract_compare", lambda run_id: queued.setdefault("completed", run_id))
    # 拦截 run_in_executor，直接同步调用
    import asyncio
    try:
        event_loop = asyncio.get_event_loop()
    except RuntimeError:
        event_loop = asyncio.new_event_loop()
        asyncio.set_event_loop(event_loop)
    original_run_in_executor = event_loop.run_in_executor
    def fake_run_in_executor(executor, fn, *args):
        fn(*args)
        # 返回一个已完成的 future
        f = asyncio.Future()
        f.set_result(None)
        return f
    monkeypatch.setattr(event_loop, "run_in_executor", fake_run_in_executor)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-05-11", "period_end": "2026-05-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("scan.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("账单.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare").json()

    assert response["status"] == "抽取中"
    assert queued.get("completed") == run["id"]


def test_labor_upload_syncs_files_to_supabase_storage(monkeypatch, tmp_path):
    from bonus_platform.engine.labor import runs as labor_runs

    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", tmp_path / "labor_runs")
    monkeypatch.setenv("SIGMA_LABOR_STORAGE_BACKEND", "supabase")
    monkeypatch.setenv("SUPABASE_URL", "https://example.supabase.co")
    monkeypatch.setenv("SUPABASE_SERVICE_ROLE_KEY", "service-role")
    monkeypatch.setenv("SIGMA_LABOR_SUPABASE_BUCKET", "labor-uat")
    synced = []

    def fake_sync(run_id, run_dir):
        synced.append((run_id, sorted(path.name for path in run_dir.iterdir() if path.is_file())))

    monkeypatch.setattr(labor_runs, "sync_labor_run_to_persistent", fake_sync)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-06-17", "period_end": "2026-06-17", "currency": "USD"},
    ).json()

    response = client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )

    assert response.status_code == 200
    assert response.json()["storage"]["backend"] == "supabase"
    assert any("metadata.json" in names and any(name.endswith(".pdf") for name in names) for _, names in synced)


def test_labor_extract_task_restores_persistent_files_before_processing(monkeypatch, tmp_path):
    from bonus_platform.engine.labor import runs as labor_runs

    labor_root = tmp_path / "labor_runs"
    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", labor_root)
    monkeypatch.setenv("SIGMA_LABOR_STORAGE_BACKEND", "supabase")
    monkeypatch.setenv("SUPABASE_URL", "https://example.supabase.co")
    monkeypatch.setenv("SUPABASE_SERVICE_ROLE_KEY", "service-role")
    monkeypatch.setenv("SIGMA_LABOR_SUPABASE_BUCKET", "labor-uat")

    snapshots = {}

    def fake_sync_to(run_id, run_dir):
        snapshots[run_id] = {
            path.relative_to(run_dir).as_posix(): path.read_bytes()
            for path in run_dir.rglob("*")
            if path.is_file()
        }

    def fake_sync_from(run_id, run_dir):
        for relative, content in snapshots.get(run_id, {}).items():
            target = run_dir / relative
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(content)
        return bool(snapshots.get(run_id))

    monkeypatch.setattr(labor_runs, "sync_labor_run_to_persistent", fake_sync_to)
    monkeypatch.setattr(app_module, "sync_labor_run_from_persistent", fake_sync_from)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-06-17", "period_end": "2026-06-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    run_dir = labor_runs.get_labor_run_dir(run["id"])
    for path in run_dir.glob("*"):
        if path.is_file() and path.name != "metadata.json":
            path.unlink()

    def fake_perform(run_id):
        metadata = labor_runs.load_labor_metadata(labor_runs.get_labor_run_dir(run_id))
        files = metadata["files"]
        assert Path(files["pdfInvoices"][0]["path"]).exists()
        assert Path(files["workbooks"][0]["path"]).exists()
        app_module.update_labor_metadata(
            run_id,
            {
                "status": "已生成差异报告",
                "stage": "生成报告",
                "diffDownloadUrl": "/api/labor/runs/fake/download/report.xlsx",
            },
        )

    monkeypatch.setattr(app_module, "_perform_labor_extract_compare", fake_perform)

    app_module._run_labor_extract_compare(run["id"])

    refreshed = client.get(f"/api/labor/runs/{run['id']}").json()
    assert refreshed["asyncTask"]["status"] == "completed"
    assert refreshed["asyncTask"]["statusLabel"] == "完成"


def test_labor_extract_endpoint_returns_queued_task_status(monkeypatch):
    monkeypatch.setattr(app_module, "_run_labor_extract_compare", lambda run_id: None)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-06-17", "period_end": "2026-06-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = response.json()
    assert body["asyncTask"]["status"] == "queued"
    assert body["asyncTask"]["statusLabel"] == "待处理"
    assert body["businessReviewStatus"] == "pending"
    assert body["manualReviewRequired"] is True
    assert body["directPaymentAllowed"] is False
    assert body["requiresHumanReview"] is True


def test_labor_personal_worker_job_carries_required_worker_version(monkeypatch):
    captured = {}
    monkeypatch.setenv("SIGMA_LABOR_REQUIRED_WORKER_VERSION", "0.3.0")
    monkeypatch.setattr(app_module, "_uses_personal_labor_worker", lambda: True)

    def fake_enqueue(
        run_id,
        *,
        owner_user_id,
        required_worker_version="",
        max_attempts=3,
        task_generation_id="",
    ):
        captured.update(
            {
                "runId": run_id,
                "ownerUserId": owner_user_id,
                "requiredWorkerVersion": required_worker_version,
                "taskGenerationId": task_generation_id,
            }
        )
        return {
            "id": "labor_job_test",
            "runId": run_id,
            "ownerUserId": owner_user_id,
            "status": "queued",
            "requiredWorkerVersion": required_worker_version,
            "taskGenerationId": task_generation_id,
        }

    monkeypatch.setattr(app_module, "enqueue_labor_worker_job", fake_enqueue)
    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-06-17", "period_end": "2026-06-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    assert captured["requiredWorkerVersion"] == app_module.OVERSEAS_LABOR_REQUIRED_WORKER_VERSION
    assert captured["taskGenerationId"]
    assert response.json()["workerTask"]["requiredWorkerVersion"] == app_module.OVERSEAS_LABOR_REQUIRED_WORKER_VERSION


def test_labor_extract_endpoint_reuses_running_task(monkeypatch):
    def fail_if_started(run_id):
        raise AssertionError("running task should not be started twice")

    monkeypatch.setattr(app_module, "_run_labor_extract_compare", fail_if_started)

    client = TestClient(app)
    run = client.post(
        "/api/labor/runs",
        json={"supplier_name": "ONESOURCE", "period_start": "2026-06-17", "period_end": "2026-06-17", "currency": "USD"},
    ).json()
    client.post(
        f"/api/labor/runs/{run['id']}/files",
        files=[
            ("pdf_files", ("invoice.pdf", b"%PDF-1.4\n", "application/pdf")),
            ("workbook_files", ("bill.xlsx", _excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ],
    )
    client.post(
        f"/api/labor/runs/{run['id']}/mapping",
        json={"sheet_name": "员工账单", "mapping": {"employeeId": "工号", "name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"}},
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "status": "抽取中",
            "asyncTask": {
                "status": "running",
                "statusLabel": "处理中",
                "message": "后台正在生成核对结果。",
            },
        },
    )

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "抽取中"
    assert body["asyncTask"]["status"] == "running"


def test_labor_extract_endpoint_clears_stale_invoice_evidence_audit_when_queued(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    app_module.update_labor_metadata(
        run["id"],
        {
            "status": "已生成差异报告",
            "invoiceEvidenceAudit": [{"source_file": "stale.pdf", "authoritative": True}],
        },
    )
    monkeypatch.setattr(app_module, "_run_labor_extract_compare", lambda run_id: None)

    response = client.post(f"/api/labor/runs/{run['id']}/extract-and-compare")

    assert response.status_code == 200
    assert response.json()["invoiceEvidenceAudit"] == []
    assert client.get(f"/api/labor/runs/{run['id']}").json()["invoiceEvidenceAudit"] == []


def test_labor_run_retains_fresh_invoice_evidence_audit_when_later_stage_fails(monkeypatch):
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    app_module.update_labor_metadata(
        run["id"],
        {"invoiceEvidenceAudit": [{"source_file": "stale.pdf", "authoritative": True}]},
    )
    fresh_audit: list[dict] = []

    def fake_quick_extract(paths, *args, **kwargs):
        fresh_audit.extend(
            [
                {
                    "source_file": Path(paths[0]).name,
                    "warehouse_id": "1",
                    "total_amount": 100.0,
                    "authoritative": True,
                    "evidence_status": "authoritative",
                    "page_evidence": [{"page": 1, "role": "invoice_total"}],
                }
            ]
        )
        return list(fresh_audit)

    def fail_after_stage1(*args, **kwargs):
        persisted = client.get(f"/api/labor/runs/{run['id']}").json()
        assert persisted["invoiceEvidenceAudit"] == fresh_audit
        raise RuntimeError("comparison failed after invoice evidence")

    monkeypatch.setattr(app_module, "quick_extract_totals", fake_quick_extract)
    monkeypatch.setattr(app_module, "compare_by_warehouse", fail_after_stage1)

    result = app_module._run_labor_extract_compare(run["id"])
    saved = client.get(f"/api/labor/runs/{run['id']}").json()

    assert result is False
    assert saved["status"] == "抽取失败"
    assert saved["invoiceEvidenceAudit"] == fresh_audit
    assert "comparison failed after invoice evidence" in saved["errorMessage"]


def test_labor_structure_fallback_promotes_unknown_language_invoice(monkeypatch):
    totals = [
        {
            "source_file": "invoice.pdf",
            "warehouse_id": "",
            "total_amount": 0.0,
            "authoritative": False,
            "pdf_type": "unknown",
            "evidence_status": "needs_review",
            "page_evidence": [{"page": 1, "role": "unknown"}],
        }
    ]
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)
    monkeypatch.setattr(
        labor_structure,
        "extract_page_texts",
        lambda _: [
            "1 A One 8 Unit10,000 20,00\n"
            "2 B Two 8 Unit10,000 20,00\n"
            "3 C Three 8 Unit10,000 20,00\n"
            "4 D Four 8 Unit10,000 20,00\n"
            "5 E Five 8 Unit10,000 20,00\n"
            "100,00 19,00 119,00"
        ],
    )

    expected_row_counts = []
    ai_extract_calls = []

    def fake_extract(paths, *args, **kwargs):
        ai_extract_calls.append([Path(path).name for path in paths])
        expected_row_counts.append(len(kwargs.get("expected_rows") or []))
        source_file = Path(paths[0]).name
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=source_file,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw=name,
                hours=8,
                amount=20,
                currency="EUR",
                confidence=0.95,
                evidence_text=name,
                warehouse_id="",
            )
            for name in ["A One", "B Two", "C Three", "D Four", "E Five"]
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf"],
        workbook_bytes=_excel_bytes_for_structure_fallback(),
    )
    stored_name = Path(client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"][0]["path"]).name
    totals[0]["source_file"] = stored_name

    saved = app_module._perform_labor_extract_compare(run["id"])

    promoted = saved["invoiceEvidenceAudit"][0]
    assert promoted["total_amount"] == 100.0
    assert promoted["warehouse_id"] == "15"
    assert promoted["authoritative"] is True
    assert len(saved["pdfExtractedRows"]) == 5
    assert saved["structureReconciliation"]["reconciledFiles"] == [stored_name]
    assert saved["warehouseComparison"]["summary"]["totalPassed"] is True
    assert ai_extract_calls == []
    assert expected_row_counts == []
    assert saved["status"] == "待币种确认"
    assert saved["files"]["diffReport"]["label"] == "币种待确认报告"
    assert "businessReport" not in saved["files"]
    assert saved["businessReportDownloadUrl"] == ""


def test_labor_run_keeps_page_role_support_pdf_audit_only_in_mixed_batch(monkeypatch):
    totals = [
        {
            "source_file": "invoice.pdf",
            "warehouse_id": "1",
            "total_amount": 150.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "page_evidence": [{"page": 1, "role": "invoice_primary"}],
        },
        {
            "source_file": "support.pdf",
            "warehouse_id": "2",
            "total_amount": 0.0,
            "authoritative": False,
            "pdf_type": "unknown",
            "evidence_status": "needs_review",
            "page_evidence": [
                {"page": 1, "role": "email_cover"},
                {"page": 2, "role": "timecard_summary"},
            ],
        },
    ]
    extracted_files: list[str] = []
    diagnostics_totals: list[dict] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)
    original_diagnostics = app_module.build_reconciliation_diagnostics

    def capture_diagnostics(*args, **kwargs):
        diagnostics_totals.extend(kwargs.get("pdf_totals") or [])
        return original_diagnostics(*args, **kwargs)

    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(paths[0]).name,
                source_page_or_row="p1",
                employee_id="WUS000001",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=150,
                currency="USD",
                warehouse_id="1",
            )
        ]

    monkeypatch.setattr(app_module, "build_reconciliation_diagnostics", capture_diagnostics)
    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf", "support.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    stored_names = [
        Path(row["path"]).name
        for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]
    ]
    totals[0]["source_file"], totals[1]["source_file"] = stored_names

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert extracted_files == [stored_names[0]]
    assert [row["warehouseId"] for row in saved["warehouseComparison"]["rows"]] == ["1"]
    assert [row["source_file"] for row in diagnostics_totals] == [stored_names[0]]
    assert [row["source_file"] for row in saved["invoiceEvidenceAudit"]] == stored_names


def test_labor_run_keeps_legacy_detail_pdf_out_of_reconciliation_and_ocr(monkeypatch):
    totals = [
        {
            "source_file": "invoice.pdf",
            "warehouse_id": "1",
            "total_amount": 150.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "page_evidence": [{"page": 1, "role": "invoice_primary"}],
        },
        {
            "source_file": "legacy_detail.pdf",
            "warehouse_id": "2",
            "total_amount": 0.0,
            "authoritative": False,
            "pdf_type": "unknown",
            "evidence_status": "needs_review",
            "page_evidence": [{"page": 1, "role": "unknown"}],
        },
    ]
    extracted_files: list[str] = []
    diagnostics_totals: list[dict] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)
    original_diagnostics = app_module.build_reconciliation_diagnostics

    def capture_diagnostics(*args, **kwargs):
        diagnostics_totals.extend(kwargs.get("pdf_totals") or [])
        return original_diagnostics(*args, **kwargs)

    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(paths[0]).name,
                source_page_or_row="p1",
                employee_id="WUS000001",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=150,
                currency="USD",
                warehouse_id="1",
            )
        ]

    monkeypatch.setattr(app_module, "build_reconciliation_diagnostics", capture_diagnostics)
    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["invoice.pdf", "legacy_detail.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    stored_names = [
        Path(row["path"]).name
        for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]
    ]
    totals[0]["source_file"], totals[1]["source_file"] = stored_names

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert extracted_files == [stored_names[0]]
    assert [row["warehouseId"] for row in saved["warehouseComparison"]["rows"]] == ["1"]
    assert [row["source_file"] for row in diagnostics_totals] == [stored_names[0]]
    assert [row["source_file"] for row in saved["invoiceEvidenceAudit"]] == stored_names


def test_labor_run_does_not_treat_partial_zero_totals_as_valid(monkeypatch):
    totals = [
        {
            "source_file": "DEPT_1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [{"page": 1, "role": "invoice_total"}],
        },
        {
            "source_file": "mystery.pdf",
            "warehouse_id": "",
            "total_amount": 0.0,
            "authoritative": False,
            "evidence_status": "needs_review",
            "total_page": None,
            "page_evidence": [{"page": 1, "role": "unknown"}],
        },
    ]
    extracted_files: list[str] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)

    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(paths[0]).name,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Review Worker",
                hours=1,
                amount=1,
                currency="USD",
                warehouse_id="",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf", "mystery.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    stored_names = [Path(row["path"]).name for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]]
    totals[0]["source_file"], totals[1]["source_file"] = stored_names

    app_module._perform_labor_extract_compare(run["id"])
    saved = client.get(f"/api/labor/runs/{run['id']}").json()

    assert saved["warehouseComparison"]["summary"]["totalPassed"] is False
    assert stored_names[0] in extracted_files
    assert stored_names[1] in extracted_files
    unresolved = next(row for row in saved["warehouseComparison"]["rows"] if row["warehouseId"] == "")
    assert unresolved["reconciliationStatus"] == "needs_review"
    assert unresolved["pdfEvidenceFile"] == stored_names[1]
    assert saved["status"] == "部分核对完成"
    assert saved["batchGuard"]["status"] == "employee_detail_incomplete"
    assert saved["comparisonSummary"]["canRelease"] is False


def test_labor_run_preserves_all_unresolved_totals_as_review_evidence(monkeypatch):
    totals = [
        {
            "source_file": f"DEPT_{warehouse}.pdf",
            "warehouse_id": warehouse,
            "total_amount": 0.0,
            "authoritative": False,
            "evidence_status": "needs_review",
            "total_page": None,
            "page_evidence": [{"page": 1, "role": "invoice_primary"}],
        }
        for warehouse in ("1", "2")
    ]
    extracted_files: list[str] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)

    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(path).name,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw=f"Review Worker {index}",
                hours=1,
                amount=1,
                currency="USD",
                warehouse_id=str(index),
            )
            for index, path in enumerate(paths, start=1)
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf", "DEPT_2.pdf"],
        workbook_bytes=_excel_bytes_with_two_warehouses(),
    )
    stored_names = [Path(row["path"]).name for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]]
    for total, source_file in zip(totals, stored_names):
        total["source_file"] = source_file

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert extracted_files == stored_names
    assert [row["reconciliationStatus"] for row in saved["warehouseComparison"]["rows"]] == [
        "needs_review",
        "needs_review",
    ]
    assert [row["pdfEvidenceFile"] for row in saved["warehouseComparison"]["rows"]] == [
        *stored_names,
    ]
    assert saved["status"] == "PDF识别未完成"
    assert saved["batchGuard"]["status"] == "pdf_recognition_incomplete"
    assert saved["comparisonSummary"]["canRelease"] is False
    assert saved["businessReportDownloadUrl"] == ""
    assert "businessReport" not in saved["files"]


def test_labor_run_formal_scope_extracts_employee_rows_for_all_payable_warehouses(monkeypatch):
    totals = [
        {
            "source_file": "DEPT_1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [{"page": 1, "role": "invoice_total"}],
        },
        {
            "source_file": "DEPT_2.pdf",
            "warehouse_id": "2",
            "total_amount": 250.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [
                {"page": 1, "role": "invoice_total"},
                {"page": 2, "role": "email_cover"},
                {"page": 3, "role": "timecard_summary"},
                {"page": 4, "role": "invoice_total"},
            ],
            "excluded_pages": [2, 3, 4],
        },
    ]
    extracted_files: list[str] = []
    allowed_pages: list[dict[str, set[int]]] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)

    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        allowed_pages.append(kwargs.get("allowed_pages_by_source") or {})
        return [
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[0]).name, source_page_or_row="p1", employee_id="WUS000001", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", warehouse_id="1"),
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[1]).name, source_page_or_row="p1", employee_id="WUS000002", employee_name_raw="Bob Worker", hours=10, amount=250, currency="USD", warehouse_id="2"),
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf", "DEPT_2.pdf"],
        workbook_bytes=_excel_bytes_with_two_warehouses(),
    )
    stored_names = [Path(row["path"]).name for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]]
    for total, source_file in zip(totals, stored_names):
        total["source_file"] = source_file

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert extracted_files == stored_names
    assert allowed_pages == [{stored_names[0]: {1}, stored_names[1]: {1}}]
    rows = {row["warehouseId"]: row for row in saved["warehouseComparison"]["rows"]}
    assert rows["1"]["reconciliationStatus"] == "passed"
    assert {row["warehouse_id"] for row in saved["pdfExtractedRows"]} == {"1", "2"}
    assert rows["2"]["reconciliationStatus"] == "amount_difference"
    assert rows["2"]["employeeRows"]


def test_labor_run_extracts_all_payable_pdfs_when_batch_is_image_only(monkeypatch):
    totals = [
        {"source_file": "DEPT_1.pdf", "warehouse_id": "1", "total_amount": 100.0, "authoritative": True, "evidence_status": "authoritative", "page_evidence": [{"page": 1, "role": "invoice_total"}]},
        {"source_file": "DEPT_2.pdf", "warehouse_id": "2", "total_amount": 200.0, "authoritative": True, "evidence_status": "authoritative", "page_evidence": [{"page": 1, "role": "invoice_total"}]},
    ]
    extracted_files: list[str] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)
    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        return [
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[0]).name, source_page_or_row="p1", employee_id="WUS000001", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", warehouse_id="1"),
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[1]).name, source_page_or_row="p1", employee_id="WUS000002", employee_name_raw="Bob Worker", hours=10, amount=200, currency="USD", warehouse_id="2"),
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf", "DEPT_2.pdf"],
        workbook_bytes=_excel_bytes_with_two_warehouses(),
    )
    monkeypatch.setattr(
        app_module,
        "_summarize_pdf_text_coverage",
        lambda paths: {
            "summary": {"fileCount": len(paths), "textReadableFileCount": 0, "imageOnlyFileCount": len(paths)},
            "files": [{"sourceFile": path.name, "hasTextLayer": False, "needsOcr": True} for path in paths],
        },
    )
    stored_names = [Path(row["path"]).name for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]]
    for total, source_file in zip(totals, stored_names):
        total["source_file"] = source_file

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert extracted_files == stored_names
    assert saved["reconciliationDiagnostics"]["signals"]["pdfDetailCoverage"]["coverageRatio"] == 1.0


def test_labor_run_retries_only_employee_detail_pdf_whose_rows_do_not_close(monkeypatch):
    totals = [
        {
            "source_file": "DEPT_1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [{"page": 1, "role": "invoice_total"}],
        },
        {
            "source_file": "DEPT_2.pdf",
            "warehouse_id": "2",
            "total_amount": 250.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [{"page": 1, "role": "invoice_total"}],
        },
    ]
    calls: list[dict] = []
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)

    def fake_extract(paths, config, *args, **kwargs):
        calls.append(
            {
                "files": [Path(path).name for path in paths],
                "config": dict(config),
                "allowed_pages": kwargs.get("allowed_pages_by_source") or {},
                "has_progress_callback": callable(kwargs.get("progress_callback")),
            }
        )
        if len(paths) > 1:
            return [
                LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[0]).name, source_page_or_row="p1", employee_id="WUS000001", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=0.99, warehouse_id="1"),
                LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[1]).name, source_page_or_row="p1", employee_id="WUS000002", employee_name_raw="Bob Worker", hours=10, amount=50, currency="USD", confidence=0.99, warehouse_id="2"),
            ]
        return [
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[0]).name, source_page_or_row="p1", employee_id="WUS000002", employee_name_raw="Bob Worker", hours=10, amount=250, currency="USD", confidence=0.99, warehouse_id="2")
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf", "DEPT_2.pdf"],
        workbook_bytes=_excel_bytes_with_two_warehouses(),
    )
    stored_names = [
        Path(row["path"]).name
        for row in client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"]
    ]
    for total, source_file in zip(totals, stored_names):
        total["source_file"] = source_file

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert [call["files"] for call in calls] == [stored_names, [stored_names[1]]]
    assert calls[1]["config"]["cache_enabled"] is False
    assert calls[1]["config"]["render_scale"] >= 2.4
    assert calls[1]["allowed_pages"] == {stored_names[1]: {1}}
    assert calls[1]["has_progress_callback"] is True
    assert saved["comparisonSummary"]["pdfAmountTotal"] == 350.0
    assert saved["extractionQuality"]["retryAttempted"] is True
    assert saved["extractionQuality"]["retryApplied"] is True


def test_labor_run_keeps_better_original_rows_when_detail_total_retry_is_worse(monkeypatch):
    totals = [
        {
            "source_file": "DEPT_2.pdf",
            "warehouse_id": "2",
            "total_amount": 250.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [{"page": 1, "role": "invoice_total"}],
        }
    ]
    calls = 0
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)

    def fake_extract(paths, config, *args, **kwargs):
        nonlocal calls
        calls += 1
        return [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=Path(paths[0]).name,
                source_page_or_row="p1",
                employee_id="WUS000002",
                employee_name_raw="Bob Worker",
                hours=10,
                amount=50.0 if calls == 1 else 25.0,
                currency="USD",
                confidence=0.99,
                warehouse_id="2",
            )
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_2.pdf"],
        workbook_bytes=_excel_bytes_with_two_warehouses(),
    )
    stored_name = Path(
        client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"][0]["path"]
    ).name
    totals[0]["source_file"] = stored_name

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert calls == 2
    assert saved["comparisonSummary"]["pdfAmountTotal"] == 50.0
    assert saved["extractionQuality"]["retryAttempted"] is True
    assert saved["extractionQuality"]["retryApplied"] is False
    assert saved["extractionQuality"]["level"] == "critical"
    assert saved["requiresHumanReview"] is True
    assert saved["comparisonSummary"]["canRelease"] is False
    assert any("员工归因不可直接采信" in issue for issue in saved["extractionQuality"]["issues"])


def test_labor_run_keeps_missing_pdf_warehouse_out_of_pdf_ocr(monkeypatch):
    totals = [
        {
            "source_file": "DEPT_1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
            "page_evidence": [{"page": 1, "role": "invoice_total"}],
        }
    ]
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)
    extracted_files: list[str] = []

    def fake_extract(paths, *args, **kwargs):
        extracted_files.extend(Path(path).name for path in paths)
        return [
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[0]).name, source_page_or_row="p1", employee_id="WUS000001", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", warehouse_id="1")
        ]

    monkeypatch.setattr(app_module, "extract_invoice_items", fake_extract)
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf"],
        workbook_bytes=_excel_bytes_with_two_warehouses(),
    )
    totals[0]["source_file"] = Path(
        client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"][0]["path"]
    ).name

    saved = app_module._perform_labor_extract_compare(run["id"])

    rows = {row["warehouseId"]: row for row in saved["warehouseComparison"]["rows"]}
    assert extracted_files == [totals[0]["source_file"]]
    assert rows["1"]["reconciliationStatus"] == "passed"
    assert rows["2"]["reconciliationStatus"] == "missing_pdf_invoice"
    assert saved["warehouseComparison"]["summary"]["totalPassed"] is False


def test_labor_run_persists_invoice_evidence_audit(monkeypatch):
    totals = [
        {
            "source_file": "DEPT_1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 2,
            "total_label": "TOTAL",
            "page_evidence": [
                {"page": 1, "role": "invoice_primary"},
                {"page": 2, "role": "invoice_total"},
            ],
            "excluded_pages": [],
        }
    ]
    monkeypatch.setattr(app_module, "quick_extract_totals", lambda *args, **kwargs: totals)
    monkeypatch.setattr(
        app_module,
        "extract_invoice_items",
        lambda paths, *args, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file=Path(paths[0]).name, source_page_or_row="p1", employee_id="WUS000001", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", warehouse_id="1")
        ],
    )
    client, run = _prepare_labor_orchestration_run(
        monkeypatch,
        pdf_names=["DEPT_1.pdf"],
        workbook_bytes=_excel_bytes_with_warehouse(),
    )
    totals[0]["source_file"] = Path(
        client.get(f"/api/labor/runs/{run['id']}").json()["files"]["pdfInvoices"][0]["path"]
    ).name

    saved = app_module._perform_labor_extract_compare(run["id"])

    assert saved["invoiceEvidenceAudit"] == totals


def test_adaptive_tolerance_for_large_amounts():
    """测试大金额的自适应容忍度"""
    from bonus_platform.engine.labor.compare import _adaptive_tolerance

    # 小金额使用基础容忍度
    assert _adaptive_tolerance(500) == 0.05

    # $1000 边界
    assert _adaptive_tolerance(1000) == 0.05
    assert _adaptive_tolerance(1001) > 0.05

    # 大金额容忍度更高
    tol_50k = _adaptive_tolerance(50000)
    tol_100k = _adaptive_tolerance(100000)
    assert tol_50k > 0.05
    assert tol_100k > tol_50k

    # 仓库19 的 $0.09 差异（$54,689）应在容忍范围内
    tol_54k = _adaptive_tolerance(54689)
    assert 0.09 <= tol_54k, f"$0.09 差异应被容忍, 但容忍度仅为 {tol_54k}"

    # 容忍度不应过高（即使 $1M 也不超过 $1）
    assert _adaptive_tolerance(1_000_000) < 1.0


def test_advanced_name_normalization():
    """测试高级姓名标准化"""
    from bonus_platform.engine.labor.parsing import normalize_employee_name_advanced

    # "Last, First" 格式
    assert normalize_employee_name_advanced("Alvarez, Rosa") == "rosa alvarez"

    # 中间名缩写
    assert normalize_employee_name_advanced("Rosa J. Alvarez") == "rosa alvarez"

    # 多余空格
    assert normalize_employee_name_advanced("  Rosa   Alvarez  ") == "rosa alvarez"

    # 空字符串
    assert normalize_employee_name_advanced("") == ""

    # 单个名字
    assert normalize_employee_name_advanced("Rosa") == "rosa"

    # 三个部分无中间名缩写
    assert normalize_employee_name_advanced("Rosa Maria Alvarez") == "rosa maria alvarez"


def test_parallel_rule_extraction(monkeypatch):
    """测试并行规则抽取 - 通过 extract_invoice_items 并行路径"""
    import bonus_platform.engine.labor.extract as extract_module
    from bonus_platform.engine.labor.extract import extract_invoice_items

    # 模拟多个页面数据
    pages = [
        {
            "source_file": "invoice1.pdf",
            "page": 1,
            "text": "05/01/2026\nAlvarez, Rosa\n8.00\nReg\nREG\n$25.00\n$200.00\n$200.00\n05/02/2026\nSmith, John\n4.00\nOT\nOT\n$37.50\n$150.00\n$150.00"
        },
        {
            "source_file": "invoice2.pdf",
            "page": 1,
            "text": "05/03/2026\nJohnson, Maria\n6.00\nReg\nREG\n$30.00\n$180.00\n$180.00"
        }
    ]

    # Mock _extract_pdf_pages 以返回多页数据
    monkeypatch.setattr(extract_module, "_extract_pdf_pages", lambda pdf_paths, **kw: pages)

    # 配置: 启用并行抽取
    ai_config = {
        "parallel_extraction_enabled": True,
        "parallel_max_workers": 4,
        "enabled": False,  # 禁用 AI，仅走规则路径
    }

    # 调用 extract_invoice_items（并行路径）
    items = extract_invoice_items(
        pdf_paths=[Path("invoice1.pdf"), Path("invoice2.pdf")],
        ai_config=ai_config,
        supplier="Test",
        period_start="2026-05-01",
        period_end="2026-05-31",
        currency="USD",
    )

    # 验证结果 - 两个文件的页面都应被并行处理
    assert len(items) == 3, f"应抽取 3 条记录，实际 {len(items)} 条"
    assert any(item.employee_name_raw == "Alvarez, Rosa" for item in items)
    assert any(item.employee_name_raw == "Smith, John" for item in items)
    assert any(item.employee_name_raw == "Johnson, Maria" for item in items)


def test_parallel_extraction_disabled(monkeypatch):
    """测试禁用并行抽取 - 并行开关关闭后走串行路径"""
    import bonus_platform.engine.labor.extract as extract_module
    from bonus_platform.engine.labor.extract import extract_invoice_items

    # 模拟多个页面数据
    pages = [
        {
            "source_file": "invoice1.pdf",
            "page": 1,
            "text": "05/01/2026\nAlvarez, Rosa\n8.00\nReg\nREG\n$25.00\n$200.00\n$200.00"
        },
        {
            "source_file": "invoice2.pdf",
            "page": 1,
            "text": "05/03/2026\nJohnson, Maria\n6.00\nReg\nREG\n$30.00\n$180.00\n$180.00"
        }
    ]

    # Mock _extract_pdf_pages 以返回多页数据
    monkeypatch.setattr(extract_module, "_extract_pdf_pages", lambda pdf_paths, **kw: pages)

    # 配置: 禁用并行抽取
    ai_config = {
        "parallel_extraction_enabled": False,
        "parallel_max_workers": 4,
        "enabled": False,  # 禁用 AI，仅走规则路径
    }

    # 调用 extract_invoice_items - 应走串行路径（for page in pages）
    items = extract_invoice_items(
        pdf_paths=[Path("invoice1.pdf"), Path("invoice2.pdf")],
        ai_config=ai_config,
        supplier="Test",
        period_start="2026-05-01",
        period_end="2026-05-31",
        currency="USD",
    )

    # 验证串行路径仍能正确抽取所有页面的结果（每页 1 人，共 2 人）
    assert len(items) == 2, f"串行路径应抽取 2 条记录，实际 {len(items)} 条"
    assert any(item.employee_name_raw == "Alvarez, Rosa" for item in items)
    assert any(item.employee_name_raw == "Johnson, Maria" for item in items)


def test_parallel_image_render_workers_config():
    """测试并行图片渲染配置默认值"""
    from bonus_platform.config import AI_CONFIG

    # 验证配置存在
    assert "parallel_extraction_enabled" in AI_CONFIG
    assert "parallel_max_workers" in AI_CONFIG
    assert "parallel_image_render_workers" in AI_CONFIG

    # 验证默认值
    assert AI_CONFIG["parallel_extraction_enabled"] is True
    assert AI_CONFIG["parallel_max_workers"] == 1
    assert AI_CONFIG["parallel_image_render_workers"] == 1


def test_parallel_image_rendering(monkeypatch):
    """测试并行图片渲染 - 多个 PDF 文件并行渲染"""
    import sys
    from unittest.mock import MagicMock
    from PIL import Image

    from bonus_platform.engine.labor.extract import _render_pdf_pages_to_images

    # 创建 3 个模拟 PDF 路径
    pdf_paths = [Path(f"invoice_{i}.pdf") for i in range(3)]

    # 追踪并行调用
    render_calls = []

    # Mock pypdfium2 - 函数内部 import pypdfium2
    mock_pdfium = MagicMock()

    def fake_pdf_document(path_str):
        """模拟 PdfDocument，返回包含 1 页的 mock 文档"""
        render_calls.append(path_str)
        mock_doc = MagicMock()
        mock_page = MagicMock()
        # 使用真实的 PIL Image，以便 save() 能产生实际 bytes
        pil_img = Image.new("RGB", (10, 10), color="white")
        mock_bitmap = MagicMock()
        mock_bitmap.to_pil.return_value = pil_img
        mock_page.render.return_value = mock_bitmap
        mock_page.close = MagicMock()
        # 文档有 1 页
        mock_doc.__len__ = MagicMock(return_value=1)
        mock_doc.__getitem__ = MagicMock(return_value=mock_page)
        mock_doc.close = MagicMock()
        return mock_doc

    mock_pdfium.PdfDocument = fake_pdf_document
    monkeypatch.setitem(sys.modules, "pypdfium2", mock_pdfium)

    # 调用并行渲染函数（3个文件 > 1，触发并行路径）
    result = _render_pdf_pages_to_images(pdf_paths, scale=1.5, max_workers=4)

    # 验证所有 3 个 PDF 都被渲染了
    assert len(render_calls) == 3, f"应渲染 3 个 PDF，实际渲染 {len(render_calls)} 个"
    assert len(result) == 3, f"应返回 3 个页面，实际 {len(result)} 个"

    # 验证每个页面都有正确的元数据
    source_files = {p["source_file"] for p in result}
    assert source_files == {"invoice_0.pdf", "invoice_1.pdf", "invoice_2.pdf"}
    for page in result:
        assert page["mime_type"] == "image/jpeg", "应使用 JPEG 格式以减少传输大小"
        assert page["base64"], "base64 不应为空"
        assert page["page"] == 1


def test_improved_name_similarity():
    """测试改进的姓名相似度"""
    from bonus_platform.engine.labor.compare import _name_similarity_improved

    # 相同姓名
    assert _name_similarity_improved("Rosa Alvarez", "Rosa Alvarez") == 1.0

    # "Last, First" vs "First Last"
    score_comma = _name_similarity_improved("Alvarez, Rosa", "Rosa Alvarez")
    assert score_comma > 0.8, f"Last/First格式匹配得分过低: {score_comma}"

    # 中间名差异
    score_middle = _name_similarity_improved("Rosa J. Alvarez", "Rosa Alvarez")
    assert score_middle > 0.8, f"中间名差异匹配得分过低: {score_middle}"

    # 拼写错误
    score_typo = _name_similarity_improved("Rosa Alvarez", "Rosa Alvarex")
    assert score_typo > 0.65, f"拼写错误匹配得分过低: {score_typo}"

    # 昵称变体
    score_nick = _name_similarity_improved("Bob Smith", "Robert Smith")
    assert score_nick > 0.5, f"昵称变体匹配得分过低: {score_nick}"

    # 完全不同的名字
    score_diff = _name_similarity_improved("Rosa Alvarez", "John Smith")
    assert score_diff < 0.5, f"不同名字匹配得分过高: {score_diff}"
