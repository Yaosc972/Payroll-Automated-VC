from pathlib import Path

from openpyxl import Workbook, load_workbook

from bonus_platform.engine.models import CalculationResult
from bonus_platform.engine.workbook_io import (
    PENDING_CONFIRMATION_HEADERS,
    build_final_workbook,
    build_pending_workbook,
)


def _pending_result() -> CalculationResult:
    return CalculationResult(
        month=202510,
        details=[],
        recruitment_summary=[],
        referral_summary=[],
        pending_confirmations=[
            {
                "源行号": 18,
                "姓名": "待确认员工",
                "工号": "zt-pending-001",
                "奖金类型": "招聘奖金",
                "角色": "招聘负责人",
                "发放对象工号": "zt-recruiter",
                "发放对象姓名": "招聘负责人",
                "发放对象状态": "离职",
                "币种": "CNY",
                "发放节点": "入职1月奖金",
                "建议发放周期": 202510,
                "建议发放金额": 300,
                "人工确认结果": "",
                "人工确认金额": "",
                "不发/暂缓原因": "",
                "系统提示": "招聘负责人非在职状态，需确认是否发放",
            }
        ],
        exceptions=[],
    )


def test_pending_workbook_is_compact_and_has_confirmation_dropdown(tmp_path: Path):
    path = tmp_path / "pending.xlsx"

    build_pending_workbook(_pending_result(), path)

    workbook = load_workbook(path)
    sheet = workbook["待确认_发放判断"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    validations = list(sheet.data_validations.dataValidation)

    assert headers == PENDING_CONFIRMATION_HEADERS
    assert "确认人" not in headers
    assert "确认时间" not in headers
    assert "币种" in headers
    assert any(
        str(validation.sqref) in {"M2", f"M2:M{sheet.max_row}"}
        and validation.formula1 == '"确认发放,不发放,暂缓"'
        for validation in validations
    )


def test_final_workbook_merges_pending_confirmation_by_currency(tmp_path: Path):
    initial_path = tmp_path / "initial.xlsx"
    confirmation_path = tmp_path / "confirmation.xlsx"
    output_path = tmp_path / "final.xlsx"
    initial = Workbook()
    recruitment = initial.active
    recruitment.title = "招聘奖金汇总"
    recruitment.append(["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    recruitment.append(["zt-recruiter", "招聘负责人", "招聘负责人", "CNY", 202510, 10, 0, 0, 0, 10])
    recruitment.append(["zt-recruiter", "招聘负责人", "招聘负责人", "USD", 202510, 20, 0, 0, 0, 20])
    referral = initial.create_sheet("内推奖金汇总")
    referral.append(["推荐人工号", "推荐人姓名", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    initial.save(initial_path)

    confirmation = Workbook()
    pending = confirmation.active
    pending.title = "待确认_发放判断"
    pending.append(PENDING_CONFIRMATION_HEADERS)
    pending.append(
        [
            18,
            "待确认员工",
            "zt-pending-001",
            "招聘奖金",
            "招聘负责人",
            "zt-recruiter",
            "招聘负责人",
            "离职",
            "USD",
            "入职1月奖金",
            202510,
            300,
            "确认发放",
            "",
            "",
            "招聘负责人非在职状态，需确认是否发放",
        ]
    )
    confirmation.save(confirmation_path)

    build_final_workbook(initial_path, confirmation_path, output_path)

    workbook = load_workbook(output_path, data_only=True)
    rows = list(workbook["最终招聘奖金汇总"].iter_rows(min_row=2, values_only=True))
    by_currency = {row[3]: row for row in rows}

    assert by_currency["CNY"][5] == 10
    assert by_currency["CNY"][9] == 10
    assert by_currency["USD"][5] == 320
    assert by_currency["USD"][9] == 320
    assert "招聘奖金汇总" not in workbook.sheetnames
    assert "内推奖金汇总" not in workbook.sheetnames
    assert "待确认_发放判断" not in workbook.sheetnames
    assert workbook["确认留痕"].sheet_state == "hidden"
    total_row = list(workbook["最终招聘奖金汇总"].iter_rows(values_only=True))[-1]
    assert total_row[1] == "合计"
    assert total_row[5] == 330
    assert total_row[9] == 330


def test_initial_summary_workbook_has_total_rows(tmp_path: Path):
    from bonus_platform.engine.workbook_io import build_result_workbook

    path = tmp_path / "initial.xlsx"
    result = CalculationResult(
        month=202510,
        details=[],
        recruitment_summary=[
            {"工号": "zt-a", "姓名": "招聘A", "角色": "招聘负责人", "币种": "CNY", "核算月份": 202510, "入职1月奖金": 10, "入职3月奖金": 20, "入职6月奖金": 0, "转正奖金": 5, "合计发放": 35},
            {"工号": "zt-b", "姓名": "招聘B", "角色": "招聘负责人", "币种": "CNY", "核算月份": 202510, "入职1月奖金": 2, "入职3月奖金": 3, "入职6月奖金": 4, "转正奖金": 1, "合计发放": 10},
        ],
        referral_summary=[
            {"推荐人工号": "zt-r", "推荐人姓名": "内推R", "币种": "CNY", "核算月份": 202510, "入职1月奖金": 75, "入职3月奖金": 25, "入职6月奖金": 0, "转正奖金": 0, "合计发放": 100}
        ],
        pending_confirmations=[],
        exceptions=[],
    )

    build_result_workbook(result, path)

    workbook = load_workbook(path, data_only=True)
    recruitment_total = list(workbook["招聘奖金汇总"].iter_rows(values_only=True))[-1]
    referral_total = list(workbook["内推奖金汇总"].iter_rows(values_only=True))[-1]

    assert recruitment_total[1] == "合计"
    assert recruitment_total[5] == 12
    assert recruitment_total[6] == 23
    assert recruitment_total[7] == 4
    assert recruitment_total[8] == 6
    assert recruitment_total[9] == 45
    assert referral_total[1] == "合计"
    assert referral_total[8] == 100


def test_final_workbook_updates_detail_rows_from_confirmation(tmp_path: Path):
    initial_path = tmp_path / "initial.xlsx"
    confirmation_path = tmp_path / "confirmation.xlsx"
    output_path = tmp_path / "final.xlsx"
    initial = Workbook()
    detail = initial.active
    detail.title = "招聘奖金明细"
    headers = [
        "唯一验证",
        "姓名",
        "工号",
        "招聘人入职1月奖金",
        "招聘人入职1月周期",
        "内推入职1月奖金",
        "内推入职1月周期",
    ]
    detail.append(headers)
    detail.append(["唯一", "待确认员工", "zt-pending-001", 300, 202510, 150, 202510])
    recruitment = initial.create_sheet("招聘奖金汇总")
    recruitment.append(["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    referral = initial.create_sheet("内推奖金汇总")
    referral.append(["推荐人工号", "推荐人姓名", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    initial.create_sheet("待确认_发放判断")
    initial.save(initial_path)

    confirmation = Workbook()
    pending = confirmation.active
    pending.title = "待确认_发放判断"
    pending.append(PENDING_CONFIRMATION_HEADERS)
    pending.append(
        [
            2,
            "待确认员工",
            "zt-pending-001",
            "招聘奖金",
            "招聘负责人",
            "zt-recruiter",
            "招聘负责人",
            "离职",
            "CNY",
            "入职1月奖金",
            202510,
            300,
            "不发放",
            "",
            "不符合发放",
            "招聘负责人非在职状态，需确认是否发放",
        ]
    )
    pending.append(
        [
            2,
            "待确认员工",
            "zt-pending-001",
            "内推奖金",
            "推荐人",
            "zt-referrer",
            "推荐人",
            "正式",
            "CNY",
            "入职1月奖金",
            202510,
            150,
            "确认发放",
            120,
            "",
            "推荐人状态缺失，需确认是否发放",
        ]
    )
    confirmation.save(confirmation_path)

    build_final_workbook(initial_path, confirmation_path, output_path)

    workbook = load_workbook(output_path, data_only=True)
    final_detail = workbook["招聘奖金明细"]
    final_headers = [final_detail.cell(1, column).value for column in range(1, final_detail.max_column + 1)]
    row = {header: final_detail.cell(2, column).value for column, header in enumerate(final_headers, start=1)}

    assert row["招聘人入职1月奖金"] == 0
    assert row["招聘人入职1月周期"] == "-"
    assert row["内推入职1月奖金"] == 120
    assert row["内推入职1月周期"] == 202510
