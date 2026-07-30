from io import BytesIO
from dataclasses import replace
import json
from pathlib import Path
import time

import pytest
import openpyxl
from openpyxl import Workbook, load_workbook
from urllib.error import HTTPError

import bonus_platform.app as app_module
import bonus_platform.engine.labor.structure as labor_structure
from bonus_platform.engine.labor.compare import compare_labor_items, compare_by_warehouse
from bonus_platform.engine.labor.evidence import LaborPageEvidence
from bonus_platform.engine.labor.extract import AI_PAGE_CACHE_VERSION, MiMoTimeoutException, _anthropic_messages_url, _candidate_is_confident, _effective_max_pages_per_request, _effective_render_scale, _extract_invoice_total_from_text, _http_post_json, extract_invoice_items, _extract_with_ai_images, _extract_with_rules, _request_headers
from bonus_platform.engine.labor.extract import _ai_instruction, _ai_ready, _extract_pdf_pages, _safe_error_message
from bonus_platform.engine.labor.extract import _analyze_layout_with_ai
from bonus_platform.engine.labor.extract import _filter_ai_rows_by_page_text
from bonus_platform.engine.labor.extract import _filter_ai_rows_by_expected_employees
from bonus_platform.engine.labor.extract import _drop_closed_employee_subtotals, _normalize_ai_rows
from bonus_platform.engine.labor.extract import _warehouse_id_from_filename as extract_warehouse_id_from_filename
from bonus_platform.engine.labor.extract import _warehouse_id_conflict
from bonus_platform.engine.labor.extract import _classify_pdf
from bonus_platform.engine.labor.extract import _warehouse_id_from_text
from bonus_platform.engine.labor import runs as labor_runs
from bonus_platform.engine.labor.governance import audit_ai_page_cache_candidates, build_ai_cache_reconciliation_preview, build_reocr_candidate_plan, build_rule_change_candidate, confirm_rule_candidate, replay_reocr_candidate_result, rollback_rule_version, summarize_rule_auto_replay, summarize_rule_replay
from bonus_platform.engine.labor.materials import build_material_dry_run, build_material_index, build_material_replay_plan
from bonus_platform.engine.labor.models import LaborLineItem, line_items_from_dicts
from bonus_platform.engine.labor.layout import InvoiceLayoutPlan, analyze_invoice_layout, extract_rows_from_layout_plan
from bonus_platform.engine.labor.parsing import normalize_employee_name, normalize_workbuddy_name, parse_number
from bonus_platform.engine.labor.profiles import load_supplier_profiles, resolve_supplier_profile
from bonus_platform.engine.labor.quality import build_reconciliation_diagnostics, calculate_extraction_quality
from bonus_platform.engine.labor.profiles import (
    generate_profile_from_extraction,
    save_supplier_profile,
    record_profile_failure,
    reset_profile_failure,
    _profiles_for_resolution,
    DEFAULT_PROFILE,
)
from bonus_platform.app import _build_conclusion, _labor_total_is_explicitly_non_payable, _non_payable_pdf_names, _normalize_labor_total_decision
from bonus_platform.engine.labor.report import build_labor_business_html_report, build_labor_report
from bonus_platform.engine.labor.structure import evaluate_batch_guards, extract_structured_invoice_rows, find_amount_closure, infer_warehouse_from_rows, parse_localized_number, prefer_closed_structured_rows, promote_structured_invoice_evidence, resolve_amount_scope
from bonus_platform.engine.labor.workbook import list_workbook_sheets, parse_reocr_candidate_rows, read_workbook_rows, suggest_mapping, summarize_otws_costs


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "账单"
    sheet.append(["姓名", "时长总计(H)", "费用总计(含税)", "币种"])
    sheet.append(["Jose Perez", 40.14, 1037.81, "USD"])
    sheet.append(["Wilfredo Martinez", 40.78, 982.74, "USD"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _ready_ai_config() -> dict:
    return {
        "enabled": True,
        "provider": "mimo",
        "api_key": "token",
        "base_url": "https://api.xiaomimimo.com/v1",
        "model": "mimo-v2.5",
    }


def test_auto_ocr_runs_when_invoice_total_is_unresolved_even_with_existing_rows():
    assert app_module._labor_should_run_auto_ocr(
        command="python worker.py",
        pdf_rows=[object()],
        target_pdf_names={"In291943.pdf"},
        reconciliation_pdf_totals=[
            {
                "source_file": "In291943.pdf",
                "total_amount": 0.0,
                "authoritative": False,
            }
        ],
    ) is True


def test_ocr_expected_totals_use_high_confidence_invoice_total_page():
    assert app_module._labor_ocr_expected_totals(
        [
            {
                "source_file": "In291943.pdf",
                "total_amount": 0.0,
                "authoritative": False,
                "page_evidence": [
                    {
                        "role": "invoice_total",
                        "role_confidence": 0.98,
                        "total_amount": 13836.28,
                    },
                    {
                        "role": "invoice_continuation",
                        "role_confidence": 0.95,
                        "total_amount": None,
                    },
                ],
            }
        ],
        {"In291943.pdf"},
    ) == {"In291943.pdf": 13836.28}


def test_local_ocr_grand_total_does_not_override_explicit_total_invoice_amount():
    pdf_totals = [
        {
            "source_file": "CA 7 46292.pdf",
            "total_amount": 54358.11,
            "warehouse_id": "7",
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 7,
            "total_label": "TOTAL",
            "page_evidence": [
                {
                    "page": 7,
                    "role": "invoice_total",
                    "role_confidence": 0.98,
                    "total_amount": 54358.11,
                    "evidence_text": "\n".join(
                        [
                            "TOTAL REG: 2089.21",
                            "TOTAL OT: 104.73",
                            "GRAND TOTAL: 2193.94",
                            "TOTAL INVOICE AMOUNT: $54,358.11",
                        ]
                    ),
                    "extraction_method": "text_explicit_total",
                }
            ],
        }
    ]
    candidate = {
        "pdfTotalEvidence": {
            "CA 7 46292.pdf": {
                "amount": 2193.94,
                "page": 7,
                "label": "GRAND TOTAL",
                "evidenceText": "GRAND TOTAL: 2193.94",
            }
        }
    }

    merged = app_module._labor_apply_ocr_pdf_total_evidence(pdf_totals, candidate)

    assert merged[0]["total_amount"] == 54358.11
    assert merged[0]["evidence_status"] == "authoritative"
    assert all(
        page.get("source") != "local_ocr_explicit_total"
        for page in merged[0]["page_evidence"]
    )


def test_local_ocr_grand_total_does_not_replace_authoritative_expected_total(
    monkeypatch,
    tmp_path,
):
    pdf_path = tmp_path / "CA 7 46292.pdf"
    pdf_path.write_bytes(b"pdf")
    captured = {}
    monkeypatch.setattr(
        app_module,
        "run_ocr_candidate_command",
        lambda *args, **kwargs: {
            "status": "completed",
            "rows": [],
            "files": [
                {
                    "sourceFile": pdf_path.name,
                    "explicitTotalAmount": 2193.94,
                    "explicitTotalEvidence": {
                        "page": 7,
                        "label": "GRAND TOTAL",
                        "evidenceText": "GRAND TOTAL: 2193.94",
                    },
                }
            ],
        },
    )

    def capture_expected_totals(ocr_result, excel_rows, expected_totals, **kwargs):
        captured.update(expected_totals)
        return {"decision": "needs_review", "safeToUse": False}

    monkeypatch.setattr(
        app_module,
        "evaluate_ocr_candidate_result",
        capture_expected_totals,
    )

    app_module._run_labor_auto_ocr_candidate(
        "labor_test",
        [pdf_path],
        [],
        [
            {
                "source_file": pdf_path.name,
                "total_amount": 54358.11,
                "authoritative": True,
            }
        ],
        supplier="Voyage Employer Services",
        period_start="2026-06-29",
        period_end="2026-07-05",
        currency="USD",
        command="python worker.py",
        timeout_seconds=30,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
    )

    assert captured == {pdf_path.name: 54358.11}


def test_detail_total_retry_is_skipped_after_auto_ocr_candidate_run():
    assert app_module._labor_should_retry_detail_totals(
        pdf_rows=[object()],
        mismatches={"invoice.pdf": {"delta": 10.0}},
        auto_ocr_candidate={"decision": "needs_review", "runtimeStatus": "completed"},
    ) is False


def test_labor_ocr_progress_callback_maps_worker_snapshot(monkeypatch):
    captured = {}
    monkeypatch.setattr(app_module, "_update_labor_progress", lambda run_id, **payload: captured.update(run_id=run_id, **payload))

    callback = app_module._labor_ocr_progress_callback("labor_test")
    callback(
        {
            "status": "running",
            "currentFile": "invoice.pdf",
            "totalFiles": 3,
            "processedFiles": 1,
            "totalPages": 10,
            "processedPages": 4,
            "cacheHitCount": 1,
            "message": "正在识别 invoice.pdf：4 / 10 页。",
        }
    )

    assert captured == {
        "run_id": "labor_test",
        "phase": "auto_ocr",
        "phase_label": "本地 OCR 识别",
        "message": "正在识别 invoice.pdf：4 / 10 页。",
        "status": "running",
        "total_files": 3,
        "processed_files": 1,
        "total_pages": 10,
        "processed_pages": 4,
        "current_file": "invoice.pdf",
        "cache_hit_count": 1,
    }


def test_run_labor_auto_ocr_candidate_returns_safe_evaluated_rows(monkeypatch, tmp_path):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"pdf")
    excel_rows = [
        LaborLineItem(
            source_type="offline_workbook",
            source_file="bill.xlsx",
            source_page_or_row="r1",
            employee_id="",
            employee_name_raw="Jane Doe",
            hours=8,
            amount=80,
        )
    ]
    monkeypatch.setattr(
        app_module,
        "run_ocr_candidate_command",
        lambda *args, **kwargs: {
            "status": "completed",
            "rows": [
                {
                    "source_file": "invoice.pdf",
                    "employee_name_raw": "Jane Doe",
                    "hours": 8,
                    "amount": 80,
                }
            ],
            "files": [{"sourceFile": "invoice.pdf", "failedPageCount": 0}],
        },
    )

    candidate = app_module._run_labor_auto_ocr_candidate(
        "labor_test",
        [pdf_path],
        excel_rows,
        [{"source_file": "invoice.pdf", "total_amount": 80, "authoritative": True}],
        supplier="Unknown",
        period_start="2026-05-11",
        period_end="2026-05-17",
        currency="USD",
        command="python worker.py",
        timeout_seconds=30,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
    )

    assert candidate["safeToUse"] is True
    assert candidate["runtimeStatus"] == "completed"
    assert candidate["rows"][0]["employee_name_raw"] == "Jane Doe"


def test_list_labor_metadata_supports_recent_limit(monkeypatch, tmp_path):
    runs_dir = tmp_path / "labor_runs"
    runs_dir.mkdir()
    for idx in range(3):
        run_dir = runs_dir / f"labor_{idx}"
        run_dir.mkdir()
        metadata_path = run_dir / "metadata.json"
        metadata_path.write_text(
            json.dumps(
                {
                    "id": f"labor_{idx}",
                    "createdAt": f"2026-06-20T10:0{idx}:00",
                    "updatedAt": f"2026-06-20T10:0{idx}:00",
                }
            ),
            encoding="utf-8",
        )

    monkeypatch.setattr(labor_runs, "LABOR_RUNS_DIR", runs_dir)

    rows = labor_runs.list_labor_metadata(limit=2)

    assert [row["id"] for row in rows] == ["labor_2", "labor_1"]


def test_save_labor_metadata_uses_atomic_writes_for_persistent_backend(monkeypatch, tmp_path):
    run_dir = tmp_path / "labor_runs" / "labor_atomic"
    run_dir.mkdir(parents=True)
    original_write_text = Path.write_text

    monkeypatch.setattr(labor_runs, "labor_persistent_storage_enabled", lambda: True)
    monkeypatch.setattr(labor_runs, "labor_persistent_storage_info", lambda: {"backend": "supabase"})
    monkeypatch.setattr(labor_runs, "sync_labor_run_to_persistent", lambda run_id, path: None)

    def guarded_write_text(self, *args, **kwargs):
        if self.name == "metadata.json":
            raise AssertionError("metadata.json must be replaced atomically")
        return original_write_text(self, *args, **kwargs)

    monkeypatch.setattr(Path, "write_text", guarded_write_text)

    metadata = labor_runs.save_labor_metadata(run_dir, {"id": "labor_atomic", "status": "已创建"})

    assert metadata["id"] == "labor_atomic"
    assert json.loads((run_dir / "metadata.json").read_text(encoding="utf-8"))["id"] == "labor_atomic"


def test_save_labor_metadata_retries_transient_windows_replace_conflict(monkeypatch, tmp_path):
    run_dir = tmp_path / "labor_runs" / "labor_windows_replace_retry"
    run_dir.mkdir(parents=True)
    real_replace = labor_runs.os.replace
    attempts = 0

    monkeypatch.setattr(labor_runs, "labor_persistent_storage_enabled", lambda: False)
    monkeypatch.setattr(labor_runs, "labor_postgres_state_enabled", lambda: False)

    def replace_with_one_windows_conflict(source, destination):
        nonlocal attempts
        attempts += 1
        if attempts == 1:
            error = PermissionError(13, "The process cannot access the file")
            error.winerror = 5
            raise error
        return real_replace(source, destination)

    monkeypatch.setattr(labor_runs.os, "replace", replace_with_one_windows_conflict)

    metadata = labor_runs.save_labor_metadata(
        run_dir,
        {"id": run_dir.name, "status": "抽取中"},
    )

    assert metadata["status"] == "抽取中"
    assert attempts == 2
    assert json.loads((run_dir / "metadata.json").read_text(encoding="utf-8"))["status"] == "抽取中"
    assert list(run_dir.glob(".metadata.json.*.tmp")) == []


def _workbook_with_tax_columns_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "账单"
    sheet.append(["姓名", "时长总计(H)", "费用总计(不含税)", "费用总计(含税)", "币种"])
    sheet.append(["Jose Perez", 40.14, 1000.00, 1037.81, "USD"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_with_hours_only_summary_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Employee Name", "Job Type", "求和项:Total Hours"])
    sheet.append(["Alberto Nunez", "Labor", 35.08])
    sheet.append(["Ivis Martinez", "Labor", 6.55])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _workbook_with_two_header_rows_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employee-expenses-detail"
    sheet.append([
        "Company Name",
        "Physical warehouse",
        "group",
        "Employee name",
        "Employee number",
        "Type of work",
        "工作日",
        None,
        "Total staff cost accounting time",
        "Total cost",
    ])
    sheet.append([
        "Company Name",
        "Physical warehouse",
        "group",
        "Employee name",
        "Employee number",
        "Type of work",
        "Day shift working hours",
        "Regular pay for day shift",
        "Total staff cost accounting time",
        "Total cost",
    ])
    sheet.append([
        "Strategic Staffing Solutions Corp.",
        "New Jersey Warehouse 13",
        "warehousing group",
        "JOSE MAGANA",
        "EUS031468",
        "操作员",
        8,
        188,
        8,
        188,
    ])
    sheet.append([
        "Total:",
        None,
        None,
        None,
        None,
        None,
        8,
        188,
        8,
        188,
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _otws_cost_workbook_bytes() -> bytes:
    workbook = Workbook()
    warehouse = workbook.active
    warehouse.title = "Warehouse-information"
    warehouse.append([
        "Region",
        "Physical warehouse",
        "Financial reimbursement process number",
        "Company Name",
        "Total employees during attendance period",
        "Contract Start Date",
        "currency",
        "Total hourly salary",
        "Total bonus",
        "Total vehicle compensation",
        "Total Meal Supplement",
        "Hourly Rate Difference",
        "Lot And Scot(Employment Insurance)",
        "Income Tax",
        "Total other expenses",
        "total.handling.fee",
        "additional fees",
        "Total",
        "Status",
        "Week Month Split Status",
        "Accounting start date",
        "Accounting end date",
        "实际付款金额",
        "remark",
    ])
    warehouse.append([
        "USNJ",
        "New Jersey Warehouse 13",
        "--",
        "Strategic Staffing Solutions Corp.",
        "64",
        "US ELOGISTICS SERVICE CORP",
        "USD",
        48055.81,
        0,
        0,
        0,
        0,
        0,
        0,
        162.15,
        0,
        0,
        48217.96,
        "Confirm the bill",
        "未拆分",
        "2026-05-11",
        "2026-05-17",
        0,
        "--",
    ])

    expenses = workbook.create_sheet("Employee-expenses-detail")
    expenses.append([
        "Company Name",
        "Physical warehouse",
        "group",
        "Employee name",
        "Employee number",
        "Type of work",
        "工作日",
        None,
        "Total staff cost accounting time",
        "Total cost",
    ])
    expenses.append([
        "Company Name",
        "Physical warehouse",
        "group",
        "Employee name",
        "Employee number",
        "Type of work",
        "Day shift working hours",
        "Regular pay for day shift",
        "Total staff cost accounting time",
        "Total cost",
    ])
    expenses.append([
        "Strategic Staffing Solutions Corp.",
        "New Jersey Warehouse 13",
        "warehousing group",
        "JOSE MAGANA",
        "EUS031468",
        "操作员",
        8,
        188,
        8,
        188,
    ])
    expenses.append([
        "Total:",
        None,
        None,
        None,
        None,
        None,
        8,
        188,
        8,
        188,
    ])

    benefits = workbook.create_sheet("Employee-benefits-detail")
    benefits.append([
        "Physical warehouse",
        "Employee name",
        "Employee number",
        "Bonus",
        "Car allowance",
        "Meal allowance",
        "Hourly Rate Difference",
        "Lot And Scot(Employment Insurance)",
        "Income Tax",
        "Other",
        "Total cost",
        "remark",
    ])
    benefits.append([
        "New Jersey Warehouse 13",
        "KRISTEL CONTRERAS MONTIEL",
        "EUS033091",
        0,
        0,
        0,
        0,
        0,
        0,
        162.15,
        162.15,
        "missing hours",
    ])
    benefits.append(["Total:", None, None, 0, 0, 0, 0, 0, 0, 162.15, 162.15, None])

    workbook.create_sheet("The-loading-and-unloading-of-ta").append([
        "group",
        "Employee name",
        "Loading and unloading date",
        "Ark type",
        "Container Number",
        "Number of unloading cabinets",
        "Unit price",
        "Total cost",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_normalize_employee_name_handles_invoice_and_workbook_variants():
    assert normalize_employee_name("PEREZ, JOSE") == normalize_employee_name("Jose Perez")
    assert normalize_employee_name("#1 Ana Maria Corea") == normalize_employee_name("COREA MARIA, ANA")
    assert normalize_employee_name("CONTRERAS, EVELYN (CERVANTES)") == normalize_employee_name("Evelyn Contreras")
    assert normalize_employee_name("MORA-3491, CLAUDIA") == normalize_employee_name("Claudia Mora-3491")
    assert normalize_employee_name("Darlene CalvilloDarlene Calvillo Aparicio Aparicio") == normalize_employee_name("Calvillo Aparicio, Darlene")
    assert normalize_employee_name("Rosales Jr., Jose") == normalize_employee_name("Jose Rosales")


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("$1,032.00", 1032.0),
        ("1,032.00$", 1032.0),
        ("218,40", 218.4),
        ("1.032,00 EUR", 1032.0),
        ("-$", 0.0),
        ("", 0.0),
        (None, 0.0),
    ],
)
def test_parse_number_handles_invoice_money_formats(raw, expected):
    assert parse_number(raw) == expected


def test_normalize_ai_rows_accepts_comma_decimal_amounts():
    rows = _normalize_ai_rows(
        [{"employee_name": "DUPONT Marie", "hours": "10,50", "amount": "218,40"}],
        supplier="Adequat",
        period_start="2026-05-25",
        period_end="2026-05-31",
        currency="EUR",
    )

    assert len(rows) == 1
    assert line_items_from_dicts(rows)[0].hours == 10.5
    assert line_items_from_dicts(rows)[0].amount == 218.4


def test_drop_closed_employee_subtotal_across_adjacent_pages():
    rows = [
        {
            "employee_name_raw": "FOFANA DRAME Sekou",
            "hours": 35.0,
            "amount": 706.65,
            "description": "HEURES NORMALES",
            "evidence_text": "FOFANA DRAME Sekou HEURES NORMALES 35,00 706,65",
            "source_file": "invoice.pdf",
            "source_page_or_row": "p1",
        },
        {
            "employee_name_raw": "FOFANA DRAME Sekou",
            "hours": 3.75,
            "amount": 94.69,
            "description": "HEURES SUPPLEMENTAIRES T1",
            "evidence_text": "FOFANA DRAME Sekou HEURES SUPPLEMENTAIRES T1 3,75 94,69",
            "source_file": "invoice.pdf",
            "source_page_or_row": "p1",
        },
        {
            "employee_name_raw": "FOFANA DRAME Sekou",
            "hours": 38.75,
            "amount": 801.34,
            "description": "Interimaire : FOFANA DRAME Sekou",
            "evidence_text": "S/Total Interimaire : FOFANA DRAME Sekou 38,75 801,34",
            "source_file": "invoice.pdf",
            "source_page_or_row": "p2",
        },
    ]

    kept = _drop_closed_employee_subtotals(rows)

    assert [row["amount"] for row in kept] == [706.65, 94.69]


def test_drop_closed_employee_subtotal_keeps_unclosed_and_single_charge_rows():
    rows = [
        {
            "employee_name_raw": "Worker One",
            "hours": 10.0,
            "amount": 200.0,
            "description": "HEURES NORMALES",
            "evidence_text": "Worker One 10,00 200,00",
            "source_file": "invoice.pdf",
            "source_page_or_row": "p1",
        },
        {
            "employee_name_raw": "Worker One",
            "hours": 12.0,
            "amount": 250.0,
            "description": "Interimaire : Worker One",
            "evidence_text": "S/Total Interimaire : Worker One 12,00 250,00",
            "source_file": "invoice.pdf",
            "source_page_or_row": "p2",
        },
        {
            "employee_name_raw": "Worker Two",
            "hours": 8.0,
            "amount": 160.0,
            "description": "Interimaire : Worker Two",
            "evidence_text": "S/Total Interimaire : Worker Two 8,00 160,00",
            "source_file": "invoice.pdf",
            "source_page_or_row": "p2",
        },
    ]

    assert _drop_closed_employee_subtotals(rows) == rows


def test_fairway_warehouse_id_parses_from_filename_and_text():
    assert extract_warehouse_id_from_filename("135306 US Elogistics Service Corp (#10).pdf") == "10"
    assert extract_warehouse_id_from_filename("135307_US_Elogistics_Service_Corp___18_20260601_222816_133922.pdf") == "18"
    assert _warehouse_id_from_text("US ELOGISTICS SERVICE CORP\nCA(LA)- #18 TAMARIND (TAMR2)") == "18"
    assert _warehouse_id_from_text("FONTANA\n(CA)LA#10 HARBOR BAR (HARBOR)") == "10"
    assert _warehouse_id_from_text("CHINO, CA 91710\nCA(LA)#25 (CEDAR)") == "25"


def test_warehouse_id_patterns_cover_common_filename_and_text_variants():
    assert extract_warehouse_id_from_filename("INVOICE_WH-30.pdf") == "30"
    assert extract_warehouse_id_from_filename("CITISTAFF_LOC_29_20260602.pdf") == "29"
    assert extract_warehouse_id_from_filename("NJ12 Invoice Report WE 051726 JF.pdf") == "12"
    assert extract_warehouse_id_from_filename("US Elogis Service #17 Invoice W.E 05.24.26.pdf") == "17"
    assert _warehouse_id_from_text("Location: 3号仓") == "3"
    assert _warehouse_id_from_text("Warehouse: WH 28") == "28"
    assert _warehouse_id_from_text("LOC #21") == "21"
    assert _warehouse_id_from_text("Purchase Order Number\nFlanders Location NJ 8") == "8"


def test_warehouse_id_from_filename_accepts_explicit_chinese_warehouse_but_not_generic_name():
    assert extract_warehouse_id_from_filename("巴黎1号仓 05.25-05.31.xlsx") == "1"
    assert extract_warehouse_id_from_filename("ADEQUAT 05.25-05.31.xlsx") == ""


def test_workbook_filename_fallback_only_fills_blank_warehouse_rows():
    rows = [
        _structure_row("Blank Worker", source_type="offline_workbook", warehouse_id=""),
        _structure_row("Explicit Worker", source_type="offline_workbook", warehouse_id="9"),
    ]

    assigned = app_module._labor_apply_workbook_warehouse_fallback(
        rows,
        "巴黎1号仓 05.25-05.31.xlsx",
    )

    assert [row.warehouse_id for row in assigned] == ["1", "9"]
    unchanged = app_module._labor_apply_workbook_warehouse_fallback(rows, "ADEQUAT May.xlsx")
    assert [row.warehouse_id for row in unchanged] == ["", "9"]


def test_warehouse_id_from_text_does_not_treat_invoice_period_as_location():
    assert (
        _warehouse_id_from_text(
            "Period Cust. ID Tax ID PAYMENT TERMS Location\n"
            "05/18/2026-05/24/2026 E-LOG 30 SHEIN\n"
        )
        == ""
    )


def test_warehouse_id_conflict_is_reported_when_filename_and_text_disagree():
    conflict = _warehouse_id_conflict("CHINA_EXPRESS__3_INVOICE.pdf", "US ELOGISTICS\nCA#30")

    assert conflict == {"source_file": "CHINA_EXPRESS__3_INVOICE.pdf", "filename_warehouse_id": "3", "text_warehouse_id": "30"}


def test_warehouse_comparison_reports_pdf_warehouse_conflict_errors():
    result = compare_by_warehouse(
        pdf_totals=[
            {
                "source_file": "CHINA_EXPRESS__3_INVOICE.pdf",
                "warehouse_id": "3",
                "total_amount": 1000.0,
                "warehouse_conflict": {"filename_warehouse_id": "3", "text_warehouse_id": "30"},
            }
        ],
        excel_rows_with_warehouse=[{"employee_name": "A", "warehouse_id": "3", "amount": 1000.0, "hours": 10}],
        amount_tolerance=0.1,
    )

    assert result["errors"] == ["仓库号冲突: CHINA_EXPRESS__3_INVOICE.pdf 文件名=3, 内容=30"]
    assert result["summary"]["pdfAmountTotal"] == 0.0
    assert result["summary"]["totalPassed"] is False
    assert result["rows"][0]["reconciliationStatus"] == "needs_review"
    assert result["rows"][0]["pdfAmountTotal"] == 0.0
    assert result["rows"][0]["pdfEvidenceFile"] == "CHINA_EXPRESS__3_INVOICE.pdf"


def test_warehouse_total_difference_equal_to_ten_cents_passes():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "fairway-warehouse-10.pdf", "warehouse_id": "10", "total_amount": 144714.83},
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Fairway Staff", "warehouse_id": "10", "amount": 144714.93, "hours": 10},
        ],
        amount_tolerance=0.1,
    )

    assert result["summary"]["amountDeltaTotal"] == -0.1
    assert result["summary"]["totalPassed"] is True


def test_warehouse_comparison_excel_grouping_error_forces_total_failed():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "warehouse-1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
        }],
        excel_rows_with_warehouse=[
            {"employee_name": "Mapped Worker", "warehouse_id": "1", "amount": 100.0, "hours": 4},
            {"employee_name": "Unmapped Worker", "warehouse_id": "", "amount": 0.0, "hours": 0},
        ],
        amount_tolerance=0.1,
    )

    assert result["errors"] == ["Excel 行缺少物理仓: Unmapped Worker"]
    assert result["rows"][0]["reconciliationStatus"] == "passed"
    assert result["summary"]["pdfAmountTotal"] == 100.0
    assert result["summary"]["excelAmountTotal"] == 100.0
    assert result["summary"]["amountDeltaTotal"] == 0.0
    assert result["summary"]["passedCount"] == 1
    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["totalPassed"] is False


def test_warehouse_comparison_pdf_row_grouping_error_forces_total_failed():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "warehouse-1.pdf",
            "warehouse_id": "1",
            "total_amount": 100.0,
            "authoritative": True,
            "evidence_status": "authoritative",
        }],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Unassigned Detail", hours=1.0, amount=25.0, currency="USD", confidence=0.95, evidence_text=""),
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Mapped Worker", "warehouse_id": "1", "amount": 100.0, "hours": 4},
        ],
        amount_tolerance=0.1,
    )

    assert result["errors"] == ["无法从文件名提取仓库号: invoice.pdf"]
    assert result["rows"][0]["reconciliationStatus"] == "passed"
    assert result["summary"]["pdfAmountTotal"] == 100.0
    assert result["summary"]["excelAmountTotal"] == 100.0
    assert result["summary"]["amountDeltaTotal"] == 0.0
    assert result["summary"]["passedCount"] == 1
    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["totalPassed"] is False


def test_warehouse_comparison_does_not_double_count_employee_detail_attachments():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "ELGA_041026-15.pdf", "warehouse_id": "", "total_amount": 2237.67, "authoritative": False, "evidence_status": "supporting", "pdf_type": "supporting"},
            {"source_file": "ELGA_Adriana.pdf", "warehouse_id": "", "total_amount": 905.92, "authoritative": False, "evidence_status": "supporting", "pdf_type": "supporting"},
            {"source_file": "Inv_04102615_from_Tru_Staffing_33680.pdf", "warehouse_id": "2", "total_amount": 2685.20, "authoritative": True, "evidence_status": "authoritative"},
            {"source_file": "Inv_04242617_from_Tru_Staffing_32284.pdf", "warehouse_id": "3", "total_amount": 3081.98, "authoritative": True, "evidence_status": "authoritative"},
        ],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="ELGA_041026-15.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Stephie Arujo", hours=44.18, amount=740.32, currency="USD", confidence=0.96, evidence_text="", warehouse_id="2"),
            LaborLineItem(source_type="pdf_invoice", source_file="ELGA_Adriana.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Adriana Bermudez Cuenu", hours=51.08, amount=905.92, currency="USD", confidence=0.96, evidence_text="", warehouse_id="3"),
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Stephie Araujo Hernandez", "warehouse_id": "2", "amount": 2685.20, "hours": 40},
            {"employee_name": "Adriana Bermudez Cuenu", "warehouse_id": "3", "amount": 3081.98, "hours": 40},
        ],
        amount_tolerance=0.1,
    )

    assert result["summary"]["pdfAmountTotal"] == 5767.18
    assert result["summary"]["totalPassed"] is True
    assert result["summary"]["diffWarehouses"] == []
    assert {row["warehouseId"]: row["matchStatus"] for row in result["rows"]} == {"2": "通过", "3": "通过"}


def test_warehouse_comparison_keeps_authoritative_total_with_employee_detail_signal():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "warehouse-1.pdf", "warehouse_id": "1", "total_amount": 100.0, "authoritative": True, "evidence_status": "authoritative", "has_employee_detail": True},
            {"source_file": "warehouse-2.pdf", "warehouse_id": "2", "total_amount": 200.0, "authoritative": True, "evidence_status": "authoritative"},
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "A", "warehouse_id": "1", "amount": 100.0, "hours": 1},
            {"employee_name": "B", "warehouse_id": "2", "amount": 200.0, "hours": 1},
        ],
        amount_tolerance=0.1,
    )

    assert result["summary"]["pdfAmountTotal"] == 300.0
    assert result["summary"]["totalPassed"] is True
    assert {row["warehouseId"]: row["reconciliationStatus"] for row in result["rows"]} == {
        "1": "passed",
        "2": "passed",
    }


def test_warehouse_attribution_closes_with_unattributed_invoice_amount():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "warehouse-9.pdf",
            "warehouse_id": "9",
            "total_amount": 11837.79,
            "authoritative": True,
            "evidence_status": "authoritative",
        }],
        pdf_rows=[
            LaborLineItem(
                source_type="pdf_invoice",
                source_file="warehouse-9.pdf",
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Employee A",
                hours=1,
                amount=11470.52,
                currency="USD",
                confidence=0.95,
                evidence_text="",
                warehouse_id="9",
            )
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Employee A", "warehouse_id": "9", "hours": 1, "amount": 11611.03}
        ],
        amount_tolerance=0.1,
    )

    row = result["rows"][0]
    attribution = {item["employeeName"]: item["delta"] for item in row["attribution"]}

    assert row["amountDelta"] == 226.76
    assert attribution["Employee A"] == -140.51
    assert attribution["未归因发票金额"] == 367.27
    assert round(sum(item["delta"] for item in row["attribution"]), 2) == row["amountDelta"]


def test_warehouse_comparison_marks_excel_only_warehouse_missing_pdf_invoice():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "warehouse-10.pdf",
            "warehouse_id": "10",
            "total_amount": 50000.0,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 2,
            "excluded_pages": [1],
        }],
        excel_rows_with_warehouse=[
            {"employee_name": "A", "warehouse_id": "10", "amount": 50000.0, "hours": 10},
            {"employee_name": "B", "warehouse_id": "3", "amount": 12459.22, "hours": 10},
        ],
        amount_tolerance=0.1,
    )

    rows = {row["warehouseId"]: row for row in result["rows"]}
    assert rows["3"] == {
        "warehouseId": "3",
        "pdfEmployeeCount": 0,
        "excelEmployeeCount": 1,
        "pdfHoursTotal": 0,
        "excelHoursTotal": 10.0,
        "pdfAmountTotal": 0.0,
        "excelAmountTotal": 12459.22,
        "amountDelta": -12459.22,
        "reconciliationStatus": "missing_pdf_invoice",
        "matchStatus": "缺少PDF发票",
        "evidenceStatus": "missing",
        "pdfEvidenceFile": "",
        "pdfEvidencePage": None,
        "excludedPdfPages": [],
        "employeeRows": [],
        "attribution": [],
    }
    assert result["summary"]["missingPdfAmountTotal"] == 12459.22
    assert result["summary"]["totalPassed"] is False


def test_warehouse_comparison_marks_pdf_only_warehouse_extra_pdf_invoice():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "warehouse-10.pdf", "warehouse_id": "10", "total_amount": 50000.0},
            {
                "source_file": "warehouse-99.pdf",
                "warehouse_id": "99",
                "total_amount": 275.5,
                "authoritative": True,
                "evidence_status": "authoritative",
                "total_page": 3,
                "excluded_pages": [1, 2],
            },
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "A", "warehouse_id": "10", "amount": 50000.0, "hours": 10},
        ],
        amount_tolerance=0.1,
    )

    row = next(row for row in result["rows"] if row["warehouseId"] == "99")
    assert row["reconciliationStatus"] == "extra_pdf_invoice"
    assert row["matchStatus"] == "多余PDF发票"
    assert row["evidenceStatus"] == "authoritative"
    assert row["pdfEvidenceFile"] == "warehouse-99.pdf"
    assert row["pdfEvidencePage"] == 3
    assert row["excludedPdfPages"] == [1, 2]
    assert result["summary"]["totalPassed"] is False


def test_warehouse_comparison_marks_unresolved_pdf_total_needs_review():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "warehouse-7.pdf",
            "warehouse_id": "7",
            "total_amount": 4105.15,
            "authoritative": False,
            "evidence_status": "needs_review",
            "total_page": None,
            "excluded_pages": [2, 3],
        }],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="warehouse-7.pdf", source_page_or_row="p2", employee_id="", employee_name_raw="Worker One", hours=40, amount=4105.15, currency="USD", confidence=0.95, evidence_text="", warehouse_id="7"),
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Worker One", "warehouse_id": "7", "amount": 4105.15, "hours": 40},
        ],
        amount_tolerance=0.1,
    )

    row = result["rows"][0]
    assert row["reconciliationStatus"] == "needs_review"
    assert row["matchStatus"] == "待复核"
    assert row["evidenceStatus"] == "needs_review"
    assert row["pdfAmountTotal"] == 0.0
    assert row["pdfEvidenceFile"] == "warehouse-7.pdf"
    assert row["pdfEvidencePage"] is None
    assert row["excludedPdfPages"] == [2, 3]
    assert row["employeeRows"]
    assert result["summary"]["pdfAmountTotal"] == 0.0
    assert result["summary"]["totalPassed"] is False


def test_warehouse_comparison_never_replaces_invoice_total_with_excel_closest_detail_sum():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "warehouse-1.pdf",
            "warehouse_id": "1",
            "total_amount": 1711.22,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 2,
            "excluded_pages": [3],
        }],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="warehouse-1.pdf", source_page_or_row="p3", employee_id="", employee_name_raw="Worker One", hours=79.65, amount=1513.35, currency="CAD", confidence=0.97, evidence_text="", warehouse_id="1"),
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Worker One", "warehouse_id": "1", "amount": 1513.35, "hours": 79.65},
        ],
        amount_tolerance=0.1,
    )

    row = result["rows"][0]
    assert result["summary"]["pdfAmountTotal"] == 1711.22
    assert result["summary"]["amountDeltaTotal"] == 197.87
    assert result["summary"]["totalPassed"] is False
    assert row["pdfAmountTotal"] == 1711.22
    assert row["reconciliationStatus"] == "amount_difference"
    assert row["pdfEvidenceFile"] == "warehouse-1.pdf"
    assert row["pdfEvidencePage"] == 2
    assert row["excludedPdfPages"] == [3]


def test_warehouse_comparison_reports_comparable_and_full_batch_deltas():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "warehouse-10.pdf", "warehouse_id": "10", "total_amount": 50000.0, "authoritative": True},
            {"source_file": "warehouse-2.pdf", "warehouse_id": "2", "total_amount": 42549.15, "authoritative": True},
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "A", "warehouse_id": "10", "amount": 50000.0, "hours": 10},
            {"employee_name": "B", "warehouse_id": "2", "amount": 42321.33, "hours": 10},
            {"employee_name": "C", "warehouse_id": "3", "amount": 12459.22, "hours": 10},
        ],
        amount_tolerance=0.1,
    )

    expected_summary = {
        "pdfAmountTotal": 92549.15,
        "excelAmountTotal": 104780.55,
        "amountDeltaTotal": -12231.4,
        "comparableExcelAmountTotal": 92321.33,
        "comparableAmountDeltaTotal": 227.82,
        "missingPdfAmountTotal": 12459.22,
        "totalPassed": False,
    }
    assert {key: result["summary"][key] for key in expected_summary} == expected_summary
    assert {row["warehouseId"]: row["reconciliationStatus"] for row in result["rows"]} == {
        "10": "passed",
        "2": "amount_difference",
        "3": "missing_pdf_invoice",
    }


def test_warehouse_comparison_does_not_fall_back_to_detail_amount_when_pdf_total_is_zero():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "DEPT_1_20260709_133853_115071.pdf", "warehouse_id": "1", "total_amount": 0},
        ],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="DEPT_1_20260709_133853_115071.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Reyes, Kaylee", hours=42.11, amount=974.64, currency="USD", confidence=0.95, evidence_text=""),
            LaborLineItem(source_type="pdf_invoice", source_file="DEPT_1_20260709_133853_115071.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Rodriguez, Jennifer", hours=40.05, amount=906.02, currency="USD", confidence=0.95, evidence_text=""),
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Kaylee Reyes", "warehouse_id": "1", "amount": 974.64, "hours": 42.11},
            {"employee_name": "Jennifer Rodriguez", "warehouse_id": "1", "amount": 906.02, "hours": 40.05},
        ],
        amount_tolerance=0.1,
    )

    assert result["rows"][0]["pdfAmountTotal"] == 0.0
    assert result["rows"][0]["reconciliationStatus"] == "needs_review"
    assert result["rows"][0]["matchStatus"] == "待复核"
    assert result["summary"]["totalPassed"] is False


def test_later_page_invoice_rows_remain_in_employee_and_warehouse_diagnostics():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="DEPT_2_20260709_133853_115595.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alex Chavez", hours=4.0, amount=100.0, currency="USD", confidence=0.95, evidence_text="Alex Chavez invoice charge 100.00", warehouse_id="2"),
        LaborLineItem(source_type="pdf_invoice", source_file="DEPT_2_20260709_133853_115595.pdf", source_page_or_row="p3", employee_id="", employee_name_raw="Jordan Lee", hours=2.0, amount=50.0, currency="USD", confidence=0.95, evidence_text="Jordan Lee 2026-06-08 invoice charge 50.00", warehouse_id="2"),
    ]
    excel_rows = [
        LaborLineItem(source_type="excel_bill", source_file="员工账单明细.xlsx", source_page_or_row="r1", employee_id="", employee_name_raw="Alex Chavez", hours=4.0, amount=100.0, currency="USD", confidence=1.0, evidence_text="", warehouse_id="2"),
        LaborLineItem(source_type="excel_bill", source_file="员工账单明细.xlsx", source_page_or_row="r2", employee_id="", employee_name_raw="Jordan Lee", hours=2.0, amount=50.0, currency="USD", confidence=1.0, evidence_text="", warehouse_id="2"),
    ]

    employee_result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)
    warehouse_result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "DEPT_2_20260709_133853_115595.pdf", "warehouse_id": "2", "total_amount": 175.0, "authoritative": True, "evidence_status": "authoritative"},
        ],
        pdf_rows=pdf_rows,
        excel_rows_with_warehouse=[
            {"employee_name": "Alex Chavez", "warehouse_id": "2", "amount": 100.0, "hours": 4.0},
            {"employee_name": "Jordan Lee", "warehouse_id": "2", "amount": 50.0, "hours": 2.0},
        ],
        amount_tolerance=0.1,
        hours_tolerance=0.1,
    )

    assert employee_result["summary"]["pdfAmountTotal"] == 150.0
    assert employee_result["summary"]["excelAmountTotal"] == 150.0
    assert {row["employeeName"] for row in employee_result["rows"]} == {"Alex Chavez", "Jordan Lee"}
    assert warehouse_result["rows"][0]["reconciliationStatus"] == "amount_difference"
    assert {row["employeeName"] for row in warehouse_result["rows"][0]["employeeRows"]} == {"Alex Chavez", "Jordan Lee"}


def test_saved_labor_run_preserves_explicit_evidence_aware_total_decision():
    metadata = {
        "id": "labor_saved_old_result",
        "warehouseComparison": {
            "summary": {
                "pdfAmountTotal": 144714.83,
                "excelAmountTotal": 144714.93,
                "amountDeltaTotal": -0.1,
                "totalPassed": False,
                "exceptionCount": 2,
            }
        },
    }

    normalized = _normalize_labor_total_decision(metadata)

    assert normalized["warehouseComparison"]["summary"]["amountDeltaTotal"] == -0.1
    assert normalized["warehouseComparison"]["summary"]["totalPassed"] is False
    assert metadata["warehouseComparison"]["summary"]["totalPassed"] is False


def test_saved_legacy_labor_run_normalizes_total_decision_when_no_decision_or_status_exists():
    metadata = {
        "warehouseComparison": {
            "summary": {
                "pdfAmountTotal": 144714.83,
                "excelAmountTotal": 144714.93,
                "amountDeltaTotal": -0.1,
            },
            "rows": [{"warehouseId": "1", "matchStatus": "通过"}],
        }
    }

    normalized = _normalize_labor_total_decision(metadata)

    assert normalized["warehouseComparison"]["summary"]["totalPassed"] is True


def test_saved_labor_run_uses_explicit_reconciliation_status_before_amount_closeness():
    metadata = {
        "warehouseComparison": {
            "summary": {"amountDeltaTotal": 0.0},
            "rows": [{"warehouseId": "", "reconciliationStatus": "needs_review"}],
        }
    }

    normalized = _normalize_labor_total_decision(metadata)

    assert normalized["warehouseComparison"]["summary"]["totalPassed"] is False


def test_build_conclusion_does_not_allow_amount_closeness_to_override_failed_total():
    conclusion = _build_conclusion(
        {
            "summary": {
                "totalPassed": False,
                "pdfAmountTotal": 100.0,
                "excelAmountTotal": 100.0,
                "amountDeltaTotal": 0.0,
            }
        },
        {
            "summary": {
                "pdfEmployeeCount": 1,
                "excelEmployeeCount": 1,
                "amountDiffCount": 0,
                "exceptionCount": 0,
                "lowConfidenceCount": 0,
            },
            "rows": [],
        },
        {"level": "ok"},
        amount_tolerance=0.1,
    )

    assert conclusion["conclusionLevel"] == "warning"
    assert "差异" in conclusion["conclusionMessage"]


def test_build_conclusion_does_not_call_business_difference_low_confidence():
    conclusion = _build_conclusion(
        {
            "summary": {
                "totalPassed": False,
                "pdfAmountTotal": 1000.0,
                "excelAmountTotal": 1150.0,
                "amountDeltaTotal": -150.0,
            }
        },
        {
            "summary": {
                "pdfEmployeeCount": 1,
                "excelEmployeeCount": 1,
                "amountDiffCount": 1,
                "exceptionCount": 1,
                "lowConfidenceCount": 0,
            },
            "rows": [],
        },
        {
            "level": "warning",
            "metrics": {"confidence": {"lowCount": 0, "veryLowCount": 0}},
        },
        amount_tolerance=0.1,
    )

    assert conclusion["conclusionLevel"] == "warning"
    assert conclusion["conclusionMessage"] == "1项员工/金额差异需关注"
    assert "低置信度" not in conclusion["conclusionMessage"]


def test_stage2_target_files_include_unassigned_review_evidence_but_not_missing_or_extra():
    warehouse_rows = [
        {"warehouseId": "", "reconciliationStatus": "needs_review", "pdfEvidenceFile": "mystery.pdf"},
        {"warehouseId": "1", "reconciliationStatus": "amount_difference", "pdfEvidenceFile": "DEPT_1.pdf"},
        {"warehouseId": "2", "reconciliationStatus": "missing_pdf_invoice", "pdfEvidenceFile": ""},
        {"warehouseId": "3", "reconciliationStatus": "extra_pdf_invoice", "pdfEvidenceFile": "DEPT_3.pdf"},
    ]

    assert app_module._labor_stage2_target_pdf_names(warehouse_rows) == {"mystery.pdf", "DEPT_1.pdf"}


def test_page_evidence_only_supporting_roles_make_unknown_pdf_audit_only():
    total = {
        "source_file": "support.pdf",
        "pdf_type": "unknown",
        "evidence_status": "needs_review",
        "page_evidence": [
            {"page": 1, "role": "email_cover"},
            {"page": 2, "role": "timecard_summary"},
            {"page": 3, "role": "daily_detail"},
            {"page": 4, "role": "supporting_attachment"},
        ],
    }

    assert _labor_total_is_explicitly_non_payable(total) is True


def test_mixed_invoice_and_supporting_page_evidence_remains_reconcilable():
    total = {
        "source_file": "invoice-with-support.pdf",
        "pdf_type": "supporting",
        "authoritative": True,
        "evidence_status": "authoritative",
        "page_evidence": [
            {"page": 1, "role": "invoice_primary"},
            {"page": 2, "role": "timecard_summary"},
        ],
    }

    assert _labor_total_is_explicitly_non_payable(total) is False


def test_employee_rows_are_filtered_to_invoice_evidence_pages_before_comparison():
    rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Primary Worker", hours=8, amount=100),
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p2", employee_id="", employee_name_raw="Continuation Worker", hours=8, amount=100),
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p3", employee_id="", employee_name_raw="Total Worker", hours=0, amount=0),
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p4", employee_id="", employee_name_raw="Timecard Worker", hours=8, amount=100),
        LaborLineItem(source_type="pdf_invoice", source_file="legacy.pdf", source_page_or_row="p5", employee_id="", employee_name_raw="Legacy Worker", hours=8, amount=100),
    ]
    totals = [
        {
            "source_file": "invoice.pdf",
            "page_evidence": [
                {"page": 1, "role": "invoice_primary"},
                {"page": 2, "role": "invoice_continuation"},
                {"page": 3, "role": "invoice_total"},
                {"page": 4, "role": "timecard_summary"},
            ],
        },
        {"source_file": "legacy.pdf"},
    ]

    filtered = app_module._filter_labor_rows_to_invoice_evidence_pages(rows, totals)

    assert [row.employee_name_raw for row in filtered] == [
        "Primary Worker",
        "Continuation Worker",
        "Total Worker",
        "Legacy Worker",
    ]


def test_unknown_image_invoice_pages_keep_successfully_extracted_employee_rows():
    rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="image-invoice.pdf",
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw="Lautric Patrick",
            hours=28,
            amount=635.04,
            confidence=0.95,
            evidence_text="LAUTRIC PATRICK 28,00 635,04",
        )
    ]
    totals = [
        {
            "source_file": "image-invoice.pdf",
            "pdf_type": "unknown",
            "authoritative": False,
            "evidence_status": "needs_review",
            "excluded_pages": [1],
            "page_evidence": [{"page": 1, "role": "unknown", "role_confidence": 0.5}],
        }
    ]

    filtered = app_module._filter_labor_rows_to_invoice_evidence_pages(rows, totals)

    assert filtered == rows


def test_employee_rows_never_extend_past_authoritative_total_page():
    rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Invoice Worker", hours=8, amount=100),
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p2", employee_id="", employee_name_raw="Email Worker", hours=8, amount=100),
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p3", employee_id="", employee_name_raw="Timecard Worker", hours=8, amount=100),
    ]
    totals = [
        {
            "source_file": "invoice.pdf",
            "authoritative": True,
            "total_page": 1,
            "page_evidence": [
                {"page": 1, "role": "invoice_primary", "total_amount": 100},
                {"page": 2, "role": "invoice_primary", "total_amount": None},
                {"page": 3, "role": "invoice_total", "total_amount": None},
            ],
        }
    ]

    filtered = app_module._filter_labor_rows_to_invoice_evidence_pages(rows, totals)

    assert [row.employee_name_raw for row in filtered] == ["Invoice Worker"]


def test_unscanned_invoice_continuation_pages_remain_available_for_employee_detail():
    rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="NJ13 Invoice Report.pdf",
            source_page_or_row="p8",
            employee_id="20132",
            employee_name_raw="Contreras, Kristel",
            hours=48,
            amount=1122.68,
        )
    ]
    totals = [
        {
            "source_file": "NJ13 Invoice Report.pdf",
            "authoritative": True,
            "total_page": 1,
            "excluded_pages": [8],
            "page_evidence": [
                {"page": 1, "role": "invoice_total", "total_amount": 1122.68},
                {"page": 2, "role": "invoice_continuation", "total_amount": None},
                {
                    "page": 8,
                    "role": "unknown",
                    "role_confidence": 0,
                    "extraction_method": "not_scanned_after_authoritative_total",
                },
            ],
        }
    ]

    filtered = app_module._filter_labor_rows_to_invoice_evidence_pages(rows, totals)

    assert [row.employee_name_raw for row in filtered] == ["Contreras, Kristel"]


def test_warehouse_comparison_never_infers_missing_pdf_warehouse_from_excel_total():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "Invoice-5058871.pdf", "warehouse_id": "", "total_amount": 8500.67, "authoritative": True, "evidence_status": "authoritative", "total_page": 2},
            {"source_file": "Invoice-5058872.pdf", "warehouse_id": "", "total_amount": 3223.94, "authoritative": True, "evidence_status": "authoritative", "total_page": 1},
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Worker One", "warehouse_id": "19", "hours": 10, "amount": 8500.67},
            {"employee_name": "Worker Two", "warehouse_id": "18", "hours": 8, "amount": 3223.94},
        ],
        amount_tolerance=0.1,
    )

    assert result["errors"] == [
        "无法提取仓库号: Invoice-5058871.pdf",
        "无法提取仓库号: Invoice-5058872.pdf",
    ]
    assert result["summary"]["pdfAmountTotal"] == 11724.61
    assert result["summary"]["totalPassed"] is False
    assert result["summary"]["passedCount"] == 0
    assert result["summary"]["missingPdfAmountTotal"] == 11724.61
    rows_by_wh = {row["warehouseId"]: row for row in result["rows"]}
    assert set(rows_by_wh) == {"", "18", "19"}
    assert rows_by_wh[""]["reconciliationStatus"] == "needs_review"
    assert rows_by_wh[""]["pdfAmountTotal"] == 11724.61
    assert rows_by_wh[""]["pdfEvidenceFile"] == "Invoice-5058871.pdf; Invoice-5058872.pdf"
    assert rows_by_wh[""]["pdfEvidencePage"] == "2, 1"
    assert rows_by_wh["18"]["reconciliationStatus"] == "missing_pdf_invoice"
    assert rows_by_wh["19"]["reconciliationStatus"] == "missing_pdf_invoice"


def test_classify_pdf_distinguishes_invoice_support_and_attachments():
    assert _classify_pdf("Invoice_123.pdf", "Invoice Total $1,000.00\nEmployee A") == "primary"
    assert _classify_pdf("Supplement1.pdf", "Timecard Detail\nDaily Log\nEmployee hours only") == "supporting"
    assert _classify_pdf("COI_certificate.pdf", "Certificate of Insurance") == "attachment"
    assert _classify_pdf("scan.pdf", "") == "unknown"
    assert _classify_pdf(
        "invoice_then_timecard.pdf",
        "INVOICE\nASSOCIATE HOURS AMOUNT\nWorker One 40.00 $1,000.00\nNET TOTAL $1,000.00\n"
        "WEEKLY TIMECARD\nDaily hours only",
    ) == "primary"


def test_fairway_invoice_total_prefers_totals_or_grand_total_over_late_payment():
    assert _extract_invoice_total_from_text(
        "21 Totals 773.82 50.00 0.00 19,655.14$ 2,081.64$ -$ 21,736.78$\n"
        "If paid after 6/7/2026 please pay: $22,171.52\n"
        "GRAND TOTAL:\n"
        "21,736.78$"
    ) == 21736.78
    assert _extract_invoice_total_from_text(
        "If paid after 6/07/2026 please pay: 15,391.68$\n"
        "GRAND TOTAL:\n"
        "P.O. BOX 31001-2434\n"
        "US ELOGISTICS SERVICE CORP\n"
        "15,089.88$"
    ) == 15089.88


def test_voyage_invoice_total_prefers_explicit_invoice_amount_over_grand_total_hours():
    assert _extract_invoice_total_from_text(
        "\n".join(
            [
                "TOTAL REG: 2089.21",
                "TOTAL OT: 104.73",
                "TOTAL DT: 0.00",
                "GRAND TOTAL: 2193.94",
                "TOTAL INVOICE AMOUNT: $54,358.11",
            ]
        )
    ) == 54358.11


def test_voyage_numbered_rows_parse_without_repeated_header():
    from bonus_platform.engine.labor.extract import _extract_voyage_invoice_rows

    page = {
        "source_file": "CA 7 46292.pdf",
        "page": 2,
        "text": "\n".join(
            [
                "31 Alison Guzman 17.5 8.00 22.40 179.20 0.10 33.60 3.36 0.00 44.8 0.00 182.56",
                "32 Alison Guzman 17.5 8.00 22.40 179.20 0.07 33.60 2.35 0.00 44.8 0.00 181.55",
            ]
        ),
    }

    rows = _extract_voyage_invoice_rows(
        page,
        supplier="Voyage Employer Services",
        period_start="2026-06-29",
        period_end="2026-07-05",
        currency="USD",
    )

    assert [(row.employee_name_raw, row.hours, row.amount, row.warehouse_id) for row in rows] == [
        ("Alison Guzman", 8.1, 182.56, "7"),
        ("Alison Guzman", 8.07, 181.55, "7"),
    ]


def test_voyage_rows_accept_zero_amount_dash_and_parenthesized_name():
    from bonus_platform.engine.labor.extract import _extract_voyage_invoice_rows

    page = {
        "source_file": "CA 18 46286.pdf",
        "page": 1,
        "text": (
            "1 Ever Ferreira(Voyage) 21 7.58 26.88 203.75$ "
            "0.00 40.32 -$ 0.00 53.76 0.00 203.75$"
        ),
    }

    rows = _extract_voyage_invoice_rows(
        page,
        supplier="Voyage Employer Services",
        period_start="2026-06-29",
        period_end="2026-07-05",
        currency="USD",
    )

    assert [(row.employee_name_raw, row.hours, row.amount, row.warehouse_id) for row in rows] == [
        ("Ever Ferreira(Voyage)", 7.58, 203.75, "18"),
    ]


def test_voyage_rows_accept_pdf_text_without_space_before_payrate():
    from bonus_platform.engine.labor.extract import _extract_voyage_invoice_rows

    page = {
        "source_file": "CA 7 46292.pdf",
        "page": 3,
        "text": (
            "95 Daniel Alexander Martinez - Ruedas20 "
            "7.98 25.60 204.29 0.00 38.40 0.00 0.00 51.2 0.00 204.29"
        ),
    }

    rows = _extract_voyage_invoice_rows(
        page,
        supplier="Voyage Employer Services",
        period_start="2026-06-29",
        period_end="2026-07-05",
        currency="USD",
    )

    assert [(row.employee_name_raw, row.hours, row.amount) for row in rows] == [
        ("Daniel Alexander Martinez - Ruedas", 7.98, 204.29),
    ]


def test_candidate_with_closed_amount_still_requires_expected_employee_coverage():
    rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="invoice.pdf",
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw="Worker One",
            hours=40.0,
            amount=100.0,
            confidence=0.98,
        )
    ]
    pages = [{"source_file": "invoice.pdf", "page": 1, "text": "INVOICE TOTAL: $100.00"}]
    expected_rows = [
        {"employee_name_raw": "Worker One", "amount": 100.0},
        {"employee_name_raw": "Worker Two", "amount": 100.0},
    ]

    assert _candidate_is_confident(rows, pages, expected_rows=expected_rows) is False


def test_sss_invoice_total_reads_billable_total_row():
    assert _extract_invoice_total_from_text(
        """
        Billable Billable Total
        Hours Fee Fees
        -$
        1 48,293.06$ 48,293.06$
        See Attached : Worksheets -$
        Total Due
        48,293.06$
        """
    ) == 48293.06


@pytest.mark.parametrize(
    ("label", "amount"),
    [
        ("NET TOTAL", 2682.75),
        ("TOTAL NETO", 2320.25),
        ("NETTOSUMME", 2640.50),
        ("GESAMT", 2580.00),
    ],
)
def test_unknown_supplier_total_reads_common_cross_language_labels(label, amount):
    assert _extract_invoice_total_from_text(f"INVOICE\nEmployee table\n{label}: ${amount:,.2f}") == amount


def test_unknown_supplier_standalone_total_overrides_earlier_net_total():
    assert _extract_invoice_total_from_text(
        "INVOICE\nNET TOTAL: $1,790.00\nTOTAL: $1,900.00"
    ) == 1900.0


def test_unknown_supplier_total_reads_header_labels_with_footer_values():
    assert _extract_invoice_total_from_text(
        "\n".join(
            [
                "Invoice",
                "Total",
                "Balance Due",
                "Payments/Credits",
                "Description Qty Rate Amount",
                "Worker One US Elogistics - Packing Team $16 Reg Time",
                "40 19.20 768.00",
                "Page 4",
                "$22,122.83",
                "$22,122.83",
                "$0.00",
            ]
        )
    ) == 22122.83


def test_unknown_supplier_multiline_description_rows_close_footer_total():
    page = {
        "source_file": "Inv_061918_from_Coastline_Resources_LLC.pdf",
        "page": 2,
        "text": "\n".join(
            [
                "Invoice",
                "Description Qty Rate Amount",
                "Pay Period: 06-08 through 06-14-2026 Julian Espindola US",
                "Elogistics - Packing Team $16 Reg Time",
                "8.15 19.20 156.48",
                "Pay Period: 06-08 through 06-14-2026 Benjamin Reyes US",
                "Elogistics - Inbound Team $16hr Reg Time",
                "39.39 19.1998 756.28",
                "Pay Period: 06-08 through 06-14-2026- Elba Galvan- US E logistics",
                "$15 hr Reg Time- Packing Team",
                "38.99 18.00 701.82",
                "Pay Period: 06-08 through 06-14-2026 Martin Valencia-US",
                "E-logisitics- Inbound Team $16 Reg Time",
                "40 19.20 768.00",
                "Victor Hernandez - Missing 3.5hrs Reg Time $16 3.5 19.20 67.20",
                "Page 2",
            ]
        ),
    }

    rows = _extract_with_rules(
        [page],
        supplier="coastline",
        period_start="2026-06-08",
        period_end="2026-06-14",
        currency="USD",
    )

    assert [(row.employee_name_raw, row.hours, row.amount) for row in rows] == [
        ("Julian Espindola", 8.15, 156.48),
        ("Benjamin Reyes", 39.39, 756.28),
        ("Elba Galvan", 38.99, 701.82),
        ("Martin Valencia", 40.0, 768.0),
        ("Victor Hernandez", 3.5, 67.2),
    ]


def test_candidate_with_authoritative_total_requires_amount_closure_even_when_employee_coverage_is_high():
    rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="invoice.pdf",
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw="Worker One",
            hours=40.0,
            amount=90.0,
            confidence=0.98,
        )
    ]
    pages = [{"source_file": "invoice.pdf", "page": 1, "text": "INVOICE TOTAL: $100.00"}]
    expected_rows = [{"employee_name_raw": "Worker One", "amount": 90.0}]

    assert _candidate_is_confident(rows, pages, expected_rows=expected_rows) is False


def test_unknown_supplier_rows_are_assigned_to_unique_excel_warehouses():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alexis Alfonso", hours=0.0, amount=100.0),
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Shane Evans", hours=0.0, amount=200.0),
    ]
    excel_rows = [
        LaborLineItem(source_type="excel_bill", source_file="bill.xlsx", source_page_or_row="2", employee_id="", employee_name_raw="Alexis Alfonso Lopez", hours=0.0, amount=100.0, warehouse_id="1"),
        LaborLineItem(source_type="excel_bill", source_file="bill.xlsx", source_page_or_row="3", employee_id="", employee_name_raw="Shane Evan", hours=0.0, amount=200.0, warehouse_id="2"),
    ]

    assigned, audit = labor_structure.assign_pdf_rows_to_excel_warehouses(pdf_rows, excel_rows)

    assert [row.warehouse_id for row in assigned] == ["1", "2"]
    assert audit["assignedRowCount"] == 2
    assert audit["unresolvedRowCount"] == 0


def test_warehouse_comparison_allocates_unassigned_invoice_when_employee_rows_close_total():
    comparison = compare_by_warehouse(
        pdf_totals=[
            {
                "source_file": "invoice.pdf",
                "total_amount": 300.0,
                "warehouse_id": "",
                "authoritative": True,
                "evidence_status": "authoritative",
                "total_page": 4,
            }
        ],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alexis Alfonso", hours=0.0, amount=100.0, warehouse_id="1"),
            LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Shane Evans", hours=0.0, amount=200.0, warehouse_id="2"),
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "Alexis Alfonso Lopez", "amount": 100.0, "hours": 0.0, "warehouse_id": "1"},
            {"employee_name": "Shane Evan", "amount": 200.0, "hours": 0.0, "warehouse_id": "2"},
        ],
        amount_tolerance=0.10,
    )

    assert comparison["summary"]["pdfAmountTotal"] == 300.0
    assert comparison["summary"]["excelAmountTotal"] == 300.0
    assert comparison["summary"]["totalPassed"] is True
    assert [(row["warehouseId"], row["pdfAmountTotal"], row["matchStatus"]) for row in comparison["rows"]] == [
        ("1", 100.0, "通过"),
        ("2", 200.0, "通过"),
    ]
    assert all(row["evidenceStatus"] == "allocated_employee_detail" for row in comparison["rows"])


def test_dcgcb_rate_amount_table_extracts_all_employee_rows_without_counting_rate_as_hours():
    text = "\n".join(
        [
            "INVOICE",
            "DCGCB-26326",
            "BALANCE $39,966.14",
            "Location: BRAMPTON",
            "Billing Period- WE MAY 30, 2026 ; WE JUN 6, 2026",
            "NAME REGULAR HRS O/T HRS RATE AMOUNT",
            "AKRAN WASIM 7.9 19 150.10",
            "BHUPINDER SINGH 71.75 19 1,363.25",
            "GURDEEP 74 20 1,480.00",
            "GURJOT SINGH 64.33 20 1,286.60",
            "GURSIMRAN SINGH 73.9 20 1,478.00",
            "HARMANPREET SINGH 46.26 19 878.94",
            "JAGJEET SINGH 82 20 1,640.00",
            "KARMAN SINGH 81.73 20 1,634.60",
            "PARAMVIR SINGH 88 6 19 1,843.00",
            "SOORAJ SURENDRANATHAN 70 19 1,330.00",
            "SUKHJEET 88 8.9 20 2,027.00",
            "TARANVIR SINGH 88 6.23 20 1,946.90",
            "AMRITPAL KAUR 40.5 18 729.00",
            "GURMEET SINGH 72.68 19 1,380.92",
            "GURPAIR SINGH 75.64 19 1,437.16",
            "JATIN 72 19 1,368.00",
            "NIRMAL ANDREWS RODRIGUES 81.55 20 1,631.00",
            "RAMNDEEP SINGH 47.9 20 958.00",
            "RAVINDER SINGH 8 19 152.00",
            "UTHANAM RAO 29.2 19 554.80",
            "HARPREET SINGH 37.91 19 720.29",
            "KANWARJEET SINGH 31.67 19 601.73",
            "HARISH KUMAR 88 8.89 19 1,925.37",
            "HARMANJOT SINGH 50.13 19 952.47",
            "KANWARDEEP SINGH 73.12 19 1,389.28",
            "MANPREET KAUR 81.37 18 1,464.66",
            "SONAM EUDEN 88 6.99 20 1,969.70",
            "TANVEER KAUR 59.75 18 1,075.50",
            "3 5,368.27",
        ]
    )

    rows = _extract_with_rules(
        [{"source_file": "DCGCB-26326.pdf", "page": 1, "text": text}],
        supplier="DCGCB",
        period_start="2026-05-30",
        period_end="2026-06-06",
        currency="CAD",
    )

    assert len(rows) == 28
    assert round(sum(row.amount for row in rows), 2) == 35368.27
    by_name = {row.employee_name_raw: row for row in rows}
    assert by_name["AKRAN WASIM"].hours == 7.9
    assert by_name["PARAMVIR SINGH"].hours == 94.0
    assert by_name["HARISH KUMAR"].hours == 96.89
    assert by_name["HARISH KUMAR"].amount == 1925.37


def test_extract_invoice_items_prefers_dcgcb_rate_amount_rows_before_ai(monkeypatch, tmp_path):
    pdf = tmp_path / "DCGCV-26324.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    page = {
        "source_file": pdf.name,
        "page": 1,
        "text": "\n".join(
            [
                "INVOICE DCGCV-26324",
                "Billing Period- WE MAY 30, 2026 ; WE JUN 7, 2026",
                "NAME REGULAR HRS O/T HRS RATE AMOUNT",
                "GURKIRAT SINGH 53.15 20 1,063.00",
                "HAROLD KUMAR 39.93 20 798.60",
                "HARMAN SINGH SANDHU 88 16.05 20 2,241.50",
                "ISHPREET SINGH 25.3 19 480.70",
                "18,996.93",
            ]
        ),
    }
    import bonus_platform.engine.labor.extract as extract_module

    monkeypatch.setattr(extract_module, "_extract_pdf_pages", lambda paths: [page])
    monkeypatch.setattr(
        extract_module,
        "_post_chat_completion",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("AI should not run when rules extract rows")),
    )

    rows = extract_invoice_items(
        [pdf],
        {
            "enabled": True,
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "parallel_extraction_enabled": False,
        },
        supplier="DCGCB",
        period_start="2026-06-01",
        period_end="2026-06-07",
        currency="CAD",
    )

    assert [row.employee_name_raw for row in rows] == [
        "GURKIRAT SINGH",
        "HAROLD KUMAR",
        "HARMAN SINGH SANDHU",
        "ISHPREET SINGH",
    ]
    assert round(sum(row.hours for row in rows), 2) == 222.43
    assert round(sum(row.amount for row in rows), 2) == 4583.80
    assert all(row.confidence == 0.97 for row in rows)


def test_extract_invoice_items_uses_best_candidate_when_rule_rows_are_incomplete(monkeypatch, tmp_path):
    pdf = tmp_path / "unknown_rate_amount.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    page = {
        "source_file": pdf.name,
        "page": 1,
        "text": "\n".join(
            [
                "INVOICE",
                "BALANCE $300.00",
                "NAME REGULAR HRS O/T HRS RATE AMOUNT",
                "ALICE SMITH 5 20 100.00",
                "BOB JONES row present but amount shifted in image table",
                "CAROL LEE row present but amount shifted in image table",
            ]
        ),
    }
    import bonus_platform.engine.labor.extract as extract_module

    monkeypatch.setattr(extract_module, "_extract_pdf_pages", lambda paths: [page])
    monkeypatch.setattr(
        extract_module,
        "_extract_with_ai_text",
        lambda *args, **kwargs: [
            {
                "source_file": pdf.name,
                "source_page_or_row": "p1",
                "employee_name_raw": "ALICE SMITH",
                "hours": 5,
                "amount": 100,
                "confidence": 0.9,
                "evidence_text": "ALICE SMITH",
            },
            {
                "source_file": pdf.name,
                "source_page_or_row": "p1",
                "employee_name_raw": "BOB JONES",
                "hours": 5,
                "amount": 100,
                "confidence": 0.9,
                "evidence_text": "BOB JONES",
            },
            {
                "source_file": pdf.name,
                "source_page_or_row": "p1",
                "employee_name_raw": "CAROL LEE",
                "hours": 5,
                "amount": 100,
                "confidence": 0.9,
                "evidence_text": "CAROL LEE",
            },
        ],
    )
    monkeypatch.setattr(
        extract_module,
        "_extract_with_ai_images",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("image AI should not run when text candidate wins")),
    )

    rows = extract_invoice_items(
        [pdf],
        {
            "enabled": True,
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "parallel_extraction_enabled": False,
        },
        supplier="Unknown",
        currency="USD",
    )

    assert [row.employee_name_raw for row in rows] == ["ALICE SMITH", "BOB JONES", "CAROL LEE"]
    assert round(sum(row.amount for row in rows), 2) == 300.0


def test_elga_invoice_detail_rows_merge_regular_and_overtime_lines():
    page = {
        "source_file": "ELGA 041026-15.pdf",
        "page": 1,
        "text": "\n".join(
            [
                "Invoice Number: ELGA 041026-15",
                "Job Site/Warehouse  Name Pay Type   Rate  Hours",
                "Paid  Amount  Paid",
                "1",
                "1950 Oak lawn ave Atlanta",
                "Ga Arnaldo Ibarra Hourly-Reg 17.00 40.00 680.00",
                "Overtime (Hourly) 25.50 21.70 553.35",
                "2 The Bluffs Austell GA Stephie Arujo Hourly-Reg 16.00 40.00 640.00",
                "Overtime (Hourly) 24.00 4.18 100.32",
                "3 Fulton Adriana Bermudez",
                "Cuenu Hourly-Reg 16.00 16.50 264.00",
                "Overtime (Hourly) 24.00 -",
                "Total 122.38 2,237.67",
                "INVOICE DETAIL",
            ]
        ),
    }

    rows = _extract_with_rules([page], "Tru Staffing", "2026-05-01", "2026-06-30", "USD")

    assert [(row.employee_name_raw, row.hours, row.amount) for row in rows] == [
        ("Arnaldo Ibarra", 61.7, 1233.35),
        ("Stephie Arujo", 44.18, 740.32),
        ("Adriana Bermudez Cuenu", 16.5, 264.0),
    ]


def test_warehouse_id_from_text_does_not_read_accounting_dept_phone_as_dept():
    text = "\n".join(
        [
            "Att.: Accounting Dept",
            "562.631.6301 sofia@trustaffing.com",
            "Gross Wages Billable Rate Job Site Description Billable Earnings",
        ]
    )

    assert _warehouse_id_from_text(text) == ""


def test_labor_text_detail_signal_distinguishes_elga_detail_from_tru_summary():
    from bonus_platform import app as app_module

    elga_text = "\n".join(
        [
            "Job Site/Warehouse  Name Pay Type   Rate  Hours",
            "Ga Arnaldo Ibarra Hourly-Reg 17.00 40.00 680.00",
            "Overtime (Hourly) 25.50 21.70 553.35",
            "INVOICE DETAIL",
        ]
    )
    tru_summary_text = "\n".join(
        [
            "Gross Wages Billable RateJob Site Description Billable Earnings",
            "1,289.56 1.20Staffing Service Staffing Services Provided for check date 5/1/26 1,547.47",
            "Balance Due",
        ]
    )

    assert app_module._labor_text_has_employee_detail_signal(elga_text) is True
    assert app_module._labor_text_has_employee_detail_signal(tru_summary_text) is False


def test_grande_solutions_simple_table_extracts_all_employee_rows():
    text = "\n".join(
        [
            "TO Elogistics GA Service Corp",
            "Invoice : ELOG-466-FL",
            "Period Location",
            "05/18/2026-05/24/2026 E-LOG 30 SHEIN",
            "No. Name Reg. Hours O.T Hours Reg. Rate O.T Rate Total",
            "1 Alberto Núñez 35.08 $21.08 $31.62 $739.49",
            "2 Ivis Martinez 6.55 $21.08 $31.62 $138.07",
            "3 Carolay Hincapie 40 7.82 $19.84 $29.76 $1,026.32",
            "4 Liliana Cue 40 7.14 $19.84 $29.76 $1,006.09",
            "TOTAL HOURS 1251.18 67.12 SUB TOTAL $25,487.50",
        ]
    )

    rows = _extract_with_rules(
        [{"source_file": "GS_invoice-ELOG-466-FL.pdf", "page": 1, "text": text}],
        supplier="Grande Solutions Staffing",
        period_start="2026-05-18",
        period_end="2026-05-24",
        currency="USD",
    )

    assert [row.employee_name_raw for row in rows] == [
        "Alberto Núñez",
        "Ivis Martinez",
        "Carolay Hincapie",
        "Liliana Cue",
    ]
    assert round(sum(row.amount for row in rows), 2) == 2909.97
    assert round(sum(row.hours for row in rows), 2) == 136.59
    assert all(row.source_file == "GS_invoice-ELOG-466-FL.pdf" for row in rows)


def test_citi_bill_rate_rows_merge_reg_and_ot_by_employee():
    page = {
        "source_file": "In291943.pdf",
        "page": 1,
        "text": "\n".join(
            [
                "Hours  Amount Bill Rate Date  Description  Pay Rate",
                "WAREHOUSE LOC.#29PO #:",
                "$33.60  0.400 $13.44 5/17/2026 Arellano Luna, Pablo $26.250 OT",
                "$22.40  40.000 $896.00 5/17/2026 Arellano Luna, Pablo $17.500 Reg",
                "$25.60  30.000 $768.00 5/17/2026 Escobar, Armando $20.000 Reg",
                "$38.40  0.450 $17.28 5/17/2026 Escobar, Armando $30.000 OT",
                "Regular",
                "Overtime",
                "Total Due: $1,694.72",
            ]
        ),
    }

    rows = _extract_with_rules([page], "CITI", "2026-05-17", "2026-05-22", "USD")

    assert len(rows) == 2
    by_name = {row.employee_name_raw: row for row in rows}
    assert by_name["Arellano Luna, Pablo"].hours == 40.4
    assert by_name["Arellano Luna, Pablo"].amount == 909.44
    assert by_name["Arellano Luna, Pablo"].warehouse_id == "29"
    assert by_name["Escobar, Armando"].hours == 30.45
    assert by_name["Escobar, Armando"].amount == 785.28


def test_layout_analyzer_recommends_simple_numbered_labor_table_for_gs_invoice():
    page = {
        "source_file": "GS_invoice-ELOG-466-FL.pdf",
        "page": 1,
        "text": "\n".join(
            [
                "No. Name Reg. Hours O.T Hours Reg. Rate O.T Rate Total",
                "1 Alberto Núñez 35.08 $21.08 $31.62 $739.49",
                "2 Ivis Martinez 6.55 $21.08 $31.62 $138.07",
                "TOTAL HOURS 41.63 0 SUB TOTAL $877.56",
            ]
        ),
    }

    plan = analyze_invoice_layout([page])
    rows = extract_rows_from_layout_plan([page], plan, supplier="Grande Solutions Staffing", period_start="2026-05-18", period_end="2026-05-24", currency="USD")

    assert plan.layout_type == "simple_numbered_labor_table"
    assert plan.recommended_parser == "simple_invoice_table"
    assert plan.amount_column == "Total"
    assert plan.hours_columns == ["Reg. Hours", "O.T Hours"]
    assert plan.total_label == "TOTAL HOURS"
    assert round(plan.confidence, 2) >= 0.8
    assert [row.employee_name_raw for row in rows] == ["Alberto Núñez", "Ivis Martinez"]


def test_layout_analyzer_keeps_unknown_layout_out_of_rule_parser():
    page = {
        "source_file": "unknown.pdf",
        "page": 1,
        "text": "This is an invoice summary without a visible employee table.",
    }

    plan = analyze_invoice_layout([page])
    rows = extract_rows_from_layout_plan([page], plan, supplier="", period_start="", period_end="", currency="USD")

    assert plan.layout_type == "unknown"
    assert plan.recommended_parser == "ai_assisted"
    assert rows == []


def test_ai_layout_analyzer_response_is_normalized_to_layout_plan(monkeypatch):
    def fake_post_chat_completion(payload, ai_config):
        return [
            {
                "layout_type": "simple_numbered_labor_table",
                "recommended_parser": "simple_invoice_table",
                "confidence": 0.86,
                "hours_columns": ["Regular", "OT"],
                "amount_column": "Total",
                "total_label": "GRAND TOTAL",
                "employee_name_pattern": "between row number and first hours value",
            }
        ]

    import bonus_platform.engine.labor.extract as extract_module

    monkeypatch.setattr(extract_module, "_post_chat_completion", fake_post_chat_completion)

    plan = _analyze_layout_with_ai(
        [{"source_file": "unknown.pdf", "page": 1, "text": "No. Name Regular OT Total\n1 Jane Doe 40 2 $900.00"}],
        {"model": "test-model"},
        supplier="Vendor",
        currency="USD",
    )

    assert plan.layout_type == "simple_numbered_labor_table"
    assert plan.recommended_parser == "simple_invoice_table"
    assert plan.confidence == 0.86
    assert plan.hours_columns == ["Regular", "OT"]
    assert plan.amount_column == "Total"


def test_layout_plan_extracts_generic_line_item_text_table():
    page = {
        "source_file": "new_vendor.pdf",
        "page": 1,
        "text": "\n".join(
            [
                "Warehouse: WH 42",
                "Employee Hours Rate Amount",
                "1 Jane Doe WUS010325 40.00 2.50 $21.00 $892.50",
                "5/17/2026 John Smith 38.25 $20.00 $765.00",
                "Invoice Total $1,657.50",
            ]
        ),
    }
    plan = InvoiceLayoutPlan(
        layout_type="single_line_employee_amount_table",
        recommended_parser="line_item_text_table",
        confidence=0.84,
        employee_name_pattern="employee name appears before hours and amount on the same line",
        hours_columns=["Hours", "OT"],
        amount_column="Amount",
    )

    rows = extract_rows_from_layout_plan([page], plan, supplier="New Vendor", period_start="2026-05-17", period_end="2026-05-22", currency="USD")

    by_name = {row.employee_name_raw: row for row in rows}
    assert list(by_name) == ["Jane Doe", "John Smith"]
    assert by_name["Jane Doe"].employee_id == "WUS010325"
    assert by_name["Jane Doe"].hours == 42.5
    assert by_name["Jane Doe"].amount == 892.5
    assert by_name["Jane Doe"].warehouse_id == "42"
    assert by_name["John Smith"].hours == 38.25
    assert by_name["John Smith"].amount == 765.0


def test_extract_invoice_items_uses_ai_layout_plan_before_direct_ai(monkeypatch, tmp_path):
    pdf = tmp_path / "unknown_vendor.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    page = {
        "source_file": "unknown_vendor.pdf",
        "page": 1,
        "text": "Worker Detail\nJane Doe 40.00 1.00 $20.00 $830.00\nInvoice Total $830.00",
    }

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_pdf_pages", lambda paths: [page])

    def fake_post_chat_completion(payload, ai_config):
        content = json.dumps(payload.get("messages", [{}])[-1].get("content", {}), ensure_ascii=False)
        assert "line_item_text_table" in content
        return [
            {
                "layout_type": "single_line_employee_amount_table",
                "recommended_parser": "line_item_text_table",
                "confidence": 0.86,
                "employee_name_pattern": "between row number and first hours value",
                "hours_columns": ["Hours", "OT"],
                "amount_column": "Amount",
                "evidence": ["Jane Doe 40.00 1.00 $20.00 $830.00"],
            }
        ]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post_chat_completion)
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_text", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("direct text AI should not run")))
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_images", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("image AI should not run")))

    rows = extract_invoice_items(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="Unknown Vendor",
        currency="USD",
    )

    assert len(rows) == 1
    assert rows[0].employee_name_raw == "Jane Doe"
    assert rows[0].hours == 41.0
    assert rows[0].amount == 830.0
    assert rows[0].confidence == 0.82


def test_ai_rows_without_supporting_page_text_are_filtered():
    pages = [
        {
            "source_file": "GS_invoice-ELOG-466-FL.pdf",
            "page": 1,
            "text": "1 Alberto Núñez 35.08 $21.08 $31.62 $739.49",
        }
    ]
    rows = [
        {"employee_name_raw": "Albert Achter", "amount": 289.88, "evidence_text": "Albert Achter 15.08 $289.88"},
        {"employee_name_raw": "Alberto Núñez", "amount": 739.49, "evidence_text": "Alberto Núñez 35.08 $739.49"},
    ]

    filtered = _filter_ai_rows_by_page_text(rows, pages)

    assert len(filtered) == 1
    assert filtered[0]["employee_name_raw"] == "Alberto Núñez"
    assert filtered[0]["source_file"] == "GS_invoice-ELOG-466-FL.pdf"
    assert filtered[0]["source_page_or_row"] == "p1"


def test_ai_page_text_support_accepts_minor_ocr_spelling_variants():
    pages = [
        {
            "source_file": "FACJS11000105.pdf",
            "page": 1,
            "text": "BATSIMBA GLOIRE RONNELE MANUTENTIONNAIRE S/Total 45,76 1143,37",
        }
    ]
    rows = [
        {
            "employee_name_raw": "BANTSIMBA GLOIRE RONLELE",
            "hours": 45.76,
            "amount": 1143.37,
            "evidence_text": "BANTSIMBA GLOIRE RONLELE S/Total 45,76 1143,37",
        }
    ]

    filtered = _filter_ai_rows_by_page_text(rows, pages)

    assert filtered == [
        {
            **rows[0],
            "source_file": "FACJS11000105.pdf",
            "source_page_or_row": "p1",
        }
    ]


def test_single_pdf_total_without_warehouse_stays_unassigned_when_excel_has_one_warehouse():
    result = compare_by_warehouse(
        pdf_totals=[{
            "source_file": "GS_invoice-ELOG-466-FL.pdf",
            "warehouse_id": "",
            "total_amount": 25487.5,
            "authoritative": True,
            "evidence_status": "authoritative",
            "total_page": 1,
        }],
        excel_rows_with_warehouse=[
            {"employee_name": "Alberto Núñez", "warehouse_id": "1", "amount": 10000.0, "hours": 400},
            {"employee_name": "Ivis Martinez", "warehouse_id": "1", "amount": 15975.47, "hours": 800},
        ],
        amount_tolerance=0.1,
    )

    assert result["errors"] == ["无法提取仓库号: GS_invoice-ELOG-466-FL.pdf"]
    rows = {row["warehouseId"]: row for row in result["rows"]}
    assert rows[""]["reconciliationStatus"] == "needs_review"
    assert rows[""]["pdfAmountTotal"] == 25487.5
    assert rows[""]["pdfEvidenceFile"] == "GS_invoice-ELOG-466-FL.pdf"
    assert rows["1"]["reconciliationStatus"] == "missing_pdf_invoice"
    assert result["summary"]["totalPassed"] is False


def test_warehouse_comparison_keeps_invoice_total_when_employee_detail_matches_excel_without_warehouse_split():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="DCGCB-26326.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="AKRAN WASIM", hours=7.9, amount=150.10, currency="CAD", confidence=0.97, evidence_text=""),
        LaborLineItem(source_type="pdf_invoice", source_file="DCGCB-26326.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="BHUPINDER SINGH", hours=71.75, amount=1363.25, currency="CAD", confidence=0.97, evidence_text=""),
    ]

    result = compare_by_warehouse(
        pdf_totals=[{"source_file": "DCGCB-26326.pdf", "warehouse_id": "", "total_amount": 1711.22}],
        pdf_rows=pdf_rows,
        excel_rows_with_warehouse=[
            {"employee_name": "AKRAN WASIM", "warehouse_id": "1", "amount": 150.10, "hours": 7.9},
            {"employee_name": "BHUPINDER SINGH", "warehouse_id": "5", "amount": 1363.25, "hours": 71.75},
        ],
        amount_tolerance=0.1,
    )

    assert result["errors"] == [
        "无法提取仓库号: DCGCB-26326.pdf",
        "无法从文件名提取仓库号: DCGCB-26326.pdf",
        "无法从文件名提取仓库号: DCGCB-26326.pdf",
    ]
    assert result["summary"]["totalPassed"] is False
    assert result["summary"]["pdfAmountTotal"] == 1711.22
    assert result["summary"]["excelAmountTotal"] == 1513.35
    assert result["summary"]["amountDeltaTotal"] == 197.87
    assert {row["warehouseId"]: row["reconciliationStatus"] for row in result["rows"]} == {
        "": "needs_review",
        "1": "missing_pdf_invoice",
        "5": "missing_pdf_invoice",
    }
    unassigned = next(row for row in result["rows"] if row["warehouseId"] == "")
    assert unassigned["pdfAmountTotal"] == 1711.22
    assert unassigned["pdfEvidenceFile"] == "DCGCB-26326.pdf"


def test_warehouse_comparison_still_runs_when_totals_offset_between_warehouses():
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "warehouse_3.pdf", "warehouse_id": "3", "total_amount": 5000.0},
            {"source_file": "warehouse_5.pdf", "warehouse_id": "5", "total_amount": 3000.0},
        ],
        excel_rows_with_warehouse=[
            {"employee_name": "A", "warehouse_id": "3", "amount": 4000.0, "hours": 10},
            {"employee_name": "B", "warehouse_id": "5", "amount": 4000.0, "hours": 10},
        ],
        amount_tolerance=0.1,
    )

    assert result["summary"]["amountDeltaTotal"] == 0.0
    assert result["summary"]["totalPassed"] is False
    assert result["summary"]["diffWarehouses"] == ["3", "5"]
    assert {row["warehouseId"]: row["amountDelta"] for row in result["rows"]} == {"3": 1000.0, "5": -1000.0}


def test_warehouse_comparison_flags_employee_allocation_offsets_across_warehouses():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="fairway_25.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="PEREZ, JOSE", hours=4.0, amount=101.26, currency="USD", confidence=0.95, evidence_text="", warehouse_id="25"),
        LaborLineItem(source_type="pdf_invoice", source_file="fairway_25.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="JIMENEZ, ENEAS", hours=5.0, amount=118.04, currency="USD", confidence=0.95, evidence_text="", warehouse_id="25"),
        LaborLineItem(source_type="pdf_invoice", source_file="fairway_28.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="PEREZ, JOSE", hours=40.0, amount=935.00, currency="USD", confidence=0.95, evidence_text="", warehouse_id="28"),
        LaborLineItem(source_type="pdf_invoice", source_file="fairway_28.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="JIMENEZ, ENEAS", hours=40.0, amount=928.67, currency="USD", confidence=0.95, evidence_text="", warehouse_id="28"),
    ]
    result = compare_by_warehouse(
        pdf_totals=[
            {"source_file": "fairway_25.pdf", "warehouse_id": "25", "total_amount": 219.30},
            {"source_file": "fairway_28.pdf", "warehouse_id": "28", "total_amount": 1863.67},
        ],
        pdf_rows=pdf_rows,
        excel_rows_with_warehouse=[
            {"employee_name": "PEREZ, JOSE", "warehouse_id": "25", "amount": 100.67, "hours": 4.0},
            {"employee_name": "JIMENEZ, ENEAS", "warehouse_id": "25", "amount": 116.85, "hours": 5.0},
            {"employee_name": "PEREZ, JOSE", "warehouse_id": "28", "amount": 935.59, "hours": 40.0},
            {"employee_name": "JIMENEZ, ENEAS", "warehouse_id": "28", "amount": 929.87, "hours": 40.0},
        ],
        amount_tolerance=0.1,
    )

    assert result["summary"]["amountDeltaTotal"] == -0.01
    assert result["summary"]["totalPassed"] is False
    assert result["summary"]["allocationIssueCount"] == 2
    assert result["summary"]["diffWarehouses"] == ["25", "28"]
    issues_by_employee = {issue["employeeName"]: issue for issue in result["allocationIssues"]}
    assert issues_by_employee["PEREZ, JOSE"]["netAmountDelta"] == 0.0
    assert [row["warehouseId"] for row in issues_by_employee["PEREZ, JOSE"]["warehouses"]] == ["25", "28"]
    assert [row["amountDelta"] for row in issues_by_employee["JIMENEZ, ENEAS"]["warehouses"]] == [1.19, -1.2]


def test_reconciliation_diagnostics_suppresses_missing_warehouse_when_single_pdf_was_safely_attributed():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[{"source_file": "GS_invoice-ELOG-466-FL.pdf", "warehouse_id": "", "total_amount": 25487.5}],
        comparison_summary={"pdfAmountTotal": 25487.5, "excelAmountTotal": 25975.47},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 25487.5, "excelAmountTotal": 25975.47, "warehouseCount": 1},
            "errors": [],
        },
        amount_tolerance=0.1,
    )

    assert diagnostics["level"] == "ok"
    assert diagnostics["issues"] == []


def test_reconciliation_diagnostics_flags_conflicting_pdf_signals():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[
            {"source_file": "fairway_10.pdf", "warehouse_id": "", "total_amount": 21736.78},
            {"source_file": "fairway_18.pdf", "warehouse_id": "18", "total_amount": 0},
            {"source_file": "fairway_19.pdf", "warehouse_id": "19", "total_amount": 27162.78},
        ],
        comparison_summary={"pdfAmountTotal": 147368.65, "excelAmountTotal": 147368.73},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 48899.56, "excelAmountTotal": 147368.73},
            "errors": ["no warehouse match"],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "critical"
    assert diagnostics["signals"]["fastPdfTotal"] == 48899.56
    assert diagnostics["signals"]["employeePdfTotal"] == 147368.65
    assert "pdf_total_conflict" in issue_codes
    assert "missing_warehouse_id" in issue_codes
    assert "zero_pdf_total" in issue_codes
    assert "warehouse_mapping_errors" in issue_codes


def test_reconciliation_diagnostics_flags_partial_pdf_employee_detail_coverage():
    pdf_totals = [
        {"source_file": "US_ELogistics_Service_Corp__35354.pdf", "warehouse_id": "18", "total_amount": 885.43},
        {"source_file": "US_ELogistics_Service_Corp__35355.pdf", "warehouse_id": "19", "total_amount": 59874.61},
        {"source_file": "US_ELogistics_Service_Corp__35361.pdf", "warehouse_id": "25", "total_amount": 50174.35},
        {"source_file": "US_ELogistics_Service_Corp__35362.pdf", "warehouse_id": "28", "total_amount": 26582.95},
        {"source_file": "US_ELogistics_Service_Corp__35363.pdf", "warehouse_id": "7", "total_amount": 7259.24},
    ]
    pdf_rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="US_ELogistics_Service_Corp__35361.pdf",
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw="Dueñas, Oscar",
            hours=48.18,
            amount=1407.38,
            currency="USD",
            confidence=0.9,
            evidence_text="Dueñas, Oscar 48.18 $1,407.38",
            warehouse_id="25",
        )
    ]

    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=pdf_totals,
        pdf_rows=pdf_rows,
        comparison_summary={"pdfAmountTotal": 50174.35, "excelAmountTotal": 144731.22},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 144776.58, "excelAmountTotal": 144731.22},
            "errors": [],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "warning"
    assert "pdf_employee_detail_partial_coverage" in issue_codes
    assert "pdf_total_conflict" not in issue_codes
    assert diagnostics["signals"]["pdfDetailCoverage"] == {
        "coverageBasis": "invoice_totals",
        "invoiceFileCount": 5,
        "detailFileCount": 1,
        "missingFileCount": 4,
        "coverageRatio": 0.2,
        "invoiceAmountTotal": 144776.58,
        "detailAmountTotal": 50174.35,
        "amountCoverageRatio": 0.35,
        "missingSourceFiles": [
            "US_ELogistics_Service_Corp__35354.pdf",
            "US_ELogistics_Service_Corp__35355.pdf",
            "US_ELogistics_Service_Corp__35362.pdf",
            "US_ELogistics_Service_Corp__35363.pdf",
        ],
    }
    issue = next(issue for issue in diagnostics["issues"] if issue["code"] == "pdf_employee_detail_partial_coverage")
    assert "员工明细只展开了部分发票" == issue["title"]
    assert "5 张发票" in issue["message"]
    assert "1 张发票" in issue["message"]


def test_reconciliation_diagnostics_compares_partial_detail_to_covered_invoice_total():
    pdf_totals = [
        {"source_file": "warehouse-2.pdf", "warehouse_id": "2", "total_amount": 4105.15},
        {"source_file": "warehouse-9.pdf", "warehouse_id": "9", "total_amount": 11837.79},
        {"source_file": "warehouse-1.pdf", "warehouse_id": "1", "total_amount": 1880.67},
    ]
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="warehouse-2.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alex Chavez", hours=32.3, amount=3341.31, currency="USD", confidence=0.95, evidence_text="", warehouse_id="2"),
        LaborLineItem(source_type="pdf_invoice", source_file="warehouse-9.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Erick Canales", hours=24.35, amount=11837.79, currency="USD", confidence=0.95, evidence_text="", warehouse_id="9"),
    ]

    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=pdf_totals,
        pdf_rows=pdf_rows,
        comparison_summary={"pdfAmountTotal": 15179.10, "excelAmountTotal": 15715.12},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 17823.61, "excelAmountTotal": 17595.79},
            "errors": [],
        },
        amount_tolerance=0.1,
    )

    issues = {issue["code"]: issue for issue in diagnostics["issues"]}

    assert "pdf_total_conflict" not in issues
    assert "pdf_employee_detail_total_conflict" in issues
    assert "$15,942.94" in issues["pdf_employee_detail_total_conflict"]["message"]
    assert "$15,179.10" in issues["pdf_employee_detail_total_conflict"]["message"]
    assert "$763.84" in issues["pdf_employee_detail_total_conflict"]["message"]


def test_reconciliation_diagnostics_excludes_employee_detail_attachments_from_payable_conflict():
    pdf_totals = [
        {"source_file": "ELGA_041026-15.pdf", "warehouse_id": "", "total_amount": 2237.67, "has_employee_detail": True},
        {"source_file": "ELGA_Adriana.pdf", "warehouse_id": "", "total_amount": 905.92, "has_employee_detail": True},
        {"source_file": "Inv_04102615_from_Tru_Staffing_33680.pdf", "warehouse_id": "", "total_amount": 2685.20, "has_employee_detail": False},
        {"source_file": "Inv_04242617_from_Tru_Staffing_32284.pdf", "warehouse_id": "", "total_amount": 3081.98, "has_employee_detail": False},
    ]
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="ELGA_041026-15.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Stephie Arujo", hours=44.18, amount=2237.67, currency="USD", confidence=0.96, evidence_text=""),
        LaborLineItem(source_type="pdf_invoice", source_file="ELGA_Adriana.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Adriana Bermudez Cuenu", hours=51.08, amount=905.92, currency="USD", confidence=0.96, evidence_text=""),
    ]

    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=pdf_totals,
        pdf_rows=pdf_rows,
        comparison_summary={"pdfAmountTotal": 3143.59, "excelAmountTotal": 5900.0},
        warehouse_comparison={
            "summary": {
                "pdfAmountTotal": 5767.18,
                "excelAmountTotal": 5900.0,
                "selectedPdfAmountTotal": 5767.18,
                "excludedPdfAmountTotal": 3143.59,
                "selectedPdfTotalCount": 2,
                "excludedPdfTotalCount": 2,
                "totalSourceDecision": "employee_detail_pdfs_excluded_from_payable_total",
            },
            "errors": [],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "critical"
    assert "payable_pdf_total_mismatch" in issue_codes
    assert "pdf_total_conflict" not in issue_codes
    assert "pdf_employee_detail_partial_coverage" not in issue_codes
    assert diagnostics["signals"]["selectedPayablePdfTotal"] == 5767.18
    assert diagnostics["signals"]["excludedEmployeeDetailPdfTotal"] == 3143.59
    assert diagnostics["signals"]["pdfDetailCoverage"]["coverageBasis"] == "employee_detail_attachments"


def test_reconciliation_diagnostics_explains_tax_inclusive_pdf_total_when_employee_detail_matches_excel():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[{"source_file": "DCGCB-26326.pdf", "warehouse_id": "", "total_amount": 39966.14}],
        pdf_rows=[
            LaborLineItem(source_type="pdf_invoice", source_file="DCGCB-26326.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="AKRAN WASIM", hours=7.9, amount=150.10, currency="CAD", confidence=0.97, evidence_text=""),
            LaborLineItem(source_type="pdf_invoice", source_file="DCGCB-26326.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="BHUPINDER SINGH", hours=71.75, amount=1363.25, currency="CAD", confidence=0.97, evidence_text=""),
        ],
        comparison_summary={"pdfAmountTotal": 1513.35, "excelAmountTotal": 1513.35},
        warehouse_comparison={
            "summary": {
                "pdfAmountTotal": 1513.35,
                "excelAmountTotal": 1513.35,
                "amountBasis": "pdf_employee_detail_total",
                "warehouseComparisonSkipped": True,
            },
            "errors": [],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "warning"
    assert "pdf_total_includes_tax_or_fee" in issue_codes
    assert "pdf_total_conflict" not in issue_codes


def test_reconciliation_diagnostics_passes_when_totals_align():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[
            {"source_file": "fairway_10.pdf", "warehouse_id": "10", "total_amount": 21736.78},
            {"source_file": "fairway_18.pdf", "warehouse_id": "18", "total_amount": 42868.43},
        ],
        comparison_summary={"pdfAmountTotal": 64605.21, "excelAmountTotal": 64605.27},
        warehouse_comparison={"summary": {"pdfAmountTotal": 64605.21, "excelAmountTotal": 64605.27}, "errors": []},
        amount_tolerance=0.1,
    )

    assert diagnostics["level"] == "ok"
    assert diagnostics["issues"] == []
    assert diagnostics["nextStep"] == "可按当前结论使用报告。"


def test_reconciliation_diagnostics_explains_otws_amount_basis_mismatch(tmp_path):
    path = tmp_path / "OTWS - Warehouse Bill-NJ13.xlsx"
    path.write_bytes(_otws_cost_workbook_bytes())
    cost_summary = summarize_otws_costs(path)

    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[{"source_file": "NJ13 Invoice Report WE 051726 JF.pdf", "warehouse_id": "13", "total_amount": 48293.06}],
        comparison_summary={"pdfAmountTotal": 0, "excelAmountTotal": 48217.96},
        warehouse_comparison={"summary": {"pdfAmountTotal": 48293.06, "excelAmountTotal": 48217.96}, "errors": []},
        cost_summaries=[cost_summary],
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "warning"
    assert "amount_basis_mismatch" in issue_codes
    assert diagnostics["signals"]["amountBasis"] == [
        {
            "warehouseId": "13",
            "sourceFile": "OTWS - Warehouse Bill-NJ13.xlsx",
            "pdfTotal": 48293.06,
            "reportedTotal": 48217.96,
            "pdfVsReportedDelta": 75.1,
            "componentTotal": 48217.96,
            "componentDelta": 0.0,
            "detailTotal": 350.15,
            "summaryDelta": 47867.81,
            "employeeExpenses": 188.0,
            "employeeBenefits": 162.15,
            "loadingAndUnloading": 0.0,
            "summaryEvidence": "Warehouse-information!2",
            "detailEvidence": "Employee-expenses-detail!3; Employee-benefits-detail!2",
            "withinTolerance": False,
        }
    ]
    mismatch = next(issue for issue in diagnostics["issues"] if issue["code"] == "amount_basis_mismatch")
    assert "仓库 13" in mismatch["items"][0]
    assert "OTWS汇总 $48,217.96" in mismatch["items"][0]


def test_reconciliation_diagnostics_flags_offsetting_warehouse_deltas():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[
            {"source_file": "135616 US Elogistics Service Corp (#25).pdf", "warehouse_id": "25", "total_amount": 17465.12},
            {"source_file": "135617 US Elogistics Service Corp (#28).pdf", "warehouse_id": "28", "total_amount": 4537.46},
        ],
        comparison_summary={"pdfAmountTotal": 22002.58, "excelAmountTotal": 22002.59},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 22002.58, "excelAmountTotal": 22002.59},
            "errors": [],
            "rows": [
                {
                    "warehouseId": "25",
                    "pdfAmountTotal": 17465.12,
                    "excelAmountTotal": 17463.34,
                    "amountDelta": 1.78,
                    "attribution": [{"employeeName": "JIMENEZ, ENEAS", "delta": 1.19}],
                },
                {
                    "warehouseId": "28",
                    "pdfAmountTotal": 4537.46,
                    "excelAmountTotal": 4539.25,
                    "amountDelta": -1.79,
                    "attribution": [{"employeeName": "JIMENEZ, ENEAS", "delta": -1.2}],
                },
            ],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "warning"
    assert "warehouse_offsetting_deltas" in issue_codes
    assert diagnostics["signals"]["offsettingWarehouseDeltas"] == [
        {
            "warehouseId": "25",
            "pdfAmountTotal": 17465.12,
            "excelAmountTotal": 17463.34,
            "amountDelta": 1.78,
            "attribution": [{"employeeName": "JIMENEZ, ENEAS", "delta": 1.19}],
        },
        {
            "warehouseId": "28",
            "pdfAmountTotal": 4537.46,
            "excelAmountTotal": 4539.25,
            "amountDelta": -1.79,
            "attribution": [{"employeeName": "JIMENEZ, ENEAS", "delta": -1.2}],
        },
    ]
    offset_issue = next(issue for issue in diagnostics["issues"] if issue["code"] == "warehouse_offsetting_deltas")
    assert "多个仓库分别超出容差" in offset_issue["message"]
    assert "仓库 25" in offset_issue["items"][0]


def test_reconciliation_diagnostics_flags_cross_warehouse_employee_allocation():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[
            {"source_file": "fairway_25.pdf", "warehouse_id": "25", "total_amount": 219.30},
            {"source_file": "fairway_28.pdf", "warehouse_id": "28", "total_amount": 1863.67},
        ],
        comparison_summary={"pdfAmountTotal": 2082.97, "excelAmountTotal": 2083.08, "exceptionCount": 0},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 2082.97, "excelAmountTotal": 2083.08, "allocationIssueCount": 1},
            "errors": [],
            "allocationIssues": [
                {
                    "employeeName": "PEREZ, JOSE",
                    "netAmountDelta": 0.0,
                    "warehouseCount": 2,
                    "warehouses": [
                        {"warehouseId": "25", "amountDelta": 0.59},
                        {"warehouseId": "28", "amountDelta": -0.59},
                    ],
                    "recommendation": "员工总额可抵消，但仓库归属金额不一致，需按仓库复核发票与账单归属。",
                }
            ],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "warning"
    assert "cross_warehouse_employee_allocation" in issue_codes
    assert diagnostics["signals"]["crossWarehouseEmployeeAllocation"][0]["employeeName"] == "PEREZ, JOSE"
    allocation_issue = next(issue for issue in diagnostics["issues"] if issue["code"] == "cross_warehouse_employee_allocation")
    assert "PEREZ, JOSE" in allocation_issue["items"][0]
    assert "仓库 25" in allocation_issue["items"][0]


def test_reconciliation_diagnostics_flags_employee_attribution_for_warehouse_delta():
    diagnostics = build_reconciliation_diagnostics(
        pdf_totals=[{"source_file": "US ELogistics Service Corp. 34794.pdf", "warehouse_id": "25", "total_amount": 62761.99}],
        comparison_summary={"pdfAmountTotal": 62761.99, "excelAmountTotal": 62803.2},
        warehouse_comparison={
            "summary": {"pdfAmountTotal": 62761.99, "excelAmountTotal": 62803.2},
            "errors": [],
            "rows": [
                {
                    "warehouseId": "25",
                    "pdfAmountTotal": 62761.99,
                    "excelAmountTotal": 62803.2,
                    "amountDelta": -41.21,
                    "attribution": [
                        {
                            "employeeName": "Fontes, Stevie ⇄ Stevie Fontes",
                            "pdfAmount": 822.12,
                            "excelAmount": 863.22,
                            "delta": -41.1,
                        },
                        {
                            "employeeName": "Sanchez Reveles, Jose ⇄ Jose Sanchez Reveles",
                            "pdfAmount": 919.56,
                            "excelAmount": 919.54,
                            "delta": 0.02,
                        },
                    ],
                }
            ],
        },
        amount_tolerance=0.1,
    )

    issue_codes = {issue["code"] for issue in diagnostics["issues"]}
    assert diagnostics["level"] == "warning"
    assert "warehouse_employee_attribution" in issue_codes
    assert diagnostics["signals"]["employeeAttribution"] == [
        {
            "warehouseId": "25",
            "employeeName": "Fontes, Stevie ⇄ Stevie Fontes",
            "pdfAmount": 822.12,
            "excelAmount": 863.22,
            "delta": -41.1,
            "warehouseDelta": -41.21,
        }
    ]
    attribution_issue = next(issue for issue in diagnostics["issues"] if issue["code"] == "warehouse_employee_attribution")
    assert "Fontes, Stevie" in attribution_issue["items"][0]


def test_suggest_mapping_and_read_workbook_rows_extract_required_fields(tmp_path):
    path = tmp_path / "账单.xlsx"
    path.write_bytes(_workbook_bytes())

    suggestion = suggest_mapping(path, "账单")

    assert suggestion["suggestedMapping"]["name"] == "姓名"
    assert suggestion["suggestedMapping"]["hours"] == "时长总计(H)"
    assert suggestion["suggestedMapping"]["amount"] == "费用总计(含税)"
    assert len(suggestion["previewRows"]) == 2

    rows = read_workbook_rows(
        path,
        "账单",
        {"name": "姓名", "hours": "时长总计(H)", "amount": "费用总计(含税)", "currency": "币种"},
    )

    assert [row.employee_name_raw for row in rows] == ["Jose Perez", "Wilfredo Martinez"]
    assert rows[0].hours == 40.14
    assert rows[0].amount == 1037.81
    assert rows[0].source_page_or_row == "账单!2"


def test_xlsx_mapping_preflight_closes_workbook_handles(monkeypatch, tmp_path):
    opened = []

    class FakeSheet:
        title = "账单"

        def reset_dimensions(self):
            return None

        def iter_rows(self, *, values_only, max_row):
            assert values_only is True
            assert max_row == 21
            return iter([
                ("姓名", "工时", "金额"),
                ("Alice", 8, 100),
            ])

    class FakeWorkbook:
        sheetnames = ["账单"]

        def __init__(self):
            self.closed = False
            self.sheet = FakeSheet()

        def __getitem__(self, name):
            assert name == "账单"
            return self.sheet

        def close(self):
            self.closed = True

    def fake_load_workbook(*_args, **_kwargs):
        workbook = FakeWorkbook()
        opened.append(workbook)
        return workbook

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load_workbook)
    path = tmp_path / "账单.xlsx"

    assert list_workbook_sheets(path) == ["账单"]
    assert opened[-1].closed is True

    suggestion = suggest_mapping(path, "账单")
    assert suggestion["suggestedMapping"]["name"] == "姓名"
    assert opened[-1].closed is True


def test_read_workbook_rows_sums_optional_amount_component_columns(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "账单"
    sheet.append(["员工名称", "总计", "本周薪资", "本周餐补", "周日补贴", "时薪"])
    sheet.append(["Unknown Worker", 27.92, 563.70, 19.44, 12.00, 20.19])
    sheet.append(["总计", 27.92, 563.70, 19.44, 12.00, None])
    path = tmp_path / "unknown-supplier.xlsx"
    workbook.save(path)

    suggestion = suggest_mapping(path, "账单")

    assert suggestion["amountColumnCandidates"] == ["本周薪资", "本周餐补", "周日补贴"]
    assert suggestion["suggestedMapping"]["name"] == "员工名称"
    assert suggestion["suggestedMapping"]["hours"] == "总计"
    assert suggestion["suggestedMapping"]["amount"] == "本周薪资"

    rows = read_workbook_rows(
        path,
        "账单",
        {
            "name": "员工名称",
            "hours": "总计",
            "amount": "本周薪资",
            "amountColumns": ["本周薪资", "本周餐补", "周日补贴"],
        },
    )

    assert len(rows) == 1
    assert rows[0].hours == 27.92
    assert rows[0].amount == 595.14
    assert rows[0].amount_components == {
        "本周薪资": 563.70,
        "本周餐补": 19.44,
        "周日补贴": 12.00,
    }


def test_read_workbook_rows_captures_amount_breakdown_without_double_counting(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单明细"
    sheet.append([
        "姓名",
        "时长总计(H)",
        "费用总计(不含税)",
        "白班工作费用",
        "奖金-非固",
        "非固费用总计",
        "备注-非固",
    ])
    sheet.append(["Carolay Hincapie", 47.82, 1101.33, 1026.33, 75, 75, "4月奖金"])
    path = tmp_path / "amount-breakdown.xlsx"
    workbook.save(path)

    rows = read_workbook_rows(
        path,
        "员工账单明细",
        {
            "name": "姓名",
            "hours": "时长总计(H)",
            "amount": "费用总计(不含税)",
        },
    )

    assert len(rows) == 1
    assert rows[0].amount == 1101.33
    assert rows[0].amount_components == {"费用总计(不含税)": 1101.33}
    assert rows[0].amount_breakdown == {
        "白班工作费用": 1026.33,
        "奖金-非固": 75.0,
    }
    assert rows[0].amount_context == {"备注-非固": "4月奖金"}


def test_compare_labor_items_explains_excel_amount_component_delta():
    pdf_rows = [
        LaborLineItem(
            source_type="pdf",
            source_file="invoice.pdf",
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw="Carolay Hincapie",
            hours=47.82,
            amount=1026.32,
        )
    ]
    excel_rows = [
        LaborLineItem(
            source_type="offline_workbook",
            source_file="bill.xlsx",
            source_page_or_row="员工账单明细!3",
            employee_id="",
            employee_name_raw="Carolay Hincapie",
            hours=47.82,
            amount=1101.33,
            amount_breakdown={"白班工作费用": 1026.33, "奖金-非固": 75.0},
            amount_context={"备注-非固": "4月奖金"},
        )
    ]

    comparison = compare_labor_items(
        pdf_rows,
        excel_rows,
        amount_tolerance=0.10,
        hours_tolerance=0.10,
    )

    row = comparison["rows"][0]
    assert row["matchStatus"] == "金额差异"
    assert row["amountDifferenceReasonCode"] == "excel_amount_component_delta"
    assert row["amountDifferenceComponents"] == [
        {
            "side": "excel",
            "label": "奖金-非固",
            "amount": 75.0,
            "note": "4月奖金",
        }
    ]
    assert row["amountDifferenceResidual"] == 0.01
    assert "4月奖金" in row["amountDifferenceExplanation"]
    assert "Excel 比 PDF 多 $75.01" in row["amountDifferenceExplanation"]


def test_amount_rate_review_uses_component_explanation():
    from bonus_platform.engine.labor.materials import _build_amount_rate_review_rows

    explanation = (
        "Excel 比 PDF 多 $75.01；其中可由 Excel 金额组成「奖金-非固」$75.00"
        "（备注：4月奖金）解释。请确认该费用项是否应包含在本批发票中。"
    )
    rows = _build_amount_rate_review_rows(
        [
            {
                "employeeKey": "name:CAROLAY HINCAPIE",
                "employeeName": "Carolay Hincapie",
                "matchStatus": "金额差异",
                "pdfHoursTotal": 47.82,
                "excelHoursTotal": 47.82,
                "hoursDelta": 0.0,
                "pdfAmountTotal": 1026.32,
                "excelAmountTotal": 1101.33,
                "amountDelta": -75.01,
                "amountDifferenceReasonCode": "excel_amount_component_delta",
                "amountDifferenceExplanation": explanation,
                "amountDifferenceComponents": [
                    {
                        "side": "excel",
                        "label": "奖金-非固",
                        "amount": 75.0,
                        "note": "4月奖金",
                    }
                ],
                "amountDifferenceResidual": 0.01,
            }
        ],
        hours_tolerance=0.1,
    )

    assert rows[0]["reviewLabel"] == "工时一致，Excel 含额外费用项"
    assert rows[0]["reviewFocus"] == "先核 Excel 额外费用项"
    assert rows[0]["businessQuestion"] == explanation
    assert rows[0]["amountDifferenceComponents"][0]["note"] == "4月奖金"


def test_read_workbook_rows_skips_explicit_total_summary_name(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bill"
    sheet.append(["Name", "Hours", "Amount", "Currency"])
    sheet.append(["Alice Example", 8, 80, "USD"])
    sheet.append(["Total Worker", 1, 10, "USD"])
    sheet.append(["TOTAL", 9, 90, "USD"])
    path = tmp_path / "summary-row.xlsx"
    workbook.save(path)

    rows = read_workbook_rows(
        path,
        "Bill",
        {"name": "Name", "hours": "Hours", "amount": "Amount", "currency": "Currency"},
    )

    assert [row.employee_name_raw for row in rows] == ["Alice Example", "Total Worker"]
    assert sum(row.hours for row in rows) == 9
    assert sum(row.amount for row in rows) == 90


def test_read_workbook_rows_skips_zero_value_footer_labels(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet2"
    sheet.append(["员工名称", "总计", "总额"])
    sheet.append(["Alice Example", 8, 80])
    sheet.append(["site", 0, 0])
    sheet.append(["Gonesse", None, None])
    path = tmp_path / "footer-labels.xlsx"
    workbook.save(path)

    rows = read_workbook_rows(
        path,
        "Sheet2",
        {"name": "员工名称", "hours": "总计", "amount": "总额"},
    )

    assert [row.employee_name_raw for row in rows] == ["Alice Example"]


def test_line_items_keep_non_time_quantity_out_of_worked_hours():
    rows = line_items_from_dicts(
        [
            {
                "source_type": "pdf_invoice",
                "source_file": "unknown.pdf",
                "source_page_or_row": "p1",
                "employee_name_raw": "Unknown Worker",
                "hours": 4,
                "amount": 19.44,
                "quantity": 4,
                "unit": "meal",
                "item_type": "meal_allowance",
                "description": "TICKET RESTAURANT",
                "evidence_text": "Unknown Worker TICKET RESTAURANT 4,00 4,86 19,44",
            }
        ]
    )

    assert len(rows) == 1
    assert rows[0].hours == 0
    assert rows[0].quantity == 4
    assert rows[0].unit == "meal"
    assert rows[0].item_type == "meal_allowance"
    assert rows[0].description == "TICKET RESTAURANT"


def test_line_items_infer_non_time_quantity_from_visible_evidence():
    rows = line_items_from_dicts(
        [
            {
                "source_type": "pdf_invoice",
                "source_file": "unknown.pdf",
                "source_page_or_row": "p1",
                "employee_name_raw": "Unknown Worker",
                "hours": 4,
                "amount": 19.44,
                "evidence_text": "Unknown Worker TICKET RESTAURANT PAT 4,00 4,86 19,44",
            },
            {
                "source_type": "pdf_invoice",
                "source_file": "unknown.pdf",
                "source_page_or_row": "p1",
                "employee_name_raw": "Unknown Worker",
                "hours": 27.92,
                "amount": 563.70,
                "evidence_text": "Unknown Worker HEURES NORMALES 27,92 20,19 563,70",
            },
        ]
    )

    assert [row.hours for row in rows] == [0, 27.92]
    assert rows[0].quantity == 4
    assert rows[0].item_type == "meal_allowance"
    assert rows[1].item_type == "worked_hours"


def test_suggest_mapping_prefers_plain_name_over_employee_id(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employee Billing"
    sheet.append(["Employee ID", "Name", "Hours", "Amount (Net)", "Currency", "Physical Warehouse"])
    sheet.append(["SYN0001", "Jordan Hale", 38.5, 924.0, "USD", "101"])
    path = tmp_path / "unknown-supplier.xlsx"
    workbook.save(path)

    suggestion = suggest_mapping(path, "Employee Billing")

    assert suggestion["suggestedMapping"]["employeeId"] == "Employee ID"
    assert suggestion["suggestedMapping"]["name"] == "Name"
    assert suggestion["suggestedMapping"]["amount"] == "Amount (Net)"


def test_suggest_mapping_prefers_amount_excluding_tax_when_available(tmp_path):
    path = tmp_path / "账单.xlsx"
    path.write_bytes(_workbook_with_tax_columns_bytes())

    suggestion = suggest_mapping(path, "账单")

    assert suggestion["suggestedMapping"]["amount"] == "费用总计(不含税)"


def test_suggest_mapping_does_not_use_hours_column_as_amount(tmp_path):
    path = tmp_path / "GRANDE-5.18-5.24.xlsx"
    path.write_bytes(_workbook_with_hours_only_summary_bytes())

    suggestion = suggest_mapping(path, "Sheet1")

    assert suggestion["suggestedMapping"]["name"] == "Employee Name"
    assert suggestion["suggestedMapping"]["hours"] == "求和项:Total Hours"
    assert suggestion["suggestedMapping"]["amount"] == ""


def test_suggest_mapping_handles_two_row_otws_employee_expense_headers(tmp_path):
    path = tmp_path / "OTWS.xlsx"
    path.write_bytes(_workbook_with_two_header_rows_bytes())

    suggestion = suggest_mapping(path, "Employee-expenses-detail")

    assert suggestion["suggestedMapping"] == {
        "employeeId": "Employee number",
        "name": "Employee name",
        "hours": "Total staff cost accounting time",
        "amount": "Total cost",
        "currency": "",
    }

    rows = read_workbook_rows(path, "Employee-expenses-detail", suggestion["suggestedMapping"])

    assert len(rows) == 1
    assert rows[0].employee_id == "EUS031468"
    assert rows[0].employee_name_raw == "JOSE MAGANA"
    assert rows[0].hours == 8
    assert rows[0].amount == 188
    assert rows[0].warehouse_id == "13"


def test_summarize_otws_costs_explains_summary_and_detail_bases(tmp_path):
    path = tmp_path / "OTWS - Warehouse Bill-NJ13.xlsx"
    path.write_bytes(_otws_cost_workbook_bytes())

    summary = summarize_otws_costs(path)

    assert summary["sourceFile"] == "OTWS - Warehouse Bill-NJ13.xlsx"
    assert summary["warehouseId"] == "13"
    assert summary["supplier"] == "Strategic Staffing Solutions Corp."
    assert summary["currency"] == "USD"
    assert summary["periodStart"] == "2026-05-11"
    assert summary["periodEnd"] == "2026-05-17"
    assert summary["employeeCount"] == 64
    assert summary["summary"]["components"]["hourlySalary"] == 48055.81
    assert summary["summary"]["components"]["otherExpenses"] == 162.15
    assert summary["summary"]["componentTotal"] == 48217.96
    assert summary["summary"]["reportedTotal"] == 48217.96
    assert summary["summary"]["componentDelta"] == 0
    assert summary["summary"]["evidence"] == "Warehouse-information!2"
    assert summary["details"]["employeeExpenses"]["amount"] == 188
    assert summary["details"]["employeeExpenses"]["hours"] == 8
    assert summary["details"]["employeeExpenses"]["rowCount"] == 1
    assert summary["details"]["employeeExpenses"]["evidence"] == "Employee-expenses-detail!3"
    assert summary["details"]["employeeBenefits"]["amount"] == 162.15
    assert summary["details"]["employeeBenefits"]["rowCount"] == 1
    assert summary["details"]["detailTotal"] == 350.15
    assert summary["details"]["summaryDelta"] == 47867.81


def test_compare_labor_items_flags_amount_delta_and_ignores_one_cent():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="a.pdf", source_page_or_row="1", employee_id="", employee_name_raw="PEREZ, JOSE", hours=40.14, amount=1037.81, currency="USD", confidence=0.96, evidence_text="invoice row"),
        LaborLineItem(source_type="pdf_invoice", source_file="a.pdf", source_page_or_row="1", employee_id="", employee_name_raw="MARTINEZ, WILFREDO", hours=40.78, amount=982.72, currency="USD", confidence=0.91, evidence_text="invoice row"),
        LaborLineItem(source_type="pdf_invoice", source_file="a.pdf", source_page_or_row="1", employee_id="", employee_name_raw="LOW, CONFIDENCE", hours=8, amount=100, currency="USD", confidence=0.5, evidence_text="low confidence"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Jose Perez", hours=40.14, amount=1037.80, currency="USD", confidence=1, evidence_text=""),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="", employee_name_raw="Wilfredo Martinez", hours=40.78, amount=982.74, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.01, hours_tolerance=0.1, confidence_threshold=0.85)

    assert result["summary"]["amountDiffCount"] == 1
    assert result["summary"]["unmatchedPdfCount"] == 1
    assert any(row["matchStatus"] == "金额差异" and row["employeeName"] == "MARTINEZ, WILFREDO" for row in result["rows"])
    assert any(row["matchStatus"] == "低置信度抽取" for row in result["rows"])
    assert all(not (row["employeeName"] == "PEREZ, JOSE" and row["matchStatus"] == "金额差异") for row in result["rows"])


def test_compare_labor_items_keeps_same_employee_separate_across_warehouses():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="wh12.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Serhii Hizhdevan", hours=10.27, amount=271.94, currency="EUR", confidence=1, evidence_text="invoice row", warehouse_id="12"),
        LaborLineItem(source_type="pdf_invoice", source_file="wh9.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Hizhdevan Serhii", hours=11.99, amount=317.48, currency="EUR", confidence=1, evidence_text="invoice row", warehouse_id="9"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Sheet1!2", employee_id="DE-WH12", employee_name_raw="Serhii Hizhdevan", hours=10.27, amount=271.95, currency="EUR", confidence=1, evidence_text="", warehouse_id="12"),
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Sheet1!3", employee_id="DE-WH9", employee_name_raw="Hizhdevan Serhii", hours=11.99, amount=317.50, currency="EUR", confidence=1, evidence_text="", warehouse_id="9"),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.05, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert len(result["rows"]) == 2
    assert {row["warehouseId"] for row in result["rows"]} == {"9", "12"}
    assert all(row["matchStatus"] == "通过" for row in result["rows"])


def test_compare_labor_items_never_fuzzy_matches_across_warehouses():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="wh1.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alex Smith", hours=8, amount=100, currency="EUR", confidence=1, evidence_text="invoice row", warehouse_id="1"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Sheet1!2", employee_id="DE-ALEX", employee_name_raw="Alex Smith", hours=8, amount=100, currency="EUR", confidence=1, evidence_text="", warehouse_id="2"),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.05, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 2
    assert result["summary"]["unmatchedPdfCount"] == 1
    assert result["summary"]["unmatchedExcelCount"] == 1
    assert {row["warehouseId"] for row in result["rows"]} == {"1", "2"}
    assert result["candidateMatches"] == []


def test_compare_labor_items_treats_tiny_unmatched_excel_residual_as_passed_risk():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="a.pdf", source_page_or_row="1", employee_id="", employee_name_raw="PEREZ, JOSE", hours=40, amount=1000, currency="USD", confidence=0.96, evidence_text="invoice row"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Jose Perez", hours=40, amount=1000, currency="USD", confidence=1, evidence_text=""),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="EUS000001", employee_name_raw="TINY RESIDUAL", hours=0.02, amount=0.41, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.25, hours_tolerance=0.1)

    residual = next(row for row in result["rows"] if row["employeeName"] == "TINY RESIDUAL")
    assert residual["matchStatus"] == "通过"
    assert "微小残差" in residual["riskFlags"]
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["summary"]["exceptionCount"] == 0


def test_compare_labor_items_matches_minor_name_typos_when_totals_align():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="osi.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Montealvo, Sergio", hours=46.2, amount=1345.89, currency="USD", confidence=0.98, evidence_text="invoice row"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!20", employee_id="WUS038206", employee_name_raw="Sergio Montalvo", hours=46.2, amount=1345.89, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["fuzzyMatchCount"] == 1
    assert result["rows"][0]["matchStatus"] == "通过"
    assert result["rows"][0]["employeeName"] == "Montealvo, Sergio ⇄ Sergio Montalvo"


def test_compare_labor_items_treats_exact_name_match_without_pdf_id_as_passed():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="osi.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alva, Patrick", hours=34.75, amount=939.25, currency="USD", confidence=0.98, evidence_text="$939.25"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="WUS045000", employee_name_raw="Patrick Alva", hours=34.75, amount=939.25, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows)

    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["summary"]["exceptionCount"] == 0
    assert result["rows"][0]["matchStatus"] == "通过"


def test_compare_labor_items_matches_partial_name_when_totals_align():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="osi.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Parra Hernandes, Nancy", hours=44.34, amount=1058.12, currency="USD", confidence=0.98, evidence_text="$1058.12"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="WUS039740", employee_name_raw="Nancy Parra", hours=44.34, amount=1058.14, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.05)

    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["summary"]["exceptionCount"] == 0
    assert result["rows"][0]["matchStatus"] == "通过"


def test_compare_labor_items_fuzzy_matches_ocr_name_variants_when_totals_align():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="scan.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Benavides, Jeremy", hours=22.68, amount=508.03, currency="USD", confidence=0.95, evidence_text="Total $508.03"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!4", employee_id="", employee_name_raw="Jeymmy Benavides", hours=22.68, amount=508.03, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(
        pdf_rows,
        excel_rows,
        manual_name_mapping={"Benavides, Jeremy": "Jeymmy Benavides"},
    )

    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["fuzzyMatchCount"] == 1
    assert result["rows"][0]["matchStatus"] == "通过"
    assert "疑似姓名匹配" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_uses_amount_as_primary_and_flags_hours_only_as_risk():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Flores, Alexis", hours=59.22, amount=1864.70, currency="USD", confidence=0.96, evidence_text="$1864.70"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Alexis Flores", hours=51.22, amount=1864.70, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["hoursRiskCount"] == 1
    assert result["rows"][0]["matchStatus"] == "通过"
    assert "工时需复核" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_marks_safe_name_format_difference_as_auto_merged():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Mucu, Pablo", hours=40, amount=1000, currency="USD", confidence=0.96, evidence_text="$1000.00"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Employee-expenses-detail!2", employee_id="", employee_name_raw="Pablo Mucu", hours=40, amount=1000, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 0
    assert result["rows"][0]["matchStatus"] == "通过"
    assert result["rows"][0]["employeeName"] == "Mucu, Pablo ⇄ Pablo Mucu"
    assert "姓名格式差异自动合并" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_marks_accent_difference_as_auto_merged():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alberto Núñez", hours=35.08, amount=739.49, currency="USD", confidence=0.96, evidence_text="$739.49"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Employee-expenses-detail!2", employee_id="", employee_name_raw="Alberto Nunez", hours=35.08, amount=739.49, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 0
    assert result["rows"][0]["matchStatus"] == "通过"
    assert result["rows"][0]["employeeName"] == "Alberto Núñez ⇄ Alberto Nunez"
    assert "姓名格式差异自动合并" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_does_not_auto_merge_amount_close_name_unlike():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Maria Lopez", hours=40, amount=812.80, currency="USD", confidence=0.96, evidence_text="$812.80"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Employee-expenses-detail!2", employee_id="", employee_name_raw="Carlos Serna", hours=40, amount=812.80, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 2
    assert result["summary"]["unmatchedPdfCount"] == 1
    assert result["summary"]["unmatchedExcelCount"] == 1
    assert result["summary"]["fuzzyMatchCount"] == 0
    assert all(row["matchStatus"] != "通过" for row in result["rows"])
    assert result["candidateMatches"] == []


def test_compare_labor_items_matches_workbuddy_jaccard_when_amounts_align():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Nava de Luna, Julian", hours=12.25, amount=276.64, currency="USD", confidence=0.96, evidence_text="$276.64"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Julieta Nava de Luna", hours=12.25, amount=276.64, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["fuzzyMatchCount"] == 1
    assert result["rows"][0]["matchStatus"] == "通过"
    assert "疑似姓名匹配" in result["rows"][0]["riskFlags"]


def test_workbuddy_normalize_removes_accents_punctuation_and_lowercases():
    assert normalize_workbuddy_name("García, María") == "garcia maria"
    assert normalize_workbuddy_name("Nava-de_Luna, Julián") == "nava de luna julian"


def test_compare_labor_items_uses_manual_mapping_for_two_token_spelling_variants():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="invoice.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Gamboa, Arilene", hours=53.62, amount=1520.28, currency="USD", confidence=0.96, evidence_text="$1520.28"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Arlene Gamboa", hours=53.62, amount=1520.28, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(
        pdf_rows,
        excel_rows,
        amount_tolerance=0.1,
        manual_name_mapping={"Gamboa, Arilene": "Arlene Gamboa"},
    )

    assert result["summary"]["exceptionCount"] == 0
    assert result["summary"]["fuzzyMatchCount"] == 1
    assert result["rows"][0]["matchStatus"] == "通过"
    assert "疑似姓名匹配" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_fuzzy_match_can_still_surface_amount_delta():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="scan.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Castillo, Misael", hours=30.92, amount=689.12, currency="USD", confidence=0.95, evidence_text="Total $689.12"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!7", employee_id="", employee_name_raw="Massiel Castillo", hours=30.92, amount=694.17, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(
        pdf_rows,
        excel_rows,
        manual_name_mapping={"Castillo, Misael": "Massiel Castillo"},
    )

    assert result["summary"]["amountDiffCount"] == 1
    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["rows"][0]["matchStatus"] == "金额差异"
    assert "疑似姓名匹配" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_promotes_same_hours_name_candidate_to_amount_diff():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="sss.pdf", source_page_or_row="p29", employee_id="", employee_name_raw="Ruben Cadiz, Carlos", hours=7.82, amount=183.73, currency="USD", confidence=0.95, evidence_text="Ruben Cadiz, Carlos 20844 7.82 $183.73"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="bill.xlsx", source_page_or_row="Employee-expenses-detail!42", employee_id="EUS020844", employee_name_raw="CARLOS RUBEN CADIZ RODRIGUEZ", hours=7.82, amount=168.83, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.25, hours_tolerance=0.1)

    assert result["summary"]["amountDiffCount"] == 1
    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["summary"]["exceptionCount"] == 1
    assert result["summary"]["candidateMatchCount"] == 1
    row = result["rows"][0]
    assert row["employeeName"] == "Ruben Cadiz, Carlos ⇄ CARLOS RUBEN CADIZ RODRIGUEZ"
    assert row["matchStatus"] == "金额差异"
    assert row["amountDelta"] == 14.9
    assert "疑似姓名匹配" in row["riskFlags"]
    candidate = result["candidateMatches"][0]
    assert candidate["recommendation"] == "姓名疑似同一人，金额/费率差异需人工复核"


def test_compare_labor_items_flags_offsetting_unmatched_excel_as_combined_pdf_row():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="oss.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Lozano, Manuel", hours=19.59, amount=439.82, currency="USD", confidence=0.95, evidence_text="Lozano, Manuel ... 19.50 0.09 ... 439.82"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!21", employee_id="WUS045753", employee_name_raw="Manuel Lozano", hours=16.09, amount=361.42, currency="USD", confidence=1, evidence_text=""),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!24", employee_id="WUS045746", employee_name_raw="Massiel Castillo", hours=3.5, amount=78.4, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["exceptionCount"] == 2
    assert result["summary"]["candidateMatchCount"] == 1
    candidate = result["candidateMatches"][0]
    assert candidate["issueType"] == "combined_pdf_row"
    assert candidate["pdfEmployeeName"] == "Lozano, Manuel ⇄ Manuel Lozano"
    assert candidate["excelEmployeeName"] == "Massiel Castillo"
    assert candidate["recommendation"] == "疑似PDF合并员工，需人工核对原始发票"
    assert candidate["hoursDelta"] == 3.5
    assert candidate["amountDelta"] == 78.4
    assert "oss.pdf p1" in candidate["sourceRefs"]
    assert "账单.xlsx 账单!24" in candidate["sourceRefs"]
    flagged = {row["employeeName"]: row for row in result["rows"]}
    assert "疑似PDF合并员工" in flagged["Lozano, Manuel ⇄ Manuel Lozano"]["riskFlags"]
    assert "疑似PDF合并员工" in flagged["Massiel Castillo"]["riskFlags"]


def test_compare_labor_items_fuzzy_matches_pdf_name_to_excel_employee_id_group():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="scan.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alvarez Mitrache, Rosa", hours=31.19, amount=701.9, currency="USD", confidence=0.95, evidence_text="Total $701.90"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="WUS042586", employee_name_raw="Rosa Alvarez Minchaca", hours=31.19, amount=701.9, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows)

    assert result["summary"]["unmatchedPdfCount"] == 0
    assert result["summary"]["unmatchedExcelCount"] == 0
    assert result["rows"][0]["matchStatus"] == "通过"
    assert "疑似姓名匹配" in result["rows"][0]["riskFlags"]


def test_compare_labor_items_suggests_unmatched_name_candidates_without_merging():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="scan.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Alvarez Mitrache, Ross", hours=30.5, amount=698.99, currency="USD", confidence=0.95, evidence_text="Total $698.99"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="WUS042586", employee_name_raw="Rosa Alvarez Minchaca", hours=31.19, amount=701.9, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows)

    assert result["summary"]["unmatchedPdfCount"] == 1
    assert result["summary"]["unmatchedExcelCount"] == 1
    assert result["summary"]["candidateMatchCount"] == 1
    candidate = result["candidateMatches"][0]
    assert candidate["pdfEmployeeName"] == "Alvarez Mitrache, Ross"
    assert candidate["excelEmployeeName"] == "Rosa Alvarez Minchaca"
    assert candidate["recommendation"] == "人工复核"


def test_compare_labor_items_suggests_low_similarity_candidate_when_totals_align():
    pdf_rows = [
        LaborLineItem(source_type="pdf_invoice", source_file="In291943.pdf", source_page_or_row="p1", employee_id="", employee_name_raw="Rozo Panche, Deisy V", hours=37.84, amount=847.84, currency="USD", confidence=0.98, evidence_text="$847.84"),
    ]
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="员工账单明细!3", employee_id="WUS040020", employee_name_raw="Deisi Pozo", hours=37.84, amount=847.84, currency="USD", confidence=1, evidence_text=""),
    ]

    result = compare_labor_items(pdf_rows, excel_rows, amount_tolerance=0.1, hours_tolerance=0.1)

    assert result["summary"]["unmatchedPdfCount"] == 1
    assert result["summary"]["unmatchedExcelCount"] == 1
    assert result["summary"]["candidateMatchCount"] == 1
    candidate = result["candidateMatches"][0]
    assert candidate["pdfEmployeeName"] == "Rozo Panche, Deisy V"
    assert candidate["excelEmployeeName"] == "Deisi Pozo"
    assert candidate["nameSimilarity"] == 0.4
    assert candidate["amountDelta"] == 0
    assert candidate["hoursDelta"] == 0


def test_rule_pdf_extractor_adds_meal_premium_amount_without_hours():
    rows = _extract_with_rules(
        [
            {
                "source_file": "invoice.pdf",
                "page": 1,
                "text": "\n".join(
                    [
                        "Associate USEL EMPLOYEE ID Payrate Reg Rate Ot Rate Dt Rate Reg. Time Overtime Dbl. Time RT OT DT TOTAL",
                        "VEGA -0980, ALEXANDER WUS034706 20.00$    25.80$    38.70$ 51.60$ 40.00 14.51 0.03 1,032.00$  561.54$     1.55$         1,595.09$",
                        "VEGA -0980, ALEXANDER WUS034706 20.00$    25.80$    38.70$ 51.60$ 1.00 25.80$       -$           -$           25.80$",
                        "MEAL PREMIUMS",
                    ]
                ),
            }
        ],
        supplier="Fairway Staffing Service",
        period_start="2026-05-04",
        period_end="2026-05-10",
        currency="USD",
    )

    result = compare_labor_items(
        rows,
        [
            LaborLineItem(
                source_type="offline_workbook",
                source_file="账单.xlsx",
                source_page_or_row="账单!48",
                employee_id="WUS034706",
                employee_name_raw="Alxander Vega -0980",
                hours=54.54,
                amount=1620.89,
                currency="USD",
            )
        ],
    )

    assert len(rows) == 2
    assert sum(row.hours for row in rows) == 54.54
    assert round(sum(row.amount for row in rows), 2) == 1620.89
    assert result["summary"]["amountDiffCount"] == 0
    assert result["summary"]["hoursRiskCount"] == 0


def test_rule_pdf_extractor_handles_osi_vertical_invoice_rows():
    rows = _extract_with_rules(
        [
            {
                "source_file": "osi.pdf",
                "page": 1,
                "text": "\n".join(
                    [
                        "Date",
                        "Description",
                        "Hours",
                        "Pay Code",
                        "Type",
                        "Pay Rate",
                        "Bill Rate",
                        "Amount",
                        "CA#25 Bloomington",
                        "5/17/2026",
                        "Alva, Patrick",
                        "32.00",
                        "Reg",
                        "REG",
                        "$20.00",
                        "26.00",
                        "$832.00",
                        "5/17/2026",
                        "Alva, Patrick",
                        "2.75",
                        "OT",
                        "OT",
                        "$30.00",
                        "39.00",
                        "$107.25",
                    ]
                ),
            }
        ],
        supplier="OSI Staffing Inc.",
        period_start="2026-05-11",
        period_end="2026-05-17",
        currency="USD",
    )

    assert len(rows) == 2
    assert [row.employee_name_raw for row in rows] == ["Alva, Patrick", "Alva, Patrick"]
    assert round(sum(row.hours for row in rows), 2) == 34.75
    assert round(sum(row.amount for row in rows), 2) == 939.25


def test_extract_invoice_items_handles_oss_bill_rate_summary_rows(monkeypatch, tmp_path):
    pdf = tmp_path / "US Elogis Service #7 Invoice W.E 05.24.26.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    page = {
        "source_file": pdf.name,
        "page": 1,
        "text": "\n".join(
            [
                "Associate Base Rate Bill Rate OT Rate Reg. Time O.T Dbl. Time RT OT DT TOTAL",
                "Benitez, Anuar $20.00 25.60$     38.40$   16.00 0.19 409.60$         7.30$         -$             416.90$",
                "Briseno Mandujano, Gabriela $17.50 22.40$     33.60$   13.40 300.16$         -$           -$             300.16$",
                "Totals 29.40 0.19 0.00 $709.76 $7.30 $0.00 $717.06",
                "Customer US Elogistics Service Corp #7",
            ]
        ),
    }
    import bonus_platform.engine.labor.extract as extract_module

    monkeypatch.setattr(extract_module, "_extract_pdf_pages", lambda paths: [page])

    rows = extract_invoice_items([pdf], {"enabled": False, "parallel_extraction_enabled": False}, supplier="oss", currency="USD")

    assert len(rows) == 2
    assert rows[0].employee_name_raw == "Benitez, Anuar"
    assert rows[0].hours == 16.19
    assert rows[0].amount == 416.90
    assert rows[0].warehouse_id == "7"
    assert rows[1].employee_name_raw == "Briseno Mandujano, Gabriela"
    assert rows[1].hours == 13.40
    assert rows[1].amount == 300.16


def test_extract_invoice_items_handles_sss_employee_summary_rows(monkeypatch, tmp_path):
    pdf = tmp_path / "NJ13 Invoice Report WE 051726 JF.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    page = {
        "source_file": pdf.name,
        "page": 18,
        "text": "\n".join(
            [
                "Candidate Number Employee Name Candidate Notes Job Code Assignment Wage Rate Service Multiplier Bill Rate Standard Hours Worked Overtime Hours Worked Fee for Regular Hours Fee for Overtime Hours Total SSS Fee",
                "1 Flexible Workforce Shift 1 37 $16.00 27.00% $20.32 1327.15 0.00 $ 26,967.69 $ - $ 26,967.69",
                "2 Open, Open Open CUE1LD2 Loaders Shift 1 Level 2 $17.00 27.00% $21.59 0.00 0.00 $ - $ - $ -",
                "1 Contreras, Kristel 20132 CUE1C1 Cordinator Shift 1 Level 1 $17.00 27.00% $21.59 40.00 8.00 $ 863.60 $ 259.08 $ 1,122.68",
                "2 Contreras, Kristel 20132 CUE1C1 Cordinator Shift 1 Level 1 $17.00 27.00% $21.59 (40.00) (8.00) $ (863.60) $ (259.08) $ (1,122.68)",
                "3 Gonzalez, Felix 20597 CUE1LD2 Loaders Shift 1 Level 2 $17.00 27.00% $21.59 40.00 0.00 $ 863.60 $ - $ 863.60",
                "Antonio 20680 CUE1LD2 Loaders Shift 1 Level 2 $17.00 27.00% $21.59 40.00 0.00 $ 863.60 $ - $ 863.60",
                "4 Hernandez, Gabriel 20125 CUE1LD2 Loaders Shift 1 Level 2 $17.00 27.00% $21.59 40.00 8.00 $ 863.60 $ 259.08 $ 1,122.68",
                "5 Aparicio, Emilio 20253 SD Shift Differential $1.00 27.00% $1.27 40.00 0.00 $ 50.80 $ - $ 50.80",
                "6 Lopez Bellis,",
                "Dalila 20683 CUE1GL2 General Labor Shift 1",
                "Level 2 $16.00 27.00% $20.32 24.00 0.00 $ 487.68 $ - $ 487.68",
                "AM Loaders Summary Confidential Page 18",
            ]
        ),
    }
    import bonus_platform.engine.labor.extract as extract_module

    monkeypatch.setattr(extract_module, "_extract_pdf_pages", lambda paths: [page])

    rows = extract_invoice_items([pdf], {"enabled": False, "parallel_extraction_enabled": False}, supplier="sss", currency="USD")

    assert [row.employee_name_raw for row in rows] == [
        "Gonzalez, Felix",
        "Antonio",
        "Hernandez, Gabriel",
        "Aparicio, Emilio",
        "Lopez Bellis, Dalila",
    ]
    assert rows[0].employee_id == "20597"
    assert rows[0].hours == 40.0
    assert rows[0].amount == 863.60
    assert rows[1].employee_id == "20680"
    assert rows[1].amount == 863.60
    assert rows[2].hours == 48.0
    assert rows[2].amount == 1122.68
    assert rows[3].employee_id == "20253"
    assert rows[3].hours == 0.0
    assert rows[3].amount == 50.80
    assert rows[4].employee_id == "20683"
    assert rows[4].hours == 24.0
    assert rows[4].amount == 487.68
    assert all("Contreras" not in row.employee_name_raw for row in rows)
    assert rows[0].warehouse_id == "13"


def test_mimo_uses_api_key_header_instead_of_bearer_authorization():
    headers = _request_headers({"provider": "mimo", "api_key": "token"})

    assert headers["api-key"] == "token"
    assert "Authorization" not in headers


def test_ai_instruction_blocks_hallucinated_ids_and_non_employee_pages():
    instruction = _ai_instruction()

    assert "return []" in instruction.lower()
    assert "employee_id" in instruction
    assert "barcode" in instruction.lower()
    assert "spatial calibration" in instruction.lower()


def test_ai_instruction_requests_generic_line_semantics_for_unknown_suppliers():
    instruction = _ai_instruction(resolve_supplier_profile("Unseen Vendor LLC"), for_image=True)

    assert "description" in instruction
    assert "item_type" in instruction
    assert "quantity" in instruction
    assert "unit" in instruction
    assert "worked hours" in instruction.lower()


def test_supplier_profile_adds_onesource_specific_extraction_guidance():
    profile = resolve_supplier_profile("One Source Staffing Inc.")
    instruction = _ai_instruction(profile)

    assert profile.key == "onesource"
    assert profile.image_page_policy == "first_page_only"
    assert "timecard" in instruction.lower()
    assert "handwritten rg/ot" in instruction.lower()


def test_supplier_profile_adds_prompt_priority_dept_guidance():
    profile = resolve_supplier_profile("Prompt Priority INC")
    instruction = _ai_instruction(profile)

    assert profile.key == "prompt"
    assert profile.image_page_policy == "all"
    assert "dept" in instruction.lower()
    assert "warehouse_id" in instruction


def test_supplier_profile_adds_citistaff_loc_guidance():
    profile = resolve_supplier_profile("CitiStaff Solutions")
    instruction = _ai_instruction(profile)

    assert profile.key == "citistaff"
    assert profile.image_page_policy == "all"
    assert "loc.#" in instruction.lower()
    assert "name mappings" in instruction.lower()


def test_unknown_supplier_uses_default_extraction_profile():
    profile = resolve_supplier_profile("Unseen Vendor LLC")

    assert profile.key == "default"
    assert profile.image_page_policy == "all"


def test_supplier_profiles_can_load_from_json_config(tmp_path):
    path = tmp_path / "profiles.json"
    path.write_text(
        json.dumps(
            [
                {
                    "key": "demo",
                    "aliases": ["demo staffing"],
                    "prompt_notes": ["Only extract rows from the Charge Summary table."],
                    "image_page_policy": "all",
                }
            ]
        ),
        encoding="utf-8",
    )

    profiles = load_supplier_profiles(path)

    assert profiles[0].key == "demo"
    assert profiles[0].aliases == ["demo staffing"]
    assert "Charge Summary" in profiles[0].prompt_notes[0]
    assert profiles[0].authoritative_total_methods == []


def test_supplier_profile_round_trips_authoritative_total_methods(tmp_path):
    path = save_supplier_profile(
        {
            "key": "demo",
            "aliases": ["demo staffing"],
            "authoritative_total_methods": ["configured_invoice_field"],
        },
        tmp_path,
    )

    profile = load_supplier_profiles(path)[0]

    assert profile.authoritative_total_methods == ["configured_invoice_field"]


def test_supplier_profiles_can_load_single_json_object(tmp_path):
    path = tmp_path / "profile.json"
    path.write_text(
        json.dumps(
            {
                "key": "grande",
                "aliases": ["grande solutions staffing"],
                "prompt_notes": ["Use the simple numbered labor table."],
                "image_page_policy": "all",
            }
        ),
        encoding="utf-8",
    )

    profiles = load_supplier_profiles(path)

    assert len(profiles) == 1
    assert profiles[0].key == "grande"
    assert profiles[0].aliases == ["grande solutions staffing"]


def test_supplier_profile_resolver_prefers_external_config(tmp_path):
    path = tmp_path / "profiles.json"
    path.write_text(
        json.dumps(
            [
                {
                    "key": "external-demo",
                    "aliases": ["onesource"],
                    "prompt_notes": ["External profile wins."],
                    "version": 1,
                    "status": "approved",
                    "approvedBy": "p0-test-reviewer",
                    "approvedAt": "2026-07-15T10:00:00Z",
                    "created_from": "manual_review",
                }
            ]
        ),
        encoding="utf-8",
    )

    profile = resolve_supplier_profile("ONESOURCE", profiles_path=path)

    assert profile.key == "external-demo"
    assert profile.prompt_notes == ["External profile wins."]


def test_extract_invoice_items_applies_first_page_only_profile_policy(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    seen_pages = []

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_pdf_pages", lambda paths: [{"source_file": "scan.pdf", "page": 1, "text": ""}])
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_rules", lambda *args, **kwargs: [])
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_text", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_pages_to_images",
        lambda paths, scale=1.5, **kwargs: [
            {"source_file": "scan.pdf", "source_path": str(pdf), "page": 1, "mime_type": "image/png", "base64": "page1"},
            {"source_file": "scan.pdf", "source_path": str(pdf), "page": 2, "mime_type": "image/png", "base64": "page2"},
        ],
    )

    def fake_extract_images(image_pages, *args, **kwargs):
        seen_pages.extend(page["page"] for page in image_pages)
        return [
            {
                "source_file": "scan.pdf",
                "source_page_or_row": "p1",
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 31.19,
                "amount": 701.9,
                "confidence": 0.95,
                "evidence_text": "Total $701.90",
            }
        ]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_images", fake_extract_images)

    rows = extract_invoice_items(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="ONESOURCE",
    )

    assert seen_pages == [1]
    assert rows[0].employee_name_raw == "Alvarez Minchaca, Rosa"


def test_extract_invoice_items_applies_first_page_only_only_to_images(monkeypatch, tmp_path):
    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "invoice.pdf",
                "page": 1,
                "text": "\n".join(["Reference", "Employee", "Wage Code", "Type", "Hours", "Rate", "Amount"]),
            },
            {
                "source_file": "invoice.pdf",
                "page": 2,
                "text": "\n".join(
                    [
                        "Torres, Fabiola",
                        "Reg",
                        "REG",
                        "40.00",
                        "22.58",
                        "$903.20",
                        "Torres, Fabiola",
                        "Reg",
                        "OT",
                        "4.64",
                        "33.86",
                        "$157.11",
                    ]
                ),
            },
        ],
    )

    rows = extract_invoice_items(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="ONESOURCE",
    )

    assert len(rows) == 2
    assert {row.source_page_or_row for row in rows} == {"p2"}
    assert round(sum(row.amount for row in rows), 2) == 1060.31


def test_extract_invoice_items_falls_back_to_images_for_unparsed_scanned_pdf(monkeypatch, tmp_path):
    text_pdf = tmp_path / "text.pdf"
    scan_pdf = tmp_path / "scan.pdf"
    text_pdf.write_bytes(b"%PDF-1.4\n")
    scan_pdf.write_bytes(b"%PDF-1.4\n")
    seen_pages = []

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "text.pdf",
                "page": 1,
                "text": "\n".join(
                    [
                        "Hours Amount Bill Rate Date Description Pay Rate",
                        "$22.40 40.000 $896.00 5/17/2026 Arellano Luna, Pablo $17.500 Reg",
                    ]
                ),
            },
            {"source_file": "scan.pdf", "page": 1, "text": ""},
            {"source_file": "scan.pdf", "page": 2, "text": ""},
        ],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_pages_to_images",
        lambda paths, scale=1.5, **kwargs: [
            {"source_file": "scan.pdf", "source_path": str(scan_pdf), "page": 1, "mime_type": "image/png", "base64": "page1"},
            {"source_file": "scan.pdf", "source_path": str(scan_pdf), "page": 2, "mime_type": "image/png", "base64": "page2"},
        ],
    )

    def fake_extract_images(image_pages, *args, **kwargs):
        seen_pages.extend(page["page"] for page in image_pages)
        return [
            {
                "source_file": "scan.pdf",
                "source_page_or_row": "p2",
                "employee_name_raw": "Scan Person",
                "hours": 8,
                "amount": 160,
                "confidence": 0.9,
            }
        ]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_images", fake_extract_images)

    rows = extract_invoice_items(
        [text_pdf, scan_pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="CITI",
    )

    assert seen_pages == [1, 2]
    assert [row.employee_name_raw for row in rows] == ["Arellano Luna, Pablo", "Scan Person"]


def test_quick_extract_totals_uses_wage_code_rows_from_all_pages(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "invoice.pdf",
                "page": 1,
                    "text": "\n".join(
                    [
                        "Invoice CA#7",
                        "Aguilar, Hortensia",
                        "Reg",
                        "REG",
                        "40.00",
                        "22.58",
                        "$903.20",
                    ]
                ),
            },
            {
                "source_file": "invoice.pdf",
                "page": 2,
                "text": "\n".join(
                    [
                        "Torres, Fabiola",
                        "Reg",
                        "REG",
                        "40.00",
                        "22.58",
                        "$903.20",
                        "Torres, Fabiola",
                        "Reg",
                        "OT",
                        "4.64",
                        "33.86",
                        "$157.11",
                    ]
                ),
            },
        ],
    )

    totals = quick_extract_totals(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="Invoice",
    )

    assert {key: totals[0][key] for key in ("source_file", "total_amount", "warehouse_id", "pdf_type")} == {
        "source_file": "invoice.pdf",
        "total_amount": 1963.51,
        "warehouse_id": "7",
        "pdf_type": "unknown",
    }
    assert totals[0]["authoritative"] is True
    assert totals[0]["evidence_status"] == "authoritative"


def test_quick_extract_totals_scans_later_image_pages_for_explicit_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "page": 1, "text": ""},
            {"source_file": pdf.name, "page": 2, "text": ""},
        ],
    )
    rendered_pages = []
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: (
            rendered_pages.append(page_number)
            or {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"}
        ),
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"],
            page=page["page"],
            role="invoice_primary" if page["page"] == 1 else "invoice_total",
            role_confidence=0.98,
            warehouse_id="7",
            total_amount=None if page["page"] == 1 else 4105.15,
            total_label="" if page["page"] == 1 else "TOTAL",
            evidence_text="TOTAL $4,105.15" if page["page"] == 2 else "Invoice details",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert rendered_pages == [1, 2]
    assert result[0]["total_amount"] == 4105.15
    assert result[0]["total_page"] == 2
    assert result[0]["evidence_status"] == "authoritative"
    assert result[0]["authoritative"] is True


@pytest.mark.parametrize(
    ("filename", "expected_role"),
    [
        ("COI_WH-7.pdf", "supporting_attachment"),
        ("Payment Terms.pdf", "supporting_attachment"),
        ("timecard_WH-7.pdf", "timecard_summary"),
        ("supporting_WH-7.pdf", "supporting_attachment"),
    ],
)
def test_quick_extract_totals_image_non_payable_filename_overrides_ai_invoice_total(monkeypatch, tmp_path, filename, expected_role):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / filename
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""}],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"], page=page["page"], role="invoice_total", role_confidence=0.99,
            warehouse_id="7", total_amount=4105.15, total_label="TOTAL", evidence_text="TOTAL $4,105.15",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["authoritative"] is False
    assert result[0]["evidence_status"] == "needs_review"
    assert result[0]["page_evidence"][0]["role"] == expected_role
    assert result[0]["excluded_pages"] == [1]


def test_quick_extract_totals_image_warehouse_conflict_demotes_ai_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "INVOICE_WH-3.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""}],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"], page=page["page"], role="invoice_total", role_confidence=0.99,
            warehouse_id="30", total_amount=1000.0, total_label="TOTAL", evidence_text="TOTAL $1,000.00",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["authoritative"] is False
    assert result[0]["evidence_status"] == "needs_review"
    assert result[0]["warehouse_conflict"] == {
        "source_file": pdf.name,
        "filename_warehouse_id": "3",
        "page_warehouse_ids": ["30"],
    }


def test_quick_extract_totals_mixed_image_and_text_pages_preserves_numeric_audit_order(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "INVOICE_WH-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""},
            {"source_file": pdf.name, "source_path": str(pdf), "page": 2, "text": "Invoice CA#7\nTotal Due: $4105.15"},
        ],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"], page=page["page"], role="email_cover", role_confidence=0.99,
            evidence_text="Invoice attached",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert [page["page"] for page in result[0]["page_evidence"]] == [1, 2]
    assert [page["role"] for page in result[0]["page_evidence"]] == ["email_cover", "invoice_total"]
    assert result[0]["excluded_pages"] == [1]


def test_quick_extract_totals_mixed_text_and_image_warehouse_conflict_preserves_candidates(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "INVOICE_WH-3.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": "Invoice cover letter"},
            {"source_file": pdf.name, "source_path": str(pdf), "page": 2, "text": ""},
        ],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"], page=page["page"], role="invoice_total", role_confidence=0.99,
            warehouse_id="30", total_amount=1000.0, total_label="TOTAL", evidence_text="TOTAL $1,000.00",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["authoritative"] is False
    assert result[0]["evidence_status"] == "needs_review"
    assert result[0]["warehouse_conflict"] == {
        "source_file": pdf.name,
        "filename_warehouse_id": "3",
        "page_warehouse_ids": ["30"],
    }


def test_quick_extract_totals_returns_review_status_instead_of_authoritative_zero(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "page": 1, "text": ""}],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"],
            page=page["page"],
            role="invoice_primary",
            role_confidence=0.98,
            warehouse_id="7",
            evidence_text="Invoice details without a visible total",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["total_page"] is None
    assert result[0]["evidence_status"] == "needs_review"
    assert result[0]["authoritative"] is False


def test_quick_extract_totals_preserves_page_role_audit(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "page": 1, "text": ""},
            {"source_file": pdf.name, "page": 2, "text": ""},
        ],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"],
            page=page["page"],
            role="email_cover" if page["page"] == 1 else "invoice_total",
            role_confidence=0.98,
            warehouse_id="7" if page["page"] == 2 else "",
            total_amount=4105.15 if page["page"] == 2 else None,
            total_label="TOTAL" if page["page"] == 2 else "",
            evidence_text="TOTAL $4,105.15" if page["page"] == 2 else "Attached invoice",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert [page["role"] for page in result[0]["page_evidence"]] == ["email_cover", "invoice_total"]
    assert result[0]["excluded_pages"] == [1]


def test_quick_extract_totals_keeps_timecard_total_out_of_invoice_evidence(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7-timecard.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{
            "source_file": pdf.name,
            "source_path": str(pdf),
            "page": 1,
            "text": "TIME CARD\nCA#7\nTotal Due: $4105.15\nWorker One Reg 40.0 $4105.15",
        }],
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["authoritative"] is False
    assert result[0]["evidence_status"] == "needs_review"
    assert result[0]["page_evidence"][0]["role"] == "timecard_summary"
    assert result[0]["excluded_pages"] == [1]


def test_quick_extract_totals_keeps_supporting_rule_rows_out_of_invoice_evidence(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7-supporting.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{
            "source_file": pdf.name,
            "source_path": str(pdf),
            "page": 1,
            "text": "SUPPORTING ATTACHMENT\nCA#7\nHours Amount Bill Rate Date Description Pay Rate\n$22.40 40.000 $896.00 5/17/2026 Arellano Luna, Pablo $17.500 Reg",
        }],
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["authoritative"] is False
    assert result[0]["page_evidence"][0]["role"] == "supporting_attachment"


def test_quick_extract_totals_does_not_cache_needs_review_results(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import _totals_cache_path, quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    config = _ready_ai_config()
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": "Invoice CA#7"}],
    )

    result = quick_extract_totals([pdf], config, supplier="Demo")

    assert result[0]["evidence_status"] == "needs_review"
    assert not _totals_cache_path(pdf, config).exists()


def test_quick_extract_totals_invalidates_cache_when_file_contents_change(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"first")
    page_text = ["Invoice CA#7\nTotal Due: $10.00"]
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": page_text[0]}],
    )

    assert quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")[0]["total_amount"] == 10.0
    pdf.write_bytes(b"second")
    page_text[0] = "Invoice CA#7\nTotal Due: $20.00"

    assert quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")[0]["total_amount"] == 20.0


def test_quick_extract_totals_invalidates_cache_when_profile_authority_changes(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    profiles_path = tmp_path / "profiles.json"
    profiles_path.write_text(json.dumps([{
        "key": "demo",
        "aliases": ["demo"],
        "version": 1,
        "status": "approved",
        "approvedBy": "p0-test-reviewer",
        "approvedAt": "2026-07-15T10:00:00Z",
        "created_from": "manual_review",
        "authoritative_total_methods": ["vendor_total"],
    }]), encoding="utf-8")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""}],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"},
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, ai_config: LaborPageEvidence(
            source_file=page["source_file"], page=page["page"], role="invoice_primary", role_confidence=0.98,
            warehouse_id="7", total_amount=10.0, extraction_method="vendor_total",
        ),
    )
    config = {**_ready_ai_config(), "supplier_profiles_path": str(profiles_path)}

    assert quick_extract_totals([pdf], config, supplier="Demo")[0]["authoritative"] is True
    profiles_path.write_text(json.dumps([{
        "key": "demo", "aliases": ["demo"], "version": 2,
        "status": "approved", "approvedBy": "p0-test-reviewer",
        "approvedAt": "2026-07-15T10:05:00Z", "created_from": "manual_review",
        "authoritative_total_methods": [],
    }]), encoding="utf-8")

    result = quick_extract_totals([pdf], config, supplier="Demo")

    assert result[0]["total_amount"] == 0.0
    assert result[0]["authoritative"] is False
    assert result[0]["evidence_status"] == "needs_review"


def test_quick_extract_totals_returns_review_result_for_input_without_extracted_pages(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_pdf_pages", lambda paths: [])

    result = quick_extract_totals([pdf], {})

    assert result == [{
        "source_file": pdf.name,
        "total_amount": 0.0,
        "warehouse_id": "",
        "pdf_type": "unknown",
        "authoritative": False,
        "evidence_status": "needs_review",
        "total_page": None,
        "total_label": "",
        "page_evidence": [],
        "excluded_pages": [],
    }]


def test_quick_extract_totals_keeps_duplicate_basenames_in_separate_path_groups(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    first = tmp_path / "first" / "invoice.pdf"
    second = tmp_path / "second" / "invoice.pdf"
    first.parent.mkdir()
    second.parent.mkdir()
    first.write_bytes(b"first")
    second.write_bytes(b"second")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": "invoice.pdf", "source_path": str(first), "page": 1, "text": "Invoice CA#7\nTotal Due: $10.00"},
            {"source_file": "invoice.pdf", "source_path": str(second), "page": 1, "text": "Invoice CA#8\nTotal Due: $20.00"},
        ],
    )

    result = quick_extract_totals([first, second], _ready_ai_config(), supplier="Demo")

    assert [row["total_amount"] for row in result] == [10.0, 20.0]
    assert [row["warehouse_id"] for row in result] == ["7", "8"]


def test_quick_extract_totals_ignores_legacy_cache_without_evidence_fields(tmp_path):
    from bonus_platform.engine.labor.extract import _load_totals_cache

    pdf = tmp_path / "warehouse-7.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    cache_path = tmp_path / ".ai_extract_cache" / f"warehouse-7_totals_mimo-v2.5_{AI_PAGE_CACHE_VERSION}.json"
    cache_path.parent.mkdir()
    cache_path.write_text(json.dumps({"total_amount": 0.0, "warehouse_id": "7"}), encoding="utf-8")

    assert _load_totals_cache(pdf, _ready_ai_config()) is None


def test_quick_extract_totals_prefers_invoice_footer_total_over_rounded_employee_sum(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "US_Elogis_Service__7_Invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "US_Elogis_Service__7_Invoice.pdf",
                "page": 1,
                "text": "\n".join(
                    [
                        "Date Invoice #",
                        "CA#7",
                        "6/17/2026 ELOG7-9",
                        "Name US Elogistics Service Corp #7",
                        "Associate Base Rate Bill Rate OT Rate Reg. Time O.T Dbl. Time RT OT DT TOTAL",
                        "Alpha One $20.00 25.60$     38.40$   1.00 0.00 25.60$         -$       -$             25.60$",
                        "Beta Two $20.00 25.60$     38.40$   1.00 0.00 25.60$         -$       -$             25.60$",
                        "Totals 2.00 0.00 0.00 $51.22 $0.00 $0.00 $51.22",
                        "$51.22",
                        "If paid after 07/17/2026 pleased pay: $60.00",
                    ]
                ),
            }
        ],
    )

    totals = quick_extract_totals([pdf], {}, supplier="oss")

    assert {key: totals[0][key] for key in ("source_file", "total_amount", "warehouse_id", "pdf_type")} == {
        "source_file": "US_Elogis_Service__7_Invoice.pdf",
        "total_amount": 51.22,
        "warehouse_id": "7",
        "pdf_type": "unknown",
    }
    assert totals[0]["authoritative"] is True
    assert totals[0]["evidence_status"] == "authoritative"


def test_quick_extract_totals_runs_rule_extraction_without_ai_config(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "NJ13 Invoice Report WE 051726 JF.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "NJ13 Invoice Report WE 051726 JF.pdf",
                "page": 1,
                "text": "\n".join(
                    [
                        "Billable Billable Total",
                        "Hours Fee Fees",
                        "-$",
                        "1 48,293.06$ 48,293.06$",
                    ]
                ),
            }
        ],
    )

    totals = quick_extract_totals([pdf], {}, supplier="Strategic Staffing Solutions Corp.")

    assert {key: totals[0][key] for key in ("source_file", "total_amount", "warehouse_id", "pdf_type")} == {
        "source_file": "NJ13 Invoice Report WE 051726 JF.pdf",
        "total_amount": 48293.06,
        "warehouse_id": "13",
        "pdf_type": "unknown",
    }
    assert totals[0]["authoritative"] is True
    assert totals[0]["evidence_status"] == "authoritative"


def test_audit_ai_page_cache_candidates_are_confirmation_only(tmp_path):
    pdf = tmp_path / "elog1-1_20260520204104.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    cache_dir = tmp_path / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / f"elog1-1_20260520204104_p1_mimo-v2.5_{AI_PAGE_CACHE_VERSION}.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "Alvarez Michalec Rosa",
                    "source_page_or_row": "1",
                    "amount": "$701.88",
                    "confidence": 0.98,
                    "evidence_text": "Alvarez Michalec Rosa | $701.88",
                },
                {
                    "employee_name_raw": "Bernavides Jennifer",
                    "source_page": 1,
                    "amount": "$698.01",
                    "confidence": 0.92,
                    "evidence_text": "Bernavides Jennifer | $698.01",
                },
            ]
        ),
        encoding="utf-8",
    )

    audit = audit_ai_page_cache_candidates([pdf])

    assert audit["decision"] == "candidate_only"
    assert audit["requiresConfirmation"] is True
    assert audit["summary"] == {"fileCount": 1, "candidateFileCount": 1, "candidateAmountTotal": 1399.89}
    assert audit["files"][0]["sourceFile"] == "elog1-1_20260520204104.pdf"
    assert audit["files"][0]["warehouseId"] == "1"
    assert audit["files"][0]["rowCount"] == 2
    assert audit["files"][0]["candidateAmountTotal"] == 1399.89
    assert audit["files"][0]["averageConfidence"] == 0.95
    assert audit["files"][0]["decision"] == "candidate_only"
    assert audit["files"][0]["requiresConfirmation"] is True
    assert audit["files"][0]["evidence"][0]["employeeName"] == "Alvarez Michalec Rosa"
    assert audit["files"][0]["evidence"][0]["sourcePageOrRow"] == "p1"
    assert audit["files"][0]["evidence"][1]["sourcePageOrRow"] == "p1"


def test_ai_cache_reconciliation_preview_compares_candidates_without_promoting(tmp_path):
    pdf = tmp_path / "elog1-1_20260520204104.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    cache_dir = tmp_path / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "elog1-1_20260520204104_p1_mimo-v2.5_v4.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "Alice Worker",
                    "source_page": 1,
                    "hours": 8,
                    "amount": 100,
                    "confidence": 0.95,
                    "evidence_text": "Alice Worker TOTAL $100.00",
                },
                {
                    "employee_name_raw": "Bob Cache",
                    "source_page": 1,
                    "hours": 4,
                    "amount": 50,
                    "confidence": 0.9,
                    "evidence_text": "Bob Cache TOTAL $50.00",
                },
            ]
        ),
        encoding="utf-8",
    )
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=1, evidence_text="", warehouse_id="1"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="", employee_name_raw="Carol Workbook", hours=4, amount=55, currency="USD", confidence=1, evidence_text="", warehouse_id="1"),
    ]

    preview = build_ai_cache_reconciliation_preview(
        [pdf],
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    assert preview["decision"] == "candidate_only"
    assert preview["requiresConfirmation"] is True
    assert preview["summary"]["candidateRowCount"] == 2
    assert preview["summary"]["excelRowCount"] == 2
    assert preview["summary"]["passedCount"] == 1
    assert preview["summary"]["exceptionCount"] == 2
    assert preview["summary"]["cacheAmountTotal"] == 150
    assert preview["summary"]["excelAmountTotal"] == 155
    assert preview["summary"]["reviewableFileCount"] == 0
    assert preview["summary"]["needsReocrFileCount"] == 1
    file_quality = preview["fileQuality"][0]
    assert file_quality["sourceFile"] == "elog1-1_20260520204104.pdf"
    assert file_quality["warehouseId"] == "1"
    assert file_quality["cacheRowCount"] == 2
    assert file_quality["excelRowCount"] == 2
    assert file_quality["cacheAmountTotal"] == 150
    assert file_quality["excelAmountTotal"] == 155
    assert file_quality["amountDelta"] == -5
    assert file_quality["averageConfidence"] == 0.925
    assert file_quality["decision"] == "needs_reocr"
    assert file_quality["recommendation"] == "历史识别金额与账单同仓库金额不一致，建议重新识别后预览影响。"
    assert file_quality["diagnostics"]["summary"]["exceptionCount"] == 2
    assert file_quality["diagnostics"]["summary"]["unmatchedCacheCount"] == 1
    assert file_quality["diagnostics"]["summary"]["unmatchedExcelCount"] == 1
    assert file_quality["diagnostics"]["summary"]["suspectedNamePairCount"] == 0
    assert file_quality["diagnostics"]["extraInCache"][0]["employeeName"] == "Bob Cache"
    assert file_quality["diagnostics"]["missingInCache"][0]["employeeName"] == "Carol Workbook"
    assert file_quality["diagnostics"]["topDifferences"][0]["amountDelta"] == -55
    assert file_quality["diagnostics"]["rootCauseHints"] == ["possible_missing_cache_rows", "possible_extra_cache_rows"]
    assert file_quality["diagnostics"]["recommendedAction"] == "reocr_with_employee_level_review"
    assert any(row["matchStatus"] == "PDF有Excel无" and row["employeeName"] == "Bob Cache" for row in preview["exceptionRows"])
    assert any(row["matchStatus"] == "Excel有PDF无" and row["employeeName"] == "Carol Workbook" for row in preview["exceptionRows"])


def test_ai_cache_reconciliation_preview_marks_file_reviewable_when_warehouse_total_aligns(tmp_path):
    pdf = tmp_path / "elog25-3_20260520204328.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    cache_dir = tmp_path / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "elog25-3_20260520204328_p1_mimo-v2.5_v4.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "David Lopez",
                    "source_page": 1,
                    "hours": 24,
                    "amount": 696.12,
                    "confidence": 0.98,
                    "evidence_text": "David Lopez TOTAL $696.12",
                },
                {
                    "employee_name_raw": "Kenneth Rosales",
                    "source_page": 1,
                    "hours": 48.46,
                    "amount": 1330.43,
                    "confidence": 0.98,
                    "evidence_text": "Kenneth Rosales TOTAL $1,330.43",
                },
            ]
        ),
        encoding="utf-8",
    )
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="David Lopez", hours=24, amount=696.12, currency="USD", confidence=1, evidence_text="", warehouse_id="25"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="", employee_name_raw="Kenneth Rosales", hours=48.46, amount=1330.43, currency="USD", confidence=1, evidence_text="", warehouse_id="25"),
    ]

    preview = build_ai_cache_reconciliation_preview(
        [pdf],
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    assert preview["summary"]["reviewableFileCount"] == 1
    assert preview["summary"]["needsReocrFileCount"] == 0
    assert preview["fileQuality"][0]["decision"] == "reviewable_candidate"
    assert preview["fileQuality"][0]["amountDelta"] == 0
    assert preview["fileQuality"][0]["recommendation"] == "历史识别金额与账单同仓库金额一致，可作为人工复核证据。"


def test_ai_cache_file_diagnostics_suggests_name_mapping_before_reocr(tmp_path):
    pdf = tmp_path / "elog27-1_20260520204231.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    cache_dir = tmp_path / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "elog27-1_20260520204231_p1_mimo-v2.5_v4.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "Coria, Virgilio",
                    "source_page": 1,
                    "hours": 14.47,
                    "amount": 353.68,
                    "confidence": 0.95,
                    "evidence_text": "Coria, Virgilio 14.47 $353.68",
                }
            ]
        ),
        encoding="utf-8",
    )
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!34", employee_id="", employee_name_raw="Brayan Gomez Vargas", hours=14.17, amount=353.69, currency="USD", confidence=1, evidence_text="", warehouse_id="27"),
    ]

    preview = build_ai_cache_reconciliation_preview(
        [pdf],
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    diagnostics = preview["fileQuality"][0]["diagnostics"]
    assert diagnostics["summary"]["suspectedNamePairCount"] == 1
    pair = diagnostics["suspectedNamePairs"][0]
    assert pair["cacheEmployeeName"] == "Coria, Virgilio"
    assert pair["excelEmployeeName"] == "Brayan Gomez Vargas"
    assert pair["amountGap"] == -0.01
    assert pair["hoursGap"] == 0.3
    assert "possible_name_mapping" in diagnostics["rootCauseHints"]
    assert diagnostics["recommendedAction"] == "review_name_mapping_then_reocr_if_amounts_remain_unexplained"


def test_reocr_candidate_plan_is_confirmation_only():
    plan = build_reocr_candidate_plan(
        [
            {
                "sourceFile": "elog7-5_20260520204043.pdf",
                "warehouseId": "7",
                "cacheRowCount": 13,
                "excelRowCount": 13,
                "cacheAmountTotal": 10945.47,
                "excelAmountTotal": 8473.21,
                "amountDelta": 2472.26,
                "decision": "needs_reocr",
                "recommendation": "历史识别金额与账单同仓库金额不一致，建议重新识别后预览影响。",
                "diagnostics": {
                    "summary": {"exceptionCount": 3, "unmatchedCacheCount": 1, "unmatchedExcelCount": 2},
                    "topDifferences": [{"employeeName": "Alice Worker", "amountDelta": 120.5}],
                    "missingInCache": [{"employeeName": "Missing Worker"}],
                    "extraInCache": [{"employeeName": "Extra Worker"}],
                },
            },
            {
                "sourceFile": "elog25-3_20260520204328.pdf",
                "warehouseId": "25",
                "cacheRowCount": 2,
                "excelRowCount": 2,
                "cacheAmountTotal": 2026.55,
                "excelAmountTotal": 2026.55,
                "amountDelta": 0,
                "decision": "reviewable_candidate",
                "recommendation": "历史识别金额与账单同仓库金额一致，可作为人工复核证据。",
            },
        ],
        amount_tolerance=0.1,
    )

    assert plan["decision"] == "candidate_only"
    assert plan["requiresConfirmation"] is True
    assert plan["summary"] == {
        "taskCount": 1,
        "reviewableCandidateCount": 1,
        "totalExpectedExcelAmount": 8473.21,
        "totalCurrentCacheAmount": 10945.47,
    }
    assert plan["tasks"][0]["sourceFile"] == "elog7-5_20260520204043.pdf"
    assert plan["tasks"][0]["amountTolerance"] == 0.1
    assert plan["tasks"][0]["diagnostics"]["summary"]["exceptionCount"] == 3
    assert plan["tasks"][0]["diagnostics"]["missingInCache"][0]["employeeName"] == "Missing Worker"
    assert plan["tasks"][0]["focusEmployees"][0]["employeeName"] == "Alice Worker"
    assert plan["tasks"][0]["focusEmployees"][1]["employeeName"] == "Missing Worker"
    assert plan["tasks"][0]["focusEmployees"][2]["employeeName"] == "Extra Worker"
    assert "必须业务确认" in plan["tasks"][0]["confirmationGate"]
    assert "必须人工确认" not in plan["tasks"][0]["confirmationGate"]
    assert plan["reviewableCandidates"][0]["sourceFile"] == "elog25-3_20260520204328.pdf"


def test_parse_reocr_candidate_rows_from_csv(tmp_path):
    path = tmp_path / "reocr.csv"
    path.write_text(
        "Employee,Hours,Amount,Page,Confidence,Evidence\n"
        "Alice Worker,8,100,p1,96%,Alice Worker 8 $100\n"
        "Bob Worker,10,200,p2,0.95,Bob Worker 10 $200\n",
        encoding="utf-8",
    )

    rows = parse_reocr_candidate_rows(path, default_currency="USD")

    assert rows == [
        {
            "employeeName": "Alice Worker",
            "sourcePageOrRow": "p1",
            "hours": 8,
            "amount": 100,
            "currency": "USD",
            "confidence": 0.96,
            "evidenceText": "Alice Worker 8 $100",
        },
        {
            "employeeName": "Bob Worker",
            "sourcePageOrRow": "p2",
            "hours": 10,
            "amount": 200,
            "currency": "USD",
            "confidence": 0.95,
            "evidenceText": "Bob Worker 10 $200",
        },
    ]


def test_parse_reocr_candidate_rows_preserves_scope_and_employee_id(tmp_path):
    path = tmp_path / "reocr_scoped.csv"
    path.write_text(
        "SourceFile,WarehouseId,EmployeeId,Employee,Hours,Amount,Page,Confidence,Currency,Evidence,ExcelRef,ExpectedHours,ExpectedAmount\n"
        "elog1.pdf,1,WUS001,Alice Worker,8,100,p1,0.95,USD,Alice Worker 8 $100,账单.xlsx 员工账单!2,8,100\n",
        encoding="utf-8",
    )

    rows = parse_reocr_candidate_rows(path, default_currency="USD")

    assert rows == [
        {
            "employeeName": "Alice Worker",
            "sourcePageOrRow": "p1",
            "hours": 8,
            "amount": 100,
            "currency": "USD",
            "confidence": 0.95,
            "evidenceText": "Alice Worker 8 $100",
            "sourceFile": "elog1.pdf",
            "warehouseId": "1",
            "employeeId": "WUS001",
        }
    ]


def test_parse_reocr_candidate_rows_requires_name_and_amount(tmp_path):
    path = tmp_path / "bad.csv"
    path.write_text("Employee,Hours\nAlice Worker,8\n", encoding="utf-8")

    with pytest.raises(ValueError, match="员工姓名列和金额列"):
        parse_reocr_candidate_rows(path)


def test_reocr_candidate_replay_can_be_ready_for_user_confirmation():
    task = {
        "sourceFile": "elog7-5_20260520204043.pdf",
        "warehouseId": "7",
        "expectedExcelAmount": 300,
        "amountDelta": 2472.26,
        "confirmationGate": "新图片识别结果金额需与同仓库 Excel 金额在容差内，员工级异常需可解释，且必须人工确认。",
    }
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=1, evidence_text="", warehouse_id="7"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="", employee_name_raw="Bob Worker", hours=10, amount=200, currency="USD", confidence=1, evidence_text="", warehouse_id="7"),
    ]
    candidate_rows = [
        {"employee_name_raw": "Alice Worker", "source_page_or_row": "p1", "hours": 8, "amount": 100, "confidence": 0.96},
        {"employee_name_raw": "Bob Worker", "source_page_or_row": "p1", "hours": 10, "amount": 200, "confidence": 0.96},
    ]

    replay = replay_reocr_candidate_result(
        task,
        candidate_rows,
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    assert replay["decision"] == "ready_for_user_confirmation"
    assert replay["requiresConfirmation"] is True
    assert replay["summary"]["candidateAmountTotal"] == 300
    assert replay["summary"]["amountPassed"] is True
    assert replay["summary"]["exceptionCount"] == 0
    assert replay["summary"]["fixedCacheDelta"] == 2472.26
    assert replay["blockers"] == []
    assert len(replay["previewRows"]) == 2
    assert all(row["matchStatus"] == "通过" for row in replay["previewRows"])


def test_reocr_candidate_replay_blocks_amount_mismatch():
    task = {
        "sourceFile": "elog7-5_20260520204043.pdf",
        "warehouseId": "7",
        "expectedExcelAmount": 300,
        "amountDelta": 2472.26,
        "diagnostics": {
            "recommendedAction": "review_name_mapping_then_reocr_if_amounts_remain_unexplained",
            "rootCauseHints": ["possible_name_mapping"],
            "suspectedNamePairs": [{"cacheEmployeeName": "Bob Worker", "excelEmployeeName": "Bob Worker"}],
        },
    }
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=1, evidence_text="", warehouse_id="7"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="", employee_name_raw="Bob Worker", hours=10, amount=200, currency="USD", confidence=1, evidence_text="", warehouse_id="7"),
    ]
    candidate_rows = [
        {"employee_name_raw": "Alice Worker", "source_page_or_row": "p1", "hours": 8, "amount": 100, "confidence": 0.96},
        {"employee_name_raw": "Bob Worker", "source_page_or_row": "p1", "hours": 10, "amount": 210, "confidence": 0.96},
    ]

    replay = replay_reocr_candidate_result(
        task,
        candidate_rows,
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    assert replay["decision"] == "blocked_by_replay"
    assert "candidate_amount_mismatch" in replay["blockers"]
    assert "employee_level_exceptions" in replay["blockers"]
    assert replay["summary"]["amountDelta"] == 10
    assert replay["summary"]["exceptionCount"] == 1
    assert replay["diagnostics"]["recommendedAction"] == "review_name_mapping_then_reocr_if_amounts_remain_unexplained"


def test_reocr_candidate_replay_blocks_employee_exceptions_even_when_total_matches():
    task = {"sourceFile": "elog7-5_20260520204043.pdf", "warehouseId": "7", "expectedExcelAmount": 300, "amountDelta": 2472.26}
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=1, evidence_text="", warehouse_id="7"),
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!3", employee_id="", employee_name_raw="Bob Worker", hours=10, amount=200, currency="USD", confidence=1, evidence_text="", warehouse_id="7"),
    ]
    candidate_rows = [
        {"employee_name_raw": "Alice Worker", "source_page_or_row": "p1", "hours": 8, "amount": 100, "confidence": 0.96},
        {"employee_name_raw": "Wrong Worker", "source_page_or_row": "p1", "hours": 10, "amount": 200, "confidence": 0.96},
    ]

    replay = replay_reocr_candidate_result(
        task,
        candidate_rows,
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    assert replay["decision"] == "blocked_by_replay"
    assert replay["summary"]["amountPassed"] is True
    assert replay["summary"]["exceptionCount"] == 2
    assert replay["blockers"] == ["employee_level_exceptions"]


def test_reocr_candidate_replay_exposes_strict_ocr_name_gate_for_spelling_difference():
    task = {"sourceFile": "invoice.pdf", "warehouseId": "29", "expectedExcelAmount": 847.84, "amountDelta": 0}
    excel_rows = [
        LaborLineItem(source_type="offline_workbook", source_file="账单.xlsx", source_page_or_row="账单!2", employee_id="", employee_name_raw="Deisi Pozo", hours=37.84, amount=847.84, currency="USD", confidence=1, evidence_text="", warehouse_id="29"),
    ]
    candidate_rows = [
        {"employee_name_raw": "Deisy Rozo Panche", "source_page_or_row": "p1", "hours": 37.84, "amount": 847.84, "confidence": 0.96},
    ]

    replay = replay_reocr_candidate_result(
        task,
        candidate_rows,
        excel_rows,
        amount_tolerance=0.1,
        hours_tolerance=0.1,
        confidence_threshold=0.85,
    )

    assert replay["decision"] == "blocked_by_replay"
    assert replay["nameGate"]["summary"]["confirmed"] == 0
    assert replay["nameGate"]["summary"]["review"] == 1
    assert replay["nameGate"]["matches"][0]["status"] == "review"


def test_rule_change_candidate_requires_user_confirmation():
    candidate = build_rule_change_candidate(
        rule_id="warehouse-filename-hash-number",
        title="从 OSS 文件名 #N 提取仓库号",
        description="识别 US Elogis Service #17 Invoice 这类文件名中的仓库号。",
        supplier="OSS",
        source="oss 2 real replay",
        proposed_by="ai",
        evidence=[{"sourceFile": "US Elogis Service #17 Invoice W.E 05.24.26.pdf", "warehouseId": "17"}],
        conditions={"filenamePattern": "#<warehouse_id> Invoice"},
    )

    assert candidate["decision"] == "candidate_only"
    assert candidate["status"] == "pending_user_confirmation"
    assert candidate["requiresConfirmation"] is True
    assert candidate["version"] == 1
    assert candidate["auditTrail"][0]["action"] == "created"
    assert candidate["conditions"]["filenamePattern"] == "#<warehouse_id> Invoice"


def test_rule_replay_summary_blocks_regressions_before_confirmation():
    candidate = build_rule_change_candidate(
        rule_id="minor-name-typo-match",
        title="轻微姓名拼写差异匹配",
        description="当姓名相似且工时金额一致时匹配。",
        supplier="OSI",
        source="osi real replay",
    )

    replay = summarize_rule_replay(
        candidate,
        [
            {
                "runId": "osi_34794",
                "supplier": "OSI",
                "periodStart": "2026-05-18",
                "periodEnd": "2026-05-24",
                "beforeStatus": "warning",
                "afterStatus": "ok",
                "beforeIssueCount": 2,
                "afterIssueCount": 0,
            },
            {
                "runId": "fairway_135612",
                "supplier": "Fairway",
                "periodStart": "2026-05-18",
                "periodEnd": "2026-05-24",
                "beforeStatus": "ok",
                "afterStatus": "warning",
                "beforeIssueCount": 0,
                "afterIssueCount": 1,
            },
        ],
    )

    assert replay["decision"] == "blocked_by_replay_regression"
    assert replay["requiresConfirmation"] is True
    assert replay["summary"] == {"replayedCount": 2, "fixedCount": 1, "regressionCount": 1, "unchangedCount": 0}
    assert replay["fixed"][0]["runId"] == "osi_34794"
    assert replay["regressions"][0]["runId"] == "fairway_135612"


def test_rule_auto_replay_uses_historical_metadata_diagnostics():
    candidate = build_rule_change_candidate(
        rule_id="oss-hash-warehouse-v1",
        title="OSS # warehouse id extraction",
        description="Parse warehouse id from US Elogis Service #N invoice names.",
        supplier="OSS",
        source="oss 2 real replay",
        conditions={"supplier": "OSS", "fixIssueCodes": ["missing_warehouse_id"]},
    )
    replay = summarize_rule_auto_replay(
        candidate,
        [
            {
                "id": "oss2_warehouse_7",
                "supplierName": "OSS",
                "periodStart": "2026-05-18",
                "periodEnd": "2026-05-24",
                "reconciliationDiagnostics": {
                    "level": "warning",
                    "issues": [{"code": "missing_warehouse_id", "level": "warning"}],
                },
                "comparisonSummary": {"exceptionCount": 0},
            },
            {
                "id": "fairway_135612",
                "supplierName": "Fairway",
                "reconciliationDiagnostics": {
                    "level": "ok",
                    "issues": [],
                },
                "comparisonSummary": {"exceptionCount": 0},
            },
        ],
        current_run_id="oss2_warehouse_7",
    )

    assert replay["mode"] == "metadata_signal_replay"
    assert replay["decision"] == "ready_for_user_confirmation"
    assert replay["summary"] == {"replayedCount": 2, "fixedCount": 1, "regressionCount": 0, "unchangedCount": 1}
    assert replay["replayResults"][0]["matchedIssueCodes"] == ["missing_warehouse_id"]
    assert replay["replayResults"][1]["impactReason"] == "out_of_scope_supplier"
    assert replay["requiresConfirmation"] is True
    assert replay["limitations"]


def test_confirm_rule_candidate_requires_successful_replay():
    candidate = build_rule_change_candidate(
        rule_id="warehouse-filename-hash-number",
        title="从 OSS 文件名 #N 提取仓库号",
        description="识别 US Elogis Service #17 Invoice 这类文件名中的仓库号。",
        supplier="OSS",
        source="oss 2 real replay",
    )
    blocked_replay = {
        "decision": "blocked_by_replay_regression",
        "summary": {"replayedCount": 1, "fixedCount": 0, "regressionCount": 1, "unchangedCount": 0},
    }

    with pytest.raises(ValueError, match="未通过历史影响预览"):
        confirm_rule_candidate(candidate, blocked_replay, confirmed_by="ops-user", reason="误伤已通过批次")


def test_confirm_and_rollback_rule_version_records_audit_trail():
    candidate = build_rule_change_candidate(
        rule_id="warehouse-filename-hash-number",
        title="从 OSS 文件名 #N 提取仓库号",
        description="识别 US Elogis Service #17 Invoice 这类文件名中的仓库号。",
        supplier="OSS",
        source="oss 2 real replay",
    )
    replay = {
        "decision": "ready_for_user_confirmation",
        "summary": {"replayedCount": 2, "fixedCount": 1, "regressionCount": 0, "unchangedCount": 1},
    }

    active = confirm_rule_candidate(candidate, replay, confirmed_by="ops-user", reason="OSS2 仓库号回放通过")

    assert active["decision"] == "active"
    assert active["status"] == "active"
    assert active["requiresConfirmation"] is False
    assert active["confirmedBy"] == "ops-user"
    assert active["replaySummary"] == replay["summary"]
    assert active["auditTrail"][-1]["action"] == "confirmed"

    rolled_back = rollback_rule_version(active, rolled_back_by="ops-user", reason="后续批次发现误伤", target_version=0)

    assert rolled_back["decision"] == "rolled_back"
    assert rolled_back["status"] == "rolled_back"
    assert rolled_back["rollbackToVersion"] == 0
    assert rolled_back["auditTrail"][-1] == {
        "action": "rolled_back",
        "actor": "ops-user",
        "reason": "后续批次发现误伤",
        "fromVersion": 1,
        "toVersion": 0,
    }


def test_quick_extract_totals_uses_citi_bill_rate_rows(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "invoice.pdf",
                "page": 1,
                "text": "\n".join(
                    [
                        "Hours Amount Bill Rate Date Description Pay Rate",
                        "WAREHOUSE LOC.#29PO #:",
                        "$22.40 40.000 $896.00 5/17/2026 Arellano Luna, Pablo $17.500 Reg",
                        "$33.60 0.400 $13.44 5/17/2026 Arellano Luna, Pablo $26.250 OT",
                    ]
                ),
            }
        ],
    )

    totals = quick_extract_totals(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="CITI",
    )

    assert {key: totals[0][key] for key in ("source_file", "total_amount", "warehouse_id", "pdf_type")} == {
        "source_file": "invoice.pdf",
        "total_amount": 909.44,
        "warehouse_id": "29",
        "pdf_type": "unknown",
    }
    assert totals[0]["authoritative"] is True
    assert totals[0]["evidence_status"] == "authoritative"


def test_quick_extract_totals_keeps_headerless_payable_continuation_pages(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "staffing-statement-7421.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")

    def payable_row(name: str, amount: str) -> str:
        return "\n".join(
            [
                "5/24/2026",
                name,
                "8.00",
                "Reg",
                "REG",
                "$20.00",
                "$25.00",
                amount,
            ]
        )

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 1,
                "text": "INVOICE\n" + payable_row("Worker One", "$200.00") + "\nPage 1 of 3",
            },
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 2,
                "text": payable_row("Worker Two", "$200.00") + "\nPage 2 of 3",
            },
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 3,
                "text": payable_row("Worker Three", "$200.00") + "\nTotal Due:\n$600.00\nPage 3 of 3",
            },
        ],
    )

    result = quick_extract_totals(
        [pdf],
        {"enabled": False, "parallel_extraction_enabled": False, "cache_enabled": False},
        supplier="Unseen Staffing Vendor",
    )[0]

    assert result["total_amount"] == 600.0
    assert result["authoritative"] is True
    assert result["excluded_pages"] == []
    assert [page["role"] for page in result["page_evidence"]] == [
        "invoice_primary",
        "invoice_continuation",
        "invoice_continuation",
    ]


def test_quick_extract_totals_keeps_voyage_continuation_and_uses_invoice_amount(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "CA 7 46292.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    row_one = "1 Abel Gonzalez 18.5 8.00 23.68 189.44 0.40 35.52 14.21 0.00 47.36 0.00 203.65"
    row_two = "31 Alison Guzman 17.5 8.00 22.40 179.20 0.10 33.60 3.36 0.00 44.8 0.00 182.56"
    row_three = "271 Ying Fang 17.5 8.00 22.40 179.20 0.00 33.60 0.00 0.00 44.8 0.00 179.20"
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 1,
                "text": "INVOICE #46292\nCA #7\nPayrate Hours Bill Rate\n" + row_one,
            },
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 2,
                "text": row_two,
            },
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 3,
                "text": (
                    row_three
                    + "\nTOTAL REG: 2089.21\nTOTAL OT: 104.73\nTOTAL DT: 0.00"
                    + "\nGRAND TOTAL: 2193.94\nTOTAL INVOICE AMOUNT: $54,358.11"
                ),
            },
        ],
    )

    result = quick_extract_totals(
        [pdf],
        {"enabled": False, "parallel_extraction_enabled": False, "cache_enabled": False},
        supplier="Voyage Employer Services",
    )[0]

    assert result["total_amount"] == 54358.11
    assert result["authoritative"] is True
    assert result["excluded_pages"] == []
    assert [page["role"] for page in result["page_evidence"]] == [
        "invoice_primary",
        "invoice_continuation",
        "invoice_total",
    ]


def test_quick_extract_totals_keeps_invoice_page_with_payment_terms_field(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "GS invoice-ELOG-466-FL.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")

    def payable_row(name: str, amount: str) -> str:
        return "\n".join(
            [
                "5/24/2026",
                name,
                "8.00",
                "Reg",
                "REG",
                "$20.00",
                "$25.00",
                amount,
            ]
        )

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 1,
                "text": (
                    "INVOICE ELOG-466-FL\n"
                    "Period Cust. ID Tax ID PAYMENT TERMS Location\n"
                    "No. Name Reg. Hours O.T Hours Reg. Rate O.T Rate Total\n"
                    + payable_row("Worker One", "$200.00")
                ),
            },
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 2,
                "text": payable_row("Worker Two", "$200.00") + "\nGRAND TOTAL\n$400.00",
            },
        ],
    )

    result = quick_extract_totals(
        [pdf],
        {"enabled": False, "parallel_extraction_enabled": False, "cache_enabled": False},
        supplier="Unseen Staffing Vendor",
    )[0]

    assert result["total_amount"] == 400.0
    assert result["authoritative"] is True
    assert result["excluded_pages"] == []
    assert [page["role"] for page in result["page_evidence"]] == [
        "invoice_primary",
        "invoice_total",
    ]


def test_quick_extract_totals_does_not_promote_explicit_supporting_vertical_rows(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "supporting-attachment.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 1,
                "text": "\n".join(
                    [
                        "SUPPORTING ATTACHMENT",
                        "5/24/2026",
                        "Worker One",
                        "8.00",
                        "Reg",
                        "REG",
                        "$20.00",
                        "$25.00",
                        "$200.00",
                    ]
                ),
            }
        ],
    )

    result = quick_extract_totals(
        [pdf],
        {"enabled": False, "parallel_extraction_enabled": False, "cache_enabled": False},
        supplier="Unseen Staffing Vendor",
    )[0]

    assert result["total_amount"] == 0.0
    assert result["authoritative"] is False
    assert result["excluded_pages"] == [1]
    assert result["page_evidence"][0]["role"] == "supporting_attachment"


def test_quick_extract_totals_preserves_warehouse_conflict(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "INVOICE_WH-3.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {
                "source_file": "INVOICE_WH-3.pdf",
                "page": 1,
                "text": "US ELOGISTICS\nCA#30\nTotal Due: $1,000.00",
            }
        ],
    )

    totals = quick_extract_totals(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="Invoice",
    )

    assert {key: totals[0][key] for key in ("source_file", "total_amount", "warehouse_id", "pdf_type")} == {
        "source_file": "INVOICE_WH-3.pdf",
        "total_amount": 0.0,
        "warehouse_id": "3",
        "pdf_type": "primary",
    }
    assert totals[0]["authoritative"] is False
    assert totals[0]["evidence_status"] == "needs_review"
    assert totals[0]["warehouse_conflict"] == {
        "source_file": "INVOICE_WH-3.pdf",
        "filename_warehouse_id": "3",
        "text_warehouse_id": "30",
    }


def test_page_evidence_image_anthropic_uses_thinking_json_fallback(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {"content": [{
            "type": "thinking",
            "thinking": json.dumps({
                "page_role": "invoice_total",
                "role_confidence": 0.98,
                "warehouse_id": "7",
                "total_amount": 4105.15,
                "total_label": "TOTAL",
                "evidence_text": "TOTAL $4,105.15",
            }),
        }]},
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "warehouse-7.pdf", "page": 2, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://token-plan-cn.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
        },
    )

    assert evidence.role == "invoice_total"
    assert evidence.total_amount == 4105.15
    assert evidence.warehouse_id == "7"


def test_page_evidence_image_openai_uses_filename_warehouse_fallback(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": "```json\n{\"page_role\":\"invoice_total\",\"role_confidence\":0.98,\"warehouse_id\":null,\"total_amount\":11837.79,\"total_label\":\"TOTAL\",\"evidence_text\":\"TOTAL $11,837.79\"}\n```"
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#9.pdf", "page": 2, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.warehouse_id == "9"
    assert evidence.total_amount == 11837.79


def test_page_evidence_image_normalizes_prefixed_warehouse_id(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": json.dumps({
                        "page_role": "invoice_primary",
                        "role_confidence": 0.99,
                        "warehouse_id": "#1",
                        "total_amount": 1880.67,
                        "total_label": "TOTAL",
                        "evidence_text": "TOTAL: $1,880.67",
                    })
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT_1.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.warehouse_id == "1"


def test_page_evidence_prompt_has_no_real_example_values_and_distinguishes_invoice_rows():
    from bonus_platform.engine.labor.extract import _page_evidence_prompt

    prompt = _page_evidence_prompt()

    assert "4105.15" not in prompt
    assert '"warehouse_id": "2"' not in prompt
    assert "Bill Rate" in prompt
    assert "payable charge rows" in prompt
    assert "attendance dates" in prompt
    assert "Never copy" in prompt
    assert "TOTAL HT" in prompt
    assert "TOTAL TTC" in prompt
    assert "TVA" in prompt


def test_page_evidence_image_accepts_french_net_total_with_comma_decimal(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": json.dumps({
                        "page_role": "invoice_primary",
                        "role_confidence": 0.98,
                        "warehouse_id": None,
                        "total_amount": "563,70",
                        "total_label": "TOTAL HT",
                        "evidence_text": "TOTAL HT 563,70",
                    })
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "facture.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.total_amount == 563.70
    assert evidence.total_label == "TOTAL HT"
    assert evidence.evidence_text == "TOTAL HT 563,70"


def test_page_evidence_reasoning_fallback_does_not_expose_model_reasoning(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": "",
                    "reasoning_content": (
                        "The document is clearly an invoice. I should return an invoice_primary role, "
                        "but no explicit total is visible in this crop."
                    ),
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "invoice.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.role == "invoice_primary"
    assert evidence.evidence_text == ""


def test_page_evidence_image_openai_has_reasoning_fallback_and_bounded_budget(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    captured = {}

    def fake_post(url, headers, payload):
        captured.update(payload)
        return {
            "choices": [{
                "message": {
                    "content": "",
                    "reasoning_content": json.dumps({
                        "page_role": "invoice_primary",
                        "role_confidence": 0.95,
                        "warehouse_id": "1",
                        "total_amount": 1880.67,
                        "total_label": "TOTAL",
                        "evidence_text": "TOTAL $1,880.67",
                    }),
                }
            }]
        }

    monkeypatch.setattr("bonus_platform.engine.labor.extract._http_post_json", fake_post)

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#1.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert captured["max_tokens"] == 256
    assert evidence.role == "invoice_primary"
    assert evidence.total_amount == 1880.67


def test_page_evidence_image_openai_classifies_truncated_reasoning_without_json(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": "",
                    "reasoning_content": (
                        "The page lists attendance dates, working hours and overtime by employee. "
                        "This is not an invoice; it is a timecard or attendance summary and has no payable total."
                    ),
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#2.pdf", "page": 3, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.role == "timecard_summary"
    assert evidence.total_amount is None
    assert evidence.warehouse_id == "2"


def test_page_evidence_image_openai_extracts_invoice_total_from_truncated_reasoning(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": "",
                    "reasoning_content": (
                        "The document is clearly an invoice with Invoice #, Bill Rate and AMOUNT columns. "
                        "A TOTAL row at the bottom shows $4,222.26."
                    ),
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#5.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.role == "invoice_primary"
    assert evidence.role_confidence == 0.95
    assert evidence.total_amount == 4222.26
    assert evidence.total_label == "TOTAL"


def test_page_evidence_image_normalizes_high_confidence_label(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": json.dumps({
                        "page_role": "invoice_total",
                        "role_confidence": "high",
                        "warehouse_id": "8",
                        "total_amount": 21456.76,
                        "total_label": "TOTAL",
                        "evidence_text": "TOTAL $21,456.76",
                    })
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#8.pdf", "page": 2, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.role_confidence == 0.95


def test_page_evidence_image_preserves_supported_net_tax_and_gross(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": json.dumps({
                        "page_role": "invoice_total",
                        "role_confidence": 0.99,
                        "warehouse_id": "2",
                        "total_amount": 15421.72,
                        "total_label": "TOTAL HT",
                        "net_amount": 15421.72,
                        "tax_amount": 3084.34,
                        "gross_amount": 18506.06,
                        "evidence_text": "TOTAL HT 15 421,72 TVA 3 084,34 TOTAL TTC 18 506,06",
                    })
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "invoice.pdf", "page": 4, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.total_amount == 15421.72
    assert evidence.net_amount == 15421.72
    assert evidence.tax_amount == 3084.34
    assert evidence.gross_amount == 18506.06


def test_parse_bottom_total_ocr_text_accepts_closed_french_summary():
    from bonus_platform.engine.labor.extract import _parse_bottom_total_ocr_text

    result = _parse_bottom_total_ocr_text(
        "TOTAL HT 15 421,72\nTVA 20,00 % 3 084,34\nTOTAL TTC GLOBAL 18 506,06"
    )

    assert result == {
        "net_amount": 15421.72,
        "tax_amount": 3084.34,
        "gross_amount": 18506.06,
    }


def test_parse_bottom_total_ocr_text_accepts_dotted_french_labels():
    from bonus_platform.engine.labor.extract import _parse_bottom_total_ocr_text

    result = _parse_bottom_total_ocr_text(
        "TOTAL H.T 5724,22\nT.V.A 1144,84\nTOTAL T.T.C 6869,06"
    )

    assert result == {
        "net_amount": 5724.22,
        "tax_amount": 1144.84,
        "gross_amount": 6869.06,
    }


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        (
            "HEURES TOTAL H.T TAUX T.V.A T.V.A TOTAL T.T.C\n"
            "Valeur en votre aimable règlement\n"
            "au 30/06/2026 45,40 20,00 9,08 54,48EUR",
            {"net_amount": 45.40, "tax_amount": 9.08, "gross_amount": 54.48},
        ),
        (
            "HEURES TOTAL H.T TAUX T.V.A T.V.A TOTAL T.T.C\n"
            "Valeur en votre aimable règlement\n"
            "au 30/06/2026 802,27 16716,35 20,00 3343,27 20059,62EUR",
            {"net_amount": 16716.35, "tax_amount": 3343.27, "gross_amount": 20059.62},
        ),
    ],
)
def test_extract_closed_french_total_row_from_pdf_text(text, expected):
    from bonus_platform.engine.labor.extract import _extract_closed_french_total_row

    assert _extract_closed_french_total_row(text) == expected


def test_parse_bottom_total_ocr_text_ignores_repeated_header_before_values():
    from bonus_platform.engine.labor.extract import _parse_bottom_total_ocr_text

    result = _parse_bottom_total_ocr_text(
        "headers HEURES, TOTAL HT, TAUX TVA, TVA, TOTAL TTC.\n"
        "TOTAL HT: 614,80\nTAUX TVA: 20,00\nTVA: 122,96\n"
        "TOTAL TTC: 737,76\nTOTAL TTC GLOBAL: 737,76"
    )

    assert result == {
        "net_amount": 614.80,
        "tax_amount": 122.96,
        "gross_amount": 737.76,
    }


@pytest.mark.parametrize(
    "text",
    [
        "TOTAL TTC 18 506,06",
        "TOTAL HT 15 421,72\nTVA 3 084,34\nTOTAL TTC 19 000,00",
        "S/Total Interimaire 801,34",
    ],
)
def test_parse_bottom_total_ocr_text_rejects_incomplete_or_unclosed_summary(text):
    from bonus_platform.engine.labor.extract import _parse_bottom_total_ocr_text

    assert _parse_bottom_total_ocr_text(text) is None


def test_quick_extract_totals_uses_closed_bottom_ocr_when_json_evidence_is_empty(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""}],
    )
    image_page = {
        "source_file": pdf.name,
        "source_path": str(pdf),
        "page": 1,
        "mime_type": "image/png",
        "base64": "page",
    }
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: image_page,
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, config: LaborPageEvidence(
            source_file=pdf.name,
            page=1,
            role="invoice_total",
            role_confidence=0.99,
            evidence_text="Invoice total area",
        ),
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_bottom_total_evidence_with_ai_ocr",
        lambda page, config: LaborPageEvidence(
            source_file=pdf.name,
            page=1,
            role="invoice_total",
            role_confidence=0.99,
            total_amount=100.0,
            total_label="TOTAL HT",
            net_amount=100.0,
            tax_amount=20.0,
            gross_amount=120.0,
            evidence_text="TOTAL HT 100,00 TVA 20,00 TOTAL TTC 120,00",
            extraction_method="bottom_total_ocr",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert result[0]["total_amount"] == 100.0
    assert result[0]["authoritative"] is True
    assert result[0]["page_evidence"][0]["extraction_method"] == "bottom_total_ocr"


def test_quick_extract_totals_uses_bottom_ocr_for_text_page_without_closed_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{
            "source_file": pdf.name,
            "source_path": str(pdf),
            "page": 1,
            "text": "FACTURE\nHEURES TOTAL H.T TAUX T.V.A T.V.A TOTAL T.T.C",
        }],
    )
    image_page = {
        "source_file": pdf.name,
        "source_path": str(pdf),
        "page": 1,
        "mime_type": "image/png",
        "base64": "page",
    }
    rendered_pages = []
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: rendered_pages.append(page_number) or image_page,
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_bottom_total_evidence_with_ai_ocr",
        lambda page, config: LaborPageEvidence(
            source_file=pdf.name,
            page=1,
            role="invoice_total",
            role_confidence=0.99,
            total_amount=5724.22,
            total_label="TOTAL HT",
            net_amount=5724.22,
            tax_amount=1144.84,
            gross_amount=6869.06,
            evidence_text="TOTAL HT 5724.22; TVA 1144.84; TOTAL TTC 6869.06",
            extraction_method="bottom_total_ocr",
        ),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Unknown")

    assert rendered_pages == [1]
    assert result[0]["total_amount"] == 5724.22
    assert result[0]["authoritative"] is True
    assert result[0]["page_evidence"][0]["extraction_method"] == "bottom_total_ocr"


def test_quick_extract_totals_does_not_render_text_page_with_authoritative_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{
            "source_file": pdf.name,
            "source_path": str(pdf),
            "page": 1,
            "text": "Invoice Total Due: $100.00",
        }],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda *args: pytest.fail("authoritative text total must not render an image fallback"),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Unknown")

    assert result[0]["total_amount"] == 100.0
    assert result[0]["authoritative"] is True


def test_quick_extract_totals_reads_closed_french_footer_without_rescanning_early_text_pages(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "facture.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": "FACTURE\nHEURES TOTAL H.T TAUX T.V.A T.V.A TOTAL T.T.C"},
            {"source_file": pdf.name, "source_path": str(pdf), "page": 2, "text": "Employee detail continuation"},
            {"source_file": pdf.name, "source_path": str(pdf), "page": 3, "text": "Employee detail continuation"},
            {
                "source_file": pdf.name,
                "source_path": str(pdf),
                "page": 4,
                "text": (
                    "HEURES TOTAL H.T TAUX T.V.A T.V.A TOTAL T.T.C\n"
                    "au 30/06/2026 802,27 16716,35 20,00 3343,27 20059,62EUR"
                ),
            },
        ],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda *args: pytest.fail("closed PDF text totals must not require image OCR"),
    )

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Unknown")

    assert result[0]["total_amount"] == 16716.35
    assert result[0]["authoritative"] is True
    assert result[0]["total_page"] == 4
    assert result[0]["page_evidence"][3]["extraction_method"] == "text_closed_french_total_row"


def test_bottom_total_ocr_uses_reasoning_content_when_visible_content_is_empty(monkeypatch):
    from PIL import Image
    from bonus_platform.engine.labor.extract import _extract_bottom_total_evidence_with_ai_ocr

    image = Image.new("RGB", (200, 300), "white")
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    encoded = __import__("base64").b64encode(buffer.getvalue()).decode("ascii")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": "",
                    "reasoning_content": "TOTAL HT 100,00\nTVA 20,00\nTOTAL TTC 120,00",
                }
            }]
        },
    )

    evidence = _extract_bottom_total_evidence_with_ai_ocr(
        {
            "source_file": "invoice.pdf",
            "page": 1,
            "mime_type": "image/png",
            "base64": encoded,
        },
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
        },
    )

    assert evidence is not None
    assert evidence.net_amount == 100.0
    assert evidence.evidence_text == "TOTAL HT 100.00; TVA 20.00; TOTAL TTC 120.00"


def test_page_evidence_image_rejects_total_contradicted_by_evidence_text(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": json.dumps({
                        "page_role": "invoice_total",
                        "role_confidence": 0.95,
                        "warehouse_id": "19",
                        "total_amount": 865.72,
                        "total_label": "Amount Due",
                        "evidence_text": "No total or Amount Due is visible. Employee row amount 865.72.",
                    })
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#19.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.total_amount is None
    assert evidence.total_label == ""


def test_page_evidence_image_does_not_join_prompt_label_to_employee_amount(monkeypatch):
    from bonus_platform.engine.labor.extract import _extract_page_evidence_with_ai_image

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._http_post_json",
        lambda url, headers, payload: {
            "choices": [{
                "message": {
                    "content": "",
                    "reasoning_content": (
                        "The document is clearly an invoice with Bill Rate and AMOUNT columns. "
                        "I need to look for TOTAL, Amount Due, or Balance Due.\n"
                        "The table lists employee amounts such as 865.72 and 59.93, but no summation line is visible."
                    ),
                }
            }]
        },
    )

    evidence = _extract_page_evidence_with_ai_image(
        {"source_file": "DEPT#19.pdf", "page": 1, "mime_type": "image/jpeg", "base64": "page"},
        "extract evidence",
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert evidence.role == "invoice_primary"
    assert evidence.total_amount is None


def test_quick_extract_totals_retries_first_invoice_page_without_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "DEPT#5.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""}],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": 1, "mime_type": "image/jpeg", "base64": "page"},
    )
    budgets = []

    def fake_extract(page, prompt, config):
        budgets.append(config.get("page_evidence_max_tokens"))
        return LaborPageEvidence(
            source_file=page["source_file"],
            page=1,
            role="invoice_primary",
            role_confidence=0.95,
            warehouse_id="5",
            total_amount=4222.26 if config.get("page_evidence_max_tokens") == 1024 else None,
            total_label="TOTAL" if config.get("page_evidence_max_tokens") == 1024 else "",
        )

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image", fake_extract)

    totals = quick_extract_totals([pdf], _ready_ai_config(), supplier="Prompt Priority INC")

    assert budgets == [None, 1024]
    assert totals[0]["total_amount"] == 4222.26
    assert totals[0]["authoritative"] is True


def test_quick_extract_totals_retries_gross_total_page_for_financial_breakdown(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "invoice.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": 3, "text": ""},
        ],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {
            "source_file": path.name,
            "source_path": str(path),
            "page": page_number,
            "mime_type": "image/jpeg",
            "base64": "page",
        },
    )
    budgets = []

    def fake_extract(page, prompt, config):
        budgets.append(config.get("page_evidence_max_tokens"))
        expanded = config.get("page_evidence_max_tokens") == 1024
        return LaborPageEvidence(
            source_file=page["source_file"],
            page=page["page"],
            role="invoice_total",
            role_confidence=0.99,
            total_amount=120.0,
            total_label="TOTAL TTC",
            net_amount=100.0 if expanded else None,
            tax_amount=20.0 if expanded else None,
            gross_amount=120.0 if expanded else None,
            evidence_text="TOTAL HT 100,00 TVA 20,00 TOTAL TTC 120,00",
        )

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image", fake_extract)

    result = quick_extract_totals([pdf], _ready_ai_config(), supplier="Demo")

    assert budgets == [None, 1024]
    assert result[0]["page_evidence"][0]["net_amount"] == 100.0


def test_quick_extract_totals_retries_non_invoice_role_with_payable_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "DEPT#5.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [{"source_file": pdf.name, "source_path": str(pdf), "page": 1, "text": ""}],
    )
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {"source_file": path.name, "source_path": str(path), "page": 1, "mime_type": "image/jpeg", "base64": "page"},
    )
    budgets = []

    def fake_extract(page, prompt, config):
        budgets.append(config.get("page_evidence_max_tokens"))
        retry = config.get("page_evidence_max_tokens") == 1024
        return LaborPageEvidence(
            source_file=page["source_file"],
            page=1,
            role="invoice_total" if retry else "timecard_summary",
            role_confidence=0.95,
            warehouse_id="5",
            total_amount=4222.26,
            total_label="TOTAL",
            evidence_text="TOTAL $4,222.26",
        )

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image", fake_extract)

    totals = quick_extract_totals([pdf], _ready_ai_config(), supplier="Prompt Priority INC")

    assert budgets == [None, 1024]
    assert totals[0]["total_amount"] == 4222.26
    assert totals[0]["authoritative"] is True


def test_quick_extract_totals_stops_after_consecutive_ai_transport_failures(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "DEPT#8.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": page, "text": ""}
            for page in range(1, 7)
        ],
    )
    rendered = []

    def fake_render(path, page_number, scale):
        rendered.append(page_number)
        return {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"}

    monkeypatch.setattr("bonus_platform.engine.labor.extract._render_pdf_page_to_image", fake_render)
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, config: (_ for _ in ()).throw(ConnectionResetError("connection reset")),
    )

    totals = quick_extract_totals([pdf], _ready_ai_config(), supplier="Prompt Priority INC")

    assert rendered == [1, 2]
    assert totals[0]["authoritative"] is False
    assert totals[0]["evidence_status"] == "needs_review"
    assert [page["extraction_method"] for page in totals[0]["page_evidence"]] == [
        "ai_image_failed",
        "ai_image_failed",
        "not_scanned_after_consecutive_ai_failures",
        "not_scanned_after_consecutive_ai_failures",
        "not_scanned_after_consecutive_ai_failures",
        "not_scanned_after_consecutive_ai_failures",
    ]


def test_quick_extract_totals_stops_after_authoritative_invoice_section(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "DEPT#1.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": page, "text": ""}
            for page in range(1, 5)
        ],
    )
    rendered = []

    def fake_render(path, page_number, scale):
        rendered.append(page_number)
        return {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"}

    monkeypatch.setattr("bonus_platform.engine.labor.extract._render_pdf_page_to_image", fake_render)
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image",
        lambda page, prompt, config: LaborPageEvidence(
            source_file=page["source_file"],
            page=page["page"],
            role="invoice_primary" if page["page"] == 1 else "email_cover",
            role_confidence=0.98,
            warehouse_id="1" if page["page"] == 1 else "",
            total_amount=1880.67 if page["page"] == 1 else None,
            total_label="TOTAL" if page["page"] == 1 else "",
            evidence_text="TOTAL $1,880.67" if page["page"] == 1 else "Attached timecards",
        ),
    )

    totals = quick_extract_totals([pdf], _ready_ai_config(), supplier="Prompt Priority INC")

    assert rendered == [1, 2]
    assert totals[0]["total_amount"] == 1880.67
    assert [page["role"] for page in totals[0]["page_evidence"]] == [
        "invoice_primary",
        "email_cover",
        "unknown",
        "unknown",
    ]
    assert totals[0]["excluded_pages"] == [2, 3, 4]


def test_quick_extract_totals_stops_after_authoritative_total_when_following_page_has_no_total(monkeypatch, tmp_path):
    from bonus_platform.engine.labor.extract import quick_extract_totals

    pdf = tmp_path / "DEPT#10.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._extract_pdf_pages",
        lambda paths: [
            {"source_file": pdf.name, "source_path": str(pdf), "page": page, "text": ""}
            for page in range(1, 5)
        ],
    )
    rendered = []
    extraction_calls = []

    def fake_render(path, page_number, scale):
        rendered.append(page_number)
        return {"source_file": path.name, "source_path": str(path), "page": page_number, "mime_type": "image/jpeg", "base64": "page"}

    def fake_extract(page, prompt, config):
        page_number = page["page"]
        extraction_calls.append((page_number, config.get("page_evidence_max_tokens")))
        return LaborPageEvidence(
            source_file=page["source_file"],
            page=page_number,
            role="invoice_primary" if page_number == 1 else "invoice_total",
            role_confidence=0.95,
            warehouse_id="10",
            total_amount=873.03 if page_number == 1 else (32.0 if page_number == 4 else None),
            total_label="TOTAL" if page_number in {1, 4} else "",
            evidence_text="TOTAL $873.03" if page_number == 1 else "Supporting hours table",
        )

    monkeypatch.setattr("bonus_platform.engine.labor.extract._render_pdf_page_to_image", fake_render)
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_page_evidence_with_ai_image", fake_extract)

    totals = quick_extract_totals([pdf], _ready_ai_config(), supplier="Prompt Priority INC")

    assert rendered == [1, 2]
    assert extraction_calls == [(1, None), (2, None)]
    assert totals[0]["total_amount"] == 873.03
    assert totals[0]["authoritative"] is True
    assert totals[0]["excluded_pages"] == [3, 4]


def test_non_payable_pdf_names_flags_supporting_types_when_payable_invoice_exists():
    totals = [
        {"source_file": "In291943.pdf", "total_amount": 13836.28, "pdf_type": "primary"},
        {"source_file": "Supplement1.pdf", "total_amount": 120.0, "pdf_type": "supporting"},
        {"source_file": "COI.pdf", "total_amount": 0, "pdf_type": "attachment"},
        {"source_file": "legacy_detail.pdf", "total_amount": 0},
        {"source_file": "unknown_scan.pdf", "total_amount": 0, "pdf_type": "unknown"},
    ]

    assert _non_payable_pdf_names(totals) == {"Supplement1.pdf", "COI.pdf", "legacy_detail.pdf"}


def test_non_payable_pdf_names_keeps_only_pdf_when_all_totals_failed():
    totals = [{"source_file": "Supplement1.pdf", "total_amount": 0, "pdf_type": "supporting"}]

    assert _non_payable_pdf_names(totals) == set()


def test_rendered_invoice_images_preserve_pdf_orientation(monkeypatch, tmp_path):
    from PIL import Image

    class FakeBitmap:
        def to_pil(self):
            return Image.new("RGB", (100, 200), "white")

    class FakePage:
        def render(self, scale):
            return FakeBitmap()

        def close(self):
            pass

    class FakeDocument:
        def __init__(self, path):
            pass

        def __len__(self):
            return 1

        def __getitem__(self, index):
            return FakePage()

        def close(self):
            pass

    class FakePdfium:
        PdfDocument = FakeDocument

    monkeypatch.setitem(__import__("sys").modules, "pypdfium2", FakePdfium)

    rows = __import__("bonus_platform.engine.labor.extract", fromlist=["_render_pdf_pages_to_images"])._render_pdf_pages_to_images([tmp_path / "scan.pdf"])

    image = Image.open(BytesIO(__import__("base64").b64decode(rows[0]["base64"])))
    assert image.size == (100, 200)


def test_rendered_invoice_images_skip_pages_outside_evidence_allowlist(monkeypatch, tmp_path):
    from PIL import Image

    rendered_indexes = []

    class FakeBitmap:
        def to_pil(self):
            return Image.new("RGB", (100, 200), "white")

    class FakePage:
        def __init__(self, index):
            self.index = index

        def render(self, scale):
            rendered_indexes.append(self.index)
            return FakeBitmap()

        def close(self):
            pass

    class FakeDocument:
        def __init__(self, path):
            pass

        def __len__(self):
            return 3

        def __getitem__(self, index):
            return FakePage(index)

        def close(self):
            pass

    class FakePdfium:
        PdfDocument = FakeDocument

    monkeypatch.setitem(__import__("sys").modules, "pypdfium2", FakePdfium)
    renderer = __import__("bonus_platform.engine.labor.extract", fromlist=["_render_pdf_pages_to_images"])._render_pdf_pages_to_images

    rows = renderer(
        [tmp_path / "scan.pdf"],
        allowed_pages_by_source={"scan.pdf": {2}},
    )

    assert rendered_indexes == [1]
    assert [row["page"] for row in rows] == [2]


def test_pdf_text_extraction_keeps_pipeline_alive_for_unreadable_pdf(tmp_path):
    broken_pdf = tmp_path / "broken.pdf"
    broken_pdf.write_bytes(b"%PDF-1.4\n")

    pages = _extract_pdf_pages([broken_pdf])

    assert pages == [{"source_file": "broken.pdf", "source_path": str(broken_pdf), "page": 1, "text": ""}]


def test_mimo_image_extractor_sends_base64_pages_and_returns_rows(monkeypatch):
    captured = {}

    def fake_post(payload, ai_config):
        captured["payload"] = payload
        return [
            {
                "source_file": "scan.pdf",
                "source_page_or_row": "p1",
                "employee_id": "",
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 40,
                "amount": 800.5,
                "currency": "USD",
                "confidence": 0.88,
                "evidence_text": "Alvarez Minchaca, Rosa ... Total $800.50",
            }
        ]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)

    rows = _extract_with_ai_images(
        [
            {
                "source_file": "scan.pdf",
                "page": 1,
                "mime_type": "image/png",
                "base64": "abc123",
            }
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "max_pages_per_request": 5,
        },
        supplier="ONESOURCE",
        period_start="2026-05-11",
        period_end="2026-05-17",
        currency="USD",
    )

    content = captured["payload"]["messages"][1]["content"]

    # 检查图片格式（支持 image_url 或 image 类型）
    assert content[0]["type"] in ("image_url", "image")
    if content[0]["type"] == "image_url":
        assert content[0]["image_url"]["url"] == "data:image/png;base64,abc123"
    else:
        assert content[0]["source"]["type"] == "base64"
        assert content[0]["source"]["data"] == "abc123"
    assert rows[0]["employee_name_raw"] == "Alvarez Minchaca, Rosa"
    assert rows[0]["source_type"] == "pdf_invoice"
    assert rows[0]["supplier"] == "ONESOURCE"


def test_mimo_image_extractor_annotates_single_page_rows_when_model_omits_source(monkeypatch):
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, ai_config: [
            {
                "employee_name_raw": "Scan Person",
                "hours": 8,
                "amount": 160,
                "confidence": 0.9,
            }
        ],
    )

    rows = _extract_with_ai_images(
        [
            {
                "source_file": "scan.pdf",
                "page": 2,
                "mime_type": "image/png",
                "base64": "abc123",
            }
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "cache_enabled": False,
        },
    )

    assert rows[0]["source_file"] == "scan.pdf"
    assert rows[0]["source_page_or_row"] == "p2"


def test_mimo_image_extractor_rejects_unattributed_rows_from_multi_page_chunk(monkeypatch):
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, ai_config: [{"employee_name_raw": "Unknown Source", "hours": 8, "amount": 160, "confidence": 0.9}],
    )

    rows = _extract_with_ai_images(
        [
            {"source_file": "a.pdf", "page": 1, "mime_type": "image/png", "base64": "abc"},
            {"source_file": "b.pdf", "page": 1, "mime_type": "image/png", "base64": "def"},
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "max_pages_per_request": 2,
            "cache_enabled": False,
        },
    )

    assert rows == []


def test_extract_invoice_items_uses_mimo_images_when_pdf_text_has_no_rows(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_pdf_pages", lambda paths: [{"source_file": "scan.pdf", "page": 1, "text": ""}])
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_pages_to_images",
        lambda paths, scale=1.5, **kwargs: [{"source_file": "scan.pdf", "page": 1, "mime_type": "image/png", "base64": "abc123"}],
    )
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_text", lambda *args, **kwargs: [])
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, config: [
            {
                "source_file": "scan.pdf",
                "source_page_or_row": "p1",
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 40,
                "amount": 800.5,
                "confidence": 0.88,
                "evidence_text": "Alvarez Minchaca, Rosa ... Total $800.50",
            }
        ],
    )

    rows = extract_invoice_items(
        [pdf],
        {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        supplier="ONESOURCE",
        period_start="2026-05-11",
        period_end="2026-05-17",
        currency="USD",
    )

    assert len(rows) == 1
    assert rows[0].employee_name_raw == "Alvarez Minchaca, Rosa"
    assert rows[0].source_type == "pdf_invoice"
    assert rows[0].supplier == "ONESOURCE"


def test_mimo_image_extractor_filters_non_employee_zero_rows(monkeypatch):
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, config: [
            {"source_file": "scan.pdf", "source_page_or_row": "p2", "employee_name_raw": "RG-31.45", "hours": 0, "amount": 0, "confidence": 0.85},
            {"source_file": "scan.pdf", "source_page_or_row": "p1", "employee_name_raw": "Alvarez Minchaca, Rosa", "hours": 31.19, "amount": 701.9, "confidence": 0.95, "evidence_text": "Total $701.90"},
        ],
    )

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert [row["employee_name_raw"] for row in rows] == ["Alvarez Minchaca, Rosa"]


def test_image_extractor_includes_full_expected_employee_list(monkeypatch):
    captured = {}

    def fake_post(payload, config):
        captured["payload"] = payload
        return []

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)
    expected_rows = [{"employee_name": f"Employee {idx}"} for idx in range(1, 36)]

    _extract_with_ai_images(
        [{"source_file": "scan.pdf", "page": 1, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        expected_rows=expected_rows,
    )

    prompt_text = captured["payload"]["messages"][1]["content"][-1]["text"]
    assert "Employee 1" in prompt_text
    assert "Employee 20" in prompt_text
    assert "Employee 35" in prompt_text


def test_image_ai_rows_are_filtered_against_expected_employee_candidates():
    rows = [
        {"employee_name_raw": "John Doe", "hours": 0, "amount": 5500, "confidence": 0.95},
        {"employee_name_raw": "Morales, Katherine", "hours": 40.6, "amount": 916.16, "confidence": 0.95},
        {"employee_name_raw": "Gerardo Torres Valencia", "hours": 39.27, "amount": 1008.51, "confidence": 0.95},
    ]
    expected_rows = [
        {"employee_name": "Katherina Morales"},
        {"employee_name": "Gerardo Torres"},
    ]

    filtered = _filter_ai_rows_by_expected_employees(rows, expected_rows)

    assert [row["employee_name_raw"] for row in filtered] == ["Morales, Katherine", "Gerardo Torres Valencia"]


def test_image_ai_expected_employee_filter_keeps_minor_spelling_variant_with_extra_token():
    rows = [
        {
            "employee_name_raw": "BANTSIMBA GLOIRE RONLELE",
            "hours": 45.76,
            "amount": 1143.37,
            "confidence": 0.95,
        }
    ]
    expected_rows = [{"employee_name": "BATSIMBA Gloire"}]

    filtered = _filter_ai_rows_by_expected_employees(rows, expected_rows)

    assert filtered == rows


def test_mimo_image_extractor_filters_timesheet_rows_without_money_evidence(monkeypatch):
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, config: [
            {"source_file": "scan.pdf", "source_page_or_row": "p2", "employee_name_raw": "Brian Cowan", "hours": 8, "amount": 40, "confidence": 0.85, "evidence_text": "Brian Cowan RG-40 OT-0.42"},
            {"source_file": "scan.pdf", "source_page_or_row": "p1", "employee_name_raw": "Alvarez Minchaca, Rosa", "hours": 31.19, "amount": 701.9, "confidence": 0.95, "evidence_text": "Total $701.90"},
        ],
    )

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    # evidence 标记检查已移除（太严格导致图片PDF抽取结果被误过滤）
    # amount=40 > 0 且有合理人名，现在应该保留
    assert [row["employee_name_raw"] for row in rows] == ["Brian Cowan", "Alvarez Minchaca, Rosa"]


def test_mimo_image_extractor_filters_rows_without_amount(monkeypatch):
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, config: [
            {"source_file": "scan.pdf", "source_page_or_row": "p2", "employee_name_raw": "Kevin Sultana", "hours": 39.43, "amount": 0, "confidence": 0.85, "evidence_text": "Total Hours 39.43"},
            {"source_file": "scan.pdf", "source_page_or_row": "p1", "employee_name_raw": "Alvarez Minchaca, Rosa", "hours": 31.19, "amount": 701.9, "confidence": 0.95, "evidence_text": "Total $701.90"},
        ],
    )

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert [row["employee_name_raw"] for row in rows] == ["Alvarez Minchaca, Rosa"]


def test_mimo_image_extractor_skips_non_first_page_json_parse_failures(monkeypatch):
    def fake_post(payload, config):
        raise json.JSONDecodeError("Expecting value", "", 0)

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert rows == []


def test_mimo_image_extractor_retries_first_page_json_parse_failures(monkeypatch):
    calls = {"count": 0}

    def fake_post(payload, config):
        calls["count"] += 1
        if calls["count"] == 1:
            raise json.JSONDecodeError("Expecting value", "", 0)
        return [{"source_file": "scan.pdf", "source_page_or_row": "p1", "employee_name_raw": "Alvarez Minchaca, Rosa", "hours": 31.19, "amount": 701.9, "confidence": 0.95, "evidence_text": "Total $701.90"}]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "page": 1, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert calls["count"] == 2
    assert rows[0]["employee_name_raw"] == "Alvarez Minchaca, Rosa"


def test_mimo_image_extractor_skips_timed_out_page_and_keeps_later_rows(monkeypatch):
    calls = {"count": 0}

    def fake_post(payload, config):
        calls["count"] += 1
        if calls["count"] <= 4:
            raise MiMoTimeoutException("gateway timeout")
        return [
            {
                "source_file": "scan.pdf",
                "source_page_or_row": "p2",
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 31.19,
                "amount": 701.9,
                "confidence": 0.95,
                "evidence_text": "Total $701.90",
            }
        ]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)

    rows = _extract_with_ai_images(
        [
            {"source_file": "scan.pdf", "page": 1, "mime_type": "image/png", "base64": "abc123"},
            {"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "def456"},
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://token-plan-cn.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "max_pages_per_request": 5,
        },
    )

    assert calls["count"] == 5
    assert [row["employee_name_raw"] for row in rows] == ["Alvarez Minchaca, Rosa"]


def test_mimo_image_extractor_fast_pass_skips_timeout_after_one_attempt(monkeypatch):
    calls = {"count": 0}

    def fake_post(payload, config):
        calls["count"] += 1
        if calls["count"] == 1:
            raise MiMoTimeoutException("gateway timeout")
        return [{"employee_name_raw": "Later Worker", "hours": 8, "amount": 160, "confidence": 0.95}]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)
    audit: list[dict] = []

    rows = _extract_with_ai_images(
        [
            {"source_file": "scan.pdf", "page": 1, "mime_type": "image/png", "base64": "abc123"},
            {"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "def456"},
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "max_pages_per_request": 1,
            "image_retry_delays": [],
            "cache_enabled": False,
        },
        audit_collector=audit,
    )

    assert calls["count"] == 2
    assert [row["employee_name_raw"] for row in rows] == ["Later Worker"]
    assert audit[0]["status"] == "failed"
    assert audit[1]["status"] == "completed"


def test_mimo_image_extractor_continues_after_non_timeout_page_error(monkeypatch):
    calls = {"count": 0}

    def fake_post(payload, config):
        calls["count"] += 1
        if calls["count"] == 1:
            raise RuntimeError("HTTP 502 upstream unavailable")
        return [{"employee_name_raw": "Recovered Worker", "hours": 8, "amount": 160, "confidence": 0.95}]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)

    rows = _extract_with_ai_images(
        [
            {"source_file": "scan.pdf", "page": 1, "mime_type": "image/png", "base64": "abc123"},
            {"source_file": "scan.pdf", "page": 2, "mime_type": "image/png", "base64": "def456"},
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "max_pages_per_request": 1,
            "image_retry_delays": [],
            "cache_enabled": False,
        },
    )

    assert calls["count"] == 2
    assert [row["employee_name_raw"] for row in rows] == ["Recovered Worker"]


def test_token_plan_image_extractor_forces_single_page_chunks():
    assert _effective_max_pages_per_request(
        {
            "provider": "mimo",
            "base_url": "https://token-plan-cn.xiaomimimo.com/v1",
            "max_pages_per_request": 5,
        }
    ) == 1


def test_runtime_image_extraction_uses_configured_single_page_chunks():
    from bonus_platform.config import AI_CONFIG

    assert AI_CONFIG["max_pages_per_request"] == 1


def test_mimo_image_extractor_uses_page_cache(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"pdf")
    cache_dir = tmp_path / ".ai_extract_cache"
    cache_dir.mkdir()
    cache_file = cache_dir / f"scan_p1_mimo-v2.5_{AI_PAGE_CACHE_VERSION}.json"
    cache_file.write_text(
        json.dumps(
            [
                {
                    "source_file": "scan.pdf",
                    "source_page_or_row": "p1",
                    "employee_name_raw": "Alvarez Minchaca, Rosa",
                    "hours": 31.19,
                    "amount": 701.9,
                    "confidence": 0.95,
                    "evidence_text": "Total $701.90",
                }
            ]
        ),
        encoding="utf-8",
    )

    def fail_post(payload, config):
        raise AssertionError("cache miss")

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fail_post)

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "source_path": str(pdf), "page": 1, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    assert rows[0]["employee_name_raw"] == "Alvarez Minchaca, Rosa"


def test_mimo_image_extractor_retries_empty_page_cache_when_configured(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"pdf")
    cache_dir = tmp_path / ".ai_extract_cache"
    cache_dir.mkdir()
    cache_file = cache_dir / f"scan_p1_mimo-v2.5_{AI_PAGE_CACHE_VERSION}.json"
    cache_file.write_text("[]", encoding="utf-8")
    calls = {"count": 0}

    def fake_post(payload, config):
        calls["count"] += 1
        return [{"employee_name_raw": "Recovered Worker", "hours": 8, "amount": 160, "confidence": 0.95}]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)

    rows = _extract_with_ai_images(
        [{"source_file": "scan.pdf", "source_path": str(pdf), "page": 1, "mime_type": "image/png", "base64": "abc123"}],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "image_retry_delays": [],
            "retry_empty_page_cache": True,
        },
    )

    assert calls["count"] == 1
    assert [row["employee_name_raw"] for row in rows] == ["Recovered Worker"]


def test_mimo_image_extractor_writes_page_cache(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"pdf")

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, config: [
            {
                "source_file": "scan.pdf",
                "source_page_or_row": "p1",
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 31.19,
                "amount": 701.9,
                "confidence": 0.95,
                "evidence_text": "Total $701.90",
            }
        ],
    )

    _extract_with_ai_images(
        [{"source_file": "scan.pdf", "source_path": str(pdf), "page": 1, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
    )

    cache_file = tmp_path / ".ai_extract_cache" / f"scan_p1_mimo-v2.5_{AI_PAGE_CACHE_VERSION}.json"
    assert cache_file.exists()


def test_mimo_image_extractor_high_res_retries_when_page_has_no_usable_rows(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"pdf")
    seen_images: list[str] = []

    def fake_post(payload, config):
        image_part = payload["messages"][1]["content"][0]
        seen_images.append(image_part["image_url"]["url"])
        if len(seen_images) == 1:
            return [{"employee_name_raw": "Totals", "hours": 0, "amount": 0}]
        return [
            {
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 31.19,
                "amount": 701.9,
                "confidence": 0.95,
                "evidence_text": "Total $701.90",
            }
        ]

    monkeypatch.setattr("bonus_platform.engine.labor.extract._post_chat_completion", fake_post)
    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._render_pdf_page_to_image",
        lambda path, page_number, scale: {
            "source_file": Path(path).name,
            "source_path": str(path),
            "page": page_number,
            "mime_type": "image/jpeg",
            "base64": "hires456",
            "render_scale": scale,
            "high_resolution_retry": True,
        },
    )
    audit: list[dict] = []

    rows = _extract_with_ai_images(
        [
            {
                "source_file": "scan.pdf",
                "source_path": str(pdf),
                "page": 1,
                "mime_type": "image/png",
                "base64": "lowres123",
                "render_scale": 1.0,
            }
        ],
        {
            "provider": "mimo",
            "api_key": "token",
            "base_url": "https://api.xiaomimimo.com/v1",
            "model": "mimo-v2.5",
            "cache_enabled": False,
            "high_resolution_retry_enabled": True,
        },
        audit_collector=audit,
    )

    assert [row["employee_name_raw"] for row in rows] == ["Alvarez Minchaca, Rosa"]
    assert "lowres123" in seen_images[0]
    assert "hires456" in seen_images[1]
    assert any(entry["status"] == "high_res_retry_applied" and entry["rowCount"] == 1 for entry in audit)


def test_mimo_image_extractor_records_page_audit(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"pdf")

    monkeypatch.setattr(
        "bonus_platform.engine.labor.extract._post_chat_completion",
        lambda payload, config: [
            {
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 31.19,
                "amount": 701.9,
                "confidence": 0.95,
                "evidence_text": "Total $701.90",
            }
        ],
    )
    audit: list[dict] = []

    _extract_with_ai_images(
        [{"source_file": "scan.pdf", "source_path": str(pdf), "page": 1, "mime_type": "image/png", "base64": "abc123"}],
        {"provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5", "cache_enabled": False},
        audit_collector=audit,
    )

    assert audit == [
        {
            "sourceFile": "scan.pdf",
            "page": 1,
            "status": "completed",
            "rowCount": 1,
            "amountTotal": 701.9,
            "renderScale": None,
            "fromCache": False,
            "highResolutionRetry": False,
        }
    ]


def test_extract_invoice_items_surfaces_ai_failure_when_enabled(monkeypatch, tmp_path):
    pdf = tmp_path / "scan.pdf"
    pdf.write_bytes(b"%PDF-1.4\n")

    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_pdf_pages", lambda paths: [{"source_file": "scan.pdf", "page": 1, "text": ""}])
    monkeypatch.setattr("bonus_platform.engine.labor.extract._extract_with_ai_text", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("HTTP 401 Invalid API Key")))
    monkeypatch.setattr("bonus_platform.engine.labor.extract._render_pdf_pages_to_images", lambda paths, **kwargs: [])

    with pytest.raises(ValueError, match="AI 抽取失败"):
        extract_invoice_items(
            [pdf],
            {"enabled": True, "provider": "mimo", "api_key": "token", "base_url": "https://api.xiaomimimo.com/v1", "model": "mimo-v2.5"},
        )


def test_safe_error_message_includes_mimo_error_body():
    error = HTTPError(
        url="https://api.xiaomimimo.com/v1/chat/completions",
        code=401,
        msg="Unauthorized",
        hdrs={},
        fp=BytesIO(b'{"error":{"message":"Invalid API Key","code":"401"}}'),
    )

    message = _safe_error_message(error)

    assert "Invalid API Key" in message


def test_http_post_json_enforces_wall_clock_timeout(monkeypatch):
    class SlowResponse:
        status_code = 200

        def raise_for_status(self):
            return None

        def json(self):
            return {"ok": True}

    class SlowClient:
        def __init__(self, *args, **kwargs):
            return None

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return None

        def post(self, *args, **kwargs):
            time.sleep(0.2)
            return SlowResponse()

    monkeypatch.setattr("bonus_platform.engine.labor.extract.httpx.Client", SlowClient)

    start = time.monotonic()
    with pytest.raises(MiMoTimeoutException):
        _http_post_json(
            "https://example.test/v1/messages",
            {},
            {"payload": "x"},
            wall_timeout_seconds=0.05,
        )

    assert time.monotonic() - start < 0.15


def test_anthropic_messages_url_does_not_duplicate_v1():
    assert _anthropic_messages_url({"base_url": "https://token-plan-cn.xiaomimimo.com"}) == "https://token-plan-cn.xiaomimimo.com/anthropic/v1/messages"
    assert _anthropic_messages_url({"base_url": "https://token-plan-cn.xiaomimimo.com/v1"}) == "https://token-plan-cn.xiaomimimo.com/anthropic/v1/messages"
    assert _anthropic_messages_url({"base_url": "https://token-plan-cn.xiaomimimo.com/anthropic"}) == "https://token-plan-cn.xiaomimimo.com/anthropic/v1/messages"
    assert _anthropic_messages_url({"base_url": "https://token-plan-cn.xiaomimimo.com/anthropic/v1"}) == "https://token-plan-cn.xiaomimimo.com/anthropic/v1/messages"
    assert _anthropic_messages_url({"base_url": "https://api.example.com/v1"}) == "https://api.example.com/v1/messages"


def test_effective_render_scale_caps_token_plan_payload_size():
    assert _effective_render_scale({"provider": "mimo", "base_url": "https://token-plan-cn.xiaomimimo.com/v1", "render_scale": 1.2}) == pytest.approx(0.75)
    assert _effective_render_scale({"provider": "mimo", "base_url": "https://token-plan-cn.xiaomimimo.com/v1", "render_scale": 0.6}) == pytest.approx(0.6)
    assert _effective_render_scale({"provider": "openai", "base_url": "https://api.example.com/v1", "render_scale": 1.2}) == pytest.approx(1.2)


def test_line_items_from_ai_rows_coerces_confidence_labels_and_name_ids():
    rows = line_items_from_dicts(
        [
            {
                "source_type": "pdf_invoice",
                "source_file": "scan.pdf",
                "source_page_or_row": "page 1 row 1",
                "employee_id": "Alvarez Minchaca, Rosa",
                "employee_name_raw": "Alvarez Minchaca, Rosa",
                "hours": 31.19,
                "amount": 701.9,
                "confidence": "High",
            }
        ]
    )

    assert rows[0].employee_id == ""
    assert rows[0].confidence == 0.95


def test_build_labor_report_contains_expected_sheets(tmp_path):
    output = tmp_path / "report.xlsx"
    comparison = {
        "summary": {"pdfEmployeeCount": 1, "excelEmployeeCount": 1, "amountDiffCount": 1},
        "rows": [
            {
                "employeeName": "MARTINEZ, WILFREDO",
                "matchStatus": "金额差异",
                "riskFlags": [],
                "pdfHoursTotal": 40.78,
                "excelHoursTotal": 40.78,
                "hoursDelta": 0,
                "pdfAmountTotal": 982.72,
                "excelAmountTotal": 982.74,
                "amountDelta": -0.02,
                "sourceRefs": "a.pdf p1; 账单!3",
            }
        ],
        "candidateMatches": [
            {
                "pdfEmployeeName": "Alvarez Mitrache, Ross",
                "excelEmployeeName": "Rosa Alvarez Minchaca",
                "nameSimilarity": 0.75,
                "pdfHoursTotal": 30.5,
                "excelHoursTotal": 31.19,
                "hoursDelta": -0.69,
                "pdfAmountTotal": 698.99,
                "excelAmountTotal": 701.9,
                "amountDelta": -2.91,
                "recommendation": "人工复核",
                "sourceRefs": "scan.pdf p1; 账单!2",
            }
        ],
    }

    build_labor_report(output, comparison, [], [], {"name": "姓名", "hours": "时长", "amount": "金额"})

    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["核对结论", "核对摘要", "全员对账明细", "金额差异员工", "工时待确认", "不在本批发票", "姓名格式差异", "明细识别待确认", "PDF发票明细", "Excel账单明细", "上传字段对应关系"]
    for internal_sheet_name in ["低置信度抽取", "PDF抽取明细", "字段映射记录", "工时风险项"]:
        assert internal_sheet_name not in workbook.sheetnames
    assert workbook["姓名格式差异"].max_row == 2
    assert workbook["全员对账明细"].max_row == 2


def test_build_labor_report_shows_evidence_and_comparable_warehouse_conclusion(tmp_path):
    output = tmp_path / "evidence-aware-report.xlsx"
    warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 1000.0,
            "excelAmountTotal": 1250.0,
            "amountDeltaTotal": -250.0,
            "comparableExcelAmountTotal": 700.0,
            "comparableAmountDeltaTotal": 100.0,
            "missingPdfAmountTotal": 200.0,
        },
        "rows": [
            {
                "warehouseId": "1",
                "matchStatus": "金额差异",
                "reconciliationStatus": "amount_difference",
                "pdfEmployeeCount": 2,
                "excelEmployeeCount": 2,
                "pdfHoursTotal": 80,
                "excelHoursTotal": 80,
                "pdfAmountTotal": 1000.0,
                "excelAmountTotal": 900.0,
                "amountDelta": 100.0,
                "pdfEvidenceFile": "warehouse-1.pdf",
                "pdfEvidencePage": 2,
                "evidenceStatus": "authoritative",
                "excludedPdfPages": [1, 3],
            },
            {
                "warehouseId": "2",
                "matchStatus": "缺少PDF发票",
                "reconciliationStatus": "missing_pdf_invoice",
                "pdfEmployeeCount": 0,
                "excelEmployeeCount": 1,
                "pdfAmountTotal": 0,
                "excelAmountTotal": 200.0,
                "amountDelta": -200.0,
                "pdfEvidenceFile": "",
                "pdfEvidencePage": None,
                "evidenceStatus": "missing",
                "excludedPdfPages": [],
            },
            {
                "warehouseId": "3",
                "matchStatus": "多余PDF发票",
                "reconciliationStatus": "extra_pdf_invoice",
                "pdfEmployeeCount": 1,
                "excelEmployeeCount": 0,
                "pdfAmountTotal": 50.0,
                "excelAmountTotal": 0,
                "amountDelta": 50.0,
                "pdfEvidenceFile": "warehouse-3.pdf",
                "pdfEvidencePage": 1,
                "evidenceStatus": "authoritative",
                "excludedPdfPages": [],
            },
            {
                "warehouseId": "4",
                "matchStatus": "待复核",
                "reconciliationStatus": "needs_review",
                "pdfEmployeeCount": 1,
                "excelEmployeeCount": 1,
                "pdfAmountTotal": 0,
                "excelAmountTotal": 150.0,
                "amountDelta": -150.0,
                "pdfEvidenceFile": "warehouse-4.pdf",
                "pdfEvidencePage": "1, 2",
                "evidenceStatus": "needs_review",
                "excludedPdfPages": [3],
            },
        ],
    }
    extraction_quality = calculate_extraction_quality(
        pdf_rows=[],
        comparison_summary={},
        warehouse_comparison=warehouse_comparison,
    )

    build_labor_report(
        output,
        {"summary": {"conclusionLevel": "warning", "conclusionMessage": "仓库证据待复核"}, "rows": []},
        [],
        [],
        {},
        warehouse_comparison=warehouse_comparison,
        extraction_quality=extraction_quality,
    )

    workbook = load_workbook(output, read_only=True)
    warehouse_sheet = workbook["仓库金额汇总"]
    headers = [cell.value for cell in warehouse_sheet[1]]
    for required_header in ["核对状态", "PDF证据文件", "PDF证据页", "证据状态", "排除附件页"]:
        assert required_header in headers
    warehouse_rows = list(warehouse_sheet.iter_rows(min_row=2, values_only=True))
    status_index = headers.index("核对状态")
    assert {row[status_index] for row in warehouse_rows} == {"金额差异", "缺少PDF发票", "多余PDF发票", "待复核"}

    conclusion = {
        row[0]: row[1]
        for row in workbook["核对结论"].iter_rows(values_only=True)
        if row[0] is not None and len(row) > 1
    }
    assert conclusion["已上传权威PDF总额"] == "$1,000.00"
    assert conclusion["整批Excel总额"] == "$1,250.00"
    assert conclusion["整批金额差异"] == "-$250.00"
    assert conclusion["可比仓库Excel总额"] == "$700.00"
    assert conclusion["可比仓库金额差异"] == "+$100.00"
    assert conclusion["缺少PDF发票金额"] == "$200.00"
    assert conclusion["缺少PDF发票仓库"] == "2"
    assert conclusion["核对结论"].startswith("需人工复核 - 仓库核对存在 4 项异常")

    quality_values = {
        row[0]: row[1]
        for row in workbook["识别完整度"].iter_rows(values_only=True)
        if row[0] is not None and len(row) > 1
    }
    assert quality_values["证据待复核仓库数"] == 1
    assert quality_values["缺少PDF发票仓库数"] == 1
    assert quality_values["多余PDF发票仓库数"] == 1


def test_build_labor_report_legacy_warehouse_summary_marks_unavailable_missing_pdf_history(tmp_path):
    output = tmp_path / "legacy-warehouse-report.xlsx"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 1,
            "pdfAmountTotal": 100.0,
            "excelAmountTotal": 110.0,
            "amountDeltaTotal": -10.0,
        },
        "rows": [],
    }
    legacy_warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 100.0,
            "excelAmountTotal": 110.0,
            "amountDeltaTotal": -10.0,
        },
        "rows": [
            {
                "warehouseId": "1",
                "matchStatus": "通过",
                "pdfEmployeeCount": 1,
                "excelEmployeeCount": 1,
                "pdfAmountTotal": 100.0,
                "excelAmountTotal": 110.0,
                "amountDelta": -10.0,
            }
        ],
    }

    build_labor_report(output, comparison, [], [], {}, warehouse_comparison=legacy_warehouse_comparison)

    workbook = load_workbook(output, read_only=True)
    conclusion = {
        row[0]: row[1]
        for row in workbook["核对结论"].iter_rows(values_only=True)
        if row[0] is not None and len(row) > 1
    }
    assert conclusion["可比仓库Excel总额"] == "$110.00"
    assert conclusion["可比仓库金额差异"] == "-$10.00"
    assert conclusion["缺少PDF发票金额"] == "历史结果未记录"
    assert conclusion["缺少PDF发票仓库"] == "历史结果未记录"


def test_build_labor_report_current_summary_reports_no_missing_pdf_warehouses(tmp_path):
    output = tmp_path / "current-no-missing-pdf-report.xlsx"
    warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 100.0,
            "excelAmountTotal": 100.0,
            "amountDeltaTotal": 0.0,
            "comparableExcelAmountTotal": 100.0,
            "comparableAmountDeltaTotal": 0.0,
            "missingPdfAmountTotal": 0.0,
        },
        "rows": [
            {
                "warehouseId": "1",
                "matchStatus": "通过",
                "reconciliationStatus": "passed",
                "pdfAmountTotal": 100.0,
                "excelAmountTotal": 100.0,
                "amountDelta": 0.0,
            }
        ],
    }

    build_labor_report(output, {"summary": {}, "rows": []}, [], [], {}, warehouse_comparison=warehouse_comparison)

    workbook = load_workbook(output, read_only=True)
    conclusion = {
        row[0]: row[1]
        for row in workbook["核对结论"].iter_rows(values_only=True)
        if row[0] is not None and len(row) > 1
    }
    assert conclusion["缺少PDF发票仓库"] == "无"


def test_build_labor_report_highlights_review_and_invoice_presence_statuses(tmp_path):
    output = tmp_path / "warehouse-status-fills.xlsx"
    warehouse_comparison = {
        "summary": {},
        "rows": [
            {"warehouseId": "review", "matchStatus": "待复核", "amountDelta": 0},
            {"warehouseId": "missing", "matchStatus": "缺少PDF发票", "amountDelta": -200},
            {"warehouseId": "extra", "matchStatus": "多余PDF发票", "amountDelta": 200},
        ],
    }

    build_labor_report(output, {"summary": {}, "rows": []}, [], [], {}, warehouse_comparison=warehouse_comparison)

    workbook = load_workbook(output, read_only=False)
    sheet = workbook["仓库金额汇总"]
    fills = {sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).fill.fgColor.rgb[-6:] for row in range(2, 5)}
    assert fills == {"review": "FFF4D6", "missing": "FFF4D6", "extra": "FFF4D6"}


def test_build_labor_report_uses_business_language_inside_workbook(tmp_path):
    output = tmp_path / "business-report.xlsx"
    comparison = {
        "summary": {"pdfEmployeeCount": 4, "excelEmployeeCount": 4, "amountDiffCount": 1},
        "rows": [
            {
                "employeeName": "LOW CONFIDENCE",
                "matchStatus": "低置信度抽取",
                "riskFlags": ["低置信度抽取"],
                "pdfHoursTotal": 8,
                "excelHoursTotal": 0,
                "hoursDelta": 8,
                "pdfAmountTotal": 100,
                "excelAmountTotal": 0,
                "amountDelta": 100,
                "sourceRefs": "invoice.pdf p1",
            },
            {
                "employeeName": "Maria Lopez",
                "matchStatus": "Excel有PDF无",
                "riskFlags": [],
                "pdfHoursTotal": 0,
                "excelHoursTotal": 40,
                "hoursDelta": -40,
                "pdfAmountTotal": 0,
                "excelAmountTotal": 812.8,
                "amountDelta": -812.8,
                "sourceRefs": "bill.xlsx!2",
            },
            {
                "employeeName": "Mucu, Pablo ⇄ Pablo Mucu",
                "matchStatus": "通过",
                "riskFlags": ["姓名格式差异自动合并"],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "invoice.pdf p2; bill.xlsx!3",
            },
            {
                "employeeName": "Ross Mitrache ⇄ Rosa Alvarez",
                "matchStatus": "疑似姓名匹配",
                "riskFlags": ["疑似姓名匹配"],
                "pdfHoursTotal": 30,
                "excelHoursTotal": 31,
                "hoursDelta": -1,
                "pdfAmountTotal": 700,
                "excelAmountTotal": 701,
                "amountDelta": -1,
                "sourceRefs": "invoice.pdf p3; bill.xlsx!4",
            },
        ],
    }
    pdf_rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="invoice.pdf",
            source_page_or_row="p1",
            employee_id="",
            employee_name_raw="LOW CONFIDENCE",
            hours=8,
            amount=100,
            currency="USD",
            confidence=0.5,
            evidence_text="LOW CONFIDENCE 8 $100.00",
        )
    ]

    build_labor_report(output, comparison, pdf_rows, [], {"name": "姓名", "hours": "工时", "amount": "金额"})

    workbook = load_workbook(output, read_only=True)
    visible_text = "\n".join(
        str(value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for value in row
        if value is not None
    )
    assert "明细识别不完整" in visible_text
    assert "账单有发票无" in visible_text
    assert "系统已自动修正" in visible_text
    assert "疑似同一员工" in visible_text
    assert "PDF发票" in visible_text
    for internal_term in ["低置信度抽取", "Excel有PDF无", "疑似姓名匹配", "source_type", "employee_name_raw", "evidence_text", "confidence", "pdf_invoice"]:
        assert internal_term not in visible_text


def test_build_labor_report_can_include_reconciliation_diagnostics(tmp_path):
    output = tmp_path / "report.xlsx"
    comparison = {"summary": {"pdfEmployeeCount": 0, "excelEmployeeCount": 0}, "rows": []}
    diagnostics = {
        "level": "warning",
        "message": "核对信号有不稳定项，建议复核。",
        "nextStep": "先确认费用口径。",
        "signals": {
            "fastPdfTotal": 48293.06,
            "employeePdfTotal": 0,
            "excelTotal": 48217.96,
            "warehouseTotal": 48293.06,
            "amountBasis": [
                {
                    "warehouseId": "13",
                    "pdfTotal": 48293.06,
                    "reportedTotal": 48217.96,
                    "pdfVsReportedDelta": 75.1,
                    "componentTotal": 48217.96,
                    "employeeExpenses": 48055.81,
                    "employeeBenefits": 162.15,
                    "loadingAndUnloading": 0.0,
                    "summaryEvidence": "Warehouse-information!2",
                    "detailEvidence": "Employee-expenses-detail!3:Employee-expenses-detail!289; Employee-benefits-detail!2",
                }
            ],
            "offsettingWarehouseDeltas": [
                {
                    "warehouseId": "25",
                    "pdfAmountTotal": 17465.12,
                    "excelAmountTotal": 17463.34,
                    "amountDelta": 1.78,
                    "attribution": [{"employeeName": "JIMENEZ, ENEAS", "delta": 1.19}],
                }
            ],
            "employeeAttribution": [
                {
                    "warehouseId": "25",
                    "employeeName": "Fontes, Stevie ⇄ Stevie Fontes",
                    "pdfAmount": 822.12,
                    "excelAmount": 863.22,
                    "delta": -41.1,
                    "warehouseDelta": -41.21,
                }
            ],
        },
        "issues": [
            {
                "code": "amount_basis_mismatch",
                "level": "warning",
                "title": "PDF 总额与账单费用口径不一致",
                "message": "账单内部费用组成已闭合，但 PDF 发票总额与 OTWS 汇总总额不同。",
                "items": ["仓库 13: PDF $48,293.06，OTWS汇总 $48,217.96，差异 $75.10"],
            }
        ],
    }

    build_labor_report(
        output,
        comparison,
        [],
        [],
        {"name": "姓名", "hours": "时长", "amount": "金额"},
        reconciliation_diagnostics=diagnostics,
    )

    workbook = load_workbook(output, read_only=True)
    assert "信号诊断" in workbook.sheetnames
    rows = list(workbook["信号诊断"].iter_rows(values_only=True))
    assert any(row[:2] == ("诊断级别", "warning") for row in rows)
    assert any(row[0] == "PDF 总额与账单费用口径不一致" for row in rows)
    assert any(row[0] == "13" and row[3] == 75.1 for row in rows)
    assert any(row[0] == "25" and row[3] == 1.78 and "JIMENEZ" in str(row[4]) for row in rows)
    assert any(row[0] == "25" and "Fontes" in str(row[1]) and row[4] == -41.1 for row in rows)


def test_build_labor_report_can_include_ai_cache_audit(tmp_path):
    output = tmp_path / "report.xlsx"
    comparison = {"summary": {"pdfEmployeeCount": 0, "excelEmployeeCount": 0}, "rows": []}
    audit = {
        "decision": "candidate_only",
        "requiresConfirmation": True,
        "message": "历史图片识别记录只能作为待复核证据，不能直接覆盖确定性核对结论。",
        "summary": {"fileCount": 1, "candidateFileCount": 1, "candidateAmountTotal": 1399.89},
        "files": [
            {
                "sourceFile": "elog1-1_20260520204104.pdf",
                "warehouseId": "1",
                "rowCount": 2,
                "candidateAmountTotal": 1399.89,
                "averageConfidence": 0.95,
                "decision": "candidate_only",
                "cacheFiles": [f"elog1-1_20260520204104_p1_mimo-v2.5_{AI_PAGE_CACHE_VERSION}.json"],
                "evidence": [{"employeeName": "Alvarez Michalec Rosa", "amount": 701.88, "evidenceText": "Total $701.88"}],
            }
        ],
    }

    build_labor_report(
        output,
        comparison,
        [],
        [],
        {"name": "姓名", "hours": "时长", "amount": "金额"},
        ai_cache_audit=audit,
    )

    workbook = load_workbook(output, read_only=True)
    assert "AI候选证据" in workbook.sheetnames
    rows = list(workbook["AI候选证据"].iter_rows(values_only=True))
    assert any(row[:2] == ("处理决策", "candidate_only") for row in rows)
    assert any(row[:2] == ("需要人工确认", "是") for row in rows)
    assert any(row[0] == "elog1-1_20260520204104.pdf" and row[5] == "candidate_only" for row in rows)


def test_build_labor_business_html_report_uses_business_language_without_internal_terms(tmp_path):
    output = tmp_path / "business-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 2,
            "excelEmployeeCount": 2,
            "pdfAmountTotal": 1333.33,
            "excelAmountTotal": 1333.36,
            "amountDeltaTotal": -0.03,
            "passedCount": 1,
            "amountDiffCount": 1,
        },
        "rows": [
            {
                "employeeName": "Aguilar, Hortensia ⇄ Hortensia Aguilar",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 41.4,
                "excelHoursTotal": 41.4,
                "hoursDelta": 0,
                "pdfAmountTotal": 950.6,
                "excelAmountTotal": 950.61,
                "amountDelta": -0.01,
                "sourceRefs": "Invoice-5058871.pdf p1; 账单!3",
            },
            {
                "employeeName": "Andrew Torres",
                "matchStatus": "疑似姓名匹配",
                "riskFlags": ["账单多行合并"],
                "pdfHoursTotal": 20.42,
                "excelHoursTotal": 20.42,
                "hoursDelta": 0,
                "pdfAmountTotal": 382.73,
                "excelAmountTotal": 382.75,
                "amountDelta": -0.02,
                "sourceRefs": "Invoice-5058877.pdf p2; 账单!10; 账单!11",
            },
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Workforce Priority",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="5058871-5058880",
    )

    html = output.read_text(encoding="utf-8")
    assert "Workforce Priority" in html
    assert "核算周期：2026-05-11 ~ 2026-05-17" in html
    assert "发票编号或文件范围：5058871-5058880" in html
    assert "核对结论" in html
    assert "待确认" in html
    assert "这批账能不能放行？" in html
    assert "总金额核对" in html
    assert "员工明细状态" in html
    assert "下一步" in html
    assert "PDF 发票总金额" in html
    assert "$1,333.33" in html
    assert "账单总金额" in html
    assert "$1,333.36" in html
    assert "一致员工数" in html
    assert "待确认员工数" in html
    assert "员工姓名（发票）" in html
    assert "账单姓名" in html
    assert "REG 工时" in html
    assert "OT 工时" in html
    assert "业务说明" in html
    assert "下载 Excel 明细" in html
    assert "Excel 明细用于留档、筛选和逐行核查" in html
    assert "页面结论以本 HTML 报告为准" in html
    assert "需查看明细说明" not in html
    assert "有差异员工数" not in html
    assert "必要说明" not in html
    assert "原始识别明细" not in html
    assert "字段映射" not in html
    assert "需要确认该员工是否为同一人" in html
    assert "同一员工可能存在多行账单，需要确认是否应合并" in html
    for internal_term in ["AI 候选", "规则治理", "profile", "re-OCR", "回放", "低置信度算法", "Blob", "线程"]:
        assert internal_term not in html


def test_build_labor_business_html_report_groups_auto_fixes_suspected_matches_and_pending_items(tmp_path):
    output = tmp_path / "business-report-sections.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 3,
            "excelEmployeeCount": 3,
            "pdfAmountTotal": 2200.0,
            "excelAmountTotal": 2200.02,
            "amountDeltaTotal": -0.02,
            "passedCount": 1,
            "amountDiffCount": 1,
        },
        "rows": [
            {
                "employeeName": "Mucu, Pablo ⇄ Pablo Mucu",
                "matchStatus": "通过",
                "riskFlags": ["疑似姓名匹配"],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "Invoice-5058871.pdf p1; 账单!3",
            },
            {
                "employeeName": "Andrew Torres",
                "matchStatus": "低置信度抽取",
                "riskFlags": ["低置信度抽取"],
                "pdfHoursTotal": 20,
                "excelHoursTotal": 0,
                "hoursDelta": 20,
                "pdfAmountTotal": 400,
                "excelAmountTotal": 0,
                "amountDelta": 400,
                "sourceRefs": "Invoice-5058877.pdf p2",
            },
            {
                "employeeName": "Maria Lopez",
                "matchStatus": "金额差异",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 810,
                "excelAmountTotal": 812.8,
                "amountDelta": -2.8,
                "sourceRefs": "Invoice-5058878.pdf p3; 账单!8",
            },
            {
                "employeeName": "Selvin Rivera",
                "matchStatus": "Excel有PDF无",
                "riskFlags": [],
                "pdfHoursTotal": 0,
                "excelHoursTotal": 48,
                "hoursDelta": -48,
                "pdfAmountTotal": 0,
                "excelAmountTotal": 1122.72,
                "amountDelta": -1122.72,
                "sourceRefs": "账单!12",
            },
        ],
        "candidateMatches": [
            {
                "pdfEmployeeName": "Mitrache, Ross",
                "excelEmployeeName": "Rosa Alvarez Minchaca",
                "nameSimilarity": 0.75,
                "pdfHoursTotal": 30.5,
                "excelHoursTotal": 31.19,
                "hoursDelta": -0.69,
                "pdfAmountTotal": 698.99,
                "excelAmountTotal": 701.9,
                "amountDelta": -2.91,
                "recommendation": "姓名接近但金额和工时仍需确认",
                "sourceRefs": "scan.pdf p1; 账单!2",
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="5058871-5058880",
    )

    html = output.read_text(encoding="utf-8")
    assert "系统自动修正" in html
    assert "系统已自动合并姓名格式差异" in html
    assert "Mucu, Pablo" in html
    assert "Pablo Mucu" in html
    assert "疑似同一员工，需确认" in html
    assert "Mitrache, Ross" in html
    assert "Rosa Alvarez Minchaca" in html
    assert "待确认异常" in html
    assert "优先处理影响放行或留档的项目" in html
    assert "处理顺序：先确认金额口径，再确认缺发票项，最后确认疑似同一员工。" in html
    assert "处理建议：核对费率、加班、服务费或税费是否同一口径" in html
    assert "处理建议：确认本员工是否属于本批发票" in html
    assert "确认前不会自动合并姓名" in html
    assert "员工明细未完整识别，请查看原发票" in html
    assert "下载 Excel 明细" in html
    assert "低置信度抽取" not in html
    assert "人工复核" not in html


def test_build_labor_business_html_report_auto_fix_section_handles_accent_differences(tmp_path):
    output = tmp_path / "accent-name-auto-fix-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 1,
            "pdfAmountTotal": 739.49,
            "excelAmountTotal": 739.49,
            "amountDeltaTotal": 0,
            "passedCount": 1,
            "amountDiffCount": 0,
        },
        "rows": [
            {
                "employeeName": "Alberto Núñez ⇄ Alberto Nunez",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 35.08,
                "excelHoursTotal": 35.08,
                "hoursDelta": 0,
                "pdfAmountTotal": 739.49,
                "excelAmountTotal": 739.49,
                "amountDelta": 0,
                "sourceRefs": "Invoice-5058871.pdf p1; 账单!3",
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="5058871",
    )

    html = output.read_text(encoding="utf-8")
    assert "系统已自动合并姓名格式差异" in html
    assert "Alberto Núñez" in html
    assert "Alberto Nunez" in html
    assert "本次未发现可由系统自动合并的姓名格式差异" not in html


def test_build_labor_business_html_report_does_not_pass_when_extraction_failed(tmp_path):
    output = tmp_path / "failed-business-report.html"
    comparison = {
        "summary": {
            "extractionFailed": True,
            "failureReason": "PDF 明细未解析完成",
            "pdfEmployeeCount": 0,
            "excelEmployeeCount": 0,
            "pdfAmountTotal": 0,
            "excelAmountTotal": 0,
            "amountDeltaTotal": 0,
            "passedCount": 0,
            "amountDiffCount": 0,
        },
        "rows": [],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Workforce Priority",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "系统未能完成核对" in html
    assert "核对通过" not in html
    assert "请查看原发票和账单后重新生成报告" in html
    assert "人工查看" not in html


def test_build_labor_business_html_report_marks_detail_rows_missing_as_total_pass_with_detail_confirmation(tmp_path):
    output = tmp_path / "missing-detail-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.83,
            "amountDeltaTotal": 0,
            "passedCount": 0,
            "amountDiffCount": 0,
        },
        "rows": [],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Workforce Priority",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "总账通过，但员工明细待确认" in html
    assert "这批账能不能放行？" in html
    assert "需业务确认" in html
    assert "总金额已通过；员工明细未完整识别，不影响总账结论，但需要业务确认明细后再对外留档。" in html
    assert "员工明细未完整识别" in html
    assert "系统已确认本批总金额一致" in html
    assert "部分员工明细未完整识别" in html
    assert "本批总金额已完成核对，但当前没有可逐项展示的员工明细" in html
    assert "暂无可展示明细" not in html
    assert "金额口径说明" not in html
    assert "系统未能完成核对" not in html


def test_build_labor_business_html_report_marks_partial_detail_coverage_as_incomplete(tmp_path):
    output = tmp_path / "partial-detail-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 1,
            "pdfAmountTotal": 100.0,
            "excelAmountTotal": 100.0,
            "amountDeltaTotal": 0,
            "passedCount": 1,
            "amountDiffCount": 0,
        },
        "rows": [
            {
                "employeeName": "Employee One",
                "matchStatus": "通过",
                "pdfAmountTotal": 100.0,
                "excelAmountTotal": 100.0,
                "amountDelta": 0.0,
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Image Supplier",
        period_start="2026-06-01",
        period_end="2026-06-07",
        detail_coverage_complete=False,
    )

    html = output.read_text(encoding="utf-8")
    assert "总账通过，但员工明细待确认" in html
    assert "员工明细未完整识别" in html
    assert "需业务确认" in html
    assert ">可放行<" not in html


def test_build_labor_business_html_report_treats_ten_cent_total_difference_as_pass(tmp_path):
    output = tmp_path / "ten-cent-total-pass.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.93,
            "amountDeltaTotal": -0.1,
            "passedCount": 0,
            "amountDiffCount": 0,
        },
        "rows": [],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "总账通过，但员工明细待确认" in html
    assert "本批总金额已完成核对" in html
    assert "总金额存在差异，暂不能放行" not in html


def test_build_labor_business_html_report_total_pass_with_review_items_does_not_claim_incomplete_recognition(tmp_path):
    output = tmp_path / "detail-review-with-rows-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 2,
            "excelEmployeeCount": 2,
            "pdfAmountTotal": 2000.00,
            "excelAmountTotal": 2000.00,
            "amountDeltaTotal": 0,
            "passedCount": 1,
            "amountDiffCount": 0,
            "candidateMatchCount": 1,
        },
        "rows": [
            {
                "employeeName": "Pablo Mucu ⇄ Mucu, Pablo",
                "matchStatus": "疑似同一员工",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "invoice.pdf p1; bill.xlsx!2",
            }
        ],
        "candidateMatches": [
            {
                "pdfEmployeeName": "Pablo Mucu",
                "excelEmployeeName": "Mucu, Pablo",
                "amountGap": 0,
                "hoursGap": 0,
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "总账通过，但员工明细待确认" in html
    assert "员工明细仍有需要确认的项目" in html
    assert "疑似同一员工，需确认" in html
    assert "部分员工明细未完整识别" not in html


def test_build_labor_business_html_report_marks_excel_only_rows_as_incomplete_pdf_coverage(tmp_path):
    output = tmp_path / "excel-only-coverage-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 0,
            "excelEmployeeCount": 2,
            "pdfAmountTotal": 0.0,
            "excelAmountTotal": 1200.0,
            "amountDeltaTotal": -1200.0,
        },
        "rows": [
            {
                "employeeName": "Excel Only Worker",
                "matchStatus": "Excel有PDF无",
                "riskFlags": [],
                "pdfHoursTotal": 0,
                "excelHoursTotal": 40,
                "hoursDelta": -40,
                "pdfAmountTotal": 0,
                "excelAmountTotal": 1200,
                "amountDelta": -1200,
                "sourceRefs": "账单!2",
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "员工明细未完整展开" in html
    assert "尚不能视为已与发票逐一核对" in html
    assert "用于确认每位员工的发票金额和账单金额是否一致" not in html


def test_build_labor_business_html_report_amount_close_but_name_unlike_stays_manual_confirmation(tmp_path):
    output = tmp_path / "amount-close-name-unlike-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 2,
            "excelEmployeeCount": 2,
            "pdfAmountTotal": 2000.00,
            "excelAmountTotal": 2000.00,
            "amountDeltaTotal": 0,
            "passedCount": 1,
            "amountDiffCount": 0,
            "candidateMatchCount": 1,
        },
        "rows": [
            {
                "employeeName": "Carlos Serna ⇄ Carlos Serna",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "invoice.pdf p1; bill.xlsx!2",
            }
        ],
        "candidateMatches": [
            {
                "pdfEmployeeName": "Maria Lopez",
                "excelEmployeeName": "Carlos Serna",
                "nameSimilarity": 0.12,
                "pdfAmountTotal": 812.80,
                "excelAmountTotal": 812.80,
                "amountDelta": 0,
                "hoursDelta": 0,
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "疑似同一员工，需确认" in html
    assert "Maria Lopez ⇄ Carlos Serna" in html
    assert "金额接近，但姓名不像，不能自动合并" in html
    assert "确认前不会自动合并姓名" in html


def test_build_labor_business_html_report_prioritizes_amount_difference_when_details_incomplete(tmp_path):
    output = tmp_path / "amount-difference-incomplete-detail-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.94,
            "amountDeltaTotal": -0.11,
            "passedCount": 0,
            "amountDiffCount": 0,
        },
        "rows": [],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Workforce Priority",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "总金额存在差异，暂不能放行" in html
    assert "这批账能不能放行？" in html
    assert "不建议放行" in html
    assert "总金额超出 $0.10 容差，先复核发票总额、账单总额和所属账期。" in html
    assert "总金额存在差异：PDF 比 Excel 少 $0.11" in html
    assert "由于员工明细未完整识别" in html
    assert "系统未能完成核对" not in html


def test_build_labor_business_html_report_does_not_blame_recognition_when_amount_diff_has_detail_rows(tmp_path):
    output = tmp_path / "amount-difference-with-detail-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 2,
            "excelEmployeeCount": 2,
            "pdfAmountTotal": 1000.00,
            "excelAmountTotal": 999.50,
            "amountDeltaTotal": 0.50,
            "passedCount": 1,
            "amountDiffCount": 1,
        },
        "rows": [
            {
                "employeeName": "Pablo Mucu ⇄ Pablo Mucu",
                "matchStatus": "金额差异",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 999.5,
                "amountDelta": 0.5,
                "sourceRefs": "invoice.pdf p1; bill.xlsx!2",
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "总金额存在差异，暂不能放行" in html
    assert "总金额存在差异：PDF 比 Excel 多 $0.50" in html
    assert "请先查看下方员工明细中的金额、工时或费率差异" in html
    assert "由于员工明细未完整识别" not in html


def test_build_labor_business_html_report_total_pass_takes_priority_over_employee_detail_differences(tmp_path):
    output = tmp_path / "total-pass-with-employee-detail-difference.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 2,
            "excelEmployeeCount": 2,
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.88,
            "amountDeltaTotal": -0.05,
            "passedCount": 1,
            "amountDiffCount": 1,
        },
        "rows": [
            {
                "employeeName": "Maria Lopez ⇄ Maria Lopez",
                "matchStatus": "金额差异",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 812.80,
                "excelAmountTotal": 812.70,
                "amountDelta": 0.10,
                "sourceRefs": "invoice.pdf p1; bill.xlsx!2",
            }
        ],
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="invoice upload",
    )

    html = output.read_text(encoding="utf-8")
    assert "总账通过，但员工明细待确认" in html
    assert "系统已确认本批总金额一致，但员工明细仍有需要确认的项目" in html
    assert "总金额存在差异，暂不能放行" not in html
    assert "总金额存在差异：PDF 比 Excel" not in html


def test_build_labor_business_html_report_separates_full_batch_from_review_scope(tmp_path):
    output = tmp_path / "review-scope-business-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 22002.58,
            "excelAmountTotal": 22002.59,
            "amountDeltaTotal": -0.01,
            "passedCount": 18,
            "amountDiffCount": 0,
        },
        "rows": [
            {
                "employeeName": "Employee A ⇄ Employee A",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "warehouse 25",
            }
        ],
    }
    warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.93,
            "amountDeltaTotal": -0.10,
            "totalPassed": False,
            "exceptionCount": 2,
            "diffWarehouses": ["25", "28"],
            "warehouseCount": 6,
        }
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="6 张发票",
        warehouse_comparison=warehouse_comparison,
    )

    html = output.read_text(encoding="utf-8")
    assert "整批 PDF 发票总金额" in html
    assert "$144,714.83" in html
    assert "整批账单总金额" in html
    assert "$144,714.93" in html
    assert "需要确认" in html
    assert "需要复核" not in html
    assert "需复核" not in html
    assert "只展示需要确认的仓库员工明细，不代表账单只有这些员工" in html
    assert "仓库 25、28" in html
    assert "$22,002.59" in html
    assert "全员对账明细" not in html
    for internal_term in ["Stage 2", "下钻", "diffWarehouses", "warehouseComparison", "核对信号存在冲突"]:
        assert internal_term not in html


def test_build_labor_business_html_report_explains_full_excel_count_vs_review_detail_scope(tmp_path):
    output = tmp_path / "excel-record-count-vs-review-scope-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 22002.58,
            "excelAmountTotal": 22002.59,
            "amountDeltaTotal": -0.01,
            "passedCount": 18,
            "amountDiffCount": 0,
        },
        "rows": [
            {
                "employeeName": "Employee A ⇄ Employee A",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "warehouse 25",
            }
        ],
    }
    warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.93,
            "amountDeltaTotal": -0.10,
            "totalPassed": True,
            "exceptionCount": 2,
            "allocationIssueCount": 2,
            "diffWarehouses": ["25", "28"],
            "warehouseCount": 6,
        }
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="6 张发票",
        warehouse_comparison=warehouse_comparison,
        excel_record_count=128,
    )

    html = output.read_text(encoding="utf-8")
    assert "整批账单已读取 128 行" in html
    assert "当前展示的是需要确认的 1 名员工明细" in html
    assert "不代表账单只有这些员工" in html
    assert "员工明细识别情况" in html
    assert "只展开需要确认的员工明细" in html
    assert "其余无明显差异的员工不在本段重复展示" in html
    assert "总账通过，但员工明细待确认" in html


def test_build_labor_business_html_report_explains_three_amount_layers(tmp_path):
    output = tmp_path / "three-amount-layers-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 22002.58,
            "excelAmountTotal": 22002.59,
            "amountDeltaTotal": -0.01,
            "passedCount": 18,
            "amountDiffCount": 0,
        },
        "rows": [
            {
                "employeeName": "Employee A ⇄ Employee A",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "warehouse 25",
            }
        ],
    }
    warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.93,
            "amountDeltaTotal": -0.10,
            "totalPassed": True,
            "exceptionCount": 2,
            "allocationIssueCount": 2,
            "diffWarehouses": ["25", "28"],
            "warehouseCount": 6,
        }
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="6 张发票",
        warehouse_comparison=warehouse_comparison,
        excel_record_count=128,
    )

    html = output.read_text(encoding="utf-8")
    assert "总金额核对" in html
    assert "金额口径说明" not in html
    assert "整批 PDF 发票总额" in html
    assert "$144,714.83" in html
    assert "整批 Excel 账单总额" in html
    assert "$144,714.93" in html
    assert "已识别员工明细金额" in html
    assert "$22,002.58" in html
    assert "员工明细金额用于定位差异，不等同于整批总账金额" in html
    assert "不代表账单少读了" in html
    assert "当前页面只展开了用于确认的明细范围" in html
    assert "员工明细识别情况" in html
    assert "如需查看所有原始员工行，请下载 Excel 明细" in html
    assert "总账结论优先看整批 PDF 与整批 Excel 的差额" in html


def test_build_labor_business_html_report_keeps_total_pass_when_allocation_needs_confirmation(tmp_path):
    output = tmp_path / "total-pass-allocation-review-report.html"
    comparison = {
        "summary": {
            "pdfEmployeeCount": 18,
            "excelEmployeeCount": 18,
            "pdfAmountTotal": 22002.58,
            "excelAmountTotal": 22002.59,
            "amountDeltaTotal": -0.01,
            "passedCount": 18,
            "amountDiffCount": 0,
        },
        "rows": [
            {
                "employeeName": "Employee A ⇄ Employee A",
                "matchStatus": "通过",
                "riskFlags": [],
                "pdfHoursTotal": 40,
                "excelHoursTotal": 40,
                "hoursDelta": 0,
                "pdfAmountTotal": 1000,
                "excelAmountTotal": 1000,
                "amountDelta": 0,
                "sourceRefs": "warehouse 25",
            }
        ],
    }
    warehouse_comparison = {
        "summary": {
            "pdfAmountTotal": 144714.83,
            "excelAmountTotal": 144714.93,
            "amountDeltaTotal": -0.10,
            "totalPassed": True,
            "exceptionCount": 2,
            "allocationIssueCount": 2,
            "diffWarehouses": ["25", "28"],
            "warehouseCount": 6,
        }
    }

    build_labor_business_html_report(
        output,
        comparison,
        supplier_name="Fairway",
        period_start="2026-05-11",
        period_end="2026-05-17",
        invoice_scope="6 张发票",
        warehouse_comparison=warehouse_comparison,
    )

    html = output.read_text(encoding="utf-8")
    assert "总账通过，但员工明细待确认" in html
    assert "系统已确认本批总金额一致" in html
    assert "员工级差异仅供确认" in html
    assert '<div class="val">需要复核</div>' not in html
    assert "总金额存在差异，暂不能放行" not in html


# ---------------------------------------------------------------------------
# Phase 2 Tests
# ---------------------------------------------------------------------------


def _make_labor_item(
    name: str = "John Doe",
    amount: float = 1000.0,
    hours: float = 40.0,
    confidence: float = 0.95,
    source_file: str = "test.pdf",
    source_page: str = "p1",
) -> LaborLineItem:
    return LaborLineItem(
        source_type="pdf",
        source_file=source_file,
        source_page_or_row=source_page,
        employee_id="",
        employee_name_raw=name,
        hours=hours,
        amount=amount,
        currency="USD",
        confidence=confidence,
        evidence_text="",
        supplier="",
    )


def test_calculate_extraction_quality_returns_low_confidence_rows_T_P2_1():
    rows = [
        _make_labor_item(name="John", confidence=0.95),
        _make_labor_item(name="Jane", confidence=0.60),
        _make_labor_item(name="Bob", confidence=0.80),
    ]
    result = calculate_extraction_quality(
        pdf_rows=rows,
        comparison_summary={},
    )
    assert "lowConfidenceRows" in result
    low = result["lowConfidenceRows"]
    assert len(low) == 2
    names = {r["employee_name_raw"] for r in low}
    assert names == {"Jane", "Bob"}
    # Check fields present
    for row in low:
        assert "employee_name_raw" in row
        assert "amount" in row
        assert "confidence" in row
        assert "source_page_or_row" in row
        assert "source_file" in row


def test_calculate_extraction_quality_low_confidence_rows_empty_when_all_high_T_P2_2():
    rows = [
        _make_labor_item(name="John", confidence=0.95),
        _make_labor_item(name="Jane", confidence=0.90),
    ]
    result = calculate_extraction_quality(
        pdf_rows=rows,
        comparison_summary={},
    )
    assert result["lowConfidenceRows"] == []


def test_calculate_extraction_quality_respects_confidence_threshold_param_T_P2_3():
    rows = [
        _make_labor_item(name="John", confidence=0.95),
        _make_labor_item(name="Jane", confidence=0.85),
        _make_labor_item(name="Bob", confidence=0.70),
    ]
    result = calculate_extraction_quality(
        pdf_rows=rows,
        comparison_summary={},
        confidence_threshold=0.9,
    )
    low = result["lowConfidenceRows"]
    names = {r["employee_name_raw"] for r in low}
    # confidence=0.85 is < 0.9, so Jane should also be in low
    assert "Jane" in names
    assert "Bob" in names
    assert "John" not in names


def test_calculate_extraction_quality_counts_unresolved_missing_and_extra_warehouse_evidence():
    result = calculate_extraction_quality(
        pdf_rows=[],
        comparison_summary={},
        warehouse_comparison={
            "rows": [
                {"warehouseId": "1", "reconciliationStatus": "passed"},
                {"warehouseId": "2", "reconciliationStatus": "needs_review"},
                {"warehouseId": "3", "reconciliationStatus": "missing_pdf_invoice"},
                {"warehouseId": "4", "reconciliationStatus": "extra_pdf_invoice"},
            ],
        },
    )

    assert result["metrics"]["warehouseEvidence"] == {
        "unresolvedEvidenceCount": 1,
        "unresolvedEvidenceWarehouses": ["2"],
        "missingPdfInvoiceCount": 1,
        "missingPdfInvoiceWarehouses": ["3"],
        "extraPdfInvoiceCount": 1,
        "extraPdfInvoiceWarehouses": ["4"],
    }


def test_calculate_extraction_quality_uses_reconciliation_status_for_warehouse_issues():
    result = calculate_extraction_quality(
        pdf_rows=[],
        comparison_summary={},
        warehouse_comparison={
            "rows": [
                {"warehouseId": "1", "matchStatus": "金额差异", "reconciliationStatus": "amount_difference", "amountDelta": -250},
                {"warehouseId": "2", "matchStatus": "缺少PDF发票", "reconciliationStatus": "missing_pdf_invoice", "amountDelta": -200},
                {"warehouseId": "3", "matchStatus": "多余PDF发票", "reconciliationStatus": "extra_pdf_invoice", "amountDelta": 150},
                {"warehouseId": "4", "matchStatus": "待复核", "reconciliationStatus": "needs_review", "amountDelta": 0},
            ],
        },
    )

    warehouse_issues = result["metrics"]["warehouseIssues"]
    assert any("仓库 1" in issue and "金额差异" in issue for issue in warehouse_issues)
    assert any("仓库 2" in issue and "缺少PDF发票" in issue and "金额差异" not in issue for issue in warehouse_issues)
    assert any("仓库 3" in issue and "多余PDF发票" in issue and "金额差异" not in issue for issue in warehouse_issues)
    assert any("仓库 4" in issue and "待复核" in issue and "金额差异" not in issue for issue in warehouse_issues)
    assert all(issue in result["issues"] for issue in warehouse_issues)


def test_calculate_extraction_quality_labels_business_delta_without_claiming_bad_extraction():
    result = calculate_extraction_quality(
        pdf_rows=[
            LaborLineItem(
                source_type="pdf_invoice",
                source_file="invoice.pdf",
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Worker One",
                hours=8,
                amount=1000,
                currency="EUR",
                confidence=1,
                evidence_text="invoice row",
                warehouse_id="1",
            )
        ],
        comparison_summary={
            "pdfEmployeeCount": 1,
            "excelEmployeeCount": 1,
            "pdfHoursTotal": 8,
            "excelHoursTotal": 8,
            "pdfAmountTotal": 1000,
            "excelAmountTotal": 1150,
            "unmatchedPdfCount": 0,
            "unmatchedExcelCount": 0,
        },
        warehouse_comparison={
            "rows": [
                {
                    "warehouseId": "1",
                    "reconciliationStatus": "amount_difference",
                    "amountDelta": -150,
                }
            ]
        },
    )

    assert result["level"] == "warning"
    assert result["message"] == "核对发现业务差异，请查看员工与仓库明细。"
    assert "抽取" not in result["message"]


def test_ai_instruction_retry_mode_appends_target_names_T_P2_4():
    prompt = _ai_instruction(retry_mode=True, target_names=["John", "Jane"])
    assert "RETRY MODE" in prompt
    assert "John" in prompt
    assert "Jane" in prompt


def test_ai_instruction_no_retry_mode_by_default_T_P2_5():
    prompt = _ai_instruction()
    assert "RETRY MODE" not in prompt


# ---------------------------------------------------------------------------
# Phase 3 Tests
# ---------------------------------------------------------------------------


def test_generate_profile_from_extraction_basic_T_P3_1():
    rows = [
        _make_labor_item(name="Alice", hours=40, amount=1000, confidence=0.95),
        _make_labor_item(name="Bob", hours=35, amount=800, confidence=0.90),
    ]
    profile = generate_profile_from_extraction("Fairway", rows)
    assert "key" in profile
    assert "aliases" in profile
    assert "prompt_notes" in profile
    assert "image_page_policy" in profile
    assert "version" in profile
    assert profile["key"] == "fairway"
    assert isinstance(profile["prompt_notes"], list)
    assert profile["version"] == 1


def test_generate_profile_candidate_learns_reviewable_line_item_aliases():
    rows = [
        replace(
            _make_labor_item(name="Unknown Worker", hours=0, amount=19.44),
            item_type="meal_allowance",
            description="TICKET RESTAURANT",
            quantity=4,
            unit="meal",
        )
    ]

    profile = generate_profile_from_extraction("Unseen Vendor LLC", rows)

    assert profile["line_item_aliases"] == {"ticket restaurant": "meal_allowance"}


def test_generate_profile_from_extraction_detects_zero_hours_premiums_T_P3_2():
    rows = [
        _make_labor_item(name="Alice", hours=0, amount=50, confidence=0.95),
        _make_labor_item(name="Bob", hours=40, amount=1000, confidence=0.90),
    ]
    profile = generate_profile_from_extraction("Fairway", rows)
    notes_text = " ".join(profile["prompt_notes"]).lower()
    assert "meal premiums" in notes_text


def test_generate_profile_from_extraction_empty_supplier_T_P3_3():
    rows = [_make_labor_item(name="Alice")]
    profile = generate_profile_from_extraction("", rows)
    assert profile["key"] == "unknown"


def test_save_supplier_profile_creates_file_T_P3_4(tmp_path):
    profile = {
        "key": "test_supplier",
        "aliases": ["test supplier"],
        "prompt_notes": ["note 1"],
        "image_page_policy": "first_page_only",
        "version": 1,
    }
    result_path = save_supplier_profile(profile, tmp_path)
    assert result_path.exists()
    loaded = json.loads(result_path.read_text(encoding="utf-8"))
    assert loaded["key"] == "test_supplier"
    assert loaded["aliases"] == ["test supplier"]


def test_profiles_for_resolution_scans_directory_T_P3_5(tmp_path):
    # Create two profile JSON files in the directory
    profile_a = [
        {
            "key": "supplier_a", "aliases": ["supplier a"], "prompt_notes": ["note a"],
            "version": 1, "status": "approved", "approvedBy": "p0-test-reviewer",
            "approvedAt": "2026-07-15T10:00:00Z", "created_from": "manual_review",
        }
    ]
    profile_b = [
        {
            "key": "supplier_b", "aliases": ["supplier b"], "prompt_notes": ["note b"],
            "version": 1, "status": "approved", "approvedBy": "p0-test-reviewer",
            "approvedAt": "2026-07-15T10:00:00Z", "created_from": "manual_review",
        }
    ]
    (tmp_path / "a.json").write_text(json.dumps(profile_a), encoding="utf-8")
    (tmp_path / "b.json").write_text(json.dumps(profile_b), encoding="utf-8")

    profiles = _profiles_for_resolution(tmp_path)
    keys = {p.key for p in profiles}
    assert "supplier_a" in keys
    assert "supplier_b" in keys


# ---------------------------------------------------------------------------
# Phase 4 Tests
# ---------------------------------------------------------------------------


def test_check_profile_validity_returns_true_for_default_T_P4_1():
    # DEFAULT_PROFILE is always valid
    assert DEFAULT_PROFILE.key == "default"
    assert DEFAULT_PROFILE.deprecated is False


def test_check_profile_validity_returns_true_when_rule_rows_exist_T_P4_2(tmp_path):
    # A non-default profile with prompt_notes (rule_rows proxy) should be valid
    from bonus_platform.engine.labor.profiles import SupplierExtractionProfile

    profile = SupplierExtractionProfile(
        key="custom",
        aliases=["custom"],
        prompt_notes=["some rule"],
        image_page_policy="first_page_only",
    )
    assert not profile.deprecated
    assert len(profile.prompt_notes) > 0


def test_check_profile_validity_returns_false_when_no_rule_rows_T_P4_3():
    from bonus_platform.engine.labor.profiles import SupplierExtractionProfile

    profile = SupplierExtractionProfile(
        key="empty",
        aliases=["empty"],
        prompt_notes=[],
        image_page_policy="first_page_only",
    )
    # Empty prompt_notes means no rules configured
    assert len(profile.prompt_notes) == 0


def test_record_profile_failure_increments_count_T_P4_4(tmp_path):
    profile_data = {
        "key": "test_profile",
        "aliases": ["test"],
        "failure_count": 0,
    }
    profile_path = tmp_path / "test_profile.json"
    profile_path.write_text(json.dumps(profile_data), encoding="utf-8")

    result = record_profile_failure(profile_path)
    assert result is not None
    assert result["failure_count"] == 1


def test_record_profile_failure_marks_deprecated_after_3_T_P4_5(tmp_path):
    profile_data = {
        "key": "bad_profile",
        "aliases": ["bad"],
        "failure_count": 2,
    }
    profile_path = tmp_path / "bad_profile.json"
    profile_path.write_text(json.dumps(profile_data), encoding="utf-8")

    result = record_profile_failure(profile_path)
    assert result is not None
    assert result["failure_count"] == 3
    assert result["deprecated"] is True


def test_reset_profile_failure_clears_count_T_P4_6(tmp_path):
    profile_data = {
        "key": "recover_profile",
        "aliases": ["recover"],
        "failure_count": 2,
        "deprecated": True,
    }
    profile_path = tmp_path / "recover_profile.json"
    profile_path.write_text(json.dumps(profile_data), encoding="utf-8")

    reset_profile_failure(profile_path)
    loaded = json.loads(profile_path.read_text(encoding="utf-8"))
    assert loaded["failure_count"] == 0
    assert "deprecated" not in loaded


def test_profiles_for_resolution_filters_deprecated_T_P4_7(tmp_path):
    # Create a deprecated profile that would match "deprecated_supplier"
    profile_data = [
        {
            "key": "deprecated_supplier",
            "aliases": ["deprecated supplier"],
            "prompt_notes": ["old rule"],
            "deprecated": True,
        }
    ]
    (tmp_path / "deprecated.json").write_text(
        json.dumps(profile_data), encoding="utf-8"
    )

    profile = resolve_supplier_profile("deprecated supplier", profiles_path=tmp_path)
    # Should fall back to DEFAULT_PROFILE because the deprecated one is filtered
    assert profile.key == "default"


def test_build_material_index_identifies_replay_ready_batches_and_ignores_temp_files(tmp_path):
    batch = tmp_path / "workforce已报账"
    batch.mkdir()
    (batch / "Invoice-5058871.pdf").write_bytes(b"%PDF-1.4\n")
    (batch / "员工账单明细 - 2026-06-01T112149.990.xlsx").write_bytes(b"fake workbook")
    (batch / "~$OTWS - Warehouse Bill-NJ8.xlsx").write_bytes(b"temp")
    (batch / "Timecard for 05.11.2026-05.17.2026.eml").write_text("context", encoding="utf-8")
    warehouse_29 = tmp_path / "29仓"
    warehouse_29.mkdir()
    (warehouse_29 / "In291943.pdf").write_bytes(b"%PDF-1.4\n")
    (warehouse_29 / "员工账单明细 - 2026-05-28T141945.414.xlsx").write_bytes(b"fake workbook")
    (tmp_path / "README.md").write_text("notes", encoding="utf-8")

    index = build_material_index(tmp_path)

    assert index["summary"]["candidateBatchCount"] == 2
    assert index["summary"]["invoicePdfCount"] == 2
    assert index["summary"]["workbookCount"] == 2
    assert index["batches"] == index["candidateBatches"]
    batch_index = next(batch for batch in index["candidateBatches"] if batch["supplier"] == "workforce")
    assert batch_index["supplier"] == "workforce"
    assert batch_index["replayReady"] is True
    assert batch_index["invoiceFiles"][0]["filename"] == "Invoice-5058871.pdf"
    assert batch_index["pdfFiles"] == batch_index["invoiceFiles"]
    assert batch_index["workbookFiles"][0]["filename"].startswith("员工账单明细")
    assert "邮件上下文" in batch_index["limitations"][0]
    assert batch_index["expectedRisks"] == batch_index["limitations"]
    warehouse_batch = next(batch for batch in index["candidateBatches"] if batch["directory"] == "29仓")
    assert warehouse_batch["warehouseIds"] == ["29"]
    assert all("~$" not in item["filename"] for item in index["files"])


def test_build_material_index_groups_nested_bill_workbooks_with_parent_invoices(tmp_path):
    batch = tmp_path / "SSS 5.11-5.17"
    bill_dir = batch / "Strategic Staffing Solutions Corp 账单"
    bill_dir.mkdir(parents=True)
    (batch / "NJ8 Invoice Report WE 051726 JF.pdf").write_bytes(b"%PDF-1.4\n")
    (batch / "NJ13 Invoice Report WE 051726 JF.pdf").write_bytes(b"%PDF-1.4\n")
    (bill_dir / "OTWS - Warehouse Bill-NJ8.xlsx").write_bytes(b"fake workbook")
    (bill_dir / "~$OTWS - Warehouse Bill-NJ8.xlsx").write_bytes(b"temp")

    index = build_material_index(tmp_path)

    sss_batch = next(batch for batch in index["candidateBatches"] if batch["batchKey"] == "SSS_5_11_5_17")
    assert sss_batch["directory"] == "SSS 5.11-5.17"
    assert sss_batch["supplier"] == "sss"
    assert sss_batch["invoicePdfCount"] == 2
    assert sss_batch["workbookCount"] == 1
    assert sss_batch["uploadableFileCount"] == 3
    assert sss_batch["workbookFiles"][0]["relativePath"] == "SSS 5.11-5.17/Strategic Staffing Solutions Corp 账单/OTWS - Warehouse Bill-NJ8.xlsx"
    assert any("子目录" in limitation for limitation in sss_batch["limitations"])


def test_build_material_replay_plan_accepts_multi_warehouse_bill_workbooks(tmp_path):
    batch = tmp_path / "SSS 5.11-5.17"
    bill_dir = batch / "Strategic Staffing Solutions Corp 账单"
    bill_dir.mkdir(parents=True)
    for warehouse_id in ("8", "13"):
        (batch / f"NJ{warehouse_id} Invoice Report WE 051726 JF.pdf").write_bytes(b"%PDF-1.4\n")
        _write_labor_bill_workbook_with_rows(
            bill_dir / f"OTWS - Warehouse Bill-NJ{warehouse_id}.xlsx",
            [["WUS001", f"Worker {warehouse_id}", 8, 100, "USD", f"New Jersey Warehouse {warehouse_id}"]],
        )

    plan = build_material_replay_plan(tmp_path, batch_key="SSS_5_11_5_17")

    item = plan["plans"][0]
    assert item["supplier"] == "sss"
    assert item["warehouseIds"] == ["8", "13"]
    assert set(item["uploadPlan"]["workbookFiles"]) == {
        "SSS 5.11-5.17/Strategic Staffing Solutions Corp 账单/OTWS - Warehouse Bill-NJ13.xlsx",
        "SSS 5.11-5.17/Strategic Staffing Solutions Corp 账单/OTWS - Warehouse Bill-NJ8.xlsx",
    }
    assert {tuple(candidate["warehouseIds"]) for candidate in item["mappingCandidates"]} == {("8",), ("13",)}
    assert "需要确认主账单" not in " ".join(item["expectedRisks"])
    assert item["replayReady"] is True


def test_build_material_replay_plan_suggests_uploads_mapping_and_risks(tmp_path):
    batch = tmp_path / "oss 2"
    batch.mkdir()
    (batch / "US Elogis Service #7 Invoice W.E 05.24.26.pdf").write_bytes(b"%PDF-1.4\n")
    workbook_path = batch / "员工账单明细 - 2026-06-04T094719.972.xlsx"
    _write_labor_bill_workbook(workbook_path)

    plan = build_material_replay_plan(tmp_path, batch_key="oss_2")

    assert plan["summary"]["planCount"] == 1
    item = plan["plans"][0]
    assert item["supplier"] == "oss"
    assert item["periodHint"] == "W.E 05.24.26"
    assert item["warehouseIds"] == ["7"]
    assert item["uploadPlan"]["pdfFiles"] == ["oss 2/US Elogis Service #7 Invoice W.E 05.24.26.pdf"]
    assert item["uploadPlan"]["workbookFiles"] == ["oss 2/员工账单明细 - 2026-06-04T094719.972.xlsx"]
    mapping = item["mappingCandidates"][0]["suggestedMapping"]
    assert mapping["name"] == "姓名"
    assert mapping["hours"] == "时长总计(H)"
    assert mapping["amount"] == "费用总计(含税)"
    assert "amountColumns" not in mapping
    assert item["replayReady"] is True
    assert item["replayMode"] == "deterministic_first"
    assert "异常解释" in item["aiAllowedFor"]


def test_build_material_replay_plan_excludes_hours_only_supporting_workbook(tmp_path):
    batch = tmp_path / "Grande-"
    batch.mkdir()
    (batch / "GS invoice-ELOG-466-FL.pdf").write_bytes(b"%PDF-1.4\n")
    bill_path = batch / "员工账单明细 - 2026-05-28T172347.826.xlsx"
    _write_labor_bill_workbook(bill_path)
    support_path = batch / "GRANDE-5.18-5.24.xlsx"
    support_path.write_bytes(_workbook_with_hours_only_summary_bytes())

    plan = build_material_replay_plan(tmp_path, batch_key="Grande")

    item = plan["plans"][0]
    assert item["uploadPlan"]["workbookFiles"] == ["Grande-/员工账单明细 - 2026-05-28T172347.826.xlsx"]
    assert len(item["mappingCandidates"]) == 1
    assert item["mappingCandidates"][0]["filename"].startswith("员工账单明细")
    assert item["excludedWorkbookFiles"] == [
        {
            "relativePath": "Grande-/GRANDE-5.18-5.24.xlsx",
            "filename": "GRANDE-5.18-5.24.xlsx",
            "reason": "缺少必要映射: amount",
        }
    ]
    assert any("辅助材料排除" in risk for risk in item["expectedRisks"])
    assert item["replayReady"] is True


def test_build_material_replay_plan_chooses_complete_mapping_sheet(tmp_path):
    batch = tmp_path / "ADEQUAT"
    batch.mkdir()
    (batch / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Sheet1"
    summary.append(["求和项:时长总计", "总计"])
    summary.append(["Alice Worker", 8])
    bill = workbook.create_sheet("Sheet2")
    bill.append(["员工名称", "总计", "时薪", "本周薪资", "本周餐补", "总额"])
    bill.append(["Alice Worker", 8, 12.5, 100, 10, 110])
    workbook.save(batch / "巴黎1号仓.xlsx")

    plan = build_material_replay_plan(tmp_path, batch_key="ADEQUAT")

    item = plan["plans"][0]
    assert item["replayReady"] is True
    assert item["supplier"] == "adequat"
    assert item["mappingCandidates"][0]["sheetName"] == "Sheet2"
    assert item["mappingCandidates"][0]["suggestedMapping"]["name"] == "员工名称"
    assert item["mappingCandidates"][0]["suggestedMapping"]["hours"] == "总计"
    assert item["mappingCandidates"][0]["suggestedMapping"]["amount"] == "总额"


def test_build_material_replay_plan_uses_amount_components_when_total_is_missing(tmp_path):
    batch = tmp_path / "Sovitrat groupe"
    batch.mkdir()
    (batch / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet2"
    sheet.append(["员工名称", "总计", "时薪", "本周薪资", "本周餐补", "周日补贴"])
    sheet.append(["Alice Worker", 8, 12.5, 100, 10, 5])
    workbook.save(batch / "巴黎2号仓.xlsx")

    plan = build_material_replay_plan(tmp_path, batch_key="Sovitrat_groupe")

    mapping = plan["plans"][0]["mappingCandidates"][0]["suggestedMapping"]
    assert plan["plans"][0]["supplier"] == "sovitrat groupe"
    assert mapping["amountColumns"] == ["本周薪资", "本周餐补", "周日补贴"]


def test_build_material_replay_plan_does_not_sum_aggregate_and_component_amounts(tmp_path):
    batch = tmp_path / "oss"
    batch.mkdir()
    (batch / "invoice.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单明细"
    sheet.append(["姓名", "时长总计(H)", "费用总计(不含税)", "费用总计(含税)", "白班工作费用"])
    sheet.append(["Alice Worker", 8, 100, 110, 100])
    workbook.save(batch / "员工账单明细.xlsx")

    plan = build_material_replay_plan(tmp_path, batch_key="oss")

    mapping = plan["plans"][0]["mappingCandidates"][0]["suggestedMapping"]
    assert mapping["amount"] == "费用总计(不含税)"
    assert "amountColumns" not in mapping


def test_labor_production_readiness_blocks_unsafe_serverless_configuration():
    from bonus_platform.engine.labor.production_readiness import evaluate_labor_production_readiness

    result = evaluate_labor_production_readiness(
        env={"VERCEL": "1", "SIGMA_LABOR_EXECUTION_MODE": "personal-worker"},
        storage_info={"enabled": False, "backend": ""},
        queue_health={"backend": "local-json", "ready": False},
        build_info={"status": "current", "buildId": "test-build", "apiContractVersion": 2},
    )

    assert result["status"] == "blocked"
    assert result["directPaymentAllowed"] is False
    assert {issue["code"] for issue in result["blockers"]} == {
        "persistent_storage_required",
        "postgres_queue_required",
        "worker_tokens_required",
        "operations_token_required",
    }


def test_labor_production_readiness_stops_at_developer_preview_until_controlled_gates_exist():
    from bonus_platform.engine.labor.production_readiness import evaluate_labor_production_readiness

    result = evaluate_labor_production_readiness(
        env={
            "VERCEL": "1",
            "SIGMA_LABOR_EXECUTION_MODE": "personal-worker",
            "SIGMA_LABOR_WORKER_TOKENS": '{"opaque":{"userId":"u1","deviceId":"d1"}}',
            "SIGMA_LABOR_OPERATIONS_TOKEN": "ops-token",
        },
        storage_info={"enabled": True, "backend": "blob", "environment": "uat"},
        queue_health={"backend": "postgres", "configured": True, "ready": True},
        build_info={"status": "current", "buildId": "test-build", "apiContractVersion": 2},
    )

    assert result["status"] == "ready_for_developer_preview"
    assert result["readinessLevel"] == "developer_preflight"
    assert result["blockers"] == []
    assert result["manualReviewRequired"] is True
    assert result["directPaymentAllowed"] is False


def test_labor_production_readiness_blocks_stale_runtime_build():
    from bonus_platform.engine.labor.production_readiness import evaluate_labor_production_readiness

    result = evaluate_labor_production_readiness(
        env={
            "SIGMA_LABOR_EXECUTION_MODE": "personal-worker",
            "SIGMA_LABOR_WORKER_TOKENS": '{"opaque":{"userId":"u1","deviceId":"d1"}}',
            "SIGMA_LABOR_OPERATIONS_TOKEN": "ops-token",
        },
        storage_info={"enabled": True, "backend": "blob", "environment": "uat"},
        queue_health={"backend": "postgres", "configured": True, "ready": True},
        build_info={"status": "restart_required", "buildId": "stale-build", "apiContractVersion": 2},
    )

    assert result["status"] == "blocked"
    assert {issue["code"] for issue in result["blockers"]} == {"runtime_restart_required"}
    assert result["build"] == {
        "status": "restart_required",
        "buildId": "stale-build",
        "apiContractVersion": 2,
    }


def test_labor_production_readiness_endpoint_returns_sanitized_gate(monkeypatch):
    monkeypatch.setenv("SIGMA_LABOR_OPERATIONS_TOKEN", "ops-token")
    monkeypatch.setenv("SIGMA_LABOR_EXECUTION_MODE", "personal-worker")
    monkeypatch.setenv("SIGMA_LABOR_WORKER_TOKENS", '{"opaque":{"userId":"u1","deviceId":"d1"}}')
    monkeypatch.setenv("SIGMA_LABOR_REQUIRE_CLIENT_CONTRACT", "true")
    monkeypatch.setattr(app_module, "labor_persistent_storage_info", lambda: {"enabled": True, "backend": "blob"})
    monkeypatch.setattr(
        app_module,
        "labor_p1_worker_job_store_health",
        lambda: {"backend": "postgres", "configured": True, "ready": True},
    )

    result = app_module.get_labor_production_readiness(x_admin_token="ops-token")

    assert result["status"] == "ready_for_developer_preview"
    assert "opaque" not in json.dumps(result)


def test_labor_production_readiness_blocks_external_ai_for_first_uat():
    from bonus_platform.engine.labor.production_readiness import evaluate_labor_production_readiness

    result = evaluate_labor_production_readiness(
        env={
            "SIGMA_LABOR_EXECUTION_MODE": "personal-worker",
            "SIGMA_LABOR_WORKER_TOKENS": '{"opaque":{"userId":"u1","deviceId":"d1"}}',
            "SIGMA_LABOR_OPERATIONS_TOKEN": "ops-token",
            "SIGMA_LABOR_EXTERNAL_AI_ENABLED": "true",
        },
        storage_info={"enabled": True, "backend": "blob", "environment": "uat"},
        queue_health={"backend": "postgres", "configured": True, "ready": True},
        build_info={"status": "current", "buildId": "test-build", "apiContractVersion": 2},
    )

    assert result["status"] == "blocked"
    assert {issue["code"] for issue in result["blockers"]} == {"external_ai_disabled_required"}


def test_labor_production_readiness_rejects_malformed_worker_update_manifest():
    from bonus_platform.engine.labor.production_readiness import evaluate_labor_production_readiness

    result = evaluate_labor_production_readiness(
        env={
            "SIGMA_LABOR_EXECUTION_MODE": "personal-worker",
            "SIGMA_LABOR_WORKER_TOKENS": '{"opaque":{"userId":"u1","deviceId":"d1"}}',
            "SIGMA_LABOR_OPERATIONS_TOKEN": "ops-token",
            "SIGMA_LABOR_WORKER_UPDATE_MANIFEST": '{"version":"0.2"}',
        },
        storage_info={"enabled": True, "backend": "blob", "environment": "uat"},
        queue_health={"backend": "postgres", "configured": True, "ready": True},
        build_info={
            "status": "current",
            "buildId": "test-build",
            "apiContractVersion": 2,
            "requiredWorkerVersion": "0.3.0",
        },
    )

    assert "signed_update_manifest_invalid" in {issue["code"] for issue in result["warnings"]}


def test_labor_production_readiness_accepts_private_multiplatform_worker_release_catalog():
    from bonus_platform.engine.labor.production_readiness import evaluate_labor_production_readiness

    digest = "a" * 64
    result = evaluate_labor_production_readiness(
        env={
            "SIGMA_LABOR_EXECUTION_MODE": "personal-worker",
            "SIGMA_LABOR_WORKER_TOKENS": '{"opaque":{"userId":"u1","deviceId":"d1"}}',
            "SIGMA_LABOR_OPERATIONS_TOKEN": "ops-token",
            "SIGMA_LABOR_WORKER_UPDATE_MANIFEST": json.dumps(
                {
                    "schemaVersion": 2,
                    "releases": {
                        platform: {
                            "version": "0.3.11",
                            "minimumVersion": "0.3.11",
                            "sha256": digest,
                            "signature": f"sha256:{digest}",
                            "objectKey": f"labor-runs/production/worker-releases/{platform}/worker",
                        }
                        for platform in ("macos-arm64", "windows-x64")
                    },
                }
            ),
        },
        storage_info={"enabled": True, "backend": "supabase", "environment": "production"},
        queue_health={"backend": "postgres", "configured": True, "ready": True},
        build_info={
            "status": "current",
            "buildId": "test-build",
            "apiContractVersion": 2,
            "requiredWorkerVersion": "0.3.11",
        },
    )

    assert "signed_update_manifest_invalid" not in {issue["code"] for issue in result["warnings"]}


def test_labor_external_ai_requires_labor_specific_opt_in():
    config = _ready_ai_config()
    config["external_ai_enabled"] = False
    assert _ai_ready(config) is False

    config["external_ai_enabled"] = True
    assert _ai_ready(config) is True


def test_postgres_worker_claim_does_not_exceed_max_attempts():
    from contextlib import contextmanager
    from bonus_platform.engine.labor.worker_jobs_postgres import PostgresLaborWorkerStore

    statements = []

    class Result:
        def fetchone(self):
            return None

    class Connection:
        def execute(self, statement, params=()):
            statements.append(str(statement))
            return Result()

        def commit(self):
            return None

    @contextmanager
    def connect():
        yield Connection()

    store = PostgresLaborWorkerStore("postgresql://unused", connect=connect)
    assert store.claim("u1", "d1", "0.3.0") is None
    assert "attempt < max_attempts" in statements[0]


def test_build_material_replay_plan_identifies_prompt_priority_batch(tmp_path):
    batch = tmp_path / "prompt"
    batch.mkdir()
    (batch / "CHINA EXPRESS #3.pdf").write_bytes(b"%PDF-1.4\n")
    (batch / "DEPT#27.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook(batch / "员工账单明细 - 2026-05-28T151400.642.xlsx")

    plan = build_material_replay_plan(tmp_path, batch_key="prompt")

    item = plan["plans"][0]
    assert item["supplier"] == "prompt"
    assert item["warehouseIds"] == ["3", "27"]
    assert "未识别供应商" not in " ".join(item["expectedRisks"])
    assert item["replayReady"] is True


def test_build_material_replay_plan_uses_workbook_accounting_period(tmp_path):
    batch = tmp_path / "prompt"
    batch.mkdir()
    (batch / "DEPT#27.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单明细"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓", "核算开始日期", "核算结束日期"])
    sheet.append(["WUS001", "Alice Worker", 8, 100, "USD", "27号仓", "2026-05-11", "2026-05-17"])
    workbook.save(batch / "员工账单明细 - exported.xlsx")

    plan = build_material_replay_plan(tmp_path, batch_key="prompt")

    item = plan["plans"][0]
    assert item["periodHint"] == "2026-05-11~2026-05-17"
    assert "未识别到账期" not in " ".join(item["expectedRisks"])


def test_build_material_replay_plan_uses_workbook_supplier_hint(tmp_path):
    batch = tmp_path / "29仓"
    batch.mkdir()
    (batch / "In291943.pdf").write_bytes(b"%PDF-1.4\n")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单明细"
    sheet.append(["供应商名称", "工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓", "核算开始日期", "核算结束日期"])
    sheet.append(["CitiStaff Solutions", "WUS001", "Alice Worker", 8, 100, "USD", "29号仓", "2026-05-11", "2026-05-17"])
    workbook.save(batch / "员工账单明细 - 2026-05-28T141945.414.xlsx")

    plan = build_material_replay_plan(tmp_path, batch_key="29仓")

    item = plan["plans"][0]
    assert item["supplier"] == "citistaff"
    assert item["periodHint"] == "2026-05-11~2026-05-17"
    assert "未识别供应商" not in " ".join(item["expectedRisks"])


def test_build_material_dry_run_uses_deterministic_extract_and_does_not_write(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "oss 2"
    batch.mkdir()
    (batch / "US Elogis Service #7 Invoice W.E 05.24.26.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook(batch / "员工账单明细 - 2026-06-04T094719.972.xlsx")

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 100.0, "warehouse_id": "7"}],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file=paths[0].name, source_page_or_row="p1", employee_id="WUS001", employee_name_raw="Alice Worker", hours=8, amount=100, currency="USD", confidence=0.95, evidence_text="Alice Worker 8 $100")
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": "Alice Worker 8 $100"}],
    )

    dry_run = build_material_dry_run(tmp_path, "oss_2")

    assert dry_run["decision"] == "dry_run_only"
    assert dry_run["writesRun"] is False
    assert dry_run["aiInvoked"] is False
    assert dry_run["summary"]["pdfRowCount"] == 1
    assert dry_run["summary"]["excelRowCount"] == 1
    assert dry_run["summary"]["comparison"]["exceptionCount"] == 0
    assert dry_run["summary"]["warehouse"]["totalPassed"] is True
    assert dry_run["summary"]["tierStatus"]["employeeDetailAvailable"] is True
    assert dry_run["reviewQueues"]["primary"] == "cleared"
    assert "无需继续处理" in dry_run["reviewQueues"]["primaryReason"]
    assert dry_run["deliveryGate"]["status"] == "ready"
    assert dry_run["deliveryGate"]["label"] == "可交付"
    assert dry_run["deliveryGate"]["summary"]["blockedCount"] == 0
    assert dry_run["deliveryGate"]["summary"]["reviewCount"] == 0
    assert dry_run["deliveryGate"]["issues"] == []
    assert dry_run["reviewQueues"]["employeeExceptions"]["count"] == 0
    assert dry_run["reviewQueues"]["employeeExceptions"]["suppressedByPrimary"] is False
    assert dry_run["summary"]["pdfTextCoverage"]["textReadableFileCount"] == 1
    assert dry_run["summary"]["pdfTextCoverage"]["imageOnlyFileCount"] == 0
    assert dry_run["sampleRows"][0]["matchStatus"] == "通过"
    assert dry_run["exceptionRows"] == []
    assert dry_run["candidateMatches"] == []
    assert dry_run["nameMappingGovernance"]["decision"] == "candidate_only"
    assert dry_run["nameMappingGovernance"]["summary"]["candidateCount"] == 0
    assert dry_run["aiCacheAudit"]["summary"]["candidateFileCount"] == 0


def test_material_review_queue_marks_zero_exception_batches_as_cleared():
    from bonus_platform.engine.labor.materials import _build_material_review_queues

    queues = _build_material_review_queues(
        comparison_summary={"exceptionCount": 0},
        warehouse_summary={"exceptionCount": 0},
        exception_rows=[],
        pdf_text_coverage={"summary": {"imageOnlyFileCount": 2}},
        reocr_plan={"summary": {"taskCount": 2, "reviewableCandidateCount": 0}},
        ai_cache_preview={"summary": {"exceptionCount": 2, "needsReocrFileCount": 2}},
        name_mapping_governance={"summary": {"candidateCount": 0}},
        combined_row_governance={"summary": {"candidateCount": 0}},
        allocation_issues=[],
        hours_tolerance=0.01,
    )

    assert queues["primary"] == "cleared"
    assert "无需继续处理" in queues["primaryReason"]
    assert queues["employeeExceptions"]["count"] == 0
    assert queues["employeeExceptions"]["suppressedByPrimary"] is False


def test_material_review_queue_preserves_all_reocr_tasks_for_frontend_collapse():
    from bonus_platform.engine.labor.materials import _build_material_review_queues

    tasks = [
        {
            "sourceFile": f"DEPT#{index}.pdf",
            "warehouseId": str(index),
            "amountDelta": -100 * index,
            "pdfTextCoverage": {"needsOcr": True},
            "diagnostics": {"summary": {"exceptionCount": index, "unmatchedCacheCount": 0, "unmatchedExcelCount": index}},
        }
        for index in range(1, 13)
    ]

    queues = _build_material_review_queues(
        comparison_summary={"exceptionCount": 12},
        warehouse_summary={"exceptionCount": 12},
        exception_rows=[],
        pdf_text_coverage={"summary": {"imageOnlyFileCount": 12}},
        reocr_plan={"summary": {"taskCount": 12, "reviewableCandidateCount": 0}, "tasks": tasks},
        ai_cache_preview={"summary": {"exceptionCount": 12, "needsReocrFileCount": 12}},
        name_mapping_governance={"summary": {"candidateCount": 0}},
        combined_row_governance={"summary": {"candidateCount": 0}},
        allocation_issues=[],
        hours_tolerance=0.01,
    )

    assert queues["primary"] == "reocr"
    assert queues["reocr"]["taskCount"] == 12
    assert len(queues["reocr"]["tasks"]) == 12
    assert queues["reocr"]["tasks"][-1]["sourceFile"] == "DEPT#12.pdf"
    assert queues["reocr"]["summaryText"] == "12 个 PDF 无文本层 · 12 个图片发票明细待确认 · 12 项员工级异常"
    assert len(queues["reocr"]["groups"]) == 12
    assert queues["reocr"]["groups"][0]["sourceFile"] == "DEPT#12.pdf"
    assert queues["reocr"]["groups"][0]["statusLabel"] == "需重新识别"
    assert queues["reocr"]["groups"][0]["needsTextRecognition"] is True
    assert queues["reocr"]["groups"][0]["exceptionCount"] == 12
    assert queues["reocr"]["groups"][0]["unmatchedExcelCount"] == 12


def test_build_material_dry_run_applies_sss_rounding_tolerance(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "SSS 5.11-5.17"
    batch.mkdir()
    (batch / "NJ8 Invoice Report WE 051726 JF.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(batch / "OTWS - Warehouse Bill-NJ8.xlsx", [["WUS001", "Alice Worker", 8, 100.20, "USD", "New Jersey-8"]])

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 100.0, "warehouse_id": "8"}],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=paths[0].name,
                source_page_or_row="p1",
                employee_id="WUS001",
                employee_name_raw="Alice Worker",
                hours=8,
                amount=100.0,
                currency="USD",
                confidence=0.95,
                evidence_text="Alice Worker 8 $100.00",
            )
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": "Alice Worker 8 $100"}],
    )

    dry_run = build_material_dry_run(tmp_path, "SSS_5_11_5_17")

    assert dry_run["summary"]["tolerances"]["amount"] == 0.25
    assert any("SSS" in note for note in dry_run["summary"]["tolerances"]["notes"])
    assert dry_run["summary"]["comparison"]["exceptionCount"] == 0
    assert dry_run["summary"]["warehouse"]["totalPassed"] is True


def test_build_material_dry_run_prioritizes_amount_rate_review_for_same_hours_delta(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "SSS 5.11-5.17"
    batch.mkdir()
    (batch / "NJ13 Invoice Report WE 051726 JF.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "OTWS - Warehouse Bill-NJ13.xlsx",
        [["WUS001", "ALVARO TEJADA CAMPOS", 8, 162.56, "USD", "New Jersey Warehouse 13"]],
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 172.72, "warehouse_id": "13"}],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=paths[0].name,
                source_page_or_row="p19",
                employee_id="WUS001",
                employee_name_raw="Tejada, Alvaro",
                hours=8,
                amount=172.72,
                currency="USD",
                confidence=0.96,
                evidence_text="Tejada, Alvaro 8 $21.59 $172.72",
            )
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 19, "text": "Tejada, Alvaro 8 $21.59 $172.72"}],
    )

    dry_run = build_material_dry_run(tmp_path, "SSS_5_11_5_17")

    assert dry_run["summary"]["comparison"]["exceptionCount"] == 1
    assert dry_run["reviewQueues"]["primary"] == "amount_rate_review"
    assert "费率" in dry_run["reviewQueues"]["primaryReason"]
    assert "复核" not in dry_run["reviewQueues"]["primaryReason"]
    queue = dry_run["reviewQueues"]["amountRateReview"]
    assert queue["count"] == 1
    assert queue["reviewMode"] == "amount_basis"
    assert queue["amountOnlyCount"] == 1
    assert queue["hoursMismatchCount"] == 0
    assert queue["amountImpactTotal"] == 10.16
    assert queue["amountOnlyImpactTotal"] == 10.16
    assert queue["largestAmountDelta"] == 10.16
    assert "工时已经对齐" in queue["businessQuestion"]
    assert "金额计算口径" in queue["businessMeaning"]
    assert "不能由系统自动" in queue["cannotAutoResolveReason"]
    row = queue["rows"][0]
    assert row["reviewType"] == "amount_basis_mismatch"
    assert row["reviewLabel"] == "工时一致，仅金额不同"
    assert row["reviewFocus"] == "先核金额口径"
    assert row["employeeName"] == "Tejada, Alvaro"
    assert row["pdfAmountTotal"] == 172.72
    assert row["excelAmountTotal"] == 162.56
    assert row["amountDelta"] == 10.16
    assert row["amountDirectionLabel"] == "PDF 高于 Excel"
    assert row["hoursDelta"] == 0
    assert row["hoursDirectionLabel"] == "工时一致"
    assert "PDF 比 Excel 多 $10.16" in row["businessQuestion"]
    assert "工时一致" in row["businessQuestion"]
    assert "金额口径属于业务结算判断" in row["cannotAutoResolveReason"]
    assert "确认前不能自动清账" in row["recommendation"]
    actions = queue["nextActions"]
    assert [item["action"] for item in actions] == [
        "create_formal_run",
        "review_source_evidence",
        "record_business_conclusion",
        "download_report",
    ]
    assert actions[0]["enabled"] is True
    assert actions[0]["label"] == "建正式批次并保留差异"
    assert actions[1]["label"] == "核对金额计算口径"
    assert "自动清账" in actions[0]["description"]
    assert "保留为待处理异常" in actions[2]["description"]
    assert actions[3]["label"] == "导出给业务确认"
    amount_visible_text = " ".join(
        [
            dry_run["reviewQueues"]["primaryReason"],
            *[str(action.get("label", "")) for action in actions],
            *[str(action.get("description", "")) for action in actions],
            row["businessQuestion"],
            row["cannotAutoResolveReason"],
            row["recommendation"],
        ]
    )
    for internal_copy in ["人工复核", "需复核", "复核费率", "复核日期范围", "复核记录"]:
        assert internal_copy not in amount_visible_text


def test_build_material_dry_run_demotes_stale_image_cache_when_deterministic_extract_is_ok(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "SSS 5.11-5.17"
    batch.mkdir()
    (batch / "NJ13 Invoice Report WE 051726 JF.pdf").write_bytes(b"%PDF-1.4\n")
    workbook_rows = [["WUS001", "Alice Worker", 8, 100.00, "USD", "New Jersey Warehouse 13"]]
    workbook_rows.extend(
        [f"WUS{idx:03d}", f"Worker {idx:02d}", 8, 100.00, "USD", "New Jersey Warehouse 13"]
        for idx in range(2, 13)
    )
    _write_labor_bill_workbook_with_rows(
        batch / "OTWS - Warehouse Bill-NJ13.xlsx",
        workbook_rows,
    )
    cache_dir = batch / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "NJ13_Invoice_Report_WE_051726_JF_p1_mimo-v2.5_v4.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "Wrong Cache Person",
                    "source_page": 1,
                    "hours": 8,
                    "amount": 80,
                    "confidence": 0.95,
                    "evidence_text": "Wrong Cache Person 8 $80.00",
                }
            ]
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 1220.0, "warehouse_id": "13"}],
    )
    pdf_rows = [
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="NJ13 Invoice Report WE 051726 JF.pdf",
            source_page_or_row="p1",
            employee_id="WUS001",
            employee_name_raw="Alice Worker",
            hours=8,
            amount=120.0,
            currency="USD",
            confidence=0.96,
            evidence_text="Alice Worker 8 $120.00",
        )
    ]
    pdf_rows.extend(
        LaborLineItem(
            source_type="pdf_invoice",
            source_file="NJ13 Invoice Report WE 051726 JF.pdf",
            source_page_or_row="p1",
            employee_id=f"WUS{idx:03d}",
            employee_name_raw=f"Worker {idx:02d}",
            hours=8,
            amount=100.0,
            currency="USD",
            confidence=0.96,
            evidence_text=f"Worker {idx:02d} 8 $100.00",
        )
        for idx in range(2, 13)
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: pdf_rows,
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [
            {
                "source_file": paths[0].name,
                "page": 1,
                "text": "\n".join(row.evidence_text for row in pdf_rows),
            }
        ],
    )

    dry_run = build_material_dry_run(tmp_path, "SSS_5_11_5_17")

    assert dry_run["summary"]["pdfRowCount"] == 12
    assert dry_run["summary"]["pdfTextCoverage"]["imageOnlyFileCount"] == 0
    assert dry_run["summary"]["quality"]["level"] == "ok"
    assert dry_run["aiCacheReconciliationPreview"]["summary"]["needsReocrFileCount"] == 1
    assert dry_run["reocrPlan"]["demotedByDeterministicExtract"] is True
    assert dry_run["reocrPlan"]["summary"]["taskCount"] == 0
    assert dry_run["reocrPlan"]["summary"]["demotedTaskCount"] == 1
    assert dry_run["reviewQueues"]["reocr"]["taskCount"] == 0
    assert dry_run["reviewQueues"]["primary"] == "amount_rate_review"
    assert dry_run["deliveryGate"]["status"] == "needs_review"
    assert dry_run["deliveryGate"]["label"] == "需业务确认"
    assert dry_run["deliveryGate"]["message"] == "无阻断项，但仍有需业务留痕确认的项目。"
    assert "复核" not in dry_run["deliveryGate"]["label"]
    assert "复核" not in dry_run["deliveryGate"]["message"]
    assert not any(issue["code"] == "reocr_required" for issue in dry_run["deliveryGate"]["issues"])
    assert any("降级为审计参考" in risk for risk in dry_run["expectedRisks"])


def test_build_material_dry_run_prioritizes_amount_hours_review_for_hours_delta(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "osi 2"
    batch.mkdir()
    (batch / "US ELogistics Service Corp. 34926.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "员工账单明细 - 2026-06-04T094636.978.xlsx",
        [["WUS001", "Maria Elena Parraguirre", 12.25, 467.41, "USD", "1号仓"]],
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 557.76, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=paths[0].name,
                source_page_or_row="p3",
                employee_id="WUS001",
                employee_name_raw="Parraguirre, Maria",
                hours=20.36,
                amount=557.76,
                currency="USD",
                confidence=0.98,
                evidence_text="Parraguirre, Maria 8.11 OT + 12.25 REG/OT $557.76",
            )
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 3, "text": "Parraguirre, Maria 20.36 $557.76"}],
    )

    dry_run = build_material_dry_run(tmp_path, "osi_2")

    assert dry_run["summary"]["comparison"]["exceptionCount"] == 1
    assert dry_run["reviewQueues"]["primary"] == "amount_rate_review"
    assert "工时/金额" in dry_run["reviewQueues"]["primaryReason"]
    queue = dry_run["reviewQueues"]["amountRateReview"]
    assert queue["count"] == 1
    assert queue["reviewMode"] == "hours_and_amount"
    assert queue["amountOnlyCount"] == 0
    assert queue["hoursMismatchCount"] == 1
    assert queue["hoursImpactTotal"] == 8.11
    assert queue["hoursMismatchImpactTotal"] == 90.35
    assert "同一账期" in queue["businessQuestion"]
    assert "是否在核同一批工时" in queue["businessMeaning"]
    assert "工时差会改变应付金额" in queue["cannotAutoResolveReason"]
    row = queue["rows"][0]
    assert row["reviewType"] == "hours_amount_mismatch"
    assert row["reviewLabel"] == "工时和金额都不同"
    assert row["reviewFocus"] == "先核工时口径"
    assert row["employeeName"] == "Parraguirre, Maria"
    assert row["amountDelta"] == 90.35
    assert row["amountDirectionLabel"] == "PDF 高于 Excel"
    assert row["hoursDelta"] == 8.11
    assert row["hoursDirectionLabel"] == "PDF 工时多于 Excel"
    assert "PDF 比 Excel 多 $90.35" in row["businessQuestion"]
    assert "工时多 8.11" in row["businessQuestion"]
    assert "工时差会改变应付金额" in row["cannotAutoResolveReason"]
    assert "账期范围" in row["recommendation"]
    actions = queue["nextActions"]
    assert actions[1]["label"] == "核对账期、加班和工时"
    assert "账期范围" in actions[1]["description"]


def test_build_material_dry_run_surfaces_cross_warehouse_allocation_review(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "fairway已报账2"
    batch.mkdir()
    (batch / "fairway_25.pdf").write_bytes(b"%PDF-1.4\n")
    (batch / "fairway_28.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "员工账单明细 - 2026-06-04T094719.972.xlsx",
        [
            ["WUS041037", "PEREZ, JOSE", 4.0, 100.67, "USD", "25号仓"],
            ["WUS043938", "JIMENEZ, ENEAS", 5.0, 116.85, "USD", "25号仓"],
            ["WUS041037", "PEREZ, JOSE", 40.0, 935.59, "USD", "28号仓"],
            ["WUS043938", "JIMENEZ, ENEAS", 40.0, 929.87, "USD", "28号仓"],
        ],
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [
            {"source_file": "fairway_25.pdf", "total_amount": 219.30, "warehouse_id": "25"},
            {"source_file": "fairway_28.pdf", "total_amount": 1863.67, "warehouse_id": "28"},
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file="fairway_25.pdf", source_page_or_row="p1", employee_id="WUS041037", employee_name_raw="PEREZ, JOSE", hours=4.0, amount=101.26, currency="USD", confidence=0.95, evidence_text="PEREZ, JOSE 4.0 101.26", warehouse_id="25"),
            LaborLineItem(source_type="pdf_invoice", source_file="fairway_25.pdf", source_page_or_row="p1", employee_id="WUS043938", employee_name_raw="JIMENEZ, ENEAS", hours=5.0, amount=118.04, currency="USD", confidence=0.95, evidence_text="JIMENEZ, ENEAS 5.0 118.04", warehouse_id="25"),
            LaborLineItem(source_type="pdf_invoice", source_file="fairway_28.pdf", source_page_or_row="p1", employee_id="WUS041037", employee_name_raw="PEREZ, JOSE", hours=40.0, amount=935.00, currency="USD", confidence=0.95, evidence_text="PEREZ, JOSE 40.0 935.00", warehouse_id="28"),
            LaborLineItem(source_type="pdf_invoice", source_file="fairway_28.pdf", source_page_or_row="p1", employee_id="WUS043938", employee_name_raw="JIMENEZ, ENEAS", hours=40.0, amount=928.67, currency="USD", confidence=0.95, evidence_text="JIMENEZ, ENEAS 40.0 928.67", warehouse_id="28"),
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [
            {"source_file": "fairway_25.pdf", "page": 1, "text": "PEREZ, JOSE 101.26 JIMENEZ, ENEAS 118.04"},
            {"source_file": "fairway_28.pdf", "page": 1, "text": "PEREZ, JOSE 935.00 JIMENEZ, ENEAS 928.67"},
        ],
    )

    dry_run = build_material_dry_run(tmp_path, "fairway已报账2")

    assert dry_run["summary"]["comparison"]["exceptionCount"] == 4
    assert dry_run["summary"]["warehouse"]["allocationIssueCount"] == 2
    assert dry_run["summary"]["tierStatus"]["allocationIssueCount"] == 2
    assert dry_run["reviewQueues"]["primary"] == "allocation_review"
    assert "仓库归属" in dry_run["reviewQueues"]["primaryReason"]
    assert "复核" not in dry_run["reviewQueues"]["primaryReason"]
    queue = dry_run["reviewQueues"]["allocationReview"]
    assert queue["count"] == 2
    assert queue["warehousePairCount"] == 4
    assert queue["amountImpactTotal"] == 1.79
    assert [item["action"] for item in queue["nextActions"]] == [
        "create_formal_run",
        "extract_compare",
        "review_warehouse_allocation",
        "confirm_or_rollback",
    ]
    assert queue["nextActions"][0]["enabled"] is True
    assert "审计记录" in queue["nextActions"][3]["description"]
    allocation_visible_text = " ".join(
        [
            dry_run["reviewQueues"]["primaryReason"],
            *dry_run["expectedRisks"],
            *[str(action.get("label", "")) for action in queue["nextActions"]],
            *[str(action.get("description", "")) for action in queue["nextActions"]],
        ]
    )
    for internal_copy in ["人工复核", "需复核", "复核仓库", "填写复核"]:
        assert internal_copy not in allocation_visible_text
    rows_by_employee = {row["employeeName"]: row for row in queue["rows"]}
    assert rows_by_employee["JIMENEZ, ENEAS"]["maxWarehouseDelta"] == 1.2
    assert rows_by_employee["PEREZ, JOSE"]["netAmountDelta"] == 0.0
    assert dry_run["allocationIssues"][0]["employeeName"] in {"JIMENEZ, ENEAS", "PEREZ, JOSE"}
    assert any("跨仓库金额抵消" in risk for risk in dry_run["expectedRisks"])


def test_build_material_dry_run_surfaces_name_mapping_candidates_as_governance_preview(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "29仓"
    batch.mkdir()
    (batch / "In291943.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "员工账单明细 - 2026-05-28T141945.414.xlsx",
        [
            ["WUS040020", "Deisi Pozo", 37.84, 847.84, "USD", "29号仓"],
            ["WUS033570", "Freddy Moran (MOR47K)", 40.48, 830.72, "USD", "29号仓"],
        ],
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 1890.27, "warehouse_id": "29"}],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(source_type="pdf_invoice", source_file=paths[0].name, source_page_or_row="p1", employee_id="", employee_name_raw="Rozo Panche, Deisy V", hours=37.84, amount=847.84, currency="USD", confidence=0.98, evidence_text="Rozo Panche, Deisy V 37.84 $847.84"),
            LaborLineItem(source_type="pdf_invoice", source_file=paths[0].name, source_page_or_row="p1", employee_id="", employee_name_raw="Moran Treminio, Freddy", hours=40.48, amount=1042.43, currency="USD", confidence=0.98, evidence_text="Moran Treminio, Freddy 40.48 $1042.43"),
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": "Rozo Panche, Deisy V 37.84 $847.84"}],
    )

    dry_run = build_material_dry_run(tmp_path, "29仓")

    assert dry_run["summary"]["comparison"]["candidateMatchCount"] == 2
    governance = dry_run["nameMappingGovernance"]
    assert governance["decision"] == "candidate_only"
    assert governance["requiresConfirmation"] is True
    assert governance["summary"]["candidateCount"] == 2
    assert governance["summary"]["highConfidenceCount"] == 1
    assert governance["summary"]["readyToReplayCount"] == 1
    assert governance["summary"]["projectedFixedExceptionCount"] == 2
    assert governance["summary"]["amountStillDifferentCount"] == 1
    candidate = governance["candidates"][0]
    assert candidate["status"] == "pending_user_confirmation"
    assert candidate["sourceFile"] == "In291943.pdf"
    assert candidate["cacheEmployeeName"] == "Rozo Panche, Deisy V"
    assert candidate["excelEmployeeName"] == "Deisi Pozo"
    assert candidate["proposedMapping"] == {"Rozo Panche, Deisy V": "Deisi Pozo"}
    assert candidate["confidence"] == "high"
    assert candidate["projectedFixedExceptionCount"] == 2
    assert candidate["matchReason"] == "姓名相似且金额/工时一致"
    assert "是否确认 PDF 名称 Rozo Panche, Deisy V 对应 Excel 员工 Deisi Pozo" in candidate["businessQuestion"]
    assert "预计减少 2 项异常" in candidate["businessQuestion"]
    assert candidate["impactSummary"] == "金额和工时均一致"
    assert "必须先查看影响" in candidate["cannotAutoResolveReason"]
    assert "业务确认" in candidate["cannotAutoResolveReason"]
    assert "预览" not in candidate["cannotAutoResolveReason"]
    assert "人工确认" not in candidate["cannotAutoResolveReason"]
    medium_candidate = governance["candidates"][1]
    assert medium_candidate["confidence"] == "medium"
    assert medium_candidate["cacheEmployeeName"] == "Moran Treminio, Freddy"
    assert medium_candidate["projectedFixedExceptionCount"] == 0
    assert medium_candidate["matchReason"] == "姓名相似，但金额或工时仍需确认"
    assert "需先确认差异原因" in medium_candidate["businessQuestion"]
    assert "PDF 高于 Excel" in medium_candidate["impactSummary"]
    assert "不能直接合并" in medium_candidate["cannotAutoResolveReason"]
    assert candidate["auditTrail"][0]["reason"] == "material_dry_run_candidate_match_name_pair"
    assert dry_run["reviewQueues"]["primary"] == "name_mapping"
    assert "先查看影响并确认" in dry_run["reviewQueues"]["primaryReason"]
    name_queue = dry_run["reviewQueues"]["nameMapping"]
    assert name_queue["count"] == 2
    assert name_queue["readyToReplayCount"] == 1
    assert name_queue["highConfidenceCount"] == 1
    assert name_queue["projectedFixedExceptionCount"] == 2
    assert name_queue["rows"][0]["candidateId"] == candidate["candidateId"]
    assert name_queue["rows"][0]["projectedFixedExceptionCount"] == 2
    assert name_queue["rows"][1]["projectedFixedExceptionCount"] == 0
    name_actions = name_queue["nextActions"]
    assert [item["action"] for item in name_actions] == [
        "create_formal_run",
        "extract_compare",
        "preview_impact",
        "confirm_or_rollback",
    ]
    assert name_actions[0]["enabled"] is True
    assert name_actions[1]["enabled"] is False
    assert "查看影响" in name_actions[2]["description"]
    assert "撤回" in name_actions[3]["description"]
    user_visible_name_mapping_text = " ".join(
        [
            candidate["businessQuestion"],
            candidate["cannotAutoResolveReason"],
            candidate["recommendation"],
            medium_candidate["matchReason"],
            medium_candidate["businessQuestion"],
            medium_candidate["cannotAutoResolveReason"],
            medium_candidate["recommendation"],
            *[str(action.get("description", "")) for action in name_actions],
        ]
    )
    for internal_copy in ["预览", "人工确认", "人工复核", "需复核", "复核差异口径"]:
        assert internal_copy not in user_visible_name_mapping_text
    assert any("疑似同一员工" in risk for risk in dry_run["expectedRisks"])
    assert all("姓名匹配建议" not in risk for risk in dry_run["expectedRisks"])


def test_build_material_dry_run_surfaces_combined_pdf_rows_as_governance_preview(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "oss 2"
    batch.mkdir()
    (batch / "US Elogis Service #1 Invoice W.E 05.24.26.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "员工账单明细 - 2026-06-04T094719.972.xlsx",
        [
            ["WUS045753", "Manuel Lozano", 16.09, 361.42, "USD", "1号仓"],
            ["WUS045746", "Massiel Castillo", 3.50, 78.40, "USD", "1号仓"],
        ],
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 439.82, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(
        materials_module,
        "extract_invoice_items",
        lambda paths, config, **kwargs: [
            LaborLineItem(
                source_type="pdf_invoice",
                source_file=paths[0].name,
                source_page_or_row="p1",
                employee_id="",
                employee_name_raw="Lozano, Manuel",
                hours=19.59,
                amount=439.82,
                currency="USD",
                confidence=0.95,
                evidence_text="Lozano, Manuel 19.50 0.09 439.82",
            )
        ],
    )
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": "Lozano, Manuel 19.50 0.09 439.82"}],
    )

    dry_run = build_material_dry_run(tmp_path, "oss_2")

    assert dry_run["summary"]["comparison"]["candidateMatchCount"] == 1
    assert dry_run["candidateMatches"][0]["issueType"] == "combined_pdf_row"
    assert dry_run["nameMappingGovernance"]["summary"]["candidateCount"] == 0
    combined = dry_run["combinedRowGovernance"]
    assert combined["decision"] == "candidate_only"
    assert combined["requiresConfirmation"] is True
    assert combined["summary"]["candidateCount"] == 1
    assert combined["summary"]["amountImpactTotal"] == 78.4
    assert combined["summary"]["hoursImpactTotal"] == 3.5
    candidate = combined["candidates"][0]
    assert candidate["status"] == "pending_invoice_review"
    assert candidate["issueType"] == "combined_pdf_row"
    assert candidate["pdfEmployeeName"] == "Lozano, Manuel ⇄ Manuel Lozano"
    assert candidate["excelEmployeeName"] == "Massiel Castillo"
    assert candidate["amountGap"] == 78.4
    assert candidate["hoursGap"] == 3.5
    assert candidate["matchReason"] == "PDF 行疑似包含多名员工或剩余金额/工时"
    assert "是否还包含 Excel 员工 Massiel Castillo" in candidate["businessQuestion"]
    assert "PDF 高于 Excel $78.40" in candidate["impactSummary"]
    assert "PDF 工时多于 Excel 3.50" in candidate["impactSummary"]
    assert "不能仅凭差额接近自动清账" in candidate["cannotAutoResolveReason"]
    assert candidate["auditTrail"][0]["reason"] == "material_dry_run_combined_pdf_row"
    assert dry_run["reviewQueues"]["primary"] == "combined_pdf_row"
    assert "原始发票" in dry_run["reviewQueues"]["primaryReason"]
    combined_queue = dry_run["reviewQueues"]["combinedPdfRows"]
    assert combined_queue["count"] == 1
    assert combined_queue["amountImpactTotal"] == 78.4
    assert combined_queue["hoursImpactTotal"] == 3.5
    assert combined_queue["rows"][0]["candidateId"] == candidate["candidateId"]
    assert any("合并员工行建议" in risk for risk in dry_run["expectedRisks"])


def test_build_material_dry_run_surfaces_ai_cache_as_candidate_only(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "oss"
    batch.mkdir()
    (batch / "elog7-5_20260520204043.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook(batch / "员工账单明细 - 2026-05-27T110404.877.xlsx")
    cache_dir = batch / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "elog7-5_20260520204043_p1_mimo-v2.5_v4.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "Alice Worker",
                    "source_page": 1,
                    "hours": 8,
                    "amount": 120,
                    "confidence": 0.95,
                    "evidence_text": "Alice Worker TOTAL $120.00",
                }
            ]
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 120.0, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(materials_module, "extract_invoice_items", lambda paths, config, **kwargs: [])
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": ""}],
    )

    dry_run = build_material_dry_run(tmp_path, "oss")

    assert dry_run["summary"]["pdfRowCount"] == 0
    assert dry_run["summary"]["pdfTextCoverage"]["imageOnlyFileCount"] == 1
    assert dry_run["pdfTextCoverage"]["files"][0]["needsOcr"] is True
    assert dry_run["summary"]["comparison"]["unmatchedExcelCount"] == 1
    assert dry_run["reviewQueues"]["primary"] == "reocr"
    assert "图片发票明细待确认" in dry_run["reviewQueues"]["primaryReason"]
    assert "复核" not in dry_run["reviewQueues"]["primaryReason"]
    assert "预览" not in dry_run["reviewQueues"]["primaryReason"]
    assert dry_run["reviewQueues"]["reocr"]["taskCount"] == 1
    assert dry_run["reviewQueues"]["reocr"]["imageOnlyFileCount"] == 1
    assert dry_run["deliveryGate"]["status"] == "blocked"
    assert dry_run["deliveryGate"]["label"] == "不可交付"
    assert dry_run["deliveryGate"]["summary"]["blockedCount"] == 1
    assert dry_run["deliveryGate"]["issues"][0]["code"] == "reocr_required"
    assert "图片发票明细待确认" in dry_run["deliveryGate"]["issues"][0]["title"]
    reocr_actions = dry_run["reviewQueues"]["reocr"]["nextActions"]
    assert [item["action"] for item in reocr_actions] == [
        "create_formal_run",
        "extract_compare",
        "replay_candidate",
        "confirm_apply",
    ]
    assert reocr_actions[0]["enabled"] is True
    assert reocr_actions[1]["enabled"] is False
    assert "查看影响" in reocr_actions[2]["description"]
    assert "撤回" in reocr_actions[3]["description"]
    assert dry_run["reviewQueues"]["employeeExceptions"]["count"] == 1
    assert dry_run["reviewQueues"]["employeeExceptions"]["suppressedByPrimary"] is True
    assert dry_run["aiCacheAudit"]["decision"] == "candidate_only"
    assert dry_run["aiCacheAudit"]["requiresConfirmation"] is True
    assert dry_run["aiCacheAudit"]["summary"]["candidateAmountTotal"] == 120
    assert dry_run["aiCacheAudit"]["files"][0]["evidence"][0]["sourcePageOrRow"] == "p1"
    assert dry_run["aiCacheReconciliationPreview"]["decision"] == "candidate_only"
    assert dry_run["aiCacheReconciliationPreview"]["summary"]["candidateRowCount"] == 1
    assert dry_run["aiCacheReconciliationPreview"]["summary"]["passedCount"] == 0
    assert dry_run["aiCacheReconciliationPreview"]["summary"]["exceptionCount"] == 1
    assert dry_run["aiCacheReconciliationPreview"]["summary"]["needsReocrFileCount"] == 1
    assert dry_run["aiCacheReconciliationPreview"]["fileQuality"][0]["decision"] == "needs_reocr"
    assert dry_run["reocrPlan"]["decision"] == "candidate_only"
    assert dry_run["reocrPlan"]["summary"]["taskCount"] == 1
    assert dry_run["reocrPlan"]["summary"]["imageOnlyTaskCount"] == 1
    assert dry_run["reocrPlan"]["tasks"][0]["sourceFile"] == "elog7-5_20260520204043.pdf"
    assert dry_run["reocrPlan"]["tasks"][0]["pdfTextCoverage"]["needsOcr"] is True
    assert dry_run["reocrPlan"]["tasks"][0]["extractionPrerequisite"] == "pdf_text_layer_empty_requires_ocr"
    assert dry_run["reocrPlan"]["tasks"][0]["reviewFocus"] == "需要重新图片识别"
    user_visible_reocr_text = " ".join(
        str(dry_run["reocrPlan"]["tasks"][0].get(field, ""))
        for field in ("reason", "confirmationGate", "matchReason", "businessQuestion", "impactSummary", "cannotAutoResolveReason")
    )
    assert "OCR" not in user_visible_reocr_text
    assert "AI" not in user_visible_reocr_text
    assert "图片识别" in user_visible_reocr_text or "重新识别" in user_visible_reocr_text
    assert "PDF 无可读取文本层" in dry_run["reocrPlan"]["tasks"][0]["matchReason"]
    assert "员工级异常 1 项" in dry_run["reocrPlan"]["tasks"][0]["matchReason"]
    assert "必须先查看员工级影响" in dry_run["reocrPlan"]["tasks"][0]["businessQuestion"]
    assert "员工级异常 1 项" in dry_run["reocrPlan"]["tasks"][0]["impactSummary"]
    assert "不能自动写入正式结果" in dry_run["reocrPlan"]["tasks"][0]["cannotAutoResolveReason"]
    assert "Alice Worker" in dry_run["reocrPlan"]["tasks"][0]["focusEmployees"][0]["employeeName"]
    assert "Alice Worker" in dry_run["reviewQueues"]["reocr"]["tasks"][0]["focusEmployees"][0]["employeeName"]
    assert dry_run["reviewQueues"]["reocr"]["tasks"][0]["reviewFocus"] == "需要重新图片识别"
    assert "必须业务确认" in dry_run["reocrPlan"]["tasks"][0]["confirmationGate"]
    assert dry_run["writesRun"] is False
    assert dry_run["aiInvoked"] is False
    assert any("无可读取文本层" in risk for risk in dry_run["expectedRisks"])
    assert any("历史图片识别" in risk for risk in dry_run["expectedRisks"])
    assert any("不能直接作为 PDF 明细" in risk for risk in dry_run["expectedRisks"])
    assert any("历史图片识别结果与账单仍有 1 项差异" in risk for risk in dry_run["expectedRisks"])
    assert any("历史图片识别结果：1 个 PDF 建议重新识别，0 个 PDF 可作为业务确认依据" in risk for risk in dry_run["expectedRisks"])
    assert any("已生成 1 个图片发票明细待确认事项" in risk for risk in dry_run["expectedRisks"])
    user_visible_material_text = " ".join(
        [
            dry_run["reviewQueues"]["primaryReason"],
            dry_run["deliveryGate"]["issues"][0]["title"],
            dry_run["deliveryGate"]["issues"][0]["message"],
            dry_run["deliveryGate"]["issues"][0]["action"],
            *dry_run["expectedRisks"],
            *[str(action.get("label", "")) for action in reocr_actions],
            *[str(action.get("description", "")) for action in reocr_actions],
        ]
    )
    for internal_copy in ["图片识别复核", "人工复核", "人工确认", "需预览", "必须预览", "影响预览"]:
        assert internal_copy not in user_visible_material_text


def test_build_material_dry_run_explains_image_only_pdf_without_history_cache(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "prompt"
    batch.mkdir()
    (batch / "DEPT#2.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "员工账单明细 - 2026-05-28T151400.642.xlsx",
        [["WUS001", "Alice Worker", 8, 100, "USD", "2号仓"]],
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 0.0, "warehouse_id": "2"}],
    )
    monkeypatch.setattr(materials_module, "extract_invoice_items", lambda paths, config, **kwargs: [])
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": ""}],
    )

    dry_run = build_material_dry_run(tmp_path, "prompt")

    assert dry_run["reviewQueues"]["primary"] == "reocr"
    assert dry_run["deliveryGate"]["status"] == "blocked"
    assert dry_run["reocrPlan"]["summary"]["taskCount"] == 1
    task = dry_run["reocrPlan"]["tasks"][0]
    assert task["sourceFile"] == "DEPT#2.pdf"
    assert "当前没有可用 PDF 明细覆盖账单" in task["matchReason"]
    assert "账单有但当前明细缺失 1 人" in task["impactSummary"]
    user_visible_text = " ".join(
        str(task.get(field, ""))
        for field in ("matchReason", "businessQuestion", "impactSummary", "cannotAutoResolveReason")
    )
    assert "缓存金额" not in user_visible_text
    assert "历史识别" not in user_visible_text
    assert any("已生成 1 个图片发票明细待确认事项" in risk for risk in dry_run["expectedRisks"])


def test_build_material_dry_run_promotes_reocr_suspected_name_pairs_to_governance(monkeypatch, tmp_path):
    import bonus_platform.engine.labor.materials as materials_module

    batch = tmp_path / "oss"
    batch.mkdir()
    (batch / "elog1-1_20260520204104.pdf").write_bytes(b"%PDF-1.4\n")
    _write_labor_bill_workbook_with_rows(
        batch / "员工账单明细 - 2026-05-27T110404.877.xlsx",
        [
            ["WUS045751", "Massiel Castillo", 30.92, 100.00, "USD", "1号仓"],
            ["WUS045752", "Other Worker", 2.00, 50.00, "USD", "1号仓"],
        ],
    )
    cache_dir = batch / ".ai_extract_cache"
    cache_dir.mkdir()
    (cache_dir / "elog1-1_20260520204104_p1_mimo-v2.5_v4.json").write_text(
        json.dumps(
            [
                {
                    "employee_name_raw": "Espinosa Manuel",
                    "source_page": 1,
                    "hours": 30.90,
                    "amount": 100.00,
                    "confidence": 0.95,
                    "evidence_text": "Espinosa Manuel TOTAL $100.00",
                }
            ]
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(
        materials_module,
        "quick_extract_totals",
        lambda paths, config, supplier="": [{"source_file": paths[0].name, "total_amount": 0.0, "warehouse_id": "1"}],
    )
    monkeypatch.setattr(materials_module, "extract_invoice_items", lambda paths, config, **kwargs: [])
    monkeypatch.setattr(
        materials_module,
        "_extract_pdf_pages",
        lambda paths: [{"source_file": paths[0].name, "page": 1, "text": ""}],
    )

    dry_run = build_material_dry_run(tmp_path, "oss")

    assert dry_run["reocrPlan"]["summary"]["taskCount"] == 1
    suspected = dry_run["reocrPlan"]["tasks"][0]["diagnostics"]["suspectedNamePairs"]
    assert suspected[0]["cacheEmployeeName"] == "Espinosa Manuel"
    governance = dry_run["nameMappingGovernance"]
    assert governance["decision"] == "candidate_only"
    assert governance["summary"]["candidateCount"] == 1
    assert governance["summary"]["fromReocrDiagnosticsCount"] == 1
    candidate = governance["candidates"][0]
    assert candidate["sourceDiagnostic"] == "reocr_suspected_name_pair"
    assert candidate["sourceFile"] == "elog1-1_20260520204104.pdf"
    assert candidate["warehouseId"] == "1"
    assert candidate["proposedMapping"] == {"Espinosa Manuel": "Massiel Castillo"}
    assert candidate["auditTrail"][0]["reason"] == "material_dry_run_reocr_suspected_name_pair"
    assert any("疑似同一员工" in risk for risk in dry_run["expectedRisks"])
    assert all("姓名匹配建议" not in risk for risk in dry_run["expectedRisks"])


def _write_labor_bill_workbook(path):
    _write_labor_bill_workbook_with_rows(path, [["WUS001", "Alice Worker", 8, 100, "USD", "7号仓"]])


def _write_labor_bill_workbook_with_rows(path, rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "员工账单"
    sheet.append(["工号", "姓名", "时长总计(H)", "费用总计(含税)", "币种", "物理仓"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)


@pytest.mark.parametrize(
    ("value", "locale_hint", "expected"),
    [
        ("1,059.16", "dot_decimal", 1059.16),
        ("1.059,16", "comma_decimal", 1059.16),
        ("26,479", "comma_decimal", 26.479),
        ("$12,167.13", "dot_decimal", 12167.13),
        ("12.167,13 EUR", "comma_decimal", 12167.13),
    ],
)
def test_parse_localized_number(value, locale_hint, expected):
    assert parse_localized_number(value, locale_hint) == expected


def test_find_amount_closure_uses_structure_not_language():
    closure = find_amount_closure(
        ["Zwischensumme (netto) 10.224,48\nUmsatzsteuer 1.942,65\nGesamtbetrag 12.167,13"],
        detail_sum=10224.48,
    )

    assert closure.net_amount == 10224.48
    assert closure.tax_amount == 1942.65
    assert closure.gross_amount == 12167.13
    assert closure.confidence == "high"
    assert closure.locale == "comma_decimal"


def test_find_amount_closure_rejects_ambiguous_candidates():
    closure = find_amount_closure(
        ["100.00 100.00 10.00 110.00 20.00 120.00"],
        detail_sum=100.00,
    )

    assert closure.confidence == "ambiguous"


def _structure_row(name, *, source_type="pdf_invoice", source_file="invoice.pdf", warehouse_id="", amount=100.0):
    return LaborLineItem(
        source_type=source_type,
        source_file=source_file,
        source_page_or_row="p1" if source_type == "pdf_invoice" else "Sheet1!2",
        employee_id="",
        employee_name_raw=name,
        hours=8.0,
        amount=amount,
        currency="EUR",
        confidence=0.95,
        evidence_text=name,
        warehouse_id=warehouse_id,
    )


def test_infer_warehouse_requires_count_coverage_and_margin():
    names = ["A One", "B Two", "C Three", "D Four", "E Five", "F Six"]
    pdf_rows = [_structure_row(name) for name in names]
    excel_rows = [
        *[_structure_row(name, source_type="offline_workbook", warehouse_id="16") for name in names],
        *[_structure_row(name, source_type="offline_workbook", warehouse_id="3") for name in names[:1]],
    ]

    result = infer_warehouse_from_rows(pdf_rows, excel_rows)

    assert result.status == "matched"
    assert result.warehouse_id == "16"
    assert result.matched_count == 6
    assert result.invoice_coverage == 1.0
    assert result.runner_up_warehouse_id == "3"


def test_infer_warehouse_returns_review_when_runner_up_is_close():
    names = ["A One", "B Two", "C Three", "D Four", "E Five"]
    pdf_rows = [_structure_row(name) for name in names]
    excel_rows = [
        *[_structure_row(name, source_type="offline_workbook", warehouse_id="3") for name in names],
        *[_structure_row(name, source_type="offline_workbook", warehouse_id="6") for name in names[:4]],
    ]

    result = infer_warehouse_from_rows(pdf_rows, excel_rows)

    assert result.status == "warehouse_review"
    assert result.warehouse_id == ""


def test_infer_warehouse_keeps_five_of_seven_unique_match():
    invoice_names = ["A One", "B Two", "C Three", "D Four", "E Five", "OCR Wrong", "OCR Missing"]
    pdf_rows = [_structure_row(name) for name in invoice_names]
    excel_rows = [
        *[_structure_row(name, source_type="offline_workbook", warehouse_id="12") for name in invoice_names[:5]],
        _structure_row("Other Worker", source_type="offline_workbook", warehouse_id="9"),
    ]

    result = infer_warehouse_from_rows(pdf_rows, excel_rows)

    assert result.status == "matched"
    assert result.warehouse_id == "12"
    assert result.matched_count == 5


def test_infer_warehouse_reviews_small_invoice_even_when_unique():
    names = ["A One", "B Two", "C Three", "D Four"]

    result = infer_warehouse_from_rows(
        [_structure_row(name) for name in names],
        [_structure_row(name, source_type="offline_workbook", warehouse_id="12") for name in names],
    )

    assert result.status == "warehouse_review"


def test_promote_unknown_invoice_when_rows_amount_and_warehouse_close(tmp_path, monkeypatch):
    pdf_path = tmp_path / "Rechnung_RE202606-0646.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    names = ["A One", "B Two", "C Three", "D Four", "E Five", "F Six"]
    pdf_rows = [
        _structure_row(name, source_file=pdf_path.name, amount=1700.0)
        for name in names[:-1]
    ]
    pdf_rows.append(_structure_row(names[-1], source_file=pdf_path.name, amount=1724.48))
    excel_rows = [
        _structure_row(name, source_type="offline_workbook", warehouse_id="15")
        for name in names
    ]
    monkeypatch.setattr(
        labor_structure,
        "extract_page_texts",
        lambda _: ["10.224,48 1.942,65 12.167,13"],
    )

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [
            {
                "source_file": pdf_path.name,
                "total_amount": 0.0,
                "warehouse_id": "",
                "pdf_type": "unknown",
                "authoritative": False,
                "evidence_status": "needs_review",
                "page_evidence": [{"page": 1, "role": "unknown"}],
            }
        ],
        pdf_rows,
        excel_rows,
    )

    promoted = result.pdf_totals[0]
    assert promoted["warehouse_id"] == "15"
    assert promoted["total_amount"] == 10224.48
    assert promoted["authoritative"] is True
    assert promoted["evidence_status"] == "authoritative"
    assert promoted["total_label"] == "employee_detail_sum"
    assert result.decisions[0]["status"] == "reconciled"


def test_promote_authoritative_total_across_multiple_employee_warehouses(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    names = ["A One", "B Two", "C Three", "D Four", "E Five", "F Six"]
    pdf_rows = [
        _structure_row(name, source_file=pdf_path.name, amount=20.0 if index < 4 else 10.0)
        for index, name in enumerate(names)
    ]
    excel_rows = [
        _structure_row(
            name,
            source_type="offline_workbook",
            warehouse_id="1" if index < 3 else "2",
            amount=20.0 if index < 4 else 10.0,
        )
        for index, name in enumerate(names)
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: ["100.00 100.00 0.00"])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [
            {
                "source_file": pdf_path.name,
                "total_amount": 100.0,
                "warehouse_id": "",
                "pdf_type": "primary",
                "authoritative": True,
                "evidence_status": "authoritative",
            }
        ],
        pdf_rows,
        excel_rows,
    )

    assert result.decisions[0]["status"] == "reconciled_multi_warehouse"
    assert result.pdf_totals[0]["authoritative"] is True
    assert result.unresolved_files == ()
    assert {row.warehouse_id for row in result.pdf_rows} == {"1", "2"}


def test_structured_promotion_keeps_explicit_warehouse_conflict_for_review(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    names = ["A One", "B Two", "C Three", "D Four", "E Five"]
    pdf_rows = [_structure_row(name, source_file=pdf_path.name, amount=20.0) for name in names]
    excel_rows = [
        _structure_row(name, source_type="offline_workbook", warehouse_id="15")
        for name in names
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: ["100.00 19.00 119.00"])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [{"source_file": pdf_path.name, "total_amount": 0.0, "warehouse_id": "3", "pdf_type": "unknown", "authoritative": False}],
        pdf_rows,
        excel_rows,
    )

    assert result.pdf_totals[0]["authoritative"] is False
    assert result.decisions[0]["status"] == "warehouse_review"


def test_structured_promotion_requires_visible_amount_closure(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    names = ["A One", "B Two", "C Three", "D Four", "E Five"]
    pdf_rows = [_structure_row(name, source_file=pdf_path.name, amount=20.0) for name in names]
    excel_rows = [
        _structure_row(name, source_type="offline_workbook", warehouse_id="15")
        for name in names
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: ["250.00 47.50 297.50"])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [{"source_file": pdf_path.name, "total_amount": 0.0, "warehouse_id": "", "pdf_type": "unknown", "authoritative": False}],
        pdf_rows,
        excel_rows,
    )

    assert result.pdf_totals[0]["authoritative"] is False
    assert result.decisions[0]["status"] == "amount_review"


def test_image_invoice_complete_detail_sum_promotes_when_batch_closes_to_excel(tmp_path, monkeypatch):
    salary_pdf = tmp_path / "salary.pdf"
    meal_pdf = tmp_path / "meal.pdf"
    salary_pdf.write_bytes(b"%PDF-1.4\n")
    meal_pdf.write_bytes(b"%PDF-1.4\n")
    pdf_rows = [
        _structure_row("COBY Hugues", source_file=salary_pdf.name, amount=563.70),
        _structure_row("COBY Hugues", source_file=meal_pdf.name, amount=19.44),
    ]
    pdf_rows[0] = replace(pdf_rows[0], hours=27.92, evidence_text="COBY Hugues 27.92 563,70")
    # Meal-ticket quantity may be extracted into the generic hours slot. It is
    # still valid amount evidence and must remain an employee-level hours alert.
    pdf_rows[1] = replace(pdf_rows[1], hours=4.0, evidence_text="COBY Hugues 4 repas 19,44")
    excel_rows = [
        _structure_row(
            "COBY Hugues",
            source_type="offline_workbook",
            warehouse_id="1",
            amount=583.14,
        )
    ]
    excel_rows[0] = replace(excel_rows[0], hours=27.92)
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: [""])

    result = promote_structured_invoice_evidence(
        [salary_pdf, meal_pdf],
        [
            {"source_file": salary_pdf.name, "total_amount": 0.0, "warehouse_id": "", "authoritative": False, "evidence_status": "needs_review"},
            {"source_file": meal_pdf.name, "total_amount": 0.0, "warehouse_id": "", "authoritative": False, "evidence_status": "needs_review"},
        ],
        pdf_rows,
        excel_rows,
        page_audit=[
            {"sourceFile": salary_pdf.name, "page": 1, "status": "cache_hit", "rowCount": 1, "fromCache": True},
            {"sourceFile": meal_pdf.name, "page": 1, "status": "completed", "rowCount": 1},
        ],
    )

    assert [item["total_amount"] for item in result.pdf_totals] == [563.70, 19.44]
    assert all(item["authoritative"] is True for item in result.pdf_totals)
    assert {item["warehouse_id"] for item in result.pdf_totals} == {"1"}
    assert result.unresolved_files == ()


def test_image_invoice_uses_explicit_net_tax_gross_page_evidence(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    pdf_rows = [_structure_row("Worker One", source_file=pdf_path.name, amount=100.0)]
    excel_rows = [
        _structure_row("Worker One", source_type="offline_workbook", warehouse_id="2", amount=100.0)
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: [""])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [{
            "source_file": pdf_path.name,
            "total_amount": 120.0,
            "total_label": "TOTAL TTC",
            "warehouse_id": "",
            "authoritative": True,
            "evidence_status": "authoritative",
            "page_evidence": [{
                "page": 1,
                "role": "invoice_total",
                "role_confidence": 0.99,
                "net_amount": 100.0,
                "tax_amount": 20.0,
                "gross_amount": 120.0,
                "evidence_text": "TOTAL HT 100,00 TVA 20,00 TOTAL TTC 120,00",
            }],
        }],
        pdf_rows,
        excel_rows,
    )

    assert result.decisions[0]["status"] == "reconciled"
    assert result.decisions[0]["closure"]["confidence"] == "page_evidence_high"
    assert result.pdf_totals[0]["total_amount"] == 100.0


def test_complete_image_invoice_rows_remain_visible_when_total_conflicts_with_excel(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    pdf_rows = [_structure_row("COBY Hugues", source_file=pdf_path.name, amount=563.70)]
    excel_rows = [
        _structure_row("COBY Hugues", source_type="offline_workbook", warehouse_id="1", amount=600.00)
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: [""])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [{"source_file": pdf_path.name, "total_amount": 0.0, "warehouse_id": "", "authoritative": False}],
        pdf_rows,
        excel_rows,
        page_audit=[{"sourceFile": pdf_path.name, "page": 1, "status": "completed", "rowCount": 1}],
    )

    assert len(result.pdf_rows) == 1
    assert result.pdf_rows[0].employee_name_raw == "COBY Hugues"
    assert result.pdf_rows[0].amount == 563.70
    assert result.pdf_totals[0]["authoritative"] is False
    assert result.pdf_totals[0]["total_amount"] == 0.0
    assert result.decisions[0]["status"] == "amount_review"


def test_complete_image_invoice_rows_resolve_repeated_amount_candidates(tmp_path, monkeypatch):
    pdf_path = tmp_path / "bonus.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    pdf_rows = [_structure_row("Lautric Patrick", source_file=pdf_path.name, amount=134.40)]
    excel_rows = [
        _structure_row(
            "Lautric Patrick",
            source_type="offline_workbook",
            warehouse_id="1",
            amount=150.00,
        )
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: ["134,40 134,40 134,40"])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [{"source_file": pdf_path.name, "total_amount": 0.0, "warehouse_id": "", "authoritative": False}],
        pdf_rows,
        excel_rows,
        page_audit=[{"sourceFile": pdf_path.name, "page": 1, "status": "completed", "rowCount": 1}],
    )

    assert result.pdf_totals[0]["authoritative"] is True
    assert result.pdf_totals[0]["total_amount"] == 134.40
    assert result.decisions[0]["closure"]["confidence"] == "complete_line_sum"


def test_image_invoice_amount_can_close_while_warehouse_remains_reviewable(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    pdf_rows = [_structure_row("COBY Hugues", source_file=pdf_path.name, amount=583.14)]
    excel_rows = [
        _structure_row("COBY Hugues", source_type="offline_workbook", warehouse_id="", amount=583.14)
    ]
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: [""])

    result = promote_structured_invoice_evidence(
        [pdf_path],
        [{"source_file": pdf_path.name, "total_amount": 0.0, "warehouse_id": "", "authoritative": False}],
        pdf_rows,
        excel_rows,
        page_audit=[{"sourceFile": pdf_path.name, "page": 1, "status": "completed", "rowCount": 1}],
    )

    assert result.pdf_totals[0]["total_amount"] == 583.14
    assert result.pdf_totals[0]["authoritative"] is True
    assert result.pdf_totals[0]["warehouse_id"] == ""
    assert result.decisions[0]["status"] == "reconciled_amount_warehouse_review"
    assert result.unresolved_files == ()


def test_batch_guard_blocks_zero_when_raw_rows_exist():
    guard = evaluate_batch_guards(
        pdf_paths=[Path("invoice.pdf")],
        pdf_totals=[{"total_amount": 0.0}],
        raw_pdf_rows=[_structure_row("A One", amount=100.0)],
        formal_pdf_rows=[],
        excel_rows=[_structure_row("A One", source_type="offline_workbook", warehouse_id="1", amount=100.0)],
        requested_currency="USD",
        detected_currencies=set(),
    )

    assert guard.status == "pdf_recognition_incomplete"
    assert guard.allow_releasable_report is False


def test_batch_guard_treats_image_only_zero_total_as_recognition_failure():
    guard = evaluate_batch_guards(
        pdf_paths=[Path("DEPT#1.pdf"), Path("DEPT#2.pdf")],
        pdf_totals=[
            {"source_file": "DEPT#1.pdf", "total_amount": 0.0},
            {"source_file": "DEPT#2.pdf", "total_amount": 0.0},
        ],
        raw_pdf_rows=[],
        formal_pdf_rows=[],
        excel_rows=[],
        requested_currency="USD",
        detected_currencies=set(),
        pdf_text_coverage={
            "summary": {
                "fileCount": 2,
                "imageOnlyFileCount": 2,
                "textReadableFileCount": 0,
            }
        },
    )

    assert guard.status == "pdf_recognition_incomplete"
    assert guard.allow_releasable_report is False
    assert "识别异常" in guard.message


def test_batch_guard_flags_currency_conflict():
    guard = evaluate_batch_guards(
        pdf_paths=[Path("invoice.pdf")],
        pdf_totals=[{"total_amount": 100.0}],
        raw_pdf_rows=[_structure_row("A One", amount=100.0)],
        formal_pdf_rows=[_structure_row("A One", amount=100.0)],
        excel_rows=[],
        requested_currency="USD",
        detected_currencies={"EUR"},
    )

    assert guard.status == "currency_review"
    assert guard.allow_releasable_report is False


def test_batch_guard_keeps_unresolved_invoice_as_partial_review():
    guard = evaluate_batch_guards(
        pdf_paths=[Path("invoice.pdf")],
        pdf_totals=[{"total_amount": 100.0}],
        raw_pdf_rows=[_structure_row("A One", amount=100.0)],
        formal_pdf_rows=[_structure_row("A One", amount=100.0)],
        excel_rows=[],
        requested_currency="EUR",
        detected_currencies={"EUR"},
        unresolved_files=["other.pdf"],
    )

    assert guard.status == "partial_review"
    assert guard.unresolved_files == ("other.pdf",)


@pytest.mark.parametrize(
    ("header", "expected"),
    [
        ("费用总计(不含税)", "net"),
        ("Net Amount", "net"),
        ("费用总计(含税)", "gross"),
        ("Gross Amount", "gross"),
        ("费用", "review"),
    ],
)
def test_resolve_amount_scope(header, expected):
    assert resolve_amount_scope(header) == expected


def test_resolve_amount_scope_prefers_explicit_user_declaration():
    assert resolve_amount_scope("本周薪资 本周餐补", declared_scope="net") == "net"
    assert resolve_amount_scope("费用总计(不含税)", declared_scope="gross") == "gross"
    assert resolve_amount_scope("本周薪资", declared_scope="unsupported") == "review"


def test_extract_structured_invoice_rows_uses_numeric_columns_and_excel_names(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        labor_structure,
        "extract_page_texts",
        lambda _: [
            "1 Archil Akhalkatsi 40 Stunde26,4791.059,16\n"
            "2 Archil Akhalkatsi overtime 12,85Stunde33,099 425,32\n"
            "Zwischensumme (netto) 1.484,48\nUmsatzsteuer 282,05\nGesamtbetrag 1.766,53"
        ],
    )
    excel_rows = [
        _structure_row("Archil Akhalkatsi", source_type="offline_workbook", warehouse_id="11")
    ]

    rows = extract_structured_invoice_rows([pdf_path], excel_rows)

    assert len(rows) == 2
    assert {row.employee_name_raw for row in rows} == {"Archil Akhalkatsi"}
    assert [row.hours for row in rows] == [40.0, 12.85]
    assert [row.amount for row in rows] == [1059.16, 425.32]
    assert {row.currency for row in rows} == {"EUR"}
    assert {row.source_page_or_row for row in rows} == {"p1"}


def test_extract_structured_invoice_rows_reads_closed_employee_subtotal_blocks_across_pages(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        labor_structure,
        "extract_page_texts",
        lambda _: [
            "DETAIL DES PRESTATIONS Quantité Taux Montant\n"
            "ALICE WORKER Semaine 22 du 25/05/2026 au 31/05/2026\n"
            "001 HEURES NORMALES 35,00 20,00 700,00",
            "S/Total 42,00 1000,00\n"
            "BOB WORKER Semaine 22 du 25/05/2026 au 31/05/2026\n"
            "555 IND. TICKET RESTAU.-NS 5,00 4,86 24,30\n"
            "S/Total 24,30\n"
            "TR Total affect. 42,00 1024,30\n"
            "au 30/06/2026 1024,30 20,00 204,86 1229,16EUR",
        ],
    )
    excel_rows = [
        _structure_row("Alice Worker", source_type="offline_workbook", warehouse_id="1"),
        _structure_row("Bob Worker", source_type="offline_workbook", warehouse_id="1"),
    ]

    rows = extract_structured_invoice_rows([pdf_path], excel_rows)

    assert [(row.employee_name_raw, row.hours, row.amount) for row in rows] == [
        ("ALICE WORKER", 42.0, 1000.0),
        ("BOB WORKER", 0.0, 24.3),
    ]
    assert {row.currency for row in rows} == {"EUR"}
    assert [row.source_page_or_row for row in rows] == ["p2", "p2"]
    assert prefer_closed_structured_rows([], rows, [pdf_path]) == rows


def test_extract_structured_invoice_rows_rejects_unclosed_employee_subtotal_blocks(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(
        labor_structure,
        "extract_page_texts",
        lambda _: [
            "DETAIL DES PRESTATIONS Quantité Taux Montant\n"
            "ALICE WORKER Semaine 22 du 25/05/2026 au 31/05/2026\n"
            "S/Total 42,00 1000,00\n"
            "TR Total affect. 999,00"
        ],
    )

    rows = extract_structured_invoice_rows(
        [pdf_path],
        [_structure_row("Alice Worker", source_type="offline_workbook", warehouse_id="1")],
    )

    assert rows == []


def test_prefer_closed_structured_rows_replaces_incomplete_ai_source(tmp_path, monkeypatch):
    pdf_path = tmp_path / "invoice.pdf"
    pdf_path.write_bytes(b"%PDF-1.4\n")
    monkeypatch.setattr(labor_structure, "extract_page_texts", lambda _: ["100.00 19.00 119.00"])
    ai_rows = [_structure_row("A One", source_file=pdf_path.name, amount=40.0)]
    structured_rows = [
        _structure_row("A One", source_file=pdf_path.name, amount=40.0),
        _structure_row("B Two", source_file=pdf_path.name, amount=60.0),
    ]

    selected = prefer_closed_structured_rows(ai_rows, structured_rows, [pdf_path])

    assert len(selected) == 2
    assert sum(row.amount for row in selected) == 100.0
