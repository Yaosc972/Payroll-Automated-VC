import asyncio
from io import BytesIO
import json
import sys
import time
from types import ModuleType, SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl import Workbook
import pytest

import bonus_platform.app as app_module
from bonus_platform.engine.fbu_performance.engines.base import EmployeeData
from bonus_platform.engine.fbu_performance.runs import FBURosterStore, FBURunManager


pytestmark = pytest.mark.usefixtures("bypass_fbu_access_gate")


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _roster_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["header"] * 123)
    row = [""] * 123
    row[0] = "Ana Roster"
    row[3] = "E001"
    row[19] = "HRAS人力综合条线"
    row[20] = "FBU HRBP Dept."
    row[89] = "US-West"
    row[122] = "蓝领"
    sheet.append(row)
    return _workbook_bytes(workbook)


def _attendance_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["header"] * 118)
    row = [""] * 118
    row[0] = "2026-04-01"
    row[1] = "Ana Attendance"
    row[2] = "E001"
    row[21] = "08:00"
    row[117] = 8
    sheet.append(row)
    return _workbook_bytes(workbook)


def _attendance_bytes_for_rows(rows: list[tuple[str, float]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "sheet1"
    sheet.append(["header"] * 149)
    for attendance_date, base_hours in rows:
        row = [""] * 149
        row[0] = attendance_date
        row[1] = "Ana Attendance"
        row[2] = "E001"
        row[21] = "08:00"
        row[23] = 8
        row[37] = base_hours
        row[117] = base_hours
        sheet.append(row)
    return _workbook_bytes(workbook)


def test_upload_job_result_exposes_authoritative_material_core_fields():
    result = app_module._fbu_upload_job_result(
        "salary",
        {"success": True},
        [
            {"kind": "currentSalary", "originalFilename": "june.xlsx"},
            {"kind": "transferHistory", "originalFilename": "transfers.xlsx"},
        ],
    )

    assert result["coreUpdates"] == {
        "current_salary_file": "june.xlsx",
        "salary_file": "june.xlsx",
        "transfer_file": "transfers.xlsx",
    }


def test_fbu_run_list_uses_lightweight_summary_manager(monkeypatch):
    class SummaryOnlyManager:
        def list_run_summaries(self):
            return [{
                "run_id": "run_123",
                "created_at": "2026-07-30T10:00:00",
                "calc_month": "2026-06",
                "status": "completed",
                "current_step": 5,
                "total_employees": 303,
                "total_bonus": 109290.74,
                "roster_file": "roster.xlsx",
                "roster_source": "base",
                "sections": {"results": {"present": True, "count": 303}},
            }]

    monkeypatch.setattr(app_module, "fbu_run_manager", SummaryOnlyManager())
    client = TestClient(app_module.app)

    response = client.get("/api/fbu-performance/runs")

    assert response.status_code == 200
    assert response.json()["runs"] == [{
        "run_id": "run_123",
        "created_at": "2026-07-30T10:00:00",
        "calc_month": "2026-06",
        "status": "completed",
        "current_step": 5,
        "total_employees": 303,
        "total_bonus": 109290.74,
        "roster_file": "roster.xlsx",
        "roster_source": "base",
        "sections": {"results": {"present": True, "count": 303}},
    }]
    assert response.headers["server-timing"].startswith("fbu;dur=")
    assert response.headers["x-sigma-request-id"]


def test_fbu_large_json_responses_are_gzipped(monkeypatch):
    class LargeSummaryManager:
        def list_run_summaries(self):
            return [{
                "run_id": "run_123",
                "created_at": "2026-07-30T10:00:00",
                "calc_month": "2026-06",
                "status": "completed",
                "padding": "x" * 5_000,
            }]

    monkeypatch.setattr(app_module, "fbu_run_manager", LargeSummaryManager())
    client = TestClient(app_module.app)

    response = client.get(
        "/api/fbu-performance/runs",
        headers={"Accept-Encoding": "gzip"},
    )

    assert response.status_code == 200
    assert response.headers["content-encoding"] == "gzip"
    assert "Accept-Encoding" in response.headers["vary"]
    assert response.json()["runs"][0]["run_id"] == "run_123"


def test_versioned_fbu_static_assets_are_immutable():
    client = TestClient(app_module.app)

    response = client.get(
        "/fbu-performance.js?v=versioned-test-asset",
        headers={"Accept-Encoding": "gzip"},
    )

    assert response.status_code == 200
    assert response.headers["content-encoding"] == "gzip"
    assert response.headers["cache-control"] == "public, max-age=31536000, immutable"


def test_fbu_html_ignores_stale_validators_and_is_never_cached():
    client = TestClient(app_module.app)
    initial = client.get("/fbu-performance.html")

    assert initial.status_code == 200
    assert "fbu-performance.js?v=" in initial.text

    stale_headers = {
        "If-None-Match": initial.headers.get("etag", '"stale-fbu-html"'),
        "If-Modified-Since": initial.headers.get(
            "last-modified",
            "Thu, 30 Jul 2026 00:00:00 GMT",
        ),
    }
    refreshed = client.get("/fbu-performance.html", headers=stale_headers)

    assert refreshed.status_code == 200
    assert refreshed.headers["cache-control"] == "no-store"
    assert refreshed.headers["pragma"] == "no-cache"
    assert "fbu-performance.js?v=" in refreshed.text


def test_fbu_result_pages_read_precomputed_view_instead_of_full_results(monkeypatch):
    result_row = {
        "employee_id": "zt1",
        "source_employee_id": "zt1",
        "name": "Test User",
        "department": "FBU-新泽西",
        "position": "Picker",
        "job_type": "warehouse",
        "calculation_path": "标准绩效基数路径",
        "performance_base": 1_000,
        "performance_ratio": 0.1,
        "performance_coefficient": 1,
        "performance_bonus": 100,
    }

    class ResultViewManager:
        def __init__(self):
            self.requested_sections = []

        def get_run(self, run_id, sections=None):
            self.requested_sections.append(sections)
            return SimpleNamespace(
                run_id=run_id,
                status="completed",
                results=[result_row],
                results_view_data={"rows": [result_row]},
            )

    manager = ResultViewManager()
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    client = TestClient(app_module.app)

    response = client.get(
        "/api/fbu-performance/runs/run_123/results",
        params={"page": 1, "page_size": 50, "group": "all"},
    )

    assert response.status_code == 200
    assert response.json()["results"][0]["employee_id"] == "zt1"
    assert manager.requested_sections == [{"results_view_data"}]


def test_fbu_result_pages_backfill_compact_view_for_legacy_runs(monkeypatch):
    result_row = {
        "employee_id": "zt1",
        "name": "Legacy User",
        "department": "FBU-新泽西",
        "position": "Picker",
        "job_type": "warehouse",
        "performance_base": 1_000,
        "performance_ratio": 0.1,
        "performance_coefficient": 1,
        "performance_bonus": 100,
    }

    class LegacyResultManager:
        def __init__(self):
            self.requested_sections = []
            self.backfilled = []

        def get_run(self, run_id, sections=None):
            self.requested_sections.append(sections)
            if sections == {"results_view_data"}:
                return SimpleNamespace(
                    run_id=run_id,
                    status="completed",
                    results=[],
                    results_view_data={},
                )
            return SimpleNamespace(
                run_id=run_id,
                status="completed",
                results=[result_row],
                results_view_data={},
            )

        def backfill_results_view_data(self, run_id, results_view_data):
            self.backfilled.append((run_id, results_view_data))

    manager = LegacyResultManager()
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    client = TestClient(app_module.app)

    response = client.get(
        "/api/fbu-performance/runs/run_legacy/results",
        params={"page": 1, "page_size": 50, "group": "all"},
    )

    assert response.status_code == 200
    assert response.json()["results"][0]["employee_id"] == "zt1"
    assert manager.requested_sections == [
        {"results_view_data"},
        {"results"},
    ]
    assert manager.backfilled[0][0] == "run_legacy"
    assert manager.backfilled[0][1]["rows"][0]["employee_id"] == "zt1"


def test_fbu_attendance_step_reads_compact_view_without_daily_rows(monkeypatch):
    attendance_view = {
        "employees": [{
            "employee_id": "zt1",
            "name": "Test User",
            "total_base_hours": 80,
        }],
        "summary": {"total_employees": 1},
    }

    class AttendanceViewManager:
        def __init__(self):
            self.requested_sections = []

        def get_run(self, run_id, sections=None):
            self.requested_sections.append(sections)
            return SimpleNamespace(
                run_id=run_id,
                created_at="2026-07-31T10:00:00",
                calc_month="2026-06",
                status="completed",
                attendance_view_data=attendance_view,
                hourly_rate_policy_data={"rows": []},
            )

    manager = AttendanceViewManager()
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    client = TestClient(app_module.app)

    response = client.get(
        "/api/fbu-performance/runs/run_123",
        params={"include": "core,attendance_view_data"},
    )

    assert response.status_code == 200
    assert response.json()["attendance_data"] == attendance_view
    assert "attendance_view_data" not in response.json()
    assert response.json()["loaded_sections"] == ["attendance_view_data"]
    assert manager.requested_sections == [{"attendance_view_data"}]


def test_fbu_core_detail_excludes_legacy_roster_payload(monkeypatch):
    legacy_roster = {
        "employees": [{"employee_id": f"zt{index:06d}"} for index in range(500)],
        "summary": {"total_employees": 500},
    }

    class LegacyCoreManager:
        def get_run(self, run_id, sections=None):
            return SimpleNamespace(
                run_id=run_id,
                created_at="2026-07-31T10:00:00",
                calc_month="2026-06",
                status="step1",
                current_step=1,
                attendance_file="attendance.xlsx",
                roster_file="roster.xlsx",
                roster_source="base",
                roster_data=legacy_roster,
                attendance_view_data={},
                hourly_rate_policy_data={},
                total_employees=0,
                total_bonus=0,
                match_rate=0,
                error="",
            )

    monkeypatch.setattr(app_module, "fbu_run_manager", LegacyCoreManager())
    response = TestClient(app_module.app).get(
        "/api/fbu-performance/runs/run_legacy",
        params={"include": "core"},
    )

    assert response.status_code == 200
    assert response.json()["roster_file"] == "roster.xlsx"
    assert "roster_data" not in response.json()


def test_fbu_hourly_rate_policy_detail_returns_visible_rows_and_hidden_employee_ids(
    monkeypatch,
):
    policy_data = {
        "rows": [
            {
                "row_id": "zt-visible|2026-06-01",
                "employee_id": "zt-visible",
                "shift_pattern": "全夜班",
                "visible": True,
            },
            {
                "row_id": "zt-hidden|2026-06-01",
                "employee_id": "zt-hidden",
                "shift_pattern": "全白班",
                "visible": False,
            },
            {
                "row_id": "zt-hidden|2026-06-15",
                "employee_id": "zt-hidden",
                "shift_pattern": "全白班",
                "visible": False,
            },
        ],
        "summary": {
            "total_periods": 3,
            "visible_count": 1,
            "all_night_count": 1,
            "mixed_count": 0,
            "manual_count": 0,
        },
    }

    class PolicyViewManager:
        def get_run(self, run_id, sections=None):
            return SimpleNamespace(
                run_id=run_id,
                created_at="2026-07-31T10:00:00",
                calc_month="2026-06",
                status="step1",
                current_step=1,
                attendance_view_data={},
                hourly_rate_policy_data=policy_data,
            )

    monkeypatch.setattr(app_module, "fbu_run_manager", PolicyViewManager())
    response = TestClient(app_module.app).get(
        "/api/fbu-performance/runs/run_123",
        params={"include": "core,hourly_rate_policy_data"},
    )

    assert response.status_code == 200
    policy = response.json()["hourly_rate_policy_data"]
    assert [row["employee_id"] for row in policy["rows"]] == ["zt-visible"]
    assert policy["hidden_employee_ids"] == ["zt-hidden"]
    assert policy["summary"]["total_periods"] == 3


def test_fbu_attendance_step_backfills_compact_view_for_legacy_runs(monkeypatch):
    attendance = {
        "employees": [{
            "employee_id": "zt1",
            "total_base_hours": 80,
            "attendance_daily_rows": [
                {"date": "2026-06-01", "base_hours": 8},
            ],
        }],
        "summary": {"total_employees": 1},
    }

    class LegacyAttendanceManager:
        def __init__(self):
            self.requested_sections = []
            self.backfilled = []

        def get_run(self, run_id, sections=None):
            self.requested_sections.append(sections)
            if sections == {"attendance_view_data"}:
                return SimpleNamespace(
                    run_id=run_id,
                    created_at="2026-07-31T10:00:00",
                    calc_month="2026-06",
                    status="completed",
                    attendance_view_data={"employees": []},
                )
            return SimpleNamespace(
                run_id=run_id,
                created_at="2026-07-31T10:00:00",
                calc_month="2026-06",
                status="completed",
                attendance_data=attendance,
            )

        def backfill_attendance_view_data(self, run_id, attendance_view_data):
            self.backfilled.append((run_id, attendance_view_data))

    manager = LegacyAttendanceManager()
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    client = TestClient(app_module.app)

    response = client.get(
        "/api/fbu-performance/runs/run_legacy",
        params={"include": "core,attendance_view_data"},
    )

    assert response.status_code == 200
    employee = response.json()["attendance_data"]["employees"][0]
    assert "attendance_daily_rows" not in employee
    assert manager.requested_sections == [
        {"attendance_view_data"},
        {"attendance_data"},
    ]
    assert manager.backfilled[0][0] == "run_legacy"
    assert "attendance_daily_rows" not in manager.backfilled[0][1]["employees"][0]


def test_fbu_upload_parsing_does_not_block_request_event_loop():
    async def blocking_parser():
        time.sleep(0.08)
        return {"success": True}

    async def scenario():
        task = asyncio.create_task(
            app_module._run_fbu_upload_operation(blocking_parser())
        )
        started = time.perf_counter()
        await asyncio.sleep(0.01)
        event_loop_delay = time.perf_counter() - started
        result = await task
        return event_loop_delay, result

    loop = asyncio.new_event_loop()
    try:
        event_loop_delay, result = loop.run_until_complete(scenario())
    finally:
        loop.close()

    assert event_loop_delay < 0.04
    assert result == {"success": True}


def test_fbu_run_detail_can_include_only_requested_sections(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    manager = FBURunManager(str(tmp_path))
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    run = manager.create_run("2026-06")
    manager.save_step_data(
        run.run_id,
        1,
        {"summary": {"total_employees": 1}, "employees": [{"employee_id": "zt1"}]},
    )
    manager.save_step_data(
        run.run_id,
        2,
        {"summary": {"total_employees": 1}, "employees": [{"employee_id": "zt1", "hourly_rate": 20}]},
    )

    client = TestClient(app_module.app)
    response = client.get(
        f"/api/fbu-performance/runs/{run.run_id}",
        params={"include": "attendance_data"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["run_id"] == run.run_id
    assert payload["attendance_data"]["employees"][0]["employee_id"] == "zt1"
    assert "salary_data" not in payload
    assert "results" not in payload
    assert payload["loaded_sections"] == ["attendance_data"]


def test_fbu_results_endpoint_supports_server_pagination_search_and_group(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    manager = FBURunManager(str(tmp_path))
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    run = manager.create_run("2026-06")
    manager.update_run(
        run.run_id,
        status="completed",
        current_step=5,
        total_employees=3,
        total_bonus=600,
        results=[
            {
                "employee_id": "zt1",
                "source_employee_id": "zt1",
                "name": "Ana",
                "department": "Warehouse",
                "area": "New Jersey",
                "job_type": "warehouse",
                "calculation_path": "标准绩效基数路径",
                "performance_base": 1000,
                "performance_bonus": 100,
            },
            {
                "employee_id": "zt2",
                "source_employee_id": "zt2",
                "name": "Alex",
                "department": "Warehouse",
                "area": "New Jersey",
                "job_type": "warehouse",
                "calculation_path": "96工时制自动基数路径",
                "performance_base": 2000,
                "performance_bonus": 200,
            },
            {
                "employee_id": "zt3",
                "source_employee_id": "zt3",
                "name": "Beth",
                "department": "HR",
                "area": "New Jersey",
                "job_type": "functional",
                "calculation_path": "标准绩效基数路径",
                "performance_base": 3000,
                "performance_bonus": 300,
            },
        ],
    )
    client = TestClient(app_module.app)

    response = client.get(
        f"/api/fbu-performance/runs/{run.run_id}/results",
        params={"page": 1, "page_size": 1, "q": "a", "group": "warehouse"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["pagination"] == {
        "page": 1,
        "page_size": 1,
        "total": 2,
        "pages": 2,
    }
    assert len(payload["results"]) == 1
    assert payload["results"][0]["employee_id"] == "zt1"
    assert payload["summary"]["total_employees"] == 3
    assert payload["summary"]["groups"]["warehouse"] == {
        "count": 2,
        "total_bonus": 300.0,
    }
    assert payload["summary"]["total_performance_base"] == 6000.0
    assert payload["summary"]["special_base_count"] == 1
    assert payload["summary"]["calculation_paths"] == [
        {
            "path": "标准绩效基数路径",
            "count": 2,
            "total_base": 4000.0,
            "total_bonus": 400.0,
        },
        {
            "path": "96工时制自动基数路径",
            "count": 1,
            "total_base": 2000.0,
            "total_bonus": 200.0,
        },
    ]


def _adjustment_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "调薪拆分"
    sheet.append(["header"] * 32)
    row = [""] * 32
    row[3] = "zt001"
    row[4] = "Ana Roster"
    row[9] = "4.26-4.30"
    row[28] = 500
    row[31] = "调薪后"
    sheet.append(row)
    return _workbook_bytes(workbook)


def _oehr_adjustment_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "调薪"
    sheet.append([
        "审批主题",
        "姓名",
        "工号",
        "审批状态",
        "二级部门",
        "三级部门",
        "四级部门",
        "五级部门",
        "六级部门",
        "七级部门",
        "八级部门",
        "调薪类型",
        "职级",
        "调薪原因",
        "调薪生效日期",
        "调薪后薪酬制度",
        "调薪后成本归属",
        "调薪后币种",
        "基本工资标准",
        "绩效奖金计算方式",
        "月度绩效奖金基数",
        "月度绩效奖金比例(%)",
        "是否考勤豁免人员",
        "时薪标准",
        "备注",
    ])
    sheet.append([
        "张海冰的调薪申请",
        "张海冰",
        "zt0021990",
        "已完成",
        "FBU仓储事业部",
        "美洲区",
        "新泽西区",
        "新泽西21号仓（SN）",
        "理货组",
        "",
        "",
        "非窗口期调薪",
        "",
        "转正调薪",
        "2026/04/26",
        "时薪制",
        "理货组",
        "美元(USD)",
        18,
        "固定比例核算",
        0,
        5,
        "否",
        18,
        "P1-2转正，增加绩效占比",
    ])
    return _workbook_bytes(workbook)


def _performance_report_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "4月绩效报表"
    sheet.append(["header"] * 19)
    row = [""] * 19
    row[3] = "zt001"
    row[16] = 95
    row[17] = "符合预期"
    row[18] = 1
    sheet.append(row)
    return _workbook_bytes(workbook)


def _performance_supplement_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "离职线下绩效考核表"
    sheet.append(["工号", "姓名", "绩效得分", "绩效等级", "绩效系数", "备注"])
    sheet.append(["zt001", "Ana", 60, "待改进", 0.5, "不应覆盖OEHR"])
    sheet.append(["zt0019943", "洪梓腾", 88.75, "符合预期-", 0.93, "5.29离职，线下绩效考核表"])
    return _workbook_bytes(workbook)


def test_base_roster_is_reused_by_new_fbu_activity(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)

    roster_response = client.post(
        "/api/fbu-performance/roster",
        files={"file": ("roster.xlsx", _roster_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert roster_response.status_code == 200
    assert roster_response.json()["roster"]["total_employees"] == 1

    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]
    assert run_response.json()["roster_source"] == "base"
    assert run_response.json()["activity"]["run_id"] == run_id
    assert run_response.json()["activity"]["roster_data"]["summary"]["total_employees"] == 1

    listed = client.get("/api/fbu-performance/runs")
    assert listed.status_code == 200
    listed_run = next(row for row in listed.json()["runs"] if row["run_id"] == run_id)
    assert "diagnostics" not in listed_run
    assert "sections" in listed_run

    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}")
    assert run_detail.status_code == 200
    roster_data = run_detail.json()["roster_data"]
    assert roster_data["summary"]["total_employees"] == 1
    assert roster_data["employees"][0]["employee_id"] == "E001"
    assert roster_data["employees"][0]["name"] == "Ana Roster"
    assert roster_data["employees"][0]["department"] == "HRAS人力综合条线-FBU HRBP Dept."

    attendance_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={"file": ("attendance.xlsx", _attendance_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert attendance_response.status_code == 200
    attendance_payload = attendance_response.json()
    employee = attendance_payload["preview"]["employees"][0]
    assert employee["name"] == "Ana Roster"
    assert employee["department"] == "HRAS人力综合条线-FBU HRBP Dept."
    assert employee["area"] == "US-West"
    assert employee["job_type"] == "functional"
    assert employee["roster_matched"] is True
    assert attendance_payload["result_file"]["type"] == "attendance"
    assert attendance_payload["result_file"]["filename"].startswith("考勤汇总_2026-04_")
    result_path = app_module.EXPORT_DIR / attendance_payload["result_file"]["filename"]
    assert not result_path.exists()
    download_response = client.get(attendance_payload["result_file"]["download_url"])
    assert download_response.status_code == 200
    assert download_response.content[:2] == b"PK"
    assert result_path.exists()


def test_fbu_activities_record_creator_region_and_allow_duplicate_month_region(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    monkeypatch.setattr(
        app_module,
        "_fbu_current_user",
        lambda request: {
            "id": "ou_zhangsan",
            "name": "张三",
            "avatarUrl": "https://example.test/zhangsan.png",
        },
    )

    client = TestClient(app_module.app)
    payload = {"calc_month": "2026-05", "region_code": "us_nj"}
    first = client.post("/api/fbu-performance/runs", json=payload)
    second = client.post("/api/fbu-performance/runs", json=payload)

    assert first.status_code == 200
    assert second.status_code == 200
    assert first.json()["run_id"] != second.json()["run_id"]
    assert first.json()["activity_name"] == "绩效奖金核算-202605-FBU新泽西区"
    assert first.json()["region_name"] == "FBU新泽西区"
    assert first.json()["created_by_user_id"] == "ou_zhangsan"
    assert first.json()["created_by_name"] == "张三"

    listing = client.get("/api/fbu-performance/runs").json()
    assert listing["current_user"]["id"] == "ou_zhangsan"
    assert any(region["name"] == "FBU加州区" for region in listing["regions"])
    assert len(listing["runs"]) == 2


def test_fbu_current_user_reads_nested_production_session_profile(monkeypatch):
    auth_module = ModuleType("bonus_platform.auth")
    auth_module.current_user_from_request = lambda _request: {
        "user": {
            "id": "ou_lisi",
            "name": "李四",
            "email": "lisi@example.com",
            "avatarUrl": "https://example.com/lisi.png",
        },
        "modules": [],
    }
    monkeypatch.setitem(sys.modules, "bonus_platform.auth", auth_module)

    assert app_module._fbu_current_user(SimpleNamespace()) == {
        "id": "ou_lisi",
        "name": "李四",
        "avatarUrl": "https://example.com/lisi.png",
    }


def test_fbu_activity_rejects_non_americas_region(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    response = TestClient(app_module.app).post(
        "/api/fbu-performance/runs",
        json={"calc_month": "2026-05", "region_code": "uk"},
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "请选择有效的FBU美洲划分区域"


def test_fbu_bulk_delete_removes_selected_runs(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    first_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    second_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-05"}).json()["run_id"]

    response = client.post("/api/fbu-performance/runs/bulk-delete", json={"run_ids": [first_id, second_id]})

    assert response.status_code == 200
    assert response.json()["deleted_count"] == 2
    assert client.get(f"/api/fbu-performance/runs/{first_id}").status_code == 404
    assert client.get(f"/api/fbu-performance/runs/{second_id}").status_code == 404


def test_fbu_attendance_upload_accepts_previous_month_context_without_counting_it(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    attendance_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            "previous_attendance": (
                "attendance-202603.xlsx",
                _attendance_bytes_for_rows([
                    ("2026-03-01", 8),
                    ("2026-03-29", 8),
                    ("2026-03-30", 8),
                    ("2026-03-31", 8),
                    ("2026-03-28", 8),
                ]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert attendance_response.status_code == 200
    payload = attendance_response.json()
    employee = payload["preview"]["employees"][0]
    assert [row["date"] for row in employee["attendance_daily_rows"]] == [
        "2026-03-29",
        "2026-03-30",
        "2026-03-31",
        "2026-04-01",
    ]
    assert employee["total_base_hours"] == 8
    assert payload["preview"]["summary"]["total_base_hours"] == 8
    context = payload["preview"]["summary"]["attendance_context"]
    assert context["required"] is True
    assert context["status"] == "complete"
    assert context["required_start"] == "2026-03-29"
    assert context["required_end"] == "2026-03-31"
    assert context["covered_dates"] == ["2026-03-29", "2026-03-30", "2026-03-31"]
    assert context["missing_dates"] == []
    assert "已识别96工时制跨月首段" in context["message"]

    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    assert run_detail["attendance_file"] == "attendance-202604.xlsx"
    assert run_detail["previous_attendance_file"] == "attendance-202603.xlsx"


def test_fbu_attendance_upload_rejects_file_without_calculation_month_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202605.xlsx",
                _attendance_bytes_for_rows([("2026-05-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "考勤日报未包含 2026-04 的数据，请确认活动月份或重新上传文件"
    run = app_module.fbu_run_manager.get_run(run_id)
    assert run.attendance_data == {}
    assert run.current_step == 0


def test_failed_attendance_reupload_preserves_previous_file_and_preview(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    valid_bytes = _attendance_bytes_for_rows([("2026-04-01", 8)])
    first_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={"file": ("attendance-202604.xlsx", valid_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert first_response.status_code == 200
    original_preview = app_module.fbu_run_manager.get_run(run_id).attendance_data

    failed_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202605.xlsx",
                _attendance_bytes_for_rows([("2026-05-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert failed_response.status_code == 400
    run = app_module.fbu_run_manager.get_run(run_id)
    assert run.attendance_file == "attendance-202604.xlsx"
    assert run.attendance_data == original_preview
    assert run.status == "step1"
    assert (tmp_path / run_id / "attendance.xlsx").read_bytes() == valid_bytes


def test_fbu_attendance_upload_can_add_previous_context_after_current_file(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    first_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert first_response.status_code == 200

    context_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "previous_attendance": (
                "attendance-202603.xlsx",
                _attendance_bytes_for_rows([
                    ("2026-03-01", 8),
                    ("2026-03-29", 8),
                    ("2026-03-30", 8),
                    ("2026-03-31", 8),
                    ("2026-03-28", 8),
                ]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert context_response.status_code == 200
    payload = context_response.json()
    assert payload["preview"]["summary"]["attendance_context"]["status"] == "complete"
    assert payload["preview"]["employees"][0]["total_base_hours"] == 8
    assert [row["date"] for row in payload["preview"]["employees"][0]["attendance_daily_rows"]] == [
        "2026-03-29",
        "2026-03-30",
        "2026-03-31",
        "2026-04-01",
    ]
    assert not (tmp_path / run_id / "attendance_with_context.xlsx").exists()
    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    assert run_detail["attendance_file"] == "attendance-202604.xlsx"
    assert run_detail["previous_attendance_file"] == "attendance-202603.xlsx"


def test_fbu_attendance_upload_reports_missing_previous_context_dates(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    context_workbook_reads = 0
    original_load_workbook = app_module.load_workbook

    def count_context_workbook_reads(*args, **kwargs):
        nonlocal context_workbook_reads
        context_workbook_reads += 1
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(app_module, "load_workbook", count_context_workbook_reads)

    client = TestClient(app_module.app)
    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    attendance_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert attendance_response.status_code == 200
    context = attendance_response.json()["preview"]["summary"]["attendance_context"]
    assert context["required"] is True
    assert context["status"] == "missing"
    assert context_workbook_reads == 0
    assert context["required_start"] == "2026-03-29"
    assert context["required_end"] == "2026-03-31"
    assert context["covered_dates"] == []
    assert context["missing_dates"] == ["2026-03-29", "2026-03-30", "2026-03-31"]
    assert "缺少上一月 2026-03-29 至 2026-03-31 考勤" in context["message"]


def test_fbu_attendance_upload_reuses_parsed_dates_for_context_summary(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]

    original_load_workbook = app_module.load_workbook
    reopened_workbooks = []

    def track_second_workbook_scan(*args, **kwargs):
        reopened_workbooks.append(args[0])
        return original_load_workbook(*args, **kwargs)

    monkeypatch.setattr(app_module, "load_workbook", track_second_workbook_scan)
    response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200, response.text
    assert response.json()["preview"]["summary"]["attendance_context"]["status"] == "missing"
    assert reopened_workbooks == []


def test_fbu_current_attendance_upload_does_not_load_previous_attendance_section(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    manager = FBURunManager(str(tmp_path))
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    requested_sections = []
    original_get_run = manager.get_run

    def get_run(target_run_id, sections=None):
        requested_sections.append(None if sections is None else set(sections))
        return original_get_run(target_run_id, sections=sections)

    monkeypatch.setattr(manager, "get_run", get_run)
    response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200, response.text
    assert all(
        sections is not None and "attendance_data" not in sections
        for sections in requested_sections
    )


def test_fbu_attendance_upload_logs_granular_stage_timings(caplog, monkeypatch, tmp_path):
    caplog.set_level("INFO", logger="bonus_platform.fbu")
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200, response.text
    performance_records = [
        record.message
        for record in caplog.records
        if record.message.startswith("FBU attendance import ")
    ]
    assert len(performance_records) == 1
    message = performance_records[0]
    for field in (
        "mode=current",
        "bytes=",
        "run_load_ms=",
        "stage_ms=",
        "roster_ms=",
        "workbook_ms=",
        "context_ms=",
        "state_ms=",
        "file_ms=",
        "result_ms=",
        "total_ms=",
    ):
        assert field in message


def test_fbu_attendance_upload_persists_run_state_once(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    manager = FBURunManager(str(tmp_path))
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    save_calls = []
    original_save_runs = manager._save_runs

    def save_runs(changed_run_id=None, changed_fields=None):
        save_calls.append((changed_run_id, set(changed_fields or [])))
        return original_save_runs(changed_run_id, changed_fields)

    monkeypatch.setattr(manager, "_save_runs", save_runs)
    response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 200, response.text
    assert len(save_calls) == 1
    assert {
        "attendance_file",
        "attendance_data",
        "attendance_view_data",
        "hourly_rate_policy_data",
        "current_step",
        "status",
    }.issubset(save_calls[0][1])


def _enable_fake_fbu_direct_upload(monkeypatch):
    monkeypatch.setattr(
        app_module,
        "fbu_persistent_storage_enabled",
        lambda: True,
        raising=False,
    )
    monkeypatch.setattr(
        app_module,
        "create_fbu_signed_upload",
        lambda run_id, relative_path: {
            "signedUrl": f"https://example.supabase.co/storage/v1/object/upload/sign/{relative_path}",
            "objectPath": f"fbu-performance-runs/production/{run_id}/{relative_path}",
            "relativePath": relative_path,
        },
        raising=False,
    )


def test_fbu_attendance_direct_upload_plan_accepts_file_above_vercel_limit(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-plan",
        json={
            "files": [{
                "kind": "attendance",
                "fileName": "考勤日报表-20260520.xlsx",
                "fileSize": 5_799_733,
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }],
        },
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["runId"] == run_id
    assert payload["uploads"][0]["size"] == 5_799_733
    assert payload["uploads"][0]["kind"] == "attendance"
    assert payload["uploads"][0]["signedUrl"].startswith("https://example.supabase.co/")


def test_fbu_generic_direct_upload_job_processes_salary_material_and_persists_status(
    caplog,
    monkeypatch,
    tmp_path,
):
    caplog.set_level("INFO", logger="bonus_platform.fbu")
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)
    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_salary_preview",
        lambda self, path: {
            "employees": [{"employee_id": "E001", "hourly_rate": 18, "ratio": 0.05}],
            "summary": {"total_employees": 1},
        },
    )

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-06"}).json()["run_id"]
    content = b"salary-file"
    plan_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/uploads/plan",
        json={
            "files": [{
                "kind": "previousSalary",
                "fileName": "may-salary.xlsx",
                "fileSize": len(content),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }],
        },
    )

    assert plan_response.status_code == 200, plan_response.text
    plan = plan_response.json()
    assert plan["job"]["status"] == "uploading"
    assert plan["uploads"][0]["kind"] == "previousSalary"
    upload_path = tmp_path / run_id / plan["uploads"][0]["relativePath"]
    upload_path.parent.mkdir(parents=True, exist_ok=True)
    upload_path.write_bytes(content)

    start_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/uploads/{plan['job']['jobId']}/start",
        json={"clientUploadMs": 1234},
    )
    assert start_response.status_code == 202, start_response.text
    assert start_response.json()["job"]["status"] == "completed"
    assert any(
        "FBU upload request" in record.message
        and "client_upload_ms=1234.0" in record.message
        and f"bytes={len(content)}" in record.message
        for record in caplog.records
    )
    assert any(
        "FBU upload processing" in record.message
        and "materialize_ms=" in record.message
        and "parse_ms=" in record.message
        for record in caplog.records
    )

    status_response = client.get(
        f"/api/fbu-performance/runs/{run_id}/uploads/{plan['job']['jobId']}",
    )
    assert status_response.status_code == 200
    status = status_response.json()["job"]
    assert status["status"] == "completed"
    assert status["stage"] == "completed"
    assert status["progress"] == 100
    assert status["result"]["step"] == "salary"

    run = app_module.fbu_run_manager.get_run(
        run_id,
        sections={"previous_salary_data"},
    )
    assert run.previous_salary_file == "may-salary.xlsx"
    assert run.previous_salary_data["employees"][0]["employee_id"] == "E001"


def test_fbu_attendance_upload_job_promotes_direct_object_without_reupload(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    manager = FBURunManager(str(tmp_path))
    monkeypatch.setattr(app_module, "fbu_run_manager", manager)
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    content = _attendance_bytes_for_rows([("2026-04-01", 8)])
    plan = client.post(
        f"/api/fbu-performance/runs/{run_id}/uploads/plan",
        json={
            "files": [{
                "kind": "attendance",
                "fileName": "attendance-202604.xlsx",
                "fileSize": len(content),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }],
        },
    ).json()
    upload = plan["uploads"][0]
    upload_path = tmp_path / run_id / upload["relativePath"]
    upload_path.parent.mkdir(parents=True, exist_ok=True)
    upload_path.write_bytes(content)

    promoted = []
    persisted = []
    original_persist_files = manager.persist_files

    def promote_persisted_file(target_run_id, source_relative_path, destination_relative_path):
        promoted.append((target_run_id, source_relative_path, destination_relative_path))
        return True

    def persist_files(target_run_id, relative_paths):
        persisted.extend(relative_paths)
        return original_persist_files(target_run_id, relative_paths)

    monkeypatch.setattr(manager, "promote_persisted_file", promote_persisted_file, raising=False)
    monkeypatch.setattr(manager, "persist_files", persist_files)

    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/uploads/{plan['job']['jobId']}/start",
        json={"clientUploadMs": 100},
    )

    assert response.status_code == 202, response.text
    assert response.json()["job"]["status"] == "completed"
    assert promoted == [(run_id, upload["relativePath"], "attendance.xlsx")]
    assert "attendance.xlsx" not in persisted


def test_fbu_upload_job_status_marks_stalled_processing_as_recoverable(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)
    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-06"}).json()["run_id"]
    plan = client.post(
        f"/api/fbu-performance/runs/{run_id}/uploads/plan",
        json={
            "files": [{
                "kind": "performance",
                "fileName": "performance.xlsx",
                "fileSize": 10,
            }],
        },
    ).json()
    job_path = tmp_path / run_id / "jobs" / f"{plan['job']['jobId']}.json"
    job = json.loads(job_path.read_text(encoding="utf-8"))
    job.update({
        "status": "processing",
        "stage": "parsing",
        "updatedAt": "2026-01-01T00:00:00",
    })
    job_path.write_text(json.dumps(job), encoding="utf-8")

    response = client.get(
        f"/api/fbu-performance/runs/{run_id}/uploads/{plan['job']['jobId']}",
    )

    assert response.status_code == 200
    assert response.json()["job"]["recoverable"] is True
    assert response.json()["job"]["canRetry"] is True


def test_fbu_calculation_runs_to_completion_inside_request_lifecycle(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-06"}).json()["run_id"]
    monkeypatch.setattr(
        app_module,
        "calculate_fbu_performance",
        lambda target_run_id: {
            "success": True,
            "run_id": target_run_id,
            "total_employees": 303,
            "total_bonus": 109290.74,
        },
    )

    start = client.post(
        f"/api/fbu-performance/runs/{run_id}/calculation-jobs",
    )

    assert start.status_code == 202, start.text
    assert start.json()["job"]["status"] == "completed"
    job_id = start.json()["job"]["jobId"]
    status = client.get(
        f"/api/fbu-performance/runs/{run_id}/calculation-jobs/{job_id}",
    )
    assert status.status_code == 200
    job = status.json()["job"]
    assert job["status"] == "completed"
    assert job["result"]["total_employees"] == 303
    assert job["result"]["total_bonus"] == 109290.74


def test_fbu_attendance_direct_upload_completion_materializes_and_parses_file(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    content = _attendance_bytes()
    plan_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-plan",
        json={
            "files": [{
                "kind": "attendance",
                "fileName": "attendance-202604.xlsx",
                "fileSize": len(content),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }],
        },
    )
    assert plan_response.status_code == 200, plan_response.text
    plan = plan_response.json()
    upload = plan["uploads"][0]
    uploaded_path = tmp_path / run_id / upload["relativePath"]
    uploaded_path.parent.mkdir(parents=True, exist_ok=True)
    uploaded_path.write_bytes(content)

    complete_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-complete",
        json={"planId": plan["planId"]},
    )

    assert complete_response.status_code == 200, complete_response.text
    payload = complete_response.json()
    assert payload["success"] is True
    assert payload["preview"]["employees"][0]["employee_id"] == "E001"
    assert (tmp_path / run_id / "attendance.xlsx").read_bytes() == content

    repeated_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-complete",
        json={"planId": plan["planId"]},
    )
    assert repeated_response.status_code == 200, repeated_response.text
    assert repeated_response.json()["preview"] == payload["preview"]


def test_fbu_previous_attendance_can_be_added_through_direct_upload(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    current_response = client.post(
        "/api/fbu-performance/import-attendance",
        data={"calc_month": "2026-04", "run_id": run_id},
        files={
            "file": (
                "attendance-202604.xlsx",
                _attendance_bytes_for_rows([("2026-04-01", 8)]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert current_response.status_code == 200, current_response.text

    current_path = tmp_path / run_id / "attendance.xlsx"
    current_content = current_path.read_bytes()
    current_path.unlink()
    original_materialize = app_module.fbu_run_manager.materialize_file

    def materialize_from_persistent(target_run_id, relative_path):
        if relative_path == "attendance.xlsx":
            restored = tmp_path / target_run_id / relative_path
            restored.parent.mkdir(parents=True, exist_ok=True)
            restored.write_bytes(current_content)
            return restored
        return original_materialize(target_run_id, relative_path)

    monkeypatch.setattr(
        app_module.fbu_run_manager,
        "materialize_file",
        materialize_from_persistent,
    )

    previous_content = _attendance_bytes_for_rows([
        ("2026-03-29", 8),
        ("2026-03-30", 8),
        ("2026-03-31", 8),
    ])
    plan_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-plan",
        json={
            "files": [{
                "kind": "previous_attendance",
                "fileName": "attendance-202603.xlsx",
                "fileSize": len(previous_content),
                "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }],
        },
    )
    assert plan_response.status_code == 200, plan_response.text
    plan = plan_response.json()
    upload = plan["uploads"][0]
    uploaded_path = tmp_path / run_id / upload["relativePath"]
    uploaded_path.parent.mkdir(parents=True, exist_ok=True)
    uploaded_path.write_bytes(previous_content)

    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-complete",
        json={"planId": plan["planId"]},
    )

    assert response.status_code == 200, response.text
    context = response.json()["preview"]["summary"]["attendance_context"]
    assert context["status"] == "complete"
    assert context["covered_dates"] == ["2026-03-29", "2026-03-30", "2026-03-31"]
    assert (tmp_path / run_id / "previous_attendance.xlsx").read_bytes() == previous_content


def test_fbu_attendance_direct_upload_plan_rejects_duplicate_kinds(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    _enable_fake_fbu_direct_upload(monkeypatch)

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"}).json()["run_id"]
    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/attendance-direct-upload-plan",
        json={
            "files": [
                {"kind": "attendance", "fileName": "a.xlsx", "fileSize": 10},
                {"kind": "attendance", "fileName": "b.xlsx", "fileSize": 10},
            ],
        },
    )

    assert response.status_code == 400
    assert response.json()["detail"] == "同类考勤文件不能重复上传。"


def test_fbu_performance_supplement_upload_merges_missing_employee_without_overwrite(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)

    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    performance_response = client.post(
        "/api/fbu-performance/import-performance",
        data={"run_id": run_id},
        files={"file": ("performance.xlsx", _performance_report_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert performance_response.status_code == 200
    assert performance_response.json()["preview"]["summary"]["total_employees"] == 1

    supplement_response = client.post(
        "/api/fbu-performance/import-performance",
        data={"run_id": run_id},
        files={"file": ("resigned-performance.xlsx", _performance_supplement_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert supplement_response.status_code == 200
    payload = supplement_response.json()
    summary = payload["preview"]["summary"]
    assert summary["source_type"] == "merged_performance"
    assert summary["supplement_added"] == 1
    assert summary["supplement_skipped_existing"] == 1
    employees = {employee["employee_id"]: employee for employee in payload["preview"]["employees"]}
    assert employees["zt001"]["score"] == 95
    assert employees["zt001"]["coefficient"] == 1
    assert employees["zt0019943"]["name"] == "洪梓腾"
    assert employees["zt0019943"]["coefficient"] == 0.93
    assert employees["zt0019943"]["performance_source"] == "绩效补录"

    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    saved = {employee["employee_id"]: employee for employee in run_detail["performance_data"]["employees"]}
    assert set(saved) == {"zt001", "zt0019943"}


def test_fbu_performance_supplement_can_be_entered_from_page(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)

    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    performance_response = client.post(
        "/api/fbu-performance/import-performance",
        data={"run_id": run_id},
        files={"file": ("performance.xlsx", _performance_report_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert performance_response.status_code == 200

    supplement_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/performance-supplement",
        json={
            "employee_id": "zt0019943",
            "name": "洪梓腾",
            "score": 88.75,
            "level": "符合预期-",
            "coefficient": 0.93,
            "note": "5.29离职，线下绩效考核表",
        },
    )

    assert supplement_response.status_code == 200
    payload = supplement_response.json()
    summary = payload["preview"]["summary"]
    assert summary["source_type"] == "merged_performance"
    assert summary["supplement_added"] == 1
    employees = {employee["employee_id"]: employee for employee in payload["preview"]["employees"]}
    assert employees["zt001"]["score"] == 95
    assert employees["zt0019943"]["name"] == "洪梓腾"
    assert employees["zt0019943"]["score"] == 88.75
    assert employees["zt0019943"]["level"] == "符合预期-"
    assert employees["zt0019943"]["coefficient"] == 0.93
    assert employees["zt0019943"]["performance_source"] == "绩效补录"
    assert employees["zt0019943"]["note"] == "5.29离职，线下绩效考核表"

    duplicate_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/performance-supplement",
        json={
            "employee_id": "zt001",
            "name": "Ana",
            "score": 60,
            "level": "待改进",
            "coefficient": 0.5,
            "note": "不应覆盖OEHR",
        },
    )

    assert duplicate_response.status_code == 200
    duplicate_payload = duplicate_response.json()
    duplicate_summary = duplicate_payload["preview"]["summary"]
    assert duplicate_summary["supplement_added"] == 0
    assert duplicate_summary["supplement_skipped_existing"] == 1
    duplicate_employees = {employee["employee_id"]: employee for employee in duplicate_payload["preview"]["employees"]}
    assert duplicate_employees["zt001"]["score"] == 95
    assert duplicate_employees["zt001"]["coefficient"] == 1

    second_supplement_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/performance-supplement",
        json={
            "employee_id": "zt009999",
            "name": "临时补录员工",
            "coefficient": 0.88,
            "note": "继续补录第二人",
        },
    )

    assert second_supplement_response.status_code == 200
    second_payload = second_supplement_response.json()
    second_employees = {employee["employee_id"]: employee for employee in second_payload["preview"]["employees"]}
    assert second_employees["zt0019943"]["name"] == "洪梓腾"
    assert second_employees["zt009999"]["name"] == "临时补录员工"
    assert second_employees["zt009999"]["coefficient"] == 0.88
    assert second_employees["zt009999"]["performance_source"] == "绩效补录"

    compact_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/performance-supplement",
        json={
            "employee_id": "zt008888",
            "name": "紧凑响应员工",
            "coefficient": 0.91,
            "response_mode": "employee",
        },
    )
    assert compact_response.status_code == 200
    compact_payload = compact_response.json()
    assert "preview" not in compact_payload
    assert compact_payload["employee"]["employee_id"] == "zt008888"
    assert compact_payload["employee"]["coefficient"] == 0.91
    assert compact_payload["summary"]["supplement_added"] >= 1


def test_fbu_adjustment_upload_is_saved_as_optional_run_data(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)

    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    adjustment_response = client.post(
        "/api/fbu-performance/import-adjustments",
        data={"run_id": run_id},
        files={"file": ("adjustments.xlsx", _adjustment_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert adjustment_response.status_code == 200
    payload = adjustment_response.json()
    assert payload["preview"]["summary"]["total_employees"] == 1
    assert payload["preview"]["summary"]["active_performance_base"] == 500
    assert payload["result_file"]["type"] == "adjustments"
    assert payload["result_file"]["filename"].startswith("调薪拆分_2026-04_")
    download_response = client.get(payload["result_file"]["download_url"])
    assert download_response.status_code == 200
    assert download_response.content[:2] == b"PK"

    run_detail = client.get(f"/api/fbu-performance/runs/{run_id}").json()
    assert run_detail["adjustment_file"] == "adjustments.xlsx"
    assert run_detail["adjustment_data"]["employees"][0]["segments"][0]["reason"] == "调薪后"


def test_fbu_oehr_adjustment_upload_exports_event_preview(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    client = TestClient(app_module.app)

    run_response = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-04"})
    assert run_response.status_code == 200
    run_id = run_response.json()["run_id"]

    adjustment_response = client.post(
        "/api/fbu-performance/import-adjustments",
        data={"run_id": run_id},
        files={"file": ("oehr_adjustments.xlsx", _oehr_adjustment_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert adjustment_response.status_code == 200
    payload = adjustment_response.json()
    assert payload["preview"]["summary"]["total_events"] == 1
    assert payload["preview"]["summary"]["auto_split_ready"] == 1
    assert payload["result_file"]["type"] == "adjustments"
    download_response = client.get(payload["result_file"]["download_url"])
    assert download_response.status_code == 200
    assert download_response.content[:2] == b"PK"


def test_fbu_salary_history_upload_persists_three_sources_and_blocks_unmatched_changes(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    previews = iter([
        {"employees": [{"employee_id": "E001", "hourly_rate": 18, "ratio": 0.05}]},
        {"employees": [{"employee_id": "E001", "hourly_rate": 21, "ratio": 0.09}]},
    ])
    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_salary_preview",
        lambda self, path: next(previews),
    )
    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_adjustments_preview",
        lambda self, path: {"employees": [], "events": [], "summary": {"total_events": 0}},
    )

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-05"}).json()["run_id"]
    response = client.post(
        "/api/fbu-performance/import-salary-history",
        data={"run_id": run_id},
        files={
            "previous_salary": ("april.xlsx", b"previous", "application/octet-stream"),
            "current_salary": ("may.xlsx", b"current", "application/octet-stream"),
            "adjustments": ("adjustments.xlsx", b"adjustments", "application/octet-stream"),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["verification"]["summary"]["blocking_count"] == 1
    run = app_module.fbu_run_manager.get_run(run_id)
    assert run.previous_salary_file == "april.xlsx"
    assert run.current_salary_file == "may.xlsx"
    assert run.adjustment_file == "adjustments.xlsx"
    assert run.salary_data["employees"][0]["verification_status"] == "blocking"

    calculate_response = client.post(f"/api/fbu-performance/calculate/{run_id}")
    assert calculate_response.status_code == 409
    assert "薪资历史核验仍有 1 条" in calculate_response.json()["detail"]

    confirm_response = client.post(
        f"/api/fbu-performance/runs/{run_id}/salary-verification/confirm",
        json={"employee_id": "E001", "choice": "previous", "note": "薪酬组确认"},
    )
    assert confirm_response.status_code == 200
    confirmed = confirm_response.json()
    assert confirmed["verification"]["summary"]["blocking_count"] == 0
    assert confirmed["preview"]["employees"][0]["hourly_rate"] == 18
    assert confirmed["preview"]["employees"][0]["ratio"] == 0.05
    assert confirmed["preview"]["employees"][0]["resolution"] == "manual_use_previous"


def test_fbu_salary_history_compact_response_does_not_duplicate_employee_rows(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    previews = iter([
        {"employees": [{"employee_id": "E001", "hourly_rate": 18, "ratio": 0.05}]},
        {"employees": [{"employee_id": "E001", "hourly_rate": 21, "ratio": 0.09}]},
    ])
    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_salary_preview",
        lambda self, path: next(previews),
    )
    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_adjustments_preview",
        lambda self, path: {"employees": [], "events": [], "summary": {"total_events": 0}},
    )

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-05"}).json()["run_id"]
    response = client.post(
        "/api/fbu-performance/import-salary-history",
        data={"run_id": run_id, "response_mode": "compact"},
        files={
            "previous_salary": ("april.xlsx", b"previous", "application/octet-stream"),
            "current_salary": ("may.xlsx", b"current", "application/octet-stream"),
            "adjustments": ("adjustments.xlsx", b"adjustments", "application/octet-stream"),
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["preview"]["employees"][0]["verification_status"] == "blocking"
    assert "employees" not in payload["verification"]
    assert payload["verification"]["summary"]["blocking_count"] == 1


def test_fbu_salary_verification_supports_compact_employee_response(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-05"}).json()["run_id"]
    run = app_module.fbu_run_manager.get_run(run_id)
    employee = {
        "employee_id": "E001",
        "hourly_rate": 21,
        "ratio": 0.09,
        "previous_hourly_rate": 18,
        "previous_ratio": 0.05,
        "current_hourly_rate": 21,
        "current_ratio": 0.09,
        "verification_status": "blocking",
    }
    run.salary_data = {"employees": [dict(employee)], "summary": {"total_employees": 1}}
    run.salary_verification_data = {
        "employees": [dict(employee)],
        "issues": [{"employee_id": "E001"}],
        "summary": {"blocking_count": 1, "resolved_count": 0},
    }

    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/salary-verification/confirm",
        json={"employee_id": "E001", "choice": "previous", "response_mode": "employee"},
    )

    assert response.status_code == 200
    payload = response.json()
    assert "preview" not in payload
    assert "verification" not in payload
    assert payload["employee"]["employee_id"] == "E001"
    assert payload["employee"]["hourly_rate"] == 18
    assert payload["verification_summary"]["blocking_count"] == 0
    assert payload["salary_summary"]["total_employees"] == 1


def test_fbu_salary_verification_batches_multiple_rows_and_is_idempotent(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-05"}).json()["run_id"]
    run = app_module.fbu_run_manager.get_run(run_id)
    employees = [
        {
            "employee_id": employee_id,
            "hourly_rate": 21,
            "ratio": 0.09,
            "previous_hourly_rate": 18,
            "previous_ratio": 0.05,
            "current_hourly_rate": 21,
            "current_ratio": 0.09,
            "verification_status": "blocking",
        }
        for employee_id in ("E001", "E002", "E003")
    ]
    run.salary_data = {"employees": [dict(row) for row in employees], "summary": {"total_employees": 3}}
    run.salary_verification_data = {
        "employees": [dict(row) for row in employees],
        "issues": [{"employee_id": row["employee_id"]} for row in employees],
        "summary": {"blocking_count": 3, "resolved_count": 0},
    }
    save_calls: list[str] = []
    original_save = app_module.fbu_run_manager._save_runs

    def capture_save(changed_run_id=None, changed_fields=None):
        save_calls.append(changed_run_id)
        return original_save(changed_run_id, changed_fields)

    monkeypatch.setattr(app_module.fbu_run_manager, "_save_runs", capture_save)
    response = client.post(
        f"/api/fbu-performance/runs/{run_id}/salary-verification/confirm",
        json={
            "confirmations": [
                {"employee_id": "E001", "choice": "previous"},
                {"employee_id": "E002", "choice": "current"},
            ],
            "response_mode": "employees",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert {row["employee_id"] for row in payload["employees"]} == {"E001", "E002"}
    assert payload["verification_summary"]["blocking_count"] == 1
    assert save_calls == [run_id]

    duplicate = client.post(
        f"/api/fbu-performance/runs/{run_id}/salary-verification/confirm",
        json={"employee_id": "E001", "choice": "previous", "response_mode": "employee"},
    )
    assert duplicate.status_code == 200
    assert duplicate.json()["employee"]["resolution"] == "manual_use_previous"


def test_fbu_salary_history_materials_upload_individually_and_reconcile_when_complete(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))

    def parse_salary_preview(_self, path):
        if str(path).endswith("previous_salary.xlsx"):
            return {
                "employees": [{"employee_id": "E001", "hourly_rate": 18, "ratio": 0.05}],
                "summary": {"source": "previous"},
            }
        return {
            "employees": [{"employee_id": "E001", "hourly_rate": 21, "ratio": 0.09}],
            "summary": {"source": "current"},
        }

    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_salary_preview",
        parse_salary_preview,
    )
    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_adjustments_preview",
        lambda self, path: {"employees": [], "events": [], "summary": {"total_events": 0}},
    )

    client = TestClient(app_module.app)
    run_id = client.post("/api/fbu-performance/runs", json={"calc_month": "2026-05"}).json()["run_id"]

    previous_response = client.post(
        "/api/fbu-performance/import-salary-history-material",
        data={"run_id": run_id, "material_type": "previousSalary"},
        files={"file": ("april.xlsx", b"previous", "application/octet-stream")},
    )
    assert previous_response.status_code == 200
    previous_payload = previous_response.json()
    assert previous_payload["ready_for_reconciliation"] is False
    assert previous_payload["missing_materials"] == ["currentSalary", "salaryAdjustments"]
    assert previous_payload["preview"] == {}

    current_response = client.post(
        "/api/fbu-performance/import-salary-history-material",
        data={"run_id": run_id, "material_type": "currentSalary"},
        files={"file": ("may.xlsx", b"current", "application/octet-stream")},
    )
    assert current_response.status_code == 200
    current_payload = current_response.json()
    assert current_payload["ready_for_reconciliation"] is False
    assert current_payload["missing_materials"] == ["salaryAdjustments"]

    adjustments_response = client.post(
        "/api/fbu-performance/import-salary-history-material",
        data={"run_id": run_id, "material_type": "salaryAdjustments"},
        files={"file": ("adjustments.xlsx", b"adjustments", "application/octet-stream")},
    )
    assert adjustments_response.status_code == 200
    adjustments_payload = adjustments_response.json()
    assert adjustments_payload["ready_for_reconciliation"] is True
    assert adjustments_payload["missing_materials"] == []
    assert adjustments_payload["verification"]["summary"]["blocking_count"] == 1

    run = app_module.fbu_run_manager.get_run(run_id)
    assert run.previous_salary_file == "april.xlsx"
    assert run.current_salary_file == "may.xlsx"
    assert run.salary_file == "may.xlsx"
    assert run.adjustment_file == "adjustments.xlsx"
    assert run.salary_data["employees"][0]["verification_status"] == "blocking"


def test_fbu_salary_verification_rejects_choice_for_missing_snapshot(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    run = app_module.fbu_run_manager.create_run(calc_month="2026-06")
    blocked = {
        "employee_id": "E001",
        "hourly_rate": 23,
        "ratio": 0.13,
        "previous_hourly_rate": 23,
        "previous_ratio": 0.13,
        "current_hourly_rate": None,
        "current_ratio": None,
        "verification_status": "blocking",
        "resolution": "missing_current_snapshot",
    }
    verification = {
        "employees": [blocked],
        "issues": [{"employee_id": "E001", "reason": "当月薪资档案缺少该员工"}],
        "summary": {"total_employees": 1, "resolved_count": 0, "blocking_count": 1},
    }
    app_module.fbu_run_manager.update_run(
        run.run_id,
        salary_verification_data=verification,
        salary_data={"employees": [blocked], "summary": verification["summary"]},
    )

    client = TestClient(app_module.app)
    invalid = client.post(
        f"/api/fbu-performance/runs/{run.run_id}/salary-verification/confirm",
        json={"employee_id": "E001", "choice": "current"},
    )
    assert invalid.status_code == 400
    assert invalid.json()["detail"] == "当月薪资快照缺失，不能选择该值"

    valid = client.post(
        f"/api/fbu-performance/runs/{run.run_id}/salary-verification/confirm",
        json={"employee_id": "E001", "choice": "previous"},
    )
    assert valid.status_code == 200
    assert valid.json()["preview"]["employees"][0]["hourly_rate"] == 23


def test_fbu_salary_verification_can_ignore_missing_current_as_zero_and_reimport_wins(
    monkeypatch,
    tmp_path,
):
    monkeypatch.setattr(app_module, "FBU_PERFORMANCE_RUNS_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))
    monkeypatch.setattr(app_module, "fbu_roster_store", FBURosterStore(str(tmp_path)))
    run = app_module.fbu_run_manager.create_run(calc_month="2026-06")
    blocked = {
        "employee_id": "E001",
        "name": "测试员工",
        "hourly_rate": 23,
        "ratio": 0.13,
        "previous_hourly_rate": 23,
        "previous_ratio": 0.13,
        "current_hourly_rate": None,
        "current_ratio": None,
        "verification_status": "blocking",
        "resolution": "missing_current_snapshot",
        "effective_segments": [],
    }
    verification = {
        "employees": [blocked],
        "issues": [{"employee_id": "E001", "reason": "当月薪资档案缺少该员工"}],
        "summary": {"total_employees": 1, "resolved_count": 0, "blocking_count": 1},
    }
    app_module.fbu_run_manager.update_run(
        run.run_id,
        previous_salary_file="may.xlsx",
        current_salary_file="june.xlsx",
        salary_file="june.xlsx",
        adjustment_file="adjustments.xlsx",
        previous_salary_data={
            "employees": [{"employee_id": "E001", "hourly_rate": 23, "ratio": 0.13}],
            "summary": {"total_employees": 1},
        },
        current_salary_data={"employees": [], "summary": {"total_employees": 0}},
        adjustment_data={
            "events": [{
                "employee_id": "E001",
                "approval_status": "已完成",
                "effective_date": "2026-06-01",
                "hourly_rate": 25,
                "performance_ratio": 0.15,
            }],
            "summary": {"total_events": 1},
        },
        salary_verification_data=verification,
        salary_data={"employees": [blocked], "summary": verification["summary"]},
    )

    client = TestClient(app_module.app)
    ignored = client.post(
        f"/api/fbu-performance/runs/{run.run_id}/salary-verification/confirm",
        json={"employee_id": "E001", "choice": "ignore_current"},
    )

    assert ignored.status_code == 200
    ignored_row = ignored.json()["preview"]["employees"][0]
    assert ignored_row["hourly_rate"] == 0
    assert ignored_row["ratio"] == 0
    assert ignored_row["current_hourly_rate"] is None
    assert ignored_row["current_ratio"] is None
    assert ignored_row["verification_status"] == "resolved"
    assert ignored_row["resolution"] == "manual_ignore_current_missing"
    assert ignored.json()["verification"]["summary"]["blocking_count"] == 0

    monkeypatch.setattr(
        app_module.FBUPerformanceParser,
        "parse_salary_preview",
        lambda self, path: {
            "employees": [{
                "employee_id": "E001",
                "name": "测试员工",
                "hourly_rate": 25,
                "ratio": 0.15,
            }],
            "summary": {"total_employees": 1},
        },
    )
    supplemented = client.post(
        "/api/fbu-performance/import-salary-history-material",
        data={"run_id": run.run_id, "material_type": "currentSalary"},
        files={"file": ("june-supplement.xlsx", b"supplement", "application/octet-stream")},
    )

    assert supplemented.status_code == 200
    supplemented_row = supplemented.json()["preview"]["employees"][0]
    assert supplemented_row["hourly_rate"] == 25
    assert supplemented_row["ratio"] == 0.15
    assert supplemented_row["verification_status"] == "resolved"
    assert supplemented_row["resolution"] == "effective_before_month_use_current"


def test_fbu_adjustment_template_download_returns_workbook(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)

    client = TestClient(app_module.app)

    response = client.get("/api/fbu-performance/templates/adjustments/download")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.content[:2] == b"PK"

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["调薪拆分"]
    main_sheet = workbook["调薪拆分"]
    assert main_sheet.cell(3, 1).value == "zt0000001"
    assert main_sheet.cell(3, 2).value == "花名一"
    assert [main_sheet.cell(6, col).value for col in range(1, 7)] == [
        "工号",
        "姓名",
        "分段期间",
        "分段绩效基数",
        "核算标识",
        "备注",
    ]
    assert main_sheet.cell(7, 1).value is None
    validations = list(main_sheet.data_validations.dataValidation)
    assert len(validations) == 1
    assert validations[0].type == "list"
    assert validations[0].formula1 == '"调薪前,调薪后"'
    assert "E7:E1000" in str(validations[0].sqref)


def test_fbu_export_escapes_formula_like_values_and_selects_header(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_results(
        run.run_id,
        [
            EmployeeData(
                employee_id="zt001",
                name="=SUM(A1:A2)",
                department="+Formula Dept",
                area="@Area",
                hourly_rate=18,
                performance_ratio=0.05,
                performance_base=1000,
                performance_coefficient=1,
                performance_bonus=50,
            )
        ],
    )

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/export-excel?type=results")

    assert response.status_code == 200
    filename = response.json()["filename"]
    workbook = load_workbook(tmp_path / filename, data_only=False)
    assert workbook.sheetnames == ["汇总表", "1.仓库管理人员", "2.非仓人员", "3.区长"]
    summary = workbook["汇总表"]
    assert summary["A1"].value == "新泽西区绩效考核与奖金核算"
    assert not str(summary["A1"].value).startswith("=")
    assert summary.sheet_view.selection[0].activeCell == "A2"
    sheet = workbook["1.仓库管理人员"]
    assert sheet.sheet_view.selection[0].activeCell == "A3"
    values = [cell.value for row in sheet.iter_rows() for cell in row if isinstance(cell.value, str)]
    assert not any(value.startswith("=") for value in values)
    assert "'=SUM(A1:A2)" in values


def test_fbu_results_export_marks_district_manager_fixed_base_path(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_results(
        run.run_id,
        [
            EmployeeData(
                employee_id="zt15638",
                name="万其鑫",
                position="区域经理",
                job_type="district_manager",
                hourly_rate=40.384615,
                performance_ratio=0,
                fixed_performance_base=3000,
                performance_base=3000,
                uploaded_coefficient=1.35,
                coefficient_override_reason="页面维护",
                performance_coefficient=1.35,
                performance_bonus=4050,
            )
        ],
    )

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/export-excel?type=results")

    assert response.status_code == 200
    filename = response.json()["filename"]
    workbook = load_workbook(tmp_path / filename, data_only=False)
    assert workbook.sheetnames == ["汇总表", "1.仓库管理人员", "2.非仓人员", "3.区长"]
    sheet = workbook["3.区长"]
    headers = [cell.value for cell in sheet[3]]
    row = [cell.value for cell in sheet[4]]
    values_by_header = dict(zip(headers, row))
    assert values_by_header["岗位"] == "区域经理"
    assert values_by_header["绩效奖金基数"] == 3000
    assert sheet["N4"].value is None
    assert values_by_header["绩效基数计算过程"].startswith("区长固定基数路径：$3,000.00")
    assert "$3,000.00 × 1.35 = $4,050.00" in values_by_header["奖金计算过程"]
    assert "0.0%" not in values_by_header["奖金计算过程"]
    assert "4月" not in values_by_header["奖金计算过程"]
    assert sheet.row_dimensions[4].height <= 24
    assert sheet.freeze_panes == "A4"


def test_fbu_results_merge_shift_split_rows_for_final_view_and_export(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_results(
        run.run_id,
        [
            EmployeeData(
                employee_id="zt0020984-1",
                source_employee_id="zt0020984",
                name="黄东俊",
                area="新泽西区",
                department="新泽西21号仓（SN）",
                position="仓库组长",
                job_type="warehouse",
                hourly_rate=18,
                performance_ratio=0.05,
                base_hours=195.385,
                base_salary=3516.93,
                performance_base=3516.93,
                performance_score=107.3,
                performance_level="符合预期+",
                performance_coefficient=1.25,
                performance_bonus=219.81,
            ),
            EmployeeData(
                employee_id="zt0020984",
                source_employee_id="zt0020984",
                name="黄东俊",
                area="新泽西区",
                department="新泽西21号仓（SN）",
                position="仓库组长",
                job_type="warehouse",
                hourly_rate=19,
                performance_ratio=0.05,
                base_hours=32.14,
                base_salary=610.66,
                performance_base=610.66,
                performance_score=107.3,
                performance_level="符合预期+",
                performance_coefficient=1.25,
                performance_bonus=38.17,
            ),
        ],
    )

    client = TestClient(app_module.app)
    detail_response = client.get(f"/api/fbu-performance/runs/{run.run_id}")

    assert detail_response.status_code == 200
    final_results = detail_response.json()["results"]
    assert len(final_results) == 1
    assert final_results[0]["employee_id"] == "zt0020984"
    assert final_results[0]["raw_employee_ids"] == ["zt0020984-1", "zt0020984"]
    assert "hourly_rate" not in final_results[0]
    assert "attendance_daily_rows" not in final_results[0]
    assert final_results[0]["performance_score"] == 107.3
    assert final_results[0]["performance_level"] == "符合预期+"
    assert final_results[0]["position"] == "仓库组长"
    assert final_results[0]["performance_base"] == 4127.59
    assert final_results[0]["performance_bonus"] == 257.98
    assert len(final_results[0]["base_calculation_details"]) == 2
    assert [
        detail["display_label"]
        for detail in final_results[0]["base_calculation_details"]
    ] == ["白班", "夜班拆行"]

    export_response = client.get(f"/api/fbu-performance/runs/{run.run_id}/export-excel?type=results")

    assert export_response.status_code == 200
    filename = export_response.json()["filename"]
    workbook = load_workbook(tmp_path / filename, data_only=False)
    assert workbook.sheetnames == ["汇总表", "1.仓库管理人员", "2.非仓人员", "3.区长"]
    summary = workbook["汇总表"]
    assert summary["A2"].value == "绩效周期"
    assert summary["C3"].value == "仓库管理人员"
    assert summary["D3"].value == 257.98
    sheet = workbook["1.仓库管理人员"]
    assert sheet["A2"].value == "员工信息"
    assert sheet["L2"].value == "本月绩效考核结果(OEHR)"
    assert sheet["O2"].value == "本月应发绩效工资"
    headers = [cell.value for cell in sheet[3]]
    assert "时薪($)" not in headers
    assert "绩效得分" in headers
    assert "系统计算绩效基数" in headers
    assert "Period adjustment" in headers
    assert "4月绩效基数" in headers
    assert headers[-2:] == ["绩效基数计算过程", "奖金计算过程"]
    rows = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))
    values_by_header = dict(zip(headers, rows[0]))
    assert values_by_header["员工工号"] == "zt0020984"
    assert values_by_header["职位"] == "仓库组长"
    assert values_by_header["绩效得分"] == 107.3
    assert values_by_header["4月绩效基数"] == 4127.59
    assert sheet["P4"].value == "白夜班拆分行已合并"
    assert sheet.freeze_panes == "A4"
    assert "计算时薪：$18.0000" in values_by_header["绩效基数计算过程"]
    assert "计算时薪：$19.0000" in values_by_header["绩效基数计算过程"]
    assert "$3,516.93 × 5.0% × 1.25 = $219.81" in values_by_header["奖金计算过程"]
    assert "$610.66 × 5.0% × 1.25 = $38.17" in values_by_header["奖金计算过程"]
    assert "4月" not in values_by_header["奖金计算过程"]
    assert sheet.row_dimensions[4].height <= 72
    assert sheet["O4"].value == 257.98


def test_fbu_results_export_backfills_position_from_uploaded_roster_preview(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.update_run(
        run.run_id,
        attendance_data={
            "employees": [
                {
                    "employee_id": "zt0003518",
                    "name": "朱杏仪",
                    "position": "高级仓库专员",
                }
            ]
        },
    )
    app_module.fbu_run_manager.save_results(
        run.run_id,
        [
            EmployeeData(
                employee_id="zt0003518",
                name="朱杏仪",
                job_type="warehouse",
                hourly_rate=18,
                performance_ratio=0.06,
                performance_base=5456.75,
                performance_score=112.4,
                performance_level="超出预期",
                performance_coefficient=1.35,
                performance_bonus=442,
            )
        ],
    )

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/export-excel?type=results")

    assert response.status_code == 200
    filename = response.json()["filename"]
    workbook = load_workbook(tmp_path / filename, data_only=False)
    sheet = workbook["1.仓库管理人员"]
    headers = [cell.value for cell in sheet[3]]
    values_by_header = dict(zip(headers, list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]))
    assert values_by_header["职位"] == "高级仓库专员"


def test_fbu_results_export_marks_96_hour_base_in_red(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_results(
        run.run_id,
        [
            EmployeeData(
                employee_id="zt12979",
                name="赵婉妍",
                job_type="warehouse",
                performance_ratio=0.05,
                performance_base=2000,
                performance_score=100,
                performance_coefficient=1,
                performance_bonus=100,
                work_hour_rule="96工时制",
            )
        ],
    )

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/export-excel?type=results")

    assert response.status_code == 200
    filename = response.json()["filename"]
    workbook = load_workbook(tmp_path / filename, data_only=False)
    sheet = workbook["1.仓库管理人员"]
    headers = [cell.value for cell in sheet[3]]
    base_col = headers.index("4月绩效基数") + 1
    base_cell = sheet.cell(row=4, column=base_col)
    assert base_cell.value == 2000
    assert base_cell.font.color.rgb == "00C00000"
    assert base_cell.fill.fgColor.rgb == "00FCE4D6"


def test_fbu_calculate_uses_saved_rule_lists_even_when_current_step_is_behind(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-05")
    attendance_employee = {
        "employee_id": "zt12979",
        "source_employee_id": "zt12979",
        "name": "赵婉妍",
        "department": "FBU仓储事业部-美洲区-新泽西区",
        "area": "新泽西区",
        "position": "USNJ Deputy General Manager 新泽西区副总经理",
        "job_type": "warehouse",
        "day_shift": {
            "计薪出勤": 100,
            "OT1.5": 40,
            "OT2.0": 0,
            "病假": 0,
            "病假清算": 0,
            "年假": 0,
            "节假日": 0,
        },
        "night_shift": {
            "计薪出勤": 0,
            "OT1.5": 0,
            "OT2.0": 0,
            "病假": 0,
            "病假清算": 0,
            "年假": 0,
            "节假日": 0,
        },
        "has_night_shift": False,
        "attendance_daily_rows": [
            {"date": "2026-05-01", "shift_type": "白班", "base_hours": 100, "ot15_hours": 40, "holiday_hours": 0},
        ],
    }
    app_module.fbu_run_manager.save_step_data(run.run_id, 1, {"employees": [attendance_employee]})
    app_module.fbu_run_manager.save_step_data(run.run_id, 3, {
        "employees": [
            {
                "employee_id": "zt12979",
                "score": 100,
                "level": "符合预期",
                "coefficient": 1,
            }
        ]
    })
    app_module.fbu_run_manager.update_run(
        run.run_id,
        base_override_data={
            "employees": [
                {
                    "employee_id": "zt12979",
                    "source_employee_id": "zt12979",
                    "name": "赵婉妍",
                    "rule_type": "96工时制",
                    "fixed_performance_base": None,
                    "allocation_month": "2026-05",
                    "status": "启用",
                    "include_in_calculation": True,
                }
            ]
        },
    )
    app_module.fbu_run_manager.save_step_data(run.run_id, 2, {
        "employees": [
            {
                "employee_id": "zt12979",
                "hourly_rate": 20,
                "ratio": 0.05,
                "fixed_performance_base": 0,
            }
        ]
    })
    assert app_module.fbu_run_manager.get_run(run.run_id).current_step == 2

    save_calls: list[str] = []
    original_save = app_module.fbu_run_manager._save_runs

    def capture_save(changed_run_id=None, changed_fields=None):
        save_calls.append(changed_run_id)
        return original_save(changed_run_id, changed_fields)

    monkeypatch.setattr(app_module.fbu_run_manager, "_save_runs", capture_save)

    client = TestClient(app_module.app)
    response = client.post(f"/api/fbu-performance/calculate/{run.run_id}?response_mode=compact")

    assert response.status_code == 200
    activity = response.json()["activity"]
    assert "attendance_data" not in activity
    assert "salary_data" not in activity
    assert "performance_data" not in activity
    assert activity["results"]
    assert activity["diagnostics"] is not None
    result = app_module.fbu_run_manager.get_run(run.run_id).results[0]
    assert result["work_hour_rule"] == "96工时制"
    assert result["calculation_path"] == "96工时制自动基数路径"
    assert save_calls == [run.run_id]


def test_fbu_diagnostics_reports_matching_issues_and_exports(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_step_data(run.run_id, 1, {
        "employees": [
            {"employee_id": "zt001", "name": "Ana"},
            {"employee_id": "zt002", "name": "Ben"},
        ]
    })
    app_module.fbu_run_manager.save_step_data(run.run_id, 2, {
        "employees": [
            {"employee_id": "zt001", "name": "Ana", "hourly_rate": 18, "ratio": 0.05},
            {"employee_id": "zt003", "name": "Cara", "hourly_rate": 0, "ratio": 0},
        ]
    })
    app_module.fbu_run_manager.save_step_data(run.run_id, 3, {
        "employees": [
            {"employee_id": "zt001", "name": "Ana"},
            {"employee_id": "zt004", "name": "Dora"},
        ]
    })
    app_module.fbu_run_manager.save_step_data(run.run_id, 4, {
        "employees": [
            {"employee_id": "zt005", "name": "Eli", "segments": [{"reason": "调薪后", "performance_base": 0}]},
        ]
    })

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/diagnostics")

    assert response.status_code == 200
    diagnostics = response.json()
    assert diagnostics["summary"]["attendance_count"] == 2
    assert diagnostics["summary"]["matched_salary_count"] == 1
    assert diagnostics["summary"]["matched_performance_count"] == 1
    issue_types = {issue["type"] for issue in diagnostics["issues"]}
    assert "考勤有薪资无" in issue_types
    assert "薪资有考勤无" in issue_types
    assert "绩效有考勤无" in issue_types
    assert "拆分有考勤无" in issue_types
    assert "拆分有薪资无" in issue_types

    export_response = client.get(f"/api/fbu-performance/runs/{run.run_id}/export-excel?type=diagnostics")
    assert export_response.status_code == 200
    workbook = load_workbook(tmp_path / export_response.json()["filename"], data_only=True)
    assert workbook["数据诊断"].cell(3, 1).value == "严重程度"


def test_fbu_diagnostics_matches_shift_split_suffix_to_source_employee(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_step_data(run.run_id, 1, {
        "employees": [
            {"employee_id": "zt0020984-1", "name": "黄东俊", "has_night_shift": False},
            {"employee_id": "zt0020984", "name": "黄东俊", "has_night_shift": True},
            {"employee_id": "zt009-1", "name": "缺绩效员工", "has_night_shift": False},
        ]
    })
    app_module.fbu_run_manager.save_step_data(run.run_id, 2, {
        "employees": [
            {"employee_id": "zt0020984", "name": "黄东俊", "hourly_rate": 18, "ratio": 0.05},
            {"employee_id": "zt009", "name": "缺绩效员工", "hourly_rate": 18, "ratio": 0.05},
        ]
    })
    app_module.fbu_run_manager.save_step_data(run.run_id, 3, {
        "employees": [
            {"employee_id": "zt0020984", "name": "黄东俊"},
        ]
    })

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/diagnostics")

    assert response.status_code == 200
    diagnostics = response.json()
    assert diagnostics["summary"]["attendance_count"] == 3
    assert diagnostics["summary"]["matched_salary_count"] == 3
    assert diagnostics["summary"]["matched_performance_count"] == 2
    assert diagnostics["summary"]["can_calculate_count"] == 3
    assert diagnostics["summary"]["error_count"] == 0
    issue_pairs = {(issue["employee_id"], issue["type"]) for issue in diagnostics["issues"]}
    assert ("zt0020984-1", "考勤有薪资无") not in issue_pairs
    assert ("zt0020984-1", "考勤有绩效无") not in issue_pairs
    assert ("zt009-1", "考勤有绩效无") not in issue_pairs
    assert ("zt009", "考勤有绩效无") in issue_pairs


def test_fbu_diagnostics_does_not_flag_future_steps_before_upload(monkeypatch, tmp_path):
    monkeypatch.setattr(app_module, "EXPORT_DIR", tmp_path)
    monkeypatch.setattr(app_module, "fbu_run_manager", FBURunManager(str(tmp_path)))

    run = app_module.fbu_run_manager.create_run(calc_month="2026-04")
    app_module.fbu_run_manager.save_step_data(run.run_id, 1, {
        "employees": [
            {"employee_id": "zt001", "name": "Ana"},
            {"employee_id": "zt002", "name": "Ben"},
        ]
    })

    client = TestClient(app_module.app)
    response = client.get(f"/api/fbu-performance/runs/{run.run_id}/diagnostics")

    assert response.status_code == 200
    diagnostics = response.json()
    issue_types = {issue["type"] for issue in diagnostics["issues"]}
    assert "考勤有薪资无" not in issue_types
    assert "考勤有绩效无" not in issue_types
    assert diagnostics["summary"]["issue_count"] == 0
    assert diagnostics["summary"]["error_count"] == 0
