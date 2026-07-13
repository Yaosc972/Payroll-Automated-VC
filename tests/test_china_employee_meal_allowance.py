from io import BytesIO
from datetime import datetime
from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient

import bonus_platform.app as app_module
from bonus_platform.app import app
import bonus_platform.engine.china_employee_payroll.meal_allowance as meal_allowance_module
from bonus_platform.engine.china_employee_payroll import (
    calculate_meal_allowance,
    parse_attendance_workbooks,
    parse_wx_attendance_workbooks,
)


HEADERS = [
    "员工",
    "员工邮箱",
    "工号",
    "考勤卡号",
    "人员状态",
    "二级组织",
    "三级组织",
    "四级组织",
    "五级组织",
    "员工考勤地点",
    "考勤日期",
    "首打卡(含补签)",
    "末打卡(含补签)",
    "当前班次",
    "日期类型",
    "备注",
]


def _write_attendance(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤记录"
    ws.append([f"sys_{i}" for i in range(len(HEADERS))])
    ws.append(HEADERS)
    for row in rows:
        ws.append([row.get(header) for header in HEADERS])
    wb.save(path)


WX_HEADERS_TOP = [
    "姓名",
    "基本信息",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "考勤组信息",
    "",
    "",
    "",
    "第 1 次上班",
    "",
    "",
    "",
    "第 1 次下班",
    "",
    "",
    "",
]

WX_HEADERS = [
    "",
    "工号",
    "部门",
    "组织架构",
    "性别",
    "人员类型",
    "员工状态",
    "入职日期",
    "离职日期",
    "日期",
    "星期",
    "班次",
    "考勤类型",
    "上班 1 打卡时间",
    "上班 1 打卡结果",
    "上班 1 修改原因",
    "上班 1 打卡地点",
    "下班 1 打卡时间",
    "下班 1 打卡结果",
    "下班 1 修改原因",
    "下班 1 打卡地点",
]


def _write_wx_attendance(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WX_HEADERS_TOP)
    ws.append(WX_HEADERS)
    for row in rows:
        ws.append([row.get(header or "姓名") for header in WX_HEADERS])
    wb.save(path)


def _wx_row(**overrides):
    row = {
        "姓名": "WX员工",
        "工号": "WX00001",
        "部门": "客户支持项目组",
        "组织架构": "WX-PBU技术部-客户支持项目组",
        "性别": "男",
        "人员类型": "正式",
        "员工状态": "在职",
        "入职日期": "2026-04-01",
        "离职日期": "-",
        "日期": datetime(2026, 5, 6),
        "星期": "星期三",
        "班次": "202305新考勤 09:10-18:10;可早到早走40 分钟",
        "考勤类型": "需要打卡",
        "上班 1 打卡时间": "09:02",
        "上班 1 打卡结果": "正常",
        "上班 1 修改原因": "-",
        "上班 1 打卡地点": "14F门禁考勤机",
        "下班 1 打卡时间": "21:05",
        "下班 1 打卡结果": "正常",
        "下班 1 修改原因": "-",
        "下班 1 打卡地点": "14F门禁考勤机",
    }
    row.update(overrides)
    return row


def _row(**overrides):
    row = {
        "员工": "张三",
        "员工邮箱": "zhangsan@example.com",
        "工号": "zt10001",
        "考勤卡号": "zt10001",
        "人员状态": "正式",
        "二级组织": "HQU技术部",
        "三级组织": "技术部",
        "四级组织": "平台技术组",
        "五级组织": "研发组",
        "员工考勤地点": "集团深圳",
        "考勤日期": datetime(2026, 5, 6),
        "首打卡(含补签)": datetime(2026, 5, 6, 9, 0),
        "末打卡(含补签)": datetime(2026, 5, 6, 21, 5),
        "当前班次": "深圳灵活打卡8:30-9:10",
        "日期类型": "工作日",
        "备注": "",
    }
    row.update(overrides)
    return row


def _is_relative_to(path, parent):
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def test_meal_allowance_calculates_only_confirmed_rule_matches(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(员工="张三", 工号="zt10001", 考勤日期=datetime(2026, 5, 6)),
            _row(员工="李四", 工号="zt10002", 考勤日期=datetime(2026, 5, 7), **{"末打卡(含补签)": datetime(2026, 5, 7, 20, 59)}),
            _row(员工="王五", 工号="wx10003", 考勤日期=datetime(2026, 5, 8), **{"末打卡(含补签)": datetime(2026, 5, 8, 21, 30)}),
            _row(员工="赵六", 工号="zt10004", 考勤日期=datetime(2026, 5, 9), 日期类型="休息日", **{"末打卡(含补签)": datetime(2026, 5, 9, 22, 0)}),
        ],
    )

    parsed = parse_attendance_workbooks([file_path])
    result = calculate_meal_allowance(parsed)

    assert result["summary"]["totalAmount"] == 20
    assert result["summary"]["payableDayCount"] == 1
    rows = {row["employeeId"]: row for row in result["results"]}
    assert rows["zt10001"]["amount"] == 20
    assert rows["zt10002"]["amount"] == 0
    assert "末打卡(含补签)不在21:00-次日08:00范围" in rows["zt10002"]["warnings"]
    assert rows["wx10003"]["amount"] == 0
    assert "工号不是zt开头" in rows["wx10003"]["warnings"]
    assert rows["zt10004"]["amount"] == 0
    assert "日期类型不是工作日" in rows["zt10004"]["warnings"]


def test_meal_allowance_uses_offline_confirmed_organization_scope(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(
                员工="ABU技术",
                工号="zt11001",
                二级组织="ABU航空事业部",
                三级组织="ABU技术部",
                四级组织="操作履约组",
            ),
            _row(
                员工="HQU运维",
                工号="zt11002",
                二级组织="HQU技术部",
                三级组织="运维部",
                四级组织="应用运维组",
            ),
            _row(
                员工="HQU安全",
                工号="zt11003",
                二级组织="HQU技术部",
                三级组织="信息安全部",
                四级组织="",
            ),
            _row(
                员工="非技术组织",
                工号="zt11004",
                二级组织="ABU航空事业部",
                三级组织="产品部",
                四级组织="产品一组",
            ),
        ],
    )

    result = calculate_meal_allowance(parse_attendance_workbooks([file_path]))
    rows = {row["employeeId"]: row for row in result["results"]}

    assert rows["zt11001"]["amount"] == 20
    assert rows["zt11002"]["amount"] == 20
    assert rows["zt11003"]["amount"] == 20
    assert rows["zt11004"]["amount"] == 0
    assert "组织不在核算对象范围" in rows["zt11004"]["warnings"]


def test_meal_allowance_handles_next_day_early_clock_out_and_exclusions(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(员工="夜班员工", 工号="zt20001", **{"末打卡(含补签)": datetime(2026, 5, 7, 7, 30)}),
            _row(员工="潘江浩", 工号="zt20002", **{"末打卡(含补签)": datetime(2026, 5, 6, 22, 0)}),
            _row(
                员工="特殊组织",
                工号="zt20003",
                二级组织="HRAS人力综合条线",
                三级组织="战略运营部",
                四级组织="BI组",
                **{"末打卡(含补签)": datetime(2026, 5, 6, 22, 0)},
            ),
        ],
    )

    result = calculate_meal_allowance(parse_attendance_workbooks([file_path]))
    rows = {row["employeeId"]: row for row in result["results"]}

    assert rows["zt20001"]["amount"] == 20
    assert rows["zt20002"]["amount"] == 0
    assert "员工在不计算补贴名单" in rows["zt20002"]["warnings"]
    assert rows["zt20003"]["amount"] == 20


def test_meal_allowance_business_trip_remark_still_requires_late_clock_out(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(员工="公出晚归", 工号="zt30001", 备注="公出 客户现场", **{"末打卡(含补签)": datetime(2026, 5, 6, 21, 1)}),
            _row(员工="出差未晚归", 工号="zt30002", 备注="出差", **{"末打卡(含补签)": datetime(2026, 5, 6, 20, 30)}),
        ],
    )

    result = calculate_meal_allowance(parse_attendance_workbooks([file_path]))
    rows = {row["employeeId"]: row for row in result["results"]}

    assert rows["zt30001"]["amount"] == 20
    assert rows["zt30002"]["amount"] == 0


def test_meal_allowance_business_trip_remark_after_21_counts_without_late_punch(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(员工="公出超过21点", 工号="zt31001", 备注="公出 08:30-22:00[已通过]", **{"末打卡(含补签)": datetime(2026, 5, 6, 9, 27)}),
            _row(员工="出差未超过21点", 工号="zt31002", 备注="出差 08:30-20:30[已通过]", **{"末打卡(含补签)": datetime(2026, 5, 6, 20, 30)}),
        ],
    )

    result = calculate_meal_allowance(parse_attendance_workbooks([file_path]))
    rows = {row["employeeId"]: row for row in result["results"]}

    assert rows["zt31001"]["amount"] == 20
    assert rows["zt31002"]["amount"] == 0


def test_meal_allowance_does_not_include_non_hras_strategy_bi_group(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(
                员工="LBU战略BI",
                工号="zt32001",
                二级组织="LBU速运事业部",
                三级组织="战略运营部",
                四级组织="BI组",
            ),
            _row(
                员工="HRAS战略BI",
                工号="zt32002",
                二级组织="HRAS人力综合条线",
                三级组织="战略运营部",
                四级组织="BI组",
            ),
        ],
    )

    result = calculate_meal_allowance(parse_attendance_workbooks([file_path]))
    rows = {row["employeeId"]: row for row in result["results"]}

    assert rows["zt32001"]["amount"] == 0
    assert "组织不在核算对象范围" in rows["zt32001"]["warnings"]
    assert rows["zt32002"]["amount"] == 20


def test_parse_multiple_attendance_exports_without_duplicate_employee_dates(tmp_path):
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _write_attendance(first, [_row(工号="zt40001", 考勤日期=datetime(2026, 5, 1))])
    _write_attendance(second, [_row(工号="zt40001", 考勤日期=datetime(2026, 5, 2))])

    parsed = parse_attendance_workbooks([first, second])

    assert parsed.summary()["fileCount"] == 2
    assert parsed.summary()["rowCount"] == 2
    assert parsed.summary()["employeeCount"] == 1
    assert parsed.duplicate_keys == []


def test_hr_parser_opens_workbook_in_read_only_mode(tmp_path, monkeypatch):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(file_path, [_row(工号="zt41001", 考勤日期=datetime(2026, 5, 1))])
    calls = []

    def recording_load_workbook(*args, **kwargs):
        calls.append(kwargs.get("read_only"))
        return load_workbook(*args, **kwargs)

    monkeypatch.setattr(meal_allowance_module, "load_workbook", recording_load_workbook)

    parsed = parse_attendance_workbooks([file_path])

    assert calls == [True]
    assert parsed.summary()["rowCount"] == 1
    assert parsed.rows[0]["工号"] == "zt41001"


def test_wx_parser_opens_workbook_in_read_only_mode(tmp_path, monkeypatch):
    file_path = tmp_path / "wx.xlsx"
    _write_wx_attendance(file_path, [_wx_row(工号="WX41001", 日期=datetime(2026, 5, 1))])
    calls = []

    def recording_load_workbook(*args, **kwargs):
        calls.append(kwargs.get("read_only"))
        return load_workbook(*args, **kwargs)

    monkeypatch.setattr(meal_allowance_module, "load_workbook", recording_load_workbook)

    parsed = parse_wx_attendance_workbooks([file_path])

    assert calls == [True]
    assert parsed.summary()["rowCount"] == 1
    assert parsed.rows[0]["工号"] == "WX41001"


def test_china_employee_payroll_run_storage_does_not_trigger_dev_reload():
    if not _is_relative_to(app_module.OUTPUT_DIR, app_module.PROJECT_ROOT):
        return

    assert not _is_relative_to(app_module.CHINA_EMPLOYEE_PAYROLL_RUNS_DIR, app_module.PROJECT_ROOT)


def test_workbench_access_keeps_china_employee_payroll_open_when_developing_modules_hidden(monkeypatch):
    monkeypatch.setenv("SIGMA_HIDE_DEVELOPING_MODULES", "1")

    response = TestClient(app).get("/api/workbench/access")

    assert response.status_code == 200
    blocked_keys = {item["key"] for item in response.json()["blockedModules"]}
    assert "cn_employee_payroll" not in blocked_keys
    assert "domestic_labor" in blocked_keys
    assert "fbu_performance" not in blocked_keys


def test_meal_allowance_api_accepts_multiple_attendance_exports(tmp_path):
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    _write_attendance(first, [_row(工号="zt50001", 考勤日期=datetime(2026, 5, 1))])
    _write_attendance(second, [_row(工号="zt50001", 考勤日期=datetime(2026, 5, 2))])

    client = TestClient(app)
    with first.open("rb") as first_file, second.open("rb") as second_file:
        response = client.post(
            "/api/china-employee-payroll/meal-allowance",
            files=[
                ("attendance_files", ("first.xlsx", first_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                ("attendance_files", ("second.xlsx", second_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ],
        )

    assert response.status_code == 200
    data = response.json()
    assert data["summary"]["fileCount"] == 2
    assert data["summary"]["payableDayCount"] == 2
    assert data["summary"]["totalAmount"] == 40
    assert "dailyRows" not in data
    assert all("daily" not in row for row in data["results"])


def test_meal_allowance_export_xlsx_filters_payable_rows_and_keeps_source_sheet(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(员工="应发员工", 工号="zt51001", 考勤日期=datetime(2026, 5, 1)),
            _row(员工="未命中员工", 工号="zt51002", 考勤日期=datetime(2026, 5, 2), **{"末打卡(含补签)": datetime(2026, 5, 2, 20, 30)}),
        ],
    )

    client = TestClient(app)
    with file_path.open("rb") as upload:
        response = client.post(
            "/api/china-employee-payroll/meal-allowance",
            files={
                "attendance_files": (
                    "attendance.xlsx",
                    upload,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    run_id = response.json()["runId"]

    export_response = client.get(f"/api/china-employee-payroll/meal-allowance/{run_id}/export")

    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export_response.content), data_only=True)
    assert workbook.sheetnames == ["餐补核算结果", "人事系统考勤源"]

    result_sheet = workbook["餐补核算结果"]
    assert result_sheet["A1"].value == "核算月份"
    assert result_sheet["B1"].value == "工号"
    assert result_sheet["A2"].value == "2026年5月"
    exported_names = [cell.value for cell in result_sheet["C"][1:]]
    assert "应发员工" in exported_names
    assert "未命中员工" not in exported_names
    assert result_sheet.cell(result_sheet.max_row, 1).value == "合计"
    assert result_sheet.cell(result_sheet.max_row, 8).value == 1
    assert result_sheet.cell(result_sheet.max_row, 9).value == 20

    source_sheet = workbook["人事系统考勤源"]
    assert source_sheet["A2"].value == "员工"
    source_names = [cell.value for cell in source_sheet["A"][2:]]
    assert "应发员工" in source_names
    assert "未命中员工" in source_names


def test_meal_allowance_runs_can_be_listed_and_reopened(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(file_path, [_row(员工="批次员工", 工号="zt52001", 考勤日期=datetime(2026, 5, 3))])

    client = TestClient(app)
    with file_path.open("rb") as upload:
        response = client.post(
            "/api/china-employee-payroll/meal-allowance",
            files={
                "attendance_files": (
                    "attendance.xlsx",
                    upload,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    run_id = response.json()["runId"]

    runs_response = client.get("/api/china-employee-payroll/meal-allowance/runs")
    assert runs_response.status_code == 200
    runs = runs_response.json()["runs"]
    current_run = next(run for run in runs if run["runId"] == run_id)
    assert current_run["monthLabel"] == "2026年5月"
    assert current_run["payableEmployeeCount"] == 1
    assert current_run["totalAmount"] == 20

    run_response = client.get(f"/api/china-employee-payroll/meal-allowance/runs/{run_id}")
    assert run_response.status_code == 200
    data = run_response.json()
    assert data["runId"] == run_id
    assert data["summary"]["totalAmount"] == 20
    assert data["results"][0]["employeeName"] == "批次员工"
    assert "daily" not in data["results"][0]


def test_meal_allowance_api_response_stays_compact_for_browser(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(
        file_path,
        [
            _row(
                工号=f"zt{i:05d}",
                员工=f"员工{i}",
                考勤日期=datetime(2026, 5, day),
                **{"末打卡(含补签)": datetime(2026, 5, day, 21, 30)},
            )
            for i in range(1, 26)
            for day in range(1, 21)
        ],
    )

    client = TestClient(app)
    with file_path.open("rb") as upload:
        response = client.post(
            "/api/china-employee-payroll/meal-allowance",
            files={
                "attendance_files": (
                    "attendance.xlsx",
                    upload,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    data = response.json()
    assert data["summary"]["rowCount"] == 500
    assert len(response.content) < 100_000
    assert "dailyRows" not in data
    assert all("daily" not in row for row in data["results"])


def test_wx_meal_allowance_uses_wx_template_and_rules(tmp_path):
    file_path = tmp_path / "wx.xlsx"
    _write_wx_attendance(
        file_path,
        [
            _wx_row(姓名="晚归员工", 工号="WX00056", 日期=datetime(2026, 5, 6), **{"下班 1 打卡时间": "21:01"}),
            _wx_row(姓名="跨零点员工", 工号="WX00060", 日期=datetime(2026, 5, 12), **{"下班 1 打卡时间": "00:10"}),
            _wx_row(姓名="未晚归员工", 工号="WX00057", 日期=datetime(2026, 5, 7), **{"下班 1 打卡时间": "20:59"}),
            _wx_row(姓名="休息员工", 工号="WX00058", 日期=datetime(2026, 5, 8), 班次="休息 ", **{"下班 1 打卡时间": "22:10"}),
            _wx_row(姓名="体系外员工", 工号="WX00059", 日期=datetime(2026, 5, 9), 组织架构="WX-客服部", **{"下班 1 打卡时间": "21:30"}),
        ],
    )

    result = calculate_meal_allowance(parse_wx_attendance_workbooks([file_path]))
    rows = {row["employeeId"]: row for row in result["results"]}

    assert result["summary"]["sourceType"] == "wx"
    assert result["summary"]["sourceLabel"] == "WX技术部考勤"
    assert result["summary"]["totalAmount"] == 40
    assert result["summary"]["payableDayCount"] == 2
    assert rows["WX00056"]["amount"] == 20
    assert rows["WX00060"]["amount"] == 20
    assert rows["WX00057"]["amount"] == 0
    assert "末打卡(含补签)不在21:00-次日08:00范围" in rows["WX00057"]["warnings"]
    assert rows["WX00058"]["amount"] == 0
    assert "班次为休息或非工作日" in rows["WX00058"]["warnings"]
    assert rows["WX00059"]["amount"] == 0
    assert "组织不在核算对象范围" in rows["WX00059"]["warnings"]


def test_wx_meal_allowance_api_exports_separate_wx_source_sheet(tmp_path):
    file_path = tmp_path / "wx.xlsx"
    _write_wx_attendance(
        file_path,
        [
            _wx_row(姓名="WX应发", 工号="WX01001", 日期=datetime(2026, 5, 10), **{"下班 1 打卡时间": "21:30"}),
            _wx_row(姓名="WX未命中", 工号="WX01002", 日期=datetime(2026, 5, 11), **{"下班 1 打卡时间": "18:30"}),
        ],
    )

    client = TestClient(app)
    with file_path.open("rb") as upload:
        response = client.post(
            "/api/china-employee-payroll/meal-allowance",
            data={"source_type": "wx"},
            files={
                "attendance_files": (
                    "wx.xlsx",
                    upload,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 200
    data = response.json()
    assert data["summary"]["sourceType"] == "wx"
    assert data["summary"]["totalAmount"] == 20
    run_id = data["runId"]

    export_response = client.get(f"/api/china-employee-payroll/meal-allowance/{run_id}/export")

    assert export_response.status_code == 200
    assert "WX" in export_response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(export_response.content), data_only=True)
    assert workbook.sheetnames == ["餐补核算结果", "WX技术部考勤源"]

    result_sheet = workbook["餐补核算结果"]
    exported_names = [cell.value for cell in result_sheet["C"][1:]]
    assert "WX应发" in exported_names
    assert "WX未命中" not in exported_names
    assert result_sheet.cell(result_sheet.max_row, 8).value == 1
    assert result_sheet.cell(result_sheet.max_row, 9).value == 20

    source_sheet = workbook["WX技术部考勤源"]
    assert source_sheet["A1"].value == "姓名"
    assert source_sheet["B2"].value == "工号"
    source_names = [cell.value for cell in source_sheet["A"][2:]]
    assert "WX应发" in source_names
    assert "WX未命中" in source_names


def test_meal_allowance_api_rejects_unknown_source_type(tmp_path):
    file_path = tmp_path / "attendance.xlsx"
    _write_attendance(file_path, [_row()])

    client = TestClient(app)
    with file_path.open("rb") as upload:
        response = client.post(
            "/api/china-employee-payroll/meal-allowance",
            data={"source_type": "mixed"},
            files={
                "attendance_files": (
                    "attendance.xlsx",
                    upload,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

    assert response.status_code == 400
    assert "考勤来源类型不支持" in response.json()["detail"]
