from __future__ import annotations

from copy import deepcopy
from io import BytesIO
from pathlib import Path
import zipfile

from openpyxl import Workbook, load_workbook
import pytest

from bonus_platform.engine.social_insurance.report import store_template
from bonus_platform.engine.social_insurance.report_package import (
    build_export_preflight,
    build_missing_export,
    generate_report_package,
    resolve_package_download,
)
from bonus_platform.engine.social_insurance.runs import confirm_run, create_run
from bonus_platform.engine.social_insurance.template_library import match_template
from bonus_platform.engine.social_insurance.template_schemas import TEMPLATE_SCHEMAS
from tests.test_social_insurance_mvp import _record


def _shenzhen_record(*, identity: str = "440305199001011234", name: str = "模板测试员工") -> dict:
    record = _record(identity=identity, name=name)
    record["coverageSource"] = {
        "socialPlace": "深圳",
        "socialMedicalStatus": "社保待审核，医保待审核",
    }
    return record


def _run(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, record: dict) -> dict:
    runs = tmp_path / "runs"
    library = tmp_path / "template-library"
    library.mkdir()
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_RUNS_DIR", str(runs))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_TEMPLATE_LIBRARY_DIR", str(library))
    return create_run(
        records=[record],
        period_start="2026-07-16",
        period_end="2026-08-15",
        confirmation_date="2026-08-20",
        subject="深圳市前海云途物流有限公司",
        source="fixture",
    )


def _template_bytes(route: str) -> bytes:
    schema = TEMPLATE_SCHEMAS[route]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = schema["sheet"]
    for column, field in enumerate(schema["fields"], start=1):
        sheet.cell(row=schema["headerRow"], column=column, value=field["header"])
        sheet.cell(row=schema["dataStartRow"], column=column).number_format = "@"
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def test_template_registry_preserves_exact_field_order_for_all_offline_routes():
    assert len(TEMPLATE_SCHEMAS) == 8
    assert [field["name"] for field in TEMPLATE_SCHEMAS["shenzhen-social-medical"]["fields"]][:7] == [
        "证件号码", "姓名", "户籍", "入深户时间", "民族", "手机号码", "通讯地址",
    ]
    assert len(TEMPLATE_SCHEMAS["shenzhen-social-medical"]["fields"]) == 20
    assert len(TEMPLATE_SCHEMAS["zhejiang-social-medical"]["fields"]) == 21
    assert len(TEMPLATE_SCHEMAS["chengdu-medical"]["fields"]) == 22


def test_configured_shenzhen_template_is_trusted_even_without_subject_name(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    library = tmp_path / "template-library"
    library.mkdir()
    configured = tmp_path / "current-government-template.xls"
    configured.write_bytes(b"configured-template")
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_TEMPLATE_LIBRARY_DIR", str(library))
    monkeypatch.setenv("SIGMA_SOCIAL_INSURANCE_TEMPLATE_FILE", str(configured))

    matched = match_template("shenzhen-social-medical", "深圳易可达科技有限公司")

    assert matched is not None
    assert matched["filename"] == configured.name
    assert matched["matchQuality"] == "configured"
    assert matched["subjectMatched"] is True


def test_preflight_groups_missing_required_fields_by_route(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    record = _shenzhen_record()
    record["report"] = deepcopy(record["report"])
    record["report"]["户口所在地行政区划代码"] = ""
    run = _run(tmp_path, monkeypatch, record)

    preflight = build_export_preflight(run["id"])

    assert preflight["summary"]["templateRoutes"] == 1
    assert preflight["summary"]["blockedRoutes"] == 1
    group = preflight["groups"][0]
    assert group["route"] == "shenzhen-social-medical"
    assert group["status"] == "missing-fields"
    assert group["missingEmployees"] == 1
    assert group["missingFields"] == [{"field": "户口所在地行政区划代码", "count": 1}]

    workbook = load_workbook(build_missing_export(run["id"]), read_only=True, data_only=True)
    sheet = workbook["待补资料"]
    assert sheet["F2"].value == "户口所在地行政区划代码"
    assert sheet["D2"].value is None
    workbook.close()


def test_confirmed_run_generates_route_workbook_and_review_materials_in_one_zip(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    route = "shenzhen-social-medical"
    run = _run(tmp_path, monkeypatch, _shenzhen_record(name="完整字段员工"))
    run = confirm_run(run["id"])
    run = store_template(run["id"], "深圳当前版模板.xlsx", _template_bytes(route), route=route)

    preflight = build_export_preflight(run["id"])
    assert preflight["summary"] == {
        "templateRoutes": 1,
        "readyRoutes": 1,
        "blockedRoutes": 0,
        "manualRoutes": 0,
        "employeeCount": 1,
    }
    assert preflight["groups"][0]["template"]["source"] == "uploaded"

    generated = generate_report_package(run["id"])
    package = generated["reportPackage"]
    assert package["partial"] is False
    assert package["generatedRoutes"][0]["route"] == route

    with zipfile.ZipFile(resolve_package_download(run["id"])) as archive:
        names = archive.namelist()
        report_name = next(name for name in names if name.startswith("报盘文件/") and name.endswith(".xlsx"))
        assert "审核资料/待补资料.xlsx" in names
        assert any(name.startswith("审核资料/社保增员审核清单_") for name in names)
        assert "生成说明.xlsx" in names
        workbook = load_workbook(BytesIO(archive.read(report_name)), read_only=True, data_only=True)
        sheet = workbook[TEMPLATE_SCHEMAS[route]["sheet"]]
        assert sheet.cell(row=TEMPLATE_SCHEMAS[route]["dataStartRow"], column=2).value == "完整字段员工"
        workbook.close()
