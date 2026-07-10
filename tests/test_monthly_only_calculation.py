from io import BytesIO
from datetime import datetime
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl import load_workbook

from bonus_platform.app import app
from bonus_platform.config import DEFAULT_RULE_WORKBOOK
from bonus_platform.engine.calculator import calculate
from bonus_platform.engine.models import ImportRow
from bonus_platform.engine.rules import load_rulebook


def _monthly_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入_月度数据"
    sheet.append(["核算月份", "工号", "姓名"])
    sheet.append([202510, "zt-test-001", "月度测试"])
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_calculation_ignores_undeclared_history_upload():
    client = TestClient(app)

    response = client.post(
        "/api/calculate",
        files={
            "file": ("monthly.xlsx", _monthly_workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "history_file": ("legacy.txt", b"not a workbook", "text/plain"),
        },
    )

    assert response.status_code == 200
    assert response.json()["month"] == 202510
    assert "historySource" not in response.json()
    assert "historyRows" not in response.json()


def test_download_template_contains_only_monthly_input_sheets():
    client = TestClient(app)

    response = client.get("/api/template")
    workbook = load_workbook(BytesIO(response.content), data_only=False, read_only=True)

    assert response.status_code == 200
    assert workbook.sheetnames == ["使用说明", "导入_月度数据"]


def test_download_template_hides_legacy_override_columns_from_monthly_import():
    client = TestClient(app)

    response = client.get("/api/template")
    workbook = load_workbook(BytesIO(response.content), data_only=False, read_only=True)
    sheet = workbook["导入_月度数据"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]

    assert response.status_code == 200
    assert not [header for header in headers if header and str(header).endswith("_覆盖")]


def test_download_template_label_formula_matches_domestic_location_variants():
    client = TestClient(app)

    response = client.get("/api/template")
    workbook = load_workbook(BytesIO(response.content), data_only=False, read_only=True)
    sheet = workbook["导入_月度数据"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_cols = {header: index for index, header in enumerate(headers, start=1)}

    assert response.status_code == 200
    assert header_cols["招聘负责人人员状态"] == header_cols["招聘负责人姓名"] + 1
    assert header_cols["招聘负责人最后工作日"] == header_cols["招聘负责人姓名"] + 2
    assert header_cols["协助招聘人人员状态"] == header_cols["协助招聘人姓名"] + 1
    assert header_cols["协助招聘人最后工作日"] == header_cols["协助招聘人姓名"] + 2
    assert header_cols["推荐人最后工作日"] == header_cols["推荐人人员状态"] + 1
    label_formula = sheet["S2"].value
    region_formula = sheet.cell(2, header_cols["奖金地区类型"]).value
    assert "非中国" in label_formula
    assert "香港" in label_formula
    assert "广东" in label_formula
    assert "深圳" in label_formula
    assert "香港" in region_formula
    assert "广东" in region_formula
    assert "德国" in region_formula


def test_current_month_calculation_ignores_legacy_override_fields():
    values = {
        "核算月份": 202510,
        "姓名": "覆盖测试",
        "工号": "zt-override-001",
        "工作地": "中国大陆",
        "标签分类": "国内",
        "职级": "P1-3",
        "ABC类别": "C类",
        "招聘渠道": "招聘网站",
        "招聘负责人工号": "zt-recruiter",
        "招聘负责人姓名": "招聘负责人",
        "招聘启动日期": 45500,
        "候选人入职时间": 45925,
        "转正日期": 46017,
        "招聘人入职1月发放金额_覆盖": 99999,
        "招聘人入职1月发放周期_覆盖": 202510,
    }

    with_overrides = calculate([ImportRow(source_row=9, values=values)], load_rulebook(DEFAULT_RULE_WORKBOOK))
    without_overrides = calculate(
        [ImportRow(source_row=9, values={key: value for key, value in values.items() if not key.endswith("_覆盖")})],
        load_rulebook(DEFAULT_RULE_WORKBOOK),
    )

    assert with_overrides.details[0].recruiter_1m_bonus == without_overrides.details[0].recruiter_1m_bonus
    assert with_overrides.details[0].recruiter_1m_bonus != 99999


def test_pending_source_row_uses_uploaded_excel_row_number():
    rows = [
        ImportRow(
            source_row=12,
            values={
                "核算月份": 202510,
                "姓名": "源行测试",
                "工号": "zt-row-001",
                "工作地": "中国大陆",
                "标签分类": "国内",
                "职级": "P1-3",
                "ABC类别": "C类",
                "招聘渠道": "招聘网站",
                "招聘负责人工号": "zt-recruiter",
                "招聘负责人姓名": "招聘负责人",
                "招聘负责人人员状态": "",
                "招聘启动日期": 45800,
                "候选人入职时间": 45925,
                "转正日期": 46017,
            },
        )
    ]

    result = calculate(rows, load_rulebook(DEFAULT_RULE_WORKBOOK))

    assert result.pending_confirmations
    assert result.pending_confirmations[0]["源行号"] == 12


def test_calculation_accepts_excel_date_as_calculation_month():
    result = calculate(
        [
            ImportRow(
                source_row=2,
                values={
                    "核算月份": datetime(2026, 6, 1),
                    "姓名": "月份格式测试",
                    "工号": "zt-month-date-001",
                },
            )
        ],
        load_rulebook(DEFAULT_RULE_WORKBOOK),
    )

    assert result.month == 202606


def test_calculation_accepts_iso_text_dates_for_actual_cycle():
    result = calculate(
        [
            ImportRow(
                source_row=2,
                values={
                    "核算月份": 202606,
                    "姓名": "文本日期测试",
                    "工号": "zt-text-date-001",
                    "工作地": "中国大陆",
                    "标签分类": "国内",
                    "职级": "P1-3",
                    "ABC类别": "C类",
                    "招聘渠道": "招聘网站",
                    "招聘启动日期": "2026-05-20",
                    "候选人入职时间": "2026-06-20",
                },
            )
        ],
        load_rulebook(DEFAULT_RULE_WORKBOOK),
    )

    assert result.details[0].actual_cycle_days == 31
    assert result.details[0].adjusted_total_bonus > 0


def test_uncalculable_current_import_row_remains_in_exception_list():
    result = calculate(
        [
            ImportRow(
                source_row=8,
                values={
                    "核算月份": 202606,
                    "姓名": "缺字段测试",
                    "工号": "zt-missing-field-001",
                    "工作地": "中国大陆",
                    "标签分类": "国内",
                    "职级": "P1-3",
                    "ABC类别": "",
                    "招聘渠道": "招聘网站",
                    "招聘启动日期": "",
                    "候选人入职时间": datetime(2026, 6, 22),
                },
            )
        ],
        load_rulebook(DEFAULT_RULE_WORKBOOK),
    )

    exception_types = {row["异常类型"] for row in result.exceptions}
    assert "招聘奖金缺关键字段" in exception_types
    assert "未匹配招聘周期" in exception_types
    assert "未匹配招聘奖金标准" in exception_types


def test_uncalculable_historical_row_does_not_pollute_current_month_exception_list():
    result = calculate(
        [
            ImportRow(
                source_row=9,
                values={
                    "核算月份": 202606,
                    "姓名": "历史缺字段测试",
                    "工号": "zt-historical-missing-001",
                    "工作地": "中国大陆",
                    "标签分类": "国内",
                    "职级": "P1-3",
                    "ABC类别": "",
                    "招聘渠道": "招聘网站",
                    "招聘启动日期": "",
                    "候选人入职时间": datetime(2025, 6, 22),
                },
            )
        ],
        load_rulebook(DEFAULT_RULE_WORKBOOK),
    )

    assert result.month == 202606
    assert result.exceptions == []


def test_uncalculable_row_with_recruitment_node_due_this_month_remains_in_exception_list():
    result = calculate(
        [
            ImportRow(
                source_row=10,
                values={
                    "核算月份": 202606,
                    "姓名": "本月节点缺字段测试",
                    "工号": "zt-current-node-missing-001",
                    "工作地": "中国大陆",
                    "标签分类": "国内",
                    "职级": "P1-3",
                    "ABC类别": "",
                    "招聘渠道": "招聘网站",
                    "招聘启动日期": "",
                    "候选人入职时间": datetime(2026, 5, 20),
                },
            )
        ],
        load_rulebook(DEFAULT_RULE_WORKBOOK),
    )

    exception_types = {row["异常类型"] for row in result.exceptions}
    assert "招聘奖金缺关键字段" in exception_types
    assert "未匹配招聘周期" in exception_types
    assert "未匹配招聘奖金标准" in exception_types
