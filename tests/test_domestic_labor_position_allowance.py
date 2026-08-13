"""岗位补贴核算规则与2026年7月线下结果回归。"""
from decimal import Decimal
from io import BytesIO
import time

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from bonus_platform.app import app
from bonus_platform.engine.domestic_labor.engines.gangwei_butie import GangWeiBuTieEngine


pytestmark = pytest.mark.usefixtures("bypass_domestic_labor_access_gate")


JULY_VERIFIED_ROWS = [
    ("OWHN14187", "陈康", "民航初级安检员", 23, {}, 1300.00),
    ("OWHN7962", "陈莉", "内部中级安检员", 23, {}, 450.00),
    ("OWHN10300", "陈晓龙", "安检组长", 25, {"休年假小时": 22, "调休时数": 35.5, "入离职缺勤时数": 24}, 474.00),
    ("OWHN14223", "方宇杰", "民航中级安检员", 23, {}, 1500.00),
    ("OWHN7936", "黄志才", "内部初级安检员", 23, {"排休请假时数": 8, "旷工时数": 8}, 300.00),
    ("OWHN9372", "李宝兰", "HRBP专员", 23, {"调休时数": 19.5}, 700.00),
    ("OWHN10229", "李奎", "民航中级安检员", 23, {"排休请假时数": 24}, 1500.00),
    ("OWHN9378", "罗华兰", "高级招聘专员", 23, {"调休时数": 4.5}, 700.00),
    ("OWHN12479", "宁绍松", "内部初级安检员", 23, {"排休请假时数": 8}, 300.00),
    ("OWHN8926", "邱润隆", "内部高级安检员", 23, {"排休请假时数": 16, "休年假小时": 8}, 650.00),
    ("OWHN10301", "沈烨炀", "民航中级安检员", 23, {"排休请假时数": 16}, 1500.00),
    ("OWHN9377", "王雯雯", "高级HRBP专员", 23, {"调休时数": 21.5}, 700.00),
    ("OWHN10564", "文勇", "内部初级安检员", 23, {}, 300.00),
    ("OWHN2214", "向明亮", "内部高级安检员", 23, {"排休请假时数": 8}, 650.00),
    ("OWHN6471", "许国庆", "内部中级安检员", 23, {"排休请假时数": 16}, 450.00),
    ("OWHN14047", "张婷", "民航中级安检员", 23, {"排休请假时数": 16}, 1500.00),
    ("OWHN12893", "郑东升", "民航初级安检员", 23, {"排休请假时数": 24}, 1300.00),
    ("OWHN6731", "钟昌凡", "叉车司机", 23, {"排休请假时数": 8}, 800.00),
]


def _row(
    employee_id="OWHN001",
    employee_name="张三",
    position="内部初级安检员",
    scheduled_days=23,
    work_area="东莞",
    **fields,
):
    return {
        "工号": employee_id,
        "姓名": employee_name,
        "工作地区": work_area,
        "岗位名称": position,
        "排班天数": scheduled_days,
        "实际在职工作日天数": scheduled_days,
        "职级": "应忽略的职级",
        **fields,
    }


def test_july_verified_roster_reconciles_all_18_people_and_total():
    engine = GangWeiBuTieEngine()

    results = [
        engine.calculate(_row(employee_id, name, position, scheduled_days, **absence_fields))
        for employee_id, name, position, scheduled_days, absence_fields, _expected in JULY_VERIFIED_ROWS
    ]

    assert [result.amount for result in results] == [row[-1] for row in JULY_VERIFIED_ROWS]
    assert sum(Decimal(str(result.amount)) for result in results) == Decimal("15074.00")
    assert all(result.warnings == [] for result in results)


@pytest.mark.parametrize(
    ("position", "expected_standard"),
    [
        ("内部初级安检员", 300),
        ("内部中级安检员", 450),
        ("内部高级安检员", 650),
        ("民航初级安检员", 1300),
        ("民航中级安检员", 1500),
        ("HRBP专员", 700),
        ("高级HRBP专员", 700),
        ("高级招聘专员", 700),
        ("叉车司机", 800),
    ],
)
def test_verified_position_standard_ignores_job_grade(position, expected_standard):
    result = GangWeiBuTieEngine().calculate(_row(position=position, 职级="P9-9"))

    assert result.amount == expected_standard
    assert result.details["岗位补贴标准"] == expected_standard
    assert result.details["职级参与计算"] is False


def test_womens_day_leave_is_converted_to_eight_hours_before_threshold():
    result = GangWeiBuTieEngine().calculate(_row(女神假天数=7))

    assert result.details["女神假折算时数"] == 56
    assert result.details["缺勤合计时数"] == 56
    assert result.details["扣减天数"] == 7
    assert result.amount == 208.70


def test_less_than_56_hours_does_not_reduce_position_allowance():
    result = GangWeiBuTieEngine().calculate(_row(事假时数=55.5))

    assert result.details["缺勤合计时数"] == 55.5
    assert result.details["扣减天数"] == 0
    assert result.amount == 300


def test_entry_exit_absence_is_derived_from_schedule_and_actual_work_days():
    result = GangWeiBuTieEngine().calculate(_row(事假时数=40, 实际在职工作日天数=20))

    assert result.details["缺勤明细"]["入离职缺勤时数"] == 24
    assert result.details["入离职缺勤时数来源"] == "按排班天数与实际在职工作日天数自动计算"
    assert result.details["缺勤合计时数"] == 64
    assert result.details["扣减天数"] == 8
    assert result.amount == 195.65


def test_reported_entry_exit_absence_takes_precedence_over_derived_value():
    result = GangWeiBuTieEngine().calculate(
        _row(实际在职工作日天数=20, 入离职缺勤时数=72)
    )

    assert result.details["缺勤明细"]["入离职缺勤时数"] == 72
    assert result.details["入离职缺勤时数来源"] == "月考勤已有值"
    assert result.amount == 182.61


def test_special_group_leaders_are_recognized_by_person_not_grade():
    engine = GangWeiBuTieEngine()

    chen = engine.calculate(_row("OWHN10300", "陈晓龙", "安检组长", 25))
    jia = engine.calculate(_row("OWHN-JW", "贾万", "操作组长", 25, work_area="晋江"))

    assert chen.amount == 800
    assert jia.amount == 800
    assert chen.details["资格依据"] == "特殊安检组长名单"
    assert jia.details["资格依据"] == "特殊安检组长名单"
    assert chen.warnings == []
    assert jia.details["exceptions"][0]["code"] == "POSITION_ALLOWANCE_SPECIAL_STANDARD_PENDING"
    assert "暂算" in jia.warnings[0]


@pytest.mark.parametrize("position", ["安检员", "民航高级安检员", "揽收充电司机"])
def test_eligible_but_unverified_standard_is_explicitly_held_for_review(position):
    result = GangWeiBuTieEngine().calculate(_row(position=position))

    assert result.amount == 0
    assert result.details["资格判断"] == "有资格，标准待确认"
    assert result.details["exceptions"][0]["code"] == "POSITION_ALLOWANCE_STANDARD_PENDING"
    assert "标准待确认" in result.warnings[0]


def test_non_eligible_position_returns_zero_without_false_warning():
    result = GangWeiBuTieEngine().calculate(_row(position="操作员"))

    assert result.amount == 0
    assert result.warnings == []
    assert result.details["资格判断"] == "不享有岗位补贴"


def _position_allowance_workbook(attendance_month: str = "202607") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{attendance_month}月考勤"
    headers = [
        "考勤月份", "一级部门名称", "工作地区", "工号", "姓名", "岗位名称", "排班天数",
        "实际在职工作日天数",
        "事假时数", "排休请假时数", "病假时数", "旷工时数", "休年假小时", "女神假天数",
        "其他假时数（带薪）", "调休时数", "入离职缺勤时数",
    ]
    sheet.append(headers)
    for employee_id, name, position, scheduled_days, absence_fields, _expected in JULY_VERIFIED_ROWS:
        base = {
            "考勤月份": attendance_month,
            "一级部门名称": "莞深操作",
            "工作地区": "东莞",
            "工号": employee_id,
            "姓名": name,
            "岗位名称": position,
            "排班天数": scheduled_days,
            "实际在职工作日天数": scheduled_days,
            **absence_fields,
        }
        sheet.append([base.get(header, 0) for header in headers])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _single_position_allowance_workbook(*, include_actual_work_days: bool) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "202607月考勤"
    headers = [
        "考勤月份", "一级部门名称", "工作地区", "工号", "姓名", "岗位名称", "排班天数",
        "事假时数", "入离职缺勤时数",
    ]
    if include_actual_work_days:
        headers.insert(7, "实际在职工作日天数")
    row = {
        "考勤月份": "202607",
        "一级部门名称": "莞深操作",
        "工作地区": "东莞",
        "工号": "OWHN-AUTO",
        "姓名": "自动折算示例",
        "岗位名称": "内部初级安检员",
        "排班天数": 23,
        "实际在职工作日天数": 20,
        "事假时数": 40,
        "入离职缺勤时数": 0,
    }
    sheet.append(headers)
    sheet.append([row.get(header) for header in headers])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _run_position_allowance_workbook(workbook_bytes: bytes) -> tuple[TestClient, str, dict]:
    client = TestClient(app)
    response = client.post(
        "/api/domestic-labor/runs",
        files={
            "file": (
                "岗位补贴入离职折算.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"engines": "gangwei_butie", "attendance_month": "202607"},
    )
    assert response.status_code == 200
    run_id = response.json()["run_id"]
    for _ in range(30):
        metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
        if metadata["status"] in {"已完成", "失败"}:
            break
        time.sleep(0.1)
    assert metadata["status"] == "已完成"
    return client, run_id, metadata


def test_position_allowance_api_derives_entry_exit_absence_from_monthly_attendance():
    client, run_id, metadata = _run_position_allowance_workbook(
        _single_position_allowance_workbook(include_actual_work_days=True)
    )
    try:
        result = metadata["results"][0]
        detail = result["subject_details"]["gangwei_butie"]["details"]
        assert result["gangwei_butie"] == 195.65
        assert detail["缺勤明细"]["入离职缺勤时数"] == 24
        assert detail["入离职缺勤时数来源"] == "按排班天数与实际在职工作日天数自动计算"
    finally:
        client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_position_allowance_missing_actual_work_days_warns_without_blocking_calculation():
    client, run_id, metadata = _run_position_allowance_workbook(
        _single_position_allowance_workbook(include_actual_work_days=False)
    )
    try:
        result = metadata["results"][0]
        detail = result["subject_details"]["gangwei_butie"]["details"]
        exceptions = result["subject_details"]["gangwei_butie"]["exceptions"]
        assert result["gangwei_butie"] == 300
        assert detail["缺勤明细"]["入离职缺勤时数"] == 0
        assert detail["入离职缺勤时数来源"] == "实际在职工作日天数缺失，暂按0小时"
        assert exceptions[0]["code"] == "POSITION_ALLOWANCE_ENTRY_EXIT_ABSENCE_PENDING"
        assert exceptions[0]["level"] == "info"
    finally:
        client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_position_allowance_does_not_block_months_before_july_2026():
    client = TestClient(app)
    response = client.post(
        "/api/domestic-labor/runs",
        files={
            "file": (
                "5月岗位补贴.xlsx",
                _position_allowance_workbook("202605"),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"engines": "gangwei_butie", "attendance_month": "202605"},
    )

    assert response.status_code == 200
    run_id = response.json()["run_id"]
    try:
        for _ in range(30):
            metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
            if metadata["status"] in {"已完成", "失败"}:
                break
            time.sleep(0.1)
        assert metadata["status"] == "已完成"
    finally:
        client.delete(f"/api/domestic-labor/runs/{run_id}")


def test_position_allowance_api_and_export_reconcile_july_baseline():
    client = TestClient(app)
    response = client.post(
        "/api/domestic-labor/runs",
        files={"file": ("7月岗位补贴回归.xlsx", _position_allowance_workbook(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"engines": "gangwei_butie", "attendance_month": "202607"},
    )

    assert response.status_code == 200
    run_id = response.json()["run_id"]
    try:
        for _ in range(30):
            metadata = client.get(f"/api/domestic-labor/runs/{run_id}").json()
            if metadata["status"] in {"已完成", "失败"}:
                break
            time.sleep(0.1)

        assert metadata["status"] == "已完成"
        assert metadata["summary"]["total_gangwei_butie"] == 15074
        assert [row["gangwei_butie"] for row in metadata["results"]] == [row[-1] for row in JULY_VERIFIED_ROWS]

        export = client.get(f"/api/domestic-labor/runs/{run_id}/export").json()
        workbook = load_workbook(export["file_path"], data_only=True)
        assert workbook.sheetnames == ["岗位补贴核算结果"]
        sheet = workbook["岗位补贴核算结果"]
        assert [cell.value for cell in sheet[1]] == [
            "工号", "姓名", "工作地区", "部门", "岗位", "资格判断", "岗位补贴标准", "排班天数",
            "缺勤合计时数", "扣减天数", "岗位补贴计发天数", "应发岗位补贴", "核算状态", "需处理事项",
        ]
        assert sum(sheet.cell(row=row_index, column=12).value for row_index in range(2, 20)) == 15074
        workbook.close()
    finally:
        client.delete(f"/api/domestic-labor/runs/{run_id}")
