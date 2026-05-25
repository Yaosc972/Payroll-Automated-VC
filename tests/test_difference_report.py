from pathlib import Path

from openpyxl import Workbook, load_workbook

from bonus_platform.engine.compare import build_difference_report


def test_difference_report_reads_payroll_style_multirow_offline_headers(tmp_path: Path):
    platform_path = tmp_path / "platform.xlsx"
    offline_path = tmp_path / "offline.xlsx"
    output_path = tmp_path / "diff.xlsx"

    platform = Workbook()
    recruitment = platform.active
    recruitment.title = "招聘奖金汇总"
    recruitment.append(["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    recruitment.append(["zt-rec", "招聘人", "招聘负责人", "人民币", 202510, 100, 0, 0, 0, 100])
    referral = platform.create_sheet("内推奖金汇总")
    referral.append(["推荐人工号", "推荐人姓名", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"])
    referral.append(["zt-ref", "推荐人", "人民币", 202510, 150, 0, 0, 0, 150])
    detail = platform.create_sheet("招聘奖金明细")
    detail.append(["工号", "姓名", "招聘人入职1月奖金", "内推入职1月奖金"])
    detail.append(["zt-ee", "员工", 100, 150])
    platform.save(platform_path)

    offline = Workbook()
    rec_sum = offline.active
    rec_sum.title = "附件二-招聘奖金汇总表"
    rec_sum.append(["重复验证", "工号", "姓名", "职位", "币种", "人员状态", "最后工作日", 202510, None, None, None, None])
    rec_sum.append([None, None, None, None, None, None, None, "入职1个月发放奖金", "入职3个月发放奖金", "入职6个月发放奖金", "转正奖金", "合计发放"])
    rec_sum.append([1, "zt-rec", "招聘人", "职位", "人民币", "正式", None, 80, 0, 0, 0, 80])
    ref_sum = offline.create_sheet("附件三-内推奖金汇总")
    ref_sum.append(["验证", "内推人员工号", "内推人员姓名", "职位", "职级", "人员状态", "币种", 202510])
    ref_sum.append([None, None, None, None, None, None, None, "本月发放"])
    ref_sum.append(["唯一", "zt-ref", "推荐人", "职位", "P1-3", "正式", "人民币", 120])
    off_detail = offline.create_sheet("附件一-招聘奖金核算明细")
    off_detail.append(["公式", "花名册", "花名册", "公式", "公式", "公式", "公式"])
    off_detail.append(["验证唯一", "姓名", "工号", "招聘负责人入职奖金明细", None, "内推奖明细", None])
    off_detail.append([None, None, None, "入职1个月发放奖金", "入职1个月发放周期", "入职1个月发放奖金", "入职1个月发放周期"])
    off_detail.append(["唯一", "员工", "zt-ee", 80, 202510, 120, 202510])
    offline.save(offline_path)

    metrics = build_difference_report(platform_path, offline_path, output_path)

    assert metrics["recruitmentSummaryDiffCount"] == 2
    assert metrics["referralSummaryDiffCount"] == 1
    assert metrics["recruitmentDetailDiffCount"] == 1
    assert metrics["referralDetailDiffCount"] == 1
    workbook = load_workbook(output_path, data_only=True, read_only=True)
    assert workbook["招聘明细差异"].cell(2, 6).value == 20
