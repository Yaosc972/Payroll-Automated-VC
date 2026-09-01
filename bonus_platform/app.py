from __future__ import annotations

import asyncio
import csv
import hashlib
from concurrent.futures import ThreadPoolExecutor
from contextlib import ExitStack, asynccontextmanager, contextmanager
from datetime import date, datetime, timedelta
import json
import logging
import mimetypes
import os
import re
import secrets
import threading
import time
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%H:%M:%S",
)
# httpx request INFO logs include full signed Storage URLs. Keep application
# telemetry at INFO while preventing short-lived credentials from entering logs.
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
from pathlib import Path
from dataclasses import replace
import shutil
from tempfile import NamedTemporaryFile, gettempdir
from time import monotonic, perf_counter
from typing import Any, Callable, Optional
from urllib.parse import quote, urlencode, urlparse
from uuid import uuid4
from fastapi import BackgroundTasks, Body, Cookie, Depends, FastAPI, File, Form, Header, HTTPException, Request, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware

from bonus_platform.time_utils import utcnow_naive
import httpx
from openpyxl import Workbook, load_workbook

logger = logging.getLogger("bonus_platform.labor")
fbu_logger = logging.getLogger("bonus_platform.fbu")
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from starlette.datastructures import MutableHeaders
from starlette.middleware.gzip import GZipMiddleware

from .auth import (
    current_user_from_request,
    labor_auth_health,
    labor_auth_required,
    user_can_enter_module,
    user_is_system_admin,
)
from .config import AI_CONFIG, AUTH_CONFIG, DEFAULT_IMPORT_TEMPLATE, DEFAULT_RULE_WORKBOOK, EXPORT_DIR, OUTPUT_DIR, MAX_PREVIEW_ROWS, DOMESTIC_LABOR_RUNS_DIR, FBU_PERFORMANCE_RUNS_DIR, LABOR_RUNS_DIR, PROJECT_ROOT, ensure_data_files
from .engine.domestic_labor.parser import MultiFilePayrollDataLoader
from .engine.domestic_labor.engines import (
    QuanQinJiangEngine,
    CanBuEngine,
    WaiSuBuTieEngine,
    GongLingJiangEngine,
    YeBanBuTieEngine,
    GangWeiBuTieEngine,
    GaoWenBuTieEngine,
)
from .engine.domestic_labor.templates import generate_template, get_template_info, ENGINE_TEMPLATES
from .engine.domestic_labor.exporter import ExcelExporter
from .engine.domestic_labor.night_shift_config import (
    build_night_shift_config_snapshot,
    config_counts,
    copy_night_shift_config,
    generate_night_shift_config_workbook,
    list_night_shift_config_revisions,
    load_night_shift_config,
    parse_night_shift_config_workbook,
    save_night_shift_config,
)
from .engine.domestic_labor.rule_package import get_rule_package
from .engine.domestic_labor.runs import (
    create_payroll_run, update_payroll_metadata, load_payroll_metadata, load_payroll_status,
    list_payroll_metadata, get_payroll_run_dir, attach_payroll_file, safe_payroll_filename,
    delete_payroll_run, materialize_payroll_file, persist_payroll_file,
)
from .engine.domestic_labor.persistent_storage import (
    create_domestic_labor_signed_upload,
    domestic_labor_persistent_storage_enabled,
)
from .engine.china_employee_payroll import calculate_meal_allowance, parse_attendance_workbooks, parse_wx_attendance_workbooks
from .engine.calculator import calculate
from .engine.compare import build_difference_report
from .engine.labor.compare import amount_within_tolerance, compare_labor_items, compare_by_warehouse
from .engine.labor.extract import _ai_ready, _warehouse_id_from_filename, _warehouse_id_from_text, extract_invoice_items, quick_extract_totals
from .engine.labor.governance import (
    audit_ai_page_cache_candidates,
    build_ai_cache_reconciliation_preview,
    build_reocr_candidate_plan,
    build_rule_change_candidate,
    replay_ai_cache_candidate_result,
    replay_reocr_candidate_result,
    confirm_rule_candidate,
    rollback_rule_version,
    summarize_rule_auto_replay,
    summarize_rule_replay,
)
from .engine.labor.quality import calculate_extraction_quality, calculate_quality_score, build_reconciliation_diagnostics
from .engine.labor.ocr_candidate_runtime import evaluate_ocr_candidate_result, run_ocr_candidate_command
from .engine.labor.presentation import build_labor_presentation, validate_labor_presentation
from .engine.labor.report import build_labor_business_html_report, build_labor_governance_report, build_labor_projection_report, build_labor_report
from .engine.labor.structure import evaluate_batch_guards, extract_structured_invoice_rows, prefer_closed_structured_rows, promote_structured_invoice_evidence, resolve_amount_scope
from .engine.labor.materials import (
    _attach_text_coverage_to_reocr_plan,
    _build_material_combined_row_governance,
    _build_material_review_queues,
    _summarize_pdf_text_coverage,
    build_material_dry_run,
    build_material_index,
    build_material_replay_plan,
)
from .engine.labor.profiles import (
    SupplierExtractionProfile,
    generate_profile_from_extraction,
    is_runtime_approved_profile,
    load_supplier_profiles,
    resolve_supplier_profile,
    supplier_alias_match_score,
)
from .engine.labor.worker_jobs import (
    LaborWorkerLeaseError,
    claim_labor_worker_job,
    clear_labor_worker_result_acceptance,
    complete_labor_worker_job,
    complete_labor_worker_auxiliary_job,
    complete_labor_worker_preflight_job,
    fail_labor_worker_job,
    get_latest_labor_worker_job,
    get_labor_worker_job,
    heartbeat_labor_worker_job,
    list_labor_worker_jobs,
    enqueue_labor_worker_job,
    labor_worker_job_store_health as labor_p1_worker_job_store_health,
    mark_labor_worker_result_accepted,
)
from .engine.labor.operations import build_labor_operations_snapshot
from .engine.labor.production_readiness import evaluate_labor_production_readiness
from .engine.labor.state_postgres import (
    LaborStateConflict,
    LaborStateNotFound,
    LaborStateOwnerMismatch,
    create_pending_labor_file_states,
    finalize_labor_file_state,
    finalize_labor_file_states,
    get_labor_file_state,
    labor_postgres_state_enabled,
    labor_postgres_state_health,
    list_labor_file_states,
    soft_delete_labor_run_state,
)
from .engine.labor.worker_identity_postgres import (
    LaborWorkerDeviceNotFound,
    LaborWorkerIdentityError,
    LaborWorkerIdentityInvalid,
    exchange_labor_worker_activation,
    issue_labor_worker_activation,
    labor_worker_identity_health,
    list_labor_worker_devices,
    resolve_labor_worker_token,
    revoke_labor_worker_device,
)
from .engine.labor.build_info import LaborBuildMonitor
from .engine.labor.worker_version import parse_stable_worker_version, worker_version_at_least
from .engine.labor.worker_archive import LaborWorkerArchiveError, build_worker_input_archive, merge_worker_result_archive

# --- FBU Performance engine imports ---
from .engine.fbu_performance.engines.base import (
    ADJUSTMENT_SPLIT_PATH,
    DISTRICT_MANAGER_FIXED_BASE_PATH,
    NINETY_SIX_HOUR_FIXED_BASE_PATH,
    STANDARD_PERFORMANCE_BASE_PATH,
)
from .engine.fbu_performance.engines.attendance import AttendanceProcessor
from .engine.fbu_performance.parser import (
    FBUPerformanceParser,
    build_hourly_rate_policy_data,
    normalize_shift_employee_id,
    update_hourly_rate_policy_data,
)
from .engine.fbu_performance.runs import (
    DEFAULT_FBU_REGION_CODE,
    FBU_AMERICAS_REGIONS,
    build_attendance_view_data,
    build_final_result_rows,
    build_results_view_data,
    has_attendance_employees,
    FBURosterStore,
    FBURun,
    FBURunManager,
)
from .engine.fbu_performance.runs import FBURuleListStore
from .engine.admin_store import (
    claim_admin_notification,
    count_audit_logs,
    create_session,
    delete_session,
    enqueue_admin_notification,
    get_admin_state,
    get_current_user,
    get_session_auth_context,
    get_session_user_id,
    init_admin_store,
    list_audit_logs,
    list_pending_admin_notifications,
    list_users,
    mark_admin_notification_failed,
    mark_admin_notification_sent,
    record_audit_event,
    set_feature_permission,
    set_module_enabled,
    set_module_role_access,
    set_user_roles,
    upsert_feishu_user,
)
from .permission_notifications import (
    admin_user_url,
    build_feedback_submitted_card,
    build_new_user_card,
    build_permission_change_card,
    build_permission_change_payload,
    build_workbench_announcement_card,
)
from .engine.feedback_store import (
    ANNOUNCEMENT_KIND_LABELS,
    CATEGORY_LABELS,
    create_announcement,
    create_feedback,
    get_feedback,
    get_feedback_attachment,
    list_announcements,
    list_feedback_for_admin,
    list_feedback_for_user,
)
from .engine.fbu_performance.persistent_storage import (
    FBU_RUN_SECTION_FIELDS,
    create_fbu_signed_upload,
    fbu_persistent_storage_enabled,
)
from .engine.fbu_performance.upload_jobs import FBUUploadJobStore
from .engine.social_insurance.router import router as social_insurance_router
from .engine.social_insurance.prefetch import (
    start_social_insurance_prefetch_scheduler,
    stop_social_insurance_prefetch_scheduler,
)


SUPPORTING_PDF_RE = re.compile(r"(?:supplement|support|time\s*card|timecard|detail|backup|appendix)", re.IGNORECASE)
NON_PAYABLE_PDF_TYPES = {"supporting", "attachment"}
LABOR_INVOICE_PAGE_ROLES = {"invoice_primary", "invoice_continuation", "invoice_total"}
LABOR_SUPPORTING_PAGE_ROLES = {"email_cover", "timecard_summary", "daily_detail", "supporting_attachment"}
LABOR_TELEMETRY_DIR = OUTPUT_DIR / "labor_telemetry"
LABOR_TELEMETRY_FILE = LABOR_TELEMETRY_DIR / "events.jsonl"
LABOR_TELEMETRY_SCHEMA_VERSION = 1
OVERSEAS_LABOR_MODULE_VERSION = "0.5-uat"
OVERSEAS_LABOR_API_CONTRACT_VERSION = 2
OVERSEAS_LABOR_REQUIRED_WORKER_VERSION = "0.3.12"
_LABOR_BUILD_MONITOR = LaborBuildMonitor(PROJECT_ROOT)
CURRENT_USER_CACHE_TTL_SECONDS = 60
_CURRENT_USER_CACHE: dict[str, tuple[float, str | None, dict[str, Any]]] = {}
PERMISSION_NOTIFICATION_CONFIG = {
    "admin_open_id": str(os.environ.get("FEISHU_PERMISSION_ADMIN_OPEN_ID") or "").strip(),
    "feedback_admin_name": str(os.environ.get("FEISHU_FEEDBACK_ADMIN_NAME") or "姚硕灿").strip(),
    "public_url": str(os.environ.get("SIGMA_WORKBENCH_PUBLIC_URL") or "").strip().rstrip("/"),
}
permission_notification_logger = logging.getLogger("bonus_platform.permission_notifications")
audit_logger = logging.getLogger("bonus_platform.audit")


def _clear_current_user_cache() -> None:
    _CURRENT_USER_CACHE.clear()


def _get_cached_current_user(
    user_id: str,
    permission_revision: str | None = None,
) -> dict[str, Any]:
    cached = _CURRENT_USER_CACHE.get(user_id)
    now = monotonic()
    if (
        cached
        and now - cached[0] < CURRENT_USER_CACHE_TTL_SECONDS
        and (permission_revision is None or cached[1] == permission_revision)
    ):
        return cached[2]
    current = get_current_user(user_id)
    _CURRENT_USER_CACHE[user_id] = (now, permission_revision, current)
    return current


def _labor_current_user_from_request(request: Request) -> dict[str, Any] | None:
    if getattr(request.state, "labor_auth_resolved", False):
        current = getattr(request.state, "labor_current_user", None)
        return current if isinstance(current, dict) else None

    request.state.labor_auth_resolved = True
    session_token = str(request.cookies.get(SESSION_COOKIE_NAME) or "").strip()
    if not session_token:
        return None
    try:
        user_id, permission_revision = get_session_auth_context(session_token)
        current = _get_cached_current_user(user_id, permission_revision)
    except KeyError:
        return None
    user = current.get("user") if isinstance(current, dict) else None
    if not isinstance(user, dict) or user.get("status") not in {"active", "pending"}:
        return None
    request.state.labor_current_user = current
    request.state.labor_user_id = user_id
    return current


def _non_payable_pdf_names(pdf_totals: list[dict]) -> set[str]:
    has_payable_invoice = any(
        float(total.get("total_amount") or 0) > 0
        and str(total.get("pdf_type") or "") not in NON_PAYABLE_PDF_TYPES
        for total in pdf_totals
    )
    if not has_payable_invoice:
        return set()
    return {
        str(total.get("source_file") or "")
        for total in pdf_totals
        if str(total.get("pdf_type") or "") in NON_PAYABLE_PDF_TYPES
        or (
            float(total.get("total_amount") or 0) == 0
            and SUPPORTING_PDF_RE.search(str(total.get("source_file") or ""))
        )
    }


def _labor_total_is_explicitly_non_payable(total: dict) -> bool:
    evidence_status = str(total.get("evidence_status") or "").strip().lower()
    pdf_type = str(total.get("pdf_type") or "").strip().lower()
    page_roles = {
        str(page.get("role") or page.get("page_role") or "").strip().lower()
        for page in (total.get("page_evidence") or [])
        if isinstance(page, dict)
    }
    if page_roles.intersection(LABOR_INVOICE_PAGE_ROLES):
        return False
    if page_roles.intersection(LABOR_SUPPORTING_PAGE_ROLES):
        return True
    return (
        total.get("non_payable") is True
        or evidence_status in {"supporting", "non_payable", "excluded"}
        or pdf_type in NON_PAYABLE_PDF_TYPES
    )


def _labor_stage2_target_pdf_names(warehouse_rows: list[dict]) -> set[str]:
    target_names: set[str] = set()
    for row in warehouse_rows:
        if row.get("reconciliationStatus") not in {"amount_difference", "needs_review"}:
            continue
        evidence_files = str(row.get("pdfEvidenceFile") or "")
        target_names.update(name.strip() for name in evidence_files.split(";") if name.strip())
    return target_names


def _labor_row_page_number(row: LaborLineItem) -> int | None:
    match = re.search(r"\d+", str(row.source_page_or_row or ""))
    return int(match.group()) if match else None


def _labor_invoice_evidence_pages(pdf_totals: list[dict]) -> dict[str, set[int]]:
    audited_pages_by_file: dict[str, set[int]] = {}
    for total in pdf_totals:
        page_evidence = total.get("page_evidence")
        if not isinstance(page_evidence, list) or not page_evidence:
            continue
        source_file = str(total.get("source_file") or "")
        if not source_file:
            continue
        authoritative_total_page = None
        if total.get("authoritative") is True and total.get("total_page") is not None:
            authoritative_total_page = int(total["total_page"])
        excluded_pages = {
            int(page)
            for page in (total.get("excluded_pages") or total.get("excludedPdfPages") or [])
            if page is not None
        }
        unscanned_pages = {
            int(page.get("page"))
            for page in page_evidence
            if isinstance(page, dict)
            and page.get("page") is not None
            and str(page.get("extraction_method") or page.get("extractionMethod") or "").strip().lower()
            == "not_scanned_after_authoritative_total"
        }
        scanned_pages = [
            page
            for page in page_evidence
            if isinstance(page, dict)
            and page.get("page") is not None
            and int(page.get("page")) not in unscanned_pages
        ]
        last_scanned_role = ""
        if scanned_pages:
            last_scanned = max(scanned_pages, key=lambda page: int(page.get("page") or 0))
            last_scanned_role = str(last_scanned.get("role") or last_scanned.get("page_role") or "").strip().lower()
        allowed_pages = {
            int(page.get("page"))
            for page in page_evidence
            if page.get("page") is not None
            and str(page.get("role") or page.get("page_role") or "") in LABOR_INVOICE_PAGE_ROLES
            and int(page.get("page")) not in excluded_pages
            and (
                authoritative_total_page is None
                or int(page.get("page")) <= authoritative_total_page
            )
        }
        # The quick total scan deliberately stops once it has an authoritative
        # total plus a following invoice-continuation page. Those later pages
        # were never classified as supporting material; they remain eligible
        # for the stricter employee-row extraction and amount-closure gates.
        if unscanned_pages and last_scanned_role in LABOR_INVOICE_PAGE_ROLES:
            allowed_pages.update(unscanned_pages)
        page_roles = {
            str(page.get("role") or page.get("page_role") or "").strip().lower()
            for page in page_evidence
            if isinstance(page, dict)
        }
        # An image-only invoice may have valid high-confidence employee rows
        # while the cheaper page-role classifier can only return ``unknown``.
        # Unknown is reviewable evidence, not proof that the page is supporting
        # material. Only install a restrictive page map when classification has
        # positively identified invoice or supporting roles.
        if (
            allowed_pages
            or page_roles.intersection(LABOR_SUPPORTING_PAGE_ROLES)
            or _labor_total_is_explicitly_non_payable(total)
        ):
            audited_pages_by_file[source_file] = allowed_pages
    return audited_pages_by_file


def _filter_labor_rows_to_invoice_evidence_pages(
    pdf_rows: list[LaborLineItem],
    pdf_totals: list[dict],
) -> list[LaborLineItem]:
    audited_pages_by_file = _labor_invoice_evidence_pages(pdf_totals)

    if not audited_pages_by_file:
        return list(pdf_rows)
    return [
        row
        for row in pdf_rows
        if row.source_file not in audited_pages_by_file
        or _labor_row_page_number(row) in audited_pages_by_file[row.source_file]
    ]


def _labor_authoritative_detail_totals(
    pdf_totals: list[dict],
    target_pdf_names: set[str],
) -> dict[str, dict]:
    totals_by_source: dict[str, dict] = {}
    for total in pdf_totals:
        source_file = str(total.get("source_file") or "").strip()
        amount = round(float(total.get("total_amount") or 0), 2)
        if (
            not source_file
            or source_file not in target_pdf_names
            or amount <= 0
            or total.get("authoritative") is not True
            or not str(total.get("warehouse_id") or "").strip()
            or total.get("warehouse_conflict")
        ):
            continue
        entry = totals_by_source.setdefault(
            source_file,
            {"expectedAmount": 0.0, "warehouseId": str(total.get("warehouse_id") or "")},
        )
        entry["expectedAmount"] = round(float(entry["expectedAmount"]) + amount, 2)
    return totals_by_source


def _labor_ocr_expected_totals(
    pdf_totals: list[dict],
    target_pdf_names: set[str],
    *,
    excel_rows: list[LaborLineItem] | None = None,
) -> dict[str, float]:
    expected: dict[str, float] = {}
    for total in pdf_totals:
        source_file = str(total.get("source_file") or "").strip()
        if not source_file or source_file not in target_pdf_names:
            continue
        amount = round(float(total.get("total_amount") or 0), 2)
        if amount > 0 and total.get("authoritative") is True:
            expected[source_file] = amount
            continue
        page_amounts = {
            round(float(page.get("total_amount") or 0), 2)
            for page in total.get("page_evidence", [])
            if str(page.get("role") or "") == "invoice_total"
            and float(page.get("role_confidence") or 0) >= 0.9
            and float(page.get("total_amount") or 0) > 0
        }
        if len(page_amounts) == 1:
            expected[source_file] = page_amounts.pop()
    if excel_rows:
        excel_by_warehouse: dict[str, float] = {}
        for row in excel_rows:
            warehouse_id = str(row.warehouse_id or "").strip()
            if warehouse_id:
                excel_by_warehouse[warehouse_id] = round(
                    excel_by_warehouse.get(warehouse_id, 0.0) + float(row.amount or 0),
                    2,
                )
        sources_by_warehouse: dict[str, list[str]] = {}
        for total in pdf_totals:
            source_file = str(total.get("source_file") or "").strip()
            warehouse_id = str(total.get("warehouse_id") or "").strip()
            if source_file in target_pdf_names and warehouse_id:
                sources_by_warehouse.setdefault(warehouse_id, []).append(source_file)
        for warehouse_id, source_files in sources_by_warehouse.items():
            excel_amount = excel_by_warehouse.get(warehouse_id, 0.0)
            if len(source_files) == 1 and excel_amount > 0:
                expected.setdefault(source_files[0], excel_amount)
    return expected


def _labor_should_run_auto_ocr(
    *,
    command: str,
    pdf_rows: list,
    target_pdf_names: set[str],
    reconciliation_pdf_totals: list[dict],
) -> bool:
    if not command.strip() or not target_pdf_names:
        return False
    authoritative_sources = {
        str(total.get("source_file") or "").strip()
        for total in reconciliation_pdf_totals
        if total.get("authoritative") is True and float(total.get("total_amount") or 0) > 0
    }
    return not pdf_rows or not target_pdf_names.issubset(authoritative_sources)


def _labor_should_retry_detail_totals(
    *,
    pdf_rows: list,
    mismatches: dict[str, dict],
    auto_ocr_candidate: dict,
) -> bool:
    return bool(pdf_rows and mismatches) and auto_ocr_candidate.get("runtimeStatus") != "completed"


def _run_labor_auto_ocr_candidate(
    run_id: str,
    pdf_paths: list[Path],
    excel_rows: list[LaborLineItem],
    reconciliation_pdf_totals: list[dict],
    *,
    supplier: str,
    period_start: str,
    period_end: str,
    currency: str,
    command: str,
    timeout_seconds: int,
    amount_tolerance: float,
    hours_tolerance: float,
    task_generation_id: str = "",
) -> dict:
    if not pdf_paths or not command.strip():
        return {}
    ocr_result = run_ocr_candidate_command(
        pdf_paths,
        command=command,
        supplier=supplier,
        period_start=period_start,
        period_end=period_end,
        currency=currency,
        timeout_seconds=timeout_seconds,
        progress_callback=_labor_ocr_progress_callback(run_id, task_generation_id=task_generation_id),
    )
    fallback_expected_totals = _labor_ocr_expected_totals(
        reconciliation_pdf_totals,
        {path.name for path in pdf_paths},
        excel_rows=excel_rows,
    )
    pdf_total_evidence: dict[str, dict] = {}
    for file_payload in ocr_result.get("files", []):
        source_file = str(file_payload.get("sourceFile") or "").strip()
        amount = round(float(file_payload.get("explicitTotalAmount") or 0), 2)
        evidence = file_payload.get("explicitTotalEvidence") or {}
        if not source_file or amount <= 0 or not isinstance(evidence, dict):
            continue
        pdf_total_evidence[source_file] = {
            "amount": amount,
            "page": int(evidence.get("page") or 0),
            "label": str(evidence.get("label") or "TOTAL"),
            "currency": str(evidence.get("currency") or currency or "").upper(),
            "evidenceText": str(evidence.get("evidenceText") or ""),
        }
    expected_totals = dict(fallback_expected_totals)
    for source_file, evidence in pdf_total_evidence.items():
        expected_totals.setdefault(source_file, evidence["amount"])
    candidate = evaluate_ocr_candidate_result(
        ocr_result,
        excel_rows,
        expected_totals,
        amount_tolerance=amount_tolerance,
        hours_tolerance=hours_tolerance,
    )
    candidate["runtimeStatus"] = ocr_result.get("status", "")
    candidate["runtimeFiles"] = ocr_result.get("files", [])
    candidate["runtimeError"] = ocr_result.get("error", "")
    candidate["runtimeCacheCleanup"] = ocr_result.get("cacheCleanup", {})
    candidate["pdfTotalEvidence"] = pdf_total_evidence
    return candidate


def _labor_apply_ocr_pdf_total_evidence(
    pdf_totals: list[dict],
    candidate: dict,
) -> list[dict]:
    evidence_by_source = candidate.get("pdfTotalEvidence") or {}
    if not isinstance(evidence_by_source, dict) or not evidence_by_source:
        return list(pdf_totals)
    merged: list[dict] = []
    for total in pdf_totals:
        source_file = str(total.get("source_file") or "").strip()
        evidence = evidence_by_source.get(source_file)
        if not isinstance(evidence, dict) or float(evidence.get("amount") or 0) <= 0:
            merged.append(dict(total))
            continue
        if (
            total.get("authoritative") is True
            and float(total.get("total_amount") or 0) > 0
        ):
            merged.append(dict(total))
            continue
        amount = round(float(evidence["amount"]), 2)
        page = int(evidence.get("page") or 0)
        updated = dict(total)
        updated.update(
            {
                "total_amount": amount,
                "authoritative": True,
                "evidence_status": "ocr_explicit_total",
                "total_label": str(evidence.get("label") or "TOTAL"),
                "total_page": page or None,
                "pdf_type": str(total.get("pdf_type") or "primary"),
            }
        )
        page_evidence = list(updated.get("page_evidence") or [])
        page_evidence.append(
            {
                "page": page,
                "role": "invoice_total",
                "role_confidence": 0.99,
                "total_amount": amount,
                "evidence_text": str(evidence.get("evidenceText") or ""),
                "source": "local_ocr_explicit_total",
            }
        )
        updated["page_evidence"] = page_evidence
        merged.append(updated)
    return merged


def _labor_can_apply_review_ocr_rows(candidate: dict, target_sources: set[str]) -> bool:
    if candidate.get("runtimeStatus") != "completed" or not target_sources:
        return False
    rows = candidate.get("rows") or []
    row_sources = {
        str(row.get("source_file") or "").strip()
        for row in rows
        if isinstance(row, dict) and str(row.get("source_file") or "").strip()
    }
    evidence_sources = {
        str(source_file).strip()
        for source_file, evidence in (candidate.get("pdfTotalEvidence") or {}).items()
        if str(source_file).strip()
        and isinstance(evidence, dict)
        and float(evidence.get("amount") or 0) > 0
    }
    runtime_by_source = {
        str(item.get("sourceFile") or "").strip(): item
        for item in (candidate.get("runtimeFiles") or [])
        if isinstance(item, dict) and str(item.get("sourceFile") or "").strip()
    }
    for source_file in target_sources:
        runtime = runtime_by_source.get(source_file) or {}
        page_count = int(runtime.get("pageCount") or 0)
        if (
            source_file not in row_sources
            or source_file not in evidence_sources
            or page_count <= 0
            or int(runtime.get("failedPageCount") or 0) > 0
            or int(runtime.get("successfulPageCount") or 0) != page_count
        ):
            return False
    return True


def _labor_detail_total_mismatches(
    pdf_rows: list[LaborLineItem],
    authoritative_totals: dict[str, dict],
    amount_tolerance: float,
) -> dict[str, dict]:
    actual_by_source = {source_file: 0.0 for source_file in authoritative_totals}
    for row in pdf_rows:
        if row.source_file in actual_by_source:
            actual_by_source[row.source_file] = round(
                actual_by_source[row.source_file] + float(row.amount or 0),
                2,
            )

    mismatches: dict[str, dict] = {}
    for source_file, expected in authoritative_totals.items():
        expected_amount = round(float(expected.get("expectedAmount") or 0), 2)
        actual_amount = round(actual_by_source.get(source_file, 0.0), 2)
        delta = round(expected_amount - actual_amount, 2)
        if not amount_within_tolerance(delta, amount_tolerance):
            mismatches[source_file] = {
                **expected,
                "actualAmount": actual_amount,
                "delta": delta,
            }
    return mismatches


def _merge_labor_detail_retry_rows(
    original_rows: list[LaborLineItem],
    retry_rows: list[LaborLineItem],
    authoritative_totals: dict[str, dict],
) -> tuple[list[LaborLineItem], set[str]]:
    original_by_source: dict[str, list[LaborLineItem]] = {
        source_file: [] for source_file in authoritative_totals
    }
    retry_by_source: dict[str, list[LaborLineItem]] = {
        source_file: [] for source_file in authoritative_totals
    }
    for row in original_rows:
        if row.source_file in original_by_source:
            original_by_source[row.source_file].append(row)
    for row in retry_rows:
        if row.source_file in retry_by_source:
            retry_by_source[row.source_file].append(row)

    replaced_sources: set[str] = set()
    for source_file, expected in authoritative_totals.items():
        expected_amount = round(float(expected.get("expectedAmount") or 0), 2)
        original_amount = round(sum(float(row.amount or 0) for row in original_by_source[source_file]), 2)
        retry_amount = round(sum(float(row.amount or 0) for row in retry_by_source[source_file]), 2)
        if retry_by_source[source_file] and abs(expected_amount - retry_amount) < abs(expected_amount - original_amount):
            replaced_sources.add(source_file)

    if not replaced_sources:
        return list(original_rows), set()
    merged_rows = [row for row in original_rows if row.source_file not in replaced_sources]
    merged_rows.extend(row for row in retry_rows if row.source_file in replaced_sources)
    return merged_rows, replaced_sources


def _warehouse_id_from_text_path(pdf_path: Path, diff_wh: list) -> bool:
    """检查 PDF 内容中的仓库号是否在差异仓库列表中。

    用于文件名无法提取仓库号时（如 US ELogistics 格式），从 PDF 内容中匹配。
    """
    try:
        from .engine.labor.extract import _extract_pdf_pages
        pages = _extract_pdf_pages([pdf_path], max_pages=1)
        if pages:
            wh = _warehouse_id_from_text(pages[0].get("text", ""))
            return wh in diff_wh
    except Exception:
        pass
    return False


def _labor_text_has_employee_detail_signal(text: str) -> bool:
    normalized = " ".join(str(text or "").split())
    if not normalized:
        return False
    if re.search(r"Job Site/Warehouse\s+Name\s+Pay Type", normalized, re.IGNORECASE) and re.search(
        r"Hourly-Reg", normalized, re.IGNORECASE
    ):
        return True
    if re.search(r"\bName\s+Pay Type\s+Rate\s+Hours\b", normalized, re.IGNORECASE) and re.search(
        r"Overtime\s*\(Hourly\)", normalized, re.IGNORECASE
    ):
        return True
    if re.search(r"\bAssociate\s+Base Rate\s+Bill Rate\s+OT Rate\s+Reg\.?\s*Time\b", normalized, re.IGNORECASE):
        return True
    return False


def _labor_pdf_has_employee_detail_signal(pdf_path: Path) -> bool:
    try:
        from .engine.labor.extract import _extract_pdf_pages

        pages = _extract_pdf_pages([pdf_path], max_pages=2)
        text = "\n".join(page.get("text", "") for page in pages)
        return _labor_text_has_employee_detail_signal(text)
    except Exception:
        return False
    return False
from .engine.labor.runs import (
    attach_labor_file,
    begin_labor_mapping_preflight,
    begin_labor_metadata_task,
    compare_and_update_labor_metadata,
    create_labor_run,
    get_labor_run_dir,
    list_labor_metadata,
    load_labor_metadata,
    labor_run_metadata_lock,
    safe_labor_filename,
    safe_labor_storage_filename,
    update_labor_metadata,
    update_labor_metadata_record_only,
    update_labor_metadata_for_mapping_preflight,
    update_labor_metadata_for_task,
)
from .engine.labor.jobs import (
    enqueue_labor_reconciliation_job,
    labor_worker_job_store_health as legacy_labor_worker_job_store_health,
    labor_worker_jobs_enabled,
)
from .engine.labor.blob_storage import (
    blob_get_bytes,
    blob_list_prefix,
    blob_put_bytes,
    canonicalize_labor_metadata_for_blob,
    create_labor_blob_presigned_url,
    labor_blob_signed_urls_enabled,
    labor_blob_storage_enabled,
    sync_labor_run_from_blob,
)
from .engine.labor.persistent_storage import (
    create_labor_supabase_signed_upload,
    labor_persistent_storage_enabled,
    labor_persistent_storage_health,
    labor_persistent_storage_info,
    labor_supabase_storage_enabled,
    sync_labor_run_from_persistent,
    sync_labor_run_to_persistent,
)
from .engine.labor.persistent_storage import (
    create_labor_supabase_signed_download,
    create_labor_supabase_signed_upload_for_object,
    delete_labor_run_from_persistent,
    get_labor_supabase_private_object,
    labor_p1_object_key,
    labor_persistent_environment,
    labor_storage_backend,
    labor_supabase_object_metadata,
    persist_labor_private_output,
    put_labor_supabase_private_object,
)
from .engine.labor.audit import append_labor_audit_event, read_labor_audit_events
from .engine.labor.hardening import LaborHardeningPolicy, LaborResourceLimitError, LaborTaskLimiter, labor_storage_info
_LABOR_TASK_LIMITER = LaborTaskLimiter()
from .engine.labor.lifecycle import cleanup_expired_labor_runs, delete_labor_run_directory, labor_run_is_active
from .engine.labor.ocr_worker_cache import cleanup_ocr_cache, labor_ocr_cache_dir
from .engine.labor.ocr_targeted_retry import build_targeted_ocr_retry_plan, merge_targeted_ocr_retry_rows
from .engine.labor.workbook import list_workbook_sheets, parse_reocr_candidate_rows, read_workbook_rows, suggest_mapping, summarize_otws_costs
from .engine.rules import load_rulebook
from .engine.runs import (
    attach_file_record,
    create_run_dir,
    get_run_dir,
    list_run_metadata,
    load_metadata,
    new_run_id,
    rule_info,
    run_file_url,
    save_metadata,
    update_metadata,
)
from .engine.table_data import build_final_table_data, build_table_data, load_table_data, merge_diff_rows, save_table_data
from .engine.workbook_io import build_final_workbook, build_pending_workbook, build_result_workbook, read_import_rows
from .engine.overseas_payroll.router import page_router as overseas_payroll_page_router, router as overseas_payroll_router
from .engine.overseas_payroll.tasks import (
    finalize_output as finalize_overseas_payroll_output,
    load_task as load_overseas_payroll_task,
    local_file_path as overseas_payroll_local_file_path,
    prepare_output as prepare_overseas_payroll_output,
    store_local_file as store_local_overseas_payroll_file,
    task_input_downloads as overseas_payroll_task_input_downloads,
    update_task as update_overseas_payroll_task,
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    ensure_data_files()
    if not os.environ.get("VERCEL"):
        init_admin_store()
    if not _is_vercel_runtime():
        _recover_stuck_labor_runs()
    _cleanup_expired_labor_data()
    social_insurance_scheduler_allowed = not _is_vercel_runtime()
    if social_insurance_scheduler_allowed:
        start_social_insurance_prefetch_scheduler()
    try:
        yield
    finally:
        if social_insurance_scheduler_allowed:
            stop_social_insurance_prefetch_scheduler()


app = FastAPI(title="招聘奖金与内推奖金核算平台", lifespan=lifespan)
app.include_router(social_insurance_router)
app.include_router(overseas_payroll_router)
app.include_router(overseas_payroll_page_router)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "null",
        "http://127.0.0.1:8006",
        "http://localhost:8006",
    ],
    allow_methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
)
STATIC_DIR = Path(__file__).resolve().parent / "static"

SAFE_ID_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_-]{0,63}$")
SESSION_COOKIE_NAME = "sigma_session"
FEISHU_STATE_COOKIE_NAME = "sigma_feishu_state"
FEISHU_API_BASE_URL = "https://open.feishu.cn/open-apis"

_BUSINESS_AUDIT_MODULE_NAMES = {
    "recruitment": "全球招聘奖金核算",
    "employee": "中国区正式工薪酬核算",
    "domestic": "中国区外包工薪酬核算",
    "fbu": "FBU美洲绩效奖金核算",
    "overseas": "海外劳务报账核对",
}
_BUSINESS_AUDIT_PAGE_MODULES = {
    "/recruitment.html": "recruitment",
    "/china-employee-payroll.html": "employee",
    "/employee-payroll.html": "employee",
    "/domestic-labor.html": "domestic",
    "/labor.html": "domestic",
    "/fbu-performance.html": "fbu",
    "/overseas-labor.html": "overseas",
}
_BUSINESS_AUDIT_ACTION_LABELS = {
    "module_open": "进入模块",
    "run_create": "新建批次",
    "data_import": "导入材料",
    "calculation_submit": "提交核算",
    "review_confirm": "复核确认",
    "result_export": "导出结果",
    "run_delete": "删除批次",
    "config_update": "更新配置",
    "business_operation": "业务操作",
}


def _business_audit_module_id(path: str) -> str | None:
    if path == "/api/calculate" or path.startswith("/api/runs"):
        return "recruitment"
    prefixes = (
        ("/api/china-employee-payroll/", "employee"),
        ("/api/domestic-labor/", "domestic"),
        ("/api/fbu-performance/", "fbu"),
        ("/api/labor/", "overseas"),
    )
    return next((module_id for prefix, module_id in prefixes if path.startswith(prefix)), None)


def _business_audit_run_id(path: str) -> str:
    match = re.search(r"/runs/([^/]+)", path)
    if match and match.group(1) not in {"calculate", "bulk-delete", "direct-upload-plan"}:
        return match.group(1)[:128]
    meal_match = re.search(r"/meal-allowance/([^/]+)/export$", path)
    return meal_match.group(1)[:128] if meal_match else ""


def _business_activity_audit_event(method: str, path: str, status_code: int) -> dict[str, str] | None:
    """Classify successful, user-meaningful activity without inspecting request data."""
    method = method.upper()
    path = path.rstrip("/") or "/"
    if not 200 <= status_code < 400:
        return None

    page_module_id = _BUSINESS_AUDIT_PAGE_MODULES.get(path)
    if method == "GET" and page_module_id:
        return {
            "action": "module_open",
            "target_type": "module",
            "target_id": page_module_id,
            "detail": f"进入{_BUSINESS_AUDIT_MODULE_NAMES[page_module_id]}",
        }

    module_id = _business_audit_module_id(path)
    if not module_id:
        return None
    if path.startswith("/api/labor/worker/") or path in {
        "/api/labor/telemetry",
        "/api/labor/maintenance/cleanup",
    }:
        return None

    run_id = _business_audit_run_id(path)
    run_suffix = f" · 批次 ID：{run_id}" if run_id else ""

    if method == "GET":
        is_template = "/templates/" in path
        is_export = path.endswith(("/export", "/export-excel")) or "/download/" in path
        if not is_template and is_export:
            return {
                "action": "result_export",
                "target_type": "module",
                "target_id": module_id,
                "detail": f"导出结果{run_suffix}",
            }
        return None

    if method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None
    if method == "POST" and (
        path.endswith("/upload-intents")
        or path.endswith("/direct-upload-plan")
        or path.endswith("/uploads/plan")
        or path.endswith("/attendance-direct-upload-plan")
    ):
        return None

    if method == "DELETE" or path.endswith("/bulk-delete"):
        action, detail = "run_delete", "删除业务批次"
    elif path in {
        "/api/domestic-labor/runs",
        "/api/fbu-performance/runs",
        "/api/labor/runs",
        "/api/labor/material-runs",
    }:
        action, detail = "run_create", "新建核算批次"
    elif (
        any(marker in path for marker in ("/calculate", "/compare", "extract-and-compare"))
        or path in {"/api/calculate", "/api/china-employee-payroll/meal-allowance"}
    ):
        action = "calculation_submit"
        detail = "提交核对" if module_id == "overseas" else "提交核算"
    elif any(marker in path for marker in ("/import", "/files", "direct-upload-complete", "upload-intents")):
        action, detail = "data_import", "导入或上传业务材料"
    elif any(marker in path for marker in ("/confirm", "/finalize", "/business-review", "/replay", "/apply", "/rollback")):
        action, detail = "review_confirm", "提交复核或确认"
    elif any(marker in path for marker in ("/mapping", "/roster", "/rule-list", "policies", "adjustments", "supplement")):
        action, detail = "config_update", "更新核算配置"
    else:
        action, detail = "business_operation", "执行业务操作"
    return {
        "action": action,
        "target_type": "module",
        "target_id": module_id,
        "detail": f"{detail}{run_suffix}",
    }


def _business_activity_audit_outcome_event(method: str, path: str, status_code: int) -> dict[str, str] | None:
    """Return a success, permission-denial, or failure event for a selected business action."""
    candidate = _business_activity_audit_event(method, path, 200)
    if candidate is None:
        return None
    success_event = _business_activity_audit_event(method, path, status_code)
    if success_event is not None:
        return success_event
    if status_code < 400:
        return None
    if status_code in {401, 403}:
        detail = "登录会话失效" if status_code == 401 else "权限拒绝"
        action = "access_denied"
    else:
        detail = f"{_BUSINESS_AUDIT_ACTION_LABELS.get(candidate['action'], '业务操作')}失败"
        action = "operation_failed"
    run_id = _business_audit_run_id(path)
    run_suffix = f" · 批次 ID：{run_id}" if run_id else ""
    return {
        "action": action,
        "target_type": candidate["target_type"],
        "target_id": candidate["target_id"],
        "detail": f"{detail}{run_suffix} · HTTP {status_code}",
    }


def _validate_safe_id(value: str, field_name: str = "id") -> str:
    if not SAFE_ID_RE.fullmatch(value or ""):
        raise HTTPException(status_code=400, detail=f"无效的 {field_name}。")
    return value


def _current_user_id(sigma_session: Optional[str] = Cookie(default=None)) -> str:
    if not sigma_session:
        raise HTTPException(status_code=401, detail="未登录。")
    try:
        return get_session_user_id(sigma_session)
    except KeyError as exc:
        raise HTTPException(status_code=401, detail="登录已失效。") from exc


def _require_admin_user(actor_user_id: str = Depends(_current_user_id)) -> str:
    try:
        current = get_current_user(actor_user_id)
    except KeyError as exc:
        raise HTTPException(status_code=401, detail="未识别的用户身份。") from exc
    if not any(role["id"] == "admin" for role in current["roles"]):
        raise HTTPException(status_code=403, detail="需要系统管理员权限。")
    return actor_user_id


def _user_can_enter_module(user_id: str, module_id: str) -> bool:
    try:
        current = _get_cached_current_user(user_id)
    except KeyError:
        return False
    return any(
        module.get("id") == module_id and module.get("enabled") and module.get("canEnter")
        for module in current.get("modules", [])
    )


def _payload_bool(payload: dict, key: str) -> bool:
    value = payload.get(key)
    if not isinstance(value, bool):
        raise HTTPException(status_code=400, detail=f"{key} 必须是 boolean。")
    return value


def _feishu_api_error(prefix: str, payload: dict[str, Any]) -> HTTPException:
    code = payload.get("code")
    message = payload.get("msg") or payload.get("message") or "unknown_error"
    return HTTPException(status_code=502, detail=f"{prefix}失败：{code} {message}")


def _feishu_post_json(path: str, payload: dict[str, Any], token: str | None = None) -> dict[str, Any]:
    headers = {"Content-Type": "application/json; charset=utf-8"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        response = httpx.post(f"{FEISHU_API_BASE_URL}{path}", json=payload, headers=headers, timeout=10)
        response.raise_for_status()
        return response.json()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail="飞书接口请求失败。") from exc
    except ValueError as exc:
        raise HTTPException(status_code=502, detail="飞书接口返回格式异常。") from exc


def _feishu_get_json(path: str, token: str) -> dict[str, Any]:
    try:
        response = httpx.get(
            f"{FEISHU_API_BASE_URL}{path}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=10,
        )
        response.raise_for_status()
        return response.json()
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=502, detail="飞书用户信息请求失败。") from exc
    except ValueError as exc:
        raise HTTPException(status_code=502, detail="飞书用户信息返回格式异常。") from exc


def _get_feishu_app_access_token() -> str:
    payload = _feishu_post_json(
        "/auth/v3/app_access_token/internal",
        {"app_id": AUTH_CONFIG["feishu_app_id"], "app_secret": AUTH_CONFIG["feishu_app_secret"]},
    )
    if payload.get("code") != 0:
        raise _feishu_api_error("获取飞书 app_access_token", payload)
    token = str(payload.get("app_access_token") or "")
    if not token:
        raise HTTPException(status_code=502, detail="飞书 app_access_token 为空。")
    return token


def _get_feishu_tenant_access_token() -> str:
    payload = _feishu_post_json(
        "/auth/v3/tenant_access_token/internal",
        {"app_id": AUTH_CONFIG["feishu_app_id"], "app_secret": AUTH_CONFIG["feishu_app_secret"]},
    )
    if payload.get("code") != 0:
        raise _feishu_api_error("获取飞书 tenant_access_token", payload)
    token = str(payload.get("tenant_access_token") or "")
    if not token:
        raise HTTPException(status_code=502, detail="飞书 tenant_access_token 为空。")
    return token


def _permission_notification_public_url() -> str:
    configured = str(PERMISSION_NOTIFICATION_CONFIG.get("public_url") or "").strip().rstrip("/")
    if configured:
        return configured
    vercel_host = str(
        os.environ.get("VERCEL_PROJECT_PRODUCTION_URL")
        or os.environ.get("VERCEL_URL")
        or ""
    ).strip().strip("/")
    return f"https://{vercel_host}" if vercel_host else "http://localhost:8000"


def _send_feishu_permission_card(notification: dict[str, Any]) -> str:
    public_url = _permission_notification_public_url()
    kind = str(notification.get("kind") or "")
    payload = dict(notification.get("payload") or {})
    if kind == "new_user_pending":
        card = build_new_user_card(
            payload,
            admin_user_url(public_url, str(payload.get("userId") or "")),
        )
    elif kind == "permission_changed":
        card = build_permission_change_card(payload, f"{public_url}/")
    elif kind == "feedback_submitted":
        feedback_id = str(payload.get("feedbackId") or "")
        user_open_id = str(payload.get("userOpenId") or "")
        card = build_feedback_submitted_card(
            payload,
            f"{public_url}/admin.html?feedback={quote(feedback_id, safe='')}#feedbackCenter",
            f"https://applink.feishu.cn/client/chat/open?openId={quote(user_open_id, safe='')}",
        )
    elif kind == "workbench_announcement":
        announcement_id = str(payload.get("announcementId") or "")
        card = build_workbench_announcement_card(
            payload,
            f"{public_url}/?announcement={quote(announcement_id, safe='')}",
        )
    else:
        raise ValueError("unsupported_permission_notification_kind")

    result = _feishu_post_json(
        "/im/v1/messages?receive_id_type=open_id",
        {
            "receive_id": str(notification.get("recipientOpenId") or ""),
            "msg_type": "interactive",
            "content": json.dumps(card, ensure_ascii=False, separators=(",", ":")),
        },
        token=_get_feishu_tenant_access_token(),
    )
    if result.get("code") != 0:
        raise _feishu_api_error("发送飞书权限卡片", result)
    data = result.get("data") if isinstance(result.get("data"), dict) else {}
    message_id = str(data.get("message_id") or "").strip()
    if not message_id:
        raise RuntimeError("feishu_permission_card_message_id_missing")
    return message_id


def _dispatch_pending_permission_notifications() -> None:
    """Best-effort delivery; queued rows preserve retries without blocking auth or grants."""
    try:
        notifications = list_pending_admin_notifications(limit=50)
    except Exception as exc:  # noqa: BLE001 - notification failures must not break core flows.
        permission_notification_logger.warning("permission notification outbox unavailable: %s", type(exc).__name__)
        return
    for notification in notifications:
        if not claim_admin_notification(notification["id"]):
            continue
        try:
            message_id = _send_feishu_permission_card(notification)
            mark_admin_notification_sent(notification["id"], message_id)
        except Exception as exc:  # noqa: BLE001 - external messaging is isolated from auth/permissions.
            try:
                mark_admin_notification_failed(notification["id"], str(exc))
            except Exception:
                permission_notification_logger.warning(
                    "permission notification delivery and status update failed: %s",
                    type(exc).__name__,
                )


def _queue_new_user_permission_notification(user: dict[str, Any]) -> dict[str, Any] | None:
    recipient_open_id = str(PERMISSION_NOTIFICATION_CONFIG.get("admin_open_id") or "").strip()
    if (
        not recipient_open_id
        or not user.get("_created")
        or user.get("status") != "pending"
        or list(user.get("roleIds") or [])
    ):
        return None
    try:
        return enqueue_admin_notification(
            event_key=f"new-user:{user['id']}",
            kind="new_user_pending",
            recipient_open_id=recipient_open_id,
            payload={
                "userId": user["id"],
                "userName": user.get("name") or "未命名用户",
                "name": user.get("name") or "未命名用户",
                "email": user.get("email") or "",
                "createdAt": utcnow_naive().replace(microsecond=0).isoformat() + "Z",
            },
        )
    except Exception as exc:  # noqa: BLE001 - login must succeed if notification enqueue fails.
        permission_notification_logger.warning("new user notification enqueue failed: %s", type(exc).__name__)
        return None


def _queue_permission_change_notification(
    before: dict[str, Any],
    after: dict[str, Any],
    actor: dict[str, Any],
) -> dict[str, Any] | None:
    payload = build_permission_change_payload(before, after, actor)
    if not payload["recipientOpenId"]:
        return None
    changed_fields = (
        payload["addedRoles"],
        payload["removedRoles"],
        payload["addedModules"],
        payload["removedModules"],
    )
    if not any(changed_fields):
        return None
    try:
        return enqueue_admin_notification(
            event_key=f"permission-change:{payload['userId']}:{uuid4().hex}",
            kind="permission_changed",
            recipient_open_id=payload["recipientOpenId"],
            payload=payload,
        )
    except Exception as exc:  # noqa: BLE001 - grants must commit even if notification enqueue fails.
        permission_notification_logger.warning("permission change notification enqueue failed: %s", type(exc).__name__)
        return None


def _feedback_notification_recipient_open_id() -> str:
    """Resolve one active system administrator by exact display name."""
    target_name = str(PERMISSION_NOTIFICATION_CONFIG.get("feedback_admin_name") or "").strip()
    if not target_name:
        return ""
    matches = [
        user
        for user in list_users()
        if str(user.get("name") or "").strip() == target_name
        and user.get("status") == "active"
        and "admin" in list(user.get("roleIds") or [])
        and str(user.get("feishuOpenId") or "").strip()
    ]
    if len(matches) != 1:
        permission_notification_logger.warning(
            "feedback notification recipient resolution failed name=%s matches=%s",
            target_name,
            len(matches),
        )
        return ""
    return str(matches[0]["feishuOpenId"]).strip()


def _queue_feedback_notification(feedback: dict[str, Any]) -> dict[str, Any] | None:
    recipient_open_id = _feedback_notification_recipient_open_id()
    if not recipient_open_id:
        return None
    try:
        return enqueue_admin_notification(
            event_key=f"feedback:{feedback['id']}",
            kind="feedback_submitted",
            recipient_open_id=recipient_open_id,
            payload={
                "feedbackId": feedback["id"],
                "userName": feedback.get("userName") or "业务用户",
                "userOpenId": feedback.get("userOpenId") or "",
                "categoryLabel": feedback.get("categoryLabel") or "平台反馈",
                "moduleName": feedback.get("moduleName") or "HRAS 全球薪酬核算工作台",
                "description": feedback.get("description") or "",
                "attachmentCount": feedback.get("attachmentCount") or 0,
                "createdAt": feedback.get("createdAt") or "",
            },
        )
    except Exception as exc:  # noqa: BLE001 - feedback must persist even if notification enqueue fails.
        permission_notification_logger.warning("feedback notification enqueue failed: %s", type(exc).__name__)
        return None


def _queue_announcement_notifications(announcement: dict[str, Any]) -> int:
    queued = 0
    for user in list_users():
        recipient_open_id = str(user.get("feishuOpenId") or "").strip()
        if user.get("status") != "active" or not recipient_open_id:
            continue
        try:
            enqueue_admin_notification(
                event_key=f"announcement:{announcement['id']}:{user['id']}",
                kind="workbench_announcement",
                recipient_open_id=recipient_open_id,
                payload={
                    "announcementId": announcement["id"],
                    "kind": announcement.get("kind") or "feature",
                    "kindLabel": announcement.get("kindLabel") or "功能更新",
                    "title": announcement.get("title") or "HRAS 工作台更新",
                    "content": announcement.get("content") or "",
                    "moduleName": announcement.get("moduleName") or "HRAS 全球薪酬核算工作台",
                    "publishedAt": announcement.get("publishedAt") or "",
                },
            )
            queued += 1
        except Exception as exc:  # noqa: BLE001 - one recipient cannot block the release notice.
            permission_notification_logger.warning(
                "announcement notification enqueue failed: user=%s error=%s",
                user.get("id"),
                type(exc).__name__,
            )
    return queued


def _get_feishu_user_access_token(code: str, app_access_token: str) -> dict[str, Any]:
    payload = _feishu_post_json(
        "/authen/v1/access_token",
        {"grant_type": "authorization_code", "code": code},
        token=app_access_token,
    )
    if payload.get("code") != 0:
        raise _feishu_api_error("获取飞书 user_access_token", payload)
    data = payload.get("data")
    if not isinstance(data, dict) or not data.get("access_token"):
        raise HTTPException(status_code=502, detail="飞书 user_access_token 返回为空。")
    return data


def _get_feishu_contact_user(open_id: str, tenant_access_token: str) -> dict[str, Any]:
    payload = _feishu_get_json(
        f"/contact/v3/users/{quote(open_id, safe='')}?user_id_type=open_id",
        tenant_access_token,
    )
    if payload.get("code") != 0:
        raise _feishu_api_error("获取飞书通讯录用户", payload)
    data = payload.get("data")
    if not isinstance(data, dict):
        return {}
    user = data.get("user")
    return user if isinstance(user, dict) else data


def _get_feishu_user_info(user_access_token: str) -> dict[str, Any]:
    payload = _feishu_get_json("/authen/v1/user_info", user_access_token)
    if payload.get("code") != 0:
        raise _feishu_api_error("获取飞书用户信息", payload)
    data = payload.get("data")
    if not isinstance(data, dict):
        raise HTTPException(status_code=502, detail="飞书用户信息为空。")
    return data


def _first_nonempty_string(*values: Any) -> str | None:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return None


def _extract_feishu_avatar_url(payload: dict[str, Any]) -> str | None:
    avatar = payload.get("avatar")
    avatar_fields: dict[str, Any] = avatar if isinstance(avatar, dict) else {}
    return _first_nonempty_string(
        payload.get("avatar_url"),
        payload.get("avatarUrl"),
        payload.get("avatar_thumb"),
        payload.get("avatarThumb"),
        payload.get("avatar_middle"),
        payload.get("avatarMiddle"),
        payload.get("avatar_big"),
        payload.get("avatarBig"),
        avatar_fields.get("avatar_origin"),
        avatar_fields.get("avatar_640"),
        avatar_fields.get("avatar_240"),
        avatar_fields.get("avatar_72"),
    )


def _feishu_identity_from_payloads(token_data: dict[str, Any], user_info: dict[str, Any]) -> dict[str, str | None]:
    merged = {**token_data, **user_info}
    open_id = str(merged.get("open_id") or merged.get("openId") or "").strip()
    if not open_id:
        raise HTTPException(status_code=502, detail="飞书用户 open_id 为空。")
    avatar_url = _extract_feishu_avatar_url(merged)
    return {
        "feishu_open_id": open_id,
        "feishu_union_id": str(merged.get("union_id") or merged.get("unionId") or "").strip() or None,
        "email": str(merged.get("email") or merged.get("enterprise_email") or "").strip() or None,
        "avatar_url": avatar_url,
        "name": str(merged.get("name") or merged.get("en_name") or merged.get("nickname") or open_id).strip(),
    }


def _overseas_labor_access_config() -> dict:
    access = os.environ.get("SIGMA_OVERSEAS_LABOR_ACCESS", "uat").strip().lower() or "uat"
    upload_policy = _labor_hardening_policy()
    disabled_values = {"disabled", "off", "false", "0", "deny", "closed"}
    uat_values = {"uat", "uat_trial", "trial"}
    enabled_values = {"enabled", "on", "true", "1", "production", "prod", "full", "online"}
    can_use = access not in disabled_values
    access_mode = "disabled" if not can_use else ("uat_trial" if access in uat_values else ("production" if access in enabled_values else access))
    stage = "UAT试用版" if access_mode == "uat_trial" else "生产试运行"
    message = (
        "仅限薪酬/报账核对 UAT 试用，结果需人工复核后再对外结论。"
        if access_mode == "uat_trial"
        else "已启用持久化上传与后台异步核对，结果仍需人工复核后再对外结论。"
    )
    allowed_roles = [
        role.strip()
        for role in os.environ.get("SIGMA_OVERSEAS_LABOR_UAT_ROLES", "Payroll Admin,Compensation UAT").split(",")
        if role.strip()
    ]
    build = _labor_build_snapshot()
    runtime_current = build["status"] == "current"
    request_scoped_runtime = _uses_request_scoped_labor_runtime()
    personal_worker_enabled = _uses_personal_labor_worker()
    formal_task_execution_ready = not request_scoped_runtime or personal_worker_enabled
    formal_task_can_queue = runtime_current and formal_task_execution_ready
    if not runtime_current:
        formal_task_reason_code = "LABOR_SERVICE_RESTART_REQUIRED"
        formal_task_message = "服务版本需要重启或重新部署，暂不能创建正式核对任务。"
    elif not formal_task_execution_ready:
        formal_task_reason_code = "LABOR_PERSONAL_WORKER_REQUIRED"
        formal_task_message = "当前请求作用域环境未启用本人核对助手，暂不能创建正式核对任务。"
    else:
        formal_task_reason_code = ""
        formal_task_message = (
            "当前 UAT 通过本人核对助手执行正式核对任务。"
            if personal_worker_enabled
            else "当前持久化环境可执行正式核对任务。"
        )
    public_build = _labor_public_build_snapshot(build)
    return {
        "module": "overseas_labor_invoice_audit",
        "stage": stage,
        "version": OVERSEAS_LABOR_MODULE_VERSION,
        "access": access_mode,
        "apiContractVersion": OVERSEAS_LABOR_API_CONTRACT_VERSION,
        "buildId": public_build["buildId"],
        "build": public_build,
        "runtimeGate": {
            "canStartFormalTask": runtime_current,
            "runtimeSourceCurrent": runtime_current,
            "reasonCode": "" if runtime_current else "LABOR_SERVICE_RESTART_REQUIRED",
            "message": (
                "当前服务代码与启动快照一致。"
                if runtime_current
                else "服务启动后海外劳务代码已变化，请重启或重新部署后再开始正式操作。"
            ),
        },
        "formalTaskGate": {
            "canQueue": formal_task_can_queue,
            "executionMode": (
                "personal_worker"
                if personal_worker_enabled
                else "blocked"
                if request_scoped_runtime
                else "local"
            ),
            "reasonCode": formal_task_reason_code,
            "message": formal_task_message,
        },
        "p1": {
            "required": _labor_p1_required(),
            "uploadMode": "signed_private_direct" if _labor_p1_required() else "server_multipart",
            "legacyMultipartAllowed": not _labor_p1_required(),
        },
        "uploadLimits": {
            "maxPdfFiles": upload_policy.max_pdf_files,
            "maxWorkbookFiles": upload_policy.max_workbook_files,
        },
        "reconciliationScope": "employee_detail_required",
        "manualReviewRequired": True,
        "directPaymentAllowed": False,
        "canUse": can_use,
        "allowedRoles": allowed_roles,
        "message": message if can_use else "海外劳务报账核对模块暂未开放。",
    }


def _labor_build_snapshot() -> dict:
    snapshot = _LABOR_BUILD_MONITOR.snapshot(
        env=os.environ,
        module_version=OVERSEAS_LABOR_MODULE_VERSION,
        api_contract_version=OVERSEAS_LABOR_API_CONTRACT_VERSION,
        required_worker_version=_labor_required_worker_version(),
    )
    if not _labor_supplier_profiles_in_release_bundle():
        snapshot = dict(snapshot)
        snapshot["status"] = "unverified"
        snapshot["missingSentinels"] = sorted(
            {
                *(snapshot.get("missingSentinels") or []),
                "supplier_profile_release_bundle",
            }
        )
    return snapshot


def _labor_supplier_profiles_in_release_bundle() -> bool:
    configured = str(AI_CONFIG.get("supplier_profiles_path") or "").strip()
    if not configured:
        return False
    candidate = Path(configured).expanduser()
    if not candidate.is_absolute():
        candidate = PROJECT_ROOT / candidate
    if not candidate.exists():
        return False
    bundle = (PROJECT_ROOT / "data" / "supplier_profiles").resolve()
    try:
        candidate.resolve().relative_to(bundle)
    except (OSError, ValueError):
        return False
    return True


def _labor_public_build_snapshot(build: dict) -> dict:
    return {
        "schemaVersion": int(build.get("schemaVersion") or 0),
        "moduleVersion": str(build.get("moduleVersion") or OVERSEAS_LABOR_MODULE_VERSION),
        "apiContractVersion": int(build.get("apiContractVersion") or OVERSEAS_LABOR_API_CONTRACT_VERSION),
        "buildId": str(build.get("buildId") or "")[:64],
        "status": str(build.get("status") or "unverified"),
        "requiredWorkerVersion": str(build.get("requiredWorkerVersion") or _labor_required_worker_version()),
    }


def _labor_required_worker_version() -> str:
    candidates = [OVERSEAS_LABOR_REQUIRED_WORKER_VERSION]
    configured = str(os.environ.get("SIGMA_LABOR_REQUIRED_WORKER_VERSION") or "").strip()
    if configured:
        candidates.append(configured)
    persisted = _load_persisted_labor_worker_release_manifest()
    if isinstance(persisted, dict):
        candidates.append(str(persisted.get("requiredWorkerVersion") or ""))
    valid = []
    for candidate in candidates:
        try:
            valid.append((parse_stable_worker_version(candidate), candidate))
        except ValueError:
            continue
    return max(valid)[1] if valid else OVERSEAS_LABOR_REQUIRED_WORKER_VERSION


def _labor_requires_employee_detail(metadata: dict) -> bool:
    scope = str(metadata.get("reconciliationScope") or "").strip().lower()
    return not (scope == "total_only_diagnostic" and metadata.get("diagnosticOnly") is True)


def _labor_uat_review_state(conclusion: dict) -> dict:
    level = str(conclusion.get("conclusionLevel") or "").strip().lower()
    machine_status = "passed" if conclusion.get("canRelease") is True and level == "pass" else "blocked" if level == "critical" else "needs_review"
    return {
        "machineCheckStatus": machine_status,
        "businessReviewStatus": "pending",
        "manualReviewRequired": True,
        "directPaymentAllowed": False,
        "requiresHumanReview": True,
    }


def _labor_apply_reconciliation_diagnostics_gate(batch_guard: dict, diagnostics: dict) -> dict:
    gated = dict(batch_guard)
    if gated.get("allowReleasableReport") is not True:
        return gated
    level = str((diagnostics or {}).get("level") or "").strip().lower()
    if level == "ok":
        return gated
    if level not in {"warning", "critical"}:
        level = "critical"
    gated.update(
        {
            "status": (
                "reconciliation_diagnostics_blocked"
                if level == "critical"
                else "reconciliation_diagnostics_review"
            ),
            "message": str((diagnostics or {}).get("message") or "核对诊断证据不完整，不能机器通过。"),
            "allowReleasableReport": False,
        }
    )
    return gated


def _labor_hardening_policy() -> LaborHardeningPolicy:
    return LaborHardeningPolicy.from_env()


def _labor_audit_path() -> Path:
    configured = str(os.environ.get("LABOR_AUDIT_PATH") or "").strip()
    return Path(configured).expanduser() if configured else OUTPUT_DIR / "labor_audit" / "events.jsonl"


def _cleanup_expired_labor_data() -> dict:
    policy = _labor_hardening_policy()
    run_cleanup = cleanup_expired_labor_runs(
        LABOR_RUNS_DIR,
        retention_days=policy.run_retention_days,
        audit_path=_labor_audit_path(),
        delete_persistent=delete_labor_run_from_persistent,
    )
    cache_cleanup = cleanup_ocr_cache(
        labor_ocr_cache_dir(),
        retention_days=policy.ocr_cache_retention_days,
        max_bytes=policy.ocr_cache_max_bytes,
    )
    append_labor_audit_event(
        _labor_audit_path(),
        action="cache_cleanup",
        outcome="success",
        reason_code="scheduled_cleanup",
        details={
            "deletedCacheEntryCount": (
                int(cache_cleanup.get("expiredEntryCount") or 0)
                + int(cache_cleanup.get("corruptEntryCount") or 0)
                + int(cache_cleanup.get("capacityEvictedEntryCount") or 0)
            ),
            "reclaimedBytes": int(cache_cleanup.get("reclaimedBytes") or 0),
            "remainingBytes": int(cache_cleanup.get("remainingBytes") or 0),
        },
    )
    return {"runs": run_cleanup, "ocrCache": cache_cleanup}


def _env_flag(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _hide_developing_modules() -> bool:
    return _env_flag("SIGMA_HIDE_DEVELOPING_MODULES", False)


def _workbench_access_config() -> dict:
    return {
        "hideDevelopingModules": _hide_developing_modules(),
        "blockedModules": [],
        "authRequired": labor_auth_required(),
    }


def _is_vercel_runtime() -> bool:
    return bool(os.environ.get("VERCEL") or os.environ.get("VERCEL_ENV") or os.environ.get("VERCEL_URL"))


def _mock_auth_enabled() -> bool:
    return not _is_vercel_runtime() and _env_flag("SIGMA_ENABLE_MOCK_LOGIN", False)


def _uses_request_scoped_labor_runtime() -> bool:
    workbench_home = str(os.environ.get("SIGMA_WORKBENCH_HOME") or "")
    storage_backend = os.environ.get("SIGMA_LABOR_STORAGE_BACKEND", "").strip().lower()
    return _is_vercel_runtime() or (workbench_home.startswith("/tmp/") and storage_backend == "blob")


def _assert_local_labor_material_tool_available() -> None:
    if not _uses_request_scoped_labor_runtime():
        return
    raise HTTPException(
        status_code=409,
        detail=_labor_request_error(
            message="当前请求作用域环境不执行本地材料扫描、复制或整批解析。",
            error_code="LABOR_LOCAL_MATERIAL_TOOL_DISABLED",
            next_action="请在受控本地/内网持久化环境运行材料验证；Vercel 正式长任务必须交给 Personal Worker。",
        ),
    )


def _uses_ephemeral_serverless_storage() -> bool:
    workbench_home = str(os.environ.get("SIGMA_WORKBENCH_HOME") or "")
    return _is_vercel_runtime() and workbench_home.startswith("/tmp/") and not labor_persistent_storage_enabled()


def _uses_vercel_labor_light_uat() -> bool:
    access = os.environ.get("SIGMA_OVERSEAS_LABOR_ACCESS", "uat").strip().lower() or "uat"
    return bool(os.environ.get("VERCEL")) and access in {"uat", "uat_trial", "trial"}


def _uses_personal_labor_worker() -> bool:
    return os.environ.get("SIGMA_LABOR_EXECUTION_MODE", "").strip().lower() in {
        "personal-worker", "personal_worker", "desktop-worker", "desktop_worker"
    }


def _with_personal_worker_status(metadata: dict) -> dict:
    if not _uses_personal_labor_worker():
        return metadata
    run_id = str(metadata.get("id") or "")
    mapping_preflight = (
        metadata.get("mappingPreflight")
        if isinstance(metadata.get("mappingPreflight"), dict)
        else {}
    )
    mapping_status = str(mapping_preflight.get("status") or "")
    mapping_active = mapping_status in {"queued", "running", "retry_wait"}
    current_generation = (
        str(mapping_preflight.get("taskGenerationId") or "").strip()
        if mapping_active
        else _labor_task_generation_id(metadata)
    )
    try:
        job = get_latest_labor_worker_job(
            run_id,
            task_generation_id=current_generation,
            job_type="mapping_preflight" if mapping_active else "reconcile",
        )
    except Exception as exc:  # noqa: BLE001 - worker status is supplemental run metadata.
        logger.warning(
            "Labor run worker status lookup failed: run_id=%s error_type=%s",
            run_id,
            type(exc).__name__,
        )
        return metadata
    if not job or job.get("status") == "succeeded":
        return metadata
    if mapping_active:
        return {
            **metadata,
            "workerTask": _public_labor_worker_job(job),
        }
    status = str(job.get("status") or "")
    labels = {
        "queued": ("waiting_for_personal_worker", "等待本人核对助手上线"),
        "retry_wait": ("retry_wait", "核对助手将在稍后重试"),
        "running": ("running", "本人核对助手正在处理"),
        "failed": ("failed", "本人核对助手处理失败"),
    }
    task_status, message = labels.get(status, (status, status))
    payload = dict(metadata)
    payload["workerTask"] = _public_labor_worker_job(job)
    payload["asyncTask"] = {
        **(payload.get("asyncTask") if isinstance(payload.get("asyncTask"), dict) else {}),
        "status": task_status,
        "statusLabel": message,
        "message": str(job.get("errorMessage") or message),
    }
    if status == "failed":
        payload.update(
            {
                "status": "失败",
                "errorCode": str(job.get("errorCode") or "LABOR_WORKER_FAILED"),
                "errorMessage": str(job.get("errorMessage") or message),
                "retryable": False,
            }
        )
    return payload


def _labor_request_error(
    *,
    message: str,
    error_code: str,
    retryable: bool = False,
    next_action: str = "",
    requires_reupload: bool = False,
    requires_human_review: bool = False,
) -> dict:
    return {
        "message": message,
        "errorCode": error_code,
        "retryable": retryable,
        "requiresReupload": requires_reupload,
        "requiresHumanReview": requires_human_review,
        "nextAction": next_action,
    }


def _raise_labor_run_missing(exc: FileNotFoundError) -> None:
    if _uses_ephemeral_serverless_storage():
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前 Vercel UAT 环境不保存上传批次。上传后跨请求会丢失文件，暂不支持在线抽取并比对。",
                error_code="LABOR_UAT_UPLOAD_NOT_PERSISTED",
                next_action="请改用本地/内网持久化环境，或使用“测试材料验证”。",
                requires_reupload=True,
            ),
        ) from exc
    raise HTTPException(
        status_code=404,
        detail=_labor_request_error(
            message="劳务核对批次记录未找到。",
            error_code="LABOR_RUN_NOT_FOUND",
            next_action="请返回「新建核对批次」重新创建并上传材料。",
            requires_reupload=True,
        ),
    ) from exc


def _developing_module_block(path: str) -> dict | None:
    blocked_paths: dict[str, dict] = {}
    for key, config in blocked_paths.items():
        if path in config["page_paths"] or any(path.startswith(prefix) for prefix in config["api_prefixes"]):
            return {"key": key, **config}
    return None


def _fbu_access_response(request: Request) -> Response | None:
    path = request.url.path
    is_fbu_api = path.startswith("/api/fbu-performance/")
    is_fbu_page = path.rstrip("/") == "/fbu-performance.html"
    if not (is_fbu_api or is_fbu_page):
        return None

    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    user_id = None
    if session_token:
        try:
            user_id = get_session_user_id(session_token)
        except KeyError:
            user_id = None
    if not user_id:
        if is_fbu_api:
            return JSONResponse({"detail": "未登录。"}, status_code=401)
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head><meta charset="utf-8"><title>需要登录</title></head>
              <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                <h1>请先登录西格玛工作台</h1>
                <p>FBU美洲绩效奖金核算仅对已授权用户开放。</p>
                <p><a href="/login.html?next=/fbu-performance.html">前往登录</a></p>
              </body>
            </html>
            """,
            status_code=401,
        )

    if not _user_can_enter_module(user_id, "fbu"):
        if is_fbu_api:
            return JSONResponse({"detail": "当前用户没有FBU美洲绩效奖金核算权限，或模块尚未开放。"}, status_code=403)
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head><meta charset="utf-8"><title>无权限访问</title></head>
              <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                <h1>无权限访问：FBU美洲绩效奖金核算</h1>
                <p>请联系系统管理员开放模块并授予FBU绩效核算管理员角色。</p>
                <p><a href="/">返回西格玛工作台</a></p>
              </body>
            </html>
            """,
            status_code=403,
        )
    return None


def _domestic_labor_access_response(request: Request) -> Response | None:
    path = request.url.path
    is_domestic_api = path.startswith("/api/domestic-labor/")
    is_domestic_page = path.rstrip("/") in {"/domestic-labor.html", "/labor.html"}
    if not (is_domestic_api or is_domestic_page):
        return None

    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    user_id = None
    if session_token:
        try:
            user_id = get_session_user_id(session_token)
        except KeyError:
            user_id = None
    if not user_id:
        if is_domestic_api:
            return JSONResponse({"detail": "未登录。"}, status_code=401)
        return RedirectResponse(
            url=f"/login.html?next={quote(path)}",
            status_code=302,
        )

    if not _user_can_enter_module(user_id, "domestic"):
        if is_domestic_api:
            return JSONResponse({"detail": "当前用户没有中国区外包工薪酬核算权限，或模块尚未开放。"}, status_code=403)
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head><meta charset="utf-8"><title>无权限访问</title></head>
              <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                <h1>无权限访问：中国区外包工薪酬核算</h1>
                <p>请联系系统管理员开放模块并授予国内外包工核算管理员角色。</p>
                <p><a href="/">返回西格玛工作台</a></p>
              </body>
            </html>
            """,
            status_code=403,
        )
    return None


def _protected_static_page_access_response(request: Request) -> Response | None:
    path = request.url.path.rstrip("/")
    page_config = {
        "/recruitment.html": {"module_id": "recruitment", "label": "全球招聘奖金核算"},
        "/china-employee-payroll.html": {"module_id": "employee", "label": "中国区正式工薪酬核算"},
        "/employee-payroll.html": {"module_id": "employee", "label": "中国区正式工薪酬核算"},
        "/overseas-compensation.html": {"module_ids": ("fbu", "overseas"), "label": "海外薪酬核算"},
        "/admin.html": {"admin_only": True, "label": "后台管理"},
    }.get(path)
    if not page_config:
        return None

    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    user_id = None
    if session_token:
        try:
            user_id = get_session_user_id(session_token)
        except KeyError:
            user_id = None
    if not user_id:
        return RedirectResponse(
            url=f"/login.html?next={quote(path)}",
            status_code=302,
        )

    if page_config.get("admin_only"):
        try:
            current = _get_cached_current_user(user_id)
        except KeyError:
            current = {}
        can_enter = any(role.get("id") == "admin" for role in current.get("roles", []))
    elif page_config.get("module_ids"):
        can_enter = any(_user_can_enter_module(user_id, module_id) for module_id in page_config["module_ids"])
    else:
        can_enter = _user_can_enter_module(user_id, str(page_config["module_id"]))
    if can_enter:
        return None

    label = page_config["label"]
    detail = "该页面仅系统管理员可访问。" if page_config.get("admin_only") else "请联系系统管理员开放模块并授予对应模块管理员角色。"
    return HTMLResponse(
        f"""
        <!doctype html>
        <html lang="zh-CN">
          <head><meta charset="utf-8"><title>无权限访问</title></head>
          <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
            <h1>无权限访问：{label}</h1>
            <p>{detail}</p>
            <p><a href="/">返回西格玛工作台</a></p>
          </body>
        </html>
        """,
        status_code=403,
    )


def _overseas_labor_access_response(request: Request) -> Response | None:
    path = request.url.path
    is_labor_api = path.startswith("/api/labor/") and path != "/api/labor/access"
    is_labor_page = path.rstrip("/") == "/overseas-labor.html"
    if not (is_labor_api or is_labor_page):
        return None
    if is_labor_api and not _labor_browser_auth_path(path):
        return None
    access = _overseas_labor_access_config()
    if not access["canUse"]:
        if is_labor_api:
            return JSONResponse({"detail": access["message"], "access": access}, status_code=403)
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head><meta charset="utf-8"><title>模块未开放</title></head>
              <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                <h1>海外劳务报账核对暂未开放</h1>
                <p>当前生产权限未开放此 UAT 模块。请联系薪酬自动化管理员调整权限后再访问。</p>
                <p><a href="/">返回西格玛工作台</a></p>
              </body>
            </html>
            """,
            status_code=403,
        )
    if isinstance(getattr(request.state, "labor_current_user", None), dict):
        return None
    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    user_id = None
    if session_token:
        try:
            user_id = get_session_user_id(session_token)
        except KeyError:
            user_id = None
    if not user_id:
        if not labor_auth_required():
            return None
        if is_labor_api:
            return JSONResponse({"detail": "未登录。", "access": access}, status_code=401)
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head><meta charset="utf-8"><title>需要登录</title></head>
              <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                <h1>请先登录西格玛工作台</h1>
                <p>海外劳务报账核对仅对已授权用户开放。</p>
                <p><a href="/login.html?next=/overseas-labor.html">前往登录</a></p>
              </body>
            </html>
            """,
            status_code=401,
        )
    if not _user_can_enter_module(user_id, "overseas"):
        if is_labor_api:
            return JSONResponse({"detail": "当前用户没有海外劳务报账核对权限。", "access": access}, status_code=403)
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head><meta charset="utf-8"><title>无权限访问</title></head>
              <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                <h1>无权限访问：海外劳务报账核对</h1>
                <p>该模块保持 UAT 试点，仅海外报账管理员或系统管理员可进入。</p>
                <p><a href="/">返回西格玛工作台</a></p>
              </body>
            </html>
            """,
            status_code=403,
        )
    return None


def _labor_runtime_guarded_mutation(path: str, method: str) -> bool:
    if method.upper() not in {"POST", "PUT", "PATCH", "DELETE"}:
        return False
    return (
        path == "/api/labor/runs"
        or path.startswith("/api/labor/runs/")
        or path == "/api/labor/material-runs"
        or path == "/api/labor/worker/devices"
        or path.startswith("/api/labor/worker/devices/")
        or bool(re.fullmatch(r"/api/labor/worker/jobs/[^/]+/(?:result|mapping-preflight-result|complete)", path))
    )


def _labor_client_contract_guarded_mutation(path: str, method: str) -> bool:
    if method.upper() not in {"POST", "PUT", "PATCH", "DELETE"}:
        return False
    return path == "/api/labor/runs" or path.startswith("/api/labor/runs/") or path == "/api/labor/material-runs"


def _labor_p1_formal_mutation(path: str, method: str) -> bool:
    if method.upper() not in {"POST", "PUT", "PATCH", "DELETE"}:
        return False
    if path == "/api/labor/runs":
        return True
    if re.fullmatch(
        r"/api/labor/runs/[^/]+/(?:files|mapping|mapping-preflight|business-review|extract-and-compare|upload-intents(?:/batch-finalize|/[^/]+/finalize)?)",
        path,
    ):
        return True
    return bool(re.fullmatch(r"/api/labor/worker/jobs/[^/]+/(?:result|mapping-preflight-result|complete)", path))


def _labor_p1_formal_request(path: str, method: str) -> bool:
    if _labor_p1_formal_mutation(path, method):
        return True
    normalized_method = method.upper()
    if path == "/api/labor/worker/jobs/claim":
        return normalized_method == "POST"
    match = re.fullmatch(
        r"/api/labor/worker/jobs/[^/]+/(heartbeat|input|input-manifest|input-file|mapping-preflight-result|result|events|complete|fail)",
        path,
    )
    if not match:
        return False
    action = match.group(1)
    if action in {"input", "input-manifest", "input-file"}:
        return normalized_method == "GET"
    return normalized_method == "POST"


def _labor_p1_required() -> bool:
    return _env_flag("SIGMA_LABOR_P1_REQUIRED", False)


_LABOR_P1_READINESS_CACHE: dict[str, object] = {}
_LABOR_P1_READINESS_CACHE_LOCK = threading.Lock()
_LABOR_P1_READINESS_CACHE_ENV_KEYS = (
    "SIGMA_LABOR_P1_REQUIRED",
    "SIGMA_LABOR_AUTH_REQUIRED",
    "SIGMA_LABOR_STATE_BACKEND",
    "SIGMA_LABOR_JOB_BACKEND",
    "SIGMA_LABOR_STORAGE_BACKEND",
    "SIGMA_LABOR_STORAGE_ENV",
    "SIGMA_LABOR_EXECUTION_MODE",
    "SIGMA_LABOR_OPERATIONS_TOKEN",
    "SIGMA_LABOR_EXTERNAL_AI_ENABLED",
    "SIGMA_LABOR_REQUIRE_CLIENT_CONTRACT",
    "SIGMA_LABOR_WORKER_TOKENS",
    "SIGMA_LABOR_WORKER_UPDATE_MANIFEST",
    "ADMIN_DATABASE_URL",
    "SIGMA_LABOR_DATABASE_URL",
    "SIGMA_LABOR_JOB_DATABASE_URL",
    "FEISHU_APP_ID",
    "FEISHU_REDIRECT_URI",
    "SESSION_COOKIE_SECURE",
)


def _labor_p1_readiness_cache_key() -> tuple[object, ...]:
    environment_fingerprint = hashlib.sha256(
        "\0".join(os.environ.get(key, "") for key in _LABOR_P1_READINESS_CACHE_ENV_KEYS).encode("utf-8")
    ).hexdigest()
    return (
        environment_fingerprint,
        id(labor_p1_worker_job_store_health),
        id(labor_persistent_storage_health),
        id(labor_worker_identity_health),
        id(labor_auth_health),
        id(labor_postgres_state_health),
        id(_labor_build_snapshot),
    )


def _collect_labor_p1_readiness_snapshot() -> dict:
    try:
        queue_health = labor_p1_worker_job_store_health()
    except Exception as exc:  # noqa: BLE001 - readiness must return a sanitized blocker.
        queue_health = {"backend": "postgres", "configured": False, "ready": False, "error": str(exc)[:240]}
    storage_info = labor_persistent_storage_info()
    storage_health = labor_persistent_storage_health(probe=_labor_p1_required())
    worker_identity_health = labor_worker_identity_health()
    return evaluate_labor_production_readiness(
        env=os.environ,
        storage_info=storage_info,
        queue_health=queue_health,
        build_info=_labor_build_snapshot(),
        auth_health=labor_auth_health(),
        state_health=labor_postgres_state_health(),
        storage_health=storage_health,
        worker_identity_health=worker_identity_health,
    )


def _labor_p1_readiness_snapshot(*, force_refresh: bool = False) -> dict:
    try:
        configured_ttl = float(os.environ.get("SIGMA_LABOR_P1_READINESS_CACHE_SECONDS", "30"))
    except ValueError:
        configured_ttl = 30.0
    ready_ttl = max(0.0, min(configured_ttl, 300.0))
    cache_key = _labor_p1_readiness_cache_key()
    now = time.monotonic()

    if not force_refresh and ready_ttl > 0:
        cached = _LABOR_P1_READINESS_CACHE.get("value")
        checked_at = float(_LABOR_P1_READINESS_CACHE.get("checkedAt") or 0.0)
        cached_key = _LABOR_P1_READINESS_CACHE.get("key")
        cached_ready = bool(((cached or {}).get("p1") or {}).get("ready")) if isinstance(cached, dict) else False
        cache_ttl = ready_ttl if cached_ready else min(ready_ttl, 5.0)
        if isinstance(cached, dict) and cached_key == cache_key and now - checked_at < cache_ttl:
            return dict(cached)

    with _LABOR_P1_READINESS_CACHE_LOCK:
        now = time.monotonic()
        if not force_refresh and ready_ttl > 0:
            cached = _LABOR_P1_READINESS_CACHE.get("value")
            checked_at = float(_LABOR_P1_READINESS_CACHE.get("checkedAt") or 0.0)
            cached_key = _LABOR_P1_READINESS_CACHE.get("key")
            cached_ready = bool(((cached or {}).get("p1") or {}).get("ready")) if isinstance(cached, dict) else False
            cache_ttl = ready_ttl if cached_ready else min(ready_ttl, 5.0)
            if isinstance(cached, dict) and cached_key == cache_key and now - checked_at < cache_ttl:
                return dict(cached)
        result = _collect_labor_p1_readiness_snapshot()
        _LABOR_P1_READINESS_CACHE.update(
            {
                "key": cache_key,
                "checkedAt": now,
                "value": dict(result),
            }
        )
        return result


def _labor_p1_gate_response(request: Request) -> Response | None:
    if not _labor_p1_required() or not _labor_p1_formal_request(request.url.path, request.method):
        return None
    readiness = _labor_p1_readiness_snapshot()
    if bool((readiness.get("p1") or {}).get("ready")):
        return None
    blockers = readiness.get("blockers") if isinstance(readiness.get("blockers"), list) else []
    blocker_codes = [str(item.get("code") or "") for item in blockers if isinstance(item, dict) and item.get("code")]
    return JSONResponse(
        {
            "detail": {
                **_labor_request_error(
                    message="P1 基础设施尚未通过统一 readiness，正式上传和核对任务已锁定。",
                    error_code="LABOR_P1_NOT_READY",
                    retryable=True,
                    next_action="请由管理员修复 readiness 阻断项后重试。",
                ),
                "blockerCodes": blocker_codes,
            }
        },
        status_code=503,
    )


def _labor_p1_signed_upload_required_detail() -> dict:
    return _labor_request_error(
        message="P1 正式环境仅接受私有对象存储签名直传。",
        error_code="LABOR_P1_SIGNED_UPLOAD_REQUIRED",
        next_action="请刷新页面后重新选择文件，由页面完成私有直传。",
        requires_reupload=True,
    )


def _labor_p1_legacy_upload_response(request: Request) -> Response | None:
    if (
        not _labor_p1_required()
        or request.method.upper() != "POST"
        or not re.fullmatch(r"/api/labor/runs/[^/]+/files", request.url.path)
    ):
        return None
    return JSONResponse(
        {"detail": _labor_p1_signed_upload_required_detail()},
        status_code=409,
    )


def _labor_p1_device_auth_response(request: Request) -> Response | None:
    path = request.url.path
    if (
        not _labor_p1_required()
        or not (path == "/api/labor/worker/devices" or path.startswith("/api/labor/worker/devices/"))
    ):
        return None
    try:
        trusted_auth_ready = bool(labor_auth_health().get("ready"))
    except Exception:  # noqa: BLE001 - device bootstrap must fail closed without exposing auth internals.
        trusted_auth_ready = False
    if trusted_auth_ready:
        return None
    return JSONResponse(
        {
            "detail": _labor_request_error(
                message="P1 Worker 设备只能由可信企业登录用户激活。",
                error_code="LABOR_P1_TRUSTED_AUTH_REQUIRED",
                next_action="请先完成飞书登录、Postgres 角色库和安全会话配置，再激活核对助手。",
            )
        },
        status_code=503,
    )


def _labor_browser_auth_path(path: str) -> bool:
    if path.rstrip("/") in {"/overseas-labor.html", "/overseas-payroll.html"}:
        return True
    if path.startswith("/api/overseas-payroll/worker/"):
        return False
    if path.startswith("/api/overseas-payroll/"):
        return True
    if path == "/api/tools" or path.startswith("/api/tool/"):
        return True
    if not path.startswith("/api/labor/"):
        return False
    if path == "/api/labor/worker/devices" or path.startswith("/api/labor/worker/devices/"):
        return True
    if path in {
        "/api/labor/worker/release",
        "/api/labor/worker/release/download",
        "/api/labor/worker/release/upload-intent",
    }:
        return True
    public_or_service_prefixes = (
        "/api/labor/access",
        "/api/labor/production-readiness",
        "/api/labor/operations",
        "/api/labor/maintenance/",
        "/api/labor/worker/",
    )
    return not any(path == prefix or path.startswith(prefix) for prefix in public_or_service_prefixes)


def _workbench_home_auth_response(request: Request) -> Response | None:
    if not labor_auth_required() or request.url.path not in {"/", "/index.html"}:
        return None
    current = current_user_from_request(request)
    if current is not None:
        request.state.workbench_current_user = current
        return None
    return RedirectResponse(
        url=f"/login.html?next={request.url.path}",
        status_code=302,
    )


def _labor_auth_access_response(request: Request) -> Response | None:
    if not labor_auth_required() or not _labor_browser_auth_path(request.url.path):
        return None
    is_overseas_payroll = request.url.path.rstrip("/") == "/overseas-payroll.html" or (
        request.url.path.startswith("/api/overseas-payroll/")
        or request.url.path == "/api/tools"
        or request.url.path.startswith("/api/tool/")
    )
    module_label = "海外薪资工作台" if is_overseas_payroll else "海外劳务报账核对"
    current = _labor_current_user_from_request(request)
    if current is None:
        if request.url.path.startswith("/api/"):
            return JSONResponse(
                {
                    "detail": _labor_request_error(
                        message="请先登录西格玛工作台。",
                        error_code="LABOR_AUTH_REQUIRED",
                        next_action=f"登录后重新进入{module_label}。",
                    )
                },
                status_code=401,
            )
        return RedirectResponse(
            url=f"/login.html?next={request.url.path}",
            status_code=302,
        )
    if not user_can_enter_module(current, "overseas"):
        if request.url.path.startswith("/api/"):
            return JSONResponse(
                {
                    "detail": _labor_request_error(
                        message=f"当前用户没有{module_label}权限。",
                        error_code="LABOR_MODULE_FORBIDDEN",
                        next_action="请联系系统管理员授予海外报账管理员角色。",
                    )
                },
                status_code=403,
            )
        return HTMLResponse(
            f"<h1>无权限访问：{module_label}</h1><p>请联系系统管理员授予海外报账管理员角色。</p>",
            status_code=403,
        )
    request.state.labor_current_user = current
    return None


def _labor_request_actor(request: Request) -> tuple[str, bool]:
    current = getattr(request.state, "labor_current_user", None)
    if not isinstance(current, dict):
        current = _labor_current_user_from_request(request) if labor_auth_required() else None
    if not isinstance(current, dict):
        return "local-default", False
    user = current.get("user") if isinstance(current.get("user"), dict) else {}
    user_id = str(user.get("id") or "").strip()
    if not user_id:
        return "local-default", False
    return user_id, user_is_system_admin(current)


def _labor_action_actor(request: Request | None, payload: dict, *legacy_fields: str) -> str:
    if labor_auth_required() and request is not None:
        actor_user_id, _ = _labor_request_actor(request)
        return actor_user_id
    for field in legacy_fields:
        value = str(payload.get(field) or "").strip()
        if value:
            return value
    return "local-default"


def _labor_run_owner_access_response(request: Request) -> Response | None:
    if not labor_auth_required():
        return None
    match = re.match(r"^/api/labor/runs/([0-9A-Za-z_-]+)(?:/|$)", request.url.path)
    if not match:
        return None
    run_id = match.group(1)
    actor_user_id, is_admin = _labor_request_actor(request)
    if is_admin:
        return None
    try:
        metadata = load_labor_metadata(get_labor_run_dir(run_id))
    except FileNotFoundError:
        return None
    if str(metadata.get("ownerUserId") or "") == actor_user_id:
        return None
    return JSONResponse(
        {
            "detail": _labor_request_error(
                message="劳务核对批次记录未找到。",
                error_code="LABOR_RUN_NOT_FOUND",
                next_action="请返回批次列表确认该批次仍存在。",
            )
        },
        status_code=404,
    )


def _labor_client_contract_matches(request: Request, build: dict) -> bool:
    return (
        request.headers.get("x-sigma-labor-api-contract", "") == str(OVERSEAS_LABOR_API_CONTRACT_VERSION)
        and request.headers.get("x-sigma-labor-ui-version", "") == OVERSEAS_LABOR_MODULE_VERSION
        and request.headers.get("x-sigma-labor-ui-build", "") == str(build.get("buildId") or "")
    )


def _labor_assert_runtime_current() -> dict:
    build = _labor_build_snapshot()
    if build["status"] != "current":
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="服务启动后海外劳务代码已变化，请重启或重新部署后再开始正式操作。",
                error_code="LABOR_SERVICE_RESTART_REQUIRED",
                next_action="重启当前服务，确认页面与 API build 一致后重试。",
            ),
        )
    return build


@app.middleware("http")
async def overseas_labor_access_gate(request: Request, call_next):
    path = request.url.path
    protected_page_response = _protected_static_page_access_response(request)
    if protected_page_response is not None:
        return protected_page_response
    fbu_access_response = _fbu_access_response(request)
    if fbu_access_response is not None:
        return fbu_access_response
    domestic_access_response = _domestic_labor_access_response(request)
    if domestic_access_response is not None:
        return domestic_access_response
    home_auth_response = _workbench_home_auth_response(request)
    if home_auth_response is not None:
        return home_auth_response
    auth_response = _labor_auth_access_response(request)
    if auth_response is not None:
        return auth_response
    owner_response = _labor_run_owner_access_response(request)
    if owner_response is not None:
        return owner_response
    if _hide_developing_modules():
        blocked = _developing_module_block(path)
        if blocked:
            detail = f"{blocked['label']}在当前 UAT 环境暂未开放。"
            if path.startswith("/api/"):
                return JSONResponse(
                    {
                        "detail": detail,
                        "access": _workbench_access_config(),
                    },
                    status_code=403,
                )
            return HTMLResponse(
                f"""
                <!doctype html>
                <html lang="zh-CN">
                  <head><meta charset="utf-8"><title>模块未开放</title></head>
                  <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                    <h1>{blocked['label']}暂未开放</h1>
                    <p>当前 UAT 环境仅开放已上线或试用模块，开发中模块已关闭入口。</p>
                    <p><a href="/">返回西格玛工作台</a></p>
                  </body>
                </html>
                """,
                status_code=403,
            )
    overseas_access_response = _overseas_labor_access_response(request)
    if overseas_access_response is not None:
        return overseas_access_response
    is_labor_api = (
        path.startswith("/api/labor/") and path != "/api/labor/access"
    ) or path.startswith("/api/overseas-payroll/") or path == "/api/tools" or path.startswith("/api/tool/")
    is_labor_page = path.rstrip("/") in {"/overseas-labor.html", "/overseas-payroll.html"}
    if is_labor_api or is_labor_page:
        access = _overseas_labor_access_config()
        if not access["canUse"]:
            if is_labor_api:
                return JSONResponse({"detail": access["message"], "access": access}, status_code=403)
            return HTMLResponse(
                """
                <!doctype html>
                <html lang="zh-CN">
                  <head><meta charset="utf-8"><title>模块未开放</title></head>
                  <body style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;padding:48px;color:#0f172a;">
                    <h1>海外劳务报账核对暂未开放</h1>
                    <p>当前生产权限未开放此 UAT 模块。请联系薪酬自动化管理员调整权限后再访问。</p>
                    <p><a href="/">返回西格玛工作台</a></p>
                  </body>
                </html>
                """,
                status_code=403,
            )
    if _labor_runtime_guarded_mutation(path, request.method):
        try:
            build = _labor_assert_runtime_current()
        except HTTPException as exc:
            return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
        if (
            _labor_client_contract_guarded_mutation(path, request.method)
            and not _labor_client_contract_matches(request, build)
        ):
            return JSONResponse(
                {
                    "detail": _labor_request_error(
                        message="当前页面与海外劳务 API 版本不一致，正式操作已锁定。",
                        error_code="LABOR_CLIENT_UPGRADE_REQUIRED",
                        next_action="刷新页面并确认页面显示的 build 与服务一致后重试。",
                    )
                },
                status_code=409,
            )
    p1_response = _labor_p1_gate_response(request)
    if p1_response is not None:
        return p1_response
    device_auth_response = _labor_p1_device_auth_response(request)
    if device_auth_response is not None:
        return device_auth_response
    legacy_upload_response = _labor_p1_legacy_upload_response(request)
    if legacy_upload_response is not None:
        return legacy_upload_response
    response = await call_next(request)
    if path in {
        "/overseas-labor.html",
        "/overseas-labor.js",
        "/overseas-payroll.html",
        "/overseas-payroll.js",
        "/api/labor/access",
    }:
        build = _labor_build_snapshot()
        response.headers["Cache-Control"] = "no-store"
        response.headers["X-Sigma-Labor-Build"] = str(build["buildId"])[:64]
    return response


@app.middleware("http")
async def audit_authenticated_business_activity(request: Request, call_next):
    candidate = _business_activity_audit_event(request.method, request.url.path, 200)
    if candidate is None:
        return await call_next(request)

    actor_user_id = ""
    session_token = request.cookies.get(SESSION_COOKIE_NAME)
    if session_token:
        try:
            actor_user_id = get_session_user_id(session_token)
        except KeyError:
            actor_user_id = ""

    response = await call_next(request)
    event = _business_activity_audit_outcome_event(request.method, request.url.path, response.status_code)
    if actor_user_id and event:
        try:
            await asyncio.to_thread(record_audit_event, actor_user_id, **event)
        except Exception as exc:
            audit_logger.warning(
                "Audit event persistence failed action=%s module=%s error=%s",
                event["action"],
                event["target_id"],
                type(exc).__name__,
            )
    return response


class _FBUGZipMiddleware:
    """Compress only the FBU API and its large text assets."""

    def __init__(self, app, *, minimum_size: int = 1024, compresslevel: int = 5):
        self.app = app
        self.compressed_app = GZipMiddleware(
            app,
            minimum_size=minimum_size,
            compresslevel=compresslevel,
        )

    async def __call__(self, scope, receive, send):
        path = str(scope.get("path") or "")
        should_refresh_html = path == "/fbu-performance.html"
        if should_refresh_html:
            scope = dict(scope)
            scope["headers"] = [
                (name, value)
                for name, value in scope.get("headers", [])
                if name.lower() not in {b"if-none-match", b"if-modified-since"}
            ]
        should_compress = (
            path.startswith("/api/fbu-performance")
            or path in {
                "/fbu-performance.html",
                "/fbu-performance.js",
                "/styles.css",
            }
        )
        should_cache = (
            bool(scope.get("query_string"))
            and path in {"/fbu-performance.js", "/styles.css"}
        ) or path == "/assets/sigma-platform-logo-20260731.png"

        async def send_with_cache_headers(message):
            if message["type"] == "http.response.start":
                headers = MutableHeaders(scope=message)
                if should_cache:
                    headers["Cache-Control"] = "public, max-age=31536000, immutable"
                elif should_refresh_html:
                    headers["Cache-Control"] = "no-store"
                    headers["Pragma"] = "no-cache"
            await send(message)

        target = self.compressed_app if should_compress else self.app
        await target(scope, receive, send_with_cache_headers)


app.add_middleware(_FBUGZipMiddleware, minimum_size=1024, compresslevel=5)


@app.middleware("http")
async def add_fbu_request_timing(request: Request, call_next):
    if not request.url.path.startswith("/api/fbu-performance"):
        return await call_next(request)
    request_id = (
        str(request.headers.get("x-sigma-request-id") or "").strip()
        or secrets.token_hex(8)
    )
    started = perf_counter()
    response = await call_next(request)
    elapsed_ms = (perf_counter() - started) * 1000
    response.headers["Server-Timing"] = f"fbu;dur={elapsed_ms:.1f}"
    response.headers["X-Sigma-Request-ID"] = request_id
    fbu_logger.info(
        "FBU request id=%s method=%s path=%s status=%s duration_ms=%.1f bytes=%s",
        request_id,
        request.method,
        request.url.path,
        response.status_code,
        elapsed_ms,
        response.headers.get("content-length", ""),
    )
    return response


@app.get("/api/health")
def health() -> dict:
    ensure_data_files()
    return {"status": "ok", "rule_workbook": str(DEFAULT_RULE_WORKBOOK)}


@app.get("/api/auth/mock-users")
def api_auth_mock_users() -> dict:
    if not _mock_auth_enabled():
        raise HTTPException(status_code=404, detail="Not Found")
    users = get_admin_state()["users"]
    return {
        "users": [
            {"id": user["id"], "name": user["name"], "email": user.get("email"), "roleIds": user.get("roleIds", [])}
            for user in users
        ]
    }


@app.post("/api/auth/mock-login")
def api_auth_mock_login(response: Response, payload: dict = Body(...)) -> dict:
    if not _mock_auth_enabled():
        raise HTTPException(status_code=404, detail="Not Found")
    user_id = _validate_safe_id(str(payload.get("userId") or payload.get("user_id") or ""), "user_id")
    try:
        token = create_session(user_id)
        current = get_current_user(user_id)
        _clear_current_user_cache()
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="用户不存在。") from exc
    response.set_cookie(
        SESSION_COOKIE_NAME,
        token,
        httponly=True,
        samesite="lax",
        secure=bool(AUTH_CONFIG["session_cookie_secure"]),
        max_age=7 * 24 * 60 * 60,
        path="/",
    )
    return current


@app.post("/api/auth/logout")
def api_auth_logout(response: Response, sigma_session: Optional[str] = Cookie(default=None)) -> dict:
    if sigma_session:
        delete_session(sigma_session)
        _clear_current_user_cache()
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return {"status": "ok"}


@app.get("/api/auth/logout")
def api_auth_logout_redirect(next: str = "login.html?next=%2F", sigma_session: Optional[str] = Cookie(default=None)) -> RedirectResponse:
    if sigma_session:
        delete_session(sigma_session)
        _clear_current_user_cache()
    redirect_target = next if next.startswith("/") or next.startswith("login.html") else "login.html?next=%2F"
    response = RedirectResponse(redirect_target, status_code=302)
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return response


@app.get("/api/auth/feishu/config")
def api_auth_feishu_config() -> dict:
    configured = bool(AUTH_CONFIG["feishu_app_id"] and AUTH_CONFIG["feishu_redirect_uri"])
    return {
        "configured": configured,
        "redirectUri": AUTH_CONFIG["feishu_redirect_uri"] if configured else "",
    }


@app.get("/api/auth/feishu/login")
def api_auth_feishu_login() -> RedirectResponse:
    if not AUTH_CONFIG["feishu_app_id"] or not AUTH_CONFIG["feishu_redirect_uri"]:
        raise HTTPException(status_code=503, detail="飞书应用未配置。")
    state = secrets.token_urlsafe(24)
    params = urlencode({
        "app_id": AUTH_CONFIG["feishu_app_id"],
        "redirect_uri": AUTH_CONFIG["feishu_redirect_uri"],
        "state": state,
    })
    response = RedirectResponse(f"{AUTH_CONFIG['feishu_auth_url']}?{params}", status_code=302)
    response.set_cookie(
        FEISHU_STATE_COOKIE_NAME,
        state,
        httponly=True,
        samesite="lax",
        secure=bool(AUTH_CONFIG["session_cookie_secure"]),
        max_age=10 * 60,
        path="/",
    )
    return response


@app.get("/api/auth/feishu/callback")
@app.get("/api/auth/lark/callback")
def api_auth_feishu_callback(
    background_tasks: BackgroundTasks,
    response: Response,
    code: str = "",
    state: str = "",
    sigma_feishu_state: Optional[str] = Cookie(default=None),
) -> RedirectResponse:
    if not code or not state or not sigma_feishu_state or state != sigma_feishu_state:
        raise HTTPException(status_code=400, detail="飞书登录 state 校验失败。")
    if not AUTH_CONFIG["feishu_app_id"] or not AUTH_CONFIG["feishu_app_secret"]:
        raise HTTPException(status_code=503, detail="飞书应用密钥未配置。")

    app_access_token = _get_feishu_app_access_token()
    token_data = _get_feishu_user_access_token(code, app_access_token)
    user_info = _get_feishu_user_info(str(token_data["access_token"]))
    identity = _feishu_identity_from_payloads(token_data, user_info)
    if not identity.get("avatar_url"):
        try:
            tenant_access_token = _get_feishu_tenant_access_token()
            contact_user = _get_feishu_contact_user(str(identity["feishu_open_id"]), tenant_access_token)
            identity["avatar_url"] = _extract_feishu_avatar_url(contact_user)
        except HTTPException:
            identity["avatar_url"] = None
    user = upsert_feishu_user(**identity)
    session_token = create_session(user["id"], action="feishu_login")
    notification = _queue_new_user_permission_notification(user)
    if notification:
        background_tasks.add_task(_dispatch_pending_permission_notifications)
    _clear_current_user_cache()

    redirect = RedirectResponse("/", status_code=302)
    redirect.delete_cookie(FEISHU_STATE_COOKIE_NAME, path="/")
    redirect.set_cookie(
        SESSION_COOKIE_NAME,
        session_token,
        httponly=True,
        samesite="lax",
        secure=bool(AUTH_CONFIG["session_cookie_secure"]),
        max_age=7 * 24 * 60 * 60,
        path="/",
    )
    return redirect


@app.get("/api/me")
def api_me(actor_user_id: str = Depends(_current_user_id)) -> dict:
    return _get_cached_current_user(actor_user_id)


@app.post("/api/workbench/feedback", status_code=201)
async def api_create_workbench_feedback(
    request: Request,
    background_tasks: BackgroundTasks,
    category: str = Form(...),
    module_id: str = Form(..., alias="moduleId"),
    description: str = Form(...),
    page_path: str = Form("", alias="pagePath"),
    attachments: list[UploadFile] = File(default=[]),
    actor_user_id: str = Depends(_current_user_id),
) -> dict:
    if len(attachments) > 3:
        raise HTTPException(status_code=400, detail="每次最多上传 3 张截图。")
    allowed_types = {"image/png", "image/jpeg", "image/webp"}
    attachment_rows = []
    for upload in attachments:
        content_type = str(upload.content_type or "").lower()
        if content_type not in allowed_types:
            raise HTTPException(status_code=400, detail="截图仅支持 PNG、JPG 或 WebP 格式。")
        content = await upload.read(5 * 1024 * 1024 + 1)
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="单张截图不能超过 5MB。")
        if not content:
            raise HTTPException(status_code=400, detail="截图文件为空。")
        attachment_rows.append({
            "filename": upload.filename or "screenshot",
            "contentType": content_type,
            "content": content,
        })
    current = _get_cached_current_user(actor_user_id)
    try:
        feedback = create_feedback(
            user=current["user"],
            category=category,
            module_id=module_id,
            description=description,
            page_path=page_path,
            user_agent=request.headers.get("user-agent", ""),
            attachments=attachment_rows,
        )
    except ValueError as exc:
        messages = {
            "invalid_feedback_category": "请选择有效的反馈类型。",
            "invalid_feedback_module": "请选择有效的所属模块。",
            "invalid_feedback_description": "问题描述需要填写 4 至 4000 个字符。",
        }
        raise HTTPException(status_code=400, detail=messages.get(str(exc), "反馈内容无效。")) from exc
    queued = _queue_feedback_notification(feedback)
    if queued:
        background_tasks.add_task(_dispatch_pending_permission_notifications)
    record_audit_event(
        actor_user_id,
        "submit_feedback",
        "feedback",
        feedback["id"],
        f"{feedback['moduleName']} · 附件 {feedback['attachmentCount']} 张",
    )
    return {"feedback": feedback}


@app.get("/api/workbench/feedback/mine")
def api_my_workbench_feedback(
    limit: int = 50,
    actor_user_id: str = Depends(_current_user_id),
) -> dict:
    return {"feedback": list_feedback_for_user(actor_user_id, limit=limit)}


def _feedback_for_viewer(feedback_id: str, actor_user_id: str) -> dict[str, Any]:
    try:
        feedback = get_feedback(feedback_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="反馈记录不存在。") from exc
    current = _get_cached_current_user(actor_user_id)
    is_admin = any(role.get("id") == "admin" for role in current.get("roles", []))
    if feedback.get("userId") != actor_user_id and not is_admin:
        raise HTTPException(status_code=404, detail="反馈记录不存在。")
    return feedback


@app.get("/api/workbench/feedback/{feedback_id}")
def api_workbench_feedback_detail(
    feedback_id: str,
    actor_user_id: str = Depends(_current_user_id),
) -> dict:
    return {"feedback": _feedback_for_viewer(feedback_id, actor_user_id)}


@app.get("/api/workbench/feedback/{feedback_id}/attachments/{attachment_id}")
def api_workbench_feedback_attachment(
    feedback_id: str,
    attachment_id: str,
    actor_user_id: str = Depends(_current_user_id),
) -> Response:
    _feedback_for_viewer(feedback_id, actor_user_id)
    try:
        attachment = get_feedback_attachment(feedback_id, attachment_id)
    except KeyError as exc:
        raise HTTPException(status_code=404, detail="截图附件不存在。") from exc
    return Response(
        content=attachment["content"],
        media_type=attachment["content_type"],
        headers={
            "Content-Disposition": f"inline; filename*=UTF-8''{quote(str(attachment['filename']), safe='')}"
        },
    )


@app.get("/api/workbench/announcements")
def api_workbench_announcements(
    limit: int = 50,
    actor_user_id: str = Depends(_current_user_id),
) -> dict:
    return {"announcements": list_announcements(limit=limit)}


@app.get("/api/workbench/module-contacts/{contact_key}")
def api_workbench_module_contact(
    contact_key: str,
    actor_user_id: str = Depends(_current_user_id),
) -> dict:
    contacts = {
        "overseas-payroll": {
            "name": "夏盈盈",
            "department": "海外薪酬组",
        },
    }
    contact = contacts.get(contact_key)
    if contact is None:
        raise HTTPException(status_code=404, detail="未找到模块对接人。")
    matched_user = next(
        (
            user
            for user in list_users()
            if str(user.get("name") or "").strip() == contact["name"]
            and str(user.get("status") or "") == "active"
        ),
        None,
    )
    avatar_url = str((matched_user or {}).get("avatarUrl") or "").strip()
    if not avatar_url.startswith("https://"):
        avatar_url = ""
    public_avatar_url = (
        f"/api/workbench/module-contacts/{contact_key}/avatar"
        if avatar_url
        else ""
    )
    return {"contact": {**contact, "avatarUrl": public_avatar_url}}


@app.get("/api/workbench/module-contacts/{contact_key}/avatar")
def api_workbench_module_contact_avatar(
    contact_key: str,
    actor_user_id: str = Depends(_current_user_id),
) -> Response:
    contact_names = {"overseas-payroll": "夏盈盈"}
    contact_name = contact_names.get(contact_key)
    if contact_name is None:
        raise HTTPException(status_code=404, detail="未找到模块对接人。")
    matched_user = next(
        (
            user
            for user in list_users()
            if str(user.get("name") or "").strip() == contact_name
            and str(user.get("status") or "") == "active"
        ),
        None,
    )
    avatar_url = str((matched_user or {}).get("avatarUrl") or "").strip()
    parsed_avatar_url = urlparse(avatar_url)
    avatar_host = str(parsed_avatar_url.hostname or "").casefold()
    if (
        parsed_avatar_url.scheme != "https"
        or not avatar_host.endswith(".feishucdn.com")
    ):
        raise HTTPException(status_code=404, detail="对接人头像暂不可用。")
    try:
        upstream = httpx.get(avatar_url, timeout=5.0, follow_redirects=True)
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=404, detail="对接人头像暂不可用。") from exc
    content_type = str(upstream.headers.get("content-type") or "").split(";", 1)[0].strip().lower()
    if upstream.status_code != 200 or content_type not in {"image/jpeg", "image/png", "image/webp"}:
        raise HTTPException(status_code=404, detail="对接人头像暂不可用。")
    if not upstream.content or len(upstream.content) > 2 * 1024 * 1024:
        raise HTTPException(status_code=404, detail="对接人头像暂不可用。")
    return Response(
        content=upstream.content,
        media_type=content_type,
        headers={"Cache-Control": "private, max-age=3600"},
    )


@app.get("/api/admin/feedback")
def api_admin_feedback(
    limit: int = 100,
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    return {"feedback": list_feedback_for_admin(limit=limit)}


@app.get("/api/admin/announcements")
def api_admin_announcements(
    limit: int = 100,
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    return {"announcements": list_announcements(limit=limit)}


@app.post("/api/admin/announcements", status_code=201)
def api_publish_workbench_announcement(
    background_tasks: BackgroundTasks,
    payload: dict = Body(...),
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    current = _get_cached_current_user(actor_user_id)
    try:
        announcement = create_announcement(
            actor=current["user"],
            kind=str(payload.get("kind") or ""),
            title=str(payload.get("title") or ""),
            content=str(payload.get("content") or ""),
            module_id=str(payload.get("moduleId") or "home"),
            visual_style=str(payload.get("visualStyle") or "sunny"),
        )
    except ValueError as exc:
        messages = {
            "invalid_announcement_kind": "请选择功能更新或问题提示。",
            "invalid_announcement_module": "请选择有效的关联模块。",
            "invalid_announcement_style": "请选择有效的便签样式。",
            "invalid_announcement_title": "公告标题需要填写 2 至 120 个字符。",
            "invalid_announcement_content": "公告内容需要填写 2 至 8000 个字符。",
        }
        raise HTTPException(status_code=400, detail=messages.get(str(exc), "公告内容无效。")) from exc
    queued_recipients = 0
    if bool(payload.get("pushToFeishu", True)):
        queued_recipients = _queue_announcement_notifications(announcement)
        if queued_recipients:
            background_tasks.add_task(_dispatch_pending_permission_notifications)
    record_audit_event(
        actor_user_id,
        "publish_announcement",
        "announcement",
        announcement["id"],
        f"{announcement['kindLabel']} · {announcement['moduleName']} · 飞书推送 {queued_recipients} 人",
    )
    return {"announcement": announcement, "queuedRecipients": queued_recipients}


@app.get("/api/admin/state")
def api_admin_state(actor_user_id: str = Depends(_require_admin_user)) -> dict:
    return get_admin_state()


@app.get("/api/admin/users")
def api_admin_users(actor_user_id: str = Depends(_require_admin_user)) -> dict:
    return {"users": get_admin_state()["users"]}


@app.put("/api/admin/users/{user_id}/roles")
def api_set_user_roles(
    user_id: str,
    background_tasks: BackgroundTasks,
    payload: dict = Body(...),
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    try:
        user_id = _validate_safe_id(user_id, "user_id")
        role_ids = payload.get("roleIds") or payload.get("role_ids") or []
        if not isinstance(role_ids, list) or len(role_ids) > 20:
            raise HTTPException(status_code=400, detail="无效的角色列表。")
        role_ids = [_validate_safe_id(str(role_id), "role_id") for role_id in role_ids]
        before = get_current_user(user_id)
        result = {"user": set_user_roles(user_id, role_ids, actor_user_id=actor_user_id)}
        after = get_current_user(user_id)
        actor = get_current_user(actor_user_id)["user"]
        notification = _queue_permission_change_notification(before, after, actor)
        if notification:
            background_tasks.add_task(_dispatch_pending_permission_notifications)
        _clear_current_user_cache()
        return result
    except ValueError as exc:
        if str(exc) == "cannot_remove_own_admin":
            raise HTTPException(status_code=400, detail="不能移除当前登录账号的系统管理员角色。") from exc
        if str(exc) == "cannot_remove_last_admin":
            raise HTTPException(status_code=400, detail="系统至少需要保留一个系统管理员。") from exc
        raise HTTPException(status_code=400, detail="无效的角色授权。") from exc
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.patch("/api/admin/modules/{module_id}")
def api_set_module_enabled(
    module_id: str,
    payload: dict = Body(...),
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    try:
        module_id = _validate_safe_id(module_id, "module_id")
        result = {"module": set_module_enabled(module_id, _payload_bool(payload, "enabled"), actor_user_id=actor_user_id)}
        _clear_current_user_cache()
        return result
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.put("/api/admin/modules/{module_id}/roles/{role_id}")
def api_set_module_role_access(
    module_id: str,
    role_id: str,
    payload: dict = Body(...),
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    try:
        module_id = _validate_safe_id(module_id, "module_id")
        role_id = _validate_safe_id(role_id, "role_id")
        result = {
            "moduleAccess": set_module_role_access(
                module_id, role_id, _payload_bool(payload, "canEnter"), actor_user_id=actor_user_id
            )
        }
        _clear_current_user_cache()
        return result
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.put("/api/admin/roles/{role_id}/features/{feature_id}")
def api_set_feature_permission(
    role_id: str,
    feature_id: str,
    payload: dict = Body(...),
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    try:
        role_id = _validate_safe_id(role_id, "role_id")
        feature_id = _validate_safe_id(feature_id, "feature_id")
        result = {
            "rolePermissions": set_feature_permission(
                role_id, feature_id, _payload_bool(payload, "enabled"), actor_user_id=actor_user_id
            )
        }
        _clear_current_user_cache()
        return result
    except KeyError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@app.get("/api/admin/audit-logs")
def api_admin_audit_logs(
    page: int = 1,
    page_size: int = 12,
    limit: Optional[int] = None,
    actor_user_id: str = Depends(_require_admin_user),
) -> dict:
    normalized_page_size = max(1, min(limit if limit is not None else page_size, 200))
    total = count_audit_logs()
    total_pages = max(1, (total + normalized_page_size - 1) // normalized_page_size)
    normalized_page = min(max(1, page), total_pages)
    return {
        "logs": list_audit_logs(
            limit=normalized_page_size,
            offset=(normalized_page - 1) * normalized_page_size,
        ),
        "pagination": {
            "page": normalized_page,
            "pageSize": normalized_page_size,
            "total": total,
            "totalPages": total_pages,
        },
    }

@app.get("/api/labor/access")
def labor_access() -> dict:
    return _overseas_labor_access_config()


@app.get("/api/labor/storage-health")
def labor_storage_health(probe: bool = False) -> dict:
    return labor_persistent_storage_health(probe=probe)


@app.get("/api/labor/worker-health")
def labor_worker_health(probe: bool = False) -> dict:
    return legacy_labor_worker_job_store_health(probe=probe)


@app.get("/api/labor/storage-info")
def get_labor_storage_info() -> dict:
    return labor_storage_info(
        _labor_hardening_policy(),
        run_dir=LABOR_RUNS_DIR,
        cache_dir=labor_ocr_cache_dir(),
        audit_path=_labor_audit_path(),
        storage_backend=labor_storage_backend() or "local",
        storage_environment=labor_persistent_environment(),
        persistent_enabled=labor_persistent_storage_enabled(),
    )


@app.get("/api/labor/audit")
def get_labor_audit_events(request: Request, run_id: str = "", limit: int = 50) -> dict:
    safe_limit = min(max(int(limit), 1), 200)
    owner_filter = ""
    if labor_auth_required():
        actor_user_id, is_admin = _labor_request_actor(request)
        if not is_admin:
            owner_filter = actor_user_id
    events = read_labor_audit_events(
        _labor_audit_path(),
        limit=safe_limit,
        owner_user_id=owner_filter,
        run_id=run_id.strip(),
    )
    return {"events": list(reversed(events[-safe_limit:])), "limit": safe_limit}


def _labor_upload_limit_http_error(exc: LaborResourceLimitError) -> HTTPException:
    return HTTPException(
        status_code=413,
        detail=_labor_request_error(
            message=str(exc),
            error_code=exc.code,
            next_action="请减少文件数量、文件大小或 PDF 页数后重新上传。",
            requires_reupload=True,
        ),
    )


def _record_labor_resource_rejection(run_id: str, owner_user_id: str, code: str, *, limit: int) -> None:
    append_labor_audit_event(
        _labor_audit_path(),
        action="resource_limit_rejected",
        run_id=run_id,
        owner_user_id=owner_user_id,
        actor_user_id=owner_user_id,
        outcome="rejected",
        reason_code=code,
        details={"limit": limit},
    )


def _labor_pdf_page_count(path: Path) -> int:
    try:
        from pypdf import PdfReader

        return len(PdfReader(str(path)).pages)
    except Exception:
        try:
            return 1 if path.read_bytes().startswith(b"%PDF") else 0
        except OSError:
            return 0


@app.post("/api/labor/maintenance/cleanup")
def cleanup_labor_data(x_admin_token: str = Header(default="")) -> dict:
    expected = os.environ.get("SIGMA_LABOR_OPERATIONS_TOKEN", "").strip()
    if not expected or x_admin_token != expected:
        raise HTTPException(status_code=401, detail="缺少有效的海外劳务运维访问令牌。")
    _labor_assert_runtime_current()
    return _cleanup_expired_labor_data()


@app.get("/api/workbench/access")
def workbench_access() -> dict:
    return _workbench_access_config()


@app.post("/api/calculate")
async def calculate_bonus(
    file: UploadFile = File(...),
) -> dict:
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xlsm）。")
    if not DEFAULT_RULE_WORKBOOK.exists():
        raise HTTPException(status_code=500, detail=f"找不到规则模板：{DEFAULT_RULE_WORKBOOK}")

    upload_path = await _save_upload(file)

    try:
        rows = read_import_rows(upload_path)
        rules = load_rulebook(DEFAULT_RULE_WORKBOOK)
        result = calculate(rows, rules)
        output_path = _output_path(file.filename)
        pending_path = _output_path(file.filename, suffix="待确认表")
        build_result_workbook(result, output_path)
        if result.pending_confirmations:
            build_pending_workbook(result, pending_path)
        else:
            pending_path = None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"计算失败：{exc}") from exc
    finally:
        upload_path.unlink(missing_ok=True)

    payload = _calculation_payload(result)
    return {
        **payload,
        "downloadUrl": f"/api/download/{output_path.name}",
        "pendingDownloadUrl": f"/api/download/{pending_path.name}" if pending_path else "",
        "filename": output_path.name,
    }


@app.post("/api/runs/calculate")
async def calculate_run(
    file: UploadFile = File(...),
) -> dict:
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx 或 .xlsm）。")
    if not DEFAULT_RULE_WORKBOOK.exists():
        raise HTTPException(status_code=500, detail=f"找不到规则模板：{DEFAULT_RULE_WORKBOOK}")

    temp_upload_path = await _save_upload(file)
    try:
        rows = read_import_rows(temp_upload_path)
        rules = load_rulebook(DEFAULT_RULE_WORKBOOK)
        result = calculate(rows, rules)
        run_id = new_run_id(result.month)
        run_dir = create_run_dir(run_id)
        run_sequence = _next_recruitment_run_sequence(result.month)
        input_path = run_dir / _recruitment_output_name(result.month, run_sequence, "原始导入")
        shutil.move(str(temp_upload_path), input_path)
        output_path = run_dir / _recruitment_output_name(result.month, run_sequence, "初算结果")
        pending_path = run_dir / _recruitment_output_name(result.month, run_sequence, "待确认表")
        build_result_workbook(result, output_path)
        if result.pending_confirmations:
            build_pending_workbook(result, pending_path)
        else:
            pending_path = None
        save_table_data(run_dir, build_table_data(run_id, result))
    except ValueError as exc:
        temp_upload_path.unlink(missing_ok=True)
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        temp_upload_path.unlink(missing_ok=True)
        raise HTTPException(status_code=500, detail=f"计算失败：{exc}") from exc

    payload = _calculation_payload(result)
    status = "待确认" if payload["pendingCount"] else "已初算"
    files = {
        "input": attach_file_record(run_id, input_path, "原始导入"),
        "initialResult": attach_file_record(run_id, output_path, "初算结果"),
        "pending": attach_file_record(run_id, pending_path, "待确认表"),
    }
    metadata = save_metadata(
        run_dir,
        {
            "id": run_id,
            "month": result.month,
            "runSequence": run_sequence,
            "displayName": _recruitment_run_display_name(result.month, run_sequence, status),
            "shortCode": _recruitment_run_short_code(result.month, run_sequence),
            "status": status,
            "sourceFilename": file.filename,
            "files": files,
            "ruleInfo": rule_info(),
            **payload,
            "downloadUrl": files["initialResult"]["downloadUrl"],
            "pendingDownloadUrl": files["pending"].get("downloadUrl", ""),
        },
    )
    return metadata


@app.get("/api/runs")
def list_runs() -> dict:
    return {"runs": list_run_metadata()}


@app.get("/api/runs/{run_id}")
def get_run(run_id: str) -> dict:
    try:
        return load_metadata(get_run_dir(run_id))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="批次不存在。") from exc


@app.get("/api/runs/{run_id}/table-data")
def get_run_table_data(run_id: str) -> dict:
    try:
        return load_table_data(get_run_dir(run_id))
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="批次不存在。") from exc


@app.get("/api/labor/runs")
def list_labor_runs(request: Request, limit: int = 50) -> dict:
    bounded_limit = max(1, min(int(limit or 50), 200))
    if labor_auth_required():
        actor_user_id, is_admin = _labor_request_actor(request)
        rows = list_labor_metadata(
            limit=bounded_limit,
            owner_user_id="" if is_admin else actor_user_id,
        )
    else:
        rows = list_labor_metadata(limit=bounded_limit)
    return {
        "runs": [
            _summarize_labor_run_for_list(_normalize_labor_total_decision(row))
            for row in rows
        ]
    }


def _summarize_labor_run_for_list(row: dict) -> dict:
    return {
        "id": row.get("id") or "",
        "status": row.get("status") or "",
        "supplierName": row.get("supplierName") or "",
        "periodStart": row.get("periodStart") or "",
        "periodEnd": row.get("periodEnd") or "",
        "currency": row.get("currency") or "",
        "createdAt": row.get("createdAt") or "",
        "updatedAt": row.get("updatedAt") or "",
        "stage": row.get("stage") or "",
        "diffDownloadUrl": row.get("diffDownloadUrl") or "",
        "comparisonSummary": row.get("comparisonSummary") or {},
        "readinessGate": row.get("readinessGate") or {},
    }


def _normalize_labor_total_decision(metadata: dict) -> dict:
    normalized = _normalize_labor_review_state(metadata)
    warehouse_comparison = normalized.get("warehouseComparison")
    if not isinstance(warehouse_comparison, dict):
        return normalized
    warehouse_summary = warehouse_comparison.get("summary")
    if not isinstance(warehouse_summary, dict):
        return normalized
    if "amountDeltaTotal" not in warehouse_summary:
        return normalized

    normalized_warehouse = dict(warehouse_comparison)
    normalized_summary = dict(warehouse_summary)
    amount_delta = round(float(normalized_summary.get("amountDeltaTotal") or 0), 2)
    normalized_summary["amountDeltaTotal"] = amount_delta
    warehouse_rows = normalized_warehouse.get("rows") or []
    reconciliation_statuses = [
        str(row.get("reconciliationStatus") or "")
        for row in warehouse_rows
        if isinstance(row, dict) and row.get("reconciliationStatus")
    ]
    if "totalPassed" not in normalized_summary:
        if reconciliation_statuses:
            normalized_summary["totalPassed"] = all(status == "passed" for status in reconciliation_statuses)
        else:
            normalized_summary["totalPassed"] = amount_within_tolerance(
                amount_delta,
                AI_CONFIG["amount_tolerance"],
            )
    normalized_warehouse["summary"] = normalized_summary
    normalized["warehouseComparison"] = normalized_warehouse
    return normalized


def _normalize_labor_review_state(metadata: dict) -> dict:
    normalized = dict(metadata)
    review_status = str(normalized.get("businessReviewStatus") or "pending").strip().lower()
    if review_status not in {"pending", "approved", "rejected"}:
        review_status = "pending"
    normalized["businessReviewStatus"] = review_status
    normalized["manualReviewRequired"] = True
    normalized["directPaymentAllowed"] = False
    normalized["requiresHumanReview"] = (
        bool(normalized.get("requiresHumanReview")) if review_status == "approved" else True
    )
    return normalized


@app.get("/api/labor/suppliers")
def list_labor_suppliers(request: Request) -> dict:
    suppliers: dict[str, dict] = {}

    def add_supplier(name: str, source: str, aliases: list[str] | None = None) -> None:
        display_name = str(name or "").strip()
        if not display_name or display_name.lower() == "unknown":
            return
        if source == "history" and _looks_like_invalid_supplier_suggestion(display_name):
            return
        key = _normalize_supplier_for_profile(display_name)
        record = suppliers.setdefault(
            key,
            {
                "name": display_name,
                "sources": [],
                "aliases": [],
            },
        )
        if source and source not in record["sources"]:
            record["sources"].append(source)
        for alias in aliases or []:
            alias_value = str(alias or "").strip()
            if alias_value and alias_value not in record["aliases"] and alias_value.lower() != display_name.lower():
                record["aliases"].append(alias_value)

    history_owner = ""
    if labor_auth_required():
        actor_user_id, is_admin = _labor_request_actor(request)
        history_owner = "" if is_admin else actor_user_id
    history_runs = list_labor_metadata(owner_user_id=history_owner)
    for run in history_runs:
        add_supplier(str(run.get("supplierName") or run.get("supplier") or ""), "history")

    profiles_path = AI_CONFIG.get("supplier_profiles_path")
    if profiles_path:
        path = Path(profiles_path)
        profile_files = sorted(path.glob("*.json")) if path.is_dir() else ([path] if path.is_file() else [])
        for profile_file in profile_files:
            try:
                for profile in load_supplier_profiles(profile_file):
                    add_supplier(profile.key, "profile", profile.aliases)
            except (OSError, ValueError, json.JSONDecodeError) as exc:
                logger.warning("供应商建议加载失败，跳过: %s, error=%s", profile_file, exc)

    values = sorted(
        suppliers.values(),
        key=lambda row: (0 if "history" in row["sources"] else 1, row["name"].lower()),
    )
    return {"suppliers": values}


def _looks_like_invalid_supplier_suggestion(value: str) -> bool:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return True
    if re.fullmatch(r"[\d\s_\-#]+", normalized):
        return True
    if re.search(r"\d+\s*仓", normalized) or re.fullmatch(r"仓库\s*\d+", normalized):
        return True
    if any(token in normalized for token in ("验证", "测试", "test", "demo")):
        return True
    return False


LABOR_TELEMETRY_SUMMARY_FIELDS = {
    "cacheHit",
    "pdfAmountTotal",
    "excelAmountTotal",
    "amountDeltaTotal",
    "hoursDeltaTotal",
    "exceptionCount",
    "amountDiffCount",
    "hoursRiskCount",
    "lowConfidenceCount",
    "candidateMatchCount",
    "notInInvoiceCount",
    "pdfEmployeeCount",
    "excelEmployeeCount",
    "warehouseExceptionCount",
    "governanceCandidateCount",
    "readinessStatus",
}
LABOR_TELEMETRY_CONTEXT_FIELDS = {
    "batchKey",
    "fileCount",
    "pdfCount",
    "workbookCount",
    "sheetName",
    "mappingFields",
    "button",
    "path",
}


def _telemetry_text(value, max_length: int = 240) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[\r\n\t]+", " ", text)
    return text[:max_length]


def _telemetry_number(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not number == number or number in (float("inf"), float("-inf")):
        return None
    return round(number, 4)


def _sanitize_labor_telemetry(payload: dict) -> dict:
    if not isinstance(payload, dict):
        payload = {}
    event = {
        "schemaVersion": LABOR_TELEMETRY_SCHEMA_VERSION,
        "createdAt": utcnow_naive().isoformat(timespec="seconds") + "Z",
        "source": "overseas-labor-ui",
        "event": _telemetry_text(payload.get("event") or payload.get("eventType"), 96),
        "runId": _telemetry_text(payload.get("runId"), 96),
        "supplier": _telemetry_text(payload.get("supplier"), 160),
        "step": _telemetry_text(payload.get("step"), 64),
        "status": _telemetry_text(payload.get("status"), 64),
        "errorCode": _telemetry_text(payload.get("errorCode"), 80),
        "errorMessage": _telemetry_text(payload.get("errorMessage"), 300),
        "summary": {},
        "context": {},
    }
    duration_ms = _telemetry_number(payload.get("durationMs"))
    if duration_ms is not None:
        event["durationMs"] = max(0, round(duration_ms))

    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    for key in LABOR_TELEMETRY_SUMMARY_FIELDS:
        value = summary.get(key)
        if value is None:
            continue
        if isinstance(value, bool):
            event["summary"][key] = value
        elif isinstance(value, (int, float)):
            event["summary"][key] = _telemetry_number(value)
        elif isinstance(value, (str, bool)):
            event["summary"][key] = _telemetry_text(value, 120)

    context = payload.get("context") if isinstance(payload.get("context"), dict) else {}
    for key in LABOR_TELEMETRY_CONTEXT_FIELDS:
        value = context.get(key)
        if value is None:
            continue
        if isinstance(value, bool):
            event["context"][key] = value
        elif isinstance(value, (int, float)):
            event["context"][key] = _telemetry_number(value)
        elif isinstance(value, list):
            event["context"][key] = [_telemetry_text(item, 80) for item in value[:12]]
        elif isinstance(value, (str, bool)):
            event["context"][key] = _telemetry_text(value, 160)
    return event


@app.post("/api/labor/telemetry")
def record_labor_telemetry(payload: dict = Body(default={})) -> dict:
    event = _sanitize_labor_telemetry(payload)
    if not event["event"]:
        raise HTTPException(status_code=400, detail="缺少事件名称。")
    LABOR_TELEMETRY_DIR.mkdir(parents=True, exist_ok=True)
    with LABOR_TELEMETRY_FILE.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(event, ensure_ascii=False, separators=(",", ":")) + "\n")
    return {"ok": True}


@app.get("/api/labor/telemetry/export")
def export_labor_telemetry() -> FileResponse:
    if not LABOR_TELEMETRY_FILE.exists():
        raise HTTPException(status_code=404, detail="暂无试用操作数据。")
    return FileResponse(
        LABOR_TELEMETRY_FILE,
        media_type="application/x-ndjson",
        filename="overseas_labor_telemetry.jsonl",
    )


def _labor_worker_identity(authorization: str, *, worker_version: str = "") -> dict[str, str]:
    prefix = "Bearer "
    token = authorization[len(prefix) :].strip() if authorization.startswith(prefix) else ""
    if labor_postgres_state_enabled():
        try:
            identity = resolve_labor_worker_token(
                token,
                worker_version=worker_version,
                refresh_ttl_seconds=_labor_worker_token_ttl_seconds(),
                recovery_grace_seconds=_labor_worker_token_recovery_grace_seconds(),
            )
            return {"userId": str(identity["userId"]), "deviceId": str(identity["deviceId"])}
        except LaborWorkerIdentityInvalid as exc:
            raise HTTPException(status_code=401, detail="Worker 身份令牌无效或已失效。") from exc
        except Exception as exc:  # noqa: BLE001 - backend details stay in server logs.
            raise HTTPException(
                status_code=503,
                detail=_labor_request_error(
                    message="Worker 身份服务暂不可用。",
                    error_code="LABOR_WORKER_IDENTITY_UNAVAILABLE",
                    retryable=True,
                    next_action="请稍后重新连接；持续失败请联系管理员。",
                ),
            ) from exc
    try:
        configured = json.loads(os.environ.get("SIGMA_LABOR_WORKER_TOKENS", "{}"))
    except json.JSONDecodeError:
        configured = {}
    identity = configured.get(token) if token and isinstance(configured, dict) else None
    if not isinstance(identity, dict) or not identity.get("userId") or not identity.get("deviceId"):
        raise HTTPException(status_code=401, detail="Worker 身份令牌无效或已失效。")
    return {"userId": str(identity["userId"]), "deviceId": str(identity["deviceId"])}


def _labor_worker_token_ttl_seconds() -> int:
    try:
        value = int(str(os.environ.get("SIGMA_LABOR_WORKER_TOKEN_TTL_SECONDS") or 90 * 24 * 60 * 60))
    except (TypeError, ValueError):
        value = 90 * 24 * 60 * 60
    return max(300, min(value, 365 * 24 * 60 * 60))


def _labor_worker_token_recovery_grace_seconds() -> int:
    try:
        value = int(
            str(
                os.environ.get("SIGMA_LABOR_WORKER_TOKEN_RECOVERY_GRACE_SECONDS")
                or 7 * 24 * 60 * 60
            )
        )
    except (TypeError, ValueError):
        value = 7 * 24 * 60 * 60
    return max(0, min(value, 30 * 24 * 60 * 60))


def _labor_worker_activation_url(request: Request, activation_code: str) -> str:
    configured = str(os.environ.get("SIGMA_LABOR_PUBLIC_URL") or "").strip().rstrip("/")
    candidate = configured or str(request.base_url).strip().rstrip("/")
    parsed = urlparse(candidate)
    is_secure = parsed.scheme == "https" or (
        parsed.scheme == "http" and parsed.hostname in {"localhost", "127.0.0.1"}
    )
    if not is_secure or not parsed.netloc or parsed.path not in {"", "/"}:
        raise HTTPException(
            status_code=503,
            detail=_labor_request_error(
                message="Worker 公网服务地址未安全配置。",
                error_code="LABOR_WORKER_PUBLIC_URL_INVALID",
                next_action="请由管理员配置 HTTPS 的 SIGMA_LABOR_PUBLIC_URL。",
            ),
        )
    api_url = f"{parsed.scheme}://{parsed.netloc}"
    return f"sigma-overseas-labor-worker://activate?{urlencode({'apiUrl': api_url, 'code': activation_code})}"


def _labor_worker_device_http_error(exc: Exception) -> HTTPException:
    if isinstance(exc, LaborWorkerDeviceNotFound):
        return HTTPException(status_code=404, detail="Worker 设备不存在或已撤销。")
    if isinstance(exc, LaborWorkerIdentityInvalid):
        return HTTPException(status_code=401, detail="Worker 身份令牌无效或已失效。")
    if isinstance(exc, (LaborWorkerIdentityError, ValueError)):
        return HTTPException(status_code=409, detail=str(exc))
    return HTTPException(
        status_code=503,
        detail=_labor_request_error(
            message="Worker 设备身份服务暂不可用。",
            error_code="LABOR_WORKER_IDENTITY_UNAVAILABLE",
            retryable=True,
            next_action="请稍后重试；持续失败请联系管理员。",
        ),
    )


@app.get("/api/labor/worker/devices")
def get_current_labor_worker_devices(request: Request) -> dict:
    _labor_p1_state_required()
    owner_user_id, _ = _labor_request_actor(request)
    try:
        devices = list_labor_worker_devices(owner_user_id=owner_user_id)
    except Exception as exc:  # noqa: BLE001
        raise _labor_worker_device_http_error(exc) from exc
    return {"devices": devices}


@app.post("/api/labor/worker/devices")
def activate_current_labor_worker_device(request: Request, payload: dict = Body(default={})) -> dict:
    _labor_p1_state_required()
    owner_user_id, _ = _labor_request_actor(request)
    display_name = str(payload.get("displayName") or "个人核对助手").strip()[:120]
    platform = str(payload.get("platform") or "macos-arm64").strip()[:80]
    worker_version = str(payload.get("workerVersion") or "").strip()[:40]
    try:
        issued = issue_labor_worker_activation(
            owner_user_id=owner_user_id,
            actor_user_id=owner_user_id,
            display_name=display_name,
            platform=platform,
            worker_version=worker_version,
            ttl_seconds=5 * 60,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_worker_device_http_error(exc) from exc
    return {
        "device": issued["device"],
        "expiresIn": issued["expiresIn"],
        "expiresAt": issued["expiresAt"],
        "activationUrl": _labor_worker_activation_url(request, str(issued["activationCode"])),
    }


@app.post("/api/labor/worker/devices/{device_id}/rotate")
def rotate_current_labor_worker_device(
    request: Request,
    device_id: str,
    payload: dict = Body(default={}),
) -> dict:
    _labor_p1_state_required()
    owner_user_id, _ = _labor_request_actor(request)
    display_name = str(payload.get("displayName") or "个人核对助手").strip()[:120]
    platform = str(payload.get("platform") or "macos-arm64").strip()[:80]
    worker_version = str(payload.get("workerVersion") or "").strip()[:40]
    try:
        issued = issue_labor_worker_activation(
            owner_user_id=owner_user_id,
            actor_user_id=owner_user_id,
            display_name=display_name,
            platform=platform,
            worker_version=worker_version,
            ttl_seconds=5 * 60,
            device_id=device_id,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_worker_device_http_error(exc) from exc
    return {
        "device": issued["device"],
        "expiresIn": issued["expiresIn"],
        "expiresAt": issued["expiresAt"],
        "activationUrl": _labor_worker_activation_url(request, str(issued["activationCode"])),
    }


@app.post("/api/labor/worker/activate")
def exchange_current_labor_worker_activation(
    response: Response,
    payload: dict = Body(default={}),
) -> dict:
    _labor_p1_state_required()
    activation_code = str(payload.get("activationCode") or "").strip()[:256]
    worker_version = str(payload.get("workerVersion") or "").strip()[:40]
    try:
        issued = exchange_labor_worker_activation(
            activation_code,
            worker_version=worker_version,
            ttl_seconds=_labor_worker_token_ttl_seconds(),
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_worker_device_http_error(exc) from exc
    response.headers["Cache-Control"] = "no-store"
    response.headers["Pragma"] = "no-cache"
    return issued


@app.delete("/api/labor/worker/devices/{device_id}")
def revoke_current_labor_worker_device(request: Request, device_id: str) -> dict:
    _labor_p1_state_required()
    owner_user_id, _ = _labor_request_actor(request)
    try:
        device = revoke_labor_worker_device(
            owner_user_id=owner_user_id,
            actor_user_id=owner_user_id,
            device_id=device_id,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_worker_device_http_error(exc) from exc
    return {"device": device, "revoked": True}


def _public_labor_worker_job(job: dict | None) -> dict | None:
    if not job:
        return None
    allowed = {
        "id", "runId", "jobType", "status", "attempt", "maxAttempts", "availableAt",
        "leaseExpiresAt", "heartbeatAt", "requiredWorkerVersion", "progress",
        "errorCode", "errorMessage", "createdAt", "updatedAt", "startedAt", "finishedAt",
        "taskGenerationId",
    }
    return {key: value for key, value in job.items() if key in allowed}


def _labor_assert_worker_version(worker_version: str, job: dict | None = None) -> str:
    current_required = _labor_required_worker_version()
    job_required = str((job or {}).get("requiredWorkerVersion") or "").strip()
    required_version = current_required
    if job_required and not worker_version_at_least(current_required, job_required):
        required_version = job_required
    if not worker_version_at_least(worker_version, required_version):
        raise HTTPException(
            status_code=426,
            detail={
                **_labor_request_error(
                    message=f"当前 Worker 版本无效或低于最低要求 {required_version}。",
                    error_code="LABOR_WORKER_UPGRADE_REQUIRED",
                    next_action="请先升级海外报账核对助手，再重新连接。",
                ),
                "requiredWorkerVersion": required_version,
            },
        )
    return required_version


def _labor_operations_events() -> list[dict]:
    if not LABOR_TELEMETRY_FILE.exists():
        return []
    rows = []
    for line in LABOR_TELEMETRY_FILE.read_text(encoding="utf-8").splitlines()[-5000:]:
        try:
            event = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(event, dict):
            rows.append(event)
    return rows


@app.get("/api/labor/operations")
def get_labor_operations(x_admin_token: str = Header(default="")) -> dict:
    expected = os.environ.get("SIGMA_LABOR_OPERATIONS_TOKEN", "").strip()
    if not expected or x_admin_token != expected:
        raise HTTPException(status_code=401, detail="缺少有效的海外劳务运维访问令牌。")
    storage_backend = labor_storage_backend() or "local"
    disk = shutil.disk_usage(LABOR_RUNS_DIR.parent)
    local_capacity_relevant = storage_backend == "local"
    snapshot = build_labor_operations_snapshot(
        list_labor_worker_jobs(),
        _labor_operations_events(),
        storage={
            "backend": storage_backend,
            "freeBytes": disk.free,
            "totalBytes": disk.total,
            "minimumFreeBytes": (
                int(os.environ.get("LABOR_MINIMUM_FREE_BYTES", 2 * 1024**3))
                if local_capacity_relevant
                else 0
            ),
            "capacityScope": "persistent_storage" if local_capacity_relevant else "server_cache_only",
        },
    )
    snapshot["recentJobs"] = [_public_labor_worker_job(job) for job in snapshot["recentJobs"]]
    return snapshot


@app.post("/api/labor/worker/jobs/claim")
def claim_personal_labor_worker_job(
    authorization: str = Header(default=""),
    x_worker_version: str = Header(default=""),
) -> dict:
    identity = _labor_worker_identity(authorization, worker_version=x_worker_version)
    _labor_assert_runtime_current()
    _labor_assert_worker_version(x_worker_version)
    job = claim_labor_worker_job(
        owner_user_id=identity["userId"],
        device_id=identity["deviceId"],
        worker_version=x_worker_version,
    )
    if job and str(job.get("jobType") or "reconcile") == "mapping_preflight":
        generation = str(job.get("taskGenerationId") or "").strip()
        update_labor_metadata_for_mapping_preflight(
            str(job.get("runId") or ""),
            expected_task_generation_id=generation,
            updates=lambda current: {
                "mappingPreflight": {
                    **(current.get("mappingPreflight") if isinstance(current.get("mappingPreflight"), dict) else {}),
                    "status": "running",
                    "statusLabel": "本人核对助手正在读取 Excel",
                    "message": "正在读取工作表、列名和样例数据。",
                    "workerDeviceId": identity["deviceId"],
                }
            },
            actor_user_id=identity["userId"],
            action="mapping_preflight_claimed",
            reason_code="mapping_preflight_worker_claimed",
        )
    return {"job": _public_labor_worker_job(job)}


_LABOR_WORKER_RELEASE_PLATFORMS = {"macos-arm64", "windows-x64"}
_LABOR_WORKER_RELEASE_MANIFEST_FILENAME = "manifest.json"


def _labor_worker_release_platform(value: str) -> str:
    platform = str(value or "macos-arm64").strip().lower()
    if platform not in _LABOR_WORKER_RELEASE_PLATFORMS:
        raise HTTPException(status_code=400, detail="核对助手平台仅支持 macos-arm64 或 windows-x64。")
    return platform


def _labor_worker_release_manifest_path() -> str:
    return (
        f"labor-runs/{labor_persistent_environment()}/owners/system/"
        f"worker-releases/{_LABOR_WORKER_RELEASE_MANIFEST_FILENAME}"
    )


def _load_persisted_labor_worker_release_manifest() -> dict:
    pathname = _labor_worker_release_manifest_path()
    try:
        if labor_blob_signed_urls_enabled():
            content = blob_get_bytes(pathname)
        elif labor_persistent_storage_enabled():
            content = get_labor_supabase_private_object(pathname)
        else:
            return {}
        parsed = json.loads(content.decode("utf-8")) if content else {}
        return parsed if isinstance(parsed, dict) else {}
    except Exception as exc:  # noqa: BLE001 - an unavailable manifest falls back to the deployment baseline.
        logger.warning("labor worker release manifest read failed: %s", type(exc).__name__)
        return {}


def _persist_labor_worker_release_manifest(manifest: dict) -> None:
    payload = json.dumps(manifest, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    pathname = _labor_worker_release_manifest_path()
    if labor_blob_signed_urls_enabled():
        blob_put_bytes(pathname, payload, content_type="application/json")
        return
    if labor_persistent_storage_enabled():
        put_labor_supabase_private_object(pathname, payload, content_type="application/json")
        return
    raise RuntimeError("未配置核对助手发布清单的持久化存储。")


def _labor_worker_release_catalog() -> dict[str, dict]:
    try:
        manifest = json.loads(os.environ.get("SIGMA_LABOR_WORKER_UPDATE_MANIFEST", "{}"))
    except json.JSONDecodeError:
        manifest = {}
    if not isinstance(manifest, dict):
        manifest = {}
    releases = manifest.get("releases")
    if isinstance(releases, dict):
        catalog = {
            platform: dict(release)
            for platform, release in releases.items()
            if platform in _LABOR_WORKER_RELEASE_PLATFORMS and isinstance(release, dict)
        }
    else:
        # Version 1 manifests represented the original macOS ARM64 release directly.
        catalog = {"macos-arm64": manifest} if manifest else {}
    persisted = _load_persisted_labor_worker_release_manifest()
    persisted_releases = persisted.get("releases") if isinstance(persisted, dict) else {}
    if isinstance(persisted_releases, dict):
        catalog.update(
            {
                platform: dict(release)
                for platform, release in persisted_releases.items()
                if platform in _LABOR_WORKER_RELEASE_PLATFORMS and isinstance(release, dict)
            }
        )
    return catalog


def _labor_worker_release_config(platform: str = "macos-arm64") -> dict:
    manifest = _labor_worker_release_catalog().get(_labor_worker_release_platform(platform), {})
    return {
        key: manifest[key]
        for key in (
            "version",
            "url",
            "sha256",
            "signature",
            "minimumVersion",
            "objectKey",
            "blobPathname",
            "filename",
            "sizeBytes",
        )
        if key in manifest
    }


def _labor_worker_release_filename(platform: str, version: str) -> str:
    if platform == "windows-x64":
        return f"Σ海外报账核对助手-{version}-windows-x64.exe"
    return f"Σ海外报账核对助手-{version}-arm64.dmg"


def _labor_worker_release_content_types(platform: str) -> list[str]:
    if platform == "windows-x64":
        return ["application/x-msdownload", "application/vnd.microsoft.portable-executable", "application/octet-stream"]
    return ["application/x-apple-diskimage", "application/octet-stream"]


def _labor_public_worker_release(platform: str = "macos-arm64") -> dict:
    platform = _labor_worker_release_platform(platform)
    manifest = _labor_worker_release_config(platform)
    version = str(manifest.get("version") or "").strip()
    url = str(manifest.get("url") or "").strip()
    object_key = str(manifest.get("objectKey") or "").strip()
    blob_pathname = str(manifest.get("blobPathname") or "").strip()
    filename = str(manifest.get("filename") or _labor_worker_release_filename(platform, version)).strip()
    try:
        size_bytes = max(0, int(manifest.get("sizeBytes") or 0))
    except (TypeError, ValueError):
        size_bytes = 0
    complete = bool(
        version
        and str(manifest.get("sha256") or "").strip()
        and str(manifest.get("signature") or "").strip()
        and (object_key or blob_pathname or url.startswith("https://"))
    )
    return {
        "available": complete,
        "platform": platform,
        "storageEnvironment": labor_persistent_environment(),
        "version": version,
        "minimumVersion": str(manifest.get("minimumVersion") or _labor_required_worker_version()),
        "requiredWorkerVersion": _labor_required_worker_version(),
        "sha256": str(manifest.get("sha256") or ""),
        "signature": str(manifest.get("signature") or ""),
        "filename": filename[:180],
        "sizeBytes": size_bytes,
        "downloadUrl": (
            (
                "/api/labor/worker/release/download"
                if platform == "macos-arm64"
                else f"/api/labor/worker/release/download?platform={platform}"
            )
            if object_key or blob_pathname
            else url
        ),
    }


@app.get("/api/labor/worker/release")
def get_personal_labor_worker_release(request: Request, platform: str = "macos-arm64") -> dict:
    is_admin = False
    if labor_auth_required():
        _, is_admin = _labor_request_actor(request)
    platform = _labor_worker_release_platform(platform)
    release = _labor_public_worker_release(platform)
    release["canUpload"] = bool(is_admin)
    if is_admin:
        persisted = _load_persisted_labor_worker_release_manifest()
        pending = persisted.get("pendingReleases") if isinstance(persisted, dict) else {}
        pending = pending if isinstance(pending, dict) else {}
        selected = pending.get(platform) if isinstance(pending.get(platform), dict) else {}
        pending_version = str(selected.get("version") or "")
        matching = sorted(
            candidate
            for candidate in _LABOR_WORKER_RELEASE_PLATFORMS
            if isinstance(pending.get(candidate), dict)
            and str(pending[candidate].get("version") or "") == pending_version
        ) if pending_version else []
        release.update(
            {
                "pendingVersion": pending_version,
                "pendingVersions": {
                    candidate: str(candidate_release.get("version") or "")
                    for candidate, candidate_release in pending.items()
                    if candidate in _LABOR_WORKER_RELEASE_PLATFORMS
                    and isinstance(candidate_release, dict)
                    and str(candidate_release.get("version") or "")
                },
                "pendingPlatforms": matching,
                "missingPlatforms": sorted(_LABOR_WORKER_RELEASE_PLATFORMS.difference(matching)),
            }
        )
    return release


@app.get("/api/labor/worker/release/download")
def download_personal_labor_worker_release(request: Request, platform: str = "macos-arm64") -> Response:
    if labor_auth_required():
        _labor_request_actor(request)
    platform = _labor_worker_release_platform(platform)
    manifest = _labor_worker_release_config(platform)
    release = _labor_public_worker_release(platform)
    if not release["available"]:
        raise HTTPException(status_code=503, detail="核对助手安装包尚未配置，请联系管理员。")
    object_key = str(manifest.get("objectKey") or "").strip()
    blob_pathname = str(manifest.get("blobPathname") or "").strip()
    if blob_pathname:
        try:
            signed_url = create_labor_blob_presigned_url(blob_pathname, operation="get", expires_in=7200)
        except Exception as exc:  # noqa: BLE001 - keep storage internals out of browser errors.
            logger.warning("labor worker Blob release signing failed: %s", type(exc).__name__)
            raise HTTPException(status_code=503, detail="核对助手安装包暂时无法下载，请稍后重试。") from exc
        return RedirectResponse(
            url=signed_url,
            status_code=307,
            headers={"cache-control": "no-store", "referrer-policy": "no-referrer"},
        )
    if object_key:
        try:
            signed = create_labor_supabase_signed_download(
                object_key,
                filename=str(release.get("filename") or "worker.dmg"),
                expires_in=7200,
            )
        except Exception as exc:  # noqa: BLE001 - keep storage internals out of browser errors.
            logger.warning("labor worker release signing failed: %s", type(exc).__name__)
            raise HTTPException(status_code=503, detail="核对助手安装包暂时无法下载，请稍后重试。") from exc
        return RedirectResponse(url=str(signed["signedUrl"]), status_code=307)
    return RedirectResponse(url=str(release["downloadUrl"]), status_code=307)


@app.post("/api/labor/worker/release/upload-intent")
def create_personal_labor_worker_release_upload_intent(request: Request, payload: dict = Body(...)) -> dict:
    _, is_admin = _labor_request_actor(request)
    if not is_admin:
        raise HTTPException(status_code=403, detail="仅系统管理员可以上传核对助手安装包。")
    platform = _labor_worker_release_platform(str(payload.get("platform") or "macos-arm64"))
    version = str(payload.get("version") or "").strip()
    try:
        parse_stable_worker_version(version)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    required_version = _labor_required_worker_version()
    if not worker_version_at_least(version, required_version):
        raise HTTPException(status_code=409, detail=f"安装包版本不能低于当前最低要求 {required_version}。")
    filename = str(payload.get("filename") or "").strip()
    expected_filename = _labor_worker_release_filename(platform, version)
    if filename != expected_filename:
        raise HTTPException(status_code=400, detail=f"安装包文件名必须是 {expected_filename}。")
    try:
        size_bytes = int(payload.get("sizeBytes") or 0)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="安装包大小格式无效。") from exc
    if size_bytes <= 0 or size_bytes > 300 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="核对助手安装包必须小于 300 MB。")
    sha256 = str(payload.get("sha256") or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", sha256):
        raise HTTPException(status_code=400, detail="安装包 SHA-256 格式无效。")
    object_key = f"labor-runs/{labor_persistent_environment()}/owners/system/worker-releases/{platform}/{filename}"
    content_types = _labor_worker_release_content_types(platform)
    content_type = content_types[0]
    if labor_blob_signed_urls_enabled():
        try:
            signed_url = create_labor_blob_presigned_url(
                object_key,
                operation="put",
                expires_in=7200,
                maximum_size_in_bytes=size_bytes,
                allowed_content_types=content_types,
            )
        except Exception as exc:  # noqa: BLE001 - keep private storage details out of the response.
            logger.warning("labor worker Blob release upload intent failed: %s", type(exc).__name__)
            raise HTTPException(status_code=503, detail="暂时无法创建安装包上传通道，请稍后重试。") from exc
        return {
            "signedUrl": signed_url,
            "method": "PUT",
            "headers": {"content-type": content_type},
            "blobPathname": object_key,
            "platform": platform,
            "expiresIn": 7200,
            "version": version,
            "filename": filename,
            "sizeBytes": size_bytes,
            "sha256": sha256,
            "private": True,
        }
    try:
        intent = create_labor_supabase_signed_upload_for_object(
            object_key,
            file_kind="worker_release",
            content_type=content_type,
        )
    except Exception as exc:  # noqa: BLE001 - keep private storage details out of the response.
        logger.warning("labor worker release upload intent failed: %s", type(exc).__name__)
        raise HTTPException(status_code=503, detail="暂时无法创建安装包上传通道，请稍后重试。") from exc
    return {
        "signedUrl": intent["signedUrl"],
        "method": "PUT",
        "headers": intent.get("headers") or {"content-type": content_type},
        "objectKey": object_key,
        "platform": platform,
        "expiresIn": int(intent.get("expiresIn") or 7200),
        "version": version,
        "filename": filename,
        "sizeBytes": size_bytes,
        "sha256": sha256,
        "private": True,
    }


def _verify_labor_worker_release_artifact(
    *,
    platform: str,
    version: str,
    filename: str,
    size_bytes: int,
) -> dict:
    object_key = (
        f"labor-runs/{labor_persistent_environment()}/owners/system/"
        f"worker-releases/{platform}/{filename}"
    )
    if labor_blob_signed_urls_enabled():
        matches = [
            entry
            for entry in blob_list_prefix(object_key)
            if str(entry.get("pathname") or "") == object_key
        ]
        if not matches:
            raise HTTPException(status_code=409, detail="私有存储中未找到刚上传的安装包，请重新上传。")
        observed_size = int(matches[0].get("size") or matches[0].get("sizeBytes") or 0)
        location = {"blobPathname": object_key}
    elif labor_persistent_storage_enabled():
        observed = labor_supabase_object_metadata(object_key)
        observed_size = int(observed.get("sizeBytes") or 0)
        location = {"objectKey": object_key}
    else:
        raise RuntimeError("未配置核对助手安装包的私有持久化存储。")
    if observed_size != size_bytes:
        raise HTTPException(
            status_code=409,
            detail=f"安装包大小与上传前不一致（预期 {size_bytes}，实际 {observed_size}），请重新上传。",
        )
    return {**location, "sizeBytes": observed_size}


@app.post("/api/labor/worker/release/finalize")
def finalize_personal_labor_worker_release(request: Request, payload: dict = Body(...)) -> dict:
    _, is_admin = _labor_request_actor(request)
    if not is_admin:
        raise HTTPException(status_code=403, detail="仅系统管理员可以发布核对助手安装包。")
    platform = _labor_worker_release_platform(str(payload.get("platform") or "macos-arm64"))
    version = str(payload.get("version") or "").strip()
    try:
        parse_stable_worker_version(version)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    required_version = _labor_required_worker_version()
    if not worker_version_at_least(version, required_version):
        raise HTTPException(status_code=409, detail=f"安装包版本不能低于当前最低要求 {required_version}。")
    filename = str(payload.get("filename") or "").strip()
    expected_filename = _labor_worker_release_filename(platform, version)
    if filename != expected_filename:
        raise HTTPException(status_code=400, detail=f"安装包文件名必须是 {expected_filename}。")
    try:
        size_bytes = int(payload.get("sizeBytes") or 0)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="安装包大小格式无效。") from exc
    if size_bytes <= 0 or size_bytes > 300 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="核对助手安装包必须小于 300 MB。")
    sha256 = str(payload.get("sha256") or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", sha256):
        raise HTTPException(status_code=400, detail="安装包 SHA-256 格式无效。")
    try:
        artifact = _verify_labor_worker_release_artifact(
            platform=platform,
            version=version,
            filename=filename,
            size_bytes=size_bytes,
        )
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001 - keep storage internals out of the response.
        logger.warning("labor worker release verification failed: %s", type(exc).__name__)
        raise HTTPException(status_code=503, detail="暂时无法确认安装包上传结果，请稍后重试。") from exc

    existing = _load_persisted_labor_worker_release_manifest()
    releases = {
        release_platform: dict(release)
        for release_platform, release in (existing.get("releases") or {}).items()
        if release_platform in _LABOR_WORKER_RELEASE_PLATFORMS and isinstance(release, dict)
    } if isinstance(existing, dict) and isinstance(existing.get("releases"), dict) else {}
    pending_releases = {
        release_platform: dict(release)
        for release_platform, release in (existing.get("pendingReleases") or {}).items()
        if release_platform in _LABOR_WORKER_RELEASE_PLATFORMS and isinstance(release, dict)
    } if isinstance(existing, dict) and isinstance(existing.get("pendingReleases"), dict) else {}
    published_at = utcnow_naive().isoformat(timespec="seconds") + "Z"
    release = {
        "version": version,
        "minimumVersion": version,
        "sha256": sha256,
        "signature": f"sha256:{sha256}",
        "filename": filename,
        "sizeBytes": int(artifact["sizeBytes"]),
        "publishedAt": published_at,
        **{
            key: str(artifact[key])
            for key in ("objectKey", "blobPathname")
            if artifact.get(key)
        },
    }
    pending_releases[platform] = release
    matching_platforms = {
        candidate
        for candidate in _LABOR_WORKER_RELEASE_PLATFORMS
        if str((pending_releases.get(candidate) or {}).get("version") or "") == version
    }
    missing_platforms = sorted(_LABOR_WORKER_RELEASE_PLATFORMS.difference(matching_platforms))
    published = not missing_platforms
    active_required_version = str(existing.get("requiredWorkerVersion") or required_version)
    if published:
        for candidate in sorted(_LABOR_WORKER_RELEASE_PLATFORMS):
            releases[candidate] = pending_releases.pop(candidate)
        active_required_version = version
    manifest = {
        "schemaVersion": 4,
        "requiredWorkerVersion": active_required_version,
        "publishedAt": published_at if published else str(existing.get("publishedAt") or ""),
        "updatedAt": published_at,
        "releases": releases,
        "pendingReleases": pending_releases,
    }
    try:
        _persist_labor_worker_release_manifest(manifest)
    except Exception as exc:  # noqa: BLE001 - a failed manifest write must not activate the release.
        logger.warning("labor worker release manifest publish failed: %s", type(exc).__name__)
        raise HTTPException(status_code=503, detail="安装包已上传，但发布清单写入失败；旧版本仍然有效，请重试发布。") from exc
    return {
        "available": published,
        "published": published,
        "platform": platform,
        "storageEnvironment": labor_persistent_environment(),
        "version": version,
        "minimumVersion": active_required_version,
        "requiredWorkerVersion": active_required_version,
        "pendingPlatforms": sorted(matching_platforms),
        "missingPlatforms": missing_platforms,
        "sha256": sha256,
        "signature": f"sha256:{sha256}",
        "filename": filename,
        "downloadUrl": (
            (
                "/api/labor/worker/release/download"
                if platform == "macos-arm64"
                else f"/api/labor/worker/release/download?platform={platform}"
            )
            if published
            else ""
        ),
    }


@app.get("/api/labor/worker/version")
def get_personal_labor_worker_version(
    currentVersion: str = "",
    platform: str = "macos-arm64",
    authorization: str = Header(default=""),
) -> dict:
    _labor_worker_identity(authorization, worker_version=currentVersion)
    platform = _labor_worker_release_platform(platform)
    release = _labor_public_worker_release(platform)
    version = str(release.get("version") or "")
    allowed = {
        key: release[key]
        for key in ("version", "sha256", "signature", "minimumVersion", "downloadUrl")
        if release.get(key)
    }
    required_version = _labor_required_worker_version()
    complete_manifest = release.get("available") is True
    try:
        manifest_version = parse_stable_worker_version(version)
        current_version = parse_stable_worker_version(currentVersion)
    except ValueError:
        manifest_version = None
        current_version = None
    allowed["requiredWorkerVersion"] = required_version
    allowed["platform"] = platform
    allowed["upgradeRequired"] = not worker_version_at_least(currentVersion, required_version)
    allowed["updateAvailable"] = bool(
        complete_manifest
        and manifest_version is not None
        and current_version is not None
        and manifest_version > current_version
    )
    return allowed


@app.get("/api/labor/worker/health")
def get_personal_labor_worker_health(x_admin_token: str = Header(default="")) -> dict:
    expected = os.environ.get("SIGMA_LABOR_OPERATIONS_TOKEN", "").strip()
    if not expected or x_admin_token != expected:
        raise HTTPException(status_code=401, detail="缺少有效的海外劳务运维访问令牌。")
    return labor_p1_worker_job_store_health()


@app.get("/api/labor/production-readiness")
def get_labor_production_readiness(x_admin_token: str = Header(default="")) -> dict:
    expected = os.environ.get("SIGMA_LABOR_OPERATIONS_TOKEN", "").strip()
    if not expected or x_admin_token != expected:
        raise HTTPException(status_code=401, detail="缺少有效的海外劳务运维访问令牌。")
    return _labor_p1_readiness_snapshot(force_refresh=True)


@app.post("/api/labor/worker/jobs/{job_id}/heartbeat")
def heartbeat_personal_labor_worker_job(
    job_id: str,
    payload: dict = Body(default={}),
    authorization: str = Header(default=""),
) -> dict:
    try:
        identity = _labor_worker_identity(authorization)
        existing = get_labor_worker_job(job_id)
        if str(existing.get("jobType") or "reconcile") == "overseas_payroll":
            _, job, _ = _overseas_payroll_worker_context(
                job_id,
                authorization,
                progress=payload.get("progress") if isinstance(payload, dict) else None,
            )
            return {"job": _public_labor_worker_job(job)}
        if (
            str(existing.get("status") or "") == "succeeded"
            and str(existing.get("ownerUserId") or "") == identity["userId"]
            and str(existing.get("claimedDeviceId") or "") == identity["deviceId"]
        ):
            return {"job": _public_labor_worker_job(existing)}
        with _leased_worker_job(
            job_id,
            authorization,
            progress=payload.get("progress") if isinstance(payload, dict) else None,
            identity=identity,
        ) as (_, job, _, generation):
            if str(job.get("jobType") or "reconcile") != "mapping_preflight":
                heartbeat_at = str(job.get("heartbeatAt") or utcnow_naive().isoformat())
                _update_worker_generation_metadata(
                    str(job.get("runId") or ""),
                    generation,
                    lambda current: {
                        "progress": {
                            **(current.get("progress") if isinstance(current.get("progress"), dict) else {}),
                            "lastUpdatedAt": heartbeat_at,
                        }
                    },
                )
            return {"job": _public_labor_worker_job(job)}
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


def _overseas_payroll_worker_context(
    job_id: str,
    authorization: str,
    *,
    progress: dict | None = None,
) -> tuple[dict[str, str], dict, str]:
    identity = _labor_worker_identity(authorization)
    existing = get_labor_worker_job(job_id)
    if str(existing.get("jobType") or "") != "overseas_payroll":
        raise LaborWorkerLeaseError("Worker 任务不是海外薪资处理任务。")
    generation = str(existing.get("taskGenerationId") or "").strip()
    if not generation or generation != str(existing.get("runId") or "").strip():
        raise LaborWorkerLeaseError("海外薪资 Worker 任务代次无效。")
    job = heartbeat_labor_worker_job(
        job_id,
        owner_user_id=identity["userId"],
        device_id=identity["deviceId"],
        progress=progress,
        expected_task_generation_id=generation,
    )
    return identity, job, generation


@app.get("/api/overseas-payroll/worker/jobs/{job_id}/manifest")
def get_overseas_payroll_worker_manifest(
    job_id: str,
    authorization: str = Header(default=""),
) -> dict:
    try:
        identity, job, _ = _overseas_payroll_worker_context(
            job_id,
            authorization,
            progress={"status": "running", "message": "正在准备海外薪资文件"},
        )
        task_id = str(job.get("runId") or "")
        task = update_overseas_payroll_task(
            task_id,
            owner_user_id=identity["userId"],
            updater=lambda current: (
                current
                if current.get("status") == "succeeded"
                else {
                    **current,
                    "status": "processing",
                    "statusLabel": "本人核对助手正在下载文件",
                    "progress": {"status": "running", "message": "正在下载待处理文件"},
                }
            ),
        )
        return {
            "task": {
                "id": task["id"],
                "toolId": task["toolId"],
                "toolName": task["toolName"],
                "country": task["country"],
                "status": task["status"],
                "alreadyCompleted": task.get("status") == "succeeded",
            },
            "files": overseas_payroll_task_input_downloads(task),
        }
    except (FileNotFoundError, LaborWorkerLeaseError, ValueError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/overseas-payroll/worker/tasks/{task_id}/files/{file_id}")
def download_local_overseas_payroll_worker_input(
    task_id: str,
    file_id: str,
    jobId: str,
    authorization: str = Header(default=""),
) -> FileResponse:
    try:
        identity, job, _ = _overseas_payroll_worker_context(jobId, authorization)
        if str(job.get("runId") or "") != task_id:
            raise LaborWorkerLeaseError("Worker 输入任务不匹配。")
        task = load_overseas_payroll_task(task_id, owner_user_id=identity["userId"])
        file = next((item for item in task.get("files", []) if item.get("id") == file_id), None)
        if not file:
            raise FileNotFoundError("Worker 输入文件不存在。")
        path = overseas_payroll_local_file_path(task_id, file_id)
        if not path.is_file():
            raise FileNotFoundError("Worker 输入文件不存在。")
        return FileResponse(path, filename=file["filename"], media_type=file.get("contentType") or "application/octet-stream")
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/overseas-payroll/worker/jobs/{job_id}/output-intent")
def create_overseas_payroll_worker_output_intent(
    job_id: str,
    payload: dict = Body(...),
    authorization: str = Header(default=""),
) -> dict:
    try:
        identity, job, _ = _overseas_payroll_worker_context(
            job_id,
            authorization,
            progress={"status": "running", "message": "正在保存海外薪资处理结果"},
        )
        task, intent = prepare_overseas_payroll_output(
            str(job.get("runId") or ""),
            owner_user_id=identity["userId"],
            filename=str(payload.get("filename") or "result.xlsx"),
            size_bytes=int(payload.get("sizeBytes") or 0),
            sha256=str(payload.get("sha256") or ""),
            content_type=str(payload.get("contentType") or "application/octet-stream"),
        )
        return {"taskId": task["id"], "intent": intent}
    except (FileNotFoundError, LaborWorkerLeaseError, ValueError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.put("/api/overseas-payroll/worker/tasks/{task_id}/output/{output_id}/content")
async def upload_local_overseas_payroll_worker_output(
    task_id: str,
    output_id: str,
    request: Request,
    jobId: str,
    authorization: str = Header(default=""),
) -> dict:
    try:
        identity, job, _ = _overseas_payroll_worker_context(jobId, authorization)
        if str(job.get("runId") or "") != task_id:
            raise LaborWorkerLeaseError("Worker 输出任务不匹配。")
        task = load_overseas_payroll_task(task_id, owner_user_id=identity["userId"])
        output = task.get("output") if isinstance(task.get("output"), dict) else None
        if not output or output.get("id") != output_id:
            raise FileNotFoundError("Worker 输出记录不存在。")
        content = await request.body()
        if len(content) != int(output.get("sizeBytes") or 0):
            raise ValueError("Worker 输出大小与任务清单不一致。")
        store_local_overseas_payroll_file(task_id, output_id, content, output=True)
        return {"ok": True}
    except (FileNotFoundError, LaborWorkerLeaseError, ValueError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/overseas-payroll/worker/jobs/{job_id}/output-finalize")
def finalize_overseas_payroll_worker_output(
    job_id: str,
    payload: dict = Body(default={}),
    authorization: str = Header(default=""),
) -> dict:
    try:
        identity, job, _ = _overseas_payroll_worker_context(job_id, authorization)
        task = finalize_overseas_payroll_output(
            str(job.get("runId") or ""),
            owner_user_id=identity["userId"],
            summary=str(payload.get("summary") or "处理完成"),
        )
        return {"taskId": task["id"], "status": task["status"]}
    except (FileNotFoundError, LaborWorkerLeaseError, ValueError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@contextmanager
def _leased_worker_job(
    job_id: str,
    authorization: str,
    *,
    progress: dict | None = None,
    identity: dict[str, str] | None = None,
):
    identity = identity or _labor_worker_identity(authorization)
    job = get_labor_worker_job(job_id)
    run_id = str(job.get("runId") or "")
    run_dir = get_labor_run_dir(run_id)
    with labor_run_metadata_lock(run_id):
        if not labor_postgres_state_enabled() and not (run_dir / "metadata.json").exists():
            sync_labor_run_from_persistent(run_id, run_dir)
        try:
            metadata = load_labor_metadata(run_dir)
            if str(job.get("jobType") or "reconcile") == "mapping_preflight":
                preflight = (
                    metadata.get("mappingPreflight")
                    if isinstance(metadata.get("mappingPreflight"), dict)
                    else {}
                )
                generation = str(preflight.get("taskGenerationId") or "").strip()
            else:
                generation = _labor_task_generation_id(metadata)
        except FileNotFoundError:
            if str(job.get("taskGenerationId") or "").strip():
                raise LaborWorkerLeaseError("服务端缺少当前任务代次元数据，Worker 任务不能继续。")
            generation = ""
        job = heartbeat_labor_worker_job(
            job_id,
            owner_user_id=identity["userId"],
            device_id=identity["deviceId"],
            progress=progress,
            expected_task_generation_id=generation,
        )
        yield identity, job, run_dir, generation


def _update_worker_generation_metadata(
    run_id: str,
    generation: str,
    updates: Callable[[dict], dict],
) -> tuple[dict, bool]:
    if generation:
        return update_labor_metadata_for_task(
            run_id,
            expected_task_generation_id=generation,
            updates=updates,
        )
    with labor_run_metadata_lock(run_id):
        run_dir = get_labor_run_dir(run_id)
        current = load_labor_metadata(run_dir)
        if _labor_task_generation_id(current):
            return current, False
        return update_labor_metadata(run_id, updates(dict(current))), True


def _mark_worker_result_rejected(run_id: str, generation: str, exc: LaborWorkerArchiveError) -> None:
    _update_worker_generation_metadata(
        run_id,
        generation,
        lambda current: {
            **_labor_invalidate_official_result(
                current,
                status="抽取失败",
                reason_code=str(exc.code),
                message=str(exc),
            ),
            "status": "抽取失败",
            "stage": "Worker 结果未接受",
            "errorCode": f"LABOR_{str(exc.code).upper()}",
            "errorMessage": str(exc),
            "retryable": True,
            "machineCheckStatus": "needs_review",
            "manualReviewRequired": True,
            "directPaymentAllowed": False,
            "requiresHumanReview": True,
            "asyncTask": {
                **(current.get("asyncTask") if isinstance(current.get("asyncTask"), dict) else {}),
                "status": "failed",
                "statusLabel": "结果未接受",
                "message": str(exc),
                "failedAt": utcnow_naive().isoformat(),
                "taskGenerationId": generation,
            },
        },
    )


def _worker_result_error_detail(exc: LaborWorkerArchiveError) -> dict:
    return _labor_request_error(
        message=str(exc),
        error_code=f"LABOR_{str(exc.code).upper()}",
        retryable=True,
        next_action="请让本人核对助手重新下载当前批次材料，并重新生成完整核对结果。",
    )


def _worker_result_acceptance_evidence(run_dir: Path) -> tuple[str, int, str]:
    metadata = load_labor_metadata(run_dir)
    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    diff_report = files.get("diffReport") if isinstance(files.get("diffReport"), dict) else {}
    try:
        report_size = int(diff_report.get("sizeBytes") or 0)
    except (TypeError, ValueError):
        report_size = 0
    filename = Path(str(diff_report.get("filename") or "")).name
    report_sha256 = str(diff_report.get("sha256") or "").strip().lower()
    result_input_fingerprint = str(metadata.get("resultInputFingerprint") or "").strip().lower()
    report_path = (run_dir / filename).resolve() if filename else run_dir
    try:
        report_path.relative_to(run_dir.resolve())
    except ValueError:
        return "", 0, ""
    if (
        not filename
        or report_size <= 0
        or not re.fullmatch(r"[0-9a-f]{64}", report_sha256)
        or not re.fullmatch(r"[0-9a-f]{64}", result_input_fingerprint)
        or result_input_fingerprint != _labor_result_input_fingerprint(metadata)
    ):
        return "", 0, ""
    if report_path.is_file():
        if report_path.stat().st_size != report_size:
            return "", 0, ""
        digest = hashlib.sha256()
        with report_path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        if digest.hexdigest() != report_sha256:
            return "", 0, ""
    else:
        object_key = str(diff_report.get("objectKey") or "").strip()
        run_id = str(metadata.get("id") or run_dir.name).strip()
        try:
            verified_size = int(diff_report.get("storageVerifiedSizeBytes") or 0)
        except (TypeError, ValueError):
            verified_size = 0
        durable_private_report = (
            str(diff_report.get("storageBackend") or "").strip().lower() == "supabase"
            and diff_report.get("storagePrivate") is True
            and diff_report.get("storageVerified") is True
            and bool(str(diff_report.get("storageVerifiedAt") or "").strip())
            and object_key.startswith("labor-runs/")
            and f"/runs/{run_id}/outputs/" in object_key
            and verified_size == report_size
        )
        if not durable_private_report:
            return "", 0, ""
    return (
        report_sha256,
        report_size,
        result_input_fingerprint,
    )


def _publish_worker_result_to_authoritative_state(
    run_id: str,
    run_dir: Path,
    *,
    generation: str,
) -> dict:
    """Promote a validated Worker result from temporary disk into the P1 authority."""

    try:
        merged = json.loads((run_dir / "metadata.json").read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise LaborWorkerArchiveError(
            "Worker 合并结果 metadata.json 无法读取，不能写入权威状态。",
            code="worker_result_state_payload_invalid",
            formal_result_rejected=True,
        ) from exc
    if not isinstance(merged, dict):
        raise LaborWorkerArchiveError(
            "Worker 合并结果 metadata.json 格式无效。",
            code="worker_result_state_payload_invalid",
            formal_result_rejected=True,
        )
    if not labor_postgres_state_enabled():
        return merged
    fingerprint = str(merged.get("resultInputFingerprint") or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", fingerprint):
        raise LaborWorkerArchiveError(
            "Worker 合并结果缺少有效输入指纹，不能写入权威状态。",
            code="worker_result_state_fingerprint_invalid",
            formal_result_rejected=True,
        )
    published, committed = compare_and_update_labor_metadata(
        run_id,
        expected_task_generation_id=generation,
        expected_fingerprint=fingerprint,
        fingerprint=_labor_result_input_fingerprint,
        updates=merged,
        conflict_updates=lambda current: {
            **_labor_invalidate_official_result(
                current,
                status="抽取失败",
                reason_code="worker_result_state_conflict",
                message="Worker 结果对应的输入或任务代次已变化，正式结果已拒绝。",
            ),
            "status": "抽取失败",
            "errorCode": "LABOR_WORKER_RESULT_STATE_CONFLICT",
            "errorMessage": "Worker 结果对应的输入或任务代次已变化，正式结果已拒绝。",
            "retryable": True,
        },
    )
    if not committed:
        raise LaborWorkerArchiveError(
            "Worker 结果对应的输入或任务代次已变化，不能写入当前批次。",
            code="worker_result_state_conflict",
            formal_result_rejected=True,
        )
    return published


def _persist_worker_result_outputs(run_id: str, run_dir: Path) -> dict:
    """Persist formal Worker reports privately and attach verified object evidence."""

    try:
        metadata = json.loads((run_dir / "metadata.json").read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise LaborWorkerArchiveError(
            "Worker 输出清单无法读取，不能持久化正式报告。",
            code="worker_output_manifest_invalid",
            formal_result_rejected=True,
        ) from exc
    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    owner_user_id = str(metadata.get("ownerUserId") or "").strip()
    if not owner_user_id:
        raise LaborWorkerArchiveError(
            "Worker 输出缺少服务端 owner，不能写入私有存储。",
            code="worker_output_owner_missing",
            formal_result_rejected=True,
        )
    output_specs = (
        ("diffReport", "diff_report", "diffDownloadUrl"),
        ("businessReport", "business_report", "businessReportDownloadUrl"),
        ("projectionReport", "projection_report", "projectionReportDownloadUrl"),
        ("governanceReport", "governance_report", "governanceReportDownloadUrl"),
    )
    persisted_count = 0
    for file_key, output_kind, top_level_url in output_specs:
        record = files.get(file_key) if isinstance(files.get(file_key), dict) else None
        if not record:
            continue
        filename = Path(str(record.get("filename") or "").replace("\\", "/")).name
        if not filename:
            continue
        source = (run_dir / filename).resolve()
        try:
            source.relative_to(run_dir.resolve())
        except ValueError as exc:
            raise LaborWorkerArchiveError(
                "Worker 输出文件路径越界，不能写入私有存储。",
                code="worker_output_path_invalid",
                formal_result_rejected=True,
            ) from exc
        try:
            storage_evidence = persist_labor_private_output(
                owner_user_id=owner_user_id,
                run_id=run_id,
                output_kind=output_kind,
                path=source,
                expected_sha256=str(record.get("sha256") or ""),
                expected_size_bytes=int(record.get("sizeBytes") or 0),
            )
        except Exception as exc:  # noqa: BLE001 - normalize private-storage internals.
            raise LaborWorkerArchiveError(
                f"Worker 正式输出 {filename} 未能写入私有存储。",
                code="worker_output_storage_failed",
                formal_result_rejected=True,
            ) from exc
        download_url = f"/api/labor/runs/{run_id}/download/{filename}"
        record.update(storage_evidence)
        record["downloadUrl"] = download_url
        files[file_key] = record
        if top_level_url in metadata or file_key in {"diffReport", "businessReport"}:
            metadata[top_level_url] = download_url
        persisted_count += 1
    if not isinstance(files.get("diffReport"), dict) or persisted_count <= 0:
        raise LaborWorkerArchiveError(
            "Worker 正式差异报告未写入私有存储。",
            code="worker_output_report_missing",
            formal_result_rejected=True,
        )
    metadata["files"] = files
    temporary = run_dir / f".metadata.{uuid4().hex}.tmp"
    temporary.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temporary, run_dir / "metadata.json")
    return metadata


def _mark_worker_result_upload_started(run_id: str, generation: str) -> None:
    _, applied = _update_worker_generation_metadata(
        run_id,
        generation,
        lambda current: {
            **_labor_invalidate_official_result(
                current,
                status="抽取中",
                reason_code="worker_result_upload_started",
                message="正在校验新的 Worker 结果；上一版正式证据已失效。",
            ),
            "status": "抽取中",
            "stage": "校验 Worker 结果",
            "retryable": True,
            "asyncTask": {
                **(current.get("asyncTask") if isinstance(current.get("asyncTask"), dict) else {}),
                "status": "running",
                "statusLabel": "校验结果",
                "message": "正在校验本人核对助手提交的新结果。",
                "taskGenerationId": generation,
            },
        },
    )
    if not applied:
        raise LaborWorkerLeaseError("Worker 任务代次已失效，不能接纳新结果。")


@app.get("/api/labor/worker/jobs/{job_id}/input")
def download_personal_labor_worker_input(job_id: str, authorization: str = Header(default="")) -> Response:
    try:
        with _leased_worker_job(job_id, authorization) as (_, job, run_dir, generation):
            payload = build_worker_input_archive(run_dir, expected_task_generation_id=generation)
            return Response(
                payload,
                media_type="application/zip",
                headers={"content-disposition": f'attachment; filename="{job["runId"]}-input.zip"'},
            )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except (LaborWorkerLeaseError, LaborWorkerArchiveError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


def _worker_input_manifest(
    run_dir: Path,
    *,
    owner_user_id: str = "",
    job_type: str = "reconcile",
) -> dict:
    metadata = load_labor_metadata(run_dir)
    mapping_preflight_only = str(job_type or "reconcile").strip() == "mapping_preflight"
    if labor_postgres_state_enabled():
        owner = str(owner_user_id or metadata.get("ownerUserId") or "").strip()
        ready_files = list_labor_file_states(
            run_id=str(metadata.get("id") or run_dir.name),
            owner_user_id=owner,
            ready_only=True,
        )
        if ready_files:
            direct_entries = []
            direct_pdf_records = []
            direct_workbook_records = []
            for file_state in ready_files:
                if mapping_preflight_only and str(file_state.get("fileKind") or "") != "workbook":
                    continue
                file_id = re.sub(r"[^0-9A-Za-z_-]+", "_", str(file_state.get("id") or "file"))[:128]
                original_name = Path(str(file_state.get("originalFilename") or "file").replace("\\", "/")).name
                relative = Path("inputs") / (file_id or "file") / (original_name or "file")
                relative_path = relative.as_posix()
                signed = create_labor_supabase_signed_download(
                    str(file_state.get("objectKey") or ""),
                    filename=original_name,
                    expires_in=600,
                )
                record = {
                    "id": str(file_state.get("id") or ""),
                    "fileKind": str(file_state.get("fileKind") or ""),
                    "filename": original_name,
                    "originalFilename": original_name,
                    "path": relative_path,
                    "objectKey": str(file_state.get("objectKey") or ""),
                    "contentType": str(file_state.get("contentType") or "application/octet-stream"),
                    "sizeBytes": int(file_state.get("sizeBytes") or 0),
                    "sha256": str(file_state.get("sha256") or ""),
                    "uploadState": "ready",
                    "storageBackend": "supabase",
                }
                if record["fileKind"] == "pdf_invoice":
                    direct_pdf_records.append(record)
                elif record["fileKind"] == "workbook":
                    direct_workbook_records.append(record)
                direct_entries.append(
                    {
                        "fileId": record["id"],
                        "fileKind": record["fileKind"],
                        "path": relative_path,
                        "size": record["sizeBytes"],
                        "sha256": record["sha256"],
                        "contentType": record["contentType"],
                        "signedUrl": signed["signedUrl"],
                        "expiresIn": signed["expiresIn"],
                    }
                )
            manifest_metadata = canonicalize_labor_metadata_for_blob(run_dir, metadata)
            manifest_files = dict(manifest_metadata.get("files") or {})
            manifest_files["pdfInvoices"] = [] if mapping_preflight_only else direct_pdf_records
            manifest_files["workbooks"] = direct_workbook_records
            if direct_workbook_records:
                manifest_files["workbook"] = direct_workbook_records[0]
            manifest_metadata["files"] = manifest_files
            return {
                "metadata": manifest_metadata,
                "files": direct_entries,
                "downloadMode": "signed_private",
                "hashVerification": "worker_sha256",
            }
    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    paths: list[Path] = []
    manifest_keys = ("workbooks",) if mapping_preflight_only else ("pdfInvoices", "workbooks")
    for key in manifest_keys:
        records = files.get(key) if isinstance(files.get(key), list) else []
        for record in records:
            if not isinstance(record, dict) or not record.get("path"):
                continue
            path = Path(str(record["path"]))
            if path.is_file():
                paths.append(path)
    entries = []
    for path in paths:
        try:
            relative = path.resolve().relative_to(run_dir.resolve()).as_posix()
        except ValueError:
            continue
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        entries.append({"path": relative, "size": path.stat().st_size, "sha256": digest.hexdigest()})
    return {
        "metadata": canonicalize_labor_metadata_for_blob(run_dir, metadata),
        "files": entries,
        "downloadMode": "api_proxy",
    }


@app.get("/api/labor/worker/jobs/{job_id}/input-manifest")
def get_personal_labor_worker_input_manifest(job_id: str, authorization: str = Header(default="")) -> dict:
    try:
        with _leased_worker_job(job_id, authorization) as (identity, job, run_dir, _):
            return _worker_input_manifest(
                run_dir,
                owner_user_id=identity["userId"],
                job_type=str(job.get("jobType") or "reconcile"),
            )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except LaborWorkerLeaseError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/labor/worker/jobs/{job_id}/input-file")
def download_personal_labor_worker_input_file(
    job_id: str,
    relativePath: str,
    authorization: str = Header(default=""),
) -> FileResponse:
    try:
        with _leased_worker_job(job_id, authorization) as (_, job, leased_run_dir, _):
            run_dir = leased_run_dir.resolve()
            path = (run_dir / relativePath).resolve()
            path.relative_to(run_dir)
            allowed = {
                entry["path"]
                for entry in _worker_input_manifest(
                    run_dir,
                    job_type=str(job.get("jobType") or "reconcile"),
                )["files"]
            }
            if relativePath.replace("\\", "/") not in allowed or not path.is_file():
                raise HTTPException(status_code=404, detail="任务文件不存在或不允许下载。")
            return FileResponse(
                path,
                filename=path.name,
                media_type=mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            )
    except (ValueError, OSError) as exc:
        raise HTTPException(status_code=400, detail="任务文件路径无效。") from exc
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/labor/worker/jobs/{job_id}/mapping-preflight-result")
def upload_personal_labor_mapping_preflight_result(
    job_id: str,
    payload: dict = Body(...),
    authorization: str = Header(default=""),
    x_worker_version: str = Header(default=""),
) -> dict:
    _labor_assert_worker_version(x_worker_version)
    try:
        with _leased_worker_job(job_id, authorization) as (identity, job, run_dir, generation):
            _labor_assert_worker_version(x_worker_version, job)
            if str(job.get("jobType") or "reconcile") != "mapping_preflight":
                raise LaborWorkerLeaseError("正式核对任务不能提交字段预检结果。")
            metadata = load_labor_metadata(run_dir)
            fingerprint = str(payload.get("inputFingerprint") or "").strip().lower()
            if (
                not re.fullmatch(r"[0-9a-f]{64}", fingerprint)
                or fingerprint != _labor_mapping_input_fingerprint(metadata)
            ):
                raise LaborWorkerLeaseError("字段预检对应的 Excel 输入已经变化，结果已拒绝。")
            normalized = _normalize_mapping_preflight_result(metadata, payload)
            completed_at = utcnow_naive().isoformat(timespec="seconds") + "Z"
            updated, applied = update_labor_metadata_for_mapping_preflight(
                str(job.get("runId") or ""),
                expected_task_generation_id=generation,
                updates=lambda current: {
                    "mappingPreflight": {
                        **(current.get("mappingPreflight") if isinstance(current.get("mappingPreflight"), dict) else {}),
                        "status": "completed",
                        "statusLabel": "Excel 字段预检完成",
                        "message": "工作表、列名和样例已由本人核对助手读取，等待用户确认字段映射。",
                        "completedAt": completed_at,
                        "workerDeviceId": identity["deviceId"],
                        "workbooks": normalized["workbooks"],
                        "sheets": normalized["sheets"],
                        "errorCode": "",
                        "errorMessage": "",
                    }
                },
                actor_user_id=identity["userId"],
                action="mapping_preflight_completed",
                reason_code="mapping_preflight_worker_result",
            )
            if not applied:
                raise LaborWorkerLeaseError("字段预检任务代次已失效，不能覆盖当前批次。")
            return {
                "ok": True,
                "mappingPreflight": updated.get("mappingPreflight") or {},
            }
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/labor/worker/jobs/{job_id}/result")
async def upload_personal_labor_worker_result(
    job_id: str,
    result_archive: UploadFile = File(...),
    authorization: str = Header(default=""),
    x_worker_version: str = Header(default=""),
) -> dict:
    _labor_assert_worker_version(x_worker_version)
    try:
        with _leased_worker_job(job_id, authorization) as (identity, job, _, generation):
            _labor_assert_worker_version(x_worker_version, job)
            if str(job.get("jobType") or "reconcile") != "reconcile":
                raise LaborWorkerLeaseError("字段预检任务不能提交正式核对结果包。")
            _mark_worker_result_upload_started(str(job["runId"]), generation)
            clear_labor_worker_result_acceptance(
                job_id,
                owner_user_id=identity["userId"],
                device_id=identity["deviceId"],
                expected_task_generation_id=generation,
            )
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    try:
        payload = await result_archive.read(200 * 1024 * 1024 + 1)
    except Exception as read_exc:  # noqa: BLE001 - fail closed after revoking prior acceptance.
        rejection = LaborWorkerArchiveError(
            "Worker 结果包读取失败，上一版正式结果已失效，请重试。",
            code="worker_result_upload_read_failed",
            formal_result_rejected=True,
        )
        _mark_worker_result_rejected(str(job["runId"]), generation, rejection)
        raise HTTPException(status_code=503, detail=_worker_result_error_detail(rejection)) from read_exc
    if len(payload) > 200 * 1024 * 1024:
        rejection = LaborWorkerArchiveError(
            "Worker 结果包超过 200 MB 限制，上一版正式结果已失效。",
            code="worker_result_archive_too_large",
            formal_result_rejected=True,
        )
        _mark_worker_result_rejected(str(job["runId"]), generation, rejection)
        raise HTTPException(status_code=413, detail=_worker_result_error_detail(rejection))
    try:
        with _leased_worker_job(job_id, authorization) as (identity, job, run_dir, generation):
            _labor_assert_worker_version(x_worker_version, job)
            try:
                merged = merge_worker_result_archive(
                    run_dir,
                    payload,
                    expected_task_generation_id=generation,
                )
            except LaborWorkerArchiveError as exc:
                _mark_worker_result_rejected(str(job["runId"]), generation, exc)
                raise HTTPException(
                    status_code=409 if exc.formal_result_rejected else 400,
                    detail=_worker_result_error_detail(exc),
                ) from exc
            try:
                if labor_postgres_state_enabled():
                    _persist_worker_result_outputs(str(job["runId"]), run_dir)
                    _publish_worker_result_to_authoritative_state(
                        str(job["runId"]),
                        run_dir,
                        generation=generation,
                    )
                else:
                    sync_labor_run_to_persistent(str(job["runId"]), run_dir)
                report_sha256, report_size, input_fingerprint = _worker_result_acceptance_evidence(run_dir)
                mark_labor_worker_result_accepted(
                    job_id,
                    owner_user_id=identity["userId"],
                    device_id=identity["deviceId"],
                    expected_task_generation_id=generation,
                    result_report_sha256=report_sha256,
                    result_report_size_bytes=report_size,
                    result_input_fingerprint=input_fingerprint,
                )
            except Exception as acceptance_exc:  # noqa: BLE001 - compensate cross-store acceptance failure.
                logger.exception(
                    "[%s] Worker 结果接纳失败 type=%s code=%s",
                    job["runId"],
                    type(acceptance_exc).__name__,
                    str(getattr(acceptance_exc, "code", ""))[:80],
                )
                rejection = LaborWorkerArchiveError(
                    "Worker 结果写入后未能完成服务端接纳，正式结果已撤销，请重试。",
                    code="worker_result_acceptance_failed",
                    formal_result_rejected=True,
                )
                try:
                    clear_labor_worker_result_acceptance(
                        job_id,
                        owner_user_id=identity["userId"],
                        device_id=identity["deviceId"],
                        expected_task_generation_id=generation,
                    )
                except Exception as clear_exc:  # noqa: BLE001 - run evidence still fails closed below.
                    logger.error("[%s] Worker 接纳失败后清除任务凭证失败: %s", job["runId"], clear_exc)
                _mark_worker_result_rejected(str(job["runId"]), generation, rejection)
                if not labor_postgres_state_enabled():
                    try:
                        sync_labor_run_to_persistent(str(job["runId"]), run_dir)
                    except Exception as compensation_sync_exc:  # noqa: BLE001 - local readiness remains fail-closed.
                        logger.error(
                            "[%s] Worker 接纳失败补偿状态持久化失败: %s",
                            job["runId"],
                            compensation_sync_exc,
                        )
                raise HTTPException(
                    status_code=503,
                    detail=_worker_result_error_detail(rejection),
                ) from acceptance_exc
            return {"ok": True, "mergedFiles": merged}
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/labor/worker/jobs/{job_id}/events")
def record_personal_labor_worker_events(
    job_id: str,
    payload: dict = Body(default={}),
    authorization: str = Header(default=""),
) -> dict:
    try:
        with _leased_worker_job(job_id, authorization) as (_, job, _, _):
            events = payload.get("events") if isinstance(payload.get("events"), list) else []
            if len(events) > 1000:
                raise HTTPException(status_code=413, detail="单次 Worker 指标事件不能超过 1000 条。")
            LABOR_TELEMETRY_DIR.mkdir(parents=True, exist_ok=True)
            accepted = 0
            with LABOR_TELEMETRY_FILE.open("a", encoding="utf-8") as handle:
                for raw in events:
                    if not isinstance(raw, dict):
                        continue
                    event = _sanitize_labor_telemetry(raw)
                    if event["event"] not in {"ocr_cache", "model_call"}:
                        continue
                    event["source"] = "personal-desktop-worker"
                    event["runId"] = str(job.get("runId") or "")
                    handle.write(json.dumps(event, ensure_ascii=False, separators=(",", ":")) + "\n")
                    accepted += 1
            return {"ok": True, "accepted": accepted}
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/labor/worker/jobs/{job_id}/complete")
def complete_personal_labor_worker_job(
    job_id: str,
    authorization: str = Header(default=""),
    x_worker_version: str = Header(default=""),
) -> dict:
    _labor_assert_worker_version(x_worker_version)
    try:
        existing = get_labor_worker_job(job_id)
        if str(existing.get("jobType") or "reconcile") == "overseas_payroll":
            identity, leased_job, generation = _overseas_payroll_worker_context(job_id, authorization)
            task = load_overseas_payroll_task(
                str(leased_job.get("runId") or ""),
                owner_user_id=identity["userId"],
            )
            if task.get("status") != "succeeded":
                raise LaborWorkerLeaseError("海外薪资处理结果尚未通过服务端校验，不能完成任务。")
            job = complete_labor_worker_auxiliary_job(
                job_id,
                owner_user_id=identity["userId"],
                device_id=identity["deviceId"],
                job_type="overseas_payroll",
                expected_task_generation_id=generation,
            )
            return {"job": _public_labor_worker_job(job)}
        with _leased_worker_job(job_id, authorization) as (identity, leased_job, run_dir, generation):
            _labor_assert_worker_version(x_worker_version, leased_job)
            if str(leased_job.get("jobType") or "reconcile") == "mapping_preflight":
                metadata = load_labor_metadata(run_dir)
                preflight = (
                    metadata.get("mappingPreflight")
                    if isinstance(metadata.get("mappingPreflight"), dict)
                    else {}
                )
                if (
                    str(preflight.get("status") or "") != "completed"
                    or str(preflight.get("taskGenerationId") or "").strip() != generation
                    or str(preflight.get("inputFingerprint") or "").strip().lower()
                    != _labor_mapping_input_fingerprint(metadata)
                ):
                    raise LaborWorkerLeaseError("字段预检结果尚未被服务端接纳，不能完成任务。")
                job = complete_labor_worker_preflight_job(
                    job_id,
                    owner_user_id=identity["userId"],
                    device_id=identity["deviceId"],
                    expected_task_generation_id=generation,
                )
                return {"job": _public_labor_worker_job(job)}
            report_sha256, report_size, input_fingerprint = _worker_result_acceptance_evidence(run_dir)
            if (
                not re.fullmatch(r"[0-9a-f]{64}", report_sha256)
                or report_size <= 0
                or not re.fullmatch(r"[0-9a-f]{64}", input_fingerprint)
            ):
                rejection = LaborWorkerArchiveError(
                    "当前 Worker 正式结果或报告文件已变化，不能完成任务。",
                    code="worker_result_acceptance_evidence_mismatch",
                    formal_result_rejected=True,
                )
                clear_labor_worker_result_acceptance(
                    job_id,
                    owner_user_id=identity["userId"],
                    device_id=identity["deviceId"],
                    expected_task_generation_id=generation,
                )
                _mark_worker_result_rejected(str(leased_job["runId"]), generation, rejection)
                if not labor_postgres_state_enabled():
                    try:
                        sync_labor_run_to_persistent(str(leased_job["runId"]), run_dir)
                    except Exception as sync_exc:  # noqa: BLE001 - local result remains fail-closed.
                        logger.error("[%s] Worker 完成前证据失效状态持久化失败: %s", leased_job["runId"], sync_exc)
                raise LaborWorkerLeaseError(str(rejection))
            job = complete_labor_worker_job(
                job_id,
                owner_user_id=identity["userId"],
                device_id=identity["deviceId"],
                expected_task_generation_id=generation,
                expected_result_report_sha256=report_sha256,
                expected_result_report_size_bytes=report_size,
                expected_result_input_fingerprint=input_fingerprint,
            )
            return {"job": _public_labor_worker_job(job)}
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.post("/api/labor/worker/jobs/{job_id}/fail")
def fail_personal_labor_worker_job(
    job_id: str,
    payload: dict = Body(default={}),
    authorization: str = Header(default=""),
) -> dict:
    try:
        existing = get_labor_worker_job(job_id)
        if str(existing.get("jobType") or "reconcile") == "overseas_payroll":
            identity, leased_job, generation = _overseas_payroll_worker_context(job_id, authorization)
            job = fail_labor_worker_job(
                job_id,
                owner_user_id=identity["userId"],
                device_id=identity["deviceId"],
                error_code=str(payload.get("errorCode") or "OVERSEAS_PAYROLL_WORKER_FAILED"),
                error_message=str(payload.get("errorMessage") or "海外薪资 Worker 任务失败。"),
                retryable=bool(payload.get("retryable")),
                expected_task_generation_id=generation,
            )
            status = str(job.get("status") or "failed")
            update_overseas_payroll_task(
                str(leased_job.get("runId") or ""),
                owner_user_id=identity["userId"],
                updater=lambda current: {
                    **current,
                    **(
                        {}
                        if current.get("status") == "succeeded"
                        else {
                            "status": "queued" if status == "retry_wait" else "failed",
                            "statusLabel": "等待核对助手重试" if status == "retry_wait" else "处理失败",
                            "error": str(job.get("errorMessage") or "海外薪资处理失败。")[:500],
                        }
                    ),
                },
            )
            return {"job": _public_labor_worker_job(job)}
        with _leased_worker_job(job_id, authorization) as (identity, leased_job, _, generation):
            job = fail_labor_worker_job(
                job_id,
                owner_user_id=identity["userId"],
                device_id=identity["deviceId"],
                error_code=str(payload.get("errorCode") or "LABOR_WORKER_FAILED"),
                error_message=str(payload.get("errorMessage") or "Worker 任务失败。"),
                retryable=bool(payload.get("retryable")),
                expected_task_generation_id=generation,
            )
            run_id = str(leased_job.get("runId") or "")
            status = str(job.get("status") or "")
            if str(leased_job.get("jobType") or "reconcile") == "mapping_preflight":
                update_labor_metadata_for_mapping_preflight(
                    run_id,
                    expected_task_generation_id=generation,
                    updates=lambda current: {
                        "mappingPreflight": {
                            **(current.get("mappingPreflight") if isinstance(current.get("mappingPreflight"), dict) else {}),
                            "status": status,
                            "statusLabel": "等待重试" if status == "retry_wait" else "字段预检失败",
                            "message": str(job.get("errorMessage") or "Worker 字段预检失败。"),
                            "errorCode": str(job.get("errorCode") or "LABOR_WORKER_FAILED"),
                            "errorMessage": str(job.get("errorMessage") or "Worker 字段预检失败。"),
                        }
                    },
                    actor_user_id=identity["userId"],
                    action="mapping_preflight_retry_wait" if status == "retry_wait" else "mapping_preflight_failed",
                    reason_code=str(job.get("errorCode") or "LABOR_WORKER_FAILED"),
                )
                return {"job": _public_labor_worker_job(job)}
            try:
                update_labor_metadata_for_task(
                    run_id,
                    expected_task_generation_id=generation,
                    updates={
                        "status": "抽取中" if status == "retry_wait" else "失败",
                        "asyncTask": {
                            "status": status,
                            "statusLabel": "等待重试" if status == "retry_wait" else "处理失败",
                            "message": str(job.get("errorMessage") or "Worker 任务失败。"),
                            **({"taskGenerationId": generation} if generation else {}),
                        },
                        "errorCode": str(job.get("errorCode") or "LABOR_WORKER_FAILED"),
                        "errorMessage": str(job.get("errorMessage") or "Worker 任务失败。"),
                        "retryable": status == "retry_wait",
                        "businessReviewStatus": "pending",
                        "manualReviewRequired": True,
                        "directPaymentAllowed": False,
                        "requiresHumanReview": True,
                    },
                )
            except FileNotFoundError:
                if generation:
                    raise
            return {"job": _public_labor_worker_job(job)}
    except (FileNotFoundError, LaborWorkerLeaseError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@app.get("/api/labor/material-index")
def labor_material_index(root: str = "") -> dict:
    _assert_local_labor_material_tool_available()
    material_root = root or os.environ.get("LABOR_REFERENCE_MATERIALS_DIR") or "/Users/zt27532/Documents/报账核对工具"
    try:
        return build_material_index(material_root)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except NotADirectoryError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/labor/material-replay-plan")
def labor_material_replay_plan(root: str = "", batchKey: str = "") -> dict:
    _assert_local_labor_material_tool_available()
    material_root = root or os.environ.get("LABOR_REFERENCE_MATERIALS_DIR") or "/Users/zt27532/Documents/报账核对工具"
    try:
        return build_material_replay_plan(material_root, batch_key=batchKey)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except NotADirectoryError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/labor/material-dry-run")
def labor_material_dry_run(payload: dict = Body(...)) -> dict:
    _assert_local_labor_material_tool_available()
    material_root = str(payload.get("root") or os.environ.get("LABOR_REFERENCE_MATERIALS_DIR") or "/Users/zt27532/Documents/报账核对工具")
    batch_key = str(payload.get("batchKey") or payload.get("batch_key") or "").strip()
    if not batch_key:
        raise HTTPException(status_code=400, detail="请提供 batchKey。")
    try:
        return build_material_dry_run(
            material_root,
            batch_key,
            amount_tolerance=float(payload.get("amountTolerance") or payload.get("amount_tolerance") or AI_CONFIG["amount_tolerance"]),
            hours_tolerance=float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"]),
            confidence_threshold=float(payload.get("confidenceThreshold") or payload.get("confidence_threshold") or AI_CONFIG["confidence_threshold"]),
            currency=str(payload.get("currency") or "USD"),
        )
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except NotADirectoryError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/labor/material-runs")
def create_labor_run_from_material(request: Request, payload: dict = Body(...)) -> dict:
    _assert_local_labor_material_tool_available()
    material_root = str(payload.get("root") or os.environ.get("LABOR_REFERENCE_MATERIALS_DIR") or "/Users/zt27532/Documents/报账核对工具")
    batch_key = str(payload.get("batchKey") or payload.get("batch_key") or "").strip()
    if not batch_key:
        raise HTTPException(status_code=400, detail="请提供 batchKey。")
    try:
        plan_payload = build_material_replay_plan(material_root, batch_key=batch_key)
    except (FileNotFoundError, NotADirectoryError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    plans = plan_payload.get("plans") or []
    if not plans:
        raise HTTPException(status_code=400, detail=f"未找到可创建批次的材料批次: {batch_key}")
    plan = plans[0]
    root_path = Path(plan_payload["root"]).resolve()
    supplier = str(payload.get("supplierName") or payload.get("supplier_name") or plan.get("supplier") or "").strip()
    if not supplier or supplier == "unknown":
        supplier = str(plan.get("directory") or plan.get("batchKey") or "unknown").strip()
    if labor_auth_required():
        owner_user_id, _ = _labor_request_actor(request)
    else:
        owner_user_id = str(payload.get("ownerUserId") or payload.get("owner_user_id") or "local-default")
    run = create_labor_run(
        {
            "supplierName": supplier,
            "periodStart": str(payload.get("periodStart") or payload.get("period_start") or "").strip(),
            "periodEnd": str(payload.get("periodEnd") or payload.get("period_end") or "").strip(),
            "currency": str(payload.get("currency") or "USD").strip() or "USD",
            "notes": str(payload.get("notes") or f"Created from material batch {plan.get('batchKey', '')}"),
            "ownerUserId": owner_user_id,
        }
    )
    if not labor_postgres_state_enabled():
        append_labor_audit_event(
            _labor_audit_path(),
            action="run_created",
            run_id=run["id"],
            owner_user_id=owner_user_id,
            actor_user_id=owner_user_id,
            outcome="success",
        )
    run_dir = get_labor_run_dir(run["id"])
    try:
        files, copied_sources = _copy_material_plan_files(run["id"], run_dir, root_path, plan)
    except (FileNotFoundError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    mapping_candidate = next(
        (
            item
            for item in plan.get("mappingCandidates", []) or []
            if item.get("sheetName") and item.get("suggestedMapping") and not item.get("error")
        ),
        {},
    )
    copied_workbooks = [item for item in copied_sources if item.get("kind") == "workbook"]
    mapping_by_relative_path = {
        str(item.get("relativePath") or ""): item
        for item in plan.get("mappingCandidates", []) or []
        if item.get("sheetName") and item.get("suggestedMapping") and not item.get("error")
    }
    workbook_mappings = []
    for copied in copied_workbooks:
        candidate = mapping_by_relative_path.get(str(copied.get("relativePath") or ""))
        if not candidate:
            continue
        workbook_mappings.append(
            {
                "relativePath": copied.get("relativePath", ""),
                "filename": Path(str(copied.get("copiedPath") or "")).name,
                "sheetName": candidate["sheetName"],
                "mapping": candidate["suggestedMapping"],
            }
        )
    updates = {
        "status": "已确认字段" if mapping_candidate else "已上传文件",
        "files": files,
        "materialReplayNextStep": {
            "action": "extract_compare" if mapping_candidate else "confirm_mapping",
            "label": "抽取并比对" if mapping_candidate else "确认字段映射",
            "enabled": bool(mapping_candidate),
            "description": (
                "材料批次已复制并预填字段映射，可直接执行抽取核对。"
                if mapping_candidate
                else "材料批次已复制，请先确认账单工作表和字段映射。"
            ),
        },
        "materialReplaySource": {
            "root": str(root_path),
            "batchKey": plan.get("batchKey", ""),
            "directory": plan.get("directory", ""),
            "supplier": plan.get("supplier", ""),
            "periodHint": plan.get("periodHint", ""),
            "uploadPlan": plan.get("uploadPlan", {}),
            "copiedSources": copied_sources,
            "mappingCandidate": mapping_candidate,
            "expectedRisks": plan.get("expectedRisks", []),
            "createdAt": utcnow_naive().isoformat(),
        },
        "workbookMappings": workbook_mappings,
    }
    if mapping_candidate:
        updates["workbookSheet"] = mapping_candidate["sheetName"]
        updates["excelMapping"] = mapping_candidate["suggestedMapping"]
    return update_labor_metadata(run["id"], updates)


@app.post("/api/labor/runs")
def create_labor_run_endpoint(request: Request, payload: dict = Body(...)) -> dict:
    supplier = str(payload.get("supplier_name") or payload.get("supplierName") or "").strip()
    period_start = str(payload.get("period_start") or payload.get("periodStart") or "").strip()
    period_end = str(payload.get("period_end") or payload.get("periodEnd") or "").strip()
    if not supplier:
        raise HTTPException(status_code=400, detail="请填写供应商名称。")
    if not period_start or not period_end:
        raise HTTPException(status_code=400, detail="请填写账期开始和结束日期。")
    if labor_auth_required():
        owner_user_id, _ = _labor_request_actor(request)
    else:
        owner_user_id = str(payload.get("ownerUserId") or payload.get("owner_user_id") or "local-default")
    try:
        run = create_labor_run(
            {
                "supplierName": supplier,
                "periodStart": period_start,
                "periodEnd": period_end,
                "currency": str(payload.get("currency") or "USD").strip() or "USD",
                "notes": str(payload.get("notes") or ""),
                "requireEmployeeDetail": True,
                "reconciliationScope": "employee_detail_required",
                "ownerUserId": owner_user_id,
            }
        )
    except Exception as exc:
        logger.exception("labor run creation failed")
        raise HTTPException(
            status_code=503,
            detail={
                "message": "海外劳务批次创建失败，持久化存储暂不可用。",
                "nextAction": "请管理员检查私有存储和状态数据库配置后重试。",
                "errorType": type(exc).__name__,
            },
        ) from exc
    if not labor_postgres_state_enabled():
        append_labor_audit_event(
            _labor_audit_path(),
            action="run_created",
            run_id=run["id"],
            owner_user_id=owner_user_id,
            actor_user_id=owner_user_id,
            outcome="success",
        )
    return run


@app.get("/api/labor/runs/{run_id}")
def get_labor_run(run_id: str) -> dict:
    try:
        metadata = load_labor_metadata(get_labor_run_dir(run_id))
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    return _with_labor_readiness(
        _with_personal_worker_status(_check_stale_extracting(_normalize_labor_total_decision(metadata)))
    )


@app.post("/api/labor/runs/{run_id}/business-review")
def review_labor_run_result(request: Request, run_id: str, payload: dict = Body(...)) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    decision = str(payload.get("decision") or "").strip().lower()
    if decision not in {"approved", "rejected"}:
        raise HTTPException(status_code=400, detail="业务复核结论仅支持 approved 或 rejected。")
    reason = str(payload.get("reason") or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="请填写业务复核理由。")
    expected_fingerprint = str(metadata.get("resultInputFingerprint") or "").strip().lower()
    current_fingerprint = _labor_result_input_fingerprint(metadata)
    if not re.fullmatch(r"[0-9a-f]{64}", expected_fingerprint) or expected_fingerprint != current_fingerprint:
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前核对结果已失效或不属于最新输入材料，不能完成业务复核。",
                error_code="LABOR_REVIEW_INPUT_FINGERPRINT_MISMATCH",
                next_action="请重新生成完整核对结果后再复核。",
            ),
        )
    actor_user_id, _ = _labor_request_actor(request)
    reviewed_at = utcnow_naive().isoformat()
    updated = update_labor_metadata(
        run_id,
        {
            "businessReviewStatus": decision,
            "businessReviewReason": reason[:1000],
            "businessReviewedBy": actor_user_id,
            "businessReviewedAt": reviewed_at,
            "manualReviewRequired": True,
            "requiresHumanReview": decision != "approved",
            "directPaymentAllowed": False,
        },
        actor_user_id=actor_user_id,
        action=f"business_review_{decision}",
        reason_code="manual_business_review",
        audit_details={"decision": decision, "resultInputFingerprint": expected_fingerprint},
    )
    return _normalize_labor_total_decision(updated)


@app.delete("/api/labor/runs/{run_id}")
def delete_labor_run(request: Request, run_id: str, reason: str = "user_requested") -> dict:
    run_dir = get_labor_run_dir(run_id)
    try:
        load_labor_metadata(run_dir)
        actor_user_id, _ = _labor_request_actor(request)
        result = delete_labor_run_directory(
            run_dir,
            audit_path=_labor_audit_path(),
            reason_code=reason or "user_requested",
            actor_user_id=actor_user_id,
            delete_persistent=delete_labor_run_from_persistent,
            delete_authoritative=(
                lambda deleted_run_id, actor, reason_code: soft_delete_labor_run_state(
                    deleted_run_id,
                    actor_user_id=actor,
                    reason_code=reason_code,
                )
            )
            if labor_postgres_state_enabled()
            else None,
            record_audit=not labor_postgres_state_enabled(),
        )
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    except RuntimeError as exc:
        if str(exc) == "ACTIVE_RUN":
            raise HTTPException(
                status_code=409,
                detail=_labor_request_error(
                    message="核对任务正在运行，不能删除批次。",
                    error_code="LABOR_ACTIVE_RUN_DELETE_BLOCKED",
                    retryable=True,
                    next_action="请等待任务完成或失败后再删除。",
                ),
            ) from exc
        raise
    return {"message": f"已删除海外劳务核对批次: {run_id}", **result}


_LABOR_P1_FILE_KINDS = {"pdf_invoice", "workbook"}
_LABOR_P1_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


def _labor_p1_state_required() -> None:
    if labor_postgres_state_enabled():
        return
    raise HTTPException(
        status_code=503,
        detail=_labor_request_error(
            message="P1 直传要求启用 Postgres 权威状态库。",
            error_code="LABOR_P1_STATE_REQUIRED",
            retryable=True,
            next_action="请由管理员完成 P1 状态库配置后重试。",
        ),
    )


def _labor_p1_file_spec(raw: object, *, policy: LaborHardeningPolicy) -> dict:
    if not isinstance(raw, dict):
        raise HTTPException(status_code=400, detail="上传文件信息格式错误。")
    file_kind = str(raw.get("fileKind") or "").strip().lower()
    if file_kind not in _LABOR_P1_FILE_KINDS:
        raise HTTPException(status_code=400, detail="文件类型仅支持 pdf_invoice 或 workbook。")
    original_filename = Path(str(raw.get("filename") or "").replace("\\", "/")).name.strip()
    if not original_filename:
        raise HTTPException(status_code=400, detail="上传文件缺少文件名。")
    extension = Path(original_filename).suffix.lower()
    if file_kind == "pdf_invoice" and extension != ".pdf":
        raise HTTPException(status_code=400, detail="供应商发票请上传 PDF 文件。")
    if file_kind == "workbook" and extension not in _LABOR_P1_WORKBOOK_EXTENSIONS:
        raise HTTPException(status_code=400, detail="线下账单请上传 Excel 文件（.xlsx / .xlsm / .xls）。")
    try:
        size_bytes = int(raw.get("sizeBytes") or 0)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="上传文件大小格式错误。") from exc
    size_limit = policy.max_pdf_bytes if file_kind == "pdf_invoice" else policy.max_workbook_bytes
    if size_bytes <= 0:
        raise HTTPException(status_code=400, detail="上传文件大小必须大于 0。")
    if size_bytes > size_limit:
        code = "LABOR_PDF_SIZE_LIMIT_EXCEEDED" if file_kind == "pdf_invoice" else "LABOR_WORKBOOK_SIZE_LIMIT_EXCEEDED"
        raise _labor_upload_limit_http_error(
            LaborResourceLimitError(code, f"文件 {original_filename} 超过大小限制。", limit=size_limit)
        )
    digest = str(raw.get("sha256") or "").strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", digest):
        raise HTTPException(status_code=400, detail="上传文件缺少有效的 SHA-256。")
    default_content_type = (
        "application/pdf"
        if file_kind == "pdf_invoice"
        else mimetypes.guess_type(original_filename)[0] or "application/octet-stream"
    )
    content_type = str(raw.get("contentType") or default_content_type).split(";", 1)[0].strip().lower()
    if file_kind == "pdf_invoice" and content_type not in {"application/pdf", "application/octet-stream"}:
        raise HTTPException(status_code=400, detail="PDF 文件的 Content-Type 不合法。")
    return {
        "fileKind": file_kind,
        "filename": original_filename[:255],
        "contentType": content_type or default_content_type,
        "sizeBytes": size_bytes,
        "sha256": digest,
    }


def _labor_p1_state_http_error(exc: Exception) -> HTTPException:
    if isinstance(exc, (LaborStateNotFound, LaborStateOwnerMismatch)):
        return HTTPException(
            status_code=404,
            detail=_labor_request_error(
                message="劳务核对批次或文件记录未找到。",
                error_code="LABOR_FILE_NOT_FOUND",
                next_action="请刷新批次后重试。",
            ),
        )
    if isinstance(exc, LaborStateConflict):
        return HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message=str(exc),
                error_code="LABOR_FILE_STATE_CONFLICT",
                retryable=True,
                next_action="请刷新文件状态后重试。",
            ),
        )
    if isinstance(exc, ValueError):
        return HTTPException(status_code=400, detail=str(exc))
    return HTTPException(
        status_code=503,
        detail=_labor_request_error(
            message="P1 文件状态服务暂不可用。",
            error_code="LABOR_P1_STATE_UNAVAILABLE",
            retryable=True,
            next_action="请稍后重试；持续失败请联系管理员查看服务日志。",
        ),
    )


def _labor_p1_storage_http_error() -> HTTPException:
    return HTTPException(
        status_code=503,
        detail=_labor_request_error(
            message="P1 私有文件存储暂不可用。",
            error_code="LABOR_P1_STORAGE_UNAVAILABLE",
            retryable=True,
            next_action="请稍后重试；持续失败请联系管理员检查私有存储 readiness。",
        ),
    )


@app.post("/api/labor/runs/{run_id}/upload-intents")
def create_labor_upload_intents(
    request: Request,
    run_id: str,
    payload: dict = Body(...),
) -> dict:
    _labor_p1_state_required()
    try:
        metadata = load_labor_metadata(get_labor_run_dir(run_id))
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    owner_user_id = str(metadata.get("ownerUserId") or "").strip()
    actor_user_id, _ = _labor_request_actor(request)
    raw_files = payload.get("files") if isinstance(payload, dict) else None
    if not isinstance(raw_files, list) or not raw_files:
        raise HTTPException(status_code=400, detail="请至少提交一个上传文件。")
    policy = _labor_hardening_policy()
    specs = [_labor_p1_file_spec(item, policy=policy) for item in raw_files]
    pdf_count = sum(item["fileKind"] == "pdf_invoice" for item in specs)
    workbook_count = sum(item["fileKind"] == "workbook" for item in specs)
    if pdf_count > policy.max_pdf_files:
        raise _labor_upload_limit_http_error(
            LaborResourceLimitError(
                "LABOR_PDF_FILE_COUNT_LIMIT_EXCEEDED",
                f"每个批次最多上传 {policy.max_pdf_files} 个 PDF 文件。",
                limit=policy.max_pdf_files,
            )
        )
    if workbook_count > policy.max_workbook_files:
        raise _labor_upload_limit_http_error(
            LaborResourceLimitError(
                "LABOR_WORKBOOK_FILE_COUNT_LIMIT_EXCEEDED",
                f"每个批次最多上传 {policy.max_workbook_files} 个 Excel 文件。",
                limit=policy.max_workbook_files,
            )
        )
    prepared = []
    for spec in specs:
        file_id = f"labor_file_{uuid4().hex}"
        object_key = labor_p1_object_key(
            owner_user_id=owner_user_id,
            run_id=run_id,
            file_id=file_id,
            filename=spec["filename"],
        )
        prepared.append(
            {
                "state": {
                    "file_id": file_id,
                    "file_kind": spec["fileKind"],
                    "object_key": object_key,
                    "original_filename": spec["filename"],
                    "content_type": spec["contentType"],
                    "size_bytes": spec["sizeBytes"],
                    "sha256": spec["sha256"],
                },
                "fileId": file_id,
                "fileKind": spec["fileKind"],
                "contentType": spec["contentType"],
            }
        )
    def sign_prepared(item: dict) -> dict:
        return {
            "fileId": item["fileId"],
            **create_labor_supabase_signed_upload_for_object(
                item["state"]["object_key"],
                file_kind=item["fileKind"],
                content_type=item["contentType"],
            ),
        }

    try:
        with ThreadPoolExecutor(max_workers=min(8, len(prepared))) as executor:
            intents = list(executor.map(sign_prepared, prepared))
    except Exception as exc:  # noqa: BLE001 - never expose signed-storage internals.
        raise _labor_p1_storage_http_error() from exc
    try:
        create_pending_labor_file_states(
            run_id=run_id,
            owner_user_id=owner_user_id,
            actor_user_id=actor_user_id,
            files=[item["state"] for item in prepared],
            max_files_by_kind={
                "pdf_invoice": policy.max_pdf_files,
                "workbook": policy.max_workbook_files,
            },
        )
    except Exception as exc:  # noqa: BLE001 - mapped to stable public state errors below.
        diagnostic = getattr(exc, "diag", None)
        logger.error(
            "Labor P1 upload intent state write failed: "
            "run_id=%s file_count=%d pdf_count=%d workbook_count=%d "
            "error_type=%s sqlstate=%s table=%s constraint=%s",
            run_id,
            len(prepared),
            pdf_count,
            workbook_count,
            type(exc).__name__,
            str(getattr(exc, "sqlstate", "") or "")[:16],
            str(getattr(diagnostic, "table_name", "") or "")[:128],
            str(getattr(diagnostic, "constraint_name", "") or "")[:128],
        )
        raise _labor_p1_state_http_error(exc) from exc
    return {"runId": run_id, "intents": intents}


@app.post("/api/labor/runs/{run_id}/upload-intents/batch-finalize")
def batch_finalize_labor_upload_intents(
    request: Request,
    run_id: str,
    payload: dict = Body(...),
) -> dict:
    _labor_p1_state_required()
    metadata = _labor_metadata_or_404(run_id)
    owner_user_id = str(metadata.get("ownerUserId") or "").strip()
    actor_user_id, _ = _labor_request_actor(request)
    raw_file_ids = payload.get("fileIds") if isinstance(payload, dict) else None
    if not isinstance(raw_file_ids, list) or not raw_file_ids:
        raise HTTPException(status_code=400, detail="请至少提交一个待确认文件。")
    file_ids = [str(value or "").strip() for value in raw_file_ids]
    if any(not value for value in file_ids):
        raise HTTPException(status_code=400, detail="待确认文件 ID 不能为空。")
    if len(file_ids) > 64:
        raise HTTPException(status_code=400, detail="单次确认文件数量不能超过 64。")
    if len(set(file_ids)) != len(file_ids):
        raise HTTPException(status_code=400, detail="待确认文件 ID 不能重复。")
    try:
        file_states = list_labor_file_states(
            run_id=run_id,
            owner_user_id=owner_user_id,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_state_http_error(exc) from exc
    state_by_id = {
        str(item.get("id") or ""): item
        for item in file_states
    }
    selected_states = []
    for file_id in file_ids:
        file_state = state_by_id.get(file_id)
        if not file_state:
            raise _labor_p1_state_http_error(LaborStateNotFound("上传文件记录不存在。"))
        selected_states.append(file_state)

    def observe(file_state: dict) -> dict:
        observed = labor_supabase_object_metadata(str(file_state.get("objectKey") or ""))
        expected_type = str(file_state.get("contentType") or "").split(";", 1)[0].strip().lower()
        observed_type = str(observed.get("contentType") or "").split(";", 1)[0].strip().lower()
        if (
            expected_type
            and observed_type
            and expected_type != "application/octet-stream"
            and expected_type != observed_type
        ):
            raise HTTPException(
                status_code=409,
                detail=_labor_request_error(
                    message="对象存储中的文件类型与上传意图不一致。",
                    error_code="LABOR_FILE_CONTENT_TYPE_MISMATCH",
                    next_action="请删除当前文件并重新上传。",
                    requires_reupload=True,
                ),
            )
        return {
            "file_id": str(file_state.get("id") or ""),
            "observed_size_bytes": int(observed.get("sizeBytes") or 0),
            "reported_sha256": str(file_state.get("sha256") or ""),
        }

    try:
        with ThreadPoolExecutor(max_workers=min(8, len(selected_states))) as executor:
            observed_files = list(executor.map(observe, selected_states))
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_storage_http_error() from exc
    try:
        finalized = finalize_labor_file_states(
            run_id=run_id,
            owner_user_id=owner_user_id,
            actor_user_id=actor_user_id,
            files=observed_files,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_state_http_error(exc) from exc
    return {
        "runId": run_id,
        "files": finalized,
        "hashVerification": "worker_required",
    }


@app.post("/api/labor/runs/{run_id}/upload-intents/{file_id}/finalize")
def finalize_labor_upload_intent(
    request: Request,
    run_id: str,
    file_id: str,
    payload: dict = Body(default={}),
) -> dict:
    _labor_p1_state_required()
    metadata = _labor_metadata_or_404(run_id)
    owner_user_id = str(metadata.get("ownerUserId") or "").strip()
    actor_user_id, _ = _labor_request_actor(request)
    try:
        file_state = get_labor_file_state(
            run_id=run_id,
            owner_user_id=owner_user_id,
            file_id=file_id,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_state_http_error(exc) from exc
    try:
        observed = labor_supabase_object_metadata(str(file_state.get("objectKey") or ""))
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_storage_http_error() from exc
    expected_type = str(file_state.get("contentType") or "").split(";", 1)[0].strip().lower()
    observed_type = str(observed.get("contentType") or "").split(";", 1)[0].strip().lower()
    if expected_type and observed_type and expected_type != "application/octet-stream" and expected_type != observed_type:
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="对象存储中的文件类型与上传意图不一致。",
                error_code="LABOR_FILE_CONTENT_TYPE_MISMATCH",
                next_action="请删除当前文件并重新上传。",
                requires_reupload=True,
            ),
        )
    try:
        finalized = finalize_labor_file_state(
            run_id=run_id,
            owner_user_id=owner_user_id,
            actor_user_id=actor_user_id,
            file_id=file_id,
            observed_size_bytes=int(observed.get("sizeBytes") or 0),
            reported_sha256=str(file_state.get("sha256") or ""),
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_state_http_error(exc) from exc
    return {
        "runId": run_id,
        "file": finalized,
        "hashVerification": "worker_required",
        "clientObjectKeyIgnored": bool(payload.get("objectKey")) if isinstance(payload, dict) else False,
    }


@app.get("/api/labor/runs/{run_id}/files/{file_id}/signed-download")
def get_labor_file_signed_download(run_id: str, file_id: str) -> dict:
    _labor_p1_state_required()
    metadata = _labor_metadata_or_404(run_id)
    owner_user_id = str(metadata.get("ownerUserId") or "").strip()
    try:
        file_state = get_labor_file_state(
            run_id=run_id,
            owner_user_id=owner_user_id,
            file_id=file_id,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_state_http_error(exc) from exc
    if str(file_state.get("uploadState") or "") != "ready":
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="文件尚未完成上传确认。",
                error_code="LABOR_FILE_NOT_READY",
                retryable=True,
                next_action="请等待上传完成后重试。",
            ),
        )
    try:
        return create_labor_supabase_signed_download(
            str(file_state.get("objectKey") or ""),
            filename=str(file_state.get("originalFilename") or ""),
            expires_in=120,
        )
    except Exception as exc:  # noqa: BLE001
        raise _labor_p1_storage_http_error() from exc


@app.post("/api/labor/runs/{run_id}/files")
async def upload_labor_files(
    request: Request,
    run_id: str,
    pdf_files: Optional[list[UploadFile]] = File(default=None),
    workbook_files: Optional[list[UploadFile]] = File(default=None),
) -> dict:
    if _labor_p1_required():
        raise HTTPException(
            status_code=409,
            detail=_labor_p1_signed_upload_required_detail(),
        )
    try:
        run_dir = get_labor_run_dir(run_id)
        metadata = load_labor_metadata(run_dir)
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    pdf_files = list(pdf_files or [])
    workbook_files = list(workbook_files or [])
    files = dict(metadata.get("files", {}))
    existing_pdf_records = list(files.get("pdfInvoices") or [])
    existing_workbook_records = list(files.get("workbooks") or [])
    if not pdf_files and not workbook_files:
        raise HTTPException(status_code=400, detail="请至少选择一个待上传文件。")
    if not pdf_files and not existing_pdf_records:
        raise HTTPException(status_code=400, detail="请至少上传一张 PDF 发票。")
    if not workbook_files and not existing_workbook_records:
        raise HTTPException(status_code=400, detail="请上传线下账单 Excel 文件。")
    policy = _labor_hardening_policy()
    owner_user_id = str(metadata.get("ownerUserId") or "local-default")
    if len(existing_pdf_records) + len(pdf_files) > policy.max_pdf_files:
        _record_labor_resource_rejection(
            run_id,
            owner_user_id,
            "LABOR_PDF_FILE_COUNT_LIMIT_EXCEEDED",
            limit=policy.max_pdf_files,
        )
        raise _labor_upload_limit_http_error(
            LaborResourceLimitError(
                "LABOR_PDF_FILE_COUNT_LIMIT_EXCEEDED",
                f"每个批次最多上传 {policy.max_pdf_files} 个 PDF 文件。",
                limit=policy.max_pdf_files,
            )
        )
    if len(existing_workbook_records) + len(workbook_files) > policy.max_workbook_files:
        _record_labor_resource_rejection(
            run_id,
            owner_user_id,
            "LABOR_WORKBOOK_FILE_COUNT_LIMIT_EXCEEDED",
            limit=policy.max_workbook_files,
        )
        raise _labor_upload_limit_http_error(
            LaborResourceLimitError(
                "LABOR_WORKBOOK_FILE_COUNT_LIMIT_EXCEEDED",
                f"每个批次最多上传 {policy.max_workbook_files} 个 Excel 文件。",
                limit=policy.max_workbook_files,
            )
        )
    _EXCEL_EXTS = (".xlsx", ".xlsm", ".xls")
    existing_names = {
        str(record.get("originalFilename") or "").strip().casefold()
        for record in existing_pdf_records + existing_workbook_records
        if isinstance(record, dict)
    }
    incoming_names: set[str] = set()
    for upload in [*pdf_files, *workbook_files]:
        normalized_name = str(upload.filename or "").strip().casefold()
        if normalized_name in existing_names or normalized_name in incoming_names:
            raise HTTPException(status_code=409, detail=f"同名文件已在当前批次中：{upload.filename}。请勿重复上传或先新建批次。")
        incoming_names.add(normalized_name)
    for upload in pdf_files:
        if not upload.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="供应商发票请上传 PDF 文件。")
    for upload in workbook_files:
        if not upload.filename.lower().endswith(_EXCEL_EXTS):
            raise HTTPException(status_code=400, detail=f"线下账单请上传 Excel 文件（.xlsx / .xlsm / .xls）。收到：{upload.filename}")
    pdf_records = []
    workbook_records = []
    saved_paths: list[Path] = []
    total_pdf_pages = sum(
        _labor_pdf_page_count(Path(str(record.get("path") or "")))
        for record in existing_pdf_records
        if isinstance(record, dict) and str(record.get("path") or "") and Path(str(record.get("path") or "")).exists()
    )
    uploaded_bytes = 0
    try:
        for upload in pdf_files:
            path = await _save_upload_to(
                upload,
                run_dir / safe_labor_filename(upload.filename),
                max_bytes=policy.max_pdf_bytes,
                limit_code="LABOR_PDF_SIZE_LIMIT_EXCEEDED",
            )
            saved_paths.append(path)
            uploaded_bytes += path.stat().st_size
            total_pdf_pages += _labor_pdf_page_count(path)
            if total_pdf_pages > policy.max_pdf_pages:
                raise LaborResourceLimitError(
                    "LABOR_PDF_PAGE_LIMIT_EXCEEDED",
                    f"每个批次 PDF 合计最多 {policy.max_pdf_pages} 页。",
                    limit=policy.max_pdf_pages,
                    details={"pdfPageCount": total_pdf_pages},
                )
            record = attach_labor_file(run_id, path, "PDF发票")
            record["originalFilename"] = upload.filename
            pdf_records.append(record)
        for upload in workbook_files:
            path = await _save_upload_to(
                upload,
                run_dir / safe_labor_filename(upload.filename),
                max_bytes=policy.max_workbook_bytes,
                limit_code="LABOR_WORKBOOK_SIZE_LIMIT_EXCEEDED",
            )
            saved_paths.append(path)
            uploaded_bytes += path.stat().st_size
            record = attach_labor_file(run_id, path, "线下账单")
            record["originalFilename"] = upload.filename
            workbook_records.append(record)
    except LaborResourceLimitError as exc:
        for path in saved_paths:
            path.unlink(missing_ok=True)
        _record_labor_resource_rejection(run_id, owner_user_id, exc.code, limit=exc.limit)
        raise _labor_upload_limit_http_error(exc) from exc
    files["pdfInvoices"] = existing_pdf_records + pdf_records
    files["workbooks"] = existing_workbook_records + workbook_records
    # 兼容旧字段：第一个文件也写入 workbook
    if files["workbooks"]:
        files["workbook"] = files["workbooks"][0]
    actor_user_id, _ = _labor_request_actor(request)
    audit_details = {
        "pdfFileCount": len(pdf_records),
        "workbookFileCount": len(workbook_records),
        "pdfFileCountTotal": len(files["pdfInvoices"]),
        "workbookFileCountTotal": len(files["workbooks"]),
        "pdfPageCount": total_pdf_pages,
        "uploadedBytes": uploaded_bytes,
    }
    updated = update_labor_metadata(
        run_id,
        _labor_invalidate_official_result(
            metadata,
            status="已上传文件",
            reason_code="input_files_changed",
            message="PDF 或 Excel 已更新，旧核对结果已失效，必须重新确认映射并执行整批核对。",
            files=files,
        ),
        actor_user_id=actor_user_id,
        action="files_uploaded",
        audit_details=audit_details,
    )
    if not labor_postgres_state_enabled():
        append_labor_audit_event(
            _labor_audit_path(),
            action="files_uploaded",
            run_id=run_id,
            owner_user_id=owner_user_id,
            actor_user_id=actor_user_id,
            outcome="success",
            details=audit_details,
        )
    return updated


def _labor_ready_private_records(metadata: dict, key: str) -> list[dict]:
    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    records = files.get(key) if isinstance(files.get(key), list) else []
    ready = []
    for record in records:
        if not isinstance(record, dict):
            continue
        try:
            size_bytes = int(record.get("sizeBytes") or 0)
        except (TypeError, ValueError):
            continue
        if (
            str(record.get("uploadState") or "").strip().lower() == "ready"
            and str(record.get("objectKey") or "").strip()
            and re.fullmatch(r"[0-9a-f]{64}", str(record.get("sha256") or "").strip().lower())
            and size_bytes > 0
        ):
            ready.append(dict(record))
    return ready


def _mapping_preflight_snapshot(metadata: dict) -> dict:
    preflight = metadata.get("mappingPreflight") if isinstance(metadata.get("mappingPreflight"), dict) else {}
    if (
        str(preflight.get("status") or "") != "completed"
        or str(preflight.get("inputFingerprint") or "").strip().lower()
        != _labor_mapping_input_fingerprint(metadata)
    ):
        status = str(preflight.get("status") or "not_started")
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="本人核对助手尚未完成 Excel 字段预检。",
                error_code="LABOR_MAPPING_PREFLIGHT_PENDING",
                retryable=True,
                next_action=(
                    "请保持本人核对助手在线，等待其读取工作表和列名后再试。"
                    if status in {"queued", "running"}
                    else "请点击“读取工作表”，启动本人核对助手字段预检。"
                ),
            ),
        )
    return preflight


def _bounded_mapping_preview_value(value):
    if value is None or isinstance(value, (bool, int, float)):
        return value
    return str(value)[:500]


def _normalize_mapping_preflight_result(metadata: dict, payload: dict) -> dict:
    expected = _labor_ready_private_records(metadata, "workbooks")
    expected_by_id = {str(item.get("id") or "").strip(): item for item in expected}
    if not expected_by_id or "" in expected_by_id:
        raise LaborWorkerLeaseError("当前批次缺少可验证的私有 Excel 文件清单。")
    raw_workbooks = payload.get("workbooks") if isinstance(payload.get("workbooks"), list) else []
    if len(raw_workbooks) != len(expected_by_id) or len(raw_workbooks) > 20:
        raise LaborWorkerLeaseError("字段预检结果未覆盖当前批次全部 Excel 文件。")
    normalized_workbooks = []
    seen_ids: set[str] = set()
    union_sheets: list[str] = []
    for raw_workbook in raw_workbooks:
        if not isinstance(raw_workbook, dict):
            raise LaborWorkerLeaseError("字段预检 Excel 结果格式无效。")
        file_id = str(raw_workbook.get("fileId") or "").strip()
        if file_id not in expected_by_id or file_id in seen_ids:
            raise LaborWorkerLeaseError("字段预检结果包含未知或重复的 Excel 文件。")
        seen_ids.add(file_id)
        raw_sheets = raw_workbook.get("sheets") if isinstance(raw_workbook.get("sheets"), list) else []
        if not raw_sheets or len(raw_sheets) > 100:
            raise LaborWorkerLeaseError("字段预检工作表数量无效。")
        normalized_sheets = []
        seen_sheet_names: set[str] = set()
        for raw_sheet in raw_sheets:
            if not isinstance(raw_sheet, dict):
                raise LaborWorkerLeaseError("字段预检工作表结果格式无效。")
            name = str(raw_sheet.get("name") or "").strip()[:200]
            suggestion = raw_sheet.get("suggestion") if isinstance(raw_sheet.get("suggestion"), dict) else {}
            if not name or name in seen_sheet_names:
                raise LaborWorkerLeaseError("字段预检包含空白或重复工作表名称。")
            seen_sheet_names.add(name)
            raw_headers = suggestion.get("headers")
            if not isinstance(raw_headers, list):
                raise LaborWorkerLeaseError("字段预检列名列表格式无效。")
            headers = [str(value)[:200] for value in raw_headers[:200]]
            suggested = suggestion.get("suggestedMapping") if isinstance(suggestion.get("suggestedMapping"), dict) else {}
            suggested_mapping = {
                key: str(suggested.get(key) or "")[:200]
                for key in ("employeeId", "name", "hours", "amount", "currency")
            }
            raw_amount_candidates = suggestion.get("amountColumnCandidates")
            if not isinstance(raw_amount_candidates, list):
                raise LaborWorkerLeaseError("字段预检金额候选列表格式无效。")
            amount_candidates = [
                str(value)[:200]
                for value in raw_amount_candidates[:100]
            ]
            raw_preview_rows = suggestion.get("previewRows")
            if not isinstance(raw_preview_rows, list):
                raise LaborWorkerLeaseError("字段预检样例行列表格式无效。")
            preview_rows = []
            for raw_row in raw_preview_rows[:20]:
                if not isinstance(raw_row, dict):
                    raise LaborWorkerLeaseError("字段预检样例行格式无效。")
                preview_rows.append(
                    {
                        str(key)[:200]: _bounded_mapping_preview_value(value)
                        for key, value in list(raw_row.items())[:200]
                    }
                )
            normalized_sheets.append(
                {
                    "name": name,
                    "suggestion": {
                        "sheetName": name,
                        "headers": headers,
                        "suggestedMapping": suggested_mapping,
                        "amountColumnCandidates": amount_candidates,
                        "previewRows": preview_rows,
                    },
                }
            )
            if name not in union_sheets:
                union_sheets.append(name)
        expected_record = expected_by_id[file_id]
        normalized_workbooks.append(
            {
                "fileId": file_id,
                "filename": Path(
                    str(expected_record.get("originalFilename") or expected_record.get("filename") or "workbook")
                ).name,
                "sheets": normalized_sheets,
            }
        )
    return {"workbooks": normalized_workbooks, "sheets": union_sheets}


@app.post("/api/labor/runs/{run_id}/mapping-preflight")
def start_labor_mapping_preflight(request: Request, run_id: str) -> dict:
    if not _uses_personal_labor_worker():
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前环境未启用本人核对助手字段预检。",
                error_code="LABOR_MAPPING_PREFLIGHT_WORKER_REQUIRED",
                next_action="本地持久化环境请直接读取 Excel；P1 环境请先启用本人核对助手。",
            ),
        )
    metadata = _labor_metadata_or_404(run_id)
    if not _labor_ready_private_records(metadata, "workbooks"):
        raise HTTPException(status_code=400, detail="请先完成线下账单 Excel 的私有上传。")
    owner_user_id = str(metadata.get("ownerUserId") or "local-default")
    actor_user_id, _ = _labor_request_actor(request)
    fingerprint = _labor_mapping_input_fingerprint(metadata)
    generation = uuid4().hex
    queued, started = begin_labor_mapping_preflight(
        run_id,
        task_generation_id=generation,
        input_fingerprint=fingerprint,
        actor_user_id=actor_user_id,
    )
    preflight = queued.get("mappingPreflight") if isinstance(queued.get("mappingPreflight"), dict) else {}
    current_generation = str(preflight.get("taskGenerationId") or generation).strip()
    job = get_latest_labor_worker_job(
        run_id,
        task_generation_id=current_generation,
        job_type="mapping_preflight",
        statuses={"queued", "running", "retry_wait"},
    )
    if started:
        try:
            job = enqueue_labor_worker_job(
                run_id,
                owner_user_id=owner_user_id,
                required_worker_version=_labor_required_worker_version(),
                task_generation_id=current_generation,
                job_type="mapping_preflight",
            )
        except Exception as exc:
            update_labor_metadata_for_mapping_preflight(
                run_id,
                expected_task_generation_id=current_generation,
                updates=lambda current: {
                    "mappingPreflight": {
                        **(current.get("mappingPreflight") if isinstance(current.get("mappingPreflight"), dict) else {}),
                        "status": "failed",
                        "statusLabel": "字段预检提交失败",
                        "errorCode": "LABOR_MAPPING_PREFLIGHT_HANDOFF_FAILED",
                        "errorMessage": str(exc)[:300],
                    }
                },
                actor_user_id=actor_user_id,
                action="mapping_preflight_handoff_failed",
                reason_code="mapping_preflight_handoff_failed",
            )
            raise
    return {
        "mappingPreflight": preflight,
        "workerTask": _public_labor_worker_job(job),
        "started": started,
    }


@app.get("/api/labor/runs/{run_id}/mapping-preflight-status")
def get_labor_mapping_preflight_status(run_id: str) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    preflight = metadata.get("mappingPreflight") if isinstance(metadata.get("mappingPreflight"), dict) else {}
    public_fields = (
        "status",
        "statusLabel",
        "message",
        "taskGenerationId",
        "requestedAt",
        "completedAt",
        "errorCode",
        "errorMessage",
    )
    status = {
        key: preflight.get(key)
        for key in public_fields
        if preflight.get(key) not in (None, "")
    }
    status.setdefault("status", "not_started")
    if status["status"] == "completed":
        sheets = preflight.get("sheets")
        workbooks = preflight.get("workbooks")
        status["sheets"] = [
            str(sheet)
            for sheet in (sheets if isinstance(sheets, list) else [])
            if str(sheet).strip()
        ]
        status["workbooks"] = workbooks if isinstance(workbooks, list) else []
    return {"mappingPreflight": status}


@app.post("/api/labor/runs/{run_id}/direct-upload-plan")
def create_labor_direct_upload_plan(run_id: str, payload: dict = Body(...)) -> dict:
    if not labor_supabase_storage_enabled():
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前环境未启用 Supabase 直传。",
                error_code="LABOR_DIRECT_UPLOAD_UNAVAILABLE",
                next_action="请使用普通上传，或联系管理员检查 Supabase Storage 配置。",
                retryable=False,
            ),
        )
    try:
        run_dir = get_labor_run_dir(run_id)
        load_labor_metadata(run_dir)
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    pdf_files = payload.get("pdfFiles") or []
    workbook_files = payload.get("workbookFiles") or []
    if not isinstance(pdf_files, list) or not isinstance(workbook_files, list):
        raise HTTPException(status_code=400, detail="上传文件清单格式不正确。")
    if not pdf_files:
        raise HTTPException(status_code=400, detail="请至少上传一张 PDF 发票。")
    if not workbook_files:
        raise HTTPException(status_code=400, detail="请上传线下账单 Excel 文件。")

    uploads = []
    try:
        for item in pdf_files:
            uploads.append(_build_labor_direct_upload_item(run_id, item, group="pdfInvoices"))
        for item in workbook_files:
            uploads.append(_build_labor_direct_upload_item(run_id, item, group="workbooks"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        logger.exception("labor direct upload plan failed")
        raise HTTPException(
            status_code=503,
            detail=_labor_request_error(
                message="生成 Supabase 直传地址失败。",
                error_code="LABOR_DIRECT_UPLOAD_PLAN_FAILED",
                retryable=True,
                next_action="请稍后重试；若连续失败，请联系管理员检查 Supabase Storage 配置。",
            ),
        ) from exc
    return {"runId": run_id, "uploads": uploads}


@app.post("/api/labor/runs/{run_id}/direct-upload-complete")
def complete_labor_direct_upload(run_id: str, payload: dict = Body(...)) -> dict:
    if not labor_supabase_storage_enabled():
        raise HTTPException(status_code=409, detail="当前环境未启用 Supabase 直传。")
    try:
        run_dir = get_labor_run_dir(run_id)
        metadata = load_labor_metadata(run_dir)
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    uploads = payload.get("uploads") or []
    if not isinstance(uploads, list) or not uploads:
        raise HTTPException(status_code=400, detail="缺少已上传文件清单。")
    try:
        sync_labor_run_from_persistent(run_id, run_dir)
    except Exception as exc:
        logger.exception("labor direct upload sync failed")
        raise HTTPException(
            status_code=503,
            detail=_labor_request_error(
                message="文件已直传，但服务器同步文件失败。",
                error_code="LABOR_DIRECT_UPLOAD_SYNC_FAILED",
                retryable=True,
                next_action="请稍后重试“上传文件”；若连续失败，请联系管理员检查 Supabase Storage。",
            ),
        ) from exc

    pdf_records = []
    workbook_records = []
    try:
        for item in uploads:
            group = str(item.get("group") or "").strip()
            relative_path = _validate_labor_direct_upload_relative_path(item.get("relativePath"))
            path = run_dir / relative_path
            if not path.exists():
                raise FileNotFoundError(path.name)
            if group == "pdfInvoices":
                record = attach_labor_file(run_id, path, "PDF发票")
                record["originalFilename"] = str(item.get("originalFilename") or item.get("filename") or path.name)
                pdf_records.append(record)
            elif group == "workbooks":
                workbook_records.append(attach_labor_file(run_id, path, "线下账单"))
            else:
                raise ValueError("直传文件分组不正确。")
    except (ValueError, FileNotFoundError) as exc:
        raise HTTPException(
            status_code=400,
            detail=_labor_request_error(
                message="直传文件校验失败。",
                error_code="LABOR_DIRECT_UPLOAD_FILE_MISSING",
                next_action="请重新选择 PDF 发票和 Excel 账单并上传。",
                requires_reupload=True,
            ),
        ) from exc
    if not pdf_records or not workbook_records:
        raise HTTPException(status_code=400, detail="请上传 PDF 发票和 Excel 账单。")
    files = dict(metadata.get("files", {}))
    files["pdfInvoices"] = pdf_records
    files["workbooks"] = workbook_records
    files["workbook"] = workbook_records[0]
    return update_labor_metadata_record_only(run_id, {"status": "已上传文件", "files": files})


@app.get("/api/labor/runs/{run_id}/workbook-sheets")
def labor_workbook_sheets(run_id: str) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    if _labor_ready_private_records(metadata, "workbooks"):
        preflight = _mapping_preflight_snapshot(metadata)
        return {
            "sheets": list(preflight.get("sheets") or []),
            "fileCount": len(preflight.get("workbooks") or []),
            "source": "personal_worker_preflight",
        }
    paths = _labor_workbook_paths(metadata)
    try:
        if len(paths) == 1:
            return {"sheets": list_workbook_sheets(paths[0])}
        # 多文件：返回去重的 sheet 名（用户按名称选择，读取时合并所有文件）
        all_sheets: list[str] = []
        seen: set[str] = set()
        for p in paths:
            for name in list_workbook_sheets(p):
                if name not in seen:
                    seen.add(name)
                    all_sheets.append(name)
        return {"sheets": all_sheets, "fileCount": len(paths)}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"读取 Excel 工作表失败：{exc}") from exc


@app.post("/api/labor/runs/{run_id}/field-suggestions")
def labor_field_suggestions(run_id: str, payload: dict = Body(...)) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    sheet_name = str(payload.get("sheet_name") or payload.get("sheetName") or "").strip()
    if not sheet_name:
        raise HTTPException(status_code=400, detail="请选择 Excel 工作表。")
    if _labor_ready_private_records(metadata, "workbooks"):
        preflight = _mapping_preflight_snapshot(metadata)
        for workbook in preflight.get("workbooks") or []:
            if not isinstance(workbook, dict):
                continue
            for sheet in workbook.get("sheets") or []:
                if isinstance(sheet, dict) and str(sheet.get("name") or "") == sheet_name:
                    suggestion = sheet.get("suggestion") if isinstance(sheet.get("suggestion"), dict) else {}
                    return dict(suggestion)
        raise HTTPException(status_code=400, detail="字段预检结果中找不到所选工作表，请重新读取。")
    paths = _labor_workbook_paths(metadata)
    try:
        # 从第一个文件读取字段映射建议（所有文件结构应一致）
        return suggest_mapping(paths[0], sheet_name)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/labor/runs/{run_id}/mapping")
def save_labor_mapping(
    run_id: str,
    payload: dict = Body(...),
    request: Request = None,
) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    sheet_name = str(payload.get("sheet_name") or payload.get("sheetName") or "").strip()
    mapping = payload.get("mapping") or {}
    manual_name_mapping = payload.get("manualNameMapping") or payload.get("manual_name_mapping") or payload.get("manualMapping") or {}
    if not sheet_name:
        raise HTTPException(status_code=400, detail="请选择 Excel 工作表。")
    for field in ("name", "hours", "amount"):
        if not mapping.get(field):
            raise HTTPException(status_code=400, detail="字段映射缺少姓名、工时或金额。")
    amount_columns = mapping.get("amountColumns") or mapping.get("amount_columns") or []
    if amount_columns and not isinstance(amount_columns, list):
        raise HTTPException(status_code=400, detail="叠加金额列格式无效。")
    if isinstance(amount_columns, list):
        mapping["amountColumns"] = list(
            dict.fromkeys(
                str(value).strip()
                for value in [mapping.get("amount"), *amount_columns]
                if str(value or "").strip()
            )
        )
    amount_scope = str(mapping.get("amountScope") or mapping.get("amount_scope") or "auto").strip().lower()
    mapping["amountScope"] = amount_scope if amount_scope in {"auto", "net", "gross"} else "auto"
    if _labor_ready_private_records(metadata, "workbooks"):
        preflight = _mapping_preflight_snapshot(metadata)
        workbook_headers: list[set[str]] = []
        missing_sheet_workbooks: list[str] = []
        for workbook in preflight.get("workbooks") or []:
            if not isinstance(workbook, dict):
                continue
            selected_sheet = next(
                (
                    item
                    for item in workbook.get("sheets") or []
                    if isinstance(item, dict) and str(item.get("name") or "") == sheet_name
                ),
                None,
            )
            if selected_sheet is None:
                missing_sheet_workbooks.append(str(workbook.get("filename") or "Excel"))
                continue
            suggestion = (
                selected_sheet.get("suggestion")
                if isinstance(selected_sheet.get("suggestion"), dict)
                else {}
            )
            workbook_headers.append(
                {
                    str(value)
                    for value in suggestion.get("headers") or []
                    if str(value or "").strip()
                }
            )
        if missing_sheet_workbooks or not workbook_headers:
            raise HTTPException(
                status_code=400,
                detail=f"部分 Excel 不包含所选工作表 {sheet_name}，请重新读取并选择共同的明细表。",
            )
        selected_columns = {
            str(value).strip()
            for value in [
                mapping.get("employeeId"),
                mapping.get("name"),
                mapping.get("hours"),
                mapping.get("amount"),
                mapping.get("currency"),
                *(mapping.get("amountColumns") or []),
            ]
            if str(value or "").strip()
        }
        unknown_columns = sorted(
            column
            for column in selected_columns
            if any(column not in headers for headers in workbook_headers)
        )
        if unknown_columns:
            raise HTTPException(
                status_code=400,
                detail=f"所选工作表不包含字段：{'、'.join(unknown_columns)}。请等待字段建议加载完成后重新确认。",
            )
    actor_user_id, _ = _labor_request_actor(request) if request is not None else (
        str(metadata.get("ownerUserId") or "local-default"),
        False,
    )
    return update_labor_metadata(
        run_id,
        {
            **_labor_invalidate_official_result(
                metadata,
                status="已确认字段",
                reason_code="workbook_mapping_changed",
                message="Excel 工作表或字段映射已更新，旧核对结果已失效，必须重新执行整批核对。",
            ),
            "workbookSheet": sheet_name,
            "excelMapping": mapping,
            "manualNameMapping": manual_name_mapping,
        },
        actor_user_id=actor_user_id,
        action="workbook_mapping_changed",
        reason_code="workbook_mapping_changed",
    )


@app.post("/api/labor/runs/{run_id}/extract-and-compare")
async def extract_and_compare_labor_run(run_id: str, request: Request = None) -> dict:
    request_scoped_runtime = _uses_request_scoped_labor_runtime()
    personal_worker_enabled = _uses_personal_labor_worker()
    if request_scoped_runtime and not personal_worker_enabled:
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前 Vercel UAT 仅支持页面试用和测试材料验证，不启动正式在线抽取任务。",
                error_code="LABOR_UAT_EXTRACT_DISABLED",
                next_action="请使用“测试材料验证”查看样例流程；正式抽取请在本地/内网持久化环境执行。",
            ),
        )
    metadata = _check_stale_extracting(_labor_metadata_or_404(run_id))
    async_task = metadata.get("asyncTask") if isinstance(metadata.get("asyncTask"), dict) else {}
    if metadata.get("status") == "抽取中" and async_task.get("status") in {"queued", "running"}:
        return _with_labor_readiness(_normalize_labor_total_decision(metadata))
    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    pdf_records = [record for record in files.get("pdfInvoices", []) if isinstance(record, dict)]
    workbook_records = [record for record in files.get("workbooks", []) if isinstance(record, dict)]

    def available(record: dict) -> bool:
        local_path = Path(str(record.get("path") or ""))
        if local_path.is_file():
            return True
        return bool(
            personal_worker_enabled
            and str(record.get("uploadState") or "").strip().lower() == "ready"
            and str(record.get("objectKey") or "").strip()
            and re.fullmatch(r"[0-9a-f]{64}", str(record.get("sha256") or "").strip().lower())
            and int(record.get("sizeBytes") or 0) > 0
        )

    available_pdfs = [record for record in pdf_records if available(record)]
    available_workbooks = [record for record in workbook_records if available(record)]
    if not available_pdfs or not available_workbooks:
        missing_parts = []
        if not available_pdfs:
            missing_parts.append("PDF 发票")
        if not available_workbooks:
            missing_parts.append("Excel 账单")
        missing_text = "、".join(missing_parts)
        raise HTTPException(
            status_code=400,
            detail=_labor_request_error(
                message=f"请先上传本期 {missing_text}。",
                error_code="LABOR_FILES_REQUIRED",
                next_action="请返回「上传文件」步骤，上传供应商 PDF 发票和线下账单 Excel 后再生成核对结果。",
                requires_reupload=True,
            ),
        )
    mapping = metadata.get("excelMapping") or {}
    sheet_name = metadata.get("workbookSheet") or ""
    if not sheet_name or not mapping:
        raise HTTPException(
            status_code=400,
            detail=_labor_request_error(
                message="请先确认 Excel 工作表和字段映射。",
                error_code="LABOR_MAPPING_REQUIRED",
                next_action="请在「字段映射」步骤选择工作表，并确认姓名、工时、金额字段。",
            ),
        )
    owner_user_id = str(metadata.get("ownerUserId") or "local-default")
    actor_user_id = (
        _labor_request_actor(request)[0]
        if request is not None and labor_auth_required()
        else owner_user_id
    )
    reservation_token = ""
    if not personal_worker_enabled:
        try:
            reservation_token = _LABOR_TASK_LIMITER.reserve(
                run_id,
                owner_user_id,
                policy=_labor_hardening_policy(),
                active_runs=[_check_stale_extracting(item) for item in list_labor_metadata()],
            )
        except LaborResourceLimitError as exc:
            _record_labor_resource_rejection(run_id, owner_user_id, exc.code, limit=exc.limit)
            raise HTTPException(
                status_code=409 if exc.code == "LABOR_RUN_ALREADY_ACTIVE" else 429,
                detail=_labor_request_error(
                    message=str(exc),
                    error_code=exc.code,
                    retryable=True,
                    next_action="请等待正在运行的核对任务完成后再重试。",
                ),
            ) from exc
    queued_at = datetime.now().isoformat(timespec="seconds")
    task_generation_id = uuid4().hex
    try:
        queued, started = begin_labor_metadata_task(
            run_id,
            task_generation_id=task_generation_id,
            actor_user_id=actor_user_id,
            updates=lambda current: {
                **_labor_invalidate_official_result(
                    current,
                    status="抽取中",
                    reason_code="new_extraction_started",
                    message="已开始重新核对；上一版机器结果和报告已失效，必须等待本次整批核对完成。",
                ),
                "inputSnapshotRejected": False,
                "status": "抽取中",
                "stage": "等待后台处理",
                "progress": {
                    "status": "queued",
                    "phase": "queued",
                    "phaseLabel": "等待后台处理",
                    "message": "核对任务已提交，等待后台开始处理。",
                    "startedAt": queued_at,
                    "lastUpdatedAt": queued_at,
                },
                "asyncTask": {
                    "status": "queued",
                    "statusLabel": "待处理",
                    "message": "核对任务已提交，等待后台处理。",
                    "queuedAt": queued_at,
                    "taskGenerationId": task_generation_id,
                },
                "errorMessage": "",
                "errorCode": "",
                "failureType": "",
                "retryable": False,
                "requiresReupload": False,
                "businessReviewStatus": "pending",
                "manualReviewRequired": True,
                "directPaymentAllowed": False,
                "requiresHumanReview": True,
                "nextAction": "",
                "diffDownloadUrl": "",
                "invoiceEvidenceAudit": [],
            },
        )
    except Exception:
        if reservation_token:
            _LABOR_TASK_LIMITER.release(run_id, reservation_token)
        raise
    if not started:
        if reservation_token:
            _LABOR_TASK_LIMITER.release(run_id, reservation_token)
        return _with_labor_readiness(_normalize_labor_total_decision(queued))
    try:
        append_labor_audit_event(
            _labor_audit_path(),
            action="extraction_started",
            run_id=run_id,
            owner_user_id=owner_user_id,
            actor_user_id=actor_user_id,
            outcome="success",
        )
        if personal_worker_enabled:
            # Keep the run generation check and queue handoff in one run -> job
            # critical section. Otherwise an older request can resume after a
            # newer generation enqueues and incorrectly supersede the new job.
            with labor_run_metadata_lock(run_id):
                current = load_labor_metadata(get_labor_run_dir(run_id))
                if _labor_task_generation_id(current) != task_generation_id:
                    return _with_labor_readiness(
                        _normalize_labor_total_decision(_with_personal_worker_status(current))
                    )
                job = enqueue_labor_worker_job(
                    run_id,
                    owner_user_id=owner_user_id,
                    required_worker_version=_labor_required_worker_version(),
                    task_generation_id=task_generation_id,
                )
            return _with_labor_readiness(
                {
                    **_normalize_labor_total_decision(queued),
                    "workerTask": _public_labor_worker_job(job),
                }
            )
        # 在独立线程中运行，不阻塞事件循环
        asyncio.get_event_loop().run_in_executor(
            None,
            _run_labor_extract_compare_with_release,
            run_id,
            task_generation_id,
            reservation_token,
        )
        return queued
    except Exception as exc:
        if reservation_token:
            _LABOR_TASK_LIMITER.release(run_id, reservation_token)
        _mark_labor_task_handoff_failed(run_id, task_generation_id, exc)
        raise


def _mark_labor_task_handoff_failed(run_id: str, task_generation_id: str, exc: BaseException) -> None:
    message = f"核对任务提交失败：{str(exc)[:300]}"
    failure = _labor_retryable_system_failure(
        message=message,
        stage="任务提交失败",
        error_code="LABOR_TASK_HANDOFF_FAILED",
    )
    failure["asyncTask"] = {
        "status": "failed",
        "statusLabel": "提交失败",
        "message": message,
        "failedAt": utcnow_naive().isoformat(),
        "taskGenerationId": task_generation_id,
    }
    try:
        update_labor_metadata_for_task(
            run_id,
            expected_task_generation_id=task_generation_id,
            updates=failure,
        )
    except Exception as update_exc:  # noqa: BLE001 - preserve the original handoff failure.
        logger.error("[%s] 任务提交失败状态写入失败: %s", run_id, update_exc)


@app.get("/api/labor/runs/{run_id}/governance")
def labor_rule_governance(run_id: str) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    return _normalized_labor_governance(metadata.get("ruleGovernance"))


@app.post("/api/labor/runs/{run_id}/reocr-candidates/replay")
def replay_labor_reocr_candidate(run_id: str, payload: dict = Body(...)) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    task = payload.get("task") if isinstance(payload.get("task"), dict) else {}
    candidate_rows = payload.get("candidateRows") or payload.get("candidate_rows") or []
    if not task:
        raise HTTPException(status_code=400, detail="请提供图片识别复核任务 task。")
    if not isinstance(candidate_rows, list) or not candidate_rows:
        raise HTTPException(status_code=400, detail="请提供新的图片识别员工明细 candidateRows。")
    sheet_name = str(metadata.get("workbookSheet") or "").strip()
    mapping = metadata.get("excelMapping") or {}
    if not sheet_name or not mapping:
        raise HTTPException(status_code=400, detail="请先确认 Excel 工作表和字段映射。")
    try:
        excel_rows = _labor_excel_rows_from_metadata(metadata)
        replay = replay_reocr_candidate_result(
            task,
            candidate_rows,
            excel_rows,
            amount_tolerance=float(payload.get("amountTolerance") or payload.get("amount_tolerance") or AI_CONFIG["amount_tolerance"]),
            hours_tolerance=float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"]),
            confidence_threshold=float(payload.get("confidenceThreshold") or payload.get("confidence_threshold") or AI_CONFIG["confidence_threshold"]),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    replay["runId"] = run_id
    replay["replayedAt"] = utcnow_naive().isoformat()
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    replays = list(governance.get("replays") or [])
    replays.append(replay)
    governance["replays"] = replays[-50:]
    update_labor_metadata(run_id, {"reocrReplayGovernance": governance})
    return replay


@app.post("/api/labor/runs/{run_id}/reocr-candidates/replay-file")
async def replay_labor_reocr_candidate_file(
    run_id: str,
    candidate_file: UploadFile = File(...),
    task: str = Form(...),
    amount_tolerance: Optional[float] = Form(default=None),
    hours_tolerance: Optional[float] = Form(default=None),
    confidence_threshold: Optional[float] = Form(default=None),
) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    if not candidate_file.filename:
        raise HTTPException(status_code=400, detail="请上传图片识别结果文件。")
    if not candidate_file.filename.lower().endswith((".csv", ".xlsx", ".xlsm", ".xls")):
        raise HTTPException(status_code=400, detail="图片识别结果文件仅支持 .csv / .xlsx / .xlsm / .xls。")
    try:
        task_data = json.loads(task)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="task 必须是 JSON 对象。") from exc
    if not isinstance(task_data, dict) or not task_data:
        raise HTTPException(status_code=400, detail="task 必须是非空 JSON 对象。")

    run_dir = get_labor_run_dir(run_id)
    candidate_path = await _save_upload_to(candidate_file, run_dir / safe_labor_filename(candidate_file.filename))
    try:
        candidate_rows = parse_reocr_candidate_rows(candidate_path, default_currency=str(metadata.get("currency") or ""))
        excel_rows = _labor_excel_rows_from_metadata(metadata)
        replay = replay_reocr_candidate_result(
            task_data,
            candidate_rows,
            excel_rows,
            amount_tolerance=float(amount_tolerance if amount_tolerance is not None else AI_CONFIG["amount_tolerance"]),
            hours_tolerance=float(hours_tolerance if hours_tolerance is not None else AI_CONFIG["hours_tolerance"]),
            confidence_threshold=float(confidence_threshold if confidence_threshold is not None else AI_CONFIG["confidence_threshold"]),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    replay["runId"] = run_id
    replay["replayedAt"] = utcnow_naive().isoformat()
    replay["candidateFile"] = candidate_file.filename
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    replays = list(governance.get("replays") or [])
    replays.append(replay)
    governance["replays"] = replays[-50:]
    files = dict(metadata.get("files", {}))
    candidate_records = list(files.get("reocrCandidateFiles") or [])
    candidate_records.append(attach_labor_file(run_id, candidate_path, "图片识别结果文件"))
    files["reocrCandidateFiles"] = candidate_records[-20:]
    update_labor_metadata(run_id, {"reocrReplayGovernance": governance, "files": files})
    replay["parsedCandidateRows"] = candidate_rows[:50]
    replay["parsedCandidateRowCount"] = len(candidate_rows)
    return replay


@app.post("/api/labor/runs/{run_id}/reocr-candidates/replay-file-batch")
async def replay_labor_reocr_candidate_file_batch(
    run_id: str,
    candidate_file: UploadFile = File(...),
    amount_tolerance: Optional[float] = Form(default=None),
    hours_tolerance: Optional[float] = Form(default=None),
    confidence_threshold: Optional[float] = Form(default=None),
) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    if not candidate_file.filename:
        raise HTTPException(status_code=400, detail="请上传图片识别批量结果文件。")
    if not candidate_file.filename.lower().endswith((".csv", ".xlsx", ".xlsm", ".xls")):
        raise HTTPException(status_code=400, detail="图片识别批量结果文件仅支持 .csv / .xlsx / .xlsm / .xls。")

    run_dir = get_labor_run_dir(run_id)
    candidate_path = await _save_upload_to(candidate_file, run_dir / safe_labor_filename(candidate_file.filename))
    try:
        candidate_rows = parse_reocr_candidate_rows(candidate_path, default_currency=str(metadata.get("currency") or ""))
        grouped_rows = _group_reocr_candidate_rows_by_scope(candidate_rows)
        excel_rows = _labor_excel_rows_from_metadata(metadata)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not grouped_rows:
        raise HTTPException(status_code=400, detail="批量候选文件缺少 SourceFile/WarehouseId，无法按 PDF 分组回放。")

    upload_coverage = _build_reocr_upload_coverage(metadata, grouped_rows)
    replays = []
    errors = []
    for (source_file, warehouse_id), rows in grouped_rows.items():
        try:
            task = _reocr_task_for_batch_candidate(metadata, excel_rows, source_file, warehouse_id)
            replay = replay_reocr_candidate_result(
                task,
                rows,
                excel_rows,
                amount_tolerance=float(amount_tolerance if amount_tolerance is not None else task.get("amountTolerance") or AI_CONFIG["amount_tolerance"]),
                hours_tolerance=float(hours_tolerance if hours_tolerance is not None else AI_CONFIG["hours_tolerance"]),
                confidence_threshold=float(confidence_threshold if confidence_threshold is not None else AI_CONFIG["confidence_threshold"]),
            )
            replay["runId"] = run_id
            replay["replayedAt"] = utcnow_naive().isoformat()
            replay["candidateFile"] = candidate_file.filename
            replay["batchUpload"] = True
            replays.append(replay)
        except ValueError as exc:
            errors.append({"sourceFile": source_file, "warehouseId": warehouse_id, "error": str(exc)})

    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    existing_replays = list(governance.get("replays") or [])
    existing_replays.extend(replays)
    governance["replays"] = existing_replays[-50:]
    files = dict(metadata.get("files", {}))
    candidate_records = list(files.get("reocrCandidateFiles") or [])
    record = attach_labor_file(run_id, candidate_path, "图片识别批量结果文件")
    record["summary"] = {
        "groupCount": len(grouped_rows),
        "replayedCount": len(replays),
        "readyCount": sum(1 for replay in replays if replay.get("decision") == "ready_for_user_confirmation"),
        "blockedCount": sum(1 for replay in replays if replay.get("decision") != "ready_for_user_confirmation"),
        "errorCount": len(errors),
        "parsedRowCount": len(candidate_rows),
        "plannedTaskCount": upload_coverage["plannedTaskCount"],
        "coveredTaskCount": upload_coverage["coveredTaskCount"],
        "missingTaskCount": upload_coverage["missingTaskCount"],
        "extraScopeCount": upload_coverage["extraScopeCount"],
    }
    record["coverage"] = upload_coverage
    candidate_records.append(record)
    files["reocrCandidateFiles"] = candidate_records[-20:]
    update_labor_metadata(run_id, {"reocrReplayGovernance": governance, "files": files})
    return {
        "decision": "batch_file_replay_completed" if replays else "batch_file_replay_failed",
        "requiresConfirmation": True,
        "summary": record["summary"],
        "coverage": upload_coverage,
        "results": replays,
        "errors": errors,
        "candidateFile": record,
    }


@app.post("/api/labor/runs/{run_id}/reocr-candidates/replay-cache")
def replay_labor_reocr_candidate_cache(run_id: str, payload: dict = Body(...)) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    task = payload.get("task") if isinstance(payload.get("task"), dict) else {}
    source_file = str(task.get("sourceFile") or payload.get("sourceFile") or payload.get("source_file") or "").strip()
    if not task:
        raise HTTPException(status_code=400, detail="请提供图片识别复核任务 task。")
    if not source_file:
        raise HTTPException(status_code=400, detail="请提供 sourceFile。")
    pdf_path = _labor_pdf_path_for_source(metadata, source_file)
    try:
        replay = replay_ai_cache_candidate_result(
            task,
            pdf_path,
            _labor_excel_rows_from_metadata(metadata),
            amount_tolerance=float(payload.get("amountTolerance") or payload.get("amount_tolerance") or AI_CONFIG["amount_tolerance"]),
            hours_tolerance=float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"]),
            confidence_threshold=float(payload.get("confidenceThreshold") or payload.get("confidence_threshold") or AI_CONFIG["confidence_threshold"]),
            currency=str(metadata.get("currency") or payload.get("currency") or "USD"),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if not replay.get("summary", {}).get("candidateRowCount"):
        raise HTTPException(status_code=400, detail="该 PDF 没有可预览的本地历史图片识别记录。")

    replay["runId"] = run_id
    replay["replayedAt"] = utcnow_naive().isoformat()
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    replays = list(governance.get("replays") or [])
    replays.append(replay)
    governance["replays"] = replays[-50:]
    update_labor_metadata(run_id, {"reocrReplayGovernance": governance})
    return replay


@app.post("/api/labor/runs/{run_id}/reocr-candidates/replay-cache-batch")
def replay_labor_reocr_candidate_cache_batch(run_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    reocr_plan = metadata.get("reocrPlan") if isinstance(metadata.get("reocrPlan"), dict) else {}
    candidates = payload.get("candidates") if isinstance(payload.get("candidates"), list) else reocr_plan.get("reviewableCandidates") or []
    if not candidates:
        raise HTTPException(status_code=400, detail="当前批次没有可批量回放的缓存候选。")

    excel_rows = _labor_excel_rows_from_metadata(metadata)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    replays = list(governance.get("replays") or [])
    results: list[dict] = []
    errors: list[dict] = []
    for candidate in candidates:
        if not isinstance(candidate, dict):
            continue
        task = _cache_replay_task_from_candidate(candidate)
        source_file = str(task.get("sourceFile") or "")
        try:
            replay = replay_ai_cache_candidate_result(
                task,
                _labor_pdf_path_for_source(metadata, source_file),
                excel_rows,
                amount_tolerance=float(payload.get("amountTolerance") or payload.get("amount_tolerance") or task.get("amountTolerance") or AI_CONFIG["amount_tolerance"]),
                hours_tolerance=float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"]),
                confidence_threshold=float(payload.get("confidenceThreshold") or payload.get("confidence_threshold") or AI_CONFIG["confidence_threshold"]),
                currency=str(metadata.get("currency") or payload.get("currency") or "USD"),
            )
            if not replay.get("summary", {}).get("candidateRowCount"):
                raise ValueError("该 PDF 没有可预览的本地历史图片识别记录。")
            replay["runId"] = run_id
            replay["replayedAt"] = utcnow_naive().isoformat()
            replays.append(replay)
            results.append(replay)
        except (HTTPException, ValueError) as exc:
            errors.append(
                {
                    "sourceFile": source_file,
                    "warehouseId": task.get("warehouseId", ""),
                    "error": exc.detail if isinstance(exc, HTTPException) else str(exc),
                }
            )

    governance["replays"] = replays[-50:]
    update_labor_metadata(run_id, {"reocrReplayGovernance": governance})
    ready_count = sum(1 for replay in results if replay.get("decision") == "ready_for_user_confirmation")
    blocked_count = sum(1 for replay in results if replay.get("decision") != "ready_for_user_confirmation")
    return {
        "decision": "batch_cache_replay_completed" if results else "batch_cache_replay_failed",
        "requiresConfirmation": True,
        "summary": {
            "candidateCount": len(candidates),
            "replayedCount": len(results),
            "readyCount": ready_count,
            "blockedCount": blocked_count,
            "errorCount": len(errors),
        },
        "results": results,
        "errors": errors,
    }


@app.post("/api/labor/runs/{run_id}/reocr-candidates/template")
def generate_labor_reocr_candidate_template(run_id: str, payload: dict = Body(...)) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    task = payload.get("task") if isinstance(payload.get("task"), dict) else {}
    source_file = str(task.get("sourceFile") or payload.get("sourceFile") or "").strip()
    warehouse_id = str(task.get("warehouseId") or payload.get("warehouseId") or "").strip()
    if not source_file:
        raise HTTPException(status_code=400, detail="请提供图片识别复核任务 sourceFile。")
    excel_rows = _labor_excel_rows_from_metadata(metadata)
    scoped_rows = [
        row for row in excel_rows
        if not warehouse_id or str(row.warehouse_id or "").strip() == warehouse_id
    ]
    if not scoped_rows:
        raise HTTPException(status_code=400, detail="没有找到该仓库的 Excel 员工明细，无法生成模板。")
    run_dir = get_labor_run_dir(run_id)
    template_path = run_dir / safe_labor_filename(f"{Path(source_file).stem}_reocr_candidate_template.csv", "图片识别结果模板")
    with template_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "SourceFile",
                "WarehouseId",
                "EmployeeId",
                "Employee",
                "Hours",
                "Amount",
                "Page",
                "Confidence",
                "Currency",
                "Evidence",
                "ExcelRef",
                "ExpectedHours",
                "ExpectedAmount",
            ],
        )
        writer.writeheader()
        for row in scoped_rows:
            writer.writerow(
                {
                    "SourceFile": source_file,
                    "WarehouseId": warehouse_id or row.warehouse_id,
                    "EmployeeId": row.employee_id,
                    "Employee": row.employee_name_raw,
                    "Hours": f"{row.hours:.2f}",
                    "Amount": f"{row.amount:.2f}",
                    "Page": "",
                    "Confidence": "0.95",
                    "Currency": row.currency or metadata.get("currency") or "",
                    "Evidence": f"请粘贴 {source_file} 中该员工的 OCR 原文证据",
                    "ExcelRef": f"{row.source_file} {row.source_page_or_row}".strip(),
                    "ExpectedHours": f"{row.hours:.2f}",
                    "ExpectedAmount": f"{row.amount:.2f}",
                }
            )
    files = dict(metadata.get("files", {}))
    template_records = list(files.get("reocrCandidateTemplates") or [])
    template_records.append(attach_labor_file(run_id, template_path, "图片识别结果模板"))
    files["reocrCandidateTemplates"] = template_records[-20:]
    updated = update_labor_metadata(run_id, {"files": files})
    return updated["files"]["reocrCandidateTemplates"][-1]


@app.post("/api/labor/runs/{run_id}/reocr-candidates/template-batch")
def generate_labor_reocr_candidate_template_batch(run_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    reocr_plan = metadata.get("reocrPlan") if isinstance(metadata.get("reocrPlan"), dict) else {}
    tasks = payload.get("tasks") if isinstance(payload.get("tasks"), list) else reocr_plan.get("tasks") or []
    if not tasks:
        raise HTTPException(status_code=400, detail="当前批次没有可导出的图片识别复核任务。")
    excel_rows = _labor_excel_rows_from_metadata(metadata)
    run_dir = get_labor_run_dir(run_id)
    template_path = run_dir / safe_labor_filename("reocr_candidate_batch_template.csv", "图片识别批量结果模板")
    row_count = 0
    task_count = 0
    missing_tasks: list[dict] = []
    with template_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "SourceFile",
                "WarehouseId",
                "EmployeeId",
                "Employee",
                "Hours",
                "Amount",
                "Page",
                "Confidence",
                "Currency",
                "Evidence",
                "ExcelRef",
                "ExpectedHours",
                "ExpectedAmount",
            ],
        )
        writer.writeheader()
        for raw_task in tasks:
            if not isinstance(raw_task, dict):
                continue
            task = _cache_replay_task_from_candidate(raw_task)
            source_file = str(task.get("sourceFile") or "").strip()
            warehouse_id = str(task.get("warehouseId") or "").strip()
            if not source_file:
                continue
            scoped_rows = [
                row for row in excel_rows
                if not warehouse_id or str(row.warehouse_id or "").strip() == warehouse_id
            ]
            if not scoped_rows:
                missing_tasks.append({"sourceFile": source_file, "warehouseId": warehouse_id, "reason": "no_excel_rows"})
                continue
            task_count += 1
            for row in scoped_rows:
                writer.writerow(
                    {
                        "SourceFile": source_file,
                        "WarehouseId": warehouse_id,
                        "EmployeeId": row.employee_id,
                        "Employee": row.employee_name_raw,
                        "Hours": f"{row.hours:.2f}",
                        "Amount": f"{row.amount:.2f}",
                        "Page": "",
                        "Confidence": "0.95",
                        "Currency": row.currency or metadata.get("currency") or "",
                        "Evidence": f"请粘贴 {source_file} 中该员工的 OCR 原文证据",
                        "ExcelRef": f"{row.source_file} {row.source_page_or_row}".strip(),
                        "ExpectedHours": f"{row.hours:.2f}",
                        "ExpectedAmount": f"{row.amount:.2f}",
                    }
                )
                row_count += 1
    files = dict(metadata.get("files", {}))
    template_records = list(files.get("reocrCandidateTemplates") or [])
    record = attach_labor_file(run_id, template_path, "图片识别批量结果模板")
    record["summary"] = {
        "taskCount": task_count,
        "rowCount": row_count,
        "missingTaskCount": len(missing_tasks),
    }
    record["missingTasks"] = missing_tasks
    template_records.append(record)
    files["reocrCandidateTemplates"] = template_records[-20:]
    updated = update_labor_metadata(run_id, {"files": files})
    return updated["files"]["reocrCandidateTemplates"][-1]


@app.post("/api/labor/runs/{run_id}/reocr-candidates/confirm")
def confirm_labor_reocr_candidate(run_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    source_file = str(payload.get("sourceFile") or payload.get("source_file") or "").strip()
    warehouse_id = str(payload.get("warehouseId") or payload.get("warehouse_id") or "").strip()
    if not source_file:
        raise HTTPException(status_code=400, detail="请提供 sourceFile。")
    replay = _latest_reocr_replay(governance["replays"], source_file, warehouse_id)
    if not replay:
        raise HTTPException(status_code=404, detail="未找到该文件的图片识别预览记录。")
    if replay.get("decision") != "ready_for_user_confirmation":
        raise HTTPException(status_code=400, detail="图片识别结果未通过影响预览，不能确认。")

    now = utcnow_naive().isoformat()
    candidate_id = f"reocr_{run_id}_{_safe_record_id(source_file)}_{_safe_record_id(warehouse_id or 'all')}"
    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    active = {
        "candidateId": candidate_id,
        "decision": "active",
        "status": "active",
        "requiresConfirmation": False,
        "sourceFile": source_file,
        "warehouseId": warehouse_id,
        "confirmedBy": actor,
        "confirmationReason": str(payload.get("reason") or "").strip(),
        "confirmedAt": now,
        "replay": replay,
        "auditTrail": [
            {
                "action": "confirmed",
                "actor": actor,
                "reason": str(payload.get("reason") or "").strip(),
                "replayedAt": replay.get("replayedAt", ""),
            }
        ],
    }
    _upsert_reocr_record(governance["activeCandidates"], active)
    files = dict(metadata.get("files", {}))
    response = dict(active)
    if payload.get("generateReport", payload.get("generate_report", True)):
        preview = _build_reocr_projected_preview(metadata, active)
        report_path = get_labor_run_dir(run_id) / safe_labor_storage_filename("labor_reocr_preview_report.xlsx", "image_result_preview")
        build_labor_projection_report(report_path, preview)
        files["reocrPreviewReport"] = attach_labor_file(run_id, report_path, "图片识别结果预览报告")
        response["preview"] = preview
        response["reportFile"] = files["reocrPreviewReport"]
    updated = update_labor_metadata(
        run_id,
        {"reocrReplayGovernance": governance, "files": files},
        actor_user_id=actor,
        action="reocr_candidate_confirmed",
        audit_details={"candidateId": candidate_id},
    )
    response["activeCandidate"] = _find_reocr_record(updated["reocrReplayGovernance"]["activeCandidates"], candidate_id)
    return response


@app.post("/api/labor/runs/{run_id}/reocr-candidates/confirm-batch")
def confirm_labor_reocr_candidate_batch(run_id: str, payload: dict = Body(default={}), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    reason = str(payload.get("reason") or "").strip()
    requested = {
        (Path(str(item.get("sourceFile") or "")).name, str(item.get("warehouseId") or ""))
        for item in payload.get("candidates", [])
        if isinstance(item, dict) and item.get("sourceFile")
    }
    existing_scopes = {
        (Path(str(item.get("sourceFile") or "")).name, str(item.get("warehouseId") or ""))
        for item in governance["activeCandidates"]
        if item.get("decision") in {"active", "applied"} or item.get("status") in {"active", "applied"}
    }
    now = utcnow_naive().isoformat()
    confirmed = []
    skipped = []
    for replay in governance["replays"]:
        source_file = Path(str(replay.get("sourceFile") or "")).name
        warehouse_id = str(replay.get("warehouseId") or "")
        scope = (source_file, warehouse_id)
        if not source_file:
            continue
        if requested and scope not in requested:
            continue
        if scope in existing_scopes:
            skipped.append({"sourceFile": source_file, "warehouseId": warehouse_id, "reason": "already_confirmed"})
            continue
        if replay.get("decision") != "ready_for_user_confirmation":
            skipped.append({"sourceFile": source_file, "warehouseId": warehouse_id, "reason": "replay_not_ready"})
            continue
        candidate_id = f"reocr_{run_id}_{_safe_record_id(source_file)}_{_safe_record_id(warehouse_id or 'all')}"
        active = {
            "candidateId": candidate_id,
            "decision": "active",
            "status": "active",
            "requiresConfirmation": False,
            "sourceFile": source_file,
            "warehouseId": warehouse_id,
            "confirmedBy": actor,
            "confirmationReason": reason,
            "confirmedAt": now,
            "replay": replay,
            "auditTrail": [
                {
                    "action": "batch_confirmed",
                    "actor": actor,
                    "reason": reason,
                    "replayedAt": replay.get("replayedAt", ""),
                }
            ],
        }
        _upsert_reocr_record(governance["activeCandidates"], active)
        existing_scopes.add(scope)
        confirmed.append(active)

    files = dict(metadata.get("files", {}))
    response: dict = {
        "decision": "batch_confirmed" if confirmed else "no_ready_reocr_candidates",
        "requiresConfirmation": False,
        "summary": {
            "confirmedCount": len(confirmed),
            "skippedCount": len(skipped),
            "requestedCount": len(requested),
        },
        "activeCandidates": confirmed,
        "skipped": skipped,
    }
    if confirmed and payload.get("generateReport", payload.get("generate_report", True)):
        preview = _build_reocr_batch_apply_preview(metadata, confirmed)
        report_path = get_labor_run_dir(run_id) / safe_labor_storage_filename("labor_reocr_batch_preview_report.xlsx", "image_batch_preview")
        build_labor_projection_report(report_path, preview)
        files["reocrPreviewReport"] = attach_labor_file(run_id, report_path, "图片识别批量结果预览报告")
        response["preview"] = preview
        response["reportFile"] = files["reocrPreviewReport"]
    updated = update_labor_metadata(
        run_id,
        {"reocrReplayGovernance": governance, "files": files},
        actor_user_id=actor,
        action="reocr_candidates_batch_confirmed",
        audit_details={"confirmedCount": len(confirmed)},
    )
    response["readinessGate"] = _build_labor_readiness_gate(updated)
    return response


@app.post("/api/labor/runs/{run_id}/reocr-candidates/{candidate_id}/apply")
def apply_labor_reocr_candidate(run_id: str, candidate_id: str, payload: dict = Body(default={}), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    active = _find_reocr_record(governance["activeCandidates"], candidate_id)
    if not active:
        raise HTTPException(status_code=404, detail="未找到已确认的图片识别结果。")
    replay = active.get("replay") if isinstance(active.get("replay"), dict) else {}
    if replay.get("decision") != "ready_for_user_confirmation":
        raise HTTPException(status_code=400, detail="图片识别结果未通过影响预览，不能采纳。")
    comparison_rows = replay.get("comparisonRows") if isinstance(replay.get("comparisonRows"), list) else replay.get("previewRows")
    if not isinstance(comparison_rows, list) or not comparison_rows:
        raise HTTPException(status_code=400, detail="图片识别结果缺少完整核对明细，不能采纳。")
    comparison_summary = replay.get("comparison") if isinstance(replay.get("comparison"), dict) else {}
    if not comparison_summary:
        raise HTTPException(status_code=400, detail="图片识别结果缺少核对摘要，不能采纳。")

    actor = _labor_action_actor(request, payload, "appliedBy", "applied_by")
    reason = str(payload.get("reason") or "").strip()
    now = utcnow_naive().isoformat()
    previous_snapshot = {
        "status": metadata.get("status", ""),
        "comparisonSummary": metadata.get("comparisonSummary", {}),
        "comparisonRows": metadata.get("comparisonRows", []),
        "candidateMatches": metadata.get("candidateMatches", []),
        "presentation": metadata.get("presentation", {}),
        "batchGuard": metadata.get("batchGuard", {}),
        "reconciliationDiagnostics": metadata.get("reconciliationDiagnostics", {}),
        "machineCheckStatus": metadata.get("machineCheckStatus", "needs_review"),
        "businessReviewStatus": metadata.get("businessReviewStatus", "pending"),
        "manualReviewRequired": metadata.get("manualReviewRequired", True),
        "directPaymentAllowed": metadata.get("directPaymentAllowed", False),
        "requiresHumanReview": metadata.get("requiresHumanReview", True),
        "reocrAdoption": metadata.get("reocrAdoption", {}),
        "files": metadata.get("files", {}),
        "diffDownloadUrl": metadata.get("diffDownloadUrl", ""),
    }
    adopted_summary = _labor_reocr_pending_revalidation_summary(
        {
            **comparison_summary,
            "notInInvoiceCount": sum(1 for row in comparison_rows if row.get("matchStatus") == "Excel有PDF无"),
            "reocrCandidateApplied": True,
        }
    )
    adopted_presentation = _build_validated_labor_presentation(
        adopted_summary,
        comparison_rows,
        replay.get("candidateMatches", []),
        excel_record_count=len(metadata.get("excelRows") or []),
    )
    preflight = _build_reocr_apply_preflight(metadata, [active], adopted_summary, comparison_rows)
    applied = {
        **active,
        "decision": "applied",
        "status": "applied",
        "appliedBy": actor,
        "applicationReason": reason,
        "appliedAt": now,
        "preflight": preflight,
        "previousSnapshot": previous_snapshot,
        "auditTrail": [
            *(active.get("auditTrail") or []),
            {"action": "applied", "actor": actor, "reason": reason, "appliedAt": now},
        ],
    }
    _upsert_reocr_record(governance["activeCandidates"], applied)
    adoption = {
        "candidateId": candidate_id,
        "sourceFile": active.get("sourceFile", ""),
        "warehouseId": active.get("warehouseId", ""),
        "status": "applied",
        "decision": "applied",
        "appliedBy": actor,
        "applicationReason": reason,
        "appliedAt": now,
        "summary": {
            "rowCount": len(comparison_rows),
            "exceptionCount": adopted_summary.get("exceptionCount", 0),
            "candidateAmountTotal": replay.get("summary", {}).get("candidateAmountTotal", 0),
            "expectedExcelAmount": replay.get("summary", {}).get("expectedExcelAmount", 0),
        },
        "preflight": preflight,
    }
    files = dict(metadata.get("files", {}))
    report_warning = ""
    if payload.get("generateReport", payload.get("generate_report", True)):
        try:
            report_file = _generate_reocr_adopted_diff_report(
                run_id,
                metadata,
                comparison_summary=adopted_summary,
                comparison_rows=comparison_rows,
                candidate_matches=replay.get("candidateMatches", []),
                presentation=adopted_presentation,
            )
            files["diffReport"] = report_file
            adoption["reportFile"] = report_file
            adoption["diffDownloadUrl"] = report_file.get("downloadUrl", "")
        except Exception as exc:  # noqa: BLE001 - 采纳结果应保留，报告失败单独提示。
            logger.warning(f"[{run_id}] 图片识别结果采纳后生成正式差异报告失败: {exc}", exc_info=True)
            report_warning = str(exc)
            adoption["reportWarning"] = report_warning
    updated = update_labor_metadata(
        run_id,
        {
            "status": "部分核对完成",
            "comparisonSummary": adopted_summary,
            "comparisonRows": comparison_rows,
            "candidateMatches": replay.get("candidateMatches", []),
            "presentation": adopted_presentation,
            "reocrReplayGovernance": governance,
            "reocrAdoption": adoption,
            "files": files,
            "diffDownloadUrl": adoption.get("diffDownloadUrl", metadata.get("diffDownloadUrl", "")),
            "batchGuard": _labor_reocr_revalidation_batch_guard(),
            "reconciliationDiagnostics": _labor_reocr_revalidation_diagnostics(),
            **_labor_uat_review_state(adopted_summary),
        },
        actor_user_id=actor,
        action="reocr_candidate_applied",
        audit_details={"candidateId": candidate_id},
    )
    readiness = _build_labor_readiness_gate(updated)
    return {
        **adoption,
        "comparisonSummary": updated.get("comparisonSummary", {}),
        "reportFile": adoption.get("reportFile"),
        "reportWarning": report_warning,
        "readinessGate": readiness,
        "activeCandidate": _find_reocr_record(updated["reocrReplayGovernance"]["activeCandidates"], candidate_id),
    }


@app.post("/api/labor/runs/{run_id}/reocr-candidates/batch-preview")
def preview_labor_reocr_batch_apply(run_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    candidate_ids = payload.get("candidateIds") or payload.get("candidate_ids") or []
    candidates = _select_reocr_apply_candidates(governance, candidate_ids)
    return _build_reocr_batch_apply_preview(metadata, candidates)


@app.post("/api/labor/runs/{run_id}/reocr-candidates/batch-apply")
def apply_labor_reocr_batch(run_id: str, payload: dict = Body(default={}), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    candidate_ids = payload.get("candidateIds") or payload.get("candidate_ids") or []
    candidates = _select_reocr_apply_candidates(governance, candidate_ids)
    preview = _build_reocr_batch_apply_preview(metadata, candidates)
    if preview["summary"]["candidateCount"] == 0:
        raise HTTPException(status_code=400, detail="没有可批量采纳的图片识别结果。")
    if preview["summary"]["duplicateScopeCount"] > 0:
        raise HTTPException(status_code=400, detail="候选中存在重复文件/仓库范围，不能批量采纳。")

    actor = _labor_action_actor(request, payload, "appliedBy", "applied_by")
    reason = str(payload.get("reason") or "").strip()
    now = utcnow_naive().isoformat()
    previous_snapshot = {
        "status": metadata.get("status", ""),
        "comparisonSummary": metadata.get("comparisonSummary", {}),
        "comparisonRows": metadata.get("comparisonRows", []),
        "candidateMatches": metadata.get("candidateMatches", []),
        "presentation": metadata.get("presentation", {}),
        "batchGuard": metadata.get("batchGuard", {}),
        "reconciliationDiagnostics": metadata.get("reconciliationDiagnostics", {}),
        "machineCheckStatus": metadata.get("machineCheckStatus", "needs_review"),
        "businessReviewStatus": metadata.get("businessReviewStatus", "pending"),
        "manualReviewRequired": metadata.get("manualReviewRequired", True),
        "directPaymentAllowed": metadata.get("directPaymentAllowed", False),
        "requiresHumanReview": metadata.get("requiresHumanReview", True),
        "reocrAdoption": metadata.get("reocrAdoption", {}),
        "files": metadata.get("files", {}),
        "diffDownloadUrl": metadata.get("diffDownloadUrl", ""),
    }
    applied_ids = {candidate["candidateId"] for candidate in candidates}
    updated_candidates = []
    for candidate in governance["activeCandidates"]:
        if candidate.get("candidateId") in applied_ids:
            updated_candidates.append(
                {
                    **candidate,
                    "decision": "applied",
                    "status": "applied",
                    "appliedBy": actor,
                    "applicationReason": reason,
                    "appliedAt": now,
                    "previousSnapshot": previous_snapshot,
                    "auditTrail": [
                        *(candidate.get("auditTrail") or []),
                        {"action": "batch_applied", "actor": actor, "reason": reason, "appliedAt": now},
                    ],
                }
            )
        else:
            updated_candidates.append(candidate)
    governance["activeCandidates"] = updated_candidates

    adoption = {
        "candidateIds": sorted(applied_ids),
        "status": "applied",
        "decision": "batch_applied",
        "appliedBy": actor,
        "applicationReason": reason,
        "appliedAt": now,
        "summary": preview["summary"],
        "coverage": preview["coverage"],
        "preflight": preview["preflight"],
    }
    files = dict(metadata.get("files", {}))
    report_warning = ""
    if payload.get("generateReport", payload.get("generate_report", True)):
        try:
            report_file = _generate_reocr_adopted_diff_report(
                run_id,
                metadata,
                comparison_summary=preview["comparisonSummary"],
                comparison_rows=preview["comparisonRows"],
                candidate_matches=preview["candidateMatches"],
                presentation=preview["presentation"],
            )
            files["diffReport"] = report_file
            adoption["reportFile"] = report_file
            adoption["diffDownloadUrl"] = report_file.get("downloadUrl", "")
        except Exception as exc:  # noqa: BLE001
            logger.warning(f"[{run_id}] 图片识别批量结果采纳后生成正式差异报告失败: {exc}", exc_info=True)
            report_warning = str(exc)
            adoption["reportWarning"] = report_warning
    updated = update_labor_metadata(
        run_id,
        {
            "status": "部分核对完成",
            "comparisonSummary": preview["comparisonSummary"],
            "comparisonRows": preview["comparisonRows"],
            "candidateMatches": preview["candidateMatches"],
            "presentation": preview["presentation"],
            "reocrReplayGovernance": governance,
            "reocrAdoption": adoption,
            "files": files,
            "diffDownloadUrl": adoption.get("diffDownloadUrl", metadata.get("diffDownloadUrl", "")),
            "batchGuard": _labor_reocr_revalidation_batch_guard(),
            "reconciliationDiagnostics": _labor_reocr_revalidation_diagnostics(),
            **_labor_uat_review_state(preview["comparisonSummary"]),
        },
        actor_user_id=actor,
        action="reocr_candidates_batch_applied",
        audit_details={"candidateIds": sorted(applied_ids)},
    )
    readiness = _build_labor_readiness_gate(updated)
    return {
        **adoption,
        "comparisonSummary": updated.get("comparisonSummary", {}),
        "reportFile": adoption.get("reportFile"),
        "reportWarning": report_warning,
        "readinessGate": readiness,
    }


@app.post("/api/labor/runs/{run_id}/reocr-candidates/{candidate_id}/rollback")
def rollback_labor_reocr_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    active = _find_reocr_record(governance["activeCandidates"], candidate_id)
    if not active:
        raise HTTPException(status_code=404, detail="未找到已确认的图片识别结果。")

    actor = _labor_action_actor(request, payload, "rolledBackBy", "rolled_back_by")
    reason = str(payload.get("reason") or payload.get("rollbackReason") or payload.get("rollback_reason") or "").strip()
    now = utcnow_naive().isoformat()
    rolled_back = {
        **active,
        "decision": "rolled_back",
        "status": "rolled_back",
        "rolledBackBy": actor,
        "rollbackReason": reason,
        "rolledBackAt": now,
        "auditTrail": [
            *(active.get("auditTrail") or []),
            {
                "action": "rolled_back",
                "actor": actor,
                "reason": reason,
                "rolledBackAt": now,
            },
        ],
    }
    _remove_reocr_record(governance["activeCandidates"], candidate_id)
    _upsert_reocr_record(governance["rolledBackCandidates"], rolled_back)
    updates = {"reocrReplayGovernance": governance}
    previous_snapshot = active.get("previousSnapshot") if isinstance(active.get("previousSnapshot"), dict) else {}
    if active.get("status") == "applied" and previous_snapshot:
        restored_files = previous_snapshot.get("files") if isinstance(previous_snapshot.get("files"), dict) else {}
        updates.update(
            {
                "status": previous_snapshot.get("status", "部分核对完成"),
                "comparisonSummary": previous_snapshot.get("comparisonSummary", {}),
                "comparisonRows": previous_snapshot.get("comparisonRows", []),
                "candidateMatches": previous_snapshot.get("candidateMatches", []),
                "presentation": previous_snapshot.get("presentation", {}),
                "files": restored_files,
                "diffDownloadUrl": previous_snapshot.get("diffDownloadUrl", ""),
                "batchGuard": previous_snapshot.get("batchGuard", {}),
                "reconciliationDiagnostics": previous_snapshot.get("reconciliationDiagnostics", {}),
                "machineCheckStatus": previous_snapshot.get("machineCheckStatus", "needs_review"),
                "businessReviewStatus": previous_snapshot.get("businessReviewStatus", "pending"),
                "manualReviewRequired": previous_snapshot.get("manualReviewRequired", True),
                "directPaymentAllowed": previous_snapshot.get("directPaymentAllowed", False),
                "requiresHumanReview": previous_snapshot.get("requiresHumanReview", True),
                "reocrAdoption": {
                    **(previous_snapshot.get("reocrAdoption") if isinstance(previous_snapshot.get("reocrAdoption"), dict) else {}),
                    "status": "rolled_back",
                    "rolledBackBy": actor,
                    "rollbackReason": reason,
                    "rolledBackAt": now,
                },
            }
        )
    updated = update_labor_metadata(
        run_id,
        updates,
        actor_user_id=actor,
        action="reocr_candidate_rolled_back",
        audit_details={"candidateId": candidate_id},
    )
    return _find_reocr_record(updated["reocrReplayGovernance"]["rolledBackCandidates"], candidate_id) or rolled_back


@app.post("/api/labor/runs/{run_id}/rule-candidates")
def create_labor_rule_candidate(run_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_labor_governance(metadata.get("ruleGovernance"))
    rule_id = _rule_id_from_payload(payload)
    if _find_rule_record(governance["candidates"], rule_id):
        raise HTTPException(status_code=409, detail=f"规则候选已存在：{rule_id}")
    actor = _labor_action_actor(request, payload, "proposedBy", "proposed_by")
    candidate = build_rule_change_candidate(
        rule_id=rule_id,
        title=str(payload.get("title") or "未命名规则候选").strip(),
        description=str(payload.get("description") or "").strip(),
        supplier=str(payload.get("supplier") or metadata.get("supplierName") or "").strip(),
        source=str(payload.get("source") or f"labor_run:{run_id}").strip(),
        proposed_by=actor,
        evidence=payload.get("evidence") if isinstance(payload.get("evidence"), list) else [],
        conditions=payload.get("conditions") if isinstance(payload.get("conditions"), dict) else {},
    )
    candidate["runId"] = run_id
    candidate["createdAt"] = utcnow_naive().isoformat()
    governance["candidates"].append(candidate)
    updated = update_labor_metadata(
        run_id,
        {"ruleGovernance": governance},
        actor_user_id=actor,
        action="rule_candidate_created",
        audit_details={"ruleId": rule_id},
    )
    return _normalized_labor_governance(updated.get("ruleGovernance"))


@app.post("/api/labor/runs/{run_id}/rule-candidates/{rule_id}/replay-summary")
def replay_labor_rule_candidate(run_id: str, rule_id: str, payload: dict = Body(...)) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_labor_governance(metadata.get("ruleGovernance"))
    candidate = _find_rule_record(governance["candidates"], rule_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"规则候选不存在：{rule_id}")
    replay_results = payload.get("replayResults") or payload.get("replay_results") or []
    if not isinstance(replay_results, list):
        raise HTTPException(status_code=400, detail="replayResults 必须是数组。")
    replay_summary = summarize_rule_replay(candidate, replay_results)
    replay_summary["runId"] = run_id
    replay_summary["summarizedAt"] = utcnow_naive().isoformat()
    replay_summary["preflight"] = _build_rule_replay_preflight(replay_summary)
    governance["replaySummaries"][rule_id] = replay_summary
    updated = update_labor_metadata(run_id, {"ruleGovernance": governance})
    return updated["ruleGovernance"]["replaySummaries"][rule_id]


@app.post("/api/labor/runs/{run_id}/rule-candidates/{rule_id}/auto-replay")
def auto_replay_labor_rule_candidate(run_id: str, rule_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_labor_governance(metadata.get("ruleGovernance"))
    candidate = _find_rule_record(governance["candidates"], rule_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"规则候选不存在：{rule_id}")
    try:
        limit = int(payload.get("limit") or 20)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="limit 必须是数字。") from exc
    historical_runs = list_labor_metadata()
    replay_summary = summarize_rule_auto_replay(
        candidate,
        historical_runs,
        current_run_id=run_id,
        limit=max(1, min(limit, 100)),
    )
    replay_summary["runId"] = run_id
    replay_summary["summarizedAt"] = utcnow_naive().isoformat()
    replay_summary["preflight"] = _build_rule_replay_preflight(replay_summary)
    governance["replaySummaries"][rule_id] = replay_summary
    updated = update_labor_metadata(run_id, {"ruleGovernance": governance})
    return updated["ruleGovernance"]["replaySummaries"][rule_id]


@app.post("/api/labor/runs/{run_id}/rule-candidates/{rule_id}/confirm")
def confirm_labor_rule_candidate(run_id: str, rule_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_labor_governance(metadata.get("ruleGovernance"))
    candidate = _find_rule_record(governance["candidates"], rule_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"规则候选不存在：{rule_id}")
    replay_summary = governance["replaySummaries"].get(rule_id)
    if not replay_summary:
        raise HTTPException(status_code=400, detail="规则候选缺少历史回放摘要，不能确认生效。")
    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    try:
        active = confirm_rule_candidate(
            candidate,
            replay_summary,
            confirmed_by=actor,
            reason=str(payload.get("reason") or "").strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    active["runId"] = run_id
    active["confirmedAt"] = utcnow_naive().isoformat()
    active["preflight"] = replay_summary.get("preflight") or _build_rule_replay_preflight(replay_summary)
    _upsert_rule_record(governance["activeRules"], active)
    _upsert_rule_record(governance["candidates"], {**candidate, "status": "confirmed", "decision": "confirmed"})
    updated = update_labor_metadata(
        run_id,
        {"ruleGovernance": governance},
        actor_user_id=actor,
        action="rule_candidate_confirmed",
        audit_details={"ruleId": rule_id},
    )
    return _find_rule_record(updated["ruleGovernance"]["activeRules"], rule_id)


@app.post("/api/labor/runs/{run_id}/rule-candidates/{rule_id}/rollback")
def rollback_labor_rule_candidate(run_id: str, rule_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_labor_governance(metadata.get("ruleGovernance"))
    active = _find_rule_record(governance["activeRules"], rule_id)
    if not active:
        raise HTTPException(status_code=404, detail=f"已确认规则不存在：{rule_id}")
    actor = _labor_action_actor(request, payload, "rolledBackBy", "rolled_back_by")
    try:
        target_version = payload.get("targetVersion", payload.get("target_version"))
        rolled_back = rollback_rule_version(
            active,
            rolled_back_by=actor,
            reason=str(payload.get("reason") or "").strip(),
            target_version=int(target_version) if target_version is not None else None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    rolled_back["runId"] = run_id
    rolled_back["rolledBackAt"] = utcnow_naive().isoformat()
    _upsert_rule_record(governance["rolledBackRules"], rolled_back)
    _remove_rule_record(governance["activeRules"], rule_id)
    updated = update_labor_metadata(
        run_id,
        {"ruleGovernance": governance},
        actor_user_id=actor,
        action="rule_candidate_rolled_back",
        audit_details={"ruleId": rule_id},
    )
    return _find_rule_record(updated["ruleGovernance"]["rolledBackRules"], rule_id)


@app.post("/api/labor/runs/{run_id}/name-mapping-candidates/{candidate_id}/confirm")
def confirm_labor_name_mapping_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    recalculate_requested = bool(
        payload.get("recalculate") or payload.get("generateReport") or payload.get("generate_report")
    )
    if recalculate_requested and labor_run_is_active(metadata):
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前批次已有核对任务正在运行，不能同时重算姓名映射结果。",
                error_code="LABOR_RUN_ALREADY_ACTIVE",
                retryable=True,
                next_action="请等待当前任务完成或失败后，再确认并重新生成报告。",
            ),
        )
    governance = _normalized_name_mapping_governance(metadata.get("nameMappingGovernance"))
    candidate = _find_name_mapping_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"姓名映射候选不存在：{candidate_id}")
    proposed = candidate.get("proposedMapping") if isinstance(candidate.get("proposedMapping"), dict) else {}
    if not proposed:
        raise HTTPException(status_code=400, detail="姓名映射候选缺少 proposedMapping。")
    replay_summary = governance["replaySummaries"].get(candidate_id)
    if not replay_summary:
        raise HTTPException(status_code=400, detail="姓名映射候选缺少影响回放摘要，不能确认生效。")
    if replay_summary.get("decision") != "ready_for_user_confirmation":
        raise HTTPException(status_code=400, detail="姓名映射候选未通过影响回放，不能确认生效。")

    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    reason = str(payload.get("reason") or "").strip()
    active = {
        **candidate,
        "decision": "active",
        "status": "active",
        "requiresConfirmation": False,
        "confirmedBy": actor,
        "confirmationReason": reason,
        "confirmedAt": utcnow_naive().isoformat(),
        "auditTrail": [
            *(candidate.get("auditTrail") or []),
            {"action": "confirmed", "actor": actor, "reason": reason, "replaySummary": replay_summary.get("summary", {})},
        ],
        "replaySummary": replay_summary.get("summary", {}),
        "preflight": replay_summary.get("preflight") or _build_profile_replay_preflight(candidate, replay_summary),
    }
    manual_mapping = dict(metadata.get("manualNameMapping") or {})
    manual_mapping.update({str(key): str(value) for key, value in proposed.items()})
    _upsert_name_mapping_record(governance["activeMappings"], active)
    _upsert_name_mapping_record(governance["candidates"], {**candidate, "status": "confirmed", "decision": "confirmed"})
    updated = update_labor_metadata(
        run_id,
        {
            **_labor_invalidate_official_result(
                metadata,
                status="已确认姓名映射",
                reason_code="name_mapping_changed",
                message="姓名映射已更新，旧核对结果已失效，必须重新执行整批核对。",
            ),
            "manualNameMapping": manual_mapping,
            "nameMappingGovernance": governance,
        },
        actor_user_id=actor,
        action="name_mapping_candidate_confirmed",
        audit_details={"candidateId": candidate_id},
    )
    recalculated_run = None
    if recalculate_requested:
        try:
            submitted = asyncio.run(extract_and_compare_labor_run(run_id))
            recalculated_run = _labor_metadata_or_404(run_id)
            if str(recalculated_run.get("status") or "") == "已确认姓名映射":
                recalculated_run = submitted
            updated = recalculated_run
        except Exception as exc:  # noqa: BLE001 - mapping confirmation should remain auditable even if report refresh fails.
            logger.warning(f"[{run_id}] 姓名映射确认后重算失败: {exc}")
    response = _find_name_mapping_record(updated["nameMappingGovernance"]["activeMappings"], candidate_id) or active
    response["manualNameMapping"] = updated.get("manualNameMapping", {})
    if recalculated_run:
        response["recalculatedRun"] = {
            "status": recalculated_run.get("status", ""),
            "comparisonSummary": recalculated_run.get("comparisonSummary", {}),
            "diffDownloadUrl": recalculated_run.get("diffDownloadUrl", ""),
            "diffReport": (recalculated_run.get("files") or {}).get("diffReport", {}),
        }
    return response


@app.post("/api/labor/runs/{run_id}/allocation-candidates/{candidate_id}/confirm")
def confirm_labor_allocation_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_allocation_governance(metadata.get("allocationGovernance"))
    candidate = _find_allocation_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"跨仓库归属候选不存在：{candidate_id}")

    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    reason = str(payload.get("reason") or "").strip()
    decision_note = str(payload.get("decisionNote") or payload.get("decision_note") or "").strip()
    now = utcnow_naive().isoformat()
    active = {
        **candidate,
        "decision": "confirmed",
        "status": "confirmed",
        "requiresConfirmation": False,
        "confirmedBy": actor,
        "confirmationReason": reason,
        "decisionNote": decision_note,
        "confirmedAt": now,
        "auditTrail": [
            *(candidate.get("auditTrail") or []),
            {
                "action": "confirmed",
                "actor": actor,
                "reason": reason,
                "decisionNote": decision_note,
                "confirmedAt": now,
            },
        ],
    }
    _upsert_allocation_record(governance["activeAllocations"], active)
    _upsert_allocation_record(governance["candidates"], {**candidate, "status": "confirmed", "decision": "confirmed"})
    updated = update_labor_metadata(
        run_id,
        {"allocationGovernance": governance},
        actor_user_id=actor,
        action="allocation_candidate_confirmed",
        audit_details={"candidateId": candidate_id},
    )
    readiness = _build_labor_readiness_gate(updated)
    response = _find_allocation_record(updated["allocationGovernance"]["activeAllocations"], candidate_id) or active
    response["readinessGate"] = readiness
    return response


@app.post("/api/labor/runs/{run_id}/allocation-candidates/{candidate_id}/rollback")
def rollback_labor_allocation_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_allocation_governance(metadata.get("allocationGovernance"))
    active = _find_allocation_record(governance["activeAllocations"], candidate_id)
    if not active:
        raise HTTPException(status_code=404, detail=f"已确认跨仓库归属候选不存在：{candidate_id}")

    actor = _labor_action_actor(request, payload, "rolledBackBy", "rolled_back_by")
    reason = str(payload.get("reason") or payload.get("rollbackReason") or payload.get("rollback_reason") or "").strip()
    now = utcnow_naive().isoformat()
    rolled_back = {
        **active,
        "decision": "rolled_back",
        "status": "rolled_back",
        "rolledBackBy": actor,
        "rollbackReason": reason,
        "rolledBackAt": now,
        "auditTrail": [
            *(active.get("auditTrail") or []),
            {"action": "rolled_back", "actor": actor, "reason": reason, "rolledBackAt": now},
        ],
    }
    _upsert_allocation_record(governance["rolledBackAllocations"], rolled_back)
    _remove_allocation_record(governance["activeAllocations"], candidate_id)
    _upsert_allocation_record(governance["candidates"], {**active, "status": "rolled_back", "decision": "rolled_back"})
    updated = update_labor_metadata(
        run_id,
        {"allocationGovernance": governance},
        actor_user_id=actor,
        action="allocation_candidate_rolled_back",
        audit_details={"candidateId": candidate_id},
    )
    response = _find_allocation_record(updated["allocationGovernance"]["rolledBackAllocations"], candidate_id) or rolled_back
    response["readinessGate"] = _build_labor_readiness_gate(updated)
    return response


@app.post("/api/labor/runs/{run_id}/name-mapping-candidates/{candidate_id}/auto-replay")
def auto_replay_labor_name_mapping_candidate(run_id: str, candidate_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_name_mapping_governance(metadata.get("nameMappingGovernance"))
    candidate = _find_name_mapping_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"姓名映射候选不存在：{candidate_id}")
    try:
        limit = int(payload.get("limit") or 20)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="limit 必须是数字。") from exc
    try:
        replay_summary = _summarize_name_mapping_auto_replay(
            candidate,
            metadata,
            amount_tolerance=float(payload.get("amountTolerance") or payload.get("amount_tolerance") or AI_CONFIG["amount_tolerance"]),
            hours_tolerance=float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"]),
            historical_runs=list_labor_metadata(),
            current_run_id=run_id,
            limit=max(0, min(limit, 100)),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    replay_summary["runId"] = run_id
    replay_summary["summarizedAt"] = utcnow_naive().isoformat()
    governance["replaySummaries"][candidate_id] = replay_summary
    updated = update_labor_metadata(run_id, {"nameMappingGovernance": governance})
    return updated["nameMappingGovernance"]["replaySummaries"][candidate_id]


@app.post("/api/labor/runs/{run_id}/name-mapping-candidates/{candidate_id}/rollback")
def rollback_labor_name_mapping_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_name_mapping_governance(metadata.get("nameMappingGovernance"))
    active = _find_name_mapping_record(governance["activeMappings"], candidate_id)
    if not active:
        raise HTTPException(status_code=404, detail=f"已确认姓名映射不存在：{candidate_id}")

    actor = _labor_action_actor(request, payload, "rolledBackBy", "rolled_back_by")
    reason = str(payload.get("reason") or "").strip()
    proposed = active.get("proposedMapping") if isinstance(active.get("proposedMapping"), dict) else {}
    manual_mapping = dict(metadata.get("manualNameMapping") or {})
    for key, value in proposed.items():
        if str(manual_mapping.get(str(key)) or "") == str(value):
            manual_mapping.pop(str(key), None)
    rolled_back = {
        **active,
        "decision": "rolled_back",
        "status": "rolled_back",
        "rolledBackBy": actor,
        "rollbackReason": reason,
        "rolledBackAt": utcnow_naive().isoformat(),
        "auditTrail": [
            *(active.get("auditTrail") or []),
            {"action": "rolled_back", "actor": actor, "reason": reason},
        ],
    }
    _upsert_name_mapping_record(governance["rolledBackMappings"], rolled_back)
    _remove_name_mapping_record(governance["activeMappings"], candidate_id)
    _upsert_name_mapping_record(governance["candidates"], {**active, "status": "rolled_back", "decision": "rolled_back"})
    updated = update_labor_metadata(
        run_id,
        {"manualNameMapping": manual_mapping, "nameMappingGovernance": governance},
        actor_user_id=actor,
        action="name_mapping_candidate_rolled_back",
        audit_details={"candidateId": candidate_id},
    )
    response = _find_name_mapping_record(updated["nameMappingGovernance"]["rolledBackMappings"], candidate_id) or rolled_back
    response["manualNameMapping"] = updated.get("manualNameMapping", {})
    return response


@app.post("/api/labor/runs/{run_id}/profile-candidates/{candidate_id}/confirm")
def confirm_labor_profile_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_profile_governance(metadata.get("profileGovernance"))
    candidate = _find_profile_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"Profile 候选不存在：{candidate_id}")
    replay_summary = governance["replaySummaries"].get(candidate_id)
    if not replay_summary:
        raise HTTPException(status_code=400, detail="Profile 候选缺少历史回放摘要，不能确认生效。")
    if replay_summary.get("decision") != "ready_for_user_confirmation":
        raise HTTPException(status_code=400, detail="供应商格式建议未通过历史影响预览，不能确认生效。")
    runtime_profile = _supplier_profile_from_governance_data(
        candidate.get("profileData") if isinstance(candidate.get("profileData"), dict) else {},
        fallback_key=str(candidate.get("profileKey") or "candidate-profile"),
        fallback_aliases=[str(candidate.get("supplier") or "")],
    )
    if runtime_profile is None or not is_runtime_approved_profile(runtime_profile):
        raise HTTPException(
            status_code=400,
            detail="draft Supplier Profile 只能保留为候选；完成 Golden 复核并补齐 approved 审批元数据后才能激活。",
        )
    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    active = {
        **candidate,
        "decision": "active",
        "status": "active",
        "requiresConfirmation": False,
        "version": int((candidate.get("profileData") or {}).get("version") or candidate.get("version") or 1),
        "confirmedBy": actor,
        "confirmationReason": str(payload.get("reason") or "").strip(),
        "confirmedAt": utcnow_naive().isoformat(),
        "auditTrail": [
            *(candidate.get("auditTrail") or []),
            {
                "action": "confirmed",
                "actor": actor,
                "reason": str(payload.get("reason") or "").strip(),
                "replaySummary": replay_summary.get("summary", {}),
            },
        ],
        "replaySummary": replay_summary.get("summary", {}),
        "preflight": replay_summary.get("preflight") or _build_profile_replay_preflight(candidate, replay_summary),
    }
    _upsert_profile_record(governance["activeProfiles"], active)
    _upsert_profile_record(governance["candidates"], {**candidate, "status": "confirmed", "decision": "confirmed"})
    updated = update_labor_metadata(
        run_id,
        {"profileGovernance": governance},
        actor_user_id=actor,
        action="profile_candidate_confirmed",
        audit_details={"candidateId": candidate_id},
    )
    return _find_profile_record(updated["profileGovernance"]["activeProfiles"], candidate_id)


@app.post("/api/labor/runs/{run_id}/profile-candidates/{candidate_id}/auto-replay")
def auto_replay_labor_profile_candidate(run_id: str, candidate_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_profile_governance(metadata.get("profileGovernance"))
    candidate = _find_profile_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"Profile 候选不存在：{candidate_id}")
    try:
        limit = int(payload.get("limit") or 20)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="limit 必须是数字。") from exc
    replay_summary = _summarize_profile_auto_replay(
        candidate,
        list_labor_metadata(),
        current_run_id=run_id,
        limit=max(1, min(limit, 100)),
    )
    replay_summary["runId"] = run_id
    replay_summary["summarizedAt"] = utcnow_naive().isoformat()
    replay_summary["preflight"] = _build_profile_replay_preflight(candidate, replay_summary)
    governance["replaySummaries"][candidate_id] = replay_summary
    updated = update_labor_metadata(run_id, {"profileGovernance": governance})
    return updated["profileGovernance"]["replaySummaries"][candidate_id]


@app.post("/api/labor/runs/{run_id}/profile-candidates/{candidate_id}/rollback")
def rollback_labor_profile_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_profile_governance(metadata.get("profileGovernance"))
    active = _find_profile_record(governance["activeProfiles"], candidate_id)
    if not active:
        raise HTTPException(status_code=404, detail=f"已确认 Profile 不存在：{candidate_id}")
    current_version = int(active.get("version") or 1)
    target_version = payload.get("targetVersion", payload.get("target_version"))
    rollback_target = int(target_version) if target_version is not None else max(current_version - 1, 0)
    actor = _labor_action_actor(request, payload, "rolledBackBy", "rolled_back_by")
    rolled_back = {
        **active,
        "decision": "rolled_back",
        "status": "rolled_back",
        "rolledBackBy": actor,
        "rollbackReason": str(payload.get("reason") or "").strip(),
        "rollbackToVersion": rollback_target,
        "rolledBackAt": utcnow_naive().isoformat(),
        "auditTrail": [
            *(active.get("auditTrail") or []),
            {
                "action": "rolled_back",
                "actor": actor,
                "reason": str(payload.get("reason") or "").strip(),
                "fromVersion": current_version,
                "toVersion": rollback_target,
            },
        ],
    }
    _upsert_profile_record(governance["rolledBackProfiles"], rolled_back)
    _remove_profile_record(governance["activeProfiles"], candidate_id)
    updated = update_labor_metadata(
        run_id,
        {"profileGovernance": governance},
        actor_user_id=actor,
        action="profile_candidate_rolled_back",
        audit_details={"candidateId": candidate_id},
    )
    return _find_profile_record(updated["profileGovernance"]["rolledBackProfiles"], candidate_id)


@app.post("/api/labor/runs/{run_id}/correction-candidates/{candidate_id}/confirm")
def confirm_labor_correction_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_correction_governance(metadata.get("correctionGovernance"))
    candidate = _find_correction_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"修正候选不存在：{candidate_id}")
    replay_summary = governance["replaySummaries"].get(candidate_id)
    if not replay_summary:
        raise HTTPException(status_code=400, detail="修正候选缺少影响回放摘要，不能确认生效。")
    if replay_summary.get("decision") != "ready_for_user_confirmation":
        raise HTTPException(status_code=400, detail="修正候选未通过影响回放，不能确认生效。")
    actor = _labor_action_actor(request, payload, "confirmedBy", "confirmed_by")
    active = {
        **candidate,
        "decision": "active",
        "status": "active",
        "requiresConfirmation": False,
        "confirmedBy": actor,
        "confirmationReason": str(payload.get("reason") or "").strip(),
        "confirmedAt": utcnow_naive().isoformat(),
        "auditTrail": [
            *(candidate.get("auditTrail") or []),
            {
                "action": "confirmed",
                "actor": actor,
                "reason": str(payload.get("reason") or "").strip(),
                "replaySummary": replay_summary.get("summary", {}),
            },
        ],
        "replaySummary": replay_summary.get("summary", {}),
    }
    _upsert_correction_record(governance["activeCorrections"], active)
    _upsert_correction_record(governance["candidates"], {**candidate, "status": "confirmed", "decision": "confirmed"})
    updated = update_labor_metadata(
        run_id,
        {"correctionGovernance": governance},
        actor_user_id=actor,
        action="correction_candidate_confirmed",
        audit_details={"candidateId": candidate_id},
    )
    return _find_correction_record(updated["correctionGovernance"]["activeCorrections"], candidate_id)


@app.post("/api/labor/runs/{run_id}/correction-candidates/{candidate_id}/auto-replay")
def auto_replay_labor_correction_candidate(run_id: str, candidate_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_correction_governance(metadata.get("correctionGovernance"))
    candidate = _find_correction_record(governance["candidates"], candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail=f"修正候选不存在：{candidate_id}")
    replay_summary = _summarize_correction_auto_replay(
        candidate,
        metadata,
        amount_tolerance=float(payload.get("amountTolerance") or payload.get("amount_tolerance") or AI_CONFIG["amount_tolerance"]),
        hours_tolerance=float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"]),
    )
    replay_summary["runId"] = run_id
    replay_summary["summarizedAt"] = utcnow_naive().isoformat()
    governance["replaySummaries"][candidate_id] = replay_summary
    updated = update_labor_metadata(run_id, {"correctionGovernance": governance})
    return updated["correctionGovernance"]["replaySummaries"][candidate_id]


@app.post("/api/labor/runs/{run_id}/correction-candidates/{candidate_id}/rollback")
def rollback_labor_correction_candidate(run_id: str, candidate_id: str, payload: dict = Body(...), request: Request = None) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    governance = _normalized_correction_governance(metadata.get("correctionGovernance"))
    active = _find_correction_record(governance["activeCorrections"], candidate_id)
    if not active:
        raise HTTPException(status_code=404, detail=f"已确认修正不存在：{candidate_id}")
    actor = _labor_action_actor(request, payload, "rolledBackBy", "rolled_back_by")
    rolled_back = {
        **active,
        "decision": "rolled_back",
        "status": "rolled_back",
        "rolledBackBy": actor,
        "rollbackReason": str(payload.get("reason") or "").strip(),
        "rolledBackAt": utcnow_naive().isoformat(),
        "auditTrail": [
            *(active.get("auditTrail") or []),
            {
                "action": "rolled_back",
                "actor": actor,
                "reason": str(payload.get("reason") or "").strip(),
            },
        ],
    }
    _upsert_correction_record(governance["rolledBackCorrections"], rolled_back)
    _remove_correction_record(governance["activeCorrections"], candidate_id)
    updated = update_labor_metadata(
        run_id,
        {"correctionGovernance": governance},
        actor_user_id=actor,
        action="correction_candidate_rolled_back",
        audit_details={"candidateId": candidate_id},
    )
    return _find_correction_record(updated["correctionGovernance"]["rolledBackCorrections"], candidate_id)


@app.post("/api/labor/runs/{run_id}/corrections/projected-preview")
def preview_labor_correction_projection(run_id: str, payload: dict = Body(default={})) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    candidate_ids = payload.get("candidateIds") or payload.get("candidate_ids")
    if candidate_ids is not None and not isinstance(candidate_ids, list):
        raise HTTPException(status_code=400, detail="candidateIds 必须是数组。")
    try:
        amount_tolerance = float(payload.get("amountTolerance") or payload.get("amount_tolerance") or AI_CONFIG["amount_tolerance"])
        hours_tolerance = float(payload.get("hoursTolerance") or payload.get("hours_tolerance") or AI_CONFIG["hours_tolerance"])
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="容差必须是数字。") from exc
    governance = _normalized_correction_governance(metadata.get("correctionGovernance"))
    selected_ids = {str(item) for item in candidate_ids} if candidate_ids is not None else None
    active_corrections = [
        correction
        for correction in governance["activeCorrections"]
        if selected_ids is None or str(correction.get("candidateId") or "") in selected_ids
    ]
    if not active_corrections:
        raise HTTPException(status_code=400, detail="没有可用于预览的已确认修正。")
    try:
        preview = _build_correction_projected_preview(
            metadata,
            active_corrections,
            amount_tolerance=amount_tolerance,
            hours_tolerance=hours_tolerance,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    if payload.get("generateReport") or payload.get("generate_report"):
        run_dir = get_labor_run_dir(run_id)
        report_path = run_dir / safe_labor_filename("海外劳务工修正预览报告.xlsx", "修正预览")
        build_labor_projection_report(report_path, preview)
        files = dict(metadata.get("files", {}))
        files["correctionPreviewReport"] = attach_labor_file(run_id, report_path, "修正预览报告")
        preview["preflight"] = _build_correction_preview_preflight(preview)
        updated_governance = _attach_correction_preflight(governance, active_corrections, preview["preflight"])
        updated = update_labor_metadata(run_id, {"files": files, "correctionGovernance": updated_governance})
        preview["reportFile"] = updated["files"]["correctionPreviewReport"]
    else:
        preview["preflight"] = _build_correction_preview_preflight(preview)
    return preview


@app.post("/api/labor/runs/{run_id}/governance-report")
def generate_labor_governance_report(run_id: str) -> dict:
    metadata = _labor_metadata_or_404(run_id)
    run_dir = get_labor_run_dir(run_id)
    report_path = run_dir / safe_labor_filename("海外劳务工治理审计报告.xlsx", "治理审计")
    build_labor_governance_report(report_path, metadata)
    files = dict(metadata.get("files", {}))
    files["governanceAuditReport"] = attach_labor_file(run_id, report_path, "治理审计报告")
    updated = update_labor_metadata(run_id, {"files": files})
    return updated["files"]["governanceAuditReport"]


def _normalized_labor_governance(value) -> dict:
    governance = value if isinstance(value, dict) else {}
    replay_summaries = governance.get("replaySummaries")
    return {
        "candidates": list(governance.get("candidates") or []),
        "replaySummaries": replay_summaries if isinstance(replay_summaries, dict) else {},
        "activeRules": list(governance.get("activeRules") or []),
        "rolledBackRules": list(governance.get("rolledBackRules") or []),
    }


def _rule_id_from_payload(payload: dict) -> str:
    rule_id = str(payload.get("ruleId") or payload.get("rule_id") or "").strip()
    if rule_id:
        return rule_id
    title = str(payload.get("title") or "rule").strip().lower()
    safe_title = re.sub(r"[^a-z0-9]+", "_", title).strip("_") or "rule"
    return f"{safe_title}_{utcnow_naive().strftime('%Y%m%d%H%M%S')}"


def _find_rule_record(records: list[dict], rule_id: str) -> dict | None:
    for record in records:
        if str(record.get("ruleId") or "") == rule_id:
            return record
    return None


def _upsert_rule_record(records: list[dict], record: dict) -> None:
    rule_id = str(record.get("ruleId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("ruleId") or "") == rule_id:
            records[index] = record
            return
    records.append(record)


def _remove_rule_record(records: list[dict], rule_id: str) -> None:
    records[:] = [record for record in records if str(record.get("ruleId") or "") != rule_id]


def _normalized_profile_governance(value) -> dict:
    governance = value if isinstance(value, dict) else {}
    replay_summaries = governance.get("replaySummaries")
    return {
        "candidates": list(governance.get("candidates") or []),
        "replaySummaries": replay_summaries if isinstance(replay_summaries, dict) else {},
        "activeProfiles": list(governance.get("activeProfiles") or []),
        "rolledBackProfiles": list(governance.get("rolledBackProfiles") or []),
    }


def _build_profile_candidate(run_id: str, supplier: str, profile_data: dict, pdf_rows: list) -> dict:
    profile_key = str(profile_data.get("key") or "unknown")
    evidence = []
    for row in pdf_rows[:8]:
        evidence.append(
            {
                "sourceFile": getattr(row, "source_file", ""),
                "sourcePageOrRow": getattr(row, "source_page_or_row", ""),
                "employeeName": getattr(row, "employee_name_raw", ""),
                "amount": getattr(row, "amount", 0),
                "confidence": getattr(row, "confidence", 0),
                "evidenceText": str(getattr(row, "evidence_text", "") or "")[:200],
            }
        )
    return {
        "candidateId": f"profile_{profile_key}_{run_id}",
        "profileKey": profile_key,
        "supplier": supplier,
        "decision": "candidate_only",
        "status": "pending_user_confirmation",
        "requiresConfirmation": True,
        "source": f"labor_run:{run_id}",
        "createdAt": utcnow_naive().isoformat(),
        "profileData": profile_data,
        "evidence": evidence,
        "auditTrail": [
            {
                "action": "created",
                "actor": "system",
                "reason": "抽取质量通过后生成 Profile 建议，等待人工确认。",
            }
        ],
    }


def _upsert_profile_candidate(records: list[dict], candidate: dict) -> None:
    candidate_id = str(candidate.get("candidateId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("candidateId") or "") == candidate_id:
            records[index] = candidate
            return
    records.append(candidate)


def _summarize_profile_auto_replay(candidate: dict, historical_runs: list[dict], *, current_run_id: str, limit: int = 20) -> dict:
    candidate_supplier = str(candidate.get("supplier") or "").strip().lower()
    results: list[dict] = []
    for run in historical_runs:
        if len(results) >= limit:
            break
        run_id = str(run.get("id") or "")
        if not run_id:
            continue
        run_supplier = str(run.get("supplierName") or run.get("supplier") or "").strip()
        in_scope = not candidate_supplier or run_supplier.lower() == candidate_supplier
        diagnostics = run.get("reconciliationDiagnostics") if isinstance(run.get("reconciliationDiagnostics"), dict) else {}
        extraction_quality = run.get("extractionQuality") if isinstance(run.get("extractionQuality"), dict) else {}
        comparison_summary = run.get("comparisonSummary") if isinstance(run.get("comparisonSummary"), dict) else {}
        before_status = _metadata_status(diagnostics, extraction_quality, comparison_summary)
        issue_count = _metadata_profile_issue_count(diagnostics, extraction_quality, comparison_summary)
        if not in_scope:
            impact = "out_of_scope"
            after_status = before_status
            after_issue_count = issue_count
        elif before_status in {"critical", "failed", "error"}:
            impact = "regression_risk"
            after_status = before_status
            after_issue_count = issue_count
        else:
            impact = "compatible"
            after_status = before_status
            after_issue_count = issue_count
        results.append(
            {
                "runId": run_id,
                "currentRun": run_id == current_run_id,
                "supplier": run_supplier,
                "periodStart": str(run.get("periodStart") or ""),
                "periodEnd": str(run.get("periodEnd") or ""),
                "beforeStatus": before_status,
                "afterStatus": after_status,
                "beforeIssueCount": issue_count,
                "afterIssueCount": after_issue_count,
                "impact": impact,
            }
        )

    compatible = [row for row in results if row["impact"] == "compatible"]
    regressions = [row for row in results if row["impact"] == "regression_risk"]
    evidence_count = len(candidate.get("evidence") or [])
    decision = (
        "blocked_by_replay_regression"
        if regressions
        else "ready_for_user_confirmation"
        if compatible and evidence_count > 0
        else "needs_more_replay_evidence"
    )
    return {
        "candidateId": candidate.get("candidateId", ""),
        "decision": decision,
        "requiresConfirmation": True,
        "mode": "profile_metadata_compatibility_replay",
        "summary": {
            "replayedCount": len(results),
            "compatibleCount": len(compatible),
            "regressionCount": len(regressions),
            "unchangedCount": len(results) - len(compatible) - len(regressions),
            "evidenceCount": evidence_count,
        },
        "compatible": compatible,
        "regressions": regressions,
        "replayResults": results,
        "limitations": [
            "本回放基于历史批次 metadata 的质量状态和异常计数，不重新抽取 PDF，也不调用 AI。",
            "确认 Profile 后仍需通过后续真实批次观察是否误伤。",
        ],
    }


def _metadata_profile_issue_count(diagnostics: dict, extraction_quality: dict, comparison_summary: dict) -> int:
    diagnostic_issues = diagnostics.get("issues") if isinstance(diagnostics.get("issues"), list) else []
    quality_issues = extraction_quality.get("issues") if isinstance(extraction_quality.get("issues"), list) else []
    return len(diagnostic_issues) + len(quality_issues) + int(comparison_summary.get("exceptionCount") or 0)


def _metadata_status(diagnostics: dict, extraction_quality: dict, comparison_summary: dict) -> str:
    for source in (diagnostics, extraction_quality, comparison_summary):
        level = str(source.get("level") or source.get("status") or "").lower()
        if level:
            return level
    return "ok" if int(comparison_summary.get("exceptionCount") or 0) == 0 else "warning"


def _find_profile_record(records: list[dict], candidate_id: str) -> dict | None:
    for record in records:
        if str(record.get("candidateId") or "") == candidate_id:
            return record
    return None


def _upsert_profile_record(records: list[dict], record: dict) -> None:
    candidate_id = str(record.get("candidateId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("candidateId") or "") == candidate_id:
            records[index] = record
            return
    records.append(record)


def _remove_profile_record(records: list[dict], candidate_id: str) -> None:
    records[:] = [record for record in records if str(record.get("candidateId") or "") != candidate_id]


def _normalized_correction_governance(value) -> dict:
    governance = value if isinstance(value, dict) else {}
    replay_summaries = governance.get("replaySummaries")
    return {
        "candidates": list(governance.get("candidates") or []),
        "replaySummaries": replay_summaries if isinstance(replay_summaries, dict) else {},
        "activeCorrections": list(governance.get("activeCorrections") or []),
        "rolledBackCorrections": list(governance.get("rolledBackCorrections") or []),
    }


def _normalized_name_mapping_governance(value) -> dict:
    governance = value if isinstance(value, dict) else {}
    replay_summaries = governance.get("replaySummaries")
    candidates = list(governance.get("candidates") or [])
    return {
        "summary": _summarize_name_mapping_candidates(candidates),
        "candidates": candidates,
        "replaySummaries": replay_summaries if isinstance(replay_summaries, dict) else {},
        "activeMappings": list(governance.get("activeMappings") or []),
        "rolledBackMappings": list(governance.get("rolledBackMappings") or []),
    }


def _summarize_name_mapping_candidates(candidates: list[dict]) -> dict:
    amount_tolerance = float(AI_CONFIG.get("amount_tolerance", 0.1) or 0.1)
    hours_tolerance = float(AI_CONFIG.get("hours_tolerance", 0.1) or 0.1)
    pending = [
        candidate
        for candidate in candidates or []
        if str(candidate.get("status") or "").lower() not in {"confirmed", "rolled_back"}
        and str(candidate.get("decision") or "").lower() not in {"confirmed", "rolled_back"}
    ]
    return {
        "candidateCount": len(pending),
        "highConfidenceCount": sum(1 for candidate in pending if candidate.get("confidence") == "high"),
        "readyToReplayCount": sum(
            1
            for candidate in pending
            if candidate.get("confidence") == "high"
            and abs(float(candidate.get("amountGap") or 0)) <= amount_tolerance
            and abs(float(candidate.get("hoursGap") or 0)) <= hours_tolerance
        ),
        "projectedFixedExceptionCount": sum(int(candidate.get("projectedFixedExceptionCount") or 0) for candidate in pending),
        "amountStillDifferentCount": sum(1 for candidate in pending if abs(float(candidate.get("amountGap") or 0)) > amount_tolerance),
        "hoursStillDifferentCount": sum(1 for candidate in pending if abs(float(candidate.get("hoursGap") or 0)) > hours_tolerance),
    }


def _normalized_reocr_governance(value) -> dict:
    governance = value if isinstance(value, dict) else {}
    return {
        "replays": list(governance.get("replays") or []),
        "activeCandidates": list(governance.get("activeCandidates") or []),
        "rolledBackCandidates": list(governance.get("rolledBackCandidates") or []),
    }


def _normalized_allocation_governance(value) -> dict:
    governance = value if isinstance(value, dict) else {}
    return {
        "candidates": list(governance.get("candidates") or []),
        "activeAllocations": list(governance.get("activeAllocations") or []),
        "rolledBackAllocations": list(governance.get("rolledBackAllocations") or []),
    }


def _build_allocation_governance(run_id: str, warehouse_comparison: dict, existing: dict | None = None) -> dict:
    governance = _normalized_allocation_governance(existing)
    generated = _build_allocation_candidates(run_id, warehouse_comparison)
    protected_ids = {
        str(record.get("candidateId") or "")
        for record in [
            *governance["activeAllocations"],
            *governance["rolledBackAllocations"],
        ]
    }
    generated_by_id = {str(candidate.get("candidateId") or ""): candidate for candidate in generated}
    merged = []
    seen = set()
    for existing_candidate in governance["candidates"]:
        candidate_id = str(existing_candidate.get("candidateId") or "")
        if not candidate_id:
            continue
        status = str(existing_candidate.get("status") or "")
        decision = str(existing_candidate.get("decision") or "")
        if candidate_id in protected_ids or status in {"confirmed", "rolled_back"} or decision in {"confirmed", "rolled_back"}:
            merged.append(existing_candidate)
            seen.add(candidate_id)
            continue
        generated_candidate = generated_by_id.get(candidate_id)
        if generated_candidate:
            merged.append({**generated_candidate, "auditTrail": existing_candidate.get("auditTrail") or generated_candidate.get("auditTrail") or []})
            seen.add(candidate_id)

    for generated_candidate in generated:
        candidate_id = str(generated_candidate.get("candidateId") or "")
        if candidate_id and candidate_id not in seen and candidate_id not in protected_ids:
            merged.append(generated_candidate)

    governance["candidates"] = merged
    return governance


def _build_allocation_candidates(run_id: str, warehouse_comparison: dict) -> list[dict]:
    issues = warehouse_comparison.get("allocationIssues") if isinstance(warehouse_comparison, dict) else []
    candidates = []
    for issue in issues or []:
        employee_key = str(issue.get("employeeKey") or "")
        employee_name = str(issue.get("employeeName") or "")
        warehouses = issue.get("warehouses") if isinstance(issue.get("warehouses"), list) else []
        warehouse_ids = "_".join(str(row.get("warehouseId") or "") for row in warehouses)
        candidate_id = "allocation_" + "_".join(
            _safe_record_id(part)
            for part in (run_id, employee_key or employee_name, warehouse_ids)
            if str(part or "").strip()
        )
        candidates.append(
            {
                "candidateId": candidate_id,
                "decision": "candidate_only",
                "status": "pending_user_confirmation",
                "requiresConfirmation": True,
                "issueType": "cross_warehouse_employee_allocation",
                "employeeKey": employee_key,
                "employeeName": employee_name,
                "netAmountDelta": round(float(issue.get("netAmountDelta") or 0), 2),
                "warehouseCount": int(issue.get("warehouseCount") or len({str(row.get("warehouseId") or "") for row in warehouses})),
                "warehouses": warehouses,
                "recommendation": issue.get("recommendation") or "员工总额可抵消，但仓库归属金额不一致，需按仓库复核发票与账单归属。",
                "confirmationGate": "该候选只记录人工复核结论，不自动修改员工金额、仓库归属或正式核对结果。",
                "auditTrail": [
                    {
                        "action": "created",
                        "actor": "system",
                        "reason": "cross_warehouse_employee_allocation_detected",
                        "warehouseCount": len(warehouses),
                    }
                ],
            }
        )
    return candidates


def _find_allocation_record(records: list[dict], candidate_id: str) -> dict | None:
    for record in records:
        if str(record.get("candidateId") or "") == candidate_id:
            return record
    return None


def _upsert_allocation_record(records: list[dict], record: dict) -> None:
    candidate_id = str(record.get("candidateId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("candidateId") or "") == candidate_id:
            records[index] = record
            return
    records.append(record)


def _remove_allocation_record(records: list[dict], candidate_id: str) -> None:
    records[:] = [record for record in records if str(record.get("candidateId") or "") != candidate_id]


def _latest_reocr_replay(replays: list[dict], source_file: str, warehouse_id: str = "") -> dict | None:
    for replay in reversed(replays):
        if str(replay.get("sourceFile") or "") != source_file:
            continue
        if warehouse_id and str(replay.get("warehouseId") or "") != warehouse_id:
            continue
        return replay
    return None


def _find_reocr_record(records: list[dict], candidate_id: str) -> dict | None:
    for record in records:
        if str(record.get("candidateId") or "") == candidate_id:
            return record
    return None


def _upsert_reocr_record(records: list[dict], record: dict) -> None:
    candidate_id = str(record.get("candidateId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("candidateId") or "") == candidate_id:
            records[index] = record
            return
    records.append(record)


def _remove_reocr_record(records: list[dict], candidate_id: str) -> None:
    records[:] = [record for record in records if str(record.get("candidateId") or "") != candidate_id]


def _find_name_mapping_record(records: list[dict], candidate_id: str) -> dict | None:
    for record in records:
        if str(record.get("candidateId") or "") == candidate_id:
            return record
    return None


def _upsert_name_mapping_record(records: list[dict], record: dict) -> None:
    candidate_id = str(record.get("candidateId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("candidateId") or "") == candidate_id:
            records[index] = record
            return
    records.append(record)


def _remove_name_mapping_record(records: list[dict], candidate_id: str) -> None:
    records[:] = [record for record in records if str(record.get("candidateId") or "") != candidate_id]


def _merge_name_mapping_candidates(governance: dict, generated_candidates: list[dict]) -> list[dict]:
    existing_candidates = list(governance.get("candidates") or [])
    protected_ids = {
        str(record.get("candidateId") or "")
        for record in [
            *(governance.get("activeMappings") or []),
            *(governance.get("rolledBackMappings") or []),
        ]
    }
    generated_by_id = {str(candidate.get("candidateId") or ""): candidate for candidate in generated_candidates}
    merged: list[dict] = []
    seen: set[str] = set()

    for existing in existing_candidates:
        candidate_id = str(existing.get("candidateId") or "")
        if not candidate_id:
            continue
        status = str(existing.get("status") or "")
        decision = str(existing.get("decision") or "")
        if candidate_id in protected_ids or status in {"confirmed", "active", "rolled_back"} or decision in {"confirmed", "active", "rolled_back"}:
            merged.append(existing)
            seen.add(candidate_id)
            continue
        generated = generated_by_id.get(candidate_id)
        if generated:
            merged.append({**generated, "auditTrail": existing.get("auditTrail") or generated.get("auditTrail") or []})
            seen.add(candidate_id)

    for generated in generated_candidates:
        candidate_id = str(generated.get("candidateId") or "")
        if not candidate_id or candidate_id in seen or candidate_id in protected_ids:
            continue
        merged.append(generated)
    return merged


def _summarize_name_mapping_auto_replay(
    candidate: dict,
    metadata: dict,
    *,
    amount_tolerance: float,
    hours_tolerance: float,
    historical_runs: list[dict] | None = None,
    current_run_id: str = "",
    limit: int = 20,
) -> dict:
    proposed = candidate.get("proposedMapping") if isinstance(candidate.get("proposedMapping"), dict) else {}
    if not proposed:
        raise HTTPException(status_code=400, detail="姓名映射候选缺少 proposedMapping。")
    current_impact = _name_mapping_replay_one_run(
        candidate,
        metadata,
        amount_tolerance=amount_tolerance,
        hours_tolerance=hours_tolerance,
    )
    historical_impacts = _name_mapping_historical_impacts(
        candidate,
        historical_runs or [],
        current_run_id=current_run_id,
        limit=limit,
        amount_tolerance=amount_tolerance,
        hours_tolerance=hours_tolerance,
    )
    historical_regressions = [impact for impact in historical_impacts if int(impact.get("regressionCount") or 0) > 0]
    decision = (
        "ready_for_user_confirmation"
        if current_impact["regressionCount"] == 0 and current_impact["fixedCount"] > 0 and not historical_regressions
        else "blocked_by_replay_regression"
    )
    insufficient_count = sum(1 for impact in historical_impacts if impact.get("decision") == "insufficient_data")
    return {
        "candidateId": candidate.get("candidateId"),
        "decision": decision,
        "requiresConfirmation": True,
        "mode": "current_and_historical_name_mapping_replay",
        "summary": {
            "fixedCount": current_impact["fixedCount"],
            "regressionCount": current_impact["regressionCount"] + len(historical_regressions),
            "manualReviewCount": 0 if current_impact["fixedCount"] else 1,
            "currentExceptionCount": current_impact["currentExceptionCount"],
            "projectedExceptionCount": current_impact["projectedExceptionCount"],
            "exceptionDelta": current_impact["summaryDelta"].get("exceptionCount", 0),
            "currentMatchRate": current_impact["currentMatchRate"],
            "projectedMatchRate": current_impact["projectedMatchRate"],
            "historicalCheckedCount": len([impact for impact in historical_impacts if impact.get("decision") != "insufficient_data"]),
            "historicalInsufficientCount": insufficient_count,
            "historicalRegressionCount": len(historical_regressions),
            "reason": "姓名映射只在人工确认后写入当前批次 manualNameMapping。",
        },
        "summaryDelta": current_impact["summaryDelta"],
        "affectedRows": current_impact["affectedRows"],
        "candidateMatches": current_impact["candidateMatches"],
        "historicalImpacts": historical_impacts,
    }


def _name_mapping_replay_one_run(
    candidate: dict,
    metadata: dict,
    *,
    amount_tolerance: float,
    hours_tolerance: float,
) -> dict:
    from .engine.labor.models import line_items_from_dicts

    proposed = candidate.get("proposedMapping") if isinstance(candidate.get("proposedMapping"), dict) else {}
    pdf_rows = line_items_from_dicts(metadata.get("pdfExtractedRows") or [])
    excel_rows = line_items_from_dicts(metadata.get("excelRows") or [])
    if not pdf_rows or not excel_rows:
        raise ValueError("缺少可回放的 PDF/Excel 明细。")

    current_mapping = metadata.get("manualNameMapping") if isinstance(metadata.get("manualNameMapping"), dict) else {}
    projected_mapping = {**current_mapping, **{str(key): str(value) for key, value in proposed.items()}}
    current = compare_labor_items(
        pdf_rows,
        excel_rows,
        amount_tolerance=amount_tolerance,
        hours_tolerance=hours_tolerance,
        confidence_threshold=AI_CONFIG["confidence_threshold"],
        manual_name_mapping=current_mapping,
    )
    projected = compare_labor_items(
        pdf_rows,
        excel_rows,
        amount_tolerance=amount_tolerance,
        hours_tolerance=hours_tolerance,
        confidence_threshold=AI_CONFIG["confidence_threshold"],
        manual_name_mapping=projected_mapping,
    )
    affected_rows = _name_mapping_affected_rows(candidate, current.get("rows", []), projected.get("rows", []))
    summary_delta = _summary_delta(current.get("summary", {}), projected.get("summary", {}))
    fixed_count = sum(
        1
        for row in affected_rows
        if row.get("currentStatus") != "通过" and row.get("projectedStatus") == "通过"
    )
    regression_count = sum(
        1
        for row in affected_rows
        if row.get("currentStatus") == "通过" and row.get("projectedStatus") != "通过"
    )
    projected_exception_delta = int(summary_delta.get("exceptionCount", 0) or 0)
    if projected_exception_delta > 0:
        regression_count += projected_exception_delta
    return {
        "runId": metadata.get("id", ""),
        "supplierName": metadata.get("supplierName", ""),
        "periodStart": metadata.get("periodStart", ""),
        "periodEnd": metadata.get("periodEnd", ""),
        "decision": "compatible" if regression_count == 0 else "regression",
        "fixedCount": fixed_count,
        "regressionCount": regression_count,
        "currentExceptionCount": current.get("summary", {}).get("exceptionCount", 0),
        "projectedExceptionCount": projected.get("summary", {}).get("exceptionCount", 0),
        "currentMatchRate": current.get("summary", {}).get("matchRate", 0),
        "projectedMatchRate": projected.get("summary", {}).get("matchRate", 0),
        "summaryDelta": summary_delta,
        "affectedRows": affected_rows,
        "candidateMatches": projected.get("candidateMatches", []),
    }


def _name_mapping_historical_impacts(
    candidate: dict,
    historical_runs: list[dict],
    *,
    current_run_id: str,
    limit: int,
    amount_tolerance: float,
    hours_tolerance: float,
) -> list[dict]:
    impacts = []
    for metadata in historical_runs:
        if not isinstance(metadata, dict) or str(metadata.get("id") or "") == str(current_run_id):
            continue
        if len(impacts) >= limit:
            break
        try:
            impact = _name_mapping_replay_one_run(
                candidate,
                metadata,
                amount_tolerance=amount_tolerance,
                hours_tolerance=hours_tolerance,
            )
        except ValueError:
            impact = {
                "runId": metadata.get("id", ""),
                "supplierName": metadata.get("supplierName", ""),
                "periodStart": metadata.get("periodStart", ""),
                "periodEnd": metadata.get("periodEnd", ""),
                "decision": "insufficient_data",
                "fixedCount": 0,
                "regressionCount": 0,
                "reason": "历史批次缺少可重算的 PDF/Excel 明细。",
            }
        impacts.append(impact)
    return impacts


def _name_mapping_affected_rows(candidate: dict, current_rows: list[dict], projected_rows: list[dict]) -> list[dict]:
    cache_name = str(candidate.get("cacheEmployeeName") or "").strip()
    excel_name = str(candidate.get("excelEmployeeName") or "").strip()
    affected_names = {_compact_match_text(cache_name), _compact_match_text(excel_name)}
    affected_names.discard("")
    current_matches = [row for row in current_rows if _row_matches_any_affected_name(row, affected_names)]
    projected_matches = [row for row in projected_rows if _row_matches_any_affected_name(row, affected_names)]
    if not current_matches and cache_name:
        current_matches = [row for row in current_rows if cache_name.lower() in str(row.get("employeeName") or "").lower()]
    if not projected_matches and excel_name:
        projected_matches = [row for row in projected_rows if excel_name.lower() in str(row.get("employeeName") or "").lower()]

    rows = []
    max_len = max(len(current_matches), len(projected_matches), 1)
    for index in range(max_len):
        current = current_matches[index] if index < len(current_matches) else {}
        projected = projected_matches[index] if index < len(projected_matches) else {}
        rows.append(
            {
                "cacheEmployeeName": cache_name,
                "excelEmployeeName": excel_name,
                "currentEmployeeName": current.get("employeeName", ""),
                "projectedEmployeeName": projected.get("employeeName", ""),
                "currentStatus": current.get("matchStatus", ""),
                "projectedStatus": projected.get("matchStatus", ""),
                "currentAmountDelta": round(float(current.get("amountDelta") or 0), 2),
                "projectedAmountDelta": round(float(projected.get("amountDelta") or 0), 2),
                "currentHoursDelta": round(float(current.get("hoursDelta") or 0), 2),
                "projectedHoursDelta": round(float(projected.get("hoursDelta") or 0), 2),
                "sourceRefs": projected.get("sourceRefs") or current.get("sourceRefs") or "",
            }
        )
    return rows


def _build_name_mapping_candidates_from_reocr_plan(run_id: str, reocr_plan: dict) -> list[dict]:
    candidates: list[dict] = []
    seen: set[tuple[str, str, str, str]] = set()
    for task in reocr_plan.get("tasks", []) or []:
        diagnostics = task.get("diagnostics") if isinstance(task.get("diagnostics"), dict) else {}
        for pair in diagnostics.get("suspectedNamePairs", []) or []:
            if not isinstance(pair, dict):
                continue
            cache_name = str(pair.get("cacheEmployeeName") or "").strip()
            excel_name = str(pair.get("excelEmployeeName") or "").strip()
            source_file = str(task.get("sourceFile") or "").strip()
            warehouse_id = str(task.get("warehouseId") or "").strip()
            if not cache_name or not excel_name:
                continue
            identity = (source_file, warehouse_id, cache_name, excel_name)
            if identity in seen:
                continue
            seen.add(identity)
            candidate_id = "name_map_" + "_".join(
                _safe_record_id(part)
                for part in (run_id, source_file, warehouse_id or "all", cache_name, excel_name)
            )
            candidates.append(
                {
                    "candidateId": candidate_id,
                    "decision": "candidate_only",
                    "status": "pending_user_confirmation",
                    "requiresConfirmation": True,
                    "sourceFile": source_file,
                    "warehouseId": warehouse_id,
                    "cacheEmployeeName": cache_name,
                    "excelEmployeeName": excel_name,
                    "proposedMapping": {cache_name: excel_name},
                    "amountGap": pair.get("amountGap", 0),
                    "hoursGap": pair.get("hoursGap", 0),
                    "confidence": pair.get("confidence", "medium"),
                    "recommendation": pair.get("recommendation", "人工确认姓名映射后才能应用。"),
                    "evidence": {
                        "sourceRefs": pair.get("sourceRefs", ""),
                        "cacheAmount": pair.get("cacheAmount", 0),
                        "excelAmount": pair.get("excelAmount", 0),
                        "cacheHours": pair.get("cacheHours", 0),
                        "excelHours": pair.get("excelHours", 0),
                    },
                    "auditTrail": [
                        {
                            "action": "created",
                            "actor": "system",
                            "reason": "reocr_suspected_name_pair",
                        }
                    ],
                }
            )
    return candidates


def _source_file_from_candidate_refs(source_refs: str) -> str:
    for segment in str(source_refs or "").split(";"):
        token = segment.strip().split(" ")[0] if segment.strip() else ""
        if token.lower().endswith(".pdf"):
            return Path(token).name
    return ""


def _candidate_match_number(match: dict, field: str) -> float:
    try:
        return round(float(match.get(field) or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _candidate_impact_summary(*, amount_delta: float, hours_delta: float) -> str:
    parts: list[str] = []
    if abs(float(amount_delta or 0)) > 0.005:
        direction = "PDF 高于 Excel" if amount_delta > 0 else "PDF 少于 Excel"
        parts.append(f"{direction} ${abs(float(amount_delta)):,.2f}")
    if abs(float(hours_delta or 0)) > 0.005:
        direction = "PDF 工时多于 Excel" if hours_delta > 0 else "PDF 工时少于 Excel"
        parts.append(f"{direction} {abs(float(hours_delta)):.2f}")
    return "；".join(parts) if parts else "金额和工时均一致"


def _build_name_mapping_candidates_from_candidate_matches(run_id: str, candidate_matches: list[dict]) -> list[dict]:
    candidates: list[dict] = []
    seen: set[tuple[str, str, str, str]] = set()
    for match in candidate_matches or []:
        if not isinstance(match, dict):
            continue
        if str(match.get("issueType") or "") == "combined_pdf_row":
            continue
        pdf_name = str(match.get("pdfEmployeeName") or "").strip()
        excel_name = str(match.get("excelEmployeeName") or "").strip()
        if not pdf_name or not excel_name:
            continue
        source_refs = str(match.get("sourceRefs") or "")
        source_file = str(match.get("sourceFile") or "").strip() or _source_file_from_candidate_refs(source_refs)
        warehouse_id = str(match.get("warehouseId") or "").strip()
        identity = (source_file, warehouse_id, pdf_name, excel_name)
        if identity in seen:
            continue
        seen.add(identity)

        amount_delta = _candidate_match_number(match, "amountDelta")
        hours_delta = _candidate_match_number(match, "hoursDelta")
        exact_totals = abs(amount_delta) <= AI_CONFIG["amount_tolerance"] and abs(hours_delta) <= AI_CONFIG["hours_tolerance"]
        projected_fixed_count = 2 if exact_totals else 0
        match_reason = "姓名相似且金额/工时一致" if exact_totals else "姓名相似，但金额或工时仍需复核"
        impact_summary = _candidate_impact_summary(amount_delta=amount_delta, hours_delta=hours_delta)
        candidate_id = "name_map_" + "_".join(
            _safe_record_id(part)
            for part in (run_id, source_file or "candidate_match", warehouse_id or "all", pdf_name, excel_name)
        )
        candidates.append(
            {
                "candidateId": candidate_id,
                "decision": "candidate_only",
                "status": "pending_user_confirmation",
                "requiresConfirmation": True,
                "sourceFile": source_file,
                "warehouseId": warehouse_id,
                "cacheEmployeeName": pdf_name,
                "excelEmployeeName": excel_name,
                "proposedMapping": {pdf_name: excel_name},
                "amountGap": amount_delta,
                "hoursGap": hours_delta,
                "confidence": "high" if exact_totals else "medium",
                "projectedFixedExceptionCount": projected_fixed_count,
                "matchReason": match_reason,
                "businessQuestion": (
                    f"是否确认 PDF 名称 {pdf_name} 对应 Excel 员工 {excel_name}？"
                    f"{'确认后预计减少 ' + str(projected_fixed_count) + ' 项异常。' if projected_fixed_count else '金额或工时仍不同，需先复核差异口径。'}"
                ),
                "impactSummary": impact_summary,
                "cannotAutoResolveReason": (
                    "姓名匹配会改变员工级对账归属，必须预览影响并由人工确认后才可写入。"
                    if exact_totals
                    else "姓名相似不能解释金额或工时差异，必须先复核差异口径，不能直接确认匹配。"
                ),
                "recommendation": (
                    "金额/工时一致，建议先预览影响，再由人工确认姓名匹配。"
                    if exact_totals
                    else "姓名相似但金额或工时仍有差异，需先复核金额口径，再决定是否确认映射。"
                ),
                "evidence": {
                    "sourceRefs": source_refs,
                    "cacheAmount": match.get("pdfAmountTotal", 0),
                    "excelAmount": match.get("excelAmountTotal", 0),
                    "cacheHours": match.get("pdfHoursTotal", 0),
                    "excelHours": match.get("excelHoursTotal", 0),
                    "nameSimilarity": match.get("nameSimilarity", 0),
                },
                "auditTrail": [
                    {
                        "action": "created",
                        "actor": "system",
                        "reason": "candidate_match_name_pair",
                    }
                ],
            }
        )
    return candidates


def _safe_record_id(value: str) -> str:
    token = re.sub(r"[^0-9A-Za-z_-]+", "_", str(value or "").strip())
    return token.strip("_") or "unknown"


def _build_low_confidence_correction_candidates(run_id: str, pdf_rows: list, extraction_quality: dict) -> list[dict]:
    low_rows = extraction_quality.get("lowConfidenceRows") or []
    if not low_rows:
        return []
    by_key = {
        (
            getattr(row, "source_file", ""),
            getattr(row, "source_page_or_row", ""),
            getattr(row, "employee_name_raw", ""),
        ): row
        for row in pdf_rows
    }
    candidates = []
    for index, low in enumerate(low_rows, start=1):
        key = (
            str(low.get("source_file") or ""),
            str(low.get("source_page_or_row") or ""),
            str(low.get("employee_name_raw") or ""),
        )
        row = by_key.get(key)
        proposed = {
            "employeeName": str(low.get("employee_name_raw") or ""),
            "hours": float(low.get("hours") or getattr(row, "hours", 0) or 0),
            "amount": float(low.get("amount") or getattr(row, "amount", 0) or 0),
            "sourceFile": key[0],
            "sourcePageOrRow": key[1],
        }
        candidates.append(
            {
                "candidateId": f"correction_{run_id}_{index}",
                "decision": "candidate_only",
                "status": "pending_user_confirmation",
                "requiresConfirmation": True,
                "source": f"labor_run:{run_id}",
                "createdAt": utcnow_naive().isoformat(),
                "reason": "low_confidence_extraction",
                "original": proposed,
                "proposed": proposed,
                "confidence": float(low.get("confidence") or getattr(row, "confidence", 0) or 0),
                "evidence": {
                    "sourceFile": key[0],
                    "sourcePageOrRow": key[1],
                    "evidenceText": str(getattr(row, "evidence_text", "") or "")[:300],
                },
                "auditTrail": [
                    {
                        "action": "created",
                        "actor": "system",
                        "reason": "低置信度抽取行进入候选修正，等待人工确认。",
                    }
                ],
            }
        )
    return candidates


def _find_correction_record(records: list[dict], candidate_id: str) -> dict | None:
    for record in records:
        if str(record.get("candidateId") or "") == candidate_id:
            return record
    return None


def _upsert_correction_record(records: list[dict], record: dict) -> None:
    candidate_id = str(record.get("candidateId") or "")
    for index, existing in enumerate(records):
        if str(existing.get("candidateId") or "") == candidate_id:
            records[index] = record
            return
    records.append(record)


def _remove_correction_record(records: list[dict], candidate_id: str) -> None:
    records[:] = [record for record in records if str(record.get("candidateId") or "") != candidate_id]


def _summarize_correction_auto_replay(candidate: dict, metadata: dict, *, amount_tolerance: float, hours_tolerance: float) -> dict:
    proposed = candidate.get("proposed") if isinstance(candidate.get("proposed"), dict) else {}
    employee_name = str(proposed.get("employeeName") or "").strip()
    comparison_row = _find_comparison_row_for_correction(employee_name, metadata.get("comparisonRows") or [])
    evidence = candidate.get("evidence") if isinstance(candidate.get("evidence"), dict) else {}
    evidence_text = str(evidence.get("evidenceText") or "").strip()
    affected = [employee_name] if employee_name else []
    if not comparison_row:
        return {
            "candidateId": candidate.get("candidateId"),
            "decision": "needs_manual_review",
            "summary": {
                "affectedEmployees": affected,
                "fixedCount": 0,
                "regressionCount": 0,
                "manualReviewCount": 1,
                "reason": "未能在当前对账明细中定位该修正候选，不能自动确认影响。",
            },
            "impact": [],
        }

    excel_amount = float(comparison_row.get("excelAmountTotal") or 0)
    excel_hours = float(comparison_row.get("excelHoursTotal") or 0)
    proposed_amount = float(proposed.get("amount") or 0)
    proposed_hours = float(proposed.get("hours") or 0)
    current_amount_delta = float(comparison_row.get("amountDelta") or 0)
    current_hours_delta = float(comparison_row.get("hoursDelta") or 0)
    projected_amount_delta = round(proposed_amount - excel_amount, 2)
    projected_hours_delta = round(proposed_hours - excel_hours, 2)
    projected_passes = abs(projected_amount_delta) <= amount_tolerance and abs(projected_hours_delta) <= hours_tolerance
    current_status = str(comparison_row.get("matchStatus") or "")
    current_passed = current_status == "通过"
    regression = current_passed and not projected_passes
    risk_flags = list(comparison_row.get("riskFlags") or [])
    risk_reduced = "低置信度抽取" in risk_flags and bool(evidence_text)
    fixed = projected_passes and (risk_reduced or current_status != "通过")
    decision = "blocked_by_replay_regression" if regression else "ready_for_user_confirmation"
    impact = {
        "employeeName": comparison_row.get("employeeName") or employee_name,
        "currentStatus": current_status,
        "currentAmountDelta": round(current_amount_delta, 2),
        "currentHoursDelta": round(current_hours_delta, 2),
        "projectedAmountDelta": projected_amount_delta,
        "projectedHoursDelta": projected_hours_delta,
        "riskFlags": risk_flags,
        "riskReduced": risk_reduced,
        "evidenceSource": {
            "sourceFile": evidence.get("sourceFile"),
            "sourcePageOrRow": evidence.get("sourcePageOrRow"),
        },
    }
    return {
        "candidateId": candidate.get("candidateId"),
        "decision": decision,
        "summary": {
            "affectedEmployees": affected,
            "fixedCount": 1 if fixed else 0,
            "regressionCount": 1 if regression else 0,
            "manualReviewCount": 0,
            "reason": "候选修正仅作为人工确认项，不会静默覆盖当前核对结论。",
        },
        "impact": [impact],
    }


def _find_comparison_row_for_correction(employee_name: str, comparison_rows: list) -> dict | None:
    target = _compact_match_text(employee_name)
    if not target:
        return None
    best_row = None
    best_score = 0
    for row in comparison_rows:
        if not isinstance(row, dict):
            continue
        haystack = _compact_match_text(
            " ".join(
                str(row.get(key) or "")
                for key in ("employeeName", "employeeKey", "pdfEmployeeName", "excelEmployeeName")
            )
        )
        if not haystack:
            continue
        score = len(set(target.split()) & set(haystack.split()))
        if target in haystack or haystack in target:
            score += 100
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_score > 0 else None


def _compact_match_text(value: str) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())
    return re.sub(r"\s+", " ", text).strip()


def _build_correction_projected_preview(metadata: dict, active_corrections: list[dict], *, amount_tolerance: float, hours_tolerance: float) -> dict:
    from .engine.labor.models import LaborLineItem, line_items_from_dicts

    pdf_rows = line_items_from_dicts(metadata.get("pdfExtractedRows") or [])
    excel_rows = line_items_from_dicts(metadata.get("excelRows") or [])
    if not pdf_rows or not excel_rows:
        raise ValueError("当前批次缺少可重算的 PDF/Excel 明细。")

    applied = []
    manual_review = []
    projected_pdf_rows = list(pdf_rows)
    for correction in active_corrections:
        proposed = correction.get("proposed") if isinstance(correction.get("proposed"), dict) else {}
        target_name = str(proposed.get("employeeName") or "").strip()
        target = _compact_match_text(target_name)
        source_file = str(proposed.get("sourceFile") or "")
        source_page = str(proposed.get("sourcePageOrRow") or "")
        matched_index = None
        for index, row in enumerate(projected_pdf_rows):
            same_source = (
                (not source_file or row.source_file == source_file)
                and (not source_page or row.source_page_or_row == source_page)
            )
            if same_source and _compact_match_text(row.employee_name_raw) == target:
                matched_index = index
                break
        if matched_index is None:
            manual_review.append(
                {
                    "candidateId": correction.get("candidateId"),
                    "employeeName": target_name,
                    "reason": "未能在 PDF 抽取明细中定位要替换的原始行。",
                }
            )
            continue

        original = projected_pdf_rows[matched_index]
        replacement = LaborLineItem(
            source_type=original.source_type,
            source_file=original.source_file,
            source_page_or_row=original.source_page_or_row,
            employee_id=original.employee_id,
            employee_name_raw=str(proposed.get("employeeName") or original.employee_name_raw),
            hours=float(proposed.get("hours") or original.hours or 0),
            amount=float(proposed.get("amount") or original.amount or 0),
            currency=original.currency,
            confidence=1.0,
            evidence_text=original.evidence_text,
            supplier=original.supplier,
            period_start=original.period_start,
            period_end=original.period_end,
            warehouse_id=original.warehouse_id,
        )
        projected_pdf_rows[matched_index] = replacement
        applied.append(
            {
                "candidateId": correction.get("candidateId"),
                "employeeName": replacement.employee_name_raw,
                "sourceFile": replacement.source_file,
                "sourcePageOrRow": replacement.source_page_or_row,
                "amountDelta": round(replacement.amount - original.amount, 2),
                "hoursDelta": round(replacement.hours - original.hours, 2),
            }
        )

    if not applied:
        raise ValueError("没有修正能应用到当前批次明细，无法生成预览。")

    original_summary = metadata.get("comparisonSummary") if isinstance(metadata.get("comparisonSummary"), dict) else {}
    projected = compare_labor_items(
        projected_pdf_rows,
        excel_rows,
        amount_tolerance=amount_tolerance,
        hours_tolerance=hours_tolerance,
        confidence_threshold=AI_CONFIG["confidence_threshold"],
        manual_name_mapping=metadata.get("manualNameMapping") if isinstance(metadata.get("manualNameMapping"), dict) else None,
    )
    projected_summary = projected["summary"]
    affected_names = [_compact_match_text(item.get("employeeName", "")) for item in applied]
    affected_rows = [
        row
        for row in projected["rows"]
        if _row_matches_any_affected_name(row, affected_names)
    ]
    return {
        "decision": "preview_only",
        "reason": "只读预览不会覆盖正式核对结果或报告。",
        "appliedCorrections": applied,
        "manualReview": manual_review,
        "summaryDelta": _summary_delta(original_summary, projected_summary),
        "originalSummary": _pick_projection_summary(original_summary),
        "projectedSummary": _pick_projection_summary(projected_summary),
        "affectedRows": affected_rows,
    }


def _build_correction_preview_preflight(preview: dict) -> dict:
    affected_rows = preview.get("affectedRows") if isinstance(preview.get("affectedRows"), list) else []
    affected_employee_names = sorted(
        {
            str(row.get("employeeName") or row.get("pdfEmployeeName") or row.get("excelEmployeeName") or "").strip()
            for row in affected_rows
            if str(row.get("employeeName") or row.get("pdfEmployeeName") or row.get("excelEmployeeName") or "").strip()
        }
    )
    projected_summary = preview.get("projectedSummary") if isinstance(preview.get("projectedSummary"), dict) else {}
    delta = preview.get("summaryDelta") if isinstance(preview.get("summaryDelta"), dict) else {}
    warnings = []
    exception_count = _safe_int(projected_summary.get("exceptionCount"))
    regression_count = sum(1 for row in affected_rows if row.get("matchStatus") != "通过")
    if exception_count:
        warnings.append(f"投影结果仍有 {exception_count} 项异常，修正后仍需人工复核。")
    if regression_count:
        warnings.append(f"受影响员工中仍有 {regression_count} 人未通过，不能直接作为上线结论。")
    return {
        "willOverwriteOfficialResult": False,
        "willRegenerateDiffReport": False,
        "current": preview.get("originalSummary") if isinstance(preview.get("originalSummary"), dict) else {},
        "projected": projected_summary,
        "delta": delta,
        "affectedScopeCount": len(preview.get("appliedCorrections") or []),
        "affectedEmployeeCount": len(affected_employee_names),
        "affectedEmployees": affected_employee_names[:50],
        "coverageCompleteAfterApply": not warnings,
        "blockingAfterApply": bool(warnings),
        "postApplyWarnings": warnings,
        "formalResultFields": [],
    }


def _build_rule_replay_preflight(replay_summary: dict) -> dict:
    summary = replay_summary.get("summary") if isinstance(replay_summary.get("summary"), dict) else {}
    fixed = replay_summary.get("fixed") if isinstance(replay_summary.get("fixed"), list) else []
    regressions = replay_summary.get("regressions") if isinstance(replay_summary.get("regressions"), list) else []
    unchanged = replay_summary.get("unchanged") if isinstance(replay_summary.get("unchanged"), list) else []
    affected_suppliers = sorted(
        {
            str(row.get("supplier") or "").strip()
            for row in [*fixed, *regressions, *unchanged]
            if str(row.get("supplier") or "").strip()
        }
    )
    warnings = []
    if regressions:
        warnings.append(f"历史回放发现 {len(regressions)} 个回归批次，不能确认生效。")
    if not fixed:
        warnings.append("历史回放没有发现可修复批次，证据不足。")
    replayed_count = _safe_int(summary.get("replayedCount"))
    if replayed_count == 0:
        warnings.append("尚未完成历史批次回放。")
    return {
        "willOverwriteOfficialResult": False,
        "willRegenerateDiffReport": False,
        "current": {
            "replayedCount": replayed_count,
            "fixedCount": _safe_int(summary.get("fixedCount")),
            "regressionCount": _safe_int(summary.get("regressionCount")),
            "unchangedCount": _safe_int(summary.get("unchangedCount")),
        },
        "projected": {},
        "delta": {
            "fixedCount": _safe_int(summary.get("fixedCount")),
            "regressionCount": _safe_int(summary.get("regressionCount")),
        },
        "affectedScopeCount": replayed_count,
        "affectedEmployeeCount": 0,
        "affectedEmployees": [],
        "affectedSuppliers": affected_suppliers,
        "fixedRuns": fixed[:20],
        "regressionRuns": regressions[:20],
        "coverageCompleteAfterApply": bool(fixed) and not regressions,
        "blockingAfterApply": bool(warnings),
        "postApplyWarnings": warnings,
        "formalResultFields": [],
    }


def _build_profile_replay_preflight(candidate: dict, replay_summary: dict) -> dict:
    summary = replay_summary.get("summary") if isinstance(replay_summary.get("summary"), dict) else {}
    compatible = replay_summary.get("compatible") if isinstance(replay_summary.get("compatible"), list) else []
    regressions = replay_summary.get("regressions") if isinstance(replay_summary.get("regressions"), list) else []
    replay_results = replay_summary.get("replayResults") if isinstance(replay_summary.get("replayResults"), list) else []
    profile_data = candidate.get("profileData") if isinstance(candidate.get("profileData"), dict) else {}
    changed_fields = sorted(
        key
        for key in profile_data
        if key not in {"key", "version"} and profile_data.get(key) not in (None, "", [], {})
    )
    affected_suppliers = sorted(
        {
            str(row.get("supplier") or "").strip()
            for row in replay_results
            if str(row.get("supplier") or "").strip()
        }
    )
    warnings = []
    if regressions:
        warnings.append(f"历史回放发现 {len(regressions)} 个 Profile 回归风险批次，不能确认生效。")
    if not compatible:
        warnings.append("历史回放没有发现兼容批次，证据不足。")
    if not changed_fields:
        warnings.append("Profile 候选缺少可审计的字段变更。")
    replayed_count = _safe_int(summary.get("replayedCount"))
    if replayed_count == 0:
        warnings.append("尚未完成历史批次回放。")
    return {
        "willOverwriteOfficialResult": False,
        "willRegenerateDiffReport": False,
        "current": {
            "replayedCount": replayed_count,
            "compatibleCount": _safe_int(summary.get("compatibleCount")),
            "regressionCount": _safe_int(summary.get("regressionCount")),
            "unchangedCount": _safe_int(summary.get("unchangedCount")),
            "evidenceCount": _safe_int(summary.get("evidenceCount")),
        },
        "projected": {},
        "delta": {
            "compatibleCount": _safe_int(summary.get("compatibleCount")),
            "regressionCount": _safe_int(summary.get("regressionCount")),
        },
        "affectedScopeCount": replayed_count,
        "affectedEmployeeCount": 0,
        "affectedEmployees": [],
        "affectedSuppliers": affected_suppliers,
        "changedFields": changed_fields,
        "compatibleRuns": compatible[:20],
        "regressionRuns": regressions[:20],
        "coverageCompleteAfterApply": bool(compatible) and not regressions,
        "blockingAfterApply": bool(warnings),
        "postApplyWarnings": warnings,
        "formalResultFields": [],
    }


def _attach_correction_preflight(governance: dict, active_corrections: list[dict], preflight: dict) -> dict:
    updated = dict(governance)
    active_ids = {str(item.get("candidateId") or "") for item in active_corrections}
    enriched = []
    for correction in updated.get("activeCorrections") or []:
        if str(correction.get("candidateId") or "") in active_ids:
            enriched.append({**correction, "preflight": preflight})
        else:
            enriched.append(correction)
    updated["activeCorrections"] = enriched
    return updated


def _build_reocr_projected_preview(metadata: dict, active: dict) -> dict:
    replay = active.get("replay") if isinstance(active.get("replay"), dict) else {}
    original_summary = metadata.get("comparisonSummary") if isinstance(metadata.get("comparisonSummary"), dict) else {}
    projected_summary = replay.get("comparison") if isinstance(replay.get("comparison"), dict) else {}
    affected_rows = replay.get("previewRows") if isinstance(replay.get("previewRows"), list) else []
    exception_rows = replay.get("exceptionRows") if isinstance(replay.get("exceptionRows"), list) else []
    return {
        "decision": "preview_only",
        "reason": "图片识别结果已人工确认；本报告只读预览识别结果，不覆盖正式核对结果或原始抽取。",
        "appliedCorrections": [
            {
                "candidateId": active.get("candidateId", ""),
                "employeeName": f"图片识别：{active.get('sourceFile', '')}",
                "sourceFile": active.get("sourceFile", ""),
                "sourcePageOrRow": active.get("warehouseId", ""),
                "amountDelta": replay.get("summary", {}).get("fixedCacheDelta", 0) if isinstance(replay.get("summary"), dict) else 0,
                "hoursDelta": 0,
            }
        ],
        "manualReview": [
            {
                "candidateId": active.get("candidateId", ""),
                "employeeName": row.get("employeeName", ""),
                "reason": row.get("matchStatus", "图片识别结果仍需人工处理"),
            }
            for row in exception_rows
        ],
        "summaryDelta": _summary_delta(original_summary, projected_summary),
        "originalSummary": _pick_projection_summary(original_summary),
        "projectedSummary": _pick_projection_summary(projected_summary),
        "affectedRows": affected_rows,
    }


def _generate_reocr_adopted_diff_report(
    run_id: str,
    metadata: dict,
    *,
    comparison_summary: dict,
    comparison_rows: list[dict],
    candidate_matches: list[dict],
    presentation: dict | None = None,
) -> dict:
    from .engine.labor.models import line_items_from_dicts

    pdf_rows = line_items_from_dicts(metadata.get("pdfExtractedRows") or [])
    excel_rows = line_items_from_dicts(metadata.get("excelRows") or [])
    comparison = {
        "summary": comparison_summary,
        "rows": comparison_rows,
        "candidateMatches": candidate_matches,
    }
    presentation = presentation or _build_validated_labor_presentation(
        comparison_summary,
        comparison_rows,
        candidate_matches,
        excel_record_count=len(excel_rows),
    )
    report_path = get_labor_run_dir(run_id) / safe_labor_filename("海外劳务工报账核对报告.xlsx", "差异报告_图片识别已采纳")
    build_labor_report(
        report_path,
        comparison,
        pdf_rows,
        excel_rows,
        metadata.get("excelMapping") if isinstance(metadata.get("excelMapping"), dict) else {},
        metadata.get("warehouseComparison") if isinstance(metadata.get("warehouseComparison"), dict) else {},
        metadata.get("extractionQuality") if isinstance(metadata.get("extractionQuality"), dict) else {},
        reconciliation_diagnostics=metadata.get("reconciliationDiagnostics") if isinstance(metadata.get("reconciliationDiagnostics"), dict) else {},
        ai_cache_audit=metadata.get("aiCacheAudit") if isinstance(metadata.get("aiCacheAudit"), dict) else {},
        presentation=presentation,
    )
    return attach_labor_file(run_id, report_path, "差异报告")


def _build_validated_labor_presentation(
    comparison_summary: dict,
    comparison_rows: list[dict],
    candidate_matches: list[dict],
    *,
    excel_record_count: int,
) -> dict:
    presentation = build_labor_presentation(
        {
            "summary": comparison_summary,
            "rows": comparison_rows,
            "candidateMatches": candidate_matches,
        },
        excel_record_count=excel_record_count,
    )
    errors = validate_labor_presentation(presentation)
    if errors:
        raise ValueError("核对展示口径校验失败: " + "；".join(errors))
    return presentation


def _select_reocr_apply_candidates(governance: dict, candidate_ids: list | None) -> list[dict]:
    requested = {str(value) for value in (candidate_ids or []) if str(value).strip()}
    candidates = []
    for candidate in governance.get("activeCandidates") or []:
        candidate_id = str(candidate.get("candidateId") or "")
        if requested and candidate_id not in requested:
            continue
        if candidate.get("status") == "applied" or candidate.get("decision") == "applied":
            continue
        replay = candidate.get("replay") if isinstance(candidate.get("replay"), dict) else {}
        if replay.get("decision") != "ready_for_user_confirmation":
            continue
        candidates.append(candidate)
    return candidates


def _build_reocr_batch_apply_preview(metadata: dict, candidates: list[dict]) -> dict:
    scopes = []
    duplicate_scopes = set()
    seen_scopes = set()
    comparison_rows = []
    candidate_matches = []
    summaries = []
    for candidate in candidates:
        replay = candidate.get("replay") if isinstance(candidate.get("replay"), dict) else {}
        scope = (str(candidate.get("sourceFile") or replay.get("sourceFile") or ""), str(candidate.get("warehouseId") or replay.get("warehouseId") or ""))
        if scope in seen_scopes:
            duplicate_scopes.add(scope)
        seen_scopes.add(scope)
        scopes.append({"sourceFile": scope[0], "warehouseId": scope[1], "candidateId": candidate.get("candidateId", "")})
        rows = replay.get("comparisonRows") if isinstance(replay.get("comparisonRows"), list) else replay.get("previewRows")
        if isinstance(rows, list):
            comparison_rows.extend(rows)
        if isinstance(replay.get("candidateMatches"), list):
            candidate_matches.extend(replay.get("candidateMatches") or [])
        if isinstance(replay.get("comparison"), dict):
            summaries.append(replay.get("comparison") or {})

    coverage = _build_reocr_plan_coverage(metadata, candidates)
    comparison_summary = _summarize_reocr_batch_comparison(summaries, comparison_rows)
    comparison_summary = _labor_reocr_pending_revalidation_summary(
        {
            **comparison_summary,
            "notInInvoiceCount": sum(1 for row in comparison_rows if row.get("matchStatus") == "Excel有PDF无"),
            "reocrCandidateApplied": True,
            "reocrBatchApplied": True,
        }
    )
    presentation = _build_validated_labor_presentation(
        comparison_summary,
        comparison_rows,
        candidate_matches,
        excel_record_count=len(metadata.get("excelRows") or []),
    )
    preflight = _build_reocr_apply_preflight(metadata, candidates, comparison_summary, comparison_rows, duplicate_scope_count=len(duplicate_scopes))
    return {
        "decision": "ready_for_batch_apply" if candidates and not duplicate_scopes else "blocked_by_duplicate_scope" if duplicate_scopes else "no_applicable_candidates",
        "summary": {
            "candidateCount": len(candidates),
            "scopeCount": len(scopes),
            "duplicateScopeCount": len(duplicate_scopes),
            "plannedTaskCount": coverage["plannedTaskCount"],
            "confirmedTaskCount": coverage["confirmedTaskCount"],
            "appliedTaskCount": coverage["appliedTaskCount"],
            "missingConfirmedTaskCount": len(coverage["missingConfirmedTasks"]),
            "missingAppliedTaskCount": len(coverage["missingAppliedTasks"]),
            "rowCount": len(comparison_rows),
            "exceptionCount": comparison_summary.get("exceptionCount", 0),
            "pdfAmountTotal": comparison_summary.get("pdfAmountTotal", 0),
            "excelAmountTotal": comparison_summary.get("excelAmountTotal", 0),
            "amountDeltaTotal": comparison_summary.get("amountDeltaTotal", 0),
        },
        "coverage": coverage,
        "scopes": scopes,
        "duplicateScopes": [{"sourceFile": item[0], "warehouseId": item[1]} for item in sorted(duplicate_scopes)],
        "preflight": preflight,
        "comparisonSummary": comparison_summary,
        "comparisonRows": comparison_rows,
        "candidateMatches": candidate_matches,
        "presentation": presentation,
    }


def _labor_reocr_pending_revalidation_summary(summary: dict) -> dict:
    pending = {
        **summary,
        "conclusionLevel": "warning",
        "conclusionMessage": "已采纳人工确认的图片识别候选；重新执行整批完整核对前不能机器通过。",
        "canRelease": False,
    }
    pending.update(_labor_uat_review_state(pending))
    return pending


def _stable_labor_file_records(metadata: dict, key: str) -> list[dict]:
    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    records = files.get(key) if isinstance(files.get(key), list) else []
    stable_records = []
    for record in records:
        if not isinstance(record, dict):
            continue
        stable_record = {
            str(field): value
            for field, value in record.items()
            if field not in {"path", "downloadUrl", "url", "signedUrl"}
        }
        local_path = Path(str(record.get("path") or ""))
        if local_path.exists() and local_path.is_file():
            digest = hashlib.sha256()
            with local_path.open("rb") as handle:
                for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                    digest.update(chunk)
            stable_record["sizeBytes"] = local_path.stat().st_size
            stable_record["sha256"] = digest.hexdigest()
        stable_records.append(stable_record)
    return stable_records


def _labor_mapping_input_fingerprint(metadata: dict) -> str:
    """Bind a mapping preflight to the exact workbook objects it inspected."""

    payload = {
        "runId": str(metadata.get("id") or ""),
        "workbooks": _stable_labor_file_records(metadata, "workbooks"),
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _labor_result_input_fingerprint(metadata: dict) -> str:
    """Bind an official result to the exact input and active governance state."""

    governance_specs = (
        ("ruleGovernance", "activeRules"),
        ("nameMappingGovernance", "activeMappings"),
        ("allocationGovernance", "activeAllocations"),
        ("profileGovernance", "activeProfiles"),
        ("correctionGovernance", "activeCorrections"),
        ("reocrReplayGovernance", "activeCandidates"),
    )
    active_governance = {}
    for governance_key, active_key in governance_specs:
        governance = metadata.get(governance_key) if isinstance(metadata.get(governance_key), dict) else {}
        active_governance[governance_key] = (
            governance.get(active_key) if isinstance(governance.get(active_key), list) else []
        )

    payload = {
        "supplierName": str(metadata.get("supplierName") or metadata.get("supplier") or ""),
        "periodStart": str(metadata.get("periodStart") or ""),
        "periodEnd": str(metadata.get("periodEnd") or ""),
        "currency": str(metadata.get("currency") or ""),
        "reconciliationScope": str(metadata.get("reconciliationScope") or "employee_detail_required"),
        "pdfInvoices": _stable_labor_file_records(metadata, "pdfInvoices"),
        "workbooks": _stable_labor_file_records(metadata, "workbooks"),
        "workbookSheet": str(metadata.get("workbookSheet") or ""),
        "excelMapping": metadata.get("excelMapping") if isinstance(metadata.get("excelMapping"), dict) else {},
        "workbookMappings": metadata.get("workbookMappings") if isinstance(metadata.get("workbookMappings"), list) else [],
        "manualNameMapping": metadata.get("manualNameMapping") if isinstance(metadata.get("manualNameMapping"), dict) else {},
        "activeGovernance": active_governance,
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _labor_invalidate_official_result(
    metadata: dict,
    *,
    status: str,
    reason_code: str,
    message: str,
    files: dict | None = None,
) -> dict:
    clean_files = dict(files if isinstance(files, dict) else metadata.get("files") or {})
    clean_files.pop("diffReport", None)
    clean_files.pop("businessReport", None)
    return {
        "status": status,
        "files": clean_files,
        "comparisonSummary": {},
        "comparisonRows": [],
        "candidateMatches": [],
        "presentation": {},
        "warehouseComparison": {},
        "extractionQuality": {},
        "reconciliationDiagnostics": {
            "level": "warning",
            "message": message,
            "issues": [{"code": reason_code, "level": "warning", "message": message}],
        },
        "batchGuard": {
            "status": reason_code,
            "message": message,
            "allowReleasableReport": False,
            "unresolvedFiles": [],
        },
        "resultInputFingerprint": "",
        "machineCheckStatus": "needs_review",
        "businessReviewStatus": "pending",
        "manualReviewRequired": True,
        "directPaymentAllowed": False,
        "requiresHumanReview": True,
        "diffDownloadUrl": "",
        "businessReportDownloadUrl": "",
    }


def _labor_reocr_revalidation_batch_guard() -> dict:
    return {
        "status": "reocr_revalidation_required",
        "message": "图片识别候选已采纳，但逐文件明细覆盖、金额闭合和核对诊断尚未重新计算。",
        "allowReleasableReport": False,
        "unresolvedFiles": [],
    }


def _labor_reocr_revalidation_diagnostics() -> dict:
    return {
        "level": "warning",
        "message": "图片识别候选采纳后需要重新执行整批完整核对。",
        "nextStep": "重新运行抽取与核对，生成新的 batchGuard 和 reconciliationDiagnostics。",
        "issues": [
            {
                "code": "reocr_revalidation_required",
                "level": "warning",
                "title": "图片识别采纳结果待整批重算",
                "message": "候选结果不能仅凭异常数为 0 改写为机器通过。",
                "items": [],
            }
        ],
    }


def _build_reocr_apply_preflight(
    metadata: dict,
    candidates: list[dict],
    comparison_summary: dict,
    comparison_rows: list[dict],
    *,
    duplicate_scope_count: int = 0,
) -> dict:
    coverage = _build_reocr_plan_coverage(metadata, candidates)
    original_summary = metadata.get("comparisonSummary") if isinstance(metadata.get("comparisonSummary"), dict) else {}
    affected_employee_names = sorted(
        {
            str(row.get("employeeName") or row.get("pdfEmployeeName") or row.get("excelEmployeeName") or "").strip()
            for row in comparison_rows
            if str(row.get("employeeName") or row.get("pdfEmployeeName") or row.get("excelEmployeeName") or "").strip()
        }
    )
    post_apply_warnings = []
    missing_applied_count = len(coverage["missingAppliedTasks"])
    exception_count = _safe_int(comparison_summary.get("exceptionCount"))
    scope_count = len(
        {
            (str(candidate.get("sourceFile") or ""), str(candidate.get("warehouseId") or ""))
            for candidate in candidates
        }
    )
    if duplicate_scope_count:
        post_apply_warnings.append(f"候选中存在 {duplicate_scope_count} 个重复文件/仓库范围，不能批量采纳。")
    if missing_applied_count:
        post_apply_warnings.append(f"仍有 {missing_applied_count} 个图片识别复核任务未采纳，交付状态将保持阻断。")
    if exception_count:
        post_apply_warnings.append(f"投影结果仍有 {exception_count} 项异常，采纳后仍需人工复核。")
    if candidates:
        post_apply_warnings.append("采纳后必须重新执行整批完整核对，重新生成员工明细覆盖、金额闭合和诊断门禁。")
    if not candidates:
        post_apply_warnings.append("没有可采纳的图片识别结果。")
    preflight = {
        "willOverwriteOfficialResult": bool(candidates) and duplicate_scope_count == 0,
        "willRegenerateDiffReport": bool(candidates) and duplicate_scope_count == 0,
        "current": _pick_projection_summary(original_summary),
        "projected": _pick_projection_summary(comparison_summary),
        "delta": _summary_delta(original_summary, comparison_summary),
        "affectedScopeCount": scope_count,
        "affectedEmployeeCount": len(affected_employee_names),
        "affectedEmployees": affected_employee_names[:50],
        "coverageCompleteAfterApply": bool(coverage["coverageComplete"]),
        "blockingAfterApply": bool(candidates or duplicate_scope_count or missing_applied_count or exception_count),
        "postApplyWarnings": post_apply_warnings,
        "formalResultFields": ["comparisonSummary", "comparisonRows", "candidateMatches", "diffDownloadUrl"],
    }
    return preflight


def _build_reocr_plan_coverage(metadata: dict, selected_candidates: list[dict]) -> dict:
    planned = []
    for task in (metadata.get("reocrPlan", {}) or {}).get("tasks", []) or []:
        source_file = str(task.get("sourceFile") or "")
        warehouse_id = str(task.get("warehouseId") or "")
        if source_file:
            planned.append({"sourceFile": source_file, "warehouseId": warehouse_id})
    planned_keys = {(_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or "")) for item in planned}
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    active_candidates = governance.get("activeCandidates") or []
    confirmed_keys = {
        (_safe_scope_key(candidate.get("sourceFile", "")), str(candidate.get("warehouseId") or ""))
        for candidate in active_candidates
        if candidate.get("decision") in {"active", "applied"} or candidate.get("status") in {"active", "applied"}
    }
    applied_keys = {
        (_safe_scope_key(candidate.get("sourceFile", "")), str(candidate.get("warehouseId") or ""))
        for candidate in active_candidates
        if candidate.get("decision") == "applied" or candidate.get("status") == "applied"
    }
    selected_keys = {
        (_safe_scope_key(candidate.get("sourceFile", "")), str(candidate.get("warehouseId") or ""))
        for candidate in selected_candidates
    }
    projected_applied_keys = applied_keys | selected_keys
    missing_confirmed = [
        item
        for item in planned
        if (_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or "")) not in confirmed_keys
    ]
    missing_applied = [
        item
        for item in planned
        if (_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or "")) not in projected_applied_keys
    ]
    return {
        "plannedTaskCount": len(planned),
        "confirmedTaskCount": len(planned_keys & confirmed_keys) if planned_keys else len(confirmed_keys),
        "appliedTaskCount": len(planned_keys & projected_applied_keys) if planned_keys else len(projected_applied_keys),
        "missingConfirmedTasks": missing_confirmed,
        "missingAppliedTasks": missing_applied,
        "coverageComplete": not missing_applied if planned else bool(selected_candidates),
    }


def _with_labor_readiness(metadata: dict) -> dict:
    enriched = dict(metadata)
    enriched["readinessGate"] = _build_labor_readiness_gate(metadata)
    return enriched


def _build_labor_readiness_gate(metadata: dict) -> dict:
    summary = metadata.get("comparisonSummary") if isinstance(metadata.get("comparisonSummary"), dict) else {}
    quality = metadata.get("extractionQuality") if isinstance(metadata.get("extractionQuality"), dict) else {}
    diagnostics = metadata.get("reconciliationDiagnostics") if isinstance(metadata.get("reconciliationDiagnostics"), dict) else {}
    batch_guard = metadata.get("batchGuard") if isinstance(metadata.get("batchGuard"), dict) else {}
    issues: list[dict] = []

    status = str(metadata.get("status") or "")
    has_result = bool(summary) or status == "已生成差异报告" or bool(metadata.get("diffDownloadUrl"))
    if not has_result:
        issues.append(
            {
                "code": "no_official_result",
                "level": "blocked",
                "title": "尚未生成正式核对结果",
                "message": "需要先完成抽取、核对并生成差异报告，才能判断本批次是否具备业务确认条件。",
                "action": "点击抽取并核对，完成后下载正式报告复核。",
            }
        )

    if has_result:
        result_input_fingerprint = str(metadata.get("resultInputFingerprint") or "").strip()
        if not result_input_fingerprint:
            issues.append(
                {
                    "code": "missing_result_input_fingerprint",
                    "level": "blocked",
                    "title": "正式结果缺少输入版本绑定",
                    "message": "当前结果无法证明对应现有 PDF、Excel、字段映射和治理版本。",
                    "action": "使用当前版本重新执行完整核对。",
                }
            )
        elif result_input_fingerprint != _labor_result_input_fingerprint(metadata):
            issues.append(
                {
                    "code": "stale_result_inputs",
                    "level": "blocked",
                    "title": "正式结果已过期",
                    "message": "PDF、Excel、字段映射或已启用治理规则在结果生成后发生变化。",
                    "action": "不要沿用旧报告；重新执行整批完整核对。",
                }
            )

        if batch_guard.get("allowReleasableReport") is not True:
            issues.append(
                {
                    "code": "non_releasable_batch_guard",
                    "level": "blocked",
                    "title": "批次证据门禁未通过",
                    "message": str(batch_guard.get("message") or "批次缺少完整且可追溯的 PDF 员工明细门禁结果。"),
                    "action": "补齐逐文件员工明细和金额闭合证据后重新执行完整核对。",
                }
            )

        diagnostics_level = str(diagnostics.get("level") or "").strip().lower()
        if diagnostics_level == "critical":
            issues.append(
                {
                    "code": "critical_reconciliation_diagnostics",
                    "level": "blocked",
                    "title": "核对诊断存在冲突",
                    "message": str(diagnostics.get("message") or "核对诊断为 critical，不能进入业务确认。"),
                    "action": "处理总额、仓库、员工归因或来源冲突后重新核对。",
                }
            )
        elif diagnostics_level == "warning":
            issues.append(
                {
                    "code": "warning_reconciliation_diagnostics",
                    "level": "needs_review",
                    "title": "核对诊断仍有不稳定项",
                    "message": str(diagnostics.get("message") or "核对诊断为 warning，不能视为机器通过。"),
                    "action": "逐项复核诊断证据并重新执行完整核对。",
                }
            )
        elif diagnostics_level != "ok":
            issues.append(
                {
                    "code": "missing_reconciliation_diagnostics",
                    "level": "blocked",
                    "title": "缺少正式核对诊断",
                    "message": "当前结果没有可追溯的 reconciliationDiagnostics=ok 证据。",
                    "action": "使用当前版本重新执行完整核对。",
                }
            )

        machine_status = str(metadata.get("machineCheckStatus") or summary.get("machineCheckStatus") or "").strip().lower()
        if (
            summary.get("canRelease") is not True
            or str(summary.get("conclusionLevel") or "").strip().lower() != "pass"
            or machine_status != "passed"
        ):
            issues.append(
                {
                    "code": "machine_result_not_releasable",
                    "level": "blocked",
                    "title": "机器结果未达到业务确认前置条件",
                    "message": "正式结果必须同时满足 canRelease=true、conclusionLevel=pass 和 machineCheckStatus=passed。",
                    "action": "不要手工改写状态；修复证据后重新执行完整核对。",
                }
            )

    async_task = metadata.get("asyncTask") if isinstance(metadata.get("asyncTask"), dict) else {}
    async_status = str(async_task.get("status") or "").strip().lower()
    if status != "已生成差异报告" or async_status in {
        "queued",
        "running",
        "retry_wait",
        "failed",
        "error",
    }:
        issues.append(
            {
                "code": "run_not_finished",
                "level": "blocked",
                "title": "批次未完成",
                "message": f"当前批次状态为「{status}」，不能进入业务确认。",
                "action": "等待抽取完成，或修复失败原因后重新核对。",
            }
        )

    if not _labor_requires_employee_detail(metadata):
        issues.append(
            {
                "code": "formal_employee_detail_scope_required",
                "level": "blocked",
                "title": "正式核对范围不完整",
                "message": "total_only_diagnostic 只能用于诊断，不能进入正式业务确认。",
                "action": "切换为 employee_detail_required 并重新执行整批员工级核对。",
            }
        )

    exception_count = _safe_int(summary.get("exceptionCount"), 0) if summary else 0
    if exception_count > 0:
        issues.append(
            {
                "code": "comparison_exceptions",
                "level": "blocked",
                "title": "仍存在核对异常",
                "message": f"正式结果中还有 {exception_count} 项员工级或总额级异常。",
                "action": "先处理金额、工时、未匹配员工或仓库差异，再重新生成报告。",
            }
        )

    if str(quality.get("level") or "").lower() == "critical":
        quality_issues = quality.get("issues") if isinstance(quality.get("issues"), list) else []
        issues.append(
            {
                "code": "critical_extraction_quality",
                "level": "blocked",
                "title": "抽取质量为高风险",
                "message": "PDF/Excel 抽取质量诊断为 critical，不能作为机器通过结果。",
                "action": "优先处理低置信度抽取、OCR 缺失或供应商 Profile 问题。",
                "evidenceCount": len(quality_issues),
            }
        )
    elif str(quality.get("level") or "").lower() == "warning":
        quality_issues = quality.get("issues") if isinstance(quality.get("issues"), list) else []
        issues.append(
            {
                "code": "warning_extraction_quality",
                "level": "needs_review",
                "title": "抽取质量需要复核",
                "message": f"质量诊断中有 {len(quality_issues)} 条提示，需先复核证据再进入业务确认。",
                "action": "查看质量诊断与员工级证据，确认无漏抽或误抽。",
            }
        )

    reocr_coverage = _build_reocr_plan_coverage(metadata, [])
    confirmed_not_applied_count = _count_confirmed_reocr_not_applied(metadata)
    if reocr_coverage["plannedTaskCount"] and not reocr_coverage["coverageComplete"]:
        issues.append(
            {
                "code": "reocr_coverage_incomplete",
                "level": "blocked",
                "title": "图片识别复核未完全采纳",
                "message": (
                    f"计划 {reocr_coverage['plannedTaskCount']} 个图片识别复核任务，"
                    f"已确认 {reocr_coverage['confirmedTaskCount']} 个，已应用 {reocr_coverage['appliedTaskCount']} 个。"
                ),
                "action": "完成识别结果上传、影响预览、人工确认和采纳，或明确回滚不采纳。",
                "missingConfirmedTasks": reocr_coverage["missingConfirmedTasks"],
                "missingAppliedTasks": reocr_coverage["missingAppliedTasks"],
            }
        )
    elif confirmed_not_applied_count:
        issues.append(
            {
                "code": "confirmed_reocr_not_applied",
                "level": "blocked",
                "title": "图片识别结果已确认但未正式采纳",
                "message": f"还有 {confirmed_not_applied_count} 个图片识别结果只处于已确认状态，正式核对结果和差异报告尚未更新。",
                "action": "请批量预览并采纳，或回滚不适用的识别结果；未采纳前不能进入业务确认。",
            }
        )

    governance_pending = _count_labor_pending_governance(metadata)
    if governance_pending:
        issues.append(
            {
                "code": "pending_governance_candidates",
                "level": "needs_review",
                "title": "仍有治理候选未闭环",
                "message": f"还有 {governance_pending} 个规则、姓名映射、Profile、修正、图片识别或跨仓库归属建议未确认/未应用/未回滚。",
                "action": "逐项回放并确认、应用或回滚，确保候选建议不会静默影响正式结论。",
            }
        )

    files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
    diff_report = files.get("diffReport") if isinstance(files.get("diffReport"), dict) else {}
    diff_download_url = str(metadata.get("diffDownloadUrl") or "")
    comparison_rows = metadata.get("comparisonRows") if isinstance(metadata.get("comparisonRows"), list) else []
    pdf_employee_count = _safe_int(summary.get("pdfEmployeeCount"))
    excel_employee_count = _safe_int(summary.get("excelEmployeeCount"))
    if has_result and not comparison_rows:
        issues.append(
            {
                "code": "missing_employee_detail_result",
                "level": "blocked",
                "title": "员工级核对明细缺失",
                "message": "正式结果没有可追溯的员工级 comparisonRows。",
                "action": "重新执行整批员工级核对并生成完整明细。",
            }
        )
    if has_result and (
        pdf_employee_count <= 0
        or excel_employee_count <= 0
        or pdf_employee_count != excel_employee_count
        or len(comparison_rows) != pdf_employee_count
    ):
        issues.append(
            {
                "code": "employee_detail_counts_not_closed",
                "level": "blocked",
                "title": "员工级数量证据未闭合",
                "message": "正式结果必须包含正数且一致的 PDF/Excel 员工数，并与 comparisonRows 逐人覆盖数一致。",
                "action": "检查漏抽、重复员工和账单映射后重新执行整批核对。",
            }
        )
    if has_result and any(str(row.get("matchStatus") or "").strip() != "通过" for row in comparison_rows):
        issues.append(
            {
                "code": "employee_detail_rows_not_clean",
                "level": "blocked",
                "title": "员工级明细仍有异常",
                "message": "机器通过结果中的每一条 comparisonRows 都必须明确标记为通过。",
                "action": "处理未匹配、金额、工时或低置信度异常后重新核对。",
            }
        )
    if has_result and (not diff_download_url or not diff_report):
        issues.append(
            {
                "code": "missing_official_report",
                "level": "blocked",
                "title": "正式报告缺失",
                "message": "当前机器结果没有完整的差异报告文件记录和下载链接。",
                "action": "重新生成正式差异报告并确认文件可下载。",
            }
        )
    elif has_result and diff_report.get("downloadUrl") != diff_download_url:
        issues.append(
            {
                "code": "report_url_mismatch",
                "level": "needs_review",
                "title": "报告链接与正式文件记录不一致",
                "message": "diffDownloadUrl 与 files.diffReport.downloadUrl 不一致。",
                "action": "重新生成或重新挂载正式差异报告后再交付。",
            }
        )
    if has_result and diff_download_url and diff_report and not _labor_diff_report_file_exists(metadata, diff_report):
        issues.append(
            {
                "code": "report_file_missing",
                "level": "blocked",
                "title": "正式报告文件不存在",
                "message": "正式差异报告记录存在，但服务器上的文件不存在或已被清理。",
                "action": "请重新生成报告，或从持久化存储恢复报告文件后再交付。",
            }
        )

    blocked_count = sum(1 for issue in issues if issue["level"] == "blocked")
    review_count = sum(1 for issue in issues if issue["level"] == "needs_review")
    status_code = "blocked" if blocked_count else "needs_review" if review_count else "ready"
    labels = {
        "ready": "可进入业务确认",
        "needs_review": "需处理后复核",
        "blocked": "机器结果阻断",
    }
    return {
        "status": status_code,
        "label": labels[status_code],
        "ready": status_code == "ready",
        "machineReady": status_code == "ready",
        "businessReviewRequired": True,
        "directPaymentAllowed": False,
        "summary": {
            "issueCount": len(issues),
            "blockedCount": blocked_count,
            "reviewCount": review_count,
            "exceptionCount": exception_count,
            "reocrPlannedTaskCount": reocr_coverage["plannedTaskCount"],
            "reocrConfirmedTaskCount": reocr_coverage["confirmedTaskCount"],
            "reocrAppliedTaskCount": reocr_coverage["appliedTaskCount"],
            "confirmedReocrNotAppliedCount": confirmed_not_applied_count,
            "pendingGovernanceCount": governance_pending,
        },
        "issues": issues,
        "reocrCoverage": reocr_coverage,
    }


def _labor_diff_report_file_exists(metadata: dict, diff_report: dict) -> bool:
    if labor_postgres_state_enabled() and str(diff_report.get("objectKey") or "").strip():
        try:
            expected_size = int(diff_report.get("sizeBytes") or 0)
            verified_size = int(diff_report.get("storageVerifiedSizeBytes") or 0)
        except (TypeError, ValueError):
            return False
        return bool(
            diff_report.get("storageVerified") is True
            and expected_size > 0
            and verified_size == expected_size
            and re.fullmatch(r"[0-9a-f]{64}", str(diff_report.get("sha256") or "").strip().lower())
        )
    explicit_path = Path(str(diff_report.get("path") or ""))
    if str(diff_report.get("path") or ""):
        path = explicit_path
    else:
        filename = str(diff_report.get("filename") or "")
        if not filename:
            return False
        run_id = str(metadata.get("runId") or metadata.get("id") or "")
        if not run_id:
            return False
        run_dir = get_labor_run_dir(run_id)
        path = _resolve_labor_download_path(run_dir, filename)
    if not path.exists() or not path.is_file():
        return False
    expected_hash = str(diff_report.get("sha256") or "").strip().lower()
    try:
        expected_size = int(diff_report.get("sizeBytes"))
    except (TypeError, ValueError):
        return False
    if (
        expected_size <= 0
        or not re.fullmatch(r"[0-9a-f]{64}", expected_hash)
        or path.stat().st_size != expected_size
    ):
        return False
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest() == expected_hash


def _count_labor_pending_governance(metadata: dict) -> int:
    def pending(records: list[dict], terminal_statuses: set[str]) -> int:
        total = 0
        for record in records:
            status = str(record.get("status") or "").lower()
            decision = str(record.get("decision") or "").lower()
            if status not in terminal_statuses and decision not in terminal_statuses:
                total += 1
        return total

    rule_governance = _normalized_labor_governance(metadata.get("ruleGovernance"))
    name_governance = _normalized_name_mapping_governance(metadata.get("nameMappingGovernance"))
    profile_governance = _normalized_profile_governance(metadata.get("profileGovernance"))
    correction_governance = _normalized_correction_governance(metadata.get("correctionGovernance"))
    reocr_governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    allocation_governance = _normalized_allocation_governance(metadata.get("allocationGovernance"))
    return (
        pending(rule_governance["candidates"], {"confirmed", "rolled_back"})
        + pending(name_governance["candidates"], {"confirmed", "rolled_back"})
        + pending(profile_governance["candidates"], {"confirmed", "rolled_back"})
        + pending(correction_governance["candidates"], {"confirmed", "rolled_back"})
        + pending(reocr_governance["activeCandidates"], {"applied", "rolled_back"})
        + pending(allocation_governance["candidates"], {"confirmed", "rolled_back"})
    )


def _count_confirmed_reocr_not_applied(metadata: dict) -> int:
    governance = _normalized_reocr_governance(metadata.get("reocrReplayGovernance"))
    total = 0
    for candidate in governance["activeCandidates"]:
        status = str(candidate.get("status") or "").lower()
        decision = str(candidate.get("decision") or "").lower()
        if status in {"applied", "rolled_back"} or decision in {"applied", "rolled_back"}:
            continue
        total += 1
    return total


def _safe_int(value, default: int = 0) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return default


def _safe_scope_key(value: str) -> str:
    return Path(str(value or "")).name.strip().lower()


def _summarize_reocr_batch_comparison(summaries: list[dict], rows: list[dict]) -> dict:
    sum_keys = [
        "pdfEmployeeCount",
        "excelEmployeeCount",
        "pdfHoursTotal",
        "excelHoursTotal",
        "pdfAmountTotal",
        "excelAmountTotal",
        "amountDiffCount",
        "hoursRiskCount",
        "unmatchedPdfCount",
        "unmatchedExcelCount",
        "lowConfidenceCount",
        "fuzzyMatchCount",
        "candidateMatchCount",
        "exceptionCount",
    ]
    summary = {}
    for key in sum_keys:
        summary[key] = round(sum(float(item.get(key) or 0) for item in summaries), 2)
    summary["amountDeltaTotal"] = round(float(summary.get("pdfAmountTotal") or 0) - float(summary.get("excelAmountTotal") or 0), 2)
    summary["hoursDeltaTotal"] = round(float(summary.get("pdfHoursTotal") or 0) - float(summary.get("excelHoursTotal") or 0), 2)
    max_amount = max(abs(float(summary.get("pdfAmountTotal") or 0)), abs(float(summary.get("excelAmountTotal") or 0)), 1.0)
    max_hours = max(abs(float(summary.get("pdfHoursTotal") or 0)), abs(float(summary.get("excelHoursTotal") or 0)), 1.0)
    summary["amountDeltaPercentage"] = round(abs(summary["amountDeltaTotal"]) / max_amount * 100, 2)
    summary["hoursDeltaPercentage"] = round(abs(summary["hoursDeltaTotal"]) / max_hours * 100, 2)
    passed_count = sum(1 for row in rows if row.get("matchStatus") == "通过")
    summary["matchRate"] = round(passed_count / len(rows) * 100, 1) if rows else 0.0
    summary["averageConfidence"] = round(
        sum(float(item.get("averageConfidence") or 0) for item in summaries) / len(summaries),
        3,
    ) if summaries else 0.0
    return summary


def _pick_projection_summary(summary: dict) -> dict:
    keys = [
        "pdfAmountTotal",
        "excelAmountTotal",
        "amountDeltaTotal",
        "pdfHoursTotal",
        "excelHoursTotal",
        "hoursDeltaTotal",
        "exceptionCount",
        "lowConfidenceCount",
        "amountDiffCount",
        "hoursRiskCount",
        "matchRate",
    ]
    return {key: summary.get(key) for key in keys if key in summary}


def _summary_delta(original: dict, projected: dict) -> dict:
    delta = {}
    for key in set(_pick_projection_summary(original)) | set(_pick_projection_summary(projected)):
        original_value = original.get(key)
        projected_value = projected.get(key)
        if isinstance(original_value, (int, float)) and isinstance(projected_value, (int, float)):
            delta[key] = round(projected_value - original_value, 4)
    return delta


def _row_matches_any_affected_name(row: dict, affected_names: list[str]) -> bool:
    row_text = _compact_match_text(
        " ".join(str(row.get(key) or "") for key in ("employeeName", "employeeKey"))
    )
    return any(name and (name in row_text or row_text in name) for name in affected_names)


def _active_supplier_profile_override(supplier: str, profile_governance: dict) -> SupplierExtractionProfile | None:
    normalized_supplier = _normalize_supplier_for_profile(supplier)
    candidates: list[tuple[tuple[int, int], bool, SupplierExtractionProfile]] = []
    for record in profile_governance.get("activeProfiles") or []:
        if record.get("decision") != "active" or record.get("status") != "active":
            continue
        profile_data = record.get("profileData") if isinstance(record.get("profileData"), dict) else {}
        aliases = [str(alias) for alias in profile_data.get("aliases", []) if str(alias).strip()]
        record_supplier = str(record.get("supplier") or "").strip()
        scope_values = [record_supplier, *aliases]
        scores = [
            score
            for value in scope_values
            if value and (score := supplier_alias_match_score(normalized_supplier, value)) is not None
        ]
        if not normalized_supplier or not scores:
            continue
        profile = _supplier_profile_from_governance_data(
            profile_data,
            fallback_key=str(record.get("profileKey") or "active-profile"),
            fallback_aliases=aliases or ([supplier.lower().strip()] if supplier else []),
        )
        if profile is not None and is_runtime_approved_profile(profile):
            exact_scope = any(
                _normalize_supplier_for_profile(value) == normalized_supplier
                for value in scope_values
                if value
            )
            candidates.append((max(scores), exact_scope, profile))
    if not candidates:
        return None

    best_score = max(score for score, _exact, _profile in candidates)
    best_candidates = [item for item in candidates if item[0] == best_score]
    profile_keys = {profile.key for _score, _exact, profile in best_candidates}
    if len(profile_keys) != 1:
        logger.warning(
            "运行内 Supplier Profile 匹配冲突，已拒绝覆盖: supplier=%s, profiles=%s",
            supplier,
            sorted(profile_keys),
        )
        return None
    selected_score, exact_scope, selected = max(
        best_candidates,
        key=lambda item: item[2].version,
    )
    baseline = resolve_supplier_profile(
        supplier,
        profiles_path=AI_CONFIG.get("supplier_profiles_path"),
    )
    baseline_scores = [
        score
        for alias in baseline.aliases
        if (score := supplier_alias_match_score(normalized_supplier, alias)) is not None
    ]
    if baseline.key != "default" and baseline.key != selected.key and baseline_scores:
        baseline_score = max(baseline_scores)
        if selected_score < baseline_score or (selected_score == baseline_score and not exact_scope):
            logger.warning(
                "运行内 Supplier Profile 覆盖范围不够具体，已保留既有 Profile: supplier=%s, active=%s, baseline=%s",
                supplier,
                selected.key,
                baseline.key,
            )
            return None
    return selected


def _supplier_profile_from_governance_data(
    profile_data: dict,
    *,
    fallback_key: str,
    fallback_aliases: list[str],
) -> SupplierExtractionProfile | None:
    try:
        version = int(profile_data.get("version") or 0)
        failure_count = int(profile_data.get("failure_count") or 0)
    except (TypeError, ValueError):
        return None
    aliases = [str(alias) for alias in profile_data.get("aliases", []) if str(alias).strip()]
    return SupplierExtractionProfile(
        key=str(profile_data.get("key") or fallback_key),
        aliases=aliases or [alias for alias in fallback_aliases if alias],
        prompt_notes=[str(note) for note in profile_data.get("prompt_notes", []) if str(note).strip()],
        authoritative_total_methods=[
            str(method)
            for method in profile_data.get("authoritative_total_methods", [])
            if str(method).strip()
        ],
        line_item_aliases={
            str(label).strip().lower(): str(item_type).strip()
            for label, item_type in (profile_data.get("line_item_aliases") or {}).items()
            if str(label).strip() and str(item_type).strip()
        },
        image_page_policy=str(profile_data.get("image_page_policy") or "all"),
        version=version,
        failure_count=failure_count,
        deprecated=bool(profile_data.get("deprecated", False)),
        status=str(profile_data.get("status") or "draft").strip().lower(),
        approved_by=str(profile_data.get("approvedBy") or profile_data.get("approved_by") or "").strip(),
        approved_at=str(profile_data.get("approvedAt") or profile_data.get("approved_at") or "").strip(),
        created_from=str(profile_data.get("created_from") or "").strip(),
    )


def _normalize_supplier_for_profile(value: str) -> str:
    text = re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())
    return re.sub(r"\s+", " ", text).strip()


def _labor_progress_completed(message: str = "核对报告已生成。") -> dict:
    now = datetime.now().isoformat(timespec="seconds")
    return {
        "status": "completed",
        "phase": "completed",
        "phaseLabel": "完成",
        "message": message,
        "lastUpdatedAt": now,
        "completedAt": now,
    }


def _labor_task_generation_id(metadata: dict) -> str:
    async_task = metadata.get("asyncTask") if isinstance(metadata.get("asyncTask"), dict) else {}
    return str(metadata.get("taskGenerationId") or async_task.get("taskGenerationId") or "").strip()


def _update_labor_task_metadata(
    run_id: str,
    task_generation_id: str,
    updates: dict | Callable[[dict], dict],
) -> tuple[dict, bool]:
    generation = str(task_generation_id or "").strip()
    if generation:
        return update_labor_metadata_for_task(
            run_id,
            expected_task_generation_id=generation,
            updates=updates,
        )
    current = _labor_metadata_or_404(run_id)
    payload = updates(current) if callable(updates) else updates
    return update_labor_metadata(run_id, payload), True


def _update_labor_progress(
    run_id: str,
    *,
    phase: str,
    phase_label: str,
    message: str = "",
    status: str = "running",
    total_files: int | None = None,
    processed_files: int | None = None,
    total_pages: int | None = None,
    processed_pages: int | None = None,
    current_file: str = "",
    current_page: int | None = None,
    cache_hit_count: int | None = None,
    retry_delay_seconds: int | None = None,
    retry_attempt: int | None = None,
    task_generation_id: str = "",
) -> bool:
    """Persist reader-facing labor progress for frontend polling.

    This updates the existing metadata file only; it does not create a second
    job store or affect non-labor modules.
    """
    try:
        def build_updates(metadata: dict) -> dict:
            now = datetime.now().isoformat(timespec="seconds")
            previous_progress = metadata.get("progress") if isinstance(metadata.get("progress"), dict) else {}
            previous_task = metadata.get("asyncTask") if isinstance(metadata.get("asyncTask"), dict) else {}
            progress = dict(previous_progress)
            progress.update(
                {
                    "status": status,
                    "phase": phase,
                    "phaseLabel": phase_label,
                    "message": message or phase_label,
                    "lastUpdatedAt": now,
                    "startedAt": progress.get("startedAt") or previous_task.get("startedAt") or now,
                }
            )
            for key, value in {
                "totalFiles": total_files,
                "processedFiles": processed_files,
                "totalPages": total_pages,
                "processedPages": processed_pages,
                "currentFile": current_file,
                "currentPage": current_page,
                "cacheHitCount": cache_hit_count,
                "retryDelaySeconds": retry_delay_seconds,
                "retryAttempt": retry_attempt,
            }.items():
                if value not in (None, ""):
                    progress[key] = value

            task_status = "failed" if status == "failed" else "completed" if status == "completed" else "running"
            task_label = "失败" if status == "failed" else "完成" if status == "completed" else "处理中"
            async_task = dict(previous_task)
            async_task.update(
                {
                    "status": task_status,
                    "statusLabel": task_label,
                    "message": message or phase_label,
                    "startedAt": async_task.get("startedAt") or progress.get("startedAt") or now,
                }
            )
            return {
                "stage": phase_label,
                "progress": progress,
                "asyncTask": async_task,
            }

        _metadata, committed = _update_labor_task_metadata(
            run_id,
            task_generation_id,
            build_updates,
        )
        return committed
    except Exception as exc:  # noqa: BLE001 - progress must not break the audit job.
        logger.warning(f"[{run_id}] 更新劳务核对进度失败: {exc}")
        return False


def _labor_ai_progress_callback(
    run_id: str,
    *,
    total_files: int,
    task_generation_id: str = "",
) -> Callable[[dict], None]:
    def _handle(event: dict) -> None:
        event_name = str(event.get("event") or "")
        total_pages = _coerce_int(event.get("total_pages"))
        processed_pages = _coerce_int(event.get("processed_pages"))
        current_file = str(event.get("current_file") or "")
        current_page = _coerce_int(event.get("current_page"))
        base_payload = {
            "phase": "pdf_detail",
            "phase_label": "识别 PDF 发票明细",
            "total_files": total_files,
            "total_pages": total_pages,
            "processed_pages": processed_pages,
            "current_file": current_file,
            "current_page": current_page,
        }
        if task_generation_id:
            base_payload["task_generation_id"] = task_generation_id
        if event_name == "pdf_pages_loaded":
            _update_labor_progress(
                run_id,
                **base_payload,
                message=f"已读取 {total_files} 张 PDF 发票，共 {total_pages or 0} 页。",
            )
            return
        if event_name == "rendering_images":
            _update_labor_progress(
                run_id,
                **base_payload,
                message="正在准备图片型发票识别。",
            )
            return
        if event_name == "ai_image_start":
            _update_labor_progress(
                run_id,
                **base_payload,
                message=f"开始识别 PDF 发票明细，共 {total_pages or 0} 页。",
            )
            return
        if event_name == "ai_image_page_started":
            detail = _labor_page_progress_text(processed_pages, total_pages, current_file, current_page, active=True)
            _update_labor_progress(run_id, **base_payload, message=detail)
            return
        if event_name == "ai_image_page_retrying":
            delay = _coerce_int(event.get("retry_delay_seconds"))
            attempt = _coerce_int(event.get("retry_attempt"))
            retry_text = f"AI 识别服务响应较慢，正在重试当前页（第 {attempt or 1} 次）。"
            _update_labor_progress(
                run_id,
                **base_payload,
                message=retry_text,
                retry_delay_seconds=delay,
                retry_attempt=attempt,
            )
            return
        if event_name == "ai_image_high_res_retrying":
            detail = _labor_page_progress_text(processed_pages, total_pages, current_file, current_page, active=True)
            _update_labor_progress(run_id, **base_payload, message=f"{detail} 正在用更清晰的图片重新识别。")
            return
        if event_name in {"ai_image_page_completed", "ai_image_page_skipped"}:
            detail = _labor_page_progress_text(processed_pages, total_pages, current_file, current_page, active=False)
            if event_name == "ai_image_page_skipped":
                detail = f"{detail} 当前页识别失败，已继续处理下一页。"
            _update_labor_progress(run_id, **base_payload, message=detail)

    return _handle


def _labor_ocr_progress_callback(run_id: str, *, task_generation_id: str = "") -> Callable[[dict], None]:
    def _handle(snapshot: dict) -> None:
        progress_payload = {
            "phase": "auto_ocr",
            "phase_label": "本地 OCR 识别",
            "message": str(snapshot.get("message") or "正在调用本地 Worker 识别 PDF 发票。"),
            "status": "running",
            "total_files": _coerce_int(snapshot.get("totalFiles")),
            "processed_files": _coerce_int(snapshot.get("processedFiles")),
            "total_pages": _coerce_int(snapshot.get("totalPages")),
            "processed_pages": _coerce_int(snapshot.get("processedPages")),
            "current_file": str(snapshot.get("currentFile") or ""),
            "cache_hit_count": _coerce_int(snapshot.get("cacheHitCount")),
        }
        if task_generation_id:
            progress_payload["task_generation_id"] = task_generation_id
        _update_labor_progress(
            run_id,
            **progress_payload,
        )

    return _handle


def _summarize_pdf_page_audit(audit_rows: list[dict]) -> dict:
    total_pages = len(audit_rows)
    zero_row_pages = 0
    failed_pages = 0
    high_res_pages = 0
    extracted_rows = 0
    amount_total = 0.0
    for row in audit_rows:
        status = str(row.get("status") or "")
        row_count = _coerce_int(row.get("rowCount")) or 0
        extracted_rows += row_count
        try:
            amount_total += float(row.get("amountTotal") or 0)
        except (TypeError, ValueError):
            pass
        if row_count == 0 and status in {"completed", "cache_hit", "high_res_retry_no_rows"}:
            zero_row_pages += 1
        if "failed" in status:
            failed_pages += 1
        if row.get("highResolutionRetry") or status.startswith("high_res_retry"):
            high_res_pages += 1
    return {
        "pageCount": total_pages,
        "zeroRowPageCount": zero_row_pages,
        "failedPageCount": failed_pages,
        "highResolutionRetryPageCount": high_res_pages,
        "extractedRowCount": extracted_rows,
        "amountTotal": round(amount_total, 2),
    }


def _labor_page_progress_text(
    processed_pages: int | None,
    total_pages: int | None,
    current_file: str,
    current_page: int | None,
    *,
    active: bool,
) -> str:
    done = int(processed_pages or 0)
    total = int(total_pages or 0)
    prefix = "正在识别 PDF 发票" if active else "已识别 PDF 发票"
    progress = f"{done} / {total} 页" if total else f"{done} 页"
    location = ""
    if current_file and current_page:
        location = f"，当前文件：{current_file}，第 {current_page} 页"
    elif current_file:
        location = f"，当前文件：{current_file}"
    return f"{prefix}：{progress}{location}。"


def _coerce_int(value) -> int | None:
    try:
        if value is None or value == "":
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _run_labor_extract_compare(
    run_id: str,
    task_generation_id: str = "",
    reservation_token: str = "",
) -> bool:
    generation = str(task_generation_id or "").strip()
    try:
        logger.info(f"[{run_id}] === 抽取任务启动 ===")
        run_dir = get_labor_run_dir(run_id)
        if labor_persistent_storage_enabled():
            sync_labor_run_from_persistent(run_id, run_dir)
        current = _labor_metadata_or_404(run_id)
        current_async_task = current.get("asyncTask") if isinstance(current.get("asyncTask"), dict) else {}
        if not generation and current.get("status") == "抽取中" and current_async_task.get("status") in {"queued", "running"}:
            generation = _labor_task_generation_id(current)
        _running, current_task = _update_labor_task_metadata(
            run_id,
            generation,
            {
                "status": "抽取中",
                "stage": "后台处理中",
                "asyncTask": {
                    **current_async_task,
                    "status": "running",
                    "statusLabel": "处理中",
                    "message": "后台正在读取已上传文件并生成核对结果。",
                    "startedAt": utcnow_naive().isoformat(),
                    **({"taskGenerationId": generation} if generation else {}),
                },
            },
        )
        if generation and not current_task:
            logger.info("[%s] 任务代次已失效，跳过旧任务启动", run_id)
            return False
        _update_labor_progress(
            run_id,
            phase="queued",
            phase_label="准备核对材料",
            message="后台已接收任务，正在准备读取材料。",
            task_generation_id=generation,
        )
        performed = (
            _perform_labor_extract_compare(run_id, task_generation_id=generation)
            if generation
            else _perform_labor_extract_compare(run_id)
        )
        if isinstance(performed, dict) and performed.get("_taskGenerationRejected") is True:
            logger.info("[%s] 旧任务完成结果已忽略，当前批次属于较新任务", run_id)
            return False
        if isinstance(performed, dict) and performed.get("inputSnapshotRejected") is True:
            raise ValueError("核对期间输入发生变化，本次旧快照结果未发布；请使用当前材料重新执行整批核对。")
        completed_metadata, completed_current_task = _update_labor_task_metadata(
            run_id,
            generation,
            {
                "progress": _labor_progress_completed("核对报告已生成。"),
                "asyncTask": {
                    **(
                        performed.get("asyncTask")
                        if isinstance(performed, dict) and isinstance(performed.get("asyncTask"), dict)
                        else {}
                    ),
                    "status": "completed",
                    "statusLabel": "完成",
                    "message": "核对结果已生成。",
                    "completedAt": utcnow_naive().isoformat(),
                    **({"taskGenerationId": generation} if generation else {}),
                },
            },
        )
        if generation and not completed_current_task:
            logger.info("[%s] 旧任务完成状态已忽略，当前批次属于较新任务", run_id)
            return False
        completed_owner = str(completed_metadata.get("ownerUserId") or "local-default")
        append_labor_audit_event(
            _labor_audit_path(),
            action="extraction_completed",
            run_id=run_id,
            owner_user_id=completed_owner,
            actor_user_id=completed_owner,
            outcome="success",
        )
        logger.info(f"[{run_id}] === 抽取任务完成 ===")
        return True
    except ValueError as exc:
        logger.error(f"[{run_id}] 抽取失败(ValueError): {exc}")
        _update_labor_progress(
            run_id,
            phase="failed",
            phase_label="核对失败",
            message=str(exc),
            status="failed",
            task_generation_id=generation,
        )
        _failed_metadata, failed_current_task = _update_labor_task_metadata(
            run_id,
            generation,
            {
                "status": "抽取失败",
                "stage": "错误",
                "errorMessage": str(exc),
                "asyncTask": {
                    "status": "failed",
                    "statusLabel": "失败",
                    "message": str(exc),
                    "failedAt": utcnow_naive().isoformat(),
                    **({"taskGenerationId": generation} if generation else {}),
                },
            },
        )
        if generation and not failed_current_task:
            logger.info("[%s] 旧任务失败状态已忽略，当前批次属于较新任务", run_id)
            return False
        append_labor_audit_event(
            _labor_audit_path(),
            action="extraction_failed",
            run_id=run_id,
            outcome="failed",
            reason_code="LABOR_EXTRACT_VALUE_ERROR",
        )
        return False
    except Exception as exc:
        logger.error(f"[{run_id}] 抽取失败(Exception): {exc}", exc_info=True)
        message = f"生成劳务核对结果失败：{exc}"
        _update_labor_progress(
            run_id,
            phase="failed",
            phase_label="核对失败",
            message=message,
            status="failed",
            task_generation_id=generation,
        )
        _failed_metadata, failed_current_task = _update_labor_task_metadata(
            run_id,
            generation,
            {
                "status": "抽取失败",
                "stage": "错误",
                "errorMessage": message,
                "asyncTask": {
                    "status": "failed",
                    "statusLabel": "失败",
                    "message": message,
                    "failedAt": utcnow_naive().isoformat(),
                    **({"taskGenerationId": generation} if generation else {}),
                },
            },
        )
        if generation and not failed_current_task:
            logger.info("[%s] 旧任务异常状态已忽略，当前批次属于较新任务", run_id)
            return False
        append_labor_audit_event(
            _labor_audit_path(),
            action="extraction_failed",
            run_id=run_id,
            outcome="failed",
            reason_code="LABOR_EXTRACT_EXCEPTION",
        )
        return False
    finally:
        if reservation_token:
            _LABOR_TASK_LIMITER.release(run_id, reservation_token)


def _run_labor_extract_compare_with_release(
    run_id: str,
    task_generation_id: str = "",
    reservation_token: str = "",
) -> None:
    # Keep the endpoint compatible with test/extension runners that still expose
    # the historical ``runner(run_id)`` callable while production uses the
    # generation-aware signature. The wrapper owns a final idempotent release so
    # a substituted runner cannot leak the reservation.
    import inspect

    try:
        runner = _run_labor_extract_compare
        parameters = inspect.signature(runner).parameters.values()
        accepts_generation = any(
            parameter.name == "task_generation_id"
            or parameter.kind in {inspect.Parameter.VAR_POSITIONAL, inspect.Parameter.VAR_KEYWORD}
            for parameter in parameters
        )
        if accepts_generation:
            runner(run_id, task_generation_id, reservation_token)
        else:
            runner(run_id)
    finally:
        if reservation_token:
            _LABOR_TASK_LIMITER.release(run_id, reservation_token)


def _aggregate_excel_rows(excel_rows: list) -> list:
    """按员工和仓库聚合 Excel 行（合并同一员工同仓库的多天记录）。

    同一员工可能跨多个仓库有账单行。员工级核对可以汇总，但仓库级核对
    必须保留仓库归属，因此这里不能跨 warehouse_id 合并。
    """
    from collections import defaultdict
    from .engine.labor.models import LaborLineItem

    groups = defaultdict(list)
    for row in excel_rows:
        employee_key = (row.employee_id or row.employee_name_raw or "").strip().lower()
        warehouse_key = str(row.warehouse_id or "").strip()
        key = (employee_key, warehouse_key)
        groups[key].append(row)

    if all(len(v) == 1 for v in groups.values()):
        return excel_rows  # 无需聚合

    aggregated = []
    for key, rows in groups.items():
        if len(rows) == 1:
            aggregated.append(rows[0])
        else:
            # 合并：取第一条的元数据，hours/amount 求和
            base = rows[0]
            total_hours = sum(r.hours for r in rows)
            total_amount = sum(r.amount for r in rows)
            merged = LaborLineItem(
                employee_name_raw=base.employee_name_raw,
                employee_id=base.employee_id,
                hours=round(total_hours, 2),
                amount=round(total_amount, 2),
                currency=base.currency,
                source_file=base.source_file,
                source_page_or_row=base.source_page_or_row,
                source_type=base.source_type,
                warehouse_id=base.warehouse_id,
                supplier=base.supplier,
                period_start=base.period_start,
                period_end=base.period_end,
                confidence=base.confidence,
                evidence_text=base.evidence_text,
            )
            aggregated.append(merged)

    logger.info(f"Excel 行聚合: {len(excel_rows)} 行 → {len(aggregated)} 行 ({len(excel_rows) - len(aggregated)} 条合并)")
    return aggregated


def _labor_apply_workbook_warehouse_fallback(
    rows: list[LaborLineItem],
    source_file: str,
    batch_warehouse_id: str = "",
) -> list[LaborLineItem]:
    """Fill only blank warehouse ids from explicit batch/filename evidence."""
    fallback = str(batch_warehouse_id or "").strip() or _warehouse_id_from_filename(source_file)
    if not fallback:
        return list(rows)
    return [
        row if str(row.warehouse_id or "").strip() else replace(row, warehouse_id=fallback)
        for row in rows
    ]


def _labor_cost_summaries(workbook_paths: list[Path]) -> list[dict]:
    summaries: list[dict] = []
    for workbook_path in workbook_paths:
        try:
            summaries.append(summarize_otws_costs(workbook_path))
        except ValueError:
            continue
        except Exception as exc:
            logger.warning(f"解析 OTWS 费用摘要失败: {workbook_path.name}: {exc}")
    return summaries


def _labor_has_complete_warehouse_cost_evidence(
    pdf_totals: list[dict],
    cost_summaries: list[dict],
) -> bool:
    """Return whether OTWS-style summaries cover every authoritative invoice warehouse."""
    invoice_warehouses = {
        str(item.get("warehouse_id") or "").strip()
        for item in pdf_totals
        if item.get("authoritative") is not False
        and float(item.get("total_amount") or 0) > 0
        and str(item.get("warehouse_id") or "").strip()
    }
    summary_warehouses = {
        str(item.get("warehouseId") or "").strip()
        for item in cost_summaries
        if str(item.get("warehouseId") or "").strip()
        and float((item.get("summary") or {}).get("reportedTotal") or 0) > 0
    }
    return bool(invoice_warehouses) and invoice_warehouses.issubset(summary_warehouses)


def _labor_apply_cost_summary_amount_basis(
    warehouse_comparison: dict,
    cost_summaries: list[dict],
    *,
    amount_tolerance: float,
) -> dict:
    """Use the OTWS reported warehouse total instead of one mapped detail sheet."""
    if not cost_summaries:
        return warehouse_comparison

    summaries_by_warehouse = {
        str(item.get("warehouseId") or "").strip(): item
        for item in cost_summaries
        if str(item.get("warehouseId") or "").strip()
        and isinstance(item.get("summary"), dict)
        and item["summary"].get("reportedTotal") is not None
    }
    if not summaries_by_warehouse:
        return warehouse_comparison

    normalized = dict(warehouse_comparison)
    rows: list[dict] = []
    for original in warehouse_comparison.get("rows", []) or []:
        row = dict(original)
        warehouse_id = str(row.get("warehouseId") or "").strip()
        cost_summary = summaries_by_warehouse.get(warehouse_id)
        if not cost_summary:
            rows.append(row)
            continue

        summary = cost_summary.get("summary") or {}
        details = cost_summary.get("details") or {}
        excel_amount = round(float(summary.get("reportedTotal") or 0), 2)
        excel_hours = round(
            sum(
                float((details.get(key) or {}).get("hours") or 0)
                for key in ("employeeExpenses", "employeeBenefits", "loadingAndUnloading")
            ),
            2,
        )
        amount_delta = round(float(row.get("pdfAmountTotal") or 0) - excel_amount, 2)
        row.update(
            {
                "excelEmployeeCount": int(cost_summary.get("employeeCount") or row.get("excelEmployeeCount") or 0),
                "excelHoursTotal": excel_hours,
                "excelAmountTotal": excel_amount,
                "amountDelta": amount_delta,
                "excelEvidenceFile": str(cost_summary.get("sourceFile") or ""),
                "excelEvidenceRef": str(summary.get("evidence") or ""),
            }
        )
        if row.get("reconciliationStatus") not in {"needs_review", "missing_pdf_invoice"}:
            reconciliation_status = (
                "passed"
                if amount_within_tolerance(amount_delta, amount_tolerance)
                else "amount_difference"
            )
            row["reconciliationStatus"] = reconciliation_status
            row["matchStatus"] = "通过" if reconciliation_status == "passed" else "金额差异"
        rows.append(row)

    normalized["rows"] = rows
    summary = dict(warehouse_comparison.get("summary") or {})
    pdf_total = round(sum(float(row.get("pdfAmountTotal") or 0) for row in rows), 2)
    excel_total = round(sum(float(row.get("excelAmountTotal") or 0) for row in rows), 2)
    comparable_rows = [
        row
        for row in rows
        if row.get("reconciliationStatus") in {"passed", "amount_difference"}
    ]
    passed_count = sum(1 for row in rows if row.get("reconciliationStatus") == "passed")
    summary.update(
        {
            "pdfAmountTotal": pdf_total,
            "excelAmountTotal": excel_total,
            "amountDeltaTotal": round(pdf_total - excel_total, 2),
            "totalPassed": bool(rows)
            and not (warehouse_comparison.get("errors") or [])
            and all(row.get("reconciliationStatus") == "passed" for row in rows),
            "warehouseCount": len(rows),
            "passedCount": passed_count,
            "exceptionCount": len(rows) - passed_count,
            "diffWarehouses": [
                row.get("warehouseId")
                for row in rows
                if row.get("reconciliationStatus") != "passed"
            ],
            "comparableExcelAmountTotal": round(
                sum(float(row.get("excelAmountTotal") or 0) for row in comparable_rows),
                2,
            ),
            "comparableAmountDeltaTotal": round(
                sum(float(row.get("amountDelta") or 0) for row in comparable_rows),
                2,
            ),
            "missingPdfAmountTotal": round(
                sum(
                    float(row.get("excelAmountTotal") or 0)
                    for row in rows
                    if row.get("reconciliationStatus") == "missing_pdf_invoice"
                ),
                2,
            ),
        }
    )
    normalized["summary"] = summary
    return normalized


def _perform_labor_extract_compare(run_id: str, task_generation_id: str = "") -> dict:
    metadata = _labor_metadata_or_404(run_id)
    generation = str(task_generation_id or "").strip()
    if generation and _labor_task_generation_id(metadata) != generation:
        return {**metadata, "_taskGenerationRejected": True}
    input_fingerprint_at_start = _labor_result_input_fingerprint(metadata)
    from .engine.labor.models import line_items_from_dicts

    previous_pdf_rows = line_items_from_dicts(metadata.get("pdfExtractedRows") or [])
    run_dir = get_labor_run_dir(run_id)
    mapping = metadata.get("excelMapping") or {}
    sheet_name = metadata.get("workbookSheet") or ""
    workbook_paths = _labor_workbook_paths(metadata)
    pdf_paths = [Path(record["path"]) for record in metadata.get("files", {}).get("pdfInvoices", []) if record.get("path")]
    supplier = metadata.get("supplierName", "")
    period_start = metadata.get("periodStart", "")
    period_end = metadata.get("periodEnd", "")
    currency = metadata.get("currency", "")
    manual_name_mapping = metadata.get("manualNameMapping") or {}
    profile_governance = _normalized_profile_governance(metadata.get("profileGovernance"))
    correction_governance = _normalized_correction_governance(metadata.get("correctionGovernance"))
    name_mapping_governance = _normalized_name_mapping_governance(metadata.get("nameMappingGovernance"))
    active_supplier_profile = _active_supplier_profile_override(supplier, profile_governance)
    final_status = "已生成差异报告"
    structure_reconciliation = {
        "decisions": [],
        "unresolvedFiles": [],
        "reconciledFiles": [],
        "currencyCodes": [],
    }
    auto_ocr_candidate = {}

    try:
        # [F] Excel 解析（多文件合并）
        logger.info(f"[{run_id}] [F] 开始解析 Excel: {len(workbook_paths)} 个文件, 工作表: {sheet_name}")
        _update_labor_progress(
            run_id,
            phase="excel",
            phase_label="读取 Excel 账单",
            message="正在读取 Excel 账单。",
            total_files=len(workbook_paths),
            processed_files=0,
            task_generation_id=generation,
        )
        _update_labor_task_metadata(run_id, generation, {"stage": "解析 Excel 账单"})
        excel_rows = []
        for index, wb_path in enumerate(workbook_paths, start=1):
            _update_labor_progress(
                run_id,
                phase="excel",
                phase_label="读取 Excel 账单",
                message=f"正在读取账单文件：{wb_path.name}",
                total_files=len(workbook_paths),
                processed_files=index - 1,
                current_file=wb_path.name,
                task_generation_id=generation,
            )
            workbook_sheet, workbook_mapping = _labor_workbook_mapping(metadata, wb_path)
            rows = read_workbook_rows(wb_path, workbook_sheet, workbook_mapping)
            rows = _labor_apply_workbook_warehouse_fallback(
                rows,
                wb_path.name,
                str(metadata.get("warehouseId") or metadata.get("warehouse_id") or ""),
            )
            logger.info(f"[{run_id}] [F]   {wb_path.name}: {len(rows)} 行")
            excel_rows.extend(rows)
            _update_labor_progress(
                run_id,
                phase="excel",
                phase_label="读取 Excel 账单",
                message=f"账单文件已读取：{wb_path.name}",
                total_files=len(workbook_paths),
                processed_files=index,
                current_file=wb_path.name,
                task_generation_id=generation,
            )
        logger.info(f"[{run_id}] [F] Excel 解析完成: 共 {len(excel_rows)} 行")
        # 聚合同一员工的多天记录
        excel_rows = _aggregate_excel_rows(excel_rows)
        excel_warehouse_data = [
            {"warehouse_id": row.warehouse_id, "hours": row.hours, "amount": row.amount, "employee_name": row.employee_name_raw}
            for row in excel_rows
        ]
        cost_summaries = _labor_cost_summaries(workbook_paths)
        ai_cache_audit = audit_ai_page_cache_candidates(pdf_paths)
        pdf_page_audit: list[dict] = []
        ai_cache_reconciliation_preview = build_ai_cache_reconciliation_preview(
            pdf_paths,
            excel_rows,
            amount_tolerance=AI_CONFIG["amount_tolerance"],
            hours_tolerance=AI_CONFIG["hours_tolerance"],
            confidence_threshold=AI_CONFIG["confidence_threshold"],
            currency=currency,
        )
        reocr_plan = build_reocr_candidate_plan(
            ai_cache_reconciliation_preview.get("fileQuality", []) or [],
            amount_tolerance=AI_CONFIG["amount_tolerance"],
        )
        pdf_text_coverage = _summarize_pdf_text_coverage(pdf_paths)
        _attach_text_coverage_to_reocr_plan(reocr_plan, pdf_text_coverage)
        name_mapping_governance["candidates"] = _merge_name_mapping_candidates(
            name_mapping_governance,
            _build_name_mapping_candidates_from_reocr_plan(run_id, reocr_plan),
        )
        existing_reocr_summary = reocr_plan.get("summary", {}) if isinstance(reocr_plan, dict) else {}
        text_coverage_summary = pdf_text_coverage.get("summary", {}) if isinstance(pdf_text_coverage, dict) else {}
        all_pdfs_need_ocr = bool(pdf_paths) and int(text_coverage_summary.get("imageOnlyFileCount") or 0) >= len(pdf_paths)
        bulk_image_reocr_batch = len(pdf_paths) > 1 and all_pdfs_need_ocr
        if (
            (metadata.get("materialReplaySource") or {}).get("skipFormalAiExtraction") is True
            and int(existing_reocr_summary.get("taskCount") or 0) > 0
            and bulk_image_reocr_batch
        ):
            logger.warning(f"[{run_id}] 材料批次已判定为图片识别主路径，跳过正式 AI 抽取，直接生成待复核结果。")
            update_labor_metadata(run_id, {"stage": "待图片识别复核"})
            pdf_rows = []
            comparison = compare_labor_items(
                [],
                excel_rows,
                amount_tolerance=AI_CONFIG["amount_tolerance"],
                hours_tolerance=AI_CONFIG["hours_tolerance"],
                confidence_threshold=AI_CONFIG["confidence_threshold"],
                manual_name_mapping=manual_name_mapping,
            )
            warehouse_comparison = compare_by_warehouse(
                pdf_totals=[],
                pdf_rows=pdf_rows,
                excel_rows_with_warehouse=excel_warehouse_data,
                amount_tolerance=AI_CONFIG["amount_tolerance"],
                hours_tolerance=AI_CONFIG["hours_tolerance"],
                confidence_threshold=AI_CONFIG["confidence_threshold"],
                manual_name_mapping=manual_name_mapping,
            )
            extraction_quality = {
                "level": "critical",
                "message": "该批 PDF 已识别为图片型或无文本层，必须先完成图片识别复核。",
                "issues": [
                    "系统已在只读验证阶段生成图片识别任务；为避免长时间等待，正式抽取不会重复调用图片识别。",
                ],
                "retryAttempted": False,
                "retryApplied": False,
            }
            reconciliation_diagnostics = build_reconciliation_diagnostics(
                pdf_totals=[],
                comparison_summary=comparison["summary"],
                warehouse_comparison=warehouse_comparison,
                amount_tolerance=AI_CONFIG["amount_tolerance"],
                cost_summaries=cost_summaries,
            )
            name_mapping_governance["summary"] = _summarize_name_mapping_candidates(name_mapping_governance["candidates"])
            combined_row_governance = _build_material_combined_row_governance(
                batch_key=str((metadata.get("materialReplaySource") or {}).get("batchKey") or run_id),
                candidate_matches=comparison.get("candidateMatches", []),
                amount_tolerance=AI_CONFIG["amount_tolerance"],
                hours_tolerance=AI_CONFIG["hours_tolerance"],
            )
            allocation_governance = _build_allocation_governance(
                run_id,
                warehouse_comparison,
                metadata.get("allocationGovernance") if isinstance(metadata.get("allocationGovernance"), dict) else {},
            )
            review_queues = _build_material_review_queues(
                comparison_summary=comparison["summary"],
                warehouse_summary=warehouse_comparison.get("summary", {}),
                exception_rows=comparison.get("rows", []),
                pdf_text_coverage=pdf_text_coverage,
                reocr_plan=reocr_plan,
                ai_cache_preview=ai_cache_reconciliation_preview,
                name_mapping_governance=name_mapping_governance,
                combined_row_governance=combined_row_governance,
                allocation_issues=warehouse_comparison.get("allocationIssues", []),
                hours_tolerance=AI_CONFIG["hours_tolerance"],
            )
            report_path = run_dir / safe_labor_storage_filename("labor_reconciliation_report.xlsx", "pending_reocr_review")
            build_labor_report(
                report_path,
                comparison,
                pdf_rows,
                excel_rows,
                mapping,
                warehouse_comparison,
                extraction_quality,
                reconciliation_diagnostics=reconciliation_diagnostics,
                ai_cache_audit=ai_cache_audit,
            )
            files = dict(metadata.get("files", {}))
            files["diffReport"] = attach_labor_file(run_id, report_path, "待图片识别复核报告")
            conclusion = _build_conclusion(warehouse_comparison, comparison, extraction_quality, amount_tolerance=AI_CONFIG["amount_tolerance"])
            updated = update_labor_metadata(
                run_id,
                {
                    "status": "待图片识别复核",
                    "stage": "待图片识别复核",
                    "files": files,
                    "comparisonSummary": {**comparison["summary"], **conclusion},
                    "comparisonRows": comparison["rows"],
                    "candidateMatches": comparison.get("candidateMatches", []),
                    "warehouseComparison": warehouse_comparison,
                    "extractionQuality": extraction_quality,
                    "reconciliationDiagnostics": reconciliation_diagnostics,
                    "costSummaries": cost_summaries,
                    "aiCacheAudit": ai_cache_audit,
                    "aiCacheReconciliationPreview": ai_cache_reconciliation_preview,
                    "reocrPlan": reocr_plan,
                    "nameMappingGovernance": name_mapping_governance,
                    "combinedRowGovernance": combined_row_governance,
                    "allocationGovernance": allocation_governance,
                    "reviewQueues": review_queues,
                    "profileGovernance": profile_governance,
                    "correctionGovernance": correction_governance,
                    "pdfExtractedRows": [],
                    "excelRows": [row.to_dict() for row in excel_rows],
                    "diffDownloadUrl": files["diffReport"]["downloadUrl"],
                },
            )
            return updated

        # === Stage 1: Quick total extraction ===
        logger.info(f"[{run_id}] === Stage 1: 快速总金额抽取 ({len(pdf_paths)} 个 PDF) ===")
        _update_labor_progress(
            run_id,
            phase="pdf_total",
            phase_label="核对发票总金额",
            message=f"正在读取 {len(pdf_paths)} 张 PDF 发票的总金额。",
            total_files=len(pdf_paths),
            processed_files=0,
            task_generation_id=generation,
        )
        _update_labor_task_metadata(run_id, generation, {"stage": "Stage 1: 快速抽取总金额"})
        pdf_totals = quick_extract_totals(pdf_paths, AI_CONFIG, supplier=supplier)
        for t in pdf_totals:
            logger.info(f"[{run_id}]   PDF总金额: {t.get('source_file','?')} -> {t.get('total_amount', 0)}")
        invoice_evidence_audit = list(pdf_totals)
        _update_labor_task_metadata(run_id, generation, {"invoiceEvidenceAudit": invoice_evidence_audit})
        invoice_page_pdf_names = {
            str(total.get("source_file") or "")
            for total in pdf_totals
            if any(
                str(page.get("role") or page.get("page_role") or "").strip().lower()
                in LABOR_INVOICE_PAGE_ROLES
                for page in (total.get("page_evidence") or [])
                if isinstance(page, dict)
            )
        }
        non_payable_pdf_names = {
            str(total.get("source_file") or "")
            for total in pdf_totals
            if _labor_total_is_explicitly_non_payable(total)
        }
        non_payable_pdf_names.update(
            _non_payable_pdf_names(pdf_totals) - invoice_page_pdf_names
        )
        reconciliation_pdf_totals = [
            total
            for total in pdf_totals
            if str(total.get("source_file") or "") not in non_payable_pdf_names
        ]
        authoritative_pdf_names = {
            str(total.get("source_file") or "")
            for total in reconciliation_pdf_totals
            if str(total.get("source_file") or "") not in non_payable_pdf_names
            and total.get("authoritative") is not False
            and not total.get("warehouse_conflict")
            and float(total.get("total_amount") or 0) > 0
        }
        unresolved_pdf_names = {
            str(total.get("source_file") or "")
            for total in reconciliation_pdf_totals
            if str(total.get("source_file") or "") not in non_payable_pdf_names
            and str(total.get("source_file") or "") not in authoritative_pdf_names
        }
        has_complete_warehouse_cost_evidence = _labor_has_complete_warehouse_cost_evidence(
            reconciliation_pdf_totals,
            cost_summaries,
        )
        warehouse_comparison = compare_by_warehouse(
            pdf_totals=reconciliation_pdf_totals,
            excel_rows_with_warehouse=excel_warehouse_data,
            amount_tolerance=AI_CONFIG["amount_tolerance"],
            manual_name_mapping=manual_name_mapping,
        )
        warehouse_comparison = _labor_apply_cost_summary_amount_basis(
            warehouse_comparison,
            cost_summaries,
            amount_tolerance=AI_CONFIG["amount_tolerance"],
        )

        pdf_rows = []
        raw_pdf_rows = []
        comparison = {"summary": {}, "rows": [], "candidateMatches": []}
        extraction_quality = {"level": "ok", "message": "总金额核对通过，无需抽取员工明细。", "issues": [], "retryAttempted": False, "retryApplied": False}
        stage2_quality_issues: list[str] = []
        force_full_image_detail = bool(
            all_pdfs_need_ocr
            and len(reconciliation_pdf_totals) > 1
            and not unresolved_pdf_names
            and len(authoritative_pdf_names) == len(reconciliation_pdf_totals)
        )
        force_employee_detail = bool(
            _labor_requires_employee_detail(metadata)
            or metadata.get("materialReplaySource")
            or force_full_image_detail
        )

        if warehouse_comparison["summary"]["totalPassed"] and not force_employee_detail:
            logger.info(f"[{run_id}] ✅ Stage 1 通过: 总金额一致，无需抽取员工明细")
            _update_labor_task_metadata(run_id, generation, {"stage": "Stage 1 通过: 总金额一致"})
        else:
            # === Stage 2: Full extraction for diff warehouses ===
            if warehouse_comparison["summary"]["totalPassed"]:
                logger.info(f"[{run_id}] === Stage 2: 总金额一致，但材料回放要求员工级明细核对 ===")
                stage2_quality_issues.append("总金额已通过，但真实材料回放批次仍执行员工级明细核对，用于发现合并行、姓名映射和员工级差异。")
            else:
                logger.info(f"[{run_id}] === Stage 2: 总金额不一致，进入员工明细抽取 ===")
            _update_labor_task_metadata(run_id, generation, {"stage": "Stage 2: 抽取员工明细"})
            warehouse_rows = warehouse_comparison.get("rows", [])
            if force_employee_detail:
                target_pdf_names = {
                    path.name
                    for path in pdf_paths
                    if path.name not in non_payable_pdf_names
                }
                target_warehouse_ids = {
                    str(row.get("warehouseId") or "")
                    for row in warehouse_rows
                    if row.get("reconciliationStatus") != "missing_pdf_invoice"
                }
            else:
                target_pdf_names = _labor_stage2_target_pdf_names(warehouse_rows)
                target_warehouse_ids = {
                    str(row.get("warehouseId") or "")
                    for row in warehouse_rows
                    if row.get("reconciliationStatus") in {"amount_difference", "needs_review"}
                }
            target_pdf_names.difference_update(non_payable_pdf_names)
            filtered_pdf_paths = [
                path
                for path in pdf_paths
                if path.name in target_pdf_names and path.name not in non_payable_pdf_names
            ]
            invoice_evidence_pages = _labor_invoice_evidence_pages(reconciliation_pdf_totals)
            allowed_pages_by_source = {
                path.name: invoice_evidence_pages[path.name]
                for path in filtered_pdf_paths
                if invoice_evidence_pages.get(path.name)
            }
            has_unmapped_target = any(
                str(total.get("source_file") or "") in target_pdf_names
                and not str(total.get("warehouse_id") or "").strip()
                for total in reconciliation_pdf_totals
            )
            filtered_excel_rows = (
                excel_rows
                if force_employee_detail or has_unmapped_target
                else [row for row in excel_rows if str(row.warehouse_id or "") in target_warehouse_ids]
            )
            non_payable_pdf_paths = [path for path in pdf_paths if path.name in non_payable_pdf_names]
            if non_payable_pdf_paths:
                issue = (
                    "检测到支持材料/附件 PDF，仅保留为审计证据，不执行员工金额抽取。"
                    f" 文件: {', '.join(path.name for path in non_payable_pdf_paths)}"
                )
                stage2_quality_issues.append(issue)
                logger.warning(f"[{run_id}] {issue}")
            unavailable_target_names = target_pdf_names - {path.name for path in filtered_pdf_paths}
            if unavailable_target_names:
                issue = f"异常发票证据文件不可用，未扩大到其他 PDF: {', '.join(sorted(unavailable_target_names))}"
                stage2_quality_issues.append(issue)
                logger.warning(f"[{run_id}] {issue}")
            if not filtered_pdf_paths and not force_employee_detail:
                extraction_quality = {
                    "level": "warning",
                    "message": "存在材料完整性差异，但没有需要执行员工明细 OCR 的 PDF。",
                    "issues": list(stage2_quality_issues),
                    "retryAttempted": False,
                    "retryApplied": False,
                }
            if filtered_pdf_paths:

                logger.info(f"[{run_id}] [C/D] 开始抽取员工明细: {len(filtered_pdf_paths)} 个 PDF, {len(filtered_excel_rows)} 行 Excel")
                _update_labor_progress(
                    run_id,
                    phase="pdf_detail",
                    phase_label="识别 PDF 发票明细",
                    message=f"正在准备识别 {len(filtered_pdf_paths)} 张 PDF 发票明细。",
                    total_files=len(filtered_pdf_paths),
                    processed_files=0,
                    processed_pages=0,
                    task_generation_id=generation,
                )
                _update_labor_task_metadata(
                    run_id,
                    generation,
                    {"stage": f"Stage 2: AI 抽取 {len(filtered_pdf_paths)} 个 PDF"},
                )
                extraction_error = ""
                if all_pdfs_need_ocr and not _ai_ready(AI_CONFIG) and not str(AI_CONFIG.get("ocr_command") or "").strip():
                    stage2_quality_issues.append("当前未连接图片识别服务，图片发票无法自动读取。请启用 MiMo 或兼容的图片识别模型后重试。")
                progress_callback = _labor_ai_progress_callback(
                    run_id,
                    total_files=len(filtered_pdf_paths),
                    task_generation_id=generation,
                )
                structured_pdf_rows = extract_structured_invoice_rows(filtered_pdf_paths, excel_rows)
                closed_structured_rows = prefer_closed_structured_rows(
                    [],
                    structured_pdf_rows,
                    filtered_pdf_paths,
                    tolerance=AI_CONFIG["amount_tolerance"],
                )
                structured_sources = {str(row.source_file or "") for row in closed_structured_rows}
                target_sources = {path.name for path in filtered_pdf_paths}
                if _labor_should_run_auto_ocr(
                    command=str(AI_CONFIG.get("ocr_command") or ""),
                    pdf_rows=closed_structured_rows,
                    target_pdf_names=target_sources,
                    reconciliation_pdf_totals=reconciliation_pdf_totals,
                ):
                    logger.info(f"[{run_id}] 在旧明细抽取前提交本地 OCR 候选任务")
                    _update_labor_task_metadata(run_id, generation, {"stage": "自动识别图片发票"})
                    _update_labor_progress(
                        run_id,
                        phase="auto_ocr",
                        phase_label="本地 OCR 识别",
                        message=f"正在优先检查 {len(filtered_pdf_paths)} 张 PDF 的本地 OCR 缓存。",
                        total_files=len(filtered_pdf_paths),
                        processed_files=0,
                        task_generation_id=generation,
                    )
                    auto_ocr_candidate = _run_labor_auto_ocr_candidate(
                        run_id,
                        filtered_pdf_paths,
                        filtered_excel_rows,
                        reconciliation_pdf_totals,
                        supplier=supplier,
                        period_start=period_start,
                        period_end=period_end,
                        currency=currency,
                        command=str(AI_CONFIG.get("ocr_command") or ""),
                        timeout_seconds=max(int(AI_CONFIG.get("timeout_seconds") or 90) * 10, 300),
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                        hours_tolerance=AI_CONFIG["hours_tolerance"],
                        task_generation_id=generation,
                    )
                    reconciliation_pdf_totals = _labor_apply_ocr_pdf_total_evidence(
                        reconciliation_pdf_totals,
                        auto_ocr_candidate,
                    )
                if auto_ocr_candidate.get("safeToUse"):
                    pdf_rows = line_items_from_dicts(auto_ocr_candidate.get("rows") or [])
                    logger.info(f"[{run_id}] 早期 OCR 候选通过安全门禁，跳过旧明细抽取: {len(pdf_rows)} 行")
                elif target_sources and target_sources.issubset(structured_sources):
                    pdf_rows = closed_structured_rows
                    logger.info(f"[{run_id}] 结构行金额全部闭合，跳过 AI 明细抽取: {len(pdf_rows)} 行")
                else:
                    targeted_retry_applied = False
                    review_ocr_rows_applied = False
                    if auto_ocr_candidate:
                        targeted_plan = build_targeted_ocr_retry_plan(auto_ocr_candidate)
                        if targeted_plan.get("eligible"):
                            targeted_pages = targeted_plan.get("allowedPagesBySource") or {}
                            targeted_paths = [path for path in filtered_pdf_paths if path.name in targeted_pages]
                            candidate_rows = line_items_from_dicts(auto_ocr_candidate.get("rows") or [])
                            expected_totals = _labor_ocr_expected_totals(
                                reconciliation_pdf_totals,
                                {path.name for path in filtered_pdf_paths},
                                excel_rows=filtered_excel_rows,
                            )
                            try:
                                targeted_rows = extract_invoice_items(
                                    targeted_paths, AI_CONFIG,
                                    supplier=supplier, period_start=period_start, period_end=period_end, currency=currency,
                                    expected_rows=_expected_labor_rows(filtered_excel_rows),
                                    supplier_profile_override=active_supplier_profile,
                                    progress_callback=progress_callback,
                                    audit_collector=pdf_page_audit,
                                    allowed_pages_by_source=targeted_pages,
                                )
                                targeted_merge = merge_targeted_ocr_retry_rows(
                                    candidate_rows,
                                    targeted_rows,
                                    allowed_pages_by_source=targeted_pages,
                                    expected_totals=expected_totals,
                                    tolerance=AI_CONFIG["amount_tolerance"],
                                )
                                targeted_retry_applied = bool(targeted_merge.get("closed"))
                                auto_ocr_candidate["targetedRetry"] = {
                                    **targeted_plan,
                                    "applied": targeted_retry_applied,
                                    "retriedPageCount": sum(len(pages) for pages in targeted_pages.values()),
                                }
                                if targeted_retry_applied:
                                    pdf_rows = targeted_merge["rows"]
                                    logger.info(
                                        f"[{run_id}] 姓名待复核仅重识别 {sum(len(pages) for pages in targeted_pages.values())} 页，"
                                        "合并后金额闭合，跳过整批旧明细抽取"
                                    )
                            except Exception as exc:  # noqa: BLE001 - targeted retry must safely fall back to the full extraction path.
                                auto_ocr_candidate["targetedRetry"] = {
                                    **targeted_plan,
                                    "applied": False,
                                    "error": str(exc),
                                }
                                logger.warning(f"[{run_id}] 姓名待复核定向页识别失败，回退原明细抽取: {exc}")
                    if (
                        not targeted_retry_applied
                        and _labor_can_apply_review_ocr_rows(auto_ocr_candidate, target_sources)
                    ):
                        pdf_rows = line_items_from_dicts(auto_ocr_candidate.get("rows") or [])
                        review_ocr_rows_applied = True
                        auto_ocr_candidate["reviewRowsApplied"] = True
                        auto_ocr_candidate["reviewRowsSources"] = sorted(target_sources)
                        stage2_quality_issues.append(
                            "本地图片识别已完成全部页面并读取到票面明确总额；员工明细作为待复核证据展示，"
                            "姓名或逐行金额门禁未全部通过，确认前不能机器放行。"
                        )
                        logger.warning(
                            f"[{run_id}] OCR 全页与票面总额证据完整，保留 {len(pdf_rows)} 条待复核明细；"
                            f"阻断项: {auto_ocr_candidate.get('blockers', [])}"
                        )
                    if not targeted_retry_applied and not review_ocr_rows_applied:
                        if auto_ocr_candidate:
                            stage2_quality_issues.append(
                                "本地图片识别已完成，但金额闭合、页面完整性或姓名安全门禁未全部通过；已回退原明细抽取，候选仅供人工复核。"
                            )
                            logger.warning(
                                f"[{run_id}] 早期 OCR 候选未自动采用，回退旧明细抽取: "
                                f"{auto_ocr_candidate.get('blockers', [])}"
                            )
                        try:
                            pdf_rows = extract_invoice_items(
                                filtered_pdf_paths, AI_CONFIG,
                                supplier=supplier, period_start=period_start, period_end=period_end, currency=currency,
                                expected_rows=_expected_labor_rows(filtered_excel_rows),
                                supplier_profile_override=active_supplier_profile,
                                progress_callback=progress_callback,
                                audit_collector=pdf_page_audit,
                                allowed_pages_by_source=allowed_pages_by_source,
                            )
                        except Exception as exc:  # noqa: BLE001 - material replay must preserve governance evidence when online AI is unavailable.
                            extraction_error = str(exc)
                            if not (
                                metadata.get("materialReplaySource")
                                or all_pdfs_need_ocr
                                or has_complete_warehouse_cost_evidence
                            ):
                                raise
                            pdf_rows = []
                            logger.warning(f"[{run_id}] 员工明细抽取失败，保留已确认的仓库级金额证据: {extraction_error}")
                    pdf_rows = prefer_closed_structured_rows(
                        pdf_rows,
                        structured_pdf_rows,
                        filtered_pdf_paths,
                        tolerance=AI_CONFIG["amount_tolerance"],
                    )
                raw_pdf_rows = list(pdf_rows)
                structure_promotion = promote_structured_invoice_evidence(
                    filtered_pdf_paths,
                    pdf_totals,
                    raw_pdf_rows,
                    excel_rows,
                    tolerance=AI_CONFIG["amount_tolerance"],
                    page_audit=pdf_page_audit,
                )
                pdf_totals = _labor_apply_ocr_pdf_total_evidence(
                    structure_promotion.pdf_totals,
                    auto_ocr_candidate,
                )
                invoice_evidence_audit = list(pdf_totals)
                reconciliation_pdf_totals = [
                    total
                    for total in pdf_totals
                    if str(total.get("source_file") or "") not in non_payable_pdf_names
                ]
                structure_reconciliation = {
                    "decisions": structure_promotion.decisions,
                    "unresolvedFiles": list(structure_promotion.unresolved_files),
                    "reconciledFiles": list(structure_promotion.reconciled_files),
                    "currencyCodes": list(structure_promotion.currency_codes),
                }
                pdf_rows = _filter_labor_rows_to_invoice_evidence_pages(structure_promotion.pdf_rows, pdf_totals)
                detail_total_retry_attempted = False
                detail_total_retry_applied = False
                remaining_detail_total_mismatches: dict[str, dict] = {}
                authoritative_detail_totals = _labor_authoritative_detail_totals(
                    reconciliation_pdf_totals,
                    {path.name for path in filtered_pdf_paths},
                )
                if all_pdfs_need_ocr and previous_pdf_rows and authoritative_detail_totals:
                    previous_scoped_rows = [
                        row for row in previous_pdf_rows if row.source_file in authoritative_detail_totals
                    ]
                    current_scoped_rows = [
                        row for row in pdf_rows if row.source_file in authoritative_detail_totals
                    ]
                    pdf_rows, resumed_sources = _merge_labor_detail_retry_rows(
                        previous_scoped_rows,
                        current_scoped_rows,
                        authoritative_detail_totals,
                    )
                    if resumed_sources:
                        logger.info(
                            f"[{run_id}] 图片明细续跑采用更接近金额闭合的新结果: "
                            f"{', '.join(sorted(resumed_sources))}"
                        )
                    preserved_sources = {
                        row.source_file for row in pdf_rows
                    } - resumed_sources
                    if preserved_sources:
                        stage2_quality_issues.append(
                            "图片明细续跑保留上一轮更完整结果；本轮失败页未覆盖已有员工明细。"
                        )
                initial_detail_total_mismatches = _labor_detail_total_mismatches(
                    pdf_rows,
                    authoritative_detail_totals,
                    AI_CONFIG["amount_tolerance"],
                )
                if _labor_should_retry_detail_totals(
                    pdf_rows=pdf_rows,
                    mismatches=initial_detail_total_mismatches,
                    auto_ocr_candidate=auto_ocr_candidate,
                ):
                    detail_total_retry_attempted = True
                    retry_pdf_names = set(initial_detail_total_mismatches)
                    retry_pdf_paths = [path for path in filtered_pdf_paths if path.name in retry_pdf_names]
                    retry_warehouse_ids = {
                        str(item.get("warehouseId") or "")
                        for item in initial_detail_total_mismatches.values()
                        if str(item.get("warehouseId") or "")
                    }
                    retry_excel_rows = [
                        row
                        for row in filtered_excel_rows
                        if not retry_warehouse_ids or str(row.warehouse_id or "") in retry_warehouse_ids
                    ]
                    retry_allowed_pages = {
                        source_file: pages
                        for source_file, pages in allowed_pages_by_source.items()
                        if source_file in retry_pdf_names
                    }
                    retry_config = dict(AI_CONFIG)
                    retry_config["cache_enabled"] = False
                    retry_config["parallel_max_workers"] = 1
                    retry_config["parallel_image_render_workers"] = 1
                    try:
                        current_render_scale = float(retry_config.get("render_scale") or 0)
                    except (TypeError, ValueError):
                        current_render_scale = 0.0
                    retry_config["render_scale"] = max(current_render_scale, 2.4)
                    logger.info(
                        f"[{run_id}] 员工明细金额不闭合，高清重识别 {len(retry_pdf_paths)} 个 PDF: "
                        f"{', '.join(sorted(retry_pdf_names))}"
                    )
                    try:
                        retry_rows = extract_invoice_items(
                            retry_pdf_paths,
                            retry_config,
                            supplier=supplier,
                            period_start=period_start,
                            period_end=period_end,
                            currency=currency,
                            expected_rows=_expected_labor_rows(retry_excel_rows),
                            supplier_profile_override=active_supplier_profile,
                            retry_mode=True,
                            progress_callback=progress_callback,
                            audit_collector=pdf_page_audit,
                            allowed_pages_by_source=retry_allowed_pages,
                        )
                        retry_rows = _filter_labor_rows_to_invoice_evidence_pages(retry_rows, pdf_totals)
                        pdf_rows, replaced_sources = _merge_labor_detail_retry_rows(
                            pdf_rows,
                            retry_rows,
                            initial_detail_total_mismatches,
                        )
                        detail_total_retry_applied = bool(replaced_sources)
                        if replaced_sources:
                            logger.info(
                                f"[{run_id}] 高清重识别改善金额闭合，采用新结果: "
                                f"{', '.join(sorted(replaced_sources))}"
                            )
                    except Exception as exc:  # noqa: BLE001 - retry failure must preserve the original evidence.
                        logger.warning(f"[{run_id}] 员工明细金额闭合重识别失败，保留原结果: {exc}")

                    remaining_detail_total_mismatches = _labor_detail_total_mismatches(
                        pdf_rows,
                        authoritative_detail_totals,
                        AI_CONFIG["amount_tolerance"],
                    )
                    for source_file, mismatch in remaining_detail_total_mismatches.items():
                        stage2_quality_issues.append(
                            f"员工明细与发票权威总额不闭合：{source_file} 权威总额 "
                            f"${float(mismatch['expectedAmount']):,.2f}，明细 ${float(mismatch['actualAmount']):,.2f}，"
                            f"差异 ${abs(float(mismatch['delta'])):,.2f}；该文件员工归因不可直接采信。"
                        )
                logger.info(f"[{run_id}] [C/D] 员工明细抽取完成: {len(pdf_rows)} 条记录")

                # 正式核对只读取已批准的运行时 Profile，不得自动改写全局配置。
                # 抽取成功/失败会留在批次审计与 Profile 候选中，由治理流程显式批准、回滚。

                if not pdf_rows and metadata.get("materialReplaySource"):
                    final_status = "待图片识别复核"
                    issue = (
                        "PDF 员工明细抽取失败，已保留历史图片识别审计、图片重新识别计划和姓名匹配建议；"
                        "请先预览并确认图片识别结果后再采纳为正式结果。"
                    )
                    if extraction_error:
                        issue = f"{issue} 原因: {extraction_error}"
                    stage2_quality_issues.append(issue)
                    comparison = compare_labor_items(
                        [],
                        filtered_excel_rows,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                        hours_tolerance=AI_CONFIG["hours_tolerance"],
                        confidence_threshold=AI_CONFIG["confidence_threshold"],
                        manual_name_mapping=manual_name_mapping,
                    )
                    extraction_quality = {
                        "level": "critical",
                        "message": "PDF 员工明细暂不可用，必须先完成图片识别复核。",
                        "issues": [],
                        "retryAttempted": False,
                        "retryApplied": False,
                    }
                    _append_quality_issues(extraction_quality, stage2_quality_issues)
                    warehouse_comparison = compare_by_warehouse(
                        pdf_totals=reconciliation_pdf_totals,
                        pdf_rows=pdf_rows,
                        excel_rows_with_warehouse=excel_warehouse_data,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                        hours_tolerance=AI_CONFIG["hours_tolerance"],
                        confidence_threshold=AI_CONFIG["confidence_threshold"],
                        manual_name_mapping=manual_name_mapping,
                    )
                    warehouse_comparison = _labor_apply_cost_summary_amount_basis(
                        warehouse_comparison,
                        cost_summaries,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                    )
                    _update_labor_task_metadata(run_id, generation, {"stage": "待图片识别复核"})
                    logger.warning(f"[{run_id}] 已生成待图片识别复核结果，未改变核对结论。")
                elif not pdf_rows:
                    if has_complete_warehouse_cost_evidence:
                        final_status = "部分核对完成"
                        issue = (
                            "发票按班组汇总计费，员工页未直接提供可核验的员工金额；"
                            "已保留仓库总额与 OTWS 费用组成核对，员工金额不可据此自动分摊。"
                        )
                        if extraction_error:
                            issue = f"{issue} 员工明细抽取信息: {extraction_error}"
                        stage2_quality_issues.append(issue)
                        comparison = compare_labor_items(
                            [],
                            [],
                            amount_tolerance=AI_CONFIG["amount_tolerance"],
                            hours_tolerance=AI_CONFIG["hours_tolerance"],
                            confidence_threshold=AI_CONFIG["confidence_threshold"],
                            manual_name_mapping=manual_name_mapping,
                        )
                        extraction_quality = {
                            "level": "warning",
                            "message": "已完成仓库级金额核对；员工金额缺少发票直接证据，必须人工复核。",
                            "issues": list(stage2_quality_issues),
                            "retryAttempted": False,
                            "retryApplied": False,
                        }
                        logger.warning(f"[{run_id}] SSS/OTWS 班组汇总发票降级为仓库级核对结果。")
                    elif target_pdf_names & unresolved_pdf_names:
                        issue = "待复核发票未找到可用于员工金额比对的发票页，已保留票面证据并维持待复核结论。"
                        stage2_quality_issues.append(issue)
                        comparison = compare_labor_items(
                            [],
                            filtered_excel_rows,
                            amount_tolerance=AI_CONFIG["amount_tolerance"],
                            hours_tolerance=AI_CONFIG["hours_tolerance"],
                            confidence_threshold=AI_CONFIG["confidence_threshold"],
                            manual_name_mapping=manual_name_mapping,
                        )
                        extraction_quality = {
                            "level": "warning",
                            "message": "发票页面证据不足，必须人工复核。",
                            "issues": list(stage2_quality_issues),
                            "retryAttempted": False,
                            "retryApplied": False,
                        }
                    else:
                        raise ValueError("PDF 未抽取出员工明细。请确认发票是可复制文本 PDF，或先完成图片识别复核。")

                if pdf_rows:
                    logger.info(f"[{run_id}] [G] 开始数据比对: PDF {len(pdf_rows)} 行 vs Excel {len(filtered_excel_rows)} 行")
                    _update_labor_progress(
                        run_id,
                        phase="matching",
                        phase_label="匹配员工明细",
                        message=f"正在匹配 PDF {len(pdf_rows)} 条明细和 Excel {len(filtered_excel_rows)} 条账单。",
                        task_generation_id=generation,
                    )
                    _update_labor_task_metadata(run_id, generation, {"stage": "比对员工明细"})
                    comparison = compare_labor_items(
                        pdf_rows, filtered_excel_rows,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                        hours_tolerance=AI_CONFIG["hours_tolerance"],
                        confidence_threshold=AI_CONFIG["confidence_threshold"],
                        manual_name_mapping=manual_name_mapping,
                    )
                    extraction_quality = calculate_extraction_quality(pdf_rows, comparison["summary"], confidence_threshold=AI_CONFIG["confidence_threshold"])
                    extraction_quality["retryAttempted"] = detail_total_retry_attempted
                    extraction_quality["retryApplied"] = detail_total_retry_applied
                    if remaining_detail_total_mismatches:
                        extraction_quality["level"] = "critical"
                        extraction_quality["message"] = "部分 PDF 员工明细金额无法与权威发票总额闭合，员工归因必须人工复核。"
                    _append_quality_issues(extraction_quality, stage2_quality_issues)
                    logger.info(f"[{run_id}] [G] 比对完成: 质量={extraction_quality['level']}, 问题={len(extraction_quality.get('issues',[]))}条")

                should_retry_quality = bool(pdf_rows) and extraction_quality["level"] in ("warning", "critical")
                if should_retry_quality and detail_total_retry_attempted:
                    should_retry_quality = False
                    logger.info(f"[{run_id}] 已按金额闭合执行定向重识别，跳过重复的全量质量重试")
                if should_retry_quality and all_pdfs_need_ocr:
                    should_retry_quality = False
                    logger.info(f"[{run_id}] PDF 已识别为图片型或无文本层，跳过质量重试以避免重复大图 AI 请求")
                if should_retry_quality and any("快速总金额为 0" in issue for issue in stage2_quality_issues):
                    should_retry_quality = False
                    logger.info(f"[{run_id}] 已包含扫描/未知版式 PDF 补充抽取，跳过质量重试以避免重复大图 AI 请求")
                # 硬编码阈值：PDF > 2 个时全量重试耗时过长（每个 PDF 需 AI 处理 30-60s），
                # 超过此阈值跳过重试，避免整体超时。可通过 AI_MAX_RETRY_PDFS 环境变量覆盖。
                if should_retry_quality and len(filtered_pdf_paths) > 2:
                    should_retry_quality = False
                    logger.info(f"[{run_id}] PDF 数量 {len(filtered_pdf_paths)} > 2，跳过质量重试以避免超时")

                if should_retry_quality:
                    logger.info(f"[{run_id}] 质量为 {extraction_quality['level']}，尝试重试...")
                    _update_labor_task_metadata(run_id, generation, {"stage": "重试抽取（质量优化）"})
                    original_rows = list(pdf_rows)
                    original_comparison = dict(comparison)
                    original_quality = dict(extraction_quality)

                    # 先尝试局部重试低置信度行
                    low_conf_rows = extraction_quality.get("lowConfidenceRows") or []
                    partial_retry_done = False
                    # 硬编码阈值：低置信度行占比 ≤ 50% 时才尝试局部重试，
                    # 超过则认为整体质量太差，直接走全量重试。
                    if low_conf_rows and len(low_conf_rows) <= len(pdf_rows) * 0.5:
                        target_names = list({row["employee_name_raw"] for row in low_conf_rows if row.get("employee_name_raw")})
                        partial_result = _retry_low_confidence_rows(
                            filtered_pdf_paths, low_conf_rows, AI_CONFIG,
                            supplier=supplier, period_start=period_start, period_end=period_end, currency=currency,
                            expected_rows=_expected_labor_rows(filtered_excel_rows),
                            supplier_profile_override=active_supplier_profile,
                        )
                        # 硬编码阈值：局部重试结果行数需 ≥ 原始行数的 80%，否则认为结果不完整，降级到全量重试。
                        if partial_result and len(partial_result) >= len(pdf_rows) * 0.8:
                            partial_comparison = compare_labor_items(
                                partial_result, filtered_excel_rows,
                                amount_tolerance=AI_CONFIG["amount_tolerance"],
                                hours_tolerance=AI_CONFIG["hours_tolerance"],
                                confidence_threshold=AI_CONFIG["confidence_threshold"],
                                manual_name_mapping=manual_name_mapping,
                            )
                            partial_quality = calculate_extraction_quality(partial_result, partial_comparison["summary"], confidence_threshold=AI_CONFIG["confidence_threshold"])
                            if calculate_quality_score(partial_quality, partial_comparison["summary"]) < calculate_quality_score(extraction_quality, comparison["summary"]):
                                # 合并：保留原始高置信度行 + 局部重试的低置信度员工结果
                                high_conf_rows = [r for r in pdf_rows if r.confidence >= AI_CONFIG["confidence_threshold"]]
                                low_conf_names = {name.lower() for name in target_names}
                                retry_low_conf_rows = [r for r in partial_result if r.employee_name_raw.lower() in low_conf_names]
                                merged_rows = high_conf_rows + retry_low_conf_rows
                                merged_comparison = compare_labor_items(
                                    merged_rows, filtered_excel_rows,
                                    amount_tolerance=AI_CONFIG["amount_tolerance"],
                                    hours_tolerance=AI_CONFIG["hours_tolerance"],
                                    confidence_threshold=AI_CONFIG["confidence_threshold"],
                                    manual_name_mapping=manual_name_mapping,
                                )
                                merged_quality = calculate_extraction_quality(merged_rows, merged_comparison["summary"], confidence_threshold=AI_CONFIG["confidence_threshold"])
                                if calculate_quality_score(merged_quality, merged_comparison["summary"]) < calculate_quality_score(extraction_quality, comparison["summary"]):
                                    logger.info(f"[{run_id}] 局部重试改善了质量，采用合并结果（高置信度 {len(high_conf_rows)} 行 + 重试 {len(retry_low_conf_rows)} 行）")
                                    pdf_rows = merged_rows
                                    comparison = merged_comparison
                                    extraction_quality = merged_quality
                                    extraction_quality["retryAttempted"] = True
                                    extraction_quality["retryApplied"] = True
                                    partial_retry_done = True
                                else:
                                    logger.info(f"[{run_id}] 局部重试合并后未改善质量，降级到全量重试")
                            else:
                                logger.info(f"[{run_id}] 局部重试未改善质量，降级到全量重试")

                    # 局部重试不够好或没有低置信度行，走全量重试
                    if not partial_retry_done:
                        pdf_rows, comparison, extraction_quality = _retry_if_better(
                            filtered_pdf_paths, pdf_rows, filtered_excel_rows, extraction_quality, comparison,
                            manual_name_mapping=manual_name_mapping,
                            supplier=supplier, period_start=period_start, period_end=period_end, currency=currency,
                            supplier_profile_override=active_supplier_profile,
                        )

                filtered_retry_rows = _filter_labor_rows_to_invoice_evidence_pages(pdf_rows, pdf_totals)
                if len(filtered_retry_rows) != len(pdf_rows):
                    retry_attempted = extraction_quality.get("retryAttempted", False)
                    retry_applied = extraction_quality.get("retryApplied", False)
                    pdf_rows = filtered_retry_rows
                    comparison = compare_labor_items(
                        pdf_rows,
                        filtered_excel_rows,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                        hours_tolerance=AI_CONFIG["hours_tolerance"],
                        confidence_threshold=AI_CONFIG["confidence_threshold"],
                        manual_name_mapping=manual_name_mapping,
                    )
                    extraction_quality = calculate_extraction_quality(
                        pdf_rows,
                        comparison["summary"],
                        confidence_threshold=AI_CONFIG["confidence_threshold"],
                    )
                    extraction_quality["retryAttempted"] = retry_attempted
                    extraction_quality["retryApplied"] = retry_applied

                # === 供应商 Profile 建议（只生成候选，不静默写入生产 Profile） ===
                if extraction_quality.get("level") == "ok" and pdf_rows:
                    try:
                        profile_data = generate_profile_from_extraction(
                            supplier=supplier,
                            pdf_rows=pdf_rows,
                            extraction_quality_level=extraction_quality.get("level", "ok"),
                        )
                        profile_candidate = _build_profile_candidate(run_id, supplier, profile_data, pdf_rows)
                        _upsert_profile_candidate(profile_governance["candidates"], profile_candidate)
                        logger.info(f"[{run_id}] 已生成供应商 Profile 候选: {profile_candidate['candidateId']}")
                    except Exception as exc:
                        logger.warning(f"[{run_id}] 供应商 Profile 候选生成失败: {exc}")

                # Re-run warehouse comparison with full employee rows for Tier 3.
                # Pass pdf_totals to preserve correct total amounts for non-diff warehouses.
                if pdf_rows:
                    tier3_target_names = _labor_stage2_target_pdf_names(warehouse_rows)
                    tier3_pdf_rows = [row for row in pdf_rows if row.source_file in tier3_target_names]
                    warehouse_comparison = compare_by_warehouse(
                        pdf_totals=reconciliation_pdf_totals,
                        pdf_rows=tier3_pdf_rows,
                        excel_rows_with_warehouse=excel_warehouse_data,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                        hours_tolerance=AI_CONFIG["hours_tolerance"],
                        confidence_threshold=AI_CONFIG["confidence_threshold"],
                        manual_name_mapping=manual_name_mapping,
                    )
                    warehouse_comparison = _labor_apply_cost_summary_amount_basis(
                        warehouse_comparison,
                        cost_summaries,
                        amount_tolerance=AI_CONFIG["amount_tolerance"],
                    )

                    # Recalculate quality with warehouse comparison data, preserving retry flags.
                    retry_attempted = extraction_quality.get("retryAttempted", False)
                    retry_applied = extraction_quality.get("retryApplied", False)
                    extraction_quality = calculate_extraction_quality(pdf_rows, comparison["summary"], warehouse_comparison, confidence_threshold=AI_CONFIG["confidence_threshold"])
                    extraction_quality["retryAttempted"] = retry_attempted
                    extraction_quality["retryApplied"] = retry_applied
                    if remaining_detail_total_mismatches:
                        extraction_quality["level"] = "critical"
                        extraction_quality["message"] = "部分 PDF 员工明细金额无法与权威发票总额闭合，员工归因必须人工复核。"
                    _append_quality_issues(extraction_quality, stage2_quality_issues)

        detected_currency_codes = {
            str(row.currency or "").strip().upper()
            for row in raw_pdf_rows
            if str(row.currency or "").strip()
        }
        detected_currency_codes.update(structure_reconciliation.get("currencyCodes") or [])
        batch_guard_result = evaluate_batch_guards(
            pdf_paths=pdf_paths,
            pdf_totals=reconciliation_pdf_totals,
            raw_pdf_rows=raw_pdf_rows,
            formal_pdf_rows=pdf_rows,
            excel_rows=excel_rows,
            requested_currency=currency,
            detected_currencies=detected_currency_codes,
            unresolved_files=structure_reconciliation.get("unresolvedFiles") or [],
            pdf_text_coverage=pdf_text_coverage,
        )
        amount_scope = resolve_amount_scope(
            " ".join(
                str(value).strip()
                for value in (mapping.get("amountColumns") or [mapping.get("amount")])
                if str(value or "").strip()
            ),
            declared_scope=str(mapping.get("amountScope") or mapping.get("amount_scope") or ""),
        )
        batch_guard = {
            "status": batch_guard_result.status,
            "message": batch_guard_result.message,
            "allowReleasableReport": batch_guard_result.allow_releasable_report,
            "unresolvedFiles": list(batch_guard_result.unresolved_files),
            "amountScope": amount_scope,
            "requestedCurrency": str(currency or "").upper(),
            "detectedCurrencies": sorted(detected_currency_codes),
        }
        if amount_scope == "review" and batch_guard["status"] == "ok":
            batch_guard.update(
                {
                    "status": "amount_scope_review",
                    "message": "Excel 金额字段未明确含税或不含税口径，确认前不能放行。",
                    "allowReleasableReport": False,
                }
            )
        if auto_ocr_candidate.get("reviewRowsApplied"):
            batch_guard.update(
                {
                    "status": "ocr_candidate_review",
                    "message": (
                        "PDF 全页识别和票面总额已完成；员工明细仍有姓名或逐行金额待确认，"
                        "本批次只可人工复核，不能机器放行。"
                    ),
                    "allowReleasableReport": False,
                }
            )
        if batch_guard["status"] == "pdf_recognition_incomplete":
            final_status = "PDF识别未完成"
            extraction_quality["level"] = "critical"
            extraction_quality["message"] = batch_guard["message"]
            _append_quality_issues(extraction_quality, [batch_guard["message"]])
        elif batch_guard["status"] == "ocr_candidate_review":
            final_status = "图片识别待复核"
            extraction_quality["level"] = "critical"
            extraction_quality["message"] = batch_guard["message"]
            _append_quality_issues(extraction_quality, [batch_guard["message"]])
        elif final_status == "待图片识别复核":
            pass
        elif batch_guard["status"] == "currency_review":
            final_status = "待币种确认"
        elif batch_guard["status"] in {"partial_review", "amount_scope_review", "employee_detail_incomplete"}:
            final_status = "部分核对完成"

        # 诊断和候选证据必须先于报告生成，确保下载报告与 API 元数据一致。
        reconciliation_diagnostics = build_reconciliation_diagnostics(
            pdf_totals=reconciliation_pdf_totals,
            pdf_rows=pdf_rows,
            comparison_summary=comparison["summary"],
            warehouse_comparison=warehouse_comparison,
            amount_tolerance=AI_CONFIG["amount_tolerance"],
            cost_summaries=cost_summaries,
        )
        previous_batch_guard_status = batch_guard.get("status")
        batch_guard = _labor_apply_reconciliation_diagnostics_gate(
            batch_guard,
            reconciliation_diagnostics,
        )
        if batch_guard.get("status") != previous_batch_guard_status:
            final_status = "部分核对完成"
        correction_governance["candidates"] = _build_low_confidence_correction_candidates(
            run_id,
            pdf_rows,
            extraction_quality,
        )
        name_mapping_governance["candidates"] = _merge_name_mapping_candidates(
            name_mapping_governance,
            [
                *_build_name_mapping_candidates_from_reocr_plan(run_id, reocr_plan),
                *_build_name_mapping_candidates_from_candidate_matches(
                    run_id,
                    comparison.get("candidateMatches", []),
                ),
            ],
        )
        name_mapping_governance["summary"] = _summarize_name_mapping_candidates(name_mapping_governance["candidates"])
        combined_row_governance = _build_material_combined_row_governance(
            batch_key=str((metadata.get("materialReplaySource") or {}).get("batchKey") or run_id),
            candidate_matches=comparison.get("candidateMatches", []),
            amount_tolerance=AI_CONFIG["amount_tolerance"],
            hours_tolerance=AI_CONFIG["hours_tolerance"],
        )
        allocation_governance = _build_allocation_governance(
            run_id,
            warehouse_comparison,
            metadata.get("allocationGovernance") if isinstance(metadata.get("allocationGovernance"), dict) else {},
        )
        review_queues = _build_material_review_queues(
            comparison_summary=comparison["summary"],
            warehouse_summary=warehouse_comparison.get("summary", {}),
            exception_rows=comparison.get("rows", []),
            pdf_text_coverage=pdf_text_coverage if final_status == "待图片识别复核" or not pdf_rows else {},
            reocr_plan=reocr_plan,
            ai_cache_preview=ai_cache_reconciliation_preview,
            name_mapping_governance=name_mapping_governance,
            combined_row_governance=combined_row_governance,
            allocation_issues=warehouse_comparison.get("allocationIssues", []),
            hours_tolerance=AI_CONFIG["hours_tolerance"],
        )
        presentation = build_labor_presentation(
            comparison,
            excel_record_count=len(excel_rows),
        )
        presentation_errors = validate_labor_presentation(presentation)
        if presentation_errors:
            raise ValueError("核对展示口径校验失败: " + "；".join(presentation_errors))

        report_label = (
            "待图片识别复核报告"
            if final_status == "待图片识别复核"
            else "PDF识别诊断报告"
            if final_status == "PDF识别未完成"
            else "币种待确认报告"
            if final_status == "待币种确认"
            else "差异报告"
        )
        logger.info(f"[{run_id}] 生成{report_label}...")
        _update_labor_progress(
            run_id,
            phase="report",
            phase_label="生成核对报告",
            message=f"正在生成{report_label}。",
            task_generation_id=generation,
        )
        _update_labor_task_metadata(run_id, generation, {"stage": "生成报告"})
        report_path = run_dir / safe_labor_filename("海外劳务工报账核对报告.xlsx", report_label)
        build_labor_report(
            report_path,
            comparison,
            pdf_rows,
            excel_rows,
            mapping,
            warehouse_comparison,
            extraction_quality,
            reconciliation_diagnostics=reconciliation_diagnostics,
            ai_cache_audit=ai_cache_audit,
            presentation=presentation,
        )
        logger.info(f"[{run_id}] 报告已生成: {report_path.name}")
    except ValueError:
        raise
    files = dict(metadata.get("files", {}))
    report_label = (
        "待图片识别复核报告"
        if final_status == "待图片识别复核"
        else "PDF识别诊断报告"
        if final_status == "PDF识别未完成"
        else "币种待确认报告"
        if final_status == "待币种确认"
        else "差异报告"
    )
    files["diffReport"] = attach_labor_file(run_id, report_path, report_label)
    if batch_guard["status"] not in {
        "pdf_recognition_incomplete",
        "currency_review",
        "amount_scope_review",
        "reconciliation_diagnostics_blocked",
    }:
        business_report_path = run_dir / safe_labor_filename("海外劳务工报账核对业务报告.html", "业务报告")
        build_labor_business_html_report(
            business_report_path,
            comparison,
            supplier_name=str(metadata.get("supplierName") or metadata.get("supplier") or ""),
            period_start=str(metadata.get("periodStart") or ""),
            period_end=str(metadata.get("periodEnd") or ""),
            invoice_scope=_labor_business_invoice_scope(metadata, pdf_rows),
            warehouse_comparison=warehouse_comparison,
            excel_record_count=len(excel_rows),
            detail_coverage_complete=_labor_detail_coverage_complete(reconciliation_diagnostics),
            presentation=presentation,
        )
        files["businessReport"] = attach_labor_file(run_id, business_report_path, "业务核对报告")
    else:
        files.pop("businessReport", None)

    # 计算结论级别
    conclusion = _build_conclusion(warehouse_comparison, comparison, extraction_quality, amount_tolerance=AI_CONFIG["amount_tolerance"])
    conclusion["canRelease"] = bool(batch_guard["allowReleasableReport"] and conclusion.get("conclusionLevel") == "pass")
    if not batch_guard["allowReleasableReport"]:
        conclusion["conclusionLevel"] = (
            "critical"
            if batch_guard["status"] in {"pdf_recognition_incomplete", "reconciliation_diagnostics_blocked"}
            else "warning"
        )
        conclusion["conclusionMessage"] = batch_guard["message"]
    review_state = _labor_uat_review_state(conclusion)
    conclusion.update(review_state)

    result_updates = {
            "inputSnapshotRejected": False,
            "status": final_status,
            "stage": "待图片识别复核" if final_status == "待图片识别复核" else "生成报告",
            "progress": _labor_progress_completed("核对报告已生成。"),
            "asyncTask": {
                "status": "completed",
                "statusLabel": "完成",
                "message": "核对结果已生成。",
                "completedAt": utcnow_naive().isoformat(),
                **({"taskGenerationId": generation} if generation else {}),
            },
            "errorMessage": "",
            "errorCode": "",
            "nextAction": "",
            "files": files,
            "comparisonSummary": {**comparison["summary"], **conclusion},
            "comparisonRows": comparison["rows"],
            "candidateMatches": comparison.get("candidateMatches", []),
            "presentation": presentation,
            "warehouseComparison": warehouse_comparison,
            "extractionQuality": extraction_quality,
            "reconciliationDiagnostics": reconciliation_diagnostics,
            "costSummaries": cost_summaries,
            "aiCacheAudit": ai_cache_audit,
            "invoiceEvidenceAudit": invoice_evidence_audit,
            "pdfPageAudit": pdf_page_audit,
            "pdfPageAuditSummary": _summarize_pdf_page_audit(pdf_page_audit),
            "aiCacheReconciliationPreview": ai_cache_reconciliation_preview,
            "reocrPlan": reocr_plan,
            "nameMappingGovernance": name_mapping_governance,
            "combinedRowGovernance": combined_row_governance,
            "allocationGovernance": allocation_governance,
            "reviewQueues": review_queues,
            "profileGovernance": profile_governance,
            "correctionGovernance": correction_governance,
            "structureReconciliation": structure_reconciliation,
            "autoOcrCandidate": auto_ocr_candidate,
            "batchGuard": batch_guard,
            **review_state,
            "pdfExtractedRows": [row.to_dict() for row in pdf_rows],
            "excelRows": [row.to_dict() for row in excel_rows],
            "diffDownloadUrl": files["diffReport"]["downloadUrl"],
            "businessReportDownloadUrl": (files.get("businessReport") or {}).get("downloadUrl", ""),
        }
    result_updates["resultInputFingerprint"] = _labor_result_input_fingerprint(
        {**metadata, **result_updates}
    )
    updated, committed = compare_and_update_labor_metadata(
        run_id,
        expected_task_generation_id=generation,
        expected_fingerprint=input_fingerprint_at_start,
        fingerprint=_labor_result_input_fingerprint,
        updates=result_updates,
        conflict_updates=lambda current: {
            **_labor_invalidate_official_result(
                current,
                status="输入已变更，待重新核对",
                reason_code="inputs_changed_during_extraction",
                message="核对期间 PDF、Excel、字段映射或已启用治理状态发生变化；本次旧快照结果未发布，请重新执行整批核对。",
            ),
            "inputSnapshotRejected": True,
        },
    )
    if not committed:
        if generation and _labor_task_generation_id(updated) != generation:
            logger.info("[%s] 任务代次已更新，已拒绝发布旧任务结果", run_id)
            return {**updated, "_taskGenerationRejected": True}
        logger.warning("[%s] 核对输入在任务期间发生变化，已拒绝发布旧快照结果", run_id)
    return updated


def _labor_business_invoice_scope(metadata: dict, pdf_rows: list) -> str:
    invoice_names: list[str] = []
    for record in (metadata.get("files", {}) or {}).get("pdfInvoices", []) or []:
        original_filename = str(record.get("originalFilename") or "").strip()
        if original_filename and original_filename not in invoice_names:
            invoice_names.append(original_filename)
    for row in pdf_rows:
        source = str(getattr(row, "source_file", "") or "").strip()
        if source and source not in invoice_names:
            invoice_names.append(source)
    if not invoice_names:
        for record in (metadata.get("files", {}) or {}).get("pdfInvoices", []) or []:
            filename = str(record.get("filename") or "").strip()
            if filename and filename not in invoice_names:
                invoice_names.append(filename)
    if not invoice_names:
        return ""
    if len(invoice_names) <= 3:
        return "、".join(invoice_names)
    return f"{invoice_names[0]} 等 {len(invoice_names)} 个文件"


def _append_quality_issues(extraction_quality: dict, issues: list[str]) -> None:
    if not issues:
        return
    existing = extraction_quality.setdefault("issues", [])
    for issue in issues:
        if issue not in existing:
            existing.append(issue)


def _retry_low_confidence_rows(
    pdf_paths: list,
    low_confidence_rows: list,
    ai_config: dict,
    supplier: str,
    period_start: str,
    period_end: str,
    currency: str,
    expected_rows: list | None = None,
    supplier_profile_override=None,
) -> list | None:
    """对低置信度行做局部重试。

    从 low_confidence_rows 提取员工名单，用 retry_mode 重新抽取。
    返回合并后的新结果，失败时返回 None（降级到全量重试）。
    """
    if not low_confidence_rows:
        return None

    target_names = list({row["employee_name_raw"] for row in low_confidence_rows if row.get("employee_name_raw")})
    if not target_names:
        return None

    logger.info(f"局部重试: {len(target_names)} 个低置信度员工: {target_names[:10]}")
    try:
        retry_config = dict(ai_config)
        retry_config["cache_enabled"] = False
        retry_config["parallel_max_workers"] = 1
        retry_config["parallel_image_render_workers"] = 1

        fresh_paths = [Path(str(p)) for p in pdf_paths]
        for p in fresh_paths:
            if not p.exists():
                logger.warning(f"局部重试跳过: 文件不存在 {p}")
                return None

        retry_rows = extract_invoice_items(
            fresh_paths, retry_config,
            supplier=supplier, period_start=period_start, period_end=period_end, currency=currency,
            expected_rows=expected_rows,
            retry_mode=True,
            target_names=target_names,
            supplier_profile_override=supplier_profile_override,
        )
        if not retry_rows:
            logger.info("局部重试返回 0 条，降级到全量重试")
            return None

        logger.info(f"局部重试完成: {len(retry_rows)} 条")
        return retry_rows
    except Exception as exc:
        logger.warning(f"局部重试异常，降级到全量重试: {exc}")
        return None


def _retry_if_better(pdf_paths, pdf_rows, excel_rows, extraction_quality, comparison, **kwargs):
    manual_name_mapping = kwargs.pop("manual_name_mapping", None)
    retry_config = dict(AI_CONFIG)
    retry_config["cache_enabled"] = False
    # Serial execution for retry stability
    retry_config["parallel_max_workers"] = 1
    retry_config["parallel_image_render_workers"] = 1
    # Ensure PDF paths are fresh (not from a closed file handle)
    fresh_pdf_paths = [Path(str(p)) for p in pdf_paths]
    for p in fresh_pdf_paths:
        if not p.exists():
            logger.error(f"重试失败: PDF 文件不存在: {p}")
            extraction_quality["retryAttempted"] = True
            extraction_quality["retryApplied"] = False
            return pdf_rows, comparison, extraction_quality
    logger.info(f"重试抽取: {len(fresh_pdf_paths)} 个 PDF, cache_enabled=False, workers=1")
    try:
        retry_pdf_rows = extract_invoice_items(
            fresh_pdf_paths, retry_config,
            expected_rows=_expected_labor_rows(excel_rows), **kwargs,
        )
    except Exception as exc:
        logger.error(f"重试抽取异常，保留原始结果: {exc}", exc_info=True)
        extraction_quality["retryAttempted"] = True
        extraction_quality["retryApplied"] = False
        return pdf_rows, comparison, extraction_quality
    logger.info(f"重试抽取结果: {len(retry_pdf_rows)} 条")
    if not retry_pdf_rows:
        logger.warning("重试抽取返回 0 条，保留原始结果")
        extraction_quality["retryAttempted"] = True
        extraction_quality["retryApplied"] = False
        return pdf_rows, comparison, extraction_quality
    min_retry_rows = (len(pdf_rows) * 8 + 9) // 10
    if len(retry_pdf_rows) < min_retry_rows:
        logger.warning(
            f"重试抽取结果不完整: 原始 {len(pdf_rows)} 条，重试 {len(retry_pdf_rows)} 条，保留原始结果"
        )
        extraction_quality["retryAttempted"] = True
        extraction_quality["retryApplied"] = False
        return pdf_rows, comparison, extraction_quality
    if retry_pdf_rows:
        retry_comparison = compare_labor_items(
            retry_pdf_rows, excel_rows,
            amount_tolerance=AI_CONFIG["amount_tolerance"],
            hours_tolerance=AI_CONFIG["hours_tolerance"],
            confidence_threshold=AI_CONFIG["confidence_threshold"],
            manual_name_mapping=manual_name_mapping,
        )
        retry_quality = calculate_extraction_quality(retry_pdf_rows, retry_comparison["summary"], confidence_threshold=AI_CONFIG["confidence_threshold"])
        extraction_quality["retryAttempted"] = True
        if calculate_quality_score(retry_quality, retry_comparison["summary"]) < calculate_quality_score(extraction_quality, comparison["summary"]):
            retry_quality["retryAttempted"] = True
            retry_quality["retryApplied"] = True
            return retry_pdf_rows, retry_comparison, retry_quality
        extraction_quality["retryApplied"] = False
    return pdf_rows, comparison, extraction_quality


def _expected_labor_rows(excel_rows) -> list[dict]:
    return [
        {
            "employee_id": row.employee_id,
            "employee_name": row.employee_name_raw,
            "hours": row.hours,
            "amount": row.amount,
            "currency": row.currency,
            "source_ref": row.source_page_or_row,
        }
        for row in excel_rows
    ]


def _labor_quality_score(quality: dict, summary: dict) -> tuple:
    return (
        1 if quality.get("level") == "warning" else 0,
        len(quality.get("issues") or []),
        int(summary.get("exceptionCount") or 0),
        int(summary.get("unmatchedPdfCount") or 0) + int(summary.get("unmatchedExcelCount") or 0),
        abs(float(summary.get("amountDeltaTotal") or 0)),
    )


def _check_stale_extracting(metadata: dict) -> dict:
    """Mark run as failed if it's been stuck in '抽取中' for over 30 minutes."""
    if metadata.get("status") != "抽取中":
        return metadata
    from datetime import datetime as _dt, timedelta
    updated = metadata.get("updatedAt") or metadata.get("createdAt") or ""
    try:
        updated_dt = _dt.fromisoformat(updated)
    except (ValueError, TypeError):
        return metadata
    now = _dt.now(updated_dt.tzinfo) if updated_dt.tzinfo is not None else _dt.now()
    if now - updated_dt > timedelta(minutes=30):
        run_id = metadata.get("id")
        if run_id:
            try:
                metadata = update_labor_metadata(run_id, {
                    "status": "抽取失败",
                    "stage": "超时中断",
                    "errorMessage": "抽取超时（超过 30 分钟未完成）。请重新点击「抽取并核对」重试。",
                    "errorCode": "LABOR_EXTRACT_TIMEOUT",
                    "failureType": "system_interrupted",
                    "retryable": True,
                    "nextAction": "请重新点击「抽取并核对」重试；如连续失败，请联系管理员查看后台日志。",
                })
            except Exception:
                pass
    return metadata


def _labor_retryable_system_failure(*, message: str, stage: str, error_code: str) -> dict:
    return {
        "status": "抽取失败",
        "stage": stage,
        "failureType": "system_interrupted",
        "errorCode": error_code,
        "errorMessage": message,
        "retryable": True,
        "requiresReupload": False,
        "businessReviewStatus": "pending",
        "manualReviewRequired": True,
        "directPaymentAllowed": False,
        "requiresHumanReview": True,
        "nextAction": "无需重新上传材料。请重新点击「抽取并核对」重试；若连续失败，请联系管理员检查服务状态。",
    }


def _recover_stuck_labor_runs() -> None:
    """Recover in-process tasks that cannot survive a server restart."""
    if _uses_personal_labor_worker():
        # The durable queue and the user's desktop Worker outlive a Vercel
        # function instance. Treating every cold start as a task interruption
        # races with valid Worker result publication and can overwrite it.
        return
    try:
        rows = list_labor_metadata()
    except Exception as exc:
        logger.warning("跳过劳务核对启动恢复：%s", exc)
        return
    for metadata in rows:
        if metadata.get("status") != "抽取中":
            continue
        run_id = metadata.get("id")
        if run_id:
            try:
                update_labor_metadata(
                    run_id,
                    _labor_retryable_system_failure(
                        message="服务器已重启，抽取任务被中断。请重新点击「抽取并核对」重试。",
                        stage="系统中断",
                        error_code="LABOR_EXTRACT_INTERRUPTED",
                    ),
                )
            except Exception:
                pass


def _build_conclusion(warehouse_comparison: dict, comparison: dict, extraction_quality: dict, amount_tolerance: float = 0.05) -> dict:
    """Build conclusion level and message for the reconciliation result."""
    wc_summary = warehouse_comparison.get("summary", {})
    comp_summary = comparison.get("summary", {})
    total_passed = wc_summary.get("totalPassed", False)
    amount_delta_total = abs(wc_summary.get("amountDeltaTotal", 0))
    pdf_amount_total = abs(wc_summary.get("pdfAmountTotal", 0))
    excel_amount_total = abs(wc_summary.get("excelAmountTotal", 0))
    max_amount = max(pdf_amount_total, excel_amount_total, 1.0)
    amount_delta_pct = amount_delta_total / max_amount * 100

    pdf_employee_count = comp_summary.get("pdfEmployeeCount", 0)
    excel_employee_count = comp_summary.get("excelEmployeeCount", 0)
    amount_diff_count = comp_summary.get("amountDiffCount", 0)
    low_confidence_count = comp_summary.get("lowConfidenceCount", 0)
    exception_count = comp_summary.get("exceptionCount", 0)

    # 结论级别判定
    if extraction_quality.get("level") == "critical":
        conclusion_level = "critical"
        conclusion_message = "抽取质量存在严重问题，必须人工复核"
    elif extraction_quality.get("level") == "warning" and str(extraction_quality.get("message") or "").strip():
        conclusion_level = "warning"
        conclusion_message = str(extraction_quality.get("message")).strip()
    elif low_confidence_count > 0:
        conclusion_level = "warning"
        conclusion_message = "存在低置信度抽取，需人工复核"
    elif pdf_employee_count <= 0 or excel_employee_count <= 0 or pdf_employee_count != excel_employee_count:
        conclusion_level = "warning"
        conclusion_message = "员工明细数量未闭合，需人工复核"
    elif exception_count > 0 or amount_diff_count > 0:
        conclusion_level = "warning"
        conclusion_message = f"{exception_count or amount_diff_count}项员工/金额差异需关注"
    elif extraction_quality.get("level") == "warning":
        conclusion_level = "warning"
        conclusion_message = "核对存在风险，请查看质量提示"
    elif total_passed:
        conclusion_level = "pass"
        conclusion_message = "仓库总金额核对通过"
    else:
        conclusion_level = "warning"
        conclusion_message = f"仓库总金额差异 ${amount_delta_total:.2f} ({amount_delta_pct:.2f}%)"

    # 计算不在本批发票人数（使用实际的"Excel有PDF无"行数，而非简单减法）
    comparison_rows = comparison.get("rows", [])
    not_in_invoice_count = sum(1 for r in comparison_rows if r.get("matchStatus") == "Excel有PDF无")

    return {
        "conclusionLevel": conclusion_level,
        "conclusionMessage": conclusion_message,
        "notInInvoiceCount": not_in_invoice_count,
    }


def _labor_detail_coverage_complete(reconciliation_diagnostics: dict) -> bool | None:
    signals = reconciliation_diagnostics.get("signals") if isinstance(reconciliation_diagnostics, dict) else {}
    coverage = signals.get("pdfDetailCoverage") if isinstance(signals, dict) else {}
    if not isinstance(coverage, dict) or not coverage:
        return None
    invoice_count = int(coverage.get("invoiceFileCount") or 0)
    detail_count = int(coverage.get("detailFileCount") or 0)
    return invoice_count > 0 and detail_count >= invoice_count


@app.get("/api/labor/runs/{run_id}/download/{filename}")
def download_labor_file(run_id: str, filename: str) -> Response:
    if labor_postgres_state_enabled():
        metadata = _labor_metadata_or_404(run_id)
        files = metadata.get("files") if isinstance(metadata.get("files"), dict) else {}
        safe_filename = Path(str(filename).replace("\\", "/")).name
        for key in ("diffReport", "businessReport", "projectionReport", "governanceReport"):
            record = files.get(key) if isinstance(files.get(key), dict) else {}
            if (
                Path(str(record.get("filename") or "").replace("\\", "/")).name != safe_filename
                or not str(record.get("objectKey") or "").strip()
                or record.get("storageVerified") is not True
            ):
                continue
            try:
                signed = create_labor_supabase_signed_download(
                    str(record["objectKey"]),
                    filename=safe_filename,
                    expires_in=120,
                )
            except Exception as exc:  # noqa: BLE001 - do not expose storage internals.
                raise _labor_p1_storage_http_error() from exc
            return RedirectResponse(
                url=str(signed["signedUrl"]),
                status_code=307,
                headers={"cache-control": "no-store", "referrer-policy": "no-referrer"},
            )
    if _labor_p1_required():
        raise HTTPException(
            status_code=409,
            detail=_labor_request_error(
                message="当前报告缺少已验证的私有存储证据，P1 不允许从服务本地回退下载。",
                error_code="LABOR_P1_PRIVATE_REPORT_REQUIRED",
                next_action="请重新生成核对报告；持续失败请联系管理员检查私有存储。",
                requires_human_review=True,
            ),
        )
    try:
        run_dir = get_labor_run_dir(run_id)
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)
    path = _resolve_labor_download_path(run_dir, filename)
    if not path.exists():
        restore_attempted = False
        restore_succeeded = False
        if labor_persistent_storage_enabled():
            restore_attempted = True
            try:
                restore_succeeded = sync_labor_run_from_persistent(run_id, run_dir)
            except Exception as exc:  # noqa: BLE001 - download path must return business-safe errors.
                logger.warning("Failed to restore labor report for download: run_id=%s file=%s error=%s", run_id, Path(filename).name, exc)
        elif labor_blob_storage_enabled():
            restore_attempted = True
            try:
                restore_succeeded = sync_labor_run_from_blob(run_id, run_dir)
            except Exception as exc:  # noqa: BLE001 - keep legacy blob restore monkeypatchable in tests.
                logger.warning("Failed to restore labor report for download: run_id=%s file=%s error=%s", run_id, Path(filename).name, exc)
        path = _resolve_labor_download_path(run_dir, filename)
        if not path.exists():
            if restore_attempted and not restore_succeeded:
                raise HTTPException(
                    status_code=503,
                    detail=_labor_request_error(
                        message="报告文件暂时无法恢复。",
                        error_code="LABOR_REPORT_RESTORE_FAILED",
                        retryable=True,
                        next_action="请稍后重试下载；若连续失败，请联系管理员检查 UAT 文件持久化状态。",
                    ),
                )
            raise HTTPException(
                status_code=404,
                detail=_labor_request_error(
                    message="报告文件不存在或已被清理。",
                    error_code="LABOR_REPORT_FILE_MISSING",
                    requires_human_review=True,
                    next_action="请重新生成报告；如果批次已完成但仍无法下载，请联系管理员恢复该批次报告文件。",
                ),
            )
    try:
        metadata = load_labor_metadata(run_dir)
        owner_user_id = str(metadata.get("ownerUserId") or "local-default")
    except Exception:
        owner_user_id = "local-default"
    append_labor_audit_event(
        _labor_audit_path(),
        action="report_downloaded",
        run_id=run_id,
        owner_user_id=owner_user_id,
        actor_user_id=owner_user_id,
        outcome="success",
    )
    return FileResponse(path, filename=path.name)


def _resolve_labor_download_path(run_dir: Path, filename: str) -> Path:
    requested = Path(filename).name
    direct = run_dir / requested
    if direct.exists():
        return direct
    try:
        metadata = load_labor_metadata(run_dir)
    except FileNotFoundError:
        return direct
    for record in (metadata.get("files") or {}).values():
        if isinstance(record, dict):
            candidate_name = Path(str(record.get("filename") or "")).name
            candidate_path = Path(str(record.get("path") or ""))
            if candidate_name == requested and candidate_path.exists():
                return candidate_path
    return direct


@app.post("/api/runs/{run_id}/finalize")
async def finalize_run(
    run_id: str,
    confirmation_file: UploadFile = File(...),
) -> dict:
    if not confirmation_file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="确认结果请上传 Excel 文件（.xlsx 或 .xlsm）。")
    try:
        run_dir = get_run_dir(run_id)
        metadata = load_metadata(run_dir)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="批次不存在。") from exc

    initial_path = Path(metadata["files"]["initialResult"]["path"])
    if not initial_path.exists():
        raise HTTPException(status_code=404, detail="批次初算结果不存在，无法生成最终结果。")

    month = _coerce_month(metadata.get("month")) or 0
    run_sequence = int(metadata.get("runSequence") or 1)
    confirmation_path = await _save_upload_to(confirmation_file, run_dir / _recruitment_output_name(month, run_sequence, "确认结果"))
    final_path = run_dir / _recruitment_output_name(month, run_sequence, "最终结果")
    try:
        build_final_workbook(initial_path, confirmation_path, final_path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"生成最终结果失败：{exc}") from exc

    final_payload = _final_calculation_payload(final_path, metadata)
    save_table_data(run_dir, build_final_table_data(run_id, final_path, final_payload["month"]))

    files = dict(metadata.get("files", {}))
    files["confirmation"] = attach_file_record(run_id, confirmation_path, "确认结果")
    files["finalResult"] = attach_file_record(run_id, final_path, "最终结果")
    updated = update_metadata(
        run_id,
        {
            "status": "已最终确认",
            "displayName": _recruitment_run_display_name(final_payload["month"], run_sequence, "已最终确认"),
            "files": files,
            "finalDownloadUrl": files["finalResult"]["downloadUrl"],
            **final_payload,
        },
    )
    return updated


@app.post("/api/runs/{run_id}/compare")
async def compare_run(
    run_id: str,
    offline_file: UploadFile = File(...),
) -> dict:
    if not offline_file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="线下复核表请上传 Excel 文件（.xlsx 或 .xlsm）。")
    try:
        run_dir = get_run_dir(run_id)
        metadata = load_metadata(run_dir)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="批次不存在。") from exc

    source_record = metadata.get("files", {}).get("finalResult") or metadata.get("files", {}).get("initialResult")
    if not source_record:
        raise HTTPException(status_code=404, detail="批次结果不存在，无法生成差异报告。")

    month = _coerce_month(metadata.get("month")) or 0
    run_sequence = int(metadata.get("runSequence") or 1)
    offline_path = await _save_upload_to(offline_file, run_dir / _recruitment_output_name(month, run_sequence, "线下复核表"))
    diff_path = run_dir / _recruitment_output_name(month, run_sequence, "差异报告")
    try:
        metrics = build_difference_report(Path(source_record["path"]), offline_path, diff_path)
        merge_diff_rows(run_dir, metrics)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"生成差异报告失败：{exc}") from exc

    files = dict(metadata.get("files", {}))
    files["offlineReview"] = attach_file_record(run_id, offline_path, "线下复核表")
    files["diffReport"] = attach_file_record(run_id, diff_path, "差异报告")
    updated = update_metadata(
        run_id,
        {
            "status": "已生成差异报告",
            "displayName": _recruitment_run_display_name(month, run_sequence, "已生成差异报告"),
            "files": files,
            "diffMetrics": metrics,
            "diffDownloadUrl": files["diffReport"]["downloadUrl"],
        },
    )
    return updated


@app.post("/api/finalize")
async def finalize_bonus(
    initial_result_file: UploadFile = File(...),
    confirmation_file: UploadFile = File(...),
) -> dict:
    for upload, label in ((initial_result_file, "初算结果"), (confirmation_file, "确认结果")):
        if not upload.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail=f"{label}请上传 Excel 文件（.xlsx 或 .xlsm）。")

    initial_path = await _save_upload(initial_result_file)
    confirmation_path = await _save_upload(confirmation_file)
    try:
        output_path = _output_path(initial_result_file.filename, suffix="最终结果")
        build_final_workbook(initial_path, confirmation_path, output_path)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"生成最终结果失败：{exc}") from exc
    finally:
        initial_path.unlink(missing_ok=True)
        confirmation_path.unlink(missing_ok=True)

    return {
        "filename": output_path.name,
        "downloadUrl": f"/api/download/{output_path.name}",
    }


@app.get("/api/download/{filename}")
def download(filename: str) -> FileResponse:
    path = EXPORT_DIR / filename
    if not path.exists():
        raise HTTPException(status_code=404, detail="文件不存在或已被清理。")
    return FileResponse(path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/runs/{run_id}/download/{filename}")
def download_run_file(run_id: str, filename: str) -> FileResponse:
    try:
        run_dir = get_run_dir(run_id)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="批次不存在。") from exc
    path = run_dir / Path(filename).name
    if not path.exists():
        raise HTTPException(status_code=404, detail="文件不存在或已被清理。")
    return FileResponse(path, filename=path.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/template")
def download_template() -> FileResponse:
    if not DEFAULT_IMPORT_TEMPLATE.exists():
        raise HTTPException(status_code=404, detail="模板文件不存在。")
    return FileResponse(
        DEFAULT_IMPORT_TEMPLATE,
        filename=DEFAULT_IMPORT_TEMPLATE.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Cache-Control": "no-store"},
    )


def _output_path(original_name: str, suffix: str = "平台计算结果") -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stem = Path(original_name).stem.replace(" ", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return EXPORT_DIR / f"{stem}_{suffix}_{timestamp}.xlsx"


async def _save_upload(file: UploadFile) -> Path:
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(await file.read())
        return Path(tmp.name)


async def _save_upload_to(
    file: UploadFile,
    path: Path,
    *,
    max_bytes: int | None = None,
    limit_code: str = "LABOR_FILE_SIZE_LIMIT_EXCEEDED",
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    written = 0
    try:
        with path.open("wb") as destination:
            while True:
                chunk = await file.read(1024 * 1024)
                if not chunk:
                    break
                written += len(chunk)
                if max_bytes is not None and written > max_bytes:
                    raise LaborResourceLimitError(
                        limit_code,
                        f"文件 {file.filename} 超过允许大小。",
                        limit=max_bytes,
                    )
                destination.write(chunk)
    except Exception:
        path.unlink(missing_ok=True)
        raise
    return path


def _build_labor_direct_upload_item(run_id: str, item: Any, *, group: str) -> dict:
    if not isinstance(item, dict):
        raise ValueError("上传文件清单格式不正确。")
    original_name = Path(str(item.get("name") or "")).name
    if not original_name:
        raise ValueError("上传文件缺少文件名。")
    lower_name = original_name.lower()
    if group == "pdfInvoices" and not lower_name.endswith(".pdf"):
        raise ValueError("供应商发票请上传 PDF 文件。")
    if group == "workbooks" and not lower_name.endswith((".xlsx", ".xlsm", ".xls")):
        raise ValueError(f"线下账单请上传 Excel 文件（.xlsx / .xlsm / .xls）。收到：{original_name}")
    filename = safe_labor_storage_filename(original_name, "direct")
    relative_path = filename
    signed_upload = create_labor_supabase_signed_upload(run_id, relative_path)
    return {
        "group": group,
        "filename": filename,
        "originalFilename": original_name,
        "relativePath": relative_path,
        "objectPath": signed_upload["objectPath"],
        "signedUrl": signed_upload["signedUrl"],
        "size": int(item.get("size") or 0),
        "contentType": str(item.get("type") or "application/octet-stream"),
    }


def _validate_labor_direct_upload_relative_path(value: Any) -> Path:
    relative = str(value or "").replace("\\", "/").strip().lstrip("/")
    path = Path(relative)
    if (
        not relative
        or path.is_absolute()
        or ".." in path.parts
        or "/" in relative
        or "\\" in relative
    ):
        raise ValueError("直传文件路径不正确。")
    return path


def _copy_material_plan_files(run_id: str, run_dir: Path, root_path: Path, plan: dict) -> tuple[dict, list[dict]]:
    files: dict = {}
    copied_sources: list[dict] = []
    pdf_records = []
    for relative in plan.get("uploadPlan", {}).get("pdfFiles", []) or []:
        copied_path = _copy_one_material_file(run_dir, root_path, relative)
        pdf_records.append(attach_labor_file(run_id, copied_path, "PDF发票"))
        copied_sources.append({"kind": "pdf", "relativePath": relative, "copiedPath": str(copied_path)})
    workbook_records = []
    for relative in plan.get("uploadPlan", {}).get("workbookFiles", []) or []:
        copied_path = _copy_one_material_file(run_dir, root_path, relative)
        workbook_records.append(attach_labor_file(run_id, copied_path, "线下账单"))
        copied_sources.append({"kind": "workbook", "relativePath": relative, "copiedPath": str(copied_path)})
    if not pdf_records:
        raise ValueError("材料批次缺少可复制的 PDF 发票。")
    if not workbook_records:
        raise ValueError("材料批次缺少可复制的 Excel 账单。")
    files["pdfInvoices"] = pdf_records
    files["workbooks"] = workbook_records
    files["workbook"] = workbook_records[0]
    return files, copied_sources


def _copy_one_material_file(run_dir: Path, root_path: Path, relative_path: str) -> Path:
    source = (root_path / relative_path).resolve()
    try:
        source.relative_to(root_path)
    except ValueError as exc:
        raise ValueError(f"材料文件路径不在参考目录内: {relative_path}") from exc
    if not source.exists() or not source.is_file():
        raise FileNotFoundError(f"材料文件不存在: {relative_path}")
    target = run_dir / safe_labor_filename(source.name)
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    if source.suffix.lower() == ".pdf":
        _copy_material_ai_cache(source, target)
    return target


def _copy_material_ai_cache(source: Path, target: Path) -> None:
    cache_dir = source.parent / ".ai_extract_cache"
    if not cache_dir.exists() or not cache_dir.is_dir():
        return
    target_cache_dir = target.parent / ".ai_extract_cache"
    copied = 0
    for cache_path in cache_dir.glob(f"{source.stem}_*.json"):
        if not cache_path.is_file():
            continue
        target_name = f"{target.stem}{cache_path.name[len(source.stem):]}"
        target_cache_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(cache_path, target_cache_dir / target_name)
        copied += 1
    if copied:
        logger.info(f"已复制材料历史图片识别记录: {source.name} -> {target.name}, {copied} 个记录文件")


def _safe_output_name(original_name: str, suffix: str) -> str:
    stem = Path(original_name).stem.replace(" ", "_")
    stem = "".join(char if char.isalnum() or char in "_-" else "_" for char in stem).strip("_") or "workbook"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    return f"{stem}_{suffix}_{timestamp}.xlsx"


def _next_recruitment_run_sequence(month: int) -> int:
    same_month = [row for row in list_run_metadata() if _coerce_month(row.get("month")) == month]
    explicit_sequences = [
        int(row.get("runSequence"))
        for row in same_month
        if str(row.get("runSequence") or "").isdigit()
    ]
    if explicit_sequences:
        return max(explicit_sequences) + 1
    return len(same_month) + 1


def _recruitment_output_name(month: int, run_sequence: int, suffix: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    month_label = _recruitment_month_label(month, compact=False)
    sequence_label = f"第{max(int(run_sequence or 1), 1)}次"
    safe_suffix = "".join(char if char.isalnum() or char in "_-" else "_" for char in suffix).strip("_") or "结果"
    return f"招聘奖金核算_{month_label}_{sequence_label}_{safe_suffix}_{timestamp}.xlsx"


def _recruitment_run_display_name(month: int, run_sequence: int, status: str = "") -> str:
    parts = [
        _recruitment_month_label(month, compact=False),
        f"第{max(int(run_sequence or 1), 1)}次核算",
    ]
    if status:
        parts.append(str(status))
    return " · ".join(parts)


def _recruitment_run_short_code(month: int, run_sequence: int) -> str:
    return f"{_recruitment_month_label(month, compact=True)}-第{max(int(run_sequence or 1), 1)}次"


def _recruitment_month_label(month: int, *, compact: bool) -> str:
    month_int = _coerce_month(month)
    if not month_int:
        return "未知月份"
    year = month_int // 100
    month_no = month_int % 100
    return f"{year}{month_no:02d}" if compact else f"{year}年{month_no:02d}月"


def _calculation_payload(result) -> dict:
    recruitment_total = round(sum(row.get("合计发放", 0) for row in result.recruitment_summary), 2)
    referral_total = round(sum(row.get("合计发放", 0) for row in result.referral_summary), 2)
    pending_total = round(sum(row.get("建议发放金额", 0) for row in result.pending_confirmations), 2)
    preview = [
        {
            "姓名": detail.name,
            "工号": detail.employee_no,
            "职级": detail.grade,
            "ABC类别": detail.category,
            "招聘渠道": detail.channel,
            "招聘人入职1月奖金": detail.recruiter_1m_bonus,
            "内推入职1月奖金": detail.referral_1m_bonus,
            "异常提示": "；".join(detail.exceptions),
        }
        for detail in result.details[:MAX_PREVIEW_ROWS]
    ]
    return {
        "month": result.month,
        "importedRows": len(result.details),
        "recruitmentTotal": recruitment_total,
        "referralTotal": referral_total,
        "exceptionCount": len(result.exceptions),
        "pendingCount": len(result.pending_confirmations),
        "pendingTotal": pending_total,
        "detailPreview": preview,
        "pendingConfirmations": result.pending_confirmations[:MAX_PREVIEW_ROWS],
        "exceptions": result.exceptions[:MAX_PREVIEW_ROWS],
    }


def _final_calculation_payload(final_path: Path, fallback_metadata: dict) -> dict:
    workbook = load_workbook(final_path, data_only=True, read_only=True)
    try:
        fallback_month = _coerce_month(fallback_metadata.get("month")) or 0
        month = (
            _first_summary_month(workbook, "最终招聘奖金汇总")
            or _first_summary_month(workbook, "最终内推奖金汇总")
            or _intro_value(workbook, "核算月份")
            or fallback_month
        )
        detail_rows = _workbook_rows(workbook, "招聘奖金明细", skip_total=False)
        exception_rows = _workbook_rows(workbook, "异常清单", skip_total=False)
        preview = [
            {
                "姓名": row.get("姓名", ""),
                "工号": row.get("工号", ""),
                "职级": row.get("职级", ""),
                "ABC类别": row.get("ABC类别", ""),
                "招聘渠道": row.get("招聘渠道", ""),
                "招聘人入职1月奖金": row.get("招聘人入职1月奖金", 0),
                "内推入职1月奖金": row.get("内推入职1月奖金", 0),
                "异常提示": row.get("异常提示", ""),
            }
            for row in detail_rows[:MAX_PREVIEW_ROWS]
        ]
        return {
            "month": month,
            "importedRows": len(detail_rows),
            "recruitmentTotal": _summary_total(workbook, "最终招聘奖金汇总"),
            "referralTotal": _summary_total(workbook, "最终内推奖金汇总"),
            "exceptionCount": len(exception_rows),
            "pendingCount": 0,
            "pendingTotal": 0,
            "detailPreview": preview,
            "pendingConfirmations": [],
            "exceptions": exception_rows[:MAX_PREVIEW_ROWS],
        }
    finally:
        workbook.close()


def _workbook_rows(workbook, sheet_name: str, skip_total: bool = True) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    rows: list[dict] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = {str(header).strip(): values[index] for index, header in enumerate(headers) if header and index < len(values)}
        if not any(value not in (None, "") for value in row.values()):
            continue
        if skip_total and _is_workbook_total_row(row):
            continue
        rows.append(row)
    return rows


def _is_workbook_total_row(row: dict) -> bool:
    return any(str(value or "").strip() in {"合计", "总计"} for value in row.values())


def _summary_total(workbook, sheet_name: str) -> float:
    total = sum(_float_value(row.get("合计发放")) for row in _workbook_rows(workbook, sheet_name, skip_total=True))
    return round(total, 2)


def _first_summary_month(workbook, sheet_name: str) -> int | None:
    for row in _workbook_rows(workbook, sheet_name, skip_total=True):
        month = _coerce_month(row.get("核算月份"))
        if month:
            return month
    return None


def _intro_value(workbook, label: str) -> int | None:
    if "计算说明" not in workbook.sheetnames:
        return None
    sheet = workbook["计算说明"]
    for key, value in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        if str(key or "").strip() == label:
            return _coerce_month(value)
    return None


def _coerce_month(value) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.year * 100 + value.month
    if isinstance(value, (int, float)):
        return int(value)
    digits = "".join(char for char in str(value) if char.isdigit())
    if len(digits) >= 6:
        return int(digits[:6])
    return None


def _float_value(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _labor_metadata_or_404(run_id: str) -> dict:
    try:
        return load_labor_metadata(get_labor_run_dir(run_id))
    except FileNotFoundError as exc:
        _raise_labor_run_missing(exc)


def _labor_workbook_path(metadata: dict) -> Path:
    """返回第一个 workbook 文件路径（兼容旧逻辑）"""
    paths = _labor_workbook_paths(metadata)
    return paths[0]


def _labor_excel_rows_from_metadata(metadata: dict) -> list:
    sheet_name = str(metadata.get("workbookSheet") or "").strip()
    mapping = metadata.get("excelMapping") or {}
    if not sheet_name or not mapping:
        raise HTTPException(status_code=400, detail="请先确认 Excel 工作表和字段映射。")
    excel_rows = []
    for path in _labor_workbook_paths(metadata):
        workbook_sheet, workbook_mapping = _labor_workbook_mapping(metadata, path)
        excel_rows.extend(read_workbook_rows(path, workbook_sheet, workbook_mapping))
    return _aggregate_excel_rows(excel_rows)


def _labor_workbook_mapping(metadata: dict, path: Path) -> tuple[str, dict]:
    for item in metadata.get("workbookMappings") or []:
        if not isinstance(item, dict) or str(item.get("filename") or "") != path.name:
            continue
        item_sheet = str(item.get("sheetName") or "").strip()
        item_mapping = item.get("mapping") if isinstance(item.get("mapping"), dict) else {}
        if item_sheet and item_mapping:
            return item_sheet, item_mapping
    return str(metadata.get("workbookSheet") or "").strip(), metadata.get("excelMapping") or {}


def _labor_pdf_path_for_source(metadata: dict, source_file: str) -> Path:
    files_meta = metadata.get("files", {}) if isinstance(metadata.get("files"), dict) else {}
    records = files_meta.get("pdfInvoices") or []
    requested = Path(source_file).name
    requested_stem = Path(requested).stem
    candidates: list[Path] = []
    for record in records:
        path = Path(str(record.get("path") or ""))
        if not path.exists():
            continue
        if path.name == requested or path.stem == requested_stem or path.name.startswith(f"{requested_stem}_"):
            return path
        candidates.append(path)
    run_id = str(metadata.get("id") or "")
    run_dir = get_labor_run_dir(run_id) if run_id else None
    if run_dir:
        direct = run_dir / requested
        if direct.exists():
            return direct
        matches = sorted(path for path in run_dir.glob(f"{requested_stem}_*.pdf") if path.is_file())
        if matches:
            return matches[0]
    known = ", ".join(path.name for path in candidates[:5])
    raise HTTPException(status_code=404, detail=f"未找到可回放缓存的 PDF：{requested}。当前批次 PDF：{known or '无'}")


def _cache_replay_task_from_candidate(candidate: dict) -> dict:
    return {
        "sourceFile": candidate.get("sourceFile", ""),
        "warehouseId": candidate.get("warehouseId", ""),
        "expectedExcelAmount": candidate.get("expectedExcelAmount", 0),
        "amountDelta": candidate.get("amountDelta", 0),
        "amountTolerance": candidate.get("amountTolerance", AI_CONFIG["amount_tolerance"]),
        "diagnostics": candidate.get("diagnostics", {}),
        "confirmationGate": "缓存候选需通过员工级回放并人工确认后才能用于当前批次。",
    }


def _group_reocr_candidate_rows_by_scope(rows: list[dict]) -> dict[tuple[str, str], list[dict]]:
    grouped: dict[tuple[str, str], list[dict]] = {}
    for row in rows:
        source_file = str(row.get("sourceFile") or row.get("source_file") or "").strip()
        warehouse_id = str(row.get("warehouseId") or row.get("warehouse_id") or "").strip()
        if not source_file:
            continue
        grouped.setdefault((Path(source_file).name, warehouse_id), []).append(row)
    return grouped


def _build_reocr_upload_coverage(metadata: dict, grouped_rows: dict[tuple[str, str], list[dict]]) -> dict:
    planned = []
    for task in ((metadata.get("reocrPlan") if isinstance(metadata.get("reocrPlan"), dict) else {}) or {}).get("tasks", []) or []:
        if not isinstance(task, dict):
            continue
        source_file = str(task.get("sourceFile") or "").strip()
        warehouse_id = str(task.get("warehouseId") or "").strip()
        if source_file:
            planned.append({"sourceFile": source_file, "warehouseId": warehouse_id})
    planned_keys = {
        (_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or ""))
        for item in planned
    }
    uploaded = [
        {"sourceFile": source_file, "warehouseId": warehouse_id, "rowCount": len(rows)}
        for (source_file, warehouse_id), rows in grouped_rows.items()
    ]
    uploaded_keys = {
        (_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or ""))
        for item in uploaded
    }
    missing_tasks = [
        item
        for item in planned
        if (_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or "")) not in uploaded_keys
    ]
    extra_scopes = [
        item
        for item in uploaded
        if planned_keys and (_safe_scope_key(item["sourceFile"]), str(item.get("warehouseId") or "")) not in planned_keys
    ]
    return {
        "plannedTaskCount": len(planned),
        "uploadedScopeCount": len(uploaded),
        "coveredTaskCount": len(planned_keys & uploaded_keys) if planned_keys else len(uploaded_keys),
        "missingTaskCount": len(missing_tasks),
        "extraScopeCount": len(extra_scopes),
        "coverageComplete": not missing_tasks if planned else True,
        "missingTasks": missing_tasks,
        "extraScopes": extra_scopes,
        "uploadedScopes": uploaded,
    }


def _reocr_task_for_batch_candidate(metadata: dict, excel_rows: list, source_file: str, warehouse_id: str) -> dict:
    plan = metadata.get("reocrPlan") if isinstance(metadata.get("reocrPlan"), dict) else {}
    for task in plan.get("tasks") or []:
        if not isinstance(task, dict):
            continue
        if Path(str(task.get("sourceFile") or "")).name == Path(source_file).name and str(task.get("warehouseId") or "") == str(warehouse_id or ""):
            return _cache_replay_task_from_candidate(task)
    scoped_rows = [row for row in excel_rows if not warehouse_id or str(row.warehouse_id or "").strip() == warehouse_id]
    if not scoped_rows:
        raise ValueError("该 SourceFile/WarehouseId 未匹配到账单员工行。")
    return {
        "sourceFile": Path(source_file).name,
        "warehouseId": warehouse_id,
        "expectedExcelAmount": round(sum(float(row.amount or 0) for row in scoped_rows), 2),
        "amountDelta": 0,
        "amountTolerance": AI_CONFIG["amount_tolerance"],
        "diagnostics": {},
        "confirmationGate": "批量上传候选需通过员工级回放并人工确认后才能用于当前批次。",
    }


def _labor_workbook_paths(metadata: dict) -> list[Path]:
    """返回所有 workbook 文件路径，支持多文件上传"""
    files_meta = metadata.get("files", {})
    # 优先使用 workbooks 列表（新格式）
    records = files_meta.get("workbooks") or []
    if not records:
        # 兼容旧格式：单个 workbook 字段
        single = files_meta.get("workbook")
        if single:
            records = [single]
    if not records:
        raise HTTPException(status_code=400, detail="请先上传线下账单 Excel。")
    paths = []
    for rec in records:
        p = Path(rec.get("path", ""))
        if not p.exists():
            raise HTTPException(status_code=404, detail=f"线下账单文件不存在：{p.name}")
        paths.append(p)
    return paths


# ============================================================
# CHINA EMPLOYEE PAYROLL API  /api/china-employee-payroll/*
# ============================================================

def _path_is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def _resolve_china_employee_payroll_runs_dir() -> Path:
    candidate = OUTPUT_DIR / "china_employee_payroll_runs"
    if _path_is_relative_to(candidate, PROJECT_ROOT):
        return Path(gettempdir()) / "sigma_workbench_china_employee_payroll_runs"
    return candidate


CHINA_EMPLOYEE_PAYROLL_RUNS_DIR = _resolve_china_employee_payroll_runs_dir()
CHINA_EMPLOYEE_PAYROLL_RUNS_DIR.mkdir(parents=True, exist_ok=True)


def _compact_china_employee_payroll_result(result: dict) -> list[dict]:
    return [
        {key: value for key, value in row.items() if key != "daily"}
        for row in result.get("results", [])
    ]


@app.post("/api/china-employee-payroll/meal-allowance")
async def calculate_china_employee_meal_allowance(
    attendance_files: list[UploadFile] = File(...),
    source_type: str = Form("hr"),
) -> dict:
    source_type = (source_type or "hr").strip().lower()
    if source_type not in {"hr", "wx"}:
        raise HTTPException(400, "考勤来源类型不支持，请选择人事系统考勤或WX技术部考勤。")
    if not attendance_files:
        raise HTTPException(400, "请上传考勤记录 Excel。")
    for file in attendance_files:
        if not file.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(400, "请上传 Excel 文件（.xlsx / .xlsm）。")

    run_id = f"china_employee_payroll_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    run_dir = CHINA_EMPLOYEE_PAYROLL_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=False)
    paths = []
    for file in attendance_files:
        saved_name = safe_payroll_filename(file.filename)
        path = run_dir / saved_name
        path.write_bytes(await file.read())
        paths.append(path)

    try:
        parsed = parse_wx_attendance_workbooks(paths) if source_type == "wx" else parse_attendance_workbooks(paths)
        result = calculate_meal_allowance(parsed)
    except Exception as exc:
        shutil.rmtree(run_dir, ignore_errors=True)
        raise HTTPException(400, f"考勤记录解析或核算失败：{exc}") from exc

    metadata = {
        "runId": run_id,
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFiles": [path.name for path in paths],
        "sourceType": source_type,
        "sourceLabel": result["summary"].get("sourceLabel", ""),
        "summary": result["summary"],
        "warnings": result["warnings"],
    }
    (run_dir / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    (run_dir / "result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "runId": run_id,
        "summary": result["summary"],
        "results": _compact_china_employee_payroll_result(result),
        "files": result["files"],
        "warnings": result["warnings"],
    }


def _china_employee_payroll_month_label(summary: dict) -> str:
    raw_date = summary.get("dateStart") or ""
    try:
        parsed = datetime.fromisoformat(raw_date)
    except ValueError:
        return ""
    return f"{parsed.year}年{parsed.month}月"


def _build_china_employee_meal_allowance_export(run_dir: Path, run_id: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    metadata_path = run_dir / "metadata.json"
    result_path = run_dir / "result.json"
    if not metadata_path.exists() or not result_path.exists():
        raise HTTPException(status_code=404, detail="核算结果不存在，请重新上传并核算。")

    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    result = json.loads(result_path.read_text(encoding="utf-8"))
    summary = result.get("summary", {})
    source_type = metadata.get("sourceType") or summary.get("sourceType") or "hr"
    month_label = _china_employee_payroll_month_label(summary)
    payable_rows = [row for row in result.get("results", []) if row.get("amount", 0) > 0]

    workbook = Workbook()
    result_sheet = workbook.active
    result_sheet.title = "餐补核算结果"
    source_sheet_name = "WX技术部考勤源" if source_type == "wx" else "人事系统考勤源"
    source_sheet = workbook.create_sheet(source_sheet_name)

    green_fill = PatternFill("solid", fgColor="EAF8F0")
    total_fill = PatternFill("solid", fgColor="FFF4D6")
    header_fill = PatternFill("solid", fgColor="D8E8FF")
    border = Border(bottom=Side(style="thin", color="D7E3F4"))
    header_font = Font(color="17324D", bold=True)
    body_font = Font(color="24364B")

    headers = ["核算月份", "工号", "姓名", "人员状态", "二级组织", "三级组织", "四级组织", "补贴天数", "餐补金额"]
    header_row = 1
    for column, header in enumerate(headers, start=1):
        cell = result_sheet.cell(header_row, column, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index, row in enumerate(payable_rows, start=header_row + 1):
        values = [
            month_label,
            row.get("employeeId", ""),
            row.get("employeeName", ""),
            row.get("status", ""),
            row.get("secondOrg", ""),
            row.get("thirdOrg", ""),
            row.get("fourthOrg", ""),
            row.get("payableDays", 0),
            row.get("amount", 0),
        ]
        for column, value in enumerate(values, start=1):
            cell = result_sheet.cell(row_index, column, value)
            cell.font = body_font
            cell.fill = green_fill if row_index % 2 else PatternFill("solid", fgColor="FFFFFF")
            cell.border = border
            if column in (8, 9):
                cell.alignment = Alignment(horizontal="right")

    total_row = header_row + len(payable_rows) + 1
    total_values = ["合计", "", "", "", "", "", "", summary.get("payableDayCount", 0), summary.get("totalAmount", 0)]
    for column, value in enumerate(total_values, start=1):
        cell = result_sheet.cell(total_row, column, value)
        cell.fill = total_fill
        cell.font = header_font
        cell.border = border
        if column in (8, 9):
            cell.alignment = Alignment(horizontal="right")

    result_sheet.freeze_panes = "A2"
    result_sheet.auto_filter.ref = f"A1:I{max(header_row + len(payable_rows), header_row)}"
    widths = [14, 14, 14, 12, 22, 22, 22, 12, 12]
    for index, width in enumerate(widths, start=1):
        result_sheet.column_dimensions[get_column_letter(index)].width = width

    next_source_row = 1
    source_files = metadata.get("sourceFiles", [])
    for file_index, filename in enumerate(source_files):
        path = run_dir / Path(filename).name
        if not path.exists():
            continue
        source_workbook = load_workbook(path, read_only=False, data_only=False)
        try:
            source_ws = source_workbook[source_workbook.sheetnames[0]]
            start_row = 1 if file_index == 0 else 3
            for row in source_ws.iter_rows(min_row=start_row, values_only=True):
                for column, value in enumerate(row, start=1):
                    source_sheet.cell(next_source_row, column, value)
                next_source_row += 1
        finally:
            source_workbook.close()

    source_sheet.freeze_panes = "A3"
    if source_sheet.max_row >= 2 and source_sheet.max_column >= 1:
        source_sheet.auto_filter.ref = f"A2:{get_column_letter(source_sheet.max_column)}{source_sheet.max_row}"
    for column in range(1, min(source_sheet.max_column, 24) + 1):
        source_sheet.column_dimensions[get_column_letter(column)].width = 16

    punch_datetime_headers = {
        "首打卡(含补签)",
        "末打卡(含补签)",
        "上班 1 打卡时间",
        "下班 1 打卡时间",
    }
    source_headers = {
        str(cell.value).strip(): cell.column
        for cell in source_sheet[2]
        if cell.value not in (None, "")
    }
    for header in punch_datetime_headers:
        column = source_headers.get(header)
        if column is None:
            continue
        for row in range(3, source_sheet.max_row + 1):
            cell = source_sheet.cell(row, column)
            if isinstance(cell.value, datetime):
                cell.number_format = "yyyy/m/d h:mm:ss"

    filename_month = month_label or datetime.now().strftime("%Y%m")
    export_prefix = "WX技术部餐补核算结果" if source_type == "wx" else "技术部餐补核算结果"
    output_path = run_dir / f"{export_prefix}_{filename_month}_{run_id}.xlsx"
    workbook.save(output_path)
    return output_path


@app.get("/api/china-employee-payroll/meal-allowance/runs")
def list_china_employee_meal_allowance_runs() -> dict:
    runs = []
    for run_dir in sorted(CHINA_EMPLOYEE_PAYROLL_RUNS_DIR.glob("china_employee_payroll_*"), reverse=True):
        metadata_path = run_dir / "metadata.json"
        if not run_dir.is_dir() or not metadata_path.exists():
            continue
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        summary = metadata.get("summary", {})
        runs.append(
            {
                "runId": metadata.get("runId") or run_dir.name,
                "createdAt": metadata.get("createdAt", ""),
                "sourceFiles": metadata.get("sourceFiles", []),
                "sourceType": metadata.get("sourceType", summary.get("sourceType", "hr")),
                "sourceLabel": metadata.get("sourceLabel", summary.get("sourceLabel", "")),
                "period": f"{summary.get('dateStart', '')} 至 {summary.get('dateEnd', '')}",
                "monthLabel": _china_employee_payroll_month_label(summary),
                "rowCount": summary.get("rowCount", 0),
                "payableEmployeeCount": summary.get("payableEmployeeCount", 0),
                "payableDayCount": summary.get("payableDayCount", 0),
                "totalAmount": summary.get("totalAmount", 0),
            }
        )
    return {"runs": runs[:50]}


@app.get("/api/china-employee-payroll/meal-allowance/runs/{run_id}")
def get_china_employee_meal_allowance_run(run_id: str) -> dict:
    safe_run_id = Path(run_id).name
    run_dir = CHINA_EMPLOYEE_PAYROLL_RUNS_DIR / safe_run_id
    result_path = run_dir / "result.json"
    if not result_path.exists():
        raise HTTPException(status_code=404, detail="核算批次不存在，请重新上传并核算。")
    result = json.loads(result_path.read_text(encoding="utf-8"))
    return {
        "runId": safe_run_id,
        "summary": result["summary"],
        "results": _compact_china_employee_payroll_result(result),
        "files": result.get("files", []),
        "warnings": result.get("warnings", {}),
        "sourceType": result.get("summary", {}).get("sourceType", "hr"),
        "sourceLabel": result.get("summary", {}).get("sourceLabel", ""),
    }


@app.get("/api/china-employee-payroll/meal-allowance/{run_id}/export")
def export_china_employee_meal_allowance(run_id: str) -> FileResponse:
    safe_run_id = Path(run_id).name
    run_dir = CHINA_EMPLOYEE_PAYROLL_RUNS_DIR / safe_run_id
    if not run_dir.exists():
        raise HTTPException(status_code=404, detail="核算批次不存在，请重新上传并核算。")
    output_path = _build_china_employee_meal_allowance_export(run_dir, safe_run_id)
    return FileResponse(
        output_path,
        filename=output_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ============================================================
# DOMESTIC LABOR PAYROLL API  /api/domestic-labor/*
# ============================================================

DOMESTIC_LABOR_RUNS_DIR.mkdir(parents=True, exist_ok=True)
PAYROLL_OUTPUT_DIR = DOMESTIC_LABOR_RUNS_DIR.parent / "payroll_outputs"
PAYROLL_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

payroll_logger = logging.getLogger("bonus_platform.payroll")

DOMESTIC_LABOR_SUBJECT_NAMES = {
    "canbu": "餐补",
    "waisu_butie": "外宿补贴",
    "quanqinjiang": "全勤奖",
    "gonglingjiang": "工龄奖",
    "gangwei_butie": "岗位补贴",
    "gaowen_butie": "高温补贴",
    "yeban_butie": "夜班补贴",
}
DOMESTIC_LABOR_EXPORT_CACHE_VERSION = "20260813.1"
DOMESTIC_LABOR_EXPORT_CACHE_MANIFEST = ".export-cache.json"


def _domestic_labor_export_filename(metadata: dict) -> str:
    engines = [engine for engine in metadata.get("engines", []) if engine]
    if len(engines) == 1:
        subject_name = DOMESTIC_LABOR_SUBJECT_NAMES.get(engines[0], "薪酬")
    elif len(engines) > 1:
        subject_name = "多科目"
    else:
        subject_name = "薪酬"

    month_digits = re.sub(r"\D", "", str(metadata.get("attendanceMonth", "")))
    attendance_month = month_digits[:6] if len(month_digits) >= 6 else "未指定月份"
    exported_at = datetime.now().strftime("%Y%m%d")
    return f"{subject_name}核算结果_{attendance_month}_{exported_at}.xlsx"


def _compact_domestic_labor_metadata(metadata: dict) -> dict:
    """Return run state without the potentially large employee result payload."""
    return {
        key: value
        for key, value in metadata.items()
        if key not in {"results", "filePath"}
    }


def _domestic_labor_cached_export(run_dir: Path, metadata: dict, file_name: str) -> Path | None:
    """Return an unchanged run's export when its format version is still current."""
    out_path = run_dir / file_name
    manifest_path = run_dir / DOMESTIC_LABOR_EXPORT_CACHE_MANIFEST
    if not out_path.exists() or not manifest_path.exists():
        return None
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if manifest != {
        "version": DOMESTIC_LABOR_EXPORT_CACHE_VERSION,
        "file_name": file_name,
        "source_updated_at": metadata.get("updatedAt", ""),
        "result_count": len(metadata.get("results", [])),
    }:
        return None
    return out_path


def _save_domestic_labor_export_cache(run_dir: Path, metadata: dict, file_name: str) -> None:
    manifest_path = run_dir / DOMESTIC_LABOR_EXPORT_CACHE_MANIFEST
    tmp_path = manifest_path.with_suffix(".json.tmp")
    tmp_path.write_text(json.dumps({
        "version": DOMESTIC_LABOR_EXPORT_CACHE_VERSION,
        "file_name": file_name,
        "source_updated_at": metadata.get("updatedAt", ""),
        "result_count": len(metadata.get("results", [])),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(manifest_path)


@app.get("/api/domestic-labor/rule-package")
def get_domestic_labor_rule_package(version: str = "") -> dict:
    try:
        return get_rule_package(version)
    except KeyError as exc:
        raise HTTPException(404, f"规则包版本不存在: {version}") from exc


def _attach_domestic_engine_result(result: dict, subject: str, calculation) -> None:
    """Attach one engine result while keeping the legacy flat response fields."""
    result[subject] = calculation.amount
    result["warnings"].extend(calculation.warnings)

    details = calculation.details or {}
    subject_detail = {
        "amount": calculation.amount,
        "details": details,
        "exceptions": details.get("exceptions", []),
        "audit_explanation": details.get("audit_explanation"),
    }
    result["subject_details"][subject] = subject_detail
    result["exceptions"].extend(subject_detail["exceptions"])


def _domestic_labor_region_for_department(department: str) -> str:
    """按员工自己的二级部门选择工龄奖兼容区域。"""
    if any(keyword in department for keyword in (
        "华东枢纽", "华东揽收组", "东南枢纽", "华西区操作部",
        "华西枢纽", "华西揽收组", "闽赣揽收组", "华东B2B枢纽",
    )):
        return "wes"
    return "default"


def _run_payroll_calculation(run_id: str, file_paths: list[str], attendance_month: str,
                              engines: list, password: str = None,
                              hrbp_list: list = None,
                              validate_inputs: bool = False,
                              initial_metadata: dict | None = None,
                              night_shift_config: dict | None = None) -> dict:
    """Load Excel, run engines, and persist the terminal task state."""
    payroll_logger.info("Starting payroll calculation for %s, engines=%s", run_id, engines)
    calculation_started = monotonic()
    try:
        running_state_started = monotonic()
        update_payroll_metadata(run_id, {
            **(initial_metadata or {}),
            "status": "计算中",
        })
        payroll_logger.info(
            "Persisted running payroll state for %s in %.2fs",
            run_id,
            monotonic() - running_state_started,
        )
        workbook_load_started = monotonic()
        with MultiFilePayrollDataLoader(file_paths, password=password) as loader:
            payroll_logger.info(
                "Opened %d payroll workbooks for %s in %.2fs",
                len(file_paths),
                run_id,
                monotonic() - workbook_load_started,
            )
            input_summary = None
            if validate_inputs:
                validation_started = monotonic()
                try:
                    input_summary = loader.validate_inputs(engines, attendance_month)
                    payroll_logger.info(
                        "Payroll input parsed for %s in %.2fs: files=%s monthly=%s daily=%s housing=%s",
                        run_id,
                        monotonic() - validation_started,
                        input_summary.get("file_count", 0),
                        input_summary.get("monthly_rows", 0),
                        input_summary.get("daily_rows", 0),
                        input_summary.get("housing_rows", 0),
                    )
                except Exception as exc:
                    payroll_logger.warning("Payroll input validation failed for %s: %s", run_id, exc)
                    return update_payroll_metadata(run_id, {
                        "status": "失败",
                        "error": f"数据文件校验失败：{exc}",
                        "errorCode": "INPUT_VALIDATION_FAILED",
                    })
                collection_roster = list(
                    (initial_metadata or {}).get("collectionSeniorityRoster") or []
                )
                if (
                    "gonglingjiang" in engines
                    and input_summary.get("requires_collection_seniority_roster")
                    and (
                        not collection_roster
                        or any(not str(item.get("employee_name") or "").strip() for item in collection_roster)
                    )
                ):
                    return update_payroll_metadata(run_id, {
                        "status": "失败",
                        "error": "已识别到东莞第四纵队，请维护包含工号和姓名的揽收线工龄奖名单",
                        "errorCode": "INPUT_VALIDATION_FAILED",
                        "inputSummary": input_summary,
                    })
            monthly = loader.monthly
            daily_by_emp = loader.group_daily_by_employee()
            housing_by_emp = loader.group_housing_by_employee()
            temperature_records = loader.temperature.rows if loader.temperature else []
            high_temperature_engine = GaoWenBuTieEngine(temperature_records)

            engine_started = monotonic()
            results = []
            for row in monthly.rows:
                emp_id = str(row.get("工号", ""))
                emp_name = str(row.get("姓名", ""))
                dept = str(
                    row.get("二级部门名称")
                    or row.get("部门名称")
                    or row.get("一级部门名称")
                    or ""
                )
                region = _domestic_labor_region_for_department(dept)
                r = {"employee_id": emp_id, "employee_name": emp_name, "department": dept,
                     "quanqinjiang": 0, "canbu": 0, "waisu_butie": 0, "gonglingjiang": 0,
                     "gangwei_butie": 0,
                     "gaowen_butie": 0,
                     "yeban_butie": 0,
                     "total": 0, "warnings": [], "exceptions": [], "subject_details": {}}

                if "quanqinjiang" in engines:
                    cr = QuanQinJiangEngine().calculate(row, daily_by_emp.get(emp_id, []))
                    _attach_domestic_engine_result(r, "quanqinjiang", cr)

                if "canbu" in engines:
                    cr = CanBuEngine().calculate(row, daily_by_emp.get(emp_id, []))
                    _attach_domestic_engine_result(r, "canbu", cr)

                if "waisu_butie" in engines:
                    cr = WaiSuBuTieEngine().calculate(row, daily_by_emp.get(emp_id, []),
                                                       housing_by_emp.get(emp_id, []))
                    _attach_domestic_engine_result(r, "waisu_butie", cr)

                if "gonglingjiang" in engines:
                    cr = GongLingJiangEngine().calculate(row, hrbp_list or [], region=region)
                    _attach_domestic_engine_result(r, "gonglingjiang", cr)

                if "gangwei_butie" in engines:
                    cr = GangWeiBuTieEngine().calculate(row)
                    _attach_domestic_engine_result(r, "gangwei_butie", cr)

                if "gaowen_butie" in engines:
                    cr = high_temperature_engine.calculate(row, daily_by_emp.get(emp_id, []))
                    _attach_domestic_engine_result(r, "gaowen_butie", cr)

                if "yeban_butie" in engines:
                    cr = YeBanBuTieEngine().calculate(
                        row,
                        daily_by_emp.get(emp_id, []),
                        config=night_shift_config or {},
                    )
                    _attach_domestic_engine_result(r, "yeban_butie", cr)

                r["total"] = (
                    r["quanqinjiang"] + r["canbu"] + r["waisu_butie"]
                    + r["gonglingjiang"] + r["gangwei_butie"] + r["gaowen_butie"] + r["yeban_butie"]
                )
                r["warnings"] = "; ".join(r["warnings"]) if r["warnings"] else ""
                results.append(r)
            payroll_logger.info(
                "Calculated payroll rules for %s in %.2fs: %d employees",
                run_id,
                monotonic() - engine_started,
                len(results),
            )

            # Compute summary
            summary = {
                "total_employees": len(results),
                "total_quanqinjiang": sum(r["quanqinjiang"] for r in results),
                "total_canbu": sum(r["canbu"] for r in results),
                "total_waisu_butie": sum(r["waisu_butie"] for r in results),
                "total_gonglingjiang": sum(r["gonglingjiang"] for r in results),
                "total_gangwei_butie": sum(r["gangwei_butie"] for r in results),
                "total_gaowen_butie": sum(r["gaowen_butie"] for r in results),
                "total_yeban_butie": sum(r["yeban_butie"] for r in results),
                "grand_total": sum(r["total"] for r in results),
                "warning_count": sum(1 for r in results if r["warnings"]),
            }

            terminal_patch = {
                "status": "已完成",
                "results": results,
                "summary": summary,
            }
            if input_summary is not None:
                terminal_patch["inputSummary"] = input_summary
            persistence_started = monotonic()
            metadata = update_payroll_metadata(run_id, terminal_patch)
            payroll_logger.info(
                "Persisted completed payroll state for %s in %.2fs",
                run_id,
                monotonic() - persistence_started,
            )
            payroll_logger.info(
                "Payroll calculation completed for %s in %.2fs: %d employees",
                run_id,
                monotonic() - calculation_started,
                len(results),
            )
            return metadata
    except Exception as exc:
        payroll_logger.exception("Payroll calculation failed for %s", run_id)
        return update_payroll_metadata(run_id, {"status": "失败", "error": str(exc)})


@app.get("/api/domestic-labor/runs")
def list_domestic_labor_runs() -> dict:
    return {
        "runs": [
            _compact_domestic_labor_metadata(metadata)
            for metadata in list_payroll_metadata(compact=True)
        ]
    }


def _normalize_domestic_collection_roster(raw_roster: Any) -> list[dict]:
    if raw_roster in (None, ""):
        return []
    if not isinstance(raw_roster, list):
        raise HTTPException(400, "揽收线工龄奖名单必须为人员列表")
    collection_roster = []
    seen_ids = set()
    for item in raw_roster:
        if isinstance(item, str):
            employee_id = item.strip()
            employee_name = ""
        elif isinstance(item, dict):
            employee_id = str(
                item.get("employee_id") or item.get("employeeId") or item.get("工号") or ""
            ).strip()
            employee_name = str(
                item.get("employee_name") or item.get("employeeName") or item.get("姓名") or ""
            ).strip()
        else:
            raise HTTPException(400, "揽收线工龄奖名单人员格式错误")
        if not employee_id or employee_id in seen_ids:
            continue
        seen_ids.add(employee_id)
        collection_roster.append({"employee_id": employee_id, "employee_name": employee_name})
    return collection_roster


@app.post("/api/domestic-labor/runs")
async def create_domestic_labor_run(files: list[UploadFile] = File(None),
                                     file: UploadFile = File(None), engines: str = Body(""),
                                     attendance_month: str = Body(""),
                                     password: str = Body(""), hrbp_list: str = Body("")):
    uploaded_files = [*(files or []), *([file] if file else [])]
    if not uploaded_files:
        raise HTTPException(400, "请至少上传一个 Excel 文件")
    for uploaded_file in uploaded_files:
        if not uploaded_file.filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
            raise HTTPException(400, f"请上传 Excel 文件（.xlsx / .xlsm / .xls）：{uploaded_file.filename} 格式不支持")

    # Parse engines
    engine_list = [e.strip() for e in engines.split(",") if e.strip()]
    if not engine_list:
        raise HTTPException(400, "请至少选择一个计算引擎")
    valid_engines = set(ENGINE_TEMPLATES.keys())
    for e in engine_list:
        if e not in valid_engines:
            raise HTTPException(400, f"未知引擎: {e}")

    parsed_hrbp = []
    night_shift_config = None
    if "yeban_butie" in engine_list:
        try:
            night_shift_config = build_night_shift_config_snapshot(
                load_night_shift_config(attendance_month, required=False)
            )
        except ValueError as exc:
            raise HTTPException(400, f"夜班补贴配置校验失败：{exc}") from exc
    if hrbp_list.strip():
        try:
            parsed_hrbp = __import__("json").loads(hrbp_list)
        except Exception as exc:
            raise HTTPException(400, "第四纵队揽收发放名单格式错误") from exc
    collection_roster = _normalize_domestic_collection_roster(parsed_hrbp)

    # Save uploaded file
    DOMESTIC_LABOR_RUNS_DIR.mkdir(parents=True, exist_ok=True)
    run = create_payroll_run({
        "engines": engine_list,
        "attendanceMonth": attendance_month,
        "fileName": uploaded_files[0].filename,
        "fileNames": [uploaded_file.filename for uploaded_file in uploaded_files],
        "nightShiftConfigSnapshot": night_shift_config if "yeban_butie" in engine_list else None,
    })
    run_id = run["id"]
    run_dir = get_payroll_run_dir(run_id)

    saved_paths = []
    for uploaded_file in uploaded_files:
        saved_name = safe_payroll_filename(uploaded_file.filename)
        file_path = run_dir / saved_name
        with open(file_path, "wb") as target:
            while chunk := await uploaded_file.read(1024 * 1024):
                target.write(chunk)
        saved_paths.append(file_path)

    try:
        with MultiFilePayrollDataLoader([str(path) for path in saved_paths], password=password or None) as loader:
            input_summary = loader.validate_inputs(engine_list, attendance_month)
    except Exception as exc:
        message = f"数据文件校验失败：{exc}"
        shutil.rmtree(run_dir, ignore_errors=True)
        raise HTTPException(400, message) from exc

    requires_collection_roster = bool(input_summary.get("requires_collection_seniority_roster"))
    if requires_collection_roster and (
        not collection_roster or any(not item["employee_name"] for item in collection_roster)
    ):
        shutil.rmtree(run_dir, ignore_errors=True)
        raise HTTPException(400, "已识别到第四纵队，请维护包含工号和姓名的揽收线工龄奖名单")

    hrbp = [item["employee_id"] for item in collection_roster]

    update_payroll_metadata(run_id, {
        "status": "已上传",
        "filePath": str(saved_paths[0]),
        "filePaths": [str(path) for path in saved_paths],
        "savedFileName": saved_paths[0].name,
        "savedFileNames": [path.name for path in saved_paths],
        "fileSize": sum(path.stat().st_size for path in saved_paths),
        "inputSummary": input_summary,
        "collectionSeniorityRoster": collection_roster if "gonglingjiang" in engine_list else [],
        "collectionSeniorityRosterCount": len(collection_roster) if "gonglingjiang" in engine_list else 0,
        "nightShiftConfigSnapshot": night_shift_config if "yeban_butie" in engine_list else None,
    })
    try:
        for file_path in saved_paths:
            await asyncio.to_thread(persist_payroll_file, run_id, file_path)
    except Exception as exc:
        payroll_logger.exception("Failed to persist payroll upload for %s", run_id)
        update_payroll_metadata(run_id, {"status": "失败", "error": f"文件持久化失败: {exc}"})
        raise HTTPException(503, "上传文件未能保存到持久化存储，请稍后重试。") from exc

    # Keep the invocation alive until the calculation reaches a terminal state.
    metadata = await asyncio.to_thread(
        _run_payroll_calculation, run_id, [str(path) for path in saved_paths],
        attendance_month, engine_list, password or None, hrbp,
        night_shift_config=night_shift_config,
    )

    status = metadata.get("status", "失败")
    return {
        "run_id": run_id,
        "status": status,
        "message": "计算完成" if status == "已完成" else "计算失败",
        "error": metadata.get("error", ""),
        "input_summary": input_summary,
        "collection_seniority_roster": collection_roster if "gonglingjiang" in engine_list else [],
        "collection_seniority_roster_count": len(collection_roster) if "gonglingjiang" in engine_list else 0,
        "night_shift_config_snapshot": night_shift_config if "yeban_butie" in engine_list else None,
    }


def _domestic_labor_direct_upload_max_bytes() -> int:
    raw = os.environ.get("SIGMA_DOMESTIC_LABOR_MAX_UPLOAD_BYTES", "262144000")
    try:
        return max(1, int(raw))
    except (TypeError, ValueError):
        return 262144000


def _validate_domestic_labor_engine_list(raw: Any) -> list[str]:
    values = raw if isinstance(raw, list) else str(raw or "").split(",")
    engine_list = [str(value).strip() for value in values if str(value).strip()]
    if not engine_list:
        raise HTTPException(400, "请至少选择一个计算引擎")
    unknown = [engine for engine in engine_list if engine not in ENGINE_TEMPLATES]
    if unknown:
        raise HTTPException(400, f"未知引擎: {unknown[0]}")
    return engine_list


def _domestic_labor_direct_upload_specs(payload: dict) -> list[dict]:
    raw_files = payload.get("files")
    if raw_files is None:
        raw_files = [payload]
    if not isinstance(raw_files, list) or not raw_files:
        raise HTTPException(400, "请至少选择一个 Excel 文件。")
    if len(raw_files) > 20:
        raise HTTPException(400, "单次最多上传 20 个 Excel 文件。")

    specs = []
    total_size = 0
    for raw in raw_files:
        if not isinstance(raw, dict):
            raise HTTPException(400, "上传文件信息格式不正确。")
        original_name = Path(str(raw.get("fileName") or "")).name
        if not original_name.lower().endswith((".xlsx", ".xlsm", ".xls")):
            raise HTTPException(400, f"请上传 Excel 文件（.xlsx / .xlsm / .xls）：{original_name or '未知文件'}")
        try:
            file_size = int(raw.get("fileSize") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, f"上传文件大小无效：{original_name}") from exc
        if file_size <= 0:
            raise HTTPException(400, f"上传文件不能为空：{original_name}")
        total_size += file_size
        suffix = Path(original_name).suffix.lower()
        specs.append({
            "originalFilename": original_name,
            "filename": f"upload_{secrets.token_hex(12)}{suffix}",
            "size": file_size,
            "contentType": str(raw.get("contentType") or "application/octet-stream"),
        })

    max_bytes = _domestic_labor_direct_upload_max_bytes()
    if total_size > max_bytes:
        raise HTTPException(413, f"上传文件总大小超过当前上限 {max_bytes} 字节。")
    return specs


@app.post("/api/domestic-labor/runs/direct-upload-plan")
def create_domestic_labor_direct_upload_plan(payload: dict = Body(...)) -> dict:
    if not domestic_labor_persistent_storage_enabled():
        raise HTTPException(409, "当前环境未启用 Supabase 直传。")
    specs = _domestic_labor_direct_upload_specs(payload)
    first = specs[0]
    run = create_payroll_run({
        "status": "等待上传",
        "fileName": first["originalFilename"],
        "fileNames": [spec["originalFilename"] for spec in specs],
        "savedFileName": first["filename"],
        "savedFileNames": [spec["filename"] for spec in specs],
        "expectedFileSize": sum(spec["size"] for spec in specs),
        "expectedFiles": specs,
        "contentType": first["contentType"],
        "uploadMode": "direct",
    })
    run_id = run["id"]
    try:
        uploads = [
            {
                **create_domestic_labor_signed_upload(run_id, spec["filename"]),
                **spec,
            }
            for spec in specs
        ]
    except Exception as exc:
        payroll_logger.exception("Failed to create domestic labor signed upload for %s", run_id)
        try:
            delete_payroll_run(run_id)
            shutil.rmtree(get_payroll_run_dir(run_id), ignore_errors=True)
        except Exception:
            payroll_logger.exception("Failed to clean domestic labor upload plan %s", run_id)
        raise HTTPException(503, "生成 Supabase 直传地址失败，请稍后重试。") from exc
    return {
        "runId": run_id,
        "upload": uploads[0],
        "uploads": uploads,
    }


@app.post("/api/domestic-labor/runs/{run_id}/direct-upload-complete")
async def complete_domestic_labor_direct_upload(run_id: str, payload: dict = Body(...)) -> dict:
    if not domestic_labor_persistent_storage_enabled():
        raise HTTPException(409, "当前环境未启用 Supabase 直传。")
    try:
        metadata = load_payroll_metadata(get_payroll_run_dir(run_id))
    except FileNotFoundError as exc:
        raise HTTPException(404, "薪酬计算任务不存在。") from exc
    if metadata.get("uploadMode") != "direct":
        raise HTTPException(400, "该任务不是浏览器直传任务。")
    if metadata.get("status") == "已完成":
        return {"run_id": run_id, "status": "已完成", "message": "计算完成", "error": ""}
    if metadata.get("status") == "计算中":
        raise HTTPException(409, "该任务正在计算，请勿重复提交。")

    engine_list = _validate_domestic_labor_engine_list(payload.get("engines"))
    expected_files = metadata.get("expectedFiles") or [{
        "filename": str(metadata.get("savedFileName") or ""),
        "originalFilename": str(metadata.get("fileName") or ""),
        "size": int(metadata.get("expectedFileSize") or 0),
    }]
    materialize_started = monotonic()
    materialize_slots = asyncio.Semaphore(4)

    async def materialize_expected_file(expected: dict) -> Path:
        saved_name = str(expected.get("filename") or "")
        async with materialize_slots:
            file_path = await asyncio.to_thread(materialize_payroll_file, run_id, saved_name)
        if not file_path or not file_path.is_file():
            raise HTTPException(400, f"未找到已上传文件：{expected.get('originalFilename') or saved_name}")
        actual_size = file_path.stat().st_size
        expected_size = int(expected.get("size") or 0)
        if expected_size and actual_size != expected_size:
            file_path.unlink(missing_ok=True)
            raise HTTPException(
                400,
                f"上传文件大小不一致（{expected.get('originalFilename') or saved_name}：预期 {expected_size}，实际 {actual_size}），请重新上传。",
            )
        return file_path

    try:
        file_paths = list(await asyncio.gather(*[
            materialize_expected_file(expected)
            for expected in expected_files
        ]))
        payroll_logger.info(
            "Materialized %d domestic labor files for %s in %.2fs",
            len(file_paths),
            run_id,
            monotonic() - materialize_started,
        )
    except HTTPException:
        raise
    except Exception as exc:
        payroll_logger.exception("Failed to materialize direct domestic labor upload %s", run_id)
        raise HTTPException(503, "读取已上传文件失败，请稍后重试。") from exc

    collection_roster = _normalize_domestic_collection_roster(payload.get("hrbpList"))
    hrbp = [item["employee_id"] for item in collection_roster]
    attendance_month = str(payload.get("attendanceMonth") or "")
    password = str(payload.get("password") or "") or None
    night_shift_config = None
    if "yeban_butie" in engine_list:
        try:
            night_shift_config = build_night_shift_config_snapshot(
                load_night_shift_config(attendance_month, required=False)
            )
        except ValueError as exc:
            raise HTTPException(400, f"夜班补贴配置校验失败：{exc}") from exc
    actual_size = sum(path.stat().st_size for path in file_paths)
    initial_metadata = {
        "engines": engine_list,
        "attendanceMonth": attendance_month,
        "filePath": str(file_paths[0]),
        "filePaths": [str(path) for path in file_paths],
        "fileSize": actual_size,
        **({
            "collectionSeniorityRoster": collection_roster,
            "collectionSeniorityRosterCount": len(collection_roster),
        } if "gonglingjiang" in engine_list else {}),
        **({
            "nightShiftConfigSnapshot": night_shift_config,
        } if "yeban_butie" in engine_list else {}),
    }
    calculation_kwargs = (
        {"night_shift_config": night_shift_config}
        if "yeban_butie" in engine_list
        else {}
    )
    result = await asyncio.to_thread(
        _run_payroll_calculation,
        run_id,
        [str(path) for path in file_paths],
        attendance_month,
        engine_list,
        password,
        hrbp,
        True,
        initial_metadata,
        **calculation_kwargs,
    )
    status = result.get("status", "失败")
    if result.get("errorCode") == "INPUT_VALIDATION_FAILED":
        raise HTTPException(400, result.get("error") or "数据文件校验失败。")
    input_summary = result.get("inputSummary") or {}
    return {
        "run_id": run_id,
        "status": status,
        "message": "计算完成" if status == "已完成" else "计算失败",
        "error": result.get("error", ""),
        "input_summary": input_summary,
        "collection_seniority_roster": collection_roster if "gonglingjiang" in engine_list else [],
        "collection_seniority_roster_count": len(collection_roster) if "gonglingjiang" in engine_list else 0,
        "night_shift_config_snapshot": night_shift_config if "yeban_butie" in engine_list else None,
    }


@app.get("/api/domestic-labor/runs/{run_id}")
def get_domestic_labor_run(run_id: str, response_mode: str = "") -> dict:
    try:
        run_dir = get_payroll_run_dir(run_id)
        metadata = (
            load_payroll_status(run_dir)
            if response_mode in {"compact", "status"}
            else load_payroll_metadata(run_dir)
        )
    except FileNotFoundError as exc:
        raise HTTPException(404, "薪酬计算任务不存在。") from exc
    return metadata


@app.get("/api/domestic-labor/runs/{run_id}/results")
def get_domestic_labor_results(run_id: str) -> dict:
    try:
        metadata = load_payroll_metadata(get_payroll_run_dir(run_id))
    except FileNotFoundError as exc:
        raise HTTPException(404, "薪酬计算任务不存在。") from exc
    return {
        "run_id": run_id,
        "status": metadata.get("status"),
        "results": metadata.get("results", []),
        "summary": metadata.get("summary", {}),
    }


@app.get("/api/domestic-labor/runs/{run_id}/export")
def export_domestic_labor(run_id: str) -> dict:
    try:
        run_dir = get_payroll_run_dir(run_id)
        metadata = load_payroll_metadata(run_dir)
    except FileNotFoundError as exc:
        raise HTTPException(404, "薪酬计算任务不存在。") from exc
    results = metadata.get("results", [])
    if not results:
        raise HTTPException(400, "暂无计算结果可导出")
    file_name = _domestic_labor_export_filename(metadata)
    cached_path = _domestic_labor_cached_export(run_dir, metadata, file_name)
    if cached_path is not None:
        return {"file_path": str(cached_path), "file_name": file_name, "cached": True}

    out_path = run_dir / file_name
    with NamedTemporaryFile(prefix=".export-", suffix=".xlsx", dir=run_dir, delete=False) as temp_file:
        temp_path = Path(temp_file.name)
    exporter = ExcelExporter(str(temp_path))
    summary = metadata.get("summary", {})
    try:
        exporter.export(results, metadata.get("attendanceMonth", ""), summary)
        temp_path.replace(out_path)
        _save_domestic_labor_export_cache(run_dir, metadata, file_name)
        persist_payroll_file(run_id, out_path)
    finally:
        temp_path.unlink(missing_ok=True)
    return {"file_path": str(out_path), "file_name": file_name, "cached": False}


@app.get("/api/domestic-labor/runs/{run_id}/download/{filename}")
def download_domestic_labor_file(run_id: str, filename: str) -> FileResponse:
    try:
        run_dir = get_payroll_run_dir(run_id)
    except FileNotFoundError as exc:
        raise HTTPException(404, "薪酬计算任务不存在。") from exc
    # Check run dir first, then output dir
    path = materialize_payroll_file(run_id, filename)
    if path is None:
        path = PAYROLL_OUTPUT_DIR / Path(filename).name
    if not path.exists():
        raise HTTPException(404, "文件不存在或已被清理。")
    return FileResponse(path, filename=path.name)


@app.delete("/api/domestic-labor/runs/{run_id}")
def delete_domestic_labor_run(run_id: str) -> dict:
    try:
        run_dir = get_payroll_run_dir(run_id)
    except FileNotFoundError as exc:
        raise HTTPException(404, "薪酬计算任务不存在。") from exc
    delete_payroll_run(run_id)
    shutil.rmtree(run_dir, ignore_errors=True)
    return {"message": f"已删除任务: {run_id}"}


@app.get("/api/domestic-labor/templates")
def list_domestic_labor_templates() -> dict:
    return {"templates": [get_template_info(k) for k in ENGINE_TEMPLATES]}


@app.get("/api/domestic-labor/templates/{engine_key}/download")
def download_domestic_labor_template(engine_key: str) -> FileResponse:
    if engine_key not in ENGINE_TEMPLATES:
        raise HTTPException(404, "模板不存在")
    data = generate_template(engine_key)
    tmp = NamedTemporaryFile(delete=False, suffix=f"_{engine_key}_template.xlsx")
    tmp.write(data)
    tmp.close()
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{engine_key}_template.xlsx",
    )


@app.get("/api/domestic-labor/night-shift/config/{month}")
def get_domestic_labor_night_shift_config(month: str) -> dict:
    try:
        payload = load_night_shift_config(month, required=False)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {**payload, "counts": config_counts(payload)}


@app.put("/api/domestic-labor/night-shift/config/{month}")
def put_domestic_labor_night_shift_config(month: str, payload: dict = Body(...)) -> dict:
    try:
        saved = save_night_shift_config(month, payload)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {**saved, "counts": config_counts(saved)}


@app.get("/api/domestic-labor/night-shift/config/{month}/history")
def list_domestic_labor_night_shift_config_history(month: str) -> dict:
    try:
        revisions = list_night_shift_config_revisions(month)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"month": month, "revisions": revisions}


@app.post("/api/domestic-labor/night-shift/config/{month}/copy")
def copy_domestic_labor_night_shift_config(month: str, payload: dict = Body(default={})) -> dict:
    source_month = str(payload.get("source_month") or "")
    if not source_month:
        try:
            target = datetime.strptime(month, "%Y%m")
        except ValueError as exc:
            raise HTTPException(400, "核算月份必须为 YYYYMM") from exc
        source_month = (target.replace(day=1) - timedelta(days=1)).strftime("%Y%m")
    try:
        copied = copy_night_shift_config(source_month, month)
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except FileExistsError as exc:
        raise HTTPException(409, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {**copied, "counts": config_counts(copied)}


@app.post("/api/domestic-labor/night-shift/config/{month}/import")
async def import_domestic_labor_night_shift_config(month: str, file: UploadFile = File(...)) -> dict:
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "夜班维护文件仅支持 .xlsx / .xlsm")
    try:
        payload = parse_night_shift_config_workbook(await file.read())
        current = load_night_shift_config(month, required=False)
        saved = save_night_shift_config(month, {
            "shift_break_overrides": current.get("shift_break_overrides", []),
            "jinjiang_exclusions": payload.get("jinjiang_exclusions", []),
            "jinjiang_list_confirmed": True,
        })
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {**saved, "counts": config_counts(saved)}


@app.get("/api/domestic-labor/night-shift/config/{month}/download")
def download_domestic_labor_night_shift_config(month: str) -> FileResponse:
    try:
        payload = load_night_shift_config(month)
    except FileNotFoundError as exc:
        raise HTTPException(404, str(exc)) from exc
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    data = generate_night_shift_config_workbook(payload)
    tmp = NamedTemporaryFile(delete=False, suffix=f"_{month}_night_shift_config.xlsx")
    tmp.write(data)
    tmp.close()
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"晋江不享有夜班补贴人员名单_{month}.xlsx",
    )


@app.get("/api/domestic-labor/night-shift/config-template/download")
def download_domestic_labor_night_shift_config_template() -> FileResponse:
    data = generate_night_shift_config_workbook()
    tmp = NamedTemporaryFile(delete=False, suffix="_night_shift_config_template.xlsx")
    tmp.write(data)
    tmp.close()
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="晋江不享有夜班补贴人员名单模板.xlsx",
    )


# =========================================================================
# FBU PERFORMANCE API  /api/fbu-performance/*
# =========================================================================

fbu_run_manager = FBURunManager(str(FBU_PERFORMANCE_RUNS_DIR))
fbu_roster_store = FBURosterStore(str(FBU_PERFORMANCE_RUNS_DIR))
fbu_rule_list_store = FBURuleListStore(str(FBU_PERFORMANCE_RUNS_DIR))


def _fbu_current_user(request: Request) -> dict[str, str]:
    """Resolve the production session user while keeping local FBU development standalone."""
    current: dict | None = None
    try:
        from .auth import current_user_from_request

        candidate = current_user_from_request(request)
        if isinstance(candidate, dict):
            user_profile = candidate.get("user")
            current = {
                **candidate,
                **(user_profile if isinstance(user_profile, dict) else {}),
            }
    except (ImportError, KeyError, RuntimeError):
        current = None

    if not current:
        return {
            "id": "local-fbu-user",
            "name": "本地用户",
            "avatarUrl": "",
        }

    return {
        "id": str(current.get("id") or current.get("userId") or "local-fbu-user").strip(),
        "name": str(current.get("name") or current.get("email") or "西格玛用户").strip(),
        "avatarUrl": str(
            current.get("avatarUrl")
            or current.get("avatar_url")
            or ""
        ).strip(),
    }


def _fbu_result_file_payload(
    run_id: str,
    result_type: str,
    calc_month: str | None = None,
) -> dict:
    """Return the lazy export link without generating the workbook during import."""
    filename_prefixes = {
        "attendance": "考勤汇总",
        "salary": "薪资匹配",
        "performance": "绩效明细",
        "adjustments": "调薪拆分",
        "base_overrides": "工时规则与固定基数例外",
        "diagnostics": "数据诊断",
        "results": "核算结果",
    }
    prefix = filename_prefixes.get(result_type)
    if not prefix:
        raise HTTPException(400, f"不支持的导出类型: {result_type}")
    if not calc_month:
        run = fbu_run_manager.get_run(run_id, sections=set())
        if not run:
            raise HTTPException(404, "任务不存在")
        calc_month = run.calc_month
    filename = f"{prefix}_{calc_month}_{run_id}.xlsx"
    return {
        "type": result_type,
        "filename": filename,
        "download_url": f"/api/fbu-performance/runs/{run_id}/download/{quote(filename)}",
    }


def _merge_fbu_attendance_context(
    current_path: Path,
    previous_path: Path,
    output_path: Path,
    calc_month: str,
) -> Path:
    """Merge previous/current attendance reports so 96-hour rules can inspect cross-month weeks."""
    current_wb = load_workbook(current_path, read_only=True, data_only=True)
    previous_wb = load_workbook(previous_path, read_only=True, data_only=True)
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)
    window = _fbu_96_previous_context_window(calc_month)
    required_start = window[2] if window else None
    required_end = window[3] if window else None

    try:
        for sheet_name in current_wb.sheetnames:
            merged_sheet = merged_wb.create_sheet(sheet_name)
            current_sheet = current_wb[sheet_name]
            previous_sheet = previous_wb[sheet_name] if sheet_name in previous_wb.sheetnames else None

            if sheet_name == "sheet1" and previous_sheet is not None:
                for row in current_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                    merged_sheet.append(list(row))
                if required_start and required_end:
                    for row in previous_sheet.iter_rows(min_row=2, values_only=True):
                        parsed_date = _parse_fbu_attendance_date(row[0] if row else None)
                        if parsed_date and required_start <= parsed_date <= required_end:
                            merged_sheet.append(list(row))
                for row in current_sheet.iter_rows(min_row=2, values_only=True):
                    merged_sheet.append(list(row))
                continue

            for row in current_sheet.iter_rows(values_only=True):
                merged_sheet.append(list(row))
    finally:
        current_wb.close()
        previous_wb.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    merged_wb.save(output_path)
    return output_path


def _parse_fbu_attendance_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _fbu_attendance_dates_from_preview(preview: dict) -> set[date]:
    """Reuse normalized parser output instead of reopening the attendance workbook."""
    available_dates = set()
    for employee in preview.get("employees", []):
        for row in employee.get("attendance_daily_rows", []):
            parsed_date = _parse_fbu_attendance_date(row.get("date"))
            if parsed_date:
                available_dates.add(parsed_date)
    return available_dates


def _fbu_96_previous_context_window(calc_month: str) -> tuple[date, date, date, date] | None:
    match = re.match(r"^(\d{4})-(\d{1,2})$", str(calc_month or "").strip())
    if not match:
        return None
    month_start = date(int(match.group(1)), int(match.group(2)), 1)
    period_start = date(2026, 3, 29)
    while period_start > month_start:
        period_start -= timedelta(days=14)
    while period_start + timedelta(days=13) < month_start:
        period_start += timedelta(days=14)
    if period_start >= month_start:
        return None
    period_end = period_start + timedelta(days=13)
    required_end = month_start - timedelta(days=1)
    return period_start, period_end, period_start, required_end


def _fbu_attendance_dates_from_preview(preview: dict) -> set[date]:
    available_dates = set()
    for employee in preview.get("employees", []):
        for row in employee.get("attendance_daily_rows", []):
            parsed_date = _parse_fbu_attendance_date(row.get("date"))
            if parsed_date:
                available_dates.add(parsed_date)
    return available_dates


def _build_fbu_attendance_context_summary(
    filepath: Path,
    calc_month: str,
    previous_attendance_file: str = "",
    available_dates: set[date] | None = None,
) -> dict:
    window = _fbu_96_previous_context_window(calc_month)
    if not window:
        return {
            "required": False,
            "status": "not_required",
            "message": "当前核算月96工时制首段不需要上一月考勤上下文。",
        }

    period_start, period_end, required_start, required_end = window
    required_dates = []
    current = required_start
    while current <= required_end:
        required_dates.append(current)
        current += timedelta(days=1)

    if available_dates is None:
        available_dates = set()
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            if "sheet1" in wb.sheetnames:
                ws = wb["sheet1"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    parsed_date = _parse_fbu_attendance_date(row[0] if row else None)
                    if parsed_date:
                        available_dates.add(parsed_date)
            wb.close()
        except Exception:
            available_dates = set()

    covered_dates = [item for item in required_dates if item in available_dates]
    missing_dates = [item for item in required_dates if item not in available_dates]
    required_start_text = required_start.isoformat()
    required_end_text = required_end.isoformat()
    if missing_dates:
        message = (
            f"缺少上一月 {required_start_text} 至 {required_end_text} 考勤，"
            "96工时制首段可能无法对齐线下结果。"
        )
        status = "missing"
    else:
        message = (
            f"已识别96工时制跨月首段：{period_start.isoformat()} 至 {period_end.isoformat()}，"
            f"已使用上月考勤 {required_start_text} 至 {required_end_text}。"
        )
        status = "complete"

    return {
        "required": True,
        "status": status,
        "period_start": period_start.isoformat(),
        "period_end": period_end.isoformat(),
        "required_start": required_start_text,
        "required_end": required_end_text,
        "previous_attendance_file": previous_attendance_file,
        "covered_dates": [item.isoformat() for item in covered_dates],
        "missing_dates": [item.isoformat() for item in missing_dates],
        "message": message,
    }


def _append_fbu_previous_attendance_context_to_preview(
    preview: dict,
    previous_path: Path,
    calc_month: str,
    previous_attendance_filename: str,
) -> dict:
    """Append only the previous-month daily rows needed for 96-hour cross-month periods."""
    window = _fbu_96_previous_context_window(calc_month)
    if not window:
        preview.setdefault("summary", {})["attendance_context"] = _build_fbu_attendance_context_summary(
            previous_path,
            calc_month,
            previous_attendance_filename,
        )
        return preview

    _, _, required_start, required_end = window
    previous_rows = []
    available_dates = set()
    previous_headers = ()
    wb = load_workbook(previous_path, read_only=True, data_only=True)
    try:
        if "sheet1" in wb.sheetnames:
            ws = wb["sheet1"]
            previous_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            for row in ws.iter_rows(min_row=2, values_only=True):
                parsed_date = _parse_fbu_attendance_date(row[0] if row else None)
                if parsed_date:
                    available_dates.add(parsed_date)
                    if required_start <= parsed_date <= required_end:
                        previous_rows.append(row)
    finally:
        wb.close()

    if previous_rows:
        processor = AttendanceProcessor()
        previous_by_employee = processor.process(previous_rows, required_start.month, headers=previous_headers)
        for employee in preview.get("employees", []):
            source_id = normalize_shift_employee_id(employee.get("source_employee_id") or employee.get("employee_id"))
            previous_hours = previous_by_employee.get(source_id)
            if not previous_hours:
                continue
            shift_type = str(employee.get("shift_type") or "").strip()
            if shift_type in {"白班", "夜班"}:
                context_rows = list(previous_hours.get(f"{shift_type}_daily_rows") or [])
            else:
                context_rows = list(previous_hours.get("daily_rows") or [])
            if not context_rows:
                continue
            existing_rows = list(employee.get("attendance_daily_rows") or [])
            seen = {
                (str(row.get("date") or ""), str(row.get("shift_type") or ""))
                for row in existing_rows
            }
            for row in context_rows:
                key = (str(row.get("date") or ""), str(row.get("shift_type") or ""))
                if key not in seen:
                    existing_rows.append(row)
                    seen.add(key)
            employee["attendance_daily_rows"] = sorted(existing_rows, key=lambda item: str(item.get("date") or ""))

    preview.setdefault("summary", {})["attendance_context"] = _build_fbu_attendance_context_summary(
        previous_path,
        calc_month,
        previous_attendance_filename,
        available_dates,
    )
    return preview


_FBU_DIAGNOSTIC_SECTION_FIELDS = {
    "attendance_data",
    "salary_data",
    "performance_data",
    "adjustment_data",
    "base_override_data",
}

_FBU_CALCULATION_SECTION_FIELDS = {
    "attendance_data",
    "salary_data",
    "salary_verification_data",
    "performance_data",
    "adjustment_data",
    "transfer_data",
    "supplemental_leave_data",
    "base_override_data",
    "hourly_rate_policy_data",
    "period_adjustment_data",
}

_FBU_RUN_CORE_RESPONSE_FIELDS = frozenset({
    "run_id",
    "created_at",
    "calc_month",
    "activity_name",
    "region_code",
    "region_name",
    "created_by_user_id",
    "created_by_name",
    "created_by_avatar_url",
    "status",
    "current_step",
    "attendance_file",
    "previous_attendance_file",
    "salary_file",
    "previous_salary_file",
    "current_salary_file",
    "performance_file",
    "adjustment_file",
    "transfer_file",
    "supplemental_leave_file",
    "base_override_file",
    "roster_file",
    "roster_source",
    "total_employees",
    "total_bonus",
    "match_rate",
    "error",
})


def _fbu_run_diagnostics(run: FBURun) -> dict:
    """生成紧凑的数据匹配诊断。"""
    attendance_employees = run.attendance_data.get("employees", []) if run.attendance_data else []
    salary_employees = run.salary_data.get("employees", []) if run.salary_data else []
    performance_employees = run.performance_data.get("employees", []) if run.performance_data else []
    adjustment_employees = run.adjustment_data.get("employees", []) if run.adjustment_data else []
    base_override_employees = run.base_override_data.get("employees", []) if run.base_override_data else []

    attendance_by_id = {emp.get("employee_id"): emp for emp in attendance_employees if emp.get("employee_id")}
    salary_by_id = {emp.get("employee_id"): emp for emp in salary_employees if emp.get("employee_id")}
    performance_by_id = {emp.get("employee_id"): emp for emp in performance_employees if emp.get("employee_id")}
    adjustment_by_id = {emp.get("employee_id"): emp for emp in adjustment_employees if emp.get("employee_id")}
    base_override_by_id = {
        emp.get("employee_id"): emp
        for emp in base_override_employees
        if emp.get("employee_id") and emp.get("include_in_calculation")
    }

    attendance_ids = set(attendance_by_id)
    salary_ids = set(salary_by_id)
    performance_ids = set(performance_by_id)
    adjustment_ids = set(adjustment_by_id)
    base_override_ids = set(base_override_by_id)

    def source_id(emp_id: str) -> str:
        text = str(emp_id or "").strip()
        return text[:-2] if text.endswith("-1") else text

    salary_source_ids = {source_id(emp_id) for emp_id in salary_ids}
    performance_source_ids = {source_id(emp_id) for emp_id in performance_ids}
    adjustment_source_ids = {source_id(emp_id) for emp_id in adjustment_ids}
    attendance_source_ids = {source_id(emp_id) for emp_id in attendance_ids}

    def name_for(emp_id: str) -> str:
        for source in (attendance_by_id, salary_by_id, performance_by_id, adjustment_by_id, base_override_by_id):
            if emp_id in source:
                return source[emp_id].get("name", "")
            split_id = f"{emp_id}-1"
            if split_id in source:
                return source[split_id].get("name", "")
        return ""

    issues = []
    has_salary_data = bool(salary_ids) or bool(run.salary_file)
    has_performance_data = bool(performance_ids) or bool(adjustment_ids) or bool(run.performance_file) or bool(run.adjustment_file)

    def add_issue(kind: str, emp_id: str, detail: str, severity: str = "warning"):
        issues.append({
            "severity": severity,
            "type": kind,
            "employee_id": emp_id,
            "name": name_for(emp_id),
            "detail": detail,
        })

    if has_salary_data:
        for emp_id in sorted(emp_id for emp_id in attendance_source_ids if emp_id not in salary_source_ids):
            add_issue("考勤有薪资无", emp_id, "该员工有考勤记录，但薪资档案未匹配", "error")

    if has_performance_data:
        for emp_id in sorted(
            emp_id
            for emp_id in attendance_source_ids
            if emp_id not in performance_source_ids
            and emp_id not in adjustment_source_ids
        ):
            add_issue("考勤有绩效无", emp_id, "该员工有考勤记录，但绩效报表未匹配", "warning")

    for emp_id in sorted(emp_id for emp_id in salary_ids if source_id(emp_id) not in attendance_source_ids):
        add_issue("薪资有考勤无", emp_id, "薪资档案存在该员工，但本月考勤未出现", "info")

    for emp_id in sorted(emp_id for emp_id in performance_ids if source_id(emp_id) not in attendance_source_ids):
        add_issue("绩效有考勤无", emp_id, "绩效报表存在该员工，但本月考勤未出现", "info")

    for emp_id in sorted(emp_id for emp_id in adjustment_ids if source_id(emp_id) not in attendance_source_ids):
        add_issue("拆分有考勤无", emp_id, "调薪/转正拆分表存在该员工，但本月考勤未出现", "error")

    if has_salary_data:
        for emp_id in sorted(emp_id for emp_id in adjustment_ids if source_id(emp_id) not in salary_source_ids):
            add_issue("拆分有薪资无", emp_id, "调薪/转正拆分表存在该员工，但薪资档案未匹配", "error")
        for emp_id in sorted(emp_id for emp_id in base_override_ids if source_id(emp_id) not in salary_source_ids):
            add_issue("规则表有薪资无", emp_id, "工时规则/固定基数例外表存在该员工，但薪资档案未匹配", "error")

    for emp_id in sorted(emp_id for emp_id in base_override_ids if source_id(emp_id) not in attendance_source_ids):
        add_issue("规则表有考勤无", emp_id, "工时规则/固定基数例外表存在该员工，但本月考勤未出现", "error")

    for emp_id in sorted(attendance_ids):
        salary = salary_by_id.get(source_id(emp_id))
        if not salary:
            continue
        if (salary.get("hourly_rate") or 0) <= 0:
            add_issue("时薪为0", emp_id, "薪资档案时薪为0，绩效基数可能无法计算", "error")
        if (salary.get("ratio") or 0) <= 0:
            add_issue("绩效比例为空", emp_id, "薪资档案绩效比例为空或为0", "warning")

    for emp in adjustment_employees:
        emp_id = emp.get("employee_id", "")
        active_base = sum(
            (segment.get("performance_base") or 0)
            for segment in emp.get("segments", [])
            if "前" not in str(segment.get("reason", ""))
        )
        if active_base <= 0:
            add_issue("拆分有效基数为0", emp_id, "调薪/转正拆分表未识别到调薪后有效基数", "warning")

    severity_rank = {"error": 0, "warning": 1, "info": 2}
    issues.sort(key=lambda item: (severity_rank.get(item["severity"], 9), item["type"], item["employee_id"]))

    matched_salary = sum(1 for emp_id in attendance_ids if source_id(emp_id) in salary_source_ids)
    matched_performance = sum(1 for emp_id in attendance_ids if source_id(emp_id) in performance_source_ids)
    can_calculate = matched_salary

    return {
        "summary": {
            "attendance_count": len(attendance_ids),
            "salary_count": len(salary_ids),
            "performance_count": len(performance_ids),
            "adjustment_count": len(adjustment_ids),
            "base_override_count": len(base_override_ids),
            "matched_salary_count": matched_salary,
            "matched_performance_count": matched_performance,
            "can_calculate_count": can_calculate,
            "issue_count": len(issues),
            "error_count": sum(1 for issue in issues if issue["severity"] == "error"),
            "warning_count": sum(1 for issue in issues if issue["severity"] == "warning"),
        },
        "issues": issues,
    }


def _rule_list_roster_info(roster_lookup: dict[str, dict] | None, employee_id: str) -> dict:
    if not roster_lookup:
        return {}
    return roster_lookup.get(employee_id, {})


def _build_rule_list_override_row(
    row: dict,
    *,
    calc_month: str,
    rule_type: str,
    fixed_performance_base,
    calculation_path: str,
    roster_lookup: dict[str, dict] | None = None,
) -> dict:
    employee_id = str(row.get("employee_id") or "").strip()
    roster_info = _rule_list_roster_info(roster_lookup, employee_id)
    name = str(roster_info.get("name") or row.get("name") or "").strip()
    area = str(roster_info.get("area") or "").strip()
    department = str(roster_info.get("department") or "").strip()
    return {
        "employee_id": employee_id,
        "source_employee_id": employee_id,
        "name": name,
        "area": area,
        "department": department,
        "rule_type": rule_type,
        "fixed_performance_base": fixed_performance_base,
        "allocation_month": calc_month,
        "status": "启用",
        "include_in_calculation": True,
        "note": "页面维护",
        "calculation_path": calculation_path,
    }


def _build_base_override_data_from_rule_lists(
    calc_month: str,
    payload: dict,
    roster_lookup: dict[str, dict] | None = None,
    region_name: str = "",
) -> dict:
    target_area = str(region_name or "").removeprefix("FBU").strip()

    def belongs_to_run_region(employee_id: str, *, allow_missing: bool = False) -> bool:
        if not roster_lookup or not target_area:
            return True
        roster_info = _rule_list_roster_info(roster_lookup, employee_id)
        if not roster_info:
            return allow_missing
        employee_area = str(roster_info.get("area") or "").strip()
        return employee_area == target_area

    employees = []
    excluded_count = 0
    for row in payload.get("work_hour_employees", []):
        if not row.get("active", True):
            continue
        employee_id = str(row.get("employee_id") or "").strip()
        if not employee_id:
            continue
        if not belongs_to_run_region(employee_id):
            excluded_count += 1
            continue
        employees.append(_build_rule_list_override_row(
            row,
            calc_month=calc_month,
            rule_type="96工时制",
            fixed_performance_base=None,
            calculation_path="96工时制自动基数路径",
            roster_lookup=roster_lookup,
        ))
    for row in payload.get("fixed_base_employees", []):
        if not row.get("active", True):
            continue
        employee_id = str(row.get("employee_id") or "").strip()
        if not employee_id:
            continue
        if not belongs_to_run_region(employee_id):
            excluded_count += 1
            continue
        employees.append(_build_rule_list_override_row(
            row,
            calc_month=calc_month,
            rule_type="线下固定基数覆盖",
            fixed_performance_base=float(row.get("fixed_performance_base") or 0),
            calculation_path="线下固定基数覆盖路径",
            roster_lookup=roster_lookup,
        ))
    fixed_base_total = sum(float(row.get("fixed_performance_base") or 0) for row in employees)
    work_hour_rule_count = sum(1 for row in employees if row["rule_type"] == "96工时制")
    fixed_base_count = sum(1 for row in employees if row["rule_type"] == "线下固定基数覆盖")
    return {
        "employees": employees,
        "summary": {
            "total_rows": len(employees),
            "active_count": len(employees),
            "excluded_count": excluded_count,
            "work_hour_rule_count": work_hour_rule_count,
            "fixed_base_count": fixed_base_count,
            "active_fixed_base": fixed_base_total,
        },
    }


def _load_fbu_roster_for_run(parser: FBUPerformanceParser, run_id: str) -> Path | None:
    """加载活动花名册；没有活动花名册时复制同月基础花名册。"""
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        return None
    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    for filename in ("roster.xlsx", "roster.xls"):
        fbu_run_manager.materialize_file(run_id, filename)
    roster_path = next((path for path in [run_dir / "roster.xlsx", run_dir / "roster.xls"] if path.exists()), None)
    if roster_path is None:
        roster_path = fbu_roster_store.copy_active_to_run(run_id, run.calc_month)
        if roster_path:
            metadata = fbu_roster_store.get_metadata(run.calc_month)
            fbu_run_manager.update_run(
                run_id,
                roster_file=metadata.get("filename", "active_roster.xlsx"),
                roster_source="base",
            )
    if roster_path and roster_path.exists():
        parser.load_roster(str(roster_path))
        return roster_path
    return None


def _fbu_roster_preview_for_run(run_id: str) -> dict | None:
    run = fbu_run_manager.runs.get(run_id)
    if run and run.roster_data:
        return run.roster_data

    parser = FBUPerformanceParser()
    roster_path = _load_fbu_roster_for_run(parser, run_id)
    if not roster_path:
        return None

    employees = []
    for employee_id, row in parser.employee_roster.items():
        employees.append({
            "employee_id": employee_id,
            "name": row.get("name", ""),
            "department": row.get("department", ""),
            "area": row.get("area", ""),
            "position": row.get("position", ""),
            "personnel_status": row.get("personnel_status", ""),
            "job_type": row.get("job_type", ""),
        })

    preview = {
        "employees": employees,
        "summary": {
            "total_employees": len(employees),
        },
    }
    if run:
        fbu_run_manager.update_run(run_id, roster_data=preview)
    return preview


@app.get("/api/fbu-performance/roster")
def get_fbu_base_roster(calc_month: str) -> dict:
    """获取FBU基础花名册状态"""
    try:
        return fbu_roster_store.get_metadata(calc_month)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc


@app.post("/api/fbu-performance/roster")
async def upload_fbu_base_roster(
    file: UploadFile = File(...),
    calc_month: str = Form(...),
    run_id: str = Form(""),
) -> dict:
    """上传FBU基础花名册，供后续月度活动默认引用"""
    tmp_path: Path | None = None
    try:
        suffix = Path(file.filename or "").suffix.lower()
        if suffix not in {".xlsx", ".xls"}:
            raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的花名册")
        if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", str(calc_month or "").strip()):
            raise HTTPException(400, "核算月份格式无效，应为YYYY-MM")
        target_run = None
        if run_id:
            target_run = fbu_run_manager.get_run(run_id, sections=set())
            if not target_run:
                raise HTTPException(404, "任务不存在")
            if target_run.calc_month != calc_month:
                raise HTTPException(409, "花名册月份与活动核算月份不一致")
        content = await file.read()
        temp_dir = FBU_PERFORMANCE_RUNS_DIR / "_roster"
        temp_dir.mkdir(parents=True, exist_ok=True)
        with NamedTemporaryFile(suffix=suffix, dir=temp_dir, delete=False) as tmp_file:
            tmp_file.write(content)
            tmp_path = Path(tmp_file.name)

        parser = FBUPerformanceParser()
        roster = parser.load_roster(str(tmp_path))
        metadata = fbu_roster_store.save_active_roster(
            content=content,
            filename=file.filename,
            total_employees=len(roster),
            calc_month=calc_month,
        )
        if target_run:
            fbu_roster_store.copy_active_to_run(run_id, target_run.calc_month)
            fbu_run_manager.update_run(
                run_id,
                roster_file=metadata.get("filename", "active_roster.xlsx"),
                roster_source="base",
            )
        return {"success": True, "roster": metadata}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"花名册解析失败: {str(e)}")
    finally:
        if tmp_path:
            tmp_path.unlink(missing_ok=True)


@app.get("/api/fbu-performance/rule-lists")
def get_fbu_rule_lists(region_code: str) -> dict:
    try:
        return fbu_rule_list_store.get(region_code)
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/fbu-performance/rule-lists")
def save_fbu_rule_lists(region_code: str, body: dict = Body(...)) -> dict:
    try:
        return fbu_rule_list_store.save(region_code, body)
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/fbu-performance/runs/{run_id}/rule-lists/confirm")
def confirm_fbu_run_rule_lists(run_id: str, body: dict = Body(...)) -> dict:
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")
    try:
        saved = fbu_rule_list_store.save(run.region_code, body)
    except ValueError as e:
        raise HTTPException(400, str(e))
    parser = FBUPerformanceParser()
    _load_fbu_roster_for_run(parser, run_id)
    preview = _build_base_override_data_from_rule_lists(
        run.calc_month,
        saved,
        roster_lookup=parser.employee_roster,
        region_name=run.region_name,
    )
    fbu_run_manager.update_run(
        run_id,
        base_override_file="页面维护",
        base_override_data=preview,
    )
    return {"success": True, "run_id": run_id, "preview": preview, "rule_lists": saved}


_FBU_ATTENDANCE_DIRECT_UPLOAD_KINDS = {"attendance", "previous_attendance"}
_FBU_UPLOAD_JOB_KINDS = {
    "attendance",
    "previousAttendance",
    "previousSalary",
    "currentSalary",
    "salaryAdjustments",
    "transferHistory",
    "supplementalLeave",
    "performance",
    "adjustments",
    "baseOverrides",
}
_FBU_UPLOAD_JOB_CORE_FIELDS = {
    "attendance": ("attendance_file",),
    "previousAttendance": ("previous_attendance_file",),
    "previousSalary": ("previous_salary_file",),
    "currentSalary": ("current_salary_file", "salary_file"),
    "salaryAdjustments": ("adjustment_file",),
    "transferHistory": ("transfer_file",),
    "supplementalLeave": ("supplemental_leave_file",),
    "performance": ("performance_file",),
    "adjustments": ("adjustment_file",),
    "baseOverrides": ("base_override_file",),
}


def _fbu_upload_job_store() -> FBUUploadJobStore:
    return FBUUploadJobStore(FBU_PERFORMANCE_RUNS_DIR, fbu_run_manager)


def _fbu_upload_job_specs(payload: dict, job_id: str) -> list[dict]:
    raw_files = payload.get("files")
    if not isinstance(raw_files, list) or not raw_files:
        raise HTTPException(400, "请至少选择一份上传文件。")
    if len(raw_files) > 2:
        raise HTTPException(400, "单次最多上传两份文件。")

    specs = []
    seen_kinds = set()
    total_size = 0
    for raw in raw_files:
        if not isinstance(raw, dict):
            raise HTTPException(400, "上传文件信息格式不正确。")
        kind = str(raw.get("kind") or "").strip()
        if kind not in _FBU_UPLOAD_JOB_KINDS:
            raise HTTPException(400, f"不支持的上传材料类型：{kind or '未知'}")
        if kind in seen_kinds:
            raise HTTPException(400, "同类材料不能重复上传。")
        seen_kinds.add(kind)

        original_name = Path(str(raw.get("fileName") or "").replace("\\", "/")).name
        suffix = Path(original_name).suffix.lower()
        if suffix not in {".xlsx", ".xls"}:
            raise HTTPException(400, f"请上传 .xlsx 或 .xls 格式文件：{original_name or '未知文件'}")
        try:
            file_size = int(raw.get("fileSize") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, f"上传文件大小无效：{original_name}") from exc
        if file_size <= 0:
            raise HTTPException(400, f"上传文件不能为空：{original_name}")
        total_size += file_size
        specs.append({
            "kind": kind,
            "originalFilename": original_name,
            "relativePath": f"direct_uploads/{job_id}_{kind}{suffix}",
            "size": file_size,
            "contentType": str(raw.get("contentType") or "application/octet-stream"),
        })

    attendance_kinds = {"attendance", "previousAttendance"}
    if len(specs) > 1 and not seen_kinds.issubset(attendance_kinds):
        raise HTTPException(400, "除当月与上月考勤外，请逐份上传材料。")
    max_bytes = _fbu_attendance_direct_upload_max_bytes()
    if total_size > max_bytes:
        raise HTTPException(413, f"上传文件总大小超过当前上限 {max_bytes} 字节。")
    return specs


@app.post("/api/fbu-performance/runs/{run_id}/uploads/plan")
def create_fbu_upload_job_plan(run_id: str, payload: dict = Body(...)) -> dict:
    if not fbu_persistent_storage_enabled():
        raise HTTPException(409, "当前环境未启用 Supabase 直传。")
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    job_id = secrets.token_hex(12)
    specs = _fbu_upload_job_specs(payload, job_id)
    try:
        uploads = [
            {**create_fbu_signed_upload(run_id, spec["relativePath"]), **spec}
            for spec in specs
        ]
        job = _fbu_upload_job_store().create(run_id, specs, job_id=job_id)
    except Exception as exc:
        fbu_logger.exception("Failed to create FBU upload job %s", job_id)
        raise HTTPException(503, "生成文件直传任务失败，请稍后重试。") from exc
    return {
        "runId": run_id,
        "job": _fbu_upload_job_store().public(job),
        "uploads": uploads,
    }


def _fbu_upload_job_result(step: str, result: dict, uploads: list[dict]) -> dict:
    summary = (
        (result.get("preview") or {}).get("summary")
        or (result.get("verification") or {}).get("summary")
        or (result.get("material_preview") or {}).get("summary")
        or {}
    )
    core_updates = {}
    for upload in uploads:
        filename = str(upload.get("originalFilename") or "").strip()
        if not filename:
            continue
        for field_name in _FBU_UPLOAD_JOB_CORE_FIELDS.get(
            str(upload.get("kind") or ""),
            (),
        ):
            core_updates[field_name] = filename
    return {
        "success": bool(result.get("success", True)),
        "step": step,
        "summary": summary,
        "readyForReconciliation": bool(result.get("ready_for_reconciliation")),
        "missingMaterials": list(result.get("missing_materials") or []),
        "coreUpdates": core_updates,
    }


async def _run_fbu_upload_operation(operation):
    """Run parser-heavy async imports on a worker thread, not the request loop."""
    return await asyncio.to_thread(asyncio.run, operation)


async def _await_fbu_job_with_heartbeat(
    awaitable,
    *,
    store: FBUUploadJobStore,
    run_id: str,
    job_id: str,
    stage: str,
    progress: int,
    message: str,
):
    """Keep durable job timestamps fresh while one parse/calculation is running."""
    task = asyncio.ensure_future(awaitable)
    while True:
        try:
            return await asyncio.wait_for(asyncio.shield(task), timeout=20.0)
        except asyncio.TimeoutError:
            try:
                await asyncio.to_thread(
                    store.update,
                    run_id,
                    job_id,
                    status="processing",
                    stage=stage,
                    progress=progress,
                    message=message,
                )
            except Exception:
                fbu_logger.exception("Failed to persist FBU job heartbeat %s", job_id)


async def _process_fbu_upload_job(run_id: str, job_id: str) -> None:
    processing_started = perf_counter()
    store = _fbu_upload_job_store()
    job = store.load(run_id, job_id, refresh=True)
    if not job:
        return
    try:
        await asyncio.to_thread(
            store.update,
            run_id,
            job_id,
            status="processing",
            stage="materializing",
            progress=10,
            message="正在读取已上传文件",
            error="",
        )
        uploads = job.get("uploads")
        if not isinstance(uploads, list) or not uploads:
            raise ValueError("上传任务没有待处理文件")
        upload_kinds = [str(item.get("kind") or "") for item in uploads]
        total_bytes = sum(int(item.get("size") or 0) for item in uploads)

        materialize_started = perf_counter()
        materialized: dict[str, tuple[dict, Path]] = {}
        for item in uploads:
            kind = str(item.get("kind") or "")
            relative_path = str(item.get("relativePath") or "")
            path = await asyncio.to_thread(
                fbu_run_manager.materialize_file,
                run_id,
                relative_path,
            )
            if not path or not path.is_file():
                raise ValueError(f"未找到已上传文件：{item.get('originalFilename') or kind}")
            expected_size = int(item.get("size") or 0)
            if expected_size and path.stat().st_size != expected_size:
                raise ValueError(f"上传文件大小不一致：{item.get('originalFilename') or kind}")
            materialized[kind] = (item, path)
        materialize_ms = (perf_counter() - materialize_started) * 1000

        await asyncio.to_thread(
            store.update,
            run_id,
            job_id,
            status="processing",
            stage="parsing",
            progress=55,
            message="正在解析并核验数据",
        )

        result: dict = {}
        step = ""
        parse_started = perf_counter()
        with ExitStack() as stack:
            upload_files = {}
            for kind, (item, path) in materialized.items():
                upload = UploadFile(
                    file=stack.enter_context(path.open("rb")),
                    filename=str(item.get("originalFilename") or path.name),
                )
                upload._fbu_persisted_relative_path = str(item.get("relativePath") or "")
                upload_files[kind] = upload
            if set(upload_files).issubset({"attendance", "previousAttendance"}):
                run = fbu_run_manager.get_run(run_id, sections=set())
                if not run:
                    raise ValueError("任务不存在")
                operation = import_fbu_attendance(
                    file=upload_files.get("attendance"),
                    previous_attendance=upload_files.get("previousAttendance"),
                    calc_month=run.calc_month,
                    roster=None,
                    run_id=run_id,
                )
                result = await _await_fbu_job_with_heartbeat(
                    _run_fbu_upload_operation(operation),
                    store=store,
                    run_id=run_id,
                    job_id=job_id,
                    stage="parsing",
                    progress=55,
                    message="正在解析并核验数据",
                )
                step = "attendance"
            else:
                kind = next(iter(upload_files))
                upload = upload_files[kind]
                if kind in {"previousSalary", "currentSalary", "salaryAdjustments"}:
                    operation = import_fbu_salary_history_material(
                        run_id=run_id, material_type=kind, file=upload
                    )
                    step = "salary"
                elif kind == "transferHistory":
                    operation = import_fbu_transfer_history(run_id=run_id, file=upload)
                    step = "salary"
                elif kind == "supplementalLeave":
                    operation = import_fbu_supplemental_leave(run_id=run_id, file=upload)
                    step = "attendance"
                elif kind == "performance":
                    operation = import_fbu_performance(run_id=run_id, file=upload)
                    step = "performance"
                elif kind == "adjustments":
                    operation = import_fbu_adjustments(run_id=run_id, file=upload)
                    step = "salary"
                elif kind == "baseOverrides":
                    operation = import_fbu_base_overrides(run_id=run_id, file=upload)
                    step = "salary"
                else:
                    raise ValueError(f"不支持的上传材料类型：{kind}")
                result = await _await_fbu_job_with_heartbeat(
                    _run_fbu_upload_operation(operation),
                    store=store,
                    run_id=run_id,
                    job_id=job_id,
                    stage="parsing",
                    progress=55,
                    message="正在解析并核验数据",
                )

        parse_ms = (perf_counter() - parse_started) * 1000
        await asyncio.to_thread(
            store.update,
            run_id,
            job_id,
            status="completed",
            stage="completed",
            progress=100,
            message="上传并解析完成",
            result=_fbu_upload_job_result(step, result, uploads),
            completedAt=datetime.now().isoformat(),
            error="",
        )
        temporary_paths = [str(item.get("relativePath") or "") for item in uploads]
        try:
            await asyncio.to_thread(
                fbu_run_manager.delete_persisted_files,
                run_id,
                temporary_paths,
            )
            for _, path in materialized.values():
                path.unlink(missing_ok=True)
        except Exception:
            fbu_logger.exception("Failed to clean FBU upload job files %s", job_id)
        fbu_logger.info(
            "FBU upload processing run=%s job=%s kinds=%s bytes=%d "
            "materialize_ms=%.1f parse_ms=%.1f total_ms=%.1f",
            run_id,
            job_id,
            ",".join(upload_kinds),
            total_bytes,
            materialize_ms,
            parse_ms,
            (perf_counter() - processing_started) * 1000,
        )
    except Exception as exc:
        detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
        fbu_logger.exception("FBU upload job failed run=%s job=%s", run_id, job_id)
        try:
            await asyncio.to_thread(
                store.update,
                run_id,
                job_id,
                status="failed",
                stage="failed",
                progress=100,
                message="文件解析失败",
                error=str(detail or "未知错误"),
            )
        except Exception:
            fbu_logger.exception("Failed to persist FBU upload job failure %s", job_id)


def _prepare_fbu_upload_job(run_id: str, job_id: str) -> tuple[dict, bool]:
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")
    store = _fbu_upload_job_store()
    job = store.load(run_id, job_id, refresh=True)
    if not job:
        raise HTTPException(404, "上传任务不存在")
    public = store.public(job)
    if job.get("status") == "completed":
        return public, False
    if job.get("status") in {"queued", "processing"} and not public["recoverable"]:
        return public, False
    allowed_from = (
        ["queued", "processing"]
        if public["recoverable"]
        else ["uploading", "failed"]
    )
    job = store.update(
        run_id,
        job_id,
        allowed_from=allowed_from,
        status="queued",
        stage="queued",
        progress=5,
        message="正在排队",
        attempt=int(job.get("attempt") or 0) + 1,
        error="",
    )
    transition_applied = bool(job.pop("__transition_applied", True))
    return store.public(job), transition_applied and job.get("status") == "queued"


@app.post(
    "/api/fbu-performance/runs/{run_id}/uploads/{job_id}/start",
    status_code=202,
)
@app.post(
    "/api/fbu-performance/runs/{run_id}/uploads/{job_id}/resume",
    status_code=202,
)
async def start_fbu_upload_job(
    run_id: str,
    job_id: str,
    request: Request,
    payload: dict = Body(default=None),
) -> dict:
    request_started = perf_counter()
    try:
        job, should_process = _prepare_fbu_upload_job(run_id, job_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if should_process:
        await _process_fbu_upload_job(run_id, job_id)
        persisted = _fbu_upload_job_store().load(run_id, job_id, refresh=True)
        if persisted is None:
            raise HTTPException(500, "上传任务状态丢失，请重新上传。")
        job = _fbu_upload_job_store().public(persisted)
    uploads = job.get("uploads") if isinstance(job.get("uploads"), list) else []
    client_upload_ms = 0.0
    try:
        client_upload_ms = min(
            3_600_000.0,
            max(0.0, float((payload or {}).get("clientUploadMs") or 0)),
        )
    except (TypeError, ValueError):
        client_upload_ms = 0.0
    fbu_logger.info(
        "FBU upload request run=%s job=%s action=%s kinds=%s bytes=%d "
        "client_upload_ms=%.1f processing_ms=%.1f status=%s",
        run_id,
        job_id,
        "resume" if request.url.path.endswith("/resume") else "start",
        ",".join(str(item.get("kind") or "") for item in uploads),
        sum(int(item.get("size") or 0) for item in uploads),
        client_upload_ms,
        (perf_counter() - request_started) * 1000,
        job.get("status") or "",
    )
    return {"job": job}


@app.get("/api/fbu-performance/runs/{run_id}/uploads/{job_id}")
def get_fbu_upload_job(run_id: str, job_id: str) -> dict:
    try:
        store = _fbu_upload_job_store()
        job = store.load(run_id, job_id, refresh=True)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if not job:
        raise HTTPException(404, "上传任务不存在")
    return {"job": store.public(job)}


def _fbu_attendance_direct_upload_max_bytes() -> int:
    raw = os.environ.get("SIGMA_FBU_MAX_UPLOAD_BYTES", "262144000")
    try:
        return max(1, int(raw))
    except (TypeError, ValueError):
        return 262144000


def _fbu_attendance_direct_upload_specs(payload: dict, plan_id: str) -> list[dict]:
    raw_files = payload.get("files")
    if not isinstance(raw_files, list) or not raw_files:
        raise HTTPException(400, "请至少选择一份考勤文件。")
    if len(raw_files) > 2:
        raise HTTPException(400, "单次最多上传当月考勤和上月考勤各一份。")

    specs = []
    seen_kinds = set()
    total_size = 0
    for raw in raw_files:
        if not isinstance(raw, dict):
            raise HTTPException(400, "上传文件信息格式不正确。")
        kind = str(raw.get("kind") or "").strip()
        if kind not in _FBU_ATTENDANCE_DIRECT_UPLOAD_KINDS:
            raise HTTPException(400, "考勤文件类型不正确。")
        if kind in seen_kinds:
            raise HTTPException(400, "同类考勤文件不能重复上传。")
        seen_kinds.add(kind)

        original_name = Path(str(raw.get("fileName") or "").replace("\\", "/")).name
        suffix = Path(original_name).suffix.lower()
        if suffix not in {".xlsx", ".xls"}:
            raise HTTPException(400, f"请上传 .xlsx 或 .xls 格式的考勤文件：{original_name or '未知文件'}")
        try:
            file_size = int(raw.get("fileSize") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(400, f"上传文件大小无效：{original_name}") from exc
        if file_size <= 0:
            raise HTTPException(400, f"上传文件不能为空：{original_name}")
        total_size += file_size
        specs.append({
            "kind": kind,
            "originalFilename": original_name,
            "relativePath": f"direct_uploads/{plan_id}_{kind}{suffix}",
            "size": file_size,
            "contentType": str(raw.get("contentType") or "application/octet-stream"),
        })

    max_bytes = _fbu_attendance_direct_upload_max_bytes()
    if total_size > max_bytes:
        raise HTTPException(413, f"上传文件总大小超过当前上限 {max_bytes} 字节。")
    return specs


def _fbu_attendance_direct_plan_relative_path(plan_id: str) -> str:
    if not re.fullmatch(r"[0-9a-f]{24}", str(plan_id or "")):
        raise HTTPException(400, "直传计划编号无效。")
    return f"direct_uploads/{plan_id}.json"


def _load_fbu_attendance_direct_plan(run_id: str, plan_id: str) -> tuple[dict, Path]:
    relative_path = _fbu_attendance_direct_plan_relative_path(plan_id)
    try:
        plan_path = fbu_run_manager.materialize_file(run_id, relative_path)
    except Exception as exc:
        fbu_logger.exception("Failed to materialize FBU attendance upload plan %s", plan_id)
        raise HTTPException(503, "读取考勤直传计划失败，请稍后重试。") from exc
    if not plan_path or not plan_path.is_file():
        raise HTTPException(404, "考勤直传计划不存在或已过期。")
    try:
        payload = json.loads(plan_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise HTTPException(400, "考勤直传计划内容异常，请重新上传。") from exc
    if not isinstance(payload, dict) or payload.get("runId") != run_id or payload.get("planId") != plan_id:
        raise HTTPException(400, "考勤直传计划与当前活动不匹配。")
    return payload, plan_path


def _fbu_attendance_direct_result(run: FBURun) -> dict:
    return {
        "success": True,
        "run_id": run.run_id,
        "step": 1,
        "preview": run.attendance_data,
        "result_file": _fbu_result_file_payload(
            run.run_id,
            "attendance",
            run.calc_month,
        ),
    }


@app.post("/api/fbu-performance/runs/{run_id}/attendance-direct-upload-plan")
def create_fbu_attendance_direct_upload_plan(run_id: str, payload: dict = Body(...)) -> dict:
    if not fbu_persistent_storage_enabled():
        raise HTTPException(409, "当前环境未启用 Supabase 直传。")
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    plan_id = secrets.token_hex(12)
    specs = _fbu_attendance_direct_upload_specs(payload, plan_id)
    try:
        uploads = [
            {**create_fbu_signed_upload(run_id, spec["relativePath"]), **spec}
            for spec in specs
        ]
    except Exception as exc:
        fbu_logger.exception("Failed to create FBU attendance signed upload for %s", run_id)
        raise HTTPException(503, "生成考勤直传地址失败，请稍后重试。") from exc

    relative_plan_path = _fbu_attendance_direct_plan_relative_path(plan_id)
    plan_path = FBU_PERFORMANCE_RUNS_DIR / run_id / relative_plan_path
    plan_path.parent.mkdir(parents=True, exist_ok=True)
    plan_payload = {
        "planId": plan_id,
        "runId": run_id,
        "calcMonth": run.calc_month,
        "status": "pending",
        "createdAt": datetime.now().isoformat(),
        "uploads": specs,
    }
    plan_path.write_text(
        json.dumps(plan_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    try:
        fbu_run_manager.persist_files(run_id, [relative_plan_path])
    except Exception as exc:
        plan_path.unlink(missing_ok=True)
        fbu_logger.exception("Failed to persist FBU attendance upload plan %s", plan_id)
        raise HTTPException(503, "保存考勤直传计划失败，请稍后重试。") from exc
    return {"runId": run_id, "planId": plan_id, "uploads": uploads}


@app.post("/api/fbu-performance/runs/{run_id}/attendance-direct-upload-complete")
async def complete_fbu_attendance_direct_upload(run_id: str, payload: dict = Body(...)) -> dict:
    if not fbu_persistent_storage_enabled():
        raise HTTPException(409, "当前环境未启用 Supabase 直传。")
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    plan_id = str(payload.get("planId") or "").strip()
    plan, plan_path = _load_fbu_attendance_direct_plan(run_id, plan_id)
    if plan.get("status") == "completed":
        refreshed_run = fbu_run_manager.get_run(run_id, sections={"attendance_data"})
        if not refreshed_run or not refreshed_run.attendance_data:
            raise HTTPException(409, "考勤直传已登记，但活动结果尚未就绪，请稍后重试。")
        return _fbu_attendance_direct_result(refreshed_run)

    uploads = plan.get("uploads")
    if not isinstance(uploads, list) or not uploads:
        raise HTTPException(400, "考勤直传计划没有待处理文件。")

    materialized: dict[str, tuple[dict, Path]] = {}
    try:
        for item in uploads:
            if not isinstance(item, dict):
                raise HTTPException(400, "考勤直传计划内容异常，请重新上传。")
            kind = str(item.get("kind") or "")
            relative_path = str(item.get("relativePath") or "")
            if kind not in _FBU_ATTENDANCE_DIRECT_UPLOAD_KINDS:
                raise HTTPException(400, "考勤直传计划文件类型异常。")
            uploaded_path = fbu_run_manager.materialize_file(run_id, relative_path)
            if not uploaded_path or not uploaded_path.is_file():
                raise HTTPException(400, f"未找到已上传文件：{item.get('originalFilename') or kind}")
            expected_size = int(item.get("size") or 0)
            actual_size = uploaded_path.stat().st_size
            if expected_size and actual_size != expected_size:
                uploaded_path.unlink(missing_ok=True)
                raise HTTPException(
                    400,
                    f"上传文件大小不一致（{item.get('originalFilename') or kind}："
                    f"预期 {expected_size}，实际 {actual_size}），请重新上传。",
                )
            materialized[kind] = (item, uploaded_path)
    except HTTPException:
        raise
    except Exception as exc:
        fbu_logger.exception("Failed to materialize FBU attendance direct upload %s", plan_id)
        raise HTTPException(503, "读取已上传考勤文件失败，请稍后重试。") from exc

    with ExitStack() as stack:
        attendance_upload = None
        previous_upload = None
        if "attendance" in materialized:
            item, path = materialized["attendance"]
            attendance_upload = UploadFile(
                file=stack.enter_context(path.open("rb")),
                filename=str(item.get("originalFilename") or path.name),
            )
            attendance_upload._fbu_persisted_relative_path = str(item.get("relativePath") or "")
        if "previous_attendance" in materialized:
            item, path = materialized["previous_attendance"]
            previous_upload = UploadFile(
                file=stack.enter_context(path.open("rb")),
                filename=str(item.get("originalFilename") or path.name),
            )
            previous_upload._fbu_persisted_relative_path = str(item.get("relativePath") or "")
        result = await import_fbu_attendance(
            file=attendance_upload,
            previous_attendance=previous_upload,
            calc_month=run.calc_month,
            roster=None,
            run_id=run_id,
        )

    plan["status"] = "completed"
    plan["completedAt"] = datetime.now().isoformat()
    plan_path.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    completion_recorded = False
    relative_plan_path = _fbu_attendance_direct_plan_relative_path(plan_id)
    try:
        fbu_run_manager.persist_files(run_id, [relative_plan_path])
        completion_recorded = True
    except Exception:
        fbu_logger.exception("Failed to persist completed FBU attendance upload plan %s", plan_id)

    if completion_recorded:
        temporary_paths = [str(item["relativePath"]) for item in uploads]
        try:
            fbu_run_manager.delete_persisted_files(run_id, temporary_paths)
        except Exception:
            fbu_logger.exception("Failed to clean FBU attendance direct upload %s", plan_id)
        for _, path in materialized.values():
            path.unlink(missing_ok=True)
    return result


def _persist_fbu_uploaded_file(
    run_id: str,
    upload: UploadFile,
    destination_relative_path: str,
) -> None:
    """Reuse a direct-upload object when available, with byte upload as the fallback."""
    source_relative_path = str(
        getattr(upload, "_fbu_persisted_relative_path", "") or ""
    ).strip()
    if source_relative_path:
        if fbu_run_manager.promote_persisted_file(
            run_id,
            source_relative_path,
            destination_relative_path,
        ):
            return
        fbu_logger.warning(
            "FBU direct upload promotion fallback run=%s source=%s destination=%s",
            run_id,
            source_relative_path,
            destination_relative_path,
        )
    fbu_run_manager.persist_files(run_id, [destination_relative_path])


def _log_fbu_attendance_import(
    run_id: str,
    mode: str,
    uploaded_bytes: int,
    timings: dict[str, float],
    started_at: float,
) -> None:
    fbu_logger.info(
        "FBU attendance import run=%s mode=%s bytes=%d "
        "run_load_ms=%.1f stage_ms=%.1f roster_ms=%.1f workbook_ms=%.1f "
        "context_ms=%.1f state_ms=%.1f file_ms=%.1f result_ms=%.1f total_ms=%.1f",
        run_id,
        mode,
        uploaded_bytes,
        timings.get("run_load_ms", 0.0),
        timings.get("stage_ms", 0.0),
        timings.get("roster_ms", 0.0),
        timings.get("workbook_ms", 0.0),
        timings.get("context_ms", 0.0),
        timings.get("state_ms", 0.0),
        timings.get("file_ms", 0.0),
        timings.get("result_ms", 0.0),
        (perf_counter() - started_at) * 1000,
    )


@app.post("/api/fbu-performance/import-attendance")
async def import_fbu_attendance(
    file: UploadFile = File(None),
    previous_attendance: UploadFile = File(None),
    calc_month: str = Body(...),
    roster: UploadFile = File(None),
    run_id: str = Body(None),
) -> dict:
    """Step 1: 导入考勤日报表"""
    import_started = perf_counter()
    timings: dict[str, float] = {}
    mode = (
        "current+previous"
        if file and previous_attendance
        else "current"
        if file
        else "previous"
        if previous_attendance
        else "existing"
    )
    # 获取或创建运行记录
    phase_started = perf_counter()
    if run_id:
        run_sections = {"hourly_rate_policy_data"}
        if file is None:
            run_sections.add("attendance_data")
        run = fbu_run_manager.get_run(
            run_id,
            sections=run_sections,
        )
        if not run:
            raise HTTPException(404, "任务不存在")
    else:
        run = fbu_run_manager.create_run(calc_month=calc_month)
    timings["run_load_ms"] = (perf_counter() - phase_started) * 1000

    # 先解析临时文件，成功后再替换活动正式文件。
    phase_started = perf_counter()
    run_dir = FBU_PERFORMANCE_RUNS_DIR / run.run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    file_path = run_dir / "attendance.xlsx"
    pending_file_path = run_dir / ".attendance-upload.xlsx"
    current_attendance_filename = run.attendance_file or ""
    if file:
        pending_file_path.write_bytes(await file.read())
        current_attendance_filename = file.filename
    elif not file_path.exists():
        await asyncio.to_thread(fbu_run_manager.materialize_file, run.run_id, "attendance.xlsx")
        if not file_path.exists():
            raise HTTPException(400, "请先上传当月考勤日报表")
    attendance_parse_path = pending_file_path if file else file_path

    previous_attendance_filename = run.previous_attendance_file or ""
    previous_path = None
    pending_previous_path = None
    if previous_attendance:
        previous_suffix = Path(previous_attendance.filename or "").suffix.lower()
        if previous_suffix not in {".xlsx", ".xls"}:
            raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的上一月考勤日报表")
        previous_attendance_filename = previous_attendance.filename
        previous_path = run_dir / f"previous_attendance{previous_suffix}"
        pending_previous_path = run_dir / f".previous-attendance-upload{previous_suffix}"
        pending_previous_path.write_bytes(await previous_attendance.read())

    roster_path = None
    pending_roster_path = None
    if roster:
        roster_suffix = Path(roster.filename or "").suffix.lower()
        if roster_suffix not in {".xlsx", ".xls"}:
            raise HTTPException(400, "请上传 .xlsx 或 .xls 格式的花名册")
        roster_path = run_dir / f"roster{roster_suffix}"
        pending_roster_path = run_dir / f".roster-upload{roster_suffix}"
        pending_roster_path.write_bytes(await roster.read())
    timings["stage_ms"] = (perf_counter() - phase_started) * 1000
    uploaded_bytes = sum(
        path.stat().st_size
        for path in (pending_file_path, pending_previous_path, pending_roster_path)
        if path and path.is_file()
    )

    if previous_attendance and not file and run.attendance_data and pending_previous_path is not None:
        try:
            phase_started = perf_counter()
            preview = _append_fbu_previous_attendance_context_to_preview(
                dict(run.attendance_data),
                pending_previous_path,
                calc_month,
                previous_attendance_filename,
            )
            timings["workbook_ms"] = (perf_counter() - phase_started) * 1000
            phase_started = perf_counter()
            pending_previous_path.replace(previous_path)
            fbu_run_manager.save_attendance_import(
                run.run_id,
                preview,
                previous_attendance_file=previous_attendance_filename,
                hourly_rate_policy_data=build_hourly_rate_policy_data(
                    preview,
                    calc_month,
                    run.hourly_rate_policy_data,
                ),
            )
            timings["state_ms"] = (perf_counter() - phase_started) * 1000
            phase_started = perf_counter()
            _persist_fbu_uploaded_file(
                run.run_id,
                previous_attendance,
                previous_path.name,
            )
            timings["file_ms"] = (perf_counter() - phase_started) * 1000
            phase_started = perf_counter()
            result_file = _fbu_result_file_payload(
                run.run_id,
                "attendance",
                run.calc_month,
            )
            timings["result_ms"] = (perf_counter() - phase_started) * 1000
            _log_fbu_attendance_import(
                run.run_id,
                mode,
                uploaded_bytes,
                timings,
                import_started,
            )
            return {
                "success": True,
                "run_id": run.run_id,
                "step": 1,
                "preview": preview,
                "result_file": result_file,
            }
        except Exception as e:
            raise HTTPException(500, f"上一月考勤解析失败: {str(e)}")
        finally:
            pending_previous_path.unlink(missing_ok=True)

    # 解析并预览
    try:
        target_month = int(calc_month.split("-")[1]) if "-" in calc_month else int(calc_month)
        parser = FBUPerformanceParser()

        # 加载本活动花名册；没有时自动引用当前基础花名册
        phase_started = perf_counter()
        if pending_roster_path and pending_roster_path.exists():
            parser.load_roster(str(pending_roster_path))
        else:
            _load_fbu_roster_for_run(parser, run.run_id)
        timings["roster_ms"] = (perf_counter() - phase_started) * 1000

        phase_started = perf_counter()
        preview = parser.parse_attendance_preview(str(attendance_parse_path), target_month)
        timings["workbook_ms"] = (perf_counter() - phase_started) * 1000
        if not preview.get("employees"):
            raise HTTPException(
                400,
                f"考勤日报未包含 {calc_month} 的数据，请确认活动月份或重新上传文件",
            )
        phase_started = perf_counter()
        if previous_attendance and pending_previous_path is not None:
            preview = _append_fbu_previous_attendance_context_to_preview(
                preview,
                pending_previous_path,
                calc_month,
                previous_attendance_filename,
            )
        else:
            preview.setdefault("summary", {})["attendance_context"] = _build_fbu_attendance_context_summary(
                attendance_parse_path,
                calc_month,
                previous_attendance_filename,
                available_dates=_fbu_attendance_dates_from_preview(preview),
            )
        timings["context_ms"] = (perf_counter() - phase_started) * 1000

        phase_started = perf_counter()
        if file:
            pending_file_path.replace(file_path)
        if pending_previous_path and previous_path:
            pending_previous_path.replace(previous_path)
        if pending_roster_path and roster_path:
            pending_roster_path.replace(roster_path)

        metadata = {
            "attendance_file": current_attendance_filename,
            "previous_attendance_file": previous_attendance_filename,
        }
        if roster:
            metadata.update(roster_file=roster.filename, roster_source="activity")
        metadata["hourly_rate_policy_data"] = build_hourly_rate_policy_data(
            preview,
            calc_month,
            run.hourly_rate_policy_data,
        )
        fbu_run_manager.save_attendance_import(run.run_id, preview, **metadata)
        timings["state_ms"] = (perf_counter() - phase_started) * 1000
        phase_started = perf_counter()
        if file:
            _persist_fbu_uploaded_file(run.run_id, file, "attendance.xlsx")
        if previous_attendance and previous_path and previous_path.exists():
            _persist_fbu_uploaded_file(
                run.run_id,
                previous_attendance,
                previous_path.name,
            )
        if roster and roster_path and roster_path.exists():
            _persist_fbu_uploaded_file(run.run_id, roster, roster_path.name)
        timings["file_ms"] = (perf_counter() - phase_started) * 1000
        phase_started = perf_counter()
        result_file = _fbu_result_file_payload(
            run.run_id,
            "attendance",
            run.calc_month,
        )
        timings["result_ms"] = (perf_counter() - phase_started) * 1000
        _log_fbu_attendance_import(
            run.run_id,
            mode,
            uploaded_bytes,
            timings,
            import_started,
        )

        return {
            "success": True,
            "run_id": run.run_id,
            "step": 1,
            "preview": preview,
            "result_file": result_file,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"考勤数据解析失败: {str(e)}")
    finally:
        pending_file_path.unlink(missing_ok=True)
        if pending_previous_path:
            pending_previous_path.unlink(missing_ok=True)
        if pending_roster_path:
            pending_roster_path.unlink(missing_ok=True)


def _fbu_hourly_rate_policy_response(policy_data: dict | None) -> dict:
    if not isinstance(policy_data, dict) or not policy_data:
        return {}
    rows = [
        dict(row)
        for row in policy_data.get("rows", [])
        if isinstance(row, dict) and row.get("visible")
    ]
    hidden_employee_ids = sorted({
        str(row.get("employee_id") or "").strip()
        for row in policy_data.get("rows", [])
        if (
            isinstance(row, dict)
            and not row.get("visible")
            and str(row.get("employee_id") or "").strip()
        )
    })
    summary = dict(policy_data.get("summary") or {})
    summary["visible_count"] = len(rows)
    return {
        "rows": rows,
        "hidden_employee_ids": hidden_employee_ids,
        "summary": summary,
    }


@app.post("/api/fbu-performance/runs/{run_id}/hourly-rate-policies")
def update_fbu_hourly_rate_policies(run_id: str, body: dict = Body(...)) -> dict:
    """Update non-blocking payroll-period hourly-rate suggestions."""
    run = fbu_run_manager.get_run(
        run_id,
        sections={"attendance_data", "hourly_rate_policy_data"},
    )
    if not run:
        raise HTTPException(404, "任务不存在")
    if not run.attendance_data:
        raise HTTPException(409, "请先上传考勤日报")
    try:
        policy_data = update_hourly_rate_policy_data(
            run.hourly_rate_policy_data,
            action=str(body.get("action") or ""),
            row_id=str(body.get("row_id") or ""),
            employee_id=str(body.get("employee_id") or ""),
            selected_policy=str(body.get("selected_policy") or ""),
        )
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    fbu_run_manager.update_run(run_id, hourly_rate_policy_data=policy_data)
    return {
        "success": True,
        "run_id": run_id,
        "hourly_rate_policy_data": _fbu_hourly_rate_policy_response(policy_data),
    }


def _build_period_adjustment_preview(rows: list[dict]) -> dict:
    normalized_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("source_month") or ""),
            str(row.get("employee_id") or ""),
        ),
    )
    return {
        "rows": normalized_rows,
        "summary": {
            "count": len(normalized_rows),
            "total_amount": round(
                sum(float(row.get("amount") or 0) for row in normalized_rows),
                2,
            ),
        },
    }


@app.post("/api/fbu-performance/runs/{run_id}/period-adjustments")
def update_fbu_period_adjustments(run_id: str, body: dict = Body(...)) -> dict:
    """Maintain the optional employee-level performance-base difference."""
    run = fbu_run_manager.get_run(
        run_id,
        sections={"salary_data", "period_adjustment_data"},
    )
    if not run:
        raise HTTPException(404, "任务不存在")
    if not run.salary_data:
        raise HTTPException(409, "请先上传当月薪资档案")

    action = str(body.get("action") or "").strip()
    employee_id = normalize_shift_employee_id(body.get("employee_id"))
    if not employee_id:
        raise HTTPException(400, "请输入员工工号")

    existing_rows = [
        dict(row) for row in (run.period_adjustment_data or {}).get("rows", [])
    ]
    row_index = next(
        (
            index for index, row in enumerate(existing_rows)
            if normalize_shift_employee_id(row.get("employee_id")) == employee_id
        ),
        None,
    )

    if action == "delete":
        if row_index is None:
            raise HTTPException(404, "未找到该员工的 Period adjustment")
        existing_rows.pop(row_index)
    elif action == "upsert":
        try:
            amount = round(float(body.get("amount")), 2)
        except (TypeError, ValueError):
            raise HTTPException(400, "调整额必须为数字")
        if amount == 0:
            raise HTTPException(400, "调整额不能为0")

        source_month = str(body.get("source_month") or "").strip()
        if not re.match(r"^\d{4}-(0[1-9]|1[0-2])$", source_month):
            raise HTTPException(400, "归属月份格式应为YYYY-MM")
        reason = str(body.get("reason") or "").strip()
        if not reason:
            raise HTTPException(400, "请填写调整原因")

        salary_lookup = {
            normalize_shift_employee_id(row.get("source_employee_id") or row.get("employee_id")): row
            for row in (run.salary_data or {}).get("employees", [])
        }
        salary_row = salary_lookup.get(employee_id)
        if not salary_row:
            raise HTTPException(400, "当月薪资档案中未找到该工号")
        normalized = {
            "employee_id": employee_id,
            "name": str(salary_row.get("name") or body.get("name") or "").strip(),
            "amount": amount,
            "source_month": source_month,
            "reason": reason,
        }
        if row_index is None:
            existing_rows.append(normalized)
        else:
            existing_rows[row_index] = normalized
    else:
        raise HTTPException(400, "操作类型无效")

    preview = _build_period_adjustment_preview(existing_rows)
    fbu_run_manager.update_run(run_id, period_adjustment_data=preview)
    return {
        "success": True,
        "run_id": run_id,
        "period_adjustment_data": preview,
    }


@app.post("/api/fbu-performance/import-salary")
async def import_fbu_salary(
    run_id: str = Body(...),
    file: UploadFile = File(...),
) -> dict:
    """Step 2: 导入薪资档案"""
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    # 保存上传文件
    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / "salary.xlsx"
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # 解析并预览
    try:
        parser = FBUPerformanceParser()

        # 加载活动花名册或基础花名册快照
        _load_fbu_roster_for_run(parser, run_id)

        preview = parser.parse_salary_preview(str(file_path))

        # 保存分步数据
        fbu_run_manager.save_step_data(run_id, 2, preview, salary_file=file.filename)
        fbu_run_manager.persist_files(run_id, ["salary.xlsx"])
        result_file = _fbu_result_file_payload(run_id, "salary")

        return {
            "success": True,
            "run_id": run_id,
            "step": 2,
            "preview": preview,
            "result_file": result_file,
        }
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"薪资数据解析失败: {str(e)}")


_FBU_SALARY_HISTORY_MATERIALS = {
    "previousSalary": {
        "path": "previous_salary.xlsx",
        "file_field": "previous_salary_file",
        "data_field": "previous_salary_data",
        "label": "上月薪资档案",
        "parser": "salary",
    },
    "currentSalary": {
        "path": "salary.xlsx",
        "file_field": "current_salary_file",
        "data_field": "current_salary_data",
        "label": "当月薪资档案",
        "parser": "salary",
    },
    "salaryAdjustments": {
        "path": "adjustments.xlsx",
        "file_field": "adjustment_file",
        "data_field": "adjustment_data",
        "label": "全量调薪流程",
        "parser": "adjustments",
    },
}


def _reconcile_fbu_salary_history(
    run,
    parser: FBUPerformanceParser,
    *,
    previous_preview: dict,
    current_preview: dict,
    adjustment_preview: dict,
) -> tuple[dict, dict]:
    verification = parser.reconcile_salary_history(
        previous_preview.get("employees", []),
        current_preview.get("employees", []),
        adjustment_preview.get("events", []),
        run.calc_month,
        roster_by_id=parser.employee_roster,
    )
    resolved_salary = {
        "employees": verification["employees"],
        "summary": {
            **current_preview.get("summary", {}),
            **verification["summary"],
        },
    }
    return resolved_salary, verification


@app.post("/api/fbu-performance/import-salary-history-material")
async def import_fbu_salary_history_material(
    run_id: str = Form(...),
    material_type: str = Form(...),
    file: UploadFile = File(...),
) -> dict:
    """Upload and parse one salary-history material, then reconcile when all three exist."""
    run = fbu_run_manager.get_run(run_id, sections=set(), refresh=True)
    if not run:
        raise HTTPException(404, "任务不存在")

    material = _FBU_SALARY_HISTORY_MATERIALS.get(material_type)
    if not material:
        raise HTTPException(400, "不支持的薪资材料类型")

    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / material["path"]
    with open(file_path, "wb") as file_handle:
        file_handle.write(await file.read())

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        if material["parser"] == "adjustments":
            material_preview = parser.parse_adjustments_preview(str(file_path))
        else:
            material_preview = parser.parse_salary_preview(str(file_path))

        refreshed_run = fbu_run_manager.get_run(
            run_id,
            sections={"previous_salary_data", "current_salary_data", "adjustment_data"},
            refresh=True,
        )
        if not refreshed_run:
            raise HTTPException(404, "任务不存在")

        material_previews = {
            key: (
                material_preview
                if key == material_type
                else getattr(refreshed_run, config["data_field"], {})
            )
            for key, config in _FBU_SALARY_HISTORY_MATERIALS.items()
        }
        material_files = {
            key: (
                file.filename
                if key == material_type
                else (
                    getattr(refreshed_run, config["file_field"], "")
                    or (config["path"] if material_previews[key] else "")
                )
            )
            for key, config in _FBU_SALARY_HISTORY_MATERIALS.items()
        }
        core_updates = {
            config["file_field"]: material_files[key]
            for key, config in _FBU_SALARY_HISTORY_MATERIALS.items()
            if material_files[key]
        }
        if material_files["currentSalary"]:
            core_updates["salary_file"] = material_files["currentSalary"]
        section_updates = {
            config["data_field"]: material_previews[key]
            for key, config in _FBU_SALARY_HISTORY_MATERIALS.items()
            if material_previews[key]
        }

        missing_materials = [
            key
            for key in _FBU_SALARY_HISTORY_MATERIALS
            if not material_files[key] or not material_previews[key]
        ]
        ready_for_reconciliation = not missing_materials
        resolved_salary = {}
        verification = {}
        if ready_for_reconciliation:
            resolved_salary, verification = _reconcile_fbu_salary_history(
                refreshed_run,
                parser,
                previous_preview=material_previews["previousSalary"],
                current_preview=material_previews["currentSalary"],
                adjustment_preview=material_previews["salaryAdjustments"],
            )

        fbu_run_manager.save_salary_history_import(
            run_id,
            core_updates=core_updates,
            section_updates=section_updates,
            ready_for_reconciliation=ready_for_reconciliation,
            resolved_salary=resolved_salary,
            verification=verification,
        )
        fbu_run_manager.persist_files(run_id, [material["path"]])

        return {
            "success": True,
            "run_id": run_id,
            "step": 2,
            "material_type": material_type,
            "material_label": material["label"],
            "material_preview": material_preview,
            "ready_for_reconciliation": ready_for_reconciliation,
            "missing_materials": missing_materials,
            "preview": resolved_salary,
            "verification": verification,
            "result_file": (
                _fbu_result_file_payload(run_id, "salary")
                if ready_for_reconciliation
                else None
            ),
        }
    except ValueError as exc:
        fbu_run_manager.update_run(run_id, status="failed", error=str(exc))
        raise HTTPException(400, str(exc))
    except Exception as exc:
        fbu_run_manager.update_run(run_id, status="failed", error=str(exc))
        raise HTTPException(500, f"{material['label']}解析失败: {str(exc)}")


@app.post("/api/fbu-performance/import-salary-history")
async def import_fbu_salary_history(
    run_id: str = Form(...),
    previous_salary: UploadFile = File(...),
    current_salary: UploadFile = File(...),
    adjustments: UploadFile = File(...),
    response_mode: str = Form(""),
) -> dict:
    """Step 2: import adjacent salary snapshots and the full adjustment export."""
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    uploads = [
        (previous_salary, run_dir / "previous_salary.xlsx"),
        (current_salary, run_dir / "salary.xlsx"),
        (adjustments, run_dir / "adjustments.xlsx"),
    ]
    for upload, path in uploads:
        with open(path, "wb") as file_handle:
            file_handle.write(await upload.read())

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        previous_preview = parser.parse_salary_preview(str(run_dir / "previous_salary.xlsx"))
        current_preview = parser.parse_salary_preview(str(run_dir / "salary.xlsx"))
        adjustment_preview = parser.parse_adjustments_preview(str(run_dir / "adjustments.xlsx"))
        verification = parser.reconcile_salary_history(
            previous_preview.get("employees", []),
            current_preview.get("employees", []),
            adjustment_preview.get("events", []),
            run.calc_month,
            roster_by_id=parser.employee_roster,
        )
        resolved_salary = {
            "employees": verification["employees"],
            "summary": {
                **current_preview.get("summary", {}),
                **verification["summary"],
            },
        }
        fbu_run_manager.save_step_data(
            run_id,
            2,
            resolved_salary,
            previous_salary_file=previous_salary.filename,
            current_salary_file=current_salary.filename,
            salary_file=current_salary.filename,
            adjustment_file=adjustments.filename,
            previous_salary_data=previous_preview,
            current_salary_data=current_preview,
            salary_verification_data=verification,
            adjustment_data=adjustment_preview,
        )
        fbu_run_manager.persist_files(
            run_id,
            ["previous_salary.xlsx", "salary.xlsx", "adjustments.xlsx"],
        )
        response_verification = verification
        if response_mode == "compact":
            response_verification = {
                "issues": verification.get("issues", []),
                "summary": verification.get("summary", {}),
            }
        return {
            "success": True,
            "run_id": run_id,
            "step": 2,
            "preview": resolved_salary,
            "verification": response_verification,
            "adjustment_preview": adjustment_preview,
            "result_file": _fbu_result_file_payload(run_id, "salary"),
        }
    except ValueError as exc:
        fbu_run_manager.update_run(run_id, status="failed", error=str(exc))
        raise HTTPException(400, str(exc))
    except Exception as exc:
        fbu_run_manager.update_run(run_id, status="failed", error=str(exc))
        raise HTTPException(500, f"薪资历史核验失败: {str(exc)}")


@app.post("/api/fbu-performance/import-transfer-history")
async def import_fbu_transfer_history(
    run_id: str = Form(...),
    file: UploadFile = File(...),
) -> dict:
    """导入人事调动记录；调动日期视为正式生效日期。"""
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / "transfer_history.xlsx"
    with open(file_path, "wb") as file_handle:
        file_handle.write(await file.read())

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        preview = parser.parse_transfer_history_preview(str(file_path), run.calc_month)
        fbu_run_manager.update_run(
            run_id,
            transfer_file=file.filename,
            transfer_data=preview,
            status="step2",
            error="",
        )
        fbu_run_manager.persist_files(run_id, ["transfer_history.xlsx"])
        return {
            "success": True,
            "run_id": run_id,
            "step": 2,
            "preview": preview,
        }
    except ValueError as exc:
        fbu_run_manager.update_run(run_id, status="failed", error=str(exc))
        raise HTTPException(400, str(exc))
    except Exception as exc:
        fbu_run_manager.update_run(run_id, status="failed", error=str(exc))
        raise HTTPException(500, f"人事调动记录解析失败: {str(exc)}")


@app.post("/api/fbu-performance/import-performance")
async def import_fbu_performance(
    run_id: str = Body(...),
    file: UploadFile = File(...),
) -> dict:
    """Step 3: 导入绩效报表"""
    run = fbu_run_manager.get_run(run_id, sections={"performance_data"})
    if not run:
        raise HTTPException(404, "任务不存在")

    # 保存上传文件
    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / "performance.xlsx"
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # 解析并预览
    try:
        parser = FBUPerformanceParser()

        # 加载活动花名册或基础花名册快照
        _load_fbu_roster_for_run(parser, run_id)

        preview = parser.parse_performance_preview(str(file_path))
        existing_performance = run.performance_data if isinstance(run.performance_data, dict) else {}
        if (
            preview.get("summary", {}).get("source_type") == "performance_supplement"
            and existing_performance.get("employees")
        ):
            preview = parser.merge_performance_supplement_preview(existing_performance, preview)

        # 保存分步数据
        fbu_run_manager.save_step_data(run_id, 3, preview, performance_file=file.filename)
        fbu_run_manager.persist_files(run_id, ["performance.xlsx"])
        result_file = _fbu_result_file_payload(run_id, "performance")

        return {
            "success": True,
            "run_id": run_id,
            "step": 3,
            "preview": preview,
            "result_file": result_file,
        }
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"绩效数据解析失败: {str(e)}")


@app.post("/api/fbu-performance/runs/{run_id}/salary-verification/confirm")
def confirm_fbu_salary_verification(run_id: str, body: dict = Body(...)) -> dict:
    """Resolve one or more salary snapshot differences using explicit snapshot choices."""
    run = fbu_run_manager.get_run(
        run_id,
        sections={"salary_verification_data", "salary_data"},
        refresh=True,
    )
    if not run:
        raise HTTPException(404, "任务不存在")

    raw_confirmations = body.get("confirmations")
    if isinstance(raw_confirmations, list) and raw_confirmations:
        confirmations = raw_confirmations
    else:
        confirmations = [body]
    if len(confirmations) > 100:
        raise HTTPException(400, "单次最多确认 100 条薪资差异")

    normalized_confirmations: dict[str, dict] = {}
    for confirmation in confirmations:
        employee_id = str(confirmation.get("employee_id") or "").strip()
        choice = str(confirmation.get("choice") or "").strip()
        if not employee_id or choice not in {"previous", "current", "ignore_current"}:
            raise HTTPException(400, "请选择按上月值、按当月值或忽略当月缺失")
        normalized_confirmations[employee_id] = {
            "employee_id": employee_id,
            "choice": choice,
            "note": str(confirmation.get("note") or "").strip(),
        }

    verification = dict(run.salary_verification_data or {})
    employees = [dict(row) for row in verification.get("employees", [])]
    employees_by_id = {str(row.get("employee_id") or ""): row for row in employees}
    updated_employees: list[dict] = []
    missing_employee_ids: list[str] = []
    for employee_id, confirmation in normalized_confirmations.items():
        target = employees_by_id.get(employee_id)
        if not target:
            missing_employee_ids.append(employee_id)
            continue
        choice = confirmation["choice"]
        if choice == "ignore_current":
            if (
                target.get("resolution") != "missing_current_snapshot"
                or target.get("current_hourly_rate") is not None
            ):
                raise HTTPException(400, "仅当当月薪资快照缺失时可忽略")
            target["hourly_rate"] = 0.0
            target["ratio"] = 0.0
            target["resolution"] = "manual_ignore_current_missing"
            target["manual_note"] = (
                confirmation["note"]
                or "当月薪资档案缺失，本次按0处理；后续重新上传当月薪资后重新核验"
            )
        else:
            prefix = "previous" if choice == "previous" else "current"
            if target.get(f"{prefix}_hourly_rate") is None:
                snapshot_label = "上月" if choice == "previous" else "当月"
                raise HTTPException(400, f"{snapshot_label}薪资快照缺失，不能选择该值")
            target["hourly_rate"] = target.get(f"{prefix}_hourly_rate", target.get("hourly_rate", 0))
            target["ratio"] = target.get(f"{prefix}_ratio", target.get("ratio", 0))
            target["resolution"] = f"manual_use_{choice}"
            target["manual_note"] = confirmation["note"]
        target["verification_status"] = "resolved"
        updated_employees.append(target)

    if not updated_employees:
        raise HTTPException(404, "未找到该员工的薪资差异记录，请刷新后重试")

    confirmed_ids = {
        str(employee.get("employee_id") or "")
        for employee in updated_employees
    }
    issues = [issue for issue in verification.get("issues", []) if issue.get("employee_id") not in confirmed_ids]
    summary = dict(verification.get("summary", {}))
    blocking_count = sum(row.get("verification_status") == "blocking" for row in employees)
    summary["blocking_count"] = blocking_count
    summary["resolved_count"] = len(employees) - blocking_count
    verification.update({"employees": employees, "issues": issues, "summary": summary})
    salary_data = dict(run.salary_data or {})
    salary_data["employees"] = employees
    salary_data["summary"] = {**salary_data.get("summary", {}), **summary}
    fbu_run_manager.update_run(
        run_id,
        salary_verification_data=verification,
        salary_data=salary_data,
        status="step2",
        error="",
    )
    if body.get("response_mode") == "employees":
        return {
            "success": True,
            "run_id": run_id,
            "employees": updated_employees,
            "missing_employee_ids": missing_employee_ids,
            "verification_summary": summary,
            "salary_summary": salary_data.get("summary", {}),
        }
    if body.get("response_mode") == "employee":
        return {
            "success": True,
            "run_id": run_id,
            "employee": updated_employees[0],
            "verification_summary": summary,
            "salary_summary": salary_data.get("summary", {}),
        }
    return {"success": True, "preview": salary_data, "verification": verification}


@app.post("/api/fbu-performance/runs/{run_id}/performance-supplement")
def add_fbu_performance_supplement(run_id: str, body: dict = Body(...)) -> dict:
    """页面录入离职/线下绩效补录。"""
    run = fbu_run_manager.get_run(run_id, sections={"performance_data"})
    if not run:
        raise HTTPException(404, "任务不存在")

    employee_id = str(body.get("employee_id") or "").strip()
    if not employee_id:
        raise HTTPException(400, "请填写工号")

    score = body.get("score")
    level = str(body.get("level") or "").strip()
    coefficient = body.get("coefficient")
    if score in (None, "") and not level and coefficient in (None, ""):
        raise HTTPException(400, "请至少填写绩效得分、绩效等级或绩效系数")

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        supplement_preview = parser.build_performance_supplement_preview([
            {
                "employee_id": employee_id,
                "name": body.get("name"),
                "score": score,
                "level": level,
                "coefficient": coefficient,
                "note": body.get("note"),
            }
        ])
        existing_performance = run.performance_data if isinstance(run.performance_data, dict) else {}
        if existing_performance.get("employees"):
            preview = parser.merge_performance_supplement_preview(existing_performance, supplement_preview)
        else:
            preview = supplement_preview

        fbu_run_manager.save_step_data(
            run_id,
            3,
            preview,
            performance_file=run.performance_file or "页面绩效补录",
        )

        if body.get("response_mode") == "employee":
            updated_employee = next(
                (row for row in preview.get("employees", []) if row.get("employee_id") == employee_id),
                None,
            )
            if updated_employee is None:
                raise HTTPException(404, "未找到补录员工")
            return {
                "success": True,
                "run_id": run_id,
                "step": 3,
                "employee": updated_employee,
                "summary": preview.get("summary", {}),
            }

        return {
            "success": True,
            "run_id": run_id,
            "step": 3,
            "preview": preview,
        }
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"绩效补录保存失败: {str(e)}")


@app.post("/api/fbu-performance/import-adjustments")
async def import_fbu_adjustments(
    run_id: str = Body(...),
    file: UploadFile = File(...),
) -> dict:
    """可选：导入调薪/转正拆分表"""
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / "adjustments.xlsx"
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        preview = parser.parse_adjustments_preview(str(file_path))
        fbu_run_manager.save_step_data(run_id, 4, preview, adjustment_file=file.filename)
        fbu_run_manager.persist_files(run_id, ["adjustments.xlsx"])
        result_file = _fbu_result_file_payload(run_id, "adjustments")

        return {
            "success": True,
            "run_id": run_id,
            "step": 4,
            "preview": preview,
            "result_file": result_file,
        }
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"调薪拆分表解析失败: {str(e)}")


@app.post("/api/fbu-performance/import-supplemental-leave")
async def import_fbu_supplemental_leave(
    run_id: str = Form(...),
    file: UploadFile = File(...),
) -> dict:
    """导入薪酬补充 sickpay&年假表，并生成待确认清单。"""
    run = fbu_run_manager.get_run(run_id, sections={"attendance_data"})
    if not run:
        raise HTTPException(404, "任务不存在")

    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / "supplemental_leave.xlsx"
    pending_file_path = run_dir / f".supplemental-leave-upload-{secrets.token_hex(8)}.xlsx"
    pending_file_path.write_bytes(await file.read())

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        preview = parser.parse_supplemental_leave_preview(str(pending_file_path), run.calc_month)
        if not preview.get("rows"):
            raise HTTPException(
                400,
                "未识别到补充假勤记录，请上传包含工号、姓名、总时长和假期类型的病假年假汇总文件",
            )
        preview = parser.validate_supplemental_leave_against_attendance(
            preview,
            run.attendance_data,
        )
        pending_file_path.replace(file_path)
        fbu_run_manager.update_run(
            run_id,
            supplemental_leave_file=file.filename,
            supplemental_leave_data=preview,
        )
        fbu_run_manager.persist_files(run_id, ["supplemental_leave.xlsx"])
        return {
            "success": True,
            "run_id": run_id,
            "preview": preview,
        }
    except HTTPException:
        raise
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"补充假勤解析失败: {str(e)}")
    finally:
        pending_file_path.unlink(missing_ok=True)


@app.post("/api/fbu-performance/import-base-overrides")
async def import_fbu_base_overrides(
    run_id: str = Form(...),
    file: UploadFile = File(...),
) -> dict:
    """导入96工时制标记/线下固定绩效基数例外表。"""
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    file_path = run_dir / "base_overrides.xlsx"
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    try:
        parser = FBUPerformanceParser()
        _load_fbu_roster_for_run(parser, run_id)
        preview = parser.parse_base_overrides_preview(str(file_path), run.calc_month)
        fbu_run_manager.update_run(
            run_id,
            base_override_file=file.filename,
            base_override_data=preview,
        )
        fbu_run_manager.persist_files(run_id, ["base_overrides.xlsx"])
        result_file = _fbu_result_file_payload(run_id, "base_overrides")

        return {
            "success": True,
            "run_id": run_id,
            "preview": preview,
            "result_file": result_file,
        }
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"工时规则/固定基数例外表解析失败: {str(e)}")


@app.post("/api/fbu-performance/runs/{run_id}/supplemental-leave/batch")
def update_fbu_supplemental_leave_batch(run_id: str, body: dict) -> dict:
    """批量更新补充假勤确认状态。"""
    apply_suggestions = bool(body.get("apply_suggestions"))
    row_ids = body.get("row_ids") or []
    if not row_ids and not apply_suggestions:
        raise HTTPException(400, "请选择需要处理的行")

    parser = FBUPerformanceParser()
    updates = {}
    if "confirmation_status" in body:
        status = str(body.get("confirmation_status") or "").strip()
        if not status:
            status = None
    else:
        status = None
    if status:
        if status not in {"pending", "confirmed", "excluded"}:
            raise HTTPException(400, "确认状态无效")
        updates["confirmation_status"] = status
    if "include_in_base" in body:
        include_raw = body.get("include_in_base")
        if include_raw != "":
            if isinstance(include_raw, bool):
                updates["include_in_base"] = include_raw
            elif str(include_raw).strip().lower() in {"true", "1", "yes"}:
                updates["include_in_base"] = True
            elif str(include_raw).strip().lower() in {"false", "0", "no"}:
                updates["include_in_base"] = False
            else:
                raise HTTPException(400, "是否计入参数无效")
    if "allocation_month" in body:
        allocation_month = str(body.get("allocation_month") or "").strip()
        if allocation_month and not re.match(r"^\d{4}-\d{2}$", allocation_month):
            raise HTTPException(400, "归属月份格式应为YYYY-MM")
        if allocation_month:
            updates["allocation_month"] = allocation_month
    if "allocation_period" in body:
        allocation_period = str(body.get("allocation_period") or "").strip()
        if allocation_period:
            updates["allocation_period"] = allocation_period
    if "confirmation_note" in body:
        confirmation_note = str(body.get("confirmation_note") or "").strip()
        if confirmation_note:
            updates["confirmation_note"] = confirmation_note
    if "included_hours" in body:
        try:
            included_hours = float(body.get("included_hours"))
        except (TypeError, ValueError):
            raise HTTPException(400, "计入小时必须为数字")
        if included_hours < 0:
            raise HTTPException(400, "计入小时不能小于0")
        updates["included_hours"] = included_hours
        if "confirmation_status" not in updates and "include_in_base" not in updates:
            if included_hours > 0:
                updates["confirmation_status"] = "confirmed"
                updates["include_in_base"] = True
            else:
                updates["confirmation_status"] = "excluded"
                updates["include_in_base"] = False

    mutation_result: dict[str, Any] = {}

    def mutate(preview: dict) -> dict:
        if not preview:
            raise HTTPException(400, "尚未导入补充假勤表")
        if row_ids:
            current_row_ids = {
                row.get("row_id")
                for row in preview.get("rows", [])
                if row.get("row_id")
            }
            if any(row_id not in current_row_ids for row_id in row_ids):
                raise HTTPException(409, "补充假勤页面数据已更新，请刷新后重试")
        if apply_suggestions:
            updated, applied_count = parser.apply_supplemental_leave_all_suggestions(preview)
            mutation_result["applied_count"] = applied_count
            return updated
        return parser.apply_supplemental_leave_batch(preview, row_ids, updates)

    committed = fbu_run_manager.mutate_section(
        run_id,
        "supplemental_leave_data",
        mutate,
    )
    if not committed:
        raise HTTPException(404, "任务不存在")
    committed_preview = committed["data"]
    if body.get("response_mode") == "row" and len(row_ids) == 1:
        updated_row = next(
            (
                row
                for row in committed_preview.get("rows", [])
                if row.get("row_id") == row_ids[0]
            ),
            None,
        )
        if updated_row is None:
            raise HTTPException(404, "补充假勤记录不存在")
        return {
            "success": True,
            "run_id": run_id,
            "row": updated_row,
            "summary": committed_preview.get("summary", {}),
            "section_revision": committed["revision"],
        }
    return {
        "success": True,
        "run_id": run_id,
        "preview": committed_preview,
        "section_revision": committed["revision"],
        **mutation_result,
    }


@app.get("/api/fbu-performance/templates/adjustments/download")
def download_fbu_adjustments_template() -> FileResponse:
    """下载FBU调薪/转正拆分表模板"""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "调薪拆分"

    headers = ["工号", "姓名", "分段期间", "分段绩效基数", "核算标识", "备注"]
    header_row = 6
    data_start_row = header_row + 1
    notes = [
        "填报说明：",
        "下面三行为脱敏填写示例，不参与导入；请从第7行开始填写真实数据。",
        "必填列：工号、分段期间、分段绩效基数、核算标识。核算标识请从下拉选项选择。",
    ]
    examples = [
        ["zt0000001", "花名一", "4.1-4.15", 1200.00, "调薪前", "示例：调薪前不参与绩效奖金"],
        ["zt0000001", "花名一", "4.16-4.30", 900.00, "调薪后", "示例：调薪后按薪资档案绩效比例核算"],
    ]

    header_fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    note_font = Font(color="475569")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for offset, note in enumerate(notes, 1):
        cell = ws.cell(row=offset, column=1, value=note)
        cell.font = Font(bold=(offset == 1), color="334155") if offset == 1 else note_font

    for row_idx, row in enumerate(examples, 3):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if col_idx == 5:
                cell.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    widths = [16, 14, 18, 18, 14, 38]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = f"A{data_start_row}"
    reason_validation = DataValidation(
        type="list",
        formula1='"调薪前,调薪后"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="核算标识无效",
        error="请选择“调薪前”或“调薪后”。",
    )
    ws.add_data_validation(reason_validation)
    reason_validation.add(f"E{data_start_row}:E1000")

    output_path = EXPORT_DIR / "FBU调薪转正拆分表模板.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return FileResponse(
        output_path,
        filename=output_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/fbu-performance/templates/base-overrides/download")
def download_fbu_base_overrides_template() -> FileResponse:
    """下载FBU工时规则/线下固定基数模板。"""
    import openpyxl
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工时规则标记"

    headers = ["工号", "姓名", "规则类型", "固定绩效基数", "归属月份", "状态", "备注"]
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws["C1"].comment = Comment("96工时制只作为规则标记；平台会按考勤日报明细自动计算特殊合计时长。线下固定基数覆盖仅用于人工例外。", "Codex")
    ws["D1"].comment = Comment("96工时制必须留空；线下固定基数覆盖才填写最终基数金额。", "Codex")
    ws["E1"].comment = Comment("请使用 YYYY-MM，例如 2026-04。", "Codex")

    widths = [16, 16, 20, 18, 14, 12, 34]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"

    rule_validation = DataValidation(
        type="list",
        formula1='"96工时制,线下固定基数覆盖"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="规则类型无效",
        error="请选择“96工时制”或“线下固定基数覆盖”。",
    )
    status_validation = DataValidation(
        type="list",
        formula1='"启用,停用,排除"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="状态无效",
        error="请选择“启用”“停用”或“排除”。",
    )
    ws.add_data_validation(rule_validation)
    ws.add_data_validation(status_validation)
    rule_validation.add("C2:C1000")
    status_validation.add("F2:F1000")

    output_path = EXPORT_DIR / "FBU工时规则与固定基数模板.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return FileResponse(
        output_path,
        filename=output_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/fbu-performance/import")
async def import_fbu_performance_data(
    attendance: UploadFile = File(...),
    salary: UploadFile = File(...),
    performance: UploadFile = File(...),
    calc_month: str = Body(...),
) -> dict:
    """导入FBU绩效数据文件（保留兼容）"""
    # 创建运行记录
    run = fbu_run_manager.create_run(
        calc_month=calc_month,
        attendance_file=attendance.filename,
        salary_file=salary.filename,
        performance_file=performance.filename,
    )

    # 保存上传文件
    run_dir = FBU_PERFORMANCE_RUNS_DIR / run.run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    for file, name in [(attendance, "attendance.xlsx"), (salary, "salary.xlsx"), (performance, "performance.xlsx")]:
        file_path = run_dir / name
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)

    fbu_run_manager.update_run(run.run_id, status="imported")
    fbu_run_manager.persist_files(run.run_id, ["attendance.xlsx", "salary.xlsx", "performance.xlsx"])

    return {
        "success": True,
        "run_id": run.run_id,
        "message": "数据导入成功",
    }


@app.post("/api/fbu-performance/calculate/{run_id}")
def calculate_fbu_performance(run_id: str, response_mode: str = "") -> dict:
    """执行FBU绩效核算"""
    run = fbu_run_manager.get_run(
        run_id,
        sections=_FBU_CALCULATION_SECTION_FIELDS,
        refresh=True,
    )
    if not run:
        raise HTTPException(404, "任务不存在")

    salary_verification = run.salary_verification_data or {}
    blocking_count = salary_verification.get("summary", {}).get("blocking_count", 0)
    if blocking_count:
        raise HTTPException(409, f"薪资历史核验仍有 {blocking_count} 条待处理差异，暂不能核算")
    supplemental_pending_count = int(
        (run.supplemental_leave_data or {}).get("summary", {}).get("pending_count", 0)
        or 0
    )
    if supplemental_pending_count:
        raise HTTPException(
            409,
            f"补充假勤仍有 {supplemental_pending_count} 条待确认，处理完成后再核算",
        )

    try:
        parser = FBUPerformanceParser()

        # 判断是分步模式还是一次性导入模式
        if run.attendance_data and run.salary_data and run.performance_data:
            # 分步模式：从已保存的分步数据计算
            engine = parser.parse_all_from_step_data(
                attendance_data=run.attendance_data.get('employees', []),
                salary_data=run.salary_data.get('employees', []),
                performance_data=run.performance_data.get('employees', []),
                adjustment_data=run.adjustment_data,
                transfer_data=run.transfer_data,
                supplemental_leave_data=run.supplemental_leave_data,
                base_override_data=run.base_override_data,
                hourly_rate_policy_data=run.hourly_rate_policy_data,
                period_adjustment_data=run.period_adjustment_data,
                calc_month=run.calc_month,
            )
        else:
            # 一次性导入模式：从文件计算
            run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
            for filename in ("attendance.xlsx", "salary.xlsx", "performance.xlsx"):
                fbu_run_manager.materialize_file(run_id, filename)
            target_month = int(run.calc_month.split("-")[1]) if "-" in run.calc_month else int(run.calc_month)
            _load_fbu_roster_for_run(parser, run_id)

            engine = parser.parse_all(
                attendance_file=str(run_dir / "attendance.xlsx"),
                salary_file=str(run_dir / "salary.xlsx"),
                performance_file=str(run_dir / "performance.xlsx"),
                target_month=target_month,
                calc_month=run.calc_month,
            )

        # 保存结果
        employees = engine.get_all_employees()
        if not employees:
            raise HTTPException(
                409,
                "未生成可核算员工，请检查已上传材料是否属于当前活动月份和划分区域",
            )
        if run.supplemental_leave_data and run.base_override_data:
            offline_bases = run.base_override_data.get("offline_bases", [])
            if offline_bases:
                supplemental_leave_data = parser.apply_supplemental_leave_suggestions(
                    run.supplemental_leave_data,
                    employees,
                    offline_bases,
                    run.calc_month,
                )
                fbu_run_manager.update_run(
                    run_id,
                    supplemental_leave_data=supplemental_leave_data,
                    persist=False,
                )
        fbu_run_manager.save_results(run_id, employees)
        completed_run = fbu_run_manager.get_run(run_id, sections={"results"})
        final_results = build_final_result_rows(completed_run.results)

        total_bonus_by_source_employee = {}
        for employee in employees:
            total_key = employee.source_employee_id or employee.employee_id
            total_bonus_by_source_employee[total_key] = (
                total_bonus_by_source_employee.get(total_key, 0.0)
                + employee.performance_bonus
            )

        activity_payload = {
            **vars(completed_run),
            "results": final_results,
            "total_employees": len(final_results),
            "diagnostics": _fbu_run_diagnostics(completed_run),
        }
        if response_mode == "compact":
            activity_payload = {
                "run_id": completed_run.run_id,
                "calc_month": completed_run.calc_month,
                "status": completed_run.status,
                "current_step": completed_run.current_step,
                "results": final_results,
                "total_employees": len(final_results),
                "total_bonus": completed_run.total_bonus,
                "match_rate": completed_run.match_rate,
                "diagnostics": _fbu_run_diagnostics(completed_run),
                "error": completed_run.error,
            }

        return {
            "success": True,
            "run_id": run_id,
            "total_employees": len(final_results),
            "total_bonus": round(
                sum(round(amount, 2) for amount in total_bonus_by_source_employee.values()),
                2,
            ),
            "activity": activity_payload,
        }

    except HTTPException as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e.detail))
        raise
    except Exception as e:
        fbu_run_manager.update_run(run_id, status="failed", error=str(e))
        raise HTTPException(500, f"计算失败: {str(e)}")


async def _process_fbu_calculation_job(run_id: str, job_id: str) -> None:
    store = _fbu_upload_job_store()
    try:
        await asyncio.to_thread(
            store.update,
            run_id,
            job_id,
            status="processing",
            stage="calculating",
            progress=35,
            message="正在核算绩效奖金",
            error="",
        )
        result = await _await_fbu_job_with_heartbeat(
            asyncio.to_thread(calculate_fbu_performance, run_id),
            store=store,
            run_id=run_id,
            job_id=job_id,
            stage="calculating",
            progress=35,
            message="正在核算绩效奖金",
        )
        await asyncio.to_thread(
            store.update,
            run_id,
            job_id,
            status="completed",
            stage="completed",
            progress=100,
            message="核算完成",
            result={
                "success": bool(result.get("success")),
                "run_id": run_id,
                "total_employees": int(result.get("total_employees") or 0),
                "total_bonus": float(result.get("total_bonus") or 0),
            },
            completedAt=datetime.now().isoformat(),
            error="",
        )
    except Exception as exc:
        detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
        fbu_logger.exception("FBU calculation job failed run=%s job=%s", run_id, job_id)
        try:
            await asyncio.to_thread(
                store.update,
                run_id,
                job_id,
                status="failed",
                stage="failed",
                progress=100,
                message="核算失败",
                error=str(detail or "未知错误"),
            )
        except Exception:
            fbu_logger.exception("Failed to persist calculation job failure %s", job_id)


def _prepare_fbu_calculation_job(run_id: str, job_id: str) -> tuple[dict, bool]:
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")
    store = _fbu_upload_job_store()
    job = store.load(run_id, job_id, refresh=True)
    if not job or job.get("kind") != "calculation":
        raise HTTPException(404, "核算任务不存在")
    public = store.public(job)
    if job.get("status") == "completed":
        return public, False
    if job.get("status") in {"queued", "processing"} and not public["recoverable"]:
        return public, False
    allowed_from = (
        ["queued", "processing"]
        if public["recoverable"]
        else ["uploading", "failed"]
    )
    job = store.update(
        run_id,
        job_id,
        allowed_from=allowed_from,
        status="queued",
        stage="queued",
        progress=5,
        message="正在排队",
        attempt=int(job.get("attempt") or 0) + 1,
        error="",
    )
    transition_applied = bool(job.pop("__transition_applied", True))
    return store.public(job), transition_applied and job.get("status") == "queued"


async def _run_fbu_calculation_job_in_request(run_id: str, job_id: str) -> dict:
    job, should_process = _prepare_fbu_calculation_job(run_id, job_id)
    if should_process:
        await _process_fbu_calculation_job(run_id, job_id)
        persisted = _fbu_upload_job_store().load(run_id, job_id, refresh=True)
        if persisted is None:
            raise HTTPException(500, "核算任务状态丢失，请重新核算。")
        job = _fbu_upload_job_store().public(persisted)
    return job


@app.post(
    "/api/fbu-performance/runs/{run_id}/calculation-jobs",
    status_code=202,
)
async def create_fbu_calculation_job(
    run_id: str,
) -> dict:
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")
    store = _fbu_upload_job_store()
    job = store.create(run_id, [])
    job = store.update(run_id, job["jobId"], kind="calculation")
    completed = await _run_fbu_calculation_job_in_request(run_id, job["jobId"])
    return {"job": completed}


@app.post(
    "/api/fbu-performance/runs/{run_id}/calculation-jobs/{job_id}/resume",
    status_code=202,
)
async def resume_fbu_calculation_job(
    run_id: str,
    job_id: str,
) -> dict:
    try:
        job = await _run_fbu_calculation_job_in_request(run_id, job_id)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    return {"job": job}


@app.get("/api/fbu-performance/runs/{run_id}/calculation-jobs/{job_id}")
def get_fbu_calculation_job(run_id: str, job_id: str) -> dict:
    try:
        store = _fbu_upload_job_store()
        job = store.load(run_id, job_id, refresh=True)
    except ValueError as exc:
        raise HTTPException(400, str(exc)) from exc
    if not job or job.get("kind") != "calculation":
        raise HTTPException(404, "核算任务不存在")
    return {"job": store.public(job)}


@app.post("/api/fbu-performance/runs")
def create_fbu_performance_run(request: Request, body: dict) -> dict:
    """创建新的月度核算活动"""
    calc_month = body.get("calc_month")
    if not calc_month:
        raise HTTPException(400, "缺少核算月份")

    # 验证calc_month格式 (YYYY-MM)
    import re
    if not re.match(r'^\d{4}-\d{2}$', calc_month):
        raise HTTPException(400, "核算月份格式无效，应为YYYY-MM")

    # 验证月份范围
    try:
        year, month = calc_month.split('-')
        if not (2020 <= int(year) <= 2030 and 1 <= int(month) <= 12):
            raise HTTPException(400, "核算月份范围无效")
    except ValueError:
        raise HTTPException(400, "核算月份格式无效")

    region_code = str(body.get("region_code") or DEFAULT_FBU_REGION_CODE).strip()
    valid_region_codes = {region["code"] for region in FBU_AMERICAS_REGIONS}
    if region_code not in valid_region_codes:
        raise HTTPException(400, "请选择有效的FBU美洲划分区域")

    actor = _fbu_current_user(request)
    run = fbu_run_manager.create_run(
        calc_month=calc_month,
        region_code=region_code,
        created_by_user_id=actor["id"],
        created_by_name=actor["name"],
        created_by_avatar_url=actor["avatarUrl"],
        persist=False,
    )
    metadata = fbu_roster_store.get_metadata(run.calc_month)
    roster_path = fbu_roster_store.copy_active_to_run(run.run_id, run.calc_month)
    roster_data = None
    if roster_path:
        parser = FBUPerformanceParser()
        parser.load_roster(str(roster_path))
        employees = [
            {
                "employee_id": employee_id,
                "name": row.get("name", ""),
                "department": row.get("department", ""),
                "area": row.get("area", ""),
                "position": row.get("position", ""),
                "personnel_status": row.get("personnel_status", ""),
                "job_type": row.get("job_type", ""),
            }
            for employee_id, row in parser.employee_roster.items()
        ]
        roster_data = {"employees": employees, "summary": {"total_employees": len(employees)}}
        fbu_run_manager.update_run(
            run.run_id,
            roster_file=metadata.get("filename", "active_roster.xlsx"),
            roster_source="base",
            roster_data=roster_data,
        )
    else:
        fbu_run_manager.update_run(run.run_id)
    run = fbu_run_manager.runs.get(run.run_id) or run
    activity = vars(run).copy()
    activity["roster_data"] = roster_data or run.roster_data or {}
    activity["diagnostics"] = _fbu_run_diagnostics(run)

    return {
        "success": True,
        "run_id": run.run_id,
        "calc_month": run.calc_month,
        "activity_name": run.activity_name,
        "region_code": run.region_code,
        "region_name": run.region_name,
        "created_by_user_id": run.created_by_user_id,
        "created_by_name": run.created_by_name,
        "created_by_avatar_url": run.created_by_avatar_url,
        "status": run.status,
        "roster_file": run.roster_file,
        "roster_source": run.roster_source,
        "activity": activity,
    }


@app.get("/api/fbu-performance/runs")
def list_fbu_performance_runs(request: Request) -> dict:
    """获取FBU绩效核算任务列表"""
    return {
        "runs": fbu_run_manager.list_run_summaries(),
        "regions": [dict(region) for region in FBU_AMERICAS_REGIONS],
        "current_user": _fbu_current_user(request),
    }


def _parse_fbu_run_detail_include(include: str) -> tuple[set[str] | None, set[str]]:
    raw = str(include or "").strip()
    if not raw or raw == "all":
        return None, {"roster_data", "diagnostics"}
    requested = {
        value.strip()
        for value in raw.split(",")
        if value.strip() and value.strip() != "core"
    }
    extras = requested.intersection({"roster_data", "diagnostics"})
    sections = requested.difference(extras)
    invalid = sections.difference(FBU_RUN_SECTION_FIELDS)
    if invalid:
        raise HTTPException(400, f"不支持的活动详情区块：{', '.join(sorted(invalid))}")
    return sections, extras


@app.get("/api/fbu-performance/runs/{run_id}")
def get_fbu_performance_run(
    run_id: str,
    background_tasks: BackgroundTasks,
    include: str = "all",
) -> dict:
    """获取FBU绩效核算任务详情"""
    sections, extras = _parse_fbu_run_detail_include(include)
    run = fbu_run_manager.get_run(run_id, sections=sections)
    if not run:
        raise HTTPException(404, "任务不存在")
    attendance_view = (
        getattr(run, "attendance_view_data", {})
        if isinstance(getattr(run, "attendance_view_data", {}), dict)
        else {}
    )
    attendance_source_run = None
    if (
        sections is not None
        and "attendance_view_data" in sections
        and not has_attendance_employees(attendance_view)
    ):
        attendance_source_run = (
            fbu_run_manager.get_run(run_id, sections={"attendance_data"})
            or run
        )
        attendance_view = build_attendance_view_data(
            attendance_source_run.attendance_data
        )
        if attendance_view:
            background_tasks.add_task(
                fbu_run_manager.backfill_attendance_view_data,
                run_id,
                attendance_view,
            )
    should_backfill_hourly_policy = (
        sections is None
        or {"attendance_data", "hourly_rate_policy_data"}.issubset(sections)
        or {"attendance_view_data", "hourly_rate_policy_data"}.issubset(sections)
    )
    if should_backfill_hourly_policy and not getattr(run, "hourly_rate_policy_data", {}):
        attendance_source_run = attendance_source_run or (
            fbu_run_manager.get_run(run_id, sections={"attendance_data"})
            or run
        )
    if (
        should_backfill_hourly_policy
        and attendance_source_run
        and attendance_source_run.attendance_data
        and not getattr(run, "hourly_rate_policy_data", {})
    ):
        fbu_run_manager.backfill_hourly_rate_policy_data(
            run_id,
            build_hourly_rate_policy_data(
                attendance_source_run.attendance_data,
                run.calc_month,
            ),
        )
        run = fbu_run_manager.get_run(run_id, sections=sections) or run
    raw_payload = vars(run).copy()
    if has_attendance_employees(attendance_view):
        raw_payload["attendance_view_data"] = attendance_view
    elif sections is not None and "attendance_view_data" in sections:
        raw_payload["attendance_view_data"] = {}
    if sections is None:
        payload = raw_payload
        payload.pop("results_view_data", None)
        payload.pop("attendance_view_data", None)
        loaded_sections = sorted(
            FBU_RUN_SECTION_FIELDS.difference({
                "attendance_view_data",
                "results_view_data",
            })
        )
    else:
        payload = {
            key: value
            for key, value in raw_payload.items()
            if key in _FBU_RUN_CORE_RESPONSE_FIELDS
        }
        for field_name in sections:
            payload[field_name] = raw_payload.get(field_name)
        loaded_sections = sorted(sections)
    if "attendance_view_data" in payload:
        payload["attendance_data"] = payload.pop("attendance_view_data")
    if "hourly_rate_policy_data" in payload:
        payload["hourly_rate_policy_data"] = _fbu_hourly_rate_policy_response(
            payload.get("hourly_rate_policy_data")
        )
    if "results" in loaded_sections and payload.get("results"):
        payload["results"] = build_final_result_rows(payload["results"])
        payload["total_employees"] = len(payload["results"])
    if sections is None or "roster_data" in extras:
        payload["roster_data"] = _fbu_roster_preview_for_run(run_id)
    if sections is None or "diagnostics" in extras:
        diagnostics_run = run
        if sections is not None:
            diagnostics_run = (
                fbu_run_manager.get_run(
                    run_id,
                    sections=_FBU_DIAGNOSTIC_SECTION_FIELDS,
                )
                or run
            )
        payload["diagnostics"] = _fbu_run_diagnostics(diagnostics_run)
    payload["loaded_sections"] = loaded_sections
    section_revision_loader = getattr(fbu_run_manager, "get_section_revisions", None)
    payload["section_revisions"] = (
        section_revision_loader(run_id, set(loaded_sections))
        if callable(section_revision_loader)
        else {}
    )
    return payload


@app.get("/api/fbu-performance/runs/{run_id}/diagnostics")
def get_fbu_performance_diagnostics(run_id: str) -> dict:
    """获取FBU数据匹配诊断"""
    run = fbu_run_manager.get_run(
        run_id,
        sections=_FBU_DIAGNOSTIC_SECTION_FIELDS,
    )
    if not run:
        raise HTTPException(404, "任务不存在")
    return _fbu_run_diagnostics(run)


@app.get("/api/fbu-performance/runs/{run_id}/results")
def get_fbu_performance_results(
    run_id: str,
    background_tasks: BackgroundTasks,
    page: int = 1,
    page_size: int = 0,
    q: str = "",
    group: str = "all",
) -> dict:
    """获取FBU绩效核算结果"""
    run = fbu_run_manager.get_run(run_id, sections={"results_view_data"})
    if not run:
        raise HTTPException(404, "任务不存在")
    if run.status != "completed":
        raise HTTPException(400, "任务未完成")
    results_view = (
        run.results_view_data
        if isinstance(run.results_view_data, dict)
        else {}
    )
    results = (
        list(results_view.get("rows") or [])
        if isinstance(results_view.get("rows"), list)
        else []
    )
    if not results:
        full_run = fbu_run_manager.get_run(run_id, sections={"results"}) or run
        results = build_final_result_rows(full_run.results)
        run = full_run
        if results:
            background_tasks.add_task(
                fbu_run_manager.backfill_results_view_data,
                run_id,
                build_results_view_data(full_run.results),
            )

    def group_key(row: dict) -> str:
        if row.get("job_type") == "district_manager":
            return "district"
        if row.get("job_type") == "functional":
            return "functional"
        return "warehouse"

    groups = {
        key: {
            "count": len(rows),
            "total_bonus": round(
                sum(float(row.get("performance_bonus") or 0) for row in rows),
                2,
            ),
        }
        for key in ("warehouse", "functional", "district")
        for rows in [[row for row in results if group_key(row) == key]]
    }
    calculation_paths: dict[str, dict] = {}
    for row in results:
        path = str(row.get("calculation_path") or STANDARD_PERFORMANCE_BASE_PATH)
        bucket = calculation_paths.setdefault(
            path,
            {
                "path": path,
                "count": 0,
                "total_base": 0.0,
                "total_bonus": 0.0,
            },
        )
        bucket["count"] += 1
        bucket["total_base"] += float(row.get("performance_base") or 0)
        bucket["total_bonus"] += float(row.get("performance_bonus") or 0)
    calculation_path_rows = [
        {
            **bucket,
            "total_base": round(bucket["total_base"], 2),
            "total_bonus": round(bucket["total_bonus"], 2),
        }
        for bucket in calculation_paths.values()
    ]
    selected = results
    if group in groups:
        selected = [row for row in selected if group_key(row) == group]
    normalized_query = str(q or "").strip().casefold()
    if normalized_query:
        selected = [
            row
            for row in selected
            if any(
                normalized_query in str(row.get(field_name) or "").casefold()
                for field_name in (
                    "employee_id",
                    "source_employee_id",
                    "name",
                    "department",
                    "area",
                    "position",
                )
            )
        ]

    total = len(selected)
    normalized_page = max(1, int(page or 1))
    normalized_page_size = max(0, min(200, int(page_size or 0)))
    if normalized_page_size:
        pages = max(1, (total + normalized_page_size - 1) // normalized_page_size)
        normalized_page = min(normalized_page, pages)
        start = (normalized_page - 1) * normalized_page_size
        page_rows = selected[start:start + normalized_page_size]
    else:
        pages = 1
        page_rows = selected
    return {
        "results": page_rows,
        "pagination": {
            "page": normalized_page,
            "page_size": normalized_page_size or total,
            "total": total,
            "pages": pages,
        },
        "summary": {
            "total_employees": len(results),
            "total_bonus": round(
                sum(float(row.get("performance_bonus") or 0) for row in results),
                2,
            ),
            "total_performance_base": round(
                sum(float(row.get("performance_base") or 0) for row in results),
                2,
            ),
            "special_base_count": sum(
                1
                for row in results
                if str(row.get("calculation_path") or STANDARD_PERFORMANCE_BASE_PATH)
                != STANDARD_PERFORMANCE_BASE_PATH
            ),
            "calculation_paths": calculation_path_rows,
            "groups": groups,
        },
    }


@app.get("/api/fbu-performance/runs/{run_id}/export")
def export_fbu_performance(run_id: str) -> dict:
    """导出FBU绩效核算结果"""
    # 检查任务是否存在
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    # 检查任务状态
    if run.status != "completed":
        raise HTTPException(400, f"任务未完成，当前状态: {run.status}")

    # 导出
    output_path = fbu_run_manager.export_run(run_id, str(EXPORT_DIR))
    if not output_path:
        raise HTTPException(500, "导出失败，请检查数据完整性")

    return {"file_path": output_path, "file_name": Path(output_path).name}


@app.get("/api/fbu-performance/runs/{run_id}/export-excel")
def export_fbu_excel(run_id: str, type: str = "attendance") -> dict:
    """导出带样式的Excel文件"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def safe_excel_value(value):
        if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
            return "'" + value
        return value

    def format_fbu_job_type(job_type: str) -> str:
        if job_type == "district_manager":
            return "区长"
        if job_type == "functional":
            return "职能"
        if job_type == "warehouse":
            return "仓库"
        return job_type or ""

    def result_calculation_path(item: dict) -> str:
        if item.get("calculation_path"):
            return item["calculation_path"]
        if item.get("job_type") == "district_manager":
            return DISTRICT_MANAGER_FIXED_BASE_PATH
        if item.get("calculation_segments"):
            return ADJUSTMENT_SPLIT_PATH
        return STANDARD_PERFORMANCE_BASE_PATH

    def is_ninety_six_hour_result(item: dict) -> bool:
        return (
            item.get("work_hour_rule") == "96工时制"
            or "96工时制" in str(item.get("calculation_path") or "")
            or "96工时制" in str(item.get("base_override_type") or "")
        )

    def final_result_group_key(item: dict) -> str:
        if item.get("job_type") == "district_manager":
            return "district"
        if item.get("job_type") == "functional":
            return "functional"
        return "warehouse"

    def month_label(calc_month: str) -> str:
        try:
            return f"{int(str(calc_month).split('-')[1])}月"
        except (IndexError, TypeError, ValueError):
            return str(calc_month or "")

    def department_levels(item: dict) -> list[str]:
        area = str(item.get("area") or "").strip()
        department = str(item.get("department") or "").strip()
        parts = [part.strip() for part in department.split("-") if part.strip()]
        if area and (not parts or area not in parts):
            parts.insert(0, area)
        return (parts + ["", "", ""])[:3]

    def result_note(item: dict) -> str:
        notes = []
        if item.get("base_override_reason"):
            notes.append(str(item["base_override_reason"]))
        if item.get("coefficient_override_reason"):
            notes.append(str(item["coefficient_override_reason"]))
        if item.get("deferred_reason"):
            notes.append(str(item["deferred_reason"]))
        if item.get("merged_result"):
            notes.append("白夜班拆分行已合并")
        return "；".join(note for note in notes if note)

    def format_process_number(value, decimals: int = 2) -> str:
        try:
            return f"{float(value or 0):,.{decimals}f}"
        except (TypeError, ValueError):
            return f"{0:.{decimals}f}"

    def format_base_component_formula(component: dict) -> str:
        amount = format_process_number(component.get("amount"), 2)
        if component.get("hours") is None or component.get("hourly_rate") is None:
            return f"${amount}"
        hours = format_process_number(component.get("hours"), 2)
        hourly_rate = format_process_number(component.get("hourly_rate"), 4)
        multiplier = float(component.get("multiplier") or 1)
        multiplier_text = f" × {multiplier:.1f}" if abs(multiplier - 1) > 0.0001 else ""
        return f"{hours}h × ${hourly_rate}{multiplier_text} = ${amount}"

    def base_calculation_process(item: dict) -> str:
        details = item.get("base_calculation_details") or []
        blocks = []
        for detail in details:
            components = detail.get("components") or []
            heading = " · ".join(
                str(value) for value in (detail.get("display_label"), detail.get("path")) if value
            ) or "绩效基数"
            lines = [f"{heading}：${format_process_number(detail.get('performance_base'), 2)}"]
            hourly_rates = []
            for component in components:
                if component.get("hourly_rate") is None:
                    continue
                rate = format_process_number(component.get("hourly_rate"), 4)
                if rate not in hourly_rates:
                    hourly_rates.append(rate)
            if hourly_rates:
                lines.append(f"计算时薪：{' / '.join('$' + rate for rate in hourly_rates)}")
            for component in components:
                label = " · ".join(
                    str(value) for value in (component.get("period"), component.get("label")) if value
                ) or "基数组成"
                lines.append(f"{label}：{format_base_component_formula(component)}")
            if detail.get("note"):
                lines.append(str(detail["note"]))
            blocks.append("\n".join(lines))
        period_adjustment = float(item.get("period_adjustment") or 0)
        if period_adjustment:
            summary_lines = [
                f"系统计算绩效基数：${format_process_number(item.get('system_performance_base'), 2)}",
                f"Period adjustment：${format_process_number(period_adjustment, 2)}",
                f"最终绩效基数：${format_process_number(item.get('performance_base'), 2)}",
            ]
            blocks.insert(0, "\n".join(summary_lines))
        if blocks:
            return "\n\n".join(blocks)
        return f"绩效基数：${format_process_number(item.get('performance_base'), 2)}"

    def bonus_calculation_process(item: dict) -> str:
        segments = item.get("calculation_segments") or [item]
        lines = []
        for segment in segments:
            reason = segment.get("reason") or result_calculation_path(item)
            performance_base = format_process_number(segment.get("performance_base"), 2)
            performance_ratio = float(segment.get("performance_ratio") or 0)
            coefficient = format_process_number(segment.get("performance_coefficient"), 2)
            performance_bonus = format_process_number(segment.get("performance_bonus"), 2)
            if item.get("job_type") == "district_manager" and not segment.get("is_period_adjustment"):
                formula = f"${performance_base} × {coefficient} = ${performance_bonus}"
            else:
                formula = (
                    f"${performance_base} × {performance_ratio:.1%} × "
                    f"{coefficient} = ${performance_bonus}"
                )
            lines.append(f"{reason}：{formula}")
        return "\n".join(lines)

    def normalize_shift_employee_id(employee_id: str) -> str:
        text = str(employee_id or "").strip()
        return text[:-2] if text.endswith("-1") else text

    def result_position_lookup(run: FBURun) -> dict[str, str]:
        lookup: dict[str, str] = {}
        for source in (
            (run.attendance_data or {}).get("employees", []),
            (run.salary_data or {}).get("employees", []),
        ):
            for row in source:
                emp_id = normalize_shift_employee_id(row.get("source_employee_id") or row.get("employee_id"))
                position = str(row.get("position") or "").strip()
                if emp_id and position and emp_id not in lookup:
                    lookup[emp_id] = position
        return lookup

    def enrich_result_positions(rows: list[dict], run: FBURun) -> list[dict]:
        positions = result_position_lookup(run)
        enriched = []
        for row in rows:
            item = dict(row)
            emp_id = normalize_shift_employee_id(item.get("source_employee_id") or item.get("employee_id"))
            if not str(item.get("position") or "").strip() and emp_id in positions:
                item["position"] = positions[emp_id]
            enriched.append(item)
        return enriched

    export_sections = {
        "attendance": {"attendance_data"},
        "salary": {"salary_data"},
        "performance": {"performance_data"},
        "adjustments": {"adjustment_data"},
        "base_overrides": {"base_override_data"},
        "diagnostics": _FBU_DIAGNOSTIC_SECTION_FIELDS,
        "results": {"results", "attendance_data", "salary_data"},
    }
    # 只读取当前导出类型需要的数据。
    run = fbu_run_manager.get_run(
        run_id,
        sections=export_sections.get(type, set()),
    )
    if not run:
        raise HTTPException(404, "任务不存在")

    # 获取数据
    data = []
    title = ""
    filename = ""

    if type == "attendance" and run.attendance_data:
        data = run.attendance_data.get('employees', [])
        title = "考勤汇总"
        filename = f"考勤汇总_{run.calc_month}_{run_id}.xlsx"
    elif type == "salary" and run.salary_data:
        data = run.salary_data.get('employees', [])
        title = "薪资匹配"
        filename = f"薪资匹配_{run.calc_month}_{run_id}.xlsx"
    elif type == "performance" and run.performance_data:
        data = run.performance_data.get('employees', [])
        title = "绩效明细"
        filename = f"绩效明细_{run.calc_month}_{run_id}.xlsx"
    elif type == "adjustments" and run.adjustment_data:
        data = []
        for employee in run.adjustment_data.get('employees', []):
            for segment in employee.get("segments", []):
                data.append({
                    "record_type": "调薪拆分",
                    "employee_id": employee.get("employee_id", ""),
                    "name": employee.get("name", ""),
                    "area": employee.get("area", ""),
                    "department": employee.get("department", ""),
                    "period": segment.get("period", ""),
                    "performance_base": segment.get("performance_base", 0),
                    "reason": segment.get("reason", ""),
        })
        for event in run.adjustment_data.get("events", []):
            data.append({
                "record_type": "OEHR调薪事件",
                "employee_id": event.get("employee_id", ""),
                "name": event.get("name", ""),
                "area": event.get("area", ""),
                "department": event.get("department", ""),
                "period": event.get("effective_date", ""),
                "performance_base": event.get("performance_base", 0),
                "reason": event.get("adjustment_reason", ""),
                "approval_status": event.get("approval_status", ""),
                "adjustment_type": event.get("adjustment_type", ""),
                "hourly_rate": event.get("hourly_rate", 0),
                "performance_ratio": event.get("performance_ratio", 0),
                "note": event.get("note", ""),
            })
        title = "调薪拆分"
        filename = f"调薪拆分_{run.calc_month}_{run_id}.xlsx"
    elif type == "base_overrides" and run.base_override_data:
        data = run.base_override_data.get("employees", [])
        title = "工时规则与固定基数例外"
        filename = f"工时规则与固定基数例外_{run.calc_month}_{run_id}.xlsx"
    elif type == "diagnostics":
        diagnostics = _fbu_run_diagnostics(run)
        data = diagnostics.get("issues", [])
        title = "数据诊断"
        filename = f"数据诊断_{run.calc_month}_{run_id}.xlsx"
    elif type == "results" and run.results:
        data = enrich_result_positions(build_final_result_rows(run.results), run)
        title = "核算结果"
        filename = f"核算结果_{run.calc_month}_{run_id}.xlsx"
    else:
        raise HTTPException(400, "没有数据可导出")

    if not data:
        raise HTTPException(400, "没有数据可导出")

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # 定义样式
    # 员工信息列样式（蓝色系）
    emp_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    emp_font = Font(bold=True, color="1565C0")

    # 数据列样式（绿色系）
    data_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    data_font = Font(bold=True, color="2E7D32")

    # 计算列样式（橙色系）
    calc_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    calc_font = Font(bold=True, color="E65100")

    # 金额列样式（紫色系）
    money_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    money_font = Font(bold=True, color="6A1B9A")

    # 标题行样式
    title_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)

    # 表头样式
    header_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # 边框样式
    thin_border = Border(
        left=Side(style='thin', color='BDBDBD'),
        right=Side(style='thin', color='BDBDBD'),
        top=Side(style='thin', color='BDBDBD'),
        bottom=Side(style='thin', color='BDBDBD')
    )

    if type == "results":
        group_defs = [
            ("warehouse", "1.仓库管理人员"),
            ("functional", "2.非仓人员"),
            ("district", "3.区长"),
        ]
        grouped = {key: [] for key, _ in group_defs}
        for item in data:
            grouped.setdefault(final_result_group_key(item), []).append(item)

        dark_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        dark_font = Font(bold=True, color="FFFFFF")
        summary_fill = PatternFill(start_color="F3F6FA", end_color="F3F6FA", fill_type="solid")
        red_base_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        red_base_font = Font(bold=True, color="C00000")
        title_fill_results = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        def write_title(sheet, title_text: str, last_col: int) -> None:
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
            cell = sheet.cell(row=1, column=1, value=safe_excel_value(title_text))
            cell.font = Font(bold=True, size=14, color="111827")
            cell.fill = title_fill_results
            cell.alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 22

        def style_result_header_cell(cell, fill=dark_fill) -> None:
            cell.fill = fill
            cell.font = dark_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        def apply_body_style(cell, number_type: str = "text") -> None:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if number_type == "money":
                cell.number_format = '"$"#,##0.00'
            elif number_type == "percent":
                cell.number_format = "0.0%"
            elif number_type == "number":
                cell.number_format = "0.00"
            elif number_type == "process":
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        def write_staff_detail_sheet(sheet, rows: list[dict], sheet_title: str) -> None:
            columns = [
                ("绩效周期", "period", 8, "text"),
                ("三级部门", "dept_1", 13, "text"),
                ("四级部门", "dept_2", 13, "text"),
                ("五级部门", "dept_3", 16, "text"),
                ("姓名", "name", 14, "text"),
                ("员工工号", "employee_id", 14, "text"),
                ("职位", "position", 34, "text"),
                ("系统计算绩效基数", "system_performance_base", 16, "money"),
                ("Period adjustment", "period_adjustment", 16, "money"),
                (f"{month_label(run.calc_month)}绩效基数", "performance_base", 16, "money"),
                ("绩效比例", "performance_ratio", 11, "percent"),
                ("绩效得分", "performance_score", 12, "number"),
                ("绩效等级", "performance_level", 12, "text"),
                ("绩效系数", "performance_coefficient", 12, "number"),
                ("本月应发绩效工资", "performance_bonus", 15, "money"),
                ("备注", "note", 34, "text"),
                ("绩效基数计算过程", "base_calculation_process", 54, "process"),
                ("奖金计算过程", "bonus_calculation_process", 48, "process"),
            ]
            write_title(sheet, f"新泽西区绩效考核与奖金核算——{sheet_title.split('.', 1)[-1]}", len(columns))
            sheet.merge_cells("A2:K2")
            sheet.merge_cells("L2:N2")
            sheet.merge_cells("O2:O3")
            sheet.merge_cells("P2:P3")
            sheet.merge_cells("Q2:R2")
            group_headers = {
                "A2": "员工信息",
                "L2": "本月绩效考核结果(OEHR)",
                "O2": "本月应发绩效工资",
                "P2": "备注",
                "Q2": "计算过程",
            }
            for address, value in group_headers.items():
                style_result_header_cell(sheet[address])
                sheet[address] = safe_excel_value(value)
            for col_idx, (header, _, width, _) in enumerate(columns, 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = width
                if col_idx in (15, 16):
                    continue
                cell = sheet.cell(row=3, column=col_idx, value=safe_excel_value(header))
                style_result_header_cell(cell)
            sheet.row_dimensions[2].height = 26
            sheet.row_dimensions[3].height = 38

            for row_idx, item in enumerate(rows, 4):
                levels = department_levels(item)
                values = {
                    "period": month_label(run.calc_month),
                    "dept_1": levels[0],
                    "dept_2": levels[1],
                    "dept_3": levels[2],
                    "name": item.get("name", ""),
                    "employee_id": item.get("employee_id", ""),
                    "position": item.get("position") or format_fbu_job_type(item.get("job_type", "")),
                    "system_performance_base": item.get("system_performance_base", item.get("performance_base", 0)),
                    "period_adjustment": item.get("period_adjustment", 0),
                    "performance_base": item.get("performance_base", 0),
                    "performance_ratio": item.get("performance_ratio", 0),
                    "performance_score": item.get("performance_score", ""),
                    "performance_level": item.get("performance_level", ""),
                    "performance_coefficient": item.get("performance_coefficient", 0),
                    "performance_bonus": item.get("performance_bonus", 0),
                    "note": result_note(item),
                    "base_calculation_process": base_calculation_process(item),
                    "bonus_calculation_process": bonus_calculation_process(item),
                }
                for col_idx, (_, field, _, number_type) in enumerate(columns, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=safe_excel_value(values.get(field, "")))
                    apply_body_style(cell, number_type)
                    if field == "performance_base" and is_ninety_six_hour_result(item):
                        cell.fill = red_base_fill
                        cell.font = red_base_font
                process_line_count = max(
                    values["base_calculation_process"].count("\n") + 1,
                    values["bonus_calculation_process"].count("\n") + 1,
                )
                sheet.row_dimensions[row_idx].height = min(80, max(24, process_line_count * 8))
            sheet.freeze_panes = "A4"
            sheet.auto_filter.ref = f"A3:R{max(3, 3 + len(rows))}"
            if sheet.sheet_view.selection:
                sheet.sheet_view.selection[0].activeCell = "A3"
                sheet.sheet_view.selection[0].sqref = "A3"

        def write_district_sheet(sheet, rows: list[dict]) -> None:
            columns = [
                ("区域", "area", 12, "text"),
                ("被考核人", "name", 14, "text"),
                ("工号", "employee_id", 12, "text"),
                ("岗位", "position", 12, "text"),
                ("考核周期", "period", 12, "text"),
                ("绩效得分", "performance_score", 12, "number"),
                ("绩效等级", "performance_level", 12, "text"),
                ("绩效申诉", "appeal", 12, "text"),
                ("系统计算绩效基数", "system_performance_base", 16, "money"),
                ("Period adjustment", "period_adjustment", 16, "money"),
                ("绩效奖金基数", "performance_base", 16, "money"),
                ("绩效系数", "performance_coefficient", 12, "number"),
                ("绩效奖金", "performance_bonus", 12, "money"),
                ("备注", "note", 26, "text"),
                ("绩效基数计算过程", "base_calculation_process", 54, "process"),
                ("奖金计算过程", "bonus_calculation_process", 48, "process"),
            ]
            write_title(sheet, "海外区长-绩效奖金核算", len(columns))
            sheet.merge_cells("A2:D2")
            sheet.merge_cells("E2:H2")
            sheet.merge_cells("I2:M2")
            sheet.merge_cells("N2:N3")
            sheet.merge_cells("O2:P2")
            group_headers = {
                "A2": "区长/副区长/商务负责人信息",
                "E2": "绩效考核结果",
                "I2": "绩效奖金核算（美元）",
                "N2": "备注",
                "O2": "计算过程",
            }
            for address, value in group_headers.items():
                style_result_header_cell(sheet[address])
                sheet[address] = safe_excel_value(value)
            for col_idx, (header, _, width, _) in enumerate(columns, 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = width
                if col_idx == 14:
                    continue
                cell = sheet.cell(row=3, column=col_idx, value=safe_excel_value(header))
                style_result_header_cell(cell)
            sheet.row_dimensions[2].height = 30
            sheet.row_dimensions[3].height = 35

            for row_idx, item in enumerate(rows, 4):
                values = {
                    "area": item.get("area") or "新泽西区",
                    "name": item.get("name", ""),
                    "employee_id": item.get("employee_id", ""),
                    "position": item.get("position") or "区长",
                    "period": month_label(run.calc_month),
                    "performance_score": item.get("performance_score", ""),
                    "performance_level": item.get("performance_level", ""),
                    "appeal": "无",
                    "system_performance_base": item.get("system_performance_base", item.get("performance_base", 0)),
                    "period_adjustment": item.get("period_adjustment", 0),
                    "performance_base": item.get("performance_base", 0),
                    "performance_coefficient": item.get("performance_coefficient", 0),
                    "performance_bonus": item.get("performance_bonus", 0),
                    "note": "",
                    "base_calculation_process": base_calculation_process(item),
                    "bonus_calculation_process": bonus_calculation_process(item),
                }
                for col_idx, (_, field, _, number_type) in enumerate(columns, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=safe_excel_value(values.get(field, "")))
                    apply_body_style(cell, number_type)
                process_line_count = max(
                    values["base_calculation_process"].count("\n") + 1,
                    values["bonus_calculation_process"].count("\n") + 1,
                )
                sheet.row_dimensions[row_idx].height = min(80, max(24, process_line_count * 8))
            total_row = 4 + len(rows)
            sheet.cell(row=total_row, column=1, value="合计")
            sheet.cell(row=total_row, column=13, value=sum(float(item.get("performance_bonus") or 0) for item in rows))
            for col_idx in range(1, len(columns) + 1):
                cell = sheet.cell(row=total_row, column=col_idx)
                cell.border = thin_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 13:
                    cell.number_format = '"$"#,##0.00'
            sheet.freeze_panes = "A4"
            sheet.auto_filter.ref = f"A3:P{max(3, total_row)}"

        summary = wb.active
        summary.title = "汇总表"
        write_title(summary, "新泽西区绩效考核与奖金核算", 5)
        summary_headers = ["绩效周期", "区域", "部门", "实际奖金总数($)", "备注"]
        for col_idx, header in enumerate(summary_headers, 1):
            cell = summary.cell(row=2, column=col_idx, value=safe_excel_value(header))
            cell.fill = dark_fill
            cell.font = dark_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            summary.column_dimensions[get_column_letter(col_idx)].width = [10, 12, 22, 20, 24][col_idx - 1]
        summary.row_dimensions[2].height = 30

        total_rows = []
        for row_idx, (key, sheet_label) in enumerate(group_defs, 3):
            rows = grouped.get(key, [])
            label = sheet_label.split(".", 1)[-1]
            total_bonus = sum(float(row.get("performance_bonus") or 0) for row in rows)
            total_rows.append((label, len(rows), total_bonus))
            values = [month_label(run.calc_month), "新泽西区", label, total_bonus, "详见后附子表"]
            for col_idx, value in enumerate(values, 1):
                cell = summary.cell(row=row_idx, column=col_idx, value=safe_excel_value(value))
                cell.border = thin_border
                cell.fill = summary_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx == 4:
                    cell.number_format = '"$"#,##0.00'

        total_row_idx = 3 + len(group_defs)
        for col_idx, value in enumerate(["", "新泽西区", "总额", sum(row[2] for row in total_rows), ""], 1):
            cell = summary.cell(row=total_row_idx, column=col_idx, value=safe_excel_value(value))
            cell.border = thin_border
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx == 4:
                cell.number_format = '"$"#,##0.00'
        summary.freeze_panes = "A3"
        if summary.sheet_view.selection:
            summary.sheet_view.selection[0].activeCell = "A2"
            summary.sheet_view.selection[0].sqref = "A2"

        for key, label in group_defs:
            sheet = wb.create_sheet(label)
            if key == "district":
                write_district_sheet(sheet, grouped.get(key, []))
            else:
                write_staff_detail_sheet(sheet, grouped.get(key, []), label)

        wb.active = 0
        output_path = EXPORT_DIR / filename
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return {
            "success": True,
            "filename": filename,
            "file_path": str(output_path),
        }

    # 写入标题。保持普通单元格，避免部分表格软件打开时误触发公式编辑状态。
    title_cell = ws['A1']
    title_cell.value = safe_excel_value(f"FBU美洲绩效核算 - {title} ({run.calc_month})")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 定义列配置
    if type == "attendance":
        columns = [
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('划分区域', 'area', 'emp', 15),
            ('部门全称', 'department', 'emp', 40),
            ('岗位类型', 'job_type', 'emp', 10),
            ('夜班', 'has_night_shift', 'data', 8),
            ('计薪出勤(h)', 'total_base_hours', 'data', 14),
            ('OT1.5(h)', 'total_ot15', 'data', 12),
            ('OT2.0(h)', 'total_ot20', 'data', 12),
            ('病假(h)', 'sick', 'data', 10),
            ('年假(h)', 'annual', 'data', 10),
            ('节假日(h)', 'holiday', 'data', 10),
        ]
    elif type == "salary":
        columns = [
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('划分区域', 'area', 'emp', 15),
            ('部门全称', 'department', 'emp', 40),
            ('时薪($)', 'hourly_rate', 'money', 12),
            ('绩效比例', 'ratio', 'money', 12),
        ]
    elif type == "performance":
        columns = [
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('划分区域', 'area', 'emp', 15),
            ('部门全称', 'department', 'emp', 40),
            ('岗位类型', 'job_type', 'emp', 10),
            ('绩效得分', 'score', 'data', 12),
            ('绩效等级', 'level', 'data', 12),
            ('绩效系数', 'coefficient', 'calc', 12),
        ]
    elif type == "adjustments":
        columns = [
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('划分区域', 'area', 'emp', 15),
            ('部门全称', 'department', 'emp', 40),
            ('分段期间', 'period', 'data', 16),
            ('分段绩效基数($)', 'performance_base', 'money', 18),
            ('核算标识', 'reason', 'calc', 14),
        ]
    elif type == "base_overrides":
        columns = [
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('划分区域', 'area', 'emp', 15),
            ('部门全称', 'department', 'emp', 40),
            ('规则类型', 'rule_type', 'data', 18),
            ('固定绩效基数($)', 'fixed_performance_base', 'money', 18),
            ('归属月份', 'allocation_month', 'data', 12),
            ('状态', 'status', 'data', 12),
            ('是否计入', 'include_in_calculation', 'calc', 12),
            ('核算路径', 'calculation_path', 'calc', 22),
            ('备注', 'note', 'data', 36),
        ]
    elif type == "diagnostics":
        columns = [
            ('严重程度', 'severity', 'calc', 12),
            ('问题类型', 'type', 'data', 18),
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('说明', 'detail', 'data', 42),
        ]
    elif type == "results":
        columns = [
            ('工号', 'employee_id', 'emp', 15),
            ('姓名', 'name', 'emp', 12),
            ('划分区域', 'area', 'emp', 15),
            ('部门全称', 'department', 'emp', 40),
            ('岗位类型', 'job_type', 'emp', 10),
            ('核算路径', 'calculation_path', 'calc', 18),
            ('绩效基数($)', 'performance_base', 'calc', 14),
            ('绩效比例', 'performance_ratio', 'money', 12),
            ('绩效得分', 'performance_score', 'data', 12),
            ('绩效等级', 'performance_level', 'data', 12),
            ('绩效系数', 'performance_coefficient', 'calc', 12),
            ('绩效奖金($)', 'performance_bonus', 'money', 14),
        ]

    # 写入表头
    header_row = 3
    for col_idx, (header, _, style_type, width) in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=safe_excel_value(header))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx - 1) // 26) + chr(65 + (col_idx - 1) % 26)].width = width

    ws.row_dimensions[header_row].height = 25

    # 写入数据
    for row_idx, item in enumerate(data, header_row + 1):
        for col_idx, (_, field, style_type, _) in enumerate(columns, 1):
            value = item.get(field, '')

            # 特殊处理
            if field == 'job_type':
                value = format_fbu_job_type(value)
            elif field == 'calculation_path':
                value = result_calculation_path(item)
            elif field == 'has_night_shift':
                value = '是' if value else '否'
            elif field == 'include_in_calculation':
                value = '是' if value else '否'
            elif field == 'sick':
                # 从day_shift和night_shift获取
                day_shift = item.get('day_shift', {})
                night_shift = item.get('night_shift', {})
                value = day_shift.get('病假', 0) + night_shift.get('病假', 0)
            elif field == 'annual':
                day_shift = item.get('day_shift', {})
                night_shift = item.get('night_shift', {})
                value = day_shift.get('年假', 0) + night_shift.get('年假', 0)
            elif field == 'holiday':
                day_shift = item.get('day_shift', {})
                night_shift = item.get('night_shift', {})
                value = day_shift.get('节假日', 0) + night_shift.get('节假日', 0)

            # 格式化数值
            if isinstance(value, float):
                if field in ['hourly_rate', 'performance_base', 'performance_bonus']:
                    value = round(value, 2)
                elif field in ['ratio', 'performance_ratio']:
                    value = f"{value * 100:.1f}%"
                elif field in ['total_base_hours', 'total_ot15', 'total_ot20', 'sick', 'annual', 'holiday']:
                    value = f"{value:.2f}"
                elif field in ['score', 'coefficient', 'performance_coefficient']:
                    value = round(value, 2)

            cell = ws.cell(row=row_idx, column=col_idx, value=safe_excel_value(value))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # 应用样式
            if style_type == 'emp':
                cell.fill = emp_fill
            elif style_type == 'data':
                cell.fill = data_fill
            elif style_type == 'calc':
                cell.fill = calc_fill
            elif style_type == 'money':
                cell.fill = money_fill

        ws.row_dimensions[row_idx].height = 22

    # 添加汇总行
    summary_row = len(data) + header_row + 2
    ws.cell(row=summary_row, column=1, value=safe_excel_value("汇总")).font = Font(bold=True, size=12)

    if type == "attendance":
        total_base = sum(e.get('total_base_hours', 0) for e in data)
        total_ot15 = sum(e.get('total_ot15', 0) for e in data)
        total_ot20 = sum(e.get('total_ot20', 0) for e in data)
        ws.cell(row=summary_row, column=2, value=safe_excel_value(f"员工数: {len(data)}"))
        ws.cell(row=summary_row, column=7, value=safe_excel_value(f"{total_base:.2f}h"))
        ws.cell(row=summary_row, column=8, value=safe_excel_value(f"{total_ot15:.2f}h"))
        ws.cell(row=summary_row, column=9, value=safe_excel_value(f"{total_ot20:.2f}h"))
    elif type == "results":
        total_bonus = sum(e.get('performance_bonus', 0) for e in data)
        ws.cell(row=summary_row, column=2, value=safe_excel_value(f"员工数: {len(data)}"))
        ws.cell(row=summary_row, column=len(columns), value=safe_excel_value(f"${total_bonus:,.2f}"))

    if ws.sheet_view.selection:
        ws.sheet_view.selection[0].activeCell = f"A{header_row}"
        ws.sheet_view.selection[0].sqref = f"A{header_row}"

    # 保存文件
    output_path = EXPORT_DIR / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    return {
        "success": True,
        "filename": filename,
        "file_path": str(output_path),
    }


@app.get("/api/fbu-performance/runs/{run_id}/download/{filename}")
def download_fbu_performance_file(run_id: str, filename: str) -> FileResponse:
    """下载FBU绩效核算文件"""
    path = EXPORT_DIR / Path(filename).name
    if not path.exists():
        export_types = {
            "考勤汇总_": "attendance",
            "薪资匹配_": "salary",
            "绩效明细_": "performance",
            "调薪拆分_": "adjustments",
            "工时规则与固定基数例外_": "base_overrides",
            "数据诊断_": "diagnostics",
            "核算结果_": "results",
        }
        result_type = next(
            (value for prefix, value in export_types.items() if path.name.startswith(prefix)),
            "",
        )
        if not result_type:
            raise HTTPException(404, "文件不存在")
        export_payload = export_fbu_excel(run_id, type=result_type)
        path = EXPORT_DIR / Path(export_payload["filename"]).name
        if not path.exists():
            raise HTTPException(404, "文件不存在")
    return FileResponse(path, filename=path.name)


@app.post("/api/fbu-performance/runs/bulk-delete")
def bulk_delete_fbu_performance_runs(body: dict = Body(...)) -> dict:
    """批量删除FBU绩效核算任务"""
    run_ids = [str(run_id).strip() for run_id in body.get("run_ids", []) if str(run_id).strip()]
    if not run_ids:
        raise HTTPException(400, "请选择要删除的活动")

    deleted_ids: list[str] = []
    missing_ids: list[str] = []
    for run_id in dict.fromkeys(run_ids):
        run = fbu_run_manager.get_run(run_id, sections=set())
        if not run:
            missing_ids.append(run_id)
            continue

        run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
        if run_dir.exists():
            shutil.rmtree(run_dir, ignore_errors=True)
        fbu_run_manager.delete_run(run_id)
        deleted_ids.append(run_id)

    return {
        "deleted_count": len(deleted_ids),
        "deleted_ids": deleted_ids,
        "missing_ids": missing_ids,
    }


@app.delete("/api/fbu-performance/runs/{run_id}")
def delete_fbu_performance_run(run_id: str) -> dict:
    """删除FBU绩效核算任务"""
    # 检查任务是否存在
    run = fbu_run_manager.get_run(run_id, sections=set())
    if not run:
        raise HTTPException(404, "任务不存在")

    # 删除文件目录
    run_dir = FBU_PERFORMANCE_RUNS_DIR / run_id
    if run_dir.exists():
        shutil.rmtree(run_dir, ignore_errors=True)

    # 删除运行记录
    fbu_run_manager.delete_run(run_id)
    return {"message": f"已删除任务: {run_id}"}


app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="static")
