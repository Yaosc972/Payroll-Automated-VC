from __future__ import annotations

import argparse
from datetime import datetime
import hashlib
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook

from .materials import build_material_index, build_material_replay_plan


SCHEMA_VERSION = 1
DEFAULT_GOLDEN_SUPPLIERS = ("fairway", "oss", "osi", "sss", "workforce", "grande")
ALLOWED_FILE_TYPES = {"invoice_pdf", "workbook", "supporting"}
REQUIRED_BATCH_FIELDS = {
    "batch_key",
    "supplier_ref",
    "files",
    "expected_result",
}
REQUIRED_FILE_FIELDS = {"relative_path", "file_type", "sha256"}
REQUIRED_APPROVED_METRICS = {
    "invoice_total",
    "excel_total",
    "warehouse_count",
    "employee_count",
    "total_hours",
    "difference_category_counts",
    "manual_review_count",
    "core_error_types",
}
EXPECTED_REVIEW_STATUSES = {
    "needs_business_review",
    "provisional",
    "approved",
    "rejected",
    "unknown",
}
EMPLOYEE_DETAIL_INCOMPLETE_CODES = {
    "employee_detail_incomplete",
    "incomplete_employee_detail",
    "pdf_detail_incomplete",
    "invoice_detail_incomplete",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def supplier_ref(value: str) -> str:
    normalized = str(value or "unknown").strip().lower() or "unknown"
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:12]


def discover_golden_batches(
    materials_root: str | Path,
    *,
    batch_key: str = "",
    supplier: str = "",
) -> dict[str, Any]:
    root = Path(materials_root).expanduser()
    index = build_material_index(root)
    batches = index.get("candidateBatches") or []
    selected = []
    for batch in batches:
        if batch_key and str(batch.get("batchKey") or "") != batch_key:
            continue
        if supplier and str(batch.get("supplier") or "").lower() != supplier.lower():
            continue
        selected.append(_build_manifest_batch(root, batch))
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "materials_root": str(root),
        "source": "local_material_discovery",
        "expected_result_policy": "program output is not accepted as truth; business review required",
        "batches": selected,
    }


def build_golden_candidate_plan(
    materials_root: str | Path,
    *,
    required_suppliers: list[str] | tuple[str, ...] | None = None,
) -> dict[str, Any]:
    root = Path(materials_root).expanduser()
    required = [_normalize_supplier_key(item) for item in (required_suppliers or DEFAULT_GOLDEN_SUPPLIERS)]
    index = build_material_index(root)
    batches = index.get("candidateBatches") or []
    selected: list[dict[str, Any]] = []
    missing: list[str] = []
    for supplier in required:
        candidates = [batch for batch in batches if _candidate_matches_supplier(batch, supplier)]
        if not candidates:
            missing.append(supplier)
            selected.append(_missing_supplier_plan(root, supplier))
            continue
        selected.append(_supplier_candidate_plan(root, supplier, candidates))
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "materials_root": str(root),
        "source": "local_material_candidate_plan",
        "required_suppliers": required,
        "summary": {
            "required_supplier_count": len(required),
            "covered_supplier_count": len(required) - len(missing),
            "missing_supplier_count": len(missing),
            "missing_suppliers": missing,
        },
        "supplier_coverage": selected,
        "expected_result_template": _expected_result_template(),
        "review_instructions": [
            "Run the manifest command for each selected batch and keep outputs outside Git.",
            "Business reviewers must fill expected_result.metrics from reviewed evidence, not from program output alone.",
            "Do not change review_status to approved until invoice total, Excel total, warehouse count, employee count, total hours, difference categories, manual review count, and core error types are confirmed.",
        ],
    }


def prepare_golden_manifests(
    materials_root: str | Path,
    output_dir: str | Path,
    *,
    required_suppliers: list[str] | tuple[str, ...] | None = None,
) -> dict[str, Any]:
    root = Path(materials_root).expanduser()
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    plan = build_golden_candidate_plan(root, required_suppliers=required_suppliers)
    (destination / "coverage_plan.json").write_text(
        json.dumps(plan, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    prepared: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    missing: list[str] = []
    for row in plan.get("supplier_coverage", []):
        supplier = str(row.get("supplier") or "")
        batch_key = str(row.get("selected_batch_key") or "")
        if not batch_key:
            missing.append(supplier)
            continue
        manifest = discover_golden_batches(root, batch_key=batch_key)
        manifest_path = destination / _manifest_filename(batch_key)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        validation = validate_golden_manifest(manifest, root)
        record = {
            "supplier": supplier,
            "batch_key": batch_key,
            "manifest_path": str(manifest_path),
            "validation": validation,
        }
        if validation["ok"]:
            prepared.append(record)
        else:
            failed.append(record)
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "materials_root": str(root),
        "output_dir": str(destination),
        "source": "local_material_manifest_prepare",
        "summary": {
            "required_supplier_count": plan["summary"]["required_supplier_count"],
            "prepared_manifest_count": len(prepared),
            "missing_supplier_count": len(missing),
            "failed_manifest_count": len(failed),
        },
        "coverage_plan_path": str(destination / "coverage_plan.json"),
        "prepared_manifests": prepared,
        "failed_manifests": failed,
        "missing_suppliers": missing,
    }


def validate_golden_manifest(
    manifest: dict[str, Any],
    materials_root: str | Path | None = None,
    *,
    require_approved: bool = False,
) -> dict[str, Any]:
    raw_root = materials_root if materials_root is not None else manifest.get("materials_root")
    root = Path(raw_root).expanduser() if raw_root else None
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    if manifest.get("schema_version") != SCHEMA_VERSION:
        errors.append(
            {
                "code": "invalid_schema_version",
                "message": f"schema_version must be {SCHEMA_VERSION}",
            }
        )
    batches = manifest.get("batches") if isinstance(manifest.get("batches"), list) else []
    if root is None:
        errors.append({"code": "missing_materials_root", "message": "materials_root is required"})
    elif not root.exists():
        errors.append({"code": "missing_materials_root", "message": f"materials_root does not exist: {root}"})
    elif not root.is_dir():
        errors.append({"code": "invalid_materials_root", "message": f"materials_root is not a directory: {root}"})
    if not isinstance(manifest.get("batches"), list):
        errors.append({"code": "missing_batches", "message": "batches must be a list"})
    for batch in batches:
        if not isinstance(batch, dict):
            errors.append({"code": "invalid_batch", "message": "batch must be an object"})
            continue
        batch_key = str(batch.get("batch_key") or "")
        for field in sorted(REQUIRED_BATCH_FIELDS):
            if field not in batch or batch.get(field) in (None, "", []):
                errors.append({"code": "missing_batch_field", "batch_key": batch_key, "field": field})
        expected = batch.get("expected_result") if isinstance(batch.get("expected_result"), dict) else {}
        review_status = str(expected.get("review_status") or "")
        if review_status not in EXPECTED_REVIEW_STATUSES:
            errors.append(
                {
                    "code": "invalid_review_status",
                    "batch_key": batch_key,
                    "message": f"unexpected review_status: {review_status}",
                }
            )
        elif require_approved and review_status != "approved":
            errors.append(
                {
                    "code": "expected_not_approved",
                    "batch_key": batch_key,
                    "message": "expected result must be approved for this validation mode",
                }
            )
        if review_status != "approved":
            warnings.append(
                {
                    "code": "not_business_approved",
                    "batch_key": batch_key,
                    "message": "expected result is not approved by business",
                }
            )
        else:
            _validate_approved_metrics(batch_key, expected, errors)
        files = batch.get("files") if isinstance(batch.get("files"), list) else []
        if not files:
            errors.append({"code": "missing_files", "batch_key": batch_key, "message": "batch has no files"})
        for record in files:
            if not isinstance(record, dict):
                errors.append({"code": "invalid_file_record", "batch_key": batch_key})
                continue
            relative_path = str(record.get("relative_path") or "")
            file_type = str(record.get("file_type") or "")
            expected_hash = str(record.get("sha256") or "")
            for field in sorted(REQUIRED_FILE_FIELDS):
                if field not in record or record.get(field) in (None, ""):
                    errors.append(
                        {
                            "code": "missing_file_field",
                            "batch_key": batch_key,
                            "relative_path": relative_path,
                            "field": field,
                        }
                    )
            if file_type and file_type not in ALLOWED_FILE_TYPES:
                errors.append(
                    {
                        "code": "invalid_file_type",
                        "batch_key": batch_key,
                        "relative_path": relative_path,
                        "file_type": file_type,
                    }
                )
            if expected_hash and not _is_sha256(expected_hash):
                errors.append(
                    {
                        "code": "invalid_sha256",
                        "batch_key": batch_key,
                        "relative_path": relative_path,
                    }
                )
            if not relative_path:
                errors.append({"code": "missing_relative_path", "batch_key": batch_key})
                continue
            if root is None:
                continue
            path = root / relative_path
            if not path.exists():
                errors.append({"code": "missing_file", "batch_key": batch_key, "relative_path": relative_path})
                continue
            actual_hash = sha256_file(path)
            if expected_hash != actual_hash:
                errors.append(
                    {
                        "code": "sha256_mismatch",
                        "batch_key": batch_key,
                        "relative_path": relative_path,
                    }
                )
    return {
        "ok": not errors,
        "batch_count": len(batches),
        "file_count": sum(len(batch.get("files") or []) for batch in batches),
        "errors": errors,
        "warnings": warnings,
        "require_approved": require_approved,
    }


def validate_golden_manifest_dir(
    manifest_dir: str | Path,
    materials_root: str | Path | None = None,
    *,
    require_approved: bool = False,
) -> dict[str, Any]:
    directory = Path(manifest_dir).expanduser()
    results: list[dict[str, Any]] = []
    if not directory.exists():
        return _manifest_dir_result(directory, [], [{"code": "missing_manifest_dir", "message": f"manifest_dir does not exist: {directory}"}], require_approved)
    if not directory.is_dir():
        return _manifest_dir_result(directory, [], [{"code": "invalid_manifest_dir", "message": f"manifest_dir is not a directory: {directory}"}], require_approved)
    for path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(path)
        validation = validate_golden_manifest(
            manifest,
            materials_root=materials_root or manifest.get("materials_root"),
            require_approved=require_approved,
        )
        results.append(
            {
                "manifest_path": str(path),
                "ok": validation["ok"],
                "batch_count": validation["batch_count"],
                "file_count": validation["file_count"],
                "errors": validation["errors"],
                "warnings": validation["warnings"],
            }
        )
    if not results:
        return _manifest_dir_result(directory, [], [{"code": "missing_manifests", "message": "no *_manifest.json files found"}], require_approved)
    return _manifest_dir_result(directory, results, [], require_approved)


def build_golden_review_template(
    manifest_dir: str | Path,
    *,
    materials_root: str | Path | None = None,
    batch_key: str | None = None,
    supplier_ref: str | None = None,
) -> dict[str, Any]:
    directory = Path(manifest_dir).expanduser()
    requested_batch_key = str(batch_key or "").strip()
    requested_supplier_ref = str(supplier_ref or "").strip()
    items: list[dict[str, Any]] = []
    for path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(path)
        validation = validate_golden_manifest(manifest, materials_root or manifest.get("materials_root"))
        for batch in manifest.get("batches") or []:
            if not isinstance(batch, dict):
                continue
            current_batch_key = str(batch.get("batch_key") or "")
            if requested_batch_key and current_batch_key != requested_batch_key:
                continue
            current_supplier_ref = str(batch.get("supplier_ref") or "")
            if requested_supplier_ref and current_supplier_ref != requested_supplier_ref:
                continue
            expected = batch.get("expected_result") if isinstance(batch.get("expected_result"), dict) else {}
            review_status = str(expected.get("review_status") or "unknown")
            files = batch.get("files") if isinstance(batch.get("files"), list) else []
            items.append(
                {
                    "batch_key": current_batch_key,
                    "supplier_ref": current_supplier_ref,
                    "period_hint": str(batch.get("period_hint") or ""),
                    "review_status": review_status,
                    "file_summary": _summarize_file_types(files),
                    "file_hashes": _review_file_hashes(files),
                    "excel": _review_excel(batch.get("excel") if isinstance(batch.get("excel"), dict) else {}),
                    "expected_metrics": _expected_result_template()["metrics"],
                    "allowed_manual_review_items": list(batch.get("allowed_manual_review_items") or []),
                    "reviewer": "",
                    "reviewed_at": "",
                    "evidence_reference": "",
                    "review_notes": "",
                    "validation": {
                        "ok": validation["ok"],
                        "errors": validation["errors"],
                        "warnings": validation["warnings"],
                    },
                }
            )
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source": "golden_business_review_template",
        "manifest_dir": str(directory),
        "review_policy": "Business reviewers must fill expected_metrics from reviewed invoice, bill, and reconciliation evidence. program output is not accepted as truth.",
        "summary": {
            "batch_count": len(items),
            "file_count": sum(sum(item["file_summary"].values()) for item in items),
            "needs_business_review_count": sum(1 for item in items if item["review_status"] != "approved"),
        },
        "review_items": items,
    }


def build_golden_review_handoff(
    manifest_dir: str | Path,
    output_dir: str | Path,
    *,
    materials_root: str | Path | None = None,
    batch_key: str | None = None,
    supplier_ref: str | None = None,
) -> dict[str, Any]:
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    template = _sanitize_review_template_for_handoff(
        build_golden_review_template(
            manifest_dir,
            materials_root=materials_root,
            batch_key=batch_key,
            supplier_ref=supplier_ref,
        )
    )
    guidance = _expected_metric_guidance()
    template_path = destination / "business_review_template.json"
    workbook_path = destination / "business_review_template.xlsx"
    readme_path = destination / "BUSINESS_REVIEW_README.md"
    summary_path = destination / "handoff_summary.json"
    template_path.write_text(json.dumps(template, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    _write_business_review_workbook(workbook_path, template)
    readme_path.write_text(_business_review_readme(template, guidance), encoding="utf-8")
    payload = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source": "golden_business_review_handoff",
        "ok": True,
        "summary": template["summary"],
        "metric_guidance": guidance,
        "files": {
            "review_template": template_path.name,
            "review_workbook": workbook_path.name,
            "readme": readme_path.name,
            "summary": summary_path.name,
        },
    }
    summary_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return payload


def scan_golden_handoff_privacy(
    handoff_dir: str | Path,
    *,
    materials_root: str | Path | None = None,
    forbidden_terms: list[str] | tuple[str, ...] | None = None,
) -> dict[str, Any]:
    directory = Path(handoff_dir).expanduser()
    issues: list[dict[str, Any]] = []
    scanned_files: list[str] = []
    if not directory.exists():
        issues.append({"code": "missing_handoff_dir", "message": f"handoff_dir does not exist: {directory}"})
        return _privacy_scan_result(directory, scanned_files, issues)
    if not directory.is_dir():
        issues.append({"code": "invalid_handoff_dir", "message": f"handoff_dir is not a directory: {directory}"})
        return _privacy_scan_result(directory, scanned_files, issues)

    root_text = str(Path(materials_root).expanduser()) if materials_root else ""
    explicit_terms = [str(term) for term in (forbidden_terms or []) if str(term)]
    for path in sorted(item for item in directory.rglob("*") if item.is_file()):
        if path.suffix.lower() not in {".json", ".md", ".txt", ".xlsx", ".xlsm"}:
            continue
        relative_path = str(path.relative_to(directory))
        scanned_files.append(relative_path)
        text = _handoff_file_text(path)
        if root_text and root_text in text:
            issues.append({"code": "materials_root_path_leak", "file": relative_path})
        for term in explicit_terms:
            if term in text:
                issues.append({"code": "forbidden_term_leak", "file": relative_path, "term": term})
        for match in re.finditer(r"\b(?:EUS)?\d{5,}\b", text):
            issues.append({"code": "employee_id_like_value", "file": relative_path, "value": match.group(0)})
    return _privacy_scan_result(directory, scanned_files, issues)


def _handoff_file_text(path: Path) -> str:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, data_only=True)
        values: list[str] = []
        for sheet in workbook.worksheets:
            values.append(sheet.title)
            for row in sheet.iter_rows(values_only=True):
                values.extend(str(value) for value in row if value not in (None, ""))
        return "\n".join(values)
    return path.read_text(encoding="utf-8", errors="replace")


def validate_golden_review_template(
    review_template: str | Path | dict[str, Any],
    *,
    require_approved: bool = False,
    batch_key: str | None = None,
    supplier_ref: str | None = None,
    review_batch_ref: str | None = None,
) -> dict[str, Any]:
    template = _load_review_template(review_template)
    top_level_errors: list[dict[str, Any]] = []
    requested_batch_key = str(batch_key or "").strip()
    requested_supplier_ref = str(supplier_ref or "").strip()
    requested_review_batch_ref = str(review_batch_ref or "").strip()
    items = template.get("review_items") if isinstance(template.get("review_items"), list) else []
    if template.get("schema_version") != SCHEMA_VERSION:
        top_level_errors.append({"code": "invalid_schema_version", "message": f"schema_version must be {SCHEMA_VERSION}"})
    if not isinstance(template.get("review_items"), list):
        top_level_errors.append({"code": "missing_review_items", "message": "review_items must be a list"})
    rows: list[dict[str, Any]] = []
    for item in items:
        item_dict = item if isinstance(item, dict) else {}
        if requested_batch_key and str(item_dict.get("batch_key") or "") != requested_batch_key:
            continue
        if requested_supplier_ref and str(item_dict.get("supplier_ref") or "") != requested_supplier_ref:
            continue
        if requested_review_batch_ref and str(item_dict.get("review_batch_ref") or "") != requested_review_batch_ref:
            continue
        row = _validate_review_item(item_dict, require_approved=require_approved)
        rows.append(row)
    if requested_batch_key and not rows:
        top_level_errors.append(
            {
                "code": "batch_key_not_found",
                "batch_key": requested_batch_key,
                "message": "batch_key was not found in review_items",
            }
        )
    if requested_supplier_ref and not rows:
        top_level_errors.append(
            {
                "code": "supplier_ref_not_found",
                "supplier_ref": requested_supplier_ref,
                "message": "supplier_ref was not found in review_items",
            }
        )
    if requested_review_batch_ref and not rows:
        top_level_errors.append(
            {
                "code": "review_batch_ref_not_found",
                "review_batch_ref": requested_review_batch_ref,
                "message": "review_batch_ref was not found in review_items",
            }
        )
    error_count = len(top_level_errors) + sum(len(row["errors"]) for row in rows)
    warning_count = sum(len(row["warnings"]) for row in rows)
    return {
        "schema_version": SCHEMA_VERSION,
        "source": "golden_business_review_template_validation",
        "ok": error_count == 0,
        "require_approved": require_approved,
        "summary": {
            "review_item_count": len(rows),
            "approved_count": sum(1 for row in rows if row["review_status"] == "approved"),
            "error_count": error_count,
            "warning_count": warning_count,
        },
        "errors": top_level_errors,
        "items": rows,
    }


def apply_golden_review_template(
    manifest_dir: str | Path,
    review_template: str | Path | dict[str, Any],
    output_dir: str | Path,
    *,
    materials_root: str | Path | None = None,
    require_approved: bool = False,
    batch_key: str | None = None,
    supplier_ref: str | None = None,
    review_batch_ref: str | None = None,
) -> dict[str, Any]:
    source_dir = Path(manifest_dir).expanduser()
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    template = _load_review_template(review_template)
    requested_batch_key = str(batch_key or "").strip()
    requested_supplier_ref = str(supplier_ref or "").strip()
    requested_review_batch_ref = str(review_batch_ref or "").strip()
    template_validation = validate_golden_review_template(
        template,
        require_approved=require_approved,
        batch_key=requested_batch_key or None,
        supplier_ref=requested_supplier_ref or None,
        review_batch_ref=requested_review_batch_ref or None,
    )
    if require_approved and not template_validation["ok"]:
        return {
            "schema_version": SCHEMA_VERSION,
            "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "source": "golden_business_review_apply",
            "manifest_dir": str(source_dir),
            "review_template": str(review_template) if not isinstance(review_template, dict) else "<in-memory>",
            "output_dir": str(destination),
            "require_approved": require_approved,
            "batch_key": requested_batch_key,
            "supplier_ref": requested_supplier_ref,
            "review_batch_ref": requested_review_batch_ref,
            "ok": False,
            "summary": {
                "source_manifest_count": len(list(source_dir.glob("*_manifest.json"))),
                "review_item_count": template_validation["summary"]["review_item_count"],
                "updated_manifest_count": 0,
                "failed_manifest_count": 0,
                "skipped_count": 0,
                "template_error_count": template_validation["summary"]["error_count"],
            },
            "template_validation": template_validation,
            "updated_manifests": [],
            "failed_manifests": [],
            "skipped": [],
        }
    review_items = _review_items_by_batch(template, source_dir)
    if requested_batch_key:
        review_items = {requested_batch_key: review_items[requested_batch_key]} if requested_batch_key in review_items else {}
    if requested_supplier_ref:
        review_items = {
            key: item
            for key, item in review_items.items()
            if str(item.get("supplier_ref") or "") == requested_supplier_ref
        }
    if requested_review_batch_ref:
        review_items = {
            key: item
            for key, item in review_items.items()
            if str(item.get("review_batch_ref") or "") == requested_review_batch_ref
        }
    updated: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    failed: list[dict[str, Any]] = []
    changed_batch_count = 0

    for manifest_path in sorted(source_dir.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        changed = False
        seen_requested_batch = False
        seen_requested_supplier_ref = False
        for batch in manifest.get("batches") or []:
            if not isinstance(batch, dict):
                continue
            batch_key = str(batch.get("batch_key") or "")
            if requested_batch_key and batch_key != requested_batch_key:
                continue
            if requested_batch_key:
                seen_requested_batch = True
            current_supplier_ref = str(batch.get("supplier_ref") or "")
            if requested_supplier_ref and current_supplier_ref != requested_supplier_ref:
                continue
            if requested_supplier_ref:
                seen_requested_supplier_ref = True
            review = review_items.get(batch_key)
            if not review:
                skipped.append(
                    {
                        "manifest_path": str(manifest_path),
                        "batch_key": batch_key,
                        "reason": "missing_review_item",
                    }
                )
                continue
            _apply_review_item_to_batch(batch, review)
            changed = True
            changed_batch_count += 1
        if requested_batch_key and not seen_requested_batch:
            continue
        if requested_supplier_ref and not seen_requested_supplier_ref:
            continue

        output_path = destination / manifest_path.name
        output_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        validation = validate_golden_manifest(
            manifest,
            materials_root=materials_root or manifest.get("materials_root"),
            require_approved=require_approved and not requested_batch_key and not requested_supplier_ref and not requested_review_batch_ref,
        )
        record = {
            "source_manifest_path": str(manifest_path),
            "manifest_path": str(output_path),
            "changed": changed,
            "validation": validation,
        }
        if validation["ok"]:
            updated.append(record)
        else:
            failed.append(record)

    missing_reviews = sorted(set(review_items) - _manifest_batch_keys(source_dir))
    for batch_key in missing_reviews:
        skipped.append({"batch_key": batch_key, "reason": "missing_source_manifest"})
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source": "golden_business_review_apply",
        "manifest_dir": str(source_dir),
        "review_template": str(review_template) if not isinstance(review_template, dict) else "<in-memory>",
        "output_dir": str(destination),
        "require_approved": require_approved,
        "batch_key": requested_batch_key,
        "supplier_ref": requested_supplier_ref,
        "review_batch_ref": requested_review_batch_ref,
        "ok": not failed and not missing_reviews,
        "summary": {
            "source_manifest_count": len(list(source_dir.glob("*_manifest.json"))),
            "review_item_count": len(review_items),
            "updated_manifest_count": len(updated),
            "changed_manifest_count": sum(1 for record in updated if record.get("changed")),
            "changed_batch_count": changed_batch_count,
            "failed_manifest_count": len(failed),
            "skipped_count": len(skipped),
            "template_error_count": template_validation["summary"]["error_count"],
        },
        "template_validation": template_validation,
        "updated_manifests": updated,
        "failed_manifests": failed,
        "skipped": skipped,
    }


def run_golden_manifest_replay(
    manifest_dir: str | Path,
    output_dir: str | Path,
    *,
    materials_root: str | Path | None = None,
    batch_key: str | None = None,
    supplier_ref: str | None = None,
) -> dict[str, Any]:
    directory = Path(manifest_dir).expanduser()
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    requested_batch_key = str(batch_key or "").strip()
    requested_supplier_ref = str(supplier_ref or "").strip()
    validation_dir = (
        _write_filtered_manifest_dir(
            directory,
            destination,
            batch_key=requested_batch_key,
            supplier_ref=requested_supplier_ref,
        )
        if requested_batch_key or requested_supplier_ref
        else directory
    )
    preflight_errors: list[dict[str, Any]] = []
    if requested_batch_key and not _manifest_batch_keys(validation_dir):
        preflight_errors.append(
            {
                "code": "batch_key_not_found",
                "batch_key": requested_batch_key,
                "message": "batch_key was not found in golden manifests",
            }
        )
    if requested_supplier_ref and not _manifest_supplier_refs(validation_dir):
        preflight_errors.append(
            {
                "code": "supplier_ref_not_found",
                "supplier_ref": requested_supplier_ref,
                "message": "supplier_ref was not found in golden manifests",
            }
        )
    validation = validate_golden_manifest_dir(
        validation_dir,
        materials_root=materials_root,
        require_approved=True,
    )
    if preflight_errors:
        validation = _manifest_dir_result(validation_dir, [], preflight_errors, True)
    business_metrics = _approved_business_metrics(validation_dir)
    approved_batch_count = _approved_batch_count(validation_dir)
    summary = {
        "manifest_count": validation["summary"]["manifest_count"],
        "batch_count": validation["summary"]["batch_count"],
        "file_count": validation["summary"]["file_count"],
        "approved_batch_count": approved_batch_count,
        "error_count": validation["summary"]["error_count"],
        "warning_count": validation["summary"]["warning_count"],
        "business_metrics": business_metrics,
        "business_checklist": _business_acceptance_checklist(
            business_metrics,
            approved_batch_count=approved_batch_count,
            validation_ok=bool(validation["ok"]),
        ),
    }
    digest_payload = {
        "schema_version": SCHEMA_VERSION,
        "manifest_dir": str(directory),
        "batch_key": requested_batch_key,
        "supplier_ref": requested_supplier_ref,
        "materials_root": str(Path(materials_root).expanduser()) if materials_root else "",
        "summary": summary,
        "validation": validation,
    }
    digest = hashlib.sha256(json.dumps(digest_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
    payload = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source": "golden_manifest_replay",
        "manifest_dir": str(directory),
        "output_dir": str(destination),
        "batch_key": requested_batch_key,
        "supplier_ref": requested_supplier_ref,
        "materials_root": str(Path(materials_root).expanduser()) if materials_root else "",
        "ok": validation["ok"],
        "summary": summary,
        "deterministic_digest": digest,
        "validation": validation,
    }
    (destination / "golden_manifest_replay_summary.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (destination / "BUSINESS_REPLAY_SUMMARY.md").write_text(
        _business_replay_summary_markdown(payload),
        encoding="utf-8",
    )
    return payload


def _write_filtered_manifest_dir(source_dir: Path, destination: Path, *, batch_key: str = "", supplier_ref: str = "") -> Path:
    filtered_dir = destination / "_filtered_manifests"
    filtered_dir.mkdir(parents=True, exist_ok=True)
    for manifest_path in sorted(source_dir.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        batches = [
            batch
            for batch in manifest.get("batches") or []
            if isinstance(batch, dict)
            and (not batch_key or str(batch.get("batch_key") or "") == batch_key)
            and (not supplier_ref or str(batch.get("supplier_ref") or "") == supplier_ref)
        ]
        if not batches:
            continue
        filtered = dict(manifest)
        filtered["batches"] = batches
        (filtered_dir / manifest_path.name).write_text(
            json.dumps(filtered, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    return filtered_dir


def _approved_business_metrics(directory: Path) -> dict[str, Any]:
    invoice_total = 0.0
    excel_total = 0.0
    warehouse_count = 0
    employee_count = 0
    total_hours = 0.0
    manual_review_count = 0
    difference_category_counts: dict[str, int] = {}
    core_error_types: set[str] = set()
    total_passed_batch_count = 0
    total_difference_batch_count = 0

    for batch in _iter_manifest_batches(directory):
        expected = batch.get("expected_result") if isinstance(batch.get("expected_result"), dict) else {}
        if expected.get("review_status") != "approved":
            continue
        metrics = expected.get("metrics") if isinstance(expected.get("metrics"), dict) else {}
        batch_invoice_total = _metric_number(metrics.get("invoice_total"))
        batch_excel_total = _metric_number(metrics.get("excel_total"))
        invoice_total += batch_invoice_total
        excel_total += batch_excel_total
        warehouse_count += int(_metric_number(metrics.get("warehouse_count")))
        employee_count += int(_metric_number(metrics.get("employee_count")))
        total_hours += _metric_number(metrics.get("total_hours"))
        manual_review_count += int(_metric_number(metrics.get("manual_review_count")))
        for key, value in (metrics.get("difference_category_counts") or {}).items():
            difference_category_counts[str(key)] = difference_category_counts.get(str(key), 0) + int(_metric_number(value))
        for item in metrics.get("core_error_types") or []:
            if str(item):
                core_error_types.add(str(item))
        if abs(batch_invoice_total - batch_excel_total) <= 0.1:
            total_passed_batch_count += 1
        else:
            total_difference_batch_count += 1

    return {
        "invoice_total": round(invoice_total, 2),
        "excel_total": round(excel_total, 2),
        "amount_delta": round(invoice_total - excel_total, 2),
        "warehouse_count": warehouse_count,
        "employee_count": employee_count,
        "total_hours": round(total_hours, 2),
        "manual_review_count": manual_review_count,
        "difference_category_counts": dict(sorted(difference_category_counts.items())),
        "core_error_types": sorted(core_error_types),
        "total_passed_batch_count": total_passed_batch_count,
        "total_difference_batch_count": total_difference_batch_count,
    }


def _business_acceptance_checklist(
    metrics: dict[str, Any],
    *,
    approved_batch_count: int = 1,
    validation_ok: bool = True,
) -> dict[str, Any]:
    if approved_batch_count <= 0 or not validation_ok:
        return {
            "amount_tolerance": 0.1,
            "total_amount_passed": False,
            "business_conclusion": "尚未完成业务验收基线",
            "amount_difference_message": "等待业务确认总金额",
            "employee_detail_status": "等待业务确认",
            "pending_confirmation_count": 0,
            "auto_fixed_name_count": 0,
            "suspected_same_employee_count": 0,
            "employee_detail_message": "请先由业务确认真实材料的发票总额、账单总额、员工人数和待确认事项，再用于上线验收。",
        }
    amount_delta = _metric_number(metrics.get("amount_delta"))
    manual_review_count = int(_metric_number(metrics.get("manual_review_count")))
    difference_counts = metrics.get("difference_category_counts") if isinstance(metrics.get("difference_category_counts"), dict) else {}
    core_error_types = {str(item) for item in metrics.get("core_error_types") or []}
    total_difference_batch_count = int(_metric_number(metrics.get("total_difference_batch_count")))
    pending_confirmation_count = manual_review_count
    auto_fixed_name_count = int(_metric_number(difference_counts.get("auto_fixed_name")))
    suspected_same_employee_count = int(_metric_number(difference_counts.get("suspected_same_employee")))
    total_amount_passed = total_difference_batch_count == 0 and abs(amount_delta) <= 0.1
    employee_detail_incomplete = bool(core_error_types & EMPLOYEE_DETAIL_INCOMPLETE_CODES)
    needs_detail_confirmation = pending_confirmation_count > 0 or suspected_same_employee_count > 0 or employee_detail_incomplete

    if total_amount_passed and needs_detail_confirmation:
        business_conclusion = "总账通过，但员工明细待确认"
    elif total_amount_passed:
        business_conclusion = "总账通过"
    else:
        business_conclusion = "总金额存在差异，暂不能放行"

    checklist = {
        "amount_tolerance": 0.1,
        "total_amount_passed": total_amount_passed,
        "business_conclusion": business_conclusion,
        "amount_difference_message": _amount_difference_message(amount_delta),
        "employee_detail_status": "员工明细待确认" if needs_detail_confirmation else "员工明细已确认",
        "pending_confirmation_count": pending_confirmation_count,
        "auto_fixed_name_count": auto_fixed_name_count,
        "suspected_same_employee_count": suspected_same_employee_count,
    }
    if employee_detail_incomplete:
        if total_amount_passed:
            checklist["employee_detail_message"] = (
                "系统已确认本批总金额一致，但部分员工明细未完整识别，员工级差异仅供确认，不能直接作为最终员工明细结论。"
            )
        else:
            checklist["employee_detail_message"] = "由于员工明细未完整识别，系统暂时无法定位全部差异来源。"
    return checklist


def _amount_difference_message(amount_delta: float) -> str:
    rounded_delta = round(amount_delta, 2)
    if abs(rounded_delta) <= 0.0:
        return "PDF 与 Excel 总金额一致"
    if rounded_delta > 0:
        return f"PDF 比 Excel 多 ${abs(rounded_delta):,.2f}"
    return f"PDF 比 Excel 少 ${abs(rounded_delta):,.2f}"


def _business_replay_summary_markdown(payload: dict[str, Any]) -> str:
    summary = payload.get("summary") if isinstance(payload.get("summary"), dict) else {}
    metrics = summary.get("business_metrics") if isinstance(summary.get("business_metrics"), dict) else {}
    checklist = summary.get("business_checklist") if isinstance(summary.get("business_checklist"), dict) else {}
    return "\n".join(
        [
            "# 真实材料回归业务摘要",
            "",
            f"结论：{checklist.get('business_conclusion') or '未生成结论'}",
            f"金额差异：{checklist.get('amount_difference_message') or '未生成金额差异说明'}",
            f"总金额容差：${_metric_number(checklist.get('amount_tolerance')):,.2f}",
            "",
            "## 金额口径",
            "",
            f"- PDF 发票总额：${_metric_number(metrics.get('invoice_total')):,.2f}",
            f"- Excel 账单总额：${_metric_number(metrics.get('excel_total')):,.2f}",
            f"- 总金额差额：${_metric_number(metrics.get('amount_delta')):,.2f}",
            "",
            "## 员工明细确认",
            "",
            f"- 员工明细状态：{checklist.get('employee_detail_status') or '未生成状态'}",
            *([f"- 明细说明：{checklist.get('employee_detail_message')}"] if checklist.get("employee_detail_message") else []),
            f"- 待业务确认：{int(_metric_number(checklist.get('pending_confirmation_count')))} 项",
            f"- 系统自动修正姓名：{int(_metric_number(checklist.get('auto_fixed_name_count')))} 项",
            f"- 疑似同一员工：{int(_metric_number(checklist.get('suspected_same_employee_count')))} 项",
            "",
            "## 回归覆盖",
            "",
            f"- 已审批批次：{int(_metric_number(summary.get('approved_batch_count')))} 批",
            f"- 总账通过批次：{int(_metric_number(metrics.get('total_passed_batch_count')))} 批",
            f"- 总账差异批次：{int(_metric_number(metrics.get('total_difference_batch_count')))} 批",
            "",
            "这份摘要用于开发回归检查，判断页面结论是否仍符合业务规则；正式对业务展示仍以核对页面和 HTML 报告为准。",
            "",
        ]
    )


def _iter_manifest_batches(directory: Path) -> list[dict[str, Any]]:
    batches: list[dict[str, Any]] = []
    for manifest_path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        batches.extend(batch for batch in manifest.get("batches") or [] if isinstance(batch, dict))
    return batches


def _metric_number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _review_items_by_batch(template: dict[str, Any], source_dir: Path | None = None) -> dict[str, dict[str, Any]]:
    items = template.get("review_items") if isinstance(template.get("review_items"), list) else []
    rows: dict[str, dict[str, Any]] = {}
    ref_to_batch = _review_batch_refs_by_batch(source_dir) if source_dir else {}
    for item in items:
        if not isinstance(item, dict):
            continue
        batch_key = str(item.get("batch_key") or "").strip()
        if not batch_key:
            batch_key = ref_to_batch.get(str(item.get("review_batch_ref") or "").strip(), "")
        if batch_key:
            rows[batch_key] = item
    return rows


def _validate_review_item(item: dict[str, Any], *, require_approved: bool) -> dict[str, Any]:
    batch_key = str(item.get("batch_key") or "")
    review_batch_ref = str(item.get("review_batch_ref") or "")
    review_ref = batch_key or review_batch_ref
    review_status = str(item.get("review_status") or "")
    errors: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    if not review_ref:
        errors.append({"code": "missing_review_batch_ref", "message": "batch_key or review_batch_ref is required"})
    if review_status not in EXPECTED_REVIEW_STATUSES:
        errors.append({"code": "invalid_review_status", "batch_key": batch_key, "review_batch_ref": review_batch_ref, "message": f"unexpected review_status: {review_status}"})
    elif require_approved and review_status != "approved":
        errors.append({"code": "review_item_not_approved", "batch_key": batch_key, "review_batch_ref": review_batch_ref, "message": "review_status must be approved"})
    if review_status == "approved":
        if not str(item.get("reviewer") or "").strip():
            errors.append({"code": "missing_reviewer", "batch_key": batch_key, "review_batch_ref": review_batch_ref})
        if not str(item.get("reviewed_at") or "").strip():
            errors.append({"code": "missing_reviewed_at", "batch_key": batch_key, "review_batch_ref": review_batch_ref})
        if not str(item.get("evidence_reference") or "").strip():
            errors.append({"code": "missing_evidence_reference", "batch_key": batch_key, "review_batch_ref": review_batch_ref})
        metrics = item.get("expected_metrics") if isinstance(item.get("expected_metrics"), dict) else {}
        _validate_review_metrics(review_ref, metrics, errors)
    else:
        warnings.append({"code": "not_business_approved", "batch_key": batch_key, "review_batch_ref": review_batch_ref})
    return {
        "batch_key": batch_key,
        "review_batch_ref": review_batch_ref,
        "review_status": review_status,
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
    }


def _validate_review_metrics(batch_key: str, metrics: dict[str, Any], errors: list[dict[str, Any]]) -> None:
    for field in sorted(REQUIRED_APPROVED_METRICS):
        if field not in metrics or metrics.get(field) in (None, ""):
            errors.append({"code": "approved_expected_metric_missing", "batch_key": batch_key, "field": field})
            continue
        value = metrics.get(field)
        if field in {"invoice_total", "excel_total", "total_hours"} and not isinstance(value, (int, float)):
            errors.append({"code": "invalid_expected_metric_type", "batch_key": batch_key, "field": field, "expected_type": "number"})
        elif field in {"warehouse_count", "employee_count", "manual_review_count"} and not isinstance(value, int):
            errors.append({"code": "invalid_expected_metric_type", "batch_key": batch_key, "field": field, "expected_type": "integer"})
        elif field == "difference_category_counts" and not isinstance(value, dict):
            errors.append({"code": "invalid_expected_metric_type", "batch_key": batch_key, "field": field, "expected_type": "object"})
        elif field == "core_error_types" and not isinstance(value, list):
            errors.append({"code": "invalid_expected_metric_type", "batch_key": batch_key, "field": field, "expected_type": "array"})


def _sanitize_review_template_for_handoff(template: dict[str, Any]) -> dict[str, Any]:
    sanitized = dict(template)
    sanitized.pop("manifest_dir", None)
    rows: list[dict[str, Any]] = []
    for item in sanitized.get("review_items") or []:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        review_batch_ref = _review_batch_ref(row)
        row["review_batch_ref"] = review_batch_ref
        row.pop("batch_key", None)
        row.pop("supplier_ref", None)
        validation = row.get("validation") if isinstance(row.get("validation"), dict) else {}
        row["validation"] = {
            "ok": bool(validation.get("ok")),
            "errors": [_public_validation_issue(error, review_batch_ref=review_batch_ref) for error in validation.get("errors") or []],
            "warnings": [_public_validation_issue(warning, review_batch_ref=review_batch_ref) for warning in validation.get("warnings") or []],
        }
        rows.append(row)
    sanitized["review_items"] = rows
    return sanitized


def _review_batch_ref(item: dict[str, Any]) -> str:
    supplier = str(item.get("supplier_ref") or "unknown").strip() or "unknown"
    batch_key = str(item.get("batch_key") or "").strip()
    supplier_digest = hashlib.sha256(supplier.encode("utf-8")).hexdigest()[:12]
    digest = hashlib.sha256(f"{supplier}:{batch_key}".encode("utf-8")).hexdigest()[:12]
    return f"batch_{supplier_digest}_{digest}"


def _public_validation_issue(issue: dict[str, Any], *, review_batch_ref: str = "") -> dict[str, Any]:
    public = {
        "code": str(issue.get("code") or ""),
        "review_batch_ref": review_batch_ref,
        "field": str(issue.get("field") or ""),
        "message": str(issue.get("message") or ""),
    }
    return {key: value for key, value in public.items() if value}


def _expected_metric_guidance() -> dict[str, str]:
    return {
        "invoice_total": "发票合计金额。以业务复核后的发票总额为准；不得直接复制程序输出。",
        "excel_total": "账单合计金额。以业务复核后的 Excel/账单总额为准。",
        "warehouse_count": "本批涉及仓库数量。按业务确认的仓库口径计数。",
        "employee_count": "本批员工人数。按业务确认应参与核对的员工去重计数。",
        "total_hours": "本批总工时。按业务确认的账单或发票明细汇总。",
        "difference_category_counts": "差异分类数量，例如 amount_mismatch、hours_mismatch、invoice_only、bill_only。",
        "manual_review_count": "仍需人工复核的记录数量。确认无待复核时填 0。",
        "core_error_types": "核心异常类型列表。无异常时填空数组。",
    }


def _write_business_review_workbook(path: Path, template: dict[str, Any]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "业务确认表"
    headers = [
        "批次编号",
        "账期",
        "PDF 发票文件数",
        "Excel 账单文件数",
        "业务确认状态",
        "发票总额",
        "账单总额",
        "仓库数",
        "员工人数",
        "总工时",
        "待确认数量",
        "核心异常",
        "差异分类数量",
        "复核人",
        "复核时间",
        "依据说明",
        "备注",
    ]
    sheet.append(headers)
    for item in template.get("review_items") or []:
        if not isinstance(item, dict):
            continue
        metrics = item.get("expected_metrics") if isinstance(item.get("expected_metrics"), dict) else {}
        file_summary = item.get("file_summary") if isinstance(item.get("file_summary"), dict) else {}
        sheet.append(
            [
                item.get("review_batch_ref") or item.get("batch_key") or "",
                item.get("period_hint") or "",
                int(_metric_number(file_summary.get("invoice_pdf"))),
                int(_metric_number(file_summary.get("workbook"))),
                item.get("review_status") or "needs_business_review",
                metrics.get("invoice_total"),
                metrics.get("excel_total"),
                metrics.get("warehouse_count"),
                metrics.get("employee_count"),
                metrics.get("total_hours"),
                metrics.get("manual_review_count"),
                ", ".join(str(value) for value in metrics.get("core_error_types") or []),
                json.dumps(metrics.get("difference_category_counts") or {}, ensure_ascii=False),
                item.get("reviewer") or "",
                item.get("reviewed_at") or "",
                item.get("evidence_reference") or "",
                item.get("review_notes") or "",
            ]
        )
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 36)
    workbook.save(path)


def _business_review_readme(template: dict[str, Any], guidance: dict[str, str]) -> str:
    lines = [
        "# 海外劳务工黄金回归业务复核交接包",
        "",
        "## 复核原则",
        "",
        "- 程序输出不能作为真值，必须以业务复核后的发票、账单和核对证据为准。",
        "- 只有确认无误后，才能把 `review_status` 改为 `approved`。",
        "- 不能确认的批次保持 `needs_business_review`、`provisional` 或 `unknown`。",
        "- 不要在模板中补充员工明细、完整文件名、本地路径或其他个人敏感信息。",
        "",
        "## 本包摘要",
        "",
        f"- 批次数：{template['summary']['batch_count']}",
        f"- 文件数：{template['summary']['file_count']}",
        f"- 待业务复核：{template['summary']['needs_business_review_count']}",
        "",
        "## expected_metrics 填写口径",
        "",
    ]
    for field, description in guidance.items():
        lines.append(f"- `{field}`：{description}")
    lines.extend(
        [
            "",
            "## 交付后步骤",
            "",
            "1. 填写 `business_review_template.json` 中每个批次的 `expected_metrics`。",
            "2. 填写 `reviewer`、`reviewed_at`、`evidence_reference` 和必要备注。",
            "3. 仅在确认全部必填指标后，将 `review_status` 改为 `approved`。",
            "4. 使用 `apply-review --require-approved` 生成 reviewed manifests。",
            "5. 使用 `replay` 做 manifest-level release gate。",
            "",
        ]
    )
    return "\n".join(lines)


def _apply_review_item_to_batch(batch: dict[str, Any], review: dict[str, Any]) -> None:
    expected = batch.get("expected_result") if isinstance(batch.get("expected_result"), dict) else {}
    updated_expected = {
        "review_status": str(review.get("review_status") or expected.get("review_status") or "needs_business_review"),
        "reviewer": str(review.get("reviewer") or expected.get("reviewer") or ""),
        "reviewed_at": str(review.get("reviewed_at") or expected.get("reviewed_at") or ""),
        "evidence_reference": str(review.get("evidence_reference") or expected.get("evidence_reference") or ""),
        "metrics": dict(review.get("expected_metrics") or expected.get("metrics") or {}),
        "allowed_manual_review_items": list(review.get("allowed_manual_review_items") or batch.get("allowed_manual_review_items") or []),
        "review_notes": str(review.get("review_notes") or expected.get("review_notes") or ""),
    }
    batch["expected_result"] = updated_expected
    batch["allowed_manual_review_items"] = list(updated_expected["allowed_manual_review_items"])


def _manifest_batch_keys(directory: Path) -> set[str]:
    keys: set[str] = set()
    for manifest_path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        for batch in manifest.get("batches") or []:
            if isinstance(batch, dict) and batch.get("batch_key"):
                keys.add(str(batch["batch_key"]))
    return keys


def _manifest_supplier_refs(directory: Path) -> set[str]:
    refs: set[str] = set()
    for manifest_path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        for batch in manifest.get("batches") or []:
            if isinstance(batch, dict) and batch.get("supplier_ref"):
                refs.add(str(batch["supplier_ref"]))
    return refs


def _review_batch_refs_by_batch(directory: Path) -> dict[str, str]:
    refs: dict[str, str] = {}
    for manifest_path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        for batch in manifest.get("batches") or []:
            if not isinstance(batch, dict):
                continue
            batch_key = str(batch.get("batch_key") or "").strip()
            if not batch_key:
                continue
            refs[_review_batch_ref(batch)] = batch_key
    return refs


def _approved_batch_count(directory: Path) -> int:
    count = 0
    for manifest_path in sorted(directory.glob("*_manifest.json")):
        manifest = _load_json(manifest_path)
        for batch in manifest.get("batches") or []:
            if not isinstance(batch, dict):
                continue
            expected = batch.get("expected_result") if isinstance(batch.get("expected_result"), dict) else {}
            if expected.get("review_status") == "approved":
                count += 1
    return count


def _supplier_candidate_plan(root: Path, supplier: str, candidates: list[dict[str, Any]]) -> dict[str, Any]:
    selected = _select_candidate_batch(candidates)
    batch_key = str(selected.get("batchKey") or "")
    invoice_count = int(selected.get("invoicePdfCount") or len(selected.get("invoiceFiles") or []))
    workbook_count = int(selected.get("workbookCount") or len(selected.get("workbookFiles") or []))
    return {
        "supplier": supplier,
        "supplier_ref": supplier_ref(supplier),
        "candidate_batch_count": len(candidates),
        "selected_batch_key": batch_key,
        "selected_directory": str(selected.get("directory") or ""),
        "period_hint": str(selected.get("periodHint") or ""),
        "warehouse_count": len(selected.get("warehouseIds") or []),
        "invoice_pdf_count": invoice_count,
        "workbook_count": workbook_count,
        "review_status": "needs_business_review",
        "selection_reason": _selection_reason(selected),
        "manifest_command": _manifest_command(root, batch_key),
        "validate_command": _validate_command(root, batch_key),
        "required_business_fields": sorted(REQUIRED_APPROVED_METRICS),
    }


def _missing_supplier_plan(root: Path, supplier: str) -> dict[str, Any]:
    return {
        "supplier": supplier,
        "supplier_ref": supplier_ref(supplier),
        "candidate_batch_count": 0,
        "selected_batch_key": "",
        "selected_directory": "",
        "period_hint": "",
        "warehouse_count": 0,
        "invoice_pdf_count": 0,
        "workbook_count": 0,
        "review_status": "missing_candidate_batch",
        "selection_reason": "No candidate batch matched this supplier in the local material index.",
        "manifest_command": "",
        "validate_command": "",
        "required_business_fields": sorted(REQUIRED_APPROVED_METRICS),
    }


def _select_candidate_batch(candidates: list[dict[str, Any]]) -> dict[str, Any]:
    def score(batch: dict[str, Any]) -> tuple[int, int, int, str]:
        invoice_count = int(batch.get("invoicePdfCount") or len(batch.get("invoiceFiles") or []))
        workbook_count = int(batch.get("workbookCount") or len(batch.get("workbookFiles") or []))
        warehouse_count = len(batch.get("warehouseIds") or [])
        ready_score = 1 if invoice_count > 0 and workbook_count > 0 else 0
        return (ready_score, min(invoice_count, 5), min(workbook_count, 3), str(batch.get("batchKey") or ""))

    return sorted(candidates, key=score, reverse=True)[0]


def _selection_reason(batch: dict[str, Any]) -> str:
    invoice_count = int(batch.get("invoicePdfCount") or len(batch.get("invoiceFiles") or []))
    workbook_count = int(batch.get("workbookCount") or len(batch.get("workbookFiles") or []))
    if invoice_count and workbook_count:
        return "Selected because it has both invoice PDF and workbook evidence."
    if invoice_count:
        return "Selected as the strongest available invoice-only candidate; workbook evidence still needs review."
    if workbook_count:
        return "Selected as the strongest available workbook-only candidate; invoice evidence still needs review."
    return "Selected from material index, but file evidence is incomplete."


def _candidate_matches_supplier(batch: dict[str, Any], supplier: str) -> bool:
    values = [
        str(batch.get("supplier") or ""),
        str(batch.get("batchKey") or ""),
        str(batch.get("directory") or ""),
    ]
    normalized = [_normalize_supplier_key(value) for value in values]
    return supplier in normalized or any(value.startswith(supplier) or supplier in value for value in normalized)


def _normalize_supplier_key(value: str) -> str:
    return "".join(char for char in str(value or "").strip().lower() if char.isalnum())


def _expected_result_template() -> dict[str, Any]:
    return {
        "review_status": "needs_business_review",
        "reviewer": "",
        "reviewed_at": "",
        "evidence_reference": "",
        "metrics": {
            "invoice_total": None,
            "excel_total": None,
            "warehouse_count": None,
            "employee_count": None,
            "total_hours": None,
            "difference_category_counts": {},
            "manual_review_count": None,
            "core_error_types": [],
        },
        "allowed_manual_review_items": [],
        "review_notes": "",
    }


def _manifest_command(root: Path, batch_key: str) -> str:
    return (
        "PYTHONDONTWRITEBYTECODE=1 python3 -m bonus_platform.engine.labor.golden discover "
        f"--materials-root {json.dumps(str(root), ensure_ascii=False)} "
        f"--batch-key {json.dumps(batch_key, ensure_ascii=False)} "
        f"--output /tmp/labor_golden/{_manifest_filename(batch_key)}"
    )


def _validate_command(root: Path, batch_key: str) -> str:
    return (
        "PYTHONDONTWRITEBYTECODE=1 python3 -m bonus_platform.engine.labor.golden validate "
        f"--manifest /tmp/labor_golden/{_manifest_filename(batch_key)} "
        f"--materials-root {json.dumps(str(root), ensure_ascii=False)} "
        "--require-approved"
    )


def _manifest_filename(batch_key: str) -> str:
    safe_batch = "".join(char if char.isalnum() or char in "_-" else "_" for char in batch_key).strip("_") or "batch"
    return f"{safe_batch}_manifest.json"


def _summarize_file_types(files: list[dict[str, Any]]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for record in files:
        if not isinstance(record, dict):
            continue
        file_type = str(record.get("file_type") or "supporting")
        summary[file_type] = summary.get(file_type, 0) + 1
    return summary


def _review_file_hashes(files: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for record in files:
        if not isinstance(record, dict):
            continue
        rows.append(
            {
                "file_type": str(record.get("file_type") or ""),
                "sha256_prefix": str(record.get("sha256") or "")[:12],
            }
        )
    return rows


def _review_excel(excel: dict[str, Any]) -> dict[str, Any]:
    return {
        "sheet": str(excel.get("sheet") or ""),
        "mapping_version": str(excel.get("mapping_version") or ""),
        "suggested_mapping": dict(excel.get("suggested_mapping") or {}),
    }


def _validate_approved_metrics(batch_key: str, expected: dict[str, Any], errors: list[dict[str, Any]]) -> None:
    metrics = expected.get("metrics") if isinstance(expected.get("metrics"), dict) else {}
    if not metrics:
        errors.append(
            {
                "code": "missing_expected_metrics",
                "batch_key": batch_key,
                "message": "approved expected_result requires metrics",
            }
        )
        return
    for field in sorted(REQUIRED_APPROVED_METRICS):
        if field not in metrics or metrics.get(field) is None:
            errors.append({"code": "approved_expected_metric_missing", "batch_key": batch_key, "field": field})
            continue
        value = metrics.get(field)
        if field in {"invoice_total", "excel_total", "total_hours"} and not _is_plain_number(value):
            errors.append({"code": "approved_expected_metric_invalid", "batch_key": batch_key, "field": field})
        elif field in {"warehouse_count", "employee_count", "manual_review_count"} and not _is_plain_int(value):
            errors.append({"code": "approved_expected_metric_invalid", "batch_key": batch_key, "field": field})
        elif field == "difference_category_counts" and not isinstance(value, dict):
            errors.append({"code": "approved_expected_metric_invalid", "batch_key": batch_key, "field": field})
        elif field == "core_error_types" and not isinstance(value, list):
            errors.append({"code": "approved_expected_metric_invalid", "batch_key": batch_key, "field": field})


def _is_plain_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_plain_int(value: Any) -> bool:
    return isinstance(value, int) and not isinstance(value, bool)


def _is_sha256(value: str) -> bool:
    return len(value) == 64 and all(char in "0123456789abcdefABCDEF" for char in value)


def _build_manifest_batch(root: Path, batch: dict[str, Any]) -> dict[str, Any]:
    batch_key = str(batch.get("batchKey") or "")
    files = []
    for record in [*(batch.get("invoiceFiles") or []), *(batch.get("workbookFiles") or [])]:
        relative_path = str(record.get("relativePath") or "")
        if not relative_path:
            continue
        path = root / relative_path
        files.append(
            {
                "relative_path": relative_path,
                "file_type": _file_type(record),
                "size_bytes": int(record.get("sizeBytes") or (path.stat().st_size if path.exists() else 0)),
                "sha256": sha256_file(path) if path.exists() else "",
            }
        )
    plan = _safe_replay_plan(root, batch_key)
    mapping = _first_mapping(plan)
    return {
        "batch_key": batch_key,
        "supplier_ref": supplier_ref(str(batch.get("supplier") or "")),
        "period_hint": str(batch.get("periodHint") or mapping.get("periodHint") or ""),
        "warehouse_count": len(batch.get("warehouseIds") or []),
        "files": files,
        "excel": {
            "sheet": mapping.get("sheetName") or "",
            "mapping_version": "suggested-v1",
            "suggested_mapping": mapping.get("suggestedMapping") or {},
        },
        "expected_result": {
            "review_status": "needs_business_review",
            "metrics": {
                "invoice_total": None,
                "excel_total": None,
                "warehouse_count": None,
                "employee_count": None,
                "total_hours": None,
                "difference_category_counts": None,
                "manual_review_count": None,
                "core_error_types": [],
            },
        },
        "allowed_manual_review_items": [],
        "notes": "Discovered locally. Expected result must be filled by reviewed business evidence.",
    }


def _file_type(record: dict[str, Any]) -> str:
    category = str(record.get("category") or "")
    if category == "invoice_pdf":
        return "invoice_pdf"
    if category == "workbook_bill":
        return "workbook"
    suffix = Path(str(record.get("relativePath") or "")).suffix.lower()
    if suffix == ".pdf":
        return "invoice_pdf"
    if suffix in {".xls", ".xlsx", ".xlsm"}:
        return "workbook"
    return "supporting"


def _safe_replay_plan(root: Path, batch_key: str) -> dict[str, Any]:
    if not batch_key:
        return {}
    try:
        return build_material_replay_plan(root, batch_key=batch_key)
    except Exception as exc:  # noqa: BLE001 - discovery should keep file evidence even when workbook parsing fails.
        return {"error": str(exc), "plans": []}


def _first_mapping(plan: dict[str, Any]) -> dict[str, Any]:
    plans = plan.get("plans") if isinstance(plan.get("plans"), list) else []
    if not plans:
        return {}
    mappings = plans[0].get("mappingCandidates") if isinstance(plans[0], dict) else []
    if not mappings:
        return {}
    return mappings[0] if isinstance(mappings[0], dict) else {}


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_review_template(review_template: str | Path | dict[str, Any]) -> dict[str, Any]:
    if isinstance(review_template, dict):
        return review_template
    path = Path(review_template).expanduser()
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_review_template_workbook(path)
    return _load_json(path)


def _load_review_template_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    if "业务确认表" not in workbook.sheetnames:
        return {"schema_version": SCHEMA_VERSION, "review_items": []}
    sheet = workbook["业务确认表"]
    header_row = [str(cell.value or "").strip() for cell in sheet[1]]
    headers = {value: index for index, value in enumerate(header_row) if value}
    items: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        get = lambda key: row[headers[key]] if key in headers and headers[key] < len(row) else None
        review_batch_ref = str(get("批次编号") or "").strip()
        if not review_batch_ref:
            continue
        metrics = {
            "invoice_total": _none_if_blank(get("发票总额")),
            "excel_total": _none_if_blank(get("账单总额")),
            "warehouse_count": _none_if_blank(get("仓库数")),
            "employee_count": _none_if_blank(get("员工人数")),
            "total_hours": _none_if_blank(get("总工时")),
            "difference_category_counts": _parse_difference_category_counts(get("差异分类数量")),
            "manual_review_count": _none_if_blank(get("待确认数量")),
            "core_error_types": _parse_list_cell(get("核心异常")),
        }
        items.append(
            {
                "review_batch_ref": review_batch_ref,
                "period_hint": str(get("账期") or "").strip(),
                "review_status": str(get("业务确认状态") or "needs_business_review").strip(),
                "expected_metrics": metrics,
                "reviewer": str(get("复核人") or "").strip(),
                "reviewed_at": str(get("复核时间") or "").strip(),
                "evidence_reference": str(get("依据说明") or "").strip(),
                "review_notes": str(get("备注") or "").strip(),
            }
        )
    return {
        "schema_version": SCHEMA_VERSION,
        "source": "golden_business_review_workbook",
        "review_items": items,
    }


def _none_if_blank(value: Any) -> Any:
    return None if value == "" else value


def _parse_difference_category_counts(value: Any) -> dict[str, int]:
    if value in (None, ""):
        return {}
    if isinstance(value, dict):
        return value
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        return {}
    if not isinstance(parsed, dict):
        return {}
    return {str(key): int(_metric_number(count)) for key, count in parsed.items()}


def _parse_list_cell(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    raw = str(value).replace("，", ",").strip()
    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            parsed = []
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
    return [item.strip() for item in raw.split(",") if item.strip()]


def _manifest_dir_result(
    directory: Path,
    manifests: list[dict[str, Any]],
    top_level_errors: list[dict[str, Any]],
    require_approved: bool,
) -> dict[str, Any]:
    error_count = len(top_level_errors) + sum(len(item.get("errors") or []) for item in manifests)
    warning_count = sum(len(item.get("warnings") or []) for item in manifests)
    return {
        "ok": error_count == 0,
        "manifest_dir": str(directory),
        "summary": {
            "manifest_count": len(manifests),
            "batch_count": sum(int(item.get("batch_count") or 0) for item in manifests),
            "file_count": sum(int(item.get("file_count") or 0) for item in manifests),
            "error_count": error_count,
            "warning_count": warning_count,
        },
        "errors": top_level_errors,
        "manifests": manifests,
        "require_approved": require_approved,
    }


def _privacy_scan_result(directory: Path, scanned_files: list[str], issues: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "source": "golden_handoff_privacy_scan",
        "handoff_dir": str(directory),
        "ok": not issues,
        "summary": {
            "scanned_file_count": len(scanned_files),
            "issue_count": len(issues),
        },
        "scanned_files": scanned_files,
        "issues": issues,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Local read-only labor golden manifest tooling")
    subparsers = parser.add_subparsers(dest="command", required=True)

    discover = subparsers.add_parser("discover", help="Discover candidate batches and print or write a manifest")
    discover.add_argument("--materials-root", required=True)
    discover.add_argument("--batch-key", default="")
    discover.add_argument("--supplier", default="")
    discover.add_argument("--output", default="")

    plan = subparsers.add_parser("plan", help="Build a supplier coverage plan and business review template")
    plan.add_argument("--materials-root", required=True)
    plan.add_argument("--required-suppliers", nargs="*", default=list(DEFAULT_GOLDEN_SUPPLIERS))
    plan.add_argument("--output", default="")

    prepare = subparsers.add_parser("prepare", help="Write selected supplier manifests into a local output directory")
    prepare.add_argument("--materials-root", required=True)
    prepare.add_argument("--output-dir", required=True)
    prepare.add_argument("--required-suppliers", nargs="*", default=list(DEFAULT_GOLDEN_SUPPLIERS))
    prepare.add_argument("--output", default="")

    validate = subparsers.add_parser("validate", help="Validate a golden manifest and file hashes")
    validate.add_argument("--manifest", required=True)
    validate.add_argument("--materials-root", default="")
    validate.add_argument("--output", default="")
    validate.add_argument("--require-approved", action="store_true")

    validate_dir = subparsers.add_parser("validate-dir", help="Validate all *_manifest.json files in a directory")
    validate_dir.add_argument("--manifest-dir", required=True)
    validate_dir.add_argument("--materials-root", default="")
    validate_dir.add_argument("--output", default="")
    validate_dir.add_argument("--require-approved", action="store_true")

    review_template = subparsers.add_parser("review-template", help="Build a business review template from prepared manifests")
    review_template.add_argument("--manifest-dir", required=True)
    review_template.add_argument("--materials-root", default="")
    review_template.add_argument("--output", required=True)
    review_template.add_argument("--batch-key", default="")
    review_template.add_argument("--supplier-ref", default="")

    handoff = subparsers.add_parser("handoff", help="Build a redacted business review handoff package")
    handoff.add_argument("--manifest-dir", required=True)
    handoff.add_argument("--output-dir", required=True)
    handoff.add_argument("--materials-root", default="")
    handoff.add_argument("--output", default="")
    handoff.add_argument("--batch-key", default="")
    handoff.add_argument("--supplier-ref", default="")

    scan_handoff = subparsers.add_parser("scan-handoff", help="Scan a business review handoff package for privacy leaks")
    scan_handoff.add_argument("--handoff-dir", required=True)
    scan_handoff.add_argument("--materials-root", default="")
    scan_handoff.add_argument("--forbidden-term", action="append", default=[])
    scan_handoff.add_argument("--output", default="")

    validate_review = subparsers.add_parser("validate-review", help="Validate a completed business review template before apply-review")
    validate_review.add_argument("--review-template", required=True)
    validate_review.add_argument("--output", default="")
    validate_review.add_argument("--require-approved", action="store_true")
    validate_review.add_argument("--batch-key", default="")
    validate_review.add_argument("--supplier-ref", default="")
    validate_review.add_argument("--review-batch-ref", default="")

    apply_review = subparsers.add_parser("apply-review", help="Apply a completed business review template into copied manifests")
    apply_review.add_argument("--manifest-dir", required=True)
    apply_review.add_argument("--review-template", required=True)
    apply_review.add_argument("--output-dir", required=True)
    apply_review.add_argument("--materials-root", default="")
    apply_review.add_argument("--output", default="")
    apply_review.add_argument("--require-approved", action="store_true")
    apply_review.add_argument("--batch-key", default="")
    apply_review.add_argument("--supplier-ref", default="")
    apply_review.add_argument("--review-batch-ref", default="")

    replay = subparsers.add_parser("replay", help="Validate approved golden manifests and write a deterministic replay summary")
    replay.add_argument("--manifest-dir", required=True)
    replay.add_argument("--output-dir", required=True)
    replay.add_argument("--materials-root", default="")
    replay.add_argument("--output", default="")
    replay.add_argument("--batch-key", default="")
    replay.add_argument("--supplier-ref", default="")

    business_replay = subparsers.add_parser("business-replay", help="Run the business-readable real-material replay gate")
    business_replay.add_argument("--manifest-dir", required=True)
    business_replay.add_argument("--output-dir", default="outputs/labor_golden/business_replay_latest")
    business_replay.add_argument("--materials-root", default="")
    business_replay.add_argument("--output", default="")
    business_replay.add_argument("--batch-key", default="")
    business_replay.add_argument("--supplier-ref", default="")

    args = parser.parse_args(argv)
    if args.command == "discover":
        payload = discover_golden_batches(args.materials_root, batch_key=args.batch_key, supplier=args.supplier)
        return _write_or_print(payload, args.output)
    if args.command == "plan":
        payload = build_golden_candidate_plan(args.materials_root, required_suppliers=args.required_suppliers)
        return _write_or_print(payload, args.output)
    if args.command == "prepare":
        payload = prepare_golden_manifests(
            args.materials_root,
            args.output_dir,
            required_suppliers=args.required_suppliers,
        )
        return _write_or_print(payload, args.output)
    if args.command == "validate":
        manifest = _load_json(Path(args.manifest))
        payload = validate_golden_manifest(
            manifest,
            materials_root=args.materials_root or None,
            require_approved=args.require_approved,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "validate-dir":
        payload = validate_golden_manifest_dir(
            args.manifest_dir,
            materials_root=args.materials_root or None,
            require_approved=args.require_approved,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "review-template":
        payload = build_golden_review_template(
            args.manifest_dir,
            materials_root=args.materials_root or None,
            batch_key=args.batch_key or None,
            supplier_ref=args.supplier_ref or None,
        )
        return _write_or_print(payload, args.output)
    if args.command == "handoff":
        payload = build_golden_review_handoff(
            args.manifest_dir,
            args.output_dir,
            materials_root=args.materials_root or None,
            batch_key=args.batch_key or None,
            supplier_ref=args.supplier_ref or None,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "scan-handoff":
        payload = scan_golden_handoff_privacy(
            args.handoff_dir,
            materials_root=args.materials_root or None,
            forbidden_terms=args.forbidden_term,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "validate-review":
        payload = validate_golden_review_template(
            args.review_template,
            require_approved=args.require_approved,
            batch_key=args.batch_key or None,
            supplier_ref=args.supplier_ref or None,
            review_batch_ref=args.review_batch_ref or None,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "apply-review":
        payload = apply_golden_review_template(
            args.manifest_dir,
            args.review_template,
            args.output_dir,
            materials_root=args.materials_root or None,
            require_approved=args.require_approved,
            batch_key=args.batch_key or None,
            supplier_ref=args.supplier_ref or None,
            review_batch_ref=args.review_batch_ref or None,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "replay":
        payload = run_golden_manifest_replay(
            args.manifest_dir,
            args.output_dir,
            materials_root=args.materials_root or None,
            batch_key=args.batch_key or None,
            supplier_ref=args.supplier_ref or None,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    if args.command == "business-replay":
        payload = run_golden_manifest_replay(
            args.manifest_dir,
            args.output_dir,
            materials_root=args.materials_root or None,
            batch_key=args.batch_key or None,
            supplier_ref=args.supplier_ref or None,
        )
        _write_or_print(payload, args.output)
        return 0 if payload["ok"] else 1
    return 2


def _write_or_print(payload: dict[str, Any], output: str = "") -> int:
    text = json.dumps(payload, ensure_ascii=False, indent=2)
    if output:
        path = Path(output)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text + "\n", encoding="utf-8")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
