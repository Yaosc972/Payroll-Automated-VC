from __future__ import annotations

import re
import unicodedata
from html import escape
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import LaborLineItem


REPORT_SHEETS = ["核对结论", "识别完整度", "核对摘要", "全员对账明细", "仓库金额汇总", "金额差异员工", "工时待确认", "不在本批发票", "姓名格式差异", "明细识别待确认", "PDF发票明细", "Excel账单明细", "上传字段对应关系"]


def build_labor_report(
    output_path: Path,
    comparison: Dict[str, Any],
    pdf_rows: List[LaborLineItem],
    excel_rows: List[LaborLineItem],
    mapping: Dict[str, str],
    warehouse_comparison: Dict[str, Any] | None = None,
    extraction_quality: Dict[str, Any] | None = None,
    reconciliation_diagnostics: Dict[str, Any] | None = None,
    ai_cache_audit: Dict[str, Any] | None = None,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    # 写入核对结论（第一个sheet）
    _write_conclusion(workbook, comparison.get("summary", {}), warehouse_comparison)

    # 写入质量评分（第二个sheet）
    if extraction_quality:
        _write_quality(workbook, extraction_quality)

    _write_summary(workbook, comparison.get("summary", {}))
    rows = comparison.get("rows", [])
    _write_reconciliation_detail(workbook, rows)
    if warehouse_comparison:
        _write_warehouse_summary(workbook, warehouse_comparison)
    if reconciliation_diagnostics:
        _write_reconciliation_diagnostics(workbook, reconciliation_diagnostics)
    if ai_cache_audit:
        _write_ai_cache_audit(workbook, ai_cache_audit)
    _write_rows(workbook, "金额差异员工", _filter(rows, "金额差异"))
    _write_rows(workbook, "工时待确认", [row for row in rows if row.get("matchStatus") == "工时不一致" or "工时需复核" in row.get("riskFlags", [])])
    _write_rows(workbook, "不在本批发票", [row for row in rows if row.get("matchStatus") in {"PDF有Excel无", "Excel有PDF无", "疑似姓名匹配"}])
    _write_candidate_matches(workbook, comparison.get("candidateMatches", []))
    _write_rows(workbook, "明细识别待确认", [row for row in rows if row.get("matchStatus") == "低置信度抽取" or "低置信度抽取" in row.get("riskFlags", [])])
    _write_detail(workbook, "PDF发票明细", pdf_rows)
    _write_detail(workbook, "Excel账单明细", excel_rows)
    _write_mapping(workbook, mapping)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_labor_business_html_report(
    output_path: Path,
    comparison: Dict[str, Any],
    *,
    supplier_name: str,
    period_start: str,
    period_end: str,
    invoice_scope: str = "",
    warehouse_comparison: Dict[str, Any] | None = None,
    excel_record_count: int | None = None,
) -> None:
    summary = comparison.get("summary", {}) or {}
    warehouse_summary = (warehouse_comparison or {}).get("summary", {}) or {}
    batch_summary = warehouse_summary or summary
    rows = comparison.get("rows", []) or []
    conclusion = _business_conclusion(summary, rows)
    warehouse_needs_review = bool(
        warehouse_summary
        and (
            warehouse_summary.get("totalPassed") is False
            or _num(warehouse_summary.get("exceptionCount")) > 0
            or _num(warehouse_summary.get("allocationIssueCount")) > 0
        )
    )
    if warehouse_needs_review and conclusion == "总账通过":
        conclusion = "总账通过，但员工明细待确认"
    amount_delta = _num(batch_summary.get("amountDeltaTotal"))
    pdf_total = _num(batch_summary.get("pdfAmountTotal"))
    excel_total = _num(batch_summary.get("excelAmountTotal"))
    detail_pdf_total = _num(summary.get("pdfAmountTotal"))
    detail_excel_total = _num(summary.get("excelAmountTotal"))
    pdf_employee_count = int(_num(summary.get("pdfEmployeeCount")))
    excel_employee_count = int(_num(summary.get("excelEmployeeCount")))
    matched_count = int(_num(summary.get("passedCount")))
    diff_count = max(
        int(_num(summary.get("amountDiffCount"))),
        sum(1 for row in rows if _business_row_status(row) != "一致"),
    )
    diff_warehouses = [str(item) for item in (warehouse_summary.get("diffWarehouses") or []) if str(item)]
    warehouse_label = f"仓库 {'、'.join(diff_warehouses)}" if diff_warehouses else "需要确认的仓库"
    displayed_employee_count = excel_employee_count or len(rows)
    detail_rows_incomplete = not rows and bool(pdf_employee_count or excel_employee_count or pdf_total or excel_total)
    full_excel_scope_note = (
        f"整批账单已读取 {excel_record_count} 行，当前展示的是需要确认的 {displayed_employee_count} 名员工明细，不代表账单只有这些员工。"
        if excel_record_count and excel_record_count > displayed_employee_count
        else ""
    )
    detail_scope_title = "待确认员工明细" if warehouse_needs_review else "员工对账明细"
    detail_scope_note = (
        f"{full_excel_scope_note or '只展示需要确认的仓库员工明细，不代表账单只有这些员工。'}当前范围：{warehouse_label}，账单金额 {_money(detail_excel_total)}。"
        if warehouse_needs_review
        else "员工明细用于确认每位员工的发票金额和账单金额是否一致。"
    )
    auto_fix_html = _business_auto_fix_section(rows)
    amount_layers_html = _business_amount_layers_section(
        pdf_total=pdf_total,
        excel_total=excel_total,
        amount_delta=amount_delta,
        detail_pdf_total=detail_pdf_total,
        detail_excel_total=detail_excel_total,
    )
    employee_recognition_html = _business_employee_recognition_section(
        pdf_employee_count=pdf_employee_count,
        excel_employee_count=excel_employee_count,
        displayed_employee_count=displayed_employee_count,
        excel_record_count=excel_record_count,
        details_incomplete=detail_rows_incomplete,
        warehouse_needs_review=warehouse_needs_review,
    )
    suspected_match_html = _business_suspected_match_section(comparison.get("candidateMatches", []) or [])
    pending_exception_html = _business_pending_exception_section(rows)
    download_section_html = _business_download_section()
    decision_panel_html = _business_decision_panel(
        conclusion=conclusion,
        amount_delta=amount_delta,
        diff_count=diff_count,
        details_incomplete=detail_rows_incomplete,
        warehouse_needs_review=warehouse_needs_review,
    )

    row_html = "\n".join(_business_detail_row(index, row) for index, row in enumerate(rows, start=1))
    if not row_html:
        empty_detail_text = _business_empty_detail_text(
            conclusion=conclusion,
            detail_rows_incomplete=detail_rows_incomplete,
            amount_delta=amount_delta,
        )
        row_html = f'<tr><td colspan="11" class="empty">{escape(empty_detail_text)}</td></tr>'

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{escape(supplier_name)} 账单 vs 发票核对报告</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; background: #f5f6fa; color: #222; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif; font-size: 13px; }}
    .container {{ max-width: 1280px; margin: 0 auto; padding: 24px 16px; }}
    h1 {{ margin: 0 0 4px; font-size: 22px; }}
    .subtitle {{ color: #666; margin-bottom: 20px; }}
    .cards {{ display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 18px; }}
    .card {{ min-width: 170px; flex: 1; background: #fff; border-radius: 10px; border-left: 4px solid #2563eb; padding: 16px 18px; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
    .card.warn {{ border-left-color: #dc2626; }}
    .card.ok {{ border-left-color: #16a34a; }}
    .lbl {{ color: #777; font-size: 11px; margin-bottom: 6px; text-transform: uppercase; }}
    .val {{ font-size: 21px; font-weight: 700; font-variant-numeric: tabular-nums; }}
    .sub {{ color: #888; font-size: 11px; margin-top: 3px; }}
    .banner {{ margin: 0 0 16px; border-radius: 8px; padding: 12px 16px; font-weight: 650; background: {_conclusion_bg(conclusion)}; border: 1px solid {_conclusion_border(conclusion)}; color: {_conclusion_color(conclusion)}; }}
    .section {{ background: #fff; border-radius: 10px; box-shadow: 0 1px 4px rgba(0,0,0,.08); overflow: hidden; }}
    .section + .section {{ margin-top: 14px; }}
    .section-header {{ padding: 14px 18px; border-bottom: 1px solid #f0f0f0; }}
    .section-header h2 {{ margin: 0; font-size: 15px; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th {{ background: #f8f9fb; color: #666; padding: 9px 12px; text-align: left; white-space: nowrap; font-size: 11px; }}
    td {{ padding: 8px 12px; border-bottom: 1px solid #f5f5f5; vertical-align: top; }}
    .num {{ font-variant-numeric: tabular-nums; white-space: nowrap; }}
    .tag {{ display: inline-block; border-radius: 10px; padding: 1px 7px; font-size: 11px; font-weight: 600; }}
    .tag-ok {{ background: #dcfce7; color: #166534; }}
    .tag-warn {{ background: #fef3c7; color: #78350f; }}
    .tag-diff {{ background: #fee2e2; color: #991b1b; }}
    .empty {{ color: #777; text-align: center; padding: 24px; }}
    .business-list {{ display: grid; gap: 10px; padding: 14px 18px; }}
    .business-item {{ border: 1px solid #edf0f5; border-radius: 8px; padding: 12px 14px; background: #fcfdff; }}
    .business-item strong {{ display: block; margin-bottom: 4px; }}
    .business-item p {{ margin: 0; color: #5b667a; line-height: 1.6; }}
    .decision-panel {{ display: grid; grid-template-columns: minmax(260px, 1.15fr) repeat(3, minmax(160px, 1fr)); gap: 12px; margin: 0 0 16px; }}
    .decision-box {{ background: #fff; border-radius: 10px; border: 1px solid #e8edf5; padding: 14px 16px; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
    .decision-box.primary {{ border-left: 5px solid {_conclusion_border(conclusion)}; }}
    .decision-question {{ color: #6b7280; font-size: 12px; margin-bottom: 6px; }}
    .decision-answer {{ font-size: 20px; font-weight: 750; color: {_conclusion_color(conclusion)}; }}
    .decision-text {{ margin: 8px 0 0; color: #5b667a; line-height: 1.6; }}
    .download-note {{ padding: 14px 18px; color: #5b667a; line-height: 1.7; }}
  </style>
</head>
<body>
  <main class="container">
    <h1>{escape(supplier_name)} — 账单 vs 发票核对报告</h1>
    <div class="subtitle">供应商：{escape(supplier_name)} &nbsp;|&nbsp; 核算周期：{escape(period_start)} ~ {escape(period_end)} &nbsp;|&nbsp; 发票编号或文件范围：{escape(invoice_scope or "未提供")}</div>
    <div class="cards">
      <div class="card {'ok' if _conclusion_is_total_pass(conclusion) else 'warn'}"><div class="lbl">核对结论</div><div class="val">{escape(conclusion)}</div><div class="sub">业务结论优先展示</div></div>
      <div class="card"><div class="lbl">整批 PDF 发票总金额</div><div class="val">{_money(pdf_total)}</div><div class="sub">来自本次上传的全部发票</div></div>
      <div class="card"><div class="lbl">整批账单总金额</div><div class="val">{_money(excel_total)}</div><div class="sub">来自本次上传的全部账单</div></div>
      <div class="card {'warn' if abs(amount_delta) > 0.1 else 'ok'}"><div class="lbl">总差额</div><div class="val">{_signed_money(amount_delta)}</div><div class="sub">PDF - 账单</div></div>
      <div class="card ok"><div class="lbl">一致员工数</div><div class="val">{matched_count}</div><div class="sub">自动核对一致</div></div>
      <div class="card {'warn' if diff_count else 'ok'}"><div class="lbl">待确认员工数</div><div class="val">{diff_count}</div><div class="sub">请按下方业务说明处理</div></div>
    </div>
    <div class="banner">{escape(_business_conclusion_message(conclusion, amount_delta, matched_count, diff_count, detail_scope_note, details_incomplete=detail_rows_incomplete))}</div>
    {decision_panel_html}
    {amount_layers_html}
    {employee_recognition_html}
    {auto_fix_html}
    {suspected_match_html}
    {pending_exception_html}
    <section class="section">
      <div class="section-header"><h2>{escape(detail_scope_title)}</h2><div class="sub">{escape(detail_scope_note)}</div></div>
      <div style="overflow-x:auto">
        <table>
          <thead>
            <tr>
              <th>#</th>
              <th>发票编号</th>
              <th>员工姓名（发票）</th>
              <th>账单姓名</th>
              <th>REG 工时</th>
              <th>OT 工时</th>
              <th>发票金额</th>
              <th>账单金额</th>
              <th>差额</th>
              <th>状态</th>
              <th>业务说明</th>
            </tr>
          </thead>
          <tbody>{row_html}</tbody>
        </table>
      </div>
    </section>
    {download_section_html}
  </main>
</body>
</html>
"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(html, encoding="utf-8")


def build_labor_projection_report(output_path: Path, preview: Dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_projection_summary(workbook, preview)
    _write_projection_affected_rows(workbook, preview.get("affectedRows", []) or [])
    _write_projection_corrections(workbook, preview.get("appliedCorrections", []) or [])
    _write_projection_manual_review(workbook, preview.get("manualReview", []) or [])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_labor_governance_report(output_path: Path, metadata: Dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_governance_overview(workbook, metadata)
    _write_rule_governance(workbook, metadata.get("ruleGovernance", {}) or {})
    _write_name_mapping_governance(workbook, metadata.get("nameMappingGovernance", {}) or {})
    _write_profile_governance(workbook, metadata.get("profileGovernance", {}) or {})
    _write_correction_governance(workbook, metadata.get("correctionGovernance", {}) or {})
    _write_reocr_governance(workbook, metadata.get("reocrReplayGovernance", {}) or {})
    _write_reocr_upload_coverage(workbook, ((metadata.get("files") or {}).get("reocrCandidateFiles") or []))
    _write_governance_ai_evidence(workbook, metadata.get("aiCacheAudit", {}) or {})
    _write_governance_audit_trail(workbook, metadata)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _write_governance_overview(workbook: Workbook, metadata: Dict[str, Any]) -> None:
    rule = metadata.get("ruleGovernance", {}) or {}
    name_mapping = metadata.get("nameMappingGovernance", {}) or {}
    profile = metadata.get("profileGovernance", {}) or {}
    correction = metadata.get("correctionGovernance", {}) or {}
    reocr = metadata.get("reocrReplayGovernance", {}) or {}
    ai_cache = metadata.get("aiCacheAudit", {}) or {}
    sheet = workbook.create_sheet("治理总览")
    sheet.append(["项目", "值"])
    sheet.append(["批次ID", metadata.get("id", "")])
    sheet.append(["供应商", metadata.get("supplierName", "")])
    sheet.append(["账期", f"{metadata.get('periodStart', '')} ~ {metadata.get('periodEnd', '')}"])
    sheet.append(["当前状态", metadata.get("status", "")])
    sheet.append(["规则候选", len(rule.get("candidates") or [])])
    sheet.append(["已确认规则", len(rule.get("activeRules") or [])])
    sheet.append(["已回滚规则", len(rule.get("rolledBackRules") or [])])
    sheet.append(["姓名映射候选", len(name_mapping.get("candidates") or [])])
    sheet.append(["已确认姓名映射", len(name_mapping.get("activeMappings") or [])])
    sheet.append(["已回滚姓名映射", len(name_mapping.get("rolledBackMappings") or [])])
    sheet.append(["Profile候选", len(profile.get("candidates") or [])])
    sheet.append(["已确认Profile", len(profile.get("activeProfiles") or [])])
    sheet.append(["已回滚Profile", len(profile.get("rolledBackProfiles") or [])])
    sheet.append(["修正候选", len(correction.get("candidates") or [])])
    sheet.append(["已确认修正", len(correction.get("activeCorrections") or [])])
    sheet.append(["已回滚修正", len(correction.get("rolledBackCorrections") or [])])
    sheet.append(["图片识别预览", len(reocr.get("replays") or [])])
    sheet.append(["已确认图片识别结果", len(reocr.get("activeCandidates") or [])])
    sheet.append(["已回滚图片识别结果", len(reocr.get("rolledBackCandidates") or [])])
    sheet.append(["AI候选证据决策", ai_cache.get("decision", "")])
    sheet.append(["AI候选文件数", (ai_cache.get("summary") or {}).get("candidateFileCount", 0)])
    _format(sheet)


def _write_rule_governance(workbook: Workbook, governance: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("规则治理")
    sheet.append(["类型", "规则ID", "标题", "状态", "决策", "版本", "回放决策", "回放摘要", "说明"])
    replay = governance.get("replaySummaries") or {}
    for kind, rows in (
        ("候选", governance.get("candidates") or []),
        ("已确认", governance.get("activeRules") or []),
        ("已回滚", governance.get("rolledBackRules") or []),
    ):
        for row in rows:
            rule_id = str(row.get("ruleId") or "")
            replay_row = replay.get(rule_id) or {}
            sheet.append([
                kind,
                rule_id,
                row.get("title", ""),
                row.get("status", ""),
                row.get("decision", ""),
                row.get("version", ""),
                replay_row.get("decision", ""),
                _governance_summary_text(replay_row.get("summary") or row.get("replaySummary") or {}),
                row.get("description") or row.get("confirmationReason") or row.get("rollbackReason") or "",
            ])
    _format(sheet)


def _write_profile_governance(workbook: Workbook, governance: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Profile治理")
    sheet.append(["类型", "候选ID", "供应商", "Profile", "状态", "决策", "版本", "回放决策", "回放摘要", "说明"])
    replay = governance.get("replaySummaries") or {}
    for kind, rows in (
        ("候选", governance.get("candidates") or []),
        ("已确认", governance.get("activeProfiles") or []),
        ("已回滚", governance.get("rolledBackProfiles") or []),
    ):
        for row in rows:
            candidate_id = str(row.get("candidateId") or "")
            profile_data = row.get("profileData") if isinstance(row.get("profileData"), dict) else {}
            replay_row = replay.get(candidate_id) or {}
            sheet.append([
                kind,
                candidate_id,
                row.get("supplier", ""),
                row.get("profileKey") or profile_data.get("key", ""),
                row.get("status", ""),
                row.get("decision", ""),
                row.get("version") or profile_data.get("version", ""),
                replay_row.get("decision", ""),
                _governance_summary_text(replay_row.get("summary") or row.get("replaySummary") or {}),
                row.get("confirmationReason") or row.get("rollbackReason") or "",
            ])
    _format(sheet)


def _write_correction_governance(workbook: Workbook, governance: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("修正治理")
    sheet.append(["类型", "候选ID", "员工", "状态", "决策", "置信度", "来源", "回放决策", "回放摘要", "说明"])
    replay = governance.get("replaySummaries") or {}
    for kind, rows in (
        ("候选", governance.get("candidates") or []),
        ("已确认", governance.get("activeCorrections") or []),
        ("已回滚", governance.get("rolledBackCorrections") or []),
    ):
        for row in rows:
            candidate_id = str(row.get("candidateId") or "")
            proposed = row.get("proposed") if isinstance(row.get("proposed"), dict) else {}
            replay_row = replay.get(candidate_id) or {}
            sheet.append([
                kind,
                candidate_id,
                proposed.get("employeeName", ""),
                row.get("status", ""),
                row.get("decision", ""),
                row.get("confidence", ""),
                f"{proposed.get('sourceFile', '')} {proposed.get('sourcePageOrRow', '')}".strip(),
                replay_row.get("decision", ""),
                _governance_summary_text(replay_row.get("summary") or row.get("replaySummary") or {}),
                row.get("confirmationReason") or row.get("rollbackReason") or row.get("reason") or "",
            ])
    _format(sheet)


def _write_name_mapping_governance(workbook: Workbook, governance: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("姓名映射治理")
    sheet.append(["类型", "候选ID", "PDF/缓存姓名", "Excel姓名", "文件", "仓库", "状态", "决策", "修复数", "回归数", "异常变化", "历史已检批次", "历史缺明细批次", "历史回归批次", "金额差", "工时差", "操作者", "说明", "证据"])
    replays = governance.get("replaySummaries") if isinstance(governance.get("replaySummaries"), dict) else {}
    for kind, rows in (
        ("候选", governance.get("candidates") or []),
        ("已确认", governance.get("activeMappings") or []),
        ("已回滚", governance.get("rolledBackMappings") or []),
    ):
        for row in rows:
            evidence = row.get("evidence") if isinstance(row.get("evidence"), dict) else {}
            replay = replays.get(row.get("candidateId"), {})
            if not replay and isinstance(row.get("replaySummary"), dict):
                replay = {"summary": row.get("replaySummary")}
            summary = replay.get("summary") if isinstance(replay, dict) and isinstance(replay.get("summary"), dict) else {}
            sheet.append([
                kind,
                row.get("candidateId", ""),
                row.get("cacheEmployeeName", ""),
                row.get("excelEmployeeName", ""),
                row.get("sourceFile", ""),
                row.get("warehouseId", ""),
                row.get("status", ""),
                row.get("decision", ""),
                summary.get("fixedCount", ""),
                summary.get("regressionCount", ""),
                summary.get("exceptionDelta", ""),
                summary.get("historicalCheckedCount", ""),
                summary.get("historicalInsufficientCount", ""),
                summary.get("historicalRegressionCount", ""),
                row.get("amountGap", ""),
                row.get("hoursGap", ""),
                row.get("confirmedBy") or row.get("rolledBackBy") or "",
                row.get("confirmationReason") or row.get("rollbackReason") or row.get("recommendation") or "",
                evidence.get("sourceRefs", ""),
            ])
    _format(sheet)


def _write_reocr_governance(workbook: Workbook, governance: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("图片识别治理")
    sheet.append(["类型", "候选ID", "文件", "仓库", "状态", "决策", "候选行数", "候选金额", "账单金额", "异常数", "建议动作", "根因提示", "疑似姓名配对", "操作者", "说明"])
    for replay in governance.get("replays", []) or []:
        summary = replay.get("summary") if isinstance(replay.get("summary"), dict) else {}
        diagnostics = replay.get("diagnostics") if isinstance(replay.get("diagnostics"), dict) else {}
        sheet.append([
            "回放",
            "",
            replay.get("sourceFile", ""),
            replay.get("warehouseId", ""),
            replay.get("mode", ""),
            replay.get("decision", ""),
            summary.get("candidateRowCount", ""),
            summary.get("candidateAmountTotal", ""),
            summary.get("expectedExcelAmount", ""),
            summary.get("exceptionCount", ""),
            diagnostics.get("recommendedAction", ""),
            "；".join(str(item) for item in diagnostics.get("rootCauseHints", []) or []),
            _reocr_name_pairs_text(diagnostics),
            "",
            "；".join(str(item) for item in replay.get("blockers", []) or []),
        ])
    for kind, rows in (
        ("已确认", governance.get("activeCandidates") or []),
        ("已回滚", governance.get("rolledBackCandidates") or []),
    ):
        for row in rows:
            replay = row.get("replay") if isinstance(row.get("replay"), dict) else {}
            summary = replay.get("summary") if isinstance(replay.get("summary"), dict) else {}
            diagnostics = row.get("diagnostics") if isinstance(row.get("diagnostics"), dict) else {}
            if not diagnostics:
                diagnostics = replay.get("diagnostics") if isinstance(replay.get("diagnostics"), dict) else {}
            sheet.append([
                kind,
                row.get("candidateId", ""),
                row.get("sourceFile", ""),
                row.get("warehouseId", ""),
                row.get("status", ""),
                row.get("decision", ""),
                summary.get("candidateRowCount", ""),
                summary.get("candidateAmountTotal", ""),
                summary.get("expectedExcelAmount", ""),
                summary.get("exceptionCount", ""),
                diagnostics.get("recommendedAction", ""),
                "；".join(str(item) for item in diagnostics.get("rootCauseHints", []) or []),
                _reocr_name_pairs_text(diagnostics),
                row.get("appliedBy") or row.get("confirmedBy") or row.get("rolledBackBy") or "",
                row.get("applicationReason") or row.get("confirmationReason") or row.get("rollbackReason") or "",
            ])
    _format(sheet)


def _write_reocr_upload_coverage(workbook: Workbook, records: List[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("图片识别上传覆盖")
    sheet.append(["类型", "候选文件", "计划任务", "已覆盖", "缺失任务", "计划外范围", "覆盖完整", "文件", "仓库", "行数"])
    for record in records or []:
        summary = record.get("summary") if isinstance(record.get("summary"), dict) else {}
        coverage = record.get("coverage") if isinstance(record.get("coverage"), dict) else {}
        filename = record.get("filename", "")
        sheet.append([
            "汇总",
            filename,
            summary.get("plannedTaskCount", coverage.get("plannedTaskCount", "")),
            summary.get("coveredTaskCount", coverage.get("coveredTaskCount", "")),
            summary.get("missingTaskCount", coverage.get("missingTaskCount", "")),
            summary.get("extraScopeCount", coverage.get("extraScopeCount", "")),
            "是" if coverage.get("coverageComplete") else "否",
            "",
            "",
            summary.get("parsedRowCount", ""),
        ])
        for item in coverage.get("uploadedScopes", []) or []:
            sheet.append(["已上传范围", filename, "", "", "", "", "", item.get("sourceFile", ""), item.get("warehouseId", ""), item.get("rowCount", "")])
        for item in coverage.get("missingTasks", []) or []:
            sheet.append(["缺失计划任务", filename, "", "", "", "", "", item.get("sourceFile", ""), item.get("warehouseId", ""), ""])
        for item in coverage.get("extraScopes", []) or []:
            sheet.append(["计划外范围", filename, "", "", "", "", "", item.get("sourceFile", ""), item.get("warehouseId", ""), item.get("rowCount", "")])
    _format(sheet)


def _reocr_name_pairs_text(diagnostics: Dict[str, Any]) -> str:
    pairs = diagnostics.get("suspectedNamePairs") if isinstance(diagnostics, dict) else []
    if not isinstance(pairs, list):
        return ""
    values = []
    for pair in pairs[:5]:
        if not isinstance(pair, dict):
            continue
        values.append(
            f"{pair.get('cacheEmployeeName', '')} ⇄ {pair.get('excelEmployeeName', '')} "
            f"${pair.get('cacheAmount', 0)} / ${pair.get('excelAmount', 0)}"
        )
    return "；".join(values)


def _write_governance_ai_evidence(workbook: Workbook, audit: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("AI候选治理")
    sheet.append(["项目", "值"])
    sheet.append(["决策", audit.get("decision", "")])
    sheet.append(["需要确认", "是" if audit.get("requiresConfirmation") else "否"])
    sheet.append(["说明", audit.get("message", "")])
    sheet.append([])
    sheet.append(["文件", "仓库", "候选行数", "候选金额", "平均置信度", "决策", "证据样例"])
    for item in audit.get("files", []) or []:
        evidence = item.get("evidence", []) or []
        sample = ""
        if evidence:
            first = evidence[0]
            sample = f"{first.get('employeeName', '')}: ${float(first.get('amount') or 0):.2f} {first.get('sourcePageOrRow', '')}"
        sheet.append([
            item.get("sourceFile", ""),
            item.get("warehouseId", ""),
            item.get("rowCount", 0),
            item.get("candidateAmountTotal", 0),
            item.get("averageConfidence", 0),
            item.get("decision", ""),
            sample,
        ])
    _format(sheet)


def _write_governance_audit_trail(workbook: Workbook, metadata: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("审计记录")
    sheet.append(["对象类型", "对象ID", "动作", "操作者", "原因", "版本/摘要"])
    sources = [
        ("规则", "ruleId", (metadata.get("ruleGovernance", {}) or {}).get("candidates") or []),
        ("规则", "ruleId", (metadata.get("ruleGovernance", {}) or {}).get("activeRules") or []),
        ("规则", "ruleId", (metadata.get("ruleGovernance", {}) or {}).get("rolledBackRules") or []),
        ("姓名映射", "candidateId", (metadata.get("nameMappingGovernance", {}) or {}).get("candidates") or []),
        ("姓名映射", "candidateId", (metadata.get("nameMappingGovernance", {}) or {}).get("activeMappings") or []),
        ("姓名映射", "candidateId", (metadata.get("nameMappingGovernance", {}) or {}).get("rolledBackMappings") or []),
        ("Profile", "candidateId", (metadata.get("profileGovernance", {}) or {}).get("candidates") or []),
        ("Profile", "candidateId", (metadata.get("profileGovernance", {}) or {}).get("activeProfiles") or []),
        ("Profile", "candidateId", (metadata.get("profileGovernance", {}) or {}).get("rolledBackProfiles") or []),
        ("修正", "candidateId", (metadata.get("correctionGovernance", {}) or {}).get("candidates") or []),
        ("修正", "candidateId", (metadata.get("correctionGovernance", {}) or {}).get("activeCorrections") or []),
        ("修正", "candidateId", (metadata.get("correctionGovernance", {}) or {}).get("rolledBackCorrections") or []),
        ("图片识别", "candidateId", (metadata.get("reocrReplayGovernance", {}) or {}).get("activeCandidates") or []),
        ("图片识别", "candidateId", (metadata.get("reocrReplayGovernance", {}) or {}).get("rolledBackCandidates") or []),
    ]
    for object_type, id_key, rows in sources:
        for row in rows:
            object_id = row.get(id_key, "")
            for event in row.get("auditTrail", []) or []:
                sheet.append([
                    object_type,
                    object_id,
                    event.get("action", ""),
                    event.get("actor", ""),
                    event.get("reason", ""),
                    _governance_summary_text(event.get("replaySummary") or {}),
                ])
    _format(sheet)


def _governance_summary_text(summary: Dict[str, Any]) -> str:
    if not isinstance(summary, dict) or not summary:
        return ""
    return "；".join(f"{key}={value}" for key, value in sorted(summary.items()))


def _write_projection_summary(workbook: Workbook, preview: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("预览结论")
    sheet.append(["项目", "值"])
    sheet.append(["预览决策", preview.get("decision", "")])
    sheet.append(["说明", preview.get("reason", "")])
    sheet.append([])

    original = preview.get("originalSummary", {}) or {}
    projected = preview.get("projectedSummary", {}) or {}
    delta = preview.get("summaryDelta", {}) or {}
    sheet.append(["指标", "正式结果", "预览结果", "变化"])
    for key in sorted(set(original) | set(projected) | set(delta)):
        sheet.append([key, original.get(key, ""), projected.get(key, ""), delta.get(key, "")])
    _format(sheet)


def _write_projection_affected_rows(workbook: Workbook, rows: List[Dict[str, Any]]) -> None:
    headers = ["employeeName", "matchStatus", "pdfHoursTotal", "excelHoursTotal", "hoursDelta", "pdfAmountTotal", "excelAmountTotal", "amountDelta", "riskFlags"]
    sheet = workbook.create_sheet("预览影响员工")
    sheet.append(headers)
    for row in rows:
        values = []
        for header in headers:
            value = row.get(header, "")
            if isinstance(value, list):
                value = "；".join(str(item) for item in value)
            values.append(value)
        sheet.append(values)
    _format(sheet)
    _apply_status_fills(sheet, status_column=2, delta_column=8)


def _write_projection_corrections(workbook: Workbook, rows: List[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("已应用修正")
    sheet.append(["候选ID", "员工", "来源文件", "页码/行", "金额变化", "工时变化"])
    for row in rows:
        sheet.append([
            row.get("candidateId", ""),
            row.get("employeeName", ""),
            row.get("sourceFile", ""),
            row.get("sourcePageOrRow", ""),
            row.get("amountDelta", 0),
            row.get("hoursDelta", 0),
        ])
    _format(sheet)


def _write_projection_manual_review(workbook: Workbook, rows: List[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("需人工处理")
    sheet.append(["候选ID", "员工", "原因"])
    for row in rows:
        sheet.append([row.get("candidateId", ""), row.get("employeeName", ""), row.get("reason", "")])
    _format(sheet)


def _write_conclusion(workbook: Workbook, summary: Dict[str, Any], warehouse_comparison: Dict[str, Any] | None = None) -> None:
    """Write the conclusion sheet as the first sheet."""
    sheet = workbook.create_sheet("核对结论", 0)

    # 核对结论
    conclusion_level = summary.get("conclusionLevel", "pass")
    conclusion_message = summary.get("conclusionMessage", "")
    level_display = {"pass": "通过", "warning": "需关注", "critical": "需人工复核"}.get(conclusion_level, conclusion_level)
    sheet.append(["核对结论", f"{level_display} - {conclusion_message}"])
    sheet.append([])

    # 总金额差异
    wc_summary = (warehouse_comparison or {}).get("summary", {})
    amount_delta_total = wc_summary.get("amountDeltaTotal", summary.get("amountDeltaTotal", 0))
    amount_delta_pct = abs(amount_delta_total) / max(abs(wc_summary.get("pdfAmountTotal", 0)), abs(wc_summary.get("excelAmountTotal", 0)), 1.0) * 100
    sheet.append(["总金额差异", f"${amount_delta_total:.2f} ({amount_delta_pct:.2f}%)"])
    sheet.append([])

    # 人数覆盖
    pdf_count = summary.get("pdfEmployeeCount", 0)
    excel_count = summary.get("excelEmployeeCount", 0)
    not_in_invoice = summary.get("notInInvoiceCount", 0)
    sheet.append(["PDF覆盖人数", pdf_count])
    sheet.append(["账单总人数", excel_count])
    sheet.append(["不在本批发票", f"{not_in_invoice}人"])
    sheet.append([])

    # 仓库差异归因摘要
    if warehouse_comparison and warehouse_comparison.get("rows"):
        sheet.append(["仓库差异归因摘要"])
        sheet.append(["仓库", "PDF金额", "Excel金额", "差异", "主要差异来源"])
        for row in warehouse_comparison["rows"]:
            if abs(row.get("amountDelta", 0)) >= 0.01:
                wh_id = row.get("warehouseId", "")
                pdf_amount = row.get("pdfAmountTotal", 0)
                excel_amount = row.get("excelAmountTotal", 0)
                delta = row.get("amountDelta", 0)
                attribution = row.get("attribution", [])
                attr_summary = "；".join([f"{a['employeeName']}: ${a['delta']:.2f}" for a in attribution[:3]])
                sheet.append([f"仓库{wh_id}", f"${pdf_amount:.2f}", f"${excel_amount:.2f}", f"${delta:.2f}", attr_summary])

    _format(sheet)


def _write_quality(workbook: Workbook, extraction_quality: Dict[str, Any]) -> None:
    """Write the quality scoring sheet."""
    sheet = workbook.create_sheet("识别完整度", 1)

    # 质量级别
    level = extraction_quality.get("level", "ok")
    level_display = {"ok": "通过", "warning": "需关注", "critical": "需人工复核"}.get(level, level)
    message = extraction_quality.get("message", "")
    sheet.append(["质量级别", f"{level_display} - {message}"])
    sheet.append([])

    # 质量问题列表
    issues = extraction_quality.get("issues", [])
    if issues:
        sheet.append(["质量问题"])
        for i, issue in enumerate(issues, 1):
            sheet.append([f"{i}.", issue])
        sheet.append([])

    # 详细指标
    metrics = extraction_quality.get("metrics", {})

    # 置信度分布
    confidence = metrics.get("confidence", {})
    if confidence:
        sheet.append(["置信度分布"])
        sheet.append(["平均置信度", f"{confidence.get('average', 0):.3f}"])
        sheet.append(["低置信度记录数 (<0.85)", confidence.get("lowCount", 0)])
        sheet.append(["极低置信度记录数 (<0.5)", confidence.get("veryLowCount", 0)])
        sheet.append(["总记录数", confidence.get("totalCount", 0)])
        sheet.append([])

    # 抽取方法统计
    methods = metrics.get("extractionMethods", {})
    if methods:
        sheet.append(["抽取方法统计"])
        sheet.append(["规则抽取", methods.get("rule", 0)])
        sheet.append(["AI文本抽取", methods.get("ai_text", 0)])
        sheet.append(["AI图片抽取", methods.get("ai_image", 0)])
        sheet.append([])

    # 员工数量对比
    employee_counts = metrics.get("employeeCounts", {})
    if employee_counts:
        sheet.append(["员工数量对比"])
        sheet.append(["PDF员工数", employee_counts.get("pdf", 0)])
        sheet.append(["Excel员工数", employee_counts.get("excel", 0)])
        sheet.append(["PDF未匹配", employee_counts.get("unmatchedPdf", 0)])
        sheet.append(["Excel未匹配", employee_counts.get("unmatchedExcel", 0)])
        sheet.append([])

    # 金额/工时偏差
    totals = metrics.get("totals", {})
    if totals:
        sheet.append(["金额/工时偏差"])
        sheet.append(["PDF总工时", f"{totals.get('pdfHours', 0):.2f}"])
        sheet.append(["Excel总工时", f"{totals.get('excelHours', 0):.2f}"])
        sheet.append(["工时差异", f"{totals.get('hoursDelta', 0):.2f}"])
        sheet.append(["PDF总金额", f"${totals.get('pdfAmount', 0):.2f}"])
        sheet.append(["Excel总金额", f"${totals.get('excelAmount', 0):.2f}"])
        sheet.append(["金额差异", f"${totals.get('amountDelta', 0):.2f}"])
        sheet.append([])

    # 仓库问题
    warehouse_issues = metrics.get("warehouseIssues", [])
    if warehouse_issues:
        sheet.append(["仓库问题"])
        for issue in warehouse_issues:
            sheet.append(["", issue])
        sheet.append([])

    # 名称模式
    name_patterns = metrics.get("namePatterns", {})
    if name_patterns:
        sheet.append(["名称模式"])
        sheet.append(["包含中文", "是" if name_patterns.get("hasChinese") else "否"])
        sheet.append(["包含英文", "是" if name_patterns.get("hasEnglish") else "否"])
        sheet.append(["中英文混合", "是" if name_patterns.get("hasMixed") else "否"])

    _format(sheet)


def _write_summary(workbook: Workbook, summary: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("核对摘要")
    sheet.append(["项目", "值"])
    for key, value in summary.items():
        sheet.append([key, value])
    _format(sheet)


def _write_reconciliation_detail(workbook: Workbook, rows: List[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("全员对账明细")
    headers = ["员工", "状态", "PDF工时", "Excel工时", "工时差异", "PDF金额", "Excel金额", "金额差异", "风险标记", "来源"]
    sheet.append(headers)
    for row in rows:
        risk_flags = row.get("riskFlags", [])
        sheet.append([
            row.get("employeeName", ""),
            _display_status(row.get("matchStatus", ""), risk_flags),
            row.get("pdfHoursTotal", 0),
            row.get("excelHoursTotal", 0),
            row.get("hoursDelta", 0),
            row.get("pdfAmountTotal", 0),
            row.get("excelAmountTotal", 0),
            row.get("amountDelta", 0),
            "；".join(_sanitize_business_text(item) for item in risk_flags) if isinstance(risk_flags, list) else _sanitize_business_text(risk_flags),
            row.get("sourceRefs", ""),
        ])
    _format(sheet)
    _apply_status_fills(sheet, status_column=2, delta_column=8)


def _write_warehouse_summary(workbook: Workbook, warehouse_comparison: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("仓库金额汇总")
    sheet.append(["仓库", "状态", "PDF人数/发票数", "Excel人数", "PDF工时", "Excel工时", "PDF金额", "Excel金额", "金额差异", "主要差异来源"])
    for row in warehouse_comparison.get("rows", []):
        attribution = row.get("attribution", [])
        attr_summary = "；".join(
            f"{item.get('employeeName', '')}: ${float(item.get('delta') or 0):.2f}"
            for item in attribution[:5]
        )
        sheet.append([
            row.get("warehouseId", ""),
            row.get("matchStatus", ""),
            row.get("pdfEmployeeCount", 0),
            row.get("excelEmployeeCount", 0),
            row.get("pdfHoursTotal", 0),
            row.get("excelHoursTotal", 0),
            row.get("pdfAmountTotal", 0),
            row.get("excelAmountTotal", 0),
            row.get("amountDelta", 0),
            attr_summary,
        ])
    _format(sheet)
    _apply_status_fills(sheet, status_column=2, delta_column=9)


def _write_reconciliation_diagnostics(workbook: Workbook, diagnostics: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("信号诊断")
    sheet.append(["项目", "值"])
    sheet.append(["诊断级别", diagnostics.get("level", "")])
    sheet.append(["诊断结论", diagnostics.get("message", "")])
    sheet.append(["下一步", diagnostics.get("nextStep", "")])
    sheet.append([])

    signals = diagnostics.get("signals", {}) or {}
    sheet.append(["总额信号", "金额"])
    sheet.append(["快速PDF总额", signals.get("fastPdfTotal", 0)])
    sheet.append(["员工PDF明细总额", signals.get("employeePdfTotal", 0)])
    sheet.append(["Excel账单总额", signals.get("excelTotal", 0)])
    sheet.append(["仓库PDF总额", signals.get("warehouseTotal", 0)])
    sheet.append([])

    issues = diagnostics.get("issues", []) or []
    if issues:
        sheet.append(["诊断问题", "级别", "说明", "明细"])
        for issue in issues:
            items = issue.get("items", []) or []
            sheet.append([
                issue.get("title", issue.get("code", "")),
                issue.get("level", ""),
                issue.get("message", ""),
                "；".join(str(item) for item in items),
            ])
        sheet.append([])

    amount_basis = signals.get("amountBasis", []) or []
    if amount_basis:
        sheet.append([
            "仓库",
            "PDF总额",
            "OTWS汇总总额",
            "PDF-OTWS差异",
            "费用组成合计",
            "员工薪资明细",
            "福利/补充费用",
            "装卸费用",
            "证据",
        ])
        for item in amount_basis:
            evidence = "; ".join(value for value in (item.get("summaryEvidence", ""), item.get("detailEvidence", "")) if value)
            sheet.append([
                item.get("warehouseId", ""),
                item.get("pdfTotal", 0),
                item.get("reportedTotal", 0),
                item.get("pdfVsReportedDelta", 0),
                item.get("componentTotal", 0),
                item.get("employeeExpenses", 0),
                item.get("employeeBenefits", 0),
                item.get("loadingAndUnloading", 0),
                evidence,
            ])
        sheet.append([])

    offsetting_deltas = signals.get("offsettingWarehouseDeltas", []) or []
    if offsetting_deltas:
        sheet.append(["互相抵消的仓库差异"])
        sheet.append(["仓库", "PDF总额", "Excel总额", "差异", "主要归因"])
        for item in offsetting_deltas:
            attribution = "；".join(
                f"{row.get('employeeName', '')}: ${float(row.get('delta') or 0):.2f}"
                for row in (item.get("attribution", []) or [])[:5]
            )
            sheet.append([
                item.get("warehouseId", ""),
                item.get("pdfAmountTotal", 0),
                item.get("excelAmountTotal", 0),
                item.get("amountDelta", 0),
                attribution,
            ])
        sheet.append([])

    employee_attribution = signals.get("employeeAttribution", []) or []
    if employee_attribution:
        sheet.append(["员工主导的仓库差异"])
        sheet.append(["仓库", "员工", "PDF金额", "Excel金额", "员工差异", "仓库总差异"])
        for item in employee_attribution:
            sheet.append([
                item.get("warehouseId", ""),
                item.get("employeeName", ""),
                item.get("pdfAmount", 0),
                item.get("excelAmount", 0),
                item.get("delta", 0),
                item.get("warehouseDelta", 0),
            ])
        sheet.append([])

    allocation_issues = signals.get("crossWarehouseEmployeeAllocation", []) or []
    if allocation_issues:
        sheet.append(["员工跨仓库金额抵消"])
        sheet.append(["员工", "净差异", "仓库数", "仓库明细", "建议"])
        for item in allocation_issues:
            details = "；".join(
                (
                    f"仓库 {row.get('warehouseId', '')}: "
                    f"PDF ${float(row.get('pdfAmount') or 0):.2f}, "
                    f"Excel ${float(row.get('excelAmount') or 0):.2f}, "
                    f"差异 ${float(row.get('amountDelta') or 0):.2f}"
                )
                for row in (item.get("warehouses", []) or [])
            )
            sheet.append([
                item.get("employeeName", ""),
                item.get("netAmountDelta", 0),
                item.get("warehouseCount", 0),
                details,
                item.get("recommendation", ""),
            ])

    _format(sheet)


def _write_ai_cache_audit(workbook: Workbook, audit: Dict[str, Any]) -> None:
    sheet = workbook.create_sheet("AI候选证据")
    sheet.append(["项目", "值"])
    sheet.append(["处理决策", audit.get("decision", "")])
    sheet.append(["需要人工确认", "是" if audit.get("requiresConfirmation") else "否"])
    sheet.append(["说明", audit.get("message", "")])
    summary = audit.get("summary", {}) or {}
    sheet.append(["候选文件数", summary.get("candidateFileCount", 0)])
    sheet.append(["候选金额合计", summary.get("candidateAmountTotal", 0)])
    sheet.append([])
    sheet.append(["文件", "仓库", "候选行数", "候选金额", "平均置信度", "决策", "缓存文件", "证据样例"])
    for item in audit.get("files", []) or []:
        evidence = item.get("evidence", []) or []
        sample = ""
        if evidence:
            first = evidence[0]
            sample = f"{first.get('employeeName', '')}: ${float(first.get('amount') or 0):.2f} {first.get('evidenceText', '')}"
        sheet.append([
            item.get("sourceFile", ""),
            item.get("warehouseId", ""),
            item.get("rowCount", 0),
            item.get("candidateAmountTotal", 0),
            item.get("averageConfidence", 0),
            item.get("decision", ""),
            "；".join(str(name) for name in item.get("cacheFiles", []) or []),
            sample,
        ])
    _format(sheet)


def _write_rows(workbook: Workbook, title: str, rows: List[Dict[str, Any]]) -> None:
    headers = [
        ("employeeName", "员工"),
        ("matchStatus", "状态"),
        ("pdfHoursTotal", "PDF工时"),
        ("excelHoursTotal", "Excel工时"),
        ("hoursDelta", "工时差异"),
        ("pdfAmountTotal", "PDF金额"),
        ("excelAmountTotal", "Excel金额"),
        ("amountDelta", "金额差异"),
        ("riskFlags", "提示"),
        ("sourceRefs", "来源"),
    ]
    sheet = workbook.create_sheet(title)
    sheet.append([label for _, label in headers])
    for row in rows:
        values = []
        for key, _label in headers:
            value = row.get(key, "")
            if isinstance(value, list):
                value = "；".join(_sanitize_business_text(item) for item in value)
            elif key == "matchStatus":
                value = _sanitize_business_text(value)
            values.append(value)
        sheet.append(values)
    _format(sheet)


def _write_detail(workbook: Workbook, title: str, rows: List[LaborLineItem]) -> None:
    headers = [
        ("source_type", "来源类型"),
        ("source_file", "来源文件"),
        ("source_page_or_row", "页码/行号"),
        ("employee_id", "工号"),
        ("employee_name_raw", "原始员工姓名"),
        ("employee_name_normalized", "标准员工姓名"),
        ("hours", "工时"),
        ("amount", "金额"),
        ("currency", "币种"),
        ("confidence", "识别程度"),
        ("evidence_text", "原文证据"),
    ]
    sheet = workbook.create_sheet(title)
    sheet.append([label for _, label in headers])
    for row in rows:
        data = row.to_dict()
        sheet.append([_business_detail_value(key, data.get(key, "")) for key, _label in headers])
    _format(sheet)


def _write_mapping(workbook: Workbook, mapping: Dict[str, str]) -> None:
    sheet = workbook.create_sheet("上传字段对应关系")
    sheet.append(["字段", "Excel列"])
    for key, label in mapping.items():
        sheet.append([_business_mapping_key(key), label])
    _format(sheet)


def _write_candidate_matches(workbook: Workbook, rows: List[Dict[str, Any]]) -> None:
    headers = ["issueType", "pdfEmployeeName", "excelEmployeeName", "nameSimilarity", "pdfHoursTotal", "excelHoursTotal", "hoursDelta", "pdfAmountTotal", "excelAmountTotal", "amountDelta", "recommendation", "sourceRefs"]
    sheet = workbook.create_sheet("姓名格式差异")
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    _format(sheet)


def _filter(rows: List[Dict[str, Any]], status: str) -> List[Dict[str, Any]]:
    return [row for row in rows if row.get("matchStatus") == status]


def _business_detail_row(index: int, row: Dict[str, Any]) -> str:
    pdf_name, excel_name = _split_business_names(row.get("employeeName", ""))
    status = _business_row_status(row)
    status_class = "tag-ok" if status == "一致" else "tag-diff" if status == "有差异" else "tag-warn"
    invoice = _invoice_label(row.get("sourceRefs", ""))
    return (
        f"<tr>"
        f'<td class="num">{index}</td>'
        f"<td>{escape(invoice)}</td>"
        f"<td>{escape(pdf_name)}</td>"
        f"<td>{escape(excel_name)}</td>"
        f'<td class="num">{_hours(_num(row.get("pdfHoursTotal")))} / {_hours(_num(row.get("excelHoursTotal")))}</td>'
        f'<td class="num">-</td>'
        f'<td class="num">{_money(_num(row.get("pdfAmountTotal")))}</td>'
        f'<td class="num">{_money(_num(row.get("excelAmountTotal")))}</td>'
        f'<td class="num">{_signed_money(_num(row.get("amountDelta")))}</td>'
        f'<td><span class="tag {status_class}">{escape(status)}</span></td>'
        f"<td>{escape(_business_note(row))}</td>"
        f"</tr>"
    )


def _business_auto_fix_section(rows: List[Dict[str, Any]]) -> str:
    auto_fixed_rows = [row for row in rows if _is_auto_fixed_name_row(row)]
    if not auto_fixed_rows:
        return _business_empty_section("系统自动修正", "本次未发现可由系统自动合并的姓名格式差异。")
    items = []
    for row in auto_fixed_rows[:8]:
        pdf_name, excel_name = _split_business_names(row.get("employeeName", ""))
        items.append(
            "<article class=\"business-item\">"
            f"<strong>{escape(pdf_name)} ⇄ {escape(excel_name)}</strong>"
            "<p>系统已自动合并姓名格式差异。该类差异通常来自大小写、重音符号、空格标点或姓名前后顺序不同；金额仍按原发票和账单金额展示。</p>"
            "</article>"
        )
    return _business_section("系统自动修正", "系统已自动合并姓名格式差异", "".join(items))


def _business_amount_layers_section(
    *,
    pdf_total: float,
    excel_total: float,
    amount_delta: float,
    detail_pdf_total: float,
    detail_excel_total: float,
) -> str:
    detail_gap = detail_pdf_total - detail_excel_total
    body = (
        '<article class="business-item">'
        "<strong>总账结论优先看整批 PDF 与整批 Excel 的差额</strong>"
        f"<p>整批 PDF 发票总额 {_money(pdf_total)}，整批 Excel 账单总额 {_money(excel_total)}，总差额 {_signed_money(amount_delta)}。"
        f"已识别员工明细金额：PDF {_money(detail_pdf_total)}，Excel {_money(detail_excel_total)}，明细差额 {_signed_money(detail_gap)}。"
        "员工明细金额用于定位差异，不等同于整批总账金额；"
        "如果员工明细金额小于整批总额，不代表账单少读了，只代表当前页面只展开了用于确认的明细范围。</p>"
        "</article>"
    )
    return _business_section("总金额核对", "先看整批总额，再用员工明细定位原因。", body)


def _business_employee_recognition_section(
    *,
    pdf_employee_count: int,
    excel_employee_count: int,
    displayed_employee_count: int,
    excel_record_count: int | None,
    details_incomplete: bool,
    warehouse_needs_review: bool,
) -> str:
    record_text = f"整批账单已读取 {excel_record_count} 行。" if excel_record_count else "整批账单已完成读取。"
    if details_incomplete:
        status = "员工明细未完整展开"
        explanation = (
            f"{record_text}当前还没有可逐项展示的员工明细。"
            "这不影响总账金额判断，但不能直接作为最终员工明细结论。"
        )
    elif warehouse_needs_review:
        status = "只展开需要确认的员工明细"
        explanation = (
            f"{record_text}当前报告展开 {displayed_employee_count} 名需要确认的员工明细，"
            "其余无明显差异的员工不在本段重复展示。"
        )
    else:
        status = "员工明细已进入核对"
        explanation = (
            f"{record_text}当前报告展示 {displayed_employee_count} 名员工明细，"
            "用于确认每位员工的发票金额和账单金额是否一致。"
        )
    body = (
        '<article class="business-item">'
        f"<strong>{escape(status)}</strong>"
        f"<p>{escape(explanation)} PDF 已识别员工 {pdf_employee_count} 人，Excel 当前核对员工 {excel_employee_count} 人。"
        "如需查看所有原始员工行，请下载 Excel 明细。</p>"
        "</article>"
    )
    return _business_section("员工明细识别情况", "说明当前报告展示的是整批账单还是待确认范围。", body)


def _business_suspected_match_section(candidate_matches: List[Dict[str, Any]]) -> str:
    if not candidate_matches:
        return _business_empty_section("疑似同一员工，需确认", "本次未发现需要业务确认的疑似同一员工。")
    items = []
    for row in candidate_matches[:8]:
        pdf_name = str(row.get("pdfEmployeeName") or "-")
        excel_name = str(row.get("excelEmployeeName") or "-")
        recommendation = _business_candidate_recommendation(row)
        items.append(
            "<article class=\"business-item\">"
            f"<strong>{escape(pdf_name)} ⇄ {escape(excel_name)}</strong>"
            f"<p>疑似同一员工，需确认。发票金额 {_money(_num(row.get('pdfAmountTotal')))}，账单金额 {_money(_num(row.get('excelAmountTotal')))}，差额 {_signed_money(_num(row.get('amountDelta')))}。{escape(recommendation)} 确认前不会自动合并姓名。</p>"
            "</article>"
        )
    return _business_section("疑似同一员工，需确认", "系统只给建议，不会仅凭金额接近自动合并。", "".join(items))


def _business_candidate_recommendation(row: Dict[str, Any]) -> str:
    recommendation = row.get("recommendation")
    if recommendation:
        return _sanitize_business_text(recommendation)
    name_similarity = _num(row.get("nameSimilarity"))
    amount_gap = abs(_num(row.get("amountDelta") if row.get("amountDelta") is not None else row.get("amountGap")))
    if name_similarity > 0 and name_similarity < 0.45 and amount_gap <= 0.1:
        return "金额接近，但姓名不像，不能自动合并。"
    return "姓名接近，但仍需确认是否为同一员工。"


def _business_pending_exception_section(rows: List[Dict[str, Any]]) -> str:
    pending_rows = [row for row in rows if _business_row_status(row) != "一致"]
    if not pending_rows:
        return _business_empty_section("待确认异常", "本次没有需要确认的员工级异常。")
    items = []
    for row in pending_rows[:10]:
        pdf_name, excel_name = _split_business_names(row.get("employeeName", ""))
        items.append(
            "<article class=\"business-item\">"
            f"<strong>{escape(pdf_name)} ⇄ {escape(excel_name)}</strong>"
            f"<p>{escape(_business_note(row))}。处理建议：{escape(_business_action_suggestion(row))}。发票金额 {_money(_num(row.get('pdfAmountTotal')))}，账单金额 {_money(_num(row.get('excelAmountTotal')))}，差额 {_signed_money(_num(row.get('amountDelta')))}。</p>"
            "</article>"
        )
    return _business_section(
        "待确认异常",
        "优先处理影响放行或留档的项目。处理顺序：先确认金额口径，再确认缺发票项，最后确认疑似同一员工。",
        "".join(items),
    )


def _business_download_section() -> str:
    return (
        '<section class="section">'
        '<div class="section-header"><h2>下载 Excel 明细</h2><div class="sub">Excel 用于留档和继续筛选，页面结论以本 HTML 报告为准。</div></div>'
        '<p class="download-note">Excel 明细用于留档、筛选和逐行核查；对外沟通时，请优先引用本页面的核对结论、金额差异和待确认说明。</p>'
        '</section>'
    )


def _business_decision_panel(
    *,
    conclusion: str,
    amount_delta: float,
    diff_count: int,
    details_incomplete: bool,
    warehouse_needs_review: bool,
) -> str:
    release_decision = _business_release_decision(conclusion)
    total_status = "总账通过" if _conclusion_is_total_pass(conclusion) else "总账不一致"
    if conclusion == "系统未能完成核对":
        total_status = "未完成"
    detail_status = _business_detail_status(
        conclusion=conclusion,
        diff_count=diff_count,
        details_incomplete=details_incomplete,
        warehouse_needs_review=warehouse_needs_review,
    )
    next_step = _business_next_step(
        conclusion=conclusion,
        diff_count=diff_count,
        details_incomplete=details_incomplete,
        warehouse_needs_review=warehouse_needs_review,
    )
    return (
        '<section class="decision-panel" aria-label="核对结论与下一步">'
        '<article class="decision-box primary">'
        '<div class="decision-question">这批账能不能放行？</div>'
        f'<div class="decision-answer">{escape(release_decision)}</div>'
        f'<p class="decision-text">{escape(next_step)}</p>'
        '</article>'
        '<article class="decision-box">'
        '<div class="decision-question">总金额核对</div>'
        f'<div class="decision-answer">{escape(total_status)}</div>'
        f'<p class="decision-text">差额 {_signed_money(amount_delta)}，容差 $0.10。</p>'
        '</article>'
        '<article class="decision-box">'
        '<div class="decision-question">员工明细状态</div>'
        f'<div class="decision-answer">{escape(detail_status)}</div>'
        f'<p class="decision-text">待确认 {diff_count} 项。</p>'
        '</article>'
        '<article class="decision-box">'
        '<div class="decision-question">下一步</div>'
        f'<div class="decision-answer">{escape(_business_short_next_step(conclusion, diff_count, details_incomplete))}</div>'
        '<p class="decision-text">先看总金额，再处理员工姓名、金额和归属问题。</p>'
        '</article>'
        '</section>'
    )


def _business_release_decision(conclusion: str) -> str:
    if conclusion == "总账通过":
        return "可放行"
    if conclusion == "总账通过，但员工明细待确认":
        return "需业务确认"
    return "不建议放行"


def _business_detail_status(
    *,
    conclusion: str,
    diff_count: int,
    details_incomplete: bool,
    warehouse_needs_review: bool,
) -> str:
    if conclusion == "系统未能完成核对":
        return "未完成"
    if details_incomplete:
        return "员工明细未完整识别"
    if warehouse_needs_review:
        return "仓库或员工归属待确认"
    if diff_count:
        return "员工明细待确认"
    return "员工明细已通过"


def _business_next_step(
    *,
    conclusion: str,
    diff_count: int,
    details_incomplete: bool,
    warehouse_needs_review: bool,
) -> str:
    if conclusion == "总账通过":
        return "总金额和员工明细均未发现需要处理的问题，可按内部流程放行。"
    if conclusion == "总账通过，但员工明细待确认":
        if details_incomplete:
            return "总金额已通过；员工明细未完整识别，不影响总账结论，但需要业务确认明细后再对外留档。"
        if warehouse_needs_review:
            return "总金额已通过；请确认仓库归属或员工拆分后再对外留档。"
        return "总金额已通过；请处理下方员工明细待确认项后再对外留档。"
    if conclusion == "总金额存在差异，暂不能放行":
        return "总金额超出 $0.10 容差，先复核发票总额、账单总额和所属账期。"
    if diff_count:
        return "请先处理下方待确认项，再判断是否可以放行。"
    return "系统未能完成核对，请检查材料是否完整后重新生成报告。"


def _business_short_next_step(conclusion: str, diff_count: int, details_incomplete: bool) -> str:
    if conclusion == "总账通过":
        return "留档即可"
    if conclusion == "总账通过，但员工明细待确认":
        return "确认明细"
    if details_incomplete:
        return "补看原发票"
    if diff_count:
        return "处理差异"
    return "重新生成"


def _business_section(title: str, subtitle: str, body_html: str) -> str:
    return (
        '<section class="section">'
        f'<div class="section-header"><h2>{escape(title)}</h2><div class="sub">{escape(subtitle)}</div></div>'
        f'<div class="business-list">{body_html}</div>'
        '</section>'
    )


def _business_empty_section(title: str, message: str) -> str:
    return _business_section(title, message, f'<article class="business-item"><p>{escape(message)}</p></article>')


def _is_auto_fixed_name_row(row: Dict[str, Any]) -> bool:
    if _business_row_status(row) != "一致":
        return False
    pdf_name, excel_name = _split_business_names(row.get("employeeName", ""))
    if pdf_name == "-" or excel_name == "-" or pdf_name == excel_name:
        return False
    return _canonical_name_tokens(pdf_name) == _canonical_name_tokens(excel_name)


def _canonical_name_tokens(value: str) -> List[str]:
    without_accents = "".join(
        char for char in unicodedata.normalize("NFKD", value) if not unicodedata.combining(char)
    )
    normalized = re.sub(r"[^a-z0-9]+", " ", without_accents.lower()).strip()
    return sorted(token for token in normalized.split() if token)


def _business_conclusion(summary: Dict[str, Any], rows: List[Dict[str, Any]]) -> str:
    if (
        summary.get("systemIncomplete")
        or summary.get("extractionFailed")
        or summary.get("failed")
        or summary.get("status") in {"抽取失败", "解析失败", "核对失败"}
    ):
        return "系统未能完成核对"
    amount_delta = _num(summary.get("amountDeltaTotal"))
    has_uploaded_totals = bool(
        _num(summary.get("pdfEmployeeCount"))
        or _num(summary.get("excelEmployeeCount"))
        or _num(summary.get("pdfAmountTotal"))
        or _num(summary.get("excelAmountTotal"))
    )
    details_incomplete = not rows and has_uploaded_totals
    has_detail_review_items = any(_business_row_status(row) != "一致" for row in rows)
    if _business_total_amount_passed(amount_delta):
        if details_incomplete or has_detail_review_items:
            return "总账通过，但员工明细待确认"
        return "总账通过"
    if not _business_total_amount_passed(amount_delta):
        return "总金额存在差异，暂不能放行"
    if details_incomplete or has_detail_review_items or _num(summary.get("amountDiffCount")) > 0:
        return "需要业务确认"
    return "总账通过"


def _business_total_amount_passed(amount_delta: float, tolerance: float = 0.10) -> bool:
    return round(abs(float(amount_delta or 0)), 2) <= round(abs(float(tolerance or 0)), 2)


def _business_conclusion_message(
    conclusion: str,
    amount_delta: float,
    matched_count: int,
    diff_count: int,
    scope_note: str = "",
    *,
    details_incomplete: bool = False,
) -> str:
    if conclusion == "总账通过":
        return f"核对通过：{matched_count} 名员工金额一致，总差额 {_signed_money(amount_delta)}。"
    if conclusion == "总账通过，但员工明细待确认":
        if not details_incomplete:
            return (
                "系统已确认本批总金额一致，但员工明细仍有需要确认的项目。"
                "员工级差异仅供确认，不能直接作为最终员工明细结论。"
            )
        return (
            "系统已确认本批总金额一致，但部分员工明细未完整识别，"
            "员工级差异仅供确认，不能直接作为最终员工明细结论。"
        )
    if conclusion == "总金额存在差异，暂不能放行":
        if details_incomplete:
            return (
                f"总金额存在差异：PDF 比 Excel {_amount_delta_direction(amount_delta)} {_money(abs(amount_delta))}。"
                "由于员工明细未完整识别，系统暂时无法定位全部差异来源。"
            )
        return (
            f"总金额存在差异：PDF 比 Excel {_amount_delta_direction(amount_delta)} {_money(abs(amount_delta))}。"
            "请先查看下方员工明细中的金额、工时或费率差异。"
        )
    if conclusion == "需要确认":
        return f"整批金额接近一致，但仍有部分仓库需要确认。{scope_note}"
    if conclusion == "需要业务确认":
        return f"需要业务确认：{diff_count} 名员工或账单记录存在无法自动确认的情况。"
    return "系统未能完成核对：请查看原发票和账单后重新生成报告。"


def _business_empty_detail_text(*, conclusion: str, detail_rows_incomplete: bool, amount_delta: float) -> str:
    if conclusion == "总账通过，但员工明细待确认" and detail_rows_incomplete:
        return "本批总金额已完成核对，但当前没有可逐项展示的员工明细；请下载 Excel 明细或查看原始发票后确认员工级明细。"
    if conclusion == "总金额存在差异，暂不能放行" and detail_rows_incomplete:
        return (
            f"本批总金额存在差异：{_signed_money(amount_delta)}；当前员工明细未完整识别，"
            "系统暂时无法逐项定位全部差异来源。"
        )
    return "当前没有需要逐项展示的员工差异；如需留档或继续筛选，请下载 Excel 明细。"


def _business_row_status(row: Dict[str, Any]) -> str:
    status = str(row.get("matchStatus") or "")
    flags = row.get("riskFlags") if isinstance(row.get("riskFlags"), list) else []
    if status == "通过" and abs(_num(row.get("amountDelta"))) <= 0.1:
        return "一致"
    if status in {"疑似姓名匹配", "疑似同一员工", "低置信度抽取"} or any("合并" in str(flag) or "复核" in str(flag) for flag in flags):
        return "需要业务确认"
    return "有差异"


def _business_note(row: Dict[str, Any]) -> str:
    status = str(row.get("matchStatus") or "")
    flags = row.get("riskFlags") if isinstance(row.get("riskFlags"), list) else []
    amount_delta = abs(_num(row.get("amountDelta")))
    hours_delta = abs(_num(row.get("hoursDelta")))
    notes: List[str] = []
    if any("合并" in str(flag) for flag in flags):
        notes.append("同一员工可能存在多行账单，需要确认是否应合并")
    if status == "疑似姓名匹配":
        notes.append("需要确认该员工是否为同一人")
    if notes:
        return "；".join(notes)
    if status == "低置信度抽取":
        return "员工明细未完整识别，请查看原发票"
    if status == "PDF有Excel无":
        return "发票有账单无，需要确认该发票是否属于本期"
    if status == "Excel有PDF无":
        return "账单有发票无，需要确认该账单是否属于本期"
    if hours_delta <= 0.01 and amount_delta > 0.1:
        return "工时一致但金额不一致，需要确认费率或金额口径"
    if hours_delta > 0.01 and amount_delta <= 0.1:
        return "金额一致但工时拆分不同，需要确认 REG/OT 拆分"
    if amount_delta <= 0.1:
        return "小额四舍五入差异，可按业务容差确认"
    return "金额或工时存在差异，需要业务确认"


def _business_action_suggestion(row: Dict[str, Any]) -> str:
    status = str(row.get("matchStatus") or "")
    flags = row.get("riskFlags") if isinstance(row.get("riskFlags"), list) else []
    amount_delta = abs(_num(row.get("amountDelta")))
    hours_delta = abs(_num(row.get("hoursDelta")))
    if status == "Excel有PDF无":
        return "确认本员工是否属于本批发票"
    if status == "PDF有Excel无":
        return "确认该发票是否属于本期账单"
    if status == "疑似姓名匹配" or any("合并" in str(flag) for flag in flags):
        return "业务确认是否同一人，确认前不会自动合并姓名"
    if status == "低置信度抽取":
        return "查看原始发票，确认员工姓名、工时和金额"
    if hours_delta <= 0.01 and amount_delta > 0.1:
        return "核对费率、加班、服务费或税费是否同一口径"
    if hours_delta > 0.01:
        return "核对账期、日期和工时"
    return "确认金额和工时是否可按业务容差接受"


def _sanitize_business_text(value: Any) -> str:
    text = str(value or "")
    replacements = {
        "人工复核": "业务确认",
        "需复核": "待确认",
        "复核": "确认",
        "低置信度抽取": "明细识别不完整",
        "低置信度": "识别不完整",
        "疑似姓名匹配": "疑似同一员工",
        "姓名模糊匹配": "疑似同一员工",
        "PDF有Excel无": "发票有账单无",
        "Excel有PDF无": "账单有发票无",
        "工时不一致": "工时待确认",
        "金额差异": "金额不一致",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return text


def _business_detail_value(key: str, value: Any) -> Any:
    if key == "source_type":
        return {
            "pdf_invoice": "PDF发票",
            "excel_bill": "Excel账单",
            "excel": "Excel账单",
            "pdf": "PDF发票",
        }.get(str(value), _sanitize_business_text(value))
    if key in {"employee_name_raw", "employee_name_normalized", "evidence_text"}:
        return _sanitize_business_text(value)
    return value


def _business_mapping_key(value: Any) -> str:
    return {
        "name": "姓名",
        "hours": "工时",
        "amount": "金额",
        "employee_id": "工号",
        "employeeId": "工号",
        "currency": "币种",
    }.get(str(value), _sanitize_business_text(value))


def _split_business_names(value: Any) -> tuple[str, str]:
    text = str(value or "").strip()
    if "⇄" in text:
        left, right = text.split("⇄", 1)
        return left.strip() or "-", right.strip() or "-"
    return text or "-", text or "-"


def _invoice_label(source_refs: Any) -> str:
    text = str(source_refs or "")
    match = re.search(r"Invoice[-_\s]?(\d+)", text, re.IGNORECASE)
    if match:
        return match.group(1)
    match = re.search(r"(\d{5,})", text)
    if match:
        return match.group(1)
    return "-"


def _num(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _money(value: float) -> str:
    return f"${value:,.2f}"


def _signed_money(value: float) -> str:
    sign = "+" if value >= 0 else "-"
    return f"{sign}${abs(value):,.2f}"


def _hours(value: float) -> str:
    return f"{value:,.2f}h"


def _conclusion_bg(conclusion: str) -> str:
    return "#f0fdf4" if _conclusion_is_total_pass(conclusion) else "#fffbeb" if conclusion == "需要业务确认" else "#fff8f8"


def _conclusion_border(conclusion: str) -> str:
    return "#86efac" if _conclusion_is_total_pass(conclusion) else "#fcd34d" if conclusion == "需要业务确认" else "#fecaca"


def _conclusion_color(conclusion: str) -> str:
    return "#166534" if _conclusion_is_total_pass(conclusion) else "#92400e" if conclusion == "需要业务确认" else "#991b1b"


def _conclusion_is_total_pass(conclusion: str) -> bool:
    return conclusion in {"总账通过", "总账通过，但员工明细待确认"}


def _amount_delta_direction(amount_delta: float) -> str:
    return "多" if amount_delta > 0 else "少"


def _format(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column in range(1, sheet.max_column + 1):
        letter = get_column_letter(column)
        width = 14
        for cell in sheet[letter]:
            width = min(max(width, len(str(cell.value or "")) + 2), 48)
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"


def _display_status(status: str, risk_flags: Any) -> str:
    if status == "通过" and isinstance(risk_flags, list) and "工时需复核" in risk_flags:
        return "金额一致（工时需确认）"
    if status == "通过" and isinstance(risk_flags, list) and "姓名格式差异自动合并" in risk_flags:
        return "系统已自动修正"
    return _sanitize_business_text(status)


def _apply_status_fills(sheet, status_column: int, delta_column: int) -> None:
    ok_fill = PatternFill("solid", fgColor="EAF7EA")
    warn_fill = PatternFill("solid", fgColor="FFF4D6")
    diff_fill = PatternFill("solid", fgColor="FDEAEA")
    for row in range(2, sheet.max_row + 1):
        status = str(sheet.cell(row=row, column=status_column).value or "")
        try:
            delta = float(sheet.cell(row=row, column=delta_column).value or 0)
        except (TypeError, ValueError):
            delta = 0.0
        if "差异" in status or "不一致" in status or abs(delta) >= 0.1:
            fill = diff_fill
        elif "确认" in status or "疑似" in status or "识别不完整" in status:
            fill = warn_fill
        elif "通过" in status or "一致" in status:
            fill = ok_fill
        else:
            continue
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = fill
