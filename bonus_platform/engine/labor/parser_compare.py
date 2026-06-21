from __future__ import annotations

import argparse
import importlib.util
import json
import re
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook


ParserRunner = Callable[[Path], str]

MONEY_RE = re.compile(r"\$?\s*\d[\d,]*(?:\.\d{2})")
HOURS_RE = re.compile(r"\b(?:hours?|hrs?|工时|时长)\b|\b\d+(?:\.\d+)?\s*(?:hours?|hrs?)\b", re.IGNORECASE)
NAME_COMMA_RE = re.compile(r"\b[A-Z][a-z]+,\s*[A-Z][a-z]+\b")
NAME_WORDS_RE = re.compile(r"\b[A-Z][a-z]+\s+[A-Z][a-z]+\b")
TOTAL_RE = re.compile(r"\b(?:invoice\s+total|total|amount\s+due|balance\s+due|合计|总计|总额)\b", re.IGNORECASE)


def summarize_business_signals(text: str) -> dict[str, Any]:
    content = str(text or "")
    money_matches = MONEY_RE.findall(content)
    name_matches = NAME_COMMA_RE.findall(content) + NAME_WORDS_RE.findall(content)
    hour_matches = HOURS_RE.findall(content)
    return {
        "textCharacterCount": len(content),
        "moneyCandidateCount": len(money_matches),
        "employeeNameSignalCount": len(set(name_matches)),
        "hoursSignalCount": len(hour_matches),
        "totalKeywordCount": len(TOTAL_RE.findall(content)),
        "hasAmountSignal": bool(money_matches),
        "hasEmployeeNameSignal": bool(name_matches),
        "hasHoursSignal": bool(hour_matches),
    }


def build_parser_comparison(
    pdf_paths: list[str | Path],
    *,
    expected_results: dict[str, Any] | None = None,
    parser_defs: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    paths = [Path(path).expanduser() for path in pdf_paths]
    selected_files = [path.name for path in paths]
    parser_defs = parser_defs or _parser_definitions()
    results: list[dict[str, Any]] = []
    for parser in parser_defs:
        for path in paths:
            results.append(_evaluate_parser_on_file(parser, path, expected_results=expected_results))
    usable = [
        row
        for row in results
        if row["available"] and row["signals"]["hasAmountSignal"] and row["signals"]["textCharacterCount"] > 0
    ]
    available_parser_count = len({row["parser"] for row in results if row["available"]})
    unavailable_parser_count = len(parser_defs) - available_parser_count
    if available_parser_count:
        next_step = "先用真实材料对比解析效果，再决定是否接入正式核对流程。"
    else:
        next_step = "当前本机未安装候选解析器。建议先安装轻量解析器（pdfplumber 或 PyMuPDF4LLM）后再跑真实材料对比；正式核对流程未改变。"
    accuracy_summary = _accuracy_summary(results)
    return {
        "source": "labor_parser_comparison",
        "summary": {
            "fileCount": len(paths),
            "parserCount": len(parser_defs),
            "availableParserCount": available_parser_count,
            "unavailableParserCount": unavailable_parser_count,
            "resultCount": len(results),
            "usableResultCount": len(usable),
            "totalFileCount": len(paths),
            "sampled": False,
            "sampleSize": len(paths),
            "selectedFiles": selected_files,
        },
        "accuracySummary": accuracy_summary,
        "parserResults": results,
        "recommendation": {
            "formalFlowChanged": False,
            "nextStep": next_step,
        },
    }


def build_parser_comparison_for_materials(
    materials_root: str | Path,
    sample_size: int | None = None,
    expected_results: dict[str, Any] | None = None,
) -> dict[str, Any]:
    root = Path(materials_root).expanduser()
    pdf_paths = sorted(path for path in root.rglob("*.pdf") if path.is_file())
    total_file_count = len(pdf_paths)
    if sample_size is not None and sample_size > 0:
        pdf_paths = pdf_paths[:sample_size]
    report = build_parser_comparison(pdf_paths, expected_results=expected_results)
    report["summary"]["totalFileCount"] = total_file_count
    report["summary"]["sampled"] = len(pdf_paths) < total_file_count
    report["summary"]["sampleSize"] = len(pdf_paths)
    report["summary"]["selectedFiles"] = [path.name for path in pdf_paths]
    return report


def build_expected_results_template_for_materials(materials_root: str | Path, sample_size: int | None = None) -> dict[str, Any]:
    root = Path(materials_root).expanduser()
    pdf_paths = _select_material_pdfs(root, sample_size=sample_size)
    total_file_count = len(_material_pdfs(root))
    return {
        "source": "labor_parser_expected_results_template",
        "summary": {
            "totalFileCount": total_file_count,
            "templateFileCount": len(pdf_paths),
            "sampled": len(pdf_paths) < total_file_count,
        },
        "instructions": [
            "填写人工确认后的 PDF 发票总额；不要直接复制程序输出。",
            "暂时只填写 invoice_total。员工明细准确性会在下一阶段扩展。",
        ],
        "files": {
            path.name: {
                "invoice_total": None,
                "reviewer": "",
                "reviewed_at": "",
                "evidence_note": "",
            }
            for path in pdf_paths
        },
    }


def write_parser_comparison_report(report: dict[str, Any], output_dir: str | Path) -> None:
    destination = Path(output_dir).expanduser()
    destination.mkdir(parents=True, exist_ok=True)
    (destination / "parser_comparison_summary.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (destination / "PARSER_COMPARISON_SUMMARY.md").write_text(
        _business_markdown(report),
        encoding="utf-8",
    )


def write_expected_results_excel_template(template: dict[str, Any], output_path: str | Path) -> None:
    destination = Path(output_path).expanduser()
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "人工确认发票总额"
    sheet.append(["PDF文件", "人工确认发票总额", "复核人", "复核日期", "证据备注"])
    files = template.get("files") if isinstance(template.get("files"), dict) else {}
    for file_name, item in files.items():
        row = item if isinstance(item, dict) else {}
        sheet.append(
            [
                file_name,
                row.get("invoice_total"),
                row.get("reviewer", ""),
                row.get("reviewed_at", ""),
                row.get("evidence_note", ""),
            ]
        )
    for column, width in {"A": 42, "B": 18, "C": 14, "D": 16, "E": 36}.items():
        sheet.column_dimensions[column].width = width
    workbook.save(destination)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Compare free PDF parsers on labor invoice materials")
    parser.add_argument("--materials-root", required=True)
    parser.add_argument("--output-dir", default="outputs/labor_parser_compare/latest")
    parser.add_argument("--sample-size", type=int, default=None)
    parser.add_argument("--expected-results", default="")
    parser.add_argument("--write-expected-template", default="")
    args = parser.parse_args(argv)
    if args.write_expected_template:
        template = build_expected_results_template_for_materials(args.materials_root, sample_size=args.sample_size)
        requested_template_path = Path(args.write_expected_template).expanduser()
        template_path = _expected_results_json_path(requested_template_path)
        template_path.parent.mkdir(parents=True, exist_ok=True)
        template_path.write_text(json.dumps(template, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        excel_template_path = _expected_results_excel_path(requested_template_path)
        write_expected_results_excel_template(template, excel_template_path)
        print(f"人工确认答案模板已生成：{template_path}")
        print(f"Excel 模板已生成：{excel_template_path}")
        return 0
    expected_results = _load_expected_results(args.expected_results)
    report = build_parser_comparison_for_materials(
        args.materials_root,
        sample_size=args.sample_size,
        expected_results=expected_results,
    )
    write_parser_comparison_report(report, args.output_dir)
    summary = report["summary"]
    recommendation = report["recommendation"]
    print(
        "\n".join(
            [
                _summary_file_count_text(summary),
                f"候选解析器：{summary['parserCount']} 个，当前可运行：{summary['availableParserCount']} 个",
                _expected_results_cli_text(report.get("accuracySummary")),
                f"报告已生成：{Path(args.output_dir).expanduser() / 'PARSER_COMPARISON_SUMMARY.md'}",
                str(recommendation["nextStep"]),
            ]
        )
    )
    return 0


def _parser_definitions() -> list[dict[str, Any]]:
    return [
        {"name": "LiteParse", "module": "liteparse", "runner": _unavailable_runner},
        {"name": "PyMuPDF4LLM", "module": "pymupdf4llm", "runner": _pymupdf4llm_text},
        {"name": "Docling", "module": "docling", "runner": _unavailable_runner},
        {"name": "Marker", "module": "marker", "runner": _unavailable_runner},
        {"name": "pdfplumber", "module": "pdfplumber", "runner": _pdfplumber_text},
        {"name": "Camelot", "module": "camelot", "runner": _camelot_text},
    ]


def _material_pdfs(root: Path) -> list[Path]:
    return sorted(path for path in root.rglob("*.pdf") if path.is_file())


def _select_material_pdfs(root: Path, *, sample_size: int | None = None) -> list[Path]:
    pdf_paths = _material_pdfs(root)
    if sample_size is not None and sample_size > 0:
        return pdf_paths[:sample_size]
    return pdf_paths


def _evaluate_parser_on_file(parser: dict[str, Any], path: Path, *, expected_results: dict[str, Any] | None = None) -> dict[str, Any]:
    available = _module_available(parser["module"])
    text = ""
    error = ""
    if available:
        try:
            text = str(parser["runner"](path) or "")
        except Exception as exc:  # noqa: BLE001 - comparison report must capture parser failures.
            error = str(exc)
    signals = summarize_business_signals(text)
    amount_candidates = _money_candidates(text)
    return {
        "parser": parser["name"],
        "module": parser["module"],
        "file": path.name,
        "available": available,
        "status": _business_status(available, signals, error),
        "error": error,
        "signals": signals,
        "amountCandidates": amount_candidates[:20],
        "accuracy": _accuracy_for_file(path.name, amount_candidates, expected_results),
    }


def _module_available(module_name: str) -> bool:
    return importlib.util.find_spec(module_name) is not None


def _business_status(available: bool, signals: dict[str, Any], error: str) -> str:
    if not available:
        return "本机未安装"
    if error:
        return "解析失败"
    if not signals["textCharacterCount"]:
        return "未读出文本"
    if signals["hasAmountSignal"] and signals["hasEmployeeNameSignal"] and signals["hasHoursSignal"]:
        return "可继续评估员工明细"
    if signals["hasAmountSignal"]:
        return "可继续评估总金额"
    return "暂未发现有效线索"


def _pdfplumber_text(path: Path) -> str:
    import pdfplumber  # type: ignore

    chunks: list[str] = []
    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages[:3]:
            chunks.append(page.extract_text() or "")
    return "\n".join(chunks)


def _pymupdf4llm_text(path: Path) -> str:
    import fitz  # type: ignore
    import pymupdf4llm  # type: ignore

    with fitz.open(str(path)) as doc:
        pages = list(range(min(3, doc.page_count)))
    if not pages:
        return ""
    return str(pymupdf4llm.to_markdown(str(path), pages=pages) or "")


def _camelot_text(path: Path) -> str:
    import camelot  # type: ignore

    tables = camelot.read_pdf(str(path), pages="1-3")
    return "\n".join(table.df.to_csv(index=False) for table in tables)


def _unavailable_runner(path: Path) -> str:
    _ = path
    return ""


def _business_markdown(report: dict[str, Any]) -> str:
    rows = report.get("parserResults") if isinstance(report.get("parserResults"), list) else []
    summary = report.get("summary") if isinstance(report.get("summary"), dict) else {}
    recommendation = report.get("recommendation") if isinstance(report.get("recommendation"), dict) else {}
    parser_names = sorted({str(row.get("parser", "")) for row in rows if row.get("parser")})
    lines = [
        "# 免费解析器对比业务摘要",
        "",
        "正式核对流程未改变。本报告只用于判断哪些免费解析器值得继续接入测试。",
        "",
        "## 当前结论",
        "",
        f"- {_summary_file_count_text(summary)}",
        f"- 候选解析器：{summary.get('parserCount', 0)} 个",
        f"- 当前可直接运行：{summary.get('availableParserCount', 0)} 个",
        f"- 候选名单：{'、'.join(parser_names) if parser_names else '暂无'}",
        "",
        str(recommendation.get("nextStep") or "先用真实材料对比解析效果，再决定是否接入正式核对流程。"),
        "",
    ]
    if summary.get("sampled"):
        lines.extend(
            [
                "本报告是抽样结果，用于快速判断方向；不能替代全量真实材料复测。",
                "",
            ]
        )
    if not summary.get("availableParserCount"):
        lines.extend(
            [
                "## 对比结果",
                "",
                "当前没有可运行的候选解析器，暂不展开逐文件明细。",
                "",
                "需要先安装至少一个轻量解析器，再用这些真实 PDF 判断它是否能读出金额、姓名和工时。",
                "",
                "## 下一步",
                "",
                "建议先安装 pdfplumber 或 PyMuPDF4LLM，跑完真实材料对比后再决定是否接入正式核对流程。",
                "",
            ]
        )
        return "\n".join(lines)
    parser_summaries = _summarize_parser_rows(rows)
    accuracy_summary = report.get("accuracySummary") if isinstance(report.get("accuracySummary"), dict) else {}
    if accuracy_summary.get("expectedFileCount"):
        lines.extend(
            [
                "## 已确认答案对照",
                "",
                f"- 有人工确认答案的文件：{int(accuracy_summary.get('expectedFileCount') or 0)} 个",
                f"- 已确认总额命中：{int(accuracy_summary.get('invoiceTotalMatchedFileCount') or 0)} 个",
                f"- 已确认总额未命中：{int(accuracy_summary.get('invoiceTotalMissedFileCount') or 0)} 个",
                "",
            ]
        )
    lines.extend(
        [
            "## 解析器表现",
            "",
            "| 解析器 | 可运行文件 | 读出金额 | 人工确认总额命中 | 读出员工姓名 | 读出工时 |",
            "| --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for item in parser_summaries:
        if not item["availableFileCount"]:
            continue
        confirmed_total = (
            f"{item['invoiceTotalMatchedFileCount']} / {item['expectedInvoiceTotalFileCount']}"
            if item["expectedInvoiceTotalFileCount"]
            else "-"
        )
        lines.append(
            "| {parser} | {available} | {amount} | {confirmed_total} | {name} | {hours} |".format(
                parser=item["parser"],
                available=item["availableFileCount"],
                amount=item["amountSignalFileCount"],
                confirmed_total=confirmed_total,
                name=item["nameSignalFileCount"],
                hours=item["hoursSignalFileCount"],
            )
        )
    lines.extend(
        [
            "",
            "未安装的解析器不展开逐文件明细；先看已安装解析器在真实材料上的有效信号。",
            "",
            "## 可继续复测的样例",
            "",
            "| 解析器 | 文件 | 可读内容 |",
            "| --- | --- | --- |",
        ]
    )
    sample_count = 0
    for row in rows:
        if sample_count >= 10:
            break
        signals = row.get("signals") if isinstance(row.get("signals"), dict) else {}
        if not row.get("available") or not signals.get("hasAmountSignal"):
            continue
        readable_parts = []
        if signals.get("hasAmountSignal"):
            readable_parts.append("金额")
        if signals.get("hasEmployeeNameSignal"):
            readable_parts.append("员工姓名")
        if signals.get("hasHoursSignal"):
            readable_parts.append("工时")
        lines.append(f"| {row.get('parser', '')} | {row.get('file', '')} | {'、'.join(readable_parts)} |")
        sample_count += 1
    lines.extend(
        [
            "",
            "## 下一步",
            "",
            _parser_recommendation_text(parser_summaries),
            "",
        ]
    )
    return "\n".join(lines)


def _summarize_parser_rows(rows: list[Any]) -> list[dict[str, Any]]:
    summaries: dict[str, dict[str, Any]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        parser_name = str(row.get("parser") or "")
        if not parser_name:
            continue
        summary = summaries.setdefault(
            parser_name,
            {
                "parser": parser_name,
                "availableFileCount": 0,
                "amountSignalFileCount": 0,
                "nameSignalFileCount": 0,
                "hoursSignalFileCount": 0,
                "expectedInvoiceTotalFileCount": 0,
                "invoiceTotalMatchedFileCount": 0,
            },
        )
        if not row.get("available"):
            continue
        signals = row.get("signals") if isinstance(row.get("signals"), dict) else {}
        accuracy = row.get("accuracy") if isinstance(row.get("accuracy"), dict) else {}
        summary["availableFileCount"] += 1
        if signals.get("hasAmountSignal"):
            summary["amountSignalFileCount"] += 1
        if signals.get("hasEmployeeNameSignal"):
            summary["nameSignalFileCount"] += 1
        if signals.get("hasHoursSignal"):
            summary["hoursSignalFileCount"] += 1
        if accuracy.get("hasExpectedResult"):
            summary["expectedInvoiceTotalFileCount"] += 1
            if accuracy.get("invoiceTotalMatched"):
                summary["invoiceTotalMatchedFileCount"] += 1
    return sorted(summaries.values(), key=lambda item: (-item["amountSignalFileCount"], item["parser"]))


def _parser_recommendation_text(parser_summaries: list[dict[str, Any]]) -> str:
    usable = [item for item in parser_summaries if item.get("availableFileCount")]
    if not usable:
        return "建议先安装 pdfplumber 或 PyMuPDF4LLM，跑完真实材料对比后再决定是否接入正式核对流程。"
    best = sorted(
        usable,
        key=lambda item: (
            -int(item.get("amountSignalFileCount", 0)),
            -int(item.get("nameSignalFileCount", 0)),
            -int(item.get("hoursSignalFileCount", 0)),
            str(item.get("parser", "")),
        ),
    )[0]
    best_score = (
        int(best.get("amountSignalFileCount", 0)),
        int(best.get("nameSignalFileCount", 0)),
        int(best.get("hoursSignalFileCount", 0)),
    )
    tied = [
        item
        for item in usable
        if (
            int(item.get("amountSignalFileCount", 0)),
            int(item.get("nameSignalFileCount", 0)),
            int(item.get("hoursSignalFileCount", 0)),
        )
        == best_score
    ]
    if len(tied) > 1:
        names = "、".join(sorted(str(item["parser"]) for item in tied))
        return f"建议同时复测 {names}，它们在当前材料里的线索表现并列；正式核对流程暂时不变。"
    return (
        f"建议优先复测 {best['parser']}，因为它在当前材料里读出金额线索最多；"
        "再用另一款已安装解析器做补充对照。正式核对流程暂时不变。"
    )


def _summary_file_count_text(summary: dict[str, Any]) -> str:
    if summary.get("sampled"):
        return f"抽样检查 PDF 文件：{summary.get('fileCount', 0)} / {summary.get('totalFileCount', 0)} 个"
    return f"已检查 PDF 文件：{summary.get('fileCount', 0)} 个"


def _load_expected_results(path: str | Path) -> dict[str, Any] | None:
    if not path:
        return None
    source = Path(path).expanduser()
    if not source.exists():
        raise FileNotFoundError(f"expected results file does not exist: {source}")
    if source.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_expected_results_workbook(source)
    return json.loads(source.read_text(encoding="utf-8"))


def _load_expected_results_workbook(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["人工确认发票总额"] if "人工确认发票总额" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"files": {}}
    headers = [str(value or "").strip() for value in rows[0]]
    file_index = _header_index(headers, ["PDF文件", "文件", "File", "file"])
    total_index = _header_index(headers, ["人工确认发票总额", "发票总额", "invoice_total", "Invoice Total"])
    reviewer_index = _header_index(headers, ["复核人", "reviewer"])
    reviewed_at_index = _header_index(headers, ["复核日期", "reviewed_at"])
    note_index = _header_index(headers, ["证据备注", "evidence_note", "备注"])
    files: dict[str, dict[str, Any]] = {}
    if file_index is None or total_index is None:
        return {"files": files}
    for row in rows[1:]:
        file_name = str(row[file_index] or "").strip() if file_index < len(row) else ""
        if not file_name:
            continue
        files[file_name] = {
            "invoice_total": _optional_float(row[total_index] if total_index < len(row) else None),
            "reviewer": str(row[reviewer_index] or "").strip() if reviewer_index is not None and reviewer_index < len(row) else "",
            "reviewed_at": str(row[reviewed_at_index] or "").strip() if reviewed_at_index is not None and reviewed_at_index < len(row) else "",
            "evidence_note": str(row[note_index] or "").strip() if note_index is not None and note_index < len(row) else "",
        }
    return {"files": files}


def _expected_results_excel_path(template_path: Path) -> Path:
    if template_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return template_path
    return template_path.with_suffix(".xlsx")


def _expected_results_json_path(template_path: Path) -> Path:
    if template_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return template_path.with_suffix(".json")
    return template_path


def _header_index(headers: list[str], candidates: list[str]) -> int | None:
    normalized = {header.lower().replace(" ", "").replace("_", ""): index for index, header in enumerate(headers)}
    for candidate in candidates:
        key = candidate.lower().replace(" ", "").replace("_", "")
        if key in normalized:
            return normalized[key]
    return None


def _optional_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(str(value).replace(",", "")), 2)
    except (TypeError, ValueError):
        return None


def _expected_results_cli_text(accuracy_summary: Any) -> str:
    if not isinstance(accuracy_summary, dict) or not accuracy_summary.get("expectedFileCount"):
        return "未载入人工确认答案。"
    return f"已载入人工确认答案：{accuracy_summary.get('expectedFileCount')} 个文件"


def _money_candidates(text: str) -> list[float]:
    candidates: list[float] = []
    for raw in MONEY_RE.findall(str(text or "")):
        normalized = raw.replace("$", "").replace(",", "").strip()
        try:
            candidates.append(round(float(normalized), 2))
        except ValueError:
            continue
    return candidates


def _accuracy_for_file(file_name: str, amount_candidates: list[float], expected_results: dict[str, Any] | None) -> dict[str, Any]:
    expected = _expected_for_file(file_name, expected_results)
    if not expected:
        return {"hasExpectedResult": False}
    invoice_total = expected.get("invoice_total")
    if not isinstance(invoice_total, (int, float)):
        return {"hasExpectedResult": False}
    matched = any(abs(candidate - float(invoice_total)) <= 0.1 for candidate in amount_candidates)
    return {
        "hasExpectedResult": True,
        "expectedInvoiceTotal": round(float(invoice_total), 2),
        "invoiceTotalMatched": matched,
    }


def _expected_for_file(file_name: str, expected_results: dict[str, Any] | None) -> dict[str, Any]:
    if not isinstance(expected_results, dict):
        return {}
    files = expected_results.get("files") if isinstance(expected_results.get("files"), dict) else {}
    expected = files.get(file_name)
    return expected if isinstance(expected, dict) else {}


def _accuracy_summary(results: list[dict[str, Any]]) -> dict[str, Any]:
    expected_files = set()
    matched_files = set()
    missed_files = set()
    for row in results:
        accuracy = row.get("accuracy") if isinstance(row.get("accuracy"), dict) else {}
        if not accuracy.get("hasExpectedResult"):
            continue
        file_name = str(row.get("file") or "")
        expected_files.add(file_name)
        if accuracy.get("invoiceTotalMatched") is True:
            matched_files.add(file_name)
        elif accuracy.get("invoiceTotalMatched") is False:
            missed_files.add(file_name)
    return {
        "expectedFileCount": len(expected_files),
        "invoiceTotalMatchedFileCount": len(matched_files),
        "invoiceTotalMissedFileCount": len(missed_files - matched_files),
    }


if __name__ == "__main__":
    raise SystemExit(main())
