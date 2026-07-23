from __future__ import annotations

import csv
from pathlib import Path
import re
from typing import Any, Dict, List

from .models import LaborLineItem
from .parsing import display_name, parse_number


def _is_xls(path: Path) -> bool:
    """判断是否为旧版 .xls 格式"""
    return path.suffix.lower() == ".xls"


NAME_KEYWORDS = ("姓名", "员工姓名", "员工名称", "name", "employee", "associate")
ID_KEYWORDS = ("工号", "员工id", "employee id", "employee number", "employee_id", "id")
HOURS_KEYWORDS = ("时长", "工时", "hours", "hour", "time", "accounting time")
AMOUNT_KEYWORDS = ("费用", "金额", "合计", "总额", "薪资", "工资", "amount", "total", "pay", "cost")
AMOUNT_PREFERRED_KEYWORDS = ("费用总计(含税)", "含税", "total", "amount")
CURRENCY_KEYWORDS = ("币种", "currency")
REOCR_NAME_KEYWORDS = ("姓名", "员工", "employee", "name", "associate")
REOCR_ID_KEYWORDS = ("工号", "员工id", "employeeid", "employee id", "employee number", "employee_id")
REOCR_HOURS_KEYWORDS = ("工时", "时长", "hours", "hour", "time")
REOCR_AMOUNT_KEYWORDS = ("金额", "费用", "amount", "total", "pay", "cost")
REOCR_PAGE_KEYWORDS = ("页", "page", "source", "row")
REOCR_CONFIDENCE_KEYWORDS = ("置信", "confidence", "score")
REOCR_EVIDENCE_KEYWORDS = ("证据", "原文", "evidence", "text")
REOCR_SOURCE_FILE_KEYWORDS = ("sourcefile", "source file", "pdf", "file", "文件")
REOCR_WAREHOUSE_KEYWORDS = ("warehouseid", "warehouse id", "warehouse", "仓库", "仓")


def list_workbook_sheets(path: Path) -> List[str]:
    if _is_xls(path):
        import xlrd
        workbook = xlrd.open_workbook(str(path))
        return workbook.sheet_names()
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def suggest_mapping(path: Path, sheet_name: str) -> Dict[str, Any]:
    sheet, rows = _sheet_rows(path, sheet_name, max_rows=21)
    if not rows:
        raise ValueError("Excel 工作表为空，无法识别字段。")
    header_row_index = _detect_header_row(rows)
    headers = [display_name(value) for value in rows[header_row_index]]
    preview = [_row_dict(headers, row) for row in rows[header_row_index + 1:]]
    return {
        "sheetName": getattr(sheet, "title", None) or getattr(sheet, "name", sheet_name),
        "headers": headers,
        "suggestedMapping": {
            "employeeId": _employee_id_header(headers),
            "name": _name_header(headers),
            "hours": _hours_header(headers),
            "amount": _amount_header(headers),
            "currency": _first_header(headers, CURRENCY_KEYWORDS),
        },
        "amountColumnCandidates": _amount_component_headers(headers),
        "previewRows": [row for row in preview if any(value not in (None, "") for value in row.values())][:20],
    }


def read_workbook_rows(path: Path, sheet_name: str, mapping: Dict[str, Any]) -> List[LaborLineItem]:
    _validate_mapping(mapping)
    sheet, rows = _sheet_rows(path, sheet_name, max_rows=None)
    if not rows:
        raise ValueError("Excel 工作表为空，无法读取线下账单。")
    header_row_index = _detect_header_row(rows)
    headers = [display_name(value) for value in rows[header_row_index]]
    index = {header: position for position, header in enumerate(headers)}
    for required in ("name", "hours", "amount"):
        if mapping[required] not in index:
            raise ValueError(f"字段映射无效：找不到 {mapping[required]}")
    amount_columns = _mapped_amount_columns(mapping)
    for header in amount_columns:
        if header not in index:
            raise ValueError(f"字段映射无效：找不到 {header}")
    amount_breakdown_headers = _amount_breakdown_headers(headers, amount_columns)
    amount_context_headers = _amount_context_headers(headers)
    # 收集所有列名（用于表头检测）
    header_names = set()
    for col_name in mapping.values():
        if isinstance(col_name, list):
            header_names.update(str(item).strip().lower() for item in col_name if str(item).strip())
            continue
        if col_name:
            header_names.add(str(col_name).strip().lower())

    result: List[LaborLineItem] = []
    for offset, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        name = _value(row, index[mapping["name"]])
        # 检测并跳过表头行
        name_str = str(name).strip() if name is not None else ""
        if name_str and name_str.lower() in header_names:
            continue
        if _is_summary_name(name_str):
            continue
        if name in (None, ""):
            continue
        hours = parse_number(_value(row, index[mapping["hours"]]))
        amount_components = {
            header: parse_number(_value(row, index[header]))
            for header in amount_columns
        }
        amount = round(sum(amount_components.values()), 2)
        amount_breakdown = {
            header: parse_number(_value(row, index[header]))
            for header in amount_breakdown_headers
            if abs(parse_number(_value(row, index[header]))) > 0.0005
        }
        amount_context = {
            header: display_name(_value(row, index[header]))
            for header in amount_context_headers
            if display_name(_value(row, index[header]))
        }
        if hours == 0 and amount == 0:
            continue
        currency = ""
        employee_id = ""
        if mapping.get("employeeId") and mapping["employeeId"] in index:
            employee_id = display_name(_value(row, index[mapping["employeeId"]]))
        if mapping.get("currency") and mapping["currency"] in index:
            currency = display_name(_value(row, index[mapping["currency"]]))
        warehouse_id = ""
        for header, position in index.items():
            if _is_warehouse_header(header):
                raw_wh = _value(row, position)
                warehouse_id = _warehouse_id_from_value(raw_wh)
                break
        result.append(
            LaborLineItem(
                source_type="offline_workbook",
                source_file=path.name,
                source_page_or_row=f"{getattr(sheet, 'title', None) or getattr(sheet, 'name', sheet_name)}!{offset}",
                employee_id=employee_id,
                employee_name_raw=display_name(name),
                hours=round(hours, 2),
                amount=round(amount, 2),
                currency=currency,
                confidence=1.0,
                evidence_text="",
                warehouse_id=warehouse_id,
                amount_components=amount_components,
                amount_breakdown=amount_breakdown,
                amount_context=amount_context,
            )
        )
    return result


def parse_reocr_candidate_rows(path: Path, *, default_currency: str = "") -> List[Dict[str, Any]]:
    """Parse user-uploaded OCR candidate rows without promoting them to official results."""
    rows = _rows_as_dicts_from_first_sheet_or_csv(path)
    if not rows:
        raise ValueError("图片识别结果文件为空。")
    headers = [header for header in rows[0].keys() if not header.startswith("__")]
    employee_id_header = _first_header(headers, REOCR_ID_KEYWORDS)
    name_header = _first_header([header for header in headers if header != employee_id_header], REOCR_NAME_KEYWORDS)
    hours_header = _first_header(headers, REOCR_HOURS_KEYWORDS)
    amount_header = _first_header(headers, REOCR_AMOUNT_KEYWORDS)
    source_file_header = _first_header(headers, REOCR_SOURCE_FILE_KEYWORDS)
    page_header = _first_header([header for header in headers if header != source_file_header], REOCR_PAGE_KEYWORDS)
    warehouse_header = _first_header(headers, REOCR_WAREHOUSE_KEYWORDS)
    confidence_header = _first_header(headers, REOCR_CONFIDENCE_KEYWORDS)
    evidence_header = _first_header(headers, REOCR_EVIDENCE_KEYWORDS)
    currency_header = _first_header(headers, CURRENCY_KEYWORDS)
    if not name_header or not amount_header:
        raise ValueError("图片识别结果文件至少需要包含员工姓名列和金额列。")

    parsed = []
    for row in rows:
        name = display_name(row.get(name_header))
        amount = parse_number(row.get(amount_header))
        hours = parse_number(row.get(hours_header)) if hours_header else 0.0
        if not name and amount == 0 and hours == 0:
            continue
        if not name:
            continue
        parsed_row = {
            "employeeName": name,
            "sourcePageOrRow": display_name(row.get(page_header)) if page_header else f"row{row.get('__row_number') or len(parsed) + 1}",
            "hours": round(hours, 2),
            "amount": round(amount, 2),
            "currency": display_name(row.get(currency_header)) if currency_header else default_currency,
            "confidence": _candidate_confidence(row.get(confidence_header)) if confidence_header else 0.95,
            "evidenceText": display_name(row.get(evidence_header)) if evidence_header else _candidate_evidence(row),
        }
        source_file = display_name(row.get(source_file_header)) if source_file_header else ""
        warehouse_id = _warehouse_id_from_value(row.get(warehouse_header)) if warehouse_header else ""
        employee_id = display_name(row.get(employee_id_header)) if employee_id_header else ""
        if source_file:
            parsed_row["sourceFile"] = source_file
        if warehouse_id:
            parsed_row["warehouseId"] = warehouse_id
        if employee_id:
            parsed_row["employeeId"] = employee_id
        parsed.append(parsed_row)
    if not parsed:
        raise ValueError("图片识别结果文件未解析出有效员工明细。")
    return parsed


def summarize_otws_costs(path: Path) -> Dict[str, Any]:
    """解析 OTWS 仓库账单费用组成，用于解释 SSS 类账单口径差异。"""
    sheets = set(list_workbook_sheets(path))
    if "Warehouse-information" not in sheets:
        raise ValueError("找不到工作表：Warehouse-information")

    warehouse_rows = _rows_as_dicts(path, "Warehouse-information")
    if not warehouse_rows:
        raise ValueError("Warehouse-information 工作表为空，无法解析费用组成。")

    primary = warehouse_rows[0]
    warehouse_raw = display_name(primary.get("Physical warehouse"))
    warehouse_id = _warehouse_id_from_value(warehouse_raw)
    components = {
        "hourlySalary": parse_number(primary.get("Total hourly salary")),
        "bonus": parse_number(primary.get("Total bonus")),
        "vehicleCompensation": parse_number(primary.get("Total vehicle compensation")),
        "mealSupplement": parse_number(primary.get("Total Meal Supplement")),
        "hourlyRateDifference": parse_number(primary.get("Hourly Rate Difference")),
        "employmentInsurance": parse_number(primary.get("Lot And Scot(Employment Insurance)")),
        "incomeTax": parse_number(primary.get("Income Tax")),
        "otherExpenses": parse_number(primary.get("Total other expenses")),
        "handlingFee": parse_number(primary.get("total.handling.fee")),
        "additionalFees": parse_number(primary.get("additional fees")),
    }
    summary_total = parse_number(primary.get("Total"))
    component_total = round(sum(components.values()), 2)

    employee_detail = _sum_detail_sheet(
        path,
        "Employee-expenses-detail",
        amount_header="Total cost",
        hours_header="Total staff cost accounting time",
    )
    benefit_detail = _sum_detail_sheet(path, "Employee-benefits-detail", amount_header="Total cost")
    loading_detail = _sum_detail_sheet(path, "The-loading-and-unloading-of-ta", amount_header="Total cost")
    detail_total = round(employee_detail["amount"] + benefit_detail["amount"] + loading_detail["amount"], 2)

    return {
        "sourceFile": path.name,
        "warehouseId": warehouse_id,
        "warehouseName": warehouse_raw,
        "supplier": display_name(primary.get("Company Name")),
        "currency": display_name(primary.get("currency")),
        "periodStart": display_name(primary.get("Accounting start date")),
        "periodEnd": display_name(primary.get("Accounting end date")),
        "employeeCount": int(parse_number(primary.get("Total employees during attendance period"))),
        "summary": {
            "components": {key: round(value, 2) for key, value in components.items()},
            "componentTotal": component_total,
            "reportedTotal": round(summary_total, 2),
            "componentDelta": round(summary_total - component_total, 2),
            "evidence": "Warehouse-information!2",
        },
        "details": {
            "employeeExpenses": employee_detail,
            "employeeBenefits": benefit_detail,
            "loadingAndUnloading": loading_detail,
            "detailTotal": detail_total,
            "summaryDelta": round(summary_total - detail_total, 2),
        },
    }


def _sheet_rows(path: Path, sheet_name: str, max_rows: int | None) -> tuple[Any, List[tuple[Any, ...]]]:
    if _is_xls(path):
        return _sheet_rows_xls(path, sheet_name, max_rows)
    return _sheet_rows_xlsx(path, sheet_name, max_rows)


def _sheet_rows_xlsx(path: Path, sheet_name: str, max_rows: int | None) -> tuple[Any, List[tuple[Any, ...]]]:
    from openpyxl import load_workbook
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"找不到工作表：{sheet_name}")
        sheet = workbook[sheet_name]
        if hasattr(sheet, "reset_dimensions"):
            sheet.reset_dimensions()
        iterator = sheet.iter_rows(values_only=True, max_row=max_rows)
        rows = [row for row in iterator if any(value not in (None, "") for value in row)]
        return sheet, rows
    finally:
        workbook.close()


def _sheet_rows_xls(path: Path, sheet_name: str, max_rows: int | None) -> tuple[Any, List[tuple[Any, ...]]]:
    import xlrd
    workbook = xlrd.open_workbook(str(path))
    if sheet_name not in workbook.sheet_names():
        raise ValueError(f"找不到工作表：{sheet_name}")
    sheet = workbook.sheet_by_name(sheet_name)
    limit = max_rows if max_rows else sheet.nrows
    rows = []
    for i in range(min(limit, sheet.nrows)):
        row = tuple(sheet.cell_value(i, j) for j in range(sheet.ncols))
        if any(value not in (None, "") for value in row):
            rows.append(row)
    return sheet, rows


def _first_header(headers: List[str], keywords: tuple[str, ...]) -> str:
    for header in headers:
        lowered = header.lower()
        if any(keyword.lower() in lowered for keyword in keywords):
            return header
    return ""


def _detect_header_row(rows: List[tuple[Any, ...]]) -> int:
    best_index = 0
    best_score = -1
    for idx, row in enumerate(rows[:10]):
        headers = [display_name(value) for value in row]
        lowered = [header.lower() for header in headers if header]
        score = 0
        if _name_header(headers):
            score += 4
        if _hours_header(headers):
            score += 3
        if _amount_header(headers):
            score += 3
        if _employee_id_header(headers):
            score += 1
        if any("warehouse" in header or "仓" in header for header in lowered):
            score += 1
        if score > best_score or (score == best_score and score > 0):
            best_score = score
            best_index = idx
    return best_index


def _name_header(headers: List[str]) -> str:
    exact = {"姓名", "员工姓名", "name", "employee name", "employee_name", "associate name", "worker name"}
    for header in headers:
        if header.lower() in exact:
            return header
    for header in headers:
        lowered = header.lower()
        if "company" in lowered or "supplier" in lowered:
            continue
        if any(keyword.lower() in lowered for keyword in NAME_KEYWORDS):
            return header
    return ""


def _hours_header(headers: List[str]) -> str:
    preferred = (
        "total staff cost accounting time",
        "时长总计(h)",
        "total hours",
        "hours total",
    )
    for target in preferred:
        for header in headers:
            if header.lower() == target:
                return header
    if any(header.strip() == "总计" for header in headers) and any(_is_rate_header(header) for header in headers):
        return next(header for header in headers if header.strip() == "总计")
    return _first_header(headers, HOURS_KEYWORDS)


def _preferred_amount_header(headers: List[str]) -> str:
    for target in ("Total cost", "total cost"):
        for header in headers:
            if header == target:
                return header
    for header in headers:
        if header.strip() == "总额":
            return header
    for header in headers:
        if "不含税" in header:
            return header
    for header in headers:
        if "含税" in header and "不含税" not in header:
            return header
    return _first_header(headers, AMOUNT_PREFERRED_KEYWORDS)


def _amount_header(headers: List[str]) -> str:
    preferred = _preferred_amount_header(headers)
    if preferred and not _is_hours_measure_header(preferred):
        return preferred
    return _first_header([header for header in headers if not _is_hours_measure_header(header)], AMOUNT_KEYWORDS)


def _amount_component_headers(headers: List[str]) -> List[str]:
    monetary_signal = re.compile(
        r"(?:金额|费用|薪资|工资|餐补|补贴|奖金|交通|合计|总额|amount|total|salary|pay|cost|fee|bonus|allowance|meal|transport|expense)",
        re.IGNORECASE,
    )
    rate_signal = re.compile(r"(?:时薪|单价|费率|hourly\s+rate|bill\s+rate|pay\s+rate|unit\s+price)", re.IGNORECASE)
    return [
        header
        for header in headers
        if monetary_signal.search(header)
        and not rate_signal.search(header)
        and not _is_hours_measure_header(header)
    ]


def _amount_breakdown_headers(headers: List[str], mapped_amount_columns: List[str]) -> List[str]:
    mapped = {str(header or "").strip() for header in mapped_amount_columns}
    return [
        header
        for header in _amount_component_headers(headers)
        if header not in mapped and not _is_aggregate_amount_header(header)
    ]


def _is_aggregate_amount_header(header: str) -> bool:
    text = str(header or "").strip()
    return bool(
        re.search(
            r"(?:费用总计|非固费用总计|总费用|总金额|金额合计|成本总计|总成本|grand\s+total|total\s+(?:cost|amount|pay)|(?:cost|amount|pay)\s+total)",
            text,
            re.IGNORECASE,
        )
    )


def _amount_context_headers(headers: List[str]) -> List[str]:
    signal = re.compile(r"(?:备注|说明|描述|摘要|note|remark|memo|comment|description)", re.IGNORECASE)
    return [header for header in headers if signal.search(str(header or ""))]


def _is_hours_measure_header(header: str) -> bool:
    lowered = header.lower()
    if "时长" in header or "工时" in header:
        return True
    if "accounting time" in lowered or "total time" in lowered:
        return True
    return bool(re.search(r"\bhours?\b", lowered))


def _is_rate_header(header: str) -> bool:
    return bool(
        re.search(
            r"(?:时薪|单价|费率|hourly\s+rate|bill\s+rate|pay\s+rate|unit\s+price)",
            str(header or ""),
            re.IGNORECASE,
        )
    )


def _employee_id_header(headers: List[str]) -> str:
    exact = {"工号", "员工工号", "Employee ID", "employee id", "Employee number", "employee number"}
    for header in headers:
        if header in exact:
            return header
    for header in headers:
        lowered = header.lower()
        if "供应商" in header:
            continue
        if any(keyword.lower() in lowered for keyword in ID_KEYWORDS):
            return header
    return ""


def _is_warehouse_header(header: str) -> bool:
    lowered = header.lower()
    return header in {"物理仓", "仓库"} or "warehouse" in lowered


def _warehouse_id_from_value(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"\d{1,3}", text):
        return text
    patterns = (
        r"(\d+)号仓",
        r"\bNJ\s*[- ]?\s*(\d{1,3})\b",
        r"\bNew\s+Jersey(?:\s+Warehouse)?\s*-?\s*(\d{1,3})\b",
        r"\bWarehouse\s*-?\s*(\d{1,3})\b",
        r"\bWH\s*-?\s*(\d{1,3})\b",
    )
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def _rows_as_dicts(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    _, rows = _sheet_rows(path, sheet_name, max_rows=None)
    if not rows:
        return []
    header_row_index = _detect_header_row(rows)
    headers = [display_name(value) for value in rows[header_row_index]]
    result = []
    for offset, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
        item = _row_dict(headers, row)
        if any(value not in (None, "") for value in item.values()):
            item["__row_number"] = offset
            result.append(item)
    return result


def _rows_as_dicts_from_first_sheet_or_csv(path: Path) -> List[Dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        return _csv_rows_as_dicts(path)
    sheets = list_workbook_sheets(path)
    if not sheets:
        return []
    return _rows_as_dicts(path, sheets[0])


def _csv_rows_as_dicts(path: Path) -> List[Dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig")
    reader = csv.DictReader(text.splitlines())
    result = []
    for index, row in enumerate(reader, start=2):
        item = {display_name(key): value for key, value in row.items() if key is not None}
        if any(value not in (None, "") for value in item.values()):
            item["__row_number"] = index
            result.append(item)
    return result


def _candidate_confidence(value: Any) -> float:
    number = parse_number(value)
    if number > 1:
        number = number / 100
    if number <= 0:
        return 0.95
    return round(min(number, 1.0), 4)


def _candidate_evidence(row: Dict[str, Any]) -> str:
    parts = []
    for key, value in row.items():
        if key.startswith("__") or value in (None, ""):
            continue
        parts.append(f"{key}: {display_name(value)}")
    return " | ".join(parts)[:500]


def _sum_detail_sheet(
    path: Path,
    sheet_name: str,
    amount_header: str,
    hours_header: str | None = None,
) -> Dict[str, Any]:
    if sheet_name not in set(list_workbook_sheets(path)):
        return {"amount": 0.0, "hours": 0.0, "rowCount": 0, "evidence": ""}

    rows = _rows_as_dicts(path, sheet_name)
    amount = 0.0
    hours = 0.0
    row_count = 0
    first_row = ""
    last_row = ""
    for idx, row in enumerate(rows, start=1):
        if _is_total_row(row):
            continue
        row_amount = parse_number(row.get(amount_header))
        row_hours = parse_number(row.get(hours_header)) if hours_header else 0.0
        if row_amount == 0 and row_hours == 0 and not _has_identity_value(row):
            continue
        amount += row_amount
        hours += row_hours
        row_count += 1
        if not first_row:
            first_row = f"{sheet_name}!{int(row.get('__row_number') or idx)}"
        last_row = f"{sheet_name}!{int(row.get('__row_number') or idx)}"

    evidence = ""
    if first_row and last_row:
        evidence = first_row if first_row == last_row else f"{first_row}:{last_row}"
    return {
        "amount": round(amount, 2),
        "hours": round(hours, 2),
        "rowCount": row_count,
        "evidence": evidence,
    }


def _is_total_row(row: Dict[str, Any]) -> bool:
    for value in row.values():
        text = display_name(value).lower()
        if text in {"total:", "total", "合计"}:
            return True
    return False


def _is_summary_name(value: Any) -> bool:
    text = display_name(value).strip().casefold()
    text = re.sub(r"[\s:：]+$", "", text)
    return text in {"total", "totals", "subtotal", "sub total", "grand total", "合计", "总计", "小计"}


def _has_identity_value(row: Dict[str, Any]) -> bool:
    for header, value in row.items():
        lowered = header.lower()
        if any(token in lowered for token in ("employee", "name", "number", "warehouse", "group")):
            if display_name(value):
                return True
    return False


def _row_dict(headers: List[str], row: tuple[Any, ...]) -> Dict[str, Any]:
    return {header: _value(row, index) for index, header in enumerate(headers) if header}


def _value(row: tuple[Any, ...], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def _validate_mapping(mapping: Dict[str, Any]) -> None:
    missing = [field for field in ("name", "hours", "amount") if not mapping.get(field)]
    if missing:
        raise ValueError("字段映射缺少姓名、工时或金额，无法比对。")


def _mapped_amount_columns(mapping: Dict[str, Any]) -> List[str]:
    configured = mapping.get("amountColumns") or mapping.get("amount_columns") or []
    if isinstance(configured, str):
        configured = [configured]
    primary = str(mapping.get("amount") or "").strip()
    result: List[str] = []
    for value in [primary, *configured]:
        header = str(value or "").strip()
        if header and header not in result:
            result.append(header)
    return result
