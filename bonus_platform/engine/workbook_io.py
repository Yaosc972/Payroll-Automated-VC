from __future__ import annotations

from datetime import datetime
import re
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import CalculationResult, ImportRow, RecruitmentDetail, as_text


IMPORT_SHEET = "导入_月度数据"
DATE_FORMAT = "yyyy/m/d"
MONEY_FORMAT = '#,##0.00'
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
EXCEPTION_FILL = PatternFill("solid", fgColor="FFD966")
LIGHT_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BASE_FILL = PatternFill("solid", fgColor="EAF2F8")
RECRUITMENT_FILL = PatternFill("solid", fgColor="D9EAD3")
ASSISTANT_FILL = PatternFill("solid", fgColor="EADCF8")
REFERRAL_FILL = PatternFill("solid", fgColor="FCE4D6")
FINAL_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="F4B183")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)


DETAIL_HEADERS = [
    "唯一验证", "姓名", "工号", "人员状态", "工作地", "职级", "ABC类别", "招聘渠道",
    "招聘负责人工号", "招聘负责人姓名", "招聘负责人状态", "招聘负责人最后工作日",
    "协助招聘人工号", "协助招聘人姓名", "协助招聘人状态", "协助招聘人最后工作日",
    "招聘启动日期", "候选人入职时间", "转正日期", "标签分类", "标准入职周期", "实际入职周期",
    "入职周期差异", "渠道系数", "招聘奖金标准", "招聘人实际发放标准",
    "招聘人入职1月奖金", "招聘人入职1月周期", "招聘人入职3月奖金", "招聘人入职3月周期",
    "招聘人入职6月奖金", "招聘人入职6月周期", "招聘人转正奖金", "招聘人转正周期",
    "协助人实际发放标准", "协助人入职1月奖金", "协助人入职1月周期", "协助人入职3月奖金",
    "协助人入职3月周期", "协助人入职6月奖金", "协助人入职6月周期", "协助人转正奖金",
    "协助人转正周期", "币种", "推荐人姓名", "推荐人工号", "推荐人状态", "推荐人是否上级",
    "内推规则范围", "内推奖金标准", "内推入职1月奖金", "内推入职1月周期", "内推入职3月奖金",
    "内推入职3月周期", "内推入职6月奖金", "内推入职6月周期", "内推转正奖金", "内推转正周期",
    "异常提示",
]

RECRUITMENT_SUMMARY_HEADERS = ["工号", "姓名", "角色", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"]
REFERRAL_SUMMARY_HEADERS = ["推荐人工号", "推荐人姓名", "币种", "核算月份", "入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"]
EXCEPTION_HEADERS = ["源行号", "姓名", "工号", "异常类型", "处理建议"]
PENDING_CONFIRMATION_HEADERS = [
    "源行号", "姓名", "工号", "奖金类型", "角色", "发放对象工号", "发放对象姓名", "发放对象状态",
    "币种", "发放节点", "建议发放周期", "建议发放金额", "人工确认结果", "人工确认金额",
    "不发/暂缓原因", "系统提示",
]
CONFIRMED_VALUES = {"发放", "确认发放", "是", "Y", "y", "YES", "Yes", "yes"}


def read_import_rows(path: Path) -> List[ImportRow]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    if IMPORT_SHEET not in workbook.sheetnames:
        raise ValueError(f"上传文件缺少工作表：{IMPORT_SHEET}")

    sheet = workbook[IMPORT_SHEET]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_names = [str(header).strip() if header else "" for header in headers]
    rows: List[ImportRow] = []

    for row_number in range(2, sheet.max_row + 1):
        values: Dict[str, Any] = {}
        has_value = False
        for column, header in enumerate(header_names, start=1):
            if not header:
                continue
            value = sheet.cell(row_number, column).value
            values[header] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            rows.append(ImportRow(source_row=row_number, values=values))
    return rows


def build_result_workbook(result: CalculationResult, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_detail_sheet(workbook, result.details)
    _write_summary_sheet(workbook, "待确认_发放判断", PENDING_CONFIRMATION_HEADERS, result.pending_confirmations, exception=True, table_name="tbl_pending_confirmations")
    _write_summary_sheet(workbook, "招聘奖金汇总", RECRUITMENT_SUMMARY_HEADERS, result.recruitment_summary, table_name="tbl_recruitment_summary")
    _write_summary_sheet(workbook, "内推奖金汇总", REFERRAL_SUMMARY_HEADERS, result.referral_summary, table_name="tbl_referral_summary")
    _write_summary_sheet(workbook, "异常清单", EXCEPTION_HEADERS, result.exceptions, exception=True, table_name="tbl_exceptions")
    _write_intro_sheet(workbook, result)

    workbook.save(path)


def build_pending_workbook(result: CalculationResult, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_intro_sheet(workbook, result)
    _write_summary_sheet(workbook, "待确认_发放判断", PENDING_CONFIRMATION_HEADERS, result.pending_confirmations, exception=True, table_name="tbl_pending_confirmations")
    workbook.save(path)


def build_final_workbook(base_path: Path, confirmation_path: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(base_path)
    confirmations = read_pending_confirmations(confirmation_path)

    recruitment_headers = RECRUITMENT_SUMMARY_HEADERS
    referral_headers = REFERRAL_SUMMARY_HEADERS
    recruitment_rows = _read_sheet_dicts(workbook, "招聘奖金汇总")
    referral_rows = _read_sheet_dicts(workbook, "内推奖金汇总")

    _apply_confirmations_to_summaries(confirmations, recruitment_rows, referral_rows)
    _apply_confirmations_to_detail_sheet(workbook, confirmations)

    for sheet_name in ("招聘奖金汇总", "内推奖金汇总", "待确认_发放判断", "最终招聘奖金汇总", "最终内推奖金汇总", "确认留痕"):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    _write_summary_sheet(workbook, "最终招聘奖金汇总", recruitment_headers, recruitment_rows, table_name="tbl_final_recruitment_summary", final=True)
    _write_summary_sheet(workbook, "最终内推奖金汇总", referral_headers, referral_rows, table_name="tbl_final_referral_summary", final=True)
    _write_summary_sheet(workbook, "确认留痕", PENDING_CONFIRMATION_HEADERS, confirmations, exception=True, table_name="tbl_confirmation_audit")
    workbook["确认留痕"].sheet_state = "hidden"
    workbook.save(output_path)


def read_pending_confirmations(path: Path) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    if "待确认_发放判断" not in workbook.sheetnames:
        raise ValueError("确认结果文件缺少工作表：待确认_发放判断")
    return _read_sheet_dicts(workbook, "待确认_发放判断")


def _write_detail_sheet(workbook: Workbook, details: List[RecruitmentDetail]) -> None:
    sheet = workbook.create_sheet("招聘奖金明细")
    sheet.append(DETAIL_HEADERS)
    for detail in details:
        sheet.append(_detail_values(detail))
    _format_sheet(sheet, table_name="tbl_detail")
    _format_detail_groups(sheet)
    _apply_date_format(sheet, _columns_by_headers(sheet, ["招聘负责人最后工作日", "协助招聘人最后工作日", "招聘启动日期", "候选人入职时间", "转正日期"]))
    _apply_money_format(sheet, _columns_by_headers(sheet, [
        "招聘奖金标准", "招聘人实际发放标准", "招聘人入职1月奖金", "招聘人入职3月奖金", "招聘人入职6月奖金", "招聘人转正奖金",
        "协助人实际发放标准", "协助人入职1月奖金", "协助人入职3月奖金", "协助人入职6月奖金", "协助人转正奖金",
        "内推奖金标准", "内推入职1月奖金", "内推入职3月奖金", "内推入职6月奖金", "内推转正奖金",
    ]))
    sheet.column_dimensions["E"].width = 24
    sheet.column_dimensions["J"].width = 18
    sheet.column_dimensions["N"].width = 18
    sheet.column_dimensions[get_column_letter(sheet.max_column)].width = 40


def _detail_values(detail: RecruitmentDetail) -> List[Any]:
    return [
        "重复" if "工号重复" in detail.exceptions else "唯一",
        detail.name,
        detail.employee_no,
        detail.status,
        detail.location,
        detail.grade,
        detail.category,
        detail.channel,
        detail.recruiter_id,
        detail.recruiter_name,
        detail.recruiter_status,
        detail.recruiter_last_work_date,
        detail.assistant_id,
        detail.assistant_name,
        detail.assistant_status,
        detail.assistant_last_work_date,
        detail.start_date,
        detail.onboard_date,
        detail.probation_date,
        detail.label,
        detail.standard_cycle_days,
        detail.actual_cycle_days,
        detail.cycle_diff_days,
        detail.channel_ratio,
        detail.standard_bonus,
        detail.recruiter_bonus_base,
        detail.recruiter_1m_bonus,
        detail.recruiter_1m_period,
        detail.recruiter_3m_bonus,
        detail.recruiter_3m_period,
        detail.recruiter_6m_bonus,
        detail.recruiter_6m_period,
        detail.recruiter_probation_bonus,
        detail.recruiter_probation_period,
        detail.assistant_bonus_base,
        detail.assistant_1m_bonus,
        detail.assistant_1m_period,
        detail.assistant_3m_bonus,
        detail.assistant_3m_period,
        detail.assistant_6m_bonus,
        detail.assistant_6m_period,
        detail.assistant_probation_bonus,
        detail.assistant_probation_period,
        detail.currency,
        detail.referrer_name,
        detail.referrer_id,
        detail.referrer_status,
        detail.referrer_is_manager,
        detail.referral_rule_scope,
        detail.referral_standard_bonus,
        detail.referral_1m_bonus,
        detail.referral_1m_period,
        detail.referral_3m_bonus,
        detail.referral_3m_period,
        detail.referral_6m_bonus,
        detail.referral_6m_period,
        detail.referral_probation_bonus,
        detail.referral_probation_period,
        "；".join(detail.exceptions),
    ]


def _write_summary_sheet(
    workbook: Workbook,
    title: str,
    headers: List[str],
    rows: List[Dict[str, Any]],
    exception: bool = False,
    table_name: str = "tbl_summary",
    final: bool = False,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    if _should_append_total_row(title, headers):
        _append_total_row(sheet, headers, rows)
    _format_sheet(sheet, table_name=table_name, exception=exception, final=final)
    if _should_append_total_row(title, headers):
        _format_total_row(sheet)
    if not exception:
        _apply_money_format(sheet, list(range(len(headers) - 4, len(headers) + 1)))
    else:
        sheet.column_dimensions["D"].width = 32
        sheet.column_dimensions["E"].width = 42
    if title == "待确认_发放判断":
        _add_pending_confirmation_validation(sheet)


def _read_sheet_dicts(workbook, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    rows: List[Dict[str, Any]] = []
    for row_number in range(2, sheet.max_row + 1):
        row = {header: sheet.cell(row_number, column).value for column, header in enumerate(headers, start=1) if header}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def _apply_confirmations_to_summaries(confirmations: List[Dict[str, Any]], recruitment_rows: List[Dict[str, Any]], referral_rows: List[Dict[str, Any]]) -> None:
    for row in confirmations:
        if as_text(row.get("人工确认结果")) not in CONFIRMED_VALUES:
            continue
        amount = row.get("人工确认金额")
        if amount in (None, ""):
            amount = row.get("建议发放金额", 0)
        try:
            amount = float(amount or 0)
        except (TypeError, ValueError):
            amount = 0.0
        if not amount:
            continue
        if row.get("奖金类型") == "招聘奖金":
            _add_confirmed_recruitment(recruitment_rows, row, amount)
        elif row.get("奖金类型") == "内推奖金":
            _add_confirmed_referral(referral_rows, row, amount)


def _apply_confirmations_to_detail_sheet(workbook, confirmations: List[Dict[str, Any]]) -> None:
    if "招聘奖金明细" not in workbook.sheetnames:
        return
    sheet = workbook["招聘奖金明细"]
    header_map = {str(sheet.cell(1, column).value).strip(): column for column in range(1, sheet.max_column + 1) if sheet.cell(1, column).value}
    if not {"工号", "姓名"}.issubset(header_map):
        return
    row_map = {}
    for row_number in range(2, sheet.max_row + 1):
        key = (
            as_text(sheet.cell(row_number, header_map["工号"]).value),
            as_text(sheet.cell(row_number, header_map["姓名"]).value),
        )
        row_map[key] = row_number
    for confirmation in confirmations:
        key = (as_text(confirmation.get("工号")), as_text(confirmation.get("姓名")))
        row_number = row_map.get(key)
        if not row_number:
            continue
        amount_header, period_header = _detail_confirmation_headers(confirmation)
        if amount_header not in header_map or period_header not in header_map:
            continue
        confirmed = as_text(confirmation.get("人工确认结果")) in CONFIRMED_VALUES
        amount = _confirmation_amount(confirmation) if confirmed else 0.0
        period = confirmation.get("建议发放周期") if confirmed else "-"
        sheet.cell(row_number, header_map[amount_header]).value = amount
        sheet.cell(row_number, header_map[period_header]).value = period
        sheet.cell(row_number, header_map[amount_header]).number_format = MONEY_FORMAT


def _detail_confirmation_headers(confirmation: Dict[str, Any]) -> tuple[str, str]:
    role = as_text(confirmation.get("角色"))
    node = as_text(confirmation.get("发放节点"))
    if role == "协助招聘人":
        prefix = "协助人"
    elif role == "推荐人" or confirmation.get("奖金类型") == "内推奖金":
        prefix = "内推"
    else:
        prefix = "招聘人"
    node_map = {
        "入职1月奖金": "入职1月",
        "入职3月奖金": "入职3月",
        "入职6月奖金": "入职6月",
        "转正奖金": "转正",
    }
    node_prefix = node_map.get(node, node.replace("奖金", ""))
    return f"{prefix}{node_prefix}奖金", f"{prefix}{node_prefix}周期"


def _confirmation_amount(confirmation: Dict[str, Any]) -> float:
    amount = confirmation.get("人工确认金额")
    if amount in (None, ""):
        amount = confirmation.get("建议发放金额", 0)
    try:
        return float(amount or 0)
    except (TypeError, ValueError):
        return 0.0


def _add_confirmed_recruitment(rows: List[Dict[str, Any]], confirmation: Dict[str, Any], amount: float) -> None:
    key = (
        confirmation.get("发放对象工号"),
        confirmation.get("发放对象姓名"),
        confirmation.get("角色"),
        confirmation.get("币种"),
    )
    target = next(
        (
            row for row in rows
            if (row.get("工号"), row.get("姓名"), row.get("角色"), row.get("币种")) == key
        ),
        None,
    )
    if target is None:
        target = {
            "工号": confirmation.get("发放对象工号"),
            "姓名": confirmation.get("发放对象姓名"),
            "角色": confirmation.get("角色"),
            "币种": confirmation.get("币种"),
            "核算月份": confirmation.get("建议发放周期"),
            "入职1月奖金": 0.0,
            "入职3月奖金": 0.0,
            "入职6月奖金": 0.0,
            "转正奖金": 0.0,
            "合计发放": 0.0,
        }
        rows.append(target)
    _add_amount_to_summary_row(target, as_text(confirmation.get("发放节点")), amount)


def _add_confirmed_referral(rows: List[Dict[str, Any]], confirmation: Dict[str, Any], amount: float) -> None:
    key = (
        confirmation.get("发放对象工号"),
        confirmation.get("发放对象姓名"),
        confirmation.get("币种"),
    )
    target = next(
        (
            row for row in rows
            if (row.get("推荐人工号"), row.get("推荐人姓名"), row.get("币种")) == key
        ),
        None,
    )
    if target is None:
        target = {
            "推荐人工号": confirmation.get("发放对象工号"),
            "推荐人姓名": confirmation.get("发放对象姓名"),
            "币种": confirmation.get("币种"),
            "核算月份": confirmation.get("建议发放周期"),
            "入职1月奖金": 0.0,
            "入职3月奖金": 0.0,
            "入职6月奖金": 0.0,
            "转正奖金": 0.0,
            "合计发放": 0.0,
        }
        rows.append(target)
    _add_amount_to_summary_row(target, as_text(confirmation.get("发放节点")), amount)


def _add_amount_to_summary_row(row: Dict[str, Any], node: str, amount: float) -> None:
    field_map = {
        "入职1月奖金": "入职1月奖金",
        "入职3月奖金": "入职3月奖金",
        "入职6月奖金": "入职6月奖金",
        "转正奖金": "转正奖金",
    }
    field = field_map.get(node)
    if not field:
        return
    row[field] = float(row.get(field) or 0) + amount
    row["合计发放"] = float(row.get("合计发放") or 0) + amount


def _should_append_total_row(title: str, headers: List[str]) -> bool:
    if "汇总" not in title:
        return False
    return "合计发放" in headers and (
        "招聘奖金" in title
        or "内推奖金" in title
        or title in {"最终招聘奖金汇总", "最终内推奖金汇总"}
    )


def _append_total_row(sheet, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    money_headers = ["入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"]
    total = []
    label_header = "姓名" if "姓名" in headers else "推荐人姓名" if "推荐人姓名" in headers else headers[0]
    for header in headers:
        if header == label_header:
            total.append("合计")
        elif header in money_headers:
            total.append(round(sum(_as_float(row.get(header)) for row in rows), 2))
        else:
            total.append("")
    sheet.append(total)


def _format_total_row(sheet) -> None:
    row_number = sheet.max_row
    for cell in sheet[row_number]:
        cell.fill = TOTAL_FILL
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if isinstance(cell.value, (int, float)):
            cell.number_format = MONEY_FORMAT


def _as_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _write_intro_sheet(workbook: Workbook, result: CalculationResult) -> None:
    sheet = workbook.create_sheet("计算说明", 0)
    sheet.append(("项目", "值"))
    rows = [
        ("核算月份", result.month),
        ("导入有效行数", len(result.details)),
        ("招聘奖金汇总金额", round(sum(row.get("合计发放", 0) for row in result.recruitment_summary), 2)),
        ("内推奖金汇总金额", round(sum(row.get("合计发放", 0) for row in result.referral_summary), 2)),
        ("待确认条数", len(result.pending_confirmations)),
        ("待确认建议金额", round(sum(row.get("建议发放金额", 0) for row in result.pending_confirmations), 2)),
        ("异常条数", len(result.exceptions)),
        ("说明", "本文件由平台按规则表确定性计算生成，异常请先复核后再发放。"),
    ]
    for row in rows:
        sheet.append(row)
    _format_sheet(sheet, table_name="tbl_intro")
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 80


def _format_sheet(sheet, table_name: str, exception: bool = False, final: bool = False) -> None:
    max_row = max(sheet.max_row, 1)
    max_col = max(sheet.max_column, 1)
    header_fill = EXCEPTION_FILL if exception else (FINAL_FILL if final else HEADER_FILL)
    header_font = Font(bold=True) if exception else WHITE_FONT

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if cell.row % 2 == 0:
                cell.fill = LIGHT_FILL

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    for column in range(1, max_col + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 15

    if max_row >= 2:
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=_safe_table_name(table_name), ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)


def _format_detail_groups(sheet) -> None:
    group_fills = {
        "基础": BASE_FILL,
        "招聘": RECRUITMENT_FILL,
        "协助": ASSISTANT_FILL,
        "内推": REFERRAL_FILL,
    }
    for column in range(1, sheet.max_column + 1):
        header = as_text(sheet.cell(1, column).value)
        if header.startswith("协助"):
            fill = group_fills["协助"]
        elif header.startswith("内推") or header.startswith("推荐"):
            fill = group_fills["内推"]
        elif header.startswith("招聘") or header.startswith("候选人") or header in {"标准入职周期", "实际入职周期", "入职周期差异", "渠道系数"}:
            fill = group_fills["招聘"]
        else:
            fill = group_fills["基础"]
        cell = sheet.cell(1, column)
        cell.fill = fill
        cell.font = Font(bold=True, color="000000")
    sheet.row_dimensions[1].height = 34


def _columns_by_headers(sheet, headers: List[str]) -> List[int]:
    wanted = set(headers)
    return [column for column in range(1, sheet.max_column + 1) if sheet.cell(1, column).value in wanted]


def _add_pending_confirmation_validation(sheet) -> None:
    if sheet.max_row < 2:
        return
    result_column = PENDING_CONFIRMATION_HEADERS.index("人工确认结果") + 1
    column_letter = get_column_letter(result_column)
    validation = DataValidation(type="list", formula1='"确认发放,不发放,暂缓"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{sheet.max_row}")


def _safe_table_name(name: str) -> str:
    candidate = re.sub(r"[^A-Za-z0-9_]", "_", name)
    candidate = re.sub(r"_+", "_", candidate).strip("_")
    if not candidate or candidate[0].isdigit():
        candidate = f"tbl_{candidate}"
    return candidate[:240]


def _apply_date_format(sheet, columns: List[int]) -> None:
    for column in columns:
        for row in range(2, sheet.max_row + 1):
            value = sheet.cell(row, column).value
            if isinstance(value, datetime):
                sheet.cell(row, column).number_format = DATE_FORMAT


def _apply_money_format(sheet, columns: List[int]) -> None:
    for column in columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column).number_format = MONEY_FORMAT
