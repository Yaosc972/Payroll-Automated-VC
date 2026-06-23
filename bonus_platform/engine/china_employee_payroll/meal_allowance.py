"""技术部餐补核算：考勤记录解析与规则计算。"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "员工",
    "工号",
    "人员状态",
    "二级组织",
    "三级组织",
    "四级组织",
    "五级组织",
    "考勤日期",
    "首打卡(含补签)",
    "末打卡(含补签)",
    "当前班次",
    "日期类型",
    "备注",
]

WX_REQUIRED_COLUMNS = [
    "姓名",
    "工号",
    "部门",
    "组织架构",
    "人员类型",
    "员工状态",
    "日期",
    "班次",
    "上班 1 打卡时间",
    "下班 1 打卡时间",
]


@dataclass(frozen=True)
class MealAllowanceConfig:
    eligible_second_orgs: frozenset[str] = frozenset(
        {
            "ABU航空事业部",
            "FBU仓储事业部",
            "FES财经条线",
            "HRAS人力综合条线",
            "LBU速运事业部",
            "PBU口岸事业部",
        }
    )
    eligible_whole_second_orgs: frozenset[str] = frozenset({"HQU技术部"})
    eligible_third_orgs: frozenset[str] = frozenset(
        {"ABU技术部", "FBU技术部", "FES技术部", "技术部", "B技术部", "PBU技术部"}
    )
    special_third_orgs: frozenset[str] = frozenset({"HSSC人力共享中心", "战略运营部"})
    special_fourth_orgs: frozenset[str] = frozenset({"HRAS技术组", "BI组"})
    eligible_shifts: frozenset[str] = frozenset(
        {
            "9:30-18:30",
            "集团深圳10:00-19:00",
            "深圳灵活打卡8:30-9:10",
            "深圳南山灵活打卡8:30-9:30",
        }
    )
    excluded_employee_names: frozenset[str] = frozenset({"潘江浩", "李显荣"})
    daily_amount: int = 20


@dataclass
class ParsedAttendance:
    rows: list[dict[str, Any]]
    files: list[dict[str, Any]]
    headers: list[str]
    missing_columns: list[str]
    duplicate_keys: list[dict[str, str]] = field(default_factory=list)
    source_type: str = "hr"
    source_label: str = "人事系统考勤"

    def summary(self) -> dict[str, Any]:
        employees = {row.get("工号") for row in self.rows if row.get("工号")}
        dates = [row.get("考勤日期") for row in self.rows if row.get("考勤日期")]
        return {
            "fileCount": len(self.files),
            "rowCount": len(self.rows),
            "employeeCount": len(employees),
            "dateStart": min(dates).isoformat() if dates else "",
            "dateEnd": max(dates).isoformat() if dates else "",
            "distinctDateCount": len(set(dates)),
            "missingColumns": self.missing_columns,
            "duplicateCount": len(self.duplicate_keys),
            "sourceType": self.source_type,
            "sourceLabel": self.source_label,
        }


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def _parse_date(value: Any):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    parsed = _parse_datetime(value)
    if parsed:
        return parsed.date()
    return None


def _parse_time(value: Any) -> time | None:
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    text = _clean(value)
    if not text or text == "-":
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    parsed = _parse_datetime(value)
    if parsed:
        return parsed.time()
    return None


def _hr_workbook_rows(path: Path, *, read_only: bool) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=read_only, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
        headers = [_clean(value) for value in header_row]
        rows: list[dict[str, Any]] = []
        for source_row, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not any(_clean(value) for value in values):
                continue
            row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
            row["_sourceFile"] = path.name
            row["_sourceRow"] = source_row
            parsed_date = _parse_date(row.get("考勤日期"))
            if parsed_date:
                row["考勤日期"] = parsed_date
            rows.append(row)
        return headers, rows, {"filename": path.name, "rowCount": len(rows), "sheetName": sheet.title}
    finally:
        workbook.close()


def _workbook_rows(path: Path) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    headers, rows, file_info = _hr_workbook_rows(path, read_only=True)
    if any(column not in headers for column in REQUIRED_COLUMNS):
        return _hr_workbook_rows(path, read_only=False)
    return headers, rows, file_info


def parse_attendance_workbooks(paths: Iterable[str | Path]) -> ParsedAttendance:
    all_rows: list[dict[str, Any]] = []
    files: list[dict[str, Any]] = []
    headers: list[str] = []
    missing_columns: set[str] = set()
    seen: dict[tuple[str, str], dict[str, Any]] = {}
    duplicates: list[dict[str, str]] = []

    for raw_path in paths:
        path = Path(raw_path)
        file_headers, rows, file_info = _workbook_rows(path)
        files.append(file_info)
        if not headers:
            headers = file_headers
        missing_columns.update(column for column in REQUIRED_COLUMNS if column not in file_headers)
        for row in rows:
            emp_id = _clean(row.get("工号"))
            attendance_date = row.get("考勤日期")
            key = (emp_id, attendance_date.isoformat() if attendance_date else "")
            if emp_id and attendance_date:
                if key in seen:
                    duplicates.append(
                        {
                            "employeeId": emp_id,
                            "attendanceDate": key[1],
                            "firstFile": _clean(seen[key].get("_sourceFile")),
                            "duplicateFile": _clean(row.get("_sourceFile")),
                        }
                    )
                else:
                    seen[key] = row
            all_rows.append(row)

    return ParsedAttendance(
        rows=all_rows,
        files=files,
        headers=headers,
        missing_columns=sorted(missing_columns),
        duplicate_keys=duplicates,
        source_type="hr",
        source_label="人事系统考勤",
    )


def _wx_workbook_rows_from(path: Path, *, read_only: bool) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=read_only, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        first_header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        second_header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
        max_columns = max(len(first_header_row), len(second_header_row))
        first_header = [_clean(value) for value in first_header_row]
        second_header = [_clean(value) for value in second_header_row]
        headers = [
            (second_header[index] if index < len(second_header) else "")
            or (first_header[index] if index < len(first_header) else "")
            or f"列{index + 1}"
            for index in range(max_columns)
        ]
        rows: list[dict[str, Any]] = []
        for source_row, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not any(_clean(value) for value in values):
                continue
            raw = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
            attendance_date = _parse_date(raw.get("日期"))
            org_path = _clean(raw.get("组织架构"))
            department = _clean(raw.get("部门"))
            row = {
                **raw,
                "员工": _clean(raw.get("姓名")),
                "人员状态": _clean(raw.get("员工状态")),
                "二级组织": "WX技术部",
                "三级组织": "WX-PBU技术部",
                "四级组织": department if org_path == "WX-PBU技术部" else org_path.replace("WX-PBU技术部-", "", 1),
                "五级组织": "",
                "考勤日期": attendance_date,
                "首打卡(含补签)": raw.get("上班 1 打卡时间"),
                "末打卡(含补签)": raw.get("下班 1 打卡时间"),
                "当前班次": raw.get("班次"),
                "日期类型": "工作日" if _clean(raw.get("班次")) and "休息" not in _clean(raw.get("班次")) else "休息",
                "备注": "",
                "_sourceFile": path.name,
                "_sourceRow": source_row,
                "_sourceOrgPath": org_path,
            }
            rows.append(row)
        return headers, rows, {"filename": path.name, "rowCount": len(rows), "sheetName": sheet.title}
    finally:
        workbook.close()


def _wx_workbook_rows(path: Path) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    headers, rows, file_info = _wx_workbook_rows_from(path, read_only=True)
    if any(column not in headers for column in WX_REQUIRED_COLUMNS):
        return _wx_workbook_rows_from(path, read_only=False)
    return headers, rows, file_info


def parse_wx_attendance_workbooks(paths: Iterable[str | Path]) -> ParsedAttendance:
    all_rows: list[dict[str, Any]] = []
    files: list[dict[str, Any]] = []
    headers: list[str] = []
    missing_columns: set[str] = set()
    seen: dict[tuple[str, str], dict[str, Any]] = {}
    duplicates: list[dict[str, str]] = []

    for raw_path in paths:
        path = Path(raw_path)
        file_headers, rows, file_info = _wx_workbook_rows(path)
        files.append(file_info)
        if not headers:
            headers = file_headers
        missing_columns.update(column for column in WX_REQUIRED_COLUMNS if column not in file_headers)
        for row in rows:
            emp_id = _clean(row.get("工号"))
            attendance_date = row.get("考勤日期")
            key = (emp_id, attendance_date.isoformat() if attendance_date else "")
            if emp_id and attendance_date:
                if key in seen:
                    duplicates.append(
                        {
                            "employeeId": emp_id,
                            "attendanceDate": key[1],
                            "firstFile": _clean(seen[key].get("_sourceFile")),
                            "duplicateFile": _clean(row.get("_sourceFile")),
                        }
                    )
                else:
                    seen[key] = row
            all_rows.append(row)

    return ParsedAttendance(
        rows=all_rows,
        files=files,
        headers=headers,
        missing_columns=sorted(missing_columns),
        duplicate_keys=duplicates,
        source_type="wx",
        source_label="WX技术部考勤",
    )


def _is_zt_employee(row: dict[str, Any]) -> bool:
    return _clean(row.get("工号")).lower().startswith("zt")


def _is_eligible_org(row: dict[str, Any], config: MealAllowanceConfig) -> bool:
    if _clean(row.get("_sourceType")) == "wx":
        org_path = _clean(row.get("_sourceOrgPath")) or _clean(row.get("组织架构"))
        return org_path == "WX-PBU技术部" or org_path.startswith("WX-PBU技术部-")
    second = _clean(row.get("二级组织"))
    third = _clean(row.get("三级组织"))
    fourth = _clean(row.get("四级组织"))
    if second in config.eligible_whole_second_orgs:
        return True
    return (
        second in config.eligible_second_orgs
        and third in config.eligible_third_orgs
    ) or (
        second == "HRAS人力综合条线"
        and
        third in config.special_third_orgs
        and fourth in config.special_fourth_orgs
    )


def _is_eligible_shift(row: dict[str, Any], config: MealAllowanceConfig) -> bool:
    shift = _clean(row.get("当前班次"))
    return any(allowed in shift for allowed in config.eligible_shifts)


def _is_workday(row: dict[str, Any]) -> bool:
    if _clean(row.get("_sourceType")) == "wx":
        shift = _clean(row.get("当前班次"))
        return bool(shift) and "休息" not in shift
    return _clean(row.get("日期类型")) == "工作日"


def _last_punch_qualifies(row: dict[str, Any]) -> bool:
    if _clean(row.get("_sourceType")) == "wx":
        punch_time = _parse_time(row.get("末打卡(含补签)"))
        return bool(punch_time and (punch_time >= time(21, 0) or punch_time <= time(8, 0)))
    last_punch = _parse_datetime(row.get("末打卡(含补签)"))
    if not last_punch:
        return False
    punch_time = last_punch.time()
    return punch_time >= time(21, 0) or punch_time <= time(8, 0)


def _remark_has_approved_business_trip(row: dict[str, Any]) -> bool:
    remark = _clean(row.get("备注"))
    if not ("公出" in remark or "出差" in remark):
        return False
    for hour, minute in re.findall(r"(\d{1,2}):(\d{2})", remark):
        if time(int(hour), int(minute)) >= time(21, 0):
            return True
    return False


def _ineligible_reason(row: dict[str, Any], config: MealAllowanceConfig) -> str:
    if _clean(row.get("_sourceType")) != "wx" and not _is_zt_employee(row):
        return "工号不是zt开头"
    if _clean(row.get("员工")) in config.excluded_employee_names:
        return "员工在不计算补贴名单"
    if not _is_eligible_org(row, config):
        return "组织不在核算对象范围"
    return ""


def _non_payable_reason(row: dict[str, Any], config: MealAllowanceConfig) -> str:
    if not _is_workday(row):
        return "班次为休息或非工作日" if _clean(row.get("_sourceType")) == "wx" else "日期类型不是工作日"
    if _clean(row.get("_sourceType")) != "wx" and not _is_eligible_shift(row, config):
        return "当前班次不在餐补班次范围"
    if _last_punch_qualifies(row):
        return ""
    if _remark_has_approved_business_trip(row):
        return ""
    return "末打卡(含补签)不在21:00-次日08:00范围"


def calculate_meal_allowance(parsed: ParsedAttendance, config: MealAllowanceConfig | None = None) -> dict[str, Any]:
    config = config or MealAllowanceConfig()
    employee_results: dict[str, dict[str, Any]] = {}
    daily_rows: list[dict[str, Any]] = []

    for row in parsed.rows:
        row["_sourceType"] = parsed.source_type
        emp_id = _clean(row.get("工号"))
        if not emp_id:
            continue
        base = employee_results.setdefault(
            emp_id,
            {
                "employeeId": emp_id,
                "employeeName": _clean(row.get("员工")),
                "status": _clean(row.get("人员状态")),
                "secondOrg": _clean(row.get("二级组织")),
                "thirdOrg": _clean(row.get("三级组织")),
                "fourthOrg": _clean(row.get("四级组织")),
                "fifthOrg": _clean(row.get("五级组织")),
                "payableDays": 0,
                "amount": 0,
                "daily": [],
                "warnings": set(),
                "sourceType": parsed.source_type,
                "sourceLabel": parsed.source_label,
            },
        )
        ineligible = _ineligible_reason(row, config)
        non_payable = "" if ineligible else _non_payable_reason(row, config)
        payable = not ineligible and not non_payable
        amount = config.daily_amount if payable else 0
        if payable:
            base["payableDays"] += 1
            base["amount"] += amount
        else:
            base["warnings"].add(ineligible or non_payable)

        attendance_date = row.get("考勤日期")
        daily = {
            "employeeId": emp_id,
            "employeeName": _clean(row.get("员工")),
            "attendanceDate": attendance_date.isoformat() if attendance_date else "",
            "dateType": _clean(row.get("日期类型")),
            "shift": _clean(row.get("当前班次")),
            "firstPunch": _clean(row.get("首打卡(含补签)")),
            "lastPunch": _clean(row.get("末打卡(含补签)")),
            "remark": _clean(row.get("备注")),
            "payable": payable,
            "amount": amount,
            "reason": "符合餐补规则" if payable else (ineligible or non_payable),
            "sourceType": parsed.source_type,
            "sourceLabel": parsed.source_label,
        }
        base["daily"].append(daily)
        daily_rows.append(daily)

    results = []
    for result in employee_results.values():
        result["warnings"] = sorted(result["warnings"])
        results.append(result)
    results.sort(key=lambda item: (-item["amount"], item["employeeId"]))

    payable_results = [row for row in results if row["amount"] > 0]
    summary = {
        **parsed.summary(),
        "eligibleEmployeeCount": len([row for row in results if row["payableDays"] or "组织不在核算对象范围" not in row["warnings"]]),
        "payableEmployeeCount": len(payable_results),
        "payableDayCount": sum(row["payableDays"] for row in results),
        "totalAmount": sum(row["amount"] for row in results),
        "dailyAmount": config.daily_amount,
        "sourceType": parsed.source_type,
        "sourceLabel": parsed.source_label,
    }
    return {
        "summary": summary,
        "results": results,
        "dailyRows": daily_rows,
        "files": parsed.files,
        "warnings": {
            "missingColumns": parsed.missing_columns,
            "duplicateKeys": parsed.duplicate_keys[:50],
        },
    }
