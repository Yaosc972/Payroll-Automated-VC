from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

from datetime import timedelta

from .models import ImportRow, add_months, as_date, as_number, as_text, yyyymm


LEGACY_DETAIL_SHEET = "附件一-招聘奖金核算明细"
IMPORT_SHEET = "导入_月度数据"
TEMPLATE_DETAIL_SHEETS = ("输出_招聘奖金明细", "招聘奖金明细")

PERIOD_FIELDS = [
    "招聘人入职1月发放周期_覆盖",
    "招聘人入职3月发放周期_覆盖",
    "招聘人入职6月发放周期_覆盖",
    "招聘人转正发放周期_覆盖",
    "协助人入职1月发放周期_覆盖",
    "协助人入职3月发放周期_覆盖",
    "协助人入职6月发放周期_覆盖",
    "协助人转正发放周期_覆盖",
]

AMOUNT_FIELDS = [
    "招聘人入职1月发放金额_覆盖",
    "招聘人入职3月发放金额_覆盖",
    "招聘人入职6月发放金额_覆盖",
    "招聘人转正发放金额_覆盖",
    "协助人入职1月发放金额_覆盖",
    "协助人入职3月发放金额_覆盖",
    "协助人入职6月发放金额_覆盖",
    "协助人转正发放金额_覆盖",
]

REFERRAL_PERIOD_FIELDS = [
    "内推入职1月发放周期_覆盖",
    "内推入职3月发放周期_覆盖",
    "内推入职6月发放周期_覆盖",
    "内推转正发放周期_覆盖",
]

REFERRAL_AMOUNT_FIELDS = [
    "内推入职1月发放金额_覆盖",
    "内推入职3月发放金额_覆盖",
    "内推入职6月发放金额_覆盖",
    "内推转正发放金额_覆盖",
]

LEGACY_PERIOD_COLUMNS = [40, 42, 44, 53, 47, 49, 51, 55]
LEGACY_AMOUNT_COLUMNS = [39, 41, 43, 52, 46, 48, 50, 54]
LEGACY_REFERRAL_PERIOD_COLUMNS = [63, 65, 67, 69]
LEGACY_REFERRAL_AMOUNT_COLUMNS = [62, 64, 66, 68]

TEMPLATE_PERIOD_HEADERS = [
    "招聘人入职1月周期",
    "招聘人入职3月周期",
    "招聘人入职6月周期",
    "招聘人转正周期",
    "协助人入职1月周期",
    "协助人入职3月周期",
    "协助人入职6月周期",
    "协助人转正周期",
]

TEMPLATE_AMOUNT_HEADERS = [
    "招聘人入职1月奖金",
    "招聘人入职3月奖金",
    "招聘人入职6月奖金",
    "招聘人转正奖金",
    "协助人入职1月奖金",
    "协助人入职3月奖金",
    "协助人入职6月奖金",
    "协助人转正奖金",
]

TEMPLATE_REFERRAL_PERIOD_HEADERS = [
    "内推入职1月周期",
    "内推入职3月周期",
    "内推入职6月周期",
    "内推转正周期",
]

TEMPLATE_REFERRAL_AMOUNT_HEADERS = [
    "内推入职1月奖金",
    "内推入职3月奖金",
    "内推入职6月奖金",
    "内推转正奖金",
]


@dataclass(frozen=True)
class HistoryOverrides:
    source_type: str
    by_key: Dict[str, Dict[str, Any]]
    source_month: int | None = None


def load_history_overrides(path: Path) -> HistoryOverrides:
    workbook = load_workbook(path, data_only=True, read_only=False)
    if LEGACY_DETAIL_SHEET in workbook.sheetnames:
        return _read_legacy_payroll_history(workbook[LEGACY_DETAIL_SHEET], _month_from_path(path))

    if IMPORT_SHEET in workbook.sheetnames:
        import_overrides = _read_import_sheet_overrides(workbook[IMPORT_SHEET])
        if import_overrides.by_key:
            return import_overrides

    for sheet_name in TEMPLATE_DETAIL_SHEETS:
        if sheet_name in workbook.sheetnames:
            return _read_template_history(workbook[sheet_name])

    raise ValueError("历史奖金表无法识别：未找到薪酬组原线下明细表或平台输出明细表。")


def merge_history_overrides(rows: Iterable[ImportRow], overrides: HistoryOverrides) -> List[ImportRow]:
    merged: List[ImportRow] = []
    for row in rows:
        values = dict(row.values)
        key = _row_key(values)
        if key in overrides.by_key:
            values.update(_applicable_override_values(values, overrides, overrides.by_key[key]))
        merged.append(ImportRow(source_row=row.source_row, values=values))
    return merged


def _read_legacy_payroll_history(sheet, source_month: int | None = None) -> HistoryOverrides:
    by_key: Dict[str, Dict[str, Any]] = {}
    if source_month is None:
        source_month = _infer_legacy_source_month(sheet)
    for row_number in range(4, sheet.max_row + 1):
        key = _employee_key(sheet.cell(row_number, 3).value, sheet.cell(row_number, 2).value)
        if not key:
            continue
        override_values: Dict[str, Any] = {}
        for field, column in zip(PERIOD_FIELDS, LEGACY_PERIOD_COLUMNS):
            _copy_if_present(override_values, field, sheet.cell(row_number, column).value)
        for field, column in zip(AMOUNT_FIELDS, LEGACY_AMOUNT_COLUMNS):
            _copy_if_present(override_values, field, sheet.cell(row_number, column).value)
        for field, column in zip(REFERRAL_PERIOD_FIELDS, LEGACY_REFERRAL_PERIOD_COLUMNS):
            _copy_if_present(override_values, field, sheet.cell(row_number, column).value)
        for field, column in zip(REFERRAL_AMOUNT_FIELDS, LEGACY_REFERRAL_AMOUNT_COLUMNS):
            _copy_if_present(override_values, field, sheet.cell(row_number, column).value)
        _copy_legacy_referral_probation_amount(override_values, sheet, row_number)
        if override_values:
            by_key[key] = override_values
    return HistoryOverrides(source_type="薪酬组原线下累计表", by_key=by_key, source_month=source_month)


def _read_template_history(sheet) -> HistoryOverrides:
    header_map = {
        as_text(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(1, column).value
    }
    by_key: Dict[str, Dict[str, Any]] = {}
    for row_number in range(2, sheet.max_row + 1):
        key = _employee_key(_cell_by_header(sheet, header_map, row_number, "工号"), _cell_by_header(sheet, header_map, row_number, "姓名"))
        if not key:
            continue
        override_values: Dict[str, Any] = {}
        for field, header in zip(PERIOD_FIELDS, TEMPLATE_PERIOD_HEADERS):
            _copy_if_present(override_values, field, _cell_by_header(sheet, header_map, row_number, header))
        for field, header in zip(AMOUNT_FIELDS, TEMPLATE_AMOUNT_HEADERS):
            _copy_if_present(override_values, field, _cell_by_header(sheet, header_map, row_number, header))
        for field, header in zip(REFERRAL_PERIOD_FIELDS, TEMPLATE_REFERRAL_PERIOD_HEADERS):
            _copy_if_present(override_values, field, _cell_by_header(sheet, header_map, row_number, header))
        for field, header in zip(REFERRAL_AMOUNT_FIELDS, TEMPLATE_REFERRAL_AMOUNT_HEADERS):
            _copy_if_present(override_values, field, _cell_by_header(sheet, header_map, row_number, header))
        if override_values:
            by_key[key] = override_values
    return HistoryOverrides(source_type="平台或模板输出明细表", by_key=by_key)


def _read_import_sheet_overrides(sheet) -> HistoryOverrides:
    header_map = {
        as_text(sheet.cell(1, column).value): column
        for column in range(1, sheet.max_column + 1)
        if sheet.cell(1, column).value
    }
    if not any(field in header_map for field in PERIOD_FIELDS + AMOUNT_FIELDS):
        return HistoryOverrides(source_type="模板导入覆盖字段", by_key={})

    by_key: Dict[str, Dict[str, Any]] = {}
    for row_number in range(2, sheet.max_row + 1):
        key = _employee_key(_cell_by_header(sheet, header_map, row_number, "工号"), _cell_by_header(sheet, header_map, row_number, "姓名"))
        if not key:
            continue
        override_values: Dict[str, Any] = {}
        for field in PERIOD_FIELDS + AMOUNT_FIELDS + REFERRAL_PERIOD_FIELDS + REFERRAL_AMOUNT_FIELDS:
            _copy_if_present(override_values, field, _cell_by_header(sheet, header_map, row_number, field))
        if override_values:
            by_key[key] = override_values
    return HistoryOverrides(source_type="模板导入覆盖字段", by_key=by_key)


def _cell_by_header(sheet, header_map: Dict[str, int], row_number: int, header: str) -> Any:
    column = header_map.get(header)
    if not column:
        return None
    return sheet.cell(row_number, column).value


def _copy_if_present(target: Dict[str, Any], field: str, value: Any) -> None:
    if value not in (None, ""):
        target[field] = value


def _copy_legacy_referral_probation_amount(target: Dict[str, Any], sheet, row_number: int) -> None:
    period = sheet.cell(row_number, 69).value
    if period in (None, ""):
        return
    standard_amount = sheet.cell(row_number, 61).value
    probation_amount = sheet.cell(row_number, 68).value
    total = 0.0
    has_amount = False
    for value in (standard_amount, probation_amount):
        if isinstance(value, (int, float)):
            total += value
            has_amount = True
    if has_amount:
        target["内推转正发放金额_覆盖"] = total


def _applicable_override_values(row_values: Dict[str, Any], overrides: HistoryOverrides, override_values: Dict[str, Any]) -> Dict[str, Any]:
    if not _should_keep_only_paid_legacy_nodes(row_values, overrides):
        return override_values
    paid = _paid_legacy_nodes(override_values, overrides.source_month or 0)
    _carry_pending_legacy_referral_nodes(row_values, override_values, paid, overrides.source_month or 0)
    return paid


def _should_keep_only_paid_legacy_nodes(row_values: Dict[str, Any], overrides: HistoryOverrides) -> bool:
    if overrides.source_type != "薪酬组原线下累计表" or not overrides.source_month:
        return False
    target_month = as_number(row_values.get("核算月份"))
    return bool(target_month and int(target_month) > overrides.source_month)


def _paid_legacy_nodes(override_values: Dict[str, Any], source_month: int) -> Dict[str, Any]:
    paid: Dict[str, Any] = {}
    for period_field, amount_field in zip(PERIOD_FIELDS, AMOUNT_FIELDS):
        _copy_paid_node(override_values, paid, period_field, amount_field, source_month)
    for period_field, amount_field in zip(REFERRAL_PERIOD_FIELDS, REFERRAL_AMOUNT_FIELDS):
        _copy_paid_node(override_values, paid, period_field, amount_field, source_month)
    return paid


def _copy_paid_node(source: Dict[str, Any], target: Dict[str, Any], period_field: str, amount_field: str, source_month: int) -> None:
    raw_period = source.get(period_field)
    if as_text(raw_period) == "-":
        target[period_field] = "-"
        target[amount_field] = 0
        return
    period = _as_yyyymm(raw_period)
    if not period or period > source_month:
        return
    target[period_field] = period
    if amount_field in source:
        target[amount_field] = source[amount_field]


def _carry_pending_legacy_referral_nodes(
    row_values: Dict[str, Any],
    source: Dict[str, Any],
    target: Dict[str, Any],
    source_month: int,
) -> None:
    probation_marker = as_text(source.get("内推转正发放周期_覆盖"))
    if "入职满6个月后一次性发放" in probation_marker:
        _copy_marker_node(
            row_values,
            source,
            target,
            period_field="内推转正发放周期_覆盖",
            amount_field="内推转正发放金额_覆盖",
            months=6,
            source_month=source_month,
        )

def _copy_marker_node(
    row_values: Dict[str, Any],
    source: Dict[str, Any],
    target: Dict[str, Any],
    period_field: str,
    amount_field: str,
    months: int,
    source_month: int,
) -> None:
    amount = source.get(amount_field)
    if amount in (None, "", 0):
        return
    period = _period_after_months(row_values, months)
    if not period:
        return
    if period <= source_month:
        period = _next_month(source_month)
    target[period_field] = period
    target[amount_field] = amount


def _period_after_months(row_values: Dict[str, Any], months: int) -> int | None:
    onboard_date = as_date(row_values.get("候选人入职时间")) or as_date(row_values.get("入职日期"))
    full_period_date = add_months(onboard_date, months)
    if not full_period_date:
        return None
    return yyyymm(full_period_date - timedelta(days=1))


def _next_month(month: int) -> int:
    year, number = divmod(month, 100)
    if number == 12:
        return (year + 1) * 100 + 1
    return year * 100 + number + 1


def _month_from_path(path: Path) -> int | None:
    matches = re.findall(r"20\d{4}", path.name)
    for match in reversed(matches):
        month = int(match)
        if 1 <= month % 100 <= 12:
            return month
    return None


def _infer_legacy_source_month(sheet) -> int | None:
    months: list[int] = []
    columns = LEGACY_PERIOD_COLUMNS + LEGACY_REFERRAL_PERIOD_COLUMNS
    for row in sheet.iter_rows(min_row=4, values_only=True):
        for column in columns:
            if column - 1 >= len(row):
                continue
            month = _as_yyyymm(row[column - 1])
            if month:
                months.append(month)
    return max(months) if months else None


def _as_yyyymm(value: Any) -> int | None:
    if isinstance(value, int):
        month = value
    elif isinstance(value, float) and value.is_integer():
        month = int(value)
    elif isinstance(value, str) and value.strip().isdigit():
        month = int(value.strip())
    else:
        return None
    return month if 200001 <= month <= 209912 and 1 <= month % 100 <= 12 else None


def _row_key(values: Dict[str, Any]) -> str:
    return _employee_key(values.get("工号"), values.get("姓名"))


def _employee_key(employee_no: Any, name: Any) -> str:
    normalized_employee_no = as_text(employee_no)
    if normalized_employee_no:
        return normalized_employee_no
    normalized_name = as_text(name)
    return f"姓名:{normalized_name}" if normalized_name else ""
