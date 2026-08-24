from __future__ import annotations

from collections import Counter
from copy import copy
from datetime import datetime
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .report import build_audit_export
from .runs import (
    RunNotFoundError,
    RunValidationError,
    current_timestamp,
    get_run_dir,
    load_run,
    save_run,
)
from .template_library import match_template, public_template_match
from .template_schemas import TEMPLATE_SCHEMAS, employee_template_routes


MISSING_SHEET_NAME = "待补资料"
SUMMARY_SHEET_NAME = "生成说明"


def _route_employees(run: dict[str, Any], route: str) -> list[dict[str, Any]]:
    return [
        employee
        for employee in run.get("employees") or []
        if employee.get("decision") == "include" and route in employee_template_routes(employee)
    ]


def _uploaded_template(run: dict[str, Any], route: str) -> dict[str, Any] | None:
    templates = run.get("templates") if isinstance(run.get("templates"), dict) else {}
    metadata = templates.get(route) if isinstance(templates.get(route), dict) else None
    if metadata is None and route == "shenzhen-social-medical" and isinstance(run.get("template"), dict):
        metadata = run["template"]
    if metadata is None:
        return None
    filename = Path(str(metadata.get("filename") or "")).name
    path = get_run_dir(str(run.get("id") or "")) / filename
    if not filename or not path.is_file():
        return None
    return {
        "route": route,
        "filename": str(metadata.get("originalFilename") or filename),
        "period": str(run.get("periodEnd") or "")[:7].replace("-", ""),
        "size": path.stat().st_size,
        "path": path,
        "source": "uploaded",
        "matchQuality": "uploaded",
        "subjectMatched": True,
    }


def _template_reference(run: dict[str, Any], route: str) -> dict[str, Any] | None:
    uploaded = _uploaded_template(run, route)
    if uploaded:
        return uploaded
    matched = match_template(route, str(run.get("subject") or ""))
    if matched:
        matched["source"] = "library"
    return matched


def _public_template_reference(run: dict[str, Any], route: str) -> dict[str, Any] | None:
    uploaded = _uploaded_template(run, route)
    if uploaded:
        return {key: value for key, value in uploaded.items() if key != "path"}
    return public_template_match(route, str(run.get("subject") or ""))


def build_export_preflight(run_or_id: dict[str, Any] | str) -> dict[str, Any]:
    run = load_run(run_or_id) if isinstance(run_or_id, str) else run_or_id
    groups: list[dict[str, Any]] = []
    manual_routes: list[dict[str, Any]] = []
    for plan in run.get("processingPlan") or []:
        route = str(plan.get("route") or "")
        if plan.get("handling") == "manual":
            manual_routes.append({
                "route": route,
                "label": str(plan.get("routeLabel") or "线下办理"),
                "employeeCount": int(plan.get("employeeCount") or 0),
                "coverages": list(plan.get("coverages") or []),
            })
            continue
        if plan.get("handling") != "template" or route not in TEMPLATE_SCHEMAS:
            continue
        employees = _route_employees(run, route)
        missing = Counter()
        for employee in employees:
            report = (employee.get("templateReports") or {}).get(route) or {}
            missing.update(report.get("missingRequired") or [])
        template = _public_template_reference(run, route)
        template_usable = bool(template and (template.get("source") == "uploaded" or template.get("subjectMatched")))
        if missing:
            status = "missing-fields"
            status_label = f"缺少 {sum(missing.values())} 项必填资料"
        elif not template_usable:
            status = "template-required"
            status_label = "需要确认或导入主体模板"
        else:
            status = "ready"
            status_label = "可生成"
        groups.append({
            "route": route,
            "label": TEMPLATE_SCHEMAS[route]["label"],
            "city": TEMPLATE_SCHEMAS[route]["city"],
            "employeeCount": len(employees),
            "coverages": list(plan.get("coverages") or []),
            "missingRequiredCount": sum(missing.values()),
            "missingEmployees": sum(
                bool(((employee.get("templateReports") or {}).get(route) or {}).get("missingRequired"))
                for employee in employees
            ),
            "missingFields": [{"field": field, "count": count} for field, count in missing.most_common()],
            "template": template,
            "templateUsable": template_usable,
            "status": status,
            "statusLabel": status_label,
            "ready": status == "ready",
        })
    ready = sum(group["ready"] for group in groups)
    return {
        "runId": run.get("id"),
        "status": run.get("status"),
        "groups": groups,
        "manualRoutes": manual_routes,
        "summary": {
            "templateRoutes": len(groups),
            "readyRoutes": ready,
            "blockedRoutes": len(groups) - ready,
            "manualRoutes": len(manual_routes),
            "employeeCount": int((run.get("summary") or {}).get("included") or 0),
        },
    }


def _safe_excel_text(value: Any) -> str:
    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def build_missing_export(run_id: str) -> Path:
    run = load_run(run_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = MISSING_SHEET_NAME
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    headers = ["合同主体", "办理路径", "姓名", "工号", "证件号码", "缺失字段", "字段来源", "处理建议"]
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(color="FFFFFF", bold=True, size=10)
    line = Side(style="thin", color="D7E0EA")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_index = 2
    for employee in run.get("employees") or []:
        if employee.get("decision") != "include":
            continue
        source = employee.get("source") if isinstance(employee.get("source"), dict) else {}
        report = employee.get("report") if isinstance(employee.get("report"), dict) else {}
        for route, template_report in (employee.get("templateReports") or {}).items():
            if not isinstance(template_report, dict):
                continue
            for field in template_report.get("missingRequired") or []:
                values = [
                    run.get("subject"), template_report.get("label") or route,
                    report.get("姓名"), source.get("jobNumber"), report.get("证件号码"),
                    field, template_report.get("origins", {}).get(field) or "待补充", "补齐后重新校验",
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_index, column=column, value=_safe_excel_text(value))
                    cell.border = Border(bottom=line)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.number_format = "@"
                row_index += 1
    if row_index == 2:
        sheet.cell(row=2, column=1, value="当前没有缺失的模板必填字段")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    widths = (28, 24, 12, 14, 22, 24, 16, 22)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = f"A1:H{max(2, row_index - 1)}"
    output = get_run_dir(run_id) / "待补资料.xlsx"
    workbook.save(output)
    return output


def _soffice_binary() -> str:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_SOFFICE")
    bundled = (
        Path.home()
        / ".cache/codex-runtimes/codex-primary-runtime/dependencies/bin/override/soffice"
    )
    for candidate in (configured, shutil.which("soffice"), shutil.which("libreoffice"), str(bundled)):
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RunValidationError("未找到旧版 .xls 模板转换组件，请改为上传 .xlsx 模板")


def _convert_with_soffice(source: Path, output_dir: Path, format_name: str) -> Path:
    completed = subprocess.run(
        [_soffice_binary(), "--headless", "--convert-to", format_name, "--outdir", str(output_dir), str(source)],
        check=False,
        capture_output=True,
        text=True,
        timeout=120,
    )
    expected = output_dir / f"{source.stem}.{format_name.split(':', 1)[0]}"
    if completed.returncode != 0 or not expected.exists():
        raise RunValidationError("政务模板格式转换失败")
    return expected


def _normalized_header(value: Any) -> str:
    return re.sub(r"[\s（）()：:]+", "", str(value or "")).replace("必填", "").lower()


def _header_matches(actual: Any, field: dict[str, Any]) -> bool:
    normalized = _normalized_header(actual)
    expected = _normalized_header(field.get("header") or field["name"])
    name = _normalized_header(field["name"])
    return normalized == expected or normalized == name or (name and name in normalized)


def _fill_template(
    source: Path,
    route: str,
    employees: list[dict[str, Any]],
    output_dir: Path,
) -> Path:
    schema = TEMPLATE_SCHEMAS[route]
    with tempfile.TemporaryDirectory(prefix="social-template-") as temporary:
        temporary_dir = Path(temporary)
        source_xlsx = source
        if source.suffix.lower() == ".xls":
            source_xlsx = _convert_with_soffice(source, temporary_dir, "xlsx")
        workbook = load_workbook(source_xlsx)
        sheet = workbook[schema["sheet"]] if schema["sheet"] in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        fields = list(schema["fields"])
        header_row = int(schema["headerRow"])
        mismatches = [
            field["name"]
            for column, field in enumerate(fields, start=1)
            if not _header_matches(sheet.cell(row=header_row, column=column).value, field)
        ]
        if mismatches:
            workbook.close()
            raise RunValidationError(f"模板版本不匹配：{schema['label']} 表头缺少{'、'.join(mismatches[:3])}")
        data_start = int(schema["dataStartRow"])
        for row in range(data_start, max(data_start, sheet.max_row) + 1):
            for column in range(1, len(fields) + 1):
                sheet.cell(row=row, column=column).value = None
        style_row = data_start
        for offset, employee in enumerate(employees):
            row = data_start + offset
            values = ((employee.get("templateReports") or {}).get(route) or {}).get("values") or {}
            for column, field in enumerate(fields, start=1):
                target = sheet.cell(row=row, column=column)
                if row != style_row:
                    template_cell = sheet.cell(row=style_row, column=column)
                    if template_cell.has_style:
                        target._style = copy(template_cell._style)
                    target.number_format = template_cell.number_format
                    target.alignment = copy(template_cell.alignment)
                target.value = _safe_excel_text(values.get(field["name"]))
        safe_label = re.sub(r"[^\w\u4e00-\u9fff-]+", "", schema["label"])
        xlsx_output = output_dir / f"{safe_label}_{len(employees)}人.xlsx"
        workbook.save(xlsx_output)
        workbook.close()
        if source.suffix.lower() != ".xls":
            return xlsx_output
        xls_output = _convert_with_soffice(xlsx_output, output_dir, "xls:MS Excel 97")
        desired = output_dir / f"{safe_label}_{len(employees)}人.xls"
        if xls_output != desired:
            xls_output.replace(desired)
        xlsx_output.unlink(missing_ok=True)
        return desired


def _build_summary_workbook(run: dict[str, Any], preflight: dict[str, Any], generated: list[dict[str, Any]]) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SUMMARY_SHEET_NAME
    sheet.sheet_view.showGridLines = False
    sheet.append(["社保报盘包", "生成结果"])
    sheet.append(["合同主体", run.get("subject")])
    sheet.append(["增员周期", f"{run.get('periodStart')} 至 {run.get('periodEnd')}"])
    sheet.append(["名单确认日", run.get("confirmationDate")])
    sheet.append(["生成时间", current_timestamp()])
    sheet.append([])
    sheet.append(["办理路径", "人数", "结果", "文件/原因"])
    generated_by_route = {item["route"]: item for item in generated}
    for group in preflight["groups"]:
        result = generated_by_route.get(group["route"])
        sheet.append([
            group["label"], group["employeeCount"], "已生成" if result else "未生成",
            result["filename"] if result else group["statusLabel"],
        ])
    for item in preflight["manualRoutes"]:
        sheet.append([item["label"], item["employeeCount"], "线下办理", "详见审核清单"])
    sheet["A1"].font = Font(size=16, bold=True, color="1E3A8A")
    for cell in sheet[7]:
        cell.fill = PatternFill("solid", fgColor="1E3A8A")
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 42
    output = get_run_dir(str(run.get("id") or "")) / "生成说明.xlsx"
    workbook.save(output)
    return output


def generate_report_package(run_id: str) -> dict[str, Any]:
    run = load_run(run_id)
    if run.get("status") not in {"confirmed", "generated"}:
        raise RunValidationError("请先完成人员人工确认，再生成政务报盘包")
    preflight = build_export_preflight(run)
    run_dir = get_run_dir(run_id)
    generated: list[dict[str, Any]] = []
    generation_errors: dict[str, str] = {}
    with tempfile.TemporaryDirectory(prefix="social-package-") as temporary:
        stage = Path(temporary)
        report_dir = stage / "报盘文件"
        review_dir = stage / "审核资料"
        report_dir.mkdir(parents=True)
        review_dir.mkdir(parents=True)
        for group in preflight["groups"]:
            if not group["ready"]:
                continue
            route = group["route"]
            reference = _template_reference(run, route)
            if not reference or not reference.get("path"):
                continue
            try:
                output = _fill_template(Path(reference["path"]), route, _route_employees(run, route), report_dir)
                generated.append({"route": route, "filename": output.name, "employeeCount": group["employeeCount"]})
            except RunValidationError as exc:
                generation_errors[route] = str(exc)
        audit = build_audit_export(run_id)
        missing = build_missing_export(run_id)
        summary = _build_summary_workbook(run, preflight, generated)
        shutil.copy2(audit, review_dir / audit.name)
        shutil.copy2(missing, review_dir / missing.name)
        shutil.copy2(summary, stage / summary.name)
        package_name = (
            f"社保增员报盘包_{str(run.get('periodStart') or '').replace('-', '')}"
            f"-{str(run.get('periodEnd') or '').replace('-', '')}.zip"
        )
        package_path = run_dir / package_name
        with zipfile.ZipFile(package_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(stage.rglob("*")):
                if path.is_file():
                    archive.write(path, path.relative_to(stage))
    blocked = [
        {
            "route": group["route"],
            "label": group["label"],
            "reason": generation_errors.get(group["route"], group["statusLabel"]),
        }
        for group in preflight["groups"]
        if not any(item["route"] == group["route"] for item in generated)
    ]
    run["reportPackage"] = {
        "filename": package_name,
        "generatedAt": current_timestamp(),
        "generatedRoutes": generated,
        "blockedRoutes": blocked,
        "employeeCount": int((run.get("summary") or {}).get("included") or 0),
        "partial": bool(blocked),
    }
    if generated:
        run["status"] = "generated"
    return save_run(run)


def resolve_package_download(run_id: str) -> Path:
    run = load_run(run_id)
    package = run.get("reportPackage") if isinstance(run.get("reportPackage"), dict) else {}
    filename = Path(str(package.get("filename") or "")).name
    path = get_run_dir(run_id) / filename
    if not filename or not path.is_file():
        raise RunNotFoundError("政务报盘包不存在")
    return path
