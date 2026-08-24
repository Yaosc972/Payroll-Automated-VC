from __future__ import annotations

from datetime import datetime
import json
import os
from pathlib import Path
import shutil
import subprocess
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ... import config
from .runs import (
    REPORT_FIELDS,
    TEMPLATE_FIELDS,
    RunNotFoundError,
    RunValidationError,
    current_timestamp,
    get_run_dir,
    load_run,
    save_run,
)
from .template_schemas import schema_for_route


MAX_TEMPLATE_BYTES = 20 * 1024 * 1024
ALLOWED_TEMPLATE_SUFFIXES = {".xls", ".xlsx"}
AUDIT_SHEET_NAME = "全部增员人员"
AUDIT_STATUS_LABELS = {
    "ready": "可报盘",
    "needs_review": "待人工确认",
    "excluded": "已排除",
}

AUDIT_EXPORT_GROUPS = (
    ("处理", ("序号", "处理结果")),
    ("人员与入职", (
        "姓名", "工号", "证件号码", "性别", "手机号码", "入职日期", "离职日期",
        "在职状态", "雇佣关系",
    )),
    ("主体与参保", (
        "合同主体", "工作地点", "员工考勤地点", "社保缴纳地", "参保城市",
        "社保医保原始状态", "公积金原始状态", "社保状态", "医保状态", "公积金状态",
        "办理去向",
    )),
    ("缴费与账户", (
        "社保缴交基数", "公积金缴交基数", "公积金个人比例", "电脑号", "公积金号",
    )),
    ("户籍与通讯", (
        "户籍所在地", "户口类别", "户口具体地址", "通讯地址", "邮箱",
    )),
    ("政务模板字段", (
        "户籍", "入深户时间", "民族", "岗位类别", "个人身份", "用工形式", "学历",
        "职称", "国家职业资格或职业技能等级", "医疗缴费档次", "部门名称", "户籍地类别",
        "户口所在地行政区划代码", "就业形式", "就业前身份",
    )),
    ("校验与备注", (
        "校验状态", "人工确认", "校验原因", "校验问题", "补充增员原因", "补充说明",
        "审核备注", "变动说明", "是否虚拟员工",
    )),
)
AUDIT_EXPORT_HEADERS = tuple(
    header
    for _group, group_headers in AUDIT_EXPORT_GROUPS
    for header in group_headers
)
assert set(REPORT_FIELDS).issubset(AUDIT_EXPORT_HEADERS)


def _safe_excel_text(value: Any) -> str:
    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _audit_export_rows(run: dict[str, Any]) -> tuple[list[str], list[list[str | int]]]:
    headers = list(AUDIT_EXPORT_HEADERS)
    rows: list[list[str | int]] = []
    for index, employee in enumerate(run.get("employees") or [], start=1):
        source = employee.get("source") if isinstance(employee.get("source"), dict) else {}
        supplemental = (
            employee.get("supplemental")
            if isinstance(employee.get("supplemental"), dict)
            else {}
        )
        report = employee.get("report") if isinstance(employee.get("report"), dict) else {}
        coverage_source = (
            employee.get("coverageSource")
            if isinstance(employee.get("coverageSource"), dict)
            else {}
        )
        coverage_tasks = employee.get("coverageTasks") if isinstance(employee.get("coverageTasks"), dict) else {}
        social_task = coverage_tasks.get("social") if isinstance(coverage_tasks.get("social"), dict) else {}
        medical_task = coverage_tasks.get("medical") if isinstance(coverage_tasks.get("medical"), dict) else {}
        housing_task = coverage_tasks.get("housing") if isinstance(coverage_tasks.get("housing"), dict) else {}
        route_labels = list(dict.fromkeys(
            str(task.get("routeLabel") or "").strip()
            for task in (social_task, medical_task)
            if str(task.get("routeLabel") or "").strip()
        ))
        issues = "\n".join(
            str(issue.get("message") or "").strip()
            for issue in employee.get("issues") or []
            if isinstance(issue, dict) and str(issue.get("message") or "").strip()
        )
        values = {
            "序号": index,
            "处理结果": "纳入" if employee.get("decision") == "include" else "排除",
            "姓名": report.get("姓名"),
            "工号": source.get("jobNumber"),
            "证件号码": report.get("证件号码"),
            "性别": source.get("gender"),
            "手机号码": report.get("手机号码") or source.get("mobile"),
            "入职日期": employee.get("entryDate"),
            "离职日期": source.get("lastWorkDate"),
            "在职状态": source.get("employeeStatus"),
            "雇佣关系": source.get("employType"),
            "合同主体": source.get("subject"),
            "工作地点": source.get("place"),
            "员工考勤地点": source.get("employmentPlace"),
            "社保缴纳地": (
                source.get("socialContributionPlace")
                or source.get("socialPlace")
                or coverage_source.get("socialPlace")
                or source.get("place")
            ),
            "参保城市": (
                source.get("socialPlace")
                or coverage_source.get("socialPlace")
                or source.get("place")
            ),
            "社保医保原始状态": (
                source.get("socialMedicalStatus") or coverage_source.get("socialMedicalStatus")
            ),
            "公积金原始状态": source.get("housingStatus") or coverage_source.get("housingStatus"),
            "社保状态": social_task.get("statusLabel"),
            "医保状态": medical_task.get("statusLabel"),
            "公积金状态": housing_task.get("statusLabel"),
            "办理去向": " / ".join(route_labels),
            "社保缴交基数": report.get("社保缴交基数"),
            "公积金缴交基数": report.get("公积金缴交基数"),
            "公积金个人比例": source.get("housingContributionRate"),
            "电脑号": report.get("电脑号"),
            "公积金号": report.get("公积金号"),
            "户籍所在地": source.get("birthplace"),
            "户口类别": source.get("domicileType"),
            "户口具体地址": report.get("户口具体地址") or source.get("householdAddress"),
            "通讯地址": report.get("通讯地址") or source.get("currentAddress"),
            "邮箱": source.get("email"),
            "户籍": report.get("户籍"),
            "入深户时间": report.get("入深户时间"),
            "民族": report.get("民族") or source.get("nation"),
            "岗位类别": report.get("岗位类别"),
            "个人身份": report.get("个人身份"),
            "用工形式": report.get("用工形式"),
            "学历": report.get("学历") or source.get("education"),
            "职称": report.get("职称"),
            "国家职业资格或职业技能等级": report.get("国家职业资格或职业技能等级"),
            "医疗缴费档次": report.get("医疗缴费档次"),
            "部门名称": report.get("部门名称"),
            "户籍地类别": report.get("户籍地类别"),
            "户口所在地行政区划代码": report.get("户口所在地行政区划代码"),
            "就业形式": report.get("就业形式"),
            "就业前身份": report.get("就业前身份"),
            "校验状态": AUDIT_STATUS_LABELS.get(
                str(employee.get("status") or ""), str(employee.get("status") or "")
            ),
            "人工确认": "已确认" if employee.get("confirmed") else "待确认",
            "校验原因": employee.get("reason"),
            "校验问题": issues,
            "补充增员原因": supplemental.get("label"),
            "补充说明": supplemental.get("note"),
            "审核备注": employee.get("reviewNote"),
            "变动说明": source.get("changeDescription"),
            "是否虚拟员工": source.get("virtualEmployee"),
        }
        rows.append([
            value if header == "序号" else _safe_excel_text(value)
            for header in headers
            for value in (values.get(header),)
        ])
    return headers, rows


def build_audit_export(run_id: str) -> Path:
    run = load_run(run_id)
    headers, rows = _audit_export_rows(run)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = AUDIT_SHEET_NAME
    sheet.freeze_panes = "E5"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_view.zoomScale = 80
    sheet.sheet_properties.tabColor = "1E3A8A"

    last_column = len(headers)
    last_column_letter = get_column_letter(last_column)
    title_fill = PatternFill("solid", fgColor="1E3A8A")
    title_font = Font(color="FFFFFF", bold=True, size=16)
    metadata_fill = PatternFill("solid", fgColor="EEF4FF")
    metadata_font = Font(color="344054", size=10)
    group_fill = PatternFill("solid", fgColor="DCE7F8")
    group_font = Font(color="1E3A8A", bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="284B8F")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_line = Side(style="thin", color="D7E0EA")
    group_line = Side(style="medium", color="B7C7DD")
    body_border = Border(bottom=thin_line)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = sheet.cell(row=1, column=1, value="社保增员审核清单（业务核对版）")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    employees = run.get("employees") or []
    status_counts = {
        "ready": sum(1 for employee in employees if employee.get("status") == "ready"),
        "needs_review": sum(1 for employee in employees if employee.get("status") == "needs_review"),
        "excluded": sum(1 for employee in employees if employee.get("status") == "excluded"),
    }
    metadata_blocks = (
        (1, 8, f"批次周期：{run.get('periodStart') or '-'} 至 {run.get('periodEnd') or '-'}"),
        (9, 16, f"名单确认日：{run.get('confirmationDate') or '-'}"),
        (17, 24, f"合同主体：{run.get('subject') or '全部主体'}"),
        (25, 32, (
            f"人员：{len(employees)}　可报盘：{status_counts['ready']}　"
            f"待确认：{status_counts['needs_review']}　已排除：{status_counts['excluded']}"
        )),
        (33, 40, f"数据来源：{run.get('source') or '北森'}"),
        (41, 48, f"导出时间：{datetime.now().astimezone().strftime('%Y-%m-%d %H:%M')}"),
        (49, last_column, "说明：字段按业务核对顺序排列，政务模板字段名保持原样"),
    )
    for start_column, end_column, text in metadata_blocks:
        if start_column > last_column:
            continue
        end_column = min(end_column, last_column)
        sheet.merge_cells(
            start_row=2, start_column=start_column, end_row=2, end_column=end_column
        )
        cell = sheet.cell(row=2, column=start_column, value=text)
        cell.fill = metadata_fill
        cell.font = metadata_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 24

    group_start_columns: list[int] = []
    next_column = 1
    for group_name, group_headers in AUDIT_EXPORT_GROUPS:
        start_column = next_column
        end_column = start_column + len(group_headers) - 1
        group_start_columns.append(start_column)
        sheet.merge_cells(
            start_row=3, start_column=start_column, end_row=3, end_column=end_column
        )
        cell = sheet.cell(row=3, column=start_column, value=group_name)
        cell.fill = group_fill
        cell.font = group_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=group_line, bottom=thin_line)
        next_column = end_column + 1
    sheet.row_dimensions[3].height = 22

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=4, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if column_index in group_start_columns:
            cell.border = Border(left=group_line)
    sheet.row_dimensions[4].height = 40

    status_column = headers.index("校验状态") + 1
    confirmation_column = headers.index("人工确认") + 1
    decision_column = headers.index("处理结果") + 1
    identifier_headers = {
        "工号", "证件号码", "手机号码", "电脑号", "公积金号", "邮箱",
        "户口所在地行政区划代码",
    }
    centered_headers = {
        "序号", "处理结果", "性别", "在职状态", "雇佣关系", "社保状态", "医保状态",
        "公积金状态", "公积金个人比例", "户籍", "民族", "岗位类别", "个人身份",
        "用工形式", "学历", "职称", "医疗缴费档次", "户籍地类别", "就业形式",
        "就业前身份", "校验状态", "人工确认", "是否虚拟员工",
    }
    for row_index, values in enumerate(rows, start=5):
        if row_index % 2 == 0:
            row_fill = PatternFill("solid", fgColor="F8FAFC")
        else:
            row_fill = None
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = body_border
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(
                horizontal="center" if headers[column_index - 1] in centered_headers else "left",
                vertical="top",
                wrap_text=True,
            )
            if column_index != 1:
                cell.number_format = "@"
            if headers[column_index - 1] in identifier_headers:
                cell.quotePrefix = True
            if column_index in group_start_columns:
                cell.border = Border(left=group_line, bottom=thin_line)
        sheet.row_dimensions[row_index].height = 30
        status = str(values[status_column - 1])
        status_cell = sheet.cell(row=row_index, column=status_column)
        if status == "待人工确认":
            status_cell.fill = PatternFill("solid", fgColor="FFF4D6")
            status_cell.font = Font(color="9A6700", bold=True)
        elif status == "已排除":
            status_cell.fill = PatternFill("solid", fgColor="F2F4F7")
            status_cell.font = Font(color="667085", bold=True)
        else:
            status_cell.fill = PatternFill("solid", fgColor="E7F8F2")
            status_cell.font = Font(color="0F766E", bold=True)

        confirmation_cell = sheet.cell(row=row_index, column=confirmation_column)
        if str(confirmation_cell.value or "") == "待确认":
            confirmation_cell.fill = PatternFill("solid", fgColor="FFF4D6")
            confirmation_cell.font = Font(color="9A6700", bold=True)

        decision_cell = sheet.cell(row=row_index, column=decision_column)
        if str(decision_cell.value or "") == "排除":
            decision_cell.fill = PatternFill("solid", fgColor="F2F4F7")
            decision_cell.font = Font(color="667085", bold=True)
        else:
            decision_cell.fill = PatternFill("solid", fgColor="E7F8F2")
            decision_cell.font = Font(color="0F766E", bold=True)

    sheet.auto_filter.ref = f"A4:{last_column_letter}{max(4, len(rows) + 4)}"
    preferred_widths = {
        "序号": 7, "处理结果": 10, "姓名": 12, "工号": 13, "证件号码": 22,
        "性别": 8, "手机号码": 15, "入职日期": 13, "离职日期": 13, "在职状态": 11,
        "雇佣关系": 12, "合同主体": 30, "工作地点": 18, "员工考勤地点": 18,
        "社保缴纳地": 16, "参保城市": 16, "社保医保原始状态": 24,
        "公积金原始状态": 18, "社保状态": 12, "医保状态": 12, "公积金状态": 12,
        "办理去向": 26, "社保缴交基数": 15, "公积金缴交基数": 16,
        "公积金个人比例": 15, "电脑号": 18, "公积金号": 18, "户籍所在地": 22,
        "户口类别": 12, "户口具体地址": 32, "通讯地址": 32, "邮箱": 28, "户籍": 16,
        "入深户时间": 14, "民族": 10, "岗位类别": 18, "个人身份": 12, "用工形式": 12,
        "学历": 14, "职称": 12, "国家职业资格或职业技能等级": 26,
        "医疗缴费档次": 16, "部门名称": 20, "户籍地类别": 14,
        "户口所在地行政区划代码": 24, "就业形式": 14, "就业前身份": 14,
        "校验状态": 14, "人工确认": 12, "校验原因": 26, "校验问题": 32,
        "补充增员原因": 16, "补充说明": 26, "审核备注": 26, "变动说明": 24,
        "是否虚拟员工": 15,
    }
    for column_index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = preferred_widths.get(header, 14)

    sheet.print_title_rows = "1:4"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.35
    sheet.page_margins.bottom = 0.35

    run_dir = get_run_dir(run_id)
    filename = (
        f"社保增员审核清单_{str(run.get('periodStart') or '').replace('-', '')}"
        f"-{str(run.get('periodEnd') or '').replace('-', '')}.xlsx"
    )
    output_path = run_dir / filename
    workbook.save(output_path)
    try:
        output_path.chmod(0o600)
    except OSError:
        pass
    return output_path


def _engine_dir() -> Path:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR")
    if configured:
        return Path(configured).expanduser().resolve()
    return (config.PROJECT_ROOT / "outputs" / "social-insurance-beisen-mvp-20260814").resolve()


def _node_binary() -> str:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_NODE")
    for candidate in (configured, shutil.which("node")):
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RunValidationError("未找到报盘生成所需的 Node.js 运行时")


def store_template(run_id: str, original_name: str, content: bytes, route: str = "") -> dict[str, Any]:
    run = load_run(run_id)
    if len(content) > MAX_TEMPLATE_BYTES:
        raise RunValidationError("政务模板不能超过20MB")
    suffix = Path(original_name or "").suffix.lower()
    if suffix not in ALLOWED_TEMPLATE_SUFFIXES:
        raise RunValidationError("仅支持政务系统导出的 .xls 或 .xlsx 模板")
    if not content:
        raise RunValidationError("政务模板为空")
    normalized_route = str(route or "").strip()
    available_routes = [
        str(plan.get("route") or "")
        for plan in run.get("processingPlan") or []
        if plan.get("handling") == "template" and schema_for_route(str(plan.get("route") or ""))
    ]
    available_routes = list(dict.fromkeys(available_routes))
    if not normalized_route:
        if len(available_routes) == 1:
            normalized_route = available_routes[0]
        elif "shenzhen-social-medical" in available_routes:
            normalized_route = "shenzhen-social-medical"
        else:
            raise RunValidationError("当前批次包含多个办理路径，请选择模板对应城市")
    if normalized_route not in available_routes:
        raise RunValidationError("所选政务模板办理路径不属于当前批次")
    run_dir = get_run_dir(run_id)
    template_path = run_dir / f"government-template-{normalized_route}{suffix}"
    for stale_template in run_dir.glob(f"government-template-{normalized_route}.*"):
        if stale_template != template_path:
            stale_template.unlink(missing_ok=True)
    template_path.write_bytes(content)
    try:
        template_path.chmod(0o600)
    except OSError:
        pass
    metadata = {
        "route": normalized_route,
        "filename": template_path.name,
        "originalFilename": Path(original_name).name,
        "uploadedAt": current_timestamp(),
        "size": len(content),
    }
    run.setdefault("templates", {})[normalized_route] = metadata
    if normalized_route == "shenzhen-social-medical":
        run["template"] = metadata
    run["reportFile"] = None
    run["reportPackage"] = None
    if run.get("status") == "generated":
        run["status"] = "confirmed"
    return save_run(run)


def generate_report(run_id: str) -> dict[str, Any]:
    run = load_run(run_id)
    if run.get("status") != "confirmed":
        raise RunValidationError("请先完成人员人工确认，再生成报盘")
    template = run.get("template") if isinstance(run.get("template"), dict) else None
    if not template:
        raise RunValidationError("请先导入政务系统下载的模板")
    run_dir = get_run_dir(run_id)
    template_path = run_dir / Path(str(template.get("filename") or "")).name
    if not template_path.exists():
        raise RunNotFoundError("政务模板文件不存在")
    ready_rows = [
        {"report": employee.get("report") or {}}
        for employee in run.get("employees") or []
        if employee.get("decision") == "include"
    ]
    if not ready_rows:
        raise RunValidationError("当前批次没有纳入报盘的人员")

    payload_path = run_dir / ".approved-rows.json"
    payload_path.write_text(json.dumps({"readyRows": ready_rows}, ensure_ascii=False), encoding="utf-8")
    try:
        payload_path.chmod(0o600)
    except OSError:
        pass
    bridge = Path(__file__).with_name("report_bridge.mjs")
    command = [_node_binary(), str(bridge), str(_engine_dir()), str(template_path), str(payload_path), str(run_dir)]
    try:
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=180,
            env={**os.environ, "NODE_TLS_REJECT_UNAUTHORIZED": "1"},
        )
    except subprocess.TimeoutExpired as exc:
        raise RunValidationError("报盘生成超时，请稍后重试") from exc
    finally:
        payload_path.unlink(missing_ok=True)
    if completed.returncode != 0:
        raise RunValidationError("模板校验或报盘生成失败；请确认模板版本、单位编号和人员字段")
    try:
        last_line = next(line for line in reversed(completed.stdout.splitlines()) if line.strip())
        result = json.loads(last_line)
        xls_name = Path(str(result["xls"])).name
        xlsx_name = Path(str(result["xlsx"])).name
    except (StopIteration, KeyError, json.JSONDecodeError, TypeError) as exc:
        raise RunValidationError("报盘生成结果无效") from exc
    for filename in (xls_name, xlsx_name):
        if not (run_dir / filename).exists():
            raise RunValidationError("报盘文件未完整生成")
    run["status"] = "generated"
    run["reportFile"] = {
        "filename": xls_name,
        "previewFilename": xlsx_name,
        "employeeCount": len(ready_rows),
        "generatedAt": current_timestamp(),
    }
    return save_run(run)


def resolve_download(run_id: str, filename: str) -> Path:
    run = load_run(run_id)
    report = run.get("reportFile") if isinstance(run.get("reportFile"), dict) else {}
    allowed = {Path(str(report.get(key) or "")).name for key in ("filename", "previewFilename")}
    safe_name = Path(str(filename).replace("\\", "/")).name
    if not safe_name or safe_name not in allowed:
        raise RunNotFoundError("报盘文件不存在")
    path = get_run_dir(run_id) / safe_name
    if not path.is_file():
        raise RunNotFoundError("报盘文件不存在")
    return path


def rpa_status() -> dict[str, Any]:
    return {
        "available": False,
        "state": "not_configured",
        "label": "RPA待接入",
        "reason": "尚未取得政务系统操作录屏，也未完成登录、验证码、U盾、上传与提交验证。",
        "validatedNow": ["北森同步", "网页人工确认", "政务模板导入", "报盘生成与下载"],
        "notValidated": ["政务系统登录", "模板自动下载", "批量上传", "自动提交"],
        "finalConfirmation": "manual",
    }
