from __future__ import annotations

from datetime import datetime, timedelta
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import threading
import time
from typing import Any

import httpx
from openpyxl import load_workbook

from ... import config
from .persistent_storage import (
    SocialInsuranceStorageError,
    load_json,
    persist_json,
    persistent_storage_enabled,
)
from .rule_catalog import RULE_VERSION
from .runs import RunValidationError, TEMPLATE_FIELDS


REVIEW_FILENAME = "深圳社保增员_候选与人工确认.xlsx"
MACHINE_RECORDS_FILENAME = "候选记录.json"
REVIEW_SHEET = "候选审查"
SUBJECT_CACHE_SECONDS = max(60, int(os.environ.get("SIGMA_SOCIAL_INSURANCE_SUBJECT_CACHE_SECONDS", "10800")))
_BEISEN_QUERY_LOCK = threading.Lock()
_SUBJECT_CACHE_LOCK = threading.Lock()
_SUBJECT_CACHE: dict[tuple[str, str, str], tuple[float, list[dict[str, Any]]]] = {}
_SUBJECT_CACHE_GENERATION: dict[tuple[str, str, str], int] = {}


def _connector_url() -> str:
    return os.environ.get("SIGMA_SOCIAL_INSURANCE_CONNECTOR_URL", "").strip().rstrip("/")


def _connector_token() -> str:
    return os.environ.get("SIGMA_SOCIAL_INSURANCE_CONNECTOR_TOKEN", "").strip()


def connector_status() -> dict[str, Any]:
    remote_url = _connector_url()
    if remote_url:
        ready = bool(_connector_token()) and (not os.environ.get("VERCEL") or remote_url.startswith("https://"))
        return {"mode": "remote", "configured": True, "ready": ready}
    engine_ready = (_engine_dir() / "lib" / "beisen-client.mjs").is_file()
    return {
        "mode": "local-development" if engine_ready else "not-configured",
        "configured": engine_ready,
        "ready": engine_ready and not bool(os.environ.get("VERCEL")),
    }


def _connector_cache_namespace() -> str:
    return _connector_url() or str(_engine_dir())


def _connector_call(operation: str, payload: dict[str, Any], *, timeout: float) -> dict[str, Any]:
    base_url = _connector_url()
    if not base_url:
        raise RunValidationError("北森远程连接器未配置")
    if os.environ.get("VERCEL") and not base_url.startswith("https://"):
        raise RunValidationError("云端北森连接器必须使用 HTTPS")
    token = _connector_token()
    if not token:
        raise RunValidationError("北森远程连接器授权未配置")
    try:
        with httpx.Client(timeout=timeout) as client:
            response = client.post(
                f"{base_url}/{operation.lstrip('/')}",
                headers={
                    "authorization": f"Bearer {token}",
                    "content-type": "application/json",
                },
                json=payload,
            )
        response.raise_for_status()
        result = response.json()
    except (httpx.TimeoutException, httpx.NetworkError) as exc:
        raise RunValidationError("北森连接器响应超时，请稍后重试") from exc
    except (httpx.HTTPStatusError, httpx.RemoteProtocolError, json.JSONDecodeError) as exc:
        raise RunValidationError("北森连接器调用失败，请检查云端授权与运行状态") from exc
    if not isinstance(result, dict):
        raise RunValidationError("北森连接器返回格式无效")
    return result


def _subject_cache_root() -> Path:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_RUNS_DIR")
    runs_root = Path(configured).expanduser() if configured else config.SOCIAL_INSURANCE_RUNS_DIR
    root = runs_root / "_subject_options_cache"
    root.mkdir(parents=True, exist_ok=True, mode=0o700)
    try:
        root.chmod(0o700)
    except OSError:
        pass
    return root


def _subject_cache_path(key: tuple[str, str, str]) -> Path:
    digest = hashlib.sha256("|".join(key).encode("utf-8")).hexdigest()
    return _subject_cache_root() / f"{digest}.json"


def clear_contract_subject_cache() -> None:
    with _SUBJECT_CACHE_LOCK:
        _SUBJECT_CACHE.clear()
        _SUBJECT_CACHE_GENERATION.clear()
        for path in _subject_cache_root().glob("*.json"):
            try:
                path.unlink()
            except OSError:
                pass


def _cached_contract_subjects(key: tuple[str, str, str]) -> list[dict[str, Any]] | None:
    now = time.monotonic()
    with _SUBJECT_CACHE_LOCK:
        cached = _SUBJECT_CACHE.get(key)
        if cached is not None:
            if cached[0] > now:
                return [dict(item) for item in cached[1]]
            _SUBJECT_CACHE.pop(key, None)
        path = _subject_cache_path(key)
        try:
            if persistent_storage_enabled():
                payload = load_json("subject-cache", path.stem) or {}
            else:
                payload = json.loads(path.read_text(encoding="utf-8"))
            remaining = float(payload.get("expiresAt") or 0) - time.time()
            subjects = payload.get("subjects")
        except (OSError, ValueError, TypeError, json.JSONDecodeError, SocialInsuranceStorageError):
            return None
        if remaining <= 0 or not isinstance(subjects, list):
            try:
                path.unlink()
            except OSError:
                pass
            return None
        safe_subjects = [dict(item) for item in subjects if isinstance(item, dict)]
        _SUBJECT_CACHE[key] = (time.monotonic() + remaining, safe_subjects)
        return [dict(item) for item in safe_subjects]


def _cache_contract_subjects(key: tuple[str, str, str], subjects: list[dict[str, Any]]) -> None:
    safe_subjects = [dict(item) for item in subjects]
    with _SUBJECT_CACHE_LOCK:
        _SUBJECT_CACHE[key] = (
            time.monotonic() + SUBJECT_CACHE_SECONDS,
            safe_subjects,
        )
        _SUBJECT_CACHE_GENERATION[key] = _SUBJECT_CACHE_GENERATION.get(key, 0) + 1
        path = _subject_cache_path(key)
        temporary = path.with_suffix(".tmp")
        temporary.write_text(
            json.dumps(
                {"expiresAt": time.time() + SUBJECT_CACHE_SECONDS, "subjects": safe_subjects},
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        try:
            temporary.chmod(0o600)
        except OSError:
            pass
        temporary.replace(path)
        if persistent_storage_enabled():
            try:
                persist_json(
                    "subject-cache",
                    path.stem,
                    {"expiresAt": time.time() + SUBJECT_CACHE_SECONDS, "subjects": safe_subjects},
                )
            except SocialInsuranceStorageError as exc:
                raise RunValidationError("合同主体缓存未能保存到持久化存储") from exc


def _subject_cache_generation(key: tuple[str, str, str]) -> int:
    with _SUBJECT_CACHE_LOCK:
        return _SUBJECT_CACHE_GENERATION.get(key, 0)


def cached_beisen_contract_subjects(
    *,
    period_start: str,
    period_end: str,
) -> list[dict[str, Any]] | None:
    """Return the complete contract-subject cache without triggering a live Beisen query."""
    try:
        start = datetime.strptime(period_start, "%Y-%m-%d").date()
        end = datetime.strptime(period_end, "%Y-%m-%d").date()
    except (TypeError, ValueError) as exc:
        raise RunValidationError("增员周期必须为 YYYY-MM-DD") from exc
    if start > end:
        raise RunValidationError("增员周期开始日期不能晚于结束日期")
    fixture = _fixture_payload()
    if fixture is not None:
        return _subject_options_from_records(fixture[0])
    cache_key = (f"{_connector_cache_namespace()}::{RULE_VERSION}", start.isoformat(), end.isoformat())
    return _cached_contract_subjects(cache_key)


def _engine_dir() -> Path:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_ENGINE_DIR")
    if configured:
        return Path(configured).expanduser().resolve()
    return (config.PROJECT_ROOT / "outputs" / "social-insurance-beisen-mvp-20260814").resolve()


def _node_binary() -> str:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_NODE")
    candidates = [configured, shutil.which("node")]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)
    raise RunValidationError("未找到社保规则引擎所需的 Node.js 运行时")


def _fixture_payload() -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    configured = os.environ.get("SIGMA_SOCIAL_INSURANCE_SYNC_FIXTURE")
    if not configured:
        return None
    path = Path(configured).expanduser().resolve()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RunValidationError("北森同步测试数据不可读取") from exc
    records = payload.get("records") if isinstance(payload, dict) else None
    if not isinstance(records, list):
        raise RunValidationError("北森同步测试数据缺少 records")
    summary = payload.get("sourceSummary") if isinstance(payload.get("sourceSummary"), dict) else {}
    return records, {
        "provider": "fixture",
        **summary,
        "rawApiResponseSaved": False,
        "governmentSiteAccessed": False,
    }


def _issue_from_message(message: str) -> dict[str, str]:
    value = message.strip()
    field = ""
    for candidate in ("行政区划", "医疗缴费档次", "学历", "通讯地址", "证件号码", "手机号码", "户籍"):
        if candidate in value:
            field = "户口所在地行政区划代码" if candidate == "行政区划" else candidate
            break
    info_markers = ("历史信息变更", "通讯地址差异", "不影响报盘")
    severity = "info" if any(marker in value for marker in info_markers) else "blocking"
    return {"field": field, "severity": severity, "message": value}


def _subject_options_from_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    options: dict[str, dict[str, Any]] = {}
    for record in records:
        source = record.get("source") if isinstance(record.get("source"), dict) else {}
        value = str(source.get("subject") or source.get("subjectCode") or "").strip()
        code = str(source.get("subjectCode") or "").strip()
        if not value:
            continue
        key = code or value
        option = options.setdefault(
            key,
            {"value": value, "label": value, "code": code, "candidateCount": 0},
        )
        option["candidateCount"] += 1
    return list(options.values())


def list_beisen_contract_subjects(
    *,
    period_start: str,
    period_end: str,
    force_refresh: bool = False,
) -> list[dict[str, Any]]:
    try:
        start = datetime.strptime(period_start, "%Y-%m-%d").date()
        end = datetime.strptime(period_end, "%Y-%m-%d").date()
    except (TypeError, ValueError) as exc:
        raise RunValidationError("增员周期必须为 YYYY-MM-DD") from exc
    if start > end:
        raise RunValidationError("增员周期开始日期不能晚于结束日期")

    fixture = _fixture_payload()
    if fixture is not None:
        return _subject_options_from_records(fixture[0])
    if os.environ.get("VERCEL") and not _connector_url():
        raise RunValidationError("云端北森同步必须配置远程连接器")

    engine_dir = _engine_dir()
    remote_connector = bool(_connector_url())
    if not remote_connector and not (engine_dir / "lib" / "beisen-client.mjs").exists():
        raise RunValidationError("已验证的北森社保规则引擎未配置")
    cache_key = (f"{_connector_cache_namespace()}::{RULE_VERSION}", start.isoformat(), end.isoformat())
    initial_cache_generation = _subject_cache_generation(cache_key)
    if not force_refresh:
        cached = _cached_contract_subjects(cache_key)
        if cached is not None:
            return cached
    bridge = Path(__file__).with_name("subjects_bridge.mjs")
    modified_start = (start - timedelta(days=90)).isoformat()
    modified_stop = (datetime.now() + timedelta(days=1)).date().isoformat()
    with _BEISEN_QUERY_LOCK:
        if not force_refresh:
            cached = _cached_contract_subjects(cache_key)
            if cached is not None:
                return cached
        elif _subject_cache_generation(cache_key) != initial_cache_generation:
            cached = _cached_contract_subjects(cache_key)
            if cached is not None:
                return cached
        if remote_connector:
            result = _connector_call(
                "subjects",
                {
                    "periodStart": start.isoformat(),
                    "periodEnd": end.isoformat(),
                    "modifiedStart": modified_start,
                    "modifiedStop": modified_stop,
                    "ruleVersion": RULE_VERSION,
                },
                timeout=180.0,
            )
            raw_subjects = result.get("subjects")
        else:
            try:
                completed = subprocess.run(
                    [
                        _node_binary(),
                        str(bridge),
                        str(engine_dir),
                        modified_start,
                        modified_stop,
                        start.isoformat(),
                        end.isoformat(),
                    ],
                    check=False,
                    capture_output=True,
                    text=True,
                    timeout=180,
                    env={**os.environ, "NODE_TLS_REJECT_UNAUTHORIZED": "1"},
                )
            except subprocess.TimeoutExpired as exc:
                raise RunValidationError("北森合同主体加载超时，请稍后重试") from exc
            if completed.returncode != 0:
                raise RunValidationError("北森合同主体加载失败，请检查连接器授权")
            try:
                last_line = next(line for line in reversed(completed.stdout.splitlines()) if line.strip())
                raw_subjects = json.loads(last_line).get("subjects")
            except (StopIteration, AttributeError, json.JSONDecodeError) as exc:
                raise RunValidationError("北森合同主体返回格式无效") from exc
    if not isinstance(raw_subjects, list):
        raise RunValidationError("北森合同主体返回格式无效")

    safe_options: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in raw_subjects[:100]:
        if not isinstance(item, dict):
            continue
        value = str(item.get("value") or "").strip()[:200]
        label = str(item.get("label") or value).strip()[:200]
        code = str(item.get("code") or "").strip()[:100]
        key = code or value
        if not value or key in seen:
            continue
        seen.add(key)
        try:
            candidate_count = max(0, int(item.get("candidateCount") or 0))
        except (TypeError, ValueError):
            candidate_count = 0
        safe_options.append({
            "value": value,
            "label": label or value,
            "code": code,
            "candidateCount": candidate_count,
        })
    _cache_contract_subjects(cache_key, safe_options)
    return [dict(item) for item in safe_options]


def _read_review_workbook(path: Path) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[REVIEW_SHEET]
        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        header_index = {header: index for index, header in enumerate(headers)}
        required = {"处理状态", "判断说明", *TEMPLATE_FIELDS}
        if not required.issubset(header_index):
            raise RunValidationError("候选审查表结构与已验证版本不一致")
        output: list[dict[str, Any]] = []
        status_map = {"可报盘": "ready", "待人工确认": "needs_review", "规则排除": "excluded"}
        for values in rows:
            if not any(value not in (None, "") for value in values):
                continue
            status_label = str(values[header_index["处理状态"]] or "").strip()
            reason = str(values[header_index["判断说明"]] or "").strip()
            report = {field: str(values[header_index[field]] or "").strip() for field in TEMPLATE_FIELDS}
            issues = [] if status_label == "可报盘" else [
                _issue_from_message(message)
                for message in reason.split("；")
                if message.strip() and message.strip() != "规则校验通过"
            ]
            source = {
                "subject": str(values[header_index.get("合同主体", -1)] or "").strip() if "合同主体" in header_index else "",
                "place": str(values[header_index.get("工作地点", -1)] or "").strip() if "工作地点" in header_index else "",
                "employType": str(values[header_index.get("雇佣关系", -1)] or "").strip() if "雇佣关系" in header_index else "",
            }
            output.append({
                "status": status_map.get(status_label, "needs_review"),
                "reason": reason or "需要业务确认",
                "issues": issues,
                "report": report,
                "entryDate": str(values[header_index.get("入职日期", -1)] or "").strip() if "入职日期" in header_index else "",
                "source": source,
            })
        return output
    except RunValidationError:
        raise
    except Exception as exc:  # noqa: BLE001 - translate parser details without leaking row data.
        raise RunValidationError("无法读取北森候选结果") from exc
    finally:
        if "workbook" in locals():
            workbook.close()


def _read_offline_coverage_context(path: Path | None) -> dict[str, dict[str, str]]:
    """Read only the business-status columns from a user-provided monthly source workbook."""
    if path is None or path.suffix.lower() != ".xlsx":
        return {}
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        next(rows, None)  # technical Beisen field names
        labels = [str(value or "").strip() for value in (next(rows, None) or [])]
        label_index = {label: index for index, label in enumerate(labels) if label}
        identity_index = label_index.get("身份证号码")
        status_index = next(
            (label_index[label] for label in ("社保、医保", "社保医保") if label in label_index),
            None,
        )
        housing_index = label_index.get("公积金")
        social_place_index = label_index.get("社保缴纳地")
        if identity_index is None:
            return {}
        output: dict[str, dict[str, str]] = {}
        for values in rows:
            identity = str(values[identity_index] or "").replace(" ", "").upper()
            if not identity:
                continue
            output[identity] = {
                "socialPlace": str(values[social_place_index] or "").strip() if social_place_index is not None else "",
                "socialMedicalStatus": str(values[status_index] or "").strip() if status_index is not None else "",
                "housingStatus": str(values[housing_index] or "").strip() if housing_index is not None else "",
            }
        return output
    except Exception:  # noqa: BLE001 - optional context must never block the validated sync.
        return {}
    finally:
        if "workbook" in locals():
            workbook.close()


def sync_beisen_candidates(
    *,
    period_start: str,
    period_end: str,
    confirmation_date: str,
    subject: str,
    output_dir: Path,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    fixture = _fixture_payload()
    if fixture is not None:
        return fixture
    if os.environ.get("VERCEL") and not _connector_url():
        raise RunValidationError("云端北森同步必须配置远程连接器")

    try:
        parsed_period_end = datetime.strptime(period_end, "%Y-%m-%d").date()
        parsed_confirmation_date = datetime.strptime(confirmation_date, "%Y-%m-%d").date()
    except (TypeError, ValueError) as exc:
        raise RunValidationError("名单确认日必须为 YYYY-MM-DD") from exc
    if parsed_confirmation_date < parsed_period_end:
        raise RunValidationError("名单确认日不能早于增员周期结束日")
    if _connector_url():
        result = _connector_call(
            "sync",
            {
                "periodStart": period_start,
                "periodEnd": period_end,
                "confirmationDate": confirmation_date,
                "subject": subject,
                "ruleVersion": RULE_VERSION,
            },
            timeout=280.0,
        )
        records = result.get("records")
        if not isinstance(records, list) or not all(isinstance(item, dict) for item in records):
            raise RunValidationError("北森连接器返回的候选人员格式无效")
        source_summary = result.get("sourceSummary")
        if not isinstance(source_summary, dict):
            source_summary = {}
        return records, {
            "provider": "beisen-remote-connector",
            **{
                key: value
                for key, value in source_summary.items()
                if key not in {"rawApiResponse", "records", "employees"}
            },
            "rawApiResponseSaved": False,
            "governmentSiteAccessed": False,
        }
    engine_dir = _engine_dir()
    entrypoint = engine_dir / "mvp.mjs"
    if not entrypoint.exists():
        raise RunValidationError("已验证的北森社保规则引擎未配置")
    output_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
    try:
        output_dir.chmod(0o700)
    except OSError:
        pass
    modified_start = (datetime.fromisoformat(period_start) - timedelta(days=90)).date().isoformat()
    modified_stop = (datetime.now() + timedelta(days=1)).date().isoformat()
    cutoff = f"{parsed_confirmation_date.isoformat()}T23:59:59+08:00"
    historical_baseline_source = os.environ.get("SIGMA_SOCIAL_INSURANCE_BASELINE_SOURCE_FILE")
    historical_baseline_path: Path | None = None
    source_mode = "beisen"
    if historical_baseline_source:
        historical_baseline_path = Path(historical_baseline_source).expanduser().resolve()
        if not historical_baseline_path.is_file():
            raise RunValidationError("已配置的历史月度名单基线不存在")
        if historical_baseline_path.suffix.lower() not in {".xls", ".xlsx"}:
            raise RunValidationError("历史月度名单基线必须为北森导出的 Excel 文件")
        source_mode = "hybrid"
    command = [
        _node_binary(),
        str(entrypoint),
        "--source", source_mode,
        "--entry-start", period_start,
        "--entry-end", period_end,
        "--modified-start", modified_start,
        "--modified-stop", modified_stop,
        "--cutoff", cutoff,
        "--subject", subject,
        "--output", str(output_dir),
        "--machine-json",
    ]
    if historical_baseline_path is not None:
        command.extend(["--source-file", str(historical_baseline_path)])
    dimission = os.environ.get("SIGMA_SOCIAL_INSURANCE_DIMISSION_FILE")
    if not dimission:
        raise RunValidationError("一键同步已锁定：请先配置当次最新离职快照，避免沿用历史停保数据")
    dimission_path = Path(dimission).expanduser().resolve()
    if not dimission_path.is_file():
        raise RunValidationError("已配置的离职快照不存在")
    command.extend(["--dimission", str(dimission_path)])
    template = os.environ.get("SIGMA_SOCIAL_INSURANCE_TEMPLATE_FILE")
    if template:
        command.extend(["--template", str(Path(template).expanduser().resolve())])

    with _BEISEN_QUERY_LOCK:
        try:
            completed = subprocess.run(
                command,
                cwd=engine_dir,
                check=False,
                capture_output=True,
                text=True,
                timeout=300,
                env={**os.environ, "NODE_TLS_REJECT_UNAUTHORIZED": "1"},
            )
        except subprocess.TimeoutExpired as exc:
            raise RunValidationError("北森同步超时，请稍后重试") from exc
    if completed.returncode != 0:
        raise RunValidationError("北森同步或规则计算失败，请检查连接器授权与数据源配置")

    machine_records_path = output_dir / MACHINE_RECORDS_FILENAME
    if machine_records_path.exists():
        try:
            machine_records = json.loads(machine_records_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise RunValidationError("北森同步候选结果不可读取") from exc
        if not isinstance(machine_records, list) or not all(isinstance(item, dict) for item in machine_records):
            raise RunValidationError("北森同步候选结果格式无效")
        status_map = {"可报盘": "ready", "待人工确认": "needs_review", "规则排除": "excluded"}
        coverage_context = _read_offline_coverage_context(historical_baseline_path)
        records = []
        for item in machine_records:
            report = item.get("report") if isinstance(item.get("report"), dict) else {}
            identity = str(report.get("证件号码") or "").replace(" ", "").upper()
            records.append({
                **item,
                "status": status_map.get(str(item.get("status") or ""), str(item.get("status") or "needs_review")),
                "issues": [
                    _issue_from_message(message)
                    for message in str(item.get("reason") or "").split("；")
                    if message.strip() and message.strip() != "规则校验通过"
                ],
                "coverageSource": coverage_context.get(identity, {}),
            })
    else:
        review_path = output_dir / REVIEW_FILENAME
        if not review_path.exists():
            raise RunValidationError("北森同步未生成候选结果")
        records = _read_review_workbook(review_path)
    safe_summary: dict[str, Any] = {}
    try:
        last_line = next(line for line in reversed(completed.stdout.splitlines()) if line.strip())
        raw_summary = json.loads(last_line)
        diagnostics = raw_summary.get("diagnostics") if isinstance(raw_summary.get("diagnostics"), dict) else {}
        safe_summary = {
            "provider": "beisen-open-platform",
            "candidateCount": int(raw_summary.get("candidate_count") or len(records)),
            "diagnostics": diagnostics,
        }
    except (StopIteration, json.JSONDecodeError, TypeError, ValueError):
        safe_summary = {"provider": "beisen-open-platform", "candidateCount": len(records)}
    snapshot_match = re.search(r"(20\d{6})", dimission_path.name)
    snapshot_date = snapshot_match.group(1) if snapshot_match else ""
    normalized_snapshot_date = (
        f"{snapshot_date[:4]}-{snapshot_date[4:6]}-{snapshot_date[6:8]}"
        if snapshot_date else ""
    )
    warnings = []
    if historical_baseline_path is not None:
        warnings.append("本批已合并月度原始名单基线，避免后续离职导致历史增员人员从北森当前任职结果中消失。")
    if not normalized_snapshot_date:
        warnings.append("离职数据时点校验：离职快照日期无法从文件名识别，请确认数据已覆盖到名单确认日，避免已离职人员被误纳入。")
    elif normalized_snapshot_date < period_end:
        warnings.append(f"离职数据时点校验：离职快照标识日期为{normalized_snapshot_date}，早于周期结束日；请确认数据已覆盖到名单确认日，避免已离职人员被误纳入。")
    return records, {
        **safe_summary,
        "rawApiResponseSaved": False,
        "governmentSiteAccessed": False,
        "departureRuleSource": "configured-snapshot",
        "departureSnapshotDate": normalized_snapshot_date or None,
        "confirmationDate": parsed_confirmation_date.isoformat(),
        "historicalBaselineSeedUsed": historical_baseline_path is not None,
        "historicalBaselineSourceName": historical_baseline_path.name if historical_baseline_path is not None else None,
        "warnings": warnings,
    }
