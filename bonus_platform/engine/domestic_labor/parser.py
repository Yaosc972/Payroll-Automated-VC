"""Excel file parser for payroll data."""
import io
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence
from dataclasses import dataclass
from datetime import date, datetime, time

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


@dataclass
class SheetData:
    """Sheet data container."""
    name: str
    headers: List[str]
    rows: List[Dict[str, Any]]
    row_count: int


class ExcelParser:
    """Excel file parser."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        self.workbook = None
        self._engine = None

    def load(self, password: str = None):
        """Load workbook. Supports encrypted files with password."""
        suffix = self.file_path.suffix.lower()
        if suffix == ".xls":
            if xlrd is None:
                raise ImportError("xlrd is required to read .xls files: pip install xlrd")
            if password:
                raise ValueError("老式 .xls 加密文件暂不支持密码解密，请先另存为 .xlsx 后上传。")
            self.workbook = xlrd.open_workbook(str(self.file_path))
            self._engine = "xlrd"
            return self

        if openpyxl is None:
            raise ImportError("openpyxl is required: pip install openpyxl")
        # Payroll uses cached cell values only; preserving external-link metadata can
        # spend seconds parsing large historical link caches that are never consumed.
        if password and msoffcrypto:
            # Decrypt encrypted file
            output_buffer = io.BytesIO()
            with open(self.file_path, 'rb') as f:
                ms_file = msoffcrypto.OfficeFile(f)
                ms_file.load_key(password=password)
                ms_file.decrypt(output_buffer)
            output_buffer.seek(0)
            self.workbook = openpyxl.load_workbook(
                output_buffer,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        else:
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True,
                read_only=True,
                keep_links=False,
            )
        self._engine = "openpyxl"
        return self

    def close(self):
        """Close workbook."""
        if self.workbook and self._engine == "openpyxl":
            self.workbook.close()

    def __enter__(self):
        self.load()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def get_sheet_names(self) -> List[str]:
        """Get all sheet names."""
        if not self.workbook:
            self.load()
        if self._engine == "xlrd":
            return self.workbook.sheet_names()
        return self.workbook.sheetnames

    def parse_sheet(self, sheet_name: str) -> SheetData:
        """Parse a single sheet into structured data."""
        if not self.workbook:
            self.load()

        if self._engine == "xlrd":
            return self._parse_xls_sheet(sheet_name)

        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")

        ws = self.workbook[sheet_name]
        rows = []
        headers = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h) if h else f"col_{j}" for j, h in enumerate(row)]
                continue

            row_dict = {}
            for j, value in enumerate(row):
                if j < len(headers):
                    header = headers[j]
                    if header not in row_dict or row_dict[header] in (None, ""):
                        row_dict[header] = value
            rows.append(row_dict)

        return SheetData(
            name=sheet_name,
            headers=headers,
            rows=rows,
            row_count=len(rows),
        )

    def _parse_xls_sheet(self, sheet_name: str) -> SheetData:
        sheet_names = self.workbook.sheet_names()
        if sheet_name not in sheet_names:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {sheet_names}")

        ws = self.workbook.sheet_by_name(sheet_name)
        if ws.nrows == 0:
            return SheetData(name=sheet_name, headers=[], rows=[], row_count=0)

        headers = [
            str(ws.cell_value(0, j)).strip() if ws.cell_value(0, j) not in ("", None) else f"col_{j}"
            for j in range(ws.ncols)
        ]
        rows = []
        for i in range(1, ws.nrows):
            row_dict = {}
            for j, header in enumerate(headers):
                row_dict[header] = self._read_xls_cell(ws, i, j)
            rows.append(row_dict)

        return SheetData(
            name=sheet_name,
            headers=headers,
            rows=rows,
            row_count=len(rows),
        )

    def _read_xls_cell(self, ws, row_idx: int, col_idx: int) -> Any:
        cell = ws.cell(row_idx, col_idx)
        if xlrd and cell.ctype == xlrd.XL_CELL_DATE:
            try:
                dt = xlrd.xldate.xldate_as_datetime(cell.value, self.workbook.datemode)
                if dt.time() == time(0, 0):
                    return dt.date()
                return dt
            except (ValueError, OverflowError):
                return cell.value
        if xlrd and cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
            return int(cell.value)
        return cell.value

    def parse_monthly_attendance(self, sheet_name: str = "月考勤") -> SheetData:
        """Parse monthly attendance sheet (Sheet2)."""
        return self.parse_sheet(sheet_name)

    def parse_daily_attendance(self, sheet_name: str = "日考勤") -> SheetData:
        """Parse daily attendance sheet (Sheet1)."""
        return self.parse_sheet(sheet_name)

    def parse_housing_list(self, sheet_name: str = "住宿名单") -> SheetData:
        """Parse housing list sheet (Sheet3)."""
        return self.parse_sheet(sheet_name)

    def get_attendance_month(self, sheet_name: str = "月考勤") -> Optional[str]:
        """Extract attendance month from any data row."""
        data = self.parse_sheet(sheet_name)
        if data.rows:
            return str(data.rows[0].get("考勤月份", ""))
        return None


class PayrollDataLoader:
    """Load and organize payroll data from Excel files."""

    def __init__(self, file_path: str, password: str = None):
        self.parser = ExcelParser(file_path)
        self._password = password
        self._monthly = None
        self._daily = None
        self._housing = None

    def load(self):
        """Load all sheets."""
        self.parser.load(password=self._password)
        return self

    def close(self):
        """Close parser."""
        self.parser.close()

    def __enter__(self):
        self.load()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def _normalize_row(self, row: Dict[str, Any]) -> Dict[str, Any]:
        """Normalize row data: fill defaults, convert numeric/date fields."""
        normalized = dict(row)

        # 晋江等月报使用“排休请假”表示天数，统一到工龄奖读取的天数字段。
        if normalized.get("排休请假天数") in (None, "") and normalized.get("排休请假") not in (None, ""):
            normalized["排休请假天数"] = normalized["排休请假"]

        # 补充缺失字段默认值
        defaults = {
            "早退次数": 0,
            "工伤假天数": 0,
            "入离职缺勤时数": 0,
            "迟到早退30分钟内扣款": normalized.get("迟到早退扣款") or 0,
            "餐补标准": "/",
            "外宿补贴标准": "/",
            "外宿补贴": 0,
            "餐补": 0,
            "全勤奖": 0,
            "工龄奖": 0,
            "排休请假天数": 0,
            "医疗期天数": 0,
        }
        for k, v in defaults.items():
            if k not in normalized or normalized[k] is None:
                normalized[k] = v

        # 数值字段：字符串转float
        numeric_fields = [
            "排班天数", "正班出勤天数", "法定节假日天数", "旷工天数",
            "正班迟到次数", "早退次数", "签卡次数",
            "转正前天数", "转正后工作天数", "实际在职工作日天数",
            "夜班天数", "月夜班时数", "正常加班", "双休加班", "节假日加班",
            "工作日加班费", "公休日加班费", "节假日加班费",
            "迟到6分钟内", "迟到6-20分钟内(次)", "迟到20-30分钟内(次)",
            "早退6分钟内(次)", "早退6-20分钟内(次)", "早退20-30分钟内(次)",
            "休年假小时", "事假时数", "病假时数", "调休时数",
            "旷工时数", "排休请假时数", "排休请假天数", "哺乳假小时", "请假时数",
            "婚假天数", "陪产假天数", "工伤假天数", "医疗期天数",
            "丧假天数", "产假天数", "多胞胎假天数", "剖腹产假天数",
            "流产假天数", "产检假天数", "女神假天数",
            "出差天数", "公出天数", "工作日产假天数", "年假剩余时数",
        ]
        optional_hour_alias_fields = {"旷工时数", "排休请假时数"}
        for field in numeric_fields:
            if field in optional_hour_alias_fields and field not in normalized:
                continue
            val = normalized.get(field)
            if val is not None and val != "":
                try:
                    normalized[field] = float(val)
                except (ValueError, TypeError):
                    normalized[field] = 0
            elif val is None or val == "":
                normalized[field] = 0

        # 日期字段：字符串转date
        date_fields = ["入职日期", "最后工作日", "转正日期"]
        for field in date_fields:
            val = normalized.get(field)
            if val is not None and val != "" and val != "None":
                if isinstance(val, time):
                    # Excel时间格式空单元格读出为time(0,0)，视为空
                    normalized[field] = None
                elif isinstance(val, datetime):
                    normalized[field] = val.date()
                elif isinstance(val, date):
                    normalized[field] = val
                elif isinstance(val, str):
                    try:
                        normalized[field] = datetime.strptime(val.strip(), "%Y-%m-%d").date()
                    except (ValueError, TypeError):
                        try:
                            normalized[field] = datetime.strptime(val.strip(), "%Y/%m/%d").date()
                        except (ValueError, TypeError):
                            pass

        # 整数字段
        int_fields = ["正班迟到次数", "早退次数", "签卡次数"]
        for field in int_fields:
            val = normalized.get(field)
            if isinstance(val, float):
                normalized[field] = int(val)

        return normalized

    @staticmethod
    def _has_valid_employee_id(row: Dict[str, Any]) -> bool:
        emp_id = row.get("工号")
        if emp_id is None:
            return False
        text = str(emp_id).strip()
        return bool(text) and text.lower() not in {"none", "nan", "null"} and text != "工号"

    def _normalize_valid_rows(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        return [
            normalized
            for row in rows
            for normalized in [self._normalize_row(row)]
            if self._has_valid_employee_id(normalized)
        ]

    @property
    def monthly(self) -> SheetData:
        """Get monthly attendance data."""
        if self._monthly is None:
            sheet_names = self.parser.get_sheet_names()
            # 优先查找包含"月报"或"月考勤"的sheet
            for name in sheet_names:
                if "月报" in name or "月考勤" in name:
                    raw = self.parser.parse_sheet(name)
                    # 标准化行数据
                    normalized_rows = self._normalize_valid_rows(raw.rows)
                    self._monthly = SheetData(
                        name=raw.name,
                        headers=raw.headers,
                        rows=normalized_rows,
                        row_count=len(normalized_rows),
                    )
                    self._fill_absence_from_daily()
                    return self._monthly
            # fallback: 取第一个sheet，也做normalize
            raw = self.parser.parse_sheet(sheet_names[0])
            normalized_rows = self._normalize_valid_rows(raw.rows)
            self._monthly = SheetData(
                name=raw.name,
                headers=raw.headers,
                rows=normalized_rows,
                row_count=len(normalized_rows),
            )
            self._fill_absence_from_daily()
        return self._monthly

    @property
    def daily(self) -> SheetData:
        """Get daily attendance data."""
        if self._daily is None:
            sheet_names = self.parser.get_sheet_names()
            for name in sheet_names:
                if "日" in name or "日考勤" in name:
                    self._daily = self.parser.parse_sheet(name)
                    return self._daily
        return self._daily

    @property
    def housing(self) -> Optional[SheetData]:
        """Get housing list data."""
        if self._housing is None:
            sheet_names = self.parser.get_sheet_names()
            for name in sheet_names:
                if "住宿" in name or "宿舍" in name:
                    raw = self.parser.parse_sheet(name)
                    normalized_rows = []
                    for row in raw.rows:
                        normalized = dict(row)
                        emp_id = str(normalized.get("工号", "") or "").strip()
                        if not emp_id:
                            continue
                        normalized["工号"] = emp_id
                        for canonical, aliases in {
                            "入住时间": ("入住时间", "入宿时间"),
                            "退宿时间": ("退宿时间", "离宿时间"),
                        }.items():
                            value = next((normalized.get(alias) for alias in aliases if normalized.get(alias) not in (None, "")), None)
                            if isinstance(value, datetime):
                                value = value.date()
                            normalized[canonical] = value
                            for alias in aliases:
                                if alias in normalized and isinstance(normalized[alias], datetime):
                                    normalized[alias] = normalized[alias].date()
                        normalized_rows.append(normalized)
                    self._housing = SheetData(
                        name=raw.name,
                        headers=raw.headers,
                        rows=normalized_rows,
                        row_count=len(normalized_rows),
                    )
                    return self._housing
        return self._housing

    def get_attendance_month(self) -> Optional[str]:
        """Get attendance month from data."""
        if self.monthly and self.monthly.rows:
            return str(self.monthly.rows[0].get("考勤月份", ""))
        return None

    def group_daily_by_employee(self) -> Dict[str, List[Dict[str, Any]]]:
        """Group daily attendance by employee ID."""
        if self.daily is None:
            return {}
        result = {}
        for row in self.daily.rows:
            if not self._has_valid_employee_id(row):
                continue
            emp_id = str(row.get("工号", "")).strip()
            if emp_id not in result:
                result[emp_id] = []
            result[emp_id].append(row)
        return result

    def group_housing_by_employee(self) -> Dict[str, List[Dict[str, Any]]]:
        """Group housing records by employee ID."""
        if self.housing is None:
            return {}
        result = {}
        for row in self.housing.rows:
            emp_id = str(row.get("工号", "")).strip()
            if not emp_id:
                continue
            if emp_id not in result:
                result[emp_id] = []
            result[emp_id].append(row)
        return result

    def _fill_absence_from_daily(self):
        """从日考勤汇总缺勤时数，回填月考勤中为0/None的字段。

        月考勤中的事假时数、排休请假时数、病假时数、旷工时数等字段
        通常是SUMIF公式引用日考勤。当openpyxl读不到公式缓存值时，
        需要从日考勤数据手动汇总。
        """
        if not self._monthly:
            return
        if self.daily is None:
            return

        daily_by_emp = self.group_daily_by_employee()

        # 日考勤字段 → 月考勤字段 的映射
        field_map = {
            "正班旷职": "旷工时数",
            "事假": "事假时数",
            "病假时数": "病假时数",
            "排休请假": "排休请假时数",
        }

        for row in self._monthly.rows:
            emp_id = str(row.get("工号", ""))
            daily_rows = daily_by_emp.get(emp_id, [])
            if not daily_rows:
                continue

            for daily_field, monthly_field in field_map.items():
                current_val = row.get(monthly_field)
                if current_val is not None and current_val != 0 and current_val != 0.0:
                    continue
                total = 0
                for d in daily_rows:
                    val = d.get(daily_field)
                    if val is not None:
                        try:
                            total += float(val)
                        except (ValueError, TypeError):
                            pass
                if total > 0:
                    row[monthly_field] = total

            # 入离职缺勤时数 = (排班天数 - 实际在职工作日天数) * 8
            current_ee = row.get("入离职缺勤时数")
            if (current_ee is None or current_ee == 0 or current_ee == 0.0):
                pb = row.get("排班天数", 0) or 0
                rz = row.get("实际在职工作日天数", 0) or 0
                diff = float(pb) - float(rz)
                if diff > 0:
                    row["入离职缺勤时数"] = diff * 8


class MultiFilePayrollDataLoader(PayrollDataLoader):
    """Merge monthly, daily and housing sheets across one or more workbooks."""

    def __init__(self, file_paths: Sequence[str], password: str = None):
        paths = [str(path) for path in file_paths if path]
        if not paths:
            raise ValueError("未提供考勤数据文件")
        self.parsers = [ExcelParser(path) for path in paths]
        self.parser = self.parsers[0]
        self._password = password
        self._monthly = None
        self._daily = None
        self._housing = None
        self._scanned = False
        self._source_summary: List[Dict[str, Any]] = []
        self._present_types = set()

    def load(self):
        if len(self.parsers) == 1:
            self.parsers[0].load(password=self._password)
            return self

        with ThreadPoolExecutor(
            max_workers=min(len(self.parsers), 4),
            thread_name_prefix="domestic-labor-workbook",
        ) as executor:
            futures = [
                executor.submit(parser.load, password=self._password)
                for parser in self.parsers
            ]
            for future in futures:
                future.result()
        return self

    def close(self):
        for parser in self.parsers:
            parser.close()

    @staticmethod
    def _sheet_type(name: str, headers: Sequence[str]) -> Optional[str]:
        sheet_name = str(name or "").strip()
        header_set = {str(header or "").strip() for header in headers}
        if "住宿" in sheet_name or "宿舍" in sheet_name or (
            "工号" in header_set
            and header_set.intersection({"入住时间", "入宿时间", "退宿时间", "离宿时间"})
        ):
            return "housing"
        daily_signals = header_set.intersection({"日期", "工作状态", "正班时数", "刷卡加班", "上班一", "下班一"})
        if "日考勤" in sheet_name or ("工号" in header_set and "日期" in header_set and len(daily_signals) >= 2):
            return "daily"
        monthly_signals = header_set.intersection({
            "考勤月份", "排班天数", "实际在职工作日天数", "正班出勤天数", "入职日期",
        })
        if "月考勤" in sheet_name or "月报" in sheet_name or (
            "工号" in header_set and len(monthly_signals) >= 2
        ):
            return "monthly"
        return None

    @staticmethod
    def _merge_headers(parts: Sequence[SheetData]) -> List[str]:
        headers = []
        seen = set()
        for part in parts:
            for header in part.headers:
                if header not in seen:
                    headers.append(header)
                    seen.add(header)
        return headers

    @staticmethod
    def _normalize_housing_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
        normalized_rows = []
        for row in rows:
            normalized = dict(row)
            emp_id = str(normalized.get("工号", "") or "").strip()
            if not emp_id or emp_id.lower() in {"none", "nan", "null"} or emp_id == "工号":
                continue
            normalized["工号"] = emp_id
            for canonical, aliases in {
                "入住时间": ("入住时间", "入宿时间"),
                "退宿时间": ("退宿时间", "离宿时间"),
            }.items():
                value = next(
                    (normalized.get(alias) for alias in aliases if normalized.get(alias) not in (None, "")),
                    None,
                )
                if isinstance(value, datetime):
                    value = value.date()
                normalized[canonical] = value
            normalized_rows.append(normalized)
        return normalized_rows

    def _scan(self):
        if self._scanned:
            return
        parts = {"monthly": [], "daily": [], "housing": []}
        for parser in self.parsers:
            recognized = []
            for sheet_name in parser.get_sheet_names():
                raw = parser.parse_sheet(sheet_name)
                sheet_type = self._sheet_type(sheet_name, raw.headers)
                if not sheet_type:
                    continue
                parts[sheet_type].append(raw)
                self._present_types.add(sheet_type)
                recognized.append({
                    "sheet": sheet_name,
                    "type": sheet_type,
                    "row_count": raw.row_count,
                })
            self._source_summary.append({
                "file_name": parser.file_path.name,
                "sheets": recognized,
            })

        monthly_rows = self._normalize_valid_rows([
            row for part in parts["monthly"] for row in part.rows
        ])
        daily_rows = [
            row for part in parts["daily"] for row in part.rows if self._has_valid_employee_id(row)
        ]
        housing_rows = self._normalize_housing_rows([
            row for part in parts["housing"] for row in part.rows
        ])
        self._monthly = SheetData(
            name="月考勤",
            headers=self._merge_headers(parts["monthly"]),
            rows=monthly_rows,
            row_count=len(monthly_rows),
        )
        self._daily = SheetData(
            name="日考勤",
            headers=self._merge_headers(parts["daily"]),
            rows=daily_rows,
            row_count=len(daily_rows),
        ) if "daily" in self._present_types else None
        self._housing = SheetData(
            name="住宿名单",
            headers=self._merge_headers(parts["housing"]),
            rows=housing_rows,
            row_count=len(housing_rows),
        ) if "housing" in self._present_types else None
        self._scanned = True
        self._fill_absence_from_daily()

    @property
    def monthly(self) -> SheetData:
        self._scan()
        return self._monthly

    @property
    def daily(self) -> Optional[SheetData]:
        self._scan()
        return self._daily

    @property
    def housing(self) -> Optional[SheetData]:
        self._scan()
        return self._housing

    @staticmethod
    def _month_digits(value: Any) -> str:
        digits = "".join(char for char in str(value or "") if char.isdigit())
        return digits[:6]

    def validate_inputs(self, engines: Sequence[str], attendance_month: str = "") -> Dict[str, Any]:
        self._scan()
        if not self.monthly.rows:
            raise ValueError("未识别到月考勤数据，请检查Sheet名称或表头")

        employee_ids = [str(row.get("工号", "")).strip() for row in self.monthly.rows]
        duplicate_ids = sorted(emp_id for emp_id, count in Counter(employee_ids).items() if count > 1)
        if duplicate_ids:
            sample = "、".join(duplicate_ids[:5])
            raise ValueError(f"月考勤存在重复工号：{sample}，请勿重复上传同一份数据")

        requested_month = self._month_digits(attendance_month)
        source_months = {
            self._month_digits(row.get("考勤月份"))
            for row in self.monthly.rows
            if self._month_digits(row.get("考勤月份"))
        }
        if requested_month and source_months and source_months != {requested_month}:
            source_text = "、".join(sorted(source_months))
            raise ValueError(f"月考勤月份为{source_text}，与批次月份{requested_month}不一致")

        engine_set = set(engines)
        has_daily = self.daily is not None and bool(self.daily.rows)
        if "canbu" in engine_set:
            has_dongguan = any("东莞" in str(row.get("工作地区", "")) for row in self.monthly.rows)
            if has_dongguan and not has_daily:
                raise ValueError("东莞餐补核算缺少日考勤数据")
        if "waisu_butie" in engine_set:
            if not has_daily:
                raise ValueError("外宿补贴核算缺少日考勤数据")
            if self.housing is None:
                raise ValueError("外宿补贴核算缺少住宿名单；即使当月无人住宿，也请上传带表头的空名单")

        return {
            "file_count": len(self.parsers),
            "monthly_rows": self.monthly.row_count,
            "daily_rows": self.daily.row_count if self.daily else 0,
            "housing_rows": self.housing.row_count if self.housing else 0,
            "present_types": sorted(self._present_types),
            "sources": self._source_summary,
        }


class DongguanDataLoader:
    """Load and organize 东莞 daily attendance data."""

    def __init__(self, file_path: str, password: str = None):
        self.parser = ExcelParser(file_path)
        self._password = password
        self._monthly = None
        self._daily = None

    def load(self):
        """Load all sheets."""
        self.parser.load(password=self._password)
        return self

    def close(self):
        """Close parser."""
        self.parser.close()

    def __enter__(self):
        self.load()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    @property
    def monthly(self) -> SheetData:
        """Get monthly attendance data with 餐补标准."""
        if self._monthly is None:
            sheet_names = self.parser.get_sheet_names()
            for name in sheet_names:
                if "月考勤" in name:
                    self._monthly = self.parser.parse_sheet(name)
                    return self._monthly
        return self._monthly

    @property
    def daily(self) -> SheetData:
        """Get daily attendance data."""
        if self._daily is None:
            sheet_names = self.parser.get_sheet_names()
            for name in sheet_names:
                if "日考勤" in name:
                    self._daily = self.parser.parse_sheet(name)
                    return self._daily
        return self._daily

    def get_meal_allowance_map(self) -> Dict[str, str]:
        """Get 餐补标准 by employee ID from monthly data.

        Returns:
            Dict[employee_id, 餐补标准] - only for employees with valid 餐补标准
        """
        if self.monthly is None:
            return {}
        result = {}
        for row in self.monthly.rows:
            emp_id = str(row.get("工号", ""))
            std = str(row.get("餐补标准", ""))
            if emp_id and emp_id not in result and std and std not in ["/", "", "None", "餐补标准"]:
                result[emp_id] = std
        return result

    def get_missing_std_employees(self) -> List[Dict[str, str]]:
        """Get 东莞 employees without 餐补标准 (excluding "/").

        "/" means not eligible, which is valid. Only flag truly empty/None values.

        Returns:
            List of {employee_id, employee_name} for 东莞 employees missing 餐补标准
        """
        if self.monthly is None:
            return []
        result = []
        seen = set()
        for row in self.monthly.rows:
            emp_id = str(row.get("工号", ""))
            region = str(row.get("工作地区", ""))
            std = row.get("餐补标准")
            std_str = str(std) if std else ""
            # 只检查东莞地区，"/"表示不享有（有效状态），空/None才是缺失
            if (emp_id and emp_id not in seen
                    and region == "东莞"
                    and std_str in ["", "None"]):
                result.append({
                    "employee_id": emp_id,
                    "employee_name": str(row.get("姓名", "")),
                })
                seen.add(emp_id)
        return result

    def group_daily_by_employee(self) -> Dict[str, List[Dict[str, Any]]]:
        """Group daily attendance by employee ID."""
        if self.daily is None:
            return {}
        result = {}
        for row in self.daily.rows:
            emp_id = str(row.get("工号", ""))
            if emp_id not in result:
                result[emp_id] = []
            result[emp_id].append(row)
        return result
