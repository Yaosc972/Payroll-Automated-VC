"""Excel export for domestic labor calculation results."""
from pathlib import Path
from typing import Any, Dict, List
from datetime import date, datetime, time

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
except ImportError:
    openpyxl = None


class ExcelExporter:
    """Export calculation results to Excel."""

    BODY_STYLE_NAME = "domestic_labor_body"
    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    WARNING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    OK_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    SUBJECTS = [
        ("quanqinjiang", "全勤奖"),
        ("canbu", "餐补"),
        ("waisu_butie", "外宿补贴"),
        ("gonglingjiang", "工龄奖"),
        ("gangwei_butie", "岗位补贴"),
        ("gaowen_butie", "高温补贴"),
        ("yeban_butie", "夜班补贴"),
    ]
    CANBU_HEADERS = [
        "工号", "姓名", "工作地区", "部门", "岗位", "餐补口径",
    ]
    WAISU_HEADERS = [
        "工号", "姓名", "工作地区", "部门", "岗位", "外宿补贴口径",
        "在职天数", "住宿扣除天数", "外宿补贴天数", "缺勤时数",
        "补贴标准", "应发外宿补贴", "异常/提示",
    ]
    YEBAN_SUMMARY_HEADERS = [
        "工号", "姓名", "工作地区", "部门", "岗位", "正常核算日", "暂算需确认日", "无需补贴日",
        "异常未计金额日", "应发夜班补贴", "核算结果", "需处理事项",
    ]
    YEBAN_HEADERS = [
        "工号", "姓名", "工作地区", "岗位", "出勤日期", "班次", "当日结果", "业务原因",
        "上班打卡", "下班打卡", "计薪上班", "计薪下班", "夜班时长（小时）",
        "晚上休息扣除（小时）", "早上休息扣除（小时）", "休息扣除合计（小时）",
        "当日夜班补贴", "需处理事项",
    ]
    GANGWEI_HEADERS = [
        "工号", "姓名", "工作地区", "部门", "岗位", "资格判断", "岗位补贴标准", "排班天数",
        "缺勤合计时数", "扣减天数", "岗位补贴计发天数", "应发岗位补贴", "核算状态", "需处理事项",
    ]
    GAOWEN_SUMMARY_HEADERS = [
        "工号", "姓名", "工作地区", "部门", "岗位", "测温网点", "高温出勤天数",
        "应发高温补贴", "核算状态", "需处理事项",
    ]
    GAOWEN_HEADERS = [
        "工号", "姓名", "工作地区", "岗位", "出勤日期", "班次", "测温网点",
        "最高测温（℃）", "实际高温出勤时长（小时）", "当日结果", "业务原因", "当日高温补贴",
    ]
    GAOWEN_REASON_LABELS = {
        "calculated": "同仓同日同班次达到33℃，按实际出勤折算",
        "temperature_below_33": "同班次最高温度未达到33℃",
        "no_matching_temperature": "没有同仓同日同班次测温记录",
        "actual_attendance_zero": "正班为0且仅有0.5小时内残留刷卡，按无实际出勤处理",
        "no_actual_attendance": "没有可计发的实际出勤时长",
        "outside_high_temperature_season": "不在6月至10月高温津贴期间",
        "employee_or_position_excluded": "命中地区固定排除人员或岗位规则",
        "measurement_site_unresolved": "无法识别员工对应测温网点",
        "attendance_shift_unresolved": "无法从考勤识别白班或夜班",
        "invalid_attendance_date": "出勤日期缺失或格式错误",
    }
    YEBAN_STATUS_LABELS = {
        "calculated": "正常核算",
        "calculated_review": "已暂算，需复核",
        "calculated_pending": "已暂算，待确认",
        "manual_review": "考勤异常，不计补贴",
        "pending_rule": "待确认后重算",
        "excluded": "无需补贴",
    }
    YEBAN_REASON_LABELS = {
        "generic_rule": "按通用夜班规则计算",
        "invalid_attendance_date": "出勤日期缺失或格式错误",
        "missing_punch": "员工缺勤（考勤异常）",
        "implausible_duration": "上下班时长超出合理范围",
        "no_effective_attendance": "取整后没有有效出勤时段",
        "no_night_overlap": "当天未覆盖夜班时段",
        "no_scheduled_night_work": "当天排班无需计算夜班补贴",
        "invalid_break_period": "班次休息时间配置错误",
        "partial_break_overlap": "实际出勤只覆盖部分休息时段",
        "negative_effective_duration": "扣除休息后没有有效夜班时长",
        "three_am_shift_pending": "凌晨3点班早退口径待确认",
        "work_area_scope_pending": "工作地区口径待确认",
        "jinjiang_special_list_unconfirmed": "晋江特殊名单尚未确认",
        "jinjiang_piecework_excluded": "晋江计件岗位不享有夜班补贴",
        "jinjiang_special_list_excluded": "属于晋江不享有夜班补贴名单",
        "jinjiang_gatekeeper_excluded": "晋江门禁岗位不享有夜班补贴",
        "shift_break_config_missing": "班次休息时间尚未维护",
    }
    GENERAL_HEADERS = [
        "工号", "姓名", "工作地区", "部门", "岗位",
        "全勤奖", "餐补", "外宿补贴", "工龄奖", "岗位补贴", "高温补贴", "夜班补贴", "应发合计",
        "规则命中", "计算公式", "关键输入", "中间值", "计算步骤", "异常/提示",
    ]

    def __init__(self, output_path: str):
        if openpyxl is None:
            raise ImportError("openpyxl is required: pip install openpyxl")
        self.output_path = Path(output_path)

    def export(
        self,
        results: List[Dict[str, Any]],
        attendance_month: str,
        summary: Dict[str, Any] = None,
    ) -> str:
        """Export results to Excel.

        Args:
            results: List of calculation result records
            attendance_month: Attendance month (YYYYMM)
            summary: Summary statistics

        Returns:
            Output file path
        """
        cleaned_results = [result for result in results if self._has_valid_employee_id(result)]
        wb = openpyxl.Workbook()
        self._register_workbook_styles(wb)

        # 1. 详情页
        ws_detail = wb.active
        if self._is_yeban_only(cleaned_results):
            ws_detail.title = "核算汇总"
            self._write_yeban_summary_sheet(ws_detail, cleaned_results)
            ws_daily = wb.create_sheet("每日明细")
            self._write_yeban_detail_sheet(ws_daily, cleaned_results)
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(self.output_path))
            wb.close()
            return str(self.output_path)
        if self._is_gaowen_only(cleaned_results):
            ws_detail.title = "核算汇总"
            self._write_gaowen_summary_sheet(ws_detail, cleaned_results)
            ws_daily = wb.create_sheet("每日明细")
            self._write_gaowen_detail_sheet(ws_daily, cleaned_results)
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(self.output_path))
            wb.close()
            return str(self.output_path)
        if self._is_gangwei_only(cleaned_results):
            ws_detail.title = "岗位补贴核算结果"
            self._write_gangwei_detail_sheet(ws_detail, cleaned_results)
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(self.output_path))
            wb.close()
            return str(self.output_path)

        ws_detail.title = "计算详情"
        self._write_detail_sheet(ws_detail, cleaned_results)

        if (
            self._is_canbu_only(cleaned_results)
            or self._is_waisu_only(cleaned_results)
            or self._is_yeban_only(cleaned_results)
        ):
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(self.output_path))
            wb.close()
            return str(self.output_path)

        # 2. 汇总页
        ws_summary = wb.create_sheet("汇总统计")
        self._write_summary_sheet(ws_summary, attendance_month, summary, cleaned_results)

        # 3. 异常页
        warnings = [r for r in cleaned_results if r.get("warnings") or r.get("exceptions")]
        if warnings:
            ws_warnings = wb.create_sheet("异常记录")
            self._write_warnings_sheet(ws_warnings, warnings)

        # 保存
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(self.output_path))
        wb.close()

        return str(self.output_path)

    def _write_detail_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write detail sheet."""
        if self._is_canbu_only(results):
            self._write_canbu_detail_sheet(ws, results)
            return
        if self._is_waisu_only(results):
            self._write_waisu_detail_sheet(ws, results)
            return
        if self._is_yeban_only(results):
            self._write_yeban_detail_sheet(ws, results)
            return
        if self._is_gangwei_only(results):
            self._write_gangwei_detail_sheet(ws, results)
            return
        if self._is_gaowen_only(results):
            self._write_gaowen_detail_sheet(ws, results)
            return

        headers = self.GENERAL_HEADERS

        # 写表头
        for col, header in enumerate(headers, 1):
            self._write_header_cell(ws, 1, col, header)

        # 写数据
        for row_idx, result in enumerate(results, 2):
            audit = self._first_audit(result)
            inputs = audit.get("inputs", {}) if audit else {}
            intermediate = audit.get("intermediate_values", {}) if audit else {}
            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                inputs.get("工作地区", intermediate.get("工作地区", "")),
                result.get("department", ""),
                inputs.get("岗位名称", intermediate.get("岗位名称", "")),
                self._number(result.get("quanqinjiang", 0)),
                self._number(result.get("canbu", 0)),
                self._number(result.get("waisu_butie", 0)),
                self._number(result.get("gonglingjiang", 0)),
                self._number(result.get("gangwei_butie", 0)),
                self._number(result.get("gaowen_butie", 0)),
                self._number(result.get("yeban_butie", 0)),
                self._number(result.get("total", 0)),
                audit.get("rule_name", "") if audit else "",
                audit.get("formula", "") if audit else "",
                self._format_mapping(inputs),
                self._format_mapping(intermediate),
                self._format_steps(audit.get("steps", []) if audit else []),
                self._format_warning(result),
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_idx, col, value)

            # 有警告的行标黄
            if result.get("warnings") or result.get("exceptions"):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.WARNING_FILL

        # 设置列宽
        self._set_widths(ws, [14, 12, 12, 18, 14, 10, 10, 12, 10, 12, 12, 12, 10, 22, 24, 38, 38, 46, 42])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_canbu_detail_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write canbu detail sheet for business reconciliation."""
        daily_columns = self._canbu_daily_column_count(results)
        headers = [
            *self.CANBU_HEADERS,
            *[f"{day:02d}日餐补" for day in range(1, daily_columns + 1)],
            "餐补合计",
        ]
        for col, header in enumerate(headers, 1):
            self._write_header_cell(ws, 1, col, header)

        for row_idx, result in enumerate(results, 2):
            detail = self._subject_detail(result, "canbu")
            audit = self._subject_audit(result, "canbu")
            inputs = audit.get("inputs", {}) if audit else {}
            intermediate = audit.get("intermediate_values", {}) if audit else {}
            daily_amounts = detail.get("日餐补明细", [])
            final_amount = self._number(result.get("canbu", detail.get("amount", 0)))
            work_area = inputs.get("工作地区", intermediate.get("工作地区", ""))
            position = inputs.get("岗位名称", intermediate.get("岗位名称", ""))
            rule_name = detail.get("地区规则", audit.get("rule_name", "") if audit else "")
            daily_values = self._canbu_daily_values(work_area, daily_amounts, daily_columns)

            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                work_area,
                result.get("department", ""),
                position,
                rule_name,
                *daily_values,
                final_amount,
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_idx, col, value)

            if final_amount <= 0:
                ws.cell(row=row_idx, column=len(headers)).fill = self.WARNING_FILL
            if result.get("warnings") or result.get("exceptions"):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.WARNING_FILL

        widths = [14, 12, 12, 18, 14, 18] + [10] * daily_columns + [12]
        self._set_widths(ws, widths)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_waisu_detail_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write housing allowance details for payroll review."""
        for col, header in enumerate(self.WAISU_HEADERS, 1):
            self._write_header_cell(ws, 1, col, header)

        for row_idx, result in enumerate(results, 2):
            detail = self._subject_detail(result, "waisu_butie")
            audit = self._subject_audit(result, "waisu_butie")
            inputs = audit.get("inputs", {}) if audit else {}
            intermediate = audit.get("intermediate_values", {}) if audit else {}
            final_amount = self._number(result.get("waisu_butie", detail.get("amount", 0)))
            work_area = inputs.get("工作地区", intermediate.get("工作地区", ""))
            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                work_area,
                result.get("department", ""),
                result.get("position", inputs.get("岗位名称", intermediate.get("岗位名称", ""))),
                f"{work_area}外宿补贴" if work_area else "外宿补贴",
                detail.get("在职天数", intermediate.get("在职天数", "")),
                detail.get("住宿扣除天数", intermediate.get("住宿扣除天数", "")),
                detail.get("外宿补贴天数", intermediate.get("外宿补贴天数", "")),
                detail.get("缺勤时数", intermediate.get("缺勤时数", "")),
                detail.get("补贴标准", intermediate.get("补贴标准", "")),
                final_amount,
                self._format_waisu_note(result, detail),
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_idx, col, value)

            if final_amount <= 0 or result.get("warnings") or result.get("exceptions"):
                for col in range(1, len(self.WAISU_HEADERS) + 1):
                    ws.cell(row=row_idx, column=col).fill = self.WARNING_FILL

        self._set_widths(ws, [14, 12, 12, 18, 14, 24, 12, 14, 14, 12, 12, 14, 42])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_yeban_detail_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write one business-readable row per employee attendance day."""
        for col, header in enumerate(self.YEBAN_HEADERS, 1):
            self._write_header_cell(ws, 1, col, header)

        row_index = 2
        for result in results:
            detail = self._subject_detail(result, "yeban_butie")
            audit = self._subject_audit(result, "yeban_butie")
            inputs = audit.get("inputs", {}) if audit else {}
            daily_results = detail.get("daily_results") or [{}]
            for daily in daily_results:
                values = [
                    result.get("employee_id", ""),
                    result.get("employee_name", ""),
                    inputs.get("工作地区", ""),
                    inputs.get("岗位名称", ""),
                    self._date_value(daily.get("attendance_date")),
                    daily.get("shift_code", ""),
                    self._night_shift_status_label(daily.get("status")),
                    self._night_shift_reason_label(daily.get("reason_code")),
                    self._time_text(daily.get("raw_start")),
                    self._time_text(daily.get("raw_end")),
                    self._minutes_as_clock(daily.get("rounded_start_minutes")),
                    self._minutes_as_clock(daily.get("rounded_end_minutes")),
                    self._minutes_as_hours(daily.get("night_minutes")),
                    self._minutes_as_hours(daily.get("evening_break_minutes")),
                    self._minutes_as_hours(daily.get("morning_break_minutes")),
                    self._minutes_as_hours(daily.get("break_minutes")),
                    self._number(daily.get("amount")),
                    self._night_shift_daily_action(daily),
                ]
                for col, value in enumerate(values, 1):
                    self._write_body_cell(ws, row_index, col, value)
                ws.cell(row=row_index, column=5).number_format = "yyyy-mm-dd"
                ws.cell(row=row_index, column=13).number_format = "0.00"
                ws.cell(row=row_index, column=14).number_format = "0.00"
                ws.cell(row=row_index, column=15).number_format = "0.00"
                ws.cell(row=row_index, column=16).number_format = "0.00"
                ws.cell(row=row_index, column=17).number_format = "0.00"
                if daily.get("status") in {
                    "manual_review", "pending_rule", "calculated_review", "calculated_pending"
                }:
                    for col in range(1, len(self.YEBAN_HEADERS) + 1):
                        ws.cell(row=row_index, column=col).fill = self.WARNING_FILL
                row_index += 1

        self._set_widths(
            ws,
            [14, 12, 12, 14, 13, 12, 18, 30, 12, 12, 12, 13, 16, 21, 21, 21, 16, 38],
        )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_gangwei_detail_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write one payroll-readable position allowance row per employee."""
        for col, header in enumerate(self.GANGWEI_HEADERS, 1):
            self._write_header_cell(ws, 1, col, header)

        for row_index, result in enumerate(results, 2):
            detail = self._subject_detail(result, "gangwei_butie")
            audit = self._subject_audit(result, "gangwei_butie")
            inputs = audit.get("inputs", {}) if audit else {}
            issue_text = self._format_warning(result)
            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                inputs.get("工作地区", ""),
                result.get("department", ""),
                inputs.get("岗位名称", ""),
                detail.get("资格判断", ""),
                self._number(detail.get("岗位补贴标准", 0)),
                self._number(detail.get("排班天数", 0)),
                self._number(detail.get("缺勤合计时数", 0)),
                self._number(detail.get("扣减天数", 0)),
                self._number(detail.get("岗位补贴计发天数", 0)),
                self._number(result.get("gangwei_butie", detail.get("amount", 0))),
                "需确认" if issue_text else "核算完成",
                issue_text or "无需处理",
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_index, col, value)
            for col in range(7, 13):
                ws.cell(row=row_index, column=col).number_format = "0.00"
            if issue_text:
                for col in range(1, len(self.GANGWEI_HEADERS) + 1):
                    ws.cell(row=row_index, column=col).fill = self.WARNING_FILL
            else:
                ws.cell(row=row_index, column=13).fill = self.OK_FILL

        self._set_widths(ws, [14, 12, 12, 18, 18, 20, 16, 12, 16, 12, 18, 16, 12, 50])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_gaowen_summary_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write one employee-level high-temperature allowance summary row."""
        for col, header in enumerate(self.GAOWEN_SUMMARY_HEADERS, 1):
            self._write_header_cell(ws, 1, col, header)

        for row_index, result in enumerate(results, 2):
            detail = self._subject_detail(result, "gaowen_butie")
            audit = self._subject_audit(result, "gaowen_butie")
            inputs = audit.get("inputs", {}) if audit else {}
            issue_text = self._format_warning(result)
            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                inputs.get("工作地区", ""),
                result.get("department", ""),
                inputs.get("岗位名称", ""),
                detail.get("测温网点", ""),
                detail.get("高温出勤天数", 0),
                self._number(result.get("gaowen_butie", detail.get("amount", 0))),
                "需确认" if issue_text else "核算完成",
                issue_text or "无需处理",
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_index, col, value)
            ws.cell(row=row_index, column=8).number_format = "0.00"
            if issue_text:
                for col in range(1, len(self.GAOWEN_SUMMARY_HEADERS) + 1):
                    ws.cell(row=row_index, column=col).fill = self.WARNING_FILL
            else:
                ws.cell(row=row_index, column=9).fill = self.OK_FILL

        self._set_widths(ws, [14, 12, 12, 18, 16, 28, 16, 16, 12, 50])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_gaowen_detail_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write one business-readable high-temperature row per attendance day."""
        for col, header in enumerate(self.GAOWEN_HEADERS, 1):
            self._write_header_cell(ws, 1, col, header)

        row_index = 2
        for result in results:
            detail = self._subject_detail(result, "gaowen_butie")
            audit = self._subject_audit(result, "gaowen_butie")
            inputs = audit.get("inputs", {}) if audit else {}
            for daily in detail.get("daily_results") or [{}]:
                reason_code = str(daily.get("reason_code") or "")
                values = [
                    result.get("employee_id", ""),
                    result.get("employee_name", ""),
                    inputs.get("工作地区", ""),
                    inputs.get("岗位名称", ""),
                    self._date_value(daily.get("attendance_date")),
                    daily.get("shift", ""),
                    daily.get("site", ""),
                    self._number(daily.get("temperature")),
                    self._number(daily.get("attendance_hours")),
                    "正常核算" if daily.get("status") == "calculated" else "不计补贴",
                    self.GAOWEN_REASON_LABELS.get(reason_code, reason_code or "原因待确认"),
                    self._number(daily.get("amount")),
                ]
                for col, value in enumerate(values, 1):
                    self._write_body_cell(ws, row_index, col, value)
                ws.cell(row=row_index, column=5).number_format = "yyyy-mm-dd"
                for col in (8, 9, 12):
                    ws.cell(row=row_index, column=col).number_format = "0.00"
                row_index += 1

        self._set_widths(ws, [14, 12, 12, 16, 13, 10, 28, 16, 25, 14, 40, 16])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_yeban_summary_sheet(self, ws, results: List[Dict[str, Any]]):
        """Write one employee-level summary row with explicit follow-up actions."""
        for col, header in enumerate(self.YEBAN_SUMMARY_HEADERS, 1):
            self._write_header_cell(ws, 1, col, header)

        for row_index, result in enumerate(results, 2):
            detail = self._subject_detail(result, "yeban_butie")
            audit = self._subject_audit(result, "yeban_butie")
            inputs = audit.get("inputs", {}) if audit else {}
            calculated_days = int(detail.get("calculated_days") or 0)
            review_days = int(detail.get("review_calculated_days") or 0)
            excluded_days = int(detail.get("excluded_days") or 0)
            unpriced_days = int(detail.get("unpriced_review_days") or 0)
            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                inputs.get("工作地区", ""),
                result.get("department", ""),
                inputs.get("岗位名称", ""),
                calculated_days,
                review_days,
                excluded_days,
                unpriced_days,
                self._number(result.get("yeban_butie", detail.get("amount", 0))),
                "金额已核算",
                self._night_shift_employee_action(review_days, unpriced_days),
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_index, col, value)
            ws.cell(row=row_index, column=10).number_format = "0.00"
            if review_days or unpriced_days:
                for col in range(1, len(self.YEBAN_SUMMARY_HEADERS) + 1):
                    ws.cell(row=row_index, column=col).fill = self.WARNING_FILL
            else:
                ws.cell(row=row_index, column=11).fill = self.OK_FILL

        self._set_widths(ws, [14, 12, 12, 18, 14, 13, 16, 14, 13, 16, 14, 42])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_summary_sheet(self, ws, attendance_month: str, summary: Dict[str, Any], results: List[Dict[str, Any]]):
        """Write summary sheet."""
        ws.cell(row=1, column=1, value="AI薪酬核算汇总").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"考勤月份: {attendance_month}")
        ws.cell(row=3, column=1, value=f"生成时间: {date.today().isoformat()}")

        # 汇总统计
        row = 5
        stats = [
            ("项目", "人数", "金额"),
            ("全勤奖", self._count_nonzero(results, "quanqinjiang"), self._sum_field(results, "quanqinjiang")),
            ("餐补", self._count_nonzero(results, "canbu"), self._sum_field(results, "canbu")),
            ("外宿补贴", self._count_nonzero(results, "waisu_butie"), self._sum_field(results, "waisu_butie")),
            ("工龄奖", self._count_nonzero(results, "gonglingjiang"), self._sum_field(results, "gonglingjiang")),
            ("岗位补贴", self._count_nonzero(results, "gangwei_butie"), self._sum_field(results, "gangwei_butie")),
            ("高温补贴", self._count_nonzero(results, "gaowen_butie"), self._sum_field(results, "gaowen_butie")),
            ("夜班补贴", self._count_nonzero(results, "yeban_butie"), self._sum_field(results, "yeban_butie")),
            ("合计", len(results), self._sum_field(results, "total")),
        ]

        for r, data in enumerate(stats):
            for c, val in enumerate(data):
                cell = ws.cell(row=row + r, column=c + 1, value=val)
                cell.border = self.BORDER
                if r == 0:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
        self._set_widths(ws, [18, 12, 14])

    def _write_warnings_sheet(self, ws, warnings: List[Dict[str, Any]]):
        """Write warnings sheet."""
        headers = ["工号", "姓名", "异常等级", "异常科目", "异常说明", "建议动作", "原始提示"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        for row_idx, result in enumerate(warnings, 2):
            exception = self._first_exception(result)
            values = [
                result.get("employee_id", ""),
                result.get("employee_name", ""),
                exception.get("level", "") if exception else "",
                exception.get("subject", "") if exception else "",
                exception.get("message", "") if exception else "",
                exception.get("suggested_action", "") if exception else "",
                result.get("warnings", ""),
            ]
            for col, value in enumerate(values, 1):
                self._write_body_cell(ws, row_idx, col, value)

        self._set_widths(ws, [14, 12, 12, 12, 52, 44, 52])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    @staticmethod
    def _count_nonzero(results: List[Dict], field: str) -> int:
        return sum(1 for r in results if r.get(field, 0) > 0)

    @staticmethod
    def _sum_field(results: List[Dict], field: str) -> float:
        return round(sum(r.get(field, 0) for r in results), 2)

    @staticmethod
    def _has_valid_employee_id(result: Dict[str, Any]) -> bool:
        employee_id = str(result.get("employee_id", "") or "").strip()
        return employee_id.lower() not in {"", "none", "nan", "null"}

    @classmethod
    def _is_canbu_only(cls, results: List[Dict[str, Any]]) -> bool:
        return cls._present_subjects(results) == {"canbu"}

    @classmethod
    def _is_waisu_only(cls, results: List[Dict[str, Any]]) -> bool:
        return cls._present_subjects(results) == {"waisu_butie"}

    @classmethod
    def _is_yeban_only(cls, results: List[Dict[str, Any]]) -> bool:
        return cls._present_subjects(results) == {"yeban_butie"}

    @classmethod
    def _is_gangwei_only(cls, results: List[Dict[str, Any]]) -> bool:
        return cls._present_subjects(results) == {"gangwei_butie"}

    @classmethod
    def _is_gaowen_only(cls, results: List[Dict[str, Any]]) -> bool:
        return cls._present_subjects(results) == {"gaowen_butie"}

    @classmethod
    def _present_subjects(cls, results: List[Dict[str, Any]]) -> set:
        if not results:
            return set()
        return {
            key
            for result in results
            for key, _label in cls.SUBJECTS
            if result.get(key, 0) or key in (result.get("subject_details") or {})
        }

    @classmethod
    def _subject_detail(cls, result: Dict[str, Any], subject: str) -> Dict[str, Any]:
        payload = (result.get("subject_details") or {}).get(subject) or {}
        details = payload.get("details") or {}
        if "amount" not in details and "amount" in payload:
            details = {**details, "amount": payload.get("amount")}
        return details

    @classmethod
    def _subject_audit(cls, result: Dict[str, Any], subject: str) -> Dict[str, Any]:
        payload = (result.get("subject_details") or {}).get(subject) or {}
        details = payload.get("details") or {}
        return payload.get("audit_explanation") or details.get("audit_explanation") or {}

    @classmethod
    def _first_audit(cls, result: Dict[str, Any]) -> Dict[str, Any]:
        for key, _label in cls.SUBJECTS:
            audit = cls._subject_audit(result, key)
            if audit:
                return audit
        return {}

    @staticmethod
    def _first_exception(result: Dict[str, Any]) -> Dict[str, Any]:
        exceptions = result.get("exceptions") or []
        return exceptions[0] if exceptions else {}

    @staticmethod
    def _number(value: Any) -> Any:
        if value in ("", None):
            return ""
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return value

    @staticmethod
    def _date_value(value: Any) -> Any:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value or "").strip()
        if not text:
            return ""
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return text[:10]

    @staticmethod
    def _time_text(value: Any) -> str:
        if isinstance(value, datetime):
            return value.strftime("%H:%M")
        if isinstance(value, time):
            return value.strftime("%H:%M")
        if isinstance(value, (int, float)) and 0 <= float(value) < 1:
            total_minutes = round(float(value) * 24 * 60)
            return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
        text = str(value or "").strip()
        if not text:
            return ""
        parts = text.split(":")
        if len(parts) >= 2:
            try:
                return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
            except ValueError:
                pass
        return text

    @staticmethod
    def _minutes_as_clock(value: Any) -> str:
        if value in ("", None):
            return ""
        try:
            total_minutes = int(round(float(value)))
        except (TypeError, ValueError):
            return ""
        day_offset, minute_of_day = divmod(total_minutes, 24 * 60)
        clock = f"{minute_of_day // 60:02d}:{minute_of_day % 60:02d}"
        if day_offset == 1:
            return f"次日{clock}"
        if day_offset > 1:
            return f"第{day_offset + 1}日{clock}"
        return clock

    @staticmethod
    def _minutes_as_hours(value: Any) -> Any:
        if value in ("", None):
            return ""
        try:
            return round(float(value) / 60, 2)
        except (TypeError, ValueError):
            return ""

    @classmethod
    def _night_shift_status_label(cls, status: Any) -> str:
        text = str(status or "").strip()
        return cls.YEBAN_STATUS_LABELS.get(text, "状态待确认" if text else "")

    @classmethod
    def _night_shift_reason_label(cls, reason_code: Any) -> str:
        text = str(reason_code or "").strip()
        return cls.YEBAN_REASON_LABELS.get(text, "原因待确认" if text else "")

    @classmethod
    def _night_shift_daily_action(cls, daily: Dict[str, Any]) -> str:
        status = str(daily.get("status") or "")
        reason = cls._night_shift_reason_label(daily.get("reason_code"))
        if status == "manual_review":
            if daily.get("reason_code") == "missing_punch":
                return "员工当天缺勤，按考勤异常处理，不计夜班补贴"
            return f"补充当天考勤信息后重新核算：{reason}"
        if status == "pending_rule":
            return f"确认业务口径后重新核算：{reason}"
        if status in {"calculated_review", "calculated_pending"}:
            return f"金额已暂算，请复核：{reason}"
        return "无需处理"

    @staticmethod
    def _night_shift_employee_action(review_days: int, unpriced_days: int) -> str:
        actions = []
        if review_days:
            actions.append(f"确认{review_days}天暂算结果")
        if unpriced_days:
            actions.append(f"查看{unpriced_days}天异常未计金额原因")
        return "；".join(actions) if actions else "无需处理"

    @classmethod
    def _effective_days(cls, daily_amounts: Any, daily_standard: Any) -> Any:
        if not isinstance(daily_amounts, list):
            return ""
        standard = cls._number(daily_standard)
        if not isinstance(standard, (int, float)) or standard <= 0:
            return ""
        return round(sum(cls._number(item) or 0 for item in daily_amounts) / standard, 2)

    @classmethod
    def _canbu_daily_column_count(cls, results: List[Dict[str, Any]]) -> int:
        counts = []
        for result in results:
            detail = cls._subject_detail(result, "canbu")
            audit = cls._subject_audit(result, "canbu")
            inputs = audit.get("inputs", {}) if audit else {}
            intermediate = audit.get("intermediate_values", {}) if audit else {}
            work_area = str(inputs.get("工作地区", intermediate.get("工作地区", "")) or "")
            daily_amounts = detail.get("日餐补明细", [])
            if "东莞" in work_area and isinstance(daily_amounts, list):
                counts.append(len(daily_amounts))
        return max(counts, default=0)

    @classmethod
    def _canbu_daily_values(cls, work_area: Any, daily_amounts: Any, daily_columns: int) -> List[Any]:
        if daily_columns <= 0:
            return []
        if "东莞" not in str(work_area or "") or not isinstance(daily_amounts, list):
            return [""] * daily_columns
        values = [cls._number(value) for value in daily_amounts[:daily_columns]]
        return values + [""] * (daily_columns - len(values))

    @classmethod
    def _format_mapping(cls, mapping: Dict[str, Any]) -> str:
        if not mapping:
            return ""
        parts = []
        for key, value in mapping.items():
            if isinstance(value, list):
                text = f"{key}: {len(value)}项"
            else:
                text = f"{key}: {cls._format_value(value)}"
            parts.append(text)
        return "\n".join(parts)

    @classmethod
    def _format_value(cls, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "是" if value else "否"
        if isinstance(value, float):
            return f"{value:.2f}".rstrip("0").rstrip(".")
        return str(value)

    @staticmethod
    def _format_steps(steps: List[Any]) -> str:
        if not steps:
            return ""
        return "\n".join(f"{idx}. {step}" for idx, step in enumerate(steps, 1))

    @staticmethod
    def _format_warning(result: Dict[str, Any]) -> str:
        pieces = []
        if result.get("warnings"):
            pieces.append(str(result.get("warnings")))
        for exception in result.get("exceptions") or []:
            message = exception.get("message") if isinstance(exception, dict) else str(exception)
            if message:
                pieces.append(message)
        return "\n".join(pieces)

    @classmethod
    def _format_waisu_note(cls, result: Dict[str, Any], detail: Dict[str, Any]) -> str:
        warning = cls._format_warning(result)
        if warning:
            return warning
        return str(detail.get("reason") or "")

    def _write_header_cell(self, ws, row: int, col: int, value: Any) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.BORDER

    def _write_body_cell(self, ws, row: int, col: int, value: Any) -> None:
        cell = ws.cell(row=row, column=col, value=value)
        cell.style = self.BODY_STYLE_NAME

    def _register_workbook_styles(self, wb) -> None:
        """Register reusable styles once instead of rebuilding them per cell."""
        if self.BODY_STYLE_NAME in wb.named_styles:
            return
        wb.add_named_style(NamedStyle(
            name=self.BODY_STYLE_NAME,
            border=self.BORDER,
            alignment=Alignment(vertical='top', wrap_text=True),
        ))

    @staticmethod
    def _set_widths(ws, widths: List[int]) -> None:
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
