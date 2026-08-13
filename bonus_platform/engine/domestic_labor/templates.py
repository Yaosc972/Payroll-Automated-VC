"""Excel模板生成器 - 为每个计算引擎生成带说明和示例的模板"""
from pathlib import Path
from typing import Dict, List, Any
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 样式定义
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
REQUIRED_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
DESC_FONT = Font(name="微软雅黑", size=10, color="888888", italic=True)
EXAMPLE_FONT = Font(name="微软雅黑", size=10, color="4CAF50")
NORMAL_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# 引擎模板定义
ENGINE_TEMPLATES: Dict[str, Dict[str, Any]] = {
    "quanqinjiang": {
        "name": "全勤奖计算模板",
        "description": "用于计算员工全勤奖（100元/人/月）",
        "columns": [
            {"name": "工号", "desc": "员工工号，如 OWHN2313", "required": True, "example": "OWHN2313"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "何俊伟"},
            {"name": "考勤月份", "desc": "格式YYYYMM，如202603", "required": True, "example": "202603"},
            {"name": "入职日期", "desc": "格式YYYY-MM-DD", "required": True, "example": "2023-05-15"},
            {"name": "最后工作日", "desc": "在职员工留空，离职员工填写日期", "required": False, "example": ""},
            {"name": "旷工天数", "desc": "默认0", "required": False, "example": "0"},
            {"name": "正班迟到次数", "desc": "迟到总次数，用于兼容原月报", "required": False, "example": "0"},
            {"name": "迟到6分钟内(次)", "desc": "最多豁免3次；与6-20分钟迟到不可同时出现", "required": False, "example": "2"},
            {"name": "迟到6-20分钟内(次)", "desc": "最多豁免1次；与6分钟内迟到不可同时出现", "required": False, "example": "0"},
            {"name": "迟到20-30分钟内(次)", "desc": "出现即不享有全勤奖", "required": False, "example": "0"},
            {"name": "早退次数", "desc": "默认0", "required": False, "example": "0"},
            {"name": "签卡次数", "desc": "默认0", "required": False, "example": "0"},
            {"name": "工伤假天数", "desc": "默认0", "required": False, "example": "0"},
            {"name": "事假时数", "desc": "事假总时数（小时）", "required": False, "example": "0"},
            {"name": "病假时数", "desc": "病假总时数（小时）", "required": False, "example": "0"},
            {"name": "入离职缺勤时数", "desc": "因入离职导致的缺勤时数", "required": False, "example": "0"},
            {"name": "迟到早退30分钟内扣款", "desc": "有扣款>0则不享全勤奖", "required": False, "example": "0"},
        ],
        "example_extra": [
            {"工号": "OWHN0424", "姓名": "韩录阳", "考勤月份": "202603", "入职日期": "2021-08-10",
             "最后工作日": "", "旷工天数": "0", "正班迟到次数": "2", "迟到6分钟内(次)": "2",
             "迟到6-20分钟内(次)": "0", "迟到20-30分钟内(次)": "0", "早退次数": "0", "签卡次数": "0",
             "工伤假天数": "0", "事假时数": "0", "病假时数": "0", "入离职缺勤时数": "0", "迟到早退30分钟内扣款": "0"},
        ],
    },
    "canbu": {
        "name": "餐补计算模板",
        "description": "用于计算员工餐补（按工作地区、部门、岗位适用平台规则）",
        "columns": [
            {"name": "工号", "desc": "员工工号", "required": True, "example": "OWHN2313"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "何俊伟"},
            {"name": "工作地区", "desc": "东莞/嘉善/义乌/晋江；按地区适用平台规则", "required": True, "example": "东莞"},
            {"name": "一级部门名称", "desc": "东莞需命中寮步区或莞深操作", "required": True, "example": "莞深操作"},
            {"name": "岗位名称", "desc": "用于判断是否享有餐补", "required": True, "example": "操作员"},
            {"name": "餐补标准", "desc": "历史字段，当前不作为发放资格入口", "required": False, "example": ""},
            {"name": "出勤天数", "desc": "历史字段，东莞按日考勤工作状态和时数逐日计算", "required": False, "example": "22"},
            {"name": "正班时数合计", "desc": "历史字段，东莞按日考勤逐日计算", "required": False, "example": "176"},
            {"name": "排班天数", "desc": "嘉善/义乌餐补月报折算分母", "required": False, "example": "22"},
            {"name": "实际在职工作日天数", "desc": "嘉善/义乌餐补月报折算基础天数", "required": False, "example": "22"},
            {"name": "事假时数", "desc": "嘉善/义乌按事假时数/8折算扣减", "required": False, "example": "0"},
            {"name": "病假时数", "desc": "嘉善/义乌按病假时数/8后40%扣减", "required": False, "example": "0"},
            {"name": "旷工天数", "desc": "嘉善/义乌按旷工天数扣减", "required": False, "example": "0"},
        ],
        "example_extra": [
            {"工号": "OWHN0424", "姓名": "韩录阳", "工作地区": "东莞", "一级部门名称": "莞深操作",
             "岗位名称": "操作员", "餐补标准": "", "出勤天数": "22", "正班时数合计": "176"},
        ],
    },
    "waisu_butie": {
        "name": "外宿补贴计算模板",
        "description": "用于按地区、岗位、住宿名单、入离职和缺勤规则计算外宿补贴（默认150元/月）",
        "columns": [
            {"name": "工号", "desc": "员工工号", "required": True, "example": "OWHN2313"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "何俊伟"},
            {"name": "工作地区", "desc": "东莞/嘉善/义乌/晋江；按地区适用平台规则", "required": True, "example": "嘉善"},
            {"name": "一级部门名称", "desc": "员工一级部门", "required": False, "example": "华东操作"},
            {"name": "岗位名称", "desc": "用于判断是否享有外宿补贴", "required": True, "example": "操作员"},
            {"name": "考勤月份", "desc": "格式YYYYMM", "required": True, "example": "202603"},
            {"name": "入职日期", "desc": "格式YYYY-MM-DD", "required": True, "example": "2023-05-15"},
            {"name": "最后工作日", "desc": "在职留空", "required": False, "example": ""},
            {"name": "事假时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "病假时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "旷工时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "排休请假时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "入离职缺勤时数", "desc": "小时", "required": False, "example": "0"},
        ],
        "example_extra": [
            {"工号": "OWHN0424", "姓名": "韩录阳", "工作地区": "嘉善", "一级部门名称": "华东操作",
             "岗位名称": "操作员", "考勤月份": "202603",
             "入职日期": "2021-08-10", "最后工作日": "", "事假时数": "0", "病假时数": "0",
             "旷工时数": "0", "排休请假时数": "0", "入离职缺勤时数": "0"},
        ],
    },
    "gangwei_butie": {
        "name": "岗位补贴核算模板",
        "description": "用于按地区、岗位名称、排班天数和缺勤时数核算岗位补贴；职级不参与计算",
        "columns": [
            {"name": "工号", "desc": "员工工号", "required": True, "example": "OWHN14187"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "陈康"},
            {"name": "考勤月份", "desc": "格式YYYYMM", "required": True, "example": "202607"},
            {"name": "一级部门名称", "desc": "员工一级部门", "required": False, "example": "莞深操作"},
            {"name": "工作地区", "desc": "东莞/嘉善/义乌/晋江", "required": True, "example": "东莞"},
            {"name": "岗位名称", "desc": "按岗位名称匹配补贴资格和标准；不读取职级", "required": True, "example": "民航初级安检员"},
            {"name": "排班天数", "desc": "当月岗位补贴折算分母", "required": True, "example": "23"},
            {"name": "实际在职工作日天数", "desc": "用于自动计算入离职缺勤时数", "required": False, "example": "23"},
            {"name": "事假时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "排休请假时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "病假时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "旷工时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "休年假小时", "desc": "小时", "required": False, "example": "0"},
            {"name": "女神假天数", "desc": "天数；核算时每1天折算8小时", "required": False, "example": "0"},
            {"name": "其他假时数（带薪）", "desc": "小时", "required": False, "example": "0"},
            {"name": "调休时数", "desc": "小时", "required": False, "example": "0"},
            {"name": "入离职缺勤时数", "desc": "已有非零值优先；否则按（排班天数-实际在职工作日天数）×8自动计算", "required": False, "example": "0"},
        ],
        "example_extra": [],
    },
    "gaowen_butie": {
        "name": "高温补贴核算模板",
        "description": "用于上传月考勤、日考勤和测温登记；按同仓同日同班次测温及实际出勤逐日核算",
        "columns": [
            {"name": "工号", "desc": "员工工号", "required": True, "example": "OWHN001"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "张三"},
            {"name": "考勤月份", "desc": "格式YYYYMM", "required": True, "example": "202607"},
            {"name": "工作地区", "desc": "东莞/嘉善/义乌/晋江", "required": True, "example": "东莞"},
            {"name": "岗位名称", "desc": "地区固定岗位范围所需；职级和领色不参与", "required": True, "example": "操作员"},
            {"name": "一级部门名称", "desc": "用于识别员工实际测温网点", "required": False, "example": "寮步区"},
            {"name": "二级部门名称", "desc": "用于识别员工实际测温网点", "required": False, "example": "中国操作部"},
            {"name": "三级部门名称", "desc": "用于识别员工实际测温网点", "required": False, "example": "华南1号枢纽2"},
            {"name": "四级部门名称", "desc": "用于识别员工实际测温网点", "required": False, "example": "寮步操作组"},
            {"name": "五级部门名称", "desc": "风控支持人员用于识别实际仓库", "required": False, "example": "寮步安全组"},
        ],
        "example_extra": [],
    },
    "yeban_butie": {
        "name": "夜班补贴核算模板",
        "description": "用于上传月/日考勤；班次休息由平台维护，晋江特殊名单单独上传",
        "columns": [
            {"name": "工号", "desc": "员工工号", "required": True, "example": "OWHN2313"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "何俊伟"},
            {"name": "考勤月份", "desc": "格式YYYYMM", "required": True, "example": "202608"},
            {"name": "工作地区", "desc": "东莞/嘉善/义乌/晋江", "required": True, "example": "东莞"},
            {"name": "一级部门名称", "desc": "员工一级部门", "required": False, "example": "莞深操作"},
            {"name": "二级部门名称", "desc": "员工二级部门", "required": False, "example": "中国操作部"},
            {"name": "岗位名称", "desc": "固定线下规则所需岗位信息", "required": True, "example": "操作员"},
        ],
        "example_extra": [],
    },
    "gonglingjiang": {
        "name": "工龄奖计算模板",
        "description": "用于计算员工工龄奖（按工龄×标准，有上限）",
        "columns": [
            {"name": "工号", "desc": "员工工号", "required": True, "example": "OWHN2313"},
            {"name": "姓名", "desc": "员工姓名", "required": True, "example": "何俊伟"},
            {"name": "工作地区", "desc": "匹配东莞、嘉善、义乌或晋江地区规则", "required": True, "example": "东莞"},
            {"name": "二级部门名称", "desc": "识别操作、第四纵队揽收、FBU及兼容区域归属", "required": True, "example": "中国操作部"},
            {"name": "岗位名称", "desc": "决定是否有工龄奖资格", "required": True, "example": "操作员"},
            {"name": "入职日期", "desc": "格式YYYY-MM-DD", "required": True, "example": "2023-05-15"},
            {"name": "考勤月份", "desc": "格式YYYYMM", "required": True, "example": "202603"},
            {"name": "请假时数", "desc": "事假+病假+旷工+排休（小时）", "required": False, "example": "0"},
            {"name": "排班天数", "desc": "当月排班天数", "required": True, "example": "26"},
            {"name": "实际在职工作日天数", "desc": "用于折算入离职缺勤", "required": False, "example": "26"},
            {"name": "备注", "desc": "如有特殊备注（如全月事假）", "required": False, "example": ""},
        ],
        "example_extra": [
            {"工号": "OWHN0424", "姓名": "韩录阳", "工作地区": "东莞", "二级部门名称": "中国操作部", "岗位名称": "操作员",
             "入职日期": "2021-08-10", "考勤月份": "202603", "请假时数": "0", "排班天数": "26",
             "实际在职工作日天数": "26", "备注": ""},
        ],
    },
}


WAISU_SHEETS = {
    "月考勤": [
        {"name": "工号", "desc": "员工工号", "required": True},
        {"name": "姓名", "desc": "员工姓名", "required": True},
        {"name": "考勤月份", "desc": "格式YYYYMM，如202605", "required": True},
        {"name": "工作地区", "desc": "东莞/嘉善/义乌/晋江", "required": True},
        {"name": "一级部门名称", "desc": "员工一级部门", "required": False},
        {"name": "岗位名称", "desc": "用于判断外宿补贴资格", "required": True},
        {"name": "入职日期", "desc": "格式YYYY-MM-DD", "required": True},
        {"name": "最后工作日", "desc": "在职员工留空", "required": False},
        {"name": "排班天数", "desc": "晋江及入离职缺勤折算使用", "required": False},
        {"name": "实际在职工作日天数", "desc": "晋江及入离职缺勤折算使用", "required": False},
        {"name": "入离职缺勤时数", "desc": "小时", "required": False},
        {"name": "事假时数", "desc": "小时", "required": False},
        {"name": "病假时数", "desc": "小时", "required": False},
        {"name": "旷工时数", "desc": "小时", "required": False},
        {"name": "旷工天数", "desc": "天", "required": False},
        {"name": "排休请假时数", "desc": "小时", "required": False},
        {"name": "排休请假天数", "desc": "天", "required": False},
        {"name": "休年假小时", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "调休时数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "哺乳假小时", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "婚假天数", "desc": "嘉善/义乌及晋江缺勤口径使用", "required": False},
        {"name": "陪产假天数", "desc": "嘉善/义乌及晋江缺勤口径使用", "required": False},
        {"name": "工伤假天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "医疗期天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "丧假天数", "desc": "嘉善/义乌及晋江缺勤口径使用", "required": False},
        {"name": "产假天数", "desc": "嘉善/义乌及晋江缺勤口径使用", "required": False},
        {"name": "多胞胎假天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "剖腹产假天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "流产假天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "产检假天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
        {"name": "女神假天数", "desc": "嘉善/义乌缺勤口径包含该字段", "required": False},
    ],
    "日考勤": [
        {"name": "日期", "desc": "格式YYYY-MM-DD", "required": True},
        {"name": "工号", "desc": "员工工号", "required": True},
        {"name": "姓名", "desc": "员工姓名", "required": True},
        {"name": "工作地区", "desc": "用于地区识别", "required": True},
        {"name": "岗位名称", "desc": "月考勤未提供岗位时用于补充", "required": False},
        {"name": "上班一", "desc": "第一段上班打卡时间", "required": False},
        {"name": "下班一", "desc": "第一段下班打卡时间", "required": False},
    ],
    "住宿名单": [
        {"name": "工号", "desc": "员工工号", "required": True},
        {"name": "姓名", "desc": "员工姓名", "required": True},
        {"name": "入住时间", "desc": "格式YYYY-MM-DD；无住宿记录的员工不填写", "required": True},
        {"name": "退宿时间", "desc": "格式YYYY-MM-DD；仍在宿员工留空", "required": False},
    ],
}


YEBAN_ATTENDANCE_SHEETS = {
    "月考勤": ENGINE_TEMPLATES["yeban_butie"]["columns"],
    "日考勤": [
        {"name": "日期", "desc": "格式YYYY-MM-DD", "required": True},
        {"name": "工号", "desc": "员工工号", "required": True},
        {"name": "姓名", "desc": "员工姓名", "required": True},
        {"name": "工作地区", "desc": "用于地区规则匹配", "required": True},
        {"name": "岗位名称", "desc": "固定线下规则所需岗位信息", "required": True},
        {"name": "计时", "desc": "计时/计件属性；晋江按日识别计件转换", "required": False},
        {"name": "班次编号", "desc": "用于匹配平台班次休息表", "required": True},
        {"name": "上班一", "desc": "第一段上班打卡时间", "required": True},
        {"name": "下班一", "desc": "第一段下班打卡时间", "required": True},
    ],
}


GAOWEN_ATTENDANCE_SHEETS = {
    "月考勤": ENGINE_TEMPLATES["gaowen_butie"]["columns"],
    "日考勤": [
        {"name": "出勤日期", "desc": "格式YYYY-MM-DD", "required": True},
        {"name": "工号", "desc": "员工工号", "required": True},
        {"name": "姓名", "desc": "员工姓名", "required": True},
        {"name": "班次名称", "desc": "用于识别白班/夜班", "required": False},
        {"name": "班次时间段", "desc": "如19:00-28:00，用于识别白班/夜班", "required": True},
        {"name": "正班时数", "desc": "当日正班小时数", "required": True},
        {"name": "刷卡加班", "desc": "当日刷卡加班小时数", "required": True},
        {"name": "实际上班时数", "desc": "明确为0时不发高温补贴", "required": False},
    ],
    "测温登记": [
        {"name": "温度测温点", "desc": "测温地点说明", "required": False},
        {"name": "班次日期", "desc": "测温归属日期，格式YYYY-MM-DD", "required": True},
        {"name": "测温班次", "desc": "白班或夜班", "required": True},
        {"name": "测温网点", "desc": "平台固定网点名称", "required": True},
        {"name": "测温温度", "desc": "摄氏温度，达到33℃才计发", "required": True},
    ],
}


def _format_input_sheet(ws, columns: List[Dict[str, Any]]) -> None:
    """创建可直接录入的数据表；字段说明使用批注，避免被解析成员工数据。"""
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        requirement = "必填" if col["required"] else "选填"
        cell.comment = Comment(f"{requirement}：{col['desc']}", "Sigma Workbench")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(
            max(len(col["name"]) + 4, 14), 24
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"


def _generate_waisu_template() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, columns in WAISU_SHEETS.items():
        _format_input_sheet(wb.create_sheet(sheet_name), columns)
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _generate_yeban_template() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, columns in YEBAN_ATTENDANCE_SHEETS.items():
        _format_input_sheet(wb.create_sheet(sheet_name), columns)
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _generate_gaowen_template() -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, columns in GAOWEN_ATTENDANCE_SHEETS.items():
        _format_input_sheet(wb.create_sheet(sheet_name), columns)
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_template(engine_key: str) -> bytes:
    """生成指定引擎的Excel模板，返回文件字节流"""
    if engine_key not in ENGINE_TEMPLATES:
        raise ValueError(f"未知引擎: {engine_key}，可选: {list(ENGINE_TEMPLATES.keys())}")
    if engine_key == "waisu_butie":
        return _generate_waisu_template()
    if engine_key == "yeban_butie":
        return _generate_yeban_template()
    if engine_key == "gaowen_butie":
        return _generate_gaowen_template()

    template = ENGINE_TEMPLATES[engine_key]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template["name"]

    columns = template["columns"]

    # 第1行：表头
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col["name"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # 第2行：列说明
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col["desc"])
        cell.font = DESC_FONT
        cell.fill = REQUIRED_FILL if col["required"] else OPTIONAL_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN_BORDER

    # 第3行起：示例数据
    for row_idx, row_data in enumerate(template.get("example_extra", []), 3):
        for col_idx, col in enumerate(columns, 1):
            val = row_data.get(col["name"], "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    # 列宽自适应
    for col_idx, col in enumerate(columns, 1):
        max_len = max(len(col["name"]), len(col["desc"]) // 2, 12)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # 冻结表头
    ws.freeze_panes = "A3"

    # 保存到字节流
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_template_info(engine_key: str) -> Dict[str, Any]:
    """获取模板信息（不生成文件）"""
    if engine_key not in ENGINE_TEMPLATES:
        raise ValueError(f"未知引擎: {engine_key}")
    template = ENGINE_TEMPLATES[engine_key]
    return {
        "engine": engine_key,
        "name": template["name"],
        "description": template["description"],
        "columns": [
            {"name": c["name"], "desc": c["desc"], "required": c["required"]}
            for c in template["columns"]
        ],
    }
