"""夜班补贴平台基线与按月维护配置。"""
from __future__ import annotations

import json
import os
import re
from copy import deepcopy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Dict, Iterable, List, Mapping, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from ...config import DOMESTIC_LABOR_RUNS_DIR


NIGHT_SHIFT_CONFIG_DIR = DOMESTIC_LABOR_RUNS_DIR.parent / "domestic_labor_configs" / "night_shift"
BASELINE_SHIFT_BREAKS_PATH = Path(__file__).parent / "data" / "night_shift_breaks.json"
MONTH_PATTERN = re.compile(r"^\d{6}$")

# 线下需要按月维护的只有晋江额外排除人员。计件岗、门禁由固定规则自动排除；
# 轻松岗位等无法从考勤字段稳定识别的人员，才需要进入本名单。
JINJIANG_INPUT_SHEET = "晋江不享有名单"
JINJIANG_LEGACY_SHEET = "晋江特殊名单"
JINJIANG_REASON_OPTIONS = ("轻松岗位", "其他线下确认不享有")
BREAK_CATEGORY_EVENING = "晚上休息"
BREAK_CATEGORY_MORNING = "早上休息"
BREAK_CATEGORY_OTHER = "其他休息"
BREAK_CATEGORY_OPTIONS = (
    BREAK_CATEGORY_EVENING,
    BREAK_CATEGORY_MORNING,
    BREAK_CATEGORY_OTHER,
)
CONFIG_SHEETS = {
    JINJIANG_INPUT_SHEET: [
        ("工号", "employee_id", True, "员工工号"),
        ("姓名", "employee_name", True, "员工姓名"),
        ("排除原因", "reason", True, "从下拉项中选择"),
        ("生效日期", "start_date", True, "YYYY-MM-DD"),
        ("失效日期", "end_date", False, "留空表示持续有效"),
        ("备注", "note", False, "选择其他原因时必填线下依据"),
    ],
}

HEADER_FILL = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
THIN_BORDER = Border(**{
    side: Side(style="thin", color="D8E1EE")
    for side in ("left", "right", "top", "bottom")
})
EXAMPLE_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
SOFT_RED_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_month(month: Any) -> str:
    digits = "".join(char for char in str(month or "") if char.isdigit())[:6]
    if not MONTH_PATTERN.fullmatch(digits):
        raise ValueError("核算月份必须为 YYYYMM")
    try:
        datetime.strptime(digits, "%Y%m")
    except ValueError as exc:
        raise ValueError("核算月份无效") from exc
    return digits


def _date_text(value: Any, field_name: str, required: bool = False) -> str:
    text = _text(value)
    if not text:
        if required:
            raise ValueError(f"{field_name}不能为空")
        return ""
    try:
        return datetime.fromisoformat(text[:10]).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"{field_name}必须为 YYYY-MM-DD：{text}") from exc


def _regular_hours(value: Any) -> Optional[float]:
    text = _text(value)
    if not text:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"正班时数必须是数字：{value}") from exc
    if number < 0 or number > 24:
        raise ValueError(f"正班时数超出合理范围：{value}")
    return int(number) if number.is_integer() else number


def _default_break_category(period: Any) -> str:
    """为历史班次表补齐分类；保存后的配置始终显式携带分类。"""
    text = _text(period).replace("：", ":")
    if not text:
        return BREAK_CATEGORY_OTHER
    numbers = [int(number) for number in re.findall(r"\d+", text)]
    if len(numbers) < 4:
        return BREAK_CATEGORY_OTHER
    start_hour = numbers[-4]
    if "次日" in text:
        return BREAK_CATEGORY_MORNING if 5 <= start_hour < 8 else BREAK_CATEGORY_OTHER
    if 22 <= start_hour < 24 or 24 <= start_hour < 30 or 0 <= start_hour < 5:
        return BREAK_CATEGORY_EVENING
    if 30 <= start_hour < 32 or 5 <= start_hour < 8:
        return BREAK_CATEGORY_MORNING
    return BREAK_CATEGORY_OTHER


def normalize_break_segments(values: Iterable[Any]) -> List[Dict[str, str]]:
    """兼容历史字符串，并统一为带业务分类的休息段。"""
    segments = []
    for raw in values or []:
        if isinstance(raw, Mapping):
            period = _text(raw.get("period") or raw.get("time") or raw.get("时段"))
            category = _text(raw.get("category") or raw.get("type") or raw.get("休息类型"))
        else:
            period = _text(raw)
            category = ""
        if not period:
            continue
        category = category or _default_break_category(period)
        if category not in BREAK_CATEGORY_OPTIONS:
            raise ValueError(
                f"休息类型只能是：{' / '.join(BREAK_CATEGORY_OPTIONS)}；当前为：{category}"
            )
        segments.append({"period": period, "category": category})
    return segments


def _normalize_shift_row(raw: Mapping[str, Any]) -> Dict[str, Any]:
    shift_code = _text(raw.get("shift_code") or raw.get("班次编号"))
    if not shift_code:
        raise ValueError("班次编号不能为空")
    raw_segments = raw.get("break_segments")
    if raw_segments is None:
        periods = raw.get("break_periods")
        if periods is None:
            periods = [raw.get(f"break_{index}") for index in range(1, 4)]
        categories = raw.get("break_categories") or []
        raw_segments = [
            {
                "period": period,
                "category": categories[index] if index < len(categories) else "",
            }
            for index, period in enumerate(periods or [])
        ]
    segments = normalize_break_segments(raw_segments or [])
    return {
        "shift_category": _text(raw.get("shift_category") or raw.get("班次类别名称")),
        "shift_name": _text(raw.get("shift_name") or raw.get("班次名称")),
        "shift_code": shift_code,
        "shift_time": _text(raw.get("shift_time") or raw.get("班次时间点描述")),
        "regular_hours": _regular_hours(
            raw.get("regular_hours") if "regular_hours" in raw else raw.get("正班时数")
        ),
        # break_periods 保留给历史调用方；break_segments 是新的可审计配置源。
        "break_periods": [segment["period"] for segment in segments],
        "break_segments": segments,
        "note": _text(raw.get("note") or raw.get("备注")),
    }


def load_baseline_shift_breaks() -> List[Dict[str, Any]]:
    payload = json.loads(BASELINE_SHIFT_BREAKS_PATH.read_text(encoding="utf-8"))
    shifts = [_normalize_shift_row(row) for row in payload.get("shifts", [])]
    _raise_on_duplicates("重复班次", shifts, ("shift_code",))
    return deepcopy(shifts)


def baseline_shift_breaks_metadata() -> Dict[str, str]:
    payload = json.loads(BASELINE_SHIFT_BREAKS_PATH.read_text(encoding="utf-8"))
    return {
        "source": _text(payload.get("source")),
        "source_updated_at": _text(payload.get("source_updated_at")),
    }


def empty_night_shift_config(month: Any = "") -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "shift_break_overrides": [],
        "jinjiang_exclusions": [],
        "jinjiang_list_confirmed": False,
    }
    if month:
        payload.update({"month": normalize_month(month), "revision": 0, "exists": False})
    return payload


def _normalize_payload(payload: Optional[Mapping[str, Any]]) -> Dict[str, Any]:
    payload = payload or {}
    normalized = empty_night_shift_config()

    # 兼容本功能重构前产生的本地草稿；保存后统一落到 override 字段。
    raw_overrides = payload.get("shift_break_overrides")
    if raw_overrides is None:
        raw_overrides = payload.get("shift_breaks", [])
    normalized["shift_break_overrides"] = [
        _normalize_shift_row(raw) for raw in raw_overrides or []
    ]

    for raw in payload.get("jinjiang_exclusions", []) or []:
        employee_id = _text(raw.get("employee_id") or raw.get("工号"))
        if not employee_id:
            raise ValueError("晋江不享有名单工号不能为空")
        start_date = _date_text(
            raw.get("start_date") or raw.get("生效日期"), "生效日期", required=True
        )
        end_date = _date_text(raw.get("end_date") or raw.get("失效日期"), "失效日期")
        if end_date and end_date < start_date:
            raise ValueError(f"{employee_id} 的失效日期不能早于生效日期")
        normalized["jinjiang_exclusions"].append({
            "employee_id": employee_id,
            "employee_name": _text(raw.get("employee_name") or raw.get("姓名")),
            "reason": _text(raw.get("reason") or raw.get("排除原因")) or "特殊名单",
            "start_date": start_date,
            "end_date": end_date,
            "note": _text(raw.get("note") or raw.get("备注")),
        })

    normalized["jinjiang_list_confirmed"] = bool(payload.get("jinjiang_list_confirmed", False))
    _validate_duplicates(normalized)
    return normalized


def _duplicate_values(rows: Iterable[Mapping[str, Any]], fields: Iterable[str]) -> List[str]:
    seen = set()
    duplicates = []
    for row in rows:
        key = tuple(_text(row.get(field)).casefold() for field in fields)
        if key in seen:
            duplicates.append("/".join(key))
        seen.add(key)
    return duplicates


def _raise_on_duplicates(
    label: str,
    rows: Iterable[Mapping[str, Any]],
    fields: Iterable[str],
) -> None:
    duplicates = _duplicate_values(rows, fields)
    if duplicates:
        raise ValueError(f"{label}：{'、'.join(duplicates[:5])}")


def _validate_duplicates(payload: Mapping[str, Any]) -> None:
    _raise_on_duplicates(
        "重复班次", payload.get("shift_break_overrides", []), ("shift_code",)
    )
    _raise_on_duplicates(
        "重复晋江不享有名单",
        payload.get("jinjiang_exclusions", []),
        ("employee_id", "start_date"),
    )


def merge_shift_breaks(overrides: Iterable[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    baseline = load_baseline_shift_breaks()
    by_code = {row["shift_code"]: row for row in baseline}
    order = [row["shift_code"] for row in baseline]
    for raw in overrides or []:
        row = _normalize_shift_row(raw)
        if row["shift_code"] not in by_code:
            order.append(row["shift_code"])
        by_code[row["shift_code"]] = row
    return [deepcopy(by_code[code]) for code in order]


def _expand_payload(payload: Mapping[str, Any], exists: bool) -> Dict[str, Any]:
    normalized = _normalize_payload(payload)
    baseline = load_baseline_shift_breaks()
    effective = merge_shift_breaks(normalized["shift_break_overrides"])
    expanded = {
        key: deepcopy(value)
        for key, value in payload.items()
        if key not in {
            "shift_breaks", "baseline_shift_breaks", "effective_shift_breaks",
            "regional_positions", "continuous_shift_overrides", "counts", "exists",
        }
    }
    expanded.update(normalized)
    expanded.update({
        "exists": exists,
        "baseline": baseline_shift_breaks_metadata(),
        "baseline_shift_breaks": baseline,
        "effective_shift_breaks": effective,
        # 引擎快照使用这一稳定别名，内容是“平台基线 + 当月调整”的最终班次表。
        "shift_breaks": deepcopy(effective),
    })
    expanded["counts"] = config_counts(expanded)
    return expanded


def _config_path(month: Any) -> Path:
    return NIGHT_SHIFT_CONFIG_DIR / f"{normalize_month(month)}.json"


def _history_path(month: Any, revision: int) -> Path:
    return NIGHT_SHIFT_CONFIG_DIR / "history" / normalize_month(month) / f"r{int(revision):04d}.json"


def _atomic_write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, delete=False) as tmp:
        json.dump(payload, tmp, ensure_ascii=False, indent=2)
        tmp.write("\n")
        tmp_path = Path(tmp.name)
    os.replace(tmp_path, path)


def load_night_shift_config(month: Any, required: bool = True) -> Dict[str, Any]:
    month_text = normalize_month(month)
    path = _config_path(month_text)
    if not path.exists():
        if required:
            raise FileNotFoundError(f"{month_text} 夜班补贴配置不存在")
        return _expand_payload(empty_night_shift_config(month_text), exists=False)
    payload = json.loads(path.read_text(encoding="utf-8"))
    return _expand_payload(payload, exists=True)


def save_night_shift_config(
    month: Any,
    payload: Mapping[str, Any],
    updated_by: str = "",
    copied_from: str = "",
) -> Dict[str, Any]:
    month_text = normalize_month(month)
    normalized = _normalize_payload(payload)
    current = load_night_shift_config(month_text, required=False)
    saved = {
        "month": month_text,
        "revision": int(current.get("revision") or 0) + 1,
        "updated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "updated_by": _text(updated_by),
        "copied_from": normalize_month(copied_from) if copied_from else _text(current.get("copied_from")),
        **normalized,
    }
    _atomic_write_json(_history_path(month_text, saved["revision"]), saved)
    _atomic_write_json(_config_path(month_text), saved)
    return _expand_payload(saved, exists=True)


def load_night_shift_config_revision(month: Any, revision: int) -> Dict[str, Any]:
    path = _history_path(month, revision)
    if not path.exists():
        raise FileNotFoundError(f"{normalize_month(month)} 夜班补贴配置版本 {revision} 不存在")
    payload = json.loads(path.read_text(encoding="utf-8"))
    return _expand_payload(payload, exists=True)


def list_night_shift_config_revisions(month: Any) -> List[Dict[str, Any]]:
    history_dir = _history_path(month, 1).parent
    revisions = []
    for path in sorted(history_dir.glob("r*.json"), reverse=True):
        payload = json.loads(path.read_text(encoding="utf-8"))
        revisions.append({
            "month": payload.get("month"),
            "revision": payload.get("revision"),
            "updated_at": payload.get("updated_at"),
            "updated_by": payload.get("updated_by", ""),
            "copied_from": payload.get("copied_from", ""),
            "counts": config_counts(payload),
        })
    return revisions


def copy_night_shift_config(source_month: Any, target_month: Any, updated_by: str = "") -> Dict[str, Any]:
    source = load_night_shift_config(source_month)
    target_path = _config_path(target_month)
    target = load_night_shift_config(target_month, required=False)
    if target_path.exists() and target.get("jinjiang_list_confirmed"):
        raise FileExistsError(f"{normalize_month(target_month)} 夜班补贴配置已存在，不能覆盖")
    return save_night_shift_config(
        target_month,
        {
            "shift_break_overrides": target.get("shift_break_overrides", []),
            "jinjiang_exclusions": source["jinjiang_exclusions"],
            "jinjiang_list_confirmed": source["jinjiang_list_confirmed"],
        },
        updated_by=updated_by,
        copied_from=source["month"],
    )


def _format_sheet(ws, sheet_name: str) -> None:
    column_widths = {
        "employee_id": 16,
        "employee_name": 12,
        "reason": 24,
        "start_date": 14,
        "end_date": 14,
        "note": 36,
    }
    for index, (header, key, required, description) in enumerate(CONFIG_SHEETS[sheet_name], 1):
        cell = ws.cell(1, index, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        cell.comment = Comment(f"{'必填' if required else '选填'}：{description}", "Sigma Workbench")
        ws.column_dimensions[cell.column_letter].width = column_widths[key]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(CONFIG_SHEETS[sheet_name])).column_letter}1"


def _add_input_validation(ws) -> None:
    reason_list = ",".join(JINJIANG_REASON_OPTIONS)
    reason_validation = DataValidation(
        type="list",
        formula1=f'"{reason_list}"',
        allow_blank=False,
    )
    reason_validation.error = "请从下拉项选择：轻松岗位 / 其他线下确认不享有"
    reason_validation.errorTitle = "排除原因不正确"
    reason_validation.prompt = "计件岗和门禁无需填写，系统会自动排除。"
    reason_validation.promptTitle = "请选择排除原因"
    reason_validation.showErrorMessage = True
    reason_validation.showInputMessage = True
    ws.add_data_validation(reason_validation)
    reason_validation.add("C2:C1000")

    for column in ("D", "E"):
        date_validation = DataValidation(
            type="date",
            operator="between",
            formula1="DATE(2020,1,1)",
            formula2="DATE(2100,12,31)",
            allow_blank=column == "E",
        )
        date_validation.error = "请填写有效日期，例如 2026-08-01"
        date_validation.errorTitle = "日期格式不正确"
        date_validation.showErrorMessage = True
        ws.add_data_validation(date_validation)
        date_validation.add(f"{column}2:{column}1000")
    invalid_end_fill = FormulaRule(formula=["AND($D2<>\"\",$E2<>\"\",$E2<$D2)"], fill=SOFT_RED_FILL)
    ws.conditional_formatting.add("D2:E1000", invalid_end_fill)


def _write_exclusion_rows(ws, rows: Iterable[Mapping[str, Any]], sheet_name: str) -> None:
    for row_index, row in enumerate(rows, 2):
        for column_index, (_header, key, _required, _description) in enumerate(
            CONFIG_SHEETS[sheet_name], 1
        ):
            value = row.get(key, "")
            if key in {"start_date", "end_date"} and value:
                value = datetime.fromisoformat(_text(value)[:10]).date()
            cell = ws.cell(row_index, column_index, value)
            cell.border = THIN_BORDER
            if key in {"start_date", "end_date"}:
                cell.number_format = "yyyy-mm-dd"


def _add_inline_examples(ws) -> None:
    examples = [
        {
            "employee_id": "示例0001", "employee_name": "张三", "reason": "轻松岗位",
            "start_date": "2026-08-01", "end_date": "", "note": "线下确认的轻松岗位",
        },
        {
            "employee_id": "示例0002", "employee_name": "李四", "reason": "其他线下确认不享有",
            "start_date": "2026-08-15", "end_date": "2026-08-31", "note": "临时调整，依据：业务确认记录",
        },
    ]
    _write_exclusion_rows(ws, examples, JINJIANG_INPUT_SHEET)
    for row_index in (2, 3):
        for cell in ws[row_index]:
            cell.fill = EXAMPLE_FILL
            cell.font = Font(name="微软雅黑", size=10, italic=True, color="64748B")
        ws.cell(row_index, 1).comment = Comment(
            "示例行：可直接覆盖填写真实员工；如保留不动，平台导入时会自动忽略。",
            "Sigma Workbench",
        )


def generate_night_shift_config_workbook(payload: Optional[Mapping[str, Any]] = None) -> bytes:
    normalized = _normalize_payload(payload)
    wb = Workbook()
    ws = wb.active
    ws.title = JINJIANG_INPUT_SHEET
    _format_sheet(ws, JINJIANG_INPUT_SHEET)
    _add_input_validation(ws)
    if normalized["jinjiang_exclusions"]:
        _write_exclusion_rows(ws, normalized["jinjiang_exclusions"], JINJIANG_INPUT_SHEET)
    else:
        _add_inline_examples(ws)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_night_shift_config_workbook(data: bytes) -> Dict[str, Any]:
    wb = load_workbook(BytesIO(data), data_only=True)
    sheet_name = next(
        (name for name in (JINJIANG_INPUT_SHEET, JINJIANG_LEGACY_SHEET) if name in wb.sheetnames),
        "",
    )
    if not sheet_name:
        raise ValueError(f"维护文件缺少工作表：{JINJIANG_INPUT_SHEET}")
    ws = wb[sheet_name]
    headers = [_text(cell.value) for cell in ws[1]]
    schema = CONFIG_SHEETS[JINJIANG_INPUT_SHEET]
    expected = [column[0] for column in schema]
    missing = [header for header in expected if header not in headers]
    if missing:
        raise ValueError(f"{sheet_name} 缺少字段：{'、'.join(missing)}")
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row_by_header = {
            headers[index]: value for index, value in enumerate(values) if index < len(headers)
        }
        if not any(_text(value) for value in row_by_header.values()):
            continue
        if _text(row_by_header.get("工号")).startswith("示例"):
            continue
        rows.append({
            key: row_by_header.get(header)
            for header, key, _required, _description in schema
        })
    wb.close()
    normalized = _normalize_payload({
        "shift_break_overrides": [],
        "jinjiang_exclusions": rows,
        "jinjiang_list_confirmed": True,
    })
    for row_index, row in enumerate(normalized["jinjiang_exclusions"], 2):
        if not row["employee_name"]:
            raise ValueError(f"{sheet_name} 第 {row_index} 行姓名不能为空")
        if row["reason"] not in JINJIANG_REASON_OPTIONS:
            raise ValueError(
                f"{sheet_name} 第 {row_index} 行排除原因只能选择：{' / '.join(JINJIANG_REASON_OPTIONS)}；"
                "计件岗和门禁无需填写"
            )
        if row["reason"] == "其他线下确认不享有" and not row["note"]:
            raise ValueError(f"{sheet_name} 第 {row_index} 行选择其他原因时，备注必须填写线下依据")
    return normalized


def find_active_jinjiang_exclusion(
    entries: Iterable[Mapping[str, Any]],
    employee_id: Any,
    attendance_date: date,
) -> Optional[Dict[str, Any]]:
    employee_id_text = _text(employee_id).casefold()
    for entry in entries or []:
        if _text(entry.get("employee_id")).casefold() != employee_id_text:
            continue
        start = datetime.fromisoformat(_text(entry.get("start_date"))[:10]).date()
        end_text = _text(entry.get("end_date"))
        end = datetime.fromisoformat(end_text[:10]).date() if end_text else None
        if start <= attendance_date and (end is None or attendance_date <= end):
            return dict(entry)
    return None


def config_counts(payload: Mapping[str, Any]) -> Dict[str, int]:
    overrides = payload.get("shift_break_overrides", []) or []
    effective = payload.get("effective_shift_breaks")
    if effective is None:
        effective = merge_shift_breaks(overrides)
    return {
        "baseline_shift_count": len(load_baseline_shift_breaks()),
        "shift_break_override_count": len(overrides),
        "effective_shift_count": len(effective),
        "jinjiang_exclusion_count": len(payload.get("jinjiang_exclusions", []) or []),
    }


def build_night_shift_config_snapshot(payload: Mapping[str, Any]) -> Dict[str, Any]:
    """冻结核算所需字段，避免在任务元数据中重复保存三份班次基线。"""
    effective = payload.get("effective_shift_breaks") or merge_shift_breaks(
        payload.get("shift_break_overrides", [])
    )
    keys = (
        "month", "revision", "exists", "updated_at", "updated_by", "copied_from",
        "baseline", "shift_break_overrides", "jinjiang_exclusions", "jinjiang_list_confirmed",
    )
    snapshot = {key: deepcopy(payload.get(key)) for key in keys if key in payload}
    snapshot["shift_breaks"] = deepcopy(effective)
    snapshot["counts"] = config_counts(payload)
    return snapshot
