from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel


STANDARD_IMPORT_HEADERS = [
    "核算月份", "姓名", "工号", "证件号", "入职日期", "人员状态", "工作地", "最后工作日", "职级",
    "二级组织", "三级组织", "四级组织", "五级组织", "职位", "职族", "职类", "直线经理", "直线经理工号",
    "标签分类", "职族分类", "ABC类别", "招聘负责人工号", "招聘负责人姓名", "招聘负责人人员状态",
    "招聘负责人最后工作日", "协助招聘人工号", "协助招聘人姓名", "协助招聘人人员状态", "协助招聘人最后工作日",
    "招聘渠道", "海外offer流程编号", "海外招聘需求编号", "招聘启动日期", "候选人入职时间", "转正日期",
    "周期剔除天数", "奖金地区类型", "推荐人姓名", "推荐人工号", "推荐人职级", "推荐人人员状态",
    "推荐人最后工作日", "推荐人职位", "直接上级工号", "间接上级1工号", "间接上级2工号",
    "间接上级3工号", "间接上级4工号", "特殊地区规则", "人工备注",
]

HISTORY_DETAIL_SHEET = "附件一-招聘奖金核算明细"
STANDARD_IMPORT_SHEET = "导入_月度数据"
DATE_FORMAT = "yyyy/m/d"

INVALID_RESIGN_STATUSES = {"已废弃", "已驳回", "已撤回", "作废", "审批未通过"}
VALID_DOMESTIC_OFFER_STATUSES = {"审批通过", "审批中", "草稿"}
VALID_OVERSEAS_OFFER_STATUSES = {"已入职", "已接受offer", "审批中", "已发offer", "待发offer"}
DEVELOPED_LOCATION_TOKENS = {
    "德国", "捷克", "美国", "英国", "法国", "加拿大", "日本", "韩国", "荷兰",
    "意大利", "西班牙", "葡萄牙", "奥地利", "瑞士", "比利时", "卢森堡",
    "丹麦", "瑞典", "挪威", "芬兰", "爱尔兰", "澳大利亚", "新西兰",
    "新加坡", "以色列", "波兰",
    "germany", "czech", "united states", "usa", "u.s.", "america", "uk",
    "united kingdom", "england", "france", "canada", "japan", "korea",
    "netherlands", "holland", "italy", "spain", "portugal", "austria",
    "switzerland", "belgium", "luxembourg", "denmark", "sweden", "norway",
    "finland", "ireland", "australia", "new zealand", "singapore", "israel",
    "poland",
}
OVERSEAS_CHINA_LOCATION_TOKENS = {"非中国", "香港", "澳门", "澳門", "台湾", "臺灣"}
MAINLAND_LOCATION_TOKENS = {
    "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁", "吉林", "黑龙江",
    "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南",
    "广东", "海南", "四川", "贵州", "云南", "陕西", "甘肃", "青海", "内蒙古",
    "广西", "西藏", "宁夏", "新疆", "中国大陆", "大陆",
}
ROLE_STATUS_FIELDS = {
    "招聘负责人人员状态": "招聘负责人工号",
    "招聘负责人最后工作日": "招聘负责人工号",
    "协助招聘人人员状态": "协助招聘人工号",
    "协助招聘人最后工作日": "协助招聘人工号",
    "推荐人人员状态": "推荐人工号",
    "推荐人最后工作日": "推荐人工号",
    "推荐人职级": "推荐人工号",
    "推荐人职位": "推荐人工号",
}
EMPLOYEE_NO_FIELDS = {
    "工号", "招聘负责人工号", "协助招聘人工号", "推荐人工号", "直线经理工号",
    "直接上级工号", "间接上级1工号", "间接上级2工号", "间接上级3工号", "间接上级4工号",
}
EMPTY_EQUIVALENT_FIELDS = {
    "推荐人姓名", "推荐人工号", "推荐人职级", "推荐人人员状态", "推荐人最后工作日", "推荐人职位",
    "协助招聘人工号", "协助招聘人姓名", "协助招聘人人员状态", "协助招聘人最后工作日",
}
AUTO_ENRICHMENT_FIELDS = {
    "证件号", "海外offer流程编号", "海外招聘需求编号",
    "直线经理", "直线经理工号", "直接上级工号",
    "间接上级1工号", "间接上级2工号", "间接上级3工号", "间接上级4工号",
}


@dataclass
class CompareResult:
    summary: dict[str, Any]
    field_rates: dict[str, dict[str, Any]]
    field_differences: list[dict[str, Any]]
    missing_in_generated: list[dict[str, Any]]
    missing_in_target: list[dict[str, Any]]
    duplicate_keys: list[dict[str, Any]]
    alias_resolved_missing: list[dict[str, Any]]
    auto_enrichments: list[dict[str, Any]]


def build_auto_import_rows(
    *,
    target_month: int,
    previous_workbook: Path | str,
    ehr_roster: Path | str,
    oehr_roster: Path | str,
    domestic_offer: Path | str,
    overseas_offer: Path | str,
    ehr_resignations: Path | str | None = None,
    oehr_resignations: Path | str | None = None,
    special_approvals: Path | str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    history_rows = extract_history_rows(Path(previous_workbook), target_month)
    ehr_people = _read_ehr_people(Path(ehr_roster))
    oehr_people = _read_oehr_people(Path(oehr_roster))
    domestic_offers = _read_domestic_offers(Path(domestic_offer))
    overseas_offers = _read_overseas_offers(Path(overseas_offer))
    resignations = _read_resignations(ehr_resignations, oehr_resignations)
    cycle_exclusions = _read_cycle_exclusions(special_approvals)
    people_by_no = {**ehr_people, **oehr_people}

    diagnostics: dict[str, Any] = {
        "counts": {
            "historyRows": len(history_rows),
            "ehrPeople": len(ehr_people),
            "oehrPeople": len(oehr_people),
            "domesticOffers": len(domestic_offers),
            "overseasOffers": len(overseas_offers),
            "resignationRows": len(resignations),
            "cycleExclusionRows": len(cycle_exclusions),
            "newDomesticRows": 0,
            "newOverseasRows": 0,
        },
        "issues": [],
        "changes": [],
        "newRows": [],
    }

    rows_by_key: dict[str, dict[str, Any]] = {}
    for row in history_rows:
        key = _row_key(row)
        if not key:
            diagnostics["issues"].append({"类型": "历史台账缺少主键", "姓名": row.get("姓名", ""), "工号": row.get("工号", "")})
            continue
        rows_by_key.setdefault(key, row)

    for row in rows_by_key.values():
        _refresh_dynamic_fields(row, people_by_no, resignations, diagnostics)
        _apply_cycle_exclusion(row, cycle_exclusions, diagnostics)

    for person in ehr_people.values():
        if not _is_target_month(person.get("入职日期"), target_month):
            continue
        employee_no = _text(person.get("工号"))
        if not employee_no.lower().startswith("zt"):
            continue
        if _row_key(person) in rows_by_key:
            continue
        offer = _lookup_domestic_offer(person, domestic_offers)
        row = _new_domestic_row(target_month, person, offer, diagnostics)
        _refresh_dynamic_fields(row, people_by_no, resignations, diagnostics)
        _apply_cycle_exclusion(row, cycle_exclusions, diagnostics)
        rows_by_key[_row_key(row)] = row
        diagnostics["counts"]["newDomesticRows"] += 1
        diagnostics["newRows"].append(_new_row_record(row, "EHR花名册"))

    for person in oehr_people.values():
        if not _is_target_month(person.get("入职日期"), target_month):
            continue
        if not _text(person.get("海外offer流程编号")):
            continue
        if _row_key(person) in rows_by_key:
            continue
        offer = overseas_offers.get(_norm_key(person.get("海外offer流程编号")), {})
        row = _new_overseas_row(target_month, person, offer, diagnostics)
        _refresh_dynamic_fields(row, people_by_no, resignations, diagnostics)
        _apply_cycle_exclusion(row, cycle_exclusions, diagnostics)
        rows_by_key[_row_key(row)] = row
        diagnostics["counts"]["newOverseasRows"] += 1
        diagnostics["newRows"].append(_new_row_record(row, "OEHR花名册"))

    ordered_rows = sorted(rows_by_key.values(), key=_sort_key)
    diagnostics["counts"]["generatedRows"] = len(ordered_rows)
    return ordered_rows, diagnostics


def extract_history_rows(path: Path, target_month: int) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    if HISTORY_DETAIL_SHEET not in workbook.sheetnames:
        raise ValueError(f"历史核算表缺少工作表：{HISTORY_DETAIL_SHEET}")
    sheet = workbook[HISTORY_DETAIL_SHEET]

    mapping = {
        "姓名": 2,
        "工号": 3,
        "证件号": 4,
        "入职日期": 5,
        "人员状态": 6,
        "工作地": 7,
        "最后工作日": 8,
        "职级": 9,
        "二级组织": 10,
        "三级组织": 11,
        "四级组织": 12,
        "五级组织": 13,
        "职位": 14,
        "职族": 15,
        "职类": 16,
        "直线经理": 17,
        "直线经理工号": 18,
        "标签分类": 19,
        "职族分类": 20,
        "ABC类别": 21,
        "招聘负责人工号": 22,
        "招聘负责人姓名": 23,
        "协助招聘人工号": 24,
        "协助招聘人姓名": 25,
        "招聘渠道": 26,
        "海外offer流程编号": 27,
        "海外招聘需求编号": 28,
        "招聘启动日期": 29,
        "候选人入职时间": 30,
        "转正日期": 31,
        "周期剔除天数": 33,
        "奖金地区类型": 57,
        "推荐人姓名": 58,
        "推荐人工号": 59,
    }

    rows: list[dict[str, Any]] = []
    for row_number in range(4, sheet.max_row + 1):
        name = sheet.cell(row_number, 2).value
        employee_no = sheet.cell(row_number, 3).value
        id_no = sheet.cell(row_number, 4).value
        if not any(_text(value) for value in (name, employee_no, id_no)):
            continue
        if _text(name) in {"姓名", "合计"} or _text(employee_no) == "工号":
            continue
        row = _blank_standard_row(target_month)
        for field, column in mapping.items():
            if column <= sheet.max_column:
                row[field] = _clean_value(sheet.cell(row_number, column).value)
        row["核算月份"] = target_month
        row["招聘渠道"] = _normalize_channel(row.get("招聘渠道"))
        rows.append(row)
    return rows


def read_standard_import_rows(path: Path | str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    if STANDARD_IMPORT_SHEET not in workbook.sheetnames:
        raise ValueError(f"标准导入模板缺少工作表：{STANDARD_IMPORT_SHEET}")
    sheet = workbook[STANDARD_IMPORT_SHEET]
    headers = [_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for row_number in range(2, sheet.max_row + 1):
        row = {header: _clean_value(sheet.cell(row_number, column).value) for column, header in enumerate(headers, start=1) if header}
        if any(_text(row.get(field)) for field in ("工号", "姓名", "证件号")):
            rows.append({field: row.get(field, "") for field in STANDARD_IMPORT_HEADERS})
    return rows


def compare_import_rows(
    generated_rows: list[dict[str, Any]],
    target_rows: list[dict[str, Any]],
    *,
    headers: list[str] | None = None,
) -> CompareResult:
    headers = headers or STANDARD_IMPORT_HEADERS
    generated_index, generated_dupes = _index_rows(generated_rows)
    target_index, target_dupes = _index_rows(target_rows)
    common_keys = sorted(set(generated_index) & set(target_index))
    generated_aliases = _alias_index(generated_rows)
    missing_in_generated: list[dict[str, Any]] = []
    alias_resolved_missing: list[dict[str, Any]] = []
    for key in sorted(set(target_index) - set(generated_index)):
        target_row = target_index[key]
        alias = _loose_row_alias(target_row)
        if not _norm_key(target_row.get("工号")) and alias and alias in generated_aliases:
            resolved = dict(target_row)
            resolved["自动匹配工号"] = generated_aliases[alias].get("工号", "")
            resolved["匹配说明"] = "人工行缺工号，但自动生成表存在同名同职位同月入职记录"
            alias_resolved_missing.append(resolved)
            continue
        missing_in_generated.append(target_row)
    missing_in_target = [generated_index[key] for key in sorted(set(generated_index) - set(target_index))]

    field_rates: dict[str, dict[str, Any]] = {
        field: {"field": field, "matchCount": 0, "diffCount": 0, "total": 0, "matchRate": 0.0}
        for field in headers
    }
    differences: list[dict[str, Any]] = []
    auto_enrichments: list[dict[str, Any]] = []

    for key in common_keys:
        generated = generated_index[key]
        target = target_index[key]
        for field in headers:
            if field == "工号":
                continue
            field_rates[field]["total"] += 1
            if _is_auto_enrichment(field, generated.get(field), target.get(field)):
                field_rates[field]["matchCount"] += 1
                auto_enrichments.append({
                    "匹配键": key,
                    "工号": target.get("工号") or generated.get("工号", ""),
                    "姓名": target.get("姓名") or generated.get("姓名", ""),
                    "字段": field,
                    "自动补充值": generated.get(field, ""),
                    "人工模板值": target.get(field, ""),
                    "说明": "源文件可自动补全，人工模板为空",
                })
            elif _equal_cell_for_field(field, generated.get(field), target.get(field), generated, target):
                field_rates[field]["matchCount"] += 1
            else:
                field_rates[field]["diffCount"] += 1
                differences.append({
                    "匹配键": key,
                    "工号": target.get("工号") or generated.get("工号", ""),
                    "姓名": target.get("姓名") or generated.get("姓名", ""),
                    "字段": field,
                    "自动生成值": generated.get(field, ""),
                    "人工模板值": target.get(field, ""),
                })
    for rate in field_rates.values():
        if rate["total"]:
            rate["matchRate"] = round(rate["matchCount"] / rate["total"], 4)

    duplicate_keys = [
        {"来源": "自动生成", "匹配键": key, "重复行数": count}
        for key, count in generated_dupes.items()
    ] + [
        {"来源": "人工模板", "匹配键": key, "重复行数": count}
        for key, count in target_dupes.items()
    ]

    summary = {
        "generatedRows": len(generated_rows),
        "targetRows": len(target_rows),
        "commonRows": len(common_keys),
        "missingInGenerated": len(missing_in_generated),
        "missingInTarget": len(missing_in_target),
        "manualBlankIdMatchedByAlias": len(alias_resolved_missing),
        "autoEnrichmentCount": len(auto_enrichments),
        "fieldDiffCount": len(differences),
        "duplicateKeyCount": len(duplicate_keys),
    }
    return CompareResult(summary, field_rates, differences, missing_in_generated, missing_in_target, duplicate_keys, alias_resolved_missing, auto_enrichments)


def write_standard_import_workbook(rows: list[dict[str, Any]], output_path: Path | str) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    intro = workbook.active
    intro.title = "使用说明"
    intro.append(["自动生成招聘奖金导入数据"])
    intro.append(["说明", "该文件由平台读取上月台账和本月源文件生成，用于字段整合验证。"])
    sheet = workbook.create_sheet(STANDARD_IMPORT_SHEET)
    sheet.append(STANDARD_IMPORT_HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in STANDARD_IMPORT_HEADERS])
    _format_standard_sheet(sheet)
    workbook.save(output_path)
    return output_path


def write_validation_report(
    compare: CompareResult,
    diagnostics: dict[str, Any],
    output_path: Path | str,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "概览"
    summary.append(["指标", "数值"])
    for key, value in compare.summary.items():
        summary.append([key, value])
    for key, value in diagnostics.get("counts", {}).items():
        summary.append([key, value])

    _write_dict_sheet(workbook, "字段命中率", list(compare.field_rates.values()), ["field", "total", "matchCount", "diffCount", "matchRate"])
    _write_dict_sheet(workbook, "字段差异", compare.field_differences, ["匹配键", "工号", "姓名", "字段", "自动生成值", "人工模板值"])
    _write_dict_sheet(workbook, "自动补全人工空值", compare.auto_enrichments, ["匹配键", "工号", "姓名", "字段", "自动补充值", "人工模板值", "说明"])
    _write_dict_sheet(workbook, "人工有自动缺", compare.missing_in_generated, STANDARD_IMPORT_HEADERS)
    _write_dict_sheet(workbook, "人工缺工号疑似重复", compare.alias_resolved_missing, STANDARD_IMPORT_HEADERS + ["自动匹配工号", "匹配说明"])
    _write_dict_sheet(workbook, "自动有人工缺", compare.missing_in_target, STANDARD_IMPORT_HEADERS)
    _write_dict_sheet(workbook, "待确认事项", _pending_confirmation_items(compare, diagnostics), ["类别", "姓名", "工号", "字段", "当前结论", "建议动作"])
    _write_dict_sheet(workbook, "数据源异常", diagnostics.get("issues", []), None)
    _write_dict_sheet(workbook, "字段刷新留痕", diagnostics.get("changes", []), ["工号", "姓名", "字段", "原值", "新值", "来源"])
    _write_dict_sheet(workbook, "新增人员清单", diagnostics.get("newRows", []), ["工号", "姓名", "标签分类", "来源"])
    _write_dict_sheet(workbook, "重复主键", compare.duplicate_keys, ["来源", "匹配键", "重复行数"])

    for sheet in workbook.worksheets:
        _format_report_sheet(sheet)
    workbook.save(output_path)
    return output_path


def write_exception_workbook(diagnostics: dict[str, Any], output_path: Path | str) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "异常清单"
    _write_rows_to_sheet(sheet, diagnostics.get("issues", []), None)
    changes = workbook.create_sheet("字段刷新留痕")
    _write_rows_to_sheet(changes, diagnostics.get("changes", []), ["工号", "姓名", "字段", "原值", "新值", "来源"])
    for ws in workbook.worksheets:
        _format_report_sheet(ws)
    workbook.save(output_path)
    return output_path


def run_validation(
    *,
    target_month: int,
    previous_workbook: Path | str,
    ehr_roster: Path | str,
    oehr_roster: Path | str,
    domestic_offer: Path | str,
    overseas_offer: Path | str,
    target_template: Path | str,
    output_dir: Path | str,
    ehr_resignations: Path | str | None = None,
    oehr_resignations: Path | str | None = None,
    special_approvals: Path | str | None = None,
) -> dict[str, Any]:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    generated_rows, diagnostics = build_auto_import_rows(
        target_month=target_month,
        previous_workbook=previous_workbook,
        ehr_roster=ehr_roster,
        oehr_roster=oehr_roster,
        domestic_offer=domestic_offer,
        overseas_offer=overseas_offer,
        ehr_resignations=ehr_resignations,
        oehr_resignations=oehr_resignations,
        special_approvals=special_approvals,
    )
    target_rows = read_standard_import_rows(target_template)
    compare = compare_import_rows(generated_rows, target_rows, headers=STANDARD_IMPORT_HEADERS)
    auto_path = write_standard_import_workbook(generated_rows, output_dir / f"自动生成_{target_month}_导入模板.xlsx")
    report_path = write_validation_report(compare, diagnostics, output_dir / f"自动_vs_人工_{target_month}_字段差异报告.xlsx")
    exception_path = write_exception_workbook(diagnostics, output_dir / f"数据源匹配异常清单_{target_month}.xlsx")
    return {
        "autoImportPath": str(auto_path),
        "diffReportPath": str(report_path),
        "exceptionPath": str(exception_path),
        "summary": compare.summary,
        "diagnostics": diagnostics,
    }


def _blank_standard_row(target_month: int) -> dict[str, Any]:
    row = {field: "" for field in STANDARD_IMPORT_HEADERS}
    row["核算月份"] = target_month
    row["周期剔除天数"] = 0
    return row


def _read_ehr_people(path: Path) -> dict[str, dict[str, Any]]:
    people: dict[str, dict[str, Any]] = {}
    for record in _read_records(path):
        employee_no = _text(record.get("JobNumber"))
        if not employee_no:
            continue
        people[_norm_key(employee_no)] = {
            "核算月份": "",
            "姓名": record.get("parent_Name", ""),
            "工号": employee_no,
            "证件号": record.get("parent_IDNumber", ""),
            "入职日期": _clean_value(record.get("EntryDate")),
            "人员状态": record.get("EmployeeStatus", ""),
            "工作地": "中国大陆",
            "最后工作日": _clean_value(record.get("LastWorkDate")),
            "职级": record.get("OIdJobLevel", ""),
            "二级组织": record.get("LookupPrefix_OIdDepartment_SecondLevelOrganization", ""),
            "三级组织": record.get("LookupPrefix_OIdDepartment_ThirdLevelOrganization", ""),
            "四级组织": record.get("LookupPrefix_OIdDepartment_FourthLevelOrganization", ""),
            "五级组织": record.get("LookupPrefix_OIdDepartment_FifthLevelOrganization", ""),
            "职位": record.get("OIdJobPosition_Name", ""),
            "直线经理": record.get("POIdEmpAdmin-ExportName", ""),
            "直线经理工号": record.get("POIdEmpAdmin-StaffCode", ""),
            "标签分类": "国内",
            "候选人入职时间": _clean_value(record.get("EntryDate")),
            "转正日期": _clean_value(record.get("RegularizationDate")),
            "直接上级工号": record.get("POIdEmpAdmin-StaffCode", ""),
            "推荐人职级": record.get("OIdJobLevel", ""),
            "推荐人职位": record.get("OIdJobPosition_Name", ""),
        }
    return people


def _read_oehr_people(path: Path) -> dict[str, dict[str, Any]]:
    people: dict[str, dict[str, Any]] = {}
    for record in _read_records(path):
        employee_no = _text(record.get("工号"))
        if not employee_no:
            continue
        location = record.get("办公地址") or record.get("工作地区") or ""
        people[_norm_key(employee_no)] = {
            "姓名": record.get("姓名", ""),
            "工号": employee_no,
            "证件号": record.get("证件号码", ""),
            "入职日期": _clean_value(record.get("入职日期")),
            "人员状态": record.get("人员状态", ""),
            "工作地": location,
            "最后工作日": _clean_value(record.get("最后工作日")),
            "职级": record.get("职级", ""),
            "二级组织": record.get("二级部门", ""),
            "三级组织": record.get("三级部门", ""),
            "四级组织": record.get("四级部门", ""),
            "五级组织": record.get("五级部门", ""),
            "职位": record.get("职位", ""),
            "职族": record.get("职族", ""),
            "职类": record.get("职类", ""),
            "直线经理": record.get("直接上级", ""),
            "直线经理工号": record.get("直接上级工号", ""),
            "标签分类": "海外",
            "海外offer流程编号": record.get("offer流程编号", ""),
            "候选人入职时间": _clean_value(record.get("入职日期")),
            "转正日期": "",
            "直接上级工号": record.get("直接上级工号", ""),
            "推荐人职级": record.get("职级", ""),
            "推荐人职位": record.get("职位", ""),
        }
    return people


def _read_domestic_offers(path: Path) -> dict[str, dict[str, Any]]:
    offers: dict[str, dict[str, Any]] = {}
    for record in _read_records(path):
        id_no = _norm_key(record.get("IDNumber"))
        if not id_no:
            continue
        status = _text(record.get("ApprovalStatus"))
        if status and status not in VALID_DOMESTIC_OFFER_STATUSES:
            continue
        offer = {
            "ABC类别": record.get("LookupPrefix_RecruitmentRequirement_extzhiweifenlei_109025_933701443", ""),
            "招聘负责人工号": record.get("HrDutyUserId-StaffCode", ""),
            "招聘负责人姓名": record.get("HrDutyUserId-ExportName", ""),
            "协助招聘人工号": record.get("extxiezhuzhaopinfuzerengonghao_109025_801104894", ""),
            "协助招聘人姓名": record.get("extxiezhuzhaopinfuzeren_109025_1163526586", ""),
            "招聘渠道": _normalize_channel(record.get("extQ20210635_109025_2041445622")),
            "招聘启动日期": _clean_value(record.get("LookupPrefix_RecruitmentRequirement_CreateDate")),
            "推荐人姓名": record.get("extQ202106334_109025_847145517-ExportName", ""),
            "推荐人工号": record.get("extQ202106334_109025_847145517-StaffCode", ""),
            "_status": status,
            "_sortDate": _as_date(record.get("LookupPrefix_RecruitmentRequirement_CreateDate")) or datetime.min,
        }
        offers[id_no] = _choose_offer(offers.get(id_no), offer)
    return offers


def _read_overseas_offers(path: Path) -> dict[str, dict[str, Any]]:
    offers: dict[str, dict[str, Any]] = {}
    for record in _read_records(path):
        offer_no = _norm_key(record.get("offer编码"))
        if not offer_no:
            continue
        status = _text(record.get("状态"))
        if status and status not in VALID_OVERSEAS_OFFER_STATUSES:
            continue
        offer = {
            "ABC类别": record.get("编制ABC分类", ""),
            "招聘负责人工号": record.get("招聘负责人工号", ""),
            "招聘负责人姓名": record.get("招聘负责人", ""),
            "协助招聘人工号": record.get("协助招聘负责人工号", ""),
            "协助招聘人姓名": record.get("协助招聘负责人姓名", ""),
            "招聘渠道": _normalize_channel(record.get("招聘渠道")),
            "海外offer流程编号": record.get("offer编码", ""),
            "海外招聘需求编号": record.get("招聘需求编码", ""),
            "招聘启动日期": _clean_value(record.get("招聘启动日期")),
            "推荐人姓名": record.get("内推人姓名", ""),
            "推荐人工号": record.get("内推人工号", ""),
            "_status": status,
            "_sortDate": _as_date(record.get("招聘启动日期")) or datetime.min,
        }
        offers[offer_no] = _choose_offer(offers.get(offer_no), offer)
    return offers


def _read_resignations(
    ehr_resignations: Path | str | None,
    oehr_resignations: Path | str | None,
) -> dict[str, dict[str, Any]]:
    resignations: dict[str, dict[str, Any]] = {}
    if ehr_resignations:
        for record in _read_records(Path(ehr_resignations)):
            employee_no = _norm_key(record.get("JobNumber"))
            last_work = _clean_value(record.get("LastWorkDate"))
            status = _text(record.get("ApprovalStatus"))
            if employee_no and last_work and status not in INVALID_RESIGN_STATUSES:
                resignations[employee_no] = {"人员状态": "离职", "最后工作日": last_work, "来源": "EHR离职审批"}
    if oehr_resignations:
        for record in _read_records(Path(oehr_resignations)):
            employee_no = _norm_key(record.get("工号"))
            last_work = _clean_value(record.get("最后工作日"))
            status = _text(record.get("审批状态"))
            voided = _text(record.get("是否作废"))
            if employee_no and last_work and status not in INVALID_RESIGN_STATUSES and voided != "是":
                resignations[employee_no] = {"人员状态": "离职", "最后工作日": last_work, "来源": "OEHR离职审批"}
    return resignations


def _read_cycle_exclusions(special_approvals: Path | str | None) -> dict[str, dict[str, Any]]:
    exclusions: dict[str, dict[str, Any]] = {}
    if not special_approvals:
        return exclusions
    for record in _read_records(Path(special_approvals)):
        if "招聘周期" not in _text(record.get("事项分类")) or "剔除" not in _text(record.get("事项分类")):
            continue
        if _text(record.get("文档状态")) != "结束":
            continue
        employee_no = _employee_lookup_key(record.get("申请人工号"))
        days = _as_number(record.get("招聘周期特殊剔除时间（天）"))
        if not employee_no or days <= 0:
            continue
        current = exclusions.get(employee_no)
        if not current or days > current.get("周期剔除天数", 0):
            exclusions[employee_no] = {
                "周期剔除天数": days,
                "来源": record.get("申请单编号", "") or "特殊事项申请（HR）",
                "申请说明": record.get("申请说明", ""),
            }
    return exclusions


def _lookup_domestic_offer(person: dict[str, Any], offers: dict[str, dict[str, Any]]) -> dict[str, Any]:
    return offers.get(_norm_key(person.get("证件号")), {})


def _new_domestic_row(
    target_month: int,
    person: dict[str, Any],
    offer: dict[str, Any],
    diagnostics: dict[str, Any],
) -> dict[str, Any]:
    row = _blank_standard_row(target_month)
    _copy_fields(row, person, [
        "姓名", "工号", "证件号", "入职日期", "人员状态", "工作地", "最后工作日", "职级",
        "二级组织", "三级组织", "四级组织", "五级组织", "职位", "直线经理", "直线经理工号",
        "标签分类", "候选人入职时间", "转正日期", "直接上级工号",
    ])
    _copy_fields(row, offer, [
        "ABC类别", "招聘负责人工号", "招聘负责人姓名", "协助招聘人工号", "协助招聘人姓名",
        "招聘渠道", "招聘启动日期", "推荐人姓名", "推荐人工号",
    ])
    row["ABC类别"] = _normalize_abc(row.get("ABC类别"), row.get("标签分类"))
    row["奖金地区类型"] = _normalized_region_type(row)
    row["推荐人是否上级"] = "否"
    if not offer:
        _issue(diagnostics, "国内新增人员未匹配offer", row)
    _required_issue(diagnostics, row, ["ABC类别", "招聘负责人工号", "招聘渠道"])
    return row


def _new_overseas_row(
    target_month: int,
    person: dict[str, Any],
    offer: dict[str, Any],
    diagnostics: dict[str, Any],
) -> dict[str, Any]:
    row = _blank_standard_row(target_month)
    _copy_fields(row, person, [
        "姓名", "工号", "证件号", "入职日期", "人员状态", "工作地", "最后工作日", "职级",
        "二级组织", "三级组织", "四级组织", "五级组织", "职位", "职族", "职类", "直线经理",
        "直线经理工号", "标签分类", "海外offer流程编号", "候选人入职时间", "直接上级工号",
    ])
    _copy_fields(row, offer, [
        "ABC类别", "招聘负责人工号", "招聘负责人姓名", "协助招聘人工号", "协助招聘人姓名",
        "招聘渠道", "海外招聘需求编号", "招聘启动日期", "推荐人姓名", "推荐人工号",
    ])
    row["ABC类别"] = _normalize_abc(row.get("ABC类别"), row.get("标签分类"))
    row["奖金地区类型"] = _normalized_region_type(row)
    if not offer:
        _issue(diagnostics, "海外新增人员未匹配offer", row)
    _required_issue(diagnostics, row, ["ABC类别", "招聘负责人工号", "招聘渠道"])
    return row


def _refresh_dynamic_fields(
    row: dict[str, Any],
    people_by_no: dict[str, dict[str, Any]],
    resignations: dict[str, dict[str, Any]],
    diagnostics: dict[str, Any],
) -> None:
    employee = people_by_no.get(_norm_key(row.get("工号")), {})
    resign = resignations.get(_norm_key(row.get("工号")), {})
    if employee:
        for field in ("人员状态", "最后工作日", "转正日期", "直线经理", "直线经理工号", "直接上级工号"):
            value = employee.get(field)
            if value not in (None, ""):
                _set_with_audit(row, field, value, "花名册", diagnostics)
    if resign:
        _set_with_audit(row, "人员状态", "离职", resign.get("来源", "离职审批"), diagnostics)
        _set_with_audit(row, "最后工作日", resign.get("最后工作日"), resign.get("来源", "离职审批"), diagnostics)

    role_fields = [
        ("招聘负责人工号", "招聘负责人人员状态", "招聘负责人最后工作日"),
        ("协助招聘人工号", "协助招聘人人员状态", "协助招聘人最后工作日"),
        ("推荐人工号", "推荐人人员状态", "推荐人最后工作日"),
    ]
    for id_field, status_field, last_work_field in role_fields:
        role_no = _employee_lookup_key(row.get(id_field))
        if not role_no or role_no == "-":
            continue
        role = people_by_no.get(role_no, {})
        role_resign = resignations.get(role_no, {})
        if role:
            if role.get("人员状态") not in (None, ""):
                _set_with_audit(row, status_field, _role_status_label(role.get("人员状态")), "花名册", diagnostics)
            if role.get("最后工作日") not in (None, ""):
                _set_with_audit(row, last_work_field, role.get("最后工作日"), "花名册", diagnostics)
        if role_resign:
            if status_field != "推荐人人员状态":
                _set_with_audit(row, status_field, "离职", role_resign.get("来源", "离职审批"), diagnostics)
            _set_with_audit(row, last_work_field, role_resign.get("最后工作日"), role_resign.get("来源", "离职审批"), diagnostics)
        if id_field == "推荐人工号" and role:
            _set_with_audit(row, "推荐人职级", role.get("推荐人职级", role.get("职级", "")), "花名册", diagnostics)
            _set_with_audit(row, "推荐人职位", role.get("推荐人职位", role.get("职位", "")), "花名册", diagnostics)

    _refresh_manager_chain(row, people_by_no, diagnostics)
    normalized_abc = _normalize_abc(row.get("ABC类别"), row.get("标签分类"))
    if normalized_abc:
        _set_with_audit(row, "ABC类别", normalized_abc, "字段规范化", diagnostics)
    normalized_region = _normalized_region_type(row)
    if normalized_region:
        _set_with_audit(row, "奖金地区类型", normalized_region, "工作地规则", diagnostics)
    special_region_rule = _normalized_special_region_rule(row)
    if special_region_rule:
        _set_with_audit(row, "特殊地区规则", special_region_rule, "部门规则", diagnostics)
    for field in ("推荐人姓名", "推荐人工号"):
        if _text(row.get(field)) == "-":
            old = row.get(field, "")
            row[field] = ""
            diagnostics["changes"].append({
                "工号": row.get("工号", ""),
                "姓名": row.get("姓名", ""),
                "字段": field,
                "原值": old,
                "新值": "",
                "来源": "空值规范化",
            })


def _refresh_manager_chain(
    row: dict[str, Any],
    people_by_no: dict[str, dict[str, Any]],
    diagnostics: dict[str, Any],
) -> None:
    manager_no = _employee_lookup_key(row.get("直接上级工号") or row.get("直线经理工号"))
    if not manager_no:
        return
    _set_with_audit(row, "直接上级工号", manager_no, "花名册上级链", diagnostics)
    current_no = manager_no
    for level in range(1, 5):
        manager = people_by_no.get(current_no, {})
        next_no = _employee_lookup_key(manager.get("直接上级工号") or manager.get("直线经理工号"))
        if not next_no:
            break
        _set_with_audit(row, f"间接上级{level}工号", next_no, "花名册上级链", diagnostics)
        current_no = next_no


def _apply_cycle_exclusion(
    row: dict[str, Any],
    cycle_exclusions: dict[str, dict[str, Any]],
    diagnostics: dict[str, Any],
) -> None:
    exclusion = cycle_exclusions.get(_norm_key(row.get("工号")))
    if not exclusion:
        return
    _set_with_audit(row, "周期剔除天数", exclusion.get("周期剔除天数"), exclusion.get("来源", "特殊事项申请（HR）"), diagnostics)


def _normalize_abc(value: Any, label: Any) -> str:
    category = _text(value)
    if not category:
        return ""
    if category in {"A", "B", "C"}:
        if category == "C" and _text(label) == "海外":
            return "C1类"
        return f"{category}类"
    if category in {"A类", "B类", "C类", "C1类"}:
        return category
    return category


def _normalized_region_type(row: dict[str, Any]) -> str:
    imported = _text(row.get("奖金地区类型"))
    if imported in {"国内发展中国家", "海外发达国家", "海外发展中国家"}:
        return imported
    location = _text(row.get("工作地"))
    if _text(row.get("标签分类")) == "国内" or _is_domestic_location(location):
        return "国内发展中国家"
    location_lower = location.lower()
    if any(token in location_lower for token in DEVELOPED_LOCATION_TOKENS):
        return "海外发达国家"
    if imported == "发达国家":
        return "海外发达国家"
    if imported == "发展中国家":
        return "国内发展中国家" if _text(row.get("标签分类")) == "国内" else "海外发展中国家"
    return "海外发展中国家" if location else ""


def _is_domestic_location(location: str) -> bool:
    if not location:
        return False
    if any(token in location for token in OVERSEAS_CHINA_LOCATION_TOKENS):
        return False
    return (
        "中国大陆" in location
        or location.startswith("中国")
        or "大陆" in location
        or any(token in location for token in MAINLAND_LOCATION_TOKENS)
    )


def _read_records(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xls":
        with tempfile.TemporaryDirectory() as tmpdir:
            return _read_records(_convert_xls_to_xlsx(path, Path(tmpdir)), sheet_name)
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header_row = _detect_header_row(sheet)
    headers = [_text(sheet.cell(header_row, column).value) for column in range(1, sheet.max_column + 1)]
    records: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = _clean_value(values[index] if index < len(values) else "")
        if any(_text(value) and _text(value) != header for header, value in record.items()):
            records.append(record)
    return records


def _convert_xls_to_xlsx(path: Path, output_dir: Path) -> Path:
    soffice = shutil.which("soffice")
    if not soffice:
        raise RuntimeError("读取 .xls 需要 soffice/LibreOffice 转换工具")
    subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(path)],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    converted = output_dir / f"{path.stem}.xlsx"
    if not converted.exists():
        matches = list(output_dir.glob("*.xlsx"))
        if not matches:
            raise RuntimeError(f".xls 转换失败：{path}")
        return matches[0]
    return converted


def _detect_header_row(sheet) -> int:
    best_row = 1
    best_count = -1
    for row_number in range(1, min(sheet.max_row, 15) + 1):
        count = sum(sheet.cell(row_number, column).value not in (None, "") for column in range(1, sheet.max_column + 1))
        if count > best_count:
            best_row = row_number
            best_count = count
    return best_row


def _copy_fields(target: dict[str, Any], source: dict[str, Any], fields: Iterable[str]) -> None:
    for field in fields:
        if source.get(field) not in (None, ""):
            target[field] = source.get(field)


def _choose_offer(current: dict[str, Any] | None, candidate: dict[str, Any]) -> dict[str, Any]:
    if not current:
        return candidate
    if candidate.get("_sortDate", datetime.min) >= current.get("_sortDate", datetime.min):
        return candidate
    return current


def _set_with_audit(row: dict[str, Any], field: str, value: Any, source: str, diagnostics: dict[str, Any]) -> None:
    if value in (None, ""):
        return
    old = row.get(field, "")
    if _equal_cell(old, value):
        row[field] = value
        return
    row[field] = value
    if _text(old) or _text(value):
        diagnostics["changes"].append({
            "工号": row.get("工号", ""),
            "姓名": row.get("姓名", ""),
            "字段": field,
            "原值": old,
            "新值": value,
            "来源": source,
        })


def _issue(diagnostics: dict[str, Any], issue_type: str, row: dict[str, Any], extra: dict[str, Any] | None = None) -> None:
    item = {"类型": issue_type, "工号": row.get("工号", ""), "姓名": row.get("姓名", "")}
    if extra:
        item.update(extra)
    diagnostics["issues"].append(item)


def _required_issue(diagnostics: dict[str, Any], row: dict[str, Any], fields: list[str]) -> None:
    missing = [field for field in fields if not _text(row.get(field))]
    if missing:
        _issue(diagnostics, "关键字段缺失", row, {"缺失字段": "、".join(missing)})


def _new_row_record(row: dict[str, Any], source: str) -> dict[str, Any]:
    return {"工号": row.get("工号", ""), "姓名": row.get("姓名", ""), "标签分类": row.get("标签分类", ""), "来源": source}


def _pending_confirmation_items(compare: CompareResult, diagnostics: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in compare.missing_in_generated:
        if "人员调动" in _text(row.get("招聘渠道")):
            items.append({
                "类别": "待接入数据源",
                "姓名": row.get("姓名", ""),
                "工号": row.get("工号", ""),
                "字段": "招聘渠道",
                "当前结论": "人员调动/异动审批源未接入，暂不自动生成",
                "建议动作": "等待人员调动或异动审批导出后再验证",
            })
    for row in compare.missing_in_target:
        if _normalize_channel(row.get("招聘渠道")) == "二次入职":
            items.append({
                "类别": "待业务确认",
                "姓名": row.get("姓名", ""),
                "工号": row.get("工号", ""),
                "字段": "招聘渠道",
                "当前结论": "自动源存在二次入职记录，人工模板未包含",
                "建议动作": "确认是否应进入人工导入模板",
            })
    exclusion_days = compare.field_rates.get("周期剔除天数", {})
    diff_count = exclusion_days.get("diffCount", 0)
    if diff_count:
        cycle_source_count = diagnostics.get("counts", {}).get("cycleExclusionRows", 0)
        current_conclusion = (
            f"已接入本次OA特殊事项审批明细，PPT已列明节假日及海外特殊剔除条件，该字段仍有 {diff_count} 条差异"
            if cycle_source_count
            else f"PPT已列明节假日及海外特殊剔除条件，该字段仍有 {diff_count} 条差异"
        )
        advice = (
            "继续补充国内节假日历、完整OA审批明细，或复核人工模板差异"
            if cycle_source_count
            else "接入节假日历和OA特殊事项审批后再自动计算"
        )
        items.append({
            "类别": "待接入数据源",
            "姓名": "",
            "工号": "",
            "字段": "周期剔除天数",
            "当前结论": current_conclusion,
            "建议动作": advice,
        })
    return items


def _index_rows(rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, int]]:
    index: dict[str, dict[str, Any]] = {}
    counts: Counter[str] = Counter()
    for row in rows:
        key = _row_key(row)
        if not key:
            continue
        counts[key] += 1
        index.setdefault(key, row)
    duplicates = {key: count for key, count in counts.items() if count > 1}
    return index, duplicates


def _alias_index(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    aliases: dict[str, dict[str, Any]] = {}
    for row in rows:
        alias = _loose_row_alias(row)
        if alias:
            aliases.setdefault(alias, row)
    return aliases


def _loose_row_alias(row: dict[str, Any]) -> str:
    name = _norm_name(row.get("姓名"))
    position = _norm_key(row.get("职位"))
    month = _entry_month(row)
    if not name or not month:
        return ""
    return f"{name}|{position}|{month}"


def _entry_month(row: dict[str, Any]) -> str:
    date_value = _as_date(row.get("入职日期") or row.get("候选人入职时间"))
    if date_value:
        return f"{date_value.year:04d}{date_value.month:02d}"
    month = _text(row.get("核算月份"))
    return month if re.fullmatch(r"\d{6}", month) else ""


def _row_key(row: dict[str, Any]) -> str:
    employee_no = _norm_key(row.get("工号"))
    if employee_no:
        return f"emp:{employee_no}"
    id_no = _norm_key(row.get("证件号"))
    if id_no:
        return f"id:{id_no}"
    name = _norm_key(row.get("姓名"))
    onboard = _norm_key(_display_value(row.get("入职日期") or row.get("候选人入职时间")))
    return f"name:{name}|{onboard}" if name else ""


def _sort_key(row: dict[str, Any]) -> tuple[int, str, str]:
    label_order = 0 if _text(row.get("标签分类")) == "国内" else 1
    date_value = _as_date(row.get("入职日期") or row.get("候选人入职时间"))
    date_key = date_value.strftime("%Y%m%d") if date_value else "99999999"
    return label_order, date_key, _text(row.get("工号") or row.get("姓名"))


def _is_target_month(value: Any, target_month: int) -> bool:
    date_value = _as_date(value)
    return bool(date_value and date_value.year * 100 + date_value.month == target_month)


def _as_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and 20000 <= value <= 80000:
        try:
            return from_excel(value)
        except (TypeError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text[:19], fmt)
            except ValueError:
                continue
    return None


def _as_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = _text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _clean_value(value: Any) -> Any:
    if hasattr(value, "text"):
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_key(value: Any) -> str:
    return _text(value).lower()


def _norm_name(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value)).lower()


def _employee_lookup_key(value: Any) -> str:
    text = _norm_key(value).replace(" ", "")
    if not text or text == "-":
        return text
    if "-" in text:
        tail = text.split("-")[-1]
        if tail:
            return tail
    match = re.search(r"(zt[a-z0-9]+|lg[a-z0-9]+)$", text)
    return match.group(1) if match else text


def _role_status_label(value: Any) -> str:
    status = _text(value)
    if status in {"正式", "试用"}:
        return "正式在职"
    return status


def _normalize_channel(value: Any) -> str:
    channel = _text(value)
    if not channel or channel == "-":
        return channel
    if "内推" in channel or "内部推荐" in channel:
        return "内推"
    if "二次" in channel:
        return "二次入职"
    if "返聘" in channel:
        return "返聘"
    if "校招" in channel:
        return "校招"
    if "猎头" in channel:
        return "猎头"
    return "招聘网站"


def _equal_cell(left: Any, right: Any) -> bool:
    left_date = _as_date(left)
    right_date = _as_date(right)
    if left_date or right_date:
        return bool(left_date and right_date and left_date.date() == right_date.date())
    return _text(left) == _text(right)


def _equal_cell_for_field(
    field: str,
    left: Any,
    right: Any,
    generated: dict[str, Any],
    target: dict[str, Any],
) -> bool:
    related_id_field = ROLE_STATUS_FIELDS.get(field)
    if related_id_field:
        left_role_no = _employee_lookup_key(generated.get(related_id_field))
        right_role_no = _employee_lookup_key(target.get(related_id_field))
        if not left_role_no and not right_role_no:
            return _placeholder_or_default_role_value(left) and _placeholder_or_default_role_value(right)

    normalized_left = _normalize_compare_value(field, left, generated)
    normalized_right = _normalize_compare_value(field, right, target)
    return _equal_cell(normalized_left, normalized_right)


def _is_auto_enrichment(field: str, generated_value: Any, target_value: Any) -> bool:
    if field not in AUTO_ENRICHMENT_FIELDS:
        return False
    generated_text = _text(generated_value)
    target_text = _text(target_value)
    return bool(generated_text and target_text in {"", "-", "0"})


def _normalize_compare_value(field: str, value: Any, row: dict[str, Any]) -> Any:
    text = _text(value)
    if field in EMPTY_EQUIVALENT_FIELDS and text == "-":
        return ""
    if field in EMPLOYEE_NO_FIELDS:
        return _employee_lookup_key(value)
    if field == "ABC类别":
        return _normalize_abc(value, row.get("标签分类"))
    if field == "奖金地区类型":
        return _normalized_region_type({**row, field: value})
    if field == "特殊地区规则":
        return _normalized_special_region_rule({**row, field: value})
    if field == "周期剔除天数" and text in {"", "-"}:
        return 0
    if field in ROLE_STATUS_FIELDS and text in {"正式", "试用"}:
        return "正式在职"
    return value


def _placeholder_or_default_role_value(value: Any) -> bool:
    return _text(value) in {"", "-", "正式在职"}


def _normalized_special_region_rule(row: dict[str, Any]) -> str:
    explicit = _text(row.get("特殊地区规则"))
    if explicit:
        return explicit
    referrer_no = _employee_lookup_key(row.get("推荐人工号"))
    if not referrer_no or referrer_no == "-":
        return ""
    manager_nos = {_employee_lookup_key(row.get(field)) for field in (
        "直接上级工号",
        "间接上级1工号",
        "间接上级2工号",
        "间接上级3工号",
        "间接上级4工号",
    )}
    if referrer_no in manager_nos:
        return ""
    second = _text(row.get("二级组织"))
    third = _text(row.get("三级组织"))
    fourth = _text(row.get("四级组织"))
    if second == "FBU仓储事业部":
        if "德国" in fourth:
            return "FBU德国"
        if "捷克" in fourth:
            return "FBU捷克"
    if second == "HRAS人力综合条线" and "FBU HRBP" in third:
        if "德国" in fourth:
            return "FBU德国"
        if "捷克" in fourth:
            return "FBU捷克"
    return ""


def _display_value(value: Any) -> Any:
    date_value = _as_date(value)
    if date_value:
        return date_value.strftime("%Y/%m/%d")
    return value


def _write_dict_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]], headers: list[str] | None) -> None:
    sheet = workbook.create_sheet(title)
    _write_rows_to_sheet(sheet, rows, headers)


def _write_rows_to_sheet(sheet, rows: list[dict[str, Any]], headers: list[str] | None) -> None:
    if headers is None:
        header_set: list[str] = []
        for row in rows:
            for key in row:
                if key not in header_set:
                    header_set.append(key)
        headers = header_set or ["提示"]
    sheet.append(headers)
    if rows:
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])


def _format_standard_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="F4B183")
    border = Border(bottom=Side(style="thin", color="D9E2EC"))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for column in range(1, sheet.max_column + 1):
        header = sheet.cell(1, column).value
        width = max(12, min(24, len(_text(header)) + 4))
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = DATE_FORMAT
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _format_report_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        max_len = max(len(_text(sheet.cell(row, column).value)) for row in range(1, min(sheet.max_row, 100) + 1))
        sheet.column_dimensions[get_column_letter(column)].width = max(10, min(42, max_len + 4))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
