from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook

from .models import as_text


@dataclass(frozen=True)
class ReferralRule:
    scope: str
    region_type: str
    category: str
    grade: str
    amount: float
    currency: str
    ratio_1m: float
    ratio_3m: float
    ratio_6m: float
    ratio_probation: float
    source: str
    note: str


@dataclass
class RuleBook:
    cycle_days: Dict[Tuple[str, str, str], float]
    recruitment_bonus: Dict[Tuple[str, str, str], Tuple[float, str]]
    channel_ratio: Dict[str, float]
    referral_bonus: Dict[Tuple[str, str, str, str], ReferralRule]

    def cycle(self, label: str, category: str, grade: str) -> Optional[float]:
        return self.cycle_days.get((label, category, grade))

    def recruitment(self, label: str, category: str, grade: str) -> Tuple[float, str]:
        return self.recruitment_bonus.get((label, category, grade), (0.0, ""))

    def channel(self, channel_name: str) -> float:
        return self.channel_ratio.get(channel_name, 1.0)

    def referral(self, scope: str, region_type: str, category: str, grade: str) -> Optional[ReferralRule]:
        normalized_category = "C类" if scope in {"FBU德国", "FBU捷克"} and category == "C1类" else category
        return self.referral_bonus.get((scope, region_type, normalized_category, grade))


def _to_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def load_rulebook(path: Path) -> RuleBook:
    workbook = load_workbook(path, data_only=True, read_only=False)

    cycle_days: Dict[Tuple[str, str, str], float] = {}
    ws = workbook["规则_招聘周期"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        label, category, grade, days = row[:4]
        if label and category and grade and days not in (None, "", "-"):
            cycle_days[(as_text(label), as_text(category), as_text(grade))] = _to_float(days)

    recruitment_bonus: Dict[Tuple[str, str, str], Tuple[float, str]] = {}
    ws = workbook["规则_招聘奖金"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        label, category, grade, amount, currency = row[:5]
        if label and category and grade and amount not in (None, "", "-"):
            recruitment_bonus[(as_text(label), as_text(category), as_text(grade))] = (_to_float(amount), as_text(currency))

    channel_ratio: Dict[str, float] = {}
    ws = workbook["规则_渠道系数"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        channel, ratio = row[:2]
        if channel not in (None, ""):
            channel_ratio[as_text(channel)] = _to_float(ratio)

    referral_bonus: Dict[Tuple[str, str, str, str], ReferralRule] = {}
    ws = workbook["规则_内推奖金"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        scope, region_type, category, grade, amount, currency, p1, p3, p6, pp, source, _key, note = row[:13]
        if not (scope and region_type and category and grade):
            continue
        rule = ReferralRule(
            scope=as_text(scope),
            region_type=as_text(region_type),
            category=as_text(category),
            grade=as_text(grade),
            amount=_to_float(amount),
            currency=as_text(currency),
            ratio_1m=_to_float(p1),
            ratio_3m=_to_float(p3),
            ratio_6m=_to_float(p6),
            ratio_probation=_to_float(pp),
            source=as_text(source),
            note=as_text(note),
        )
        referral_bonus[(rule.scope, rule.region_type, rule.category, rule.grade)] = rule

    return RuleBook(
        cycle_days=cycle_days,
        recruitment_bonus=recruitment_bonus,
        channel_ratio=channel_ratio,
        referral_bonus=referral_bonus,
    )
