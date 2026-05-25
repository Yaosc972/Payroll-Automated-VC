from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_MONEY_FIELDS = ["入职1月奖金", "入职3月奖金", "入职6月奖金", "转正奖金", "合计发放"]
RECRUITMENT_KEY_FIELDS = ["工号", "姓名", "币种"]
REFERRAL_KEY_FIELDS = ["推荐人工号", "推荐人姓名", "币种"]
DETAIL_KEY_FIELDS = ["工号", "姓名"]


def build_difference_report(platform_path: Path, offline_path: Path, output_path: Path) -> Dict[str, Any]:
    platform = load_workbook(platform_path, data_only=True, read_only=True)
    offline = load_workbook(offline_path, data_only=True, read_only=True)

    platform_recruitment = _aggregate_amount_rows(
        _sheet_dicts(platform, "最终招聘奖金汇总") or _sheet_dicts(platform, "招聘奖金汇总"),
        RECRUITMENT_KEY_FIELDS,
    )
    platform_referral = _aggregate_amount_rows(
        _sheet_dicts(platform, "最终内推奖金汇总") or _sheet_dicts(platform, "内推奖金汇总"),
        REFERRAL_KEY_FIELDS,
    )
    platform_detail = _sheet_dicts(platform, "招聘奖金明细")

    offline_recruitment = _aggregate_amount_rows(
        _first_existing_sheet(offline, ["附件二-招聘奖金汇总表", "招聘奖金汇总", "最终招聘奖金汇总"]),
        RECRUITMENT_KEY_FIELDS,
    )
    offline_referral = _aggregate_amount_rows(
        _first_existing_sheet(offline, ["附件三-内推奖金汇总", "内推奖金汇总", "最终内推奖金汇总"]),
        REFERRAL_KEY_FIELDS,
    )
    offline_detail = _first_existing_sheet(offline, ["附件一-招聘奖金核算明细", "招聘奖金明细"])

    recruitment_diffs = _compare_amount_rows(platform_recruitment, offline_recruitment, RECRUITMENT_KEY_FIELDS, SUMMARY_MONEY_FIELDS)
    referral_diffs = _compare_amount_rows(platform_referral, offline_referral, REFERRAL_KEY_FIELDS, ["合计发放"])
    recruitment_detail_diffs = _compare_detail_rows(platform_detail, offline_detail, "招聘")
    referral_detail_diffs = _compare_detail_rows(platform_detail, offline_detail, "内推")

    metrics = {
        "recruitmentSummaryDiffCount": len(recruitment_diffs),
        "referralSummaryDiffCount": len(referral_diffs),
        "recruitmentDetailDiffCount": len(recruitment_detail_diffs),
        "referralDetailDiffCount": len(referral_detail_diffs),
        "recruitmentSummaryDelta": round(sum(_number(row.get("差异金额")) for row in recruitment_diffs), 2),
        "referralSummaryDelta": round(sum(_number(row.get("差异金额")) for row in referral_diffs), 2),
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, "摘要", ["项目", "值"], [{"项目": key, "值": value} for key, value in metrics.items()])
    _write_sheet(workbook, "招聘汇总差异", _summary_diff_headers(RECRUITMENT_KEY_FIELDS), recruitment_diffs)
    _write_sheet(workbook, "内推汇总差异", _summary_diff_headers(REFERRAL_KEY_FIELDS), referral_diffs)
    _write_sheet(workbook, "招聘明细差异", _detail_diff_headers(), recruitment_detail_diffs)
    _write_sheet(workbook, "内推明细差异", _detail_diff_headers(), referral_detail_diffs)
    workbook.save(output_path)
    return metrics


def _first_existing_sheet(workbook, names: Iterable[str]) -> List[Dict[str, Any]]:
    for name in names:
        rows = _sheet_dicts(workbook, name)
        if rows:
            return rows
    return []


def _sheet_dicts(workbook, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    if sheet_name == "附件一-招聘奖金核算明细":
        header_rows = list(sheet.iter_rows(min_row=1, max_row=3, values_only=True))
        if len(header_rows) >= 3 and _value_at(header_rows[2], 3):
            return _payroll_detail_rows(sheet, header_rows)
    elif sheet_name == "附件二-招聘奖金汇总表":
        header_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        if len(header_rows) >= 2 and _value_at(header_rows[1], 7):
            return _payroll_recruitment_summary_rows(sheet)
    elif sheet_name == "附件三-内推奖金汇总":
        header_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        if len(header_rows) >= 2 and _value_at(header_rows[1], 7) == "本月发放":
            return _payroll_referral_summary_rows(sheet)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    result: List[Dict[str, Any]] = []
    for values in rows[1:]:
        row = {str(header).strip(): value for header, value in zip(headers, values) if header}
        if _has_values(row):
            result.append(row)
    return result


def _value_at(values: tuple[Any, ...], index: int) -> Any:
    if index >= len(values):
        return None
    return values[index]


def _payroll_recruitment_summary_rows(sheet) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        row = {
            "工号": _value_at(values, 1),
            "姓名": _value_at(values, 2),
            "币种": _value_at(values, 4),
            "入职1月奖金": _value_at(values, 7),
            "入职3月奖金": _value_at(values, 8),
            "入职6月奖金": _value_at(values, 9),
            "转正奖金": _value_at(values, 10),
            "合计发放": _value_at(values, 11),
        }
        if _has_values(row):
            rows.append(row)
    return rows


def _payroll_referral_summary_rows(sheet) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        row = {
            "推荐人工号": _value_at(values, 1),
            "推荐人姓名": _value_at(values, 2),
            "币种": _value_at(values, 6),
            "入职1月奖金": 0,
            "入职3月奖金": 0,
            "入职6月奖金": 0,
            "转正奖金": 0,
            "合计发放": _value_at(values, 7),
        }
        if _has_values(row):
            rows.append(row)
    return rows


def _payroll_detail_rows(sheet, header_rows: List[tuple[Any, ...]]) -> List[Dict[str, Any]]:
    groups = header_rows[1]
    subheaders = header_rows[2]
    headers = [_payroll_detail_header(group, subheader, column) for column, (group, subheader) in enumerate(zip(groups, subheaders), start=1)]
    rows: List[Dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=4, values_only=True):
        row = {header: value for header, value in zip(headers, values) if header}
        if _has_values(row):
            rows.append(row)
    return rows


def _payroll_detail_header(group: Any, subheader: Any, column: int) -> str:
    group_text = str(group or "").strip()
    sub_text = str(subheader or "").strip()
    direct_map = {
        2: "姓名",
        3: "工号",
    }
    if column in direct_map:
        return direct_map[column]
    if "招聘负责人入职奖金明细" in group_text:
        return _node_header("招聘人", sub_text)
    if "协助人入职奖金明细" in group_text:
        return _node_header("协助人", sub_text)
    if "招聘人转正奖金" in group_text and "转正发放奖金" in sub_text:
        return "招聘人转正奖金"
    if "协助人转正奖金" in group_text and "转正发放奖金" in sub_text:
        return "协助人转正奖金"
    if "内推奖明细" in group_text:
        return _node_header("内推", sub_text)
    return str(group or subheader or "").strip()


def _node_header(prefix: str, sub_text: str) -> str:
    if "入职1个月发放奖金" in sub_text:
        return f"{prefix}入职1月奖金"
    if "入职3个月发放奖金" in sub_text:
        return f"{prefix}入职3月奖金"
    if "入职6个月" in sub_text and "发放奖金" in sub_text:
        return f"{prefix}入职6月奖金"
    if "转正发放奖金" in sub_text:
        return f"{prefix}转正奖金"
    return sub_text


def _has_values(row: Dict[str, Any]) -> bool:
    if any(str(value).strip() in {"合计", "总计"} for value in row.values() if value is not None):
        return False
    return any(value not in (None, "") for value in row.values())


def _aggregate_amount_rows(rows: List[Dict[str, Any]], key_fields: List[str]) -> List[Dict[str, Any]]:
    grouped: Dict[tuple[str, ...], Dict[str, Any]] = {}
    for row in rows:
        key = _row_key(row, key_fields)
        if not any(key):
            continue
        target = grouped.setdefault(key, {field: row.get(field, "") for field in key_fields})
        for field in SUMMARY_MONEY_FIELDS:
            target[field] = round(_number(target.get(field)) + _number(row.get(field)), 2)
    return list(grouped.values())


def _compare_amount_rows(platform_rows: List[Dict[str, Any]], offline_rows: List[Dict[str, Any]], key_fields: List[str], money_fields: List[str]) -> List[Dict[str, Any]]:
    platform_map = {_row_key(row, key_fields): row for row in platform_rows}
    offline_map = {_row_key(row, key_fields): row for row in offline_rows}
    diffs: List[Dict[str, Any]] = []

    for key in sorted(set(platform_map) | set(offline_map)):
        platform_row = platform_map.get(key, {})
        offline_row = offline_map.get(key, {})
        for field in money_fields:
            platform_amount = _number(platform_row.get(field))
            offline_amount = _number(offline_row.get(field))
            delta = round(platform_amount - offline_amount, 2)
            if delta == 0:
                continue
            diff = {field_name: value for field_name, value in zip(key_fields, key)}
            diff.update(
                {
                    "字段": field,
                    "平台金额": platform_amount,
                    "线下金额": offline_amount,
                    "差异金额": delta,
                    "差异类型": _classify_delta(delta),
                }
            )
            diffs.append(diff)
    return sorted(diffs, key=lambda row: abs(_number(row.get("差异金额"))), reverse=True)


def _compare_detail_rows(platform_rows: List[Dict[str, Any]], offline_rows: List[Dict[str, Any]], bonus_type: str) -> List[Dict[str, Any]]:
    fields = _detail_fields(bonus_type)
    platform_map = {_row_key(row, DETAIL_KEY_FIELDS): row for row in platform_rows}
    offline_map = {_row_key(row, DETAIL_KEY_FIELDS): row for row in offline_rows}
    diffs: List[Dict[str, Any]] = []
    for key in sorted(set(platform_map) | set(offline_map)):
        platform_row = platform_map.get(key, {})
        offline_row = offline_map.get(key, {})
        for field in fields:
            platform_amount = _number(platform_row.get(field))
            offline_amount = _number(offline_row.get(field))
            delta = round(platform_amount - offline_amount, 2)
            if delta == 0:
                continue
            diffs.append(
                {
                    "工号": key[0],
                    "姓名": key[1],
                    "字段": field,
                    "平台金额": platform_amount,
                    "线下金额": offline_amount,
                    "差异金额": delta,
                    "差异类型": _classify_delta(delta),
                }
            )
    return sorted(diffs, key=lambda row: abs(_number(row.get("差异金额"))), reverse=True)


def _detail_fields(bonus_type: str) -> List[str]:
    if bonus_type == "招聘":
        return [
            "招聘人入职1月奖金",
            "招聘人入职3月奖金",
            "招聘人入职6月奖金",
            "招聘人转正奖金",
            "协助人入职1月奖金",
            "协助人入职3月奖金",
            "协助人入职6月奖金",
            "协助人转正奖金",
        ]
    return ["内推入职1月奖金", "内推入职3月奖金", "内推入职6月奖金", "内推转正奖金"]


def _row_key(row: Dict[str, Any], key_fields: List[str]) -> tuple[str, ...]:
    return tuple(_normalize_key(row.get(field)) for field in key_fields)


def _normalize_key(value: Any) -> str:
    text = str(value or "").strip()
    if text.upper() in {"RMB", "CNY"}:
        return "人民币"
    return text


def _number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _classify_delta(delta: float) -> str:
    if abs(delta) <= 0.05:
        return "四舍五入"
    if delta > 0:
        return "平台多算"
    return "平台少算"


def _summary_diff_headers(key_fields: List[str]) -> List[str]:
    return key_fields + ["字段", "平台金额", "线下金额", "差异金额", "差异类型"]


def _detail_diff_headers() -> List[str]:
    return ["工号", "姓名", "字段", "平台金额", "线下金额", "差异金额", "差异类型"]


def _write_sheet(workbook: Workbook, title: str, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(max(sheet.max_column, 1))}{max(sheet.max_row, 1)}"
    for column in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 18
