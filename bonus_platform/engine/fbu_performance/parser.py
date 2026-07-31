"""FBU绩效核算引擎 - 数据解析"""
from __future__ import annotations
from pathlib import Path
from typing import Optional
from collections import defaultdict
from calendar import monthrange
from datetime import datetime, date, time, timedelta
import re
import openpyxl
import msoffcrypto
import io
import xlrd

from .engines.base import (
    CalculationSegment,
    EmployeeData,
    FBUPerformanceEngine,
    FIXED_BASE_OVERRIDE_PATH,
    NINETY_SIX_HOUR_AUTO_BASE_PATH,
    NINETY_SIX_HOUR_FIXED_BASE_PATH,
)
from .engines.attendance import AttendanceProcessor
from .engines.salary import SalaryProcessor
from .engines.bonus import BonusCalculator
from .engines.coefficient import CoefficientCalculator


def _cell(row, index: int, default=None):
    if index is None:
        return default
    return row[index] if len(row) > index else default


def _to_float(value, default=None):
    if value is None or value == "":
        return default
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "").replace("$", "")
        if not cleaned:
            return default
        if cleaned.endswith("%"):
            cleaned = cleaned[:-1].strip()
            return float(cleaned) / 100
        return float(cleaned)
    return float(value)


def safe_float(value) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_excel_datetime(value) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, (int, float)):
        try:
            parsed = openpyxl.utils.datetime.from_excel(value)
            if isinstance(parsed, datetime):
                return parsed
            if isinstance(parsed, date):
                return datetime(parsed.year, parsed.month, parsed.day)
        except Exception:
            return None
    text = str(value).strip()
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        pass
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def infer_period_from_sheet_name(sheet_name: str, calc_month: str) -> str:
    ranges = re.findall(r"(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})", sheet_name or "")
    if not ranges:
        return ""
    start_month, start_day, end_month, end_day = ranges[-1]
    calc_month_number = int(str(calc_month).split("-")[1])
    if int(start_month) < calc_month_number and int(end_month) == calc_month_number:
        return f"{calc_month_number}.1-{end_month}.{end_day}"
    if int(start_month) == calc_month_number and int(end_month) > calc_month_number:
        return f"{start_month}.{start_day}-{calc_month_number}.30"
    return f"{start_month}.{start_day}-{end_month}.{end_day}"


def _parse_period_range(text: str) -> tuple[int, int, int, int] | None:
    match = re.search(r"(\d{1,2})\.(\d{1,2})-(\d{1,2})\.(\d{1,2})", str(text or ""))
    if not match:
        return None
    return tuple(int(part) for part in match.groups())


def _period_label_for_calc_month(text: str, calc_month: str) -> str:
    parsed = _parse_period_range(text)
    if not parsed:
        return ""
    start_month, start_day, end_month, end_day = parsed
    calc_month_number = int(str(calc_month).split("-")[1])
    if end_month < calc_month_number or start_month > calc_month_number:
        return ""
    _, last_day = monthrange(int(str(calc_month).split("-")[0]), calc_month_number)
    clipped_start = max(start_day if start_month == calc_month_number else 1, 1)
    clipped_end = min(end_day if end_month == calc_month_number else last_day, last_day)
    if clipped_start > clipped_end:
        return ""
    return f"{calc_month_number}.{clipped_start}-{calc_month_number}.{clipped_end}"


def _next_month(calc_month: str) -> str:
    year, month = (int(part) for part in str(calc_month).split("-")[:2])
    if month == 12:
        return f"{year + 1}-01"
    return f"{year}-{month + 1:02d}"


def _monthly_96_hour_cap(calc_month: str | None) -> float:
    return sum(period["cap_hours"] for period in _iter_96_hour_periods(calc_month))


def _iter_96_hour_periods(calc_month: str | None) -> list[dict]:
    month_start = _calc_month_start(calc_month)
    if not month_start:
        return []
    _, last_day = monthrange(month_start.year, month_start.month)
    month_end = date(month_start.year, month_start.month, last_day)

    # New Jersey payroll periods are biweekly; 2026-03-29 is the known anchor
    # for the April 2026 materials. Sum the month-intersecting period caps so
    # we never apply a single 96h cap to a whole calendar month.
    period_start = date(2026, 3, 29)
    while period_start > month_start:
        period_start -= timedelta(days=14)
    while period_start + timedelta(days=13) < month_start:
        period_start += timedelta(days=14)

    periods = []
    while period_start <= month_end:
        period_end = period_start + timedelta(days=13)
        overlap_start = max(period_start, month_start)
        overlap_end = min(period_end, month_end)
        if overlap_start <= overlap_end:
            overlap_days = (overlap_end - overlap_start).days + 1
            periods.append({
                "period": (
                    f"{month_start.year}-{month_start.month:02d} / "
                    f"{overlap_start.month}.{overlap_start.day}-{overlap_end.month}.{overlap_end.day}"
                ),
                "period_start": overlap_start.isoformat(),
                "period_end": overlap_end.isoformat(),
                "cap_hours": min(96.0, overlap_days * 8.0),
            })
        period_start += timedelta(days=14)
    return periods


def _daily_row_number(row: dict, english_key: str, chinese_key: str) -> float:
    return safe_float(row.get(english_key, row.get(chinese_key, 0)))


def _daily_row_date(row: dict) -> date | None:
    parsed = parse_excel_datetime(row.get("date") or row.get("考勤日期"))
    return parsed.date() if parsed else None


def _parse_time_value(value) -> time | None:
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = str(value or "").strip()
    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return time(hour, minute)


def _daily_row_interval(row: dict) -> tuple[datetime, datetime] | None:
    row_date = _daily_row_date(row)
    if not row_date:
        return None

    shift_time = _parse_time_value(
        row.get("shift_start_time")
        or row.get("班次上班时间")
        or row.get("shift_name")
        or row.get("班次名称")
    )
    if not shift_time:
        start_at = datetime.combine(row_date, time.min)
        return start_at, start_at + timedelta(days=1)

    start_at = datetime.combine(row_date, shift_time)
    duration_hours = max(
        _daily_row_number(row, "work_hours", "工作时长"),
        _daily_row_number(row, "scheduled_hours", "应出勤时长"),
        _daily_row_number(row, "attendance_hours", "出勤时长"),
        _daily_row_number(row, "base_hours", "计薪出勤")
        + _daily_row_number(row, "ot15_hours", "OT1.5")
        + _daily_row_number(row, "ot20_hours", "OT2.0")
        + _daily_row_number(row, "sick_hours", "病假")
        + _daily_row_number(row, "annual_hours", "年假"),
    )
    if duration_hours <= 0:
        duration_hours = 8.0
    end_at = start_at + timedelta(hours=duration_hours)
    return start_at, end_at


def _calc_month_end(month_start: date | None) -> date | None:
    if not month_start:
        return None
    _, last_day = monthrange(month_start.year, month_start.month)
    return date(month_start.year, month_start.month, last_day)


def _period_label(start: date, end: date) -> str:
    return f"{start.month}.{start.day}-{end.month}.{end.day}"


def _is_night_daily_row(row: dict) -> bool:
    return str(row.get("shift_type") or row.get("班次") or "").strip() == "夜班"


def _daily_rows_performance_base(rows: list[dict], hourly_rate: float) -> float:
    amount = 0.0
    for row in rows:
        rate_delta = row.get("hourly_rate_delta")
        if rate_delta is None:
            rate_delta = 1 if _is_night_daily_row(row) else 0
        row_rate = hourly_rate + safe_float(rate_delta)
        amount += row_rate * (
            _daily_row_number(row, "base_hours", "计薪出勤")
            + _daily_row_number(row, "ot15_hours", "OT1.5") * 1.5
            + _daily_row_number(row, "ot20_hours", "OT2.0") * 2
            + _daily_row_number(row, "sick_hours", "病假")
            + _daily_row_number(row, "sick_settlement_hours", "病假清算")
            + _daily_row_number(row, "annual_hours", "年假")
            + _daily_row_number(row, "holiday_hours", "节假日")
        )
    return round(amount, 2)


HOURLY_RATE_POLICY_BASE = "base"
HOURLY_RATE_POLICY_NIGHT = "night"
HOURLY_RATE_POLICY_BY_SHIFT = "by_shift"
HOURLY_RATE_POLICY_CHOICES = {
    HOURLY_RATE_POLICY_BASE,
    HOURLY_RATE_POLICY_NIGHT,
    HOURLY_RATE_POLICY_BY_SHIFT,
}


def _actual_work_hours(row: dict) -> float:
    return (
        _daily_row_number(row, "base_hours", "计薪出勤")
        + _daily_row_number(row, "ot15_hours", "OT1.5")
        + _daily_row_number(row, "ot20_hours", "OT2.0")
    )


def _hourly_rate_policy_row_id(employee_id: str, period_start: str) -> str:
    return f"{normalize_shift_employee_id(employee_id)}|{period_start}"


def build_hourly_rate_policy_data(
    attendance_preview: dict | None,
    calc_month: str | None,
    existing: dict | None = None,
) -> dict:
    """Build non-blocking payroll-period rate suggestions from attendance facts."""
    employees = (attendance_preview or {}).get("employees", [])
    existing_rows = {
        str(row.get("row_id") or ""): row
        for row in (existing or {}).get("rows", [])
        if row.get("row_id")
    }
    periods = _iter_96_hour_full_periods(calc_month)
    grouped: dict[str, dict] = {}
    for employee in employees:
        source_id = normalize_shift_employee_id(
            employee.get("source_employee_id") or employee.get("employee_id")
        )
        if not source_id:
            continue
        target = grouped.setdefault(source_id, {
            "employee_id": source_id,
            "name": employee.get("name", ""),
            "department": employee.get("department", ""),
            "position": employee.get("position", ""),
            "daily_rows": [],
        })
        target["daily_rows"].extend(employee.get("attendance_daily_rows") or [])

    rows = []
    for employee in grouped.values():
        for period in periods:
            start = _to_date(period.get("period_start"))
            end = _to_date(period.get("period_end"))
            if not start or not end:
                continue
            period_rows = [
                row for row in employee["daily_rows"]
                if (row_date := _daily_row_date(row)) and start <= row_date <= end
            ]
            day_hours = round(sum(
                _actual_work_hours(row) for row in period_rows if not _is_night_daily_row(row)
            ), 2)
            night_hours = round(sum(
                _actual_work_hours(row) for row in period_rows if _is_night_daily_row(row)
            ), 2)
            if day_hours <= 0 and night_hours <= 0:
                continue

            if night_hours > 0 and day_hours <= 0:
                suggested_policy = HOURLY_RATE_POLICY_NIGHT
                suggestion_reason = "该工资周期仅识别到夜班实际出勤"
                shift_pattern = "全夜班"
            elif day_hours > 0 and night_hours > 0:
                suggested_policy = HOURLY_RATE_POLICY_BY_SHIFT
                suggestion_reason = "该工资周期同时识别到白班和夜班实际出勤"
                shift_pattern = "白夜混合"
            else:
                suggested_policy = HOURLY_RATE_POLICY_BASE
                suggestion_reason = "该工资周期仅识别到白班实际出勤"
                shift_pattern = "全白班"

            row_id = _hourly_rate_policy_row_id(employee["employee_id"], period["period_start"])
            previous = existing_rows.get(row_id, {})
            manual_override = bool(previous.get("manual_override"))
            selected_policy = (
                previous.get("selected_policy")
                if manual_override and previous.get("selected_policy") in HOURLY_RATE_POLICY_CHOICES
                else suggested_policy
            )
            rows.append({
                "row_id": row_id,
                "employee_id": employee["employee_id"],
                "name": employee["name"],
                "department": employee["department"],
                "position": employee["position"],
                "period_start": period["period_start"],
                "period_end": period["period_end"],
                "overlap_start": period["overlap_start"],
                "overlap_end": period["overlap_end"],
                "day_work_hours": day_hours,
                "night_work_hours": night_hours,
                "shift_pattern": shift_pattern,
                "suggested_policy": suggested_policy,
                "selected_policy": selected_policy,
                "suggestion_reason": suggestion_reason,
                "manual_override": manual_override,
                "visible": shift_pattern != "全白班" or manual_override,
            })

    visible_rows = [row for row in rows if row["visible"]]
    return {
        "rows": rows,
        "summary": {
            "total_periods": len(rows),
            "visible_count": len(visible_rows),
            "all_night_count": sum(row["shift_pattern"] == "全夜班" for row in rows),
            "mixed_count": sum(row["shift_pattern"] == "白夜混合" for row in rows),
            "manual_count": sum(bool(row["manual_override"]) for row in rows),
        },
    }


def update_hourly_rate_policy_data(
    policy_data: dict | None,
    *,
    action: str,
    row_id: str = "",
    employee_id: str = "",
    selected_policy: str = "",
) -> dict:
    data = {
        "rows": [dict(row) for row in (policy_data or {}).get("rows", [])],
        "summary": dict((policy_data or {}).get("summary", {})),
    }
    if action == "restore_all":
        for row in data["rows"]:
            row["selected_policy"] = row.get("suggested_policy", HOURLY_RATE_POLICY_BY_SHIFT)
            row["manual_override"] = False
            row["visible"] = row.get("shift_pattern") != "全白班"
    elif action == "add_employee":
        source_id = normalize_shift_employee_id(employee_id)
        matches = [row for row in data["rows"] if row.get("employee_id") == source_id]
        if not matches:
            raise ValueError("未找到该员工有实际出勤的工资周期")
        for row in matches:
            row["visible"] = True
    elif action == "update":
        if selected_policy not in HOURLY_RATE_POLICY_CHOICES:
            raise ValueError("无效的适用时薪选项")
        target = next((row for row in data["rows"] if row.get("row_id") == row_id), None)
        if not target:
            raise ValueError("未找到工资周期时薪记录")
        target["selected_policy"] = selected_policy
        target["manual_override"] = selected_policy != target.get("suggested_policy")
        target["visible"] = True
    else:
        raise ValueError("无效操作")

    visible_rows = [row for row in data["rows"] if row.get("visible")]
    data["summary"] = {
        "total_periods": len(data["rows"]),
        "visible_count": len(visible_rows),
        "all_night_count": sum(row.get("shift_pattern") == "全夜班" for row in data["rows"]),
        "mixed_count": sum(row.get("shift_pattern") == "白夜混合" for row in data["rows"]),
        "manual_count": sum(bool(row.get("manual_override")) for row in data["rows"]),
    }
    return data


def apply_hourly_rate_policies(
    attendance_data: list[dict],
    policy_data: dict | None,
) -> None:
    """Annotate daily rows with the rate delta selected for their payroll period."""
    if not (policy_data or {}).get("rows"):
        return
    policies_by_employee: dict[str, list[tuple[date, date, str]]] = defaultdict(list)
    for policy_row in (policy_data or {}).get("rows", []):
        start = _to_date(policy_row.get("period_start"))
        end = _to_date(policy_row.get("period_end"))
        employee_id = normalize_shift_employee_id(policy_row.get("employee_id"))
        if not employee_id or not start or not end:
            continue
        policies_by_employee[employee_id].append((
            start,
            end,
            str(policy_row.get("selected_policy") or HOURLY_RATE_POLICY_BY_SHIFT),
        ))
    for employee in attendance_data:
        source_id = normalize_shift_employee_id(
            employee.get("source_employee_id") or employee.get("employee_id")
        )
        for row in employee.get("attendance_daily_rows") or []:
            row_date = _daily_row_date(row)
            if not row_date:
                continue
            selected_policy = None
            for start, end, policy in policies_by_employee.get(source_id, []):
                if start and end and start <= row_date <= end:
                    selected_policy = policy
                    break
            if selected_policy == HOURLY_RATE_POLICY_BASE:
                delta = 0
            elif selected_policy == HOURLY_RATE_POLICY_NIGHT:
                delta = 1
            else:
                delta = 1 if _is_night_daily_row(row) else 0
            row["hourly_rate_delta"] = delta
            row["hourly_rate_policy"] = selected_policy or HOURLY_RATE_POLICY_BY_SHIFT


def _is_auto_adjustment_event(event: dict) -> bool:
    text = " ".join(
        str(event.get(key) or "")
        for key in ("subject", "adjustment_type", "adjustment_reason", "note")
    )
    return "转正" in text or "新增绩效占比" in text or "增加绩效占比" in text


def _date_range_overlaps_row(row: dict, start_at: datetime | None, end_at: datetime | None) -> bool:
    if not start_at:
        return False
    leave_end = end_at or start_at
    if leave_end < start_at:
        leave_end = start_at
    interval = _daily_row_interval(row)
    if not interval:
        return False
    row_start, row_end = interval
    return max(row_start, start_at) < min(row_end, leave_end)


def _daily_96_work_hours(
    row: dict,
    *,
    allow_scheduled_floor: bool = False,
    cap_to_scheduled_hours: bool = False,
    include_ot: bool = True,
) -> float:
    regular_hours = _daily_row_number(row, "base_hours", "计薪出勤")
    if include_ot:
        paid_work_hours = (
            regular_hours
            + _daily_row_number(row, "ot15_hours", "OT1.5")
            + _daily_row_number(row, "ot20_hours", "OT2.0")
        )
        raw_work_hours = _daily_row_number(row, "work_hours", "工作时长")
        work_hours = raw_work_hours if raw_work_hours > 0 else paid_work_hours
    else:
        work_hours = regular_hours
    leave_hours = _daily_96_leave_hours(row)
    scheduled_hours = _daily_row_number(row, "scheduled_hours", "应出勤时长")
    if leave_hours > 0 and scheduled_hours > 0 and leave_hours >= scheduled_hours:
        work_hours = 0.0
    elif cap_to_scheduled_hours and scheduled_hours > 0:
        work_hours = min(work_hours, scheduled_hours)

    if not allow_scheduled_floor:
        return work_hours

    scheduled_hours = max(
        scheduled_hours - leave_hours,
        0.0,
    )
    return max(work_hours, scheduled_hours)


def _daily_96_leave_hours(row: dict) -> float:
    return (
        _daily_row_number(row, "sick_hours", "病假")
        + _daily_row_number(row, "sick_settlement_hours", "病假清算")
        + _daily_row_number(row, "annual_hours", "年假")
    )


def _daily_96_holiday_hours(row: dict) -> float:
    return _daily_row_number(row, "holiday_hours", "节假日")


def _iter_96_hour_full_periods(calc_month: str | None) -> list[dict]:
    month_start = _calc_month_start(calc_month)
    if not month_start:
        return []
    _, last_day = monthrange(month_start.year, month_start.month)
    month_end = date(month_start.year, month_start.month, last_day)

    period_start = date(2026, 3, 29)
    while period_start > month_start:
        period_start -= timedelta(days=14)
    while period_start + timedelta(days=13) < month_start:
        period_start += timedelta(days=14)

    periods = []
    while period_start <= month_end:
        period_end = period_start + timedelta(days=13)
        overlap_start = max(period_start, month_start)
        overlap_end = min(period_end, month_end)
        if overlap_start <= overlap_end:
            overlap_days = (overlap_end - overlap_start).days + 1
            periods.append({
                "period": (
                    f"{month_start.year}-{month_start.month:02d} / "
                    f"{overlap_start.month}.{overlap_start.day}-{overlap_end.month}.{overlap_end.day}"
                ),
                "period_start": period_start.isoformat(),
                "period_end": period_end.isoformat(),
                "overlap_start": overlap_start.isoformat(),
                "overlap_end": overlap_end.isoformat(),
                "cap_hours": min(96.0, overlap_days * 8.0),
                "full_period_cap_hours": 96.0,
            })
        period_start += timedelta(days=14)
    return periods


def _date_range(start: date, end: date) -> list[date]:
    days = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def _target_regular_hours_for_workweek(
    rows_by_date: dict[date, list[dict]],
    week_start: date,
    target_start: date,
    target_end: date,
    *,
    cap_to_scheduled_hours: bool = False,
) -> float:
    week_end = week_start + timedelta(days=6)
    prior_work = 0.0
    target_work = 0.0
    for day in _date_range(week_start, week_end):
        day_work = sum(
            _daily_96_work_hours(row, cap_to_scheduled_hours=cap_to_scheduled_hours)
            for row in rows_by_date.get(day, [])
        )
        if day < target_start:
            prior_work += day_work
        elif target_start <= day <= target_end:
            target_work += day_work

    remaining_regular_capacity = max(40.0 - min(prior_work, 40.0), 0.0)
    return min(target_work, remaining_regular_capacity)


def _sum_96_work_hours_for_range(
    rows_by_date: dict[date, list[dict]],
    start: date,
    end: date,
    *,
    cap_to_scheduled_hours: bool = False,
) -> float:
    if start > end:
        return 0.0
    return sum(
        _daily_96_work_hours(row, cap_to_scheduled_hours=cap_to_scheduled_hours)
        for day in _date_range(start, end)
        for row in rows_by_date.get(day, [])
    )


def _sum_96_regular_hours_for_range(
    rows_by_date: dict[date, list[dict]],
    start: date,
    end: date,
) -> float:
    if start > end:
        return 0.0
    return sum(
        _daily_row_number(row, "base_hours", "计薪出勤")
        for day in _date_range(start, end)
        for row in rows_by_date.get(day, [])
    )


def _build_96_hour_special_periods_from_daily_rows(
    daily_rows: list[dict],
    calc_month: str | None,
    first_period_work_mode: str = "actual",
) -> list[dict]:
    month_start = _calc_month_start(calc_month)
    if not month_start:
        return []
    _, last_day = monthrange(month_start.year, month_start.month)
    month_end = date(month_start.year, month_start.month, last_day)

    rows_by_date: dict[date, list[dict]] = defaultdict(list)
    available_dates: set[date] = set()
    for row in daily_rows or []:
        dt = parse_excel_datetime(
            row.get("date")
            or row.get("attendance_date")
            or row.get("考勤日期")
        )
        if not dt:
            continue
        day = dt.date()
        rows_by_date[day].append(row)
        available_dates.add(day)

    if not rows_by_date:
        return []

    result = []
    for index, period in enumerate(_iter_96_hour_full_periods(calc_month)):
        full_start = date.fromisoformat(period["period_start"])
        full_end = date.fromisoformat(period["period_end"])
        overlap_start = date.fromisoformat(period["overlap_start"])
        overlap_end = date.fromisoformat(period["overlap_end"])

        leave_hours = 0.0
        holiday_hours = 0.0
        raw_work_hours = 0.0
        regular_only_work_hours = 0.0
        scheduled_floor_work_hours = 0.0
        for day in _date_range(overlap_start, overlap_end):
            for row in rows_by_date.get(day, []):
                leave_hours += _daily_96_leave_hours(row)
                holiday_hours += _daily_96_holiday_hours(row)
                raw_work_hours += _daily_96_work_hours(row)
                regular_only_work_hours += _daily_96_work_hours(row, include_ot=False)
                scheduled_floor_work_hours += _daily_96_work_hours(row, allow_scheduled_floor=True)

        if full_start < month_start:
            prior_work_hours = _sum_96_regular_hours_for_range(
                rows_by_date,
                full_start,
                overlap_start - timedelta(days=1),
            )
            target_work_hours = _sum_96_work_hours_for_range(
                rows_by_date,
                overlap_start,
                overlap_end,
            )
            remaining_regular_capacity = max(
                period["full_period_cap_hours"] - min(prior_work_hours, period["full_period_cap_hours"]),
                0.0,
            )
            regular_hours = min(target_work_hours, remaining_regular_capacity)
            mode = "跨月首段REG-双周96封顶"
        elif full_end > month_end:
            regular_hours = regular_only_work_hours
            mode = "跨月尾段REG（不含OT）"
        else:
            full_period_regular_capacity = max(period["full_period_cap_hours"] - leave_hours, 0.0)
            regular_hours = min(
                max(raw_work_hours, scheduled_floor_work_hours),
                full_period_regular_capacity,
            )
            mode = "完整双周REG封顶"

        missing_context_dates = [
            day.isoformat()
            for day in _date_range(full_start, min(overlap_start - timedelta(days=1), full_end))
            if day not in available_dates
        ]

        has_hours = any(value > 0 for value in (regular_hours, leave_hours, holiday_hours, raw_work_hours))
        if has_hours:
            included_hours = regular_hours + leave_hours + holiday_hours
            result.append({
                "period": period["period"],
                "period_start": period["overlap_start"],
                "period_end": period["overlap_end"],
                "full_period_start": period["period_start"],
                "full_period_end": period["period_end"],
                "cap_hours": period["cap_hours"],
                "mode": mode,
                "first_period_work_mode": first_period_work_mode if full_start < month_start else "",
                "regular_hours": round(regular_hours, 2),
                "raw_work_hours": round(raw_work_hours, 2),
                "regular_only_work_hours": round(regular_only_work_hours, 2),
                "scheduled_floor_work_hours": round(scheduled_floor_work_hours, 2),
                "leave_hours": round(leave_hours, 2),
                "holiday_hours": round(holiday_hours, 2),
                "included_hours": round(included_hours, 2),
                "missing_context_dates": missing_context_dates,
                "context_complete": not missing_context_dates,
                "sequence": index + 1,
            })
    return result


def _build_96_hour_periods_from_daily_rows(
    daily_rows: list[dict],
    calc_month: str | None,
) -> list[dict]:
    period_rows = []
    for period in _iter_96_hour_periods(calc_month):
        period_rows.append((
            date.fromisoformat(period["period_start"]),
            date.fromisoformat(period["period_end"]),
            {
                "period": period["period"],
                "period_start": period["period_start"],
                "period_end": period["period_end"],
                "cap_hours": period["cap_hours"],
                "base_hours": 0.0,
                "ot15_hours": 0.0,
                "ot20_hours": 0.0,
                "sick_hours": 0.0,
                "sick_settlement_hours": 0.0,
                "annual_hours": 0.0,
                "holiday_hours": 0.0,
            },
        ))
    if not period_rows:
        return []

    for row in daily_rows or []:
        dt = parse_excel_datetime(
            row.get("date")
            or row.get("attendance_date")
            or row.get("考勤日期")
        )
        if not dt:
            continue
        day = dt.date()
        for start, end, period in period_rows:
            if start <= day <= end:
                period["base_hours"] += _daily_row_number(row, "base_hours", "计薪出勤")
                period["ot15_hours"] += _daily_row_number(row, "ot15_hours", "OT1.5")
                period["ot20_hours"] += _daily_row_number(row, "ot20_hours", "OT2.0")
                period["sick_hours"] += _daily_row_number(row, "sick_hours", "病假")
                period["sick_settlement_hours"] += _daily_row_number(row, "sick_settlement_hours", "病假清算")
                period["annual_hours"] += _daily_row_number(row, "annual_hours", "年假")
                period["holiday_hours"] += _daily_row_number(row, "holiday_hours", "节假日")
                break

    result = []
    for _, _, period in period_rows:
        has_hours = any(
            safe_float(period.get(key)) > 0
            for key in (
                "base_hours",
                "ot15_hours",
                "ot20_hours",
                "sick_hours",
                "sick_settlement_hours",
                "annual_hours",
                "holiday_hours",
            )
        )
        if has_hours:
            result.append({
                key: (round(value, 2) if isinstance(value, float) else value)
                for key, value in period.items()
            })
    return result


def normalize_shift_employee_id(employee_id: str) -> str:
    text = str(employee_id or "").strip()
    return text[:-2] if text.endswith("-1") else text


def has_shift_hours(shift_hours: dict) -> bool:
    return any(safe_float(value) > 0 for value in (shift_hours or {}).values())


def _find_column(headers, names: list[str], fallback: int | None = None) -> int | None:
    normalized = {
        str(header).strip(): index
        for index, header in enumerate(headers)
        if header is not None and str(header).strip()
    }
    for name in names:
        if name in normalized:
            return normalized[name]
    return fallback


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return openpyxl.utils.datetime.from_excel(value).date()
        except Exception:
            return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%Y.%m.%d"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _calc_month_start(calc_month: str | None) -> date | None:
    if not calc_month:
        return None
    if isinstance(calc_month, datetime):
        return date(calc_month.year, calc_month.month, 1)
    if isinstance(calc_month, date):
        return date(calc_month.year, calc_month.month, 1)
    text = str(calc_month).strip()
    for fmt in ("%Y-%m", "%Y/%m", "%Y%m"):
        try:
            parsed = datetime.strptime(text, fmt)
            return date(parsed.year, parsed.month, 1)
        except ValueError:
            continue
    parsed = parse_excel_datetime(calc_month)
    if parsed:
        return date(parsed.year, parsed.month, 1)
    return None


def _normalize_calc_month(calc_month, default: str | None = None) -> str:
    month_start = _calc_month_start(calc_month)
    if not month_start and default:
        month_start = _calc_month_start(default)
    return f"{month_start.year}-{month_start.month:02d}" if month_start else ""


def _same_month(value: date | None, month_start: date | None) -> bool:
    return bool(value and month_start and value.year == month_start.year and value.month == month_start.month)


DISTRICT_MANAGER_IDS = {"zt15638"}
FUNCTIONAL_DEPARTMENT_KEYWORDS = (
    "FBU HRBP Dept.",
    "渠道管理部",
    "新泽西区渠道部",
    "新泽西区行政部",
)
DEFAULT_COEFFICIENT_POSITION_KEYWORDS = (
    "Management Trainee",
)
DEFAULT_COEFFICIENT_POSITIONS = {
    "management trainee 管培生",
}
DEFAULT_COEFFICIENT_VALUE = 1.0


def classify_job_type(employee_id: str, department: str) -> str:
    """按FBU美洲业务口径区分仓库端、职能端和区长。"""
    normalized_id = str(employee_id or "").strip().lower()
    if normalized_id in DISTRICT_MANAGER_IDS:
        return "district_manager"
    if any(keyword in (department or "") for keyword in FUNCTIONAL_DEPARTMENT_KEYWORDS):
        return "functional"
    return "warehouse"


def default_coefficient_reason(
    position: str,
    personnel_status: str,
    confirmation_date: date | None = None,
    calc_month_start: date | None = None,
    has_performance_result: bool = True,
) -> str:
    """识别无绩效流程但绩效系数默认按1的业务场景。"""
    normalized_status = str(personnel_status or "").strip()
    normalized_position = str(position or "").strip().lower()
    if normalized_position in DEFAULT_COEFFICIENT_POSITIONS or any(
        keyword.lower() in normalized_position
        for keyword in DEFAULT_COEFFICIENT_POSITION_KEYWORDS
    ):
        if normalized_status == "离职" and not has_performance_result:
            return ""
        return "岗位为管培生，绩效系数系统固定按1"
    if _same_month(confirmation_date, calc_month_start):
        if confirmation_date.day > 15:
            return "月中15号后转正，绩效系数系统固定按1"
        return ""
    if normalized_status == "试用":
        return "人员状态为试用，绩效系数系统固定按1"
    return ""


class FBUPerformanceParser:
    """FBU绩效数据解析器"""

    def __init__(self):
        self.engine = FBUPerformanceEngine()
        self.attendance_processor = AttendanceProcessor()
        self.salary_processor = SalaryProcessor()
        self.employee_roster = {}  # 花名册数据 {emp_id: {name, department, ...}}

    @staticmethod
    def load_excel(
        filepath: str,
        password: Optional[str] = None,
        *,
        read_only: bool = False,
    ) -> openpyxl.Workbook:
        """加载Excel文件（支持密码保护）"""
        if password:
            with open(filepath, "rb") as f:
                ms_file = msoffcrypto.OfficeFile(f)
                ms_file.load_key(password=password)
                decrypted = io.BytesIO()
                ms_file.decrypt(decrypted)
                decrypted.seek(0)
                return openpyxl.load_workbook(decrypted, data_only=True, read_only=read_only)
        else:
            return openpyxl.load_workbook(filepath, data_only=True, read_only=read_only)

    def load_roster(self, filepath: str) -> dict:
        """
        加载花名册数据

        Args:
            filepath: 花名册文件路径

        Returns:
            员工信息字典 {emp_id: {name, department, job_type, ...}}
        """
        path = Path(filepath)
        if path.suffix.lower() == ".xls":
            book = xlrd.open_workbook(filepath)
            sheet = book.sheet_by_index(0)
            headers = sheet.row_values(0) if sheet.nrows else []
            rows = (
                sheet.row_values(row_idx)
                for row_idx in range(1, sheet.nrows)
            )
        else:
            wb = self.load_excel(filepath)
            ws = wb[wb.sheetnames[0]]
            headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            rows = ws.iter_rows(min_row=2, values_only=True)

        # 找到关键列的索引
        col_map = {
            '姓名': _find_column(headers, ['姓名'], 0),
            '工号': _find_column(headers, ['工号'], 3),
            '人员状态': _find_column(headers, ['人员状态', '员工状态', '状态'], 7),
            '入职日期': _find_column(headers, ['入职日期'], 15),
            '实际转正日期': _find_column(headers, ['实际转正日期', '转正日期', '确认日期'], 17),
            '二级部门': _find_column(headers, ['二级部门'], 19),
            '三级部门': _find_column(headers, ['三级部门'], 20),
            '四级部门': _find_column(headers, ['四级部门'], 21),
            '五级部门': _find_column(headers, ['五级部门'], 22),
            '六级部门': _find_column(headers, ['六级部门'], 23),
            '七级部门': _find_column(headers, ['七级部门'], 24),
            '八级部门': _find_column(headers, ['八级部门'], 25),
            '职位': _find_column(headers, ['职位', '岗位', '岗位名称', '职务', 'Position', 'Job Title'], 30),
            '划分区域': _find_column(headers, ['划分区域'], 89),
            '离职日期': _find_column(headers, ['离职日期', '最后工作日'], 96),
            '领色': _find_column(headers, ['领色'], 107),
        }

        roster = {}
        for row in rows:
            if not row or _cell(row, col_map['工号']) is None:
                continue

            emp_id = str(_cell(row, col_map['工号'])).strip()
            name = str(_cell(row, col_map['姓名'])).strip() if _cell(row, col_map['姓名']) else ''
            personnel_status = str(_cell(row, col_map['人员状态'])).strip() if _cell(row, col_map['人员状态']) else ''
            hire_date = _to_date(_cell(row, col_map['入职日期']))
            confirmation_date = _to_date(_cell(row, col_map['实际转正日期']))
            resignation_date = _to_date(_cell(row, col_map['离职日期']))

            # 构建部门全称：二级-三级-四级-五级-六级-七级-八级
            dept_parts = []
            for level in ['二级部门', '三级部门', '四级部门', '五级部门', '六级部门', '七级部门', '八级部门']:
                val = _cell(row, col_map[level])
                if val and str(val).strip():
                    dept_parts.append(str(val).strip())
            department_full = '-'.join(dept_parts) if dept_parts else ''

            # 划分区域
            area = str(_cell(row, col_map['划分区域'])).strip() if _cell(row, col_map['划分区域']) else ''
            position = str(_cell(row, col_map['职位'])).strip() if _cell(row, col_map['职位']) else ''

            job_type = classify_job_type(emp_id, department_full)

            roster[emp_id] = {
                'name': name,
                'department': department_full,
                'area': area,
                'personnel_status': personnel_status,
                'hire_date': hire_date,
                'confirmation_date': confirmation_date,
                'resignation_date': resignation_date,
                'position': position,
                'job_type': job_type,
            }

        self.employee_roster = roster
        return roster

    def get_employee_info(self, emp_id: str) -> dict:
        """获取员工信息"""
        return self.employee_roster.get(emp_id, {
            'name': '',
            'department': '',
            'area': '',
            'personnel_status': '',
            'hire_date': None,
            'confirmation_date': None,
            'resignation_date': None,
            'position': '',
            'job_type': 'warehouse',
        })

    def parse_attendance(self, filepath: str, target_month: int) -> dict:
        """解析考勤数据"""
        wb = self.load_excel(filepath, read_only=True)
        ws = wb['sheet1']
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

        # 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _cell(row, 0) is not None:
                rows.append(row)
        wb.close()

        # 处理考勤数据
        return self.attendance_processor.process(rows, target_month, headers=headers)

    def parse_salary(self, filepath: str) -> dict:
        """解析薪资档案"""
        wb = self.load_excel(filepath)
        ws = wb[wb.sheetnames[0]]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

        # 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _cell(row, 0) is not None:
                rows.append(row)

        return self.salary_processor.load(rows, headers=headers)

    def parse_performance(self, filepath: str) -> dict:
        """解析绩效报表"""
        performance_data = {}
        preview = self.parse_performance_preview(filepath)
        for employee in preview.get("employees", []):
            emp_id = employee.get("employee_id")
            if not emp_id:
                continue
            performance_data[emp_id] = {
                'score': employee.get("score"),
                'level': employee.get("level"),
                'coefficient': employee.get("coefficient"),
            }

        return performance_data

    def build_employees(
        self,
        attendance_data: dict,
        salary_data: dict,
        performance_data: dict,
        employee_info: dict = None,
        adjustment_data: dict = None,
        calc_month: str | None = None,
        supplemental_leave_data: dict | None = None,
        base_override_data: dict | None = None,
    ) -> list[EmployeeData]:
        """构建员工数据"""
        if employee_info is None:
            employee_info = {}
        if adjustment_data is None:
            adjustment_data = {}

        employees = []
        month_start = _calc_month_start(calc_month)
        ninety_six_hour_ids = {
            normalize_shift_employee_id(row.get("employee_id"))
            for row in (base_override_data or {}).get("employees", [])
            if row.get("include_in_calculation")
            and "96" in str(row.get("rule_type") or "")
        }

        for emp_id, hours in attendance_data.items():
            source_emp_id = normalize_shift_employee_id(emp_id)
            # 获取薪资信息
            salary_info = salary_data.get(source_emp_id, {})
            exceptions = []
            if not salary_info:
                exceptions.append("未匹配薪资档案")
            hourly_rate = salary_info.get('hourly_rate', 0)
            ratio = salary_info.get('ratio', 0)
            fixed_performance_base = salary_info.get('fixed_performance_base', 0)

            # 获取绩效信息
            perf_info = performance_data.get(source_emp_id, {})
            has_performance_result = bool(perf_info)
            if not perf_info:
                exceptions.append("未匹配绩效报表")

            # 获取员工信息
            info = employee_info.get(source_emp_id, {})
            name = info.get('name', '') or hours.get('name', '')
            department = info.get('department', '')
            area = info.get('area', '')
            personnel_status = info.get('personnel_status', '')
            hire_date = info.get('hire_date')
            confirmation_date = info.get('confirmation_date')
            resignation_date = info.get('resignation_date')
            position = info.get('position', '')

            # 判断岗位类型（简化：有得分用仓库端，否则职能端）
            score = perf_info.get('score')
            level = perf_info.get('level')
            uploaded_coefficient = perf_info.get('coefficient')
            job_type = info.get('job_type', 'warehouse') if info else 'warehouse'
            coefficient_override_reason = default_coefficient_reason(
                position,
                personnel_status,
                confirmation_date=confirmation_date,
                calc_month_start=month_start,
                has_performance_result=has_performance_result,
            )
            if coefficient_override_reason:
                uploaded_coefficient = DEFAULT_COEFFICIENT_VALUE
                if "未匹配绩效报表" in exceptions:
                    exceptions.remove("未匹配绩效报表")
                exceptions.append(coefficient_override_reason)
            adjustment_segments = self._build_calculation_segments(
                adjustment_data.get(source_emp_id, []),
                ratio=ratio,
                coefficient=uploaded_coefficient,
            )
            if (
                adjustment_segments
                and all(segment.performance_coefficient > 0 for segment in adjustment_segments)
                and "未匹配绩效报表" in exceptions
            ):
                exceptions.remove("未匹配绩效报表")

            def make_employee(
                employee_id: str,
                shift_hours: dict,
                is_night_shift: bool,
                shift_hourly_rate: float,
                attendance_daily_rows: list[dict] | None = None,
            ) -> EmployeeData:
                return EmployeeData(
                    employee_id=employee_id,
                    source_employee_id=source_emp_id,
                    name=name,
                    department=department,
                    area=area,
                    position=position,
                    personnel_status=personnel_status,
                    hire_date=hire_date,
                    confirmation_date=confirmation_date,
                    resignation_date=resignation_date,
                    hourly_rate=shift_hourly_rate,
                    performance_ratio=ratio,
                    performance_score=score,
                    performance_level=level,
                    uploaded_coefficient=uploaded_coefficient,
                    coefficient_override_reason=coefficient_override_reason,
                    job_type=job_type,
                    fixed_performance_base=fixed_performance_base,
                    base_hours=shift_hours['计薪出勤'],
                    ot15_hours=shift_hours['OT1.5'],
                    ot20_hours=shift_hours['OT2.0'],
                    sick_hours=shift_hours['病假'],
                    sick_settlement_hours=shift_hours.get('病假清算', 0),
                    annual_hours=shift_hours['年假'],
                    holiday_hours=shift_hours['节假日'],
                    is_night_shift=is_night_shift,
                    attendance_daily_rows=list(attendance_daily_rows or []),
                    calculation_segments=list(adjustment_segments),
                    exceptions=list(exceptions),
                )

            if adjustment_segments:
                combined_hours = {
                    key: hours['白班'].get(key, 0) + hours['夜班'].get(key, 0)
                    for key in ['计薪出勤', 'OT1.5', 'OT2.0', '病假', '病假清算', '年假', '节假日']
                }
                employees.append(make_employee(emp_id, combined_hours, False, hourly_rate, hours.get('daily_rows', [])))
                continue

            annotated_rows = [
                row for row in hours.get("daily_rows", [])
                if row.get("hourly_rate_delta") is not None
                and (
                    not month_start
                    or ((row_date := _daily_row_date(row)) and _same_month(row_date, month_start))
                )
            ]
            if annotated_rows and source_emp_id not in ninety_six_hour_ids:
                rows_by_delta: dict[int, list[dict]] = defaultdict(list)
                for row in annotated_rows:
                    delta = 1 if safe_float(row.get("hourly_rate_delta")) >= 0.5 else 0
                    rows_by_delta[delta].append(row)

                def summarize_daily_rows(daily_rows: list[dict]) -> dict:
                    return {
                        "计薪出勤": sum(_daily_row_number(row, "base_hours", "计薪出勤") for row in daily_rows),
                        "OT1.5": sum(_daily_row_number(row, "ot15_hours", "OT1.5") for row in daily_rows),
                        "OT2.0": sum(_daily_row_number(row, "ot20_hours", "OT2.0") for row in daily_rows),
                        "病假": sum(_daily_row_number(row, "sick_hours", "病假") for row in daily_rows),
                        "病假清算": sum(
                            _daily_row_number(row, "sick_settlement_hours", "病假清算")
                            for row in daily_rows
                        ),
                        "年假": sum(_daily_row_number(row, "annual_hours", "年假") for row in daily_rows),
                        "节假日": sum(_daily_row_number(row, "holiday_hours", "节假日") for row in daily_rows),
                    }

                split_rates = len(rows_by_delta) > 1
                for delta in sorted(rows_by_delta):
                    rate_rows = rows_by_delta[delta]
                    rate_hours = summarize_daily_rows(rate_rows)
                    if not has_shift_hours(rate_hours):
                        continue
                    rate_employee_id = (
                        f"{source_emp_id}-1"
                        if split_rates and delta == 0
                        else source_emp_id
                    )
                    employees.append(make_employee(
                        rate_employee_id,
                        rate_hours,
                        all(_is_night_daily_row(row) for row in rate_rows),
                        hourly_rate + delta,
                        rate_rows,
                    ))
                continue

            has_day_hours = has_shift_hours(hours['白班'])
            has_night_hours = hours['has_night_shift'] and has_shift_hours(hours['夜班'])
            split_mixed_shift = has_day_hours and has_night_hours

            # 处理白班
            if has_day_hours:
                day_employee_id = f"{emp_id}-1" if split_mixed_shift else emp_id
                day_shift_hours = hours['白班']
                day_daily_rows = hours.get('白班_daily_rows', [])
                if has_shift_hours(day_shift_hours):
                    employees.append(make_employee(
                        day_employee_id,
                        day_shift_hours,
                        False,
                        hourly_rate,
                        day_daily_rows,
                    ))

            # 处理夜班
            if has_night_hours:
                night_shift_hours = hours['夜班']
                night_daily_rows = hours.get('夜班_daily_rows', [])
                if has_shift_hours(night_shift_hours):
                    employees.append(make_employee(
                        emp_id,
                        night_shift_hours,
                        True,
                        hourly_rate + 1,
                        night_daily_rows,
                    ))

        self.apply_supplemental_leave_to_employees(employees, supplemental_leave_data, calc_month)
        self.apply_base_overrides_to_employees(employees, base_override_data, calc_month)
        return employees

    @staticmethod
    def _base_override_path(rule_type: str, fixed_base: float = 0) -> str:
        if "96" in str(rule_type or "") and safe_float(fixed_base) <= 0:
            return NINETY_SIX_HOUR_AUTO_BASE_PATH
        return (
            NINETY_SIX_HOUR_FIXED_BASE_PATH
            if "96" in str(rule_type or "")
            else FIXED_BASE_OVERRIDE_PATH
        )

    @staticmethod
    def _summarize_base_overrides(rows: list[dict]) -> dict:
        active_rows = [row for row in rows if row.get("include_in_calculation")]
        return {
            "total_rows": len(rows),
            "active_count": len(active_rows),
            "excluded_count": len(rows) - len(active_rows),
            "active_fixed_base": round(sum(safe_float(row.get("fixed_performance_base")) for row in active_rows), 2),
        }

    @staticmethod
    def _find_nearest_left(headers: tuple | list, start_index: int, labels: list[str]) -> int | None:
        for index in range(start_index, -1, -1):
            text = str(headers[index] or "").strip()
            if any(label in text for label in labels):
                return index
        return None

    def _parse_summary_base_overrides(
        self,
        sheet_name: str,
        rows: list[tuple],
        header_row_index: int,
        headers: tuple | list,
        calc_month: str,
    ) -> list[dict]:
        """从线下“基数汇总”的差异备注中抽取96工时制/固定基数覆盖。"""
        note_col = _find_column(headers, ["差异备注"])
        attendance_col = _find_column(headers, ["考勤日报"])
        if note_col is None or attendance_col is None:
            return []

        fixed_base_col = self._find_nearest_left(headers, attendance_col - 1, ["合计所得"])
        employee_id_col = self._find_nearest_left(headers, fixed_base_col - 1, ["工号"]) if fixed_base_col is not None else None
        name_col = self._find_nearest_left(headers, fixed_base_col - 1, ["中文名字", "姓名"]) if fixed_base_col is not None else None
        if employee_id_col is None or fixed_base_col is None:
            return []

        parsed_rows: list[dict] = []
        for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            employee_id = str(_cell(row, employee_id_col) or "").strip()
            if not employee_id or not employee_id.lower().startswith("zt"):
                continue

            note = str(_cell(row, note_col, "") or "").strip()
            if not note:
                continue

            if "96" in note:
                rule_type = "96工时制"
            elif "固定" in note and "基数" in note:
                rule_type = "线下固定基数覆盖"
            else:
                continue

            fixed_base = _to_float(_cell(row, fixed_base_col), 0) or 0.0
            if fixed_base <= 0:
                continue

            emp_info = self.get_employee_info(normalize_shift_employee_id(employee_id))
            parsed_rows.append({
                "row_id": f"{sheet_name}:{row_number}",
                "source_sheet": sheet_name,
                "source_row": row_number,
                "employee_id": employee_id,
                "source_employee_id": normalize_shift_employee_id(employee_id),
                "name": emp_info["name"] or str(_cell(row, name_col, "") or "").strip(),
                "department": emp_info["department"],
                "area": emp_info["area"],
                "rule_type": rule_type,
                "fixed_performance_base": round(fixed_base, 2),
                "allocation_month": calc_month,
                "status": "启用",
                "include_in_calculation": True,
                "calculation_path": self._base_override_path(rule_type, fixed_base),
                "note": f"线下基数汇总差异备注: {note}",
            })

        return parsed_rows

    def _parse_special_work_hour_summary(
        self,
        sheet_name: str,
        rows: list[tuple],
        calc_month: str,
    ) -> list[dict]:
        """从“特殊工时汇总”首块抽取96工时制最终小时口径。"""
        if "特殊工时汇总" not in sheet_name:
            return []

        header_row_index = None
        headers = None
        for index, row in enumerate(rows[:10]):
            if (
                _find_column(row, ["员工工号", "工号", "employee_id"]) is not None
                and _find_column(row, ["合计时长"]) is not None
            ):
                header_row_index = index
                headers = row
                break
        if header_row_index is None or headers is None:
            return []

        employee_id_col = _find_column(headers, ["员工工号", "工号", "employee_id"])
        name_col = _find_column(headers, ["姓名", "员工姓名", "name"])
        total_hours_col = _find_column(headers, ["合计时长"])
        hourly_rate_col = _find_column(headers, ["时薪"])
        amount_col = _find_column(headers, ["时薪*时长"])
        if employee_id_col is None or total_hours_col is None:
            return []

        parsed_rows: list[dict] = []
        for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            employee_id = str(_cell(row, employee_id_col) or "").strip()
            if not employee_id:
                if parsed_rows:
                    break
                continue
            if not employee_id.lower().startswith("zt"):
                continue

            total_hours = safe_float(_cell(row, total_hours_col))
            if total_hours <= 0:
                continue

            source_id = normalize_shift_employee_id(employee_id)
            emp_info = self.get_employee_info(source_id)
            hourly_rate = safe_float(_cell(row, hourly_rate_col)) if hourly_rate_col is not None else 0.0
            special_base = safe_float(_cell(row, amount_col)) if amount_col is not None else 0.0
            parsed_rows.append({
                "row_id": f"{sheet_name}:{row_number}",
                "source_sheet": sheet_name,
                "source_row": row_number,
                "employee_id": employee_id,
                "source_employee_id": source_id,
                "name": emp_info["name"] or str(_cell(row, name_col, "") or "").strip(),
                "department": emp_info["department"],
                "area": emp_info["area"],
                "rule_type": "96工时制",
                "fixed_performance_base": 0.0,
                "allocation_month": calc_month,
                "status": "启用",
                "include_in_calculation": True,
                "calculation_path": NINETY_SIX_HOUR_AUTO_BASE_PATH,
                "special_total_hours": round(total_hours, 2),
                "special_hourly_rate": hourly_rate,
                "special_performance_base": special_base,
                "note": f"特殊工时汇总: 合计时长={total_hours:g}",
            })

        return parsed_rows

    def _parse_offline_base_summary(
        self,
        sheet_name: str,
        rows: list[tuple],
        header_row_index: int,
        headers: tuple | list,
    ) -> list[dict]:
        """从线下“基数汇总”抽取全量员工基数，仅用于差异诊断和安全建议。"""
        if "基数汇总" not in sheet_name:
            return []

        def first_exact(*names: str) -> int | None:
            wanted = {str(name).strip() for name in names}
            return next(
                (
                    index
                    for index, header in enumerate(headers)
                    if str(header or "").strip() in wanted
                ),
                None,
            )

        def first_contains(pattern: str) -> int | None:
            return next(
                (
                    index
                    for index, header in enumerate(headers)
                    if pattern in str(header or "").strip()
                ),
                None,
            )

        # “基数汇总”横向包含多块透视/验算区，表头重复。这里必须锚定第一块
        # 员工绩效基数区，不能用全局 _find_column，否则会被后面的“考勤日报”块覆盖。
        employee_id_col = first_exact("工号", "员工工号", "employee_id")
        name_col = first_exact("中文名字", "姓名", "员工姓名", "name")
        amount_col = first_contains("合计所得")
        if employee_id_col is None or amount_col is None:
            return []

        by_employee: dict[str, dict] = {}
        for row in rows[header_row_index + 1:]:
            employee_id = str(_cell(row, employee_id_col) or "").strip()
            if not employee_id or not employee_id.lower().startswith("zt"):
                continue
            amount = safe_float(_cell(row, amount_col))
            if amount == 0:
                continue

            source_id = normalize_shift_employee_id(employee_id)
            bucket = by_employee.setdefault(
                source_id,
                {
                    "employee_id": source_id,
                    "name": str(_cell(row, name_col, "") or "").strip() if name_col is not None else "",
                    "offline_performance_base": 0.0,
                    "offline_rows": 0,
                },
            )
            bucket["offline_performance_base"] = round(bucket["offline_performance_base"] + amount, 2)
            bucket["offline_rows"] += 1

        return list(by_employee.values())

    def parse_base_overrides_preview(self, filepath: str, calc_month: str) -> dict:
        """解析线下固定基数/96工时制覆盖表，生成可审计的覆盖清单。"""
        wb = self.load_excel(filepath)
        calc_month_text = _normalize_calc_month(calc_month)
        special_rows: list[dict] = []
        for special_sheet_name in wb.sheetnames:
            if "特殊工时汇总" not in special_sheet_name:
                continue
            special_ws = wb[special_sheet_name]
            special_rows.extend(
                self._parse_special_work_hour_summary(
                    special_sheet_name,
                    list(special_ws.iter_rows(values_only=True)),
                    calc_month_text or calc_month,
                )
            )

        sheet_name = next(
            (
                name for name in wb.sheetnames
                if any(keyword in name for keyword in ("固定基数", "96", "覆盖", "基数汇总", "特殊工时汇总"))
            ),
            wb.sheetnames[0],
        )
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("固定基数覆盖表为空")

        header_row_index = 0
        headers = rows[0]
        for index, row in enumerate(rows[:20]):
            if _find_column(row, ["工号", "员工工号", "employee_id"]) is not None:
                header_row_index = index
                headers = row
                break

        offline_bases = self._parse_offline_base_summary(sheet_name, rows, header_row_index, headers)
        summary_rows = self._parse_summary_base_overrides(sheet_name, rows, header_row_index, headers, calc_month_text or calc_month)
        if summary_rows or special_rows:
            merged: dict[str, dict] = {}
            for row in summary_rows:
                merged[row["source_employee_id"]] = row
            for row in special_rows:
                merged[row["source_employee_id"]] = row
            rows_for_preview = list(merged.values())
            return {
                "employees": rows_for_preview,
                "offline_bases": offline_bases,
                "summary": self._summarize_base_overrides(rows_for_preview),
            }

        columns = {
            "employee_id": _find_column(headers, ["工号", "员工工号", "employee_id"]),
            "name": _find_column(headers, ["姓名", "员工姓名", "name"]),
            "rule_type": _find_column(headers, ["规则类型", "覆盖类型", "核算路径", "类型", "rule_type"]),
            "fixed_base": _find_column(headers, ["固定绩效基数", "绩效基数", "固定基数", "fixed_performance_base"]),
            "month": _find_column(headers, ["归属月份", "核算月份", "月份", "allocation_month"]),
            "status": _find_column(headers, ["状态", "启用状态", "生效状态", "status"]),
            "note": _find_column(headers, ["备注", "原因", "说明", "note"]),
        }
        if columns["employee_id"] is None or columns["fixed_base"] is None:
            raise ValueError("固定基数覆盖表至少需要“工号”和“固定绩效基数”列")

        parsed_rows: list[dict] = []
        for row_number, row in enumerate(rows[header_row_index + 1:], start=header_row_index + 2):
            employee_id = str(_cell(row, columns["employee_id"]) or "").strip()
            if not employee_id or not employee_id.lower().startswith("zt"):
                continue
            rule_type = str(_cell(row, columns["rule_type"], "线下固定基数覆盖") or "线下固定基数覆盖").strip()
            fixed_base = _to_float(_cell(row, columns["fixed_base"]), 0) or 0.0
            is_96_hour_rule = "96" in rule_type
            if fixed_base <= 0 and not is_96_hour_rule:
                continue
            allocation_month = _normalize_calc_month(_cell(row, columns["month"], calc_month_text), calc_month_text)
            status = str(_cell(row, columns["status"], "") or "").strip()
            note = str(_cell(row, columns["note"], "") or "").strip()
            disabled = status in {"停用", "禁用", "不启用", "排除", "不计入", "否", "N", "n", "false", "FALSE"}
            include_in_calculation = allocation_month == calc_month_text and not disabled
            emp_info = self.get_employee_info(normalize_shift_employee_id(employee_id))

            parsed_rows.append({
                "row_id": f"{sheet_name}:{row_number}",
                "source_sheet": sheet_name,
                "source_row": row_number,
                "employee_id": employee_id,
                "source_employee_id": normalize_shift_employee_id(employee_id),
                "name": emp_info["name"] or str(_cell(row, columns["name"], "") or "").strip(),
                "department": emp_info["department"],
                "area": emp_info["area"],
                "rule_type": rule_type,
                "fixed_performance_base": round(fixed_base, 2),
                "allocation_month": allocation_month,
                "status": status or "启用",
                "include_in_calculation": include_in_calculation,
                "calculation_path": self._base_override_path(rule_type, fixed_base),
                "note": note,
            })

        return {
            "employees": parsed_rows,
            "offline_bases": offline_bases,
            "summary": self._summarize_base_overrides(parsed_rows),
        }

    @classmethod
    def apply_base_overrides_to_employees(
        cls,
        employees: list[EmployeeData],
        base_override_data: dict | None,
        calc_month: str | None,
    ) -> None:
        """把固定基数覆盖表应用到员工拆行；精确工号优先，原工号回退到对应行。"""
        if not base_override_data:
            return

        calc_month_text = _normalize_calc_month(calc_month)
        active_rows = [
            row for row in base_override_data.get("employees", [])
            if row.get("include_in_calculation")
            and (safe_float(row.get("fixed_performance_base")) > 0 or "96" in str(row.get("rule_type") or ""))
            and (
                not calc_month_text
                or not row.get("allocation_month")
                or _normalize_calc_month(row.get("allocation_month"), calc_month_text) == calc_month_text
            )
        ]
        if not active_rows:
            return

        by_source: dict[str, list[EmployeeData]] = defaultdict(list)
        for employee in employees:
            by_source[normalize_shift_employee_id(employee.source_employee_id)].append(employee)

        for row in active_rows:
            override_employee_id = str(row.get("employee_id") or "").strip()
            source_id = normalize_shift_employee_id(override_employee_id)
            source_employees = by_source.get(source_id)
            if not source_employees:
                continue

            target = next(
                (employee for employee in source_employees if employee.employee_id == override_employee_id),
                None,
            )
            if target is None and override_employee_id == source_id:
                target = next((employee for employee in source_employees if employee.employee_id == source_id), None)
            if target is None:
                target = next((employee for employee in source_employees if not employee.is_night_shift), source_employees[0])

            rule_type = str(row.get("rule_type") or "").strip()
            fixed_base = safe_float(row.get("fixed_performance_base"))
            if "96" in rule_type:
                target.work_hour_rule = "96工时制"
                target.work_hour_rule_cap = _monthly_96_hour_cap(calc_month)
                target.work_hour_rule_include_holiday_in_cap = source_id == "zt12979"
                first_period_work_mode = (
                    "actual"
                    if 0 < safe_float(target.performance_ratio) < 0.4
                    else "scheduled_cap"
                )
                explicit_special_total_hours = safe_float(row.get("special_total_hours"))
                rule_daily_rows = [
                    daily_row
                    for source_employee in source_employees
                    for daily_row in source_employee.attendance_daily_rows
                ]
                special_periods = _build_96_hour_special_periods_from_daily_rows(
                    rule_daily_rows,
                    calc_month,
                    first_period_work_mode=first_period_work_mode,
                )
                if explicit_special_total_hours > 0:
                    target.work_hour_rule_special_total_hours = explicit_special_total_hours
                    target.work_hour_rule_periods = special_periods or _build_96_hour_periods_from_daily_rows(
                        rule_daily_rows,
                        calc_month,
                    )
                elif special_periods:
                    target.work_hour_rule_periods = special_periods
                    target.work_hour_rule_special_total_hours = round(
                        sum(safe_float(period.get("included_hours")) for period in special_periods),
                        2,
                    )
                else:
                    target.work_hour_rule_special_total_hours = 0.0
                    target.work_hour_rule_periods = _build_96_hour_periods_from_daily_rows(
                        rule_daily_rows,
                        calc_month,
                    )

                if target.work_hour_rule_special_total_hours > 0:
                    target.exceptions.append("96工时制自动基数: 按考勤日报明细自动计算特殊合计时长，并用两位小数时薪计算")
                    if any(not period.get("context_complete", True) for period in target.work_hour_rule_periods):
                        target.exceptions.append("96工时制自动基数: 考勤日报缺少跨工资周期日期，结果可能与线下完整周期口径存在差异")
                else:
                    target.exceptions.append("96工时制自动基数: OT按直薪计算，双周封顶96h")
                if fixed_base <= 0:
                    continue

            target.base_override_amount = fixed_base
            target.base_override_type = rule_type
            target.base_override_reason = str(row.get("note") or row.get("rule_type") or "固定基数覆盖").strip()
            target.exceptions.append(
                f"固定基数覆盖: {target.base_override_reason}，覆盖后绩效基数={target.base_override_amount:.2f}"
            )

    @staticmethod
    def _supplemental_leave_overrides(
        supplemental_leave_data: dict | None,
        calc_month: str | None,
    ) -> dict[str, list[dict]]:
        """把已确认计入的补充假勤整理为保留日期的员工级条目。"""
        entries_by_source: dict[str, list[dict]] = defaultdict(list)
        if not supplemental_leave_data:
            return {}

        for row in supplemental_leave_data.get("rows", []):
            if row.get("confirmation_status") != "confirmed" or not row.get("include_in_base"):
                continue
            if calc_month and row.get("allocation_month") and row.get("allocation_month") != calc_month:
                continue

            employee_id = str(row.get("employee_id") or "").strip()
            if not employee_id:
                continue

            leave_type = str(row.get("leave_type") or row.get("raw_leave_type") or "").strip()
            hours = safe_float(row.get("included_hours", row.get("hours")))
            if hours <= 0:
                continue

            leave_hours = {
                "sick_hours": 0.0,
                "annual_hours": 0.0,
                "sick_settlement_hours": 0.0,
            }
            if leave_type == "病假":
                leave_hours["sick_hours"] = hours
            elif leave_type == "年假":
                leave_hours["annual_hours"] = hours
            elif leave_type in {"离职病假结算", "病假清算"}:
                leave_hours["sick_settlement_hours"] = hours
            else:
                continue

            entries_by_source[normalize_shift_employee_id(employee_id)].append({
                "employee_id": employee_id,
                "leave_hours": leave_hours,
                "start_at": parse_excel_datetime(row.get("start_at")),
                "end_at": parse_excel_datetime(row.get("end_at")),
            })
        return dict(entries_by_source)

    @staticmethod
    def _select_supplemental_leave_target(
        source_employees: list[EmployeeData],
        override_employee_id: str,
        start_at: datetime | None,
        end_at: datetime | None,
    ) -> EmployeeData:
        if start_at:
            dated_matches = [
                employee for employee in source_employees
                if any(_date_range_overlaps_row(row, start_at, end_at) for row in employee.attendance_daily_rows)
            ]
            if len(dated_matches) == 1:
                return dated_matches[0]
            if len(dated_matches) > 1:
                exact_match = next(
                    (employee for employee in dated_matches if employee.employee_id == override_employee_id),
                    None,
                )
                if exact_match:
                    return exact_match
                return next((employee for employee in dated_matches if not employee.is_night_shift), dated_matches[0])

        target = next(
            (employee for employee in source_employees if employee.employee_id == override_employee_id),
            None,
        )
        source_id = normalize_shift_employee_id(override_employee_id)
        if target is None and override_employee_id == source_id:
            target = next((employee for employee in source_employees if employee.employee_id == source_id), None)
        if target is None:
            target = next((employee for employee in source_employees if not employee.is_night_shift), source_employees[0])
        return target

    @classmethod
    def apply_supplemental_leave_to_employees(
        cls,
        employees: list[EmployeeData],
        supplemental_leave_data: dict | None,
        calc_month: str | None,
    ) -> None:
        """用补充假勤确认结果覆盖考勤日报中的病假、年假、病假清算。"""
        overrides = cls._supplemental_leave_overrides(supplemental_leave_data, calc_month)
        if not overrides:
            return

        by_source: dict[str, list[EmployeeData]] = defaultdict(list)
        for employee in employees:
            by_source[normalize_shift_employee_id(employee.source_employee_id)].append(employee)

        for source_id, row_overrides in overrides.items():
            source_employees = by_source.get(source_id)
            if not source_employees:
                continue

            for employee in source_employees:
                employee.sick_hours = 0.0
                employee.annual_hours = 0.0
                employee.sick_settlement_hours = 0.0

            for entry in row_overrides:
                target = cls._select_supplemental_leave_target(
                    source_employees,
                    entry["employee_id"],
                    entry.get("start_at"),
                    entry.get("end_at"),
                )
                leave_hours = entry["leave_hours"]
                target.sick_hours += leave_hours["sick_hours"]
                target.annual_hours += leave_hours["annual_hours"]
                target.sick_settlement_hours += leave_hours["sick_settlement_hours"]
                message = "病假/年假/病假清算已按补充假勤确认表覆盖考勤日报"
                if message not in target.exceptions:
                    target.exceptions.append(message)

    @staticmethod
    def _build_calculation_segments(
        raw_segments: list[dict],
        ratio: float,
        coefficient: float | None,
    ) -> list[CalculationSegment]:
        """把调薪拆分预览数据转成核算段。"""
        segments = []
        effective_coefficient = coefficient if coefficient is not None else 1.0
        for raw in raw_segments:
            reason = str(raw.get("reason", "")).strip()
            segment_ratio = _to_float(raw.get("performance_ratio"), None)
            if segment_ratio is None:
                segment_ratio = 0.0 if "前" in reason else ratio
            segment_coefficient = _to_float(raw.get("performance_coefficient"), None)
            if segment_coefficient is None:
                segment_coefficient = effective_coefficient
            segments.append(
                CalculationSegment(
                    period=str(raw.get("period", "")).strip(),
                    reason=reason,
                    performance_base=_to_float(raw.get("performance_base"), 0) or 0.0,
                    performance_ratio=segment_ratio,
                    performance_coefficient=segment_coefficient,
                    department=str(raw.get("department") or "").strip(),
                    position=str(raw.get("position") or "").strip(),
                    job_type=str(raw.get("job_type") or "").strip(),
                )
            )
        return segments

    def parse_transfer_history_preview(self, filepath: str, calc_month: str | None = None) -> dict:
        """解析人事调动记录；调动日期按正式生效日期处理。"""
        wb = self.load_excel(filepath)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("人事调动记录为空")

        header_index = None
        headers = ()
        for index, row in enumerate(rows[:20]):
            if (
                _find_column(row, ["工号", "员工工号"]) is not None
                and _find_column(row, ["调动日期"]) is not None
            ):
                header_index = index
                headers = row
                break
        if header_index is None:
            raise ValueError("未识别人事调动记录表头，请确认包含工号和调动日期")

        col_map = {
            "employee_id": _find_column(headers, ["工号", "员工工号"]),
            "name": _find_column(headers, ["姓名", "员工姓名"]),
            "effective_date": _find_column(headers, ["调动日期"]),
            "approval_status": _find_column(headers, ["审批状态"]),
            "transfer_type": _find_column(headers, ["异动类型", "调动类型"]),
            "transfer_reason": _find_column(headers, ["异动原因", "调动原因"]),
            "before_department": _find_column(headers, ["调动前部门"]),
            "before_department_levels": [
                _find_column(headers, [f"调动前{level}部门"])
                for level in ("二级", "三级", "四级", "五级", "六级", "七级", "八级")
            ],
            "before_position": _find_column(headers, ["调动前职位", "调动前岗位"]),
            "before_area": _find_column(headers, ["调动前划分区域"]),
            "before_cost_owner": _find_column(headers, ["调动前成本中心", "调动前成本归属"]),
            "after_department": _find_column(headers, ["调动后部门"]),
            "after_department_levels": [
                _find_column(headers, [f"调动后{level}部门"])
                for level in ("二级", "三级", "四级", "五级", "六级", "七级", "八级")
            ],
            "after_position": _find_column(headers, ["调动后职位", "调动后岗位"]),
            "after_area": _find_column(headers, ["调动后划分区域"]),
            "after_cost_owner": _find_column(headers, ["调动后成本中心", "调动后成本归属"]),
            "note": _find_column(headers, ["备注"]),
        }
        required = (
            "employee_id",
            "effective_date",
            "before_department",
            "before_position",
            "after_department",
            "after_position",
        )
        missing = [key for key in required if col_map[key] is None]
        if missing:
            raise ValueError("人事调动记录缺少调动前后部门或职位字段")

        month_start = _calc_month_start(calc_month)
        month_end = _calc_month_end(month_start)
        all_rows = []
        events = []
        status_counts: dict[str, int] = defaultdict(int)
        for source_row, row in enumerate(rows[header_index + 1:], start=header_index + 2):
            employee_id = str(_cell(row, col_map["employee_id"]) or "").strip()
            if not employee_id or not employee_id.lower().startswith("zt"):
                continue
            effective_date = _to_date(_cell(row, col_map["effective_date"]))
            approval_status = str(_cell(row, col_map["approval_status"]) or "").strip()
            status_counts[approval_status or "未填写"] += 1
            before_department_parts = [
                str(_cell(row, index) or "").strip()
                for index in col_map["before_department_levels"]
                if index is not None and str(_cell(row, index) or "").strip()
            ]
            after_department_parts = [
                str(_cell(row, index) or "").strip()
                for index in col_map["after_department_levels"]
                if index is not None and str(_cell(row, index) or "").strip()
            ]
            before_department = str(_cell(row, col_map["before_department"]) or "").strip()
            after_department = str(_cell(row, col_map["after_department"]) or "").strip()
            if before_department and (not before_department_parts or before_department_parts[-1] != before_department):
                before_department_parts.append(before_department)
            if after_department and (not after_department_parts or after_department_parts[-1] != after_department):
                after_department_parts.append(after_department)
            parsed = {
                "row_id": f"{ws.title}:{source_row}",
                "source_sheet": ws.title,
                "source_row": source_row,
                "employee_id": normalize_shift_employee_id(employee_id),
                "name": str(_cell(row, col_map["name"]) or "").strip(),
                "effective_date": effective_date.isoformat() if effective_date else "",
                "approval_status": approval_status,
                "transfer_type": str(_cell(row, col_map["transfer_type"]) or "").strip(),
                "transfer_reason": str(_cell(row, col_map["transfer_reason"]) or "").strip(),
                "before_department": "-".join(before_department_parts) or before_department,
                "before_position": str(_cell(row, col_map["before_position"]) or "").strip(),
                "before_area": str(_cell(row, col_map["before_area"]) or "").strip(),
                "before_cost_owner": str(_cell(row, col_map["before_cost_owner"]) or "").strip(),
                "after_department": "-".join(after_department_parts) or after_department,
                "after_position": str(_cell(row, col_map["after_position"]) or "").strip(),
                "after_area": str(_cell(row, col_map["after_area"]) or "").strip(),
                "after_cost_owner": str(_cell(row, col_map["after_cost_owner"]) or "").strip(),
                "note": str(_cell(row, col_map["note"]) or "").strip(),
            }
            all_rows.append(parsed)
            if approval_status in {"已完成", "completed"} and effective_date:
                events.append(parsed)

        events.sort(key=lambda item: (item["employee_id"], item["effective_date"], item["source_row"]))
        affected_in_month = {
            row["employee_id"]
            for row in events
            if month_start and month_end and month_start <= _to_date(row["effective_date"]) <= month_end
        }
        future_evidence = {
            row["employee_id"]
            for row in events
            if month_end and _to_date(row["effective_date"]) > month_end
        }
        return {
            "events": events,
            "rows": all_rows,
            "summary": {
                "total_rows": len(all_rows),
                "completed_count": len(events),
                "ignored_count": len(all_rows) - len(events),
                "affected_employee_count": len({row["employee_id"] for row in events}),
                "in_month_employee_count": len(affected_in_month),
                "future_evidence_employee_count": len(future_evidence),
                "status_counts": dict(status_counts),
            },
        }

    @staticmethod
    def _transfer_assignment(event: dict, prefix: str) -> dict:
        department = str(event.get(f"{prefix}_department") or "").strip()
        return {
            "department": department,
            "position": str(event.get(f"{prefix}_position") or "").strip(),
            "area": str(event.get(f"{prefix}_area") or "").strip(),
            "cost_owner": str(event.get(f"{prefix}_cost_owner") or "").strip(),
            "job_type": classify_job_type(event.get("employee_id", ""), department),
        }

    @classmethod
    def _assignment_on_date(cls, events: list[dict], target_date: date) -> dict:
        assignment = cls._transfer_assignment(events[0], "before")
        for event in events:
            effective_date = _to_date(event.get("effective_date"))
            if effective_date and effective_date <= target_date:
                assignment = cls._transfer_assignment(event, "after")
            elif effective_date and effective_date > target_date:
                break
        return assignment

    @staticmethod
    def _salary_on_date(salary_info: dict, target_date: date) -> tuple[float, float]:
        for segment in salary_info.get("effective_segments") or []:
            start = _to_date(segment.get("period_start"))
            end = _to_date(segment.get("period_end"))
            if start and end and start <= target_date <= end:
                return (
                    _to_float(segment.get("hourly_rate"), 0) or 0.0,
                    _to_float(segment.get("performance_ratio"), 0) or 0.0,
                )
        return (
            _to_float(salary_info.get("hourly_rate"), 0) or 0.0,
            _to_float(salary_info.get("ratio"), 0) or 0.0,
        )

    @classmethod
    def _apply_transfer_history(
        cls,
        transfer_data: dict | None,
        attendance_dict: dict,
        salary_dict: dict,
        performance_dict: dict,
        employee_info: dict,
        adjustment_dict: dict,
        calc_month: str | None,
    ) -> None:
        """用已完成调动记录恢复历史岗位，并按实际出勤日拆分月内调动。"""
        month_start = _calc_month_start(calc_month)
        month_end = _calc_month_end(month_start)
        if not month_start or not month_end:
            return

        events_by_id: dict[str, list[dict]] = defaultdict(list)
        for event in (transfer_data or {}).get("events", []):
            employee_id = normalize_shift_employee_id(event.get("employee_id"))
            effective_date = _to_date(event.get("effective_date"))
            if (
                employee_id
                and effective_date
                and str(event.get("approval_status") or "").strip() in {"已完成", "completed"}
            ):
                events_by_id[employee_id].append(event)

        for employee_id, events in events_by_id.items():
            if employee_id not in employee_info:
                continue
            events.sort(key=lambda item: (_to_date(item.get("effective_date")), item.get("source_row", 0)))
            month_assignment = cls._assignment_on_date(events, month_end)
            info = employee_info[employee_id]
            for field in ("department", "position", "area", "job_type"):
                if month_assignment.get(field):
                    info[field] = month_assignment[field]

            in_month_events = [
                event for event in events
                if month_start <= _to_date(event.get("effective_date")) <= month_end
            ]
            if not in_month_events:
                continue

            daily_rows = [
                row for row in attendance_dict.get(employee_id, {}).get("daily_rows", [])
                if (row_date := _daily_row_date(row)) and month_start <= row_date <= month_end
            ]
            if not daily_rows:
                continue

            performance = performance_dict.get(employee_id, {})
            salary_info = salary_dict.get(employee_id, {})
            personnel_status = info.get("personnel_status", "")
            confirmation_date = info.get("confirmation_date")
            has_performance_result = bool(performance)
            grouped: list[dict] = []
            for daily_row in sorted(daily_rows, key=lambda row: _daily_row_date(row) or month_start):
                row_date = _daily_row_date(daily_row)
                if not row_date:
                    continue
                assignment = cls._assignment_on_date(events, row_date)
                hourly_rate, performance_ratio = cls._salary_on_date(salary_info, row_date)
                override_reason = default_coefficient_reason(
                    assignment.get("position", ""),
                    personnel_status,
                    confirmation_date=confirmation_date,
                    calc_month_start=month_start,
                    has_performance_result=has_performance_result,
                )
                if override_reason:
                    coefficient = DEFAULT_COEFFICIENT_VALUE
                elif performance.get("coefficient") is not None:
                    coefficient = _to_float(performance.get("coefficient"), 0) or 0.0
                else:
                    coefficient = CoefficientCalculator.calculate(
                        job_type=assignment.get("job_type") or "warehouse",
                        score=performance.get("score"),
                        level=performance.get("level"),
                    )
                key = (
                    assignment.get("department", ""),
                    assignment.get("position", ""),
                    assignment.get("job_type", ""),
                    round(hourly_rate, 6),
                    round(performance_ratio, 6),
                    round(coefficient, 6),
                )
                if not grouped or grouped[-1]["key"] != key:
                    grouped.append({
                        "key": key,
                        "start": row_date,
                        "end": row_date,
                        "rows": [daily_row],
                        "assignment": assignment,
                        "hourly_rate": hourly_rate,
                        "performance_ratio": performance_ratio,
                        "performance_coefficient": coefficient,
                        "override_reason": override_reason,
                    })
                else:
                    grouped[-1]["end"] = row_date
                    grouped[-1]["rows"].append(daily_row)

            if len(grouped) == 1:
                only_assignment = grouped[0]["assignment"]
                for field in ("department", "position", "area", "job_type"):
                    if only_assignment.get(field):
                        info[field] = only_assignment[field]
                continue
            segments = []
            for group in grouped:
                performance_base = _daily_rows_performance_base(
                    group["rows"],
                    group["hourly_rate"],
                )
                if performance_base <= 0:
                    continue
                assignment = group["assignment"]
                position = assignment.get("position") or "未填写岗位"
                reason = f"岗位调动：{position}"
                if group["override_reason"]:
                    reason = f"{reason}；{group['override_reason']}"
                segments.append({
                    "period": _period_label(group["start"], group["end"]),
                    "reason": reason,
                    "performance_base": performance_base,
                    "performance_ratio": group["performance_ratio"],
                    "performance_coefficient": group["performance_coefficient"],
                    "department": assignment.get("department", ""),
                    "position": assignment.get("position", ""),
                    "job_type": assignment.get("job_type", ""),
                })
            if len(segments) > 1:
                adjustment_dict[employee_id] = segments

    @staticmethod
    def _apply_oehr_adjustment_events(
        events: list[dict],
        attendance_dict: dict,
        salary_dict: dict,
        employee_info: dict,
        adjustment_dict: dict,
        calc_month: str | None,
    ) -> None:
        """把OEHR转正调薪事件转成可核算的自动分段。"""
        month_start = _calc_month_start(calc_month)
        month_end = _calc_month_end(month_start)
        if not month_start or not month_end:
            return

        for event in events or []:
            emp_id = str(event.get("employee_id") or "").strip()
            if not emp_id or emp_id in adjustment_dict or not _is_auto_adjustment_event(event):
                continue
            if salary_dict.get(emp_id, {}).get("verification_status"):
                continue

            effective_date = _to_date(event.get("effective_date"))
            if not effective_date or effective_date > month_end:
                continue

            event_ratio = _to_float(event.get("performance_ratio"), 0) or 0.0
            if event_ratio <= 0:
                continue

            salary_info = salary_dict.setdefault(
                emp_id,
                {
                    "hourly_rate": 0.0,
                    "ratio": 0.0,
                    "calculation_method": "",
                    "fixed_performance_base": 0,
                },
            )
            event_hourly_rate = _to_float(event.get("hourly_rate"), 0) or 0.0
            if event_hourly_rate > 0:
                salary_info["hourly_rate"] = event_hourly_rate
            hourly_rate = _to_float(salary_info.get("hourly_rate"), 0) or event_hourly_rate
            if hourly_rate <= 0:
                continue

            salary_info["ratio"] = event_ratio

            if "转正" in " ".join(str(event.get(key) or "") for key in ("subject", "adjustment_reason", "note")):
                info = employee_info.setdefault(emp_id, {})
                if not info.get("confirmation_date"):
                    info["confirmation_date"] = effective_date

            attendance = attendance_dict.get(emp_id, {})
            daily_rows = [
                row for row in attendance.get("daily_rows", [])
                if (row_date := _daily_row_date(row)) and month_start <= row_date <= month_end
            ]
            if not daily_rows:
                continue

            if effective_date <= month_start:
                continue

            pre_rows = [
                row for row in daily_rows
                if (row_date := _daily_row_date(row)) and row_date < effective_date
            ]
            post_rows = [
                row for row in daily_rows
                if (row_date := _daily_row_date(row)) and row_date >= effective_date
            ]
            if not post_rows:
                continue

            segments = []
            pre_base = _daily_rows_performance_base(pre_rows, hourly_rate)
            if pre_base > 0:
                segments.append({
                    "period": _period_label(month_start, effective_date - timedelta(days=1)),
                    "reason": "调薪前",
                    "performance_base": pre_base,
                    "performance_ratio": 0.0,
                })

            post_base = _daily_rows_performance_base(post_rows, hourly_rate)
            if post_base <= 0:
                continue
            segments.append({
                "period": _period_label(effective_date, month_end),
                "reason": "调薪后",
                "performance_base": post_base,
                "performance_ratio": event_ratio,
            })
            adjustment_dict[emp_id] = segments

    @staticmethod
    def _apply_salary_history_segments(
        salary_dict: dict,
        attendance_dict: dict,
        adjustment_dict: dict,
    ) -> None:
        """Convert verified salary effective-date segments into calculation segments."""
        for emp_id, salary_info in salary_dict.items():
            raw_segments = salary_info.get("effective_segments") or []
            if not raw_segments or emp_id in adjustment_dict:
                continue

            daily_rows = attendance_dict.get(emp_id, {}).get("daily_rows", [])
            calculation_segments = []
            for segment in raw_segments:
                period_start = _to_date(segment.get("period_start"))
                period_end = _to_date(segment.get("period_end"))
                if not period_start or not period_end:
                    continue
                segment_rows = [
                    row for row in daily_rows
                    if (row_date := _daily_row_date(row)) and period_start <= row_date <= period_end
                ]
                performance_base = _daily_rows_performance_base(
                    segment_rows,
                    _to_float(segment.get("hourly_rate"), 0) or 0.0,
                )
                if performance_base <= 0:
                    continue
                calculation_segments.append({
                    "period": _period_label(period_start, period_end),
                    "reason": str(segment.get("reason") or "调薪拆分"),
                    "performance_base": performance_base,
                    "performance_ratio": _to_float(segment.get("performance_ratio"), 0) or 0.0,
                })

            if calculation_segments:
                adjustment_dict[emp_id] = calculation_segments

    def parse_all(
        self,
        attendance_file: str,
        salary_file: str,
        performance_file: str,
        target_month: int,
        calc_month: str | None = None,
    ) -> FBUPerformanceEngine:
        """
        解析所有数据并计算

        Args:
            attendance_file: 考勤日报表路径
            salary_file: 薪资档案路径
            performance_file: 绩效报表路径
            target_month: 目标月份

        Returns:
            计算完成的引擎实例
        """
        # 1. 解析考勤数据
        attendance_data = self.parse_attendance(attendance_file, target_month)

        # 2. 解析薪资档案
        salary_data = self.parse_salary(salary_file)

        # 3. 解析绩效报表
        performance_data = self.parse_performance(performance_file)

        # 4. 构建员工数据
        employees = self.build_employees(
            attendance_data,
            salary_data,
            performance_data,
            self.employee_roster,
            calc_month=calc_month,
        )

        # 5. 计算绩效奖金
        for emp in employees:
            BonusCalculator.calculate(emp)
            self.engine.add_employee(emp)

        return self.engine

    def build_attendance_preview_rows(self, attendance_data: dict) -> list[dict]:
        """构建考勤预览行；混班员工按白班/夜班物理拆行。"""
        employee_list = []

        def blank_shift() -> dict:
            return {
                '计薪出勤': 0,
                'OT1.5': 0,
                'OT2.0': 0,
                '病假': 0,
                '病假清算': 0,
                '年假': 0,
                '节假日': 0,
            }

        for emp_id, hours in attendance_data.items():
            source_emp_id = normalize_shift_employee_id(emp_id)
            emp_info = self.get_employee_info(source_emp_id)
            roster_matched = bool(emp_info.get('name') or emp_info.get('department') or emp_info.get('area'))
            day_hours = hours['白班']
            night_hours = hours['夜班']
            has_day = has_shift_hours(day_hours)
            has_night = bool(hours.get('has_night_shift')) and has_shift_hours(night_hours)
            split_mixed_shift = has_day and has_night

            def make_row(row_emp_id: str, shift_type: str, shift_hours: dict, is_night: bool) -> dict:
                daily_rows = hours.get(f'{shift_type}_daily_rows', [])
                return {
                    "employee_id": row_emp_id,
                    "source_employee_id": source_emp_id,
                    "name": emp_info['name'] or hours.get('name', ''),
                    "department": emp_info['department'],
                    "area": emp_info['area'],
                    "personnel_status": emp_info.get('personnel_status', ''),
                    "hire_date": emp_info.get('hire_date').isoformat() if emp_info.get('hire_date') else "",
                    "confirmation_date": emp_info.get('confirmation_date').isoformat() if emp_info.get('confirmation_date') else "",
                    "resignation_date": emp_info.get('resignation_date').isoformat() if emp_info.get('resignation_date') else "",
                    "position": emp_info.get('position', ''),
                    "job_type": emp_info['job_type'],
                    "roster_matched": roster_matched,
                    "shift_type": shift_type,
                    "has_night_shift": is_night,
                    "day_shift": blank_shift() if is_night else shift_hours,
                    "night_shift": shift_hours if is_night else blank_shift(),
                    "attendance_daily_rows": list(daily_rows),
                    "total_base_hours": shift_hours.get('计薪出勤', 0),
                    "total_ot15": shift_hours.get('OT1.5', 0),
                    "total_ot20": shift_hours.get('OT2.0', 0),
                    "sick_hours": shift_hours.get('病假', 0),
                    "sick_settlement_hours": shift_hours.get('病假清算', 0),
                    "annual_hours": shift_hours.get('年假', 0),
                    "holiday_hours": shift_hours.get('节假日', 0),
                }

            if has_day:
                employee_list.append(make_row(
                    f"{emp_id}-1" if split_mixed_shift else emp_id,
                    "白班",
                    day_hours,
                    False,
                ))
            if has_night:
                employee_list.append(make_row(emp_id, "夜班", night_hours, True))

        return employee_list

    def parse_attendance_preview(self, filepath: str, target_month: int) -> dict:
        """
        解析考勤数据并返回预览

        Args:
            filepath: 考勤日报表路径
            target_month: 目标月份

        Returns:
            预览数据 {员工明细列表, 汇总统计}
        """
        wb = self.load_excel(filepath, read_only=True)
        ws = wb['sheet1']
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

        # 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _cell(row, 0) is not None:
                rows.append(row)
        wb.close()

        # 处理考勤数据
        attendance_data = self.attendance_processor.process(rows, target_month, headers=headers)

        employee_list = self.build_attendance_preview_rows(attendance_data)

        # 汇总统计
        total_employees = len(attendance_data)
        day_shift_count = sum(1 for e in employee_list if e.get('shift_type') == '白班')
        night_shift_count = sum(1 for e in employee_list if e.get('shift_type') == '夜班')
        roster_matched = sum(1 for e in employee_list if e.get('roster_matched'))
        total_base_hours = sum(e['total_base_hours'] for e in employee_list)
        total_ot15 = sum(e['total_ot15'] for e in employee_list)
        total_ot20 = sum(e['total_ot20'] for e in employee_list)

        return {
            "employees": employee_list,
            "summary": {
                "total_employees": total_employees,
                "roster_matched": roster_matched,
                "roster_missing": len(employee_list) - roster_matched,
                "attendance_rows": len(employee_list),
                "day_shift_count": day_shift_count,
                "night_shift_count": night_shift_count,
                "total_base_hours": round(total_base_hours, 2),
                "total_ot15": round(total_ot15, 2),
                "total_ot20": round(total_ot20, 2),
            }
        }

    def parse_salary_preview(self, filepath: str) -> dict:
        """
        解析薪资档案并返回预览

        Args:
            filepath: 薪资档案路径

        Returns:
            预览数据 {员工明细列表, 汇总统计}
        """
        wb = self.load_excel(filepath)
        ws = wb[wb.sheetnames[0]]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

        # 读取数据行
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if _cell(row, 0) is not None:
                rows.append(row)

        salary_data = self.salary_processor.load(rows, headers=headers)

        # 构建明细列表
        employee_list = []
        for emp_id, info in salary_data.items():
            # 获取员工信息
            emp_info = self.get_employee_info(emp_id)

            employee_list.append({
                "employee_id": emp_id,
                "name": emp_info['name'] or info.get('name', ''),
                "department": info.get('department') or emp_info['department'],
                "area": info.get('area') or emp_info['area'],
                "position": info.get('position') or emp_info.get('position', ''),
                "personnel_status": info.get('personnel_status') or emp_info.get('personnel_status', ''),
                "cost_owner": info.get('cost_owner', ''),
                "hourly_rate": info.get('hourly_rate', 0),
                "ratio": info.get('ratio', 0),
                "calculation_method": info.get('calculation_method', ''),
                "fixed_performance_base": info.get('fixed_performance_base', 0),
            })

        # 汇总统计：薪资档案常包含离职或未维护时薪的员工，预览中拆开显示。
        total_employees = len(employee_list)
        valid_hourly_count = sum(1 for e in employee_list if e['hourly_rate'] > 0)
        zero_hourly_count = sum(1 for e in employee_list if e['hourly_rate'] == 0)
        avg_hourly_rate = (
            sum(e['hourly_rate'] for e in employee_list if e['hourly_rate'] > 0) / valid_hourly_count
            if valid_hourly_count > 0
            else 0
        )

        return {
            "employees": employee_list,
            "summary": {
                "total_employees": total_employees,
                "valid_hourly_count": valid_hourly_count,
                "zero_hourly_count": zero_hourly_count,
                "avg_hourly_rate": round(avg_hourly_rate, 2),
            }
        }

    @staticmethod
    def reconcile_salary_history(
        previous_salary: list[dict],
        current_salary: list[dict],
        adjustment_events: list[dict],
        calc_month: str,
        roster_by_id: dict[str, dict] | None = None,
    ) -> dict:
        """Resolve salary snapshot changes using completed adjustment effective dates."""
        month_start = _calc_month_start(calc_month)
        month_end = _calc_month_end(month_start)
        if not month_start or not month_end:
            raise ValueError("核算月份格式无效")
        previous_month_end = month_start - timedelta(days=1)
        previous_month_start = previous_month_end.replace(day=1)

        previous_by_id = {
            str(row.get("employee_id") or "").strip(): row
            for row in previous_salary or []
            if str(row.get("employee_id") or "").strip()
        }
        current_by_id = {
            str(row.get("employee_id") or "").strip(): row
            for row in current_salary or []
            if str(row.get("employee_id") or "").strip()
        }
        events_by_id: dict[str, list[dict]] = defaultdict(list)
        for event in adjustment_events or []:
            employee_id = str(event.get("employee_id") or "").strip()
            if employee_id and str(event.get("approval_status") or "").strip() in {"已完成", "completed"}:
                events_by_id[employee_id].append(event)

        def value(row: dict, key: str) -> float:
            return round(_to_float(row.get(key), 0) or 0.0, 10)

        def changed(left: float, right: float) -> bool:
            return abs(left - right) >= 0.0001

        employees = []
        issues = []
        ignored_historical_resigned_count = 0
        for employee_id in sorted(set(previous_by_id) | set(current_by_id)):
            previous = previous_by_id.get(employee_id)
            current = current_by_id.get(employee_id)
            if current is None:
                if str((previous or {}).get("personnel_status") or "").strip() == "离职":
                    ignored_historical_resigned_count += 1
                    continue
                row = dict(previous or {})
                previous_hourly = value(previous or {}, "hourly_rate")
                previous_ratio = value(previous or {}, "ratio")
                row.update({
                    "employee_id": employee_id,
                    "verification_status": "blocking",
                    "resolution": "missing_current_snapshot",
                    "effective_segments": [],
                    "previous_hourly_rate": previous_hourly,
                    "previous_ratio": previous_ratio,
                    "current_hourly_rate": None,
                    "current_ratio": None,
                })
                employees.append(row)
                issues.append({"employee_id": employee_id, "reason": "当月薪资档案缺少该员工"})
                continue
            if previous is None:
                row = dict(current)
                roster_row = (roster_by_id or {}).get(employee_id, {})
                hire_date = _to_date(roster_row.get("hire_date"))
                if not hire_date or hire_date <= previous_month_end:
                    row.update({
                        "verification_status": "blocking",
                        "resolution": "missing_previous_snapshot",
                        "effective_segments": [],
                        "previous_hourly_rate": None,
                        "previous_ratio": None,
                        "current_hourly_rate": value(current, "hourly_rate"),
                        "current_ratio": value(current, "ratio"),
                    })
                    employees.append(row)
                    issues.append({
                        "employee_id": employee_id,
                        "reason": "上月薪资档案缺少该员工，且花名册入职日期无法证明其为本月新员工",
                    })
                    continue
                row.update({
                    "verification_status": "resolved",
                    "resolution": "new_employee_use_current",
                    "effective_segments": [],
                    "previous_hourly_rate": 0.0,
                    "previous_ratio": 0.0,
                    "current_hourly_rate": value(current, "hourly_rate"),
                    "current_ratio": value(current, "ratio"),
                })
                employees.append(row)
                continue

            old_hourly = value(previous, "hourly_rate")
            new_hourly = value(current, "hourly_rate")
            old_ratio = value(previous, "ratio")
            new_ratio = value(current, "ratio")
            hourly_changed = changed(old_hourly, new_hourly)
            ratio_changed = changed(old_ratio, new_ratio)
            row = dict(current)
            row.update({
                "previous_hourly_rate": old_hourly,
                "previous_ratio": old_ratio,
                "current_hourly_rate": new_hourly,
                "current_ratio": new_ratio,
                "effective_segments": [],
            })
            if not hourly_changed and not ratio_changed:
                row.update({"verification_status": "resolved", "resolution": "unchanged"})
                employees.append(row)
                continue

            matching_events = []
            for event in events_by_id.get(employee_id, []):
                effective_date = _to_date(event.get("effective_date"))
                if not effective_date or effective_date < previous_month_start:
                    continue
                event_hourly = round(_to_float(event.get("hourly_rate"), 0) or 0.0, 10)
                event_ratio = round(_to_float(event.get("performance_ratio"), 0) or 0.0, 10)
                if hourly_changed and changed(event_hourly, new_hourly):
                    continue
                if ratio_changed and changed(event_ratio, new_ratio):
                    continue
                matching_events.append((effective_date, event))

            if not matching_events:
                row.update({"verification_status": "blocking", "resolution": "missing_adjustment"})
                employees.append(row)
                issues.append({"employee_id": employee_id, "reason": "薪资字段变化但未匹配已完成调薪流程"})
                continue

            effective_date, matched_event = max(matching_events, key=lambda item: item[0])
            row["matched_adjustment"] = dict(matched_event)
            row["effective_date"] = effective_date.isoformat()
            row["verification_status"] = "resolved"
            if effective_date > month_end:
                row["hourly_rate"] = old_hourly
                row["ratio"] = old_ratio
                row["resolution"] = "future_effective_use_previous"
            elif effective_date <= month_start:
                row["hourly_rate"] = new_hourly
                row["ratio"] = new_ratio
                row["resolution"] = "effective_before_month_use_current"
            else:
                row["hourly_rate"] = new_hourly
                row["ratio"] = new_ratio
                row["resolution"] = "in_month_split"
                row["effective_segments"] = [
                    {
                        "period_start": month_start.isoformat(),
                        "period_end": (effective_date - timedelta(days=1)).isoformat(),
                        "hourly_rate": old_hourly,
                        "performance_ratio": old_ratio,
                        "reason": "调薪前",
                    },
                    {
                        "period_start": effective_date.isoformat(),
                        "period_end": month_end.isoformat(),
                        "hourly_rate": new_hourly,
                        "performance_ratio": new_ratio,
                        "reason": "调薪后",
                    },
                ]
            employees.append(row)

        blocking_count = sum(1 for row in employees if row.get("verification_status") == "blocking")
        changed_count = sum(
            1 for row in employees
            if changed(_to_float(row.get("previous_hourly_rate"), 0) or 0, _to_float(row.get("current_hourly_rate"), 0) or 0)
            or changed(_to_float(row.get("previous_ratio"), 0) or 0, _to_float(row.get("current_ratio"), 0) or 0)
        )
        return {
            "employees": employees,
            "issues": issues,
            "summary": {
                "total_employees": len(employees),
                "changed_count": changed_count,
                "resolved_count": len(employees) - blocking_count,
                "blocking_count": blocking_count,
                "ignored_historical_resigned_count": ignored_historical_resigned_count,
            },
        }

    def parse_performance_preview(self, filepath: str) -> dict:
        """
        解析绩效报表并返回预览

        Args:
            filepath: 绩效报表路径

        Returns:
            预览数据 {员工明细列表, 汇总统计}
        """
        wb = self.load_excel(filepath)
        ws = wb[wb.sheetnames[0]]

        supplement_preview = self._parse_performance_supplement_preview(ws)
        if supplement_preview is not None:
            return supplement_preview

        employee_list = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if _cell(row, 3) is None:  # 工号列
                continue

            emp_id = str(_cell(row, 3)).strip()
            score = _cell(row, 16)  # 总分
            level = _cell(row, 17)  # 总等级
            coefficient = _cell(row, 18)  # 绩效系数

            # 获取员工信息
            emp_info = self.get_employee_info(emp_id)

            employee_list.append({
                "employee_id": emp_id,
                "name": emp_info['name'],
                "department": emp_info['department'],
                "area": emp_info['area'],
                "job_type": emp_info['job_type'],
                "score": _to_float(score),
                "level": str(level).strip() if level else None,
                "coefficient": _to_float(coefficient),
            })

        return {
            "employees": employee_list,
            "summary": self._summarize_performance_rows(employee_list, source_type="performance_report"),
        }

    def _parse_performance_supplement_preview(self, worksheet) -> dict | None:
        rows = list(worksheet.iter_rows(values_only=True))
        header_row_index = None
        headers = None
        for index, row in enumerate(rows[:20]):
            employee_id_col = _find_column(row, ["工号", "员工工号", "employee_id", "Employee ID"])
            score_col = _find_column(row, ["绩效得分", "绩效分数", "总分", "得分", "score", "Score"])
            level_col = _find_column(row, ["绩效等级", "总等级", "等级", "level", "Level"])
            coefficient_col = _find_column(row, ["绩效系数", "系数", "coefficient", "Coefficient"])
            if employee_id_col is not None and (
                score_col is not None or level_col is not None or coefficient_col is not None
            ):
                header_row_index = index
                headers = row
                break

        if header_row_index is None or headers is None:
            return None

        col_map = {
            "employee_id": _find_column(headers, ["工号", "员工工号", "employee_id", "Employee ID"]),
            "name": _find_column(headers, ["姓名", "员工姓名", "name", "Name"]),
            "score": _find_column(headers, ["绩效得分", "绩效分数", "总分", "得分", "score", "Score"]),
            "level": _find_column(headers, ["绩效等级", "总等级", "等级", "level", "Level"]),
            "coefficient": _find_column(headers, ["绩效系数", "系数", "coefficient", "Coefficient"]),
            "note": _find_column(headers, ["备注", "说明", "note", "Note"]),
        }

        source_type = self._performance_header_source_type(worksheet.title, headers)
        source_label = "绩效补录" if source_type == "performance_supplement" else "OEHR绩效报表"
        employee_list = []
        for row in rows[header_row_index + 1:]:
            emp_id = str(_cell(row, col_map["employee_id"]) or "").strip()
            if not emp_id:
                continue
            score = _to_float(_cell(row, col_map["score"]))
            level_value = _cell(row, col_map["level"])
            coefficient = _to_float(_cell(row, col_map["coefficient"]))
            if score is None and not level_value and coefficient is None:
                continue

            emp_info = self.get_employee_info(emp_id)
            row_name = str(_cell(row, col_map["name"]) or "").strip()
            note = str(_cell(row, col_map["note"]) or "").strip()
            employee_list.append({
                "employee_id": emp_id,
                "name": emp_info["name"] or row_name,
                "department": emp_info["department"],
                "area": emp_info["area"],
                "job_type": emp_info["job_type"],
                "score": score,
                "level": str(level_value).strip() if level_value else None,
                "coefficient": coefficient,
                "performance_source": source_label,
                "note": note,
            })

        return {
            "employees": employee_list,
            "summary": self._summarize_performance_rows(employee_list, source_type=source_type),
        }

    @staticmethod
    def _performance_header_source_type(sheet_title: str, headers: tuple | list) -> str:
        text = " ".join([str(sheet_title or ""), *(str(header or "") for header in headers)])
        supplement_markers = ("绩效补录", "补录", "线下绩效", "线下考核", "离职线下", "离职人员")
        if any(marker in text for marker in supplement_markers):
            return "performance_supplement"
        return "performance_report"

    @staticmethod
    def _summarize_performance_rows(
        employee_list: list[dict],
        source_type: str | None = None,
        extra: dict | None = None,
    ) -> dict:
        level_distribution = defaultdict(int)
        for employee in employee_list:
            level = employee.get("level")
            if level:
                level_distribution[str(level)] += 1
        scored_employees = [e for e in employee_list if e.get("score") is not None]
        avg_score = (
            sum(e["score"] for e in scored_employees) / len(scored_employees)
            if scored_employees else 0
        )
        summary = {
            "total_employees": len(employee_list),
            "scored_employees": len(scored_employees),
            "avg_score": round(avg_score, 2),
            "level_distribution": dict(level_distribution),
        }
        if source_type:
            summary["source_type"] = source_type
        if extra:
            summary.update(extra)
        return summary

    @staticmethod
    def _performance_row_has_result(employee: dict) -> bool:
        return (
            employee.get("score") is not None
            or bool(employee.get("level"))
            or employee.get("coefficient") is not None
        )

    def build_performance_supplement_preview(self, rows: list[dict]) -> dict:
        employee_list = []
        for row in rows:
            emp_id = str(row.get("employee_id") or "").strip()
            if not emp_id:
                continue
            emp_info = self.get_employee_info(emp_id)
            employee_list.append({
                "employee_id": emp_id,
                "name": emp_info["name"] or str(row.get("name") or "").strip(),
                "department": emp_info["department"],
                "area": emp_info["area"],
                "job_type": emp_info["job_type"],
                "score": _to_float(row.get("score")),
                "level": str(row.get("level") or "").strip() or None,
                "coefficient": _to_float(row.get("coefficient")),
                "performance_source": "绩效补录",
                "note": str(row.get("note") or "").strip(),
            })

        return {
            "employees": employee_list,
            "summary": self._summarize_performance_rows(employee_list, source_type="performance_supplement"),
        }

    def merge_performance_supplement_preview(self, existing_preview: dict, supplement_preview: dict) -> dict:
        existing_rows = [dict(employee) for employee in existing_preview.get("employees", [])]
        by_employee_id = {
            str(employee.get("employee_id") or "").strip(): employee
            for employee in existing_rows
            if employee.get("employee_id")
        }
        added = 0
        updated_missing = 0
        skipped_existing = 0

        for supplement_row in supplement_preview.get("employees", []):
            emp_id = str(supplement_row.get("employee_id") or "").strip()
            if not emp_id:
                continue
            existing_row = by_employee_id.get(emp_id)
            if existing_row is None:
                copied = dict(supplement_row)
                existing_rows.append(copied)
                by_employee_id[emp_id] = copied
                added += 1
                continue
            if self._performance_row_has_result(existing_row):
                skipped_existing += 1
                continue
            existing_row.update({key: value for key, value in supplement_row.items() if value not in (None, "")})
            updated_missing += 1

        return {
            "employees": existing_rows,
            "summary": self._summarize_performance_rows(
                existing_rows,
                source_type="merged_performance",
                extra={
                    "supplement_added": added,
                    "supplement_updated_missing": updated_missing,
                    "supplement_skipped_existing": skipped_existing,
                },
            ),
        }

    @staticmethod
    def _summarize_supplemental_leave(rows: list[dict]) -> dict:
        total_hours = sum(safe_float(row.get("hours")) for row in rows)
        include_rows = [row for row in rows if row.get("include_in_base")]
        pending_rows = [row for row in rows if row.get("confirmation_status") == "pending"]
        confirmed_rows = [row for row in rows if row.get("confirmation_status") == "confirmed"]
        excluded_rows = [row for row in rows if row.get("confirmation_status") == "excluded"]
        return {
            "total_rows": len(rows),
            "include_count": len(include_rows),
            "include_hours": round(
                sum(safe_float(row.get("included_hours", row.get("hours"))) for row in include_rows),
                2,
            ),
            "pending_count": len(pending_rows),
            "confirmed_count": len(confirmed_rows),
            "excluded_count": len(excluded_rows),
            "attendance_unmatched_count": sum(1 for row in rows if row.get("attendance_matched") is False),
            "termination_settlement_count": sum(1 for row in rows if row.get("is_termination_settlement")),
            "suggested_count": sum(1 for row in rows if safe_float(row.get("suggested_included_hours")) > 0),
            "total_hours": round(total_hours, 2),
        }

    @staticmethod
    def _merge_duplicate_supplemental_leave_rows(rows: list[dict]) -> list[dict]:
        """合并同一流程在源表中被拆成多行的假勤小时。"""
        merged: dict[tuple, dict] = {}
        output: list[dict] = []

        for row in rows:
            flow_id = str(row.get("flow_id") or "").strip()
            if not flow_id:
                row["source_rows"] = row.get("source_rows") or [row.get("source_row")]
                output.append(row)
                continue

            key = (
                row.get("source_sheet"),
                normalize_shift_employee_id(row.get("employee_id")),
                flow_id,
                row.get("raw_leave_type"),
                row.get("leave_type"),
                row.get("start_at"),
                row.get("end_at"),
                row.get("paid_at"),
                row.get("is_termination_settlement"),
                row.get("suggested_month"),
                row.get("suggested_period"),
                row.get("allocation_month"),
                row.get("allocation_period"),
                row.get("confirmation_status"),
                bool(row.get("include_in_base")),
            )
            source_rows = row.get("source_rows") or [row.get("source_row")]
            if key not in merged:
                next_row = dict(row)
                next_row["source_rows"] = [item for item in source_rows if item]
                merged[key] = next_row
                output.append(next_row)
                continue

            target = merged[key]
            target["hours"] = round(safe_float(target.get("hours")) + safe_float(row.get("hours")), 2)
            target["included_hours"] = round(
                safe_float(target.get("included_hours")) + safe_float(row.get("included_hours")),
                2,
            )
            target["source_rows"] = [
                *target.get("source_rows", []),
                *[item for item in source_rows if item],
            ]

        return output

    def validate_supplemental_leave_against_attendance(
        self,
        preview: dict,
        attendance_data: dict | None,
    ) -> dict:
        """校验补充假勤员工是否存在于本次考勤，未匹配行保留假勤默认计入规则。"""
        attendance_ids = {
            normalize_shift_employee_id(row.get("employee_id"))
            for row in (attendance_data or {}).get("employees", [])
            if row.get("employee_id")
        }
        if not attendance_ids:
            return {
                "rows": [dict(row, attendance_matched=None) for row in preview.get("rows", [])],
                "summary": self._summarize_supplemental_leave(preview.get("rows", [])),
            }

        rows = []
        for row in preview.get("rows", []):
            next_row = dict(row)
            employee_id = normalize_shift_employee_id(next_row.get("employee_id"))
            if employee_id in attendance_ids:
                next_row["attendance_matched"] = True
            else:
                next_row["attendance_matched"] = False
                reason = str(next_row.get("system_reason") or "").strip()
                unmatched_reason = "该员工未匹配本月考勤，系统按补充假勤规则默认计入；请确认工号或考勤拆行。"
                next_row["system_reason"] = f"{reason} {unmatched_reason}".strip() if reason else unmatched_reason
            rows.append(next_row)

        return {
            "rows": rows,
            "summary": self._summarize_supplemental_leave(rows),
        }

    def parse_supplemental_leave_preview(self, filepath: str, calc_month: str) -> dict:
        """解析薪酬补充 sickpay&年假表，生成网页确认清单。"""
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        rows: list[dict] = []
        calc_month_text = str(calc_month)
        month_start = _calc_month_start(calc_month_text)
        scan_max_col = 80

        for sheet_name in wb.sheetnames:
            sheet_text = sheet_name.lower()
            if "sickpay" not in sheet_text and "年假" not in sheet_name and "病假" not in sheet_name:
                continue

            ws = wb[sheet_name]
            if hasattr(ws, "reset_dimensions"):
                ws.reset_dimensions()

            header_row = next(
                ws.iter_rows(min_row=1, max_row=1, max_col=scan_max_col, values_only=True),
                [],
            )
            headers = [str(value or "").strip() for value in header_row]

            def col(*names: str) -> Optional[int]:
                for name in names:
                    for index, header in enumerate(headers):
                        if header == name:
                            return index
                return None

            employee_col = col("工号")
            name_col = col("姓名")
            status_col = col("人员状态")
            start_col = col("申请开始时间", "开始时间", "申请时间")
            request_time_indexes = [i for i, header in enumerate(headers) if header == "申请时间"]
            end_col = col("申请结束时间", "结束时间")
            if end_col is None and len(request_time_indexes) > 1:
                end_col = request_time_indexes[1]
            hours_col = col("总时长", "时长", "小时数")
            flow_col = col("流程号")
            paid_col = col("发放时间")
            leave_type_col = col("假期类型")

            required = [employee_col, name_col, hours_col, leave_type_col]
            if any(index is None for index in required):
                continue

            inferred_period = infer_period_from_sheet_name(sheet_name, calc_month_text)
            explicit_allocation_cols: list[tuple[int, str, str]] = []
            fallback_allocation_col: tuple[int, str, str] | None = None
            for index, header in enumerate(headers):
                if index <= (leave_type_col or 0):
                    continue
                if not header:
                    continue
                current_period = _period_label_for_calc_month(header, calc_month_text)
                parsed_period = _parse_period_range(header)
                if current_period:
                    explicit_allocation_cols.append((index, current_period, calc_month_text))
                elif parsed_period:
                    start_month, *_ = parsed_period
                    target_month = (
                        calc_month_text
                        if start_month == int(calc_month_text.split("-")[1])
                        else _next_month(calc_month_text)
                    )
                    explicit_allocation_cols.append((index, header, target_month))
                elif "剔除离职" in header and fallback_allocation_col is None:
                    fallback_allocation_col = (index, inferred_period, calc_month_text)

            current_allocation_cols = [
                item for item in explicit_allocation_cols
                if item[2] == calc_month_text and item[1]
            ]
            future_allocation_cols = [
                item for item in explicit_allocation_cols
                if item[2] != calc_month_text
            ]
            allocation_cols = current_allocation_cols or ([fallback_allocation_col] if fallback_allocation_col else [])
            max_required_col = max(index for index in [
                employee_col,
                name_col,
                status_col,
                start_col,
                end_col,
                hours_col,
                flow_col,
                paid_col,
                leave_type_col,
                *[item[0] for item in explicit_allocation_cols],
                *((fallback_allocation_col[0],) if fallback_allocation_col else ()),
            ] if index is not None) + 1
            empty_row_streak = 0

            for row_number, row in enumerate(
                ws.iter_rows(min_row=2, max_col=max_required_col, values_only=True),
                start=2,
            ):
                employee_id = str(row[employee_col] or "").strip() if employee_col < len(row) else ""
                if not employee_id:
                    empty_row_streak += 1
                    if empty_row_streak >= 200:
                        break
                    continue
                empty_row_streak = 0

                raw_leave_type = str(row[leave_type_col] or "").strip() if leave_type_col < len(row) else ""
                if not raw_leave_type:
                    continue

                start_at = parse_excel_datetime(row[start_col]) if start_col is not None and start_col < len(row) else None
                end_at = parse_excel_datetime(row[end_col]) if end_col is not None and end_col < len(row) else None
                paid_at = parse_excel_datetime(row[paid_col]) if paid_col is not None and paid_col < len(row) else None
                hours = safe_float(row[hours_col]) if hours_col < len(row) else 0.0
                included_hours = hours
                allocation_period = inferred_period
                allocation_month = calc_month_text
                has_allocation_columns = bool(explicit_allocation_cols or fallback_allocation_col)
                if allocation_cols:
                    selected_hours = [
                        safe_float(row[index]) if index < len(row) else 0.0
                        for index, _, _ in allocation_cols
                    ]
                    included_hours = sum(selected_hours)
                    allocation_period = allocation_cols[0][1] or inferred_period
                    allocation_month = allocation_cols[0][2] or calc_month_text
                elif future_allocation_cols:
                    future_hours = [
                        safe_float(row[index]) if index < len(row) else 0.0
                        for index, _, _ in future_allocation_cols
                    ]
                    if sum(future_hours) > 0:
                        included_hours = 0.0
                        allocation_period = future_allocation_cols[0][1]
                        allocation_month = future_allocation_cols[0][2]
                employee_status = str(row[status_col] or "").strip() if status_col is not None and status_col < len(row) else ""
                roster_info = self.get_employee_info(employee_id)
                termination_date = (
                    roster_info.get("termination_date")
                    or (
                        roster_info.get("resignation_date").isoformat()
                        if roster_info.get("resignation_date")
                        else ""
                    )
                )
                if not employee_status:
                    employee_status = roster_info.get("employee_status") or roster_info.get("personnel_status", "")

                is_termination = employee_status == "离职" and raw_leave_type in {"病假", "年假"} and not start_at and not end_at
                leave_type = f"离职{raw_leave_type}结算" if is_termination else raw_leave_type
                if is_termination and raw_leave_type == "年假":
                    confirmation_status = "excluded"
                    include_in_base = False
                    included_hours = 0
                    system_reason = "离职年假结算属于年假兑现，默认不计入本月绩效基数。"
                elif is_termination and raw_leave_type == "病假":
                    if has_allocation_columns:
                        confirmation_status = "confirmed"
                        include_in_base = True
                        system_reason = "离职病假清算归属到当前周期时，计入病假清算基数。"
                    else:
                        confirmation_status = "pending"
                        include_in_base = False
                        included_hours = 0
                        system_reason = "离职病假清算缺少申请时间和周期归属列，需确认本月计入小时后再计入绩效基数。"
                else:
                    confirmation_status = "confirmed"
                    include_in_base = True
                    system_reason = "普通病假/年假按申请时间默认计入当前核算周期。"
                if has_allocation_columns and included_hours <= 0:
                    include_in_base = False
                    confirmation_status = "excluded"
                    system_reason = "假勤小时归属到非当前核算月份，系统默认不计入本月绩效基数。"
                elif not has_allocation_columns and start_at and end_at:
                    start_in_month = _same_month(start_at.date(), month_start)
                    end_in_month = _same_month(end_at.date(), month_start)
                    if not start_in_month and not end_in_month:
                        included_hours = 0
                        include_in_base = False
                        confirmation_status = "excluded"
                        system_reason = "假勤申请时间不属于当前核算月份，系统默认不计入本月绩效基数。"
                    elif start_in_month != end_in_month:
                        included_hours = 0
                        include_in_base = False
                        confirmation_status = "pending"
                        system_reason = "跨月假勤缺少周期拆分列，需人工填写本月计入小时后再计入绩效基数。"

                rows.append({
                    "row_id": f"{sheet_name}:{row_number}",
                    "source_sheet": sheet_name,
                    "source_row": row_number,
                    "employee_id": employee_id,
                    "name": str(row[name_col] or "").strip() if name_col < len(row) else "",
                    "employee_status": employee_status,
                    "termination_date": termination_date,
                    "leave_type": leave_type,
                    "raw_leave_type": raw_leave_type,
                    "start_at": start_at.isoformat(timespec="minutes") if start_at else "",
                    "end_at": end_at.isoformat(timespec="minutes") if end_at else "",
                    "hours": hours,
                    "included_hours": included_hours,
                    "flow_id": str(row[flow_col] or "").strip() if flow_col is not None and flow_col < len(row) else "",
                    "paid_at": paid_at.date().isoformat() if paid_at else "",
                    "source_rows": [row_number],
                    "is_termination_settlement": is_termination,
                    "suggested_month": calc_month_text,
                    "suggested_period": inferred_period,
                    "allocation_month": allocation_month,
                    "allocation_period": allocation_period,
                    "include_in_base": include_in_base,
                    "confirmation_status": confirmation_status,
                    "confirmation_note": "",
                    "system_reason": system_reason,
                })

        rows = self._merge_duplicate_supplemental_leave_rows(rows)
        return {
            "rows": rows,
            "summary": self._summarize_supplemental_leave(rows),
        }

    def apply_supplemental_leave_batch(
        self,
        preview: dict,
        row_ids: list[str],
        updates: dict,
    ) -> dict:
        """批量更新补充假勤确认字段。"""
        allowed_fields = {
            "confirmation_status",
            "include_in_base",
            "allocation_month",
            "allocation_period",
            "confirmation_note",
            "included_hours",
        }
        selected = set(row_ids or [])
        rows = []

        for row in preview.get("rows", []):
            next_row = dict(row)
            if next_row.get("row_id") in selected:
                status_changed = "confirmation_status" in updates
                include_changed = "include_in_base" in updates
                hours_changed = "included_hours" in updates
                for key, value in updates.items():
                    if key in allowed_fields:
                        next_row[key] = value
                if status_changed and not include_changed:
                    status = str(next_row.get("confirmation_status") or "")
                    if status == "confirmed":
                        next_row["include_in_base"] = True
                    elif status in {"excluded", "pending"}:
                        next_row["include_in_base"] = False
                if status_changed and not hours_changed:
                    if next_row.get("confirmation_status") == "excluded":
                        next_row["included_hours"] = 0
                if hours_changed and not status_changed and not include_changed:
                    if safe_float(next_row.get("included_hours")) > 0:
                        next_row["confirmation_status"] = "confirmed"
                        next_row["include_in_base"] = True
            rows.append(next_row)

        return {
            "rows": rows,
            "summary": self._summarize_supplemental_leave(rows),
        }

    def apply_supplemental_leave_all_suggestions(self, preview: dict) -> tuple[dict, int]:
        """Apply all pending supplemental leave suggestions with row-specific hours."""
        rows = []
        applied_count = 0

        for row in preview.get("rows", []):
            next_row = dict(row)
            suggested_hours = safe_float(next_row.get("suggested_included_hours"))
            if (
                next_row.get("suggestion_status") == "suggested"
                and next_row.get("confirmation_status") == "pending"
                and suggested_hours > 0
            ):
                next_row["included_hours"] = suggested_hours
                next_row["confirmation_status"] = "confirmed"
                next_row["include_in_base"] = True
                applied_count += 1
            rows.append(next_row)

        return {
            "rows": rows,
            "summary": self._summarize_supplemental_leave(rows),
        }, applied_count

    def apply_supplemental_leave_suggestions(
        self,
        preview: dict,
        employees: list[EmployeeData],
        offline_bases: list[dict] | dict | None,
        calc_month: str | None = None,
    ) -> dict:
        """按线下基数差异为待确认补充假勤生成安全建议小时。

        这里只生成建议，不直接确认计入。建议必须满足：
        - 线下基数高于平台基数；
        - 待确认行能精确匹配拆行工号，包括白班 -1；
        - 非离职年假结算、非固定基数/区长路径；
        - 反推小时不超过该行原始小时。
        """
        if not preview:
            return {"rows": [], "summary": self._summarize_supplemental_leave([])}

        rows = [
            {
                **row,
                "suggested_included_hours": "",
                "suggestion_status": "",
                "suggestion_reason": "",
            }
            for row in preview.get("rows", [])
        ]
        if not rows or not employees or not offline_bases:
            return {"rows": rows, "summary": self._summarize_supplemental_leave(rows)}

        offline_by_source: dict[str, float] = {}
        if isinstance(offline_bases, dict):
            iterable = offline_bases.values()
        else:
            iterable = offline_bases
        for row in iterable:
            source_id = normalize_shift_employee_id(row.get("employee_id") or row.get("source_employee_id"))
            amount = safe_float(row.get("offline_performance_base") or row.get("amount"))
            if source_id and amount > 0:
                offline_by_source[source_id] = round(offline_by_source.get(source_id, 0.0) + amount, 2)
        if not offline_by_source:
            return {"rows": rows, "summary": self._summarize_supplemental_leave(rows)}

        platform_base_by_source: dict[str, float] = defaultdict(float)
        exact_employee_rows: dict[str, list[EmployeeData]] = defaultdict(list)
        source_has_fixed_base_path: dict[str, bool] = defaultdict(bool)
        for employee in employees:
            exact_employee_rows[employee.employee_id].append(employee)
            source_id = normalize_shift_employee_id(employee.source_employee_id or employee.employee_id)
            platform_base_by_source[source_id] += safe_float(employee.performance_base)
            if employee.base_override_amount or (employee.job_type == "district_manager" and employee.fixed_performance_base):
                source_has_fixed_base_path[source_id] = True

        pending_by_source: dict[str, list[dict]] = defaultdict(list)
        for row in rows:
            if row.get("confirmation_status") != "pending":
                continue
            if calc_month and row.get("allocation_month") and row.get("allocation_month") != calc_month:
                continue
            pending_by_source[normalize_shift_employee_id(row.get("employee_id"))].append(row)

        def block(row: dict, reason: str) -> None:
            row["suggestion_status"] = "blocked"
            row["suggestion_reason"] = reason
            row["suggested_included_hours"] = ""

        for source_id, pending_rows in pending_by_source.items():
            if source_has_fixed_base_path.get(source_id):
                for row in pending_rows:
                    block(row, "固定基数/区长路径不生成建议计入，请按固定基数规则复核。")
                continue

            offline_base = offline_by_source.get(source_id, 0.0)
            platform_base = round(platform_base_by_source.get(source_id, 0.0), 2)
            missing_amount = round(offline_base - platform_base, 2)
            if missing_amount <= 0.05:
                continue

            candidates: list[dict] = []
            for row in pending_rows:
                leave_type = str(row.get("leave_type") or row.get("raw_leave_type") or "").strip()
                if leave_type not in {"病假", "年假", "离职病假结算", "病假清算", "离职年假结算"}:
                    block(row, "非病假/年假/病假清算类型，不生成建议计入。")
                    continue
                if leave_type == "离职年假结算":
                    block(row, "离职年假结算默认不计入，不生成建议计入。")
                    continue
                if row.get("attendance_matched") is False:
                    block(row, "未匹配本次考勤，不生成建议计入。")
                    continue
                if safe_float(row.get("hours")) <= 0:
                    block(row, "原始小时为空，不生成建议计入。")
                    continue
                candidates.append(row)

            if len(candidates) != 1:
                if len(candidates) > 1:
                    for row in candidates:
                        block(row, "同一员工存在多条待确认补充假勤，需人工判断计入小时。")
                continue

            row = candidates[0]
            exact_matches = exact_employee_rows.get(str(row.get("employee_id") or "").strip(), [])
            if len(exact_matches) != 1:
                block(row, "无法按精确工号匹配唯一考勤拆行，不生成建议计入。")
                continue

            employee = exact_matches[0]
            if employee.base_override_amount or (employee.job_type == "district_manager" and employee.fixed_performance_base):
                block(row, "固定基数/区长路径不生成建议计入，请按固定基数规则复核。")
                continue

            hourly_rate = safe_float(employee.hourly_rate)
            if hourly_rate <= 0:
                block(row, "精确考勤拆行缺少时薪，不生成建议计入。")
                continue

            suggested_hours = round(missing_amount / hourly_rate, 2)
            row_hours = safe_float(row.get("hours"))
            if suggested_hours <= 0:
                continue
            if suggested_hours > row_hours + 0.05:
                block(row, "线下差异反推小时超过原始小时，需人工确认。")
                continue

            row["suggested_included_hours"] = min(suggested_hours, row_hours)
            row["suggestion_status"] = "suggested"
            row["suggestion_reason"] = (
                f"线下基数差异反推: 差额{missing_amount:.2f} / 时薪{hourly_rate:.2f}"
            )

        return {
            "rows": rows,
            "summary": self._summarize_supplemental_leave(rows),
        }

    def parse_adjustments_preview(self, filepath: str) -> dict:
        """
        解析调薪/转正拆分表并返回预览。

        支持两种格式：
        1. 平台标准模板：按表头读取“工号/姓名/分段期间/分段绩效基数/核算标识”。
        2. 线下《仓库管理绩效基数》中的“调薪拆分”sheet固定列。
        3. OEHR原始调薪管理导出：仅作为调薪事件预览，不参与计算。
        """
        wb = self.load_excel(filepath)
        if "调薪拆分" not in wb.sheetnames:
            return self._parse_oehr_adjustments_preview(wb)

        ws = wb["调薪拆分"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("调薪拆分工作表为空")

        header_row_index = 0
        headers = rows[0]
        for index, row in enumerate(rows[:20]):
            if _find_column(row, ["工号", "员工工号", "employee_id"]) is not None:
                header_row_index = index
                headers = row
                break
        template_map = {
            "employee_id": _find_column(headers, ["工号", "员工工号", "employee_id"]),
            "name": _find_column(headers, ["姓名", "员工姓名", "name"]),
            "period": _find_column(headers, ["分段期间", "期间", "核算期间", "period"]),
            "performance_base": _find_column(headers, ["分段绩效基数", "绩效基数", "分段基数", "performance_base"]),
            "reason": _find_column(headers, ["核算标识", "调薪标识", "拆分标识", "reason"]),
        }
        use_template = all(template_map[key] is not None for key in ("employee_id", "period", "performance_base", "reason"))
        grouped: dict[str, dict] = {}

        data_rows = rows[header_row_index + 1:] if use_template else rows[1:]

        for row in data_rows:
            if use_template:
                emp_id = str(_cell(row, template_map["employee_id"]) or "").strip()
                name = str(_cell(row, template_map["name"]) or "").strip()
                period = str(_cell(row, template_map["period"]) or "").strip()
                amount = _to_float(_cell(row, template_map["performance_base"]), 0)
                reason = str(_cell(row, template_map["reason"]) or "").strip()
            else:
                emp_id = str(_cell(row, 3) or "").strip()
                name = str(_cell(row, 4) or "").strip()
                period = str(_cell(row, 9) or "").strip()
                amount = _to_float(_cell(row, 28), 0)
                reason = str(_cell(row, 31) or "").strip()

            if not emp_id or not emp_id.lower().startswith("zt"):
                continue

            if amount is None or not reason:
                continue

            emp_info = self.get_employee_info(emp_id)
            entry = grouped.setdefault(
                emp_id,
                {
                    "employee_id": emp_id,
                    "name": emp_info["name"] or name,
                    "department": emp_info["department"],
                    "area": emp_info["area"],
                    "segments": [],
                },
            )
            entry["segments"].append({
                "period": period,
                "reason": reason,
                "performance_base": round(amount, 2),
            })

        employees = list(grouped.values())
        total_segments = sum(len(emp["segments"]) for emp in employees)
        active_base = sum(
            segment["performance_base"]
            for emp in employees
            for segment in emp["segments"]
            if "前" not in segment["reason"]
        )

        return {
            "employees": employees,
            "summary": {
                "total_employees": len(employees),
                "total_segments": total_segments,
                "active_performance_base": round(active_base, 2),
            },
        }

    def _parse_oehr_adjustments_preview(self, wb: openpyxl.Workbook) -> dict:
        """兼容OEHR调薪管理导出；该格式不含分段绩效基数，不能直接参与核算。"""
        required_columns = {
            "employee_id": ["工号", "员工工号"],
            "name": ["姓名", "员工姓名"],
            "adjustment_type": ["调薪类型"],
            "adjustment_reason": ["调薪原因"],
            "effective_date": ["调薪生效日期"],
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            header_index = None
            headers = None
            for index, row in enumerate(rows[:20]):
                if all(_find_column(row, names) is not None for names in required_columns.values()):
                    header_index = index
                    headers = row
                    break
            if header_index is None or headers is None:
                continue

            col_map = {
                "subject": _find_column(headers, ["审批主题"]),
                "employee_id": _find_column(headers, required_columns["employee_id"]),
                "name": _find_column(headers, required_columns["name"]),
                "approval_status": _find_column(headers, ["审批状态"]),
                "department_levels": [
                    _find_column(headers, [label])
                    for label in ["二级部门", "三级部门", "四级部门", "五级部门", "六级部门", "七级部门", "八级部门"]
                ],
                "adjustment_type": _find_column(headers, required_columns["adjustment_type"]),
                "adjustment_reason": _find_column(headers, required_columns["adjustment_reason"]),
                "effective_date": _find_column(headers, required_columns["effective_date"]),
                "pay_system": _find_column(headers, ["调薪后薪酬制度"]),
                "cost_owner": _find_column(headers, ["调薪后成本归属"]),
                "hourly_rate": _find_column(headers, ["时薪标准", "基本工资标准"]),
                "performance_base": _find_column(headers, ["月度绩效奖金基数"]),
                "performance_ratio": _find_column(headers, ["月度绩效奖金比例(%)", "月度绩效奖金比例"]),
                "note": _find_column(headers, ["备注"]),
            }

            events = []
            for row in rows[header_index + 1:]:
                emp_id = str(_cell(row, col_map["employee_id"]) or "").strip()
                if not emp_id or not emp_id.lower().startswith("zt"):
                    continue

                effective_date = _to_date(_cell(row, col_map["effective_date"]))
                ratio = _to_float(_cell(row, col_map["performance_ratio"]), 0) or 0.0
                if ratio > 1:
                    ratio = ratio / 100
                department_parts = [
                    str(_cell(row, index) or "").strip()
                    for index in col_map["department_levels"]
                    if index is not None and str(_cell(row, index) or "").strip()
                ]
                emp_info = self.get_employee_info(emp_id)
                events.append({
                    "employee_id": emp_id,
                    "name": emp_info["name"] or str(_cell(row, col_map["name"]) or "").strip(),
                    "approval_status": str(_cell(row, col_map["approval_status"]) or "").strip(),
                    "subject": str(_cell(row, col_map["subject"]) or "").strip(),
                    "department": emp_info["department"] or "-".join(department_parts),
                    "adjustment_type": str(_cell(row, col_map["adjustment_type"]) or "").strip(),
                    "adjustment_reason": str(_cell(row, col_map["adjustment_reason"]) or "").strip(),
                    "effective_date": effective_date.isoformat() if effective_date else "",
                    "pay_system": str(_cell(row, col_map["pay_system"]) or "").strip(),
                    "cost_owner": str(_cell(row, col_map["cost_owner"]) or "").strip(),
                    "hourly_rate": round(_to_float(_cell(row, col_map["hourly_rate"]), 0) or 0.0, 2),
                    "performance_base": round(_to_float(_cell(row, col_map["performance_base"]), 0) or 0.0, 2),
                    "performance_ratio": round(ratio, 4),
                    "note": str(_cell(row, col_map["note"]) or "").strip(),
                })

            if events:
                auto_split_ready = sum(1 for event in events if _is_auto_adjustment_event(event))
                return {
                    "employees": [],
                    "events": events,
                    "warnings": [
                        "OEHR调薪管理导出已识别；转正调薪/新增绩效占比会在计算时按生效日和考勤日报自动拆分，其他调薪需上传平台调薪/转正拆分表。"
                    ],
                    "summary": {
                        "total_employees": 0,
                        "total_segments": 0,
                        "active_performance_base": 0,
                        "total_events": len(events),
                        "auto_split_ready": auto_split_ready,
                        "manual_split_required": len(events) - auto_split_ready,
                    },
                }

        raise ValueError("未找到“调薪拆分”工作表或可识别的OEHR调薪管理导出")

    def parse_all_from_step_data(
        self,
        attendance_data: list,
        salary_data: list,
        performance_data: list,
        adjustment_data: list | dict = None,
        transfer_data: dict | None = None,
        calc_month: str | None = None,
        supplemental_leave_data: dict | None = None,
        base_override_data: dict | None = None,
        hourly_rate_policy_data: dict | None = None,
        period_adjustment_data: dict | None = None,
    ) -> FBUPerformanceEngine:
        """
        从分步数据计算最终结果

        Args:
            attendance_data: 考勤预览数据中的employees列表
            salary_data: 薪资预览数据中的employees列表
            performance_data: 绩效预览数据中的employees列表
            adjustment_data: 调薪拆分预览数据，可传employees列表或完整preview
            transfer_data: 已完成人事调动记录，调动日期视为正式生效日期

        Returns:
            计算完成的引擎实例
        """
        if isinstance(adjustment_data, dict):
            adjustment_employees = adjustment_data.get("employees", [])
            adjustment_events = adjustment_data.get("events", [])
        else:
            adjustment_employees = adjustment_data or []
            adjustment_events = []

        # 转换为字典格式，并保存员工信息
        employee_info = {}  # 保存员工基本信息
        apply_hourly_rate_policies(attendance_data, hourly_rate_policy_data)
        attendance_dict = {}
        for emp in attendance_data:
            emp_id = normalize_shift_employee_id(
                emp.get("source_employee_id") or emp["employee_id"]
            )
            attendance_daily_rows = list(emp.get('attendance_daily_rows') or emp.get('daily_rows') or [])
            row_shift_type = emp.get('shift_type')
            if attendance_daily_rows and row_shift_type:
                attendance_daily_rows = [
                    {**row, "shift_type": row.get("shift_type") or row_shift_type}
                    for row in attendance_daily_rows
                ]
            day_daily_rows = [
                row for row in attendance_daily_rows
                if str(row.get("shift_type") or row_shift_type or "白班") != "夜班"
            ]
            night_daily_rows = [
                row for row in attendance_daily_rows
                if str(row.get("shift_type") or row_shift_type or "") == "夜班"
            ]
            attendance = attendance_dict.setdefault(emp_id, {
                'name': emp.get('name', ''),
                '白班': {
                    '计薪出勤': 0, 'OT1.5': 0, 'OT2.0': 0, '病假': 0,
                    '病假清算': 0, '年假': 0, '节假日': 0,
                },
                '夜班': {
                    '计薪出勤': 0, 'OT1.5': 0, 'OT2.0': 0, '病假': 0,
                    '病假清算': 0, '年假': 0, '节假日': 0,
                },
                'has_night_shift': False,
                'daily_rows': [],
                '白班_daily_rows': [],
                '夜班_daily_rows': [],
            })
            for shift_key, source in (('白班', emp['day_shift']), ('夜班', emp['night_shift'])):
                for hour_key in attendance[shift_key]:
                    attendance[shift_key][hour_key] += safe_float(source.get(hour_key))
            attendance['has_night_shift'] = attendance['has_night_shift'] or bool(emp['has_night_shift'])
            attendance['daily_rows'].extend(attendance_daily_rows)
            attendance['白班_daily_rows'].extend(day_daily_rows)
            attendance['夜班_daily_rows'].extend(night_daily_rows)
            # 保存员工信息
            employee_info[emp_id] = {
                'name': emp.get('name', ''),
                'department': emp.get('department', ''),
                'area': emp.get('area', ''),
                'personnel_status': emp.get('personnel_status', ''),
                'hire_date': _to_date(emp.get('hire_date')),
                'confirmation_date': _to_date(emp.get('confirmation_date')),
                'resignation_date': _to_date(emp.get('resignation_date')),
                'position': emp.get('position', ''),
                'job_type': classify_job_type(emp_id, emp.get('department', '')),
            }

        salary_dict = {}
        for emp in salary_data:
            emp_id = emp['employee_id']
            salary_dict[emp_id] = {
                'hourly_rate': emp['hourly_rate'],
                'ratio': emp['ratio'],
                'calculation_method': emp.get('calculation_method', ''),
                'fixed_performance_base': emp.get('fixed_performance_base', 0),
                'effective_segments': list(emp.get('effective_segments') or []),
                'verification_status': emp.get('verification_status', ''),
            }

        performance_dict = {}
        for emp in performance_data:
            emp_id = emp['employee_id']
            performance_dict[emp_id] = {
                'score': emp['score'],
                'level': emp['level'],
                'coefficient': emp['coefficient'],
            }

        adjustment_dict = {}
        for emp in adjustment_employees:
            emp_id = emp['employee_id']
            adjustment_dict[emp_id] = emp.get('segments', [])
        self._apply_salary_history_segments(
            salary_dict,
            attendance_dict,
            adjustment_dict,
        )
        self._apply_oehr_adjustment_events(
            adjustment_events,
            attendance_dict,
            salary_dict,
            employee_info,
            adjustment_dict,
            calc_month,
        )
        self._apply_transfer_history(
            transfer_data,
            attendance_dict,
            salary_dict,
            performance_dict,
            employee_info,
            adjustment_dict,
            calc_month,
        )

        # 构建员工数据
        employees = self.build_employees(
            attendance_dict,
            salary_dict,
            performance_dict,
            employee_info,
            adjustment_dict,
            calc_month=calc_month,
            supplemental_leave_data=supplemental_leave_data,
            base_override_data=base_override_data,
        )

        # 计算绩效奖金
        for emp in employees:
            BonusCalculator.calculate(emp)
            self.engine.add_employee(emp)
        self._apply_period_adjustments(employees, period_adjustment_data)

        return self.engine

    @staticmethod
    def _apply_period_adjustments(
        employees: list[EmployeeData],
        period_adjustment_data: dict | None,
    ) -> None:
        """Apply each employee's base adjustment once after split-row calculation."""
        rows = (period_adjustment_data or {}).get("rows", [])
        adjustments = {
            normalize_shift_employee_id(row.get("employee_id")): row
            for row in rows
            if normalize_shift_employee_id(row.get("employee_id"))
        }
        grouped: dict[str, list[EmployeeData]] = {}
        for employee in employees:
            source_id = normalize_shift_employee_id(
                employee.source_employee_id or employee.employee_id
            )
            grouped.setdefault(source_id, []).append(employee)

        for source_id, employee_rows in grouped.items():
            for employee in employee_rows:
                employee.system_performance_base = employee.performance_base
            adjustment = adjustments.get(source_id)
            if not adjustment:
                continue

            amount = safe_float(adjustment.get("amount"))
            target = next(
                (employee for employee in employee_rows if employee.employee_id == source_id),
                employee_rows[0],
            )
            final_base = target.performance_base + amount
            if final_base < 0:
                raise ValueError(
                    f"{source_id} 的 Period adjustment 使最终绩效基数小于0，请检查调整额"
                )
            target.period_adjustment = amount
            target.period_adjustment_source_month = str(adjustment.get("source_month") or "")
            target.period_adjustment_reason = str(adjustment.get("reason") or "")
            target.performance_base = final_base
            target.performance_bonus += (
                amount
                * target.performance_ratio
                * target.performance_coefficient
            )
