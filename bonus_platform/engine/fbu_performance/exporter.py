"""FBU绩效核算引擎 - 结果导出"""
from __future__ import annotations
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .engines.base import EmployeeData


def _format_job_type(job_type: str) -> str:
    if job_type == "district_manager":
        return "区长"
    if job_type == "functional":
        return "职能"
    return "仓库"


class FBUPerformanceExporter:
    """FBU绩效结果导出器"""

    # 表头样式
    HEADER_FONT = Font(bold=True, size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    # 边框样式
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def export_to_excel(
        self,
        employees: list[EmployeeData],
        output_path: str,
        summary: Optional[dict] = None,
    ) -> str:
        """
        导出核算结果到Excel

        Args:
            employees: 员工数据列表
            output_path: 输出文件路径
            summary: 汇总信息

        Returns:
            输出文件路径
        """
        wb = openpyxl.Workbook()

        # Sheet1: 绩效奖金明细
        ws1 = wb.active
        ws1.title = "绩效奖金明细"
        self._write_detail_sheet(ws1, employees)

        # Sheet2: 汇总统计
        if summary:
            ws2 = wb.create_sheet("汇总统计")
            self._write_summary_sheet(ws2, summary)

        # 保存文件
        wb.save(output_path)
        return output_path

    def _write_detail_sheet(self, ws, employees: list[EmployeeData]):
        """写入明细表"""
        headers = [
            "核算工号", "原始工号", "姓名", "岗位类型",
            "计薪出勤时长", "OT1.5时长", "OT2.0时长",
            "病假时长", "年假时长", "节假日时长",
            "时薪", "绩效比例",
            "基础工资", "OT1.5工资", "OT2.0工资",
            "病假工资", "年假补贴", "节日补贴",
            "绩效基数", "绩效得分", "绩效等级", "上传绩效系数", "系统绩效系数",
            "绩效奖金", "异常提示",
        ]

        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT_WHITE
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER

        # 写入数据
        for row, emp in enumerate(employees, 2):
            data = [
                emp.employee_id,
                emp.source_employee_id,
                emp.name,
                _format_job_type(emp.job_type),
                emp.base_hours,
                emp.ot15_hours,
                emp.ot20_hours,
                emp.sick_hours,
                emp.annual_hours,
                emp.holiday_hours,
                emp.hourly_rate,
                emp.performance_ratio,
                emp.base_salary,
                emp.ot15_salary,
                emp.ot20_salary,
                emp.sick_pay,
                emp.annual_leave_pay,
                emp.holiday_pay,
                emp.performance_base,
                emp.performance_score or "",
                emp.performance_level or "",
                emp.uploaded_coefficient if emp.uploaded_coefficient is not None else "",
                emp.performance_coefficient,
                emp.performance_bonus,
                "；".join(emp.exceptions),
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.THIN_BORDER
                # 数值列右对齐
                if col >= 4:
                    cell.alignment = Alignment(horizontal="right")

        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def _write_summary_sheet(self, ws, summary: dict):
        """写入汇总表"""
        ws.cell(row=1, column=1, value="FBU美洲绩效奖金核算汇总").font = Font(bold=True, size=14)

        row = 3
        for key, value in summary.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
