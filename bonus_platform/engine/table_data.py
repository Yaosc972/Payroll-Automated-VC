from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

from .models import CalculationResult, RecruitmentDetail, as_date


TABLE_ROWS_FILE = "table_rows.json"


def save_table_data(run_dir: Path, payload: Dict[str, Any]) -> Dict[str, Any]:
    path = run_dir / TABLE_ROWS_FILE
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def load_table_data(run_dir: Path) -> Dict[str, Any]:
    path = run_dir / TABLE_ROWS_FILE
    if not path.exists():
        return _empty_payload()
    return json.loads(path.read_text(encoding="utf-8"))


def build_table_data(run_id: str, result: CalculationResult) -> Dict[str, Any]:
    rows: List[Dict[str, Any]] = []
    for detail in result.details:
        rows.extend(_detail_rows(run_id, result.month, detail))
    rows.extend(_pending_rows(run_id, result.month, result.pending_confirmations))
    rows.extend(_exception_rows(run_id, result.month, result.exceptions))
    return _payload(run_id, result.month, rows)


def build_final_table_data(run_id: str, workbook_path: Path, month: int) -> Dict[str, Any]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        rows: List[Dict[str, Any]] = []
        rows.extend(_final_detail_rows(run_id, month, workbook))
        rows.extend(_final_exception_rows(run_id, month, workbook))
        return _payload(run_id, month, rows)
    finally:
        workbook.close()


def merge_diff_rows(run_dir: Path, metrics: Dict[str, Any]) -> Dict[str, Any]:
    payload = load_table_data(run_dir)
    rows = [row for row in payload.get("rows", []) if row.get("type") != "差异"]
    rows.extend(_diff_rows(payload.get("runId", run_dir.name), payload.get("month"), metrics))
    updated = _payload(payload.get("runId", run_dir.name), payload.get("month"), rows)
    return save_table_data(run_dir, updated)


def _detail_rows(run_id: str, month: int, detail: RecruitmentDetail) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    recruitment_amount = round(sum(amount for period, amount in detail.monthly_recruitment_amounts if period == month), 2)
    referral_amount = round(sum(amount for period, amount in detail.monthly_referral_amounts if period == month), 2)
    if recruitment_amount or not referral_amount:
        rows.append(_base_row(run_id, month, detail, "招聘奖金", recruitment_amount, "正常"))
    if referral_amount:
        rows.append(_base_row(run_id, month, detail, "内推奖金", referral_amount, "正常"))
    return rows


def _final_detail_rows(run_id: str, month: int, workbook) -> List[Dict[str, Any]]:
    if "招聘奖金明细" not in workbook.sheetnames:
        return []
    sheet = workbook["招聘奖金明细"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    header_cols = {header: index for index, header in enumerate(headers) if header}
    rows: List[Dict[str, Any]] = []
    for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {header: values[index] for header, index in header_cols.items() if index < len(values)}
        if not _has_final_detail_values(row):
            continue
        recruitment_amount = _final_current_amount(
            row,
            month,
            [
                ("招聘人入职1月奖金", "招聘人入职1月周期"),
                ("招聘人入职3月奖金", "招聘人入职3月周期"),
                ("招聘人入职6月奖金", "招聘人入职6月周期"),
                ("招聘人转正奖金", "招聘人转正周期"),
                ("协助人入职1月奖金", "协助人入职1月周期"),
                ("协助人入职3月奖金", "协助人入职3月周期"),
                ("协助人入职6月奖金", "协助人入职6月周期"),
                ("协助人转正奖金", "协助人转正周期"),
            ],
        )
        referral_amount = _final_current_amount(
            row,
            month,
            [
                ("内推入职1月奖金", "内推入职1月周期"),
                ("内推入职3月奖金", "内推入职3月周期"),
                ("内推入职6月奖金", "内推入职6月周期"),
                ("内推转正奖金", "内推转正周期"),
            ],
        )
        if recruitment_amount or not referral_amount:
            rows.append(_final_base_row(run_id, month, row, "招聘奖金", recruitment_amount, source_row))
        if referral_amount:
            rows.append(_final_base_row(run_id, month, row, "内推奖金", referral_amount, source_row))
    return rows


def _final_base_row(run_id: str, month: int, detail: Dict[str, Any], bonus_type: str, amount: float, source_row: int) -> Dict[str, Any]:
    payer_name = detail.get("招聘负责人姓名", "") if bonus_type == "招聘奖金" else detail.get("推荐人姓名", "")
    payer_id = detail.get("招聘负责人工号", "") if bonus_type == "招聘奖金" else detail.get("推荐人工号", "")
    row = {
        "id": f"{run_id}-{source_row}-{bonus_type}",
        "runId": run_id,
        "month": month,
        "type": bonus_type,
        "status": "正常",
        "employeeName": detail.get("姓名", ""),
        "employeeNo": detail.get("工号", ""),
        "grade": detail.get("职级", ""),
        "category": detail.get("ABC类别", ""),
        "location": detail.get("工作地", ""),
        "ruleScope": detail.get("标签分类", "") if bonus_type == "招聘奖金" else detail.get("内推规则范围", ""),
        "channel": detail.get("招聘渠道", ""),
        "node": _final_current_nodes(detail, month, bonus_type),
        "amount": amount,
        "currency": detail.get("币种", ""),
        "ownerName": payer_name,
        "ownerNo": payer_id,
        "recruiterName": detail.get("招聘负责人姓名", ""),
        "recruiterNo": detail.get("招聘负责人工号", ""),
        "assistantName": detail.get("协助招聘人姓名", ""),
        "assistantNo": detail.get("协助招聘人工号", ""),
        "referrerName": detail.get("推荐人姓名", ""),
        "referrerNo": detail.get("推荐人工号", ""),
        "message": detail.get("异常提示", ""),
        "sourceRow": source_row,
        "startDate": _format_final_date(detail.get("招聘启动日期")),
        "onboardDate": _format_final_date(detail.get("候选人入职时间")),
        "probationDate": _format_final_date(detail.get("转正日期")),
        "calculation": {
            "standardBonus": _number(detail.get("招聘奖金标准")),
            "adjustedTotalBonus": _number(detail.get("招聘人实际发放标准")) + _number(detail.get("协助人实际发放标准")),
            "recruitmentStandardBonus": _number(detail.get("招聘奖金标准")),
            "referralStandardBonus": _number(detail.get("内推奖金标准")),
            "channelRatio": _number(detail.get("渠道系数")),
            "standardCycleDays": _number(detail.get("标准入职周期")),
            "actualCycleDays": _number(detail.get("实际入职周期")),
            "cycleDiffDays": _number(detail.get("入职周期差异")),
            "nodes": _final_node_details(detail),
        },
    }
    row["searchText"] = _search_text(row)
    return row


def _final_exception_rows(run_id: str, month: int, workbook) -> List[Dict[str, Any]]:
    if "异常清单" not in workbook.sheetnames:
        return []
    sheet = workbook["异常清单"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    rows = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        exception = {header: value for header, value in zip(headers, values) if header}
        if not any(value not in (None, "") for value in exception.values()):
            continue
        row = {
            "id": f"{run_id}-exception-{index}",
            "runId": run_id,
            "month": month,
            "type": "异常",
            "status": "异常",
            "employeeName": exception.get("姓名", ""),
            "employeeNo": exception.get("工号", ""),
            "grade": "",
            "category": "",
            "location": "",
            "ruleScope": "",
            "channel": "",
            "node": "",
            "amount": 0,
            "currency": "",
            "ownerName": "",
            "ownerNo": "",
            "recruiterName": "",
            "recruiterNo": "",
            "assistantName": "",
            "assistantNo": "",
            "referrerName": "",
            "referrerNo": "",
            "message": exception.get("异常类型", ""),
            "sourceRow": exception.get("源行号", ""),
            "calculation": {"nodes": [], "exception": exception},
        }
        row["searchText"] = _search_text(row)
        rows.append(row)
    return rows


def _base_row(run_id: str, month: int, detail: RecruitmentDetail, bonus_type: str, amount: float, status: str) -> Dict[str, Any]:
    payer_name = detail.recruiter_name if bonus_type == "招聘奖金" else detail.referrer_name
    payer_id = detail.recruiter_id if bonus_type == "招聘奖金" else detail.referrer_id
    row = {
        "id": f"{run_id}-{detail.row_no}-{bonus_type}",
        "runId": run_id,
        "month": month,
        "type": bonus_type,
        "status": status,
        "employeeName": detail.name,
        "employeeNo": detail.employee_no,
        "grade": detail.grade,
        "category": detail.category,
        "location": detail.location,
        "ruleScope": detail.label if bonus_type == "招聘奖金" else detail.referral_rule_scope,
        "channel": detail.channel,
        "node": _current_nodes(detail, month, bonus_type),
        "amount": amount,
        "currency": detail.currency,
        "ownerName": payer_name,
        "ownerNo": payer_id,
        "recruiterName": detail.recruiter_name,
        "recruiterNo": detail.recruiter_id,
        "assistantName": detail.assistant_name,
        "assistantNo": detail.assistant_id,
        "referrerName": detail.referrer_name,
        "referrerNo": detail.referrer_id,
        "message": "；".join(detail.exceptions),
        "sourceRow": detail.row_no,
        "startDate": _format_date(detail.start_date),
        "onboardDate": _format_date(detail.onboard_date),
        "probationDate": _format_date(detail.probation_date),
        "calculation": {
            "standardBonus": detail.standard_bonus,
            "adjustedTotalBonus": detail.adjusted_total_bonus,
            "recruitmentStandardBonus": detail.standard_bonus,
            "referralStandardBonus": detail.referral_standard_bonus,
            "channelRatio": detail.channel_ratio,
            "standardCycleDays": detail.standard_cycle_days,
            "actualCycleDays": detail.actual_cycle_days,
            "cycleDiffDays": detail.cycle_diff_days,
            "nodes": _node_details(detail),
        },
    }
    row["searchText"] = _search_text(row)
    return row


def _pending_rows(run_id: str, month: int, pending_rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for index, pending in enumerate(pending_rows, start=1):
        row = {
            "id": f"{run_id}-pending-{index}",
            "runId": run_id,
            "month": month,
            "type": pending.get("奖金类型") or "待确认",
            "status": "待确认",
            "employeeName": pending.get("姓名", ""),
            "employeeNo": pending.get("工号", ""),
            "grade": "",
            "category": "",
            "location": "",
            "ruleScope": "",
            "channel": "",
            "node": pending.get("发放节点", ""),
            "amount": pending.get("建议发放金额", 0),
            "currency": pending.get("币种", ""),
            "ownerName": pending.get("发放对象姓名", ""),
            "ownerNo": pending.get("发放对象工号", ""),
            "recruiterName": "",
            "recruiterNo": "",
            "assistantName": "",
            "assistantNo": "",
            "referrerName": "",
            "referrerNo": "",
            "message": pending.get("系统提示", ""),
            "sourceRow": pending.get("源行号", ""),
            "calculation": {"nodes": [], "pending": pending},
        }
        row["searchText"] = _search_text(row)
        rows.append(row)
    return rows


def _exception_rows(run_id: str, month: int, exception_rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for index, exception in enumerate(exception_rows, start=1):
        row = {
            "id": f"{run_id}-exception-{index}",
            "runId": run_id,
            "month": month,
            "type": "异常",
            "status": "异常",
            "employeeName": exception.get("姓名", ""),
            "employeeNo": exception.get("工号", ""),
            "grade": "",
            "category": "",
            "location": "",
            "ruleScope": "",
            "channel": "",
            "node": "",
            "amount": 0,
            "currency": "",
            "ownerName": "",
            "ownerNo": "",
            "recruiterName": "",
            "recruiterNo": "",
            "assistantName": "",
            "assistantNo": "",
            "referrerName": "",
            "referrerNo": "",
            "message": exception.get("异常类型", ""),
            "sourceRow": exception.get("源行号", ""),
            "calculation": {"nodes": [], "exception": exception},
        }
        row["searchText"] = _search_text(row)
        rows.append(row)
    return rows


def _diff_rows(run_id: str, month: int | None, metrics: Dict[str, Any]) -> List[Dict[str, Any]]:
    mapping = [
        ("招聘汇总差异", "recruitmentSummaryDiffCount", "recruitmentSummaryDelta"),
        ("内推汇总差异", "referralSummaryDiffCount", "referralSummaryDelta"),
        ("招聘明细差异", "recruitmentDetailDiffCount", ""),
        ("内推明细差异", "referralDetailDiffCount", ""),
    ]
    rows = []
    for label, count_key, delta_key in mapping:
        count = int(metrics.get(count_key) or 0)
        if count <= 0:
            continue
        amount = metrics.get(delta_key) if delta_key else 0
        row = {
            "id": f"{run_id}-diff-{count_key}",
            "runId": run_id,
            "month": month,
            "type": "差异",
            "status": "差异",
            "employeeName": "",
            "employeeNo": "",
            "grade": "",
            "category": "",
            "location": "",
            "ruleScope": "",
            "channel": "",
            "node": label,
            "amount": amount or 0,
            "currency": "",
            "ownerName": "",
            "ownerNo": "",
            "recruiterName": "",
            "recruiterNo": "",
            "assistantName": "",
            "assistantNo": "",
            "referrerName": "",
            "referrerNo": "",
            "message": f"{label} {count} 行",
            "sourceRow": "",
            "calculation": {"nodes": [], "diffMetrics": metrics},
        }
        row["searchText"] = _search_text(row)
        rows.append(row)
    return rows


def _payload(run_id: str, month: int | None, rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "runId": run_id,
        "month": month,
        "rows": rows,
        "filters": {
            "bonusTypes": _unique(row.get("type") for row in rows if row.get("type") not in {"异常", "差异"}),
            "statuses": _unique(row.get("status") for row in rows),
            "ruleScopes": _unique(row.get("ruleScope") for row in rows),
            "owners": _unique(row.get("ownerName") or row.get("ownerNo") for row in rows),
        },
        "stats": {
            "totalRows": len(rows),
            "normalRows": sum(1 for row in rows if row.get("status") == "正常"),
            "pendingRows": sum(1 for row in rows if row.get("status") == "待确认"),
            "exceptionRows": sum(1 for row in rows if row.get("status") == "异常"),
            "diffRows": sum(1 for row in rows if row.get("status") == "差异"),
            "amountRows": sum(1 for row in rows if _number(row.get("amount")) != 0),
        },
    }


def _empty_payload() -> Dict[str, Any]:
    return _payload("", None, [])


def _current_nodes(detail: RecruitmentDetail, month: int, bonus_type: str) -> str:
    nodes = _node_details(detail)
    prefix = "内推" if bonus_type == "内推奖金" else ""
    if bonus_type == "招聘奖金":
        matched = [node["label"] for node in nodes if node["type"] == "招聘奖金" and node["period"] == month and _number(node["amount"])]
    else:
        matched = [node["label"] for node in nodes if node["type"] == "内推奖金" and node["period"] == month and _number(node["amount"])]
    return "、".join(matched) or prefix or bonus_type


def _node_details(detail: RecruitmentDetail) -> List[Dict[str, Any]]:
    return [
        *_recruitment_nodes("招聘负责人", detail.recruiter_name, detail.recruiter_id, detail),
        *_assistant_nodes(detail),
        *_referral_nodes(detail),
    ]


def _recruitment_nodes(role: str, name: str, employee_no: str, detail: RecruitmentDetail) -> List[Dict[str, Any]]:
    return [
        _node("招聘奖金", role, name, employee_no, "入职1月", detail.recruiter_1m_period, detail.recruiter_1m_bonus),
        _node("招聘奖金", role, name, employee_no, "入职3月", detail.recruiter_3m_period, detail.recruiter_3m_bonus),
        _node("招聘奖金", role, name, employee_no, "入职6月", detail.recruiter_6m_period, detail.recruiter_6m_bonus),
        _node("招聘奖金", role, name, employee_no, "转正", detail.recruiter_probation_period, detail.recruiter_probation_bonus),
    ]


def _assistant_nodes(detail: RecruitmentDetail) -> List[Dict[str, Any]]:
    return [
        _node("招聘奖金", "协助招聘人", detail.assistant_name, detail.assistant_id, "入职1月", detail.assistant_1m_period, detail.assistant_1m_bonus),
        _node("招聘奖金", "协助招聘人", detail.assistant_name, detail.assistant_id, "入职3月", detail.assistant_3m_period, detail.assistant_3m_bonus),
        _node("招聘奖金", "协助招聘人", detail.assistant_name, detail.assistant_id, "入职6月", detail.assistant_6m_period, detail.assistant_6m_bonus),
        _node("招聘奖金", "协助招聘人", detail.assistant_name, detail.assistant_id, "转正", detail.assistant_probation_period, detail.assistant_probation_bonus),
    ]


def _referral_nodes(detail: RecruitmentDetail) -> List[Dict[str, Any]]:
    return [
        _node("内推奖金", "推荐人", detail.referrer_name, detail.referrer_id, "入职1月", detail.referral_1m_period, detail.referral_1m_bonus),
        _node("内推奖金", "推荐人", detail.referrer_name, detail.referrer_id, "入职3月", detail.referral_3m_period, detail.referral_3m_bonus),
        _node("内推奖金", "推荐人", detail.referrer_name, detail.referrer_id, "入职6月", detail.referral_6m_period, detail.referral_6m_bonus),
        _node("内推奖金", "推荐人", detail.referrer_name, detail.referrer_id, "转正", detail.referral_probation_period, detail.referral_probation_bonus),
    ]


def _node(node_type: str, role: str, name: str, employee_no: str, label: str, period: Any, amount: Any) -> Dict[str, Any]:
    return {
        "type": node_type,
        "role": role,
        "name": name,
        "employeeNo": employee_no,
        "label": label,
        "period": period,
        "amount": amount,
    }


def _format_date(value: Any) -> str:
    date_value = as_date(value)
    if isinstance(date_value, datetime):
        return f"{date_value.year}/{date_value.month}/{date_value.day}"
    return ""


def _has_final_detail_values(row: Dict[str, Any]) -> bool:
    return any(str(row.get(header) or "").strip() for header in ("唯一验证", "姓名", "工号", "招聘负责人姓名", "推荐人姓名"))


def _final_current_amount(row: Dict[str, Any], month: int, fields: Iterable[tuple[str, str]]) -> float:
    amount = 0.0
    for amount_header, period_header in fields:
        if _period_value(row.get(period_header)) == month:
            amount += _number(row.get(amount_header))
    return round(amount, 2)


def _period_value(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.year * 100 + value.month
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    digits = "".join(char for char in text if char.isdigit())
    if len(digits) >= 6:
        return int(digits[:6])
    return None


def _format_final_date(value: Any) -> str:
    date_value = as_date(value)
    if isinstance(date_value, datetime):
        return f"{date_value.year}/{date_value.month}/{date_value.day}"
    return str(value).strip() if value not in (None, "") else ""


def _final_current_nodes(detail: Dict[str, Any], month: int, bonus_type: str) -> str:
    nodes = _final_node_details(detail)
    matched = [
        node["label"]
        for node in nodes
        if node["type"] == bonus_type and _period_value(node["period"]) == month and _number(node["amount"])
    ]
    return "、".join(matched) or ("内推" if bonus_type == "内推奖金" else bonus_type)


def _final_node_details(detail: Dict[str, Any]) -> List[Dict[str, Any]]:
    return [
        _node("招聘奖金", "招聘负责人", detail.get("招聘负责人姓名", ""), detail.get("招聘负责人工号", ""), "入职1月", detail.get("招聘人入职1月周期"), detail.get("招聘人入职1月奖金")),
        _node("招聘奖金", "招聘负责人", detail.get("招聘负责人姓名", ""), detail.get("招聘负责人工号", ""), "入职3月", detail.get("招聘人入职3月周期"), detail.get("招聘人入职3月奖金")),
        _node("招聘奖金", "招聘负责人", detail.get("招聘负责人姓名", ""), detail.get("招聘负责人工号", ""), "入职6月", detail.get("招聘人入职6月周期"), detail.get("招聘人入职6月奖金")),
        _node("招聘奖金", "招聘负责人", detail.get("招聘负责人姓名", ""), detail.get("招聘负责人工号", ""), "转正", detail.get("招聘人转正周期"), detail.get("招聘人转正奖金")),
        _node("招聘奖金", "协助招聘人", detail.get("协助招聘人姓名", ""), detail.get("协助招聘人工号", ""), "入职1月", detail.get("协助人入职1月周期"), detail.get("协助人入职1月奖金")),
        _node("招聘奖金", "协助招聘人", detail.get("协助招聘人姓名", ""), detail.get("协助招聘人工号", ""), "入职3月", detail.get("协助人入职3月周期"), detail.get("协助人入职3月奖金")),
        _node("招聘奖金", "协助招聘人", detail.get("协助招聘人姓名", ""), detail.get("协助招聘人工号", ""), "入职6月", detail.get("协助人入职6月周期"), detail.get("协助人入职6月奖金")),
        _node("招聘奖金", "协助招聘人", detail.get("协助招聘人姓名", ""), detail.get("协助招聘人工号", ""), "转正", detail.get("协助人转正周期"), detail.get("协助人转正奖金")),
        _node("内推奖金", "推荐人", detail.get("推荐人姓名", ""), detail.get("推荐人工号", ""), "入职1月", detail.get("内推入职1月周期"), detail.get("内推入职1月奖金")),
        _node("内推奖金", "推荐人", detail.get("推荐人姓名", ""), detail.get("推荐人工号", ""), "入职3月", detail.get("内推入职3月周期"), detail.get("内推入职3月奖金")),
        _node("内推奖金", "推荐人", detail.get("推荐人姓名", ""), detail.get("推荐人工号", ""), "入职6月", detail.get("内推入职6月周期"), detail.get("内推入职6月奖金")),
        _node("内推奖金", "推荐人", detail.get("推荐人姓名", ""), detail.get("推荐人工号", ""), "转正", detail.get("内推转正周期"), detail.get("内推转正奖金")),
    ]


def _search_text(row: Dict[str, Any]) -> str:
    fields = ["employeeName", "employeeNo", "ownerName", "ownerNo", "recruiterName", "recruiterNo", "assistantName", "assistantNo", "referrerName", "referrerNo", "location", "message"]
    return " ".join(str(row.get(field) or "") for field in fields).lower()


def _unique(values: Iterable[Any]) -> List[str]:
    result = sorted({str(value).strip() for value in values if str(value or "").strip()})
    return result


def _number(value: Any) -> float:
    if value in (None, "", "-"):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0
